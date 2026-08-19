"""
Final comprehensive analysis - proper header row detection
"""
import openpyxl
from pathlib import Path
from collections import defaultdict

def get_data_with_headers(ws):
    """Extract data with proper header detection (row 2)"""
    # Row 1 is navigation, Row 2 is headers, Row 3 is "Start", Row 4+ is data
    headers = [cell.value for cell in ws[2]]
    
    data_rows = []
    for row in ws.iter_rows(min_row=4, values_only=True):  # Skip navigation, headers, and "Start"
        if row and any(cell for cell in row):
            data_rows.append(row)
    
    return headers, data_rows


def analyze_pcf_detailed(spec_file, prefix):
    """Detailed PipingCommodityFilter analysis"""
    print(f"\n{'='*100}")
    print(f"ISSUE #1: PIPING COMMODITY FILTER ANALYSIS - {prefix} (30% admin work)")
    print(f"{'='*100}")
    
    wb = openpyxl.load_workbook(spec_file, data_only=True)
    
    if 'PipingCommodityFilter' not in wb.sheetnames:
        print("  Sheet not found!")
        wb.close()
        return
    
    ws = wb['PipingCommodityFilter']
    headers, rows = get_data_with_headers(ws)
    
    print(f"\n  Total data rows: {len(rows)}")
    print(f"  Columns: {', '.join([h for h in headers if h])}")
    
    # Analyze by component type
    shortcode_idx = headers.index('ShortCode') if 'ShortCode' in headers else None
    size_from_idx = headers.index('FirstSizeFrom') if 'FirstSizeFrom' in headers else None
    size_to_idx = headers.index('FirstSizeTo') if 'FirstSizeTo' in headers else None
    fluid_idx = headers.index('FluidCode') if 'FluidCode' in headers else None
    spec_idx = headers.index('SpecName') if 'SpecName' in headers else None
    
    if shortcode_idx is not None:
        shortcode_counts = defaultdict(int)
        for row in rows:
            if shortcode_idx < len(row) and row[shortcode_idx]:
                shortcode_counts[row[shortcode_idx]] += 1
        
        print(f"\n  Entries by ShortCode:")
        for shortcode, count in sorted(shortcode_counts.items(), key=lambda x: -x[1])[:20]:
            print(f"    {shortcode:30s}: {count:4d} entries")
    
    # Sample piping entries
    print(f"\n  Sample PIPING entries (first 20):")
    print(f"  {'#':<4} {'SpecName':<10} {'ShortCode':<30} {'SizeFrom':<10} {'SizeTo':<10} {'FluidCode':<12}")
    print(f"  {'-'*4} {'-'*10} {'-'*30} {'-'*10} {'-'*10} {'-'*12}")
    
    piping_count = 0
    for idx, row in enumerate(rows, start=1):
        if shortcode_idx < len(row) and row[shortcode_idx] == 'Piping':
            spec = row[spec_idx] if spec_idx and spec_idx < len(row) else 'N/A'
            shortcode = row[shortcode_idx] if shortcode_idx < len(row) else 'N/A'
            size_from = row[size_from_idx] if size_from_idx and size_from_idx < len(row) else 'N/A'
            size_to = row[size_to_idx] if size_to_idx and size_to_idx < len(row) else 'N/A'
            fluid = row[fluid_idx] if fluid_idx and fluid_idx < len(row) else 'N/A'
            
            print(f"  {idx:<4} {str(spec):<10} {str(shortcode):<30} {str(size_from):<10} {str(size_to):<10} {str(fluid):<12}")
            
            piping_count += 1
            if piping_count >= 20:
                break
    
    wb.close()


def analyze_matl_control_detailed(spec_file, prefix):
    """Detailed PipingCommodityMatlControlData analysis"""
    print(f"\n{'='*100}")
    print(f"ISSUE #2: PIPING COMMODITY MATL CONTROL DATA ANALYSIS - {prefix} (20% admin work)")
    print(f"{'='*100}")
    print(f"\nPROBLEM: Descriptions are inappropriate - should come from CAT component descriptions")
    
    wb = openpyxl.load_workbook(spec_file, data_only=True)
    
    sheet_name = 'PipingCommodityMatlControlData'
    
    if sheet_name not in wb.sheetnames:
        print(f"  Sheet not found!")
        wb.close()
        return
    
    ws = wb[sheet_name]
    headers, rows = get_data_with_headers(ws)
    
    print(f"\n  Total data rows: {len(rows)}")
    print(f"  Columns: {', '.join([h for h in headers if h])}")
    
    # Check descriptions
    desc_idx = headers.index('Description') if 'Description' in headers else None
    commodity_idx = headers.index('CommodityCode') if 'CommodityCode' in headers else None
    shortcode_idx = headers.index('ShortCode') if 'ShortCode' in headers else None
    
    print(f"\n  Current descriptions (first 20):")
    print(f"  {'#':<4} {'CommodityCode':<45} {'ShortCode':<30} {'Description':<50}")
    print(f"  {'-'*4} {'-'*45} {'-'*30} {'-'*50}")
    
    for idx, row in enumerate(rows[:20], start=1):
        commodity = str(row[commodity_idx] if commodity_idx and commodity_idx < len(row) else 'N/A')[:45]
        shortcode = str(row[shortcode_idx] if shortcode_idx and shortcode_idx < len(row) else 'N/A')[:30]
        desc = str(row[desc_idx] if desc_idx and desc_idx < len(row) else 'N/A')[:50]
        print(f"  {idx:<4} {commodity:<45} {shortcode:<30} {desc:<50}")
    
    wb.close()


def analyze_cat_components_detailed(cat_file, prefix):
    """Detailed CAT component analysis"""
    print(f"\n{'='*100}")
    print(f"ISSUE #3: CAT COMPONENT/PART SHEETS ANALYSIS - {prefix} (50% admin work)")
    print(f"{'='*100}")
    print(f"\nPROBLEM: Part sheets should be populated from paper spec data")
    
    wb = openpyxl.load_workbook(cat_file, data_only=True)
    
    # Analyze valve sheets as examples
    valve_sheets = [name for name in wb.sheetnames if 'Valve' in name]
    
    print(f"\n  Found {len(valve_sheets)} valve sheets: {', '.join(valve_sheets)}")
    
    # Analyze GateValve in detail
    if 'GateValve' in wb.sheetnames:
        print(f"\n  === GateValve Sheet Analysis ===")
        ws = wb['GateValve']
        headers, rows = get_data_with_headers(ws)
        
        print(f"  Total rows: {len(rows)}")
        print(f"  Columns: {', '.join([h for h in headers if h][:20])}")
        
        # Show sample data
        if rows:
            print(f"\n  Sample GateValve entries (first 5):")
            for idx, row in enumerate(rows[:5], start=1):
                print(f"\n  Row {idx}:")
                for col_idx, (header, value) in enumerate(zip(headers, row)):
                    if header and value and col_idx < 15:  # First 15 columns only
                        print(f"    {header}: {str(value)[:60]}")
    
    wb.close()


def main():
    """Main analysis"""
    files = [
        ('AA1B-A3_SPEC.xlsx', 'AA1B-A3_CAT.xlsx', 'AA1B-A3'),
    ]
    
    print("="*100)
    print("SMARTPLANT 3D GAP ANALYSIS - DETAILED")
    print("="*100)
    print("\nOBJECTIVE: Reduce admin work by fixing 3 issues:")
    print("  1. PipingCommodityFilter sheet not filling properly (30%)")
    print("  2. PipingCommodityMatlControlData descriptions inappropriate (20%)")
    print("  3. CAT Part sheets not populated from spec data (50%)")
    
    for spec_file, cat_file, prefix in files:
        spec_path = Path(spec_file)
        cat_path = Path(cat_file)
        
        if not spec_path.exists() or not cat_path.exists():
            print(f"\nFiles not found for {prefix}")
            continue
        
        try:
            analyze_pcf_detailed(spec_path, prefix)
            analyze_matl_control_detailed(spec_path, prefix)
            analyze_cat_components_detailed(cat_path, prefix)
        except Exception as e:
            print(f"\nError: {e}")
            import traceback
            traceback.print_exc()
    
    print(f"\n{'='*100}")
    print("ANALYSIS COMPLETE - Ready to implement fixes")
    print(f"{'='*100}")


if __name__ == "__main__":
    main()
