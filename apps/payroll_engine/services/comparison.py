"""Comparison service — reconcile a Payroll Run against an external HR file
(ValueFrame timesheet, Sympa salary master, or any generic XLSX/CSV).

The parser is driven by ``catalog.COMPARISON_PROFILES`` so adding a new
vendor never requires editing this file — just add a profile entry.

Public entry point:
    run_comparison(run, file_obj, *, source_label, source_profile,
                   uploaded_by=None) -> PayrollComparison
"""
from __future__ import annotations
import difflib
import re
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Dict, List, Optional, Tuple

from django.db import transaction

from .. import catalog, config
from ..models import (
    PayrollComparison, PayrollComparisonRow, PayrollEmployee, PayrollRun, Payslip,
)
from .calculator import to_decimal


# ── Header normalisation ──────────────────────────────────────────
def _norm(s) -> str:
    """Collapse whitespace, lowercase, strip punctuation noise."""
    if s is None:
        return ''
    s = str(s).strip().lower()
    s = re.sub(r'\s+', ' ', s)
    return s


def _to_dec(value) -> Optional[Decimal]:
    if value in (None, ''):
        return None
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        try:
            return Decimal(str(value))
        except (InvalidOperation, ValueError):
            return None
    # String — strip currency / commas
    raw = str(value).strip().replace(',', '').replace('AED', '').replace('$', '')
    raw = raw.strip()
    if not raw or raw.lower() in ('-', 'n/a', 'na'):
        return None
    try:
        return Decimal(raw)
    except (InvalidOperation, ValueError):
        return None


# ── Workbook reading ──────────────────────────────────────────────
def _open_workbook(file_or_bytes):
    import openpyxl
    if hasattr(file_or_bytes, 'read'):
        return openpyxl.load_workbook(file_or_bytes, data_only=True, read_only=True)
    if isinstance(file_or_bytes, bytes):
        return openpyxl.load_workbook(BytesIO(file_or_bytes), data_only=True, read_only=True)
    return openpyxl.load_workbook(str(file_or_bytes), data_only=True, read_only=True)


def _locate_header_row(ws, profile: Dict) -> int:
    """Either use the profile's fixed header_row or scan until we find a
    row whose cells look like a header (at least 3 non-empty alphanum
    cells AND we can map at least one canonical field to it).
    Returns 1-based row index. Falls back to 1.
    """
    fixed = profile.get('header_row')
    if fixed:
        return int(fixed)
    max_scan = int(profile.get('header_scan_max') or 20)
    aliases = _build_alias_lookup(profile)
    for r_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if r_idx > max_scan:
            break
        non_empty = [c for c in row if c not in (None, '')]
        if len(non_empty) < 3:
            continue
        normalised = [_norm(c) for c in row]
        # Did this row map at least one canonical field?
        if any(aliases.get(h) for h in normalised if h):
            return r_idx
    return 1


# ── Alias resolution ──────────────────────────────────────────────
def _build_alias_lookup(profile: Dict) -> Dict[str, str]:
    """Return {normalised_header: canonical_field}. Profile aliases take
    precedence over universal aliases."""
    out: Dict[str, str] = {}
    # Universal first
    for field, headers in catalog.COMPARISON_FIELD_ALIASES.items():
        for h in headers:
            out.setdefault(_norm(h), field)
    # Profile overrides win
    for field, headers in (profile.get('field_aliases') or {}).items():
        for h in headers:
            out[_norm(h)] = field
    # Compose fields
    for field, compose in (profile.get('compose') or {}).items():
        for h in compose.get('parts') or []:
            out.setdefault(_norm(h), f'__compose__{field}')
    return out


def _resolve_columns(headers: List[str], profile: Dict) -> Tuple[Dict[str, int], Dict[str, str], Dict[str, dict]]:
    """Map headers → column indices.

    Returns:
        col_index   {canonical_field: 0-based col idx}
        col_header  {canonical_field: original header text}  (for audit)
        compose     {canonical_field: {'parts': [col_idx, ...], 'separator': str}}
    """
    aliases = _build_alias_lookup(profile)
    col_index: Dict[str, int] = {}
    col_header: Dict[str, str] = {}
    compose_parts: Dict[str, List[int]] = {}

    for idx, raw in enumerate(headers):
        n = _norm(raw)
        if not n:
            continue
        field = aliases.get(n)
        if not field:
            continue
        if field.startswith('__compose__'):
            target = field.replace('__compose__', '')
            compose_parts.setdefault(target, []).append(idx)
            continue
        col_index.setdefault(field, idx)
        col_header.setdefault(field, str(raw))

    compose: Dict[str, dict] = {}
    for target, indices in compose_parts.items():
        rule = (profile.get('compose') or {}).get(target) or {}
        compose[target] = {
            'parts': indices,
            'separator': rule.get('separator', ' '),
        }

    return col_index, col_header, compose


# ── External row → canonical dict ─────────────────────────────────
def _row_to_canonical(row: tuple, col_index: Dict[str, int],
                      compose: Dict[str, dict]) -> Dict:
    out: Dict = {}
    for field, idx in col_index.items():
        out[field] = row[idx] if idx < len(row) else None
    for field, rule in compose.items():
        parts = []
        for idx in rule['parts']:
            if idx < len(row) and row[idx] not in (None, ''):
                parts.append(str(row[idx]).strip())
        if parts:
            out.setdefault(field, rule['separator'].join(parts))
    return out


# ── Employee matching ────────────────────────────────────────────
def _name_key(name: str) -> str:
    """Normalise a person name for matching: lowercase, single-spaced,
    no punctuation."""
    if not name:
        return ''
    n = re.sub(r'[^a-z\s]', '', str(name).lower())
    return re.sub(r'\s+', ' ', n).strip()


def _name_token_set(name: str) -> frozenset:
    """Order-insensitive token bag for name matching.
    "Anam Binte Abbas" -> {'anam','binte','abbas'} which then matches
    SYMPA's "Anam Abbas" via Jaccard similarity even though strict
    fuzzy string ratio is below threshold.
    """
    return frozenset(t for t in _name_key(name).split(' ') if len(t) >= 2)


def _token_jaccard(a: frozenset, b: frozenset) -> float:
    if not a or not b:
        return 0.0
    inter = len(a & b)
    union = len(a | b)
    return inter / union if union else 0.0


def _match_employee(ext_emp_no: Optional[str], ext_name: Optional[str],
                    by_code: Dict[str, PayrollEmployee],
                    by_name: Dict[str, PayrollEmployee],
                    by_token_set: List[Tuple[frozenset, PayrollEmployee]],
                    all_names: List[str]) -> Tuple[Optional[PayrollEmployee], str]:
    """Try exact code → exact name → token-set Jaccard → fuzzy string.
    Returns (employee_or_None, matched_by)."""
    if ext_emp_no:
        emp = by_code.get(str(ext_emp_no).strip())
        if emp:
            return emp, 'employee_no'
    if ext_name:
        key = _name_key(ext_name)
        emp = by_name.get(key)
        if emp:
            return emp, 'name'
        if config.COMPARISON_MATCH_FUZZY:
            ext_tokens = _name_token_set(ext_name)
            # 1) Token-set Jaccard — robust to extra middle names
            best_emp, best_score = None, 0.0
            for tokens, emp in by_token_set:
                score = _token_jaccard(ext_tokens, tokens)
                if score > best_score:
                    best_score, best_emp = score, emp
            if best_emp is not None and best_score >= 0.6:
                return best_emp, f'tokens:{best_score:.2f}'
            # 2) Fall back to whole-string fuzzy ratio
            if all_names:
                match = difflib.get_close_matches(
                    key, all_names, n=1,
                    cutoff=config.COMPARISON_MATCH_THRESHOLD,
                )
                if match:
                    ratio = difflib.SequenceMatcher(None, key, match[0]).ratio()
                    return by_name[match[0]], f'fuzzy:{ratio:.2f}'
    return None, ''


# ── Diff + recommendations ───────────────────────────────────────
def _classify_variance(field: str, our: Optional[Decimal],
                       ext: Optional[Decimal]) -> Optional[Dict]:
    """Compare one field. Return a variance dict or None if within tolerance.
    The variance dict shape is documented in the model."""
    if our is None and ext is None:
        return None
    if ext is None:
        return {
            'field': field, 'our': float(our or 0), 'external': None,
            'diff': None, 'pct': None, 'severity': 'info',
            'recommendation': 'External file has no value for this field',
        }
    if our is None:
        return {
            'field': field, 'our': None, 'external': float(ext),
            'diff': None, 'pct': None, 'severity': 'warning',
            'recommendation': f'Our payroll has no {field}; consider adopting external value',
        }

    diff = (ext - our).quantize(config.QUANTUM)
    base = our if our != 0 else (ext if ext != 0 else Decimal('1'))
    pct = (abs(diff) / abs(base) * Decimal('100')).quantize(Decimal('0.01'))

    meta = catalog.comparison_field_meta(field) or {}
    kind = meta.get('kind', 'currency')

    # Tolerance check
    if kind == 'hours':
        if abs(diff) <= config.COMPARISON_HOURS_TOL_ABS:
            return None
    else:
        if abs(diff) <= config.COMPARISON_TOL_ABS and pct <= config.COMPARISON_TOL_PCT:
            return None

    # Severity
    if pct >= config.COMPARISON_SEVERITY_CRIT_PCT:
        severity = 'critical'
    elif pct >= config.COMPARISON_SEVERITY_WARN_PCT:
        severity = 'warning'
    else:
        severity = 'info'

    direction = 'higher' if diff > 0 else 'lower'
    rec = _recommendation_for(field, kind, direction, diff, pct, severity)
    return {
        'field': field,
        'our': float(our),
        'external': float(ext),
        'diff': float(diff),
        'pct': float(pct),
        'severity': severity,
        'recommendation': rec,
    }


def _recommendation_for(field: str, kind: str, direction: str,
                        diff: Decimal, pct: Decimal, severity: str) -> str:
    label = catalog.comparison_field_meta(field).get('label', field)
    abs_diff = abs(diff)
    if kind == 'hours':
        if direction == 'higher':
            return (f'External shows {abs_diff:.2f}h more — investigate missed punches '
                    f'or run "Force Refresh Hours" from the live timesheet.')
        return (f'External shows {abs_diff:.2f}h less — verify HR overrides and '
                f'check for unsynced biometric punches.')
    # Currency
    suffix = f' ({pct:.1f}%)'
    if severity == 'critical':
        prefix = 'CRITICAL: '
    elif severity == 'warning':
        prefix = 'Review: '
    else:
        prefix = ''
    if direction == 'higher':
        return (f'{prefix}External {label} is {abs_diff:.2f} higher{suffix}. '
                f'Consider raising a salary adjustment or updating the master.')
    return (f'{prefix}External {label} is {abs_diff:.2f} lower{suffix}. '
            f'Verify deductions, recent contract changes, or data entry.')


def _our_values_from_slip(slip: Optional[Payslip],
                          emp: Optional[PayrollEmployee]) -> Dict:
    """Pull the canonical comparison values from our DB. Falls back to the
    employee master if no payslip exists for the run."""
    out: Dict = {}
    if slip is not None:
        for f in ('hours', 'basic', 'housing', 'transport', 'home_leave',
                  'gross_earnings', 'total_deductions', 'net_payable'):
            out[f] = float(getattr(slip, f) or 0)
        # other_earnings = gross - (basic + housing + transport + home_leave)
        fixed = sum([slip.basic or 0, slip.housing or 0,
                     slip.transport or 0, slip.home_leave or 0])
        out['other_earnings'] = float((slip.gross_earnings or 0) - fixed)
    elif emp is not None:
        for f in ('hours', 'basic', 'housing', 'transport', 'home_leave'):
            out[f] = float(getattr(emp, f) or 0)
    return out


def _ext_values_canonical(raw: Dict) -> Dict:
    """Convert raw external row values (mixed types) into Decimal-friendly
    floats for storage, keeping only the numeric fields we know to diff."""
    out: Dict = {}
    for field in catalog.comparison_field_codes():
        val = _to_dec(raw.get(field))
        if val is not None:
            out[field] = float(val)
    return out


# ── Main entry ───────────────────────────────────────────────────
@transaction.atomic
def run_comparison(run: PayrollRun, file_obj, *,
                   source_label: str, source_profile: str = 'auto',
                   uploaded_by=None, source_filename: str = '') -> PayrollComparison:
    profile = catalog.comparison_profile(source_profile)
    wb = _open_workbook(file_obj)
    ws = wb.worksheets[0]

    header_row_idx = _locate_header_row(ws, profile)

    # Read the header
    headers: List = []
    for r_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if r_idx == header_row_idx:
            headers = list(row)
            break

    col_index, col_header, compose = _resolve_columns(headers, profile)

    # Build employee lookups once
    employees = list(PayrollEmployee.objects.all())
    by_code: Dict[str, PayrollEmployee] = {e.employee_no: e for e in employees}
    by_name: Dict[str, PayrollEmployee] = {}
    by_token_set: List[Tuple[frozenset, PayrollEmployee]] = []
    for e in employees:
        key = _name_key(e.full_name)
        if key and key not in by_name:
            by_name[key] = e
        tokens = _name_token_set(e.full_name)
        if tokens:
            by_token_set.append((tokens, e))
    all_name_keys = list(by_name.keys())

    # Payslips for this run keyed by employee_id
    slips_by_emp = {s.employee_id: s for s in run.payslips.all()}

    comparison = PayrollComparison.objects.create(
        run=run,
        source_label=source_label[:64] or source_profile,
        source_profile=source_profile,
        source_filename=source_filename[:255],
        column_mapping={**col_header, **{
            f'__compose__{f}': [str(headers[i]) for i in rule['parts']
                                if i < len(headers)]
            for f, rule in compose.items()
        }},
        uploaded_by=uploaded_by,
        summary={},
    )

    matched_employee_ids = set()
    rows_to_create: List[PayrollComparisonRow] = []
    summary = {
        'matched': 0, 'variance': 0, 'external_only': 0, 'payroll_only': 0,
        'by_field': {f: {'variances': 0, 'critical': 0, 'warning': 0}
                     for f in catalog.comparison_field_codes()},
        'fields_detected': list(col_header.keys()),
    }

    row_cap = config.COMPARISON_MAX_ROWS
    processed = 0
    # Iterate data rows
    for r_idx, raw_row in enumerate(ws.iter_rows(values_only=True), start=1):
        if r_idx <= header_row_idx:
            continue
        if processed >= row_cap:
            break
        canonical = _row_to_canonical(raw_row, col_index, compose)
        ext_emp_no = canonical.get('employee_no')
        ext_name = canonical.get('full_name')
        # Skip blank rows
        if not ext_emp_no and not ext_name:
            continue
        ext_emp_no_str = str(ext_emp_no).strip() if ext_emp_no else ''
        ext_name_str = str(ext_name).strip() if ext_name else ''
        if not ext_emp_no_str and not ext_name_str:
            continue
        processed += 1

        emp, matched_by = _match_employee(
            ext_emp_no_str, ext_name_str, by_code, by_name, by_token_set, all_name_keys,
        )
        ext_values = _ext_values_canonical(canonical)

        if emp is None:
            summary['external_only'] += 1
            rows_to_create.append(PayrollComparisonRow(
                comparison=comparison,
                external_employee_no=ext_emp_no_str,
                external_name=ext_name_str,
                external_values=ext_values,
                our_values={},
                variances=[{
                    'field': '__match__', 'severity': 'warning',
                    'recommendation': ('No match in payroll. Add this employee '
                                       'to the master roster or verify spelling.'),
                }],
                status=catalog.ComparisonStatus.EXTERNAL_ONLY,
            ))
            continue

        matched_employee_ids.add(emp.id)
        slip = slips_by_emp.get(emp.id)
        our_values = _our_values_from_slip(slip, emp)

        # Only diff fields the external file actually carries. Fields
        # we have internally but the external roster doesn't are noise,
        # not variances.
        diffable_fields = [f for f in catalog.comparison_field_codes()
                           if f in ext_values]
        variances = []
        for f in diffable_fields:
            v = _classify_variance(
                f,
                Decimal(str(our_values.get(f))) if f in our_values else None,
                Decimal(str(ext_values.get(f))) if f in ext_values else None,
            )
            if v:
                variances.append(v)
                summary['by_field'][f]['variances'] += 1
                if v['severity'] == 'critical':
                    summary['by_field'][f]['critical'] += 1
                elif v['severity'] == 'warning':
                    summary['by_field'][f]['warning'] += 1

        # Only promote to VARIANCE status if at least one warning/critical
        # variance exists. Pure 'info' rows (e.g. "ext has no value") stay
        # as MATCH so the dashboard isn't drowned in noise.
        actionable = [v for v in variances if v.get('severity') in ('warning', 'critical')]
        if actionable:
            summary['variance'] += 1
            status = catalog.ComparisonStatus.VARIANCE
        else:
            summary['matched'] += 1
            status = catalog.ComparisonStatus.MATCH

        rows_to_create.append(PayrollComparisonRow(
            comparison=comparison,
            payroll_employee=emp,
            external_employee_no=ext_emp_no_str,
            external_name=ext_name_str,
            matched_by=matched_by,
            our_values=our_values,
            external_values=ext_values,
            variances=variances,
            status=status,
        ))

    # Payroll employees not seen in the external file
    for emp in employees:
        if emp.id in matched_employee_ids:
            continue
        slip = slips_by_emp.get(emp.id)
        if slip is None:
            continue
        summary['payroll_only'] += 1
        rows_to_create.append(PayrollComparisonRow(
            comparison=comparison,
            payroll_employee=emp,
            external_employee_no='',
            external_name='',
            our_values=_our_values_from_slip(slip, emp),
            external_values={},
            variances=[{
                'field': '__match__', 'severity': 'warning',
                'recommendation': ('Employee paid this cycle but absent from '
                                   'external file — verify external roster.'),
            }],
            status=catalog.ComparisonStatus.PAYROLL_ONLY,
        ))

    PayrollComparisonRow.objects.bulk_create(rows_to_create, batch_size=500)

    summary['total_rows'] = len(rows_to_create)
    summary['header_row_used'] = header_row_idx
    comparison.summary = summary
    comparison.save(update_fields=['summary'])
    wb.close()
    return comparison
