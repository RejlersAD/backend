"""Excel import for the Payroll Engine.

Two modes:

* ``import_master_roster(file)`` — reads sheet #1 (master roster) and
  upserts PayrollEmployee rows. The fixed-earning columns (Basic,
  Housing, Transport, Home Leave) and any Other Pay / Salary Deduction
  free-form columns are loaded as PayrollAdjustment rows when
  ``include_adjustments=True``.
* ``import_full_payroll(file, year, month)`` — does the above PLUS
  immediately calls run_generator to produce a Draft PayrollRun for the
  given period.
* ``import_adjustments(file, year, month)`` — reads a roster-shaped
  sheet but only persists the Other Pay / Deduction columns as
  PayrollAdjustment rows for the target period.

Column positions are driven entirely by
``apps.payroll_engine.catalog.EXCEL_MASTER_COLUMN_MAP``.
"""
from __future__ import annotations
import datetime as _dt
from dataclasses import dataclass, field
from decimal import Decimal
from typing import Iterable

from django.db import transaction

from .. import catalog
from ..models import PayrollAdjustment, PayrollEmployee
from .calculator import to_decimal


@dataclass
class ImportSummary:
    employees_created: int = 0
    employees_updated: int = 0
    adjustments_created: int = 0
    rows_skipped: int = 0
    warnings: list = field(default_factory=list)

    def as_dict(self) -> dict:
        return {
            'employees_created':    self.employees_created,
            'employees_updated':    self.employees_updated,
            'adjustments_created':  self.adjustments_created,
            'rows_skipped':         self.rows_skipped,
            'warnings':             list(self.warnings),
        }


def _open_workbook(file_or_path):
    """Lazy openpyxl import (heavy dep)."""
    import openpyxl
    if hasattr(file_or_path, 'read'):
        return openpyxl.load_workbook(file_or_path, data_only=True)
    return openpyxl.load_workbook(str(file_or_path), data_only=True)


def _row_dict(ws, row_idx: int) -> dict:
    """Pull a row out of the master sheet keyed by canonical field name."""
    out = {}
    for col_idx, field_name in catalog.EXCEL_MASTER_COLUMN_MAP.items():
        out[field_name] = ws.cell(row_idx, col_idx).value
    return out


def _is_data_row(row: dict) -> bool:
    """A row is a real employee row if it has both an employee_no AND a name.
    Skips header / blank / TOTAL footer rows."""
    name = (row.get('full_name') or '')
    if not name or str(name).strip().upper() == catalog.EXCEL_MASTER_TOTAL_LABEL:
        return False
    emp_no = row.get('employee_no')
    if emp_no in (None, '', 0):
        return False
    return True


def _normalise_date(value):
    if value in (None, '', 0):
        return None
    if isinstance(value, _dt.datetime):
        return value.date()
    if isinstance(value, _dt.date):
        return value
    return None


def _normalise_str(value, max_len: int = 255) -> str:
    if value in (None, ''):
        return ''
    return str(value).strip()[:max_len]


def _upsert_employee(row: dict, summary: ImportSummary) -> PayrollEmployee:
    emp_no = str(row['employee_no']).strip()
    defaults = {
        'full_name':         _normalise_str(row.get('full_name')),
        'mol_no':            _normalise_str(row.get('mol_no'), 32),
        'iban':              _normalise_str(row.get('iban'), 64),
        'bank_name':         _normalise_str(row.get('bank_name'), 128),
        'routing_code':      _normalise_str(row.get('routing_code'), 64),
        'department':        _normalise_str(row.get('department'), 128),
        'discipline':        _normalise_str(row.get('discipline'), 128),
        'designation':       _normalise_str(row.get('designation'), 128),
        'grade':             _normalise_str(row.get('grade'), 128),
        'nationality_group': _normalise_str(row.get('nationality_group'), 64),
        'joining_date':      _normalise_date(row.get('joining_date')),
        'basic':             to_decimal(row.get('basic')),
        'housing':           to_decimal(row.get('housing')),
        'transport':         to_decimal(row.get('transport')),
        'home_leave':        to_decimal(row.get('home_leave')),
        'default_payment_mode': catalog.normalise_payment_mode(row.get('payment_mode')),
        'is_active':         True,
    }
    employee, created = PayrollEmployee.objects.update_or_create(
        employee_no=emp_no, defaults=defaults,
    )
    if created:
        summary.employees_created += 1
    else:
        summary.employees_updated += 1
    return employee


def _materialise_adjustments(
    row: dict,
    employee: PayrollEmployee,
    target_year: int,
    target_month: int,
    summary: ImportSummary,
) -> None:
    """Read the Other Pay / Deduction columns of a row and create
    pending PayrollAdjustment entries for the target period.
    """
    # Earning side (Other Allowance + Other Pay)
    for amount_field, detail_field, default_code, default_label in [
        ('other_allowance', None,                'other_earning', 'Other Allowance'),
        ('other_pay_amount', 'other_pay_details', 'leave_payout',  'Other Earning'),
    ]:
        amt = to_decimal(row.get(amount_field))
        if amt <= 0:
            continue
        description = ''
        if detail_field:
            description = _normalise_str(row.get(detail_field), 255)
        label = description.split(' - ')[0].strip() if description else default_label
        PayrollAdjustment.objects.create(
            employee=employee,
            target_year=target_year,
            target_month=target_month,
            kind=catalog.LineItemKind.EARNING,
            component_code=default_code,
            label=label or default_label,
            description=description,
            amount=amt,
            status=catalog.AdjustmentStatus.PENDING,
        )
        summary.adjustments_created += 1

    # Deduction side
    amt = to_decimal(row.get('deduction_amount'))
    if amt > 0:
        description = _normalise_str(row.get('deduction_details'), 255)
        label = description.split(' - ')[0].strip() if description else 'Deduction'
        PayrollAdjustment.objects.create(
            employee=employee,
            target_year=target_year,
            target_month=target_month,
            kind=catalog.LineItemKind.DEDUCTION,
            component_code='other_deduction',
            label=label or 'Deduction',
            description=description,
            amount=amt,
            status=catalog.AdjustmentStatus.PENDING,
        )
        summary.adjustments_created += 1


def _iter_data_rows(ws) -> Iterable[tuple[int, dict]]:
    start = catalog.EXCEL_MASTER_DATA_START_ROW
    for r in range(start, ws.max_row + 1):
        row = _row_dict(ws, r)
        if not _is_data_row(row):
            continue
        yield r, row


# ── Public entry points ─────────────────────────────────────────────
@transaction.atomic
def import_master_roster(file_or_path) -> ImportSummary:
    """Upsert PayrollEmployee rows from the first sheet of the workbook.
    Does NOT create adjustments. Use import_full_payroll for that.
    """
    summary = ImportSummary()
    wb = _open_workbook(file_or_path)
    ws = wb[wb.sheetnames[0]]
    for _, row in _iter_data_rows(ws):
        _upsert_employee(row, summary)
    return summary


@transaction.atomic
def import_full_payroll(file_or_path, year: int, month: int) -> ImportSummary:
    """Upsert employees AND queue per-employee adjustments for (year, month)
    based on the Other Pay / Deduction columns in the master sheet.
    Call run_generator.generate_monthly_run(year, month) afterwards to
    create the Draft PayrollRun.
    """
    summary = ImportSummary()
    wb = _open_workbook(file_or_path)
    ws = wb[wb.sheetnames[0]]
    # Wipe any prior pending adjustments for the period (idempotent re-imports)
    PayrollAdjustment.objects.filter(
        target_year=year, target_month=month,
        status=catalog.AdjustmentStatus.PENDING,
    ).delete()
    for _, row in _iter_data_rows(ws):
        employee = _upsert_employee(row, summary)
        _materialise_adjustments(row, employee, year, month, summary)
    return summary


@transaction.atomic
def import_adjustments(file_or_path, year: int, month: int) -> ImportSummary:
    """Only the Other Pay / Deduction columns get persisted. Existing
    pending adjustments for the period are wiped first (so the upload is
    authoritative). Employees that don't exist yet get auto-created from
    the row (so HR doesn't have to seed master separately).
    """
    summary = ImportSummary()
    wb = _open_workbook(file_or_path)
    ws = wb[wb.sheetnames[0]]
    PayrollAdjustment.objects.filter(
        target_year=year, target_month=month,
        status=catalog.AdjustmentStatus.PENDING,
    ).delete()
    for _, row in _iter_data_rows(ws):
        emp_no = str(row['employee_no']).strip()
        employee = PayrollEmployee.objects.filter(employee_no=emp_no).first()
        if not employee:
            employee = _upsert_employee(row, summary)
        _materialise_adjustments(row, employee, year, month, summary)
    return summary
