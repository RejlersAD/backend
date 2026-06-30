"""Excel export — generate XLSX matching the source template layout.

Two artefacts:
  * ``export_master_xlsx(run)`` — single sheet matching the
    APRIL 2026 master roster format (24 columns + TOTAL row).
  * ``export_payslip_pack(run)`` — master sheet PLUS one sheet per
    employee, in the per-payslip A4 format.
"""
from __future__ import annotations
import datetime as _dt
from io import BytesIO
from typing import Optional

from .. import catalog
from ..models import Payslip, PayrollRun
from .calculator import to_decimal

_MONTH_NAMES = [
    '', 'JANUARY', 'FEBRUARY', 'MARCH', 'APRIL', 'MAY', 'JUNE',
    'JULY', 'AUGUST', 'SEPTEMBER', 'OCTOBER', 'NOVEMBER', 'DECEMBER',
]


def _period_label(run: PayrollRun) -> str:
    return f'{_MONTH_NAMES[run.month]} {run.year}'


def _safe_sheet_title(name: str, used: set) -> str:
    base = ''.join(c for c in name if c not in r'[]:*?/\\')[:28] or 'EMP'
    base = f'!{base}'  # match source format
    title = base[:31]
    counter = 1
    while title in used:
        suffix = f'_{counter}'
        title = (base[:31 - len(suffix)] + suffix)
        counter += 1
    used.add(title)
    return title


def _write_master_sheet(ws, run: PayrollRun) -> None:
    # Title row
    ws.cell(1, 1, value='Payroll for the month of')
    ws.cell(1, 7, value=_period_label(run))

    # Header row (row 4 — matches source)
    header_labels = {
        1:  'No.', 2: 'Employee', 3: 'MOL #', 4: 'Routing Code', 5: 'IBAN #',
        6:  'Bank', 7: 'Employee Number', 8: 'Joining date', 9: 'Dep.',
        10: 'Discipline', 11: 'Position', 12: 'Category', 13: 'Demography',
        14: 'Basic', 15: 'Housing', 16: 'Transpo', 17: 'Home Leave',
        18: 'Other Allowance', 19: 'Other\nPay', 20: 'Other Pay Details',
        21: 'Salary Deduction', 22: 'Salary Deduction Details',
        23: f'FINAL Remuneration {_period_label(run)}',
        24: 'Payment Method',
        25: catalog.EXCEL_MASTER_HOURS_HEADER,
    }
    for col, label in header_labels.items():
        ws.cell(catalog.EXCEL_MASTER_HEADER_ROW, col, value=label)

    # Data rows
    row_idx = catalog.EXCEL_MASTER_DATA_START_ROW
    grand_total = 0
    for i, slip in enumerate(
        run.payslips.select_related('employee').prefetch_related('line_items').order_by('snapshot_full_name'),
        start=1,
    ):
        emp = slip.employee
        # Split line items into earning vs deduction summaries
        other_earnings = []
        deductions = []
        for li in slip.line_items.all():
            if li.kind == catalog.LineItemKind.EARNING:
                other_earnings.append(li)
            else:
                deductions.append(li)

        other_earn_amt = sum(to_decimal(li.amount) for li in other_earnings) or None
        other_earn_desc = '; '.join(
            (li.description or li.label) for li in other_earnings if li.description or li.label
        )
        deduct_amt = sum(to_decimal(li.amount) for li in deductions) or None
        deduct_desc = '; '.join(
            (li.description or li.label) for li in deductions if li.description or li.label
        )

        ws.cell(row_idx, 1, value=i)
        ws.cell(row_idx, 2, value=emp.full_name)
        ws.cell(row_idx, 3, value=emp.mol_no or None)
        ws.cell(row_idx, 4, value=emp.routing_code or None)
        ws.cell(row_idx, 5, value=emp.iban or None)
        ws.cell(row_idx, 6, value=emp.bank_name or None)
        ws.cell(row_idx, 7, value=emp.employee_no)
        ws.cell(row_idx, 8, value=emp.joining_date)
        ws.cell(row_idx, 9, value=emp.department or None)
        ws.cell(row_idx, 10, value=emp.discipline or None)
        ws.cell(row_idx, 11, value=emp.designation or None)
        ws.cell(row_idx, 12, value=emp.grade or None)
        ws.cell(row_idx, 13, value=emp.nationality_group or None)
        ws.cell(row_idx, 14, value=float(slip.basic) if slip.basic else None)
        ws.cell(row_idx, 15, value=float(slip.housing) if slip.housing else None)
        ws.cell(row_idx, 16, value=float(slip.transport) if slip.transport else None)
        ws.cell(row_idx, 17, value=float(slip.home_leave) if slip.home_leave else None)
        ws.cell(row_idx, 18, value=None)  # other_allowance — unused in current data
        ws.cell(row_idx, 19, value=float(other_earn_amt) if other_earn_amt else None)
        ws.cell(row_idx, 20, value=other_earn_desc or None)
        ws.cell(row_idx, 21, value=float(deduct_amt) if deduct_amt else None)
        ws.cell(row_idx, 22, value=deduct_desc or None)
        ws.cell(row_idx, 23, value=float(slip.net_payable))
        ws.cell(row_idx, 24, value=slip.payment_mode)
        ws.cell(row_idx, 25, value=float(slip.hours) if slip.hours is not None else None)

        grand_total += float(slip.net_payable)
        row_idx += 1

    # TOTAL footer
    ws.cell(row_idx + 1, 2, value=catalog.EXCEL_MASTER_TOTAL_LABEL)
    ws.cell(row_idx + 1, 23, value=grand_total)


def _write_payslip_sheet(ws, slip: Payslip, period_label: str) -> None:
    L = catalog.EXCEL_PAYSLIP_LAYOUT
    ws.cell(*L['title_row'], value='Payslip for the month of')
    ws.cell(*L['period_label'], value=period_label)

    ws.cell(*L['name_label'], value='Name')
    ws.cell(*L['name_value'], value=slip.snapshot_full_name)
    ws.cell(*L['joining_date_label'], value='Joining Date')
    ws.cell(*L['joining_date_value'], value=slip.snapshot_joining_date)

    ws.cell(*L['employee_no_label'], value='Employee No.')
    ws.cell(*L['employee_no_value'], value=slip.employee.employee_no)

    ws.cell(*L['title_label'], value='Title')
    ws.cell(*L['title_value'], value=slip.snapshot_designation)

    # Hours / Month (biometric “Total” hours, with HR overrides overlaid)
    if 'hours_label' in L:
        ws.cell(*L['hours_label'], value='Hours / Month')
        ws.cell(*L['hours_value'], value=float(slip.hours) if slip.hours is not None else None)

    ws.cell(*L['salary_header'], value='SALARY')
    ws.cell(*L['deductions_header'], value='DEDUCTIONS')

    # Column headers
    hdr_row = L['col_headers_row']
    ws.cell(hdr_row, 1, value='Description')
    ws.cell(hdr_row, 2, value='Amount')
    ws.cell(hdr_row, 3, value='Description')
    ws.cell(hdr_row, 4, value='Details')
    ws.cell(hdr_row, 5, value='Amount')

    # Fixed earnings
    ws.cell(*L['basic_label'],      value='Basic salary');             ws.cell(*L['basic_value'],      value=float(slip.basic))
    ws.cell(*L['housing_label'],    value='Housing Allowance');        ws.cell(*L['housing_value'],    value=float(slip.housing))
    ws.cell(*L['transport_label'],  value='Transportation Allowance'); ws.cell(*L['transport_value'],  value=float(slip.transport))
    ws.cell(*L['home_leave_label'], value='Home Leave Allowance');     ws.cell(*L['home_leave_value'], value=float(slip.home_leave))

    # Other earnings (sum + first description as detail row)
    earnings = [li for li in slip.line_items.all() if li.kind == catalog.LineItemKind.EARNING]
    other_total = sum(to_decimal(li.amount) for li in earnings)
    ws.cell(*L['others_earning_label'], value='Others:')
    ws.cell(*L['others_earning_value'], value=float(other_total) if other_total else None)
    detail_row = L['others_earning_detail_row']
    if earnings:
        desc_parts = [li.description or li.label for li in earnings]
        ws.cell(detail_row, 1, value='; '.join(desc_parts))

    # Deductions in right-hand columns
    deductions = [li for li in slip.line_items.all() if li.kind == catalog.LineItemKind.DEDUCTION]
    d_start = L['deduction_start_row']
    standard_labels = ['Absent', 'Housing Allowance Advance', 'Salary Advance', 'Sick Leave', 'Telephone']
    for i, label in enumerate(standard_labels):
        ws.cell(d_start + i, L['deduction_label_col'], value=label)
    # All free-form deductions go in the "Others" row (row 20 by default)
    others_row = d_start + len(standard_labels)
    ws.cell(others_row, L['deduction_label_col'], value='Others')
    if deductions:
        desc_parts = [li.description or li.label for li in deductions]
        amount = sum(to_decimal(li.amount) for li in deductions)
        ws.cell(others_row, L['deduction_detail_col'], value='; '.join(desc_parts))
        ws.cell(others_row, L['deduction_value_col'], value=float(amount))

    # Totals
    ws.cell(*L['total_earnings_label'],   value='Total')
    ws.cell(*L['total_earnings_value'],   value=float(slip.gross_earnings))
    ws.cell(*L['total_deductions_label'], value='Total')
    ws.cell(*L['total_deductions_value'], value=float(slip.total_deductions))

    # Net payable
    ws.cell(*L['net_label'], value='NET PAYABLE Amount')
    ws.cell(*L['net_value'], value=float(slip.net_payable))

    # Payment mode
    ws.cell(*L['payment_mode_label'], value='Salary payment through')
    ws.cell(*L['payment_mode_value'], value=slip.payment_mode)


def export_master_xlsx(run: PayrollRun) -> bytes:
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = _period_label(run)[:31]
    _write_master_sheet(ws, run)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_payslip_pack(run: PayrollRun) -> bytes:
    """Master sheet + one sheet per employee."""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = _period_label(run)[:31]
    _write_master_sheet(ws, run)
    period_label = _period_label(run)
    used = {ws.title}
    for slip in run.payslips.select_related('employee').prefetch_related('line_items').order_by('snapshot_full_name'):
        sheet_title = _safe_sheet_title(slip.snapshot_full_name or slip.employee.employee_no, used)
        sheet = wb.create_sheet(title=sheet_title)
        _write_payslip_sheet(sheet, slip, period_label)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_single_payslip(slip: Payslip) -> bytes:
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = (slip.snapshot_full_name or 'Payslip')[:31]
    _write_payslip_sheet(ws, slip, _period_label(slip.run))
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
