"""
External File Import Service — ValueFrame / Sympa / Generic XLSX
================================================================
Parses an external HR file and applies the mapped fields directly to
Draft payslips in the given PayrollRun.

Field mapping is soft-coded in catalog.EXTERNAL_IMPORT_FIELD_MAP.
Column header aliases come from catalog.COMPARISON_PROFILES.

Files are uploaded to S3 (payroll/exports/run-uploads/) for audit trail.
Falls back gracefully when S3 is unavailable (local dev).
"""
from __future__ import annotations

import io
import logging
import re
import uuid
from decimal import Decimal, InvalidOperation
from typing import Optional

logger = logging.getLogger(__name__)

# Soft-coded S3 prefix for uploaded external files.
_UPLOAD_S3_PREFIX = 'payroll/exports/run-uploads'


def _norm(text: str) -> str:
    """Normalise a header cell for alias matching: lowercase, collapse spaces."""
    return re.sub(r'\s+', ' ', str(text or '').strip().lower())


def _scan_header_row(ws, profile: dict) -> tuple[int, dict]:
    """
    Scan rows (up to profile['header_scan_max']) looking for a row whose
    cells match the profile's field aliases.

    Returns (header_row_index_1based, {col_index: field_code}).
    Returns (-1, {}) when no header row is found.
    """
    from .. import catalog as _cat

    field_aliases = profile.get('field_aliases', {})
    # Build reverse lookup: alias_normalised → field_code
    alias_lookup: dict[str, str] = {}
    for field, aliases in field_aliases.items():
        for a in aliases:
            alias_lookup[_norm(a)] = field
        alias_lookup[_norm(field)] = field  # also match the canonical name

    # Also accept global COMPARISON_FIELD_ALIASES
    for field, aliases in _cat.COMPARISON_FIELD_ALIASES.items():
        for a in aliases:
            if _norm(a) not in alias_lookup:
                alias_lookup[_norm(a)] = field

    max_scan = profile.get('header_scan_max', 20)
    best_row = -1
    best_map: dict[int, str] = {}

    for row_idx in range(1, max_scan + 1):
        row_cells = [ws.cell(row_idx, col).value for col in range(1, ws.max_column + 1)]
        col_map: dict[int, str] = {}
        for col, cell in enumerate(row_cells, 1):
            normed = _norm(cell)
            if normed in alias_lookup:
                col_map[col] = alias_lookup[normed]
        if len(col_map) > len(best_map):
            best_map = col_map
            best_row = row_idx

    return best_row, best_map


def _to_decimal(value) -> Optional[Decimal]:
    """Coerce a cell value to Decimal; return None on failure."""
    if value is None or str(value).strip() == '':
        return None
    try:
        return Decimal(str(value).replace(',', '').strip())
    except InvalidOperation:
        return None


def _upload_to_s3(file_bytes: bytes, run_cycle: str, filename: str) -> str:
    """
    Upload file to S3 and return the full S3 key.
    Returns empty string if S3 is unavailable (local dev fallback).
    """
    try:
        from apps.payroll.storage import PayrollExportStorage, S3_AVAILABLE
        if not S3_AVAILABLE:
            return ''
        safe_name = re.sub(r'[^\w.\-]', '_', filename)
        key_name  = f'run-uploads/{run_cycle}/{uuid.uuid4().hex[:8]}_{safe_name}'
        storage   = PayrollExportStorage()
        saved     = storage.save(key_name, io.BytesIO(file_bytes))
        return f'{storage.location}/{saved}'
    except Exception:
        logger.warning('external_import: S3 upload failed — storing without S3 reference')
        return ''


def apply_external_file(
    run,
    file_bytes: bytes,
    original_filename: str,
    file_type: str,
    user=None,
) -> dict:
    """
    Parse *file_bytes* as an XLSX, map columns using the profile for
    *file_type*, and apply the mapped values to Draft Payslips in *run*.

    Parameters
    ----------
    run            : PayrollRun instance (must be in DRAFT status)
    file_bytes     : Raw content of the uploaded file
    original_filename : Original name (for storage + display)
    file_type      : 'valueframe' | 'sympa' | 'generic'
    user           : Django User who triggered the upload

    Returns
    -------
    dict with keys:
        rows_matched    : int
        rows_updated    : int
        unmatched       : list[str]  — identifiers not found in the run
        updated_fields  : list[str]  — payslip field names that were written
        s3_key          : str
        upload_id       : int
    """
    import openpyxl
    from django.db import transaction

    from .. import catalog as _cat
    from ..models import PayrollRunUpload, Payslip
    from ..services.calculator import recompute_payslip_totals, recompute_run_totals

    field_map: dict = _cat.EXTERNAL_IMPORT_FIELD_MAP.get(file_type, {})
    profile: dict   = _cat.comparison_profile(file_type)

    # ── Parse workbook ────────────────────────────────────────────────────────
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        ws = wb.active
    except Exception as exc:
        raise ValueError(f'Cannot open file: {exc}') from exc

    header_row, col_map = _scan_header_row(ws, profile)
    if not col_map:
        raise ValueError(
            f'Could not detect required column headers in {original_filename}. '
            f'Expected: {list(profile.get("field_aliases", {}).keys())}'
        )

    # ── Build payslip lookup by employee_no ───────────────────────────────────
    payslip_lookup: dict[str, Payslip] = {}
    for slip in (Payslip.objects.filter(run=run)
                 .select_related('employee')
                 .only('id', 'employee__employee_no',
                       'hours', 'days', 'annual_leave_days', 'unpaid_leave_days',
                       'basic', 'housing', 'transport', 'home_leave')):
        emp_no = (slip.employee.employee_no or '').strip().lower()
        if emp_no:
            payslip_lookup[emp_no] = slip

    # ── Compose config (Sympa: first_name + surname → full_name) ──────────────
    compose_cfg = profile.get('compose', {})
    compose_full_name = compose_cfg.get('full_name', {})

    # Also build name → payslip lookup as fallback
    name_to_payslip: dict[str, Payslip] = {
        (s.employee.full_name or '').strip().lower(): s
        for s in payslip_lookup.values()
    }

    # ── Iterate data rows ─────────────────────────────────────────────────────
    rows_matched  = 0
    rows_updated  = 0
    unmatched: list[str] = []
    updated_fields_set: set[str] = set()
    to_save: list[Payslip] = []

    for row_idx in range(header_row + 1, ws.max_row + 1):
        row = {col: ws.cell(row_idx, col).value for col in col_map}
        if all(v is None for v in row.values()):
            continue  # blank row

        # Resolve employee_no
        emp_no_val = None
        for col, field in col_map.items():
            if field == 'employee_no':
                emp_no_val = str(row[col] or '').strip().lower()
                break

        # Compose full_name if available (Sympa)
        full_name_val = None
        if compose_full_name.get('parts'):
            parts: list[str] = []
            for part_label in compose_full_name['parts']:
                for col, field in col_map.items():
                    if field == f'__compose_{part_label}' or \
                       _norm(ws.cell(header_row, col).value) == _norm(part_label):
                        v = str(row.get(col) or '').strip()
                        if v:
                            parts.append(v)
                        break
            if parts:
                full_name_val = compose_full_name.get('separator', ' ').join(parts).lower()

        # Find payslip
        slip: Optional[Payslip] = None
        if emp_no_val and emp_no_val in payslip_lookup:
            slip = payslip_lookup[emp_no_val]
        elif full_name_val and full_name_val in name_to_payslip:
            slip = name_to_payslip[full_name_val]

        identifier = emp_no_val or full_name_val or f'row {row_idx}'
        if slip is None:
            if identifier and identifier not in unmatched:
                unmatched.append(identifier)
            continue

        rows_matched += 1
        changed = False

        # Apply mapped fields
        for col, comp_field in col_map.items():
            payslip_field = field_map.get(comp_field)
            if not payslip_field:
                continue
            raw_value = row.get(col)
            dec_value = _to_decimal(raw_value)
            if dec_value is None:
                continue
            if getattr(slip, payslip_field, None) != dec_value:
                setattr(slip, payslip_field, dec_value)
                updated_fields_set.add(payslip_field)
                changed = True

        if changed:
            rows_updated += 1
            to_save.append(slip)

    # ── Persist changes ───────────────────────────────────────────────────────
    updated_fields = sorted(updated_fields_set)
    with transaction.atomic():
        for slip in to_save:
            # Recompute days if hours changed
            if 'hours' in updated_fields_set:
                from ..config import hours_to_days
                slip.days = hours_to_days(slip.hours)
            recompute_payslip_totals(slip)
            slip.save()
        if to_save:
            recompute_run_totals(run)
            run.save()

    # ── Upload to S3 ──────────────────────────────────────────────────────────
    s3_key = _upload_to_s3(file_bytes, run.cycle_code, original_filename)

    # ── Audit record ─────────────────────────────────────────────────────────
    upload_obj = PayrollRunUpload.objects.create(
        run=run,
        file_type=file_type,
        original_filename=original_filename,
        s3_key=s3_key,
        uploaded_by=user,
        rows_matched=rows_matched,
        rows_updated=rows_updated,
        unmatched=unmatched,
        updated_fields=updated_fields,
        status='applied',
    )

    return {
        'upload_id':     upload_obj.id,
        'rows_matched':  rows_matched,
        'rows_updated':  rows_updated,
        'unmatched':     unmatched,
        'updated_fields': updated_fields,
        's3_key':        s3_key,
    }
