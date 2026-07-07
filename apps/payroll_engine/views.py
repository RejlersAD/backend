"""DRF views for the Payroll Engine."""
from __future__ import annotations
from datetime import datetime
from decimal import Decimal

from django.db.models import Sum, Q
from django.http import HttpResponse
from django.utils import timezone
from rest_framework import status as http_status, viewsets, mixins, filters
from rest_framework.decorators import action, api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

from . import catalog
from .config import (
    CURRENCY, EXCEL_MAX_UPLOAD_MB,
    EMPLOYEE_WRITE_ROLE_CODES, ADJUSTMENT_WRITE_ROLE_CODES,
    RUN_FORCE_OVERRIDE_ROLE_CODES,
)
from .models import (
    PayrollAdjustment, PayrollEmployee, PayrollRun, Payslip, PayslipLineItem,
    PayrollWorkflowLog, PayrollComparison, PayrollComparisonRow,
)
from .serializers import (
    CatalogSerializer, PayrollAdjustmentSerializer, PayrollEmployeeSerializer,
    PayrollRunSerializer, PayslipLineItemSerializer, PayslipSerializer,
    PayrollWorkflowLogSerializer,
    PayrollComparisonSerializer, PayrollComparisonDetailSerializer,
    PayrollComparisonRowSerializer,
)
from .services import excel_export, excel_import, run_generator, workflow
from .services import bulk_deduction, comparison as comparison_service
from .services.calculator import recompute_payslip_totals, recompute_run_totals


# ── Permissions ────────────────────────────────────────────────────
SAFE_METHODS = frozenset({'GET', 'HEAD', 'OPTIONS'})


def _user_role_codes(user) -> set:
    """Return the lowercased role codes assigned to ``user`` (best-effort)."""
    if not user or not user.is_authenticated:
        return set()
    try:
        codes = user.user_roles.values_list('role__code', flat=True)
    except Exception:
        try:
            codes = user.roles.values_list('code', flat=True)
        except Exception:
            return set()
    return {(c or '').strip().lower() for c in codes if c}


def _user_has_any_role(user, allowed_codes) -> bool:
    """True if the user is super/staff or has any matching role code."""
    if not user or not user.is_authenticated:
        return False
    if user.is_superuser or user.is_staff:
        return True
    allowed = {c.lower() for c in allowed_codes}
    return bool(_user_role_codes(user) & allowed)


def _user_has_payroll_admin_role(user) -> bool:
    """Legacy alias kept for clarity at call-sites that gate employees."""
    return _user_has_any_role(user, EMPLOYEE_WRITE_ROLE_CODES)


def _user_can_force_payroll_run(user) -> bool:
    """True if the user may override approval gates on Payroll Runs.

    Soft-coded via ``PAYROLL_RUN_FORCE_OVERRIDE_ROLES``. Reserved for
    super-admin style emergency edits to approved/released runs.
    """
    return _user_has_any_role(user, RUN_FORCE_OVERRIDE_ROLE_CODES)


class PayrollEmployeeWritePermission(IsAuthenticated):
    """Read = any authenticated user. Write = super-admin / configured RBAC roles.

    Soft-coded: extend ``PAYROLL_EMPLOYEE_WRITE_ROLES`` env var to grant
    additional role codes without touching code.
    """

    message = (
        'Only super-admins (or roles configured in PAYROLL_EMPLOYEE_WRITE_ROLES) '
        'may edit or delete payroll employee records.'
    )

    def has_permission(self, request, view):
        if not super().has_permission(request, view):
            return False
        if request.method in SAFE_METHODS:
            return True
        return _user_has_any_role(request.user, EMPLOYEE_WRITE_ROLE_CODES)


class PayrollAdjustmentWritePermission(IsAuthenticated):
    """Read = any authenticated user. Write = configured RBAC roles.

    Override via env var ``PAYROLL_ADJUSTMENT_WRITE_ROLES`` (CSV).
    Django super/staff are always allowed.
    """

    message = (
        'You do not have permission to create or modify payroll adjustments. '
        'Configure roles in PAYROLL_ADJUSTMENT_WRITE_ROLES.'
    )

    def has_permission(self, request, view):
        if not super().has_permission(request, view):
            return False
        if request.method in SAFE_METHODS:
            return True
        return _user_has_any_role(request.user, ADJUSTMENT_WRITE_ROLE_CODES)


# ── Catalog (read-only, public to authenticated users) ──────────────
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def catalog_view(request):
    # Dynamic lists — computed from live PayrollEmployee data so new values
    # appear in dropdowns automatically without any config change.
    from .models import PayrollEmployee as _PE
    _depts = list(
        _PE.objects.filter(is_active=True)
        .values_list('department', flat=True)
        .exclude(department='')
        .distinct()
        .order_by('department')
    )
    _desigs = list(
        _PE.objects.filter(is_active=True)
        .values_list('designation', flat=True)
        .exclude(designation='')
        .distinct()
        .order_by('designation')
    )
    payload = {
        'currency': CURRENCY,
        'workflow_statuses': catalog.WORKFLOW_STATUSES,
        'workflow_transitions': catalog.WORKFLOW_TRANSITIONS,
        'payment_modes': catalog.PAYMENT_MODES,
        'fixed_earnings': catalog.FIXED_EARNINGS,
        'earning_components': catalog.EARNING_COMPONENTS,
        'deduction_components': catalog.DEDUCTION_COMPONENTS,
        'line_item_kinds': catalog.LINE_ITEM_KINDS,
        'line_item_sources': catalog.LINE_ITEM_SOURCES,
        'adjustment_statuses': catalog.ADJUSTMENT_STATUSES,
        'grade_options': catalog.GRADE_OPTIONS,
        'nationality_groups': catalog.NATIONALITY_GROUPS,
        # Dynamic options derived from live employee roster
        'departments': _depts,
        'designations': _desigs,
    }
    return Response(CatalogSerializer(payload).data)


# ── Dashboard summary ───────────────────────────────────────────────
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def engine_dashboard_summary(request):
    now = timezone.now()
    year = int(request.query_params.get('year') or now.year)
    month = int(request.query_params.get('month') or now.month)

    current = PayrollRun.objects.filter(year=year, month=month).first()
    latest = PayrollRun.objects.order_by('-year', '-month').first()
    ytd_agg = PayrollRun.objects.filter(year=year).aggregate(
        ytd_net=Sum('total_net'),
        ytd_gross=Sum('total_gross'),
        ytd_deductions=Sum('total_deductions'),
    )

    pending_payslips = Payslip.objects.filter(
        run__status__in=[catalog.Status.DRAFT, catalog.Status.HR_APPROVED]
    ).count()

    payload = {
        'currency':           CURRENCY,
        'year':               year,
        'month':              month,
        'active_employees':   PayrollEmployee.objects.filter(is_active=True).count(),
        'total_employees':    PayrollEmployee.objects.count(),
        'current_run': PayrollRunSerializer(current).data if current else None,
        'current_month_gross':       str(current.total_gross if current else Decimal('0')),
        'current_month_net':         str(current.total_net if current else Decimal('0')),
        'current_month_deductions':  str(current.total_deductions if current else Decimal('0')),
        'current_month_slip_count':  current.employee_count if current else 0,
        'pending_approvals':         pending_payslips,
        'ytd_gross':         str(ytd_agg['ytd_gross'] or Decimal('0')),
        'ytd_net':           str(ytd_agg['ytd_net'] or Decimal('0')),
        'ytd_deductions':    str(ytd_agg['ytd_deductions'] or Decimal('0')),
        'latest_run': PayrollRunSerializer(latest).data if latest else None,
        'pending_adjustments': PayrollAdjustment.objects.filter(
            status=catalog.AdjustmentStatus.PENDING
        ).count(),
    }
    return Response(payload)


# ── Employee CRUD ───────────────────────────────────────────────────
# Soft-coded: PayrollEmployee fields that sync to Draft payslip snapshots.
# Extend here when a new field needs two-way sync.
_EMPLOYEE_TO_SNAPSHOT_SYNC: dict = {
    'department':  'snapshot_department',
    'designation': 'snapshot_designation',
    'joining_date': 'snapshot_joining_date',
}


class PayrollEmployeeViewSet(viewsets.ModelViewSet):
    queryset = PayrollEmployee.objects.all()
    serializer_class = PayrollEmployeeSerializer
    permission_classes = [PayrollEmployeeWritePermission]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['employee_no', 'full_name', 'department', 'designation', 'iban']
    ordering_fields = ['full_name', 'employee_no', 'department', 'basic']
    ordering = ['full_name']

    def _sync_to_draft_payslips(self, employee, request_data: dict) -> None:
        """
        When an employee's org/identity fields change, cascade the update to
        all Draft payslip snapshots for this employee so Monthly Runs and
        the Employees tab stay aligned.
        """
        snap_updates = {
            snap_field: request_data[emp_field]
            for emp_field, snap_field in _EMPLOYEE_TO_SNAPSHOT_SYNC.items()
            if emp_field in request_data
        }
        if snap_updates:
            (Payslip.objects
             .filter(employee=employee, run__status=catalog.Status.DRAFT)
             .update(**snap_updates))

    def update(self, request, *args, **kwargs):
        employee = self.get_object()
        response = super().update(request, *args, **kwargs)
        self._sync_to_draft_payslips(employee, request.data)
        return response

    def partial_update(self, request, *args, **kwargs):
        employee = self.get_object()
        response = super().partial_update(request, *args, **kwargs)
        self._sync_to_draft_payslips(employee, request.data)
        return response

    def get_queryset(self):
        qs = super().get_queryset()
        is_active = self.request.query_params.get('is_active')
        if is_active in ('true', '1'):
            qs = qs.filter(is_active=True)
        elif is_active in ('false', '0'):
            qs = qs.filter(is_active=False)
        department = self.request.query_params.get('department')
        if department:
            qs = qs.filter(department__iexact=department)
        return qs

    @action(detail=False, methods=['post'], url_path='import-xlsx')
    def import_xlsx(self, request):
        """Upload the master roster Excel and upsert employees."""
        upload = request.FILES.get('file')
        if not upload:
            return Response({'error': "Missing 'file' upload."}, status=http_status.HTTP_400_BAD_REQUEST)
        if upload.size > EXCEL_MAX_UPLOAD_MB * 1024 * 1024:
            return Response({'error': f'File exceeds {EXCEL_MAX_UPLOAD_MB} MB.'},
                            status=http_status.HTTP_400_BAD_REQUEST)
        summary = excel_import.import_master_roster(upload)
        return Response(summary.as_dict())

    @action(detail=False, methods=['get'], url_path='export-xlsx')
    def export_xlsx(self, request):
        """Download a fresh master roster (uses current employee data, not a run)."""
        # Build a temp in-memory workbook listing current employees
        from io import BytesIO
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Employees'
        headers = [
            'Employee No', 'Name', 'MOL', 'IBAN', 'Bank', 'Department',
            'Discipline', 'Designation', 'Grade', 'Nationality',
            'Joining Date', 'Hours', 'Basic', 'Housing', 'Transport', 'Home Leave',
            'Payment Mode', 'Active',
        ]
        for col, h in enumerate(headers, start=1):
            ws.cell(1, col, value=h)
        for i, emp in enumerate(PayrollEmployee.objects.all().order_by('full_name'), start=2):
            ws.cell(i, 1, emp.employee_no)
            ws.cell(i, 2, emp.full_name)
            ws.cell(i, 3, emp.mol_no)
            ws.cell(i, 4, emp.iban)
            ws.cell(i, 5, emp.bank_name)
            ws.cell(i, 6, emp.department)
            ws.cell(i, 7, emp.discipline)
            ws.cell(i, 8, emp.designation)
            ws.cell(i, 9, emp.grade)
            ws.cell(i, 10, emp.nationality_group)
            ws.cell(i, 11, emp.joining_date)
            ws.cell(i, 12, float(emp.hours) if emp.hours is not None else None)
            ws.cell(i, 13, float(emp.basic))
            ws.cell(i, 14, float(emp.housing))
            ws.cell(i, 15, float(emp.transport))
            ws.cell(i, 16, float(emp.home_leave))
            ws.cell(i, 17, emp.default_payment_mode)
            ws.cell(i, 18, emp.is_active)
        buf = BytesIO()
        wb.save(buf)
        response = HttpResponse(
            buf.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="payroll_employees.xlsx"'
        return response


# ── PayrollRun CRUD + workflow + Excel ──────────────────────────────
class PayrollRunViewSet(viewsets.ModelViewSet):
    queryset = PayrollRun.objects.all()
    serializer_class = PayrollRunSerializer
    permission_classes = [IsAuthenticated]
    ordering = ['-year', '-month']

    def get_queryset(self):
        qs = super().get_queryset()
        year = self.request.query_params.get('year')
        month = self.request.query_params.get('month')
        st = self.request.query_params.get('status')
        if year:
            qs = qs.filter(year=int(year))
        if month:
            qs = qs.filter(month=int(month))
        if st:
            qs = qs.filter(status=st)
        return qs

    def create(self, request, *args, **kwargs):
        """POST { year, month, overwrite?, working_days? } → generates the run."""
        year = int(request.data.get('year') or 0)
        month = int(request.data.get('month') or 0)
        if not (year and 1 <= month <= 12):
            return Response({'error': 'year and month (1-12) required.'}, status=400)
        overwrite = str(request.data.get('overwrite') or '').lower() in ('true', '1', 'yes')
        note = request.data.get('note') or ''
        # working_days: HR-supplied, defaults to catalog constant
        from .catalog import DEFAULT_WORKING_DAYS_PER_MONTH
        try:
            working_days = int(request.data.get('working_days') or DEFAULT_WORKING_DAYS_PER_MONTH)
            if not (1 <= working_days <= 31):
                return Response({'error': 'working_days must be between 1 and 31.'}, status=400)
        except (ValueError, TypeError):
            working_days = DEFAULT_WORKING_DAYS_PER_MONTH
        try:
            run = run_generator.generate_monthly_run(
                year, month,
                user=request.user, overwrite=overwrite, note=note,
                working_days=working_days,
            )
        except run_generator.GenerationError as exc:
            return Response({'error': str(exc)}, status=409)
        return Response(self.get_serializer(run).data, status=201)

    def destroy(self, request, *args, **kwargs):
        """Delete a payroll run.

        Standard rule: only Draft runs are deletable. Super-admins (or any
        role listed in ``PAYROLL_RUN_FORCE_OVERRIDE_ROLES``) may pass
        ``?force=true`` to delete an *approved or released* run as an
        emergency override. Every force-delete is recorded in
        ``PayrollWorkflowLog`` for audit.
        """
        run = self.get_object()
        force_flag = str(request.query_params.get('force')
                         or request.data.get('force')
                         or '').lower() in ('true', '1', 'yes')

        if not run.is_editable:
            if not force_flag:
                return Response(
                    {'error': (
                        f'Run {run.cycle_code} is "{run.status}"; only Draft runs '
                        'can be deleted. Revert it to Draft first — or, if you are a '
                        'Super Administrator, retry with force=true to override.'
                    )},
                    status=409,
                )
            if not _user_can_force_payroll_run(request.user):
                return Response(
                    {'error': (
                        'Force delete is restricted to Super Administrators '
                        f'(roles configured in PAYROLL_RUN_FORCE_OVERRIDE_ROLES).'
                    )},
                    status=403,
                )
            # Stamp an audit log row *before* the cascade wipes the run.
            note = (
                request.data.get('note')
                or f"FORCE DELETE of {run.status.upper()} run by "
                f"{getattr(request.user, 'username', 'anonymous')}"
            )
            PayrollWorkflowLog.objects.create(
                run=run,
                from_status=run.status,
                to_status='deleted',
                actor=request.user if request.user.is_authenticated else None,
                note=note,
            )

        cycle = run.cycle_code
        was_status = run.status
        super().destroy(request, *args, **kwargs)
        return Response(
            {'deleted': True, 'cycle_code': cycle, 'previous_status': was_status, 'forced': force_flag},
            status=200,
        )

    def update(self, request, *args, **kwargs):
        run = self.get_object()
        if not run.is_editable and not _user_can_force_payroll_run(request.user):
            return Response(
                {'error': (
                    f'Run {run.cycle_code} is "{run.status}" and cannot be edited. '
                    'Only Super Administrators may force-edit approved or released runs.'
                )},
                status=409,
            )
        return super().update(request, *args, **kwargs)

    def partial_update(self, request, *args, **kwargs):
        run = self.get_object()
        if not run.is_editable and not _user_can_force_payroll_run(request.user):
            return Response(
                {'error': (
                    f'Run {run.cycle_code} is "{run.status}" and cannot be edited. '
                    'Only Super Administrators may force-edit approved or released runs.'
                )},
                status=409,
            )
        return super().partial_update(request, *args, **kwargs)

    @action(detail=True, methods=['post'], url_path='force-revert')
    def force_revert(self, request, pk=None):
        """Emergency revert: move *any* run back to Draft regardless of status.

        Restricted to roles in ``PAYROLL_RUN_FORCE_OVERRIDE_ROLES``. Clears the
        HR/Finance/Released stamps so the workflow restarts cleanly. Audit-logged.
        """
        run = self.get_object()
        if not _user_can_force_payroll_run(request.user):
            return Response(
                {'error': (
                    'Force revert is restricted to Super Administrators '
                    '(roles configured in PAYROLL_RUN_FORCE_OVERRIDE_ROLES).'
                )},
                status=403,
            )
        if run.status == catalog.Status.DRAFT:
            return Response(self.get_serializer(run).data)

        from_status = run.status
        # Wipe forward stamps so the next approval cycle is clean.
        run.status = catalog.Status.DRAFT
        run.hr_approved_at = None
        run.hr_approved_by = None
        run.finance_approved_at = None
        run.finance_approved_by = None
        run.released_at = None
        run.released_by = None
        run.save(update_fields=[
            'status',
            'hr_approved_at', 'hr_approved_by',
            'finance_approved_at', 'finance_approved_by',
            'released_at', 'released_by',
            'updated_at',
        ])
        PayrollWorkflowLog.objects.create(
            run=run,
            from_status=from_status,
            to_status=catalog.Status.DRAFT,
            actor=request.user if request.user.is_authenticated else None,
            note=(request.data.get('note')
                  or f"FORCE REVERT from {from_status.upper()} by "
                     f"{getattr(request.user, 'username', 'anonymous')}"),
        )
        return Response(self.get_serializer(run).data)

    @action(detail=True, methods=['post'], url_path='regenerate')
    def regenerate(self, request, pk=None):
        run = self.get_object()
        if run.status != catalog.Status.DRAFT:
            return Response({'error': 'Only Draft runs can be regenerated.'}, status=409)
        try:
            run = run_generator.generate_monthly_run(
                run.year, run.month, user=request.user, overwrite=True,
                note=request.data.get('note') or run.notes,
            )
        except run_generator.GenerationError as exc:
            return Response({'error': str(exc)}, status=409)
        return Response(self.get_serializer(run).data)

    @action(detail=True, methods=['post'], url_path='refresh-totals')
    def refresh_totals(self, request, pk=None):
        run = self.get_object()
        run_generator.refresh_run_totals(run)
        return Response(self.get_serializer(run).data)

    @action(detail=True, methods=['post'], url_path='refresh-hours-from-timesheet')
    def refresh_hours_from_timesheet(self, request, pk=None):
        """Pull each payslip's `hours` from the live Time Sheet Summary
        ("Total" column = biometric punches with HR overrides overlaid).

        Draft runs: any authenticated user with write access.
        Approved / Finance-approved / Released runs: Super-Admin only,
        requires `?force=true` (mirrors the force-delete / force-revert
        pattern so the override is intentional and audit-logged).

        Query/body params:
          force          — bool, allow overriding non-draft runs (Super-Admin only).
          zero_missing   — bool, set hours=0 for employees with no biometric
                           data this month so payroll mirrors Attendance Total
                           exactly. Defaults to PAYROLL_REFRESH_ZERO_MISSING.
        """
        run = self.get_object()
        force = str(request.query_params.get('force', '')).lower() in ('1', 'true', 'yes')

        # zero_missing may come from query string or body; None ⇒ use config default.
        zm_raw = request.query_params.get('zero_missing')
        if zm_raw is None:
            zm_raw = request.data.get('zero_missing') if hasattr(request, 'data') else None
        zero_missing: bool | None
        if zm_raw is None or zm_raw == '':
            zero_missing = None
        else:
            zero_missing = str(zm_raw).lower() in ('1', 'true', 'yes')

        if run.status != catalog.Status.DRAFT:
            if not force:
                return Response(
                    {'error': f'Run is {run.status}. Pass ?force=true to refresh hours on an approved run.'},
                    status=409,
                )
            if not _user_can_force_payroll_run(request.user):
                return Response(
                    {'error': 'Only Super Administrators can force-refresh hours on approved runs.'},
                    status=403,
                )

        result = run_generator.refresh_run_hours_from_timesheet(run, zero_missing=zero_missing)
        return Response({
            'run': self.get_serializer(run).data,
            'forced': force and run.status != catalog.Status.DRAFT,
            **result,
        })

    @action(detail=True, methods=['post'], url_path='upload-external',
            parser_classes=None)
    def upload_external(self, request, pk=None):
        """
        POST /api/v1/payroll-engine/runs/{id}/upload-external/

        Upload a ValueFrame or Sympa XLSX and apply the data directly to
        the run's Draft payslips.

        Multipart fields:
          file       — the XLSX file
          file_type  — 'valueframe' | 'sympa' | 'generic'  (default: 'valueframe')

        Run must be in DRAFT status.
        Returns upload summary: rows_matched, rows_updated, unmatched, updated_fields, s3_key.
        """
        from rest_framework.parsers import MultiPartParser
        run = self.get_object()
        if run.status != catalog.Status.DRAFT:
            return Response(
                {'error': f'Run {run.cycle_code} is "{run.status}". Only Draft runs accept external imports.'},
                status=409,
            )
        file_obj  = request.FILES.get('file')
        file_type = (request.data.get('file_type') or 'valueframe').lower()
        if not file_obj:
            return Response({'error': "Missing 'file' upload."}, status=400)
        if file_type not in ('valueframe', 'sympa', 'generic'):
            return Response({'error': f"Unknown file_type '{file_type}'. Use valueframe, sympa, or generic."}, status=400)

        from .services import external_import
        try:
            result = external_import.apply_external_file(
                run=run,
                file_bytes=file_obj.read(),
                original_filename=file_obj.name,
                file_type=file_type,
                user=request.user,
            )
            return Response(result, status=201)
        except ValueError as exc:
            return Response({'error': str(exc)}, status=400)
        except Exception as exc:
            logger.exception('upload_external failed for run %s: %s', run.cycle_code, exc)
            return Response({'error': 'Import failed. Check server logs.'}, status=500)

    @action(detail=True, methods=['get'], url_path='uploads')
    def list_uploads(self, request, pk=None):
        """
        GET /api/v1/payroll-engine/runs/{id}/uploads/
        Returns all external file uploads for this run, newest first.
        """
        from .models import PayrollRunUpload
        from .serializers import PayrollRunUploadSerializer
        run = self.get_object()
        qs  = PayrollRunUpload.objects.filter(run=run).order_by('-uploaded_at')
        return Response({'results': PayrollRunUploadSerializer(qs, many=True).data})

    # Workflow transitions
    def _transition_action(self, request, pk, fn):
        run = self.get_object()
        try:
            fn(run, user=request.user, note=request.data.get('note') or '')
        except Exception as exc:
            return Response({'error': str(exc)}, status=400)
        return Response(self.get_serializer(run).data)

    @action(detail=True, methods=['post'], url_path='hr-approve')
    def hr_approve(self, request, pk=None):
        return self._transition_action(request, pk, workflow.hr_approve)

    @action(detail=True, methods=['post'], url_path='finance-approve')
    def finance_approve(self, request, pk=None):
        return self._transition_action(request, pk, workflow.finance_approve)

    @action(detail=True, methods=['post'], url_path='release')
    def release(self, request, pk=None):
        return self._transition_action(request, pk, workflow.release)

    @action(detail=True, methods=['post'], url_path='revert')
    def revert(self, request, pk=None):
        return self._transition_action(request, pk, workflow.revert_to_draft)

    # Excel I/O on a run
    @action(detail=True, methods=['get'], url_path='download-master-xlsx')
    def download_master_xlsx(self, request, pk=None):
        run = self.get_object()
        blob = excel_export.export_master_xlsx(run)
        resp = HttpResponse(
            blob,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        resp['Content-Disposition'] = f'attachment; filename="payroll_master_{run.cycle_code}.xlsx"'
        return resp

    @action(detail=True, methods=['get'], url_path='download-payslip-pack')
    def download_payslip_pack(self, request, pk=None):
        run = self.get_object()
        blob = excel_export.export_payslip_pack(run)
        resp = HttpResponse(
            blob,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        resp['Content-Disposition'] = f'attachment; filename="payroll_payslips_{run.cycle_code}.xlsx"'
        return resp

    @action(detail=True, methods=['post'], url_path='upload-adjustments')
    def upload_adjustments(self, request, pk=None):
        run = self.get_object()
        if run.status != catalog.Status.DRAFT:
            return Response({'error': 'Run is not in Draft.'}, status=409)
        upload = request.FILES.get('file')
        if not upload:
            return Response({'error': "Missing 'file'."}, status=400)
        if upload.size > EXCEL_MAX_UPLOAD_MB * 1024 * 1024:
            return Response({'error': f'File exceeds {EXCEL_MAX_UPLOAD_MB} MB.'}, status=400)
        summary = excel_import.import_adjustments(upload, run.year, run.month)
        # Re-generate so the adjustments materialise into line items
        run = run_generator.generate_monthly_run(
            run.year, run.month, user=request.user, overwrite=True, note=run.notes,
        )
        payload = summary.as_dict()
        payload['run'] = self.get_serializer(run).data
        return Response(payload)

    @action(detail=False, methods=['post'], url_path='import-full-xlsx')
    def import_full_xlsx(self, request):
        """Upload Excel + generate run in one shot. multipart: file, year, month."""
        upload = request.FILES.get('file')
        year = int(request.data.get('year') or 0)
        month = int(request.data.get('month') or 0)
        if not upload or not (year and 1 <= month <= 12):
            return Response({'error': "file, year, and month (1-12) required."}, status=400)
        if upload.size > EXCEL_MAX_UPLOAD_MB * 1024 * 1024:
            return Response({'error': f'File exceeds {EXCEL_MAX_UPLOAD_MB} MB.'}, status=400)
        summary = excel_import.import_full_payroll(upload, year, month)
        try:
            run = run_generator.generate_monthly_run(
                year, month, user=request.user, overwrite=True, note='Imported from Excel',
            )
        except run_generator.GenerationError as exc:
            return Response({'error': str(exc), 'import_summary': summary.as_dict()}, status=409)
        return Response({'run': self.get_serializer(run).data, 'import_summary': summary.as_dict()})

    @action(detail=True, methods=['get'], url_path='workflow-log')
    def workflow_log(self, request, pk=None):
        run = self.get_object()
        logs = run.workflow_logs.all()
        return Response(PayrollWorkflowLogSerializer(logs, many=True).data)

    @action(detail=True, methods=['post'], url_path='bulk-deduction')
    def bulk_deduction_action(self, request, pk=None):
        """Apply a percentage-based deduction to every payslip in this run.

        Body: {
          percentage: <0.01..100>,
          fields: ['housing','transport','home_leave','other_earnings']  # optional
          label:       str  # optional
          description: str  # optional
          replace_existing: bool  # default True
        }
        """
        run = self.get_object()
        if 'percentage' not in request.data:
            return Response({'error': 'percentage is required.'}, status=400)
        try:
            summary = bulk_deduction.apply_bulk_percentage_deduction(
                run,
                percentage=request.data.get('percentage'),
                fields=request.data.get('fields'),
                label=request.data.get('label') or None,
                description=request.data.get('description', '') or '',
                replace_existing=bool(request.data.get('replace_existing', True)),
            )
        except bulk_deduction.BulkDeductionError as exc:
            return Response({'error': str(exc)}, status=400)
        run.refresh_from_db()
        return Response({
            'run': self.get_serializer(run).data,
            'summary': summary.to_dict(),
        })

    @action(detail=True, methods=['post'], url_path='bulk-deduction/reverse')
    def bulk_deduction_reverse(self, request, pk=None):
        """Remove every bulk percentage-deduction line item from this run."""
        run = self.get_object()
        try:
            removed = bulk_deduction.reverse_bulk_percentage_deduction(run)
        except bulk_deduction.BulkDeductionError as exc:
            return Response({'error': str(exc)}, status=400)
        run.refresh_from_db()
        return Response({
            'run': self.get_serializer(run).data,
            'removed_line_items': removed,
        })


# ── Payslip CRUD ────────────────────────────────────────────────────
# Soft-coded mapping: Payslip snapshot field → PayrollEmployee field.
# Edit a payslip snapshot → the linked employee record also updates, and vice versa.
# Add an entry here to extend the sync to additional fields.
_PAYSLIP_SNAPSHOT_SYNC: dict = {
    'snapshot_department':  'department',
    'snapshot_designation': 'designation',
    'snapshot_joining_date': 'joining_date',
}


class PayslipViewSet(viewsets.ModelViewSet):
    queryset = Payslip.objects.select_related('employee', 'run').prefetch_related('line_items')
    serializer_class = PayslipSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        qs = super().get_queryset()
        run = self.request.query_params.get('run')
        emp = self.request.query_params.get('employee')
        st = self.request.query_params.get('status')
        q = self.request.query_params.get('q')
        if run:
            qs = qs.filter(run_id=run)
        if emp:
            qs = qs.filter(employee_id=emp)
        if st:
            qs = qs.filter(status=st)
        if q:
            qs = qs.filter(
                Q(snapshot_full_name__icontains=q)
                | Q(employee__employee_no__icontains=q)
                | Q(snapshot_department__icontains=q)
            )
        return qs.order_by('snapshot_full_name')

    def update(self, request, *args, **kwargs):
        slip = self.get_object()
        if not slip.run.is_editable:
            return Response({'error': 'Run is not editable.'}, status=409)
        resp = super().update(request, *args, **kwargs)
        slip.refresh_from_db()
        # If hours were updated, recompute days (hours ÷ HOURS_PER_WORKDAY)
        if 'hours' in request.data:
            from .config import hours_to_days
            slip.days = hours_to_days(slip.hours)
            slip.save(update_fields=['days'])
        recompute_payslip_totals(slip)
        slip.save()
        recompute_run_totals(slip.run)
        slip.run.save()
        # ── Two-way sync: snapshot changes → update the source PayrollEmployee ──
        emp_updates = {
            emp_field: request.data[snap_field]
            for snap_field, emp_field in _PAYSLIP_SNAPSHOT_SYNC.items()
            if snap_field in request.data
        }
        if emp_updates and slip.employee_id:
            PayrollEmployee.objects.filter(id=slip.employee_id).update(**emp_updates)
        return Response(self.get_serializer(slip).data)

    def destroy(self, request, *args, **kwargs):
        slip = self.get_object()
        if not slip.run.is_editable:
            return Response(
                {'error': f'Run {slip.run.cycle_code} is "{slip.run.status}"; only Draft runs allow deleting payslips.'},
                status=409,
            )
        run = slip.run
        emp_label = slip.snapshot_full_name or (slip.employee.full_name if slip.employee_id else f'payslip #{slip.pk}')
        super().destroy(request, *args, **kwargs)
        recompute_run_totals(run)
        run.save()
        return Response(
            {'deleted': True, 'employee': emp_label, 'run': PayrollRunSerializer(run).data},
            status=200,
        )

    @action(detail=True, methods=['get'], url_path='download-xlsx')
    def download_xlsx(self, request, pk=None):
        slip = self.get_object()
        blob = excel_export.export_single_payslip(slip)
        resp = HttpResponse(
            blob,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        resp['Content-Disposition'] = (
            f'attachment; filename="payslip_{slip.employee.employee_no}_{slip.run.cycle_code}.xlsx"'
        )
        return resp


# ── PayslipLineItem CRUD ────────────────────────────────────────────
class PayslipLineItemViewSet(viewsets.ModelViewSet):
    queryset = PayslipLineItem.objects.all()
    serializer_class = PayslipLineItemSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        qs = super().get_queryset()
        slip = self.request.query_params.get('payslip')
        if slip:
            qs = qs.filter(payslip_id=slip)
        return qs.order_by('kind', 'component_code')

    def perform_create(self, serializer):
        slip = serializer.validated_data['payslip']
        if not slip.run.is_editable:
            from rest_framework.exceptions import ValidationError as DRFValidationError
            raise DRFValidationError({'error': 'Run is not editable.'})
        item = serializer.save(created_by=self.request.user if self.request.user.is_authenticated else None)
        recompute_payslip_totals(item.payslip)
        item.payslip.save()
        recompute_run_totals(item.payslip.run)
        item.payslip.run.save()

    def perform_update(self, serializer):
        slip = serializer.instance.payslip
        if not slip.run.is_editable:
            from rest_framework.exceptions import ValidationError as DRFValidationError
            raise DRFValidationError({'error': 'Run is not editable.'})
        item = serializer.save()
        recompute_payslip_totals(item.payslip)
        item.payslip.save()
        recompute_run_totals(item.payslip.run)
        item.payslip.run.save()

    def perform_destroy(self, instance):
        slip = instance.payslip
        if not slip.run.is_editable:
            from rest_framework.exceptions import ValidationError as DRFValidationError
            raise DRFValidationError({'error': 'Run is not editable.'})
        instance.delete()
        recompute_payslip_totals(slip)
        slip.save()
        recompute_run_totals(slip.run)
        slip.run.save()


# ── Adjustments CRUD ────────────────────────────────────────────────
class PayrollAdjustmentViewSet(viewsets.ModelViewSet):
    queryset = PayrollAdjustment.objects.select_related('employee')
    serializer_class = PayrollAdjustmentSerializer
    permission_classes = [PayrollAdjustmentWritePermission]

    def get_queryset(self):
        qs = super().get_queryset()
        year = self.request.query_params.get('year')
        month = self.request.query_params.get('month')
        st = self.request.query_params.get('status')
        emp = self.request.query_params.get('employee')
        kind = self.request.query_params.get('kind')
        search = self.request.query_params.get('search')
        if year:
            qs = qs.filter(target_year=int(year))
        if month:
            qs = qs.filter(target_month=int(month))
        if st:
            qs = qs.filter(status=st)
        if emp:
            qs = qs.filter(employee_id=emp)
        if kind:
            qs = qs.filter(kind=kind)
        if search:
            qs = qs.filter(
                Q(employee__full_name__icontains=search)
                | Q(employee__employee_no__icontains=search)
                | Q(label__icontains=search)
                | Q(description__icontains=search)
                | Q(component_code__icontains=search)
            )
        return qs

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user if self.request.user.is_authenticated else None)

    def _block_if_applied(self, instance):
        if instance.status == catalog.AdjustmentStatus.APPLIED:
            return Response(
                {'error': (
                    f'Adjustment is already applied to payslip #{instance.applied_to_id}; '
                    'applied adjustments cannot be modified. Revert the run to Draft first.'
                )},
                status=409,
            )
        return None

    def update(self, request, *args, **kwargs):
        instance = self.get_object()
        blocker = self._block_if_applied(instance)
        if blocker is not None:
            return blocker
        return super().update(request, *args, **kwargs)

    def partial_update(self, request, *args, **kwargs):
        instance = self.get_object()
        blocker = self._block_if_applied(instance)
        if blocker is not None:
            return blocker
        return super().partial_update(request, *args, **kwargs)

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        blocker = self._block_if_applied(instance)
        if blocker is not None:
            return blocker
        return super().destroy(request, *args, **kwargs)

    @action(detail=True, methods=['post'], url_path='cancel')
    def cancel(self, request, pk=None):
        """Mark a pending adjustment as cancelled. Idempotent for already-cancelled."""
        adj = self.get_object()
        if adj.status == catalog.AdjustmentStatus.APPLIED:
            return Response(
                {'error': 'Cannot cancel an adjustment that has already been applied.'},
                status=409,
            )
        if adj.status == catalog.AdjustmentStatus.CANCELLED:
            return Response(self.get_serializer(adj).data)
        adj.status = catalog.AdjustmentStatus.CANCELLED
        adj.save(update_fields=['status', 'updated_at'])
        return Response(self.get_serializer(adj).data)

    @action(detail=True, methods=['post'], url_path='reopen')
    def reopen(self, request, pk=None):
        """Move a cancelled adjustment back to pending so it can be reused."""
        adj = self.get_object()
        if adj.status != catalog.AdjustmentStatus.CANCELLED:
            return Response(
                {'error': 'Only cancelled adjustments can be re-opened.'},
                status=409,
            )
        adj.status = catalog.AdjustmentStatus.PENDING
        adj.save(update_fields=['status', 'updated_at'])
        return Response(self.get_serializer(adj).data)

    @action(detail=False, methods=['post'], url_path='bulk-cancel')
    def bulk_cancel(self, request):
        """POST { ids: [<int>, ...] } → mark each non-applied adjustment cancelled."""
        ids = request.data.get('ids') or []
        if not isinstance(ids, list) or not ids:
            return Response({'error': 'Provide a non-empty list of ids.'}, status=400)
        qs = (PayrollAdjustment.objects
              .filter(id__in=ids)
              .exclude(status=catalog.AdjustmentStatus.APPLIED))
        updated = qs.update(status=catalog.AdjustmentStatus.CANCELLED)
        return Response({'cancelled': updated, 'requested': len(ids)})

    @action(detail=False, methods=['get'], url_path='summary')
    def summary(self, request):
        """Aggregate totals by kind/status for the active filter window.

        Honours the same query params as `list` (year, month, status, employee).
        """
        from django.db.models import Count
        qs = self.filter_queryset(self.get_queryset())
        totals_by_kind = list(
            qs.values('kind')
              .annotate(count=Count('id'), total=Sum('amount'))
              .order_by('kind')
        )
        totals_by_status = list(
            qs.values('status')
              .annotate(count=Count('id'), total=Sum('amount'))
              .order_by('status')
        )
        return Response({
            'count': qs.count(),
            'total_amount': qs.aggregate(s=Sum('amount'))['s'] or 0,
            'by_kind': totals_by_kind,
            'by_status': totals_by_status,
        })


# ── Workflow log (read-only) ────────────────────────────────────────
class PayrollWorkflowLogViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = PayrollWorkflowLog.objects.select_related('run', 'actor')
    serializer_class = PayrollWorkflowLogSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        qs = super().get_queryset()
        run = self.request.query_params.get('run')
        if run:
            qs = qs.filter(run_id=run)
        return qs.order_by('-at')


# ── Comparison (reconcile a run vs an external HR file) ─────────
class PayrollComparisonViewSet(viewsets.ModelViewSet):
    """Upload an external XLSX/CSV (ValueFrame, Sympa, generic) and diff
    it against the selected PayrollRun.
    """
    queryset = PayrollComparison.objects.select_related('run', 'uploaded_by')
    serializer_class = PayrollComparisonSerializer
    permission_classes = [IsAuthenticated]
    http_method_names = ['get', 'post', 'delete']

    def get_queryset(self):
        qs = super().get_queryset()
        run = self.request.query_params.get('run')
        if run:
            qs = qs.filter(run_id=run)
        profile = self.request.query_params.get('profile')
        if profile:
            qs = qs.filter(source_profile=profile)
        return qs.order_by('-created_at')

    def get_serializer_class(self):
        if self.action == 'retrieve':
            return PayrollComparisonDetailSerializer
        return PayrollComparisonSerializer

    def create(self, request, *args, **kwargs):
        run_id = request.data.get('run')
        if not run_id:
            return Response({'error': '`run` (PayrollRun id) is required.'}, status=400)
        upload = request.FILES.get('file')
        if not upload:
            return Response({'error': "Missing 'file' upload."}, status=400)
        if upload.size > EXCEL_MAX_UPLOAD_MB * 1024 * 1024:
            return Response({'error': f'File exceeds {EXCEL_MAX_UPLOAD_MB} MB.'}, status=400)
        try:
            run = PayrollRun.objects.get(pk=run_id)
        except PayrollRun.DoesNotExist:
            return Response({'error': f'PayrollRun {run_id} not found.'}, status=404)
        source_profile = request.data.get('source_profile') or 'auto'
        source_label = request.data.get('source_label') or \
            catalog.comparison_profile(source_profile).get('label', source_profile)
        try:
            comparison = comparison_service.run_comparison(
                run=run,
                file_obj=upload,
                source_label=source_label,
                source_profile=source_profile,
                uploaded_by=request.user if request.user.is_authenticated else None,
                source_filename=upload.name,
            )
        except Exception as exc:
            return Response(
                {'error': f'Comparison failed: {exc.__class__.__name__}: {exc}'},
                status=400,
            )
        return Response(
            PayrollComparisonDetailSerializer(comparison, context={'request': request}).data,
            status=201,
        )

    @action(detail=True, methods=['get'], url_path='rows')
    def rows(self, request, pk=None):
        comparison = self.get_object()
        rows_qs = comparison.rows.select_related('payroll_employee').all()
        status_filter = request.query_params.get('status')
        if status_filter:
            rows_qs = rows_qs.filter(status=status_filter)
        search = request.query_params.get('search')
        if search:
            rows_qs = rows_qs.filter(
                Q(external_name__icontains=search)
                | Q(external_employee_no__icontains=search)
                | Q(payroll_employee__full_name__icontains=search)
                | Q(payroll_employee__employee_no__icontains=search)
            )
        page = self.paginate_queryset(rows_qs)
        ser = PayrollComparisonRowSerializer
        if page is not None:
            return self.get_paginated_response(ser(page, many=True).data)
        return Response(ser(rows_qs, many=True).data)

    @action(detail=True, methods=['get'], url_path='export-xlsx')
    def export_xlsx(self, request, pk=None):
        comparison = self.get_object()
        from io import BytesIO
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Summary'

        s = comparison.summary or {}
        ws.cell(1, 1, value='Payroll Comparison Report').font = Font(bold=True, size=14)
        ws.cell(2, 1, value='Run');               ws.cell(2, 2, value=comparison.run.cycle_code)
        ws.cell(3, 1, value='External source');   ws.cell(3, 2, value=comparison.source_label)
        ws.cell(4, 1, value='Profile');           ws.cell(4, 2, value=comparison.source_profile)
        ws.cell(5, 1, value='Uploaded file');     ws.cell(5, 2, value=comparison.source_filename)
        ws.cell(6, 1, value='Generated at');      ws.cell(6, 2, value=comparison.created_at.strftime('%Y-%m-%d %H:%M'))
        ws.cell(8, 1, value='Matches');           ws.cell(8, 2, value=s.get('matched', 0))
        ws.cell(9, 1, value='Variances');         ws.cell(9, 2, value=s.get('variance', 0))
        ws.cell(10, 1, value='External-only');    ws.cell(10, 2, value=s.get('external_only', 0))
        ws.cell(11, 1, value='Missing from external'); ws.cell(11, 2, value=s.get('payroll_only', 0))

        # Per-field breakdown
        ws.cell(13, 1, value='Field').font = Font(bold=True)
        ws.cell(13, 2, value='Variances').font = Font(bold=True)
        ws.cell(13, 3, value='Critical').font = Font(bold=True)
        ws.cell(13, 4, value='Warning').font = Font(bold=True)
        by_field = s.get('by_field') or {}
        r = 14
        for fmeta in catalog.COMPARISON_FIELDS:
            f = fmeta['field']
            stats = by_field.get(f, {})
            ws.cell(r, 1, value=fmeta['label'])
            ws.cell(r, 2, value=stats.get('variances', 0))
            ws.cell(r, 3, value=stats.get('critical', 0))
            ws.cell(r, 4, value=stats.get('warning', 0))
            r += 1

        # Variance detail sheet
        ws2 = wb.create_sheet('Variances')
        headers = [
            'Employee No', 'Name', 'Status', 'Matched By', 'Field',
            'Our Value', 'External Value', 'Diff', 'Diff %', 'Severity',
            'Recommendation',
        ]
        for c, h in enumerate(headers, start=1):
            ws2.cell(1, c, value=h).font = Font(bold=True)
        sev_fill = {
            'critical': PatternFill('solid', fgColor='FFE4E6'),
            'warning':  PatternFill('solid', fgColor='FEF3C7'),
            'info':     PatternFill('solid', fgColor='E0F2FE'),
        }
        r = 2
        for row in comparison.rows.select_related('payroll_employee').order_by('status', 'external_name'):
            emp_no = (row.payroll_employee.employee_no if row.payroll_employee
                      else row.external_employee_no)
            emp_name = (row.payroll_employee.full_name if row.payroll_employee
                        else row.external_name)
            variances = row.variances or []
            if not variances:
                ws2.cell(r, 1, value=emp_no)
                ws2.cell(r, 2, value=emp_name)
                ws2.cell(r, 3, value=catalog.COMPARISON_STATUS_LABELS.get(row.status, {}).get('label', row.status))
                ws2.cell(r, 4, value=row.matched_by)
                r += 1
                continue
            for v in variances:
                ws2.cell(r, 1, value=emp_no)
                ws2.cell(r, 2, value=emp_name)
                ws2.cell(r, 3, value=catalog.COMPARISON_STATUS_LABELS.get(row.status, {}).get('label', row.status))
                ws2.cell(r, 4, value=row.matched_by)
                ws2.cell(r, 5, value=catalog.comparison_field_meta(v.get('field', '')).get('label', v.get('field', '')))
                ws2.cell(r, 6, value=v.get('our'))
                ws2.cell(r, 7, value=v.get('external'))
                ws2.cell(r, 8, value=v.get('diff'))
                ws2.cell(r, 9, value=v.get('pct'))
                sev = v.get('severity', '')
                ws2.cell(r, 10, value=sev)
                ws2.cell(r, 11, value=v.get('recommendation', ''))
                fill = sev_fill.get(sev)
                if fill:
                    for c in range(1, 12):
                        ws2.cell(r, c).fill = fill
                r += 1

        # Column widths
        for col_letter, width in zip('ABCDEFGHIJK',
                                     [14, 28, 22, 14, 18, 12, 12, 12, 10, 10, 60]):
            ws2.column_dimensions[col_letter].width = width

        buf = BytesIO()
        wb.save(buf)
        resp = HttpResponse(
            buf.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        fname = f'comparison_{comparison.run.cycle_code}_{comparison.source_profile}_{comparison.id}.xlsx'
        resp['Content-Disposition'] = f'attachment; filename="{fname}"'
        return resp

    @action(detail=False, methods=['get'], url_path='profiles')
    def profiles(self, request):
        """Return the list of available comparison profiles for the dropdown."""
        return Response([
            {'code': code, 'label': p['label']}
            for code, p in catalog.COMPARISON_PROFILES.items()
        ])
