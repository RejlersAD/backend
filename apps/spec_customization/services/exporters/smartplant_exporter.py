"""
SmartPlant 3D Spec/Catalog Exporter Engine
==========================================

Generates two output workbooks from a completed `PaperSpecExtractionJob`:

  • spec_customisation_<job>_SPEC.xlsx  — piping spec rules (25 sheets)
  • spec_customisation_<job>_CAT.xlsx   — component catalog (23 sheets)

Templates ship in `services/templates/` and define the authoritative sheet /
column structure.  This engine:

  1. Opens the template (preserving every reference / lookup sheet untouched).
  2. For each sheet that has a soft-coded builder in `smartplant_config`, it:
        a. Locates the `Head` row (column-header row).
        b. Locates the `End` marker — or the next empty row — for safe inject.
        c. Builds rows from the PipingClass(es) and writes by header NAME
           (never by column index — keeps it resilient to template edits).
  3. Saves to a BytesIO buffer the caller can stream as an HTTP attachment.

No data is destroyed: extracted rows are *appended* into each region just
before its `End` marker, so the template's example/reference rows remain
intact for cross-checking.
"""
from __future__ import annotations

import io
import logging
from typing import Iterable, Optional

from openpyxl import load_workbook

from . import smartplant_config as cfg
from . import workbook_preview as preview

logger = logging.getLogger(__name__)


# ─────────────────────────────────────────────────────────────────────────────
# Sheet-region introspection
# ─────────────────────────────────────────────────────────────────────────────
def _find_marker_row(ws, markers: tuple, max_search: int = 40) -> Optional[int]:
    """Return the 1-based row index where column A matches any of `markers`."""
    for r in range(1, min(ws.max_row, max_search) + 1):
        v = ws.cell(r, 1).value
        if v is not None and str(v).strip() in markers:
            return r
    return None


def _find_header_row(ws) -> Optional[int]:
    """SPEC sheets use 'Head' in column A; CAT sheets may have multiple
    'Head' rows (definition block + CommodityPart block) — we want the
    one *closest* to the data we'll insert, which is the CommodityPart-
    region `Head` (i.e. the LAST `Head` before any `Start` we'll use)."""
    last = None
    for r in range(1, min(ws.max_row, 60) + 1):
        v = ws.cell(r, 1).value
        if v is not None and str(v).strip() in cfg.HEADER_MARKERS:
            last = r
    return last


def _read_header_map(ws, header_row: int) -> dict[str, int]:
    """Map column-header name → 1-based column index."""
    out: dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(header_row, c).value
        if v is not None:
            name = str(v).strip()
            if name and name not in out:
                out[name] = c
    return out


def _find_insert_row(ws, start_row: int) -> int:
    """Locate the row to insert at: either the row with 'End' in col A, or
    the first empty row after `start_row`.  Falls back to ws.max_row + 1."""
    for r in range(start_row + 1, ws.max_row + 2):
        v = ws.cell(r, 1).value
        if v is not None and str(v).strip() in cfg.END_MARKERS:
            return r           # insert BEFORE the End marker
        if v is None and all(ws.cell(r, c).value in (None, '') for c in range(1, min(ws.max_column, 10) + 1)):
            # First fully empty row → safe append point.
            return r
    return ws.max_row + 1


def _write_row(ws, row_idx: int, header_map: dict[str, int], data: dict):
    """Write a dict of {header_name: value} into `row_idx`."""
    for name, value in data.items():
        col = header_map.get(name)
        if col is None or value in (None, ''):
            continue
        ws.cell(row_idx, col).value = value


def _insert_blank_row(ws, at_row: int):
    """openpyxl insert_rows shifts everything down, preserving formats."""
    ws.insert_rows(at_row, amount=1)


# ─────────────────────────────────────────────────────────────────────────────
# SPEC writer
# ─────────────────────────────────────────────────────────────────────────────
def _write_spec_sheet(ws, classes: Iterable, builders: list):
    header_row = _find_header_row(ws)
    if header_row is None:
        logger.warning("[SmartPlantExport] no Head marker in sheet %s — skipped", ws.title)
        return 0

    header_map = _read_header_map(ws, header_row)
    start_row = _find_marker_row(ws, cfg.START_MARKERS, max_search=header_row + 10) or header_row + 1
    insert_at = _find_insert_row(ws, start_row)

    # Collect every row first → single batched insert (avoids O(N²) shifts).
    all_rows: list[dict] = []
    for cls in classes:
        for build_fn in builders:
            for row_dict in build_fn(cls):
                if row_dict:
                    all_rows.append(row_dict)
    if not all_rows:
        return 0

    ws.insert_rows(insert_at, amount=len(all_rows))
    for offset, row_dict in enumerate(all_rows):
        _write_row(ws, insert_at + offset, header_map, row_dict)
    return len(all_rows)


def build_spec_workbook(job) -> io.BytesIO:
    """Build the SPEC workbook for a completed extraction `job`.  Returns a
    BytesIO ready to stream as an HTTP attachment."""
    wb = load_workbook(cfg.SPEC_TEMPLATE_PATH)

    # Build the preview JSON (with overrides already merged in) and write
    # it sheet-by-sheet — this guarantees the canvas and the xlsx agree.
    preview_data = preview.build_preview(job, preview.WORKBOOK_SPEC)
    summary = {}
    for sheet in preview_data['sheets']:
        sheet_name = sheet['name']
        if sheet_name not in wb.sheetnames:
            logger.warning("[SmartPlantExport] SPEC template missing sheet %s", sheet_name)
            continue
        rows = [r['cells'] for r in sheet['rows']]
        n = _write_preview_rows_to_sheet(wb[sheet_name], rows)
        if n:
            summary[sheet_name] = n

    logger.info("[SmartPlantExport] SPEC job=%s wrote rows: %s", job.id, summary)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ─────────────────────────────────────────────────────────────────────────────
# CAT writer
# ─────────────────────────────────────────────────────────────────────────────
def _write_cat_sheet(ws, items: list[tuple], builder):
    """`items` is a list of (cls, component) pairs routed to this sheet."""
    header_row = _find_header_row(ws)
    if header_row is None:
        logger.warning("[SmartPlantExport] no Head marker in CAT sheet %s", ws.title)
        return 0
    header_map = _read_header_map(ws, header_row)
    start_row = _find_marker_row(ws, cfg.START_MARKERS, max_search=header_row + 5) or header_row + 1
    insert_at = _find_insert_row(ws, start_row)

    # Collect every row first → single batched insert (avoids O(N²) shifts).
    all_rows: list[dict] = []
    for cls, comp in items:
        result = builder(cls, comp)
        if not result:
            continue
        # Builders may return a single dict OR a list of dicts (per-NPD rows).
        rows = result if isinstance(result, list) else [result]
        for row_dict in rows:
            if row_dict:
                all_rows.append(row_dict)
    if not all_rows:
        return 0

    ws.insert_rows(insert_at, amount=len(all_rows))
    for offset, row_dict in enumerate(all_rows):
        _write_row(ws, insert_at + offset, header_map, row_dict)
    return len(all_rows)


def build_cat_workbook(job) -> io.BytesIO:
    """Build the CAT workbook for a completed extraction `job`."""
    wb = load_workbook(cfg.CAT_TEMPLATE_PATH)

    preview_data = preview.build_preview(job, preview.WORKBOOK_CAT)
    summary = {}
    for sheet in preview_data['sheets']:
        sheet_name = sheet['name']
        if sheet_name not in wb.sheetnames:
            logger.warning("[SmartPlantExport] CAT template missing sheet %s", sheet_name)
            continue
        rows = [r['cells'] for r in sheet['rows']]
        n = _write_preview_rows_to_sheet(wb[sheet_name], rows)
        if n:
            summary[sheet_name] = n

    logger.info("[SmartPlantExport] CAT job=%s wrote rows: %s", job.id, summary)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ─────────────────────────────────────────────────────────────────────────────
# Generic sheet writer (used by both SPEC + CAT after preview merges overrides)
# ─────────────────────────────────────────────────────────────────────────────
def _write_preview_rows_to_sheet(ws, rows: list[dict]) -> int:
    """Insert `rows` (each a {header_name: value} dict) into `ws`.

    Uses the same Head / Start / End conventions as the original per-workbook
    writers, but is workbook-agnostic.
    """
    if not rows:
        return 0
    header_row = _find_header_row(ws)
    if header_row is None:
        logger.warning("[SmartPlantExport] no Head marker in sheet %s — skipped", ws.title)
        return 0
    header_map = _read_header_map(ws, header_row)
    start_row = _find_marker_row(ws, cfg.START_MARKERS, max_search=header_row + 10) or header_row + 1
    insert_at = _find_insert_row(ws, start_row)

    ws.insert_rows(insert_at, amount=len(rows))
    for offset, row_dict in enumerate(rows):
        _write_row(ws, insert_at + offset, header_map, row_dict)
    return len(rows)
