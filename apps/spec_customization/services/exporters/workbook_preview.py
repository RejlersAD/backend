"""
Workbook Preview Builder
========================

Generates the SPEC/CAT workbook contents as **JSON** (instead of writing to
Excel), so the frontend can render an editable "cross-check" canvas that
matches the template layout exactly.

The same soft-coded builders used by `smartplant_exporter` are reused — so
the canvas and the downloaded xlsx always agree.

Outputs the shape:

    {
        "workbook":   "spec" | "cat",
        "job_id":     "<uuid>",
        "sheets": [
            {
                "name":    "PipingMaterialsClassData",
                "headers": ["Class", "MaterialGrade", ...],
                "rows": [
                    {
                        "row_key":   "cls:<uuid>:b0:0",
                        "cells":     {"Class": "A1", "MaterialGrade": "...", ...},
                        "overridden": ["Class"],     # cells edited by user
                        "source": {                  # provenance for UI hints
                            "class_id":     "<uuid>",
                            "class_code":   "A",
                            "component_id": "<uuid>" | null
                        }
                    },
                    ...
                ]
            },
            ...
        ]
    }

The `row_key` is deterministic and stable across rebuilds; it is the same
key persisted on `WorkbookCellOverride.row_key`.
"""
from __future__ import annotations

import logging
from typing import Iterable

from openpyxl import load_workbook

from . import smartplant_config as cfg

logger = logging.getLogger(__name__)


# ─── Soft-coded knobs ────────────────────────────────────────────────────────
WORKBOOK_SPEC = 'spec'
WORKBOOK_CAT  = 'cat'

# Limit how deep we scan the template for the Head row.
_HEADER_SCAN_DEPTH = 60


def _scan_header_rows(template_path: str, sheet_name: str) -> tuple[int | None, list[str]]:
    """Locate the header row and return its 1-indexed row number + values.

    Two layouts are supported:

      1. **Marker layout** (most SP3D bulkload sheets) — col A contains a
         token in ``cfg.HEADER_MARKERS`` (e.g. ``Head``). The LAST marker
         row within the first ~60 rows wins.
      2. **Filter layout** (e.g. ``PipingCommodityFilter``) — col A blank,
         headers begin at col B on row 2. Detected by scanning the first
         ``cfg.HEADER_FALLBACK_SCAN_ROWS`` rows for one where col A is
         empty and at least ``cfg.HEADER_FALLBACK_MIN_LABELS`` cells
         (starting col B) are non-empty strings.

    Returns ``(header_row, header_values)`` or ``(None, [])`` if neither
    layout matches.
    """
    try:
        wb = load_workbook(template_path, read_only=True, data_only=True)
    except Exception:
        logger.exception("[WorkbookPreview] could not open template %s", template_path)
        return None, []
    if sheet_name not in wb.sheetnames:
        return None, []
    ws = wb[sheet_name]

    fallback_depth = max(
        _HEADER_SCAN_DEPTH,
        getattr(cfg, 'HEADER_FALLBACK_SCAN_ROWS', 10),
    )
    # Pull just the first `fallback_depth` rows once — iter_rows is O(rows)
    # in read-only mode, whereas ws.cell() in read-only is O(rows²).
    initial_rows: list[tuple] = []
    for row in ws.iter_rows(min_row=1, max_row=fallback_depth, values_only=True):
        initial_rows.append(row)

    marker_row_idx: int | None = None
    for idx, row in enumerate(initial_rows[:_HEADER_SCAN_DEPTH], start=1):
        if not row:
            continue
        v = row[0]
        if v is not None and str(v).strip() in cfg.HEADER_MARKERS:
            marker_row_idx = idx

    chosen_row_idx: int | None = marker_row_idx
    if chosen_row_idx is None:
        scan_n = getattr(cfg, 'HEADER_FALLBACK_SCAN_ROWS', 10)
        min_lbl = getattr(cfg, 'HEADER_FALLBACK_MIN_LABELS', 3)
        for idx, row in enumerate(initial_rows[:scan_n], start=1):
            if not row:
                continue
            if row[0] not in (None, ''):
                continue
            label_count = sum(
                1 for v in row[1:]
                if isinstance(v, str) and v.strip()
            )
            if label_count >= min_lbl:
                chosen_row_idx = idx
                break

    if chosen_row_idx is None or chosen_row_idx > len(initial_rows):
        return None, []

    header_row = initial_rows[chosen_row_idx - 1]
    headers = [
        (str(v).strip() if v is not None else '')
        for v in header_row
    ]
    while headers and not headers[-1]:
        headers.pop()
    return chosen_row_idx, headers


def _template_headers(template_path: str, sheet_name: str) -> list[str]:
    """Public-style helper (cached via `_get_headers`)."""
    return _scan_header_rows(template_path, sheet_name)[1]


# Module-level cache: template_path → {sheet_name → [headers]}
_HEADERS_CACHE: dict[str, dict[str, list[str]]] = {}


def _get_headers(template_path: str, sheet_name: str) -> list[str]:
    by_sheet = _HEADERS_CACHE.setdefault(template_path, {})
    if sheet_name not in by_sheet:
        by_sheet[sheet_name] = _template_headers(template_path, sheet_name)
    return by_sheet[sheet_name]


def _template_passthrough_rows(
    template_path: str,
    sheet_name: str,
    headers: list[str],
) -> list[dict]:
    """Read static data rows below the header row from the template.

    Used for sheets that have no builder registered (e.g. SP3D static
    reference sheets like ``PipingCommodityFilter``) so the canvas reflects
    the LS1E reference content.  Returns rows in preview shape with stable
    ``row_key``s of the form ``tpl:{sheet}:{row_idx}``.
    """
    if not headers:
        return []
    if not getattr(cfg, 'TEMPLATE_PASSTHROUGH_ENABLED', True):
        return []
    header_row_idx, _ = _scan_header_rows(template_path, sheet_name)
    if header_row_idx is None:
        return []
    try:
        wb = load_workbook(template_path, read_only=True, data_only=True)
    except Exception:
        return []
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    max_rows = getattr(cfg, 'TEMPLATE_PASSTHROUGH_MAX_ROWS', 5000)
    end_markers = set(cfg.END_MARKERS)
    n_cols = len(headers)

    out: list[dict] = []
    # iter_rows is O(rows) and the only efficient access pattern in
    # read-only mode.  We slice it to the first N columns we actually need.
    for row_offset, row in enumerate(
        ws.iter_rows(min_row=header_row_idx + 1, max_col=n_cols, values_only=True),
        start=1,
    ):
        if len(out) >= max_rows:
            break
        # Stop at SP3D End sentinel.
        col_a = row[0] if row else None
        if col_a is not None and str(col_a).strip() in end_markers:
            break
        if not row or all(v in (None, '') for v in row):
            continue
        cells: dict = {}
        for idx, header in enumerate(headers):
            v = row[idx] if idx < len(row) else None
            key = header or f'_col{idx}'
            if v is None:
                cells[key] = ''
            elif isinstance(v, str):
                cells[key] = v
            else:
                cells[key] = str(v)
        # Soft-coded blank-cell enrichment (SP3D-valid defaults).  Safe no-op
        # for sheets without a defaults entry in smartplant_config.
        try:
            cfg.apply_passthrough_defaults(sheet_name, cells)
        except Exception:
            logger.exception("[WorkbookPreview] apply_passthrough_defaults failed for %s",
                             sheet_name)
        out.append({
            'row_key':    f'tpl:{sheet_name}:{header_row_idx + row_offset}',
            'cells':      cells,
            'overridden': [],
            'source':     {'class_id': None, 'class_code': None, 'component_id': None},
        })
    return out


# ─── Row enumeration (reuses exporter builders) ──────────────────────────────
def _enumerate_spec_rows(job):
    """Yield (sheet_name, row_key, source_dict, cells_dict) tuples for SPEC."""
    classes = list(job.piping_classes.prefetch_related('components').order_by('class_code'))
    for sheet_name, builders in cfg.SPEC_SHEET_BUILDERS.items():
        for cls in classes:
            for b_idx, builder in enumerate(builders):
                seq = 0
                for row in builder(cls):
                    if not row:
                        continue
                    row_key = f'cls:{cls.id}:b{b_idx}:{seq}'
                    source = {
                        'class_id':     str(cls.id),
                        'class_code':   cls.class_code,
                        'component_id': None,
                    }
                    yield sheet_name, row_key, source, dict(row)
                    seq += 1


def _enumerate_cat_rows(job):
    """Yield (sheet_name, row_key, source_dict, cells_dict) tuples for CAT."""
    classes = list(job.piping_classes.prefetch_related('components').order_by('class_code'))
    # Route components → sheets exactly like the exporter does.
    by_sheet: dict[str, list] = {}
    for cls in classes:
        for comp in cls.components.all().order_by('display_order'):
            sheet = cfg.route_component_to_cat_sheet(comp)
            if sheet is None:
                continue
            by_sheet.setdefault(sheet, []).append((cls, comp))

    for sheet_name, items in by_sheet.items():
        builder = cfg.CAT_SHEET_BUILDERS.get(sheet_name)
        if builder is None:
            continue
        for cls, comp in items:
            result = builder(cls, comp)
            if not result:
                continue
            rows = result if isinstance(result, list) else [result]
            for seq, row in enumerate(rows):
                if not row:
                    continue
                row_key = f'comp:{comp.id}:{sheet_name}:{seq}'
                source = {
                    'class_id':     str(cls.id),
                    'class_code':   cls.class_code,
                    'component_id': str(comp.id),
                }
                yield sheet_name, row_key, source, dict(row)


# ─── Override application ────────────────────────────────────────────────────
def _load_overrides(job, workbook: str) -> dict[tuple[str, str], dict[str, str]]:
    """Return overrides keyed by (sheet_name, row_key) → {column_name: value}."""
    from ...models import WorkbookCellOverride  # local import to dodge cycles
    out: dict[tuple[str, str], dict[str, str]] = {}
    qs = WorkbookCellOverride.objects.filter(job=job, workbook=workbook)
    for ov in qs:
        out.setdefault((ov.sheet_name, ov.row_key), {})[ov.column_name] = ov.value
    return out


# ─── Public API ──────────────────────────────────────────────────────────────
def build_preview(job, workbook: str) -> dict:
    """Build the JSON preview for the given workbook ('spec' or 'cat').

    Applies any saved `WorkbookCellOverride` entries on top of the
    builder-generated cells.
    """
    if workbook not in (WORKBOOK_SPEC, WORKBOOK_CAT):
        raise ValueError(f"unknown workbook: {workbook!r}")

    template_path = cfg.SPEC_TEMPLATE_PATH if workbook == WORKBOOK_SPEC else cfg.CAT_TEMPLATE_PATH
    enumerate_fn  = _enumerate_spec_rows if workbook == WORKBOOK_SPEC else _enumerate_cat_rows
    overrides     = _load_overrides(job, workbook)

    # Bucket rows by sheet, preserving order.
    sheets_order: list[str] = []
    rows_by_sheet: dict[str, list[dict]] = {}

    for sheet_name, row_key, source, cells in enumerate_fn(job):
        if sheet_name not in rows_by_sheet:
            sheets_order.append(sheet_name)
            rows_by_sheet[sheet_name] = []
        sheet_overrides = overrides.get((sheet_name, row_key), {})
        # Apply overrides (string-typed; UI sends strings).
        merged = {k: ('' if v is None else v) for k, v in cells.items()}
        for k, v in sheet_overrides.items():
            merged[k] = v
        rows_by_sheet[sheet_name].append({
            'row_key':    row_key,
            'cells':      merged,
            'overridden': sorted(sheet_overrides.keys()),
            'source':     source,
        })

    # Build the final shape with template-driven column order.
    sheets_out = []
    for sheet_name in sheets_order:
        headers = _get_headers(template_path, sheet_name)
        # If the template scan failed, derive headers from the first row.
        if not headers and rows_by_sheet[sheet_name]:
            headers = list(rows_by_sheet[sheet_name][0]['cells'].keys())
        sheets_out.append({
            'name':       sheet_name,
            'headers':    headers,
            'row_count':  len(rows_by_sheet[sheet_name]),
            'rows':       rows_by_sheet[sheet_name],
        })

    # Surface every template sheet in the canvas — even sheets with no
    # auto-generated rows (e.g. `CustomInterfaces`, `PipingCommodityFilter`).
    # This keeps the UI aligned with the LS1E-A3 reference workbook so the
    # user can see the full bulkload schema, not just the populated subset.
    try:
        _wb = load_workbook(template_path, read_only=True, data_only=True)
        template_sheet_names = list(_wb.sheetnames)
    except Exception:
        logger.exception("[WorkbookPreview] could not enumerate template sheets at %s",
                         template_path)
        template_sheet_names = []

    rendered = {s['name'] for s in sheets_out}
    for sheet_name in template_sheet_names:
        if sheet_name in rendered:
            continue
        headers = _get_headers(template_path, sheet_name)
        if not headers:
            # Skip sheets whose Head row could not be detected.
            continue
        # Surface template-shipped static data (e.g. PipingCommodityFilter's
        # 700+ LS1E reference rows) and apply any saved cell overrides on top.
        passthrough = _template_passthrough_rows(template_path, sheet_name, headers)
        for row in passthrough:
            sheet_overrides = overrides.get((sheet_name, row['row_key']), {})
            if sheet_overrides:
                row['cells'].update(sheet_overrides)
                row['overridden'] = sorted(sheet_overrides.keys())
        sheets_out.append({
            'name':       sheet_name,
            'headers':    headers,
            'row_count':  len(passthrough),
            'rows':       passthrough,
        })

    return {
        'workbook':  workbook,
        'job_id':    str(job.id),
        'sheets':    sheets_out,
    }


def apply_overrides_to_row(
    sheet_name: str,
    row_key: str,
    cells: dict,
    overrides: dict[tuple[str, str], dict[str, str]],
) -> dict:
    """Merge an override slice into a generated cells dict.

    Helper for the xlsx exporter so the same overrides also flow into the
    downloadable workbook.
    """
    ov = overrides.get((sheet_name, row_key))
    if not ov:
        return cells
    merged = dict(cells)
    for k, v in ov.items():
        merged[k] = v
    return merged
