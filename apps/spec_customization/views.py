"""
Spec Customization — DRF Views.

Endpoints (mounted at /api/v1/spec-customization/):

  POST   paper-spec/upload/                 → upload PDF + queue job
  GET    paper-spec/jobs/                   → list jobs
  GET    paper-spec/jobs/<id>/              → job detail + live progress
  GET    paper-spec/jobs/<id>/classes/      → extracted classes (paginated)
  GET    paper-spec/jobs/<id>/export/       → xlsx / json export
  POST   paper-spec/jobs/<id>/cancel/       → mark job cancelled
  GET    paper-spec/classes/<id>/           → class detail + components
"""
from __future__ import annotations

import hashlib
import io
import json
import logging
from typing import Any, Dict

from django.conf import settings
from django.core.cache import cache
from django.db.models import Count
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from rest_framework import status
from rest_framework.decorators import api_view, parser_classes, permission_classes
from rest_framework.parsers import FormParser, MultiPartParser
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

from .models import (
    PaperSpecDocument,
    PaperSpecExtractionJob,
    PipingClass,
    WorkbookCellOverride,
)
from .serializers import (
    PaperSpecDocumentSerializer,
    PaperSpecExtractionJobSerializer,
    PipingClassSerializer,
    PipingClassListSerializer,
)
from .services.config import (
    SPEC_EXTRACTION_CONFIG,
    PROGRESS_CACHE_KEY_TPL,
    PARTIAL_CACHE_KEY_TPL,
)
from .services.file_normalizer import (
    SUPPORTED_FORMATS,
    get_accepted_extensions,
    normalize_to_pdf,
)
from .services.exporters import build_spec_workbook, build_cat_workbook
from .services.exporters.workbook_preview import build_preview, WORKBOOK_SPEC, WORKBOOK_CAT
from .services.exporters.smartplant_config import (
    SPEC_OUTPUT_FILENAME_TPL,
    CAT_OUTPUT_FILENAME_TPL,
)
from .services.presigned_upload import (
    PRESIGNED_UPLOAD_CONFIG,
    best_effort_delete,
    fetch_uploaded_to_temp_file,
    generate_presigned_put,
    is_presigned_upload_available,
)
from .tasks import extract_paper_spec

logger = logging.getLogger(__name__)


def _sha256_of_file(django_file) -> str:
    h = hashlib.sha256()
    for chunk in django_file.chunks():
        h.update(chunk)
    django_file.seek(0)
    return h.hexdigest()


def _page_count(django_file_path: str) -> int:
    try:
        import fitz
        with fitz.open(django_file_path) as doc:
            return doc.page_count
    except Exception:
        try:
            import PyPDF2
            with open(django_file_path, 'rb') as f:
                return len(PyPDF2.PdfReader(f).pages)
        except Exception:
            return 0


def _model_has_field(model_class, field_name: str) -> bool:
    """
    Soft-coded helper: Check if a Django model has a specific field.
    
    Used for backward compatibility during migrations - allows code to gracefully
    handle cases where schema changes haven't been applied yet (e.g., during
    Railway deployment when migrations might not run immediately).
    
    Returns:
        True if field exists, False otherwise
    """
    try:
        model_class._meta.get_field(field_name)
        return True
    except Exception:
        return False


# ─────────────────────────────────────────────────────────────────────────────
# Shared ingest pipeline — used by both legacy multipart upload AND the new
# direct-to-S3 presigned-upload "complete" endpoint. Centralised so both paths
# stay byte-for-byte equivalent (SHA dedupe, normalisation, Celery enqueue).
# ─────────────────────────────────────────────────────────────────────────────
def _ingest_paper_spec_file(
    *,
    src,                      # Django UploadedFile-like (.name, .size, .read, .seek, .chunks)
    project_id,
    title: str,
    document_number: str,
    engineer_name: str,       # BYOK attribution field
    user_api_key: str,        # BYOK user-supplied OpenAI API key (optional)
    request_user,
):
    """Returns a DRF ``Response`` — same shape as the legacy upload view."""
    sha = _sha256_of_file(src)
    original_filename = getattr(src, 'name', 'upload')
    original_size = getattr(src, 'size', 0) or 0

    if SPEC_EXTRACTION_CONFIG.get("dedupe_by_sha256", True):
        existing = (
            PaperSpecDocument.objects
            .filter(sha256_hash=sha)
            .order_by('-created_at')
            .first()
        )
        if existing:
            latest_completed = (
                existing.jobs
                .filter(status=PaperSpecExtractionJob.STATUS_COMPLETED)
                .order_by('-completed_at')
                .first()
            )
            if latest_completed:
                return Response({
                    "deduped": True,
                    "document": PaperSpecDocumentSerializer(existing).data,
                    "job":      PaperSpecExtractionJobSerializer(latest_completed).data,
                })

    normalized = normalize_to_pdf(src)
    if not normalized.success:
        return Response(
            {"error": normalized.error_message,
             "accepted_extensions": get_accepted_extensions()},
            status=status.HTTP_400_BAD_REQUEST,
        )

    doc = PaperSpecDocument.objects.create(
        file=normalized.file,
        original_filename=original_filename,
        file_size_bytes=original_size,
        sha256_hash=sha,
        project_id=project_id,
        title=title,
        document_number=document_number,
        uploaded_by=request_user if getattr(request_user, 'is_authenticated', False) else None,
    )
    doc.total_pages = _page_count(doc.file.path)
    doc.save(update_fields=['total_pages'])

    # BACKWARD COMPATIBILITY: Build job creation kwargs conditionally
    # Only set BYOK fields (migration 0006) if they exist in the schema
    job_kwargs = {
        'document': doc,
        'status': PaperSpecExtractionJob.STATUS_QUEUED,
        'created_by': request_user if getattr(request_user, 'is_authenticated', False) else None,
    }
    
    # Add BYOK fields only if migration 0006 has been applied (soft-coded check)
    if _model_has_field(PaperSpecExtractionJob, 'engineer_name'):
        job_kwargs['engineer_name'] = engineer_name
    if _model_has_field(PaperSpecExtractionJob, 'user_openai_api_key'):
        job_kwargs['user_openai_api_key'] = user_api_key  # Stored temporarily; wiped after extraction completes
    
    job = PaperSpecExtractionJob.objects.create(**job_kwargs)

    # SOFT-CODED: Check if Celery is in EAGER mode to avoid connection errors
    # In EAGER mode, tasks run synchronously and don't need a broker connection.
    # Calling .delay() in EAGER mode can still try to connect to broker (Celery quirk),
    # causing "Connection refused" errors that crash the endpoint before CORS headers
    # can be sent. Fix: call task directly in EAGER mode, use .delay() only when
    # broker is available.
    celery_eager_mode = getattr(settings, 'CELERY_TASK_ALWAYS_EAGER', False)
    
    if celery_eager_mode:
        # EAGER mode: Run task synchronously without broker connection
        logger.info("[SpecExtraction] Running task in EAGER mode (synchronous)")
        try:
            extract_paper_spec(str(job.id))
        except Exception as e:
            logger.exception("[SpecExtraction] Synchronous task failed: %s", e)
            job.status = PaperSpecExtractionJob.STATUS_FAILED
            job.error_message = f"Extraction failed: {e}"
            job.save(update_fields=['status', 'error_message'])
    else:
        # Normal mode: Queue task asynchronously via Celery
        try:
            async_result = extract_paper_spec.delay(str(job.id))
            job.celery_task_id = async_result.id or ''
            job.save(update_fields=['celery_task_id'])
        except Exception as e:
            logger.exception("[SpecExtraction] failed to queue task: %s", e)
            # Fallback to synchronous execution
            try:
                extract_paper_spec(str(job.id))
            except Exception as run_err:
                job.status = PaperSpecExtractionJob.STATUS_FAILED
                job.error_message = f"Queue failed: {e}; inline failed: {run_err}"
                job.save(update_fields=['status', 'error_message'])

    return Response({
        "deduped": False,
        "document": PaperSpecDocumentSerializer(doc).data,
        "job":      PaperSpecExtractionJobSerializer(job).data,
    }, status=status.HTTP_201_CREATED)


# ─────────────────────────────────────────────────────────────────────────────
# Upload + start extraction
# ─────────────────────────────────────────────────────────────────────────────
@api_view(['POST'])
@parser_classes([MultiPartParser, FormParser])
@permission_classes([IsAuthenticated])
def upload_paper_spec(request):
    src = request.FILES.get('file') or request.FILES.get('pdf_file')
    if not src:
        return Response({"error": "No file uploaded (expected field: 'file' or 'pdf_file')"},
                        status=status.HTTP_400_BAD_REQUEST)

    return _ingest_paper_spec_file(
        src=src,
        project_id=request.data.get('project_id') or None,
        title=request.data.get('title') or '',
        document_number=request.data.get('document_number') or '',
        engineer_name=request.data.get('engineer_name') or '',          # BYOK attribution
        user_api_key=request.data.get('user_openai_api_key') or '',     # BYOK user API key
        request_user=request.user,
    )


# ─────────────────────────────────────────────────────────────────────────────
# Direct-to-S3 presigned upload (bypasses Railway's edge body-size limit).
# Two-step flow:
#   1) POST paper-spec/upload/presign/   → returns presigned PUT URL + s3_key
#   2) Browser PUTs file directly to S3 (with progress)
#   3) POST paper-spec/upload/complete/  → backend ingests via the SAME helper
# Falls back gracefully: if S3 isn't configured or the feature flag is off,
# `presign` returns `{"enabled": false, "reason": "..."}` and the client should
# revert to the legacy multipart `upload/` endpoint.
# ─────────────────────────────────────────────────────────────────────────────
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def presign_paper_spec_upload(request):
    filename     = request.data.get('filename') or ''
    content_type = request.data.get('content_type') or 'application/octet-stream'
    try:
        size_bytes = int(request.data.get('size') or 0)
    except (TypeError, ValueError):
        size_bytes = 0

    if not filename:
        return Response({"error": "filename is required"},
                        status=status.HTTP_400_BAD_REQUEST)

    result = generate_presigned_put(
        filename=filename,
        content_type=content_type,
        size_bytes=size_bytes,
    )

    if not result.enabled:
        # 200 OK with enabled=False — this is an *expected* fallback signal,
        # not an error. Lets the frontend silently revert to multipart upload.
        return Response({
            "enabled": False,
            "reason":  result.reason,
        })

    return Response({
        "enabled":     True,
        "method":      result.method,
        "upload_url":  result.upload_url,
        "s3_key":      result.s3_key,
        "headers":     result.headers,
        "expires_in":  result.expires_in,
        "bucket":      result.bucket,
        "max_bytes":   PRESIGNED_UPLOAD_CONFIG['max_bytes'],
    })


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def complete_paper_spec_upload(request):
    """Ingest a file that was already PUT to S3 via a presigned URL."""
    s3_key            = request.data.get('s3_key') or ''
    original_filename = request.data.get('filename') or ''
    project_id        = request.data.get('project_id') or None
    title             = request.data.get('title') or ''
    document_number   = request.data.get('document_number') or ''
    engineer_name     = request.data.get('engineer_name') or ''          # BYOK attribution
    user_api_key      = request.data.get('user_openai_api_key') or ''   # BYOK user API key

    if not s3_key or not original_filename:
        return Response(
            {"error": "s3_key and filename are required"},
            status=status.HTTP_400_BAD_REQUEST,
        )

    if not is_presigned_upload_available():
        return Response(
            {"error": "presigned uploads are not available in this environment"},
            status=status.HTTP_503_SERVICE_UNAVAILABLE,
        )

    try:
        src = fetch_uploaded_to_temp_file(
            s3_key=s3_key,
            original_filename=original_filename,
        )
    except RuntimeError as exc:
        return Response({"error": str(exc)}, status=status.HTTP_400_BAD_REQUEST)

    try:
        response = _ingest_paper_spec_file(
            src=src,
            project_id=project_id,
            title=title,
            document_number=document_number,
            engineer_name=engineer_name,      # BYOK attribution
            user_api_key=user_api_key,        # BYOK user API key
            request_user=request.user,
        )
    finally:
        try:
            src.close()
        except Exception:
            pass
        # Stage object served its purpose — drop it. Failure is non-fatal.
        best_effort_delete(s3_key)

    return response


# ─────────────────────────────────────────────────────────────────────────────
# Job listing / detail
# ─────────────────────────────────────────────────────────────────────────────
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def list_jobs(request):
    qs = PaperSpecExtractionJob.objects.all().order_by('-created_at')[:50]
    return Response(PaperSpecExtractionJobSerializer(qs, many=True).data)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def job_detail(request, job_id):
    job = get_object_or_404(PaperSpecExtractionJob, pk=job_id)
    data = PaperSpecExtractionJobSerializer(job).data
    # Add live progress + partial classes from cache.
    progress = cache.get(PROGRESS_CACHE_KEY_TPL.format(job_id=str(job.id))) or {}
    partial = cache.get(PARTIAL_CACHE_KEY_TPL.format(job_id=str(job.id))) or []
    data['live_progress'] = progress
    data['partial_classes_preview'] = [
        {
            'class_code':      c.get('class_code', ''),
            'class_full_code': c.get('class_full_code', ''),
            'engine':          c.get('_engine', ''),
            'components':      len(c.get('components', []) or []),
        }
        for c in partial[:50]
    ]
    return Response(data)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def cancel_job(request, job_id):
    job = get_object_or_404(PaperSpecExtractionJob, pk=job_id)
    if job.status in (PaperSpecExtractionJob.STATUS_COMPLETED,
                      PaperSpecExtractionJob.STATUS_FAILED,
                      PaperSpecExtractionJob.STATUS_CANCELLED):
        return Response({"ok": False, "message": f"Job already {job.status}"},
                        status=status.HTTP_400_BAD_REQUEST)
    job.status = PaperSpecExtractionJob.STATUS_CANCELLED
    job.save(update_fields=['status'])
    return Response({"ok": True, "status": job.status})


# ─────────────────────────────────────────────────────────────────────────────
# Classes
# ─────────────────────────────────────────────────────────────────────────────
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def job_classes(request, job_id):
    job = get_object_or_404(PaperSpecExtractionJob, pk=job_id)
    qs = (job.piping_classes
              .annotate(components_count=Count('components'))
              .order_by('class_code'))
    return Response(PipingClassListSerializer(qs, many=True).data)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def class_detail(request, class_id):
    cls = get_object_or_404(PipingClass, pk=class_id)
    return Response(PipingClassSerializer(cls).data)


# ─────────────────────────────────────────────────────────────────────────────
# Export
# ─────────────────────────────────────────────────────────────────────────────
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_job(request, job_id):
    job = get_object_or_404(PaperSpecExtractionJob, pk=job_id)
    fmt = (request.query_params.get('format') or 'json').lower()

    classes = list(
        job.piping_classes.prefetch_related('components').order_by('class_code')
    )

    if fmt == 'xlsx':
        try:
            from openpyxl import Workbook
        except ImportError:
            return Response({"error": "openpyxl not installed on backend"},
                            status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        wb = Workbook()
        wb.remove(wb.active)
        for cls in classes:
            sheet_name = (cls.class_code or 'CLASS')[:31] or 'CLASS'
            ws = wb.create_sheet(title=sheet_name)
            ws.append(['Field', 'Value'])
            ws.append(['Class Code', cls.class_code])
            ws.append(['Full Header', cls.class_full_code])
            ws.append(['Material', cls.material_grade])
            ws.append(['Rating', cls.pressure_rating])
            ws.append(['Facing', cls.flange_facing])
            ws.append(['Corrosion Allowance', cls.corrosion_allowance])
            ws.append(['Services', '; '.join(cls.service_list or [])])
            ws.append([])
            ws.append(['P/T Rating Table'])
            ws.append(['Pressure (bar-g)', 'Temperature (°C)', 'Notes'])
            for r in (cls.pt_rating_table or []):
                ws.append([r.get('pressure_bar_g'), r.get('temperature_c'), r.get('notes', '')])
            ws.append([])
            ws.append(['Components'])
            ws.append(['Type', 'Sub-Type', 'Size From', 'Size To',
                       'Schedule/Rating', 'Material Std', 'End Conn.', 'Description', 'Notes'])
            for c in cls.components.all().order_by('display_order'):
                ws.append([
                    c.component_type, c.sub_type, c.size_from, c.size_to,
                    c.schedule_or_rating, c.material_standard, c.end_connection,
                    c.description, c.notes,
                ])
        if not classes:
            ws = wb.create_sheet(title='Empty')
            ws.append(['No piping classes extracted'])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        resp = HttpResponse(
            buf.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        fname = f"paper_spec_{job.id}.xlsx"
        resp['Content-Disposition'] = f'attachment; filename="{fname}"'
        return resp

    # JSON default
    payload: Dict[str, Any] = {
        "job_id": str(job.id),
        "status": job.status,
        "document": job.document.original_filename,
        "classes": [PipingClassSerializer(c).data for c in classes],
    }
    resp = HttpResponse(json.dumps(payload, default=str), content_type='application/json')
    resp['Content-Disposition'] = f'attachment; filename="paper_spec_{job.id}.json"'
    return resp


# ─────────────────────────────────────────────────────────────────────────────
# SmartPlant 3D — Spec / Catalog exports (two-file output)
# ─────────────────────────────────────────────────────────────────────────────
def _smartplant_response(buf, filename: str) -> HttpResponse:
    resp = HttpResponse(
        buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    resp['Content-Disposition'] = f'attachment; filename="{filename}"'
    return resp


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_smartplant_spec(request, job_id):
    """Export SmartPlant 3D SPEC workbook (rule sheets)."""
    job = get_object_or_404(PaperSpecExtractionJob, pk=job_id)
    try:
        buf = build_spec_workbook(job)
    except Exception as e:
        logger.exception("[SmartPlantExport] SPEC build failed for job %s", job_id)
        return Response({"error": f"Failed to build SPEC workbook: {e}"},
                        status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    return _smartplant_response(buf, SPEC_OUTPUT_FILENAME_TPL.format(job_id=job.id))


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_smartplant_cat(request, job_id):
    """Export SmartPlant 3D Catalog workbook (component part sheets)."""
    job = get_object_or_404(PaperSpecExtractionJob, pk=job_id)
    try:
        buf = build_cat_workbook(job)
    except Exception as e:
        logger.exception("[SmartPlantExport] CAT build failed for job %s", job_id)
        return Response({"error": f"Failed to build CAT workbook: {e}"},
                        status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    return _smartplant_response(buf, CAT_OUTPUT_FILENAME_TPL.format(job_id=job.id))


# ─────────────────────────────────────────────────────────────────────────────
# Health / config introspection (debug-friendly)
# ─────────────────────────────────────────────────────────────────────────────
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def config_view(request):
    """Expose soft-coded config to the UI so feature flags stay in sync."""
    accepted_exts = get_accepted_extensions()
    return Response({
        "chunk_size_pages":          SPEC_EXTRACTION_CONFIG["chunk_size_pages"],
        "ai_engines":                SPEC_EXTRACTION_CONFIG["ai_engines"],
        "max_ai_pages_per_job":      SPEC_EXTRACTION_CONFIG["max_ai_pages_per_job"],
        "skip_ai_if_text_chars_gte": SPEC_EXTRACTION_CONFIG["skip_ai_if_text_chars_gte"],
        "dedupe_by_sha256":          SPEC_EXTRACTION_CONFIG["dedupe_by_sha256"],
        "accepted_extensions":       accepted_exts,
        "accept_attribute":          ','.join(f'.{e}' for e in accepted_exts),
        "format_groups": sorted({
            f"{desc['group']}" for desc in SUPPORTED_FORMATS.values()
        }),
        # Soft-coded presigned-upload signalling — frontend can ask once at
        # boot and choose direct-to-S3 vs legacy multipart accordingly.
        "presigned_upload": {
            "available":           is_presigned_upload_available(),
            "max_bytes":           PRESIGNED_UPLOAD_CONFIG["max_bytes"],
            "url_expiry_seconds":  PRESIGNED_UPLOAD_CONFIG["url_expiry_seconds"],
        },
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def extraction_config_diagnostics(request):
    """
    Diagnostic endpoint to compare extraction configuration between environments.
    
    Helps diagnose why local vs production produce different extraction results
    (e.g., 1858 vs 116 records in PipingCommodityFilter sheet).
    
    Returns:
        - Chunking settings
        - AI engine configuration
        - Ensemble extraction status
        - Validation layer status
        - API key availability (without exposing actual keys)
    
    Example:
        GET /api/v1/spec-customization/diagnostics/config/
    """
    from apps.spec_customization.services.advanced_validation import ADVANCED_VALIDATION_CONFIG
    
    # Check API key availability without exposing them
    openai_key = os.getenv('OPENAI_API_KEY', '')
    gemini_key = os.getenv('GOOGLE_API_KEY', '') or os.getenv('GEMINI_API_KEY', '')
    
    config_data = {
        "environment": {
            "name": os.getenv('ENVIRONMENT', 'development'),
            "debug": os.getenv('DEBUG', 'True') == 'True',
            "aiflow_env": os.getenv('AIFLOW_ENVIRONMENT', 'unknown'),
        },
        
        "extraction_config": {
            "chunking": {
                "chunk_size_pages": SPEC_EXTRACTION_CONFIG['chunk_size_pages'],
                "page_overlap": SPEC_EXTRACTION_CONFIG['page_overlap'],
                "max_chunks_parallel": SPEC_EXTRACTION_CONFIG['max_chunks_parallel'],
            },
            "ai_engines": {
                "engines_list": SPEC_EXTRACTION_CONFIG['ai_engines'],
                "gemini_model": SPEC_EXTRACTION_CONFIG['gemini_model'],
                "openai_model": SPEC_EXTRACTION_CONFIG['openai_model'],
                "openai_max_tokens": SPEC_EXTRACTION_CONFIG['openai_max_tokens'],
                "temperature": {
                    "gemini": SPEC_EXTRACTION_CONFIG['gemini_temperature'],
                    "openai": SPEC_EXTRACTION_CONFIG['openai_temperature'],
                }
            },
            "cost_guard_rails": {
                "skip_ai_if_text_chars_gte": SPEC_EXTRACTION_CONFIG['skip_ai_if_text_chars_gte'],
                "max_ai_pages_per_job": SPEC_EXTRACTION_CONFIG['max_ai_pages_per_job'],
            }
        },
        
        "advanced_validation": {
            "ensemble_extraction": {
                "enabled": ADVANCED_VALIDATION_CONFIG.get('enable_ensemble_extraction', False),
                "consensus_threshold": ADVANCED_VALIDATION_CONFIG.get('ensemble_consensus_threshold', 0.7),
                "voting_strategy": ADVANCED_VALIDATION_CONFIG.get('ensemble_voting_strategy', 'weighted'),
                "status": (
                    "✅ ACTIVE - Gemini + OpenAI parallel processing" 
                    if ADVANCED_VALIDATION_CONFIG.get('enable_ensemble_extraction', False)
                    else "⚠️ DISABLED - Waterfall mode only"
                ),
            },
            "validation_layers": {
                "component_count": ADVANCED_VALIDATION_CONFIG.get('enable_component_count_validation', False),
                "material_standards": ADVANCED_VALIDATION_CONFIG.get('enable_material_standard_validation', False),
                "size_ranges": ADVANCED_VALIDATION_CONFIG.get('enable_size_range_validation', False),
                "template_comparison": ADVANCED_VALIDATION_CONFIG.get('enable_template_comparison', False),
                "auto_retry": ADVANCED_VALIDATION_CONFIG.get('enable_auto_retry', False),
            },
            "thresholds": {
                "min_components_per_class": ADVANCED_VALIDATION_CONFIG.get('min_components_per_class', 10),
                "warn_if_components_below": ADVANCED_VALIDATION_CONFIG.get('warn_if_components_below', 30),
            }
        },
        
        "api_keys_status": {
            "openai": {
                "configured": bool(openai_key),
                "key_prefix": openai_key[:10] + "..." if openai_key else None,
                "key_length": len(openai_key) if openai_key else 0,
            },
            "gemini": {
                "configured": bool(gemini_key),
                "key_prefix": gemini_key[:10] + "..." if gemini_key else None,
                "key_length": len(gemini_key) if gemini_key else 0,
            },
            "warning": None if (openai_key and gemini_key) else (
                "⚠️ Missing API keys - extraction may fail or use fallback only"
            )
        },
        
        "expected_behavior": {
            "mode": (
                "Multi-model ensemble (Gemini + OpenAI parallel)"
                if ADVANCED_VALIDATION_CONFIG.get('enable_ensemble_extraction', False)
                else "Waterfall (sequential fallback)"
            ),
            "expected_accuracy": (
                "95-98%"
                if ADVANCED_VALIDATION_CONFIG.get('enable_ensemble_extraction', False)
                else "85-92%"
            ),
            "expected_components_per_class": "50-200+",
            "warnings": []
        }
    }
    
    # Add warnings
    if SPEC_EXTRACTION_CONFIG['chunk_size_pages'] < 8:
        config_data["expected_behavior"]["warnings"].append(
            f"⚠️ Chunk size ({SPEC_EXTRACTION_CONFIG['chunk_size_pages']} pages) < 8 may fragment component tables"
        )
    
    if SPEC_EXTRACTION_CONFIG['openai_max_tokens'] < 12000:
        config_data["expected_behavior"]["warnings"].append(
            f"⚠️ Max tokens ({SPEC_EXTRACTION_CONFIG['openai_max_tokens']}) < 12000 may truncate large component lists"
        )
    
    if not (openai_key and gemini_key):
        config_data["expected_behavior"]["warnings"].append(
            "❌ Missing API keys - extraction will fail or produce incomplete results"
        )
    
    return Response(config_data)


# ─────────────────────────────────────────────────────────────────────────────
# Workbook canvas — preview SPEC/CAT contents as JSON + per-cell edit/clear
# ─────────────────────────────────────────────────────────────────────────────
_VALID_WORKBOOKS = {WORKBOOK_SPEC, WORKBOOK_CAT}
# Soft-coded value-length guard for free-text cell edits.
_MAX_CELL_VALUE_LEN = 2000


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def workbook_preview(request, job_id):
    """Return the SPEC or CAT workbook as JSON for the canvas.

    Query: ?workbook=spec|cat (defaults to spec).
    """
    workbook = (request.query_params.get('workbook') or WORKBOOK_SPEC).lower()
    if workbook not in _VALID_WORKBOOKS:
        return Response(
            {"error": f"workbook must be one of {sorted(_VALID_WORKBOOKS)}"},
            status=status.HTTP_400_BAD_REQUEST,
        )
    job = get_object_or_404(PaperSpecExtractionJob, pk=job_id)
    try:
        data = build_preview(job, workbook)
    except Exception as e:
        logger.exception("[WorkbookPreview] build failed for job %s / %s", job_id, workbook)
        return Response(
            {"error": f"Failed to build workbook preview: {e}"},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )
    return Response(data)


@api_view(['POST', 'DELETE'])
@permission_classes([IsAuthenticated])
def workbook_cell(request, job_id):
    """Save (POST) or clear (DELETE) a single cell override.

    Body (JSON):
        {
            "workbook":    "spec" | "cat",
            "sheet_name":  "PipingMaterialsClassData",
            "row_key":     "cls:<uuid>:b0:0",
            "column_name": "MaterialGrade",
            "value":       "new value"       # required for POST
        }
    """
    job = get_object_or_404(PaperSpecExtractionJob, pk=job_id)
    payload = request.data or {}

    workbook    = (payload.get('workbook') or '').lower().strip()
    sheet_name  = (payload.get('sheet_name')  or '').strip()
    row_key     = (payload.get('row_key')     or '').strip()
    column_name = (payload.get('column_name') or '').strip()

    if workbook not in _VALID_WORKBOOKS or not sheet_name or not row_key or not column_name:
        return Response(
            {"error": "workbook, sheet_name, row_key, column_name are all required"},
            status=status.HTTP_400_BAD_REQUEST,
        )

    if request.method == 'DELETE':
        deleted, _ = WorkbookCellOverride.objects.filter(
            job=job, workbook=workbook, sheet_name=sheet_name,
            row_key=row_key, column_name=column_name,
        ).delete()
        return Response({"cleared": bool(deleted)})

    # POST — upsert.
    value = payload.get('value', '')
    if value is None:
        value = ''
    value = str(value)
    if len(value) > _MAX_CELL_VALUE_LEN:
        return Response(
            {"error": f"value exceeds {_MAX_CELL_VALUE_LEN} characters"},
            status=status.HTTP_400_BAD_REQUEST,
        )

    obj, created = WorkbookCellOverride.objects.update_or_create(
        job=job, workbook=workbook, sheet_name=sheet_name,
        row_key=row_key, column_name=column_name,
        defaults={
            'value':     value,
            'edited_by': request.user if request.user.is_authenticated else None,
        },
    )
    return Response(
        {
            "saved":       True,
            "created":     created,
            "workbook":    obj.workbook,
            "sheet_name":  obj.sheet_name,
            "row_key":     obj.row_key,
            "column_name": obj.column_name,
            "value":       obj.value,
        },
        status=status.HTTP_201_CREATED if created else status.HTTP_200_OK,
    )


# ─────────────────────────────────────────────────────────────────────────────
# Enhanced workbook operations — batch save & row-level delete
# ─────────────────────────────────────────────────────────────────────────────
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def workbook_batch_save(request, job_id):
    """
    Batch save multiple cell overrides at once.
    
    Body (JSON):
        {
            "cells": [
                {
                    "workbook": "spec",
                    "sheet_name": "PipingMaterialsClassData",
                    "row_key": "cls:<uuid>:b0:0",
                    "column_name": "MaterialGrade",
                    "value": "A106 Gr B"
                },
                ...
            ]
        }
    
    Response:
        {
            "saved_count": 123,
            "created": 45,
            "updated": 78,
            "s3_snapshot": {...} or null
        }
    """
    from .workbook_storage_service import batch_save_cells
    
    job = get_object_or_404(PaperSpecExtractionJob, pk=job_id)
    cells = request.data.get('cells', [])
    
    if not cells or not isinstance(cells, list):
        return Response(
            {"error": "cells array is required"},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    # Validate each cell
    for i, cell in enumerate(cells):
        if not all(k in cell for k in ['workbook', 'sheet_name', 'row_key', 'column_name']):
            return Response(
                {"error": f"Cell at index {i} is missing required fields"},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        if cell['workbook'] not in _VALID_WORKBOOKS:
            return Response(
                {"error": f"Invalid workbook at index {i}: {cell['workbook']}"},
                status=status.HTTP_400_BAD_REQUEST
            )
    
    try:
        result = batch_save_cells(
            job=job,
            cells=cells,
            user=request.user if request.user.is_authenticated else None
        )
        
        return Response(result, status=status.HTTP_200_OK)
    
    except Exception as e:
        logger.exception(f"[WorkbookBatchSave] Failed for job {job_id}")
        return Response(
            {"error": str(e)},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def workbook_delete_row(request, job_id):
    """
    Delete all cell overrides for a specific row.
    
    Body (JSON):
        {
            "workbook": "spec",
            "sheet_name": "PipingMaterialsClassData",
            "row_key": "cls:<uuid>:b0:0"
        }
    
    Response:
        {
            "deleted_count": 12,
            "row_key": "cls:<uuid>:b0:0",
            "columns_deleted": ["MaterialGrade", "Class", ...]
        }
    """
    from .workbook_storage_service import delete_row
    
    job = get_object_or_404(PaperSpecExtractionJob, pk=job_id)
    payload = request.data or {}
    
    workbook = (payload.get('workbook') or '').lower().strip()
    sheet_name = (payload.get('sheet_name') or '').strip()
    row_key = (payload.get('row_key') or '').strip()
    
    if workbook not in _VALID_WORKBOOKS or not sheet_name or not row_key:
        return Response(
            {"error": "workbook, sheet_name, and row_key are required"},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    try:
        result = delete_row(
            job=job,
            workbook=workbook,
            sheet_name=sheet_name,
            row_key=row_key
        )
        
        return Response(result, status=status.HTTP_200_OK)
    
    except ValueError as e:
        return Response(
            {"error": str(e)},
            status=status.HTTP_400_BAD_REQUEST
        )
    except Exception as e:
        logger.exception(f"[WorkbookDeleteRow] Failed for job {job_id}")
        return Response(
            {"error": str(e)},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def workbook_bulk_delete_rows(request, job_id):
    """
    Delete multiple rows at once.
    
    Body (JSON):
        {
            "workbook": "spec",
            "sheet_name": "PipingMaterialsClassData",
            "row_keys": ["cls:<uuid1>:b0:0", "cls:<uuid2>:b0:0", ...]
        }
    
    Response:
        {
            "deleted_rows": 45,
            "deleted_cells": 234,
            "row_keys": [...]
        }
    """
    from .workbook_storage_service import bulk_delete_rows
    
    job = get_object_or_404(PaperSpecExtractionJob, pk=job_id)
    payload = request.data or {}
    
    workbook = (payload.get('workbook') or '').lower().strip()
    sheet_name = (payload.get('sheet_name') or '').strip()
    row_keys = payload.get('row_keys', [])
    
    if workbook not in _VALID_WORKBOOKS or not sheet_name or not row_keys:
        return Response(
            {"error": "workbook, sheet_name, and row_keys array are required"},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    if not isinstance(row_keys, list):
        return Response(
            {"error": "row_keys must be an array"},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    try:
        result = bulk_delete_rows(
            job=job,
            workbook=workbook,
            sheet_name=sheet_name,
            row_keys=row_keys
        )
        
        return Response(result, status=status.HTTP_200_OK)
    
    except ValueError as e:
        return Response(
            {"error": str(e)},
            status=status.HTTP_400_BAD_REQUEST
        )
    except Exception as e:
        logger.exception(f"[WorkbookBulkDeleteRows] Failed for job {job_id}")
        return Response(
            {"error": str(e)},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )
