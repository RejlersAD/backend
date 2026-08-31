"""Validated Excel/CSV import for temporary manual daily attendance entry."""
from __future__ import annotations

import csv
import datetime as dt
import io
import re
from dataclasses import dataclass

from django.db import transaction
from django.db.models import Q
from django.utils import timezone
from openpyxl import load_workbook

from .config import RULES
from .identity import norm_code
from .models import DailyAttendanceSummary


MAX_UPLOAD_BYTES = 10 * 1024 * 1024
CODE_ALIASES = {'employeeid', 'empid', 'employeecode', 'employeenumber', 'staffid', 'code', 'id'}
DATE_ALIASES = {'date', 'workdate', 'attendancedate', 'day'}
HOURS_ALIASES = {'hours', 'dailyhours', 'workedhours', 'hoursworked', 'totalhours'}
NAME_ALIASES = {'employeename', 'staffname', 'name'}
DEPARTMENT_ALIASES = {'department', 'dept', 'discipline'}
STATUS_ALIASES = {'status', 'attendancestatus'}
TIME_IN_ALIASES = {'timein', 'checkin', 'intime'}
TIME_OUT_ALIASES = {'timeout', 'checkout', 'outtime'}


@dataclass
class ImportRow:
    row_number: int
    employee_code: str
    date: dt.date
    hours: float
    employee_name: str = ''
    department: str = ''
    status: str = 'present'
    time_in: dt.time | None = None
    time_out: dt.time | None = None


def _key(value) -> str:
    return re.sub(r'[^a-z0-9]', '', str(value or '').lower())


def _date(value, *, year=None, month=None) -> dt.date:
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    if isinstance(value, (int, float)) and year and month and 1 <= int(value) <= 31:
        return dt.date(int(year), int(month), int(value))
    raw = str(value or '').strip()
    if raw.isdigit() and year and month:
        return dt.date(int(year), int(month), int(raw))
    for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%m/%d/%Y'):
        try:
            return dt.datetime.strptime(raw, fmt).date()
        except ValueError:
            continue
    raise ValueError(f'Invalid date "{raw}"')


def _time(value) -> dt.time | None:
    if value in (None, ''):
        return None
    if isinstance(value, dt.datetime):
        return value.time().replace(microsecond=0)
    if isinstance(value, dt.time):
        return value.replace(microsecond=0)
    raw = str(value).strip()
    if raw.lower() in {'-', '--', 'n/a', 'na'}:
        return None
    for fmt in ('%I:%M %p', '%H:%M', '%H:%M:%S'):
        try:
            return dt.datetime.strptime(raw, fmt).time()
        except ValueError:
            continue
    raise ValueError(f'Invalid time "{raw}"')


def _hours(value) -> float:
    """Parse decimal hours or COSEC duration values such as ``10:55``.

    COSEC's attendance export writes Total Hours as an HH:MM string rather
    than a decimal number.  Excel may also expose a duration-formatted cell as
    a ``time`` or ``timedelta`` value, so accept those representations too.
    """
    if isinstance(value, dt.timedelta):
        return value.total_seconds() / 3600
    if isinstance(value, dt.datetime):
        value = value.time()
    if isinstance(value, dt.time):
        return value.hour + (value.minute / 60) + (value.second / 3600)
    if isinstance(value, (int, float)):
        return float(value)

    raw = str(value or '').strip()
    if raw.lower() in {'-', '--', 'n/a', 'na'}:
        return 0.0
    match = re.fullmatch(r'(\d{1,3}):([0-5]\d)(?::([0-5]\d))?', raw)
    if match:
        hours, minutes, seconds = match.groups()
        return int(hours) + (int(minutes) / 60) + (int(seconds or 0) / 3600)
    try:
        return float(raw)
    except ValueError as exc:
        raise ValueError(f'Invalid hours "{raw}"') from exc


def _sheet_rows(upload) -> list[list]:
    name = (upload.name or '').lower()
    if upload.size > MAX_UPLOAD_BYTES:
        raise ValueError('File is larger than the 10 MB limit.')
    payload = upload.read()
    if name.endswith('.csv'):
        text = payload.decode('utf-8-sig')
        return [list(row) for row in csv.reader(io.StringIO(text))]
    if name.endswith('.xlsx'):
        book = load_workbook(io.BytesIO(payload), read_only=True, data_only=True)
        return [list(row) for row in book.active.iter_rows(values_only=True)]
    raise ValueError('Upload a .xlsx or .csv file.')


def _cosec_report_columns(rows: list[list]) -> tuple[int, dict] | None:
    """Detect Matrix COSEC's native Organization-Wise Attendance report.

    COSEC exports a two-row header (for example ``User`` + ``ID`` and
    ``Work`` + ``Hrs``), followed by date-marker rows and employee rows.
    Return the second header row index and the resolved column map.
    """
    title_found = any(
        'organizationwiseattendance' in _key(value)
        for row in rows[:10]
        for value in row
    )
    if not title_found:
        return None

    for top_idx in range(min(12, max(0, len(rows) - 1))):
        top, bottom = rows[top_idx], rows[top_idx + 1]
        width = max(len(top), len(bottom))
        combined = [
            _key(f'{(top[idx] if idx < len(top) else "") or ""} '
                 f'{(bottom[idx] if idx < len(bottom) else "") or ""}')
            for idx in range(width)
        ]
        code_idx = next((idx for idx, key in enumerate(combined) if key == 'userid'), None)
        hours_idx = next((idx for idx, key in enumerate(combined) if key == 'workhrs'), None)
        name_idx = next((idx for idx, key in enumerate(combined) if key == 'name'), None)
        if code_idx is None or hours_idx is None or name_idx is None:
            continue
        return top_idx + 1, {
            'code': code_idx,
            'name': name_idx,
            'hours': hours_idx,
            'time_in': [idx for idx, key in enumerate(combined) if key.startswith('in')],
            'time_out': [idx for idx, key in enumerate(combined) if key.startswith('out')],
        }
    raise ValueError('COSEC attendance report header could not be recognized.')


def _parse_cosec_report(rows: list[list], header_idx: int, columns: dict,
                        *, year=None, month=None) -> tuple[list[ImportRow], list[dict]]:
    parsed, errors = [], []
    current_date = None

    def value_at(values, idx, default=None):
        return values[idx] if idx is not None and idx < len(values) else default

    def clock(values, indices, *, last=False):
        candidates = []
        for idx in indices:
            raw = value_at(values, idx)
            if raw not in (None, ''):
                candidates.append(raw)
        return _time(candidates[-1 if last else 0]) if candidates else None

    for number, values in enumerate(rows[header_idx + 1:], start=header_idx + 2):
        code_value = value_at(values, columns['code'])
        if isinstance(code_value, (dt.datetime, dt.date)):
            current_date = _date(code_value)
            continue
        if current_date is None or code_value in (None, ''):
            continue
        try:
            code = norm_code(code_value)
            if not code:
                raise ValueError('Employee ID is required')
            hours_value = value_at(values, columns['hours'])
            hours = 0.0 if hours_value in (None, '') else _hours(hours_value)
            time_in = clock(values, columns['time_in'])
            time_out = clock(values, columns['time_out'], last=True)
            status = 'present' if hours > 0 else ('incomplete' if time_in and not time_out else 'absent')
            parsed.append(ImportRow(
                number, code, current_date, hours,
                employee_name=str(value_at(values, columns['name'], '') or '').strip(),
                status=status,
                time_in=time_in,
                time_out=time_out,
            ))
        except (TypeError, ValueError) as exc:
            errors.append({'row': number, 'error': str(exc)})
    return parsed, errors


def _parse(rows: list[list], *, year=None, month=None) -> tuple[list[ImportRow], list[dict]]:
    if not rows:
        raise ValueError('The uploaded file is empty.')
    cosec_layout = _cosec_report_columns(rows)
    if cosec_layout:
        return _parse_cosec_report(rows, *cosec_layout, year=year, month=month)
    headers = [_key(v) for v in rows[0]]
    code_idx = next((i for i, h in enumerate(headers) if h in CODE_ALIASES), None)
    if code_idx is None:
        raise ValueError('Missing Employee ID column.')
    date_idx = next((i for i, h in enumerate(headers) if h in DATE_ALIASES), None)
    hours_idx = next((i for i, h in enumerate(headers) if h in HOURS_ALIASES), None)
    name_idx = next((i for i, h in enumerate(headers) if h in NAME_ALIASES), None)
    dept_idx = next((i for i, h in enumerate(headers) if h in DEPARTMENT_ALIASES), None)
    status_idx = next((i for i, h in enumerate(headers) if h in STATUS_ALIASES), None)
    time_in_idx = next((i for i, h in enumerate(headers) if h in TIME_IN_ALIASES), None)
    time_out_idx = next((i for i, h in enumerate(headers) if h in TIME_OUT_ALIASES), None)
    parsed, errors = [], []

    # Long format: Employee ID | Date | Hours
    if date_idx is not None and hours_idx is not None:
        for number, values in enumerate(rows[1:], start=2):
            if not any(v not in (None, '') for v in values):
                continue
            try:
                code = norm_code(values[code_idx] if code_idx < len(values) else '')
                if not code:
                    raise ValueError('Employee ID is required')
                day = _date(values[date_idx] if date_idx < len(values) else '', year=year, month=month)
                hours = _hours(values[hours_idx] if hours_idx < len(values) else '')
                value_at = lambda idx, default='': values[idx] if idx is not None and idx < len(values) else default
                parsed.append(ImportRow(
                    number, code, day, hours,
                    employee_name=str(value_at(name_idx) or '').strip(),
                    department=str(value_at(dept_idx) or '').strip(),
                    status=str(value_at(status_idx, 'present') or 'present').strip().lower(),
                    time_in=_time(value_at(time_in_idx)),
                    time_out=_time(value_at(time_out_idx)),
                ))
            except (TypeError, ValueError) as exc:
                errors.append({'row': number, 'error': str(exc)})
        return parsed, errors

    # Wide format: Employee ID | 1 | 2 | ... or Employee ID | 2026-08-01 | ...
    if not year or not month:
        raise ValueError('Wide files require the selected year and month.')
    day_columns = []
    for idx, value in enumerate(rows[0]):
        if idx == code_idx:
            continue
        try:
            day_columns.append((idx, _date(value, year=year, month=month)))
        except ValueError:
            continue
    if not day_columns:
        raise ValueError('Missing Date and Hours columns, or daily date columns.')
    for number, values in enumerate(rows[1:], start=2):
        code = norm_code(values[code_idx] if code_idx < len(values) else '')
        if not code:
            if any(v not in (None, '') for v in values):
                errors.append({'row': number, 'error': 'Employee ID is required'})
            continue
        for idx, day in day_columns:
            value = values[idx] if idx < len(values) else None
            if value in (None, '', '-'):
                continue
            try:
                parsed.append(ImportRow(number, code, day, _hours(value)))
            except (TypeError, ValueError):
                errors.append({'row': number, 'error': f'Invalid hours for {day.isoformat()}'})
    return parsed, errors


def import_daily_attendance(upload, *, year=None, month=None) -> dict:
    parsed, errors = _parse(_sheet_rows(upload), year=year, month=month)
    max_regular_hours = float(RULES.get('max_daily_hours', 9.0))
    max_total_hours = 24.0
    valid = []
    for item in parsed:
        if item.hours < 0 or item.hours > max_total_hours:
            errors.append({'row': item.row_number, 'error': f'Hours must be between 0 and {max_total_hours:g}'})
        elif ((year and item.date.year != int(year))
              or (month and item.date.month != int(month))):
            errors.append({
                'row': item.row_number,
                'error': f'{item.date.isoformat()} is outside the selected month',
            })
        else:
            valid.append(item)

    if not valid:
        return {'created': 0, 'updated': 0, 'removed': 0, 'skipped': len(errors), 'errors': errors[:50]}

    # Last occurrence wins when a file contains the same employee/date twice.
    valid_by_key = {(item.employee_code, item.date): item for item in valid}
    valid = list(valid_by_key.values())
    existing_keys = set(
        DailyAttendanceSummary.objects.filter(source=DailyAttendanceSummary.SOURCE_MANUAL)
        .values_list('employee_code', 'date')
    )
    positive = [item for item in valid if item.hours > 0]
    zero = [item for item in valid if item.hours == 0]
    created = sum((item.employee_code, item.date) not in existing_keys for item in positive)
    updated = len(positive) - created
    removed = 0
    full_day = float(RULES.get('full_day_hours', 9.0))
    with transaction.atomic():
        for offset in range(0, len(zero), 250):
            condition = Q()
            for item in zero[offset:offset + 250]:
                condition |= Q(employee_code=item.employee_code, date=item.date)
            if condition:
                count, _ = DailyAttendanceSummary.objects.filter(
                    condition, source=DailyAttendanceSummary.SOURCE_MANUAL,
                ).delete()
                removed += count

        now = timezone.now()
        records = [DailyAttendanceSummary(
                employee_code=item.employee_code,
                date=item.date,
                source=DailyAttendanceSummary.SOURCE_MANUAL,
                paired_hours=item.hours,
                elapsed_hours=item.hours,
                # Payroll consumes effective_hours, so keep it capped at the
                # nine-hour regular day. Excess remains an informational
                # attendance record until HR approves a separate OT request.
                effective_hours=min(item.hours, max_regular_hours),
                punch_count_in=0,
                punch_count_out=0,
                paired_segments=0,
                open_shift=False,
                open_shift_credited=0,
                is_late=False,
                is_full_day=item.hours >= full_day,
                employee_name=item.employee_name,
                department=item.department,
                attendance_status=item.status,
                time_in=item.time_in,
                time_out=item.time_out,
                overtime_hours=max(0.0, item.hours - max_regular_hours),
                computed_at=now,
            ) for item in positive]
        update_fields = [
            'paired_hours', 'elapsed_hours', 'effective_hours',
            'punch_count_in', 'punch_count_out', 'paired_segments',
            'open_shift', 'open_shift_credited', 'is_late', 'is_full_day',
            'employee_name', 'department', 'attendance_status',
            'time_in', 'time_out', 'overtime_hours', 'computed_at',
        ]
        # Keep individual SQL statements small for hosted Postgres instances
        # with strict statement timeouts and heavily indexed attendance tables.
        for offset in range(0, len(records), 100):
            DailyAttendanceSummary.objects.bulk_create(
                records[offset:offset + 100],
                batch_size=100,
                update_conflicts=True,
                unique_fields=['employee_code', 'date', 'source'],
                update_fields=update_fields,
            )
    return {
        'created': created, 'updated': updated, 'removed': removed,
        'skipped': len(errors), 'errors': errors[:50],
    }
