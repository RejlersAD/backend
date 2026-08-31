"""
Excel + PDF exporters for daily and monthly time-sheet reports.

Reuses libraries that are already in requirements.txt (openpyxl, reportlab).
Returns Django HttpResponse objects with proper content-type + filename.
"""
from __future__ import annotations

import datetime as dt
import io
from typing import Iterable

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from . import config as ts_config


def _svc():
    """Soft-coded backend dispatcher (mirrors views._svc)."""
    from . import get_service
    return get_service()


# ─────────────────────────────────────────────────────────────────────────────
# Deduplication helper
# ─────────────────────────────────────────────────────────────────────────────
def _norm_code(code: str) -> str:
    """Normalise an employee_code for deduplication.

    Soft-coded via ts_config.EXPORT_DEDUP_NORM.  Currently only 'strip' is
    supported; adding 'upper' or 'zfill_6' here does not require changes in
    callers.
    """
    s = str(code or '').strip()
    if ts_config.EXPORT_DEDUP_NORM == 'upper':
        return s.upper()
    return s   # default: strip only


def _dedup_rows(rows: list[dict]) -> list[dict]:
    """Merge rows that share the same normalised employee_code.

    Root cause guarded: the biometric source can emit the same employee as
    '22393', '22393 ' (trailing space) and ' 22393' (leading space).  Each
    variant creates a separate DB entry and — without this guard — a separate
    row in every Excel / PDF export.

    When ts_config.EXPORT_DEDUP is False the list is returned unchanged so a
    developer can inspect raw (un-merged) rows for diagnostics.
    """
    if not ts_config.EXPORT_DEDUP or not rows:
        return rows

    seen: dict[str, dict] = {}
    for r in rows:
        key = _norm_code(r.get('employee_code', ''))
        if key not in seen:
            # First occurrence — clone the row and store the normalised code.
            seen[key] = dict(r)
            seen[key]['employee_code'] = key
            # Ensure days_detail is a mutable list (might be None from DB).
            seen[key]['days_detail'] = list(r.get('days_detail') or [])
        else:
            # Merge duplicate into the canonical row.
            canonical = seen[key]
            # Prefer the row that has a resolved RAD AI name.
            if r.get('radai_full_name') and not canonical.get('radai_full_name'):
                for field in ('radai_full_name', 'radai_email', 'radai_department',
                              'radai_user_id', 'radai_job_title', 'matched_by'):
                    if r.get(field):
                        canonical[field] = r[field]
            # Accumulate numeric attendance counters.
            for num_field in ('days_present', 'full_days', 'half_days',
                              'late_arrivals', 'open_shifts', 'total_hours'):
                canonical[num_field] = (canonical.get(num_field) or 0) + (r.get(num_field) or 0)
            # Merge days_detail — keep at most one entry per date (highest hours wins).
            detail_by_date: dict[str, dict] = {
                d['date']: d for d in canonical['days_detail'] if d.get('date')
            }
            for d in (r.get('days_detail') or []):
                date_key = d.get('date')
                if not date_key:
                    continue
                existing = detail_by_date.get(date_key)
                if existing is None or (d.get('hours') or 0) > (existing.get('hours') or 0):
                    detail_by_date[date_key] = d
            canonical['days_detail'] = sorted(detail_by_date.values(), key=lambda d: d['date'])

    # Recompute derived fields after merge.
    result = list(seen.values())
    for r in result:
        r['total_hours'] = round(r.get('total_hours') or 0, 2)
        dp = r.get('days_present') or 0
        r['avg_hours_per_day'] = round(r['total_hours'] / dp, 2) if dp else 0
    return result


# ─────────────────────────────────────────────────────────────────────────────
# Excel
# ─────────────────────────────────────────────────────────────────────────────
HEADER_FILL = PatternFill('solid', fgColor='003366')
HEADER_FONT = Font(name='Calibri', size=11, bold=True, color='FFFFFF')


def export_daily_excel(date: str | None = None) -> HttpResponse:
    payload = _svc().daily_report(date)
    rows = _dedup_rows(payload['rows'])
    wb = Workbook()
    ws = wb.active
    ws.title = f"Daily {payload['date']}"

    headers = ['Employee Code', 'Name', 'Email', 'Department',
               'First In', 'Last Out', 'Regular Hours',
               'Recorded OT (Unapproved)', 'Total Presence',
               'Late?', 'Full Day?', 'Matched']
    _write_header(ws, headers)
    for r in rows:
        ws.append([
            r.get('employee_code'),
            r.get('radai_full_name') or r.get('name') or '',
            r.get('radai_email') or r.get('email') or '',
            r.get('radai_department') or r.get('department') or '',
            _fmt(r.get('first_in')),
            _fmt(r.get('last_out')),
            r.get('hours_worked'),
            r.get('overtime_hours', 0),
            r.get('total_presence_hours', r.get('hours_worked')),
            'Yes' if r.get('is_late') else 'No',
            'Yes' if r.get('is_full_day') else 'No',
            r.get('matched_by') or 'unmatched',
        ])
    _autosize(ws)
    return _xlsx_response(wb, f'timesheet_daily_{payload["date"]}.xlsx')


def export_monthly_excel(year: int | None = None, month: int | None = None) -> HttpResponse:
    payload = _svc().monthly_report(year, month)
    rows = _dedup_rows(payload['rows'])
    wb = Workbook()
    ws = wb.active
    ws.title = f"Monthly {payload['year']}-{payload['month']:02d}"

    headers = ['Employee Code', 'Name', 'Email', 'Department',
               'Days Present', 'Full Days', 'Half Days', 'Late Arrivals',
               'Regular Hours', 'Recorded OT (Unapproved)', 'Total Presence',
               'Avg Regular Hours/Day', 'Matched']
    _write_header(ws, headers)
    for r in rows:
        ws.append([
            r.get('employee_code'),
            r.get('radai_full_name') or r.get('name') or '',
            r.get('radai_email') or r.get('email') or '',
            r.get('radai_department') or r.get('department') or '',
            r.get('days_present'),
            r.get('full_days'),
            r.get('half_days'),
            r.get('late_arrivals'),
            r.get('total_hours'),
            r.get('overtime_hours', 0),
            r.get('total_presence_hours', r.get('total_hours')),
            r.get('avg_hours_per_day'),
            r.get('matched_by') or 'unmatched',
        ])
    _autosize(ws)

    # Per-day drilldown on a second sheet
    ws2 = wb.create_sheet('Per-Day Detail')
    _write_header(ws2, [
        'Employee Code', 'Name', 'Date', 'First In', 'Last Out',
        'Regular Hours', 'Recorded OT (Unapproved)', 'Total Presence',
    ])
    for r in rows:
        for d in r.get('days_detail') or []:
            ws2.append([
                r.get('employee_code'),
                r.get('radai_full_name') or r.get('name') or '',
                d.get('date'),
                _fmt(d.get('first_in')),
                _fmt(d.get('last_out')),
                d.get('hours'),
                d.get('overtime_hours', 0),
                d.get('total_presence_hours', d.get('hours')),
            ])
    _autosize(ws2)
    return _xlsx_response(wb, f'timesheet_monthly_{payload["year"]}_{payload["month"]:02d}.xlsx')


def _write_header(ws, headers: Iterable[str]):
    ws.append(list(headers))
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')


def _autosize(ws):
    for col_idx, col_cells in enumerate(ws.columns, 1):
        max_len = max((len(str(c.value)) for c in col_cells if c.value is not None), default=10)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 12), 40)


def _fmt(v) -> str:
    if v is None:
        return ''
    if isinstance(v, (dt.datetime, dt.date, dt.time)):
        return v.isoformat()
    return str(v)


def _xlsx_response(wb, filename: str) -> HttpResponse:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    resp['Content-Disposition'] = f'attachment; filename="{filename}"'
    return resp


# ─────────────────────────────────────────────────────────────────────────────
# PDF (reportlab — lightweight summary, not pixel-perfect like Excel)
# ─────────────────────────────────────────────────────────────────────────────
def export_monthly_pdf(year: int | None = None, month: int | None = None) -> HttpResponse:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    )

    payload = _svc().monthly_report(year, month)
    rows = _dedup_rows(payload['rows'])   # <- dedup safety net
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=12 * mm, rightMargin=12 * mm,
                            topMargin=14 * mm, bottomMargin=14 * mm)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('title', parent=styles['Title'],
                                 fontSize=16, textColor=colors.HexColor('#003366'))

    elements = []
    elements.append(Paragraph(
        f"Time Sheet -- Monthly Report ({payload['year']}-{payload['month']:02d})",
        title_style,
    ))
    elements.append(Spacer(1, 4 * mm))
    elements.append(Paragraph(
        f"Period: {payload['start']} to {payload['end']}<br/>"
        f"Contractual workdays: {payload.get('standard_working_days', 22)} "
        f"(calendar weekdays: {payload['working_days_in_month']})<br/>"
        f"Total employees with attendance: {len(rows)}",
        styles['Normal'],
    ))
    elements.append(Spacer(1, 6 * mm))

    headers = ['Employee', 'Email', 'Dept', 'Days', 'Full',
               'Half', 'Late', 'Hours', 'Avg/Day']
    data = [headers]
    for r in rows:
        data.append([
            (r.get('radai_full_name') or r.get('name') or r.get('employee_code') or '')[:32],
            (r.get('radai_email') or r.get('email') or '')[:36],
            (r.get('radai_department') or r.get('department') or '')[:20],
            r.get('days_present'),
            r.get('full_days'),
            r.get('half_days'),
            r.get('late_arrivals'),
            r.get('total_hours'),
            r.get('avg_hours_per_day'),
        ])
    table = Table(data, repeatRows=1, hAlign='LEFT')
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#003366')),
        ('TEXTCOLOR',  (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME',   (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE',   (0, 0), (-1, -1), 8),
        ('GRID',       (0, 0), (-1, -1), 0.25, colors.HexColor('#cccccc')),
        ('ALIGN',      (3, 1), (-1, -1), 'RIGHT'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f5f7fb')]),
    ]))
    elements.append(table)
    doc.build(elements)
    buf.seek(0)
    resp = HttpResponse(buf.getvalue(), content_type='application/pdf')
    resp['Content-Disposition'] = (
        f'attachment; filename="timesheet_monthly_{payload["year"]}_{payload["month"]:02d}.pdf"'
    )
    return resp


# ---------------------------------------------------------------------------
# Summary pivot export  (employee × day grid, matching Rejlers report format)
# ---------------------------------------------------------------------------

def export_summary_excel(year: int | None = None, month: int | None = None) -> HttpResponse:
    """
    Wide-format pivot: one row per employee, one column per calendar day.
    Matches the HR pivot table in the Summary tab exactly.
    Soft-coded thresholds from ts_config.RULES.
    """
    import calendar as _cal

    payload = _svc().monthly_report(year, month)
    rows = _dedup_rows(payload['rows'])   # <- dedup before pivot
    y, m = payload['year'], payload['month']
    days_in_month = _cal.monthrange(y, m)[1]
    day_nums = list(range(1, days_in_month + 1))

    # Build a quick lookup: employee_code -> date_str -> hours (use deduped rows)
    emp_day: dict[str, dict[str, float]] = {}
    for row in rows:
        code = row.get('employee_code', '')
        emp_day[code] = {
            d['date']: d['hours']
            for d in (row.get('days_detail') or [])
        }

    wb = Workbook()
    ws = wb.active
    ws.title = f"Summary {y}-{m:02d}"

    # Header row: Name | Code | 1 | 2 | ... | 31 | Total | Days | Normal | Diff
    std_hours = ts_config.RULES.get('standard_daily_hours', 9)
    working_days = payload.get(
        'standard_working_days',
        ts_config.RULES.get('standard_monthly_working_days', 22),
    )
    normal_total = working_days * std_hours

    fixed_hdrs = ['Name', 'Employee Code', 'Department']
    day_hdrs   = [str(d) for d in day_nums]
    tail_hdrs  = ['Total Hours', 'Days Present', f'Normal ({working_days}d×{std_hours}h)', 'Difference']
    ws.append(fixed_hdrs + day_hdrs + tail_hdrs)

    # Style header
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Weekend / day-of-week fill (light amber = Saturday, light rose = Sunday)
    SAT_FILL = PatternFill('solid', fgColor='FFF3CD')
    SUN_FILL = PatternFill('solid', fgColor='FFE4E6')

    for d in day_nums:
        dow = dt.date(y, m, d).weekday()   # 0=Mon, 5=Sat, 6=Sun
        col_idx = len(fixed_hdrs) + d      # 1-based column index
        cell = ws[1][col_idx - 1]
        if dow == 5:
            cell.fill = PatternFill('solid', fgColor='FEF08A')  # amber header
        elif dow == 6:
            cell.fill = PatternFill('solid', fgColor='FCA5A5')  # rose header

    for row in rows:  # use deduped rows
        code = row.get('employee_code', '')
        name = row.get('radai_full_name') or row.get('name') or code
        dept = row.get('radai_department') or row.get('department') or ''
        day_lookup = emp_day.get(code, {})

        total_hrs = 0.0
        days_present = 0
        day_cells: list = []

        for d in day_nums:
            date_str = f'{y}-{m:02d}-{d:02d}'
            hrs = day_lookup.get(date_str)
            dow = dt.date(y, m, d).weekday()
            if hrs is not None:
                total_hrs += hrs
                days_present += 1
                day_cells.append(round(hrs, 2))
            elif dow >= 5:
                day_cells.append('WE')   # weekend
            else:
                day_cells.append('')     # absent on working day

        total_hrs = round(total_hrs, 2)
        diff = round(total_hrs - normal_total, 2)
        ws.append([name, code, dept] + day_cells + [total_hrs, days_present, normal_total, diff])

    _autosize(ws)
    # Freeze panes after the fixed columns
    ws.freeze_panes = ws.cell(row=2, column=len(fixed_hdrs) + 1)
    return _xlsx_response(wb, f'timesheet_summary_{y}_{m:02d}.xlsx')


def export_summary_pdf(year: int | None = None, month: int | None = None) -> HttpResponse:
    """
    Landscape PDF summary: one row per employee with totals.
    The day-by-day grid is too wide for PDF; this produces the roll-up table.
    """
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle

    payload = _svc().monthly_report(year, month)
    rows = _dedup_rows(payload['rows'])   # <- dedup safety net
    y, m = payload['year'], payload['month']
    std_hours = ts_config.RULES.get('standard_daily_hours', 9)
    working_days = payload.get(
        'standard_working_days',
        ts_config.RULES.get('standard_monthly_working_days', 22),
    )
    normal_total = working_days * std_hours

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=12 * mm, rightMargin=12 * mm,
                            topMargin=14 * mm, bottomMargin=14 * mm)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('title', parent=styles['Title'],
                                 fontSize=15, textColor=colors.HexColor('#003366'))
    elements = []
    elements.append(Paragraph(f"Time Sheet -- Summary Report ({y}-{m:02d})", title_style))
    elements.append(Spacer(1, 3 * mm))
    elements.append(Paragraph(
        f"Period: {payload['start']} to {payload['end']}  |  "
        f"Working days: {working_days}  |  Normal hours/employee: {normal_total}h",
        styles['Normal'],
    ))
    elements.append(Spacer(1, 5 * mm))

    headers = ['Employee', 'Code', 'Dept', 'Total Hrs', 'Days', f'Normal ({working_days}dx{std_hours}h)', 'Difference']
    data = [headers]
    for r in rows:
        total = r.get('total_hours') or 0
        diff  = round(float(total) - normal_total, 2)
        data.append([
            (r.get('radai_full_name') or r.get('name') or r.get('employee_code') or '')[:30],
            r.get('employee_code') or '',
            (r.get('radai_department') or r.get('department') or '')[:18],
            total,
            r.get('days_present'),
            normal_total,
            f'+{diff}' if diff >= 0 else str(diff),
        ])
    table = Table(data, repeatRows=1, hAlign='LEFT')
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#003366')),
        ('TEXTCOLOR',  (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME',   (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE',   (0, 0), (-1, -1), 8),
        ('GRID',       (0, 0), (-1, -1), 0.25, colors.HexColor('#cccccc')),
        ('ALIGN',      (3, 1), (-1, -1), 'RIGHT'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f5f7fb')]),
    ]))
    elements.append(table)
    doc.build(elements)
    buf.seek(0)
    resp = HttpResponse(buf.getvalue(), content_type='application/pdf')
    resp['Content-Disposition'] = f'attachment; filename="timesheet_summary_{y}_{m:02d}.pdf"'
    return resp


# ---------------------------------------------------------------------------
# Yearly export  (12-month roll-up per employee)
# ---------------------------------------------------------------------------

def export_yearly_excel(year: int | None = None) -> HttpResponse:
    """
    One sheet per month + a 12-month summary sheet.
    Each monthly sheet mirrors export_monthly_excel; the summary sheet has
    12 columns (one per month) with total hours per employee.
    """
    import calendar as _cal

    today = dt.date.today()
    y = int(year or today.year)
    months = list(range(1, 13))
    month_names = [_cal.month_name[m] for m in months]

    wb = Workbook()

    # Build all 12 monthly payloads (only months that have data are written)
    # and collect data into a cross-month employee map for the summary sheet.
    yearly: dict[str, dict] = {}   # employee_code -> { name, dept, email, m1..m12, total }

    summary_ws = wb.active
    summary_ws.title = f'Yearly {y}'
    sum_headers = ['Name', 'Employee Code', 'Department'] + month_names + ['Full Year Total']
    summary_ws.append(sum_headers)
    for cell in summary_ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center')

    for m in months:
        payload = _svc().monthly_report(y, m)
        if not payload['rows']:
            continue

        # Monthly detail sheet — deduplicate before writing
        month_rows = _dedup_rows(payload['rows'])
        ws = wb.create_sheet(month_names[m - 1][:4])
        _write_header(ws, ['Employee Code', 'Name', 'Department', 'Email',
                            'Days Present', 'Full Days', 'Half Days', 'Late Arrivals',
                            'Total Hours', 'Avg Hours/Day'])
        for r in month_rows:
            ws.append([
                r.get('employee_code'),
                r.get('radai_full_name') or r.get('name') or '',
                r.get('radai_department') or r.get('department') or '',
                r.get('radai_email') or r.get('email') or '',
                r.get('days_present'),
                r.get('full_days'),
                r.get('half_days'),
                r.get('late_arrivals'),
                r.get('total_hours'),
                r.get('avg_hours_per_day'),
            ])
            # Accumulate into yearly map (use deduped row's normalised code)
            code = r.get('employee_code', '')
            if code not in yearly:
                yearly[code] = {
                    'name':  r.get('radai_full_name') or r.get('name') or '',
                    'dept':  r.get('radai_department') or r.get('department') or '',
                    'email': r.get('radai_email') or r.get('email') or '',
                    'months': {},
                }
            yearly[code]['months'][m] = r.get('total_hours') or 0
        _autosize(ws)

    # Write yearly summary rows
    for code, info in sorted(yearly.items(), key=lambda x: x[1]['name']):
        month_totals = [round(info['months'].get(m, 0), 2) for m in months]
        full_year    = round(sum(month_totals), 2)
        summary_ws.append([info['name'], code, info['dept']] + month_totals + [full_year])
    _autosize(summary_ws)

    return _xlsx_response(wb, f'timesheet_yearly_{y}.xlsx')


def export_yearly_pdf(year: int | None = None) -> HttpResponse:
    """
    Landscape PDF with a 12-month summary table (one row per employee,
    one column per month + full-year total).
    """
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    import calendar as _cal

    today = dt.date.today()
    y = int(year or today.year)
    months = list(range(1, 13))
    month_abbr = [_cal.month_abbr[m] for m in months]

    # Collect yearly data (reuse monthly_report for each month) — dedup each month
    yearly: dict[str, dict] = {}
    for m in months:
        payload = _svc().monthly_report(y, m)
        for r in _dedup_rows(payload['rows']):
            code = r.get('employee_code', '')
            if code not in yearly:
                yearly[code] = {
                    'name':   r.get('radai_full_name') or r.get('name') or '',
                    'dept':   (r.get('radai_department') or r.get('department') or '')[:16],
                    'months': {},
                }
            yearly[code]['months'][m] = r.get('total_hours') or 0

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=8 * mm, rightMargin=8 * mm,
                            topMargin=12 * mm, bottomMargin=12 * mm)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('title', parent=styles['Title'],
                                 fontSize=14, textColor=colors.HexColor('#003366'))
    elements = []
    elements.append(Paragraph(f"Time Sheet -- Yearly Report ({y})", title_style))
    elements.append(Spacer(1, 4 * mm))

    col_w = [4.5 * mm] * 14   # uniform narrow columns for months
    col_widths = [52 * mm, 22 * mm, 24 * mm] + col_w + [18 * mm]

    headers = ['Name', 'Code', 'Dept'] + month_abbr + ['Total']
    data = [headers]
    for code, info in sorted(yearly.items(), key=lambda x: x[1]['name']):
        month_totals = [round(info['months'].get(m, 0), 2) for m in months]
        full_year    = round(sum(month_totals), 2)
        data.append([info['name'][:30], code, info['dept']] + month_totals + [full_year])

    table = Table(data, repeatRows=1, colWidths=col_widths, hAlign='LEFT')
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#003366')),
        ('TEXTCOLOR',  (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME',   (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE',   (0, 0), (-1, -1), 7),
        ('GRID',       (0, 0), (-1, -1), 0.25, colors.HexColor('#cccccc')),
        ('ALIGN',      (3, 1), (-1, -1), 'RIGHT'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f5f7fb')]),
    ]))
    elements.append(table)
    doc.build(elements)
    buf.seek(0)
    resp = HttpResponse(buf.getvalue(), content_type='application/pdf')
    resp['Content-Disposition'] = f'attachment; filename="timesheet_yearly_{y}.pdf"'
    return resp
