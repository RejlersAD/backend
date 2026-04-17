"""
Non-TEFF Metadata — DRF views.

Endpoints
---------
POST   /api/v1/non-teff/upload/              upload_non_teff_file
GET    /api/v1/non-teff/status/<job_id>/     get_non_teff_status
GET    /api/v1/non-teff/results/<job_id>/    get_non_teff_results
GET    /api/v1/non-teff/export/<job_id>/     export_non_teff_excel
"""

import io
import json
import logging
import os
import tempfile
import uuid

from django.http import HttpResponse
from rest_framework.decorators import api_view, permission_classes, parser_classes
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework import status

from .models import NonTeffExtractionJob
from .services.extractor import run_extraction_async

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# SOFT-CODED configuration constants
# ---------------------------------------------------------------------------

MAX_FILE_SIZE_BYTES = 50 * 1024 * 1024  # 50 MB

# Extension → internal format key
EXTENSION_FORMAT_MAP = {
    '.pdf':  'pdf',
    '.xlsx': 'excel',
    '.xls':  'excel',
    '.docx': 'word',
    '.doc':  'word',
    '.dwg':  'autocad',
    '.dxf':  'autocad',
}

ALLOWED_EXTENSIONS = set(EXTENSION_FORMAT_MAP.keys())


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _get_file_format(filename):
    """Return the internal format key for a given filename, or None."""
    ext = os.path.splitext(filename.lower())[1]
    return EXTENSION_FORMAT_MAP.get(ext)


def _build_excel_response(job):
    """Build an in-memory Excel file from job results and return an HttpResponse."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return None, 'openpyxl not installed'

    # Load field definitions
    config_path = os.path.join(os.path.dirname(__file__), 'config', 'non_teff_fields.json')
    with open(config_path, 'r') as f:
        field_config = json.load(f)
    fields = field_config['fields']

    items = (job.result_json or {}).get('items', [])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Non-TEFF Metadata'

    # Header style
    header_fill = PatternFill(start_color='1B4F72', end_color='1B4F72', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=11)

    # Write headers
    for col_idx, field in enumerate(fields, start=1):
        cell = ws.cell(row=1, column=col_idx, value=field['label'])
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.column_dimensions[cell.column_letter].width = field['width']

    ws.row_dimensions[1].height = 30

    # Write data rows
    for row_idx, item in enumerate(items, start=2):
        for col_idx, field in enumerate(fields, start=1):
            ws.cell(row=row_idx, column=col_idx, value=item.get(field['key'], ''))

    ws.freeze_panes = 'A2'

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer, None


# ---------------------------------------------------------------------------
# Views
# ---------------------------------------------------------------------------

@api_view(['POST'])
@permission_classes([IsAuthenticated])
@parser_classes([MultiPartParser, FormParser])
def upload_non_teff_file(request):
    """
    Upload a document for Non-TEFF metadata extraction.

    Request: multipart/form-data with field `file`.
    Response: { job_id, status }
    """
    uploaded_file = request.FILES.get('file')
    if not uploaded_file:
        return Response({'error': 'No file provided.'}, status=status.HTTP_400_BAD_REQUEST)

    if uploaded_file.size > MAX_FILE_SIZE_BYTES:
        return Response(
            {'error': f'File exceeds maximum size of {MAX_FILE_SIZE_BYTES // (1024 * 1024)} MB.'},
            status=status.HTTP_400_BAD_REQUEST,
        )

    file_format = _get_file_format(uploaded_file.name)
    if not file_format:
        return Response(
            {'error': f'Unsupported file type. Allowed: {", ".join(ALLOWED_EXTENSIONS)}'},
            status=status.HTTP_400_BAD_REQUEST,
        )

    # Persist file to a temp location for background processing
    suffix = os.path.splitext(uploaded_file.name)[1].lower()
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=suffix)
    for chunk in uploaded_file.chunks():
        tmp.write(chunk)
    tmp.flush()
    tmp_path = tmp.name
    tmp.close()

    # Create DB job record
    job = NonTeffExtractionJob.objects.create(
        file_name=uploaded_file.name,
        file_format=file_format,
        status=NonTeffExtractionJob.STATUS_PENDING,
        status_message='Queued for extraction',
        created_by=request.user,
    )

    # Kick off async extraction
    run_extraction_async(str(job.job_id), tmp_path, file_format)

    return Response({'job_id': str(job.job_id), 'status': job.status}, status=status.HTTP_202_ACCEPTED)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_non_teff_status(request, job_id):
    """
    Poll extraction status.

    Response: { job_id, status, progress, message }
    """
    try:
        job = NonTeffExtractionJob.objects.get(pk=job_id)
    except NonTeffExtractionJob.DoesNotExist:
        return Response({'error': 'Job not found.'}, status=status.HTTP_404_NOT_FOUND)
    except Exception:
        return Response({'error': 'Invalid job ID.'}, status=status.HTTP_400_BAD_REQUEST)

    return Response({
        'job_id': str(job.job_id),
        'status': job.status,
        'progress': job.progress,
        'message': job.status_message,
        'error': job.error_message,
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_non_teff_results(request, job_id):
    """
    Fetch completed extraction results.

    Response: { job_id, total, items: [...] }
    """
    try:
        job = NonTeffExtractionJob.objects.get(pk=job_id)
    except NonTeffExtractionJob.DoesNotExist:
        return Response({'error': 'Job not found.'}, status=status.HTTP_404_NOT_FOUND)
    except Exception:
        return Response({'error': 'Invalid job ID.'}, status=status.HTTP_400_BAD_REQUEST)

    if job.status != NonTeffExtractionJob.STATUS_COMPLETED:
        return Response(
            {'error': f'Results not ready. Current status: {job.status}'},
            status=status.HTTP_202_ACCEPTED,
        )

    result = job.result_json or {}
    return Response({
        'job_id': str(job.job_id),
        'file_name': job.file_name,
        'total': result.get('total', 0),
        'items': result.get('items', []),
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_non_teff_excel(request, job_id):
    """
    Export extraction results as an Excel file.
    """
    try:
        job = NonTeffExtractionJob.objects.get(pk=job_id)
    except NonTeffExtractionJob.DoesNotExist:
        return Response({'error': 'Job not found.'}, status=status.HTTP_404_NOT_FOUND)
    except Exception:
        return Response({'error': 'Invalid job ID.'}, status=status.HTTP_400_BAD_REQUEST)

    if job.status != NonTeffExtractionJob.STATUS_COMPLETED:
        return Response(
            {'error': 'Export not available — extraction not yet complete.'},
            status=status.HTTP_400_BAD_REQUEST,
        )

    buffer, error = _build_excel_response(job)
    if error:
        return Response({'error': error}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    safe_name = os.path.splitext(job.file_name)[0].replace(' ', '_')
    filename = f'NonTEFF_Metadata_{safe_name}.xlsx'

    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response
