"""Authoritative Purchase Order register import with PR integrity checks."""

from __future__ import annotations

from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO
import re
from typing import Any

from django.db import transaction
from django.db.models import Case, IntegerField, Q, Value, When
from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel

from ..models import PurchaseOrder, PurchaseRequisition, Vendor
from .pr_excel_import import _match_vendor
from .purchase_order_numbering import PurchaseOrderNumberService


MAX_IMPORT_ROWS = 5000
MAX_FILE_SIZE = 15 * 1024 * 1024
AUTHORITATIVE_SOURCE = "Procurement Department"
PO_SOURCE_PATTERN = re.compile(
    r"RAD-(GEN|PRJ)-PUR-(\d{4})_\s*(?:[A-Z]{3})?(\d{4})",
    re.IGNORECASE,
)
PR_SOURCE_PATTERN = re.compile(
    r"RAD-(GEN|PRJ)-PR-(\d{4})_(\d{4})",
    re.IGNORECASE,
)

HEADER_ALIASES = {
    "po_number": {"po number", "po no", "po no."},
    "pr_number": {"pr number", "pr no", "pr no."},
    "pr_accepted_date": {"pr accepted date"},
    "supplier": {"suppl.name", "supplier name", "vendor name"},
    "summary": {"summary of purchase", "summary of purchase /activity"},
    "project": {"project short name/ code", "project short name/code"},
    "order_date": {"ord.date", "ord. date", "order date"},
    "oa_date": {"oa date"},
    "delivery_date": {"delivery date", "delivery/ completion date"},
    "payment_terms": {"payment terms"},
    "amount": {"amount curr.", "po amount w/o vat"},
    "currency": {"curr.", "currency", "po currency"},
    "amount_including_vat": {"amount including vat", "po amount including vat"},
    "amount_aed": {"amount inc vat in aed", "amount excl vat in aed"},
    "country": {"country", "country (of vendor/sc)"},
    "remarks": {"remarks", "notes"},
}


def _clean(value: Any, limit: int | None = None) -> str:
    text = re.sub(r"\s+", " ", str(value or "")).strip()
    if text.lower() in {"none", "nan"}:
        return ""
    return text[:limit] if limit else text


def _header(value: Any) -> str:
    return _clean(value).lower()


NORMALISED_ALIASES = {
    key: {_header(alias) for alias in aliases}
    for key, aliases in HEADER_ALIASES.items()
}


def canonical_po_number(value: Any) -> str | None:
    match = PO_SOURCE_PATTERN.search(_clean(value).upper())
    if not match:
        return None
    scope, sequence, year = match.groups()
    return f"RAD-{scope.upper()}-PUR-{sequence}_{year}"


def canonical_pr_number(value: Any) -> str | None:
    match = PR_SOURCE_PATTERN.search(_clean(value).upper())
    if not match:
        return None
    scope, sequence, year = match.groups()
    return f"RAD-{scope.upper()}-PR-{sequence}_{year}"


def _parse_decimal(value: Any) -> Decimal | None:
    if value is None or value == "" or isinstance(value, bool):
        return None
    match = re.search(r"-?\d+(?:\.\d+)?", str(value).replace(",", ""))
    if not match:
        return None
    try:
        return Decimal(match.group()).quantize(Decimal("0.01"))
    except InvalidOperation:
        return None


def _parse_date(value: Any, epoch) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        try:
            converted = from_excel(value, epoch)
            return converted.date() if isinstance(converted, datetime) else converted
        except (TypeError, ValueError, OverflowError):
            return None
    text = _clean(value)
    # Delivery cells may contain qualifiers such as "or earlier".
    match = re.search(r"\d{1,2}[./-]\d{1,2}[./-]\d{4}", text)
    candidate = match.group() if match else text
    for fmt in ("%d.%m.%Y", "%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(candidate, fmt).date()
        except ValueError:
            continue
    return None


def _source_value(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    if isinstance(value, (int, float, bool)):
        return value
    return _clean(value)


def _find_header(ws):
    for row_number, row in enumerate(ws.iter_rows(min_row=1, max_row=min(30, ws.max_row), values_only=True), 1):
        columns = {}
        for index, value in enumerate(row):
            name = _header(value)
            for key, aliases in NORMALISED_ALIASES.items():
                if name in aliases:
                    columns[key] = index
                    break
        if "po_number" in columns and "pr_number" in columns and len(columns) >= 5:
            return row_number, columns
    return None, {}


@dataclass
class ParsedPORow:
    sheet: str
    row_number: int
    source_po_number: str
    po_number: str
    source_pr_number: str
    pr_number: str
    values: dict[str, Any] = field(default_factory=dict)
    source_register: dict[str, Any] = field(default_factory=dict)
    warnings: list[str] = field(default_factory=list)

    def preview(self):
        return {
            "sheet": self.sheet,
            "row_number": self.row_number,
            "source_po_number": self.source_po_number,
            "po_number": self.po_number,
            "source_pr_number": self.source_pr_number,
            "pr_number": self.pr_number,
            "supplier_name": self.values["supplier_name"],
            "summary": self.values["title"],
            "amount": str(self.values["total_amount"]),
            "currency": self.values["currency"],
            "warnings": list(self.warnings),
        }


class POExcelImportError(ValueError):
    pass


def parse_po_workbook(file_obj):
    size = getattr(file_obj, "size", None)
    if size is not None and size > MAX_FILE_SIZE:
        raise POExcelImportError("Excel file must not exceed 15 MB.")
    filename = getattr(file_obj, "name", "")
    if filename and not filename.lower().endswith((".xlsx", ".xlsm")):
        raise POExcelImportError("Only .xlsx and .xlsm files are supported.")
    try:
        if hasattr(file_obj, "seek"):
            file_obj.seek(0)
        workbook = load_workbook(BytesIO(file_obj.read()), read_only=True, data_only=True)
    except Exception as exc:
        raise POExcelImportError("The uploaded file is not a readable Excel workbook.") from exc

    parsed, errors, seen = [], [], set()
    for ws in workbook.worksheets:
        header_row, columns = _find_header(ws)
        if not header_row:
            errors.append({"sheet": ws.title, "row_number": None, "error": "PO/PR register headers were not found."})
            continue
        for row_number, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), header_row + 1):
            if len(parsed) + len(errors) >= MAX_IMPORT_ROWS:
                raise POExcelImportError(f"The workbook exceeds the {MAX_IMPORT_ROWS}-row import limit.")

            def cell(key):
                index = columns.get(key)
                return row[index] if index is not None and index < len(row) else None

            source_po = _clean(cell("po_number"), 200)
            source_pr = _clean(cell("pr_number"), 200)
            if not source_po and not source_pr:
                continue
            if source_po.lower() in {"no po", "n/a", "na", "none", "hold"}:
                errors.append({"sheet": ws.title, "row_number": row_number, "pr_number": source_pr, "error": "No PO was issued for this requisition; row skipped.", "error_code": "no_po"})
                continue
            po_number = canonical_po_number(source_po)
            pr_number = canonical_pr_number(source_pr)
            if not po_number:
                errors.append({"sheet": ws.title, "row_number": row_number, "po_number": source_po, "error": "PO number cannot be converted to RAD-{GEN|PRJ}-PUR-####_YYYY."})
                continue
            if not pr_number:
                errors.append({"sheet": ws.title, "row_number": row_number, "po_number": source_po, "error": "PR number does not follow RAD-{GEN|PRJ}-PR-####_YYYY."})
                continue
            if po_number in seen:
                errors.append({"sheet": ws.title, "row_number": row_number, "po_number": po_number, "error": "Duplicate canonical PO number in workbook."})
                continue
            seen.add(po_number)
            verified, message = PurchaseOrderNumberService.verify(po_number, pr_number)
            if not verified:
                errors.append({"sheet": ws.title, "row_number": row_number, "po_number": po_number, "pr_number": pr_number, "error": message})
                continue

            supplier = _clean(cell("supplier"), 300)
            summary = _clean(cell("summary"))
            amount = _parse_decimal(cell("amount")) or Decimal("0.00")
            amount_with_vat = _parse_decimal(cell("amount_including_vat"))
            currency = _clean(cell("currency"), 10).upper().replace(" ", "") or "AED"
            warnings = []
            if source_po != po_number:
                warnings.append(f"Source PO number normalized to {po_number}.")
            if not supplier:
                warnings.append("Supplier name is blank.")
            if not summary:
                warnings.append("Purchase summary is blank.")
            register = {
                "PO Number": _source_value(cell("po_number")),
                "PR Number": _source_value(cell("pr_number")),
                "PR Accepted Date": _source_value(cell("pr_accepted_date")),
                "Suppl.Name": _source_value(cell("supplier")),
                "Summary of Purchase": _source_value(cell("summary")),
                "Project short name/ Code": _source_value(cell("project")),
                "Ord.Date": _source_value(cell("order_date")),
                "OA date": _source_value(cell("oa_date")),
                "Delivery Date": _source_value(cell("delivery_date")),
                "Payment terms": _source_value(cell("payment_terms")),
                "Amount Curr.": _source_value(cell("amount")),
                "Curr.": _source_value(cell("currency")),
                "Amount including VAT": _source_value(cell("amount_including_vat")),
                "Amount Inc VAT in AED": _source_value(cell("amount_aed")),
                "Country": _source_value(cell("country")),
                "Remarks": _source_value(cell("remarks")),
            }
            parsed.append(ParsedPORow(
                ws.title, row_number, source_po, po_number, source_pr, pr_number,
                values={
                    "supplier_name": supplier,
                    "title": (summary or f"Purchase Order {po_number}")[:300],
                    "description": summary,
                    "category": "other",
                    "total_amount": amount,
                    "tax_amount": max((amount_with_vat or amount) - amount, Decimal("0.00")),
                    "currency": currency,
                    "payment_terms": _clean(cell("payment_terms"), 300),
                    "project_number": _clean(cell("project"), 100),
                    "po_date": _parse_date(cell("order_date"), workbook.epoch),
                    "expected_delivery": _parse_date(cell("delivery_date"), workbook.epoch),
                    "notes": _clean(cell("remarks")),
                    "status": "sent",
                },
                source_register=register,
                warnings=warnings,
            ))
    if not parsed and not errors:
        raise POExcelImportError("No Purchase Order rows were found in the workbook.")
    return parsed, errors


def _source_attachment(row: ParsedPORow, filename: str):
    return {
        "type": "po_excel_import_source",
        "source_authority": AUTHORITATIVE_SOURCE,
        "source_workbook": filename,
        "source_sheet": row.sheet,
        "source_row": row.row_number,
        "source_po_number": row.source_po_number,
        "canonical_po_number": row.po_number,
        "procurement_register": row.source_register,
    }


@transaction.atomic
def import_po_workbook(file_obj, *, user, dry_run=True):
    rows, errors = parse_po_workbook(file_obj)
    filename = getattr(file_obj, "name", "RAD-PO.xlsx").replace("\\", "/").rsplit("/", 1)[-1]
    prs = {pr.pr_number: pr for pr in PurchaseRequisition.objects.filter(pr_number__in=[row.pr_number for row in rows])}
    vendors = list(Vendor.objects.all().only("id", "vendor_code", "name", "country"))
    previews, imported = [], []

    for row in rows:
        preview = row.preview()
        pr = prs.get(row.pr_number)
        vendor_match = _match_vendor(row.values["supplier_name"], vendors)
        existing = PurchaseOrder.objects.filter(
            Q(po_number=row.po_number) | Q(po_number=row.source_po_number)
        ).annotate(
            canonical_first=Case(
                When(po_number=row.po_number, then=Value(0)),
                default=Value(1),
                output_field=IntegerField(),
            )
        ).order_by("canonical_first").first()
        legacy_duplicate = bool(
            existing
            and existing.po_number == row.po_number
            and row.source_po_number != row.po_number
            and PurchaseOrder.objects.filter(po_number=row.source_po_number).exists()
        )
        preview.update({
            "pr_linked": bool(pr),
            "pr_id": str(pr.id) if pr else None,
            "vendor_match": vendor_match,
            "operation": "overwrite" if existing else "create",
        })
        if legacy_duplicate:
            preview["warnings"].append(
                "A separate legacy month-suffixed PO also exists; the canonical PO is the overwrite target."
            )
        if not pr:
            preview["status"] = "error"
            preview["warnings"].append("The referenced PR does not exist in RADAI; PO was not imported.")
        elif not vendor_match.get("matched"):
            preview["status"] = "error"
            preview["warnings"].append("Supplier does not unambiguously match the company vendor database.")
        else:
            preview["status"] = "ready"
        previews.append(preview)
        if dry_run or preview["status"] != "ready":
            continue

        with transaction.atomic():
            po = PurchaseOrder.objects.select_for_update().filter(
                Q(po_number=row.po_number) | Q(po_number=row.source_po_number)
            ).annotate(
                canonical_first=Case(
                    When(po_number=row.po_number, then=Value(0)),
                    default=Value(1),
                    output_field=IntegerField(),
                )
            ).order_by("canonical_first").first()
            created = po is None
            if created:
                po = PurchaseOrder(po_number=row.po_number, created_by=user)
            else:
                po.po_number = row.po_number
            for key, value in row.values.items():
                if key != "supplier_name" and not (key == "po_date" and value is None):
                    setattr(po, key, value)
            po.vendor_id = vendor_match["id"]
            po.pr_reference = pr
            attachments = [item for item in (po.attachments or []) if not (isinstance(item, dict) and item.get("type") == "po_excel_import_source")]
            attachments.append(_source_attachment(row, filename))
            po.attachments = attachments
            po.save()
            if row.values.get("po_date"):
                PurchaseOrder.objects.filter(pk=po.pk).update(po_date=row.values["po_date"])
                po.po_date = row.values["po_date"]
            PurchaseRequisition.objects.filter(pk=pr.pk).update(
                po_applicable=True,
                po_number_reference=row.po_number,
                status="converted",
            )
            imported.append({
                "id": str(po.id),
                "po_number": po.po_number,
                "pr_id": str(pr.id),
                "pr_number": pr.pr_number,
                "operation": "created" if created else "overwritten",
            })

    persisted = []
    if imported:
        persisted = list(
            PurchaseOrder.objects.filter(id__in=[item["id"] for item in imported])
            .select_related("pr_reference")
            .values("id", "po_number", "pr_reference_id", "pr_reference__pr_number")
        )
    persisted_by_id = {str(item["id"]): item for item in persisted}
    verified_links = [
        item for item in imported
        if str(item["id"]) in persisted_by_id
        and str(persisted_by_id[str(item["id"])]["pr_reference_id"]) == item["pr_id"]
        and persisted_by_id[str(item["id"])]["pr_reference__pr_number"] == item["pr_number"]
    ]

    return {
        "dry_run": bool(dry_run),
        "source_authority": AUTHORITATIVE_SOURCE,
        "source_authoritative": True,
        "total_rows": len(rows),
        "ready_rows": sum(item["status"] == "ready" for item in previews),
        "created_count": sum(item["operation"] == "created" for item in imported),
        "overwritten_count": sum(item["operation"] == "overwritten" for item in imported),
        "error_count": len(errors) + sum(item["status"] == "error" for item in previews),
        "rows": previews,
        "errors": errors,
        "imported": imported,
        "database_verification": {
            "verified": not dry_run and bool(imported) and len(verified_links) == len(imported),
            "verified_count": len(verified_links),
            "expected_count": len(imported),
            "transactional": True,
            "message": (
                "All imported PO records and PR foreign keys were verified in the database."
                if imported and len(verified_links) == len(imported)
                else "Database verification is available after the import is executed."
            ),
        },
    }
