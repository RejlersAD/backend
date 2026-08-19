"""Excel import support for Purchase Requisition registers."""

from __future__ import annotations

from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from difflib import SequenceMatcher
from io import BytesIO
import re
from typing import Any

from django.db import IntegrityError, transaction
from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel

from apps.core.project_models import Project as CoreProject

from ..models import PurchaseRequisition, Vendor
from ..models_master import Project


MAX_IMPORT_ROWS = 5000
MAX_FILE_SIZE = 15 * 1024 * 1024
AUTHORITATIVE_SOURCE = "Procurement Department"
STANDARD_PR_NUMBER = re.compile(
    r"^RAD-(?:PRJ|GEN)-PR-\d{4}_\d{4}(?:_R\d+)?$",
    re.IGNORECASE,
)


HEADER_ALIASES = {
    "sn": {"sn", "s/n", "serial number"},
    "pr_number": {"pr number", "pr no", "pr no."},
    "accepted_date": {"pr accepted date", "pr date", "accepted date"},
    "po_number": {"po number", "po no", "po no."},
    "order_date": {"ord.date", "ord. date", "order date"},
    "supplier": {"suppl.name", "supplier name", "vendor name"},
    "summary": {"summary of purchase /activity", "summary of purchase/activity", "purchase summary"},
    "project": {"project short name/ code", "project short name/code", "project code"},
    "oa_date": {"oa date", "order acknowledgement date"},
    "delivery_date": {"delivery/ completion date", "delivery/completion date", "required date"},
    "payment_terms": {"payment terms"},
    "amount_excl_vat": {"po amount w/o vat", "amount excl vat", "amount excluding vat"},
    "currency": {"po currency", "currency"},
    "amount_incl_vat": {"po amount including vat", "amount including vat"},
    "amount_aed": {"amount excl vat in aed", "amount excluding vat in aed"},
    "budget_aed": {"budget in aed", "budget"},
    "initial_proposal_aed": {"initial proposal in aed", "initial proposal"},
    "final_negotiated_aed": {"final negotiated price in aed", "final negotiated price"},
    "savings_percentage": {"%savings from budget", "% savings from budget", "savings from budget"},
    "negotiated_percentage": {"% negotiated", "%negotiated", "negotiated %"},
    "country": {"country (of vendor/sc)", "vendor country", "country"},
    "po_status": {"po status", "status"},
    "icv": {"icv", "icv score"},
    "remarks": {"remarks", "notes"},
}


def _normalise_header(value: Any) -> str:
    text = str(value or "").replace("\t", " ").strip().lower()
    text = re.sub(r"\s+", " ", text)
    return text


NORMALISED_ALIASES = {
    key: {_normalise_header(alias) for alias in aliases}
    for key, aliases in HEADER_ALIASES.items()
}


def _clean_text(value: Any, max_length: int | None = None) -> str:
    if value is None:
        return ""
    text = re.sub(r"\s+", " ", str(value)).strip()
    if text.lower() in {"none", "nan"}:
        return ""
    return text[:max_length] if max_length else text


def _source_value(value: Any) -> Any:
    """Return a JSON-safe display value without changing authoritative content."""
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, Decimal):
        return str(value)
    if isinstance(value, (int, float, bool)):
        return value
    return _clean_text(value)


def _parse_date(value: Any, epoch) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        try:
            converted = from_excel(value, epoch)
            return converted.date() if isinstance(converted, datetime) else converted
        except (TypeError, ValueError, OverflowError):
            return None

    text = _clean_text(value)
    if not text or text.upper() in {"HOLD", "TBA", "N/A", "NA"}:
        return None
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d-%m-%y", "%d/%m/%Y", "%d/%m/%y", "%m/%d/%Y"):
        try:
            return datetime.strptime(text.split(" ")[0], fmt).date()
        except ValueError:
            continue
    return None


def _parse_decimal(value: Any) -> Decimal | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return None
    text = str(value).strip().replace(",", "")
    if not text or text.startswith("=") or text.upper() in {"HOLD", "TBA", "NA", "N/A"}:
        return None
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    if not match:
        return None
    try:
        return Decimal(match.group(0)).quantize(Decimal("0.01"))
    except InvalidOperation:
        return None


def _project_details(project_text: str) -> list[dict[str, str]]:
    if not project_text:
        return []
    values = [part.strip() for part in re.split(r"[,;]", project_text) if part.strip()]
    return [
        {"project_id": "", "project_name": value, "department": "", "type": "project"}
        for value in values
    ]


def _normalise_lookup(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", " ", str(value or "").lower()).strip()


def _match_vendor(supplier_name: str, vendors: list[Vendor]) -> dict[str, Any]:
    source = _normalise_lookup(supplier_name)
    if not source:
        return {"matched": False, "source": supplier_name, "reason": "No supplier name supplied."}

    ranked = []
    for vendor in vendors:
        candidate = _normalise_lookup(vendor.name)
        if not candidate:
            continue
        if source == candidate:
            score, method = 1.0, "exact"
        elif len(candidate) >= 5 and (source.startswith(candidate) or candidate.startswith(source)):
            score, method = 0.98, "company-name prefix"
        elif len(candidate) >= 8 and (candidate in source or source in candidate):
            score, method = 0.95, "company-name contains"
        else:
            score, method = SequenceMatcher(None, source, candidate).ratio(), "fuzzy"
        ranked.append((score, method, vendor))

    ranked.sort(key=lambda item: item[0], reverse=True)
    if not ranked:
        return {"matched": False, "source": supplier_name, "reason": "Vendor master is empty."}

    score, method, vendor = ranked[0]
    next_score = ranked[1][0] if len(ranked) > 1 else 0
    # Fuzzy matches require both a strong score and a clear lead over the next
    # candidate. Exact/prefix/contains matches are deterministic.
    accepted = method != "fuzzy" or (score >= 0.86 and score - next_score >= 0.04)
    if not accepted:
        return {
            "matched": False,
            "source": supplier_name,
            "reason": "No unambiguous vendor-master match.",
            "best_confidence": round(score, 3),
        }
    return {
        "matched": True,
        "source": supplier_name,
        "id": str(vendor.id),
        "vendor_code": vendor.vendor_code,
        "vendor_name": vendor.name,
        "method": method,
        "confidence": round(score, 3),
    }


def _match_projects(project_details: list[dict[str, str]], projects: list[dict[str, Any]]) -> list[dict[str, Any]]:
    by_number = {}
    for project in projects:
        key = _normalise_lookup(project.get("project_number"))
        # Procurement master data takes precedence over the core project table.
        if key and (key not in by_number or project.get("database") == "procurement_master"):
            by_number[key] = project
    matches = []
    for source_item in project_details:
        source = source_item.get("project_name", "")
        project = by_number.get(_normalise_lookup(source))
        if not project:
            matches.append({"source": source, "matched": False})
            continue
        matches.append({
            "source": source,
            "matched": True,
            "id": str(project["id"]),
            "project_number": project["project_number"],
            "project_name": project["project_name"],
            "database": project["database"],
        })
    return matches


def _company_match_status(vendor_match: dict[str, Any], project_matches: list[dict[str, Any]]) -> str:
    checks = []
    if vendor_match.get("source"):
        checks.append(bool(vendor_match.get("matched")))
    checks.extend(bool(item.get("matched")) for item in project_matches)
    if not checks:
        return "no_reference"
    if all(checks):
        return "matched"
    if any(checks):
        return "partial"
    return "unmatched"


@dataclass
class ParsedRow:
    sheet: str
    row_number: int
    pr_number: str
    values: dict[str, Any] = field(default_factory=dict)
    warnings: list[str] = field(default_factory=list)
    format_valid: bool = True

    def preview(self) -> dict[str, Any]:
        return {
            "sheet": self.sheet,
            "row_number": self.row_number,
            "pr_number": self.pr_number,
            "issued_date": self.values.get("issued_date").isoformat() if self.values.get("issued_date") else None,
            "supplier_name": self.values.get("supplier_name", ""),
            "product_service": self.values.get("product_service", ""),
            "project": self.values.get("project", ""),
            "total_price": str(self.values["total_price"]) if self.values.get("total_price") is not None else None,
            "currency": self.values.get("currency", "AED"),
            "po_number_reference": self.values.get("po_number_reference", ""),
            "warnings": self.warnings,
            "format_valid": self.format_valid,
        }


class PRExcelImportError(ValueError):
    pass


def _find_header(ws) -> tuple[int, dict[str, int]] | tuple[None, dict]:
    for row_number, row in enumerate(ws.iter_rows(min_row=1, max_row=min(ws.max_row, 30), values_only=True), 1):
        mapping: dict[str, int] = {}
        for index, value in enumerate(row):
            header = _normalise_header(value)
            for key, aliases in NORMALISED_ALIASES.items():
                if header in aliases:
                    mapping[key] = index
                    break
        if "pr_number" in mapping and len(mapping) >= 3:
            return row_number, mapping
    return None, {}


def parse_pr_workbook(file_obj) -> tuple[list[ParsedRow], list[dict[str, Any]]]:
    size = getattr(file_obj, "size", None)
    if size is not None and size > MAX_FILE_SIZE:
        raise PRExcelImportError("Excel file must not exceed 15 MB.")

    filename = getattr(file_obj, "name", "")
    if filename and not filename.lower().endswith((".xlsx", ".xlsm")):
        raise PRExcelImportError("Only .xlsx and .xlsm files are supported.")

    try:
        if hasattr(file_obj, "seek"):
            file_obj.seek(0)
        content = file_obj.read() if hasattr(file_obj, "read") else file_obj
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise PRExcelImportError("The uploaded file is not a readable Excel workbook.") from exc

    parsed: list[ParsedRow] = []
    errors: list[dict[str, Any]] = []
    seen_numbers: set[str] = set()

    for ws in workbook.worksheets:
        header_row, columns = _find_header(ws)
        if header_row is None:
            errors.append({"sheet": ws.title, "row_number": None, "error": "PR Number header was not found."})
            continue

        for row_number, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), header_row + 1):
            if len(parsed) + len(errors) >= MAX_IMPORT_ROWS:
                raise PRExcelImportError(f"The workbook exceeds the {MAX_IMPORT_ROWS}-row import limit.")

            def cell(key: str):
                index = columns.get(key)
                return row[index] if index is not None and index < len(row) else None

            pr_number = _clean_text(cell("pr_number"), 50)
            if not pr_number:
                continue
            if pr_number in seen_numbers:
                errors.append({"sheet": ws.title, "row_number": row_number, "pr_number": pr_number, "error": "Duplicate PR number in workbook."})
                continue
            seen_numbers.add(pr_number)

            summary = _clean_text(cell("summary"))
            supplier = _clean_text(cell("supplier"), 300)
            project = _clean_text(cell("project"), 200)
            po_number = _clean_text(cell("po_number"), 100)
            remarks = _clean_text(cell("remarks"))
            payment_terms = _clean_text(cell("payment_terms"))
            issued_date = _parse_date(cell("accepted_date"), workbook.epoch)
            required_date = _parse_date(cell("delivery_date"), workbook.epoch)
            amount = _parse_decimal(cell("amount_excl_vat"))
            budget = _parse_decimal(cell("budget_aed"))
            currency = _clean_text(cell("currency"), 3).upper() or "AED"
            warnings: list[str] = []
            format_valid = bool(STANDARD_PR_NUMBER.fullmatch(pr_number))
            if not format_valid:
                warnings.append(
                    "Authoritative PR number does not match the current RADAI numbering pattern; it will be preserved unchanged."
                )
            if not issued_date:
                warnings.append("PR Accepted Date is blank, HOLD, or invalid; issued date will use import date.")
            if not summary:
                warnings.append("Purchase summary is blank.")

            pricing_meta = {
                "payment_terms": payment_terms,
                "amount_including_vat": str(_parse_decimal(cell("amount_incl_vat")) or ""),
                "amount_excl_vat_aed": str(_parse_decimal(cell("amount_aed")) or ""),
                "budget_in_aed": str(budget or ""),
                "vendor_country": _clean_text(cell("country")),
                "source_po_status": _clean_text(cell("po_status")),
                "icv": _clean_text(cell("icv")),
                "import_source": "excel",
                "source_authority": AUTHORITATIVE_SOURCE,
                "source_authoritative": True,
                "source_workbook": filename.rsplit("\\", 1)[-1].rsplit("/", 1)[-1] or "PR Module.xlsx",
                "source_sheet": ws.title,
                "source_row": row_number,
            }
            pricing_meta["procurement_register"] = {
                "SN": _source_value(cell("sn")),
                "PR Number": pr_number,
                "PR Accepted Date": _source_value(cell("accepted_date")),
                "PO Number": po_number,
                "Ord.Date": _source_value(cell("order_date")),
                "Suppl.Name": supplier,
                "Summary of Purchase /Activity": summary,
                "Project short name/ Code": project,
                "OA date": _source_value(cell("oa_date")),
                "Delivery/ Completion Date": _source_value(cell("delivery_date")),
                "Payment terms": payment_terms,
                "PO Amount w/o VAT": _source_value(cell("amount_excl_vat")),
                "PO Currency": currency,
                "PO Amount including VAT": _source_value(cell("amount_incl_vat")),
                "Amount Excl VAT in AED": _source_value(cell("amount_aed")),
                "Budget in AED": _source_value(cell("budget_aed")),
                "Initial Proposal in AED": _source_value(cell("initial_proposal_aed")),
                "Final Negotiated price in AED": _source_value(cell("final_negotiated_aed")),
                "%Savings from Budget": _source_value(cell("savings_percentage")),
                "% Negotiated": _source_value(cell("negotiated_percentage")),
                "Country (of Vendor/SC)": _source_value(cell("country")),
                "PO Status": _source_value(cell("po_status")),
                "ICV": _source_value(cell("icv")),
                "Remarks": remarks,
            }
            values = {
                "issued_date": issued_date,
                "supplier_name": supplier,
                "preferred_supplier_if_any": supplier,
                "product_service": summary,
                "title": summary[:300],
                "description_reason": summary,
                "price_description": summary,
                "project_department": project,
                "project": project,
                "project_details": _project_details(project),
                "total_price": amount,
                "net_total_excl_vat": amount,
                "estimated_budget": budget,
                "currency": currency,
                "required_date": required_date,
                "po_applicable": bool(po_number),
                "po_number_reference": po_number,
                "price_remarks": remarks,
                "price_remarks_data": pricing_meta,
                "notes": remarks,
                "requisition_type": "project",
                "priority": "normal",
                "status": "draft",
            }
            parsed.append(ParsedRow(ws.title, row_number, pr_number, values, warnings, format_valid))

    if not parsed and not errors:
        raise PRExcelImportError("No Purchase Requisition rows were found in the workbook.")
    return parsed, errors


def import_pr_workbook(file_obj, *, user, dry_run: bool = True) -> dict[str, Any]:
    rows, errors = parse_pr_workbook(file_obj)
    pr_numbers = [row.pr_number for row in rows]
    existing = set(
        PurchaseRequisition.objects.filter(pr_number__in=pr_numbers).values_list("pr_number", flat=True)
    )
    vendors = list(Vendor.objects.all().only("id", "vendor_code", "name"))
    projects = [
        {
            "id": project.id,
            "project_number": project.project_number,
            "project_name": project.project_name,
            "database": "procurement_master",
        }
        for project in Project.objects.all().only("id", "project_number", "project_name")
    ]
    projects.extend(
        {
            "id": project.id,
            "project_number": project.code,
            "project_name": project.name,
            "database": "company_project_master",
        }
        for project in CoreProject.objects.all().only("id", "code", "name")
    )

    previews: list[dict[str, Any]] = []
    created: list[dict[str, Any]] = []
    skipped = 0

    for row in rows:
        preview = row.preview()
        vendor_match = _match_vendor(row.values.get("supplier_name", ""), vendors)
        project_matches = _match_projects(row.values.get("project_details", []), projects)
        match_status = _company_match_status(vendor_match, project_matches)
        preview["vendor_match"] = vendor_match
        preview["project_matches"] = project_matches
        preview["company_match_status"] = match_status
        if row.pr_number in existing:
            preview["status"] = "duplicate"
            preview["warnings"] = [*preview["warnings"], "PR number already exists in RADAI."]
            skipped += 1
            previews.append(preview)
            continue

        preview["status"] = "ready"
        previews.append(preview)
        if dry_run:
            continue

        values = dict(row.values)
        if not values.get("issued_date"):
            values["issued_date"] = date.today()

        if vendor_match.get("matched"):
            values["vendor_id"] = vendor_match["id"]

        if project_matches:
            values["project_details"] = [
                {
                    "project_id": item.get("id", ""),
                    "project_name": item.get("project_number") or item.get("source", ""),
                    "department": "",
                    "type": "project",
                    "company_project_name": item.get("project_name", ""),
                    "matched": bool(item.get("matched")),
                }
                for item in project_matches
            ]

        pricing_meta = dict(values.get("price_remarks_data") or {})
        pricing_meta["company_database_match"] = {
            "status": match_status,
            "vendor": vendor_match,
            "projects": project_matches,
            "checked_at_import": True,
        }
        values["price_remarks_data"] = pricing_meta

        try:
            with transaction.atomic():
                instance = PurchaseRequisition.objects.create(
                    pr_number=row.pr_number,
                    issued_by=user,
                    requested_by=user,
                    **values,
                )
        except IntegrityError:
            skipped += 1
            duplicate_race = PurchaseRequisition.objects.filter(pr_number=row.pr_number).exists()
            errors.append({
                "sheet": row.sheet,
                "row_number": row.row_number,
                "pr_number": row.pr_number,
                "error": (
                    "PR number was created by another import and was skipped."
                    if duplicate_race
                    else "The company database rejected this row because its PR schema is not ready for import."
                ),
                "error_code": "duplicate_race" if duplicate_race else "database_constraint",
            })
            continue
        created.append({"id": str(instance.id), "pr_number": instance.pr_number})
        existing.add(row.pr_number)

    return {
        "dry_run": dry_run,
        "source_authority": AUTHORITATIVE_SOURCE,
        "source_authoritative": True,
        "total_rows": len(rows),
        "ready_rows": sum(1 for row in previews if row["status"] == "ready"),
        "created_count": len(created),
        "skipped_count": skipped,
        "error_count": len(errors),
        "company_match_summary": {
            status_name: sum(1 for row in previews if row.get("company_match_status") == status_name)
            for status_name in ("matched", "partial", "unmatched", "no_reference")
        },
        "rows": previews,
        "created": created,
        "errors": errors,
    }
