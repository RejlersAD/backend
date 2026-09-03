"""
Procurement Management Views
API endpoints for procurement workflows
"""

from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework.pagination import PageNumberPagination
from rest_framework.exceptions import PermissionDenied, ValidationError
from django.core.exceptions import ValidationError as DjangoValidationError

# RBAC - Module-level access control (soft-coded)
from apps.rbac.permissions import HasModuleAccess
from django.db.models import Q, Count, Sum, Avg
from django.utils import timezone
from datetime import timedelta
from apps.core.project_models import Project as CoreProject
from .services.document_filenames import build_procurement_pdf_filename

from rest_framework.parsers import MultiPartParser, FormParser, JSONParser
from django.db import models as django_models
from django.http import HttpResponse
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable, Flowable
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
import io
import json

class RADAILogoMark(Flowable):
    """Resolution-independent RADAI mark for controlled PDF exports."""

    def __init__(self, size=42):
        super().__init__()
        self.width = size
        self.height = size

    def draw(self):
        canvas = self.canv
        scale = self.width / 60
        canvas.saveState()
        canvas.setStrokeColor(colors.HexColor('#2B3A55'))
        canvas.setLineWidth(7 * scale)
        canvas.setLineCap(1)
        canvas.setLineJoin(1)

        outer = canvas.beginPath()
        outer.moveTo(8 * scale, 8 * scale)
        outer.lineTo(52 * scale, 8 * scale)
        outer.lineTo(52 * scale, 52 * scale)
        canvas.drawPath(outer, stroke=1, fill=0)

        diagonal = canvas.beginPath()
        diagonal.moveTo(8 * scale, 8 * scale)
        diagonal.lineTo(34 * scale, 35 * scale)
        canvas.drawPath(diagonal, stroke=1, fill=0)
        canvas.restoreState()

from .models import (
    Vendor, 
    PurchaseRequisition, 
    PurchaseOrder, 
    Receipt, 
    PODocument, 
    PROCUREMENT_CATEGORIES,
    # Master database models
    Project,
    Budget,
    CostCenter,
)
from .serializers import (
    VendorSerializer,
    VendorICVSerializer,
    PurchaseRequisitionSerializer,
    PurchaseOrderSerializer,
    ReceiptSerializer,
    ProcurementCategorySerializer,
    PODocumentSerializer,
    # Master database serializers
    ProjectListSerializer,
    ProjectDetailSerializer,
    CostCenterSerializer,
    BudgetSerializer,
)
from .services.requisition_workflow import RequisitionWorkflowService
from .services.requisition_conversion import RequisitionConversionService
from .services.purchase_order_numbering import PurchaseOrderNumberService
from .services.purchase_order_approvals import pending_entries_for, record_decision
from .services.requisition_status import canonicalize_pr_status, stored_values_for
from .services.project_relationships import (
    build_project_reconciliation_payload,
    resolve_invoice_purchase_order,
    resolve_project_relationship,
)
from apps.project_control.access import CommercialModulePermission

# Soft-coded pagination for vendor list - supports large page_size
class VendorPagination(PageNumberPagination):
    """
    Custom pagination for vendors allowing larger page sizes
    Supports page_size query parameter up to 10000 records
    """
    page_size = 100  # Default page size
    page_size_query_param = 'page_size'  # Allow client to set page_size via query param
    max_page_size = 10000  # Maximum allowed page_size


class VendorViewSet(viewsets.ModelViewSet):
    """
    ViewSet for Vendor management
    
    =��� SECURITY: Requires 'procurement_vendors' module access (soft-coded from rbac_config.py)
    """
    
    queryset = Vendor.objects.all().order_by('-created_at')
    serializer_class = VendorSerializer
    permission_classes = [IsAuthenticated, HasModuleAccess]
    module_required = 'procurement_vendors'
    pagination_class = VendorPagination  # Use custom pagination
    
    def get_queryset(self):
        queryset = super().get_queryset()
        
        # Filter by status
        status_filter = self.request.query_params.get('status', None)
        if status_filter:
            queryset = queryset.filter(status=status_filter)
        
        # Filter by rating
        rating_filter = self.request.query_params.get('rating', None)
        if rating_filter:
            queryset = queryset.filter(rating=rating_filter)
        
        # Search
        search = self.request.query_params.get('search', None)
        if search:
            queryset = queryset.filter(
                Q(name__icontains=search) |
                Q(vendor_code__icontains=search) |
                Q(email__icontains=search)
            )
        
        return queryset
    
    @action(detail=True, methods=['post'])
    def rate_vendor(self, request, pk=None):
        """Update vendor rating"""
        vendor = self.get_object()
        rating = request.data.get('rating')
        notes = request.data.get('notes', '')
        
        if rating not in [1, 2, 3, 4, 5]:
            return Response(
                {'error': 'Rating must be between 1 and 5'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        vendor.rating = rating
        if notes:
            vendor.performance_notes = notes
        vendor.save()
        
        return Response({'message': 'Vendor rating updated successfully'})
    
    @action(detail=False, methods=['get'])
    def top_vendors(self, request):
        """Get top-rated vendors"""
        top_vendors = self.get_queryset().filter(
            status='active',
            rating__gte=4
        ).order_by('-rating', 'name')[:10]
        
        serializer = self.get_serializer(top_vendors, many=True)
        return Response(serializer.data)


class PurchaseRequisitionViewSet(viewsets.ModelViewSet):
    """
    ViewSet for Purchase Requisition management
    
    Features:
    - Multi-part file upload to S3
    - Two-tier approval workflow (PM G�� VP)
    - Advanced filtering and search
    - PDF generation aligned with template
    
    =��� SECURITY: Requires 'procurement_requisitions' module access (soft-coded from rbac_config.py)
    """
    
    queryset = PurchaseRequisition.objects.select_related(
        'issued_by',
        'vendor',
        'enterprise_project',
        'requested_by',
        'approved_by',
        'pm_name',
        'eng_manager_name',
        'manager_projects_name',
        'vp_op_name',
    ).all().order_by('-created_at')
    serializer_class = PurchaseRequisitionSerializer
    permission_classes = [IsAuthenticated, HasModuleAccess]
    module_required = 'procurement_requisitions'
    parser_classes = [FormParser, MultiPartParser, JSONParser]

    def get_permissions(self):
        # Approval stages may be assigned to any active employee. The workflow
        # service verifies that the signed-in user owns the current stage, so
        # those employees do not also need procurement module access.
        approval_actions = {
            'pending_for_me',
            'pm_approve',
            'pm_reject',
            'vp_approve',
            'vp_reject',
            'eng_manager_approve',
            'eng_manager_reject',
            'manager_projects_approve',
            'manager_projects_reject',
            'process_dynamic_approval',
            'process_dynamic_rejection',
        }
        if getattr(self, 'action', None) in approval_actions:
            return [IsAuthenticated()]

        # Conversion creates a Purchase Order and therefore requires order
        # module access in addition to authentication.
        if getattr(self, 'action', None) == 'convert_to_po':
            self.module_required = 'procurement_orders'
        return super().get_permissions()

    def get_queryset(self):
        queryset = super().get_queryset()
        
        # Filter by status
        status_filter = self.request.query_params.get('status', None)
        if status_filter:
            queryset = queryset.filter(status__in=stored_values_for(status_filter))
        
        # Filter by priority
        priority_filter = self.request.query_params.get('priority', None)
        if priority_filter:
            queryset = queryset.filter(priority=priority_filter)
        
        # Filter by department
        department = self.request.query_params.get('department', None)
        if department:
            queryset = queryset.filter(department=department)
        
        # Filter by project
        project_filter = self.request.query_params.get('project', None)
        if project_filter:
            queryset = queryset.filter(project__icontains=project_filter)
        
        # Filter by approval status
        pm_approval = self.request.query_params.get('pm_approval_status', None)
        if pm_approval:
            queryset = queryset.filter(pm_approval_status=pm_approval)
        
        vp_approval = self.request.query_params.get('vp_op_approval_status', None)
        if vp_approval:
            queryset = queryset.filter(vp_op_approval_status=vp_approval)
        
        # Search across multiple fields
        search = self.request.query_params.get('search', None)
        if search:
            queryset = queryset.filter(
                Q(pr_number__icontains=search) |
                Q(title__icontains=search) |
                Q(product_service__icontains=search) |
                Q(description_reason__icontains=search) |
                Q(supplier_name__icontains=search) |
                Q(project_department__icontains=search)
            )
        
        return queryset

    def _enforce_owner_mutation(self, pr, *, deletable=False):
        if (
            str(pr.issued_by_id) != str(self.request.user.id)
            and not RequisitionWorkflowService._is_super_admin(self.request.user)
        ):
            raise PermissionDenied('Only the requisition issuer may modify this requisition.')

        if deletable:
            return

        # Procurement may correct a requisition at any lifecycle status. The
        # serializer still protects workflow decisions and other audit fields.

    def update(self, request, *args, **kwargs):
        self._enforce_owner_mutation(self.get_object())
        return super().update(request, *args, **kwargs)

    def destroy(self, request, *args, **kwargs):
        self._enforce_owner_mutation(self.get_object(), deletable=True)
        return super().destroy(request, *args, **kwargs)
    
    def create(self, request, *args, **kwargs):
        """Create PR with file upload support"""
        files = request.FILES.getlist('attachments_files', [])
        
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        
        if files:
            serializer.validated_data['attachments_files'] = files
        
        self.perform_create(serializer)
        headers = self.get_success_headers(serializer.data)
        return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)

    @action(
        detail=False,
        methods=['post'],
        url_path='import-excel',
        parser_classes=[MultiPartParser, FormParser],
    )
    def import_excel(self, request):
        """Preview or import Purchase Requisitions from an Excel register."""
        excel_file = request.FILES.get('file')
        if not excel_file:
            return Response(
                {'error': 'No Excel file provided. Use multipart field "file".'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        dry_run = str(request.data.get('dry_run', 'true')).strip().lower() not in {
            'false', '0', 'no',
        }

        from .services.pr_excel_import import (
            PRExcelImportError,
            import_pr_workbook,
        )

        try:
            result = import_pr_workbook(
                excel_file,
                user=request.user,
                dry_run=dry_run,
            )
        except PRExcelImportError as exc:
            return Response({'error': str(exc)}, status=status.HTTP_400_BAD_REQUEST)

        if not dry_run and result['ready_rows'] > 0 and result['created_count'] == 0:
            result['error'] = (
                'No Purchase Requisitions were created. Review the row errors and '
                'company database configuration before retrying.'
            )
            return Response(result, status=status.HTTP_422_UNPROCESSABLE_ENTITY)

        response_status = status.HTTP_200_OK if dry_run else status.HTTP_201_CREATED
        return Response(result, status=response_status)

    @action(
        detail=False,
        methods=['post'],
        url_path='import-signed-pdf',
        parser_classes=[MultiPartParser, FormParser],
    )
    def import_signed_pdf(self, request):
        """Capture a signed PR PDF into an existing RADAI requisition."""
        pdf_file = request.FILES.get('file')
        if not pdf_file:
            return Response(
                {'error': 'Select a signed Purchase Requisition PDF to import.'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if pdf_file.size > 15 * 1024 * 1024:
            return Response({'error': 'PDF file must not exceed 15 MB.'}, status=status.HTTP_400_BAD_REQUEST)

        from .services.signed_pr_pdf_import import (
            SignedPRImportError,
            import_signed_pr_pdf,
            preview_signed_pr_pdf,
        )
        approvals = {
            key: str(request.data.get(f'{key}_name', '')).strip()
            for key in ('pm', 'moe', 'mop', 'vp')
        }
        approvals = {key: value for key, value in approvals.items() if value}
        try:
            pdf_bytes = pdf_file.read()
            expected_pr_number = str(request.data.get('expected_pr_number', '')).strip()
            if str(request.data.get('preview_only', '')).lower() in {'1', 'true', 'yes'}:
                result = preview_signed_pr_pdf(
                    pdf_bytes,
                    filename=pdf_file.name,
                    expected_pr_number=expected_pr_number,
                )
                return Response(result, status=status.HTTP_200_OK)

            raw_overrides = request.data.get('manual_overrides', '')
            try:
                manual_overrides = json.loads(raw_overrides) if raw_overrides else None
            except (TypeError, ValueError, json.JSONDecodeError) as exc:
                raise SignedPRImportError('Manual corrections must be a valid JSON object.') from exc
            if manual_overrides is not None and not isinstance(manual_overrides, dict):
                raise SignedPRImportError('Manual corrections must be a JSON object.')

            raw_signature_overrides = request.data.get('manual_signature_overrides', '')
            try:
                manual_signature_overrides = json.loads(raw_signature_overrides) if raw_signature_overrides else None
            except (TypeError, ValueError, json.JSONDecodeError) as exc:
                raise SignedPRImportError('Manual signature verification must be a valid JSON object.') from exc

            result = import_signed_pr_pdf(
                pdf_bytes,
                filename=pdf_file.name,
                uploaded_by=request.user,
                approvals=approvals,
                signatures_verified=None,
                approval_date=str(request.data.get('approval_date', '')).strip(),
                expected_pr_number=expected_pr_number,
                manual_overrides=manual_overrides,
                manual_signature_overrides=manual_signature_overrides,
            )
        except SignedPRImportError as exc:
            return Response({'error': str(exc)}, status=status.HTTP_400_BAD_REQUEST)
        return Response(result, status=status.HTTP_200_OK)

    @action(detail=False, methods=['post'])
    def check_pr_number(self, request):
        """
        Check if PR number already exists (for manual entry validation)
        PROCUREMENT DEPARTMENT REQUIREMENT: Alert for duplicate PR numbers
        """
        pr_number = request.data.get('pr_number', '').strip()
        pr_id = request.data.get('pr_id', None)
        
        if not pr_number:
            return Response(
                {'error': 'PR number is required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        query = PurchaseRequisition.objects.filter(pr_number__iexact=pr_number)
        if pr_id:
            query = query.exclude(pk=pr_id)
        
        exists = query.exists()
        
        if exists:
            duplicate_pr = query.first()
            return Response({
                'available': False,
                'exists': True,
                'message': f'PR number "{pr_number}" already exists',
                'duplicate_pr': {
                    'id': str(duplicate_pr.id),
                    'pr_number': duplicate_pr.pr_number,
                    'title': duplicate_pr.title or duplicate_pr.product_service,
                    'status': duplicate_pr.status,
                    'created_at': duplicate_pr.created_at.isoformat() if duplicate_pr.created_at else None
                }
            })
        
        return Response({
            'available': True,
            'exists': False,
            'message': f'PR number "{pr_number}" is available'
        })

    def _build_requisition_response(self, pr):
        """Ensure API response contains both status and approval_hierarchy."""
        serializer = self.get_serializer(pr)
        payload = dict(serializer.data)
        payload['status'] = canonicalize_pr_status(pr.status)
        payload['approval_hierarchy'] = pr.approval_workflow_config if isinstance(pr.approval_workflow_config, list) else []
        payload['convert_to_po_enabled'] = canonicalize_pr_status(pr.status) == 'approved'
        return payload

    @action(detail=False, methods=['get'], url_path='pending-for-me')
    def pending_for_me(self, request):
        """Return only PRs whose current workflow stage is assigned to this approver."""
        active_statuses = set()
        for workflow_status in RequisitionWorkflowService.ACTIVE_REVIEW_STATUSES:
            active_statuses.update(stored_values_for(workflow_status))

        queryset = self.get_queryset().filter(status__in=active_statuses)
        is_super_admin = RequisitionWorkflowService._is_super_admin(request.user)
        assigned = []

        for pr in queryset:
            workflow = pr.approval_workflow_config
            if not isinstance(workflow, list):
                continue
            pending = [
                (index, stage) for index, stage in enumerate(workflow)
                if isinstance(stage, dict)
                and str(stage.get('status', 'pending')).lower() in ('pending', 'in_review')
            ]
            if not pending:
                continue
            active_level = min(RequisitionWorkflowService._stage_level(stage, index) for index, stage in pending)
            active_stages = [stage for index, stage in pending if RequisitionWorkflowService._stage_level(stage, index) == active_level]
            if is_super_admin or any(
                str(stage.get('user_id') or stage.get('approver_id')) == str(request.user.id)
                for stage in active_stages
            ):
                assigned.append(pr)

        count = len(assigned)
        if str(request.query_params.get('count_only', '')).lower() == 'true':
            return Response({'count': count, 'pending_count': count})

        try:
            limit = max(1, min(int(request.query_params.get('limit', 50)), 100))
        except (TypeError, ValueError):
            limit = 50

        return Response({
            'count': count,
            'results': self.get_serializer(assigned[:limit], many=True).data,
        })

    @action(detail=True, methods=['post'])
    def pm_approve(self, request, pk=None):
        pr = RequisitionWorkflowService.approve(
            pk,
            request.user,
            signature=request.data.get('signature', ''),
            expected_stage_key='pm',
        )
        return Response(self._build_requisition_response(pr))
    
    @action(detail=True, methods=['post'])
    def pm_reject(self, request, pk=None):
        pr = RequisitionWorkflowService.reject(
            pk,
            request.user,
            request.data.get('reason', ''),
            expected_stage_key='pm',
        )
        return Response(self._build_requisition_response(pr))
    
    @action(detail=True, methods=['post'])
    def vp_approve(self, request, pk=None):
        pr = RequisitionWorkflowService.approve(
            pk,
            request.user,
            signature=request.data.get('signature', ''),
            expected_stage_key='vp',
        )
        return Response(self._build_requisition_response(pr))
    
    @action(detail=True, methods=['post'])
    def vp_reject(self, request, pk=None):
        pr = RequisitionWorkflowService.reject(
            pk,
            request.user,
            request.data.get('reason', ''),
            expected_stage_key='vp',
        )
        return Response(self._build_requisition_response(pr))
    
    @action(detail=True, methods=['post'])
    def eng_manager_approve(self, request, pk=None):
        pr = RequisitionWorkflowService.approve(
            pk,
            request.user,
            signature=request.data.get('signature', ''),
            expected_stage_key='eng_manager',
        )
        return Response(self._build_requisition_response(pr))
    
    @action(detail=True, methods=['post'])
    def eng_manager_reject(self, request, pk=None):
        pr = RequisitionWorkflowService.reject(
            pk,
            request.user,
            request.data.get('reason', ''),
            expected_stage_key='eng_manager',
        )
        return Response(self._build_requisition_response(pr))
    
    @action(detail=True, methods=['post'])
    def manager_projects_approve(self, request, pk=None):
        pr = RequisitionWorkflowService.approve(
            pk,
            request.user,
            signature=request.data.get('signature', ''),
            expected_stage_key='manager_projects',
        )
        return Response(self._build_requisition_response(pr))
    
    @action(detail=True, methods=['post'])
    def manager_projects_reject(self, request, pk=None):
        pr = RequisitionWorkflowService.reject(
            pk,
            request.user,
            request.data.get('reason', ''),
            expected_stage_key='manager_projects',
        )
        return Response(self._build_requisition_response(pr))
    
    @action(detail=True, methods=['post'])
    def process_dynamic_approval(self, request, pk=None):
        pr = RequisitionWorkflowService.approve(
            pk,
            request.user,
            signature=request.data.get('signature', ''),
        )
        return Response(self._build_requisition_response(pr))

    @action(detail=True, methods=['post'])
    def process_dynamic_rejection(self, request, pk=None):
        pr = RequisitionWorkflowService.reject(
            pk,
            request.user,
            request.data.get('reason', ''),
        )
        return Response(self._build_requisition_response(pr))
    
    @action(detail=True, methods=['post'])
    def recommend_vendors(self, request, pk=None):
        """AI-powered vendor recommendation based on PR details and historical data"""
        pr = self.get_object()
        
        vendors = Vendor.objects.filter(status='active', rating__gte=3)
        recommendations = []
        
        for vendor in vendors:
            score = 0.0
            reasons = []
            
            if vendor.rating:
                score += (vendor.rating / 5.0) * 0.4
                reasons.append(f"Rating: {vendor.rating}/5")
            
            past_orders = vendor.purchase_orders.count() + vendor.purchase_requisitions.count()
            if past_orders > 0:
                score += min(past_orders / 10.0, 1.0) * 0.3
                reasons.append(f"{past_orders} past orders")
            
            if vendor.is_icv_certified:
                score += 0.15
                reasons.append(f"ICV certified: {vendor.icv_percentage}%")
            
            if vendor.adnoc_approved:
                score += 0.15
                reasons.append("ADNOC approved")
            
            if score > 0.5:
                recommendations.append({
                    'vendor_id': str(vendor.id),
                    'vendor_code': vendor.vendor_code,
                    'vendor_name': vendor.name,
                    'score': round(score, 2),
                    'rating': vendor.rating,
                    'past_orders': past_orders,
                    'icv_certified': vendor.is_icv_certified,
                    'icv_percentage': float(vendor.icv_percentage) if vendor.icv_percentage else None,
                    'adnoc_approved': vendor.adnoc_approved,
                    'reasons': reasons,
                })
        
        recommendations.sort(key=lambda x: x['score'], reverse=True)
        pr.ai_vendor_recommendations = recommendations[:5]
        pr.save()
        
        return Response({
            'message': f'Generated {len(recommendations)} vendor recommendations',
            'recommendations': recommendations[:5]
        })
    
    @action(detail=False, methods=['get'], url_path='vendor-options')
    def vendor_options(self, request):
        """
        Get vendor options for dropdown selection in PR form
        Supports search by name/code and filtering by ID
        """
        from .models import Vendor
        
        # Get query parameters
        search_query = request.query_params.get('q', '').strip()
        vendor_id = request.query_params.get('id', '').strip()
        try:
            limit = int(request.query_params.get('limit', 30))
            limit = min(max(1, limit), 100)  # Clamp between 1 and 100
        except (TypeError, ValueError):
            limit = 30
        
        # Start with active vendors only
        queryset = Vendor.objects.filter(status='active')
        
        # Filter by specific vendor ID if provided
        if vendor_id:
            try:
                queryset = queryset.filter(id=vendor_id)
            except (ValueError, TypeError):
                pass
        # Otherwise filter by search query
        elif search_query:
            queryset = queryset.filter(
                Q(name__icontains=search_query) |
                Q(vendor_code__icontains=search_query) |
                Q(email__icontains=search_query)
            )
        
        # Order by rating and name
        queryset = queryset.order_by('-rating', 'name')[:limit]
        
        # Format response for frontend dropdown
        suggestions = []
        for vendor in queryset:
            suggestions.append({
                'id': vendor.id,
                'name': vendor.name,
                'vendor_code': vendor.vendor_code,
                'email': vendor.email,
                'rating': vendor.rating,
                'status': vendor.status,
                'icv_percentage': float(vendor.icv_percentage) if vendor.icv_percentage else None,
                'icv_expiry_date': vendor.icv_expiry_date.strftime('%Y-%m-%d') if vendor.icv_expiry_date else None,
                'is_icv_certified': vendor.is_icv_certified,
                'adnoc_approved': vendor.adnoc_approved,
            })
        
        return Response({
            'suggestions': suggestions,
            'count': len(suggestions)
        })

    @action(detail=False, methods=['patch'], url_path='vendor-icv')
    def vendor_icv(self, request):
        """Record missing ICV data without granting full vendor-master access."""
        vendor_id = request.data.get('vendor_id')
        if not vendor_id:
            return Response({'vendor_id': 'Select a vendor.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            vendor = Vendor.objects.get(pk=vendor_id, status='active')
        except (Vendor.DoesNotExist, ValueError, TypeError):
            return Response({'vendor_id': 'The selected active vendor was not found.'}, status=status.HTTP_404_NOT_FOUND)

        serializer = VendorICVSerializer(vendor, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(VendorSerializer(vendor, context=self.get_serializer_context()).data)
    
    @action(detail=False, methods=['get'])
    def get_approvers(self, request):
        from apps.rbac.models import UserProfile
        from django.contrib.auth import get_user_model
        User = get_user_model()
        
        role = request.query_params.get('role', '').lower()
        role_title_mapping = {
            'project_manager': ['Project Manager', 'PM', 'Manager - Projects'],
            'engineering_manager': ['Engineering Manager', 'Manager - Engineering', 'Eng Manager'],
            'manager_projects': ['Manager of Projects', 'Manager - Projects', 'Projects Manager'],
            'vp_operations': ['VP Delivery', 'Vice President Delivery', 'VP Operations', 'Vice President Operations', 'VP - Operations', 'Vice President of Operation', 'Project Delivery', 'Operations & Project Delivery'],
            'finance': ['Finance Manager', 'Finance Controller', 'Financial Controller', 'Accountant'],
            'procurement_head': ['Head of Procurement', 'Procurement Head'],
            'ict_head_admin': ['ICT Administrator', 'ICT Head Admin', 'Head of ICT'],
            'super_admin': ['Super Administrator'],
        }
        assigned_role_mapping = {
            'procurement_head': {
                'codes': ['head_of_procurement', 'head_procurement', 'procurement_head'],
                'names': ['Head of Procurement', 'Procurement Head'],
            },
            'ict_head_admin': {
                'codes': ['ict_admin', 'ict_head_admin', 'head_of_ict'],
                'names': ['ICT Administrator', 'ICT Head Admin', 'Head of ICT'],
            },
            'super_admin': {
                'codes': ['super_admin', 'superadmin'],
                'names': ['Super Administrator', 'Super Admin'],
            },
        }
        # Soft-coded: the dropdown must always let the requester pick ANY active
        # user as the approver G�� job_title is only used to surface the most
        # relevant users FIRST, never to hide the rest of the user base.
        matched_user_ids = set()

        if role and role in role_title_mapping:
            job_titles = role_title_mapping[role]
            matched_qs = UserProfile.objects.filter(
                status='active',
                is_deleted=False,
                job_title__in=job_titles
            ).select_related('user')

            if matched_qs.count() == 0:
                q_objects = Q()
                for title in job_titles:
                    q_objects |= Q(job_title__icontains=title)
                matched_qs = UserProfile.objects.filter(
                    q_objects,
                    status='active',
                    is_deleted=False
                ).select_related('user')

            matched_user_ids = set(matched_qs.values_list('user_id', flat=True))

        matched_by_job_title = len(matched_user_ids) > 0

        profiles = UserProfile.objects.filter(
            status='active', is_deleted=False, user__is_active=True,
        ).select_related('user').distinct()
        if role in assigned_role_mapping:
            assigned_role_q = Q()
            for code in assigned_role_mapping[role]['codes']:
                assigned_role_q |= Q(roles__code__iexact=code)
            for name in assigned_role_mapping[role]['names']:
                assigned_role_q |= Q(roles__name__iexact=name)
            if role == 'super_admin':
                profiles = profiles.filter(
                    Q(user__is_superuser=True)
                    | Q(assigned_role_q, roles__is_active=True)
                ).distinct()
            else:
                profiles = profiles.filter(
                    assigned_role_q,
                    roles__is_active=True,
                ).distinct()
        elif role == 'finance':
            profiles = profiles.filter(
                Q(department__icontains='finance')
                | Q(department__icontains='financial')
                | Q(department__icontains='account')
                | Q(roles__is_active=True, roles__modules__code__istartswith='finance', roles__modules__is_active=True)
                | Q(roles__is_active=True, roles__modules__code__istartswith='invoice', roles__modules__is_active=True)
                | Q(roles__is_active=True, roles__modules__code__istartswith='account', roles__modules__is_active=True)
            ).distinct()
        elif role != 'any_active':
            profiles = profiles.filter(
                Q(user__is_superuser=True)
                | Q(
                    roles__is_active=True,
                    roles__modules__code='procurement_requisitions',
                    roles__modules__is_active=True,
                )
            ).distinct()

        profiles = sorted(
            profiles,
            key=lambda p: (
                p.user_id not in matched_user_ids,
                (p.user.first_name or '').lower(),
                (p.user.last_name or '').lower(),
            )
        )

        users_list = []
        for profile in profiles:
            users_list.append({
                'id': str(profile.user.id),
                'username': profile.user.get_username(),
                'email': profile.user.email,
                'full_name': profile.user.get_full_name(),
                'first_name': profile.user.first_name,
                'last_name': profile.user.last_name,
                'job_title': profile.job_title,
                'department': profile.department,
                'employee_id': profile.employee_id,
                'job_title_match': profile.user_id in matched_user_ids,
                'is_current_user': profile.user_id == request.user.id,
            })
        
        return Response({
            'role': role,
            'count': len(users_list),
            'matched_by_job_title': matched_by_job_title,
            'users': users_list
        })
    
    @action(detail=False, methods=['get'])
    def get_product_services(self, request):
        search_query = request.query_params.get('q', '').strip()
        limit = int(request.query_params.get('limit', 50))
        
        products_qs = PurchaseRequisition.objects.filter(
            product_service__isnull=False
        ).exclude(
            product_service__exact=''
        ).values_list('product_service', flat=True).distinct()
        
        if search_query:
            products_qs = PurchaseRequisition.objects.filter(
                product_service__icontains=search_query
            ).values_list('product_service', flat=True).distinct()
        
        products_list = list(products_qs[:limit])
        
        return Response({
            'count': len(products_list),
            'suggestions': products_list
        })
    @action(detail=False, methods=['get'], url_path='check-pr-number')
    def check_pr_number_endpoint(self, request):
        number = str(request.query_params.get('number', '') or '').strip()
        exclude_id = request.query_params.get('exclude_id')
        queryset = PurchaseRequisition.objects.filter(pr_number__iexact=number)
        if exclude_id:
            queryset = queryset.exclude(pk=exclude_id)
        return Response({
            'number': number,
            'available': bool(number) and not queryset.exists(),
            'message': '' if number and not queryset.exists() else 'This PR number already exists.',
        })
    
    @action(detail=False, methods=['get'])
    def get_projects_departments(self, request):
        """Return project/department suggestions from master data and PR history."""
        search_query = request.query_params.get('q', '').strip()
        limit = int(request.query_params.get('limit', 50))
        
        suggestions = []
        # Source 1: Project master table
        try:
            projects = Project.objects.exclude(
                status__in=['cancelled', 'archived']
            ).select_related('cost_center')
            if search_query:
                projects = projects.filter(
                    Q(project_name__icontains=search_query) |
                    Q(project_number__icontains=search_query) |
                    Q(cost_center__name__icontains=search_query) |
                    Q(cost_center__code__icontains=search_query) |
                    Q(cost_center__department__icontains=search_query)
                )
            
            for project in projects[:limit]:
                department = ''
                if project.cost_center:
                    department = project.cost_center.department or project.cost_center.name
                suggestions.append({
                    'project_id': str(project.id),
                    'value': f"{project.project_name} ({project.project_number})",
                    'label': f"{project.project_number} - {project.project_name}",
                    'project_number': project.project_number,
                    'department': department,
                    'source': 'master'
                })
        except Exception:
            pass
        
        # Source 2: Historical PRs
        pr_projects = PurchaseRequisition.objects.filter(
            project_department__isnull=False
        ).exclude(
            project_department__exact=''
        )
        
        if search_query:
            pr_projects = pr_projects.filter(project_department__icontains=search_query)
        
        pr_projects = pr_projects.values_list('project_department', flat=True).distinct()[:limit]
        
        for project in pr_projects:
            if project not in [s['value'] for s in suggestions]:
                suggestions.append({
                    'value': project,
                    'label': project,
                    'source': 'historical'
                })
        
        return Response({
            'count': len(suggestions),
            'suggestions': suggestions[:limit]
        })
    
    @action(detail=False, methods=['get'])
    def get_ongoing_projects(self, request):
        """Return current projects for the Purchase Recommendation selector."""
        search_query = request.query_params.get('q', '').strip()

        try:
            limit = int(request.query_params.get('limit', 100))
        except (TypeError, ValueError):
            limit = 100
        limit = max(1, min(limit, 500))

        projects = CoreProject.objects.filter(status__in=['planning', 'active'])
        if search_query:
            projects = projects.filter(
                Q(name__icontains=search_query) |
                Q(code__icontains=search_query) |
                Q(client_name__icontains=search_query)
            )

        results = []
        for project in projects.order_by('name')[:limit]:
            custom_fields = (
                project.custom_fields
                if isinstance(project.custom_fields, dict)
                else {}
            )
            department = custom_fields.get('department', '')
            results.append({
                'id': str(project.id),
                'name': project.name,
                'code': project.code,
                'department': department,
                'status': project.status,
                'value': project.name,
                'label': f'{project.code} - {project.name}',
                'project_name': project.name,
                'project_number': project.code,
                'source': 'core_project',
            })

        return Response({
            'count': len(results),
            'projects': results,
        })
    
    @action(detail=False, methods=['get'])
    def get_suppliers(self, request):
        search_query = request.query_params.get('q', '').strip()
        limit = int(request.query_params.get('limit', 50))
        
        suggestions = []
        vendors = Vendor.objects.filter(status='active')
        if search_query:
            vendors = vendors.filter(
                Q(name__icontains=search_query) |
                Q(vendor_code__icontains=search_query) |
                Q(trade_license_number__icontains=search_query) |
                Q(tax_id__icontains=search_query) |
                Q(vat_number__icontains=search_query)
            )
        
        for vendor in vendors[:limit]:
            suggestions.append({
                'vendor_id': str(vendor.id),
                'supplier_name': vendor.name,
                'supplier_business_id': (
                    vendor.trade_license_number
                    or vendor.tax_id
                    or vendor.vat_number
                    or vendor.vendor_code
                ),
                'vendor_code': vendor.vendor_code,
                'rating': vendor.rating,
                'source': 'master'
            })
        
        if len(suggestions) < limit:
            pr_suppliers = PurchaseRequisition.objects.filter(
                supplier_name__isnull=False
            ).exclude(
                supplier_name__exact=''
            )
            
            if search_query:
                pr_suppliers = pr_suppliers.filter(
                    Q(supplier_name__icontains=search_query) |
                    Q(supplier_business_id__icontains=search_query)
                )
            
            pr_suppliers = pr_suppliers.values('supplier_name', 'supplier_business_id').distinct()[:limit]
            
            for supplier in pr_suppliers:
                if supplier['supplier_name'] not in [s['supplier_name'] for s in suggestions]:
                    suggestions.append({
                        'supplier_name': supplier['supplier_name'],
                        'supplier_business_id': supplier['supplier_business_id'],
                        'source': 'historical'
                    })
        
        return Response({
            'count': len(suggestions),
            'suggestions': suggestions[:limit]
        })
    
    @action(detail=False, methods=['get'])
    def get_po_numbers(self, request):
        search_query = request.query_params.get('q', '').strip()
        limit = int(request.query_params.get('limit', 50))
        status_filter = request.query_params.get('status', None)

        suggestions = []
        pos = PurchaseOrder.objects.select_related('vendor').order_by('-created_at')

        if status_filter:
            pos = pos.filter(status=status_filter)
        else:
            # PR references may only select a PO whose lifecycle is complete.
            pos = pos.filter(status='completed')
        
        if search_query:
            pos = pos.filter(
                Q(po_number__icontains=search_query) |
                Q(vendor__name__icontains=search_query) |
                Q(title__icontains=search_query) |
                Q(description__icontains=search_query)
            )

        for po in pos[:limit]:
            suggestions.append({
                'po_number': po.po_number,
                'supplier_name': po.vendor.name,
                'total_amount': str(po.total_amount) if po.total_amount is not None else None,
                'currency': po.currency,
                'status': po.status,
            })
        
        return Response({
            'count': len(suggestions),
            'suggestions': suggestions
        })
    @action(detail=True, methods=['post'], url_path='refer-rejection')
    def refer_rejection(self, request, pk=None):
        pr = self.get_object()
        if canonicalize_pr_status(pr.status) != 'rejected':
            raise ValidationError({'error': 'Only rejected recommendations can be referred.'})
        if str(pr.issued_by_id) != str(request.user.id) and not RequisitionWorkflowService._is_super_admin(request.user):
            raise PermissionDenied('Only the issuer may refer this rejected recommendation.')

        target = str(request.data.get('target', '') or '').strip().lower()
        if target not in {'moe', 'mop'}:
            raise ValidationError({'target': 'Select MoE or MoP.'})
        remarks = str(request.data.get('remarks', '') or '').strip()
        if len(remarks) < 10:
            raise ValidationError({'remarks': 'Add at least 10 characters explaining the discussion required.'})

        pr.resolution_referral = {
            'target': target,
            'target_label': 'Manager of Engineering' if target == 'moe' else 'Manager of Projects',
            'remarks': remarks[:2000],
            'referred_by_id': str(request.user.id),
            'referred_by_name': request.user.get_full_name() or request.user.email,
            'referred_at': timezone.now().isoformat(),
            'status': 'open',
        }
        pr.save(update_fields=['resolution_referral', 'updated_at'])
        return Response(self._build_requisition_response(pr))
    
    @action(detail=False, methods=['get'], url_path='check_po_conflict')
    def check_po_conflict(self, request):
        po_number = request.query_params.get('po_number', '').strip()
        pr_id = request.query_params.get('pr_id', '').strip()
        
        if not po_number:
            return Response({
                'conflict': False,
                'message': 'No PO number provided'
            })
        
        conflicting_prs = PurchaseRequisition.objects.filter(
            po_manually_entered=po_number
        )
        
        if pr_id:
            conflicting_prs = conflicting_prs.exclude(pk=pr_id)
        
        if conflicting_prs.exists():
            first_conflict = conflicting_prs.first()
            return Response({
                'conflict': True,
                'existing_pr': first_conflict.pr_number,
                'message': f'PO number "{po_number}" is already used in PR #{first_conflict.pr_number}'
            })
        
        return Response({
            'conflict': False,
            'message': 'PO number is available'
        })
    
    @action(detail=True, methods=['post'])
    def submit(self, request, pk=None):
        pr = self.get_object()
        workflow = request.data.get('approval_workflow_config')
        if workflow is not None and canonicalize_pr_status(pr.status) == 'draft':
            self._enforce_owner_mutation(pr)
            serializer = self.get_serializer(
                pr,
                data={'approval_workflow_config': workflow},
                partial=True,
            )
            serializer.is_valid(raise_exception=True)
            serializer.save()
        pr = RequisitionWorkflowService.submit(pk, request.user)
        return Response(self._build_requisition_response(pr))

    @action(detail=True, methods=['post'])
    def convert_to_po(self, request, pk=None):
        pr, po = RequisitionConversionService.convert(pk, request.user)
        return Response({
            'message': f'Requisition converted to purchase order {po.po_number}.',
            'requisition': self._build_requisition_response(pr),
            'purchase_order': PurchaseOrderSerializer(po, context=self.get_serializer_context()).data,
        }, status=status.HTTP_201_CREATED)
    
    @action(detail=True, methods=['post'])
    def upload_attachment(self, request, pk=None):
        pr = self.get_object()
        files = request.FILES.getlist('files', [])
        
        if not files:
            return Response(
                {'error': 'No files provided'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        serializer = self.get_serializer(pr)
        uploaded = serializer._upload_attachments(pr, files)
        
        return Response({
            'message': f'{len(uploaded)} file(s) uploaded successfully',
            'attachments': pr.attachments
        })

    @action(detail=True, methods=['get'])
    def export_pdf(self, request, pk=None):
        pr = self.get_object()

        if canonicalize_pr_status(pr.status) != 'approved':
            return Response(
                {'error': 'Only approved requisitions can be exported as PDF.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        filename = build_procurement_pdf_filename(
            pr.pr_number or pr.id,
            'pr',
            pr.issued_date or timezone.localdate(),
        )

        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=letter,
            rightMargin=36,
            leftMargin=36,
            topMargin=36,
            bottomMargin=36,
        )

        story = []
        styles = getSampleStyleSheet()

        PRIMARY_BLUE = colors.HexColor('#2563EB')
        SLATE_DARK = colors.HexColor('#0F172A')
        TEXT_MUTED = colors.HexColor('#64748B')
        BG_LIGHT = colors.HexColor('#F8FAFC')
        BORDER_COLOR = colors.HexColor('#E2E8F0')
        APPROVED_GREEN = colors.HexColor('#16A34A')

        style_company = ParagraphStyle('CompanyTitle', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=18, leading=22, textColor=PRIMARY_BLUE)
        style_doc_title = ParagraphStyle('DocTitle', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=14, leading=18, textColor=SLATE_DARK, alignment=2)
        style_badge = ParagraphStyle('Badge', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=10, leading=12, textColor=APPROVED_GREEN, alignment=2)
        style_label = ParagraphStyle('Label', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=8, leading=10, textColor=TEXT_MUTED)
        style_value = ParagraphStyle('Value', parent=styles['Normal'], fontName='Helvetica', fontSize=9, leading=12, textColor=SLATE_DARK)
        style_table_header = ParagraphStyle('TableHeader', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=8, leading=10, textColor=colors.white)
        style_table_cell = ParagraphStyle('TableCell', parent=styles['Normal'], fontName='Helvetica', fontSize=8, leading=11, textColor=SLATE_DARK)

        # Header section with the controlled RADAI logo mark.
        brand_identity = Table(
            [[
                RADAILogoMark(42),
                Paragraph(
                    '<b>RADAI PROCUREMENT</b><br/>'
                    '<font size="8" color="#64748B">Oil &amp; Gas Standard Compliance</font>',
                    style_company,
                ),
            ]],
            colWidths=[0.68 * inch, 2.57 * inch],
        )
        brand_identity.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 0),
            ('RIGHTPADDING', (0, 0), (0, 0), 8),
            ('RIGHTPADDING', (1, 0), (1, 0), 0),
            ('TOPPADDING', (0, 0), (-1, -1), 0),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
        ]))
        header_data = [
            [
                brand_identity,
                [
                    Paragraph('<b>PURCHASE REQUISITION</b>', style_doc_title),
                    Paragraph('STATUS: <b>APPROVED</b>', style_badge),
                ],
            ]
        ]
        header_table = Table(header_data, colWidths=[3.25 * inch, 4.25 * inch])
        header_table.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        story.append(header_table)
        story.append(Spacer(1, 10))
        story.append(HRFlowable(width='100%', thickness=1.5, color=PRIMARY_BLUE, spaceBefore=0, spaceAfter=12))

        requester_name = pr.issued_by.get_full_name() if pr.issued_by else 'N/A'
        project_reference = pr.project or pr.project_department or 'N/A'
        department_name = pr.department or pr.project_department or 'N/A'

        meta_data = [
            [
                Paragraph('REQUISITION DETAILS', style_label),
                Paragraph('PROJECT &amp; VENDOR INFO', style_label),
            ],
            [
                Paragraph(
                    f'<b>PR Number:</b> {pr.pr_number or "N/A"}<br/>'
                    f'<b>Date Issued:</b> {pr.issued_date or "N/A"}<br/>'
                    f'<b>Issued By:</b> {requester_name}<br/>'
                    f'<b>Department:</b> {department_name}',
                    style_value,
                ),
                Paragraph(
                    f'<b>Project Reference:</b> {project_reference}<br/>'
                    f'<b>Supplier/Vendor:</b> {pr.supplier_name or "RAD Internal"}<br/>'
                    f'<b>Business ID / License:</b> {pr.supplier_business_id or "N/A"}',
                    style_value,
                ),
            ],
        ]
        meta_table = Table(meta_data, colWidths=[3.75 * inch, 3.75 * inch])
        meta_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), BG_LIGHT),
            ('BOX', (0, 0), (-1, -1), 1, BORDER_COLOR),
            ('INNERGRID', (0, 0), (-1, -1), 0.5, BORDER_COLOR),
            ('PADDING', (0, 0), (-1, -1), 8),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ]))
        story.append(meta_table)
        story.append(Spacer(1, 14))

        items_data = [[
            Paragraph('Item', style_table_header),
            Paragraph('Description', style_table_header),
            Paragraph('API / ASME Standard', style_table_header),
            Paragraph('Qty', style_table_header),
            Paragraph('UOM', style_table_header),
            Paragraph('Unit Price', style_table_header),
            Paragraph('Discount', style_table_header),
            Paragraph('Line Total', style_table_header),
        ]]

        raw_items = pr.items if isinstance(pr.items, list) else []

        if not raw_items and pr.product_service:
            raw_items = [{
                'description': pr.product_service,
                'tag': 'Standard',
                'qty': 1,
                'uom': 'LOT',
                'unit_price': float(pr.total_price or 0),
                'discount': 0,
                'total': float(pr.total_price or 0),
            }]

        def _to_float(value, default=0.0):
            try:
                return float(value)
            except (TypeError, ValueError):
                return default

        normalized_items = []
        for item in raw_items:
            description = item.get('description') or item.get('desc') or item.get('item') or 'N/A'
            standards = item.get('tag') or item.get('standard') or item.get('standards') or item.get('api_asme_standard')
            if isinstance(standards, list):
                standards = ', '.join([str(s) for s in standards if s])
            standards = standards or 'Standard'

            qty = _to_float(item.get('quantity', item.get('qty', 1)), 1)
            uom = item.get('uom') or item.get('unit_of_measure') or item.get('unit') or 'EA'
            unit_price = _to_float(item.get('unit_price', item.get('price', 0)), 0)
            discount = _to_float(item.get('discount', item.get('line_discount', 0)), 0)
            line_total = _to_float(item.get('line_total', item.get('total', (qty * unit_price) - discount)), (qty * unit_price) - discount)

            normalized_items.append({
                'description': description,
                'standards': standards,
                'qty': qty,
                'uom': uom,
                'unit_price': unit_price,
                'discount': discount,
                'line_total': line_total,
            })

        for idx, item in enumerate(normalized_items, start=1):
            items_data.append([
                Paragraph(str(idx), style_table_cell),
                Paragraph(item['description'], style_table_cell),
                Paragraph(f"<font color='#2563EB'><b>{item['standards']}</b></font>", style_table_cell),
                Paragraph(f"{item['qty']:,.2f}".rstrip('0').rstrip('.'), style_table_cell),
                Paragraph(str(item['uom']), style_table_cell),
                Paragraph(f"{item['unit_price']:,.2f}", style_table_cell),
                Paragraph(f"{item['discount']:,.2f}", style_table_cell),
                Paragraph(f"<b>{item['line_total']:,.2f}</b>", style_table_cell),
            ])

        items_table = Table(items_data, colWidths=[0.35 * inch, 2.15 * inch, 1.6 * inch, 0.5 * inch, 0.5 * inch, 0.85 * inch, 0.8 * inch, 0.9 * inch])
        items_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), PRIMARY_BLUE),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('ALIGN', (3, 1), (-1, -1), 'RIGHT'),
            ('GRID', (0, 0), (-1, -1), 0.5, BORDER_COLOR),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, BG_LIGHT]),
            ('LEFTPADDING', (0, 0), (-1, -1), 5),
            ('RIGHTPADDING', (0, 0), (-1, -1), 5),
            ('TOPPADDING', (0, 0), (-1, -1), 5),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ]))
        story.append(items_table)
        story.append(Spacer(1, 12))

        subtotal = sum(item['line_total'] for item in normalized_items)
        vat_value = _to_float(pr.total_price, subtotal) - _to_float(pr.net_total_excl_vat, subtotal)
        if vat_value == 0:
            vat_value = subtotal * 0.05
        grand_total = _to_float(pr.total_price, subtotal + vat_value)
        currency = pr.currency or 'AED'

        totals_data = [
            [Paragraph('Subtotal (excl. VAT):', style_label), Paragraph(f"{subtotal:,.2f} {currency}", style_value)],
            [Paragraph('VAT:', style_label), Paragraph(f"{vat_value:,.2f} {currency}", style_value)],
            [Paragraph('<b>Grand Total:</b>', style_label), Paragraph(f"<b>{grand_total:,.2f} {currency}</b>", style_value)],
        ]
        totals_table = Table(totals_data, colWidths=[2.0 * inch, 1.5 * inch])
        totals_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
            ('LINEABOVE', (0, 2), (-1, 2), 1, PRIMARY_BLUE),
            ('PADDING', (0, 0), (-1, -1), 4),
        ]))

        story.append(Table([["", totals_table]], colWidths=[3.9 * inch, 3.6 * inch]))
        story.append(Spacer(1, 12))

        pm_name = pr.pm_name.get_full_name() if pr.pm_name else 'N/A'
        vp_name = pr.vp_op_name.get_full_name() if pr.vp_op_name else 'N/A'
        pm_date = pr.pm_approved_at.strftime('%Y-%m-%d %H:%M UTC') if pr.pm_approved_at else 'N/A'
        vp_date = pr.vp_op_approved_at.strftime('%Y-%m-%d %H:%M UTC') if pr.vp_op_approved_at else 'N/A'

        approval_data = [
            [
                Paragraph('<b>PROJECT MANAGER SIGN-OFF</b>', style_label),
                Paragraph('<b>EXECUTIVE / VP SIGN-OFF</b>', style_label),
            ],
            [
                Paragraph(
                    f'<b>Approver:</b> {pm_name}<br/>'
                    f'<b>Status:</b> <font color="#16A34A"><b>{pr.pm_approval_status.upper()}</b></font><br/>'
                    f'<b>Date:</b> {pm_date}<br/>'
                    f'<i>Comments: Technical and standards verification completed.</i>',
                    style_value,
                ),
                Paragraph(
                    f'<b>Approver:</b> {vp_name}<br/>'
                    f'<b>Status:</b> <font color="#16A34A"><b>{pr.vp_op_approval_status.upper()}</b></font><br/>'
                    f'<b>Date:</b> {vp_date}<br/>'
                    f'<i>Comments: Commercial terms and budget approved.</i>',
                    style_value,
                ),
            ],
        ]
        approval_table = Table(approval_data, colWidths=[3.75 * inch, 3.75 * inch])
        approval_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#F1F5F9')),
            ('BOX', (0, 0), (-1, -1), 1, BORDER_COLOR),
            ('INNERGRID', (0, 0), (-1, -1), 0.5, BORDER_COLOR),
            ('PADDING', (0, 0), (-1, -1), 8),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ]))
        story.append(approval_table)

        doc.build(story)
        buffer.seek(0)
        response.write(buffer.read())
        return response
    
    @action(detail=False, methods=['get'])
    def dashboard(self, request):
        """Return counts for the canonical PR lifecycle."""
        queryset = self.get_queryset()

        totals = {
            lifecycle_status: queryset.filter(
                status__in=stored_values_for(lifecycle_status)
            ).count()
            for lifecycle_status in (
                'draft', 'submitted', 'in_review', 'approved',
                'rejected', 'cancelled', 'converted',
            )
        }
        totals['total'] = queryset.count()
        totals['pending'] = totals['submitted'] + totals['in_review']
        
        by_priority = queryset.values('priority').annotate(
            count=Count('id')
        )
        
        recent = queryset[:10]
        recent_serializer = self.get_serializer(recent, many=True)
        
        return Response({
            'totals': {
                **totals,
            },
            'by_priority': list(by_priority),
            'recent': recent_serializer.data
        })
    
    @action(detail=True, methods=['post'])
    def approve(self, request, pk=None):
        pr = RequisitionWorkflowService.approve(
            pk,
            request.user,
            signature=request.data.get('signature', ''),
        )
        return Response(self._build_requisition_response(pr))
    
    @action(detail=True, methods=['post'])
    def reject(self, request, pk=None):
        pr = RequisitionWorkflowService.reject(
            pk,
            request.user,
            request.data.get('reason', ''),
        )
        return Response(self._build_requisition_response(pr))
    @action(detail=False, methods=['get'], url_path='export-to-excel')
    def export_to_excel(self, request):
        import io
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from django.http import HttpResponse
        
        queryset = self.get_queryset()
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Purchase Requisitions"
        
        headers = [
            'PR No.', 'PO Number', 'Description', 'Project Details',
            'Supplier Name', 'PR Value', 'Currency', 'Approval Status',
            'Priority', 'Created Date', 'Required Date'
        ]
        
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        for row_num, pr in enumerate(queryset, 2):
            approval_status = pr.get_detailed_approval_status()
            
            ws.cell(row=row_num, column=1, value=pr.pr_number)
            ws.cell(row=row_num, column=2, value=pr.po_manually_entered or pr.po_number_reference or 'N/A')
            ws.cell(row=row_num, column=3, value=pr.product_service or pr.description_reason)
            ws.cell(row=row_num, column=4, value=pr.project_department or pr.project)
            ws.cell(row=row_num, column=5, value=pr.supplier_name or 'N/A')
            ws.cell(row=row_num, column=6, value=float(pr.total_price) if pr.total_price else 0)
            ws.cell(row=row_num, column=7, value=pr.currency)
            ws.cell(row=row_num, column=8, value=approval_status['display'])
            ws.cell(row=row_num, column=9, value=pr.get_priority_display())
            ws.cell(row=row_num, column=10, value=pr.created_at.strftime('%Y-%m-%d') if pr.created_at else '')
            ws.cell(row=row_num, column=11, value=pr.required_date.strftime('%Y-%m-%d') if pr.required_date else '')
        
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=purchase_requisitions.xlsx'
        
        return response
    
    @action(detail=False, methods=['get'], url_path='pending-approvals-dashboard')
    def pending_approvals_dashboard(self, request):
        pending_prs = self.get_queryset().filter(
            status__in=['submitted', 'pm_approved', 'vp_approved']
        )
        
        dashboard_data = {
            'total_pending': pending_prs.count(),
            'under_review': [],
            'overdue': [],
            'by_level': {},
            'by_priority': {},
            'by_department': {},
            'recent_rejections': [],
            'escalated_rejections': []
        }
        
        for pr in pending_prs:
            detailed_status = pr.get_detailed_approval_status()
            
            pr_data = {
                'id': str(pr.id),
                'pr_number': pr.pr_number,
                'po_number': pr.po_manually_entered or pr.po_number_reference or 'N/A',
                'description': pr.product_service or pr.description_reason,
                'project': pr.project_department or pr.project,
                'supplier': pr.supplier_name,
                'value': float(pr.total_price) if pr.total_price else 0,
                'currency': pr.currency,
                'priority': pr.priority,
                'status': detailed_status['status'],
                'status_display': detailed_status['display'],
                'is_overdue': detailed_status.get('is_overdue', False),
                'days_pending': detailed_status.get('days_pending', 0),
                'current_level': pr.current_approval_level,
                'created_at': pr.created_at.isoformat() if pr.created_at else None
            }
            
            if detailed_status.get('is_overdue'):
                dashboard_data['overdue'].append(pr_data)
            else:
                dashboard_data['under_review'].append(pr_data)
            
            level_key = f"Level {pr.current_approval_level}"
            dashboard_data['by_level'][level_key] = dashboard_data['by_level'].get(level_key, 0) + 1
            dashboard_data['by_priority'][pr.priority] = dashboard_data['by_priority'].get(pr.priority, 0) + 1
            
            dept = pr.department or 'Unknown'
            dashboard_data['by_department'][dept] = dashboard_data['by_department'].get(dept, 0) + 1
        
        recent_rejections = self.get_queryset().filter(status='rejected').order_by('-updated_at')[:10]
        dashboard_data['recent_rejections'] = [
            {
                'id': str(pr.id),
                'pr_number': pr.pr_number,
                'description': pr.product_service or pr.description_reason,
                'rejection_reason': pr.rejection_reason,
                'rejected_at': pr.updated_at.isoformat() if pr.updated_at else None,
                'escalated': pr.rejection_escalated_to_moe or pr.rejection_escalated_to_mop
            }
            for pr in recent_rejections
        ]
        
        escalated = self.get_queryset().filter(
            Q(rejection_escalated_to_moe=True) | Q(rejection_escalated_to_mop=True)
        ).order_by('-escalation_initiated_at')
        dashboard_data['escalated_rejections'] = [
            {
                'id': str(pr.id),
                'pr_number': pr.pr_number,
                'description': pr.product_service or pr.description_reason,
                'escalated_to': 'MoE' if pr.rejection_escalated_to_moe else 'MoP',
                'escalation_notes': pr.escalation_notes,
                'resolved': pr.escalation_resolved,
                'escalated_at': pr.escalation_initiated_at.isoformat() if pr.escalation_initiated_at else None
            }
            for pr in escalated
        ]
        
        dashboard_data['summary'] = {
            'total_pending': dashboard_data['total_pending'],
            'under_review_count': len(dashboard_data['under_review']),
            'overdue_count': len(dashboard_data['overdue']),
            'urgent_count': dashboard_data['by_priority'].get('urgent', 0),
            'escalated_count': len(dashboard_data['escalated_rejections'])
        }
        
        return Response(dashboard_data)
    
    @action(detail=True, methods=['post'], url_path='escalate-rejection')
    def escalate_rejection(self, request, pk=None):
        pr = self.get_object()
        
        if pr.status != 'rejected':
            return Response(
                {'error': 'Only rejected PRs can be escalated'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        escalate_to = request.data.get('escalate_to', '').lower()
        escalation_notes = request.data.get('escalation_notes', '')
        
        if escalate_to not in ['moe', 'mop']:
            return Response(
                {'error': 'escalate_to must be "moe" or "mop"'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        if not escalation_notes:
            return Response(
                {'error': 'escalation_notes is required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        if escalate_to == 'moe':
            pr.rejection_escalated_to_moe = True
        else:
            pr.rejection_escalated_to_mop = True
        
        pr.escalation_notes = escalation_notes
        pr.escalation_initiated_by = request.user
        pr.escalation_initiated_at = timezone.now()
        pr.escalation_resolved = False
        pr.save()
        
        serializer = self.get_serializer(pr)
        return Response({
            'message': f'PR escalated to {escalate_to.upper()} successfully',
            'data': serializer.data
        })


class PurchaseOrderViewSet(viewsets.ModelViewSet):
    queryset = PurchaseOrder.objects.all().select_related(
        'vendor', 'pr_reference', 'project', 'enterprise_project'
    ).prefetch_related('receipts').order_by('-created_at', '-id')
    serializer_class = PurchaseOrderSerializer
    permission_classes = [IsAuthenticated, HasModuleAccess]
    module_required = 'procurement_orders'
    pagination_class = VendorPagination
    parser_classes = [FormParser, MultiPartParser, JSONParser]

    def create(self, request, *args, **kwargs):
        """Create a PO and pass repeated multipart attachment fields intact."""
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        files = request.FILES.getlist('attachments_files', [])
        serializer.save(**({'attachments_files': files} if files else {}))
        headers = self.get_success_headers(serializer.data)
        return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)

    def update(self, request, *args, **kwargs):
        """Update a PO with optional files from its dedicated Attachments tab."""
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        files = request.FILES.getlist('attachments_files', [])
        serializer.save(**({'attachments_files': files} if files else {}))
        return Response(serializer.data)

    def get_permissions(self):
        # Any active employee may be selected as a PO approver. Assignment is
        # enforced inside the approval actions, so procurement module access is
        # not required merely to view or action that employee's own queue.
        # Similarly, users with procurement_requisitions access should be able
        # to create and update purchase orders (often converted from PRs).
        # Helper actions like available-requisitions, available-projects, and
        # create-project are also needed during PO creation.
        if self.action in {
            'pending_for_me', 'approve', 'reject',
            'create', 'update', 'partial_update',
            'available_requisitions', 'available_projects', 'create_project',
            'reserve_number',
        }:
            return [IsAuthenticated()]
        return super().get_permissions()
    
    def get_queryset(self):
        queryset = super().get_queryset()
        
        status_filter = self.request.query_params.get('status', None)
        if status_filter:
            queryset = queryset.filter(status=status_filter)
        
        vendor_id = self.request.query_params.get('vendor', None)
        if vendor_id:
            queryset = queryset.filter(vendor_id=vendor_id)
        
        search = self.request.query_params.get('search', None)
        if search:
            queryset = queryset.filter(
                Q(po_number__icontains=search) |
                Q(title__icontains=search) |
                Q(vendor__name__icontains=search)
            )
        
        return queryset.order_by('-created_at', '-id')

    @action(detail=False, methods=['get'], url_path='pending-for-me')
    def pending_for_me(self, request):
        queue = pending_entries_for(
            request.user,
            self.get_queryset().select_related('created_by'),
        )
        if str(request.query_params.get('count_only') or '').lower() == 'true':
            return Response({'count': len(queue), 'pending_count': len(queue), 'results': []})

        results = []
        for order, index, entry in queue[:100]:
            payload = self.get_serializer(order).data
            payload.update({
                'approval_queue_id': f'{order.id}:{index}',
                'approval_stage': entry.get('stage'),
                'approval_status': entry.get('status'),
                'approver_name': entry.get('approver'),
            })
            results.append(payload)
        return Response({'count': len(queue), 'pending_count': len(queue), 'results': results})

    def _record_approval_decision(self, request, decision):
        order = self.get_object()
        note = request.data.get('note') or request.data.get('reason') or ''
        stage = request.data.get('approval_stage') or ''
        updated, entry = record_decision(order, request.user, decision, stage=stage, comment=note)
        return Response({
            'message': f"{entry['stage']} {entry['status'].lower()} successfully.",
            'approval': entry,
            'purchase_order': self.get_serializer(updated).data,
        })

    @action(detail=True, methods=['post'])
    def approve(self, request, pk=None):
        return self._record_approval_decision(request, 'approve')

    @action(detail=True, methods=['post'])
    def reject(self, request, pk=None):
        return self._record_approval_decision(request, 'reject')

    @action(detail=False, methods=['get'], url_path='available-requisitions')
    def available_requisitions(self, request):
        """Existing PRs available for Purchase Order linkage and search."""
        queryset = PurchaseRequisition.objects.select_related(
            'issued_by', 'vendor', 'requested_by', 'approved_by', 'pm_name',
            'eng_manager_name', 'manager_projects_name', 'vp_op_name',
        ).all().order_by('-created_at')

        search = str(request.query_params.get('search') or '').strip()
        if search:
            queryset = queryset.filter(
                Q(pr_number__icontains=search)
                | Q(title__icontains=search)
                | Q(product_service__icontains=search)
                | Q(supplier_name__icontains=search)
                | Q(project_department__icontains=search)
            )

        serializer = PurchaseRequisitionSerializer(
            queryset,
            many=True,
            context=self.get_serializer_context(),
        )
        return Response(serializer.data)

    @action(detail=False, methods=['get'], url_path='available-projects')
    def available_projects(self, request):
        """Unified procurement-master and company-core projects for PO linkage."""
        queryset = Project.objects.select_related(
            'cost_center', 'project_manager', 'lead_engineer'
        ).prefetch_related('team_members', 'budgets').order_by('project_number', 'project_name')
        core_queryset = CoreProject.objects.all().order_by('code', 'name')

        search = str(request.query_params.get('search') or '').strip()
        if search:
            queryset = queryset.filter(
                Q(project_number__icontains=search)
                | Q(project_name__icontains=search)
            )
            core_queryset = core_queryset.filter(
                Q(code__icontains=search) | Q(name__icontains=search)
            )

        master_projects = ProjectListSerializer(
            queryset,
            many=True,
            context=self.get_serializer_context(),
        ).data
        for project in master_projects:
            project['source'] = 'procurement_master'

        master_numbers = {
            str(project['project_number']).strip().lower()
            for project in master_projects
        }
        core_projects = [
            {
                'id': f'core:{project.id}',
                'source_project_id': str(project.id),
                'source': 'core',
                'project_number': project.code,
                'project_name': project.name,
                'status': project.status,
                'status_display': project.get_status_display(),
                'client_name': project.client_name,
                'is_active': project.status not in {'cancelled'},
            }
            for project in core_queryset
            if str(project.code).strip().lower() not in master_numbers
        ]
        combined = [*master_projects, *core_projects]
        combined.sort(key=lambda project: (
            str(project.get('project_number') or '').lower(),
            str(project.get('project_name') or '').lower(),
        ))
        return Response(combined)

    @action(detail=False, methods=['post'], url_path='create-project')
    def create_project(self, request):
        """Create a minimal master project while preparing a Purchase Order."""
        source_project_id = request.data.get('source_project_id')
        source_project = None
        if source_project_id:
            source_project = CoreProject.objects.filter(pk=source_project_id).first()
            if source_project is None:
                return Response(
                    {'source_project_id': 'The selected company project was not found.'},
                    status=status.HTTP_404_NOT_FOUND,
                )

        project_number = str(
            source_project.code if source_project else request.data.get('project_number') or ''
        ).strip()
        project_name = str(
            source_project.name if source_project else request.data.get('project_name') or request.data.get('title') or ''
        ).strip()
        errors = {}
        if not project_number:
            errors['project_number'] = 'Project number is required.'
        if not project_name:
            errors['project_name'] = 'Project title is required.'
        if errors:
            return Response(errors, status=status.HTTP_400_BAD_REQUEST)

        existing = Project.objects.filter(project_number__iexact=project_number).first()
        if existing:
            if source_project:
                payload = ProjectListSerializer(existing, context=self.get_serializer_context()).data
                payload['source'] = 'procurement_master'
                return Response(payload, status=status.HTTP_200_OK)
            return Response(
                {'project_number': f'Project number {existing.project_number} already exists.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        project_data = {
            'project_number': project_number,
            'project_name': project_name,
            'enterprise_project': source_project.pk if source_project else None,
        }
        if source_project:
            project_data.update({
                'client_name': source_project.client_name,
                'status': source_project.status if source_project.status in {
                    'planning', 'active', 'on_hold', 'completed', 'cancelled'
                } else 'planning',
            })
        serializer = ProjectListSerializer(
            data=project_data,
            context=self.get_serializer_context(),
        )
        serializer.is_valid(raise_exception=True)
        project = serializer.save()
        payload = ProjectListSerializer(project, context=self.get_serializer_context()).data
        payload['source'] = 'procurement_master'
        return Response(payload, status=status.HTTP_201_CREATED)

    @action(detail=False, methods=['post'], url_path='reserve-number')
    def reserve_number(self, request):
        """Reserve the next authoritative PO number for a selected PR."""
        pr_id = request.data.get('pr_reference')
        if not pr_id:
            return Response(
                {'pr_reference': 'Select an existing Purchase Requisition.'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        try:
            requisition = PurchaseRequisition.objects.get(pk=pr_id)
            po_number = PurchaseOrderNumberService.next_for_requisition(requisition.pr_number)
        except PurchaseRequisition.DoesNotExist:
            return Response(
                {'pr_reference': 'The selected Purchase Requisition was not found.'},
                status=status.HTTP_404_NOT_FOUND,
            )
        except ValueError as exc:
            return Response({'pr_reference': str(exc)}, status=status.HTTP_400_BAD_REQUEST)

        return Response({
            'po_number': po_number,
            'pr_reference': str(requisition.id),
            'pr_number': requisition.pr_number,
        })

    @action(
        detail=False,
        methods=['post'],
        url_path='import-excel',
        parser_classes=[MultiPartParser, FormParser],
    )
    def import_excel(self, request):
        """Preview or upsert the authoritative PO register with PR linkage."""
        upload = request.FILES.get('file')
        if not upload:
            return Response({'error': 'Select an Excel file to import.'}, status=status.HTTP_400_BAD_REQUEST)
        dry_run = str(request.data.get('dry_run', 'true')).strip().lower() not in {'false', '0', 'no'}
        from .services.po_excel_import import POExcelImportError, import_po_workbook
        try:
            result = import_po_workbook(upload, user=request.user, dry_run=dry_run)
        except POExcelImportError as exc:
            return Response({'error': str(exc)}, status=status.HTTP_400_BAD_REQUEST)
        return Response(result, status=status.HTTP_200_OK)
    
    @action(detail=True, methods=['post'])
    def send_to_vendor(self, request, pk=None):
        po = self.get_object()
        
        if po.status != 'draft':
            return Response(
                {'error': 'Only draft POs can be sent'},
                status=status.HTTP_400_BAD_REQUEST
            )

        verified, message = PurchaseOrderNumberService.verify(
            po.po_number,
            po.pr_reference.pr_number if po.pr_reference_id else None,
        )
        if not verified:
            return Response(
                {'error': f'PO number verification failed: {message}'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        
        po.status = 'sent'
        po.save()
        
        serializer = self.get_serializer(po)
        return Response(serializer.data)
    
    @action(detail=True, methods=['post'])
    def acknowledge(self, request, pk=None):
        po = self.get_object()
        
        if po.status != 'sent':
            return Response(
                {'error': 'Only sent POs can be acknowledged'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        po.status = 'acknowledged'
        po.save()
        
        serializer = self.get_serializer(po)
        return Response(serializer.data)
    @action(detail=True, methods=['get'], url_path='receiving-summary')
    def receiving_summary(self, request, pk=None):
        po = self.get_object()
        return Response(GoodsReceiptService.receiving_summary(po))
    
    @action(detail=False, methods=['get'])
    def dashboard(self, request):
        total = self.get_queryset().count()
        draft = self.get_queryset().filter(status='draft').count()
        sent = self.get_queryset().filter(status='sent').count()
        acknowledged = self.get_queryset().filter(status='acknowledged').count()
        completed = self.get_queryset().filter(status='completed').count()
        
        total_value = self.get_queryset().aggregate(
            total=Sum('total_amount')
        )['total'] or 0
        
        by_vendor = self.get_queryset().values('vendor__name').annotate(
            count=Count('id'),
            total=Sum('total_amount')
        ).order_by('-total')[:5]
        
        recent = self.get_queryset()[:5]
        recent_serializer = self.get_serializer(recent, many=True)
        
        return Response({
            'totals': {
                'total': total,
                'draft': draft,
                'sent': sent,
                'acknowledged': acknowledged,
                'completed': completed,
                'total_value': float(total_value)
            },
            'top_vendors': list(by_vendor),
            'recent': recent_serializer.data
        })
    @action(detail=False, methods=['get'], url_path='by_number/(?P<po_number>[^/.]+)')
    def by_number(self, request, po_number=None):
        if not po_number:
            return Response(
                {'error': 'PO number is required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            po = PurchaseOrder.objects.select_related('vendor').get(po_number=po_number)
            return Response({
                'id': str(po.id),
                'po_number': po.po_number,
                'status': po.status,
                'vendor_name': po.vendor.name if po.vendor else None,
                'vendor_id': str(po.vendor.id) if po.vendor else None,
                'total_amount': str(po.total_amount) if po.total_amount else '0',
                'currency': po.currency,
                'title': po.title,
                'description': po.description,
                'created_at': po.created_at.isoformat() if po.created_at else None
            })
        except PurchaseOrder.DoesNotExist:
            return Response(
                {'error': f'Purchase Order with number "{po_number}" not found'},
                status=status.HTTP_404_NOT_FOUND
            )


class ReceiptViewSet(viewsets.ModelViewSet):
    """Goods Receipt management secured by procurement receipt access."""

    queryset = Receipt.objects.all().select_related(
        'purchase_order', 'received_by'
    ).order_by('-created_at')
    serializer_class = ReceiptSerializer
    permission_classes = [IsAuthenticated, HasModuleAccess]
    module_required = 'procurement_receipts'
    http_method_names = ['get', 'post', 'patch', 'head', 'options']
    
    def get_queryset(self):
        queryset = super().get_queryset()
        
        status_filter = self.request.query_params.get('status', None)
        if status_filter:
            queryset = queryset.filter(status=status_filter)
        
        quality_filter = self.request.query_params.get('quality_check', None)
        if quality_filter == 'passed':
            queryset = queryset.filter(quality_check_passed=True)
        elif quality_filter == 'failed':
            queryset = queryset.filter(quality_check_passed=False)
        
        search = self.request.query_params.get('search', None)
        if search:
            queryset = queryset.filter(
                Q(receipt_number__icontains=search) |
                Q(purchase_order__po_number__icontains=search) |
                Q(delivery_note_number__icontains=search)
            )
        
        return queryset
    
    @action(detail=True, methods=['post'])
    def submit(self, request, pk=None):
        receipt = self.get_object()
        response_data = dict(self.get_serializer(receipt).data)
        response_data.update({
            'database_confirmed': True,
            'message': (
                f'Goods Receipt {receipt.receipt_number} was saved in the database '
                'and submitted successfully.'
            ),
        })
        return Response(response_data)

    @action(detail=True, methods=['post'])
    def accept(self, request, pk=None):
        receipt = self.get_object()
        if receipt.status != 'pending':
            raise ValidationError({'error': 'Only pending receipts can be accepted.'})
        receipt.status = 'accepted'
        receipt.quality_check_passed = True
        receipt.save(update_fields=['status', 'quality_check_passed', 'updated_at'])
        po = receipt.purchase_order
        po.status = 'completed'
        po.actual_delivery = timezone.localdate()
        po.save(update_fields=['status', 'actual_delivery', 'updated_at'])
        return Response(self.get_serializer(receipt).data)
    
    @action(detail=True, methods=['post'])
    def reject_delivery(self, request, pk=None):
        receipt = self.get_object()
        if receipt.status != 'pending':
            raise ValidationError({'error': 'Only pending receipts can be rejected.'})
        receipt.status = 'rejected'
        receipt.quality_check_passed = False
        receipt.inspection_notes = request.data.get('reason') or request.data.get('notes') or ''
        receipt.save(update_fields=['status', 'quality_check_passed', 'inspection_notes', 'updated_at'])
        return Response(self.get_serializer(receipt).data)

    @action(detail=True, methods=['post'])
    def cancel(self, request, pk=None):
        raise ValidationError({'error': 'Cancellation is not available for this Goods Receipt model.'})
    
    @action(detail=False, methods=['get'])
    def dashboard(self, request):
        total = self.get_queryset().count()
        pending = self.get_queryset().filter(status='pending').count()
        accepted = self.get_queryset().filter(status='accepted').count()
        rejected = self.get_queryset().filter(status='rejected').count()
        
        quality_passed = self.get_queryset().filter(quality_check_passed=True).count()
        quality_failed = self.get_queryset().filter(quality_check_passed=False).count()
        
        recent = self.get_queryset()[:5]
        recent_serializer = self.get_serializer(recent, many=True)
        
        return Response({
            'totals': {
                'total': total,
                'pending': pending,
                'accepted': accepted,
                'rejected': rejected,
                'quality_passed': quality_passed,
                'quality_failed': quality_failed
            },
            'recent': recent_serializer.data
        })


@action(detail=False, methods=['get'])
def get_categories(request):
    categories = [
        {'code': code, **data}
        for code, data in PROCUREMENT_CATEGORIES.items()
    ]
    serializer = ProcurementCategorySerializer(categories, many=True)
    return Response(serializer.data)


class PODocumentViewSet(viewsets.ReadOnlyModelViewSet):
    """
    ViewSet for PODocument — read-only list/detail plus the AI extraction action.
    The `extract_from_pdf` action is the primary entry point: it accepts a PDF
    upload, stores it in S3, runs the AI extractor, and returns the result.
    
    🔒 SECURITY: Requires 'procurement_orders' module access (soft-coded from rbac_config.py)
    """

    queryset = PODocument.objects.all().order_by('-created_at')
    serializer_class = PODocumentSerializer
    permission_classes = [IsAuthenticated, HasModuleAccess]
    module_required = 'procurement_orders'
    parser_classes = [MultiPartParser, FormParser]

    def get_queryset(self):
        return super().get_queryset().filter(uploaded_by=self.request.user)

    @action(detail=False, methods=['post'], url_path='extract_from_pdf',
            parser_classes=[MultiPartParser, FormParser])
    def extract_from_pdf(self, request):
        """
        POST /api/v1/procurement/po-documents/extract_from_pdf/

        Multipart form fields:
            file  G�� PDF file (required)

        Returns structured JSON with all extractable PO/PR fields plus
        S3 storage reference and extraction metadata.
        """
        import logging
        logger = logging.getLogger(__name__)

        pdf_file = request.FILES.get('file')
        if not pdf_file:
            return Response(
                {'error': 'No file provided. Send a PDF as multipart/form-data field "file".'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        allowed_types = ['application/pdf', 'application/x-pdf']
        content_type = getattr(pdf_file, 'content_type', '') or ''
        if content_type not in allowed_types and not pdf_file.name.lower().endswith('.pdf'):
            return Response(
                {'error': 'Only PDF files are supported.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        doc = PODocument.objects.create(
            original_filename=pdf_file.name,
            file_size_bytes=pdf_file.size,
            extraction_status='processing',
            uploaded_by=request.user,
        )

        try:
            from .services.po_ai_extractor import extract_po_from_pdf

            pdf_bytes = pdf_file.read()
            result = extract_po_from_pdf(
                pdf_bytes=pdf_bytes,
                original_filename=pdf_file.name,
                user_id=str(request.user.id),
            )

            doc.s3_key = result.get('s3_key', '')
            doc.s3_url = result.get('s3_url', '')
            doc.extraction_status = result.get('extraction_status', 'failed')
            doc.extraction_error = result.get('error', '') or ''

            extracted = result.get('extracted_data', {})
            doc.extracted_data = extracted
            doc.document_type = extracted.get('document_type', 'unknown')
            doc.save()

            if not result.get('success'):
                return Response(
                    {
                        'success': False,
                        'document_id': str(doc.id),
                        'error': result.get('error', 'Extraction failed'),
                    },
                    status=status.HTTP_422_UNPROCESSABLE_ENTITY,
                )

            return Response(
                {
                    'success': True,
                    'document_id': str(doc.id),
                    's3_key': doc.s3_key,
                    's3_url': doc.s3_url,
                    'extraction_status': doc.extraction_status,
                    'document_type': doc.document_type,
                    'raw_text_length': result.get('raw_text_length', 0),
                    'extracted_data': extracted,
                },
                status=status.HTTP_200_OK,
            )

        except Exception as exc:
            logger.exception('[PODocumentViewSet] Unhandled extraction error')
            doc.extraction_status = 'failed'
            doc.extraction_error = str(exc)
            doc.save()
            return Response(
                {'success': False, 'document_id': str(doc.id), 'error': str(exc)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

    @action(detail=False, methods=['post'], url_path='import_signed_pdf',
            parser_classes=[MultiPartParser, FormParser])
    def import_signed_pdf(self, request):
        """Import, reconcile, persist, and verify a signed Purchase Order PDF."""
        pdf_file = request.FILES.get('file')
        if not pdf_file:
            return Response({'error': 'Select a signed PDF file to import.'}, status=status.HTTP_400_BAD_REQUEST)
        if pdf_file.size > 15 * 1024 * 1024:
            return Response({'error': 'PDF file must not exceed 15 MB.'}, status=status.HTTP_400_BAD_REQUEST)
        from .services.signed_po_pdf_import import SignedPOImportError, import_signed_po_pdf
        try:
            result = import_signed_po_pdf(
                pdf_file.read(),
                filename=pdf_file.name,
                user=request.user,
                signature_verified=str(request.data.get('signature_verified', 'false')).lower() == 'true',
                stamp_verified=str(request.data.get('stamp_verified', 'false')).lower() == 'true',
                approved_by_name=request.data.get('approved_by_name', ''),
                approved_by_title=request.data.get('approved_by_title', ''),
                approved_date=request.data.get('approved_date', ''),
            )
        except SignedPOImportError as exc:
            return Response({'error': str(exc)}, status=status.HTTP_400_BAD_REQUEST)
        return Response(result, status=status.HTTP_200_OK)

    @action(detail=True, methods=['post'], url_path='confirm_po')
    def confirm_po(self, request, pk=None):
        doc = self.get_object()
        po_id = request.data.get('purchase_order_id')
        if not po_id:
            return Response({'error': 'purchase_order_id required'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            po = PurchaseOrder.objects.get(pk=po_id)
        except PurchaseOrder.DoesNotExist:
            return Response({'error': 'PurchaseOrder not found'}, status=status.HTTP_404_NOT_FOUND)

        doc.confirmed_po = po
        doc.save(update_fields=['confirmed_po'])
        return Response({'success': True, 'document_id': str(doc.id), 'purchase_order_id': str(po.id)})


class CostCenterViewSet(viewsets.ModelViewSet):
    """
    Cost Center API - Master organizational cost center registry.
    Soft-coded for departmental budget tracking and financial reporting.
    
    =��� SECURITY: Requires 'procurement' module access (soft-coded from rbac_config.py)
    """
    queryset = CostCenter.objects.all()
    serializer_class = CostCenterSerializer
    permission_classes = [IsAuthenticated, HasModuleAccess]
    module_required = 'procurement'
    filterset_fields = ['department', 'division', 'is_active']
    search_fields = ['code', 'name', 'department', 'division']
    ordering_fields = ['code', 'name', 'department', 'created_at']
    ordering = ['code']
    
    def get_queryset(self):
        queryset = super().get_queryset()
        
        if self.request.query_params.get('active_only') == 'true':
            queryset = queryset.filter(is_active=True)
        
        parent_id = self.request.query_params.get('parent')
        if parent_id:
            queryset = queryset.filter(parent_id=parent_id)
        
        return queryset.select_related('parent', 'manager')


class BudgetViewSet(viewsets.ModelViewSet):
    """
    Budget Allocation API - Project budget lines with spend tracking.
    Soft-coded for professional financial control and variance analysis.
    
    =��� SECURITY: Requires 'procurement' module access (soft-coded from rbac_config.py)
    """
    queryset = Budget.objects.all()
    serializer_class = BudgetSerializer
    permission_classes = [IsAuthenticated, HasModuleAccess]
    module_required = 'procurement'
    filterset_fields = ['project', 'category', 'fiscal_year', 'is_approved']
    search_fields = ['project__project_number', 'project__project_name', 'description']
    ordering_fields = ['category', 'allocated_amount', 'fiscal_year', 'created_at']
    ordering = ['project', 'category']
    
    def get_queryset(self):
        queryset = super().get_queryset()
        
        if self.request.query_params.get('approved_only') == 'true':
            queryset = queryset.filter(is_approved=True)
        
        if self.request.query_params.get('current_year') == 'true':
            from django.utils import timezone
            current_year = timezone.now().year
            queryset = queryset.filter(fiscal_year=current_year)
        
        return queryset.select_related('project', 'cost_center', 'approved_by')
    
    @action(detail=True, methods=['post'])
    def approve(self, request, pk=None):
        budget = self.get_object()
        budget.is_approved = True
        budget.approved_by = request.user
        budget.approved_at = timezone.now()
        budget.save(update_fields=['is_approved', 'approved_by', 'approved_at'])
        return Response({'success': True, 'budget_id': str(budget.id)})


class ProjectViewSet(viewsets.ModelViewSet):
    """
    Project Master Registry API - Central project database.
    Professional project-based procurement with budget tracking.
    
    Features:
      - List/Detail views with different serializers
      - Project-based PO aggregation
      - Budget tracking and variance analysis
      - Invoice reconciliation (A/P + A/R)
      - Soft-coded filtering and search
    
    =��� SECURITY: Requires 'procurement' module access (soft-coded from rbac_config.py)
    """
    queryset = Project.objects.all()
    permission_classes = [IsAuthenticated, HasModuleAccess]
    module_required = 'procurement'
    filterset_fields = ['status', 'project_type', 'is_active', 'is_billable', 'health_status']
    search_fields = ['project_number', 'project_name', 'client_name', 'description']
    ordering_fields = ['project_number', 'project_name', 'start_date', 'created_at', 'status']
    ordering = ['-created_at']
    
    def get_serializer_class(self):
        if self.action in {'retrieve', 'create', 'update', 'partial_update'}:
            return ProjectDetailSerializer
        return ProjectListSerializer
    
    def get_queryset(self):
        queryset = super().get_queryset()
        
        if self.request.query_params.get('active_only', 'true') == 'true':
            queryset = queryset.filter(is_active=True)
        
        if self.request.query_params.get('my_projects') == 'true':
            queryset = queryset.filter(
                Q(project_manager=self.request.user) |
                Q(lead_engineer=self.request.user) |
                Q(team_members=self.request.user)
            ).distinct()
        
        health = self.request.query_params.get('health')
        if health:
            queryset = queryset.filter(health_status=health)
        
        return queryset.select_related(
            'cost_center', 'project_manager', 'lead_engineer', 'enterprise_project'
        ).prefetch_related('team_members', 'budgets')

    @action(
        detail=False, methods=['get'], url_path='relationship-report',
        permission_classes=[IsAuthenticated, CommercialModulePermission],
    )
    def relationship_report(self, request):
        """Preview unresolved canonical links without mutating project data."""
        return Response(build_project_reconciliation_payload())

    @action(
        detail=False, methods=['post'], url_path='resolve-relationship',
        permission_classes=[IsAuthenticated, CommercialModulePermission],
    )
    def resolve_relationship(self, request):
        """Manually assign a canonical project with an immutable audit trail."""
        record_type = request.data.get('record_type')
        record_id = request.data.get('record_id')
        enterprise_project_id = request.data.get('enterprise_project_id')
        if not record_type or not record_id or not enterprise_project_id:
            raise ValidationError({
                'detail': 'record_type, record_id, and enterprise_project_id are required.'
            })
        try:
            result = resolve_project_relationship(
                record_type=record_type,
                record_id=record_id,
                enterprise_project_id=enterprise_project_id,
                user=request.user,
                reason=str(request.data.get('reason') or ''),
            )
        except DjangoValidationError as exc:
            detail = getattr(exc, 'message_dict', None) or getattr(exc, 'messages', None)
            raise ValidationError(detail) from exc
        return Response(result)

    @action(
        detail=False, methods=['post'], url_path='resolve-invoice-po',
        permission_classes=[IsAuthenticated, CommercialModulePermission],
    )
    def resolve_invoice_po(self, request):
        """Record an operator-selected PO and evaluate the invoice evidence."""
        invoice_id = request.data.get('invoice_id')
        purchase_order_id = request.data.get('purchase_order_id')
        allocated_amount = request.data.get('allocated_amount')
        if not invoice_id or not purchase_order_id or allocated_amount in (None, ''):
            raise ValidationError({
                'detail': 'invoice_id, purchase_order_id, and allocated_amount are required.'
            })
        try:
            result = resolve_invoice_purchase_order(
                invoice_id=invoice_id,
                purchase_order_id=purchase_order_id,
                allocated_amount=allocated_amount,
                user=request.user,
                reason=str(request.data.get('reason') or ''),
            )
        except DjangoValidationError as exc:
            detail = getattr(exc, 'message_dict', None) or getattr(exc, 'messages', None)
            raise ValidationError(detail) from exc
        return Response(result, status=status.HTTP_201_CREATED)
    
    @action(detail=True, methods=['get'])
    def purchase_orders(self, request, pk=None):
        project = self.get_object()
        pos = project.purchase_orders.all().order_by('-created_at')
        serializer = PurchaseOrderSerializer(pos, many=True)
        return Response(serializer.data)
    
    @action(detail=True, methods=['get'])
    def financial_summary(self, request, pk=None):
        project = self.get_object()
        
        po_stats = project.purchase_orders.aggregate(
            total_count=Count('id'),
            total_value=Sum('total_amount'),
            invoiced_value=Sum('total_invoiced_amount'),
        )
        
        budget_stats = project.budgets.aggregate(
            total_allocated=Sum('allocated_amount'),
            approved_count=Count('id', filter=Q(is_approved=True)),
        )
        
        summary = {
            'project_number': project.project_number,
            'project_name': project.project_name,
            'contract_value': float(project.contract_value or 0),
            'contract_currency': project.contract_currency,
            'budgets': {
                'total_allocated': float(project.get_total_budget()),
                'approved_budget_lines': budget_stats['approved_count'],
                'total_spent': float(project.get_total_spent()),
                'remaining': float(project.get_total_budget() - project.get_total_spent()),
                'utilization_percentage': float(project.get_budget_utilization()),
                'is_over_budget': project.is_over_budget(),
            },
            'purchase_orders': {
                'count': po_stats['total_count'],
                'total_value': float(po_stats['total_value'] or 0),
                'invoiced_value': float(po_stats['invoiced_value'] or 0),
                'pending_invoice': float((po_stats['total_value'] or 0) - (po_stats['invoiced_value'] or 0)),
            },
            'health_status': project.health_status,
            'progress_percentage': float(project.progress_percentage),
        }
        
        return Response(summary)
    
    @action(detail=False, methods=['get'])
    def dashboard_stats(self, request):
        queryset = self.filter_queryset(self.get_queryset())
        
        stats = {
            'total_projects': queryset.count(),
            'active_projects': queryset.filter(status='active').count(),
            'completed_projects': queryset.filter(status='completed').count(),
            'projects_on_hold': queryset.filter(status='on_hold').count(),
            'health_breakdown': {
                'green': queryset.filter(health_status='green').count(),
                'yellow': queryset.filter(health_status='yellow').count(),
                'red': queryset.filter(health_status='red').count(),
            },
            'total_contract_value': queryset.aggregate(
                total=Sum('contract_value')
            )['total'] or 0,
        }
        
        return Response(stats)
