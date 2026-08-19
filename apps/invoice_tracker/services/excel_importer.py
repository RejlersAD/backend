"""
Excel importer for CustomerInvoice.

Soft-coded design — every column alias and sheet-category mapping lives in
COLUMN_ALIASES / SHEET_CATEGORY_RULES at module top. Adding a new spreadsheet
variant or renaming a column is a one-line edit; no code changes needed.

Upsert strategy: keyed on `invoice_number`. Re-importing the same Excel
updates existing rows in place.
"""
from __future__ import annotations

import re
from dataclasses import dataclass, field
from datetime import datetime, date
from decimal import Decimal, InvalidOperation
from typing import Any, Iterable

from openpyxl import load_workbook

from ..models import CustomerInvoice, InvoiceCategory, PaymentStatus, Currency


# ─── Soft-coded mapping tables ──────────────────────────────────────────────

# Map *normalised* header text → model field name. Multiple aliases per field.
# Normalisation = lower-case + strip + collapse spaces + drop punctuation.
COLUMN_ALIASES: dict[str, list[str]] = {
    'invoice_number':           ['invoice no', 'invoice number', 'invoice #'],
    'credit_note_ref':          ['credit note'],
    'account':                  ['account', 'company', 'company project'],
    'company':                  ['company'],
    'rad_project_no':           ['rad project', 'rad project #', 'project'],
    'project_name':             ['project name', 'project'],
    'project_id':               ['project id'],
    'invoice_date':             ['invoice date', 'date'],
    'invoice_sent_date':        ['invoice sent', 'sent by'],
    'due_date':                 ['due date'],
    'payment_date':             ['payment date', 'date when payment is received'],
    'payment_terms':            ['payment terms'],
    'currency':                 ['inv cur', 'currency'],
    'ppc_value':                ['ppc value'],
    'retention':                ['retention'],
    'icv_applicable':           ['icv applicable', 'icv applicable yes no'],
    'invoice_amount':           ['invoice amount', 'amount', 'grand total'],
    'invoice_amount_aed':       ['inv amt aed', 'inv amt aed'],
    'amount_excl_vat':          ['inv amount excl vat', 'amount excl vat'],
    'grand_total':              ['grand total', 'incl vat'],
    'balance_to_be_received':   ['balance to be received'],
    'actual_payment_received':  ['actual payment received'],
    'paid_amount_excl_vat':     ['paid amount excl vat'],
    'payment_status':           ['payment status'],
    'days_overdue':             ['days overdue'],
    'bank_reference_code':      ['bank reference code'],
    'customer_inv_reference':   ['customer invoice reference'],
    'contract_clause':          ['contract clause'],
    'finance_pm_email':         ['finance pm email id', 'finance pm email'],
    'pm':                       ['pm'],
    'details':                  ['details'],
    'remarks':                  ['remarks'],
    'sent_by':                  ['sent by'],
    'sent_to_account':          ['received in which nj account', 'received in nj account'],
}

# Sheet-name pattern (regex, case-insensitive) → InvoiceCategory.
# First match wins. Catch-all at bottom defaults to EXTERNAL.
SHEET_CATEGORY_RULES: list[tuple[str, str]] = [
    (r'internal',  InvoiceCategory.INTERNAL),
    (r'external',  InvoiceCategory.EXTERNAL),
    (r'.*',        InvoiceCategory.EXTERNAL),
]

# Header row scan: search the first N rows for the row containing >= MIN_HEADER_HITS
# matched aliases. Excel masters use 2–5 blank rows before headers.
HEADER_SCAN_ROWS = 12
MIN_HEADER_HITS  = 3

# Payment-status keyword → enum value
STATUS_KEYWORDS: list[tuple[str, str]] = [
    ('paid',           PaymentStatus.PAID),
    ('partial',        PaymentStatus.PARTIAL),
    ('overdue',        PaymentStatus.OVERDUE),
    ('cancel',         PaymentStatus.CANCELLED),
    ('credit',         PaymentStatus.CREDIT_NOTE),
    ('pending',        PaymentStatus.PENDING),
]

# Currency keyword (substring, case-insensitive) → enum.
CURRENCY_KEYWORDS: list[tuple[str, str]] = [
    ('aed',  Currency.AED),
    ('usd',  Currency.USD),
    ('eur',  Currency.EUR),
    ('euro', Currency.EUR),
    ('gbp',  Currency.GBP),
    ('sgd',  Currency.SGD),
]


# ─── Result container ───────────────────────────────────────────────────────

@dataclass
class ImportResult:
    sheets_processed: int = 0
    rows_seen:        int = 0
    rows_created:     int = 0
    rows_updated:     int = 0
    rows_skipped:     int = 0
    errors:           list[str] = field(default_factory=list)
    warnings:         list[str] = field(default_factory=list)

    def as_dict(self) -> dict:
        return self.__dict__.copy()


# ─── Helpers ────────────────────────────────────────────────────────────────

_norm_re = re.compile(r'[^a-z0-9 ]+')

def _normalise(header: Any) -> str:
    if header is None:
        return ''
    s = str(header).lower().replace('\n', ' ').strip()
    s = _norm_re.sub(' ', s)
    return ' '.join(s.split())


def _build_alias_lookup() -> dict[str, str]:
    """Flatten COLUMN_ALIASES into {alias_normalised: field_name}."""
    out: dict[str, str] = {}
    for field_name, aliases in COLUMN_ALIASES.items():
        for alias in aliases:
            out[_normalise(alias)] = field_name
    return out


_ALIAS_LOOKUP = _build_alias_lookup()


def _category_for_sheet(sheet_name: str) -> str:
    for pattern, category in SHEET_CATEGORY_RULES:
        if re.search(pattern, sheet_name, flags=re.IGNORECASE):
            return category
    return InvoiceCategory.EXTERNAL


def _to_decimal(value: Any) -> Decimal | None:
    if value is None or value == '':
        return None
    if isinstance(value, (int, float, Decimal)):
        try:
            return Decimal(str(value))
        except (InvalidOperation, ValueError):
            return None
    # String — strip currency text and commas
    s = str(value).strip()
    if not s or s.startswith('#'):
        return None
    s = re.sub(r'[^0-9.\-]', '', s)
    if s in ('', '-', '.'):
        return None
    try:
        return Decimal(s)
    except (InvalidOperation, ValueError):
        return None


def _to_date(value: Any) -> date | None:
    if value is None or value == '':
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    s = str(value).strip()
    if not s:
        return None
    for fmt in ('%d/%m/%Y', '%d/%m/%y', '%Y-%m-%d', '%d-%m-%Y', '%d.%b.%y', '%d.%B.%y'):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _to_status(value: Any) -> str:
    if value is None:
        return PaymentStatus.PENDING
    s = str(value).strip().lower()
    if not s:
        return PaymentStatus.PENDING
    for kw, enum in STATUS_KEYWORDS:
        if kw in s:
            return enum
    return PaymentStatus.PENDING


def _to_currency(value: Any, fallback: str = Currency.AED) -> str:
    if value is None:
        return fallback
    s = str(value).strip().lower()
    if not s:
        return fallback
    for kw, enum in CURRENCY_KEYWORDS:
        if kw in s:
            return enum
    return fallback


def _to_bool(value: Any) -> bool:
    if value is None:
        return False
    if isinstance(value, bool):
        return value
    return str(value).strip().lower() in {'yes', 'y', 'true', '1'}


def _to_text(value: Any) -> str:
    if value is None:
        return ''
    return str(value).strip()


# Type coercer per model field
_FIELD_COERCERS: dict[str, callable] = {
    'invoice_number':           lambda v: _to_text(v),
    'invoice_date':             _to_date,
    'invoice_sent_date':        _to_date,
    'due_date':                 _to_date,
    'payment_date':             _to_date,
    'currency':                 _to_currency,
    'icv_applicable':           _to_bool,
    'payment_status':           _to_status,
    'days_overdue':             lambda v: int(v) if isinstance(v, (int, float)) and v else None,
    'ppc_value':                _to_decimal,
    'retention':                _to_decimal,
    'invoice_amount':           _to_decimal,
    'invoice_amount_aed':       _to_decimal,
    'amount_excl_vat':          _to_decimal,
    'grand_total':              _to_decimal,
    'balance_to_be_received':   _to_decimal,
    'actual_payment_received':  _to_decimal,
    'paid_amount_excl_vat':     _to_decimal,
}


def _coerce(field_name: str, value: Any) -> Any:
    coercer = _FIELD_COERCERS.get(field_name, _to_text)
    return coercer(value)


# ─── Header detection ───────────────────────────────────────────────────────

def _find_header_row(rows: list[list[Any]]) -> tuple[int, dict[int, str]]:
    """Return (row_index, {col_index: field_name}). Empty dict if not found."""
    best_row, best_map = -1, {}
    for idx in range(min(HEADER_SCAN_ROWS, len(rows))):
        row = rows[idx]
        col_map: dict[int, str] = {}
        for col_idx, cell in enumerate(row):
            field_name = _ALIAS_LOOKUP.get(_normalise(cell))
            if field_name and col_idx not in col_map.values():
                col_map[col_idx] = field_name
        if len(col_map) >= MIN_HEADER_HITS and len(col_map) > len(best_map):
            best_row, best_map = idx, col_map
    return best_row, best_map


# ─── Main entry point ───────────────────────────────────────────────────────

def import_workbook(file_path: str, *, user=None,
                    sheet_names: Iterable[str] | None = None) -> ImportResult:
    """Upsert rows from every (or selected) sheet of the workbook.

    Args:
        file_path: path to the .xlsx file (local).
        user: optional User to attribute as `created_by` on new rows.
        sheet_names: optional whitelist of sheet names to process.

    Returns:
        ImportResult with counts + per-row error/warning messages.
    """
    result = ImportResult()
    try:
        wb = load_workbook(file_path, data_only=True, read_only=True)
    except Exception as exc:
        result.errors.append(f'Cannot open workbook: {exc}')
        return result

    targets = list(sheet_names) if sheet_names else wb.sheetnames
    for sheet_name in targets:
        if sheet_name not in wb.sheetnames:
            result.warnings.append(f"Sheet not found: '{sheet_name}'")
            continue
        result.sheets_processed += 1
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        header_row_idx, col_map = _find_header_row(rows)
        if not col_map:
            result.warnings.append(
                f"Sheet '{sheet_name}': no recognisable header row in first {HEADER_SCAN_ROWS} rows."
            )
            continue

        category = _category_for_sheet(sheet_name)

        for raw in rows[header_row_idx + 1:]:
            result.rows_seen += 1
            payload: dict[str, Any] = {}
            for col_idx, field_name in col_map.items():
                if col_idx >= len(raw):
                    continue
                payload[field_name] = _coerce(field_name, raw[col_idx])

            invoice_number = payload.get('invoice_number')
            if not invoice_number:
                result.rows_skipped += 1
                continue

            payload['category'] = category
            # Default currency for internal sheets — they don't track currency
            if not payload.get('currency'):
                payload['currency'] = (
                    Currency.EUR if category == InvoiceCategory.INTERNAL else Currency.AED
                )

            try:
                obj, created = CustomerInvoice.objects.update_or_create(
                    invoice_number=invoice_number,
                    defaults=payload,
                )
                obj.recompute_overdue()
                if created and user is not None and not obj.created_by_id:
                    obj.created_by = user
                obj.save(update_fields=['days_overdue', 'created_by']
                        if obj.created_by_id else ['days_overdue'])
                if created:
                    result.rows_created += 1
                else:
                    result.rows_updated += 1
            except Exception as exc:
                result.errors.append(f"Row '{invoice_number}': {exc}")
                result.rows_skipped += 1

    return result
