"""
Cable Block Diagram Service — server-side ADNOC instrument cable layout.
========================================================================

Pure-Python equivalent of the (now-retired) frontend util that built JB
buckets, multicore tiers, and cabinet assignments.  Centralising it here
gives us:

  * one source of truth for ADNOC naming conventions (reference CRS doc
    30201-50200-H0-113-13-15-00-001 — Habshan MP Fuel Gas, Unit 113);
  * server-side Excel generation matching the manual cable schedule
    layout (17-column ADNOC sheet);
  * trivial extensibility — every threshold / pattern / regex below is a
    module-level constant.  No business logic edits required to retune.

This service is consumed by `cable_block_views.extract_cable_block_diagram`
which in turn reuses `InstrumentIndexService.extract_instruments()` for
the raw P&ID extraction pass.

Reference output samples (verbatim from the ADNOC dwg):
    JB:               113 A 16 101   113 D 15 005   113 C 15 008
    DCS Marshalling:  15-DT-01       ESD Marshalling: 15-ET-02A
    DCS System:       15-DS-01       ESD System:       15-ES-02
    Field cable:      1Px1.5mm² (IS)            1Px2.5mm² (NIS)
    Multicore:        10Px1.5mm² (IS) | 20Px1.5mm² (IS) | 30Px1.5mm² (IS)
                      10Px2.5mm² (NIS)
"""

from __future__ import annotations

import io
import logging
import re
from typing import Any, Callable

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# ════════════════════════════════════════════════════════════════════════════
# SOFT-CODED CONFIGURATION  (edit any constant — no code-flow change needed)
# ════════════════════════════════════════════════════════════════════════════

# ── Column schema (17 columns — matches ADNOC reference cable schedule) ────
CABLE_BLOCK_COLUMNS: list[dict[str, Any]] = [
    {"key": "s_no",                "label": "#",                     "width":  6},
    {"key": "tag_number",          "label": "Instrument Tag",        "width": 18},
    {"key": "service_description", "label": "Service",               "width": 36},
    {"key": "system",              "label": "System",                "width": 10},
    {"key": "is_nis",              "label": "IS / NIS",              "width":  9},
    {"key": "signal_type",         "label": "Signal Type",           "width": 14},
    {"key": "jb_no",               "label": "Junction Box",          "width": 16},
    {"key": "field_cable_no",      "label": "Field Cable Tag",       "width": 16},
    {"key": "field_cable_size",    "label": "Field Cable Size",      "width": 18},
    {"key": "multicore_cable_no",  "label": "Multicore Tag",         "width": 16},
    {"key": "multicore_size",      "label": "Multicore Size",        "width": 20},
    {"key": "marsh_cab_no",        "label": "Marshalling Cabinet",   "width": 14},
    {"key": "sys_cab_no",          "label": "System Cabinet",        "width": 14},
    {"key": "function",            "label": "Function",              "width": 10},
    {"key": "pid_no",              "label": "P&ID No.",              "width": 22},
    {"key": "rev",                 "label": "Rev",                   "width":  6},
    {"key": "remarks",             "label": "Remarks",               "width": 22},
]

# ── ADNOC patterns (placeholders resolved by `_fmt`) ───────────────────────
# NOTE: ADNOC reference (Habshan MP Fuel Gas, doc 30201-50200-H0-113-13-15-00-001)
# uses TWO distinct unit identifiers — they are NOT the same:
#   * plant_unit  = drawing unit ("113")  → first segment of JB number
#   * ies_area    = IES sub-area ("15" or "16") → second-to-last segment of JB,
#                                                 also first segment of cabinets
# Sample JBs:        "113 A 16 101"   "113 D 15 005"   "113 C 15 008"
# Sample cabinets:   "15-DT-01"       "15-ES-02"       "15-ET-02A"
JUNCTION_BOX_PATTERN        = "{plant_unit} {jb_letter} {ies_area} {seq}"
MULTICORE_TAG_PATTERN       = "{plant_unit} {jb_letter} {ies_area} {seq}"   # = JB
MARSHALLING_CABINET_PATTERN = "{ies_area}-{sys}T-{seq}"
SYSTEM_CABINET_PATTERN      = "{ies_area}-{sys}S-{seq}"
FIELD_CABLE_TAG_PATTERN     = "{plant_unit}-FC-{seq}"

# Signal-class → JB letter (A = Analog IS, D = Digital IS, C = NIS)
JB_LETTER_BY_CLASS: dict[str, str] = {
    "ANALOG_IS":  "A",
    "DIGITAL_IS": "D",
    "ANALOG_NIS": "C",
    "DIGITAL_NIS":"C",
}
JB_LETTER_FALLBACK = "A"

# System letter (D = DCS, E = ESD, F = FGS)
SYSTEM_LETTER_BY_NAME: dict[str, str] = {
    "DCS":  "D",
    "ESD":  "E",
    "F&G":  "F",
    "FGS":  "F",
    "F G":  "F",
    "FG":   "F",
}
SYSTEM_LETTER_FALLBACK = "D"

# Cable gauges (mm²) by IS/NIS
FIELD_GAUGE_BY_IS_NIS:     dict[str, str] = {"IS": "1.5", "NIS": "2.5"}
MULTICORE_GAUGE_BY_IS_NIS: dict[str, str] = {"IS": "1.5", "NIS": "2.5"}

# Multicore pair tiers — first tier whose `max_io` ≥ IO count wins
MULTICORE_PAIR_TIERS: list[dict[str, int]] = [
    {"max_io":  8, "pairs": 10},
    {"max_io": 16, "pairs": 20},
    {"max_io": 30, "pairs": 30},
]
MULTICORE_PAIR_FALLBACK = 30

# JB bucketing
MAX_INSTRUMENTS_PER_JB = 12
JB_SEQ_PAD_WIDTH       = 3
CABINET_SEQ_PAD_WIDTH  = 2
DEFAULT_CABINET_SEQ    = 1
JB_SEQ_START           = 1   # first JB in each (area, sys, class) group

# Defaults (used when project metadata is blank — match Habshan MP Fuel Gas ref)
DEFAULT_PLANT_UNIT = "113"
DEFAULT_IES_AREA   = "15"

# ── Tag parsing ────────────────────────────────────────────────────────────
# Examples that must parse:
#   "113-PT-3191"          → area="113", isa="PT"
#   "562-FZT-1501A"        → area="562", isa="FZT"
#   "PT-100"               → area="",    isa="PT"
TAG_REGEX = re.compile(
    r"^\s*(?:(?P<area>[A-Za-z0-9]+)[-_])?(?P<isa>[A-Za-z]+)[-_]?(?P<num>\d+)(?P<suffix>[A-Za-z]*)\s*$"
)

# ISA code → signal classification
ANALOG_ISA = {
    "FT", "FIT", "FE", "FQ", "FQI",
    "PT", "PIT", "PDT", "PDIT", "DPT", "DPIT",
    "TT", "TIT", "TE", "TW",
    "LT", "LIT",
    "AT", "AIT", "AE",
    "VT", "ZT", "XT",
    "FIC", "LIC", "PIC", "TIC", "HIC", "AIC",
}
DIGITAL_ISA = {
    # discrete inputs
    "PSH", "PSL", "PSHH", "PSLL", "TSH", "TSL", "TSHH", "TSLL",
    "LSH", "LSL", "LSHH", "LSLL", "FSH", "FSL", "VSH", "VSL",
    "ZSH", "ZSL", "HS",
    # discrete outputs
    "SDV", "BDV", "MOV", "ROV", "XV", "ESV", "SSV", "SOV", "HV",
}

# ESD / F&G ISA membership (drives system column when project metadata blank)
ESD_ISA = {"SDV", "BDV", "ESV", "SSV", "PSHH", "PSLL", "TSHH", "TSLL",
           "LSHH", "LSLL", "VSH", "VSL"}
FNG_ISA = {"AT", "AIT", "AE", "AAH", "AAL"}

# NIS ISA (high-power solenoids / actuators)
NIS_ISA = {"MOV", "ROV", "SOV", "HV"}

# Excel styling
EXCEL_HEADER_FILL = "1F4E78"   # ADNOC dark navy
EXCEL_HEADER_FONT_COLOR = "FFFFFF"
EXCEL_ALT_ROW_FILL = "F2F6FA"
EXCEL_BORDER_COLOR = "B7B7B7"
EXCEL_SHEET_TITLE  = "Cable Block Diagram"


# ════════════════════════════════════════════════════════════════════════════
# Helpers
# ════════════════════════════════════════════════════════════════════════════
def _fmt(pattern: str, **kw: Any) -> str:
    """Soft format — undefined placeholders stay literal (defensive)."""
    out = pattern
    for k, v in kw.items():
        out = out.replace("{" + k + "}", str(v))
    return out


def _pad(seq: int | str, width: int) -> str:
    return str(seq).zfill(width)


def _parse_tag(raw: Any) -> dict[str, str]:
    m = TAG_REGEX.match(str(raw or ""))
    if not m:
        return {"area": "", "isa": "", "num": "", "suffix": ""}
    return {
        "area":   m.group("area") or "",
        "isa":   (m.group("isa") or "").upper(),
        "num":    m.group("num") or "",
        "suffix": (m.group("suffix") or "").upper(),
    }


def _system_letter(system_name: str | None) -> str:
    key = (system_name or "").upper().strip()
    return SYSTEM_LETTER_BY_NAME.get(key, SYSTEM_LETTER_FALLBACK)


def _derive_system(inst: dict, isa: str) -> str:
    """Pick System (DCS/ESD/F&G) when not provided in extracted record."""
    existing = (inst.get("system") or "").strip()
    if existing and existing.upper() not in ("N/A", "-", "—", "NONE", "NULL"):
        return existing.upper()
    if isa in ESD_ISA:
        return "ESD"
    if isa in FNG_ISA:
        return "F&G"
    return "DCS"


def _derive_is_nis(inst: dict, isa: str) -> str:
    existing = (inst.get("is_nis") or "").strip().upper()
    if existing in ("IS", "NIS"):
        return existing
    if isa in NIS_ISA:
        return "NIS"
    return "IS"


def _derive_signal_class(isa: str) -> str:
    """Returns 'ANALOG' or 'DIGITAL' based on ISA prefix."""
    if isa in ANALOG_ISA:
        return "ANALOG"
    if isa in DIGITAL_ISA:
        return "DIGITAL"
    # default to analog for unknown transmitter-style prefixes (T/I endings)
    if isa.endswith("T") or isa.endswith("IT") or isa.endswith("IC"):
        return "ANALOG"
    return "DIGITAL"


def _pairs_for_io_count(n: int) -> int:
    for tier in MULTICORE_PAIR_TIERS:
        if n <= tier["max_io"]:
            return tier["pairs"]
    return MULTICORE_PAIR_FALLBACK


def _signal_type_label(signal_class: str, is_nis: str) -> str:
    """Used for the on-screen `Signal Type` column."""
    if signal_class == "ANALOG":
        return "Analog 4-20mA"
    return "Digital 24VDC" if is_nis == "IS" else "Digital PF"


# ════════════════════════════════════════════════════════════════════════════
# Builder
# ════════════════════════════════════════════════════════════════════════════
def build_cable_block_rows(
    instruments: list[dict],
    *,
    plant_unit: str | None = None,
    ies_area: str | None = None,
    pid_no: str = "",
    rev: str = "0",
) -> list[dict]:
    """Group instruments into JB buckets and emit one CBD row per IO point.

    Args:
        instruments: extracted instrument records (canonical keys: tag_number,
            service_description, system, is_nis, signal_type, pid_no, …).
        plant_unit: drawing unit prefix (e.g. ``"113"``). Used as the first
            segment of the JB number and field-cable tag. Defaults to
            `DEFAULT_PLANT_UNIT` when blank.
        ies_area: IES sub-area number (e.g. ``"15"`` or ``"16"``). Used as the
            second-to-last segment of the JB number AND as the prefix of all
            marshalling/system cabinet tags. Defaults to `DEFAULT_IES_AREA`.
        pid_no: drawing number to stamp on every row.
        rev: drawing revision letter/number to stamp on every row.

    Returns:
        list of dicts keyed by `CABLE_BLOCK_COLUMNS[*]["key"]`, ready for
        rendering or Excel export.
    """
    plant_unit = str(plant_unit or DEFAULT_PLANT_UNIT).strip() or DEFAULT_PLANT_UNIT
    ies_area   = str(ies_area   or DEFAULT_IES_AREA).strip()   or DEFAULT_IES_AREA

    # ── Pass 1 — classify + bucket ─────────────────────────────────────────
    buckets: dict[str, list[dict]] = {}
    bucket_meta: dict[str, dict[str, Any]] = {}
    jb_counters: dict[str, int] = {}  # (area, sys_letter, jb_letter) → next seq

    for inst in instruments:
        tag = inst.get("tag_number") or ""
        if not tag:
            continue
        parsed = _parse_tag(tag)
        isa    = parsed["isa"]

        system  = _derive_system(inst, isa)
        is_nis  = _derive_is_nis(inst, isa)
        sig_cls = _derive_signal_class(isa)
        class_key = f"{sig_cls}_{is_nis}"
        jb_letter = JB_LETTER_BY_CLASS.get(class_key, JB_LETTER_FALLBACK)
        sys_letter = _system_letter(system)

        # All buckets share the same ies_area (user-provided); JB segregation is
        # by (system × signal-class) only, matching ADNOC plant practice.
        group_key = f"{ies_area}|{sys_letter}|{jb_letter}"
        current_seq = jb_counters.get(group_key, JB_SEQ_START - 1)
        bucket_key = f"{group_key}|{current_seq}"
        bucket = buckets.get(bucket_key)
        if bucket is None or len(bucket) >= MAX_INSTRUMENTS_PER_JB:
            current_seq += 1
            jb_counters[group_key] = current_seq
            bucket_key = f"{group_key}|{current_seq}"
            bucket = []
            buckets[bucket_key] = bucket
            bucket_meta[bucket_key] = {
                "system":     system,
                "sys_letter": sys_letter,
                "jb_letter":  jb_letter,
                "is_nis":     is_nis,
                "sig_cls":    sig_cls,
                "jb_seq":     current_seq,
            }
        bucket.append({
            "inst":    inst,
            "isa":     isa,
            "system":  system,
            "is_nis":  is_nis,
            "sig_cls": sig_cls,
        })

    # ── Pass 2 — cabinet allocation (one cabinet pair per sys letter) ──────
    cabinet_counters: dict[str, int] = {}  # sys_letter → next seq
    cabinet_cache:    dict[str, dict[str, str]] = {}  # sys_letter → {marsh, sys}

    def _cabinets_for(sys_letter: str) -> dict[str, str]:
        if sys_letter in cabinet_cache:
            return cabinet_cache[sys_letter]
        next_seq = cabinet_counters.get(sys_letter, DEFAULT_CABINET_SEQ - 1) + 1
        cabinet_counters[sys_letter] = next_seq
        pad = _pad(next_seq, CABINET_SEQ_PAD_WIDTH)
        result = {
            "marsh": _fmt(MARSHALLING_CABINET_PATTERN, ies_area=ies_area, sys=sys_letter, seq=pad),
            "sys":   _fmt(SYSTEM_CABINET_PATTERN,      ies_area=ies_area, sys=sys_letter, seq=pad),
        }
        cabinet_cache[sys_letter] = result
        return result

    # ── Pass 3 — emit rows ─────────────────────────────────────────────────
    out: list[dict] = []
    serial = 0
    cable_seq = 0
    for bucket_key, members in buckets.items():
        meta = bucket_meta[bucket_key]
        jb_seq_pad = _pad(meta["jb_seq"], JB_SEQ_PAD_WIDTH)
        jb_no = _fmt(
            JUNCTION_BOX_PATTERN,
            plant_unit=plant_unit, jb_letter=meta["jb_letter"],
            ies_area=ies_area, seq=jb_seq_pad,
        )
        multicore_no = _fmt(
            MULTICORE_TAG_PATTERN,
            plant_unit=plant_unit, jb_letter=meta["jb_letter"],
            ies_area=ies_area, seq=jb_seq_pad,
        )
        mc_gauge = MULTICORE_GAUGE_BY_IS_NIS.get(meta["is_nis"], MULTICORE_GAUGE_BY_IS_NIS["IS"])
        mc_pairs = _pairs_for_io_count(len(members))
        multicore_size = f"{mc_pairs}Px{mc_gauge}mm² ({meta['is_nis']})"
        f_gauge = FIELD_GAUGE_BY_IS_NIS.get(meta["is_nis"], FIELD_GAUGE_BY_IS_NIS["IS"])
        field_cable_size = f"1Px{f_gauge}mm² ({meta['is_nis']})"
        cab = _cabinets_for(meta["sys_letter"])

        for m in members:
            serial += 1
            cable_seq += 1
            inst = m["inst"]
            out.append({
                "s_no":                str(serial),
                "tag_number":          inst.get("tag_number") or "",
                "service_description": inst.get("service_description") or "",
                "system":              m["system"],
                "is_nis":              m["is_nis"],
                "signal_type":         _signal_type_label(m["sig_cls"], m["is_nis"]),
                "jb_no":               jb_no,
                "field_cable_no":      _fmt(FIELD_CABLE_TAG_PATTERN, plant_unit=plant_unit, seq=_pad(cable_seq, 4)),
                "field_cable_size":    field_cable_size,
                "multicore_cable_no":  multicore_no,
                "multicore_size":      multicore_size,
                "marsh_cab_no":        cab["marsh"],
                "sys_cab_no":          cab["sys"],
                "function":            m["isa"],
                "pid_no":              inst.get("pid_no") or pid_no,
                "rev":                 rev,
                "remarks":             inst.get("notes") or inst.get("instrument_remark") or "",
            })

    return out


# ════════════════════════════════════════════════════════════════════════════
# Excel generation
# ════════════════════════════════════════════════════════════════════════════
def generate_excel(rows: list[dict], drawing_info: dict | None = None) -> bytes:
    """Render *rows* to a styled ADNOC-flavoured workbook.

    Returns raw .xlsx bytes — caller is responsible for caching / streaming.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = EXCEL_SHEET_TITLE
    drawing_info = drawing_info or {}

    thin = Side(style="thin", color=EXCEL_BORDER_COLOR)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_font = Font(bold=True, color=EXCEL_HEADER_FONT_COLOR, size=10)
    header_fill = PatternFill("solid", fgColor=EXCEL_HEADER_FILL)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    alt_fill = PatternFill("solid", fgColor=EXCEL_ALT_ROW_FILL)

    # ── Title strip (rows 1-3) ─────────────────────────────────────────────
    n_cols = len(CABLE_BLOCK_COLUMNS)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    ws.cell(row=1, column=1, value="INSTRUMENT CABLE BLOCK DIAGRAM").font = Font(bold=True, size=14)
    ws.cell(row=1, column=1).alignment = center

    sub = (
        f"Drawing: {drawing_info.get('drawing_number') or '-'}    "
        f"Rev: {drawing_info.get('revision') or '-'}    "
        f"Project: {drawing_info.get('project_name') or '-'}"
    )
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
    ws.cell(row=2, column=1, value=sub).alignment = center
    ws.cell(row=2, column=1).font = Font(italic=True, size=10, color="555555")

    # ── Header row (row 4) ─────────────────────────────────────────────────
    header_row = 4
    for idx, col in enumerate(CABLE_BLOCK_COLUMNS, start=1):
        cell = ws.cell(row=header_row, column=idx, value=col["label"])
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
        ws.column_dimensions[get_column_letter(idx)].width = col["width"]
    ws.row_dimensions[header_row].height = 30

    # ── Data rows ──────────────────────────────────────────────────────────
    for r_offset, row in enumerate(rows, start=1):
        excel_row = header_row + r_offset
        for c_idx, col in enumerate(CABLE_BLOCK_COLUMNS, start=1):
            cell = ws.cell(row=excel_row, column=c_idx, value=row.get(col["key"], ""))
            cell.alignment = left if col["key"] in (
                "service_description", "remarks",
            ) else center
            cell.border = border
            cell.font = Font(size=9)
            if r_offset % 2 == 0:
                cell.fill = alt_fill

    ws.freeze_panes = "A5"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
