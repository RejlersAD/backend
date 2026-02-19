"""
AI-Powered Pressure Instrument P&ID Analysis Service

This service analyzes P&ID diagrams to extract pressure instrument data using 
advanced AI (OpenAI Vision API) and populates Excel datasheets automatically.
Implements soft coding techniques for easy configuration and extensibility.
"""

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from io import BytesIO
import boto3
from django.conf import settings
import logging
import json
import base64
from openai import OpenAI
from PIL import Image
import io
import os
from pathlib import Path
from pdf2image import convert_from_bytes
from datetime import datetime

logger = logging.getLogger(__name__)


class PressureInstrumentAnalyzer:
    """
    Advanced AI-powered P&ID analyzer for pressure instruments using soft coding.
    Extracts instrument data and generates standardized Excel datasheets.
    """

    def __init__(self):
        """Initialize S3, OpenAI clients and soft-coded configurations"""
        # Initialize S3 client with smart configuration check
        try:
            aws_key = os.getenv('AWS_ACCESS_KEY_ID') or getattr(settings, 'AWS_ACCESS_KEY_ID', None)
            aws_secret = os.getenv('AWS_SECRET_ACCESS_KEY') or getattr(settings, 'AWS_SECRET_ACCESS_KEY', None)
            
            if aws_key and aws_secret:
                self.s3_client = boto3.client(
                    's3',
                    aws_access_key_id=aws_key,
                    aws_secret_access_key=aws_secret,
                    region_name='us-east-1'
                )
                logger.info("[PressureInstrument] ✅ S3 client initialized successfully")
            else:
                self.s3_client = None
                logger.warning("[PressureInstrument] ⚠️ S3 credentials not found - will use generated template")
        except Exception as e:
            self.s3_client = None
            logger.warning(f"[PressureInstrument] ⚠️ S3 initialization failed: {e}")
        
        self.template_bucket = 'rejlers-engineering-data'
        self.template_folder = 'process-department'
        self.template_file = 'Pressure_Instrument.xlsx'
        
        # Initialize OpenAI client
        api_key = os.getenv('OPENAI_API_KEY') or getattr(settings, 'OPENAI_API_KEY', None)
        if api_key:
            self.openai_client = OpenAI(api_key=api_key, timeout=120.0)
            logger.info("[PressureInstrument] ✅ OpenAI client initialized successfully")
        else:
            self.openai_client = None
            logger.error("[PressureInstrument] ❌ OpenAI API key not found - extraction will fail!")
        
        # Soft-coded instrument type configuration
        self.INSTRUMENT_TYPES = {
            'PT': {'name': 'Pressure Transmitter', 'category': 'Transmitter'},
            'PI': {'name': 'Pressure Indicator', 'category': 'Indicator'},
            'PS': {'name': 'Pressure Switch', 'category': 'Switch'},
            'PC': {'name': 'Pressure Controller', 'category': 'Controller'},
            'PG': {'name': 'Pressure Gauge', 'category': 'Gauge'},
            'PDT': {'name': 'Differential Pressure Transmitter', 'category': 'Transmitter'},
            'PDI': {'name': 'Differential Pressure Indicator', 'category': 'Indicator'},
            'PDIS': {'name': 'Differential Pressure Indicator Switch', 'category': 'Indicator'},
            'PIT': {'name': 'Pressure Indicating Transmitter', 'category': 'Transmitter'},
            'PSV': {'name': 'Pressure Safety Valve', 'category': 'Safety'},
            'PRV': {'name': 'Pressure Relief Valve', 'category': 'Safety'}
        }
        
        # Soft-coded Excel field mapping configuration
        self.EXCEL_FIELD_MAPPING = self._initialize_field_mapping()

    def _initialize_field_mapping(self):
        """
        Soft-coded Excel field mapping configuration.
        Maps extracted AI data to specific Excel cells.
        
        Returns:
            dict: Field mapping configuration
        """
        return {
            'HEADER': {
                'drawing_number': {'row': 2, 'col': 2, 'label': 'Drawing Number'},
                'drawing_title': {'row': 3, 'col': 2, 'label': 'Drawing Title'},
                'revision': {'row': 4, 'col': 2, 'label': 'Revision'},
                'project_name': {'row': 2, 'col': 5, 'label': 'Project Name'},
                'area': {'row': 3, 'col': 5, 'label': 'Area'},
                'date': {'row': 4, 'col': 5, 'label': 'Date'}
            },
            'INSTRUMENT_DATA_START_ROW': 8,
            'COLUMNS': {
                'tag_number': {'col': 1, 'label': 'Tag Number', 'width': 15},
                'pid_no': {'col': 2, 'label': 'P&ID No', 'width': 15},
                'line_no': {'col': 3, 'label': 'Line No.', 'width': 15},
                'piping_class': {'col': 4, 'label': 'Piping Class', 'width': 15},
                'equipment_no': {'col': 5, 'label': 'Equipment No.', 'width': 15},
                'service': {'col': 6, 'label': 'Service', 'width': 25},
                'fluid_state': {'col': 7, 'label': 'Fluid State', 'width': 12},
                'fluid_phase': {'col': 8, 'label': 'Fluid Phase', 'width': 12},
                'operating_pressure_min': {'col': 9, 'label': 'Operating Pressure Min', 'width': 18},
                'operating_pressure_norm': {'col': 10, 'label': 'Operating Pressure Norm', 'width': 18},
                'operating_pressure_max': {'col': 11, 'label': 'Operating Pressure Max', 'width': 18},
                'operating_temp_min': {'col': 12, 'label': 'Operating Temp Min', 'width': 18},
                'operating_temp_norm': {'col': 13, 'label': 'Operating Temp Norm', 'width': 18},
                'operating_temp_max': {'col': 14, 'label': 'Operating Temp Max', 'width': 18},
                'operating_differential_pressure': {'col': 15, 'label': 'Operating Differential Pressure', 'width': 20},
                'design_pressure_min': {'col': 16, 'label': 'Design Pressure Min', 'width': 18},
                'design_pressure_norm': {'col': 17, 'label': 'Design Pressure Norm', 'width': 18},
                'design_pressure_max': {'col': 18, 'label': 'Design Pressure Max', 'width': 18},
                'source_service': {'col': 19, 'label': 'Source Service', 'width': 20},
                'special_conditions': {'col': 20, 'label': 'Special Conditions', 'width': 25},
                'density_min': {'col': 21, 'label': 'Density Min', 'width': 12},
                'density_norm': {'col': 22, 'label': 'Density Norm', 'width': 12},
                'density_max': {'col': 23, 'label': 'Density Max', 'width': 12},
                'viscosity_min': {'col': 24, 'label': 'Viscosity Min', 'width': 12},
                'viscosity_norm': {'col': 25, 'label': 'Viscosity Norm', 'width': 12},
                'viscosity_max': {'col': 26, 'label': 'Viscosity Max', 'width': 12},
                'gauge_adaptor': {'col': 27, 'label': 'Gauge Adaptor', 'width': 15},
                'nace_requirement': {'col': 28, 'label': 'Nace Requirement', 'width': 15},
                'notes': {'col': 29, 'label': 'Notes', 'width': 30}
            }
        }

    def analyze_pid_with_ai(self, pid_image_data, drawing_info):
        """
        Analyze P&ID diagram using OpenAI Vision API to extract pressure instrument data.
        
        Args:
            pid_image_data: Image data (bytes or PIL Image)
            drawing_info: Dictionary with drawing metadata
            
        Returns:
            list: Extracted pressure instrument data
        """
        if not self.openai_client:
            logger.error("[PressureInstrument] OpenAI client not initialized")
            return []
        
        try:
            # Convert image to base64
            if isinstance(pid_image_data, bytes):
                # Check if PDF
                if pid_image_data[:4] == b'%PDF':
                    logger.info("[PressureInstrument] Converting PDF to image...")
                    # Convert PDF first page to image
                    images = convert_from_bytes(pid_image_data, first_page=1, last_page=1, dpi=300)
                    img = images[0]
                    
                    # Resize if too large (max 2048x2048 for OpenAI)
                    max_size = 2048
                    if img.width > max_size or img.height > max_size:
                        logger.info(f"[PressureInstrument] Resizing image from {img.width}x{img.height}")
                        img.thumbnail((max_size, max_size), Image.Resampling.LANCZOS)
                    
                    buffered = io.BytesIO()
                    img.save(buffered, format="PNG", optimize=True)
                    image_bytes = buffered.getvalue()
                    logger.info(f"[PressureInstrument] PDF converted to PNG: {len(image_bytes)} bytes")
                else:
                    # Already an image, just validate and potentially resize
                    try:
                        img = Image.open(io.BytesIO(pid_image_data))
                        logger.info(f"[PressureInstrument] Image loaded: {img.format} {img.width}x{img.height}")
                        
                        # Resize if too large
                        max_size = 2048
                        if img.width > max_size or img.height > max_size:
                            logger.info(f"[PressureInstrument] Resizing image from {img.width}x{img.height}")
                            img.thumbnail((max_size, max_size), Image.Resampling.LANCZOS)
                        
                        # Convert to PNG for consistency
                        buffered = io.BytesIO()
                        img.save(buffered, format="PNG", optimize=True)
                        image_bytes = buffered.getvalue()
                    except Exception as e:
                        logger.warning(f"[PressureInstrument] Could not process as image: {e}, using raw bytes")
                        image_bytes = pid_image_data
            else:
                image_bytes = pid_image_data
            
            logger.info(f"[PressureInstrument] Final image size: {len(image_bytes)} bytes")
            base64_image = base64.b64encode(image_bytes).decode('utf-8')
            logger.info(f"[PressureInstrument] Base64 encoded: {len(base64_image)} characters")
            
            # Create comprehensive AI prompt
            prompt = self._create_analysis_prompt(drawing_info)
            
            logger.info("[PressureInstrument] Sending P&ID to OpenAI Vision API for analysis...")
            
            # Call OpenAI Vision API with updated model
            # Using gpt-4o which has vision capabilities and is the latest model
            response = self.openai_client.chat.completions.create(
                model="gpt-4o",
                messages=[
                    {
                        "role": "system",
                        "content": """You are an expert P&ID analyzer and process instrumentation engineer with 20+ years of experience.
                        
                        Your specialized skills:
                        - Reading complex P&ID drawings and identifying ALL instrument symbols
                        - Recognizing instrument bubbles (circles with tags like PT-101, PI-202, etc.)
                        - Understanding ISA instrument tag nomenclature
                        - Extracting data from P&ID legends, tables, and annotations
                        - Identifying pressure instruments even with partial or unclear tags
                        
                        CRITICAL: You MUST find ALL pressure-related instruments in the diagram. Look everywhere:
                        - On process lines (inline instruments)
                        - Connected to vessels and equipment
                        - In instrument loops and control schemes
                        - In legend tables and instrument lists
                        - In line schedules and process data tables
                        - In border notes and general notes sections
                        - Even if partially visible or unclear
                        
                        EXTRACT MAXIMUM DATA from the drawing:
                        - Check line schedules for operating conditions (pressure, temperature, piping class)
                        - Look for process data tables with fluid properties (density, viscosity)
                        - Read equipment data sheets if visible in the drawing
                        - Extract design conditions from notes or specifications
                        - Identify special service requirements (H2S, NACE, corrosive, etc.)
                        
                        If you see ANY circular symbol with letters/numbers that could be a pressure instrument, 
                        INCLUDE IT in your analysis with ALL VISIBLE DATA. When in doubt, include it."""
                    },
                    {
                        "role": "user",
                        "content": [
                            {
                                "type": "text",
                                "text": prompt
                            },
                            {
                                "type": "image_url",
                                "image_url": {
                                    "url": f"data:image/png;base64,{base64_image}",
                                    "detail": "high"
                                }
                            }
                        ]
                    }
                ],
                max_tokens=10000,
                temperature=0.1
            )
            
            # Parse AI response
            ai_response = response.choices[0].message.content
            logger.info(f"[PressureInstrument] ✅ AI Response received: {len(ai_response)} characters")
            logger.info(f"[PressureInstrument] 📄 Full AI Response:\n{ai_response}")
            logger.info(f"[PressureInstrument] 📄 Response first 500 chars: {ai_response[:500]}")
            
            # Extract structured data from AI response
            instruments = self._parse_ai_response(ai_response)
            logger.info(f"[PressureInstrument] 📊 Parsed {len(instruments)} instruments from primary response")
            
            # If no instruments found, try simplified detection prompt
            if not instruments:
                logger.warning("[PressureInstrument] ⚠️ No instruments on first attempt, trying simplified detection...")
                instruments = self._retry_with_simplified_prompt(base64_image, drawing_info)
                logger.info(f"[PressureInstrument] 📊 Simplified detection found {len(instruments)} instruments")
            
            if not instruments:
                logger.warning(f"[PressureInstrument] No instruments extracted from AI response")
                logger.warning(f"[PressureInstrument] Response preview: {ai_response[:500]}")
            
            logger.info(f"[PressureInstrument] Extracted {len(instruments)} pressure instruments")
            return instruments
            
        except Exception as e:
            logger.error(f"[PressureInstrument] AI analysis error: {str(e)}")
            return []

    def _create_analysis_prompt(self, drawing_info):
        """
        Create comprehensive AI prompt for P&ID analysis.
        
        Args:
            drawing_info: Drawing metadata dictionary
            
        Returns:
            str: Formatted prompt
        """
        instrument_types_list = ', '.join([f"{k} ({v['name']})" for k, v in self.INSTRUMENT_TYPES.items()])
        
        prompt = f"""
🎯 MISSION: Scan this entire P&ID diagram and find EVERY SINGLE pressure instrument.

📋 Drawing Information:
- Drawing Number: {drawing_info.get('drawing_number', 'N/A')}
- Drawing Title: {drawing_info.get('drawing_title', 'N/A')}
- Area: {drawing_info.get('area', 'N/A')}
- Project: {drawing_info.get('project_name', 'N/A')}

🔍 VISUAL DETECTION GUIDE:
Look for these VISUAL PATTERNS on the P&ID:

1. CIRCULAR BUBBLES with any of these patterns:
   - PT-XXX, PTI-XXX, PIT-XXX (Pressure Transmitters)
   - PI-XXX, PIA-XXX (Pressure Indicators)
   - PS-XXX, PSH-XXX, PSL-XXX (Pressure Switches)
   - PC-XXX, PIC-XXX, PCV-XXX (Pressure Controllers)
   - PG-XXX (Pressure Gauges)
   - PDT-XXX, PDTI-XXX (Differential Pressure Transmitters)
   - PDI-XXX, PDIS-XXX (Differential Pressure Indicators)
   - PSV-XXX, PRV-XXX (Pressure Relief/Safety Valves)
   - Any circle with 'P' as first letter followed by numbers

2. INSTRUMENT SYMBOLS:
   - Small circles (1-2 inches diameter on drawing)
   - Circles connected to process lines with thin lines
   - Circles with internal divisions or multiple sections
   - Balloon-style tags attached to equipment/lines

3. WHERE TO LOOK:
   - Along ALL process lines and piping
   - On vessel nozzles and connections
   - Near pumps, compressors, and rotating equipment
   - Around control valves and block valves
   - In instrument connection details
   - On the legend/symbol table
   - In instrument index tables
   - Near tie-in points and boundaries

🎯 Target Instrument Types: {instrument_types_list}

For EACH pressure instrument found (minimum 1, expect 5-20), extract:

1. **Tag Number**: The instrument tag (e.g., PT-101, PI-201, PDT-301)
2. **P&ID No**: The P&ID drawing number where the instrument appears
3. **Line No.**: The process line number where the instrument is installed
4. **Piping Class**: The piping class specification (e.g., 150#, 300#, 600#)
5. **Equipment No.**: Associated equipment number if connected to equipment
6. **Service**: Description of what is being measured (e.g., "Compressor Discharge Pressure")
7. **Fluid State**: State of the fluid (e.g., "Liquid", "Gas", "Vapor", "Two-Phase")
8. **Fluid Phase**: Phase description (e.g., "Single Phase", "Multi-Phase")
9. **Operating Pressure Min**: Minimum operating pressure (bar or psi)
10. **Operating Pressure Norm**: Normal/nominal operating pressure (bar or psi)
11. **Operating Pressure Max**: Maximum operating pressure (bar or psi)
12. **Operating Temp Min**: Minimum operating temperature (°C or °F)
13. **Operating Temp Norm**: Normal/nominal operating temperature (°C or °F)
14. **Operating Temp Max**: Maximum operating temperature (°C or °F)
15. **Operating Differential Pressure**: Differential pressure across the instrument (if applicable)
16. **Design Pressure Min**: Minimum design pressure (bar or psi)
17. **Design Pressure Norm**: Normal design pressure (bar or psi)
18. **Design Pressure Max**: Maximum design pressure (bar or psi)
19. **Source Service**: Origin or source of the fluid/service
20. **Special Conditions**: Any special requirements, hazardous conditions, or safety notes
21. **Density Min**: Minimum fluid density (kg/m³ or lb/ft³)
22. **Density Norm**: Normal fluid density (kg/m³ or lb/ft³)
23. **Density Max**: Maximum fluid density (kg/m³ or lb/ft³)
24. **Viscosity Min**: Minimum fluid viscosity (cP or cSt)
25. **Viscosity Norm**: Normal fluid viscosity (cP or cSt)
26. **Viscosity Max**: Maximum fluid viscosity (cP or cSt)
27. **Gauge Adaptor**: Type of gauge adaptor or connection (e.g., "Diaphragm Seal", "Siphon")
28. **Nace Requirement**: NACE compliance requirement (e.g., "MR0175", "Not Required")
29. **Notes**: Additional remarks, specifications, or comments

**Output Format**: Return data as a JSON array with this structure:
```json
[
  {{
    "tag_number": "PT-101",
    "pid_no": "P-16093-001",
    "line_no": "1001-P-4\\"",
    "piping_class": "150#",
    "equipment_no": "C-101",
    "service": "Compressor Discharge Pressure",
    "fluid_state": "Gas",
    "fluid_phase": "Single Phase",
    "operating_pressure_min": "40.0",
    "operating_pressure_norm": "45.0",
    "operating_pressure_max": "50.0",
    "operating_temp_min": "30.0",
    "operating_temp_norm": "40.0",
    "operating_temp_max": "50.0",
    "operating_differential_pressure": "N/A",
    "design_pressure_min": "50.0",
    "design_pressure_norm": "60.0",
    "design_pressure_max": "65.0",
    "source_service": "Natural Gas Processing",
    "special_conditions": "H2S Service",
    "density_min": "15.0",
    "density_norm": "18.5",
    "density_max": "22.0",
    "viscosity_min": "0.01",
    "viscosity_norm": "0.015",
    "viscosity_max": "0.02",
    "gauge_adaptor": "Diaphragm Seal",
    "nace_requirement": "MR0175",
    "notes": "Critical alarm point"
  }}
]
```

**🚨 CRITICAL DETECTION INSTRUCTIONS:**

⚠️ IMPORTANT: A typical P&ID has 5-20+ pressure instruments. If you find 0 instruments, YOU MISSED THEM!

STEP-BY-STEP SCANNING PROCESS:
1. **SYSTEMATIC SCAN**: Scan the ENTIRE drawing from left-to-right, top-to-bottom
2. **ZOOM IN MENTALLY**: Look at EVERY circular symbol, no matter how small
3. **CHECK EVERYTHING**: Any circle or bubble could be a pressure instrument
4. **READ ALL TEXT**: Check every tag, label, and annotation for pressure instrument codes
5. **EXAMINE LEGENDS**: If there's a legend/symbol table, extract all pressure instruments listed
6. **LINE CONNECTIONS**: Follow every process line and check for instrument connections

📋 WHAT TO EXTRACT:
- IF you see a tag clearly: Extract all visible information
- IF tag is partially visible: Use best judgment and note uncertainty in \"notes\"
- IF fluid data not visible: Use \"N/A\" for unknown fields
- IF only one pressure value visible: Put it in \"norm\", use \"N/A\" for min/max

⚠️ DETECTION RULES:
1. ANY circle with text starting with 'P' followed by letters/numbers → Likely pressure instrument
2. Circles near valves, pumps, vessels → High probability of instruments
3. Small bubbles connected to lines with thin leader lines → Instrument tags
4. Look for PT, PI, PS, PC, PG, PDT, PDI, PSV, PRV prefixes
5. Tags may be: PT-101, PT101, P-101, 101-PT, or just P101
6. Some drawings use abbreviations: PTR, PIT, PIA, PSH, PSL, etc.

🎯 OUTPUT REQUIREMENTS:
- Return JSON array with AT LEAST 1 instrument (typical is 5-20)
- If diagram truly has NO pressure instruments (very rare), return empty array []
- Use \"N/A\" for any field you cannot determine
- Include ALL instruments you find, even if data is incomplete
- Better to include uncertain instruments than miss them

📦 RETURN FORMAT:
ONLY return the JSON array - NO markdown blocks, NO explanations, NO text before/after
Start with [ and end with ]
"""
        return prompt

    def _retry_with_simplified_prompt(self, base64_image, drawing_info):
        """
        Retry analysis with an ultra-simplified, aggressive detection prompt.
        Used as fallback when main analysis returns no results.
        
        Args:
            base64_image: Base64 encoded image
            drawing_info: Drawing metadata dictionary
            
        Returns:
            list: Extracted instruments or empty list
        """
        if not self.openai_client:
            return []
        
        try:
            logger.info("[PressureInstrument] Attempting simplified detection strategy...")
            
            simplified_prompt = f"""🎯 EMERGENCY FALLBACK DETECTION MODE

Analyze this P&ID diagram and extract EVERY pressure instrument with ALL AVAILABLE data.

📋 Drawing: {drawing_info.get('drawing_number', 'N/A')} - {drawing_info.get('drawing_title', 'N/A')}

🔍 TARGET INSTRUMENTS (starting with 'P'):
PT, PI, PS, PC, PG, PDT, PDI, PSV, PRV, PIT, PIA, PSH, PSL, PIC, PCV, etc.

📊 EXTRACTION STRATEGY - For EACH pressure instrument found:

**STEP 1 - INSTRUMENT IDENTIFICATION:**
- Tag Number: Read the instrument tag (PT-101, PI-202, PSV-3601-01, etc.)
- Line Number: Check what process line it's connected to (look for line tags like 1"-P-4001)
- Equipment Number: See if connected to equipment (C-101, V-202, P-301, etc.)
- Service: Read nearby text describing the service or function

**STEP 2 - PROCESS CONDITIONS (Look for tables, notes, line lists):**
- Operating Pressure: Check for pressure values near the line/equipment (bar, psi, kPa)
- Operating Temperature: Look for temperature annotations (°C, °F, K)
- Design Pressure: Usually in line schedule or equipment data sheets
- Design Temperature: Check equipment data or line specifications

**STEP 3 - FLUID PROPERTIES (From line schedules or process notes):**
- Fluid State: Gas, Liquid, Vapor, Two-Phase (look at line shading/patterns)
- Fluid Phase: Single Phase, Multi-Phase
- Piping Class: Line class specification (150#, 300#, 600#, ANSI rating)
- Density: ρ = kg/m³ or lb/ft³ (check process data tables)
- Viscosity: μ = cP or cSt (check fluid property tables)

**STEP 4 - SPECIAL REQUIREMENTS:**
- Special Conditions: H2S Service, Corrosive, High Temperature, Cryogenic, etc.
- NACE Requirement: MR0175, MR0103, or "Not Required"
- Gauge Adaptor: Diaphragm Seal, Siphon, Remote Seal, Direct Mount
- Source Service: Where fluid originates (Feed Gas, Diesel Oil, Cooling Water, etc.)

🎯 DATA EXTRACTION RULES:
- ✅ EXTRACT visible data from: Line tags, equipment tags, tables, notes, legends, schedules
- ✅ INFER reasonable values when partially visible
- ✅ CHECK borders/margins for process data tables or legends
- ❌ ONLY use "N/A" when data is truly not available in the drawing
- ❌ DON'T assume – if pressure is "45 bar", don't guess min/max, use "N/A"

📦 JSON OUTPUT FORMAT:
[
  {{
    "tag_number": "PT-101",
    "pid_no": "{drawing_info.get('drawing_number', 'N/A')}",
    "line_no": "1\"-P-4001",
    "piping_class": "150#",
    "equipment_no": "C-101",
    "service": "Compressor Discharge Pressure",
    "fluid_state": "Gas",
    "fluid_phase": "Single Phase",
    "operating_pressure_min": "40",
    "operating_pressure_norm": "45",
    "operating_pressure_max": "50",
    "operating_temp_min": "30",
    "operating_temp_norm": "40",
    "operating_temp_max": "45",
    "operating_differential_pressure": "N/A",
    "design_pressure_min": "60",
    "design_pressure_norm": "65",
    "design_pressure_max": "70",
    "source_service": "Natural Gas",
    "special_conditions": "H2S Service",
    "density_min": "15.0",
    "density_norm": "18.5",
    "density_max": "22.0",
    "viscosity_min": "0.01",
    "viscosity_norm": "0.015",
    "viscosity_max": "0.02",
    "gauge_adaptor": "Diaphragm Seal",
    "nace_requirement": "MR0175",
    "notes": "Critical alarm - monitor continuously"
  }}
]

⚠️ CRITICAL: Extract ALL instruments you find with MAXIMUM data available from the drawing.
Return ONLY the JSON array - no explanations, no markdown blocks.
"""

            response = self.openai_client.chat.completions.create(
                model="gpt-4o",
                messages=[
                    {
                        "role": "system",
                        "content": "You are an expert P&ID analyzer specializing in process instrumentation. Extract ALL pressure instruments with MAXIMUM available data from drawings including: tags, line numbers, equipment connections, operating conditions, design parameters, fluid properties, and special requirements. Return comprehensive JSON data."
                    },
                    {
                        "role": "user",
                        "content": [
                            {"type": "text", "text": simplified_prompt},
                            {
                                "type": "image_url",
                                "image_url": {
                                    "url": f"data:image/png;base64,{base64_image}",
                                    "detail": "high"
                                }
                            }
                        ]
                    }
                ],
                max_tokens=8000,
                temperature=0.2
            )
            
            retry_response = response.choices[0].message.content
            logger.info(f"[PressureInstrument] Simplified detection response: {retry_response}")
            
            instruments = self._parse_ai_response(retry_response)
            logger.info(f"[PressureInstrument] Simplified detection found {len(instruments)} instruments")
            
            return instruments
            
        except Exception as e:
            logger.error(f"[PressureInstrument] Simplified detection error: {str(e)}")
            return []

    def _parse_ai_response(self, ai_response):
        """
        Parse AI response and extract structured instrument data.
        Handles markdown code blocks, plain JSON, and various formats.
        
        Args:
            ai_response: Raw AI response text
            
        Returns:
            list: Parsed instrument data
        """
        try:
            # Remove markdown code blocks if present
            cleaned_response = ai_response.strip()
            
            # Check for markdown code blocks
            if '```json' in cleaned_response:
                # Extract content between ```json and ```
                start = cleaned_response.find('```json') + 7
                end = cleaned_response.find('```', start)
                if end > start:
                    cleaned_response = cleaned_response[start:end].strip()
                    logger.debug("[PressureInstrument] Extracted JSON from markdown block")
            elif '```' in cleaned_response:
                # Generic code block
                start = cleaned_response.find('```') + 3
                end = cleaned_response.find('```', start)
                if end > start:
                    cleaned_response = cleaned_response[start:end].strip()
                    logger.debug("[PressureInstrument] Extracted content from code block")
            
            # Try to find JSON array in the response
            start_idx = -1
            for idx, char in enumerate(cleaned_response):
                if char == '[':
                    start_idx = idx
                    break
            
            end_idx = cleaned_response.rfind(']') + 1
            
            if start_idx >= 0 and end_idx > start_idx:
                json_str = cleaned_response[start_idx:end_idx]
                logger.debug(f"[PressureInstrument] Extracted JSON string: {json_str[:200]}...")
                
                instruments = json.loads(json_str)
                
                # Validate it's a list
                if not isinstance(instruments, list):
                    logger.error(f"[PressureInstrument] Parsed data is not a list: {type(instruments)}")
                    return []
                
                logger.info(f"[PressureInstrument] Successfully parsed {len(instruments)} instruments")
                return instruments
            else:
                # Try parsing the entire cleaned response as JSON
                try:
                    instruments = json.loads(cleaned_response)
                    if isinstance(instruments, list):
                        logger.info(f"[PressureInstrument] Parsed full response as JSON list")
                        return instruments
                    elif isinstance(instruments, dict) and 'instruments' in instruments:
                        logger.info(f"[PressureInstrument] Extracted instruments from dict")
                        return instruments['instruments']
                except:
                    pass
                
                logger.warning("[PressureInstrument] No JSON array found in AI response")
                logger.debug(f"[PressureInstrument] Cleaned response: {cleaned_response[:500]}")
                return []
                
        except json.JSONDecodeError as e:
            logger.error(f"[PressureInstrument] JSON parsing error: {str(e)}")
            logger.error(f"[PressureInstrument] Error at position {e.pos}")
            logger.debug(f"[PressureInstrument] Raw response: {ai_response[:1000]}")
            return []
        except Exception as e:
            logger.error(f"[PressureInstrument] Unexpected parsing error: {str(e)}", exc_info=True)
            return []

    def download_template_from_s3(self):
        """
        Download Excel template from S3 with smart fallback to local file.
        
        Priority:
        1. Try S3 bucket
        2. Try local backend folder
        3. Return None to create from scratch
        
        Returns:
            BytesIO: Template file data or None if unavailable
        """
        # Try S3 first if client is available
        if self.s3_client:
            try:
                template_path = f"{self.template_folder}/{self.template_file}"
                logger.info(f"[PressureInstrument] 📥 Downloading template from S3: {self.template_bucket}/{template_path}")
                
                response = self.s3_client.get_object(
                    Bucket=self.template_bucket,
                    Key=template_path
                )
                
                template_data = response['Body'].read()
                logger.info(f"[PressureInstrument] ✅ Template downloaded from S3: {len(template_data)} bytes")
                
                return BytesIO(template_data)
                
            except Exception as e:
                logger.warning(f"[PressureInstrument] ⚠️ S3 template download failed: {str(e)}")
        
        # Try local file as fallback
        try:
            # Get the backend directory (project root)
            from django.conf import settings
            backend_dir = settings.BASE_DIR
            local_template_path = backend_dir / self.template_file
            
            if local_template_path.exists():
                logger.info(f"[PressureInstrument] 📁 Loading template from local file: {local_template_path}")
                
                with open(local_template_path, 'rb') as f:
                    template_data = f.read()
                
                logger.info(f"[PressureInstrument] ✅ Template loaded from local file: {len(template_data)} bytes")
                return BytesIO(template_data)
            else:
                logger.warning(f"[PressureInstrument] ⚠️ Local template not found: {local_template_path}")
        except Exception as e:
            logger.warning(f"[PressureInstrument] ⚠️ Local template load error: {str(e)}")
        
        # No template available - will create from scratch
        logger.info("[PressureInstrument] 🔄 No template available - will create from scratch")
        return None

    def populate_excel_datasheet(self, instruments_data, drawing_info):
        """
        Populate Excel datasheet with extracted instrument data using soft-coded mapping.
        Implements robust error handling for corrupted templates with automatic fallback.
        
        Args:
            instruments_data: List of instrument dictionaries
            drawing_info: Drawing metadata dictionary
            
        Returns:
            BytesIO: Populated Excel file
        """
        try:
            # Try to download template
            template = self.download_template_from_s3()
            
            wb = None
            ws = None
            template_loaded = False
            
            # Try loading the template with error handling for corrupted files
            if template:
                try:
                    logger.info("[PressureInstrument] 📋 Attempting to load template workbook...")
                    wb = openpyxl.load_workbook(template)
                    ws = wb.active
                    template_loaded = True
                    logger.info("[PressureInstrument] ✅ Template loaded successfully")
                except ValueError as ve:
                    if "Unable to read workbook" in str(ve) or "invalid XML" in str(ve):
                        logger.warning(f"[PressureInstrument] ⚠️ Template file is corrupted (invalid XML): {str(ve)[:100]}")
                        logger.info("[PressureInstrument] 🔄 Falling back to programmatic template creation")
                        template_loaded = False
                    else:
                        raise
                except Exception as e:
                    logger.warning(f"[PressureInstrument] ⚠️ Template loading error: {str(e)[:100]}")
                    logger.info("[PressureInstrument] 🔄 Falling back to programmatic template creation")
                    template_loaded = False
            
            # Create new workbook if template failed to load
            if not template_loaded:
                logger.info("[PressureInstrument] 🆕 Creating new workbook from scratch")
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Pressure Instruments"
                self._create_template_structure(ws)
                logger.info("[PressureInstrument] ✅ New template structure created")
            
            # Populate header information
            self._populate_header(ws, drawing_info)
            
            # Populate instrument data
            self._populate_instruments(ws, instruments_data)
            
            # Apply styling
            self._apply_styling(ws, len(instruments_data))
            
            # Save to BytesIO
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            logger.info(f"[PressureInstrument] ✅ Excel datasheet generated: {len(instruments_data)} instruments")
            return output
            
        except Exception as e:
            logger.error(f"[PressureInstrument] Excel generation error: {str(e)}")
            raise

    def _create_template_structure(self, ws):
        """Create Excel template structure with headers"""
        mapping = self.EXCEL_FIELD_MAPPING
        
        # Set column widths and headers
        header_row = mapping['INSTRUMENT_DATA_START_ROW'] - 1
        for field_key, field_info in mapping['COLUMNS'].items():
            col = field_info['col']
            ws.column_dimensions[get_column_letter(col)].width = field_info['width']
            ws.cell(row=header_row, column=col, value=field_info['label'])

    def _populate_header(self, ws, drawing_info):
        """Populate header section with drawing information"""
        mapping = self.EXCEL_FIELD_MAPPING['HEADER']
        
        for field_key, field_info in mapping.items():
            value = drawing_info.get(field_key, '')
            if field_key == 'date' and not value:
                value = datetime.now().strftime('%Y-%m-%d')
            
            ws.cell(row=field_info['row'], column=field_info['col'], value=value)

    def _populate_instruments(self, ws, instruments_data):
        """Populate instrument data rows"""
        start_row = self.EXCEL_FIELD_MAPPING['INSTRUMENT_DATA_START_ROW']
        columns = self.EXCEL_FIELD_MAPPING['COLUMNS']
        
        for idx, instrument in enumerate(instruments_data):
            row = start_row + idx
            
            # Populate each column
            for field_key, field_info in columns.items():
                col = field_info['col']
                value = instrument.get(field_key, '')
                
                # Convert numeric strings to numbers for pressure, temperature, density, and viscosity fields
                numeric_fields = [
                    'operating_pressure_min', 'operating_pressure_norm', 'operating_pressure_max',
                    'operating_temp_min', 'operating_temp_norm', 'operating_temp_max',
                    'operating_differential_pressure',
                    'design_pressure_min', 'design_pressure_norm', 'design_pressure_max',
                    'density_min', 'density_norm', 'density_max',
                    'viscosity_min', 'viscosity_norm', 'viscosity_max'
                ]
                
                if field_key in numeric_fields:
                    try:
                        value = float(value) if value and value != 'N/A' else value
                    except ValueError:
                        pass
                
                ws.cell(row=row, column=col, value=value)

    def _apply_styling(self, ws, data_rows):
        """Apply professional styling to the worksheet"""
        # Header styling
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Calculate total columns from mapping
        total_columns = len(self.EXCEL_FIELD_MAPPING['COLUMNS'])
        
        header_row = self.EXCEL_FIELD_MAPPING['INSTRUMENT_DATA_START_ROW'] - 1
        for col in range(1, total_columns + 1):
            cell = ws.cell(row=header_row, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Data styling
        data_alignment = Alignment(horizontal='left', vertical='center')
        border = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC')
        )
        
        start_row = self.EXCEL_FIELD_MAPPING['INSTRUMENT_DATA_START_ROW']
        for row in range(start_row, start_row + data_rows):
            for col in range(1, total_columns + 1):
                cell = ws.cell(row=row, column=col)
                cell.alignment = data_alignment
                cell.border = border
                
                # Alternate row colors
                if row % 2 == 0:
                    cell.fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')

    def generate_datasheet_from_pid(self, pid_file, drawing_info):
        """
        Main method to generate Excel datasheet from P&ID.
        
        Args:
            pid_file: P&ID file (bytes or file object)
            drawing_info: Drawing metadata dictionary
            
        Returns:
            tuple: (BytesIO of Excel file, list of instruments, success message)
        """
        try:
            # Read PID file
            if hasattr(pid_file, 'read'):
                pid_data = pid_file.read()
            else:
                pid_data = pid_file
            
            # Analyze P&ID with AI
            logger.info("[PressureInstrument] Starting P&ID analysis...")
            instruments = self.analyze_pid_with_ai(pid_data, drawing_info)
            
            if not instruments:
                logger.warning("[PressureInstrument] No instruments detected")
                return None, [], "No pressure instruments detected in the P&ID"
            
            # Generate Excel datasheet
            logger.info(f"[PressureInstrument] Generating Excel datasheet for {len(instruments)} instruments...")
            excel_file = self.populate_excel_datasheet(instruments, drawing_info)
            
            success_message = f"Successfully extracted {len(instruments)} pressure instrument(s) and generated datasheet"
            
            return excel_file, instruments, success_message
            
        except Exception as e:
            logger.error(f"[PressureInstrument] Datasheet generation error: {str(e)}")
            raise Exception(f"Failed to generate datasheet: {str(e)}")
