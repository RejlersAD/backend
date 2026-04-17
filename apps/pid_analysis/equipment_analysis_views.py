"""
Equipment Analysis Views - P&ID Equipment List Extraction
"""

import json
import logging
import os
import re
import uuid
from functools import lru_cache

from django.http import HttpResponse
from rest_framework import status as drf_status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

# Lazy import — avoids circular import at module load; models resolved at request time.
def _get_equipment_models():
    from apps.pid_analysis.models import PIDEquipmentType, PIDEquipmentItem  # noqa
    return PIDEquipmentType, PIDEquipmentItem

# Lazy import — avoids circular import at module load; models resolved at request time.
def _get_equipment_models():
    from apps.pid_analysis.models import PIDEquipmentType, PIDEquipmentItem  # noqa
    return PIDEquipmentType, PIDEquipmentItem

logger = logging.getLogger(__name__)

_CONFIG_PATH = os.path.join(
    os.path.dirname(__file__), 'config', 'equipment_type_config.json'
)


@lru_cache(maxsize=1)
def _load_config() -> dict:
    try:
        with open(_CONFIG_PATH, 'r', encoding='utf-8') as fh:
            raw = json.load(fh)
        return {k: v for k, v in raw.items() if not k.startswith('_')}
    except Exception as exc:
        logger.warning('[EquipmentList] Could not load config: %s - using defaults', exc)
        return {
            'extraction': {'context_window_chars': 120, 'description_max_words': 5},
            'type_labels': {
                'V': 'Vessel', 'P': 'Pump', 'E': 'Heat Exchanger', 'T': 'Tank',
                'K': 'Compressor', 'C': 'Column / Tower', 'H': 'Heater / Cooler',
                'D': 'Drum / Separator', 'R': 'Reactor',
            },
            'fluid_keywords': ['crude', 'gas', 'oil', 'water', 'steam'],
            'excel_columns': [
                {'key': 'sl_no',            'label': 'S. No',             'width': 6 },
                {'key': 'tag',              'label': 'Tag Number',         'width': 14},
                {'key': 'type_label',       'label': 'Equipment Type',     'width': 22},
                {'key': 'description',      'label': 'Description',        'width': 30},
                {'key': 'drawing_ref',      'label': 'Drawing Reference',  'width': 22},
                {'key': 'line_connections', 'label': 'Line Connections',   'width': 30},
                {'key': 'service_fluid',    'label': 'Service / Fluid',    'width': 20},
            ],
        }


_LINE_TAG_RE = re.compile(
    r'(?<![A-Za-z0-9])'
    r'(\d+(?:\.\d+)?)\s*["\u201c\u201d\u2019\'`]{1,2}'
    r'[\s\-_]{0,3}([A-Z]{1,4})[\s\-_]+(\d{3,6})[\s\-_]+(\d{4,8})'
    r'(?:[\s\-_]+([A-Z0-9]{1,8}))?'
    r'(?![A-Za-z0-9])',
    re.IGNORECASE,
)


def _normalize_text(text: str) -> str:
    """
    Normalize Unicode variants that CAD tools write in place of ASCII characters.
    Runs on every extracted text string before regex matching.

    Soft-coded by category — add mappings here without touching callers.
    """
    # Non-ASCII hyphens / dashes → ASCII hyphen  (most common CAD encoding issue)
    # U+2010 HYPHEN, U+2011 NON-BREAKING HYPHEN, U+2012/2013/2014 DASHES,
    # U+2212 MINUS SIGN, U+FE63 SMALL HYPHEN-MINUS, U+FF0D FULLWIDTH HYPHEN
    UNICODE_HYPHENS = '\u2010\u2011\u2012\u2013\u2014\u2212\ufe63\uff0d'
    for ch in UNICODE_HYPHENS:
        text = text.replace(ch, '-')
    # Non-breaking space → regular space
    text = text.replace('\u00a0', ' ').replace('\u202f', ' ')
    # Collapse repeated ASCII hyphens (OCR artifact: "V-803--TF" → "V-803-TF")
    import re as _re
    text = _re.sub(r'-{2,}', '-', text)
    return text


def _extract_text_from_pdf(file_obj, config=None, _page_index=None) -> str:
    """
    Extract all text from a PDF with three progressive strategies.

    _page_index (int | None): when specified (0-based), extract only that page.
    When None (default), all pages are processed — original behaviour.

    Strategy 1 — block text  : get_text('text') reading-order blocks (fast).
    Strategy 2 — spatial words: get_text('words') sorted by (y-bucket, x) to
        reconstruct left→right order regardless of CAD stream order.
    Strategy 3 — span proximity: iterate spans/chars to bond fragments that a
        CAD tool stored as separate micro-elements (e.g. "V-308" + "-TF").

    All three are run for EVERY vector PDF and their results concatenated so
    the downstream regex sees the text in every possible form.

    Scanned PDFs: OCR fallback (Tesseract) triggered when the combined vector
    text is shorter than the soft-coded min_vector_chars threshold.

    All config values are soft-coded in equipment_type_config.json.
    """
    cfg        = config or {}
    ext_cfg    = cfg.get('extraction', {})
    ocr_angles = ext_cfg.get('ocr_rotation_angles', [0, 90])
    ocr_psm_modes  = ext_cfg.get('ocr_psm_modes', [11, 6])
    ocr_scale      = float(ext_cfg.get('ocr_render_scale', 3.0))
    # Soft-coded: primary render scale for large-format pages (A0/A1 P&IDs)
    _OCR_SCALE_LARGE      = float(ext_cfg.get('ocr_render_scale_large_format', 4.0))
    # Soft-coded: additional full-page scales pooled on large pages
    _OCR_EXTRA_SCALES     = [float(s) for s in ext_cfg.get('ocr_additional_scales_large_format', [2.0])]
    _LARGE_PAGE_THRESHOLD = float(ext_cfg.get('ocr_large_page_threshold_pts', 900))
    # Soft-coded: tile-based OCR grid for large-format pages
    _TILE_ROWS     = int(ext_cfg.get('ocr_tile_rows', 3))
    _TILE_COLS     = int(ext_cfg.get('ocr_tile_cols', 4))
    _TILE_SCALE    = float(ext_cfg.get('ocr_tile_scale', 3.0))
    _TILE_PSM      = int(ext_cfg.get('ocr_tile_psm', 6))
    _TILE_OVERLAP  = float(ext_cfg.get('ocr_tile_overlap_frac', 0.12))
    # Soft-coded: vertical bucket height (pts) for spatially-sorted word pass
    _Y_BUCKET_PTS  = int(ext_cfg.get('spatial_word_y_bucket_pts', 15))
    # Soft-coded: max x-gap (pts) to bond two horizontally adjacent span fragments
    _SPAN_BOND_GAP = float(ext_cfg.get('span_bond_gap_pts', 20.0))
    # Soft-coded: threshold below which OCR fallback is triggered (chars)
    _MIN_VECTOR    = int(ext_cfg.get('min_vector_chars_for_ocr_skip', 200))
    # Soft-coded: always append OCR results even when vector text is long enough
    _ALWAYS_OCR    = bool(ext_cfg.get('always_include_ocr', True))

    text_parts: list = []
    file_bytes = None

    try:
        import fitz
        file_bytes = file_obj.read()
        doc = fitz.open(stream=file_bytes, filetype='pdf')

        for _pg_idx, page in enumerate(doc):
            if _page_index is not None and _pg_idx != _page_index:
                continue

            # ── Strategy 1: block text ──────────────────────────────────
            # TEXT_DEHYPHENATE excluded: structural hyphens in tags (V-308-TF)
            # must NOT be removed when they span a line boundary in the stream.
            blk_text = _normalize_text(page.get_text('text') or '')
            text_parts.append(blk_text)

            # ── Strategy 2: spatially-sorted word tokens ────────────────
            # get_text('words') → (x0,y0,x1,y1, word, block, line, word_no)
            words_raw = page.get_text('words')
            if words_raw:
                spatial = sorted(
                    words_raw,
                    key=lambda w: (round(w[1] / _Y_BUCKET_PTS) * _Y_BUCKET_PTS, w[0]),
                )
                spatial_text = _normalize_text(' '.join(w[4] for w in spatial))
                text_parts.append(spatial_text)

            # ── Strategy 3: span proximity bonding ─────────────────────
            # CAD tools (AutoCAD, SmartPlant, AVEVA) often write each word or
            # sub-token as an independent text span with a small positional gap.
            # get_text('words') treats a gap as a word boundary, so "V-308-TF"
            # may arrive as ["V-308", "-TF"] → joined with space → "V-308 -TF"
            # which breaks the regex.
            #
            # This pass iterates over character-level spans, sorts them by
            # (y-bucket, x) and bonds adjacent fragments whose right-edge to
            # next-left-edge gap is ≤ _SPAN_BOND_GAP pts, producing the
            # reconstructed token before adding a space.
            try:
                span_tokens: list = []  # (x0, reconstructed_text)
                raw_dict = page.get_text('rawdict')
                for blk in raw_dict.get('blocks', []):
                    for ln in blk.get('lines', []):
                        for sp in ln.get('spans', []):
                            txt = (sp.get('text') or '').strip()
                            if not txt:
                                continue
                            x0  = float(sp['bbox'][0])
                            y0  = float(sp['bbox'][1])
                            x1  = float(sp['bbox'][2])
                            row = round(y0 / _Y_BUCKET_PTS) * _Y_BUCKET_PTS
                            span_tokens.append((row, x0, x1, txt))

                span_tokens.sort(key=lambda t: (t[0], t[1]))

                bonded_parts: list = []
                buf = ''
                last_x1 = None
                last_row = None

                for row, x0, x1, txt in span_tokens:
                    if last_row is None:
                        buf = _normalize_text(txt)
                        last_x1 = x1
                        last_row = row
                    elif row == last_row and last_x1 is not None and (x0 - last_x1) <= _SPAN_BOND_GAP:
                        # Bond: adjacent on same row without visible gap
                        buf += _normalize_text(txt)
                        last_x1 = x1
                    else:
                        if buf:
                            bonded_parts.append(buf)
                        buf = _normalize_text(txt)
                        last_x1 = x1
                        last_row = row

                if buf:
                    bonded_parts.append(buf)

                if bonded_parts:
                    text_parts.append(' '.join(bonded_parts))
            except Exception as exc:
                logger.debug('[EquipmentList] Span bond pass failed: %s', exc)

        doc.close()
    except Exception as exc:
        logger.debug('[EquipmentList] PyMuPDF issue: %s', exc)

    full_text = '\n'.join(text_parts).strip()
    print(f'[EQ-DIAG] Vector text len={len(full_text)}  preview={repr(full_text[:200])}', flush=True)

    if _ALWAYS_OCR or len(full_text) < _MIN_VECTOR:
        # ── OCR fallback ─────────────────────────────────────────────────
        try:
            import fitz
            import pytesseract
            from PIL import Image, ImageEnhance, ImageFilter
            import io

            if file_bytes is None:
                file_obj.seek(0)
                file_bytes = file_obj.read()
            doc = fitz.open(stream=file_bytes, filetype='pdf')
            ocr_parts: list = []
            for _pg_idx, page in enumerate(doc):
                if _page_index is not None and _pg_idx != _page_index:
                    continue
                _r = page.rect
                _page_min_dim = min(abs(_r.width), abs(_r.height))
                _is_large = _page_min_dim > _LARGE_PAGE_THRESHOLD

                # Build list of scales to run for this page:
                # large-format pages run the primary large scale PLUS any
                # additional scales (e.g. 4.0 + 2.0) — different scales produce
                # different word-boundary decisions in Tesseract so pooling them
                # maximises tag coverage.
                if _is_large:
                    _scales_to_run = [_OCR_SCALE_LARGE] + _OCR_EXTRA_SCALES
                else:
                    _scales_to_run = [ocr_scale]

                seen_snippets: set = set()
                for _run_scale in _scales_to_run:
                    print(f'[EQ-DIAG] OCR page min_dim={_page_min_dim:.0f}  scale={_run_scale}', flush=True)
                    mat      = fitz.Matrix(_run_scale, _run_scale)
                    pix      = page.get_pixmap(matrix=mat, colorspace=fitz.csGRAY)
                    base_img = Image.open(io.BytesIO(pix.tobytes('png')))
                    _effective_dpi = int(72 * _run_scale)
                    base_img = ImageEnhance.Contrast(base_img).enhance(1.8)
                    base_img = base_img.filter(ImageFilter.SHARPEN)

                    for angle in ocr_angles:
                        rotated = base_img.rotate(-angle, expand=True) if angle != 0 else base_img
                        for psm in ocr_psm_modes:
                            ocr_text = pytesseract.image_to_string(
                                rotated, config=f'--oem 1 --psm {psm} --dpi {_effective_dpi}'
                            )
                            if not ocr_text.strip():
                                continue
                            fingerprint = ' '.join(ocr_text.split())[:200]
                            if fingerprint not in seen_snippets:
                                seen_snippets.add(fingerprint)
                                ocr_parts.append(_normalize_text(ocr_text))

                # ── Tile-based OCR pass for large-format pages ────────────────
                # Splits the page into a grid of tiles and OCRs each tile
                # independently.  Tesseract processes smaller, focused regions
                # more accurately than a single A0-scale image, so this pass
                # catches equipment tags in dense areas of the drawing that the
                # full-page pass misses.
                if _is_large and _TILE_ROWS > 0 and _TILE_COLS > 0:
                    print(f'[EQ-DIAG] Tiling {_TILE_ROWS}x{_TILE_COLS} scale={_TILE_SCALE} psm={_TILE_PSM}', flush=True)
                    _tile_dpi = int(72 * _TILE_SCALE)
                    _tmat  = fitz.Matrix(_TILE_SCALE, _TILE_SCALE)
                    _tpix  = page.get_pixmap(matrix=_tmat, colorspace=fitz.csGRAY)
                    _tfull = Image.open(io.BytesIO(_tpix.tobytes('png')))
                    _tfull = ImageEnhance.Contrast(_tfull).enhance(2.0)
                    _tfull = _tfull.filter(ImageFilter.SHARPEN)
                    _tw, _th = _tfull.size
                    for _ri in range(_TILE_ROWS):
                        for _ci in range(_TILE_COLS):
                            _x0 = max(0, int(_ci * _tw / _TILE_COLS - _tw * _TILE_OVERLAP / 2))
                            _y0 = max(0, int(_ri * _th / _TILE_ROWS - _th * _TILE_OVERLAP / 2))
                            _x1 = min(_tw, int((_ci + 1) * _tw / _TILE_COLS + _tw * _TILE_OVERLAP / 2))
                            _y1 = min(_th, int((_ri + 1) * _th / _TILE_ROWS + _th * _TILE_OVERLAP / 2))
                            _tile = _tfull.crop((_x0, _y0, _x1, _y1))
                            _tile_text = pytesseract.image_to_string(
                                _tile, config=f'--oem 1 --psm {_TILE_PSM} --dpi {_tile_dpi}'
                            )
                            if not _tile_text.strip():
                                continue
                            _fp = ' '.join(_tile_text.split())[:200]
                            if _fp not in seen_snippets:
                                seen_snippets.add(_fp)
                                ocr_parts.append(_normalize_text(_tile_text))
            doc.close()
            ocr_combined = '\n'.join(ocr_parts)
            print(f'[EQ-DIAG] OCR text len={len(ocr_combined)}  preview={repr(ocr_combined[:200])}', flush=True)
            # When always_include_ocr=true, APPEND to existing vector text.
            # When it's a pure OCR fallback (vector text was too short), REPLACE.
            if _ALWAYS_OCR and full_text:
                full_text = full_text + '\n' + ocr_combined
            else:
                full_text = ocr_combined
        except Exception as exc:
            logger.debug('[EquipmentList] Tesseract fallback issue: %s', exc)
            print(f'[EQ-DIAG] Tesseract fallback error: {exc}', flush=True)

    return full_text




# ---------------------------------------------------------------------------
# Equipment Register (18-field tabular document) extraction
# All thresholds / field-header variants are in equipment_type_config.json
# ---------------------------------------------------------------------------

_PAGE_Y_OFFSET      = 50000   # Vertical offset per PDF page so rows stay distinct
_Y_CLUSTER_TOL     = 12      # px — words within this y-distance are on the same row (vector PDF)
_Y_CLUSTER_TOL_OCR = 22      # px — wider tolerance for OCR; coords can drift more on scanned pages


def _cluster_words_into_rows(word_triples: list, y_tol: int = _Y_CLUSTER_TOL) -> list:
    """
    word_triples: list of (text, x, y) — may span multiple pages.
    Returns sorted list-of-rows, each row = [(text, x, y), ...] sorted by x.
    """
    if not word_triples:
        return []

    word_triples = sorted(word_triples, key=lambda w: (round(w[2] / _PAGE_Y_OFFSET), w[2]))
    rows: list = []
    current: list = [word_triples[0]]
    row_y = word_triples[0][2]

    for item in word_triples[1:]:
        # Treat items on different pages as always new rows
        same_page = abs(item[2] - row_y) < _PAGE_Y_OFFSET // 2
        if same_page and abs(item[2] - row_y) <= y_tol:
            current.append(item)
        else:
            rows.append(sorted(current, key=lambda w: w[1]))
            current = [item]
            row_y = item[2]

    if current:
        rows.append(sorted(current, key=lambda w: w[1]))
    return rows


def _extract_words_with_coords(file_obj, config: dict) -> tuple:
    """
    Returns (word_triples, used_ocr).
    word_triples: [(text, x, y), ...] from the PDF.
    Tries vector (PyMuPDF) first; falls back to pytesseract image_to_data.
    """
    cfg       = config.get('extraction', {})
    ocr_scale = float(cfg.get('ocr_render_scale', 3.0))
    word_list: list = []

    try:
        import fitz
        file_bytes = file_obj.read()
        doc = fitz.open(stream=file_bytes, filetype='pdf')
        for page_num, page in enumerate(doc):
            for entry in page.get_text('words'):
                x0, y0, x1, y1, word = entry[0], entry[1], entry[2], entry[3], entry[4]
                w = word.strip()
                if w:
                    word_list.append((w, x0, y0 + page_num * _PAGE_Y_OFFSET))
        doc.close()
    except Exception as exc:
        logger.debug('[EquipRegister] PyMuPDF words error: %s', exc)

    if len(word_list) > 30:
        return word_list, False  # vector PDF — use directly

    # ------- OCR fallback -------
    try:
        import fitz, pytesseract, io
        from PIL import Image, ImageEnhance, ImageFilter

        file_obj.seek(0)
        file_bytes = file_obj.read()
        doc = fitz.open(stream=file_bytes, filetype='pdf')
        ocr_words: list = []

        for page_num, page in enumerate(doc):
            mat = fitz.Matrix(ocr_scale, ocr_scale)
            pix = page.get_pixmap(matrix=mat, colorspace=fitz.csGRAY)
            img = Image.open(io.BytesIO(pix.tobytes('png')))
            img = ImageEnhance.Contrast(img).enhance(2.0)
            img = img.filter(ImageFilter.SHARPEN)

            # Try 0° first; also 90° for landscape CAD title blocks
            for angle in cfg.get('ocr_rotation_angles', [0, 90, 180, 270])[:2]:
                rotated = img.rotate(-angle, expand=True) if angle else img
                try:
                    data = pytesseract.image_to_data(
                        rotated,
                        config='--oem 1 --psm 6',
                        output_type=pytesseract.Output.DICT,
                    )
                    for i, word in enumerate(data['text']):
                        w = str(word).strip()
                        raw_conf = data['conf'][i]
                        conf = int(raw_conf) if str(raw_conf).lstrip('-').isdigit() else 0
                        if w and conf > 20:
                            x = float(data['left'][i]) / ocr_scale
                            y = float(data['top'][i]) / ocr_scale + page_num * _PAGE_Y_OFFSET
                            ocr_words.append((w, x, y))
                except Exception:
                    continue

        doc.close()
        return ocr_words, True

    except Exception as exc:
        logger.debug('[EquipRegister] OCR fallback error: %s', exc)
        return [], True


def _norm_header(text: str) -> str:
    """Normalise a column header for fuzzy matching: uppercase, collapse punctuation/spaces.

    Strips & so 'P&ID' and 'P & ID' and 'P ID' all normalise to the same 'P ID' form,
    which means variants only need to cover the stripped form once.
    """
    s = text.upper()
    s = re.sub(r'[.\-/()\[\]&,]', ' ', s)   # & added: P&ID → P ID, A&E → A E
    s = re.sub(r'\s+', ' ', s).strip()
    return s


# ── Soft-coded revision normalisation constants ──────────────────────────────
# Adjust these to cover new drawing conventions without touching logic.
#
# _REVISION_STRIP_PREFIX_RE  — removes leading "Rev"/"Revision"/"Rev." prefix
#                               common in registers that repeat the column label
#                               inside the cell (e.g. "Rev A" → "A").
# _REVISION_VALID_RE         — first match is the cleaned revision mark.
#                               Covers: single letter (A-Z), digit (0-9),
#                               letter+digit (A1, P0, R1), digit+letter (0A),
#                               3-letter IFC codes (IFC, IFD, IFA, IFR, AFD).
# _REVISION_MAX_RAW_LEN      — raw cell text longer than this is unlikely to be
#                               a revision mark; skip to avoid capturing dates or
#                               description bleed-over.
_REVISION_STRIP_PREFIX_RE = re.compile(
    r'^(?:rev(?:ision)?\s*(?:no\.?\s*|mark\s*|code\s*|\.\s*)?)',
    re.IGNORECASE,
)
_REVISION_VALID_RE = re.compile(
    r'\b(IF[CDABR]|AF[CD]|[A-Z]{1,2}[0-9]?|[0-9]{1,2}[A-Z]?)\b'
)
# Rejects date-like strings (e.g. 24/03/2025) that bleed from adjacent date columns.
_REVISION_DATE_RE = re.compile(r'\d[/\-.:]\d')
_REVISION_MAX_RAW_LEN = 20

# ── Soft-coded register extraction constants ──────────────────────────────────
# _REGISTER_HEADER_SCAN_ROWS — number of coordinate-sorted rows to scan when
#                              searching for the equipment register header.
#                              Set high (2000) so multi-page CRS/cover sheets
#                              before the equipment table are always skipped in
#                              favour of the real register header, which scores
#                              much higher (10+ columns vs 2-3 CRS columns).
_REGISTER_HEADER_SCAN_ROWS = 2000

# _REGISTER_HEADER_MIN_SCORE — minimum number of recognised fields the best
#                              header row must match before register mode is
#                              accepted.  Raised above min_cols (4) so that a
#                              CRS table whose header shares only "Rev" and
#                              "Description" with the equipment register schema
#                              is rejected and causes a fall-back to P&ID mode.
#                              Typical equipment register headers match 8-14
#                              fields; CRS headers match 2-4.
_REGISTER_HEADER_MIN_SCORE = 6

# _REGISTER_TAG_FILTER_RE    — after rows are extracted, any row whose 'tag'
#                              value does NOT contain this pattern is discarded
#                              as a footnote, note, or separator row.
#                              Equipment tags always look like "X-NNN" or
#                              "XX-NNNx" (1-2 cap letters + hyphen + 2-5 digits).
#                              This removes garbage like "units are in mm unless",
#                              "Pacakge document for" (note rows), ITEM-001
#                              fallback placeholders, and duplicate-text OCR noise.
#                              Set to None to disable filtering.
_REGISTER_TAG_FILTER_RE = re.compile(r'\b[A-Z]{1,2}-[0-9]{2,5}')

# _REGISTER_REPEATED_HDR_MARGIN — how much the header-match score must exceed
#   min_cols before a data row is treated as a repeated column-header and skipped.
#   Short variants like "R" (revision) and "NO" (sl_no) cause substring false-
#   positives against description text, so data rows routinely score 3-4.
#   Real repeated headers (same words as the header row) score 10+.
_REGISTER_REPEATED_HDR_MARGIN = 3

# _REV_PRE_TAG_WIN_CHARS     — characters to read BEFORE the matched tag in the
#                              raw OCR/vector text, used to find the revision cell
#                              in tabular PDFs where columns are newline-separated.
# _REV_PRE_TAG_TOKENS        — maximum number of newline-split tokens to inspect.
_REV_PRE_TAG_WIN_CHARS = 80
_REV_PRE_TAG_TOKENS    = 3

# _REVISION_USE_TOPMOST      — when True, the first non-empty revision value
#                              found in the register (topmost row) is applied to
#                              ALL extracted rows.  Equipment registers typically
#                              carry one document revision; individual rows should
#                              all reflect the current (topmost) revision mark.
#                              Set to False to keep per-row revision values.
_REVISION_USE_TOPMOST = True

# _HEADER_MAX_SPAN_ROWS      — maximum number of consecutive rows that can form
#                              a table column header.  CAD equipment registers
#                              commonly split long column labels across 3 lines
#                              (e.g. "Des./Set" / "Press." / "Min (PSIG)").
#                              Raising this from 2 → 3 ensures the qualifier
#                              row ("Min"/"Max") is included when building the
#                              column-x map, fixing extraction of
#                              design_pressure_min and design_pressure_max.
_HEADER_MAX_SPAN_ROWS = 3

# ── Soft-coded title-block revision extraction patterns ───────────────────────
# Used in P&ID drawing mode to extract the DOCUMENT revision from the title
# block, which applies uniformly to all equipment on the drawing.
#
# _TITLEBLOCK_REV_LABEL_RE   — explicit "REV" / "REVISION" label in title block,
#                              followed by the revision mark.  Capture group 1
#                              is the revision value.
# _TITLEBLOCK_DRAWN_CTX_RE   — title-block revision-history row pattern:
#                              a single revision mark that appears in a line
#                              containing DR(AWN)/CH(ECKED)/AP(PROVED) keywords
#                              (e.g. "A  IFR  12/04/2025  MAK  AKR  HJS").
#                              The revision mark is always the first short token.
# _TITLEBLOCK_ISOLATED_RE    — last-resort: single isolated letter/digit on its
#                              own line that appears within a few lines of the
#                              drawing number pattern (project doc-no format).
_TITLEBLOCK_REV_LABEL_RE   = re.compile(
    r'(?:^|\n)\s*REV(?:ISION)?\.?\s*[:\-]\s*([A-Z0-9]{1,3})\s*(?:\n|$)',
    re.IGNORECASE | re.MULTILINE,
)
# _TITLEBLOCK_REVTABLE_ROW_RE — matches a revision-history table row in the
# O&G title block convention:
#   REV_MARK  DD/MM/YYYY  ISSUED FOR .../APPROVED FOR .../RE-APPROVED FOR ...
# Captures groups: (1) REV mark, (2) day, (3) month, (4) year.
# Strategy 0 in _extract_titleblock_revision finds ALL matches and returns
# the mark whose date is the LATEST — correct regardless of OCR read order
# (rows may appear oldest→newest or newest→oldest depending on tile/scale).
# Handles both numeric (0, 1, 2) and alpha (A, B, C, IFC) revision marks.
_TITLEBLOCK_REVTABLE_ROW_RE = re.compile(
    r'(?:^|\s)([0-9]{1,2}[A-Z]?|[A-Z]{1,3})\s+'
    r'(\d{2})[/\-](\d{2})[/\-](\d{4})\s+'
    r'(?:ISSUED|APPROVED|RE[\s\-]?APPROVED|RETURNED|INCORPORATED)',
    re.IGNORECASE,
)
_TITLEBLOCK_DRAWN_CTX_RE   = re.compile(
    r'(?:DR[\'.]?N|DRW|DRAWN|CH[\'.]?D|CHK|CHECKED|APP?[\'.]?D|APPROVED)',
    re.IGNORECASE,
)
# Matches document numbers such as PJ6-EXD-MRI-BQDA-0023 (4–5 hyphen segments,
# last segment is 4–6 digits, each segment is 2–6 alphanumeric chars).
# The first segment may contain digits (e.g. PJ6) so [A-Z0-9]+ is used.
# Anchored with word boundary; minimum total length 10 to avoid short tags.
_TITLEBLOCK_DWG_NO_RE      = re.compile(
    r'\b([A-Z0-9]{2,6}(?:-[A-Z0-9]{2,6}){3,4})\b',
    re.IGNORECASE,
)
# Label that precedes the drawing number in the title block
_TITLEBLOCK_DWG_LABEL_RE   = re.compile(
    r'(?:DWG\.?\s*NO\.?|DRAWING\s*NO\.?|DOCUMENT\s*NO\.?|DOC\.?\s*NO\.?)',
    re.IGNORECASE,
)

# ── Soft-coded operating temperature range normalisation ─────────────────────
# Equipment registers sometimes store two operating temperatures in a single
# cell (e.g. shell/tube, inlet/outlet, or min/max condition) separated by "/".
# e.g. "105/60 °F" → "60 – 105 °F"  (ascending range, engineering convention)
_TEMP_RANGE_SEPARATOR = ' \u2013 '   # en-dash with spaces  (matches frontend constant)
_TEMP_SLASH_RE        = re.compile(
    r'^(-?\d+(?:\.\d+)?)\s*/\s*(-?\d+(?:\.\d+)?)\s*(°[FC]|DEG\s*[FC]|[FC])?$',
    re.IGNORECASE,
)


def _normalize_oper_temp(raw: str) -> str:
    """
    Normalise an operating temperature string that contains two values
    separated by "/" into a clean ascending range.

    Examples:
        "105/60 °F"  ->  "60 – 105 °F"
        "60 / 105"   ->  "60 – 105 °F"   (°F assumed)
        "175 °F"     ->  "175 °F"         (unchanged)
        ""           ->  ""               (unchanged)
    """
    if not raw or '/' not in raw:
        return raw
    m = _TEMP_SLASH_RE.match(raw.strip())
    if not m:
        return raw
    v1, v2 = float(m.group(1)), float(m.group(2))
    raw_unit = (m.group(3) or '').strip().upper()
    # Normalise unit display
    if raw_unit in ('F', 'DEGF') or raw_unit.endswith('F'):
        unit = '°F'
    elif raw_unit in ('C', 'DEGC') or raw_unit.endswith('C'):
        unit = '°C'
    else:
        unit = '°F'   # default for process equipment
    lo, hi = sorted([v1, v2])
    fmt = lambda v: str(int(v)) if v == int(v) else str(v)   # strip ".0" suffix
    return f'{fmt(lo)}{_TEMP_RANGE_SEPARATOR}{fmt(hi)} {unit}'


def _extract_titleblock_dwg_no(text: str) -> str:
    """
    Extract the drawing / document number from a P&ID title block.

    Strategy 1: look for the label 'DWG. NO.' / 'DRAWING NO.' followed by
        a document-number pattern within 120 chars.
    Strategy 2: find the most frequent multi-segment document number
        (4–5 segments, ends in 4-digit sequence) anywhere in the text.
        Filters out equipment tags (≤3 chars prefix + ≤5 digits) and
        change-request numbers.

    Returns the extracted drawing number string, or '' if not found.
    """
    # Strategy 1: label-adjacent.
    # Skip "FEED DRAWING NO." — that is a reference-document label, not the
    # main title-block drawing number.  A negative lookback of 25 chars
    # handles both "FEED DRAWING NO." and "FROM FEED DRAWING NO." forms.
    for lbl_m in _TITLEBLOCK_DWG_LABEL_RE.finditer(text):
        pre_ctx = text[max(0, lbl_m.start() - 25):lbl_m.start()]
        if re.search(r'\bFEED\b', pre_ctx, re.IGNORECASE):
            continue
        window = text[lbl_m.end():lbl_m.end() + 120]
        m = _TITLEBLOCK_DWG_NO_RE.search(window)
        if m:
            candidate = m.group(1).upper()
            # Must end with digits — excludes equipment tags like V-308-TF
            if re.search(r'-[0-9]{4,6}$', candidate):
                print(f'[EQ-DIAG][DwgNo] Found via label strategy: {candidate!r}', flush=True)
                return candidate

    # Strategy 2: most common 4–5 segment number ending in 4 digits
    from collections import Counter
    candidates: list = []
    for m in _TITLEBLOCK_DWG_NO_RE.finditer(text):
        cand = m.group(1).upper()
        if re.search(r'-[0-9]{4,6}$', cand) and len(cand) >= 10:
            candidates.append(cand)
    if candidates:
        most_common = Counter(candidates).most_common(1)[0][0]
        print(f'[EQ-DIAG][DwgNo] Found via frequency strategy: {most_common!r}', flush=True)
        return most_common

    return ''


def _extract_titleblock_revision(text: str) -> str:
    """
    Extract the DOCUMENT revision mark from a P&ID title block.

    Uses three progressive strategies (all soft-coded via module constants):

    Strategy 1 — Explicit label: looks for "REV[.] A" / "REVISION: B" patterns
        anywhere in the extracted text.  Returns the LAST such match because
        the topmost OCR text is often the legend/cover sheet; the title block
        with the current revision appears later.

    Strategy 2 — Drawn/Checked/Approved context: scans each line that contains
        DR'N / CH'D / AP'D keywords (the revision-history table in the title
        block) and extracts the first short token on that line as the revision.
        Takes the LAST such match (= the most recent / lowest-numbered row in
        the revision table, which in O&G conventions is the CURRENT revision).

    Strategy 3 — Isolated token near drawing number: looks for a standalone
        1-3 char alphanumeric token on lines adjacent to a project document
        number (pattern: two-letter-code–section–discipline–doc-no).

    All candidate values are validated by _clean_revision before return.
    Returns '' if no valid revision found.
    """
    if not text:
        return ''

    # Strategy 0: revision table row (most reliable — O&G title block convention)
    # Pattern: REV_MARK  DD/MM/YYYY  ISSUED/APPROVED FOR ...
    # Finds ALL matching rows, parses the date from each, and returns the mark
    # whose date is LATEST.  This is correct regardless of OCR read order —
    # some drawings OCR oldest→newest, others newest→oldest depending on how
    # tiles/spatial passes are merged.
    rev_row_matches = list(_TITLEBLOCK_REVTABLE_ROW_RE.finditer(text))
    if rev_row_matches:
        best_mark = ''
        best_date = (0, 0, 0)   # (year, month, day)
        for rm in rev_row_matches:
            cleaned = _clean_revision(rm.group(1))
            if not cleaned:
                continue
            try:
                day   = int(rm.group(2))
                month = int(rm.group(3))
                year  = int(rm.group(4))
                rev_date = (year, month, day)
            except (IndexError, ValueError):
                # Date parse failed — accept as fallback only
                if not best_mark:
                    best_mark = cleaned
                continue
            if rev_date > best_date:
                best_date = rev_date
                best_mark = cleaned
        if best_mark:
            print(
                f'[EQ-DIAG][TitleBlock] Rev found via table-row strategy: {best_mark!r} '
                f'(date {best_date[2]:02d}/{best_date[1]:02d}/{best_date[0]})',
                flush=True,
            )
            return best_mark

    # Strategy 1: explicit "REV: value" label (strict colon required)
    matches = _TITLEBLOCK_REV_LABEL_RE.findall(text)
    if matches:
        # Validate each match; take the last valid one
        for raw in reversed(matches):
            cleaned = _clean_revision(raw)
            if cleaned:
                print(f'[EQ-DIAG][TitleBlock] Rev found via label strategy: {cleaned!r}', flush=True)
                return cleaned

    # Strategy 2: drawn/checked/approved context rows
    lines = text.split('\n')
    last_rev_from_ctx = ''
    for line in lines:
        if _TITLEBLOCK_DRAWN_CTX_RE.search(line):
            tokens = [t.strip() for t in line.split() if t.strip()]
            for tok in tokens:
                cleaned = _clean_revision(tok)
                if cleaned:
                    last_rev_from_ctx = cleaned
                    break
    if last_rev_from_ctx:
        print(f'[EQ-DIAG][TitleBlock] Rev found via DR/CH/AP context: {last_rev_from_ctx!r}', flush=True)
        return last_rev_from_ctx

    # Strategy 3: isolated short token near project document number
    dwg_matches = list(_TITLEBLOCK_DWG_NO_RE.finditer(text))
    for dwg_m in dwg_matches:
        # Look at a ±300 char window around the drawing number
        win_start = max(0, dwg_m.start() - 300)
        win_end   = min(len(text), dwg_m.end() + 300)
        window    = text[win_start:win_end]
        for ln in window.split('\n'):
            stripped = ln.strip()
            if re.match(r'^[A-Z0-9]{1,3}$', stripped, re.IGNORECASE):
                cleaned = _clean_revision(stripped)
                if cleaned:
                    print(f'[EQ-DIAG][TitleBlock] Rev found via DWG-no proximity: {cleaned!r}', flush=True)
                    return cleaned

    print('[EQ-DIAG][TitleBlock] No document revision found in title block', flush=True)
    return ''


def _clean_revision(raw: str) -> str:
    """Normalise an extracted revision cell value to a short clean mark.

    Steps
    -----
    1. Strip whitespace; bail out if cell is too long to be a revision mark.
    2. Strip leading 'Rev'/'Revision' prefix (some registers duplicate the
       column label inside every cell, e.g. 'Rev A' → 'A').
    3. If the remaining text is already 1-3 chars and alphanumeric, return it.
    4. Otherwise scan for the first token matching _REVISION_VALID_RE.
    5. Return the match in uppercase, or '' if nothing valid found.

    Soft-coded via module-level constants:
      _REVISION_STRIP_PREFIX_RE, _REVISION_DATE_RE, _REVISION_VALID_RE, _REVISION_MAX_RAW_LEN
    """
    if not raw:
        return ''
    s = raw.strip()
    if len(s) > _REVISION_MAX_RAW_LEN:
        # Cell is too long to be a valid revision mark — likely a bleed-over
        # from an adjacent wide column (description, remarks).
        return ''
    # Reject date-like values (e.g. 24/03/2025) that bleed from adjacent columns
    if _REVISION_DATE_RE.search(s):
        return ''
    # Reject "Note N" or "(Note N)" bleed from description/remarks columns
    if re.search(r'\bnote\b', s, re.IGNORECASE):
        return ''
    # Strip 'Rev' / 'Revision' prefix
    s = _REVISION_STRIP_PREFIX_RE.sub('', s).strip()
    if not s:
        return ''
    # If already short and clean, return immediately (fast path)
    if re.match(r'^[A-Za-z0-9]{1,3}$', s):
        return s.upper()
    # Scan for first valid revision token in the (possibly noisy) remainder
    m = _REVISION_VALID_RE.search(s.upper())
    return m.group(1) if m else ''


def _find_header_range(rows: list, field_variants: dict, min_cols: int) -> tuple:
    """
    Scan first _REGISTER_HEADER_SCAN_ROWS rows for the table header row(s).
    Supports single-row and double-row headers (common in CAD documents).
    Returns (start_idx, end_idx_exclusive) or None if not found.

    Uses _REGISTER_HEADER_MIN_SCORE (≥ min_cols) so that low-scoring CRS/cover
    headers that share only "Rev" and "Description" with the equipment schema are
    rejected rather than used as a fallback.
    """
    scan_limit = min(_REGISTER_HEADER_SCAN_ROWS, len(rows))
    best_score = 0
    best_range: tuple = (0, 1)

    all_variants_norm = {
        k: [_norm_header(v) for v in variants]
        for k, variants in field_variants.items()
    }

    for start in range(scan_limit):
        for span in range(1, _HEADER_MAX_SPAN_ROWS + 1):
            end = min(start + span, len(rows))
            combined_norm = _norm_header(
                ' '.join(t for row in rows[start:end] for (t, x, y) in row)
            )
            score = sum(
                1 for variants_norm in all_variants_norm.values()
                if any(v in combined_norm for v in variants_norm)
            )
            if score > best_score:
                best_score = score
                best_range = (start, end)

    # Use the stricter _REGISTER_HEADER_MIN_SCORE threshold so that a CRS table
    # whose header only matches "Rev" + "Description" (score ≤ 3) is rejected.
    required_score = max(min_cols, _REGISTER_HEADER_MIN_SCORE)
    print(f'[EQ-DIAG][Register] header scan: total_rows={len(rows)} scan_limit={scan_limit}'
          f'  best_score={best_score}  best_range={best_range}  required={required_score}', flush=True)
    if best_score < required_score:
        return None
    return best_range


def _build_col_map(header_rows: list, field_variants: dict) -> dict:
    """
    Build mapping: field_key -> x_center from the header row(s).

    Handles multi-line CAD table headers by:
    1. Sorting all header words by (x, y) so same-column words are adjacent.
    2. Grouping into x-column clusters.
    3. Trying left-neighbor merges for short unmatched clusters (handles
       "Des./Set Press. Min" where Min lands in its own cluster).
    4. Greedy conflict resolution to avoid two clusters claiming the same field.
    """
    all_words_y = [(t.strip(), float(x), float(y))
                   for row in header_rows for (t, x, y) in row if t.strip()]
    if not all_words_y:
        return {}

    all_variants_norm = {
        k: [_norm_header(v) for v in variants]
        for k, variants in field_variants.items()
    }

    # ── Sort by x, then y ────────────────────────────────────────────────────
    sorted_by_x = sorted(all_words_y, key=lambda w: (w[1], w[2]))

    # ── Adaptive x-cluster tolerance ─────────────────────────────────────────
    distinct_xs = sorted(set(round(v[1]) for v in sorted_by_x))
    if len(distinct_xs) > 1:
        gaps = [distinct_xs[i + 1] - distinct_xs[i] for i in range(len(distinct_xs) - 1)]
        median_gap = sorted(gaps)[len(gaps) // 2]
        x_col_tol = max(median_gap * 0.8, 8.0)
    else:
        x_col_tol = 15.0

    # ── Form initial clusters ─────────────────────────────────────────────────
    col_clusters: list = []
    current: list = [sorted_by_x[0]]
    for we in sorted_by_x[1:]:
        cm = sum(w[1] for w in current) / len(current)
        if abs(we[1] - cm) <= x_col_tol:
            current.append(we)
        else:
            col_clusters.append(current)
            current = [we]
    if current:
        col_clusters.append(current)

    # ── Helper: build phrase + x-center from a cluster list ──────────────────
    def _cluster_info(cluster: list) -> tuple:
        ro = sorted(cluster, key=lambda w: (w[2], w[1]))
        phrase = _norm_header(' '.join(w[0] for w in ro))
        x_c = sum(w[1] for w in cluster) / len(cluster)
        return phrase, x_c

    # ── Helper: score a phrase against a single field ─────────────────────────
    def _score(phrase: str, field_key: str) -> int:
        best = 0
        for variant in all_variants_norm.get(field_key, []):
            if phrase == variant:
                s = len(variant) * 2
            elif len(variant) >= 3 and variant in phrase:
                s = len(variant)
            elif len(phrase) >= 3 and len(variant) >= 5 and phrase in variant:
                s = len(phrase)
            else:
                continue
            if s > best:
                best = s
        return best

    # ── Step 4: Build (score, x_center, field_key) candidates ────────────────
    # Each cluster produces ALL matching fields (not just best), then we do
    # greedy conflict-free assignment.  Also try LEFT-MERGE for short clusters
    # (catches "Min"/"Max" separated from their prefix by a gap).
    all_matches: list = []  # (score, x_center, field_key)

    for ci, cluster in enumerate(col_clusters):
        phrase, x_c = _cluster_info(cluster)

        # Also try merging with left neighbor (helps "Des./Set Press." + "Min")
        if ci > 0 and len(cluster) <= 2:
            merged = col_clusters[ci - 1] + cluster
            merged_phrase, merged_xc = _cluster_info(merged)
        else:
            merged_phrase, merged_xc = None, None

        for field_key in all_variants_norm:
            s = _score(phrase, field_key)
            use_x = x_c   # always use a local copy — do NOT mutate x_c
            # Prefer merged phrase only if it yields a strictly better score
            if merged_phrase is not None:
                ms = _score(merged_phrase, field_key)
                if ms > s:
                    s, use_x = ms, merged_xc
            if s > 0:
                all_matches.append((s, use_x, field_key))

    # ── Step 5: Greedy conflict-free assignment ───────────────────────────────
    # Sort by score desc, then by x (stable ordering for equal scores).
    all_matches.sort(key=lambda m: (-m[0], m[1]))

    field_best: dict = {}  # field_key -> (score, x_center)
    # Track which physical x-centers have already been "used" (±5pt tolerance)
    used_x: list = []

    for score, x_center, field_key in all_matches:
        if field_key in field_best:
            continue  # field already claimed
        # Check whether a different field already claimed this x_center
        already_used = any(abs(x_center - ux) < 5.0 for ux in used_x)
        if already_used:
            continue
        field_best[field_key] = (score, x_center)
        used_x.append(x_center)

    return {k: v[1] for k, v in field_best.items()}


def _assign_row_to_cols(data_row: list, col_map: dict) -> dict:
    """
    Assign each word in data_row to the nearest column by x-distance.
    Returns dict field_key -> value_string.
    """
    if not data_row or not col_map:
        return {}

    sorted_cols = sorted(col_map.items(), key=lambda c: c[1])   # (key, x)
    n_cols = len(sorted_cols)

    # Midpoints between adjacent columns
    boundaries = [
        (sorted_cols[i][1] + sorted_cols[i + 1][1]) / 2
        for i in range(n_cols - 1)
    ]

    buckets: dict = {k: [] for k in col_map}
    for (text, x, _y) in data_row:
        col_idx = 0
        for bi, bx in enumerate(boundaries):
            if x > bx:
                col_idx = bi + 1
            else:
                break
        assigned = sorted_cols[col_idx][0]
        buckets[assigned].append((x, text))

    return {
        k: ' '.join(txt for _, txt in sorted(items)).strip()
        for k, items in buckets.items()
        if items
    }


def _pid_item_to_register_schema(pid_item: dict) -> dict:
    """Map a P&ID-extraction item to the 18-field register schema."""
    return {
        'sl_no':               str(pid_item.get('sl_no', '')),
        'revision':            _clean_revision(str(pid_item.get('revision', ''))),
        'tag':                 pid_item.get('tag', ''),
        'description':         pid_item.get('description', ''),
        'design_flowrate':     pid_item.get('design_flowrate', ''),
        'oper_pressure':       pid_item.get('oper_pressure', ''),
        'oper_temperature':    pid_item.get('oper_temperature', ''),
        'design_pressure_min': pid_item.get('design_pressure_min', ''),
        'design_pressure_max': pid_item.get('design_pressure_max', ''),
        'design_temp_min':     pid_item.get('design_temp_min', ''),
        'design_temp_max':     pid_item.get('design_temp_max', ''),
        'moc':                 pid_item.get('material_class', ''),
        'insulation':          pid_item.get('insulation', ''),
        'dimension_length':    pid_item.get('dimension_length', ''),
        'dimension_diameter':  pid_item.get('dimension_diameter', ''),
        'motor_rating':        pid_item.get('motor_rating', ''),
        'pid_no':              pid_item.get('drawing_ref', ''),
        'quality_required':    pid_item.get('quality_required', ''),
        'phase':               pid_item.get('service_fluid', ''),
        'remarks':             pid_item.get('process_notes', ''),
        # Backward-compat fields kept for status/results endpoints
        'type_label':         pid_item.get('type_label', ''),
        'area':               pid_item.get('area', ''),
        'drawing_ref':        pid_item.get('drawing_ref', ''),
        'line_connections':   pid_item.get('line_connections', []),
        'nozzle_connections': pid_item.get('nozzle_connections', []),
        'service_fluid':      pid_item.get('service_fluid', ''),
        'material_class':     pid_item.get('material_class', ''),
        'process_notes':      pid_item.get('process_notes', ''),
    }


def _extract_equipment_register_rows(file_obj, config: dict):
    """
    Extract 18-field Equipment Register from a tabular CAD/PDF document.

    Uses coordinate-based table detection (PyMuPDF words + pytesseract
    image_to_data as fallback) so it works on both vector and scanned PDFs.

    Returns list of equipment dicts if the document is a register table,
    or None if the document doesn't look like a register (triggers P&ID fallback).
    """
    field_variants = config.get('equip_register_fields', {})
    min_cols       = int(config.get('equip_register_min_columns', 4))
    min_rows       = int(config.get('equip_register_min_rows', 2))
    # Soft-coded: shortest page dimension (pts) above which we treat the doc
    # as a large-format P&ID drawing and skip register detection entirely.
    # A4 landscape smallest dim = 595 pts; A3 = 842 pts; A1/A0 >> 1000 pts.
    # Equipment registers are A4; P&IDs are A1-A0. Threshold = 900 pts.
    max_drawing_min_dim = int(config.get('equip_register_skip_if_page_dim_gt', 900))

    if not field_variants:
        return None  # Config missing — skip register mode

    # ── Page-size guard ──────────────────────────────────────────────────────
    # Large-format drawings (A1/A0 P&IDs) can accidentally match headers from
    # equipment data boxes (DIAMETER, LENGTH, OPERATING PRESS, etc.).
    # Skip register mode when the smallest page dimension exceeds the threshold.
    try:
        import fitz as _fitz
        _fb = file_obj.read()
        _doc = _fitz.open(stream=_fb, filetype='pdf')
        if _doc.page_count > 0:
            _r = _doc[0].rect
            _min_dim = min(abs(_r.width), abs(_r.height))
            print(f'[EQ-DIAG][Register] page_min_dim={_min_dim:.0f}pts  threshold={max_drawing_min_dim}', flush=True)
            if _min_dim > max_drawing_min_dim:
                _doc.close()
                print('[EQ-DIAG][Register] Large-format drawing -> skipping register mode', flush=True)
                return None
        _doc.close()
        file_obj.seek(0)
    except Exception as _exc:
        logger.debug('[EquipRegister] Page-size check failed: %s', _exc)
        try:
            file_obj.seek(0)
        except Exception:
            pass

    logger.info('[EquipRegister] Starting coordinate-based table extraction')

    word_list, used_ocr = _extract_words_with_coords(file_obj, config)
    if not word_list:
        logger.info('[EquipRegister] No words extracted')
        return None

    # Use wider y-tolerance for OCR pages — coordinates are less precise
    y_tol = _Y_CLUSTER_TOL_OCR if used_ocr else _Y_CLUSTER_TOL
    rows = _cluster_words_into_rows(word_list, y_tol=y_tol)
    if len(rows) < 3:
        logger.info('[EquipRegister] Too few rows (%d)', len(rows))
        return None

    header_range = _find_header_range(rows, field_variants, min_cols)
    if header_range is None:
        logger.info('[EquipRegister] No register header detected')
        return None

    h_start, h_end = header_range
    col_map = _build_col_map(rows[h_start:h_end], field_variants)
    if len(col_map) < min_cols:
        logger.info('[EquipRegister] Too few columns mapped (%d)', len(col_map))
        return None

    logger.info('[EquipRegister] Columns detected: %s', list(col_map.keys()))

    # All variants for repeated-header detection
    all_variants_norm = {
        k: [_norm_header(v) for v in variants]
        for k, variants in field_variants.items()
    }

    equipment: list = []
    row_counter = 0

    for row in rows[h_end:]:
        row_text = ' '.join(t for (t, x, y) in row).strip()
        if not row_text or len(row_text) < 2:
            continue

        # Skip repeated header rows (some CAD drawings repeat headers each page)
        combined_norm = _norm_header(row_text)
        hdr_score = sum(
            1 for v_list in all_variants_norm.values()
            if any(v in combined_norm for v in v_list)
        )
        if hdr_score >= min_cols + _REGISTER_REPEATED_HDR_MARGIN:
            continue

        values = _assign_row_to_cols(row, col_map)
        tag_val = values.get('tag', '').strip()
        sl_val  = values.get('sl_no', '').strip()

        # ── Tag clean + rescue ────────────────────────────────────────────────
        # Full equipment tag pattern: 1-3 cap letters, hyphen, 2-5 digits,
        # optional alpha suffix (A, B, A/B/C), optional project suffix (-TF, -1F).
        # Soft-coded via _REGISTER_TAG_FULL_RE.  Applied in order:
        #  1. Strip trailing noise from column-assigned tag (e.g. "PX-851-TF MRD …")
        #  2. If no valid tag in the column value, check whether neighbouring words
        #     in the same row land a valid tag (handles x-coordinate drift that
        #     places the tag token in the revision column instead of the tag column).
        if tag_val:
            _tm = re.search(
                r'\b([A-Z]{1,3}-[0-9]{2,5}[A-Za-z]?(?:[/\-][A-Z0-9]{1,4})*)\b',
                tag_val
            )
            tag_val = _tm.group(1) if _tm else tag_val

        # Check unconditionally whether the revision column contains an equipment
        # tag that landed there due to x-coordinate drift.  When that happens the
        # real tag is in the rev cell and description text bleeds into the tag cell.
        # We accept the swap only when what remains after removing the tag from rev
        # is a plausible revision mark of 0-2 characters (e.g. "1", "A", "1A", "").
        _TAG_FULL_RE = re.compile(
            r'\b([A-Z]{1,3}-[0-9]{2,5}[A-Za-z]?(?:[/\-][A-Z0-9]{1,4})*)\b'
        )
        _rev_raw = values.get('revision', '')
        if _rev_raw:
            _rv_tag = _TAG_FULL_RE.search(_rev_raw)
            if _rv_tag:
                _cleaned_rev = _rev_raw.replace(_rv_tag.group(1), '').strip()
                if re.match(r'^[0-9A-Za-z]{0,2}$', _cleaned_rev):
                    tag_val = _rv_tag.group(1)
                    values['revision'] = _cleaned_rev
                    print(f'[EQ-DIAG][Register] rev-rescue: extracted tag {tag_val!r} '
                          f'from rev col; rev cleaned to {values["revision"]!r}', flush=True)

        if not _REGISTER_TAG_FILTER_RE.search(tag_val):
            # Tag column didn't yield a valid tag — scan the full row text (all
            # words ordered by x-position) and take the leftmost matching token.
            _row_plain = ' '.join(t for (t, x, y) in sorted(row, key=lambda w: w[1]))
            _rescue = re.search(
                r'\b([A-Z]{1,3}-[0-9]{2,5}[A-Za-z]?(?:[/\-][A-Z0-9]{1,4})*)\b',
                _row_plain
            )
            if _rescue:
                tag_val = _rescue.group(1)
                print(f'[EQ-DIAG][Register] tag-rescue: row "{_row_plain[:50]}" → {tag_val}', flush=True)
        # ── End tag clean + rescue ────────────────────────────────────────────

        # Count non-empty fields; skip blank/near-blank rows regardless of tag presence.
        # A row that has ≥ 2 mapped fields is kept even if neither tag nor sl_no are
        # populated — this handles: (a) registers where tag column wasn't matched,
        # (b) multi-line continuation rows caught here before the post-merge pass.
        populated = sum(1 for v in values.values() if v.strip())
        if populated < 2:
            continue

        row_counter += 1
        item: dict = {
            'sl_no':               sl_val or str(row_counter),
            'revision':            _clean_revision(values.get('revision', '')),
            'tag':                 tag_val or f'ITEM-{row_counter:03d}',
            'description':         values.get('description', ''),
            'design_flowrate':     values.get('design_flowrate', ''),
            'oper_pressure':       values.get('oper_pressure', ''),
            'oper_temperature':    _normalize_oper_temp(values.get('oper_temperature', '')),
            'design_pressure_min': values.get('design_pressure_min', ''),
            'design_pressure_max': values.get('design_pressure_max', ''),
            'design_temp_min':     values.get('design_temp_min', ''),
            'design_temp_max':     values.get('design_temp_max', ''),
            'moc':                 values.get('moc', ''),
            'insulation':          values.get('insulation', ''),
            'dimension_length':    values.get('dimension_length', ''),
            'dimension_diameter':  values.get('dimension_diameter', ''),
            'motor_rating':        values.get('motor_rating', ''),
            'pid_no':              values.get('pid_no', ''),
            'quality_required':    values.get('quality_required', ''),
            'phase':               values.get('phase', ''),
            'remarks':             values.get('remarks', ''),
            # Backward-compat fields
            'type_label':         '',
            'area':               '',
            'drawing_ref':        '',
            'line_connections':   [],
            'nozzle_connections': [],
            'service_fluid':      values.get('oper_pressure', ''),
            'material_class':     values.get('moc', ''),
            'process_notes':      values.get('remarks', ''),
        }
        equipment.append(item)

    # Post-merge: combine continuation rows (wrapped cells) into the preceding item.
    # A row is a continuation candidate if both tag and sl_no are empty and it has
    # ≤ 3 mapped fields — typically the second line of a wrapped description cell.
    if equipment:
        merged_equip: list = [equipment[0]]
        for item in equipment[1:]:
            is_cont = (
                not item.get('tag') and not item.get('sl_no')
                and sum(1 for v in item.values() if isinstance(v, str) and v.strip()) <= 3
            )
            if is_cont and merged_equip:
                prev = merged_equip[-1]
                for fld in ('description', 'remarks', 'moc', 'insulation'):
                    if item.get(fld):
                        prev[fld] = (prev.get(fld, '') + ' ' + item[fld]).strip()
            else:
                merged_equip.append(item)
        equipment = merged_equip

    # ── Tag-pattern filter ────────────────────────────────────────────────────
    # Remove footnote/separator rows whose 'tag' cell doesn't contain a real
    # equipment tag (e.g. "Note 1: units are in mm", "Water Treatment Package",
    # ITEM-NNN placeholders, duplicated OCR noise lines).
    # Controlled by _REGISTER_TAG_FILTER_RE; set to None to disable.
    if _REGISTER_TAG_FILTER_RE is not None:
        before_filter = len(equipment)
        equipment = [
            _item for _item in equipment
            if _REGISTER_TAG_FILTER_RE.search(_item.get('tag', ''))
        ]
        removed = before_filter - len(equipment)
        if removed:
            print(f'[EQ-DIAG][Register] Tag-filter removed {removed} non-equipment rows '
                  f'(footnotes/notes/separators), {len(equipment)} rows kept', flush=True)

    # Confirm enough populated rows to treat this as a real register
    well_populated = sum(
        1 for item in equipment
        if sum(
            1 for k, v in item.items()
            if k not in ('sl_no', 'tag', 'type_label', 'area', 'drawing_ref',
                         'line_connections', 'nozzle_connections')
            and (v if not isinstance(v, list) else v)
        ) >= 3   # lowered from 5 — sparse registers (tag + desc + one pressure) should qualify
    )

    if len(equipment) < min_rows or well_populated < min_rows:
        logger.info('[EquipRegister] Insufficient populated rows (total=%d, well_pop=%d)',
                    len(equipment), well_populated)
        return None

    # ── Topmost revision override ─────────────────────────────────────────────
    # Apply the topmost (first) non-empty revision value to ALL rows.
    # Equipment registers carry one document revision; individual row cells
    # often read wrongly due to column-coordinate drift or adjacent date bleed.
    # Controlled by _REVISION_USE_TOPMOST constant (True by default).
    if _REVISION_USE_TOPMOST and equipment:
        topmost_rev = next(
            (item['revision'] for item in equipment if item.get('revision')),
            ''
        )
        if topmost_rev:
            for item in equipment:
                item['revision'] = topmost_rev
            print(f'[EQ-DIAG][Register] Topmost-revision applied: "{topmost_rev}" → all {len(equipment)} rows', flush=True)

    logger.info('[EquipRegister] Extracted %d register rows (OCR=%s)', len(equipment), used_ocr)
    return equipment


# ---------------------------------------------------------------------------


# ---------------------------------------------------------------------------
# Data-box index — global scanner for equipment data boxes in P&ID drawings
# ---------------------------------------------------------------------------

# Soft-coded: chars to look BACK from the label to find the associated tag.
_DATABOX_TAG_LOOKBACK_CHARS = 1500


def _norm_databox_label(label: str) -> str:
    """Lowercase + collapse punctuation/whitespace for map key comparison."""
    s = label.lower()
    s = re.sub(r'[./()_\-]', ' ', s)
    s = re.sub(r'\s+', ' ', s).strip()
    return s


# Soft-coded: stop patterns for data-box value cleaning.
# When OCR merges a data-box value with adjacent text (e.g. NOTES on the same
# visual row), these patterns mark where the useful value ends.
# Order matters: more-specific patterns first.
# Soft-coded value-stop patterns for _clean_databox_value.
# Each branch marks where the useful numeric/unit value ends and junk begins.
_DATABOX_VALUE_STOPS = re.compile(
    r'(?:'
    r'\s+\d+\.\s+[A-Z]{3,}'            # note number: "  2. CORROSION..."
    r'|\bSLOPE\b'                      # slope annotation (SLOPE 1:200)
    r'|\bSOUR\s+GAS\b'                 # flow routing annotation
    r'|\bIN\s+SCOPE\b'                 # scope note
    r'|\bWILL\s+BE\b'                  # notes prose
    r'|\bSHALL\b'                      # notes prose
    r'|\bSCOPE\b'                      # scope note
    r'|\bHEADER\b'                     # pipe destination
    r'|[;|]'                           # column separator / pipe char
    r'|//'                             # // separating dual-condition values
    r'|\s{3,}'                         # 3+ spaces = OCR column gap
    r'|[\"\u201c\u201d][A-Z]{2,4}-'   # pipe tag starting with inch mark
    r')',
    re.IGNORECASE,
)

# Known data-box label starts — if the value contains one of these after the
# first word, truncate there (OCR ran two rows together).
_DATABOX_LABEL_STARTS_RE = re.compile(
    r'\b(?:NOMINAL|OPERATING|DESIGN|LENGTH|DIAMETER|MATERIAL|MOTOR|HEAT|DUTY'
    r'|CAPACITY|VOLUME|RATED|TRIM|INSULATION|WEIGHT|SHELL|INTERNALS?)\b',
    re.IGNORECASE,
)


def _clean_databox_value(raw: str) -> str:
    """
    Truncate a raw data-box extracted value at the first noise/stop pattern.

    Data-box values on large P&IDs are often followed on the same OCR line by:
    - NOTES text (e.g. "327 M3  2. CORROSION COUPON AND PROBE WILL BE...")
    - A second data-box label (OCR merged two rows)
    - Column separators (spaces, pipes, semicolons)

    Soft-coded via _DATABOX_VALUE_STOPS and _DATABOX_LABEL_STARTS_RE.
    Returns the cleaned, stripped value or '' if nothing useful remains.
    """
    if not raw:
        return ''

    # Truncate at first stop pattern
    stop_m = _DATABOX_VALUE_STOPS.search(raw)
    if stop_m:
        raw = raw[:stop_m.start()]

    # Truncate at the SECOND occurrence of a known label keyword
    # (first occurrence may be part of the value itself e.g. "SHELL DIAMETER")
    label_hits = list(_DATABOX_LABEL_STARTS_RE.finditer(raw))
    if len(label_hits) >= 2:
        raw = raw[:label_hits[1].start()]

    raw = raw.strip(' .,;:()')
    # Reject values that are pure punctuation or very short after cleaning
    if len(raw) < 1:
        return ''

    # ── OCR unit correction ──────────────────────────────────────────────
    # Tesseract commonly misreads 'M3' as 'MS', 'M²' as 'M2', 'm³' as 'm3'.
    # Soft-coded replacements applied ONLY to the unit portion (after a digit).
    _ocr_unit_fixes = [
        (re.compile(r'(\d+\.?\d*)\s*MS\b', re.IGNORECASE), r'\1 M3'),  # 327 MS → 327 M3
        (re.compile(r'\bM\s+3\b'),                              'M3'),   # M 3 → M3
        (re.compile(r'\bM3/[Hh]\b'),                            'M3/H'), # normalise
        (re.compile(r'"Fo\b'),                                   '°F'),   # "Fo → °F
        (re.compile(r'\'F\b'),                                   '°F'),   # 'F  → °F
        (re.compile(r'┬░'),                                      '°'),    # mojibake °
    ]
    for pat, repl in _ocr_unit_fixes:
        raw = pat.sub(repl, raw)

    # Remove stray OCR mojibake sequences
    raw = re.sub(r'\xc2\xb0|\u00c2\u00b0|ΓÇ[\x00-\xff]', '', raw)
    raw = raw.strip()

    # ── Dimension value normaliser ────────────────────────────────────────
    # When OCR merges a dimension value with the next label/annotation the
    # result is e.g. "5.0 M iN" or "15.0 M LONG".  Strip trailing alphabetic
    # junk that is not a recognised pressure/temperature/flow unit.
    # Triggered only when the value starts with a digit (not e.g. "CS + LINING").
    if raw and raw[0].isdigit():
        _dim_m = re.match(
            r'^(-?\d+(?:\.\d+)?)\s*(mm|M|NB|DN)\b',
            raw, re.IGNORECASE,
        )
        if _dim_m:
            _remainder = raw[_dim_m.end():].strip()
            # Keep remainder only if it looks like a valid unit/modifier (psig, °F, etc.)
            _is_valid_suffix = bool(re.match(
                r'^(?:psig|psia|psi|barg|bara|kpag|mpa|°[fc]|/[hH]|m3|kw|hp|mw)',
                _remainder, re.IGNORECASE,
            ))
            if _remainder and not _is_valid_suffix:
                raw = f'{_dim_m.group(1)} {_dim_m.group(2).upper()}'.strip()

    # ── MOC trailing noise cleanup ─────────────────────────────────────────
    # Strip OCR junk that follows a known material spec, e.g.
    # "CS + LINING ix 3.K" → "CS + LINING".
    _moc_m = re.match(
        r'^((?:CS|SS|316L?|304L?|317L|DSS|SDSS|DUPLEX|A516|INCONEL|MONEL|'
        r'HASTELLOY|GRE|FRP|HDPE|PVC|CARBON\s*STEEL|STAINLESS|ALLOY\s*STEEL)'
        r'(?:\s*[+/&]\s*(?:LINING|CLAD(?:DING)?|LINED?\b[^,;\n]{0,20}?|'
        r'RUBBER|EPOXY|FRP|GRE|HDPE|NEOPRENE))?)',
        raw, re.IGNORECASE,
    )
    if _moc_m and len(_moc_m.group(1)) < len(raw):
        _after_moc = raw[_moc_m.end():].strip()
        # Only truncate if the remainder doesn't look like a meaningful continuation
        if not re.match(r'^(?:ASTM|AISI|ISO|EN|NACE|\+|-)', _after_moc, re.IGNORECASE):
            raw = _moc_m.group(1).strip()

    return raw


def _build_databox_index(text: str, config: dict) -> dict:
    """
    Global equipment data-box scanner for P&ID drawings.

    Scans the full OCR text for known engineering labels
    (e.g. "OPERATING PRESS.", "NOMINAL CAPACITY :") and associates each
    extracted value with the nearest equipment tag found within
    _DATABOX_TAG_LOOKBACK_CHARS before the label position.

    Returns {TAG: {field_key: value_string}} merged later into per-tag
    results so narrow context-window extraction does not miss data-box
    values that appear far from the tag in OCR order.

    Label to field mappings are soft-coded in equipment_type_config.json
    under 'databox_label_map'.  Scan window is 'databox_scan_window_chars'
    in the 'extraction' section.  A list value in the map indicates a
    MIN/MAX split: the raw value is split on '/' and assigned to
    [0]=first field, [1]=second field.
    """
    ext_cfg   = config.get('extraction', {})
    db_window = int(ext_cfg.get('databox_scan_window_chars', _DATABOX_TAG_LOOKBACK_CHARS))
    label_map = config.get('databox_label_map', {})
    if not label_map:
        return {}

    _DB_TAG_RE = re.compile(
        r'\b([A-Z]{1,3}-[0-9]{2,5}[A-Za-z]?(?:-[A-Z0-9]{1,4})?)\b',
        re.IGNORECASE,
    )

    index: dict = {}

    # Sort longest label first so specific variants (e.g. "design press (min/max)")
    # are matched before shorter ones ("design press") — first-match-wins.
    sorted_variants = sorted(label_map.items(), key=lambda kv: -len(kv[0]))

    for label_variant, field_or_pair in sorted_variants:
        words = label_variant.strip().upper().split()
        if not words:
            continue
        # Allow 0-5 punctuation/space chars between words
        pat_str = r'[\s./()\-]{0,5}'.join(re.escape(w) for w in words)
        # Capture up to 80 chars after the colon (reduced from 120 to limit
        # runaway captures); stop at newline or semicolon.
        # _clean_databox_value() further truncates at noise patterns.
        pat_str = pat_str + r'[^:\n]{0,35}:\s*([^;\n]{1,80})'
        try:
            pat = re.compile(pat_str, re.IGNORECASE)
        except re.error:
            continue

        for m in pat.finditer(text):
            raw_value = _clean_databox_value(m.group(1))
            if not raw_value:
                continue

            win_start   = max(0, m.start() - db_window)
            pre_text    = text[win_start:m.start()]
            tag_matches = list(_DB_TAG_RE.finditer(pre_text))
            if not tag_matches:
                # Wider post-window: 400 chars covers data-box inline formats
                post_text   = text[m.end():min(len(text), m.end() + 400)]
                tag_matches = list(_DB_TAG_RE.finditer(post_text))
            if not tag_matches:
                continue

            tag = tag_matches[-1].group(1).upper()
            if tag not in index:
                index[tag] = {}

            if isinstance(field_or_pair, list) and len(field_or_pair) == 2:
                parts = re.split(r'\s*/\s*', raw_value, maxsplit=1)
                if len(parts) == 2:
                    v0 = _clean_databox_value(parts[0])
                    v1 = _clean_databox_value(parts[1])
                    # ── Smart numerical min/max assignment ──────────────────
                    # When both target fields end with _min / _max, use numeric
                    # ordering rather than position so that a box labelled
                    # (MAX/MIN) with "185 F / -13.2 F" and one labelled (MIN/MAX)
                    # with "-13.2 F / 185 F" both map correctly regardless of
                    # the order in which the values appear in the cell.
                    _f0, _f1 = field_or_pair[0], field_or_pair[1]
                    _is_minmax = (
                        (_f0.endswith('_min') or _f0.endswith('_max')) and
                        (_f1.endswith('_min') or _f1.endswith('_max'))
                    )
                    if _is_minmax and v0 and v1:
                        _n0 = re.search(r'-?\d+(?:\.\d+)?', v0)
                        _n1 = re.search(r'-?\d+(?:\.\d+)?', v1)
                        if _n0 and _n1:
                            _flt0, _flt1 = float(_n0.group()), float(_n1.group())
                            _min_f = _f0 if _f0.endswith('_min') else _f1
                            _max_f = _f0 if _f0.endswith('_max') else _f1
                            _min_v = v0 if _flt0 <= _flt1 else v1
                            _max_v = v0 if _flt0 >= _flt1 else v1
                            if _min_f not in index[tag]:
                                index[tag][_min_f] = _min_v
                            if _max_f not in index[tag]:
                                index[tag][_max_f] = _max_v
                        else:
                            if v0 and _f0 not in index[tag]:
                                index[tag][_f0] = v0
                            if v1 and _f1 not in index[tag]:
                                index[tag][_f1] = v1
                    else:
                        if v0 and field_or_pair[0] not in index[tag]:
                            index[tag][field_or_pair[0]] = v0
                        if v1 and field_or_pair[1] not in index[tag]:
                            index[tag][field_or_pair[1]] = v1
                else:
                    if field_or_pair[0] not in index[tag]:
                        index[tag][field_or_pair[0]] = raw_value
            else:
                if field_or_pair not in index[tag]:
                    index[tag][field_or_pair] = raw_value

    if index:
        print(
            f'[EQ-DIAG][DataBox] Indexed {len(index)} equipment: '
            + str({k: list(v.keys()) for k, v in index.items()}),
            flush=True,
        )
    return index


def _extract_equipment_items(text: str, drawing_ref: str, config: dict) -> list:
    """
    All field extraction patterns are soft-coded in equipment_type_config.json.
    Add / adjust patterns there without touching this function.

    Fields returned per item
    ------------------------
    tag, type_label, description, area, drawing_ref,
    line_connections, nozzle_connections, service_fluid,
    material_class, process_notes
    """
    ext_cfg     = config.get('extraction', {})
    type_labels = config.get('type_labels', {})
    fluid_kws   = [kw for kw in config.get('fluid_keywords', []) if not kw.startswith('_')]
    ctx_win                 = int(ext_cfg.get('context_window_chars', 160))
    desc_words              = int(ext_cfg.get('description_max_words', 6))
    desc_ctx_chars          = int(ext_cfg.get('description_context_chars', 400))
    desc_min_len            = int(ext_cfg.get('description_min_word_length', 3))
    area_ctx_chars          = int(ext_cfg.get('area_context_chars', 600))
    area_from_tag_heuristic = bool(ext_cfg.get('area_from_tag_heuristic', True))
    nozzle_ctx_chars        = int(ext_cfg.get('nozzle_context_chars', 400))
    mat_ctx_chars           = int(ext_cfg.get('material_context_chars', 400))
    service_ctx_chars       = int(ext_cfg.get('service_context_chars', 400))
    note_ctx_chars          = int(ext_cfg.get('note_context_chars', 400))
    # Standards refs, conjunctions and short noise tokens to exclude from description
    _desc_stop_words        = {
        'API','ASME','ANSI','ISO','DIN','NACE','NOTE','REF','SEE','PER',
        'AND','FOR','THE','OR','TO','OF','IN','AT','BY','NO','AS','IS','ON',
        # Drawing / title block words that appear near tags on P&IDs
        'LOCATION','MUBARRAZ','ISLAND','SCALE','NTS','DATE','DESCRIPTION',
        'REFERENCE','DOCUMENTS','DRAWINGS','DOCUMENT','DRAWING','TITLE',
        'COMPANY','PROJECT','SHEET','SIZE','ENGINEERING','CONSULTANT',
        # Company names (may appear in nearby company block)
        'REJLERS','DORSCH','HOLDING','GMBH','ABU','DHABI','UAE','HAMDAN',
        # Revision table words (near tags in OCR order)
        'ISSUED','APPROVED','REVIEW','HAZOP','CONSTRUCTION','INCORPORATED',
        'RETURNED','REAPPROVED','INFORMATION','COMMENTS',
        # P&ID noise
        'ALARM','TRIP','OPEN','HALF','SLOPE','SOUR','FLARE','HEADER',
        'WELL','FLUID','PHASE','LINE','TYPE','NOTE','NOTES','SCOPE',
    }

    # Soft-coded via tag_pattern in equipment_type_config.json.
    # The optional (?:-[A-Za-z0-9]{1,4})? captures project train/unit suffixes such
    # as -TF, -1F, -2A that are common in O&G tag numbering (e.g. V-308-TF,
    # V-805-1F).  Without this suffix, duplicate-deduplication collapses
    # equipment with the same base number but different trains into one row.
    # tag_pattern_ignorecase (default True): compile with IGNORECASE so OCR
    # lowercase variants like C-010c-TF are found; tag is always uppercased.
    _tag_pat_default = r'\b([A-Za-z]{1,2})-([0-9]{3,5}[A-Za-z]?(?:-[A-Za-z0-9]{1,4})?)\b'
    _tag_ic = bool(ext_cfg.get('tag_pattern_ignorecase', True))
    tag_re = re.compile(
        ext_cfg.get('tag_pattern', _tag_pat_default),
        re.IGNORECASE if _tag_ic else 0,
    )

    # --- Soft-coded helper patterns (read once per call) ------------------
    # Used by description strategy 1: identify bare tag lines and pure-noise tokens.
    # Must also match the extended suffix form so lines like "V-308-TF" are
    # not misidentified as description text.  IGNORECASE covers OCR lowercase.
    _tag_like_re  = re.compile(r'^[A-Za-z]{1,2}-\d{3,5}[A-Za-z]?(?:-[A-Za-z0-9]{1,4})?$', re.IGNORECASE)
    _noise_tok_re = re.compile(r'^[\d\.\+\-\/\%\(\)\[\]]{1,6}$')

    # Soft-coded reject patterns for description lines — lines matching any of
    # these are skipped entirely rather than partially filtered.
    # Covers: pipe designations (20"-PL-...), document/drawing numbers
    # (PJ6-EXD-...-0023), instrument/valve tags (FT-1234), fraction-inch
    # size tokens (3/4"), grid refs (A1-H8), and OCR fragments (|[, =£).
    _desc_line_reject_re = re.compile(
        r'(?:'
        r'\d+["\']-[A-Z]{2,4}-'          # pipe designation: 20"-PL-...
        r'|[A-Z]{2,4}-[A-Z]{2,4}-[A-Z]{2,4}-[A-Z]{2,4}-\d{4}'  # doc number
        r'|PJ\d[-_][A-Z]'                # project document prefix PJ6-...
        r'|\b[A-Z]{2,3}-\d{4,6}\b'       # instrument/valve tags FT-1234
        r'|\d+\s*/\s*\d+'                # fractions 3/4
        r'|[|=\[\]£$@#]'                 # OCR junk characters
        r'|^\d{1,3}["\']?\s*[-]\s*[A-Z]{2,4}'  # starts with size then tag type
        r'|\bFROM\s+[A-Z]|\bTO\s+[A-Z]|\bVIA\s+[A-Z]'  # flow routing text
        r'|\bLINE\s*\d|\bNOTE\s*\d|\bSHEET\s*\d'  # line/note/sheet refs
        r'|\bSLOPE\s*1?\s*[:.]'          # SLOPE 1:100 annotations
        r'|\bNTS\b|\bSCALE\b'            # scale annotations
        r')',
        re.IGNORECASE,
    )
    area_re    = re.compile(
        ext_cfg.get('area_pattern',
                    r'(?:AREA|UNIT|TRAIN|BAY|SECTION|BATTERY|MODULE|MOD|ZONE|BLOCK|SKID|PLANT|FIELD|STREAM)\s*[:\-]?\s*([A-Z0-9]{1,8})'),
        re.IGNORECASE,
    )
    nozzle_re         = re.compile(
        ext_cfg.get('nozzle_pattern', r'\bN[-]?[0-9]{1,2}[A-Z]?\b')
    )
    mat_re            = re.compile(
        ext_cfg.get('material_class_pattern',
                    r'\b(A1[A-Z]R?|B1[A-Z]|C1[A-Z]|D1[A-Z]|[A-D]2[A-Z]'
                    r'|CS|SS|316L?|304L?|317L|321|347|2205|254SMO'
                    r'|DSS|SDSS|DUPLEX|INCONEL|HASTELLOY|MONEL'
                    r'|GRE|FRP|HDPE|CPVC|PVC|PVDF|A516|A240|A312|A106)\b'),
        re.IGNORECASE,
    )
    material_label_re = re.compile(
        ext_cfg.get('material_label_pattern',
                    r'(?:MATERIAL|MTL|SHELL|BODY|CASING|LINER'
                    r'|WETTED\s*PARTS?|INTERNALS?)'
                    r'\s*[:\-/]\s*([A-Z0-9][A-Z0-9/\-\s\.]{1,28})'),
        re.IGNORECASE,
    )
    service_label_re  = re.compile(
        ext_cfg.get('service_label_pattern',
                    r'(?:SERVICE|FLUID|MEDIUM|PROCESS\s*FLUID'
                    r'|CONTENTS|PRODUCT|DUTY)'
                    r'\s*[:\.\.\-]\s*([A-Za-z][A-Za-z0-9\s/\-]{1,30})'),
        re.IGNORECASE,
    )
    note_re           = re.compile(
        ext_cfg.get('note_pattern',
                    r'(?:(?:SEE\s+)?NOTE\s*[-\s\(]?[0-9]+[\)\.]*'
                    r'|\bHOLD\b(?:\s*[-]?\s*[0-9]+)?'
                    r'|\bTBD\b|\bTBC\b'
                    r'|\bREF[.\s]+DWG[.\s]+[A-Z0-9/\-]+'
                    r'|SEE\s+(?:DWG|SPEC|DOC)[.]*\s*[A-Z0-9/\-]+)'),
        re.IGNORECASE,
    )
    # -----------------------------------------------------------------------

    instr_valve_prefixes = {
        'FT','FI','FIC','FC','PT','PI','PIC','PC','LT','LI','LIC','LC',
        'TT','TI','TIC','TC','AT','AI','FY','PY','LY','TY',
        'HV','FV','XV','PV','SDV','BDV','PSV','PRV','CV','LV','TV',
        'FE','TE','LE','PE','HS','HIC','HI',
    }

    # Soft-coded: exact tag suffix values that identify non-equipment tokens
    # (e.g. project change-request numbers like PJ-2025-CR-002 which OCR
    # garbles into P-2028-CR — "CR" is a document type, not an equipment suffix).
    _exclude_suffixes = {s.upper() for s in ext_cfg.get('exclude_tag_suffixes', ['CR', 'NCR', 'WO', 'TQ', 'MDR', 'MOM', 'MR'])}

    seen = set()
    results = []

    # ── Global data-box index (built once, merged per-tag below) ─────────
    # Scans the full OCR text for LABEL : VALUE pairs in equipment data boxes
    # so narrow context-window extraction doesn't miss values that are spatially
    # far from the tag in OCR text order (common on large-format P&IDs).
    _databox_idx = _build_databox_index(text, config)

    # ── Slash-variant tag expansion ────────────────────────────────────────
    # OCR on P&IDs sometimes reads multi-unit tags like "P-851A/B/C-TF" as a
    # single token. Expand these into individual variants (P-851A-TF,
    # P-851B-TF, P-851C-TF) and append them to the text so the main loop
    # finds each unit independently.
    #
    # Soft-coded: slash_ocr_substitutions — OCR chars that should be treated
    # as '/' in this context (e.g. '?' → '/' when OCR misreads the slash in
    # P-851A/B/C-TF as P-851A?B/C-TF). Applied locally to a copy of the text
    # so the substitution only affects slash-expansion; the main text is left
    # intact to avoid corrupting description or parameter extraction.
    _slash_ocr_subs = ext_cfg.get('slash_ocr_substitutions', [['?', '/']])
    _slash_text = text
    for _bad, _good in _slash_ocr_subs:
        # Only substitute inside plausible tag tokens: letter-digit?letter pattern
        _slash_text = re.sub(
            r'(?<=[A-Za-z])' + re.escape(_bad) + r'(?=[A-Za-z])',
            _good, _slash_text
        )
    _slash_re = re.compile(
        r'\b([A-Za-z]{1,2}-\d{3,5})([A-Za-z])/([A-Za-z])(?:/([A-Za-z]))?(?:-([A-Za-z0-9]{1,4}))?\b',
        re.IGNORECASE,
    )
    _slash_expanded: list[str] = []
    for _sm in _slash_re.finditer(_slash_text):
        _base = _sm.group(1).upper()
        _sfx  = _sm.group(5).upper() if _sm.group(5) else ''
        for _v in [_sm.group(2).upper(), _sm.group(3).upper()] + ([_sm.group(4).upper()] if _sm.group(4) else []):
            _slash_expanded.append(f'{_base}{_v}' + (f'-{_sfx}' if _sfx else ''))
    if _slash_expanded:
        text = text + '\n' + '\n'.join(_slash_expanded)
        print(f'[EQ-DIAG] Slash expansion added: {_slash_expanded}', flush=True)

    for m in tag_re.finditer(text):
        prefix = m.group(1).upper()
        tag    = m.group(0).upper()   # always uppercase — OCR may emit lowercase letters

        if prefix in instr_valve_prefixes:
            continue
        if type_labels and prefix not in type_labels:
            continue
        # Filter non-equipment project-reference suffixes
        _tag_suffix_m = re.search(r'-([A-Z]{1,4})$', tag)
        if _tag_suffix_m and _tag_suffix_m.group(1) in _exclude_suffixes:
            continue

        # ── Gate 0: tag-context annotation reject ────────────────────
        # Soft-coded via 'tag_context_reject_patterns' in equipment_type_config.json.
        # If ANY pattern matches within 80 chars of the tag, the occurrence is
        # skipped WITHOUT adding to 'seen' (so a legitimate data-box occurrence
        # of the same tag later in the OCR stream is still extracted).
        # Use-case: revision-cloud bubbles annotated directly on P&IDs contain
        # tag text that is NOT equipment on this drawing.
        _ctx_reject_pats = ext_cfg.get('tag_context_reject_patterns', [])
        if _ctx_reject_pats:
            _surrounding = text[max(0, m.start() - 80): min(len(text), m.end() + 80)]
            if any(re.search(_crpat, _surrounding) for _crpat in _ctx_reject_pats):
                print(f'[EQ-DIAG] Gate0-ctx-reject skipped: {tag!r}', flush=True)
                continue

        # ── Gate 1: connector-arrow context check ────────────────────────
        # On ADNOC/O&G P&IDs the sheet-edge continuation arrows use the
        # format: "[KEYWORD] [DESCRIPTION] [TAG]\n[DWG-NO]" (e.g.
        # "SOUR GAS TO MEA INLET SCRUBBER V-804-TF\nPJ6-EXD-MRI-BQDA-0024").
        # These are cross-sheet references to equipment that lives on a
        # different drawing and has no data box here.
        # Detection: a flow-routing keyword within lookback_chars BEFORE the
        # tag AND a multi-segment drawing number within lookahead_chars AFTER.
        # When detected, skip this occurrence WITHOUT adding to 'seen' so
        # the same tag can still be processed if it appears in its own data
        # box later in the OCR text of this very drawing.
        # Soft-coded via connector_context_enabled / connector_context_lookback_chars
        # / connector_context_lookahead_chars / connector_keywords_pattern in
        # equipment_type_config.json extraction section.
        if bool(ext_cfg.get('connector_context_enabled', True)):
            _conn_lookback  = int(ext_cfg.get('connector_context_lookback_chars', 80))
            _conn_lookahead = int(ext_cfg.get('connector_context_lookahead_chars', 120))
            _conn_kws_pat   = ext_cfg.get(
                'connector_keywords_pattern',
                r'\b(?:FROM|TO|VIA|INTO|INLET|OUTLET|SUCTION|DISCHARGE|DEST(?:INATION)?|SOURCE)\b',
            )
            _conn_before = text[max(0, m.start() - _conn_lookback): m.start()]
            _conn_after  = text[m.end(): min(len(text), m.end() + _conn_lookahead)]
            if (re.search(_conn_kws_pat, _conn_before, re.IGNORECASE)
                    and _TITLEBLOCK_DWG_NO_RE.search(_conn_after)):
                print(
                    f'[EQ-DIAG] Gate1-connector-ref skipped: {tag!r} '
                    f'(keyword in lookback + dwg-no in lookahead)',
                    flush=True,
                )
                # Do NOT add to seen — allow a data-box occurrence of the
                # same tag later in the text to be processed normally.
                continue

        if tag in seen:
            continue
        seen.add(tag)

        start = max(0, m.start() - ctx_win)
        end   = min(len(text), m.end() + ctx_win)
        ctx   = text[start:end]

        type_label = type_labels.get(prefix, 'Equipment')

        # ── Description — multi-strategy extraction ───────────────────────
        after       = text[m.end(): m.end() + desc_ctx_chars]
        description = ''

        # Strategy 0: data-box title — on P&IDs the equipment data box has the
        # format: TAG_LINE\nDESCRIPTION_LINE (e.g. "V-803-TF\nMRD OIL SLUG CATCHER").
        # Look for 2-6 consecutive ALL-CAPS words on the first non-blank line
        # after the tag that is NOT a pipe/tag/doc reference.
        # Soft-coded: description_databox_min_words, description_databox_max_words
        _db_min_w = int(ext_cfg.get('description_databox_min_words', 2))
        _db_max_w = int(ext_cfg.get('description_databox_max_words', 6))
        for _dln in (ln.strip() for ln in after.split('\n') if ln.strip()):
            if _tag_like_re.match(_dln) or _desc_line_reject_re.search(_dln):
                continue
            _dln_toks = _dln.split()
            # Require at least _db_min_w tokens, all purely alphabetic (or
            # common hyphenated words like "THREE-PHASE"), no digits
            _alpha_toks = [
                t.strip('.,;:/()"\'[]')
                for t in _dln_toks
                if re.match(r'^[A-Za-z][A-Za-z\-]{1,}$', t.strip('.,;:/()"\'[]'))
                and len(t) >= desc_min_len
                and t.upper() not in _desc_stop_words
            ]
            if len(_alpha_toks) >= _db_min_w:
                description = ' '.join(_alpha_toks[:_db_max_w]).title()
                break

        # Strategy 1: newline-segmented lines right after the tag.
        # Each line is checked for "description-likeness":
        # skip bare tag IDs, pipe designations and pure digit/symbol noise.
        if not description:
            desc_lines = []
            for _ln in (ln.strip() for ln in after.split('\n') if ln.strip()):
                if _tag_like_re.match(_ln):
                    continue
                if _desc_line_reject_re.search(_ln):
                    continue
                _toks = [t.strip('.,;:/()"\'[]') for t in _ln.split()]
                _valid = [
                    t for t in _toks
                    if len(t) >= desc_min_len
                    and not t.isdigit()
                    and not _tag_like_re.match(t)
                    and not _noise_tok_re.match(t)
                    and t.upper() not in _desc_stop_words
                    and not re.search(r'\d{2,}', t)       # skip tokens with 2+ digits
                    and not re.match(r'^[A-Z]{1,3}-\d', t)  # skip tag-like tokens
                ]
                if _valid:
                    desc_lines.append(' '.join(_valid[:5]))
                if len(desc_lines) >= 1:
                    break
            if desc_lines:
                description = ' '.join(desc_lines).title()

        # Strategy 2: ALL-CAPS word scan in narrower ctx_win (improved filter)
        if not description:
            _cap_words = re.findall(r'\b[A-Z][A-Z]{2,19}\b', after[:ctx_win])
            _filtered_caps = [
                w for w in _cap_words
                if not re.match(r'^[A-Z]{1,2}-\d', w)
                and w not in _desc_stop_words
                and len(w) >= desc_min_len
            ][:desc_words]
            if _filtered_caps:
                description = ' '.join(w.capitalize() for w in _filtered_caps[:3])

        # Strategy 3: fall back to the equipment TypeLabel
        if not description:
            description = type_label

        # ── Line connections (piping designation tokens) ───────────────────
        lc_tokens = []
        for lm in _LINE_TAG_RE.finditer(ctx):
            token = lm.group(0).strip()
            if token and token not in lc_tokens:
                lc_tokens.append(token)

        # ── Service / fluid — multi-strategy extraction ───────────────────
        _svc_start    = max(0, m.start() - service_ctx_chars)
        _svc_end      = min(len(text), m.end() + service_ctx_chars)
        _svc_ctx      = text[_svc_start:_svc_end]
        service_fluid = ''
        # Strategy 1: label-based — SERVICE: CRUDE OIL, FLUID: NITROGEN, MEDIUM: GAS
        _svc_lm = service_label_re.search(_svc_ctx)
        if _svc_lm:
            _raw_svc = _svc_lm.group(1).split('\n')[0].strip().rstrip('.,;')
            if len(_raw_svc) >= 2:
                service_fluid = _raw_svc[:35].title()
        # Strategy 2: keyword scan in wider context
        if not service_fluid:
            _svc_lower = _svc_ctx.lower()
            found_fluids = [kw for kw in fluid_kws if kw in _svc_lower]
            service_fluid = ', '.join(found_fluids[:2]).title() if found_fluids else ''
        # Strategy 3: derive from fluid code embedded in already-found line connection tags.
        # e.g. "4"-HO-5665-033842-X" → fluid code "HO" → "Hydrocarbon Oil".
        # Guaranteed to find something whenever line connections were extracted.
        if not service_fluid and lc_tokens:
            _lf_map = {k: v for k, v in config.get('line_fluid_code_map', {}).items()
                       if not str(k).startswith('_')}
            for _lc in lc_tokens:
                _fc_m = re.match(r'^[\d½¾¼]+\s*["\'?]\s*[-_]\s*([A-Z]{1,4})\s*[-_]', _lc)
                if _fc_m:
                    _fc = _fc_m.group(1).upper()
                    _mapped = _lf_map.get(_fc)
                    if _mapped:
                        service_fluid = _mapped
                        break

        # ── Area / Unit — multi-strategy extraction ───────────────────────
        # Strategy 1: search a wider context (soft-coded area_context_chars).
        # Uses capture group(1) — returns just the code, not the whole keyword match.
        _a_start = max(0, m.start() - area_ctx_chars)
        _a_end   = min(len(text), m.end() + area_ctx_chars)
        area_m   = area_re.search(text[_a_start:_a_end])
        area     = area_m.group(1).strip() if area_m else ''

        # Strategy 2: derive from serial number digits (O&G tag-number convention).
        # V-101 → "100", P-2201 → "2200", E-10001 → "10000"
        if not area and area_from_tag_heuristic:
            _digits = re.sub(r'[^0-9]', '', m.group(2))
            if len(_digits) >= 3:
                area = _digits[0] + '0' * (len(_digits) - 1)

        # ── Nozzle connections — multi-strategy extraction ────────────────
        _nzl_start    = max(0, m.start() - nozzle_ctx_chars)
        _nzl_end      = min(len(text), m.end() + nozzle_ctx_chars)
        _nzl_ctx      = text[_nzl_start:_nzl_end]
        # Strategy 1: N1 / N-1 / N2A nozzle tag pattern
        nozzle_tokens = list(dict.fromkeys(nozzle_re.findall(_nzl_ctx)))
        # Strategy 2: size-prefixed nozzle labels  e.g. 4"-N1, 6"-N2A
        for _snm in re.finditer(
            r'\b\d{1,3}\s*["\']\s*-?\s*(N[-]?[0-9]{1,2}[A-Z]?)\b',
            _nzl_ctx, re.IGNORECASE
        ):
            _tok = _snm.group(1).upper()
            if _tok not in nozzle_tokens:
                nozzle_tokens.append(_tok)
        # Strategy 3: functional orientation labels as fallback when no N-tags found
        # e.g. INLET, OUTLET, SUCTION, DISCHARGE on equipment bubbles
        if not nozzle_tokens:
            _orient_hits = re.findall(
                r'\b(INLET|OUTLET|SUCT(?:ION)?|DISCH(?:ARGE)?|VENT|DRAIN|BYPASS'
                r'|OVERFLOW|RECYCLE|RETURN|FEED|PRODUCT|OVERHEAD|BOTTOM(?:S)?)\b',
                _nzl_ctx, re.IGNORECASE,
            )
            for _ow in dict.fromkeys(w.capitalize() for w in _orient_hits):
                nozzle_tokens.append(_ow)
        nozzle_tokens = nozzle_tokens[:8]

        # ── Material / piping spec — multi-strategy extraction ────────────
        _mat_start     = max(0, m.start() - mat_ctx_chars)
        _mat_end       = min(len(text), m.end() + mat_ctx_chars)
        _mat_ctx       = text[_mat_start:_mat_end]
        material_class = ''
        # Strategy 1: label-based — MATERIAL: CS/SS316, SHELL: DSS, MTL: INCONEL
        _mat_lm = material_label_re.search(_mat_ctx)
        if _mat_lm:
            _raw_mat = _mat_lm.group(1).split('\n')[0].strip().rstrip('.,;/ ')
            if len(_raw_mat) >= 2:
                material_class = _raw_mat[:25].upper()
        # Strategy 2: pattern scan in wider context
        if not material_class:
            mat_matches    = mat_re.findall(_mat_ctx)
            material_class = mat_matches[0].upper() if mat_matches else ''
        # Strategy 3: derive material hint from pipe-class prefix in line connection tags.
        # Line tag format: SIZE"-FLUID-SEQ-PIPECLASS[-SUFFIX]
        # First 2 digits of the pipe-class code encode material in most spec books
        # (soft-coded in pipe_class_prefix_map — add project-specific mappings there).
        if not material_class and lc_tokens:
            _pc_prefix_map = {k: v for k, v in config.get('pipe_class_prefix_map', {}).items()
                              if not str(k).startswith('_')}
            for _lc in lc_tokens:
                # Pipe class is typically 5-8 digits separated by '-' or '_'
                _pc_m = re.search(r'[\-_](\d{4,8})(?:[\-_][A-Z0-9]{0,5})?(?:\s|$)', _lc)
                if _pc_m:
                    _prefix = _pc_m.group(1)[:2]
                    _mat = _pc_prefix_map.get(_prefix)
                    if _mat:
                        material_class = _mat
                        break

        # ── Process note references — wider context scan ──────────────────
        _nt_start     = max(0, m.start() - note_ctx_chars)
        _nt_end       = min(len(text), m.end() + note_ctx_chars)
        _nt_ctx       = text[_nt_start:_nt_end]
        note_matches  = list(dict.fromkeys(
            n.strip() for n in note_re.findall(_nt_ctx)
        ))[:3]
        process_notes = ', '.join(note_matches) if note_matches else ''

        # ── Process parameter extraction (P&ID data-bubble / annotation) ──
        # Uses a narrower context window so values are specific to this tag.
        _pp_ctx_chars  = int(ext_cfg.get('process_param_context_chars', 300))
        _pp_start      = max(0, m.start() - _pp_ctx_chars)
        _pp_end        = min(len(text), m.end() + _pp_ctx_chars)
        _pp_ctx        = text[_pp_start:_pp_end].upper()

        # Soft-coded regex patterns (all read from config)
        _press_val_pat  = ext_cfg.get('pressure_value_pattern',
                           r'(-?\d+(?:\.\d+)?)\s*(PSIG|PSIA|PSI|barg|bara|kPag|MPa|bar)\b')
        _temp_val_pat   = ext_cfg.get('temperature_value_pattern',
                           r'(-?\d+(?:\.\d+)?)\s*(?:°\s*[FC]|DEG\.?\s*[FC]|DEGF|DEGC)')
        _flow_lbl_pat   = ext_cfg.get('flowrate_label_pattern',
                           r'(?:Q\s*[:=]|FLOW\s*RATE|FLOWRATE|CAPACITY|DESIGN\s*FLOW|DUTY)\s*[:=/(]?')
        _flow_val_pat   = ext_cfg.get('flowrate_value_pattern',
                           r'(\d+(?:[,.]\d+)?)\s*(M3/H|M3/HR|NM3/H|NM3/HR|SM3/D|MMSCFD|BBL/D|BBL/H|BPD|GPM|T/H|KG/H|KG/HR|MW|KW|MMBTU/H)')
        _flow_bare_pat  = ext_cfg.get('flowrate_bare_value_pattern',
                           r'(\d+(?:[,.]\d+)?)\s*(M3/H|M3/HR|NM3/H|NM3/HR|SM3/D|MMSCFD|BBL/D|GPM|T/H|KG/H|MW|KW|MMBTU/H)')
        _flow_ctx_chars = int(ext_cfg.get('flowrate_context_chars', 500))
        _op_press_lbl   = ext_cfg.get('oper_pressure_label_pattern',
                           r'(?:OPER(?:ATING)?|OP\.?)\s*PRESS(?:URE)?\.?\s*[-:=/(]')
        _des_press_lbl  = ext_cfg.get('design_pressure_label_pattern',
                           r'(?:DES(?:IGN)?\.?(?:\s*/\s*SET)?|SET)\s*PRESS(?:URE)?\.?\s*[-:=/(]')
        _op_temp_lbl    = ext_cfg.get('oper_temp_label_pattern',
                           r'(?:OPER(?:ATING)?|OP\.?)\s*TEMP(?:ERATURE)?\.?\s*[-:=/(]')
        _des_temp_lbl   = ext_cfg.get('design_temp_label_pattern',
                           r'DES(?:IGN)?\.?\s*TEMP(?:ERATURE)?\.?\s*[-:=/(]')
        # Soft-coded: slash-pair dual-value patterns for "LABEL (MIN/MAX) : V1/V2 UNIT"
        # format common on O&G P&ID data boxes (single cell stores both min + max).
        _dual_temp_pat  = ext_cfg.get('dual_value_temp_pattern',
                           r'(-?\d+(?:\.\d+)?)\s*/\s*(-?\d+(?:\.\d+)?)\s*(°\s*[FC]|DEG\.?\s*[FC]|DEGF|DEGC)?')
        _dual_press_pat = ext_cfg.get('dual_value_press_pattern',
                           r'(-?\d+(?:\.\d+)?)\s*/\s*(-?\d+(?:\.\d+)?)\s*(PSIG|PSIA|PSI|barg|bara|kPag|MPa|bar)\b')
        _ins_codes      = [c.upper() for c in ext_cfg.get('insulation_codes',
                           ['HOT', 'COLD', 'PERS', 'HT', 'CT', 'TRACED', 'EHT', 'BARE', 'ACOUSTIC'])]
        _dim_len_lbl    = ext_cfg.get('dimension_length_label_pattern',
                           r'(?:LENGTH|HEIGHT|TL[-/]TL|TAN[-/]TAN|LONG|T/T)\s*[:=]?')
        _dim_dia_lbl    = ext_cfg.get('dimension_diameter_label_pattern',
                           r'(?:DIA(?:METER)?|O\.?D\.?|BORE|I\.?D\.?|NB|DN)\s*[:=]?')
        _dim_val_pat    = ext_cfg.get('dimension_value_pattern', r'(\d+(?:\.\d+)?)\s*(mm|M)?')
        _mtr_pat        = ext_cfg.get('motor_rating_pattern',
                           r'(?:MOTOR|DRIVER|RATED\s*POWER|INSTALLED\s*POWER)\s*[:=]?\s*(\d+(?:\.\d+)?)\s*(kW|KW|HP|BHP|KVA)\b')
        _mtr_bare_pat   = ext_cfg.get('motor_rating_bare_pattern', r'(\d+(?:\.\d+)?)\s*(kW|KW)\b')
        _qual_pat       = ext_cfg.get('quality_required_pattern',
                           r'(?:QUALITY|QC|NDE|NDT|INSPECT(?:ION)?)\s*[:=]?\s*([A-D](?:\s+LEVEL)?|LEVEL\s*[A-D]|(?:100%\s*)?(?:RT|MT|UT|PT|VT)(?:[+&,/\s]+(?:RT|MT|UT|PT|VT))*)')

        # ── Oper. Pressure ────────────────────────────────────────────────
        oper_pressure = ''
        _op_lbl_m = re.search(_op_press_lbl, _pp_ctx, re.IGNORECASE)
        if _op_lbl_m:
            _after_lbl = _pp_ctx[_op_lbl_m.end():]
            _pv = re.search(_press_val_pat, _after_lbl[:80], re.IGNORECASE)
            if _pv:
                oper_pressure = f'{_pv.group(1)} {_pv.group(2)}'
            elif not _pv:
                # Fallback: label matched at "(MIN/MAX)" qualifier  →  after = "MIN/MAX) : 155 psig".
                # Skip over the qualifier and grab the first pressure value after ":".
                _after_colon = re.search(r'\)\s*[:,=]\s*(.{1,60})', _after_lbl[:80], re.IGNORECASE)
                if _after_colon:
                    _pv2 = re.search(_press_val_pat, _after_colon.group(1), re.IGNORECASE)
                    if _pv2:
                        oper_pressure = f'{_pv2.group(1)} {_pv2.group(2)}'

        # ── Design Pressure min / max ─────────────────────────────────────
        design_pressure_min = ''
        design_pressure_max = ''
        _dp_lbl_ms = list(re.finditer(_des_press_lbl, _pp_ctx, re.IGNORECASE))
        _dp_vals   = []   # list of (numeric_value, unit_string)
        for _dlm in _dp_lbl_ms:
            _win80 = _pp_ctx[_dlm.end():_dlm.end() + 80]
            # Strategy A: slash-pair with unit at end e.g. "195 / -13.2 psig"
            # (label matched at "(MIN/MAX)" qualifier → after = "MIN/MAX) : 195 / -13.2 psig")
            _dp_slash = re.search(_dual_press_pat, _win80, re.IGNORECASE)
            if _dp_slash:
                _pu = _dp_slash.group(3).upper()
                _dp_vals.append((float(_dp_slash.group(1)), _pu))
                _dp_vals.append((float(_dp_slash.group(2)), _pu))
            else:
                # Strategy B: standard "VALUE UNIT" or "VALUE UNIT / VALUE UNIT"
                for _pv in re.finditer(_press_val_pat, _win80, re.IGNORECASE):
                    _dp_vals.append((float(_pv.group(1)), _pv.group(2)))
            if not _dp_vals:
                # Strategy C: unit embedded in label parens e.g. "DES./SET PRESS (PSIG) : 195 / FV".
                # Locate (UNIT) just after the label match, collect numeric values that follow;
                # non-numeric tokens like FV (Full Vacuum) are silently skipped.
                _pu_m = re.search(
                    r'\(([A-Za-z]{2,5})\)\s*[:,=\s]+(-?\d+(?:\.\d+)?)', _win80, re.IGNORECASE
                )
                if _pu_m:
                    _pu = _pu_m.group(1).upper()
                    if _pu in ('PSIG', 'PSIA', 'PSI', 'BARG', 'BARA', 'BAR', 'KPAG', 'MPA'):
                        _dp_vals.append((float(_pu_m.group(2)), _pu))
                        # Also grab second value (MIN or MAX partner), ignoring non-numerics
                        _rest = _win80[_pu_m.end():]
                        _v2 = re.search(r'[/,\s]+(-?\d+(?:\.\d+)?)', _rest[:30])
                        if _v2:
                            try:
                                _dp_vals.append((float(_v2.group(1)), _pu))
                            except ValueError:
                                pass
        if _dp_vals:
            _dp_nums  = [v for v, _u in _dp_vals]
            _dp_units = _dp_vals[0][1]  # use unit from first match
            if len(_dp_nums) == 1:
                # Single value = maximum design pressure; MIN is unspecified
                design_pressure_max = f'{_dp_nums[0]} {_dp_units}'
            else:
                design_pressure_min = f'{min(_dp_nums)} {_dp_units}'
                design_pressure_max = f'{max(_dp_nums)} {_dp_units}'
        elif not oper_pressure:
            # Fallback: any bare pressure value in narrow context
            _bare_pv = re.search(_press_val_pat, _pp_ctx, re.IGNORECASE)
            if _bare_pv:
                design_pressure_max = f'{_bare_pv.group(1)} {_bare_pv.group(2)}'

        # ── Oper. Temperature ─────────────────────────────────────────────
        oper_temperature = ''
        _ot_lbl_m = re.search(_op_temp_lbl, _pp_ctx, re.IGNORECASE)
        if _ot_lbl_m:
            _ot_after = _pp_ctx[_ot_lbl_m.end():_ot_lbl_m.end() + 80]
            # Strategy 1: standard single value e.g. "175 °F"
            _tv = re.search(_temp_val_pat, _ot_after, re.IGNORECASE)
            if _tv:
                oper_temperature = f'{_tv.group(1)} °F'
            else:
                # Strategy 2: slash-pair dual value e.g. "105/60 °F" or "105 / 60 °F"
                # Matches "LABEL (MIN/MAX) : 105/60 °F" where unit is only at end.
                # The normalizer converts "105/60 °F" → "60 – 105 °F" (ascending range).
                _ot_dual = re.search(_dual_temp_pat, _ot_after, re.IGNORECASE)
                if _ot_dual and _ot_dual.group(1) != _ot_dual.group(2):
                    _raw_unit = (_ot_dual.group(3) or '°F').strip()
                    oper_temperature = f'{_ot_dual.group(1)}/{_ot_dual.group(2)} {_raw_unit}'
                else:
                    # Strategy 3: unit embedded in label parens e.g. "OPER TEMP (°F) : 175".
                    # The label matches at "(", leaving "°F) : 175" in the window.
                    # Detect °C vs °F from label context, then grab first bare number.
                    _ot_unit = 'C' if re.search(
                        r'\(°?\s*C\)', _pp_ctx[_ot_lbl_m.start():_ot_lbl_m.end() + 8], re.IGNORECASE
                    ) else 'F'
                    _bare_t = re.search(
                        r'[°FC)\s]+[:,=\s]+(-?\d+(?:\.\d+)?)', _ot_after, re.IGNORECASE
                    )
                    if _bare_t:
                        oper_temperature = f'{_bare_t.group(1)} °{_ot_unit}'
        # Normalise dual-temperature values e.g. "105/60 °F" → "60 – 105 °F"
        oper_temperature = _normalize_oper_temp(oper_temperature)

        # ── Design Temp min / max ─────────────────────────────────────────
        design_temp_min = ''
        design_temp_max = ''
        _dt_lbl_ms = list(re.finditer(_des_temp_lbl, _pp_ctx, re.IGNORECASE))
        _dt_vals   = []  # list of numeric values (°F)
        for _dtlm in _dt_lbl_ms:
            _win80 = _pp_ctx[_dtlm.end():_dtlm.end() + 80]
            # Strategy A: slash-pair with unit at end e.g. "185 / -13.2 °F"
            # (label matches at "(MIN/MAX)" qualifier → after = "MIN/MAX) : 185 / -13.2 °F").
            # _dual_temp_pat captures the two numbers (group 3 = optional unit).
            _dt_dual = re.search(_dual_temp_pat, _win80, re.IGNORECASE)
            if _dt_dual and _dt_dual.group(1) != _dt_dual.group(2):
                _dt_vals.extend([float(_dt_dual.group(1)), float(_dt_dual.group(2))])
            else:
                # Strategy B: standard "VALUE °F" or two "VALUE °F" entries
                for _tv in re.finditer(_temp_val_pat, _win80, re.IGNORECASE):
                    _dt_vals.append(float(_tv.group(1)))
        if _dt_vals:
            if len(_dt_vals) == 1:
                # Single design temperature = maximum; MIN is unspecified
                design_temp_max = f'{_dt_vals[0]} °F'
            else:
                design_temp_min = f'{min(_dt_vals)} °F'
                design_temp_max = f'{max(_dt_vals)} °F'

        # ── Design Flowrate ───────────────────────────────────────────────
        # Use a wider context window than the general _pp_ctx (soft-coded via
        # flowrate_context_chars) since flow annotations on P&IDs often sit in
        # connected line labels some distance from the equipment symbol.
        _fl_start  = max(0, m.start() - _flow_ctx_chars)
        _fl_end    = min(len(text), m.end() + _flow_ctx_chars)
        _fl_ctx    = text[_fl_start:_fl_end].upper()
        design_flowrate = ''

        # Strategy 1: label-based (Q= / FLOW RATE: / CAPACITY: / DUTY: …)
        _fl_lbl_m = re.search(_flow_lbl_pat, _fl_ctx, re.IGNORECASE)
        if _fl_lbl_m:
            _fv = re.search(_flow_val_pat, _fl_ctx[_fl_lbl_m.end():_fl_lbl_m.end() + 80], re.IGNORECASE)
            if _fv:
                _val = _fv.group(1).replace(',', '.')
                design_flowrate = f'{_val} {_fv.group(2).upper()}'

        # Strategy 2: bare unit scan — number immediately followed by a
        # recognised flow/duty unit, no label required.  Avoids picking up
        # tag serial numbers by requiring the number > 0 and the unit token
        # is word-bounded (e.g. '100 M3/H' but not '308-TF').
        if not design_flowrate:
            for _bm in re.finditer(_flow_bare_pat, _fl_ctx, re.IGNORECASE):
                _bval = float(_bm.group(1).replace(',', '.'))
                if _bval > 0:
                    design_flowrate = f'{_bm.group(1).replace(",", ".")} {_bm.group(2).upper()}'
                    break

        # Strategy 3: for heaters / heat-exchangers derive duty from any
        # kW or MW annotation.  Only applies when no flow-unit was found.
        if not design_flowrate and prefix in {'E', 'H', 'HT', 'AG', 'CL', 'VR'}:
            _duty_m = re.search(r'(\d+(?:[,.]\d+)?)\s*(MW|KW)\b', _fl_ctx, re.IGNORECASE)
            if _duty_m:
                _dval = float(_duty_m.group(1).replace(',', '.'))
                if _dval > 0:
                    design_flowrate = f'{_dval} {_duty_m.group(2).upper()} (Duty)'

        # ── Insulation ────────────────────────────────────────────────────
        insulation = ''
        _ins_lbl_m = re.search(
            r'(?:INSUL(?:ATION)?|INS|TRACE)\s*[:=/]?\s*([A-Z]{2,10})', _pp_ctx, re.IGNORECASE
        )
        if _ins_lbl_m and _ins_lbl_m.group(1).upper() in _ins_codes:
            insulation = _ins_lbl_m.group(1).upper()
        elif not insulation:
            for _ic in _ins_codes:
                if re.search(r'\b' + _ic + r'\b', _pp_ctx):
                    insulation = _ic
                    break

        # Strategy 2: infer BARE for static equipment with no insulation label.
        # Vessels, drums, separators, tanks, columns etc. on O&G P&IDs are
        # typically BARE unless specifically annotated.  Soft-coded via
        # 'insulation_bare_default_prefixes' in equipment_type_config.json.
        if not insulation:
            _bare_pfxs = {p.upper() for p in ext_cfg.get('insulation_bare_default_prefixes', [
                'V', 'T', 'D', 'S', 'TK', 'F', 'R', 'SC', 'AB', 'CY',
                'FX', 'SK', 'SX', 'VX', 'PF',
            ])}
            if prefix.upper() in _bare_pfxs:
                insulation = 'BARE'

        # ── Dimensions ────────────────────────────────────────────────────
        # Uses a wider context window than _pp_ctx because dimension data
        # boxes on P&IDs are often in a separate table whose OCR text can be
        # far from the equipment tag text.  Both values are soft-coded via
        # dimension_length_context_chars (default 600) and
        # dimension_value_window (default 60 chars after label end).
        _dim_ctx_chars = int(ext_cfg.get('dimension_length_context_chars', 600))
        _dim_val_win   = int(ext_cfg.get('dimension_value_window', 60))
        _dim_start     = max(0, m.start() - _dim_ctx_chars)
        _dim_end       = min(len(text), m.end() + _dim_ctx_chars)
        _dim_ctx       = text[_dim_start:_dim_end].upper()
        dimension_length   = ''
        dimension_diameter = ''
        _len_lbl_m = re.search(_dim_len_lbl, _dim_ctx, re.IGNORECASE)
        if _len_lbl_m:
            _dv = re.search(_dim_val_pat, _dim_ctx[_len_lbl_m.end():_len_lbl_m.end() + _dim_val_win], re.IGNORECASE)
            if _dv and float(_dv.group(1)) > 0:
                _unit = (_dv.group(2) or 'mm').upper()
                dimension_length = f'{_dv.group(1)} {_unit}'
        _dia_lbl_m = re.search(_dim_dia_lbl, _dim_ctx, re.IGNORECASE)
        if _dia_lbl_m:
            _dv = re.search(_dim_val_pat, _dim_ctx[_dia_lbl_m.end():_dia_lbl_m.end() + _dim_val_win], re.IGNORECASE)
            if _dv and float(_dv.group(1)) > 0:
                _unit = (_dv.group(2) or 'mm').upper()
                dimension_diameter = f'{_dv.group(1)} {_unit}'

        # ── Motor Rating ─────────────────────────────────────────────────
        # Uses a WIDER context window than _pp_ctx so motor callouts attached
        # via lead lines (far from the tag in OCR text order) are still found.
        # Soft-coded via motor_rating_context_chars (default 800 chars each
        # side).  OCR at 0°/90°/180°/270° (ocr_rotation_angles) means vertical
        # and downward-oriented motor annotations are already in the text pool.
        _mtr_ctx_chars = int(ext_cfg.get('motor_rating_context_chars', 800))
        _mtr_start     = max(0, m.start() - _mtr_ctx_chars)
        _mtr_end       = min(len(text), m.end() + _mtr_ctx_chars)
        _mtr_ctx       = text[_mtr_start:_mtr_end].upper()
        motor_rating = ''
        _mr_m = re.search(_mtr_pat, _mtr_ctx, re.IGNORECASE)
        if _mr_m:
            motor_rating = f'{_mr_m.group(1)} {_mr_m.group(2).upper()}'
        elif not motor_rating:
            _mr_bare_m = re.search(_mtr_bare_pat, _mtr_ctx, re.IGNORECASE)
            if _mr_bare_m:
                motor_rating = f'{_mr_bare_m.group(1)} {_mr_bare_m.group(2).upper()}'

        # Soft-coded: non-rotating equipment has no motor.  Prefixes are
        # configurable via 'motor_na_prefixes' in equipment_type_config.json.
        # The display value ('No', '', 'N/A' …) is controlled by
        # 'motor_na_display_value' (default 'No') — change in config only.
        _motor_na_pfxs = {p.upper() for p in ext_cfg.get('motor_na_prefixes', [
            'V', 'T', 'D', 'F', 'R', 'E', 'S', 'TK', 'SC', 'AB', 'ST', 'FL',
            'CY', 'DR', 'FG', 'MS', 'SK', 'HX', 'SX', 'FX', 'VX', 'GX',
            'PF', 'DP', 'AN', 'EJ', 'MX',
        ])}
        _motor_na_val = str(ext_cfg.get('motor_na_display_value', 'No'))
        if not motor_rating and prefix.upper() in _motor_na_pfxs:
            motor_rating = _motor_na_val

        # ── Quality Required ─────────────────────────────────────────────
        quality_required = ''
        _qr_m = re.search(_qual_pat, _pp_ctx, re.IGNORECASE)
        if _qr_m:
            quality_required = _qr_m.group(1).strip().upper()

        # Strategy 2: scan wider context for explicit NACE reference.
        # P&IDs for sour-service equipment (H2S/SOUR GAS) should comply with
        # NACE MR0175 / ISO 15156.  When the drawing has NACE in the notes or
        # title block, apply it.  Threshold is soft-coded via
        # 'quality_nace_context_chars' (default 1500).
        if not quality_required:
            _qual_ctx_chars = int(ext_cfg.get('quality_nace_context_chars', 1500))
            _qual_ctx_start = max(0, m.start() - _qual_ctx_chars)
            _qual_ctx_end   = min(len(text), m.end() + _qual_ctx_chars)
            _qual_wide_ctx  = text[_qual_ctx_start:_qual_ctx_end]
            if re.search(r'\bNACE\s*MR\s*0175\b', _qual_wide_ctx, re.IGNORECASE):
                quality_required = 'NACE MR0175'
            elif re.search(r'\bNACE\b', _qual_wide_ctx, re.IGNORECASE):
                quality_required = 'NACE'

        # Strategy 3: infer from sour-service context.  If sour gas / H2S /
        # HIC is detected anywhere in the drawing, all vessels must comply
        # with NACE MR0175.  Soft-coded disable via
        # 'quality_infer_from_sour_service' = false.
        _infer_sour = bool(ext_cfg.get('quality_infer_from_sour_service', True))
        if not quality_required and _infer_sour:
            _sour_ctx = text[:min(len(text), 6000)]  # check first 6 k chars
            if re.search(r'\bSOUR\s+GAS\b|\bH2S\b|\bHIC\b|\bSSC\b',
                         _sour_ctx, re.IGNORECASE):
                quality_required = 'NACE MR0175'
            elif service_fluid and re.search(r'\bsour\b', service_fluid, re.IGNORECASE):
                quality_required = 'NACE MR0175'

        # ── Revision & SL No — pre-tag token scan ─────────────────────────
        # Tabular PDF text writes table cells as newline-separated tokens in
        # reading order.  The pattern is: ...\n[sl_no]\n[revision]\n[TAG]\n...
        # We scan _REV_PRE_TAG_TOKENS tokens immediately before the tag match.
        # Soft-coded via _REV_PRE_TAG_WIN_CHARS, _REV_PRE_TAG_TOKENS,
        # _clean_revision(), and the SL-no regex below.
        _pre_text = text[max(0, m.start() - _REV_PRE_TAG_WIN_CHARS):m.start()]
        _pre_toks = [t.strip() for t in _pre_text.split('\n') if t.strip()]
        revision  = ''
        sl_no     = ''
        if _pre_toks:
            # Closest token before the tag is most likely the revision cell
            _rev_candidate = _clean_revision(_pre_toks[-1])
            if _rev_candidate:
                revision = _rev_candidate
                # Token before the revision cell is likely the SL No
                if len(_pre_toks) >= 2 and re.match(r'^\d{1,3}$', _pre_toks[-2]):
                    sl_no = _pre_toks[-2]
            elif re.match(r'^\d{1,3}$', _pre_toks[-1]):
                # Last token is a number; could be sl_no with no revision column
                sl_no = _pre_toks[-1]

        results.append({
            'tag':                 tag,
            'type_label':          type_label,
            'description':         description,
            'revision':            revision,
            'sl_no':               sl_no,
            'area':                area,
            'drawing_ref':         drawing_ref,
            'line_connections':    lc_tokens,
            'nozzle_connections':  nozzle_tokens,
            'service_fluid':       service_fluid,
            'material_class':      material_class,
            'process_notes':       process_notes,
            # Process parameters extracted from data-bubble context
            'design_flowrate':     design_flowrate,
            'oper_pressure':       oper_pressure,
            'oper_temperature':    oper_temperature,
            'design_pressure_min': design_pressure_min,
            'design_pressure_max': design_pressure_max,
            'design_temp_min':     design_temp_min,
            'design_temp_max':     design_temp_max,
            'insulation':          insulation,
            'dimension_length':    dimension_length,
            'dimension_diameter':  dimension_diameter,
            'motor_rating':        motor_rating,
            'quality_required':    quality_required,
        })

        # ── Merge data-box index values (fill empty fields only) ──────────
        # Values extracted by _build_databox_index from the full text are
        # merged in here.  Only fills columns that the narrow context-window
        # extraction above left empty — never overwrites a found value.
        _db_vals = _databox_idx.get(tag.upper(), {})
        if _db_vals:
            _last = results[-1]
            for _fk, _fv in _db_vals.items():
                if _fk in _last:
                    # For _min/_max fields the explicit databox label extraction
                    # (e.g. "DESIGN TEMP (MAX/MIN): 185 F / -13.2 F") is more
                    # reliable than the narrow context scan — always override
                    # so a wrong single-value pickup from context is corrected.
                    if _fk.endswith(('_min', '_max')):
                        if _fv:  # only write a non-empty databox value
                            _last[_fk] = _fv
                    elif not _last[_fk]:
                        _last[_fk] = _fv
            # Normalise oper_temperature if it was filled from the databox
            # (the narrow-context path already calls _normalize_oper_temp;
            #  the databox path does not, so we apply it here).
            if _last.get('oper_temperature') and '/' in _last['oper_temperature']:
                _last['oper_temperature'] = _normalize_oper_temp(_last['oper_temperature'])
            # ── Fallback: split raw "X / Y" values that landed in a single
            # _max or _min field (happens when the (MAX/MIN) label variant
            # did not match and the shorter fallback label fired instead,
            # e.g. "design temp" → design_temp_max = "185 F / -13.2 F").
            # We detect these, split numerically and re-assign correctly.
            for _single_fk in ('design_temp_max', 'design_pressure_max', 'design_temp_min', 'design_pressure_min'):
                _sv = _last.get(_single_fk, '')
                if _sv and '/' in _sv:
                    _sparts = re.split(r'\s*/\s*', _sv, maxsplit=1)
                    if len(_sparts) == 2:
                        _sn0 = re.search(r'-?\d+(?:\.\d+)?', _sparts[0])
                        _sn1 = re.search(r'-?\d+(?:\.\d+)?', _sparts[1])
                        if _sn0 and _sn1:
                            _sf0, _sf1 = float(_sn0.group()), float(_sn1.group())
                            # Determine the _min/_max counterpart field name
                            if _single_fk.endswith('_max'):
                                _counterpart = _single_fk[:-4] + '_min'
                            else:
                                _counterpart = _single_fk[:-4] + '_max'
                            # Assign larger to _max, smaller to _min
                            _larger  = _sparts[0] if _sf0 >= _sf1 else _sparts[1]
                            _smaller = _sparts[0] if _sf0 <= _sf1 else _sparts[1]
                            if _single_fk.endswith('_max'):
                                _last[_single_fk] = _larger
                                if _counterpart in _last and not _last[_counterpart]:
                                    _last[_counterpart] = _smaller
                            else:
                                _last[_single_fk] = _smaller
                                if _counterpart in _last and not _last[_counterpart]:
                                    _last[_counterpart] = _larger
            print(
                f'[EQ-DIAG][DataBox] Merged into {tag}: '
                + str({k: v for k, v in _db_vals.items() if _last.get(k)}),
                flush=True,
            )

    results.sort(key=lambda x: x['tag'])

    # ── Gate 2: data-box presence post-filter ────────────────────────────
    # Soft-coded via 'require_at_least_one_param' and
    # 'prefer_databox_index_filter' in equipment_type_config.json.
    #
    # Two-tier strategy:
    #
    # Tier A — _databox_idx (preferred, more reliable):
    #   _build_databox_index scans label:value pairs across the FULL OCR
    #   text and associates each value with the NEAREST preceding tag. This
    #   proximity-based attribution is immune to context-window bleed: even
    #   if V-804-TF and V-308-TF both appear in a 400-char window around a
    #   "OPER PRESS" label, the LAST tag before the label in OCR text order
    #   is V-308-TF (the actual data-box owner). So _databox_idx accurately
    #   contains {V-308-TF: {oper_pressure: ..., ...}} and nothing for
    #   referenced tags like V-804-TF.
    #   When _databox_idx has any entries, use it as the SOLE authoritative
    #   list of primary equipment on this drawing.
    #   Controlled by 'prefer_databox_index_filter' (default true).
    #
    # Tier B — context-window param-field check (fallback):
    #   Used when _databox_idx is empty (label format not in databox_label_map,
    #   drawing has no data boxes, etc.). Any item with at least one non-empty
    #   core process parameter from the regex context-window pass is kept.
    #   Controlled by 'require_at_least_one_param' (default true).
    if bool(ext_cfg.get('require_at_least_one_param', True)):
        _primary_fields = ext_cfg.get(
            'param_fields_for_primary_check',
            ['oper_pressure', 'oper_temperature',
             'design_pressure_min', 'design_pressure_max',
             'design_temp_min', 'design_temp_max',
             'dimension_length', 'dimension_diameter',
             'design_flowrate'],
        )
        _prefer_db_idx = bool(ext_cfg.get('prefer_databox_index_filter', True))
        _kept, _dropped = [], []

        if _prefer_db_idx and _databox_idx:
            # ── Tier A: databox-index is authoritative ───────────────────
            # Only tags confirmed by the proximity-based label scan are primary.
            for _item in results:
                if _item['tag'].upper() in _databox_idx:
                    _kept.append(_item)
                else:
                    _dropped.append(_item['tag'])
            print(
                f'[EQ-DIAG] Gate2-TierA: kept {len(_kept)} databox-indexed tag(s), '
                f'removed {len(_dropped)} referenced tag(s): {_dropped}',
                flush=True,
            )
        else:
            # ── Tier B: context-window param check (fallback) ────────────
            for _item in results:
                _has_param = any(_item.get(f) for f in _primary_fields)
                if _has_param:
                    _kept.append(_item)
                else:
                    _dropped.append(_item['tag'])
            if _dropped:
                print(
                    f'[EQ-DIAG] Gate2-TierB: removed {len(_dropped)} no-param '
                    f'referenced tag(s): {_dropped}',
                    flush=True,
                )

        results = _kept if _kept else results  # safety: never return empty when all pass

    # ── Description reject filter ───────────────────────────────────
    # Soft-coded via 'description_reject_patterns' in equipment_type_config.json.
    # Any item whose description field matches ANY pattern is excluded.
    # Use-case: OCR picks up revision-cloud annotation text ("Revision Cloud")
    # or description-only rows that have no real equipment tag meaning.
    _desc_reject_pats = ext_cfg.get('description_reject_patterns', [])
    if _desc_reject_pats:
        _pre_dr = len(results)
        results = [
            _item for _item in results
            if not any(
                re.search(_drpat, _item.get('description', '') or '')
                for _drpat in _desc_reject_pats
            )
        ]
        _removed_dr = _pre_dr - len(results)
        if _removed_dr:
            print(
                f'[EQ-DIAG] Description-reject filter removed {_removed_dr} item(s)',
                flush=True,
            )
    # Soft-coded via config key 'minmax_correction_pairs' in the 'extraction'
    # section. Ensures that design_temp_max is always numerically >= design_temp_min
    # regardless of which code path populated them.
    _minmax_pairs = ext_cfg.get('minmax_correction_pairs', [
        ['design_temp_min',      'design_temp_max'],
        ['design_pressure_min',  'design_pressure_max'],
    ])
    for _item in results:
        for _fmin, _fmax in _minmax_pairs:
            _vmin = _item.get(_fmin, '')
            _vmax = _item.get(_fmax, '')
            if _vmin and _vmax:
                _nmin = re.search(r'-?\d+(?:\.\d+)?', _vmin)
                _nmax = re.search(r'-?\d+(?:\.\d+)?', _vmax)
                if _nmin and _nmax and float(_nmin.group()) > float(_nmax.group()):
                    # Values are inverted — swap
                    _item[_fmin], _item[_fmax] = _vmax, _vmin

    # ── Post-deduplication: canonicalize to the most-specific tag form ────
    # When the same physical equipment appears in two forms (e.g. both "V-308"
    # and "V-308-TF"), keep only the longer (more specific) form.
    # Two different equipment with the same base but different suffixes
    # (e.g. V-805-TF vs V-805-1F) are intentionally kept as separate rows.
    _suffix_re = re.compile(r'^([A-Z]{1,2}-[0-9]{3,5}[A-Z]?)-[A-Z0-9]{1,4}$')
    _full_tag_bases: set = set()
    for _item in results:
        _m = _suffix_re.match(_item['tag'])
        if _m:
            _full_tag_bases.add(_m.group(1))  # e.g. "V-308" from "V-308-TF"
    # Remove bare-base entries that have a suffixed sibling
    results = [
        _item for _item in results
        if _item['tag'] not in _full_tag_bases
    ]

    return results


def _dedup_equipment_by_tag(items: list) -> list:
    """
    Deduplicate a merged multi-page equipment list by tag.

    When the same tag appears on more than one page (e.g. it was referenced on
    page 1 AND has its own data box on page 2), keep only the entry with the
    most populated fields.  This ensures the richest extraction result wins.
    """
    by_tag: dict = {}
    for item in items:
        tag = (item.get('tag') or '').upper()
        if not tag:
            continue
        if tag not in by_tag:
            by_tag[tag] = item
        else:
            _skip_keys = {'sl_no', 'tag', 'type_label', 'area', 'drawing_ref',
                          'line_connections', 'nozzle_connections'}
            _pop_new = sum(1 for k, v in item.items()
                           if k not in _skip_keys and v and v not in ('', 'No', [], 'N/A'))
            _pop_old = sum(1 for k, v in by_tag[tag].items()
                           if k not in _skip_keys and v and v not in ('', 'No', [], 'N/A'))
            if _pop_new > _pop_old:
                by_tag[tag] = item
    result = list(by_tag.values())
    print(f'[EQ-DIAG] _dedup_equipment_by_tag: {len(items)} in → {len(result)} unique tags out', flush=True)
    return result


_result_store: dict = {}

# ── Soft-coded AI gap-fill constants ─────────────────────────────────────────
# Default list of fields the AI gap-fill pass will attempt.
# Controlled per-deploy via equipment_type_config.json extraction.ai_gap_fill_fields.
# Fields already populated by the regex pass are NEVER overwritten.
_AI_GAP_FILL_DEFAULT_FIELDS = [
    'oper_pressure',
    'oper_temperature',
    'design_pressure_min',
    'design_pressure_max',
    'design_temp_min',
    'design_temp_max',
    'design_flowrate',
    'moc',
    'insulation',
    'description',
]

# JSON response validation: AI must return keys matching these field names.
# Values must be strings (or null). Any other structure is rejected.
_AI_FILL_FIELD_SET = set(_AI_GAP_FILL_DEFAULT_FIELDS)


def _ai_gap_fill_pid_items(items: list, text: str, config: dict) -> list:
    """
    Multi-model AI gap-fill pass for P&ID drawing mode.

    For each equipment item with one or more empty target fields, re-extracts
    a wider text context window around the tag and sends it to:
      1. OpenAI GPT-4o (primary)  — with automatic Gemini fallback inside
         MultiModelAIService if the OpenAI quota is exceeded.
      2. Gemini Flash             — as an independent second-opinion when
         ai_gap_fill_provider == 'both', to fill any fields GPT-4o left null.

    The AI is prompted to return ONLY a flat JSON object.  Values are merged
    into the item ONLY when the field is still empty after regex extraction.

    Soft-coded via equipment_type_config.json extraction section:
      ai_gap_fill_enabled       : true/false  (default true)
      ai_gap_fill_provider      : "both" | "openai" | "gemini"
      ai_gap_fill_fields        : list of field keys to attempt
      ai_gap_fill_context_chars : chars each side of the tag (default 800)
      ai_gap_fill_max_tokens    : max tokens for AI response (default 350)
      ai_gap_fill_temperature   : sampling temperature (default 0)
      ai_gap_fill_min_empty_fields : minimum empty fields before AI is called
    """
    import json as _json

    ext_cfg     = config.get('extraction', {})
    enabled     = bool(ext_cfg.get('ai_gap_fill_enabled', True))
    if not enabled:
        return items

    fill_fields = list(ext_cfg.get('ai_gap_fill_fields', _AI_GAP_FILL_DEFAULT_FIELDS))
    ctx_chars   = int(ext_cfg.get('ai_gap_fill_context_chars', 800))
    max_tokens  = int(ext_cfg.get('ai_gap_fill_max_tokens', 350))
    temperature = float(ext_cfg.get('ai_gap_fill_temperature', 0))
    min_empty   = int(ext_cfg.get('ai_gap_fill_min_empty_fields', 1))
    provider    = str(ext_cfg.get('ai_gap_fill_provider', 'both')).lower()

    # Lazy-import to avoid circular imports and keep startup fast
    try:
        from apps.pid_analysis.multi_model_service import MultiModelAIService
        ai = MultiModelAIService()
    except Exception as _e:
        print(f'[EQ-DIAG][AI-FILL] Service init failed: {_e}', flush=True)
        return items

    text_upper = text.upper()

    # ── Shared prompt builder ────────────────────────────────────────────────
    _FIELD_HINTS = {
        'oper_pressure':      'Operating pressure with unit (e.g. "155 PSIG" or "10 barg")',
        'oper_temperature':   'Operating temperature with unit (e.g. "105 °F" or "40 °C")',
        'design_pressure_min': 'Minimum design/set pressure with unit (e.g. "-13.2 PSIG")',
        'design_pressure_max': 'Maximum design/set pressure with unit (e.g. "195 PSIG")',
        'design_temp_min':    'Minimum design temperature with unit (e.g. "-13 °F")',
        'design_temp_max':    'Maximum design temperature with unit (e.g. "185 °F")',
        'design_flowrate':    'Design flowrate or capacity with unit (e.g. "327 M3" or "100 M3/H")',
        'moc':                'Material of construction abbreviation (e.g. "CS", "SS316L", "DUPLEX")',
        'insulation':         'Insulation type code (e.g. "PERS", "HOT", "BARE", "TRACED")',
        'description':        'Equipment description (2-6 words, e.g. "OIL SLUG CATCHER")',
    }

    def _build_prompt(tag: str, empty: list, ctx: str) -> str:
        field_list = '\n'.join(
            f'  "{f}": {_FIELD_HINTS.get(f, f)}'
            for f in empty
        )
        return (
            f'You are an Oil & Gas P&ID data extraction assistant.\n'
            f'Extract the following fields for equipment tag {tag} from the text excerpt below.\n'
            f'Return ONLY a valid JSON object with these exact keys. '
            f'Set the value to null if the field cannot be found.\n'
            f'Do NOT include markdown fences, explanations, or keys not listed.\n\n'
            f'Fields to extract:\n{field_list}\n\n'
            f'TEXT EXCERPT:\n{ctx}'
        )

    def _call_ai(prompt: str, model_hint: str) -> dict:
        """Call the AI and return parsed dict, or {} on failure."""
        try:
            raw = ai.chat_completion(
                messages=[{'role': 'user', 'content': prompt}],
                model=model_hint,
                max_tokens=max_tokens,
                temperature=temperature,
            )
            # Strip markdown code fences if present
            cleaned = re.sub(r'^```[a-z]*\s*', '', raw.strip(), flags=re.IGNORECASE)
            cleaned = re.sub(r'\s*```$', '', cleaned.strip())
            parsed  = _json.loads(cleaned)
            if isinstance(parsed, dict):
                return parsed
        except Exception as _e:
            print(f'[EQ-DIAG][AI-FILL] parse/call error ({model_hint}): {_e}', flush=True)
        return {}

    # ── Per-item gap fill ────────────────────────────────────────────────────
    for item in items:
        tag = item.get('tag', '')
        if not tag:
            continue

        empty_fields = [f for f in fill_fields if not item.get(f)]
        if len(empty_fields) < min_empty:
            continue

        # Re-locate tag in text for context window
        idx = text_upper.find(tag.upper())
        if idx == -1:
            continue
        ctx_start = max(0, idx - ctx_chars // 2)
        ctx_end   = min(len(text), idx + ctx_chars // 2)
        ctx       = text[ctx_start:ctx_end]

        prompt = _build_prompt(tag, empty_fields, ctx)

        # ── Pass 1: primary provider (GPT-4o, with auto Gemini fallback) ──
        gpt_result = {}
        if provider in ('openai', 'both'):
            gpt_result = _call_ai(prompt, 'openai')

        # ── Pass 2: Gemini second-opinion (fills any fields GPT-4o left null) ──
        gem_result = {}
        if provider in ('gemini', 'both'):
            still_empty = [f for f in empty_fields if not gpt_result.get(f)]
            if still_empty:
                gem_result = _call_ai(_build_prompt(tag, still_empty, ctx), 'gemini')

        # ── Merge: OpenAI wins over Gemini; neither overwrites regex values ──
        filled = []
        for f in empty_fields:
            if item.get(f):
                continue                       # already filled by regex — skip
            val = gpt_result.get(f) or gem_result.get(f)
            if val and str(val).strip().lower() not in ('null', 'none', ''):
                item[f] = str(val).strip()
                filled.append(f)

        if filled:
            print(f'[EQ-DIAG][AI-FILL] {tag}: AI filled {filled}', flush=True)

    return items


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def analyze_pid_equipment(request):
    """POST /api/v1/pid/equipment/analyze/"""
    config  = _load_config()
    ext_cfg = config.get('extraction', {})
    allowed = [e.lower() for e in ext_cfg.get('allowed_extensions', ['pdf'])]
    max_mb  = float(ext_cfg.get('max_file_size_mb', 50))

    pid_file = request.FILES.get('file') or (list(request.FILES.values())[0] if request.FILES else None)
    if not pid_file:
        return Response({'error': 'No file provided', 'success': False},
                        status=drf_status.HTTP_400_BAD_REQUEST)

    ext = pid_file.name.rsplit('.', 1)[-1].lower()
    if ext not in allowed:
        return Response({'error': f'Unsupported format: .{ext}. Allowed: {", ".join(allowed)}', 'success': False},
                        status=drf_status.HTTP_400_BAD_REQUEST)

    if pid_file.size > max_mb * 1024 * 1024:
        return Response({'error': f'File exceeds {max_mb} MB limit', 'success': False},
                        status=drf_status.HTTP_400_BAD_REQUEST)

    drawing_ref = pid_file.name.rsplit('.', 1)[0]
    upload_id   = f'EQ-{uuid.uuid4().hex[:12].upper()}'

    print(f'[EQ-DIAG] Analyzing: {pid_file.name}  upload_id={upload_id}', flush=True)
    _debug_info: dict = {'file': pid_file.name, 'upload_id': upload_id}

    try:
        # ── Stage 1: try Equipment Register (18-field table) extraction ──
        equipment       = _extract_equipment_register_rows(pid_file, config)
        extraction_mode = 'register'

        if equipment is None:
            # ── Stage 2: fall back to P&ID drawing mode ──────────────────────
            # Soft-coded: multi_page_mode = "per_page" (default) processes each
            # PDF page in isolation so cross-page tag references cannot pollute
            # the results. Set to "combined" to revert to the old single-pass.
            print('[EQ-DIAG] Register mode returned None -> falling back to P&ID mode', flush=True)
            pid_file.seek(0)
            _pid_bytes = pid_file.read()   # read once; reuse for all page passes

            # Count pages
            _total_pages = 1
            try:
                import fitz as _fitz_pc
                _pc_doc = _fitz_pc.open(stream=_pid_bytes, filetype='pdf')
                _total_pages = len(_pc_doc)
                _pc_doc.close()
            except Exception as _e:
                logger.debug('[EquipmentList] Page count failed: %s', _e)

            _multi_mode    = ext_cfg.get('multi_page_mode', 'per_page')
            _tb_rev_enabled = bool(ext_cfg.get('titleblock_revision_enabled', True))
            print(f'[EQ-DIAG] PDF pages={_total_pages}  multi_page_mode={_multi_mode!r}', flush=True)

            def _process_pid_page(page_idx):
                """Extract, filter, AI-fill and revision-stamp one PDF page.
                Returns (schema_items, page_drawing_ref, page_text)."""
                import io as _io_pg
                _pf   = _io_pg.BytesIO(_pid_bytes)
                _ptxt = _extract_text_from_pdf(_pf, config, _page_index=page_idx)

                # Title-block DWG NO: try vector-only first (higher fidelity)
                _pvec_no = ''
                try:
                    import fitz as _fitz_v
                    _vdoc  = _fitz_v.open(stream=_pid_bytes, filetype='pdf')
                    _vpg   = _vdoc[page_idx]
                    _vtext = (_normalize_text(_vpg.get_text('text') or '') + '\n'
                              + _normalize_text(' '.join(
                                  w[4] for w in sorted(
                                      _vpg.get_text('words') or [],
                                      key=lambda w: (round(w[1] / 15) * 15, w[0])
                                  )
                              )))
                    _vdoc.close()
                    _pvec_no = _extract_titleblock_dwg_no(_vtext)
                    if _pvec_no:
                        print(f'[EQ-DIAG][P{page_idx+1}] Vector DWG NO: {_pvec_no!r}', flush=True)
                except Exception as _e:
                    logger.debug('[EquipmentList] Vector DWG NO (page %d): %s', page_idx, _e)

                _pg_ref = (_pvec_no
                           or _extract_titleblock_dwg_no(_ptxt)
                           or (f'{drawing_ref}_P{page_idx + 1}' if _total_pages > 1 else drawing_ref))

                _pg_raw = _extract_equipment_items(_ptxt, _pg_ref, config)
                _pg_sch = [_pid_item_to_register_schema(item) for item in _pg_raw]
                print(f'[EQ-DIAG][P{page_idx+1}] {_pg_ref!r}: {len(_pg_raw)} raw → {len(_pg_sch)} schema items', flush=True)
                print(f'[EQ-DIAG][P{page_idx+1}] Tags: {[i["tag"] for i in _pg_raw]}', flush=True)

                if _pg_sch and _ptxt:
                    _pg_sch = _ai_gap_fill_pid_items(_pg_sch, _ptxt, config)

                if _REVISION_USE_TOPMOST and _tb_rev_enabled:
                    _pg_rev = _extract_titleblock_revision(_ptxt)
                    if _pg_rev:
                        for _it in _pg_sch:
                            _it['revision'] = _pg_rev
                        print(f'[EQ-DIAG][P{page_idx+1}] Revision "{_pg_rev}" applied to {len(_pg_sch)} items', flush=True)

                return _pg_sch, _pg_ref, _ptxt

            if _total_pages > 1 and _multi_mode == 'per_page':
                # ── Multi-page: each sheet is an independent drawing ──────────
                extraction_mode = 'pid_drawing_multipage'
                _all_items: list = []
                _page_refs: list = []
                for _page_idx in range(_total_pages):
                    _pg_items, _pg_ref, _ = _process_pid_page(_page_idx)
                    _all_items.extend(_pg_items)
                    if _pg_items:
                        _page_refs.append(_pg_ref)
                equipment = _dedup_equipment_by_tag(_all_items)
                # drawing_ref: join unique per-page refs (or keep filename base)
                if _page_refs:
                    drawing_ref = ' | '.join(dict.fromkeys(_page_refs))
                text = ''  # no single text blob in multi-page mode
                _debug_info.update({
                    'total_pages': _total_pages,
                    'page_refs': _page_refs,
                    'after_dedup_count': len(equipment),
                })
                print(f'[EQ-DIAG] Multi-page: {_total_pages} pages  {len(_all_items)} raw → {len(equipment)} after dedup', flush=True)
            else:
                # ── Single-page or combined mode ─────────────────────────────
                _pg_items, _pg_ref, text = _process_pid_page(0)
                equipment = _pg_items
                drawing_ref = _pg_ref
                extraction_mode = 'pid_drawing'
                _debug_info.update({
                    'text_len': len(text),
                    'text_preview': text[:400] if text else '',
                    'raw_items_count': len(equipment),
                    'after_dedup_count': len(equipment),
                })
                print(f'[EQ-DIAG] Single-page P&ID: {_pg_ref!r}  {len(equipment)} items', flush=True)

            # Diagnostic: all tag-shaped tokens (single-page mode only)
            if text:
                import re as _re
                _all_candidates = _re.findall(r'\b[A-Z]{1,2}-[0-9]{3,5}[A-Z]?(?:-[A-Z0-9]{1,4})?\b', text)
                print(f'[EQ-DIAG] All tag-shaped tokens in text: {list(dict.fromkeys(_all_candidates))}', flush=True)

        for idx, item in enumerate(equipment, 1):
            if not item.get('sl_no'):
                item['sl_no'] = str(idx)
            item['drawing_ref'] = drawing_ref

        # ── Equipment type classification (soft-coded via config) ──────────
        _desg_codes   = config.get('designation_codes', {})
        _prefix_map   = config.get('tag_prefix_type_map', {})
        # Sort prefix keys longest-first so 'ST' matches before 'S'
        _pfx_keys     = sorted(_prefix_map.keys(), key=len, reverse=True)
        _type_re      = re.compile(r'^([A-Z]{1,4})')
        for _item in equipment:
            _tag_pfx = _type_re.match(_item.get('tag', ''))
            if _tag_pfx:
                _pfx = _tag_pfx.group(1)
                _desig = None
                for _pk in _pfx_keys:
                    if _pfx.startswith(_pk):
                        _desig = _prefix_map[_pk]
                        break
                if _desig and _desig in _desg_codes:
                    _item['equipment_type']      = _desig
                    _item['equipment_type_name'] = _desg_codes[_desig]['name']
                    _item['equipment_category']  = _desg_codes[_desig]['category']
                else:
                    _item.setdefault('equipment_type', '')
                    _item.setdefault('equipment_type_name', '')
                    _item.setdefault('equipment_category', '')

        # ── Persist extracted items to DB (upsert by upload_id + tag) ──────
        try:
            PIDEquipmentType, PIDEquipmentItem = _get_equipment_models()
            _db_user = getattr(request, 'user', None)
            _scalar_keys = {
                'revision', 'description', 'extraction_mode',
                'sl_no', 'tag', 'drawing_ref',
                'equipment_type', 'equipment_type_name', 'equipment_category',
            }
            for _item in equipment:
                _etag  = _item.get('tag', '')
                _edata = {k: v for k, v in _item.items() if k not in _scalar_keys}
                _etype_code = _item.get('equipment_type') or None
                _etype_obj  = None
                if _etype_code:
                    _etype_obj, _ = PIDEquipmentType.objects.get_or_create(
                        code=_etype_code,
                        defaults={
                            'name':        _desg_codes.get(_etype_code, {}).get('name', _etype_code),
                            'category':    _desg_codes.get(_etype_code, {}).get('category', 'MISC'),
                            'is_rotating': bool(_desg_codes.get(_etype_code, {}).get('rotating', False)),
                        },
                    )
                PIDEquipmentItem.objects.update_or_create(
                    upload_id=upload_id,
                    tag=_etag,
                    defaults={
                        'drawing_ref':     drawing_ref,
                        'revision':        _item.get('revision', ''),
                        'description':     _item.get('description', ''),
                        'extraction_mode': extraction_mode,
                        'equipment_type':  _etype_obj,
                        'data':            _edata,
                        'uploaded_by':     _db_user if _db_user and _db_user.is_authenticated else None,
                    },
                )
            print(f'[EQ-DIAG] Saved {len(equipment)} items to DB (upload_id={upload_id})', flush=True)
        except Exception as _db_exc:
            # DB save failure must NOT break the API response
            print(f'[EQ-DIAG] DB save WARNING: {_db_exc}', flush=True)

        _result_store[upload_id] = {
            'status':          'completed',
            'equipment':       equipment,
            'total':           len(equipment),
            'drawing_ref':     drawing_ref,
            'extraction_mode': extraction_mode,
        }

        _debug_info['extraction_mode'] = extraction_mode
        _debug_info['total'] = len(equipment)
        print(f'[EQ-DIAG] Done: {len(equipment)} items  mode={extraction_mode}', flush=True)
        # Per-item diagnostic — shows tag + revision + how many fields are populated
        for _item in equipment:
            _pop = sum(1 for k, v in _item.items()
                       if k not in ('sl_no','tag','type_label','area','drawing_ref',
                                    'line_connections','nozzle_connections')
                       and (v if not isinstance(v, list) else v))
            print(f'[EQ-DIAG]   rev={_item.get("revision","")}  tag={_item.get("tag","")}  desc={repr(_item.get("description",""))[:40]}  pop={_pop}', flush=True)

        resp_body: dict = {
            'success':         True,
            'upload_id':       upload_id,
            'status':          'completed',
            'equipment':       equipment,
            'total':           len(equipment),
            'drawing_ref':     drawing_ref,
            'extraction_mode': extraction_mode,
            'columns':         [c['label'] for c in config.get('excel_columns', []) if c['key'] != 'sl_no'],
        }
        # Include debug info automatically when nothing was extracted,
        # so the frontend can display diagnostic details without Docker log access.
        if len(equipment) == 0:
            resp_body['debug_info'] = _debug_info
        return Response(resp_body, status=drf_status.HTTP_200_OK)

    except Exception as exc:
        logger.error('[EquipmentList] Error: %s', exc, exc_info=True)
        _result_store[upload_id] = {'status': 'failed', 'error': str(exc)}
        return Response({'error': f'Extraction failed: {exc}', 'success': False},
                        status=drf_status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_equipment_analysis_status(request, upload_id):
    """GET /api/v1/pid/equipment/status/<upload_id>/"""
    entry = _result_store.get(upload_id)
    if not entry:
        return Response({'upload_id': upload_id, 'status': 'not_found', 'progress': 0},
                        status=drf_status.HTTP_404_NOT_FOUND)
    return Response({
        'upload_id': upload_id,
        'status':    entry.get('status', 'processing'),
        'progress':  100 if entry.get('status') == 'completed' else 50,
        'message':   entry.get('error', 'Extraction complete' if entry.get('status') == 'completed' else 'Processing...'),
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_equipment_analysis_results(request, upload_id):
    """GET /api/v1/pid/equipment/results/<upload_id>/"""
    entry = _result_store.get(upload_id)
    if not entry:
        return Response({'error': 'Results not found - re-upload the file', 'upload_id': upload_id},
                        status=drf_status.HTTP_404_NOT_FOUND)
    if entry.get('status') == 'failed':
        return Response({'error': entry.get('error', 'Extraction failed'), 'upload_id': upload_id},
                        status=drf_status.HTTP_500_INTERNAL_SERVER_ERROR)

    config = _load_config()
    return Response({
        'success':     True,
        'upload_id':   upload_id,
        'equipment':   entry.get('equipment', []),
        'total':       entry.get('total', 0),
        'drawing_ref': entry.get('drawing_ref', ''),
        'columns':     [c['label'] for c in config.get('excel_columns', []) if c['key'] != 'sl_no'],
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def download_equipment_excel(request, upload_id):
    """GET /api/v1/pid/equipment/download-excel/<upload_id>/"""
    entry = _result_store.get(upload_id)
    if not entry or entry.get('status') != 'completed':
        return Response({'error': 'Results not available - re-upload the file'},
                        status=drf_status.HTTP_404_NOT_FOUND)

    config    = _load_config()
    col_defs  = config.get('excel_columns', [])
    equipment = entry.get('equipment', [])
    drawing   = entry.get('drawing_ref', 'equipment_list')

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Equipment List'

        header_font  = Font(bold=True, color='FFFFFF', size=11)
        header_fill  = PatternFill(start_color='1E3A5F', end_color='1E3A5F', fill_type='solid')
        header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin_border  = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'),  bottom=Side(style='thin'),
        )
        alt_fill = PatternFill(start_color='EFF6FF', end_color='EFF6FF', fill_type='solid')

        headers = [c['label'] for c in col_defs]
        for col_idx, label in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=label)
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = header_align
            cell.border    = thin_border

        ws.row_dimensions[1].height = 30

        for row_idx, item in enumerate(equipment, 2):
            row_fill = alt_fill if row_idx % 2 == 0 else None
            for col_idx, col_def in enumerate(col_defs, 1):
                key   = col_def['key']
                value = item.get(key, '')
                if isinstance(value, list):
                    value = ', '.join(str(v) for v in value) if value else '-'
                elif value == '' or value is None:
                    value = '-'
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = Alignment(vertical='center', wrap_text=(key in ('line_connections', 'description')))
                cell.border    = thin_border
                if row_fill:
                    cell.fill = row_fill

        for col_idx, col_def in enumerate(col_defs, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = col_def.get('width', 18)

        ws.freeze_panes = 'A2'

        import io
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        safe_name = re.sub(r'[^\w\-]', '_', drawing)
        response  = HttpResponse(
            buf.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="{safe_name}_equipment_list.xlsx"'
        return response

    except ImportError:
        return Response({'error': 'openpyxl is not installed on the server'},
                        status=drf_status.HTTP_501_NOT_IMPLEMENTED)
    except Exception as exc:
        logger.error('[EquipmentList] Excel error: %s', exc, exc_info=True)
        return Response({'error': f'Excel generation failed: {exc}'},
                        status=drf_status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def analyze_pid_equipment_batch(request):
    """POST /api/v1/pid/equipment/analyze-batch/
    Accepts multiple files and returns combined equipment results for all drawings.
    """
    config  = _load_config()
    ext_cfg = config.get('extraction', {})
    allowed = [e.lower() for e in ext_cfg.get('allowed_extensions', ['pdf'])]
    max_mb  = float(ext_cfg.get('max_file_size_mb', 50))

    files = list(request.FILES.values())
    if not files:
        return Response({'error': 'No files provided', 'success': False},
                        status=drf_status.HTTP_400_BAD_REQUEST)

    all_equipment = []
    drawing_refs  = []
    upload_id     = f'EQB-{uuid.uuid4().hex[:12].upper()}'

    for pid_file in files:
        ext = pid_file.name.rsplit('.', 1)[-1].lower()
        if ext not in allowed:
            return Response(
                {'error': f'Unsupported format: .{ext}. Allowed: {", ".join(allowed)}', 'success': False},
                status=drf_status.HTTP_400_BAD_REQUEST,
            )
        if pid_file.size > max_mb * 1024 * 1024:
            return Response(
                {'error': f'{pid_file.name} exceeds {max_mb} MB limit', 'success': False},
                status=drf_status.HTTP_400_BAD_REQUEST,
            )

    for pid_file in files:
        drawing_ref = pid_file.name.rsplit('.', 1)[0]
        drawing_refs.append(drawing_ref)
        logger.info('[EquipmentList Batch] Analyzing: %s  upload_id=%s', pid_file.name, upload_id)
        try:
            # Try register mode first; fall back to P&ID mode
            equipment = _extract_equipment_register_rows(pid_file, config)
            if equipment is None:
                pid_file.seek(0)
                text      = _extract_text_from_pdf(pid_file, config)
                # Override drawing_ref with actual DWG NO from title block
                _tb_dwg_no = _extract_titleblock_dwg_no(text)
                if _tb_dwg_no:
                    drawing_ref          = _tb_dwg_no
                    drawing_refs[-1]     = _tb_dwg_no   # update the list entry too
                raw_items = _extract_equipment_items(text, drawing_ref, config)
                equipment = [_pid_item_to_register_schema(item) for item in raw_items]
            for idx, item in enumerate(equipment, 1):
                if not item.get('sl_no'):
                    item['sl_no'] = str(idx)
                item['drawing_ref'] = drawing_ref
            all_equipment.extend(equipment)
        except Exception as exc:
            logger.error('[EquipmentList Batch] Error on %s: %s', pid_file.name, exc, exc_info=True)
            _result_store[upload_id] = {'status': 'failed', 'error': str(exc)}
            return Response({'error': f'Extraction failed for {pid_file.name}: {exc}', 'success': False},
                            status=drf_status.HTTP_500_INTERNAL_SERVER_ERROR)

    # Re-number across drawings
    for idx, item in enumerate(all_equipment, 1):
        item['sl_no'] = idx

    _result_store[upload_id] = {
        'status':      'completed',
        'equipment':   all_equipment,
        'total':       len(all_equipment),
        'drawing_ref': ', '.join(drawing_refs),
    }

    logger.info('[EquipmentList Batch] Done: %d items from %d drawing(s)', len(all_equipment), len(files))

    return Response({
        'success':     True,
        'upload_id':   upload_id,
        'status':      'completed',
        'equipment':   all_equipment,
        'total':       len(all_equipment),
        'drawing_ref': ', '.join(drawing_refs),
        'columns':     [c['label'] for c in config.get('excel_columns', []) if c['key'] != 'sl_no'],
    }, status=drf_status.HTTP_200_OK)
