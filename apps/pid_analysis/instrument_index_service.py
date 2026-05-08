"""
Instrument Index Service — ALL-INSTRUMENT extraction from P&ID drawings
-----------------------------------------------------------------------
Multi-engine extraction waterfall (soft-coded via EXTRACTION_CONFIG):

  Engine 1  — PyMuPDF text layer  (vector PDFs, instant, free)
  Engine 2  — Gemini Vision       (primary AI — free tier, 1M context)
  Engine 3  — OpenAI Vision       (fallback AI — GPT-4o)
  Engine 4  — Tesseract OCR       (scanned PDFs with no text layer)

All pass results are merged and deduplicated by normalised tag number.

SOFT-CODED: all instrument categories live in INSTRUMENT_CATEGORIES dict
below — add/remove types without touching logic.
"""

import io
import os
import re
import json
import base64
import logging
from datetime import datetime

import time
from pdf2image import convert_from_bytes
from PIL import Image
import openpyxl

# Disable PIL's DecompressionBomb limit — large P&ID drawings (A0/A1 at 150 DPI)
# legitimately produce images above the 89 MP default threshold.
Image.MAX_IMAGE_PIXELS = None
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

logger = logging.getLogger(__name__)


# ────────────────────────────────────────────────────────────────────────────
# SOFT-CODED CONFIGURATION
# All instrument categories / type descriptions live here.
# ────────────────────────────────────────────────────────────────────────────
INSTRUMENT_CATEGORIES = {
    # ── FLOW ──────────────────────────────────────────────────────────────
    "FT":   {"name": "Flow Transmitter",                  "category": "Flow"},
    "FI":   {"name": "Flow Indicator",                    "category": "Flow"},
    "FIC":  {"name": "Flow Indicating Controller",        "category": "Flow"},
    "FIT":  {"name": "Flow Indicating Transmitter",       "category": "Flow"},
    "FCV":  {"name": "Flow Control Valve",                "category": "Flow"},
    "FE":   {"name": "Flow Element / Orifice",            "category": "Flow"},
    "FQI":  {"name": "Flow Quantity Indicator",           "category": "Flow"},
    "FG":   {"name": "Flow Glass / Sight Glass",          "category": "Flow"},
    "FO":   {"name": "Flow Orifice",                      "category": "Flow"},
    "FY":   {"name": "Flow Relay / Computer",             "category": "Flow"},
    # ── PRESSURE ──────────────────────────────────────────────────────────
    "PT":   {"name": "Pressure Transmitter",              "category": "Pressure"},
    "PI":   {"name": "Pressure Indicator",                "category": "Pressure"},
    "PIC":  {"name": "Pressure Indicating Controller",    "category": "Pressure"},
    "PIT":  {"name": "Pressure Indicating Transmitter",   "category": "Pressure"},
    "PS":   {"name": "Pressure Switch",                   "category": "Pressure"},
    "PSH":  {"name": "Pressure Switch High",              "category": "Pressure"},
    "PSL":  {"name": "Pressure Switch Low",               "category": "Pressure"},
    "PSHH": {"name": "Pressure Switch High-High",         "category": "Pressure"},
    "PSLL": {"name": "Pressure Switch Low-Low",           "category": "Pressure"},
    "PSAL": {"name": "Pressure Switch Alarm Low",         "category": "Pressure"},
    "PSAH": {"name": "Pressure Switch Alarm High",        "category": "Pressure"},
    "PSDL": {"name": "Pressure Switch Differential Low",  "category": "Pressure"},
    "PSDH": {"name": "Pressure Switch Differential High", "category": "Pressure"},
    "PCV":  {"name": "Pressure Control Valve",            "category": "Pressure"},
    "PG":   {"name": "Pressure Gauge",                    "category": "Pressure"},
    "PSV":  {"name": "Pressure Safety Valve",             "category": "Safety"},
    "PRV":  {"name": "Pressure Relief Valve",             "category": "Safety"},
    # ── TEMPERATURE ───────────────────────────────────────────────────────
    "TT":   {"name": "Temperature Transmitter",           "category": "Temperature"},
    "TI":   {"name": "Temperature Indicator",             "category": "Temperature"},
    "TIC":  {"name": "Temperature Indicating Controller", "category": "Temperature"},
    "TIT":  {"name": "Temperature Indicating Transmitter","category": "Temperature"},
    "TS":   {"name": "Temperature Switch",                "category": "Temperature"},
    "TSH":  {"name": "Temperature Switch High",           "category": "Temperature"},
    "TSL":  {"name": "Temperature Switch Low",            "category": "Temperature"},
    "TSHH": {"name": "Temperature Switch High-High",      "category": "Temperature"},
    "TSLL": {"name": "Temperature Switch Low-Low",        "category": "Temperature"},
    "TCV":  {"name": "Temperature Control Valve",         "category": "Temperature"},
    "TW":   {"name": "Thermowell",                        "category": "Temperature"},
    "TE":   {"name": "Temperature Element (Thermocouple)","category": "Temperature"},
    # ── LEVEL ─────────────────────────────────────────────────────────────
    "LT":   {"name": "Level Transmitter",                 "category": "Level"},
    "LI":   {"name": "Level Indicator",                   "category": "Level"},
    "LIC":  {"name": "Level Indicating Controller",       "category": "Level"},
    "LIT":  {"name": "Level Indicating Transmitter",      "category": "Level"},
    "LS":   {"name": "Level Switch",                      "category": "Level"},
    "LSH":  {"name": "Level Switch High",                 "category": "Level"},
    "LSL":  {"name": "Level Switch Low",                  "category": "Level"},
    "LSHH": {"name": "Level Switch High-High",            "category": "Level"},
    "LSLL": {"name": "Level Switch Low-Low",              "category": "Level"},
    "LSAL": {"name": "Level Switch Alarm Low",            "category": "Level"},
    "LSAH": {"name": "Level Switch Alarm High",           "category": "Level"},
    "LSDL": {"name": "Level Switch Differential Low",     "category": "Level"},
    "LSDH": {"name": "Level Switch Differential High",    "category": "Level"},
    "LG":   {"name": "Level Gauge",                       "category": "Level"},
    "LCV":  {"name": "Level Control Valve",               "category": "Level"},
    "LY":   {"name": "Level Relay / Computer",            "category": "Level"},
    # ── DIFFERENTIAL PRESSURE ─────────────────────────────────────────────
    "DPI":  {"name": "Differential Pressure Indicator",   "category": "Differential Pressure"},
    "DPIT": {"name": "DP Indicating Transmitter",         "category": "Differential Pressure"},
    "DPT":  {"name": "Differential Pressure Transmitter", "category": "Differential Pressure"},
    "DPAH": {"name": "DP Alarm High",                     "category": "Differential Pressure"},
    "DPAL": {"name": "DP Alarm Low",                      "category": "Differential Pressure"},
    "DPZY": {"name": "DP Position Transmitter",           "category": "Differential Pressure"},
    # ── ANALYSIS ──────────────────────────────────────────────────────────
    "AT":   {"name": "Analyzer Transmitter",              "category": "Analysis"},
    "AI":   {"name": "Analyzer Indicator",                "category": "Analysis"},
    "AIC":  {"name": "Analyzer Indicating Controller",    "category": "Analysis"},
    "AIT":  {"name": "Analyzer Indicating Transmitter",   "category": "Analysis"},
    # ── SHUTDOWN / ESD / CONTROL VALVES ───────────────────────────────────
    "SDV":  {"name": "Shutdown Valve",                    "category": "Shutdown & ESD"},
    "BDV":  {"name": "Blowdown Valve",                    "category": "Shutdown & ESD"},
    "XV":   {"name": "On/Off Valve (ESD)",                "category": "Shutdown & ESD"},
    "EV":   {"name": "Emergency Valve",                   "category": "Shutdown & ESD"},
    "HCV":  {"name": "Hand Control Valve",                "category": "Control Valves"},
    # ── MOTOR / SOLENOID OPERATED ─────────────────────────────────────────
    "MOV":  {"name": "Motor Operated Valve",              "category": "Motor & Solenoid"},
    "SOV":  {"name": "Solenoid Operated Valve",           "category": "Motor & Solenoid"},
    "AOV":  {"name": "Air Operated Valve",                "category": "Motor & Solenoid"},
    # ── POSITION / VALVE POSITION ─────────────────────────────────────────
    "ZI":   {"name": "Position Indicator",                "category": "Position"},
    "ZT":   {"name": "Position Transmitter",              "category": "Position"},
    "ZS":   {"name": "Position Switch",                   "category": "Position"},
    "ZSH":  {"name": "Position Switch High (Open)",       "category": "Position"},
    "ZSL":  {"name": "Position Switch Low (Closed)",      "category": "Position"},
    "ZCV":  {"name": "Position Control Valve",            "category": "Position"},
    "SVZY": {"name": "Solenoid Valve + Position TX",      "category": "Position"},
    "BVZY": {"name": "Ball Valve + Position TX",          "category": "Position"},
    # ── RESTRICTION / SPECIAL ─────────────────────────────────────────────
    "RO":   {"name": "Restriction Orifice",               "category": "Restriction"},
    "XPD":  {"name": "Special / Explosion Proof Device",  "category": "Special"},
    "XY":   {"name": "Relay / Computer (Special)",        "category": "Special"},
    "WI":   {"name": "Weight Indicator",                  "category": "Weight"},
    "WIT":  {"name": "Weight Indicating Transmitter",     "category": "Weight"},
    "SI":   {"name": "Speed Indicator",                   "category": "Speed"},
    "SIT":  {"name": "Speed Indicating Transmitter",      "category": "Speed"},
    "VI":   {"name": "Vibration Indicator",               "category": "Vibration"},
    "VIT":  {"name": "Vibration Indicating Transmitter",  "category": "Vibration"},
    "HI":   {"name": "Hand Indicator",                    "category": "Hand/Manual"},
    "HS":   {"name": "Hand Switch",                       "category": "Hand/Manual"},
}

# Column definitions for Excel output (default / generic)
EXCEL_COLUMNS = [
    {"key": "index_no",              "label": "Index No.",          "width": 10},
    {"key": "tag_number",            "label": "Tag Number",         "width": 18},
    {"key": "control_system_tag",    "label": "CS Tag",             "width": 18},
    {"key": "instrument_type",       "label": "Instrument Type",    "width": 35},
    {"key": "category",              "label": "Category",           "width": 22},
    {"key": "pid_no",                "label": "P&ID No.",           "width": 22},
    {"key": "service_description",   "label": "Service Description","width": 40},
    {"key": "line_number",           "label": "Line Number",        "width": 20},
    {"key": "equipment_number",      "label": "Equipment No.",      "width": 18},
    {"key": "loop_number",           "label": "Loop No.",           "width": 14},
    {"key": "fail_safe",             "label": "Fail Safe",          "width": 12},
    {"key": "signal_type",           "label": "Signal Type",        "width": 16},
    {"key": "set_point",             "label": "Set Point",          "width": 14},
    {"key": "drawing_number",        "label": "Drawing No.",        "width": 22},
    {"key": "revision",              "label": "Rev.",               "width": 8},
    {"key": "notes",                 "label": "Notes",              "width": 40},
]

# ──────────────────────────────────────────────────────────────────────────
# ADNOC Gas — 25-column schema (mirrors frontend `_ADNOC_GAS_TEMPLATE`)
# Each column is rendered through `accessor(inst)` so display logic stays
# soft-coded. Group header is a list of (label, span) tuples whose spans
# must total len(ADNOC_GAS_EXCEL_COLUMNS).
# ──────────────────────────────────────────────────────────────────────────
ADNOC_GAS_FIELD_ONLY_CODES = {
    "FE", "FG", "PG", "PSV", "PSE", "TE", "TG", "TW", "LG", "AE", "RO",
}


def _adnoc_gas_loop_no(inst):
    import re as _re
    tag = (inst.get("tag_number") or "").upper()
    m = _re.match(r"^[A-Z0-9]+-([A-Z]{1,5})-", tag)
    isa = m.group(1) if m else ""
    if isa in ADNOC_GAS_FIELD_ONLY_CODES:
        return "-"
    cs = (inst.get("control_system_tag") or "").strip()
    if cs and cs.upper() not in ("N/A", "NA", "-", "—", "NONE", "NULL"):
        return cs.upper()
    ln = (inst.get("loop_number") or "").strip()
    if ln and ln.upper() not in ("N/A", "NA", "-", "—", "NONE", "NULL"):
        return ln.upper()
    return "-"


def _adnoc_gas_line_no(inst):
    ln = (inst.get("line_number") or "").strip()
    if ln and ln.upper() not in ("N/A", "NA", "-", "—", "NONE", "NULL"):
        return ln
    eq = (inst.get("equipment_number") or "").strip()
    loc = (inst.get("location") or "").strip().lower()
    if eq and eq.upper() not in ("N/A", "NA", "-", "—") and ("vessel" in loc):
        return eq
    return "-"


def _v(inst, key, *, mono=False):
    """Soft-coded value getter: treat all empty markers as '-'."""
    raw = inst.get(key)
    if raw is None:
        return "-"
    s = str(raw).strip()
    if not s or s.upper() in ("N/A", "NA", "-", "—", "NONE", "NULL"):
        return "-"
    return s.upper() if mono else s


# ── Soft-coded "Line Number" pure column (ADNOC Gas only) ────────────────
# Manual convention puts the EQUIPMENT TAG in the Gas "Line No (Note-7)"
# column for vessel-mounted instruments — useful for traceability but it
# hides the actual pipeline line ID for those rows. This extra column
# always shows the raw `line_number` field (no equipment fallback) so
# users can see real line IDs at a glance. Edit label/width here.
_GAS_PURE_LINE_NO_LABEL = "LINE NO"
_GAS_PURE_LINE_NO_WIDTH = 26


def _adnoc_gas_pure_line_no(inst):
    """Pure line_number accessor for ADNOC Gas — no equipment fallback."""
    raw = (inst.get("line_number") or "")
    s = str(raw).strip()
    if not s or s.upper() in ("N/A", "NA", "-", "—", "NONE", "NULL"):
        return "-"
    return s.upper()


ADNOC_GAS_EXCEL_GROUP_HEADER = [
    ("",                              4),
    ("Calibration Range (Note-4)",    3),
    ("Alarm (Note-7)",                4),
    # Trailing block widened by 1 to host the pure "Line No." column.
    ("",                              15),
]

ADNOC_GAS_EXCEL_COLUMNS = [
    {"key": "tag_number",       "label": "Tag Number\n(Note-7)", "width": 18,
     "accessor": lambda i: _v(i, "tag_number", mono=True)},
    {"key": "loop_number",      "label": "Loop No.",             "width": 18,
     "accessor": _adnoc_gas_loop_no},
    {"key": "service",          "label": "Service",              "width": 40,
     "accessor": lambda i: _v(i, "service_description")},
    {"key": "instrument_type",  "label": "Instrument Type",      "width": 28,
     "accessor": lambda i: _v(i, "instrument_type")},
    # Calibration Range
    {"key": "cal_min",          "label": "Min",                  "width":  8,
     "accessor": lambda i: _v(i, "calibration_min")},
    {"key": "cal_max",          "label": "Max",                  "width":  8,
     "accessor": lambda i: _v(i, "calibration_max")},
    {"key": "cal_unit",         "label": "Unit",                 "width": 10,
     "accessor": lambda i: _v(i, "calibration_unit")},
    # Alarm
    {"key": "alarm_l",          "label": "L",                    "width": 7,
     "accessor": lambda i: _v(i, "alarm_l")},
    {"key": "alarm_ll",         "label": "LL",                   "width": 7,
     "accessor": lambda i: _v(i, "alarm_ll")},
    {"key": "alarm_h",          "label": "H",                    "width": 7,
     "accessor": lambda i: _v(i, "alarm_h")},
    {"key": "alarm_hh",         "label": "HH",                   "width": 7,
     "accessor": lambda i: _v(i, "alarm_hh")},
    # Remaining
    {"key": "location",         "label": "Location",             "width": 12,
     "accessor": lambda i: _v(i, "location")},
    {"key": "io_type",          "label": "I/O Type",             "width": 10,
     "accessor": lambda i: _v(i, "io_type") if (i.get("io_type") not in (None, "", "-")) else _v(i, "signal_type")},
    {"key": "system",           "label": "System",               "width": 10,
     "accessor": lambda i: _v(i, "system") if (i.get("system") not in (None, "", "-")) else ("DCS" if (i.get("signal_type") or "").strip() not in ("", "-", "N/A") else "-")},
    {"key": "pid_no",           "label": "PID",                  "width": 24,
     "accessor": lambda i: _v(i, "pid_no", mono=True)},
    # Pure line number (no equipment fallback) — ADNOC Gas only.
    {"key": "pure_line_no",     "label": _GAS_PURE_LINE_NO_LABEL, "width": _GAS_PURE_LINE_NO_WIDTH,
     "accessor": _adnoc_gas_pure_line_no},
    {"key": "line_number",      "label": "Line No\n(Note-7)",    "width": 26,
     "accessor": _adnoc_gas_line_no},
    {"key": "equipment_number", "label": "Equip No\n(Note-7)",   "width": 18,
     "accessor": lambda i: _v(i, "equipment_number", mono=True)},
    {"key": "purchase_order",   "label": "Purchase Order\n(Note-7)", "width": 18,
     "accessor": lambda i: _v(i, "purchase_order", mono=True)},
    {"key": "datasheet_no",     "label": "Datasheet No.\n(Note-7)",  "width": 22,
     "accessor": lambda i: _v(i, "datasheet_no", mono=True)},
    {"key": "manufacturer",     "label": "Manufacturer\n(Note-7)",   "width": 18,
     "accessor": lambda i: _v(i, "manufacturer")},
    {"key": "model_no",         "label": "Model No.\n(Note-7)",  "width": 16,
     "accessor": lambda i: _v(i, "model_no", mono=True)},
    {"key": "junction_box",     "label": "Junction Box\n(Note-7)", "width": 16,
     "accessor": lambda i: _v(i, "junction_box", mono=True)},
    {"key": "multi_cable",      "label": "Multi Cable\n(Note-7)",  "width": 14,
     "accessor": lambda i: _v(i, "multi_cable", mono=True)},
    {"key": "loop_dwg",         "label": "Loop dwg.\n(Note-7)",  "width": 16,
     "accessor": lambda i: _v(i, "loop_dwg", mono=True)},
    {"key": "remark",           "label": "Instrument Remark",    "width": 32,
     "accessor": lambda i: _v(i, "instrument_remark") if (i.get("instrument_remark") not in (None, "", "-")) else _v(i, "notes")},
]

# ──────────────────────────────────────────────────────────────────────────
# ADNOC Onshore — 18-column schema (mirrors frontend `_ADNOC_ONSHORE_TEMPLATE`)
# Soft-coded sibling of the ADNOC Gas registry. Lives in its own constants
# block so any column tweak here is guaranteed not to leak into ADNOC Gas
# or the generic default schema.
# ──────────────────────────────────────────────────────────────────────────
_ADNOC_ONSHORE_TYPE_LABELS = {
    "FT": "Flow Transmitter", "FE": "Flow Element (Orifice)",
    "FIT": "Flow Indicating Transmitter", "FI": "Flow Indicator",
    "FV": "Control Valve", "FCV": "Control Valve",
    "FZT": "Position Transmitter", "FZI": "Position Indicator",
    "FAL": "Flow Alarm Low",
    "LT": "Level Transmitter", "LIT": "Level Indicating Transmitter",
    "LI": "Level Indicator", "LG": "Level Gauge (Mag)",
    "LV": "Control Valve", "LCV": "Control Valve",
    "LIC": "Level Indicator Controller",
    "LSL": "Level Switch Low", "LSH": "Level Switch High",
    "LSLL": "Level Switch Low Low", "LSHH": "Level Switch High High",
    "LALL": "Level Alarm Low Low",
    "PG": "Pressure Gauge", "PT": "Pressure Transmitter",
    "PIT": "Pressure Indicating Transmitter", "PI": "Pressure Indicator",
    "PV": "Control Valve", "PCV": "Control Valve",
    "PIC": "Pressure Indicator Controller", "PSV": "Pressure Safety Valve",
    "PSH": "Pressure Switch High", "PSL": "Pressure Switch Low",
    "TG": "Temperature Gauge", "TT": "Temperature Transmitter",
    "TIT": "Temperature Indicating Transmitter",
    "TI": "Temperature Indicator", "TE": "Temperature Element",
    "TW": "Thermowell", "TIC": "Temperature Indicator Controller",
    "TV": "Control Valve", "HS": "Hand Switch",
    "VAH": "Vibration Alarm High", "VSH": "Vibration Switch High",
    "SDV": "Shutdown Valve", "BDV": "Blowdown Valve",
    "SOV": "Solenoid Valve", "MOV": "Motor Operated Valve",
}

# ── ADNOC Onshore tag-validation soft-codes ─────────────────────────────
# A canonical Onshore instrument tag is `<UNIT>-<ISA>-<SEQ>` where:
#   UNIT  : 2-4 digit unit/area number (e.g. "562")
#   ISA   : 2-5 letters drawn from the controlled vocabulary below
#   SEQ   : 3-4 digit loop number, optionally followed by a single
#           letter suffix (A/B/…) for redundant trains
# Anything that doesn't match — line numbers like `562-0407-VE-3"-11030-P`,
# equipment tags like `562-V-201`, drawing IDs like `50196-H5-562-PX-...` —
# must be rejected. The vocabulary lives here as the single source of
# truth; extending it is a one-line change.
#
# Canonical ADNOC Onshore instrument-tag shape (per manual datasheet):
#     <UNIT>-<ISA>-<LOOP>
#         UNIT : exactly 3 digits (e.g. "562")
#         ISA  : 2 or 3 capital letters from `_ADNOC_ONSHORE_VALID_ISA`
#         LOOP : exactly 4 digits, optional single-letter train suffix (A/B/…)
# Examples that must be KEPT  : "562-FZT-1501", "562-LG-2502", "562-PSV-8501A"
# Examples that must be DROPPED: "FE-1401-10" (legacy 3-segment form),
#                                "562-V-201"  (single-letter ISA / 3-digit loop),
#                                "562-FT-150" (loop too short),
#                                "562-0407-VE-3-11030-P" (line number),
#                                "50196-H5-562-PX-PID-00003" (drawing id)
_ADNOC_ONSHORE_TAG_RE = __import__('re').compile(
    r'^(?P<unit>\d{3})-(?P<isa>[A-Z]{2,3})-(?P<seq>\d{4}[A-Z]?)$'
)
# Soft default unit prefix (loaded from drawing context when available)
_ADNOC_ONSHORE_DEFAULT_UNIT = "562"

# Whitelisted ADNOC Onshore instrument ISA codes (ISA-5.1 + ADNOC HS/MOV
# additions). Anything outside this set is treated as equipment / line /
# document noise. Edit here to expand the accepted vocabulary.
_ADNOC_ONSHORE_VALID_ISA = {
    # Flow
    "FT", "FE", "FIT", "FI", "FQ", "FQI", "FV", "FCV", "FZT", "FZI",
    "FAL", "FAH", "FSL", "FSH", "FY", "FIC",
    # Level
    "LT", "LIT", "LI", "LG", "LV", "LCV", "LIC", "LSL", "LSH", "LSLL",
    "LSHH", "LALL", "LAHH", "LAH", "LAL", "LY",
    # Pressure
    "PT", "PIT", "PI", "PG", "PV", "PCV", "PIC", "PSV", "PRV", "PVSV",
    "PSH", "PSL", "PSHH", "PSLL", "PAH", "PAL", "PY", "PDT", "PDIT",
    "PDI", "DPT", "DPIT", "DPI",
    # Temperature
    "TT", "TIT", "TI", "TG", "TE", "TW", "TIC", "TV", "TCV", "TSH",
    "TSL", "TSHH", "TSLL", "TAH", "TAL", "TY",
    # Analyser / vibration / hand
    "AT", "AIT", "AI", "AE", "AAH", "AAL", "AY",
    "VAH", "VAL", "VSH", "VSL", "VT", "XT",
    "HS", "HV", "HIC", "HC",
    # Shutdown / blowdown / solenoids
    "SDV", "BDV", "SOV", "MOV", "ROV", "XV", "ESV", "SSV", "SSSV",
    "PSE", "RO", "ZT", "ZI", "ZSH", "ZSL",
}

# ── ADNOC Onshore Location classification (soft-coded) ──────────────────
# Maps ISA prefix → installation location used in the manual datasheet's
# "Location" column. Order matters only when multiple buckets could apply
# — first match wins. Edit these sets to retune the policy without
# touching `_apply_adnoc_onshore_style`.
_ADNOC_ONSHORE_LOCATION_LOCAL_PANEL = {
    # Local-mount visual indicators / gauges
    "PG", "LG", "TG", "TI", "PI", "LI", "FI", "FQI",
}
_ADNOC_ONSHORE_LOCATION_CONTROL_ROOM = {
    # DCS controllers / panel-mounted recorders
    "FIC", "LIC", "PIC", "TIC", "HIC", "AIC",
}
_ADNOC_ONSHORE_LOCATION_VESSEL_ISA = {
    # Direct equipment-mounted ISA codes (level/pressure/temperature
    # transmitters typically strapped to a vessel/drum/exchanger)
    "LT", "LIT", "LSL", "LSH", "LSLL", "LSHH", "LALL", "LAHH",
    "LAH", "LAL", "LY",
    "PT", "PIT", "PDT", "PDIT", "DPT", "DPIT",
    "TT", "TIT", "TE", "TW", "TY",
}
# Default fallback when no rule matches.
_ADNOC_ONSHORE_LOCATION_DEFAULT = "Field"


# ── ADNOC Onshore I/O TYPE / SYSTEM / IS-NIS classification (soft-coded) ──
# Mirrors the manual "Instrument Index" sheet columns. Edit the sets/maps
# below to retune the policy without touching `_apply_adnoc_onshore_style`.
#
# I/O TYPE — DCS terminology used in ADNOC Onshore datasheets:
#   AI / AO / DI / DO  (no -R suffix; Onshore datasheet keeps it terse)
#   ""                 — purely local devices (PG/TG/LG/PSV) get blank.
_ADNOC_ONSHORE_IO_TYPE_BY_ISA = {
    # Analog Inputs — transmitters & indicating transmitters
    "PT": "AI", "PIT": "AI", "PDT": "AI", "PDIT": "AI", "DPT": "AI", "DPIT": "AI",
    "TT": "AI", "TIT": "AI", "TE": "AI",
    "FT": "AI", "FIT": "AI", "FE": "AI", "FQ": "AI", "FQI": "AI",
    "LT": "AI", "LIT": "AI",
    "AT": "AI", "AIT": "AI", "AE": "AI",
    "VT": "AI", "XT": "AI", "ZT": "AI",
    # Controllers — DCS analog input feeding a control loop
    "FIC": "AI", "LIC": "AI", "PIC": "AI", "TIC": "AI", "HIC": "AI", "AIC": "AI",
    # Analog Outputs — control valves / positioners / I-P converters
    "FV": "AO", "FCV": "AO", "LV": "AO", "LCV": "AO",
    "PV": "AO", "PCV": "AO", "TV": "AO", "TCV": "AO",
    "FY": "AO", "PY": "AO", "TY": "AO", "LY": "AO", "AY": "AO", "HY": "AO",
    # Discrete Inputs — switches & position indicators
    "PSH": "DI", "PSL": "DI", "PSHH": "DI", "PSLL": "DI",
    "TSH": "DI", "TSL": "DI", "TSHH": "DI", "TSLL": "DI",
    "LSH": "DI", "LSL": "DI", "LSHH": "DI", "LSLL": "DI",
    "FSH": "DI", "FSL": "DI",
    "VSH": "DI", "VSL": "DI",
    "ZSH": "DI", "ZSL": "DI", "ZI": "DI",
    "FZT": "DI", "FZI": "DI",
    "AAH": "DI", "AAL": "DI", "PAH": "DI", "PAL": "DI",
    "TAH": "DI", "TAL": "DI", "VAH": "DI", "VAL": "DI",
    "LAH": "DI", "LAL": "DI", "LALL": "DI", "LAHH": "DI",
    "FAL": "DI", "FAH": "DI",
    "HS": "DI",
    # Discrete Outputs — shutdown / blowdown / on-off valves & solenoids
    "SDV": "DO", "BDV": "DO", "MOV": "DO", "ROV": "DO", "XV": "DO",
    "ESV": "DO", "SSV": "DO", "SSSV": "DO",
    "SOV": "DO", "HV": "DO",
}
# Local-only devices: no DCS I/O — column stays blank.
_ADNOC_ONSHORE_IO_LOCAL_ISA = {
    "PG", "TG", "LG", "PI", "TI", "LI", "FI", "FQI", "FG",
    "PSV", "PRV", "PVSV", "PSE", "RO", "RD",
}

# SYSTEM — owning control system. Soft-coded per ISA category.
#   DCS  — process monitoring & control (transmitters/controllers/CVs)
#   ESD  — emergency shutdown (SDV/BDV/PSHH/LSLL/TSHH switches feeding ESD)
#   F&G  — fire & gas (analyser detectors AT/AIT/AE/AAH/AAL)
#   Local — PG/TG/LG/PSV — purely local devices
_ADNOC_ONSHORE_SYSTEM_ESD_ISA = {
    "SDV", "BDV", "ESV", "SSV", "SSSV",
    "PSHH", "PSLL", "TSHH", "TSLL", "LSHH", "LSLL", "LALL", "LAHH",
    "VSH", "VSL", "VAH", "VAL",
}
_ADNOC_ONSHORE_SYSTEM_FNG_ISA = {
    "AT", "AIT", "AE", "AI", "AAH", "AAL", "AY", "AIC",
}
_ADNOC_ONSHORE_SYSTEM_LOCAL_ISA = _ADNOC_ONSHORE_IO_LOCAL_ISA  # share the local set
_ADNOC_ONSHORE_SYSTEM_DEFAULT = "DCS"

# IS / NIS — Intrinsically Safe vs Non-Intrinsically Safe (per ADNOC
# Onshore practice: low-power 4-20 mA field instruments are IS by default;
# high-power solenoids/MOV actuators are NIS; local mechanical devices
# without electrical signal are blank).
_ADNOC_ONSHORE_IS_NIS_NIS_ISA = {
    # Solenoids and motorised on/off valves carry mains/24 V DC discrete
    # outputs above the IS power-limit envelope.
    "MOV", "ROV", "SOV", "HV",
}
_ADNOC_ONSHORE_IS_NIS_BLANK_ISA = {
    # Purely local mechanical devices — no electrical interface.
    "PG", "TG", "LG", "PSV", "PRV", "PVSV", "PSE", "RO", "RD",
}
_ADNOC_ONSHORE_IS_NIS_DEFAULT = "IS"


def _adnoc_onshore_resolve_isa(inst):
    """Pull the ISA token from a canonical UNIT-ISA-LOOP tag, else ''."""
    tag = (inst.get("tag_number") or "").upper()
    m = _ADNOC_ONSHORE_TAG_RE.match(tag)
    if m:
        return m.group("isa")
    return _onshore_isa(tag)


def _adnoc_onshore_derive_io_type(inst):
    """Soft-coded Onshore I/O Type classifier.

    Returns "" for local-only instruments so the column stays blank
    (matches the manual "Instrument Index" sheet convention).
    """
    isa = _adnoc_onshore_resolve_isa(inst)
    if not isa:
        return ""
    if isa in _ADNOC_ONSHORE_IO_LOCAL_ISA:
        return ""
    return _ADNOC_ONSHORE_IO_TYPE_BY_ISA.get(isa, "")


def _adnoc_onshore_derive_system(inst):
    """Soft-coded Onshore System classifier (DCS / ESD / F&G / Local)."""
    isa = _adnoc_onshore_resolve_isa(inst)
    if not isa:
        return ""
    if isa in _ADNOC_ONSHORE_SYSTEM_LOCAL_ISA:
        return "Local"
    if isa in _ADNOC_ONSHORE_SYSTEM_ESD_ISA:
        return "ESD"
    if isa in _ADNOC_ONSHORE_SYSTEM_FNG_ISA:
        return "F&G"
    return _ADNOC_ONSHORE_SYSTEM_DEFAULT


def _adnoc_onshore_derive_is_nis(inst):
    """Soft-coded Onshore IS/NIS classifier."""
    isa = _adnoc_onshore_resolve_isa(inst)
    if not isa:
        return ""
    if isa in _ADNOC_ONSHORE_IS_NIS_BLANK_ISA:
        return ""
    if isa in _ADNOC_ONSHORE_IS_NIS_NIS_ISA:
        return "NIS"
    return _ADNOC_ONSHORE_IS_NIS_DEFAULT


def _adnoc_onshore_derive_location(inst):
    """Soft-coded Onshore Location classifier.

    Resolution order (first match wins):
      1. Existing non-empty `location` value is kept verbatim.
      2. Controllers with DCS/panel ISA → ``Control Room``.
      3. Local-panel indicators (PG/LG/TG/PI/LI/TI…) → ``Local Panel``.
      4. Vessel-mounted transmitters with an `equipment_number` → ``Vessel``.
      5. Anything else → ``Field`` (`_ADNOC_ONSHORE_LOCATION_DEFAULT`).
    """
    cur = (inst.get("location") or "").strip()
    if cur and cur not in ("-", "—", "N/A", "n/a", "NA"):
        return cur
    tag = (inst.get("tag_number") or "").upper()
    # Extract ISA from canonical `<ISA>-<LOOP>-<PAGE>`.
    isa = ""
    m = _ADNOC_ONSHORE_TAG_RE.match(tag)
    if m:
        isa = m.group("isa")
    if isa in _ADNOC_ONSHORE_LOCATION_CONTROL_ROOM:
        return "Control Room"
    if isa in _ADNOC_ONSHORE_LOCATION_LOCAL_PANEL:
        return "Local Panel"
    eq = (inst.get("equipment_number") or "").strip()
    if isa in _ADNOC_ONSHORE_LOCATION_VESSEL_ISA and eq and eq not in ("-", "—", "N/A"):
        return "Vessel"
    return _ADNOC_ONSHORE_LOCATION_DEFAULT


def _adnoc_onshore_canonicalise_tag(raw, default_unit=None):
    """Normalise *raw* into ADNOC Onshore canonical form ``UNIT-ISA-LOOP``.

    Returns ``(canonical_tag, isa)`` on success, ``(None, None)`` if
    *raw* cannot be reduced to a valid instrument tag in the manual
    datasheet shape (e.g. ``562-FZT-1501``).

    Resolution rules:
      * Strict full match against `_ADNOC_ONSHORE_TAG_RE` (3-segment).
      * Bare ``ISA-LOOP`` (e.g. ``FZT-1501``) is promoted to the
        canonical form using *default_unit* when supplied (else
        `_ADNOC_ONSHORE_DEFAULT_UNIT` is used as a soft fallback).
      * Legacy ``ISA-LOOP-PAGE`` (e.g. ``FZT-1501-04``) is reduced by
        dropping the trailing page segment.

    Args:
        raw: any value (string-like) extracted from the P&ID.
        default_unit: optional unit prefix (e.g. ``"562"``).
    """
    if raw is None:
        return None, None
    txt = str(raw).strip().upper()
    if not txt or txt in ("-", "N/A", "NA", "NONE", "NULL"):
        return None, None
    # Collapse internal whitespace and normalise dashes.
    re_mod = __import__('re')
    txt = re_mod.sub(r"\s*[-–—]\s*", "-", txt)
    txt = re_mod.sub(r"\s+", "-", txt)

    # Pick the unit prefix to use when input is missing one.
    unit_prefix = (str(default_unit).strip() if default_unit else "") or _ADNOC_ONSHORE_DEFAULT_UNIT

    # 1) Strict canonical match: UNIT-ISA-LOOP
    m = _ADNOC_ONSHORE_TAG_RE.match(txt)
    if m:
        isa = m.group("isa")
        if isa not in _ADNOC_ONSHORE_VALID_ISA:
            return None, None
        return f"{m.group('unit')}-{isa}-{m.group('seq')}", isa

    # 2) Bare ISA-LOOP → promote with unit_prefix
    bare_re = re_mod.compile(r'^([A-Z]{2,3})-(\d{4}[A-Z]?)$')
    bm = bare_re.match(txt)
    if bm and unit_prefix:
        isa = bm.group(1)
        if isa not in _ADNOC_ONSHORE_VALID_ISA:
            return None, None
        return f"{unit_prefix}-{isa}-{bm.group(2)}", isa

    # 3) Legacy 3-segment ISA-LOOP-PAGE → drop trailing page
    legacy_re = re_mod.compile(r'^([A-Z]{2,3})-(\d{4}[A-Z]?)-\d{1,3}$')
    lm = legacy_re.match(txt)
    if lm and unit_prefix:
        isa = lm.group(1)
        if isa not in _ADNOC_ONSHORE_VALID_ISA:
            return None, None
        return f"{unit_prefix}-{isa}-{lm.group(2)}", isa

    return None, None


def _onshore_isa(tag):
    import re as _re
    m = _re.match(r"^\d+\s*-\s*([A-Z]{1,5})\s*-",
                  (tag or "").upper())
    return m.group(1) if m else ""


def _adnoc_onshore_instrument_type(inst):
    cur = (inst.get("instrument_type") or "").strip()
    isa = _onshore_isa(inst.get("tag_number"))
    verbose = _ADNOC_ONSHORE_TYPE_LABELS.get(isa)
    if not cur:
        return verbose or "-"
    if verbose and cur.upper() == isa:
        return verbose
    return cur


def _adnoc_onshore_eq_or_line(inst):
    ln = (inst.get("line_number") or "").strip()
    if ln and ln.upper() not in ("N/A", "NA", "-", "—", "NONE", "NULL"):
        return ln
    eq = (inst.get("equipment_number") or "").strip()
    if eq and eq.upper() not in ("N/A", "NA", "-", "—", "NONE", "NULL"):
        return eq
    return "-"


ADNOC_ONSHORE_EXCEL_GROUP_HEADER = [
    # Leading block widened by 1 to host the new "LINE NO" column.
    ("",                                 11),  # Tag…Device Status (+ LINE NO)
    ("Inst range (Refer Gen Note 5)",    3),   # Min/Max/Unit
    ("Calibration range",                3),   # Min/Max/Unit
    ("",                                 1),   # Remarks
]

ADNOC_ONSHORE_EXCEL_COLUMNS = [
    {"key": "tag_number",        "label": "Tag No.",        "width": 18,
     "accessor": lambda i: _v(i, "tag_number", mono=True)},
    {"key": "instrument_type",   "label": "Instrument Type","width": 24,
     "accessor": _adnoc_onshore_instrument_type},
    {"key": "service",           "label": "Service Description", "width": 40,
     "accessor": lambda i: _v(i, "service_description")},
    {"key": "location",          "label": "Location",       "width": 12,
     "accessor": lambda i: _v(i, "location")},
    {"key": "equipment_or_line", "label": "Equipment / Line No.", "width": 26,
     "accessor": _adnoc_onshore_eq_or_line},
    # Pure line-number column — reuses the Gas accessor (no equipment fallback).
    {"key": "pure_line_no",      "label": _GAS_PURE_LINE_NO_LABEL, "width": _GAS_PURE_LINE_NO_WIDTH,
     "accessor": _adnoc_gas_pure_line_no},
    {"key": "pid_no",            "label": "P&ID No.",       "width": 24,
     "accessor": lambda i: _v(i, "pid_no", mono=True)},
    {"key": "io_type",           "label": "I/O Type",       "width":  8,
     "accessor": lambda i: _v(i, "io_type")},
    {"key": "is_nis",            "label": "IS/NIS",         "width":  8,
     "accessor": lambda i: _v(i, "is_nis")},
    {"key": "system",            "label": "System",         "width": 10,
     "accessor": lambda i: _v(i, "system")},
    {"key": "device_status",     "label": "Device Status",  "width": 10,
     "accessor": lambda i: _v(i, "device_status") if (i.get("device_status") not in (None, "", "-")) else "New"},
    # Inst range
    {"key": "inst_range_min",    "label": "Min",  "width": 8,
     "accessor": lambda i: _v(i, "inst_range_min")},
    {"key": "inst_range_max",    "label": "Max",  "width": 8,
     "accessor": lambda i: _v(i, "inst_range_max")},
    {"key": "inst_range_unit",   "label": "Unit", "width": 8,
     "accessor": lambda i: _v(i, "inst_range_unit")},
    # Calibration range (re-uses same backend keys as ADNOC Gas)
    {"key": "calibration_min",   "label": "Min",  "width": 8,
     "accessor": lambda i: _v(i, "calibration_min")},
    {"key": "calibration_max",   "label": "Max",  "width": 8,
     "accessor": lambda i: _v(i, "calibration_max")},
    {"key": "calibration_unit",  "label": "Unit", "width": 8,
     "accessor": lambda i: _v(i, "calibration_unit")},
    {"key": "remarks",           "label": "Remarks", "width": 32,
     "accessor": lambda i: _v(i, "instrument_remark") if (i.get("instrument_remark") not in (None, "", "-")) else _v(i, "notes")},
]

# Registry mapping project_category → (group_header, columns) tuple.
# Add new client schemas here — `generate_excel` will pick them up
# automatically. None for group_header means no merged-header strip.
EXCEL_SCHEMAS = {
    "adnoc_gas":     (ADNOC_GAS_EXCEL_GROUP_HEADER,     ADNOC_GAS_EXCEL_COLUMNS),
    "adnoc_onshore": (ADNOC_ONSHORE_EXCEL_GROUP_HEADER, ADNOC_ONSHORE_EXCEL_COLUMNS),
}

# Category colour coding for Excel rows
CATEGORY_COLOURS = {
    "Flow":               "DDEEFF",
    "Pressure":           "FFE4CC",
    "Temperature":        "FFE4E4",
    "Level":              "E4F4E4",
    "Differential Pressure": "FFF9CC",
    "Analysis":           "E8E4FF",
    "Safety":             "FFCCCC",
    "Shutdown & ESD":     "FFD9D9",
    "Control Valves":     "CCFFEE",
    "Motor & Solenoid":   "E0E0FF",
    "Position":           "FFFACC",
    "Restriction":        "DDEEDD",
    "Special":            "F0F0F0",
    "Weight":             "E8F4FF",
    "Speed":              "F4E8FF",
    "Vibration":          "FFE8F4",
    "Hand/Manual":        "F0FFE8",
}

# ────────────────────────────────────────────────────────────────────────────
# CATEGORY-AWARE TEMPLATE REGISTRY (soft-coded)
# Each entry adds extra fields that the AI prompt should try to extract for
# projects of that category, plus optional default values used to fill gaps
# when the AI doesn't return data.  Add a new client by appending here — no
# other code changes needed.
#
# `extra_fields` schema:
#   key         — the dict key written into each instrument record
#   label       — short human label used in the AI prompt
#   description — what the field captures (becomes prompt instruction)
#   default     — string used if the field is absent and no derive() returns
#   derive(inst)— optional fn(inst) -> str, computed AFTER AI extraction
# ────────────────────────────────────────────────────────────────────────────

def _derive_io_type(inst):
    """Best-effort I/O type derivation from existing fields."""
    sig = (inst.get("signal_type") or "").upper()
    cat = (inst.get("category") or "").lower()
    tag = (inst.get("tag_number") or "").upper()
    if not sig or sig in ("N/A", "-", ""):
        return ""
    if "DISCRETE" in sig or "0/1" in sig:
        # SDV / MOV / shutdown valves take DO; transmitters reading binary → DI
        if any(p in tag for p in ("SDV", "MOV", "BDV", "FCV", "CV", "XV")):
            return "DO-R"  # discrete output (with response/feedback)
        return "DI"
    if "4-20" in sig or "HART" in sig or "ANALOG" in sig:
        # Control valves / actuators take AO, transmitters AI
        if any(p in tag for p in ("FV", "PV", "TV", "LV", "CV", "FCV", "PCV", "TCV", "LCV")):
            return "AO-R"
        return "AI" if "control" not in cat else "AI-R"
    return ""

def _derive_system(inst):
    """If a control_system_tag is present, instrument is on DCS by default."""
    cs = (inst.get("control_system_tag") or "").strip().upper()
    sig = (inst.get("signal_type") or "").upper()
    if cs and cs not in ("N/A", "-", ""):
        return "DCS"
    if any(k in sig for k in ("4-20", "HART", "FIELDBUS")):
        return "DCS"
    return ""

def _derive_location(inst):
    """Tagged equipment-mounted instruments default to 'Vessel', else 'Field'."""
    eq = (inst.get("equipment_number") or "").strip()
    tag = (inst.get("tag_number") or "").upper()
    # Level instruments on a vessel/drum are typically vessel-mounted
    if eq and eq not in ("-", "N/A", ""):
        if tag.startswith(("LT", "LIT", "LG", "LSL", "LSH", "LAH", "LAL", "PT", "PIT", "PG")):
            return "Vessel"
    return "Field"


# Common ADNOC Gas extra fields (matching the manual "Manual Inst Index" sheet)
_ADNOC_GAS_EXTRA_FIELDS = [
    {"key": "calibration_min",  "label": "Cal Min",   "description": "Calibration range minimum value (as written)",                       "default": "-"},
    {"key": "calibration_max",  "label": "Cal Max",   "description": "Calibration range maximum value",                                    "default": "-"},
    {"key": "calibration_unit", "label": "Cal Unit",  "description": "Calibration range engineering unit (e.g. barg, °C, m³/h)",          "default": "-"},
    {"key": "alarm_l",          "label": "Alarm L",   "description": "Low alarm setpoint shown on drawing or instrument list",            "default": "-"},
    {"key": "alarm_ll",         "label": "Alarm LL",  "description": "Low-low (trip) alarm setpoint",                                     "default": "-"},
    {"key": "alarm_h",          "label": "Alarm H",   "description": "High alarm setpoint",                                               "default": "-"},
    {"key": "alarm_hh",         "label": "Alarm HH",  "description": "High-high (trip) alarm setpoint",                                   "default": "-"},
    {"key": "location",         "label": "Location",  "description": "Where the instrument is installed: 'Field', 'Vessel', 'Local Panel', 'Control Room'",
     "default": "-",  "derive": _derive_location},
    {"key": "io_type",          "label": "I/O Type",  "description": "DCS I/O type: AI / AO / DI / DO, suffix '-R' for redundant. e.g. 'AI', 'AO-R'",
     "default": "-",  "derive": _derive_io_type},
    {"key": "system",           "label": "System",    "description": "Owning control system: 'DCS', 'ESD', 'F&G', 'PLC', 'Local'",
     "default": "-",  "derive": _derive_system},
    {"key": "purchase_order",   "label": "PO",        "description": "Purchase order number (rarely on P&ID — leave blank if not shown)", "default": "-"},
    {"key": "datasheet_no",     "label": "Datasheet", "description": "Instrument datasheet document number",                              "default": "-"},
    {"key": "manufacturer",     "label": "Mfr",       "description": "Manufacturer / vendor name",                                        "default": "-"},
    {"key": "model_no",         "label": "Model",     "description": "Model number",                                                      "default": "-"},
    {"key": "junction_box",     "label": "JB",        "description": "Junction box tag",                                                  "default": "-"},
    {"key": "multi_cable",      "label": "Multi Cable","description": "Multi-cable identifier",                                           "default": "-"},
    {"key": "loop_dwg",         "label": "Loop Dwg",  "description": "Loop drawing reference",                                            "default": "-"},
    {"key": "instrument_remark","label": "Remark",    "description": "Free-text remark — preserve any markings near the bubble",         "default": "-"},
]


# ────────────────────────────────────────────────────────────────────────────
# ADNOC LEGEND DICTIONARIES (soft-coded from "PIPING SYMBOLS & LEGENDS" and
# "INSTRUMENT SYMBOLS AND LEGENDS" sheets — drawings 50196-500-00-30-101..108).
# These are surfaced to the AI in the prompt so it can recognise tags and
# line numbers in the drawing convention used by ADNOC Onshore (Habshan-5)
# and other ADNOC Gas projects.
# ────────────────────────────────────────────────────────────────────────────

# Instrument letter codes seen in the ADNOC INSTRUMENT SYMBOLS & LEGENDS sheets.
# Used as TYPE token in <UNIT>-<TYPE>-<LOOP> tag pattern.
_ADNOC_INSTRUMENT_TYPES = [
    # Flow
    "FE", "FT", "FQI", "FG", "FV", "FCV", "FIC", "FI", "FY",
    # Level
    "LG", "LT", "LIT", "LV", "LCV", "LSH", "LSL", "LSHH", "LSLL", "LAH", "LAL",
    # Pressure
    "PG", "PT", "PIT", "PV", "PCV", "PSV", "PSE", "PAL", "PAH", "PSH", "PSL", "PSHH", "PSLL",
    # Temperature
    "TE", "TT", "TIT", "TV", "TW", "TI", "TIC", "TY",
    # Analyser / misc
    "AT", "AIT", "AY", "AI",
    # Valves & actuators (ESD/BMS/MOV families per legend sheet 4–6)
    "XV", "XY", "XSV", "XDV", "XDY",
    "KV", "KY", "KYO", "KYC",
    "MOV", "MZSO", "MZSC", "MZLO", "MZLC", "MHSO", "MHSC", "MXL",
    "XHSO", "XHSC", "XHS",
    "XZA", "XZI", "XZL", "XZLO", "XZLC", "XZSO", "XZSC", "XZT",
    "KZLO", "KZLC", "KZSO", "KZSC",
    "DVC", "SDV", "BDV", "VSH", "VSL", "ZSH", "ZSL", "ZSC", "ZSO",
]

# Service designations from LINE NUMBERING legend (sheet 2). Used to recognise
# line numbers like 000-0001-HCX-6"-... where the 3rd token is a service code.
_ADNOC_SERVICE_DESIGNATIONS = [
    "AC", "AD", "AG", "AGS", "AGX", "AV", "BA", "BD", "BFW", "BG", "BWS",
    "CAC", "CAD", "CBA", "CCD", "CD", "CDC", "CG", "CH", "CL", "CW", "CWR", "CWS",
    "DF", "DMW", "FAG", "FD", "FGL", "FGM", "FGX", "FL", "FLC", "FW",
    "GD", "GL", "GLX", "HC", "HCX", "HR", "IA", "IW",
    "LN", "LO", "LS", "LSM", "MOH", "N2",
    "OD", "PA", "PW", "RF", "RG", "RGX", "RSM", "RSX", "RW",
    "SAC", "SAD", "SC", "SCH", "SCL", "SD", "SDM", "SE", "SL", "SLH", "SM", "SO", "SSM",
    "STH", "STL", "STS", "SW",
    "TAG", "TW", "UW", "VE", "VT", "WD", "WDX", "WF", "WFX", "WS",
]

# Common abbreviations used in ADNOC P&IDs (subset of the full ABBREVIATIONS
# table — only those the AI is likely to need to recognise).
_ADNOC_PID_ABBREVIATIONS = {
    "FC": "Fail Closed", "FO": "Fail Open", "FL": "Fail Locked", "FI": "Fail Indeterminate",
    "DBB": "Double Block and Bleed", "DCS": "Distributed Control System",
    "ESD": "Emergency Shut Down System", "F&G": "Fire & Gas System",
    "BMS": "Burner Management System", "MCC": "Motor Control Centre",
    "PLC": "Programmable Logic Controller", "SDS": "Shutdown System",
    "MOV": "Motor Operated Valve", "PSV": "Pressure Safety Valve",
    "GWR": "Guided Wave Radar", "RTU": "Remote Terminal Unit",
    "IPCS": "Integrated Protection Control System",
    "MMS": "Machine Monitoring System", "MVC": "Multi-Variable Controller",
    "SCADA": "Supervisory Control and Data Acquisition",
    "SMC": "Supervisory Monitoring Control System",
    "TS": "Twin Seal", "TSO": "Tight Shut Off",
    "VEWFD": "Very Early Warning Fire Detection", "VT": "Vessel Trim",
    "WHC": "Wellhead Choke Valve",
}


def _build_adnoc_legend_block():
    """Compact legend reference inserted into the AI prompt for ADNOC projects."""
    types_csv = ", ".join(_ADNOC_INSTRUMENT_TYPES)
    svc_csv   = ", ".join(_ADNOC_SERVICE_DESIGNATIONS)
    abbr_csv  = ", ".join(f"{k}={v}" for k, v in _ADNOC_PID_ABBREVIATIONS.items())
    return (
        "ADNOC LEGEND REFERENCE (for tag recognition):\n"
        f"- Instrument letter codes (TYPE token): {types_csv}\n"
        f"- Service designations (line-number 3rd token): {svc_csv}\n"
        f"- Abbreviations: {abbr_csv}\n"
        "- Line number convention: <size>-<seq>-<service>-<class>-<insulation>, "
        "e.g. '6\"-0001-HCX-61055-N' (size, sequence, service designation, "
        "piping class, insulation purpose).\n"
        "- Equipment numbering uses '<UNIT>-<TYPE>-<NN>' e.g. '562-V-101', "
        "'803-E-XX1', '562-P-2401A'.\n"
    )


# ─────────────────────────────────────────────────────────────────────────
# ADNOC GAS — equipment-tag inference (soft-coded). When the AI doesn't
# return `equipment_number`, the post-processor scans the PDF text for the
# drawing's primary equipment tag using these patterns + the type-letter →
# noun map below to render group headers like "LP STEAM GENERATOR (803-E-XX1)".
# Add new equipment-type letters here as needed.
# ─────────────────────────────────────────────────────────────────────────
_ADNOC_GAS_EQUIPMENT_TAG_RE = re.compile(
    r"\b(\d{3})-([EVDTPKCRFSAH])-([A-Z0-9X]{2,5})\b"
)
_ADNOC_GAS_EQUIPMENT_TYPE_NOUN = {
    "E": "Exchanger / Generator",
    "V": "Vessel",
    "D": "Drum",
    "T": "Tower",
    "P": "Pump",
    "K": "Compressor",
    "C": "Column",
    "R": "Reactor",
    "F": "Filter",
    "S": "Separator",
    "A": "Air Cooler",
    "H": "Heater",
}
# Equipment-noun keywords commonly found in ADNOC drawing titles. Used to
# pull a clean description out of `drawing_title` or page text near the tag.
_ADNOC_GAS_EQUIPMENT_NOUN_RE = re.compile(
    r"\b(?:LP\s+STEAM\s+GENERATOR|HP\s+STEAM\s+GENERATOR|STEAM\s+GENERATOR|"
    r"HEAT\s+EXCHANGER|REBOILER|CONDENSER|COOLER|HEATER|FURNACE|REACTOR|"
    r"COMPRESSOR|EXPANDER|TURBINE|PUMP|SEPARATOR|KO\s+DRUM|KNOCKOUT\s+DRUM|"
    r"FLASH\s+DRUM|REFLUX\s+DRUM|SUCTION\s+DRUM|STORAGE\s+TANK|DRUM|VESSEL|"
    r"ABSORBER|STRIPPER|CONTACTOR|REGENERATOR|SCRUBBER|FILTER|COALESCER|"
    r"COLUMN|TOWER|FRACTIONATOR|DISTILLATION\s+COLUMN)\b",
    re.IGNORECASE,
)


# ─────────────────────────────────────────────────────────────────────────
# ADNOC GAS — Loop No. canonical format.
#
# The "Loop No." column on the Manual Inst Index lists the controller (or
# indicator) tag of the loop the device belongs to, in the form
# `{UNIT}-{CTRL_ISA}-{LOOP_SEQ}` e.g. `803-FC-XXXX`, `803-PI-XXXX`.
#
# `_ADNOC_LOOP_CTRL_MAP` maps every device ISA code to the ISA code that
# heads its loop. Soft-coded — extend as new device codes appear in the
# legend. Field-only devices (FE, PG, TG, …) are NOT in this map; they're
# still rendered as '-' via `ADNOC_GAS_FIELD_ONLY_CODES`.
# ─────────────────────────────────────────────────────────────────────────
_ADNOC_LOOP_CTRL_MAP = {
    # Flow loops — controlled by FC / FIC, indicators stay FI
    "FT": "FC", "FIT": "FC", "FE": "FC", "FV": "FC", "FCV": "FC",
    "FIC": "FC", "FC": "FC", "FY": "FC", "FQI": "FQI",
    "FI": "FI", "FG": "FI",
    # Pressure loops
    "PT": "PIC", "PIT": "PIC", "PV": "PIC", "PCV": "PIC",
    "PIC": "PIC", "PC": "PIC", "PY": "PIC",
    "PI": "PI", "PG": "PI",
    # Temperature loops
    "TT": "TIC", "TIT": "TIC", "TE": "TIC", "TV": "TIC", "TCV": "TIC",
    "TIC": "TIC", "TC": "TIC", "TY": "TIC",
    "TI": "TI", "TG": "TI", "TW": "TI",
    # Level loops
    "LT": "LIC", "LIT": "LIC", "LV": "LIC", "LCV": "LIC",
    "LIC": "LIC", "LC": "LIC", "LY": "LIC",
    "LI": "LI", "LG": "LI", "LSL": "LIC", "LSH": "LIC", "LSLL": "LIC", "LSHH": "LIC",
    # Analytical
    "AT": "AIC", "AIT": "AIC", "AE": "AIC", "AY": "AIC",
    "AIC": "AIC", "AI": "AI",
    # Speed / vibration / position
    "ST": "SIC", "SI": "SI",
    "VT": "VIC", "VI": "VI",
    "ZT": "ZIC", "ZI": "ZI",
}

# Soft-coded loop-sequence placeholder (mirrors `_ADNOC_TAG_FORMAT`).
_ADNOC_LOOP_DEFAULT_SEQ = "XXXX"
# Strict regex the loop sequence must match — anything else is replaced with
# the placeholder. ADNOC convention: 3 or 4 digits + optional A/B suffix.
_ADNOC_LOOP_SEQ_RE = re.compile(r"^\d{3,4}[A-Z]?$")
# Recognise an already-canonical loop tag like "803-FC-1234A".
_ADNOC_LOOP_CANONICAL_RE = re.compile(
    r"^(\d{3})-([A-Z]{1,5})-([A-Z0-9]{2,5}[A-Z]?)$"
)


def _adnoc_normalize_loop_no(tag_number, current_loop, default_unit=""):
    """
    Return a canonical ADNOC Loop No. ``{UNIT}-{CTRL_ISA}-{LOOP_SEQ}``
    derived from the instrument's tag and any loop value the AI emitted.

    Rules (all soft-coded above):
      • Unit = unit prefix of the instrument tag (or `default_unit`).
      • CTRL ISA = `_ADNOC_LOOP_CTRL_MAP[device_isa]`. If the device ISA is
        not mapped, fall back to the device ISA itself.
      • Loop sequence = digits found inside `current_loop`, validated
        against `_ADNOC_LOOP_SEQ_RE`; otherwise `_ADNOC_LOOP_DEFAULT_SEQ`.

    If the inputs cannot produce a unit + ISA, returns the input unchanged.
    """
    tag = (tag_number or "").strip().upper()
    cur = (current_loop or "").strip().upper()

    # If `current_loop` is already canonical, keep as-is (idempotent).
    if _ADNOC_LOOP_CANONICAL_RE.match(cur):
        return cur

    # ── Extract unit + device ISA from the instrument tag ──
    m = re.match(r"^(\d{3,4})-([A-Z]{1,5})-", tag)
    if m:
        unit = m.group(1)[:3]
        device_isa = m.group(2)
    else:
        unit = (default_unit or "").strip()[:3]
        # Try to recover device ISA from a less canonical tag (e.g. "FT-803")
        m2 = re.match(r"^([A-Z]{1,5})-", tag)
        device_isa = m2.group(1) if m2 else ""

    if not unit or not device_isa:
        return cur or "-"

    ctrl_isa = _ADNOC_LOOP_CTRL_MAP.get(device_isa, device_isa)

    # ── Extract loop sequence from `current_loop` ──
    seq = ""
    if cur and cur not in ("-", "N/A", "NA", "NONE", "NULL"):
        # Pull the right-most numeric+suffix token (handles inputs like
        # "FC-1234", "803-1234A", "1234", "FIC1234").
        sm = re.search(r"(\d{3,4}[A-Z]?)\s*$", cur)
        if sm:
            cand = sm.group(1)
            if _ADNOC_LOOP_SEQ_RE.match(cand):
                seq = cand
    if not seq:
        seq = _ADNOC_LOOP_DEFAULT_SEQ

    return f"{unit}-{ctrl_isa}-{seq}"


# ─────────────────────────────────────────────────────────────────────────
# ADNOC GAS — descriptive vocabulary mirroring the manual "Manual Inst Index"
# reference sheet. Editing this dict re-styles every record returned by the
# AI: the LHS is the ISA function code (TYPE token); the RHS is the verbose
# label the engineer actually writes in the spreadsheet. Lookup is exact —
# add new keys here as new instrument codes appear in the legend.
# ─────────────────────────────────────────────────────────────────────────
_ADNOC_GAS_INSTRUMENT_TYPE_MAP = {
    # Flow
    "FE":   "Flow Element (Orifice)",
    "FT":   "Flow Transmitter (DP Type)",
    "FI":   "Flow Indicator",
    "FG":   "Flow Glass / Sight Flow",
    "FQI":  "Flow Quantity Indicator (Totalizer)",
    "FV":   "Flow Control Valve (FCV, Globe)",
    "FCV":  "Flow Control Valve (FCV, Globe)",
    "FIC":  "Flow Indicating Controller",
    "FC":   "Flow Controller",
    "FY":   "Flow Computing Relay",
    # Pressure
    "PG":   "Pressure Gauge",
    "PT":   "Pressure Transmitter",
    "PIT":  "Pressure Indicating Transmitter",
    "PI":   "Pressure Indicator",
    "PV":   "Pressure Control Valve (PCV, Globe)",
    "PCV":  "Pressure Control Valve (PCV, Self-actuated)",
    "PSV":  "Pressure Safety Valve",
    "PSE":  "Pressure Safety Element (Rupture Disc)",
    # Temperature
    "TE":   "Temperature Element (RTD with Thermowell)",
    "TT":   "Temperature Transmitter",
    "TIT":  "Temperature Indicating Transmitter",
    "TI":   "Temperature Indicator",
    "TG":   "Temperature Gauge (With Thermowell)",
    "TW":   "Thermowell",
    "TV":   "Temperature Control Valve (TCV)",
    # Level
    "LG":   "Level Gauge (Mag)",
    "LT":   "Level Transmitter (GWR)",
    "LIT":  "Level Indicating Transmitter (GWR)",
    "LI":   "Level Indicator",
    "LV":   "Level Control Valve (LCV, Globe)",
    "LCV":  "Level Control Valve (LCV, Globe)",
    "LSH":  "Level Switch High",
    "LSL":  "Level Switch Low",
    # Analyser
    "AE":   "Analyzer Element (PH)",
    "AT":   "Analyzer Transmitter (PH)",
    "AIT":  "Analyzer Indicating Transmitter",
    "AI":   "Analyzer Indicator",
    # Shutdown / on-off / motorised
    "SDV":  "Shutdown Valve",
    "BDV":  "Blowdown Valve",
    "XV":   "On/Off Valve (Solenoid Actuated)",
    "MOV":  "Motor Operated Valve",
    # Safety / fire & gas
    "VSH":  "Vibration Switch High",
}

# Service-prefix verbs used to phrase "Service" descriptions like the manual
# (e.g. "MBW To 803-E-XX1 LP Steam Generator"). Used by the prompt only —
# the AI is asked to follow this style.
_ADNOC_GAS_SERVICE_VERBS = ["From", "To", "Inlet", "Outlet", "Suction", "Discharge"]


# ─────────────────────────────────────────────────────────────────────────
# ADNOC GAS — Instrument-Type variant resolver.
#
# Some ISA codes have multiple acceptable verbose labels in the reference
# sheet depending on context (e.g. FI on a rotameter line is "Flow Meter
# (Rotameter)" but FI on a DCS panel is "Flow Indicator"). The default
# label lives in `_ADNOC_GAS_INSTRUMENT_TYPE_MAP`; this resolver overrides
# it when the row's context (location, system, line size) signals a
# variant. Soft-coded — extend per ISA code as new variants appear in the
# legend without touching `_apply_adnoc_gas_style` core logic.
#
# Each rule: (predicate(inst) → bool, override_label).
# Predicates receive the full instrument dict with already-uppercased
# `tag`/`isa` injected.
# ─────────────────────────────────────────────────────────────────────────
def _ctx_is_field(inst):
    return (inst.get("location") or "").strip().upper().startswith("FIELD")

def _ctx_is_vessel(inst):
    return (inst.get("location") or "").strip().upper().startswith("VESSEL") or \
           "VESSEL" in (inst.get("location") or "").upper()

def _ctx_no_dcs(inst):
    sys = (inst.get("system") or "").strip().upper()
    return sys in ("", "-", "N/A", "NA", "NONE", "NULL")

def _ctx_line_small(inst, max_inches=2.0):
    """True when the line size token (e.g. '3/4\"-803-...' / '1-1/2\"-...') is ≤ max_inches."""
    ln = (inst.get("line_number") or "")
    m = re.match(r'\s*(\d+(?:[\-/]\d+)*)(?:\s*[/]\s*\d+)?\s*"', ln)
    if not m:
        return False
    raw = m.group(1)
    try:
        # Handle '3/4', '1-1/2', '1/2' style fractions
        if "-" in raw:
            whole, frac = raw.split("-", 1)
            whole = float(whole)
        else:
            whole, frac = 0.0, raw
        if "/" in frac:
            n, d = frac.split("/")
            val = whole + float(n) / float(d)
        else:
            val = whole + float(frac)
        return val <= max_inches
    except Exception:
        return False

def _ctx_service_has(inst, *kw):
    s = (inst.get("service_description") or "").upper()
    return any(k.upper() in s for k in kw)

# Per-ISA variant rules. First matching rule wins.
_ADNOC_GAS_TYPE_VARIANTS = {
    "FI": [
        # Field-mounted FI on a rotameter (no DCS, small line) → rotameter
        (lambda i: _ctx_is_field(i) and _ctx_no_dcs(i), "Flow Meter (Rotameter)"),
    ],
    "FE": [
        # FE on small-bore lines (≤ 2") → integral orifice variant
        (lambda i: _ctx_line_small(i, max_inches=2.0), "Flow Element (Integral Orifice)"),
    ],
    "FV": [
        # Blowdown / vent / drain service → angle-body globe valve
        (lambda i: _ctx_service_has(i, "BLOWDOWN", "BLOW DOWN", "VENT", "DRAIN"),
         "Flow Control Valve (FCV, Angle Type)"),
    ],
    "PCV": [
        # PCV with self-actuated keyword in remark/notes
        (lambda i: _ctx_service_has(i, "SELF-ACTUATED", "SELF ACTUATED"),
         "Pressure Control Valve (PCV, Self-actuated)"),
    ],
}


def _adnoc_resolve_instrument_type(isa, inst, default_label):
    """
    Return the best instrument-type label for ``inst``: the first variant
    whose predicate matches, otherwise ``default_label``. Pure function —
    callers decide whether to overwrite an existing AI value.
    """
    for predicate, label in _ADNOC_GAS_TYPE_VARIANTS.get(isa, []):
        try:
            if predicate(inst):
                return label
        except Exception:
            continue
    return default_label


# ─────────────────────────────────────────────────────────────────────────
# ADNOC GAS — Service-description normaliser.
#
# Normalises the free-text Service column so it matches the manual's
# phrasing (e.g. "MBW To 803-E-XX1 LP Steam Generator", "LP Steam Generator
# Level"). Pure clean-up — only fixes capitalisation and whitespace, and
# weaves in the equipment context when the AI returned a bare keyword.
# ─────────────────────────────────────────────────────────────────────────
# Connector verbs that should be Title-Cased in the middle of a service phrase.
_ADNOC_GAS_SERVICE_CONNECTORS = {"to", "from", "inlet", "outlet",
                                 "suction", "discharge", "via", "at"}
# Tokens kept fully UPPER-CASE wherever they appear (process abbreviations).
_ADNOC_GAS_SERVICE_UPPER_TOKENS = {
    "LP", "MP", "HP", "VHP", "LLP",
    "MBW", "FW", "BFW", "DM", "BD",
    "PSV", "PCV", "FCV", "TCV", "LCV",
    "FT", "PT", "TT", "LT", "AT",
    "I/O", "DCS", "ESD", "F&G", "PH", "TI",
    "A/B", "A/B/C", "A/B/C/D",
}
# Title-case overrides for common multi-word service phrases (preserves
# proper-noun casing so "Steam Generator" doesn't become "Steam generator").
_ADNOC_GAS_SERVICE_TITLES = (
    "Steam Generator", "Steam to", "Steam from",
    "Continuous Blowdown", "Intermittent Blowdown",
    "Flash Drum", "KO Drum", "Knockout Drum",
    "Reflux Drum", "Suction Drum", "Storage Tank",
    "Heat Exchanger", "Air Cooler",
    "Diesel Product", "Diesel Cooler",
    "Safe Location", "Vent Header",
)


def _adnoc_titlecase_service(raw):
    """
    Manual-style title-case for ADNOC service text. Keeps process
    abbreviations upper-case, lower-cases connector verbs (then re-caps the
    first letter of the phrase), and collapses runs of whitespace.
    """
    s = (raw or "").strip()
    if not s:
        return s
    # Collapse internal whitespace.
    s = re.sub(r"\s+", " ", s)
    # Token-by-token re-casing.
    out = []
    for tok in s.split(" "):
        if not tok:
            continue
        bare = re.sub(r"[^A-Za-z0-9/&]", "", tok)
        upper = bare.upper()
        # Already-canonical equipment tag (\d{3}-[A-Z]+-\w+) → keep upper-case
        if re.match(r"^\d{3}-[A-Z]+-[A-Z0-9X/]+", tok.upper()):
            out.append(tok.upper())
            continue
        if upper in _ADNOC_GAS_SERVICE_UPPER_TOKENS:
            # Preserve trailing punctuation.
            tail = tok[len(bare):] if len(bare) < len(tok) else ""
            out.append(upper + tail)
            continue
        if upper.lower() in _ADNOC_GAS_SERVICE_CONNECTORS:
            out.append(upper.lower())
            continue
        # Default: Title case the alphabetic stem.
        out.append(tok[:1].upper() + tok[1:].lower())
    s = " ".join(out)
    # Capitalise first letter (overrides connector-lowering at start).
    if s:
        s = s[:1].upper() + s[1:]
    # Apply preferred multi-word title overrides (case-insensitive).
    for phrase in _ADNOC_GAS_SERVICE_TITLES:
        s = re.sub(re.escape(phrase), phrase, s, flags=re.IGNORECASE)
    return s


# ─────────────────────────────────────────────────────────────────────────
# ADNOC GAS — Service-description templating.
#
# Builds manual-style phrases like:
#   • "MBW To 803-E-XX1 LP Steam Generator"
#   • "Diesel Product to Diesel Product Cooler (803-EA-005)"
#   • "LP Steam to Vent (Safe Location)"
#   • "From LP Steam Generator to Continuous Blowdown"
#   • "LP Steam Generator Level"
#   • "803-E-XX1 LP Steam Generator - Pressure"
#
# Soft-coded vocabularies — extend per legend without touching core logic.
# ─────────────────────────────────────────────────────────────────────────

# Fluid labels keyed by line-number service designation token (the 3rd token
# in the ADNOC line-number scheme: e.g. '2"-803-MBWXX2-31270X-I' → 'MBW').
# Stems with trailing 'X' or digits are stripped before lookup.
# Reference: ADNOC line-numbering legend.
_ADNOC_FLUID_LABEL_MAP = {
    "MBW":   "MBW",
    "BFW":   "BFW",
    "FW":    "Feed Water",
    "DMW":   "DM Water",
    "BD":    "Blowdown",
    "BA":    "Continuous Blowdown",
    "LS":    "LP Steam",
    "MS":    "MP Steam",
    "HS":    "HP Steam",
    "VHS":   "VHP Steam",
    "STH":   "HP Steam",
    "STL":   "LP Steam",
    "STS":   "Steam",
    "CW":    "Cooling Water",
    "CWS":   "Cooling Water Supply",
    "CWR":   "Cooling Water Return",
    "IW":    "Industrial Water",
    "RW":    "Raw Water",
    "PW":    "Process Water",
    "UW":    "Utility Water",
    "WS":    "Wash Water",
    "OD":    "Open Drain",
    "CD":    "Closed Drain",
    "FD":    "Flare Drain",
    "FG":    "Fuel Gas",
    "FGL":   "LP Fuel Gas",
    "FGM":   "MP Fuel Gas",
    "FGX":   "Fuel Gas",
    "RG":    "Regeneration Gas",
    "N2":    "Nitrogen",
    "IA":    "Instrument Air",
    "PA":    "Plant Air",
    "AG":    "Acid Gas",
    "AGX":   "Acid Gas",
    "CG":    "Combustion Gas",
    "CH":    "Phosphate",
    "CHX":   "Phosphate",
    "DF":    "Diesel Fuel",
    "P":     "Diesel Product",
    "PXX":   "Diesel Product",
    "HC":    "Hydrocarbon",
    "HCX":   "Hydrocarbon",
    "HR":    "Hot Recycle",
    "LO":    "Lube Oil",
    "GD":    "Gas Distribution",
    "GL":    "Glycol",
    "VE":    "Vent",
    "VT":    "Vent",
    "FL":    "Flare",
    "FLC":   "Closed Flare",
    "SC":    "Sour Condensate",
    "SD":    "Sour Drain",
    "SE":    "Sour Effluent",
    "SLH":   "Slop Header",
    "SO":    "Slop Oil",
    "SW":    "Sour Water",
}

# Property each measurement-ISA reads on a vessel-mounted device.
# Used to build "<eq_desc> <Property>" or "<eq_tag> <eq_desc> - <Property>".
_ADNOC_VESSEL_PROPERTY_MAP = {
    "PT": "Pressure", "PIT": "Pressure", "PI": "Pressure", "PG": "Pressure",
    "TT": "Temperature", "TIT": "Temperature", "TI": "Temperature",
    "TG": "Temperature", "TE": "Temperature", "TW": "Temperature",
    "LT": "Level", "LIT": "Level", "LI": "Level", "LG": "Level",
    "AT": "Analysis", "AIT": "Analysis", "AI": "Analysis", "AE": "Analysis",
    "ST": "Speed", "VT": "Vibration",
}

# Control-valve ISA → suffix appended to the loop's service text
# ("MBW To … LP Steam Generator" + " - FCV").
_ADNOC_VALVE_SUFFIX_MAP = {
    "FV": "FCV", "FCV": "FCV",
    "PV": "PCV", "PCV": "PCV",
    "LV": "LCV", "LCV": "LCV",
    "TV": "TCV", "TCV": "TCV",
}

# Service designations that signal a destination phrase rather than a fluid
# (e.g. 'BA' → "to Continuous Blowdown", 'VE' → "to Vent (Safe Location)").
_ADNOC_DESTINATION_PHRASES = {
    "BA":  "Continuous Blowdown",
    "BD":  "Blowdown",
    "VE":  "Vent (Safe Location)",
    "VT":  "Vent (Safe Location)",
    "FL":  "Flare",
    "FLC": "Closed Flare",
    "FD":  "Flare Drain",
    "OD":  "Open Drain",
    "CD":  "Closed Drain",
}

# Recognise an ADNOC line-number and return its 3rd-token stem (the service
# designation portion). Soft-coded so changes to the line scheme stay here.
_ADNOC_LINE_NUMBER_RE = re.compile(
    r'^\s*[\d\-/"\']+\s*-\s*\d{3,4}\s*-\s*([A-Z]{1,5})[A-Z0-9]*\s*-',
    re.IGNORECASE,
)

# Sub-equipment / process-component names that, when mentioned in the AI's
# text, identify a more specific routing target than the parent vessel
# (e.g. an FT on the outlet of a Superheater coil within the LP Steam
# Generator vessel). Order = match priority. Each entry maps an UPPERCASE
# search keyword → the canonical Title-Case label used in the manual.
_ADNOC_SUBEQUIPMENT_LABELS = (
    ("SUPERHEATER",        "Superheater"),
    ("ECONOMISER",         "Economiser"),
    ("ECONOMIZER",         "Economiser"),
    ("REBOILER",           "Reboiler"),
    ("CONDENSER",          "Condenser"),
    ("DESUPERHEATER",      "Desuperheater"),
    ("DEAERATOR",          "Deaerator"),
    ("STEAM DRUM",         "Steam Drum"),
    ("MUD DRUM",           "Mud Drum"),
    ("KO DRUM",            "KO Drum"),
    ("KNOCK-OUT DRUM",     "KO Drum"),
    ("PRODUCT COOLER",     "Product Cooler"),
    ("AIR COOLER",         "Air Cooler"),
    ("AFTERCOOLER",        "Aftercooler"),
    ("INTERCOOLER",        "Intercooler"),
    ("HEAT EXCHANGER",     "Heat Exchanger"),
)

# Tokens that hint the instrument sits on an OUTLET/discharge line —
# fluid is leaving the parent equipment. Used to flip the templater's
# default "<fluid> To <eq>" pattern into "<fluid> from <eq>".
_ADNOC_OUTLET_TOKENS = (
    "OUTLET", "OUT TO", "DISCHARGE", "DISCH",
    "FROM ", " EX ", "EXIT", "EFFLUENT",
)

# Tokens that hint the instrument sits on an INLET/suction line.
_ADNOC_INLET_TOKENS = (
    "INLET", "SUCTION", "SUCT", "FEED TO", "TO INLET",
)

# Equipment-tag pattern used to spot a SECOND tag (a source / destination
# equipment) inside the AI's service text — e.g. "Diesel Product from
# 803-E-012 A/B to LP Steam Generator" should preserve "803-E-012 A/B"
# as the source. Permits trailing letter/slash suffixes (A, B, A/B).
_ADNOC_EQUIPMENT_TAG_RE = re.compile(
    r"\b(\d{3}-[A-Z]{1,3}-[A-Z0-9X]{2,5}(?:\s*[A-Z](?:/[A-Z])?)?)\b"
)

# ── Soft-coded reject patterns for the SERVICE column ──────────────────
# AI sometimes leaks title-block boilerplate (drawing-number tokens,
# system codes, raw line numbers, generic phrases like "Process line").
# Anything that matches these patterns must NOT be displayed — the
# templater is the authoritative source for those rows; if it cannot
# build a phrase, the cell stays empty.
_ADNOC_SERVICE_REJECT_PHRASES = (
    "PROCESS LINE", "INSTRUMENT SIGNAL", "INSTRUMENT LINE", "MEASUREMENT",
    "FLOW MEASUREMENT", "PRESSURE MEASUREMENT", "LEVEL MEASUREMENT",
    "TEMPERATURE MEASUREMENT", "GENERAL SERVICE", "UTILITY", "N/A",
    "NOT APPLICABLE", "NOT AVAILABLE", "NONE", "UNKNOWN", "TBD", "TBC",
    "REFER TO", "SEE DRAWING", "AS PER", "AS SHOWN",
)

# Regex patterns that, when matched, force the AI value to be discarded.
# - "System 803", "System 31270", "Sys-803" — title-block / area codes
# - bare drawing numbers (TAK300171-...)
# - bare line-number strings (e.g. "2\"-803-MBWXX2-31270X-I") with no verb
# - bare ISA codes ("FT", "PT-XXXX")
_ADNOC_SERVICE_REJECT_RE = (
    re.compile(r"\bSYSTEM\s*\d", re.IGNORECASE),
    re.compile(r"\bSYS[\s\-]*\d", re.IGNORECASE),
    re.compile(r"\bUNIT\s*\d{3,4}\b", re.IGNORECASE),
    re.compile(r"\bAREA\s*\d{2,4}\b", re.IGNORECASE),
    re.compile(r"\bTAK\d{6,}", re.IGNORECASE),
    re.compile(r"^\s*\d+[\"']\s*-\s*\d{3,4}\s*-\s*[A-Z0-9\-]+\s*$"),
    re.compile(r"^\s*[A-Z]{1,5}-[A-Z0-9X]+\s*$"),
)


def _adnoc_service_is_rejected(text):
    """Return True when the AI's service text is title-block / system-code
    boilerplate and must be cleared."""
    if not text:
        return False
    upper = text.upper().strip()
    for phrase in _ADNOC_SERVICE_REJECT_PHRASES:
        if phrase in upper:
            return True
    for rx in _ADNOC_SERVICE_REJECT_RE:
        if rx.search(text):
            return True
    return False


# ── Soft-coded patterns for CALIBRATION RANGE & ALARM extraction ────────
# These are scraped directly from the drawing's text (title block,
# equipment data box, setpoint callouts) so the columns populate even
# when the AI couldn't read the values. All patterns are case-insensitive
# and tolerate the `°`, `²`, `³` glyph variants commonly seen on P&IDs.
_ADNOC_DESIGN_PRESS_RE = re.compile(
    r"(?:TUBE\s+DESIGN\s+PRESS\.?|SHELL\s+DESIGN\s+PRESS\.?|"
    r"DESIGN\s+PRESS(?:URE)?)\s*[:\-]?\s*"
    r"(\d+\.?\d*)\s*"
    r"(KG\s*/\s*CM\s*[\^]?\s*[2²]?\s*G?|BAR\s*G?|MPA\s*G?|PSI\s*G?)",
    re.IGNORECASE,
)
_ADNOC_DESIGN_TEMP_RE = re.compile(
    r"(?:TUBE\s+DESIGN\s+TEMP\.?|SHELL\s+DESIGN\s+TEMP\.?|"
    r"DESIGN\s+TEMP(?:ERATURE)?)\s*[:\-]?\s*"
    r"(\d+\.?\d*)\s*°?\s*([CF])\b",
    re.IGNORECASE,
)
_ADNOC_PSV_SET_RE = re.compile(
    r"SET\s*@\s*(\d+\.?\d*)\s*"
    r"(KG\s*/\s*CM\s*[\^]?\s*[2²]?\s*G?|BAR\s*G?|MPA\s*G?|PSI\s*G?)",
    re.IGNORECASE,
)

# Canonical engineering-unit labels (so output is consistent regardless of
# how the unit was written on the drawing — "KG/CM2G", "kg/cm^2g",
# "kg/cm²g" all collapse to the same string).
def _adnoc_canonical_unit(raw):
    if not raw:
        return ""
    u = re.sub(r"\s+", "", raw.upper())
    if u.startswith("KG"):
        return "kg/cm²g"
    if u.startswith("BAR"):
        return "barg"
    if u.startswith("MPA"):
        return "MPag"
    if u.startswith("PSI"):
        return "psig"
    if u in ("C", "°C"):
        return "°C"
    if u in ("F", "°F"):
        return "°F"
    return raw

# ISA → calibration-range derivation rule. Soft-coded so adding a new
# instrument type only requires extending this map.
#   ("pressure", 0, "design") → range = 0..min(tube_press, shell_press), kg/cm²g
#   ("temperature", 0, "design") → range = 0..min(tube_temp, shell_temp), °C
#   ("level", 0, 100)            → range = 0..100, %
_ADNOC_CALIBRATION_RULES = {
    # Pressure measurement → 0..design_press
    "PT":  ("pressure",    0, "design"),
    "PI":  ("pressure",    0, "design"),
    "PG":  ("pressure",    0, "design"),
    "PIT": ("pressure",    0, "design"),
    # Temperature measurement → 0..design_temp
    "TT":  ("temperature", 0, "design"),
    "TI":  ("temperature", 0, "design"),
    "TG":  ("temperature", 0, "design"),
    "TE":  ("temperature", 0, "design"),
    # Level measurement → 0..100 %
    "LT":  ("level",       0, 100),
    "LG":  ("level",       0, 100),
    "LI":  ("level",       0, 100),
    "LIT": ("level",       0, 100),
    "LC":  ("level",       0, 100),
}

# ── ADNOC Onshore typical instrument / calibration ranges (soft-coded) ──
# Keyed by ISA. Used as a last-resort fallback when no design pressure /
# temperature data was scraped from the P&ID. The "Inst range" column on
# the manual datasheet captures the *physical sensor* range; the
# "Calibration range" column captures the *configured 4-20 mA span* —
# in absence of project-specific data the two start out identical and are
# refined manually downstream. Edit this table to retune defaults
# without touching `_apply_adnoc_onshore_style`.
#
# Each entry: (min, max, unit). ``""`` for *unit* skips the unit column.
_ADNOC_ONSHORE_TYPICAL_RANGES = {
    # ── Pressure ─────────────────────────────────────────────────────
    "PT":   ("0", "10",  "barg"),
    "PIT":  ("0", "10",  "barg"),
    "PI":   ("0", "10",  "barg"),
    "PG":   ("0", "10",  "barg"),
    "PDT":  ("0", "500", "mbar"),
    "PDIT": ("0", "500", "mbar"),
    "DPT":  ("0", "500", "mbar"),
    "DPIT": ("0", "500", "mbar"),
    "DPI":  ("0", "500", "mbar"),
    "PSH":  ("0", "10",  "barg"),
    "PSL":  ("0", "10",  "barg"),
    "PSHH": ("0", "10",  "barg"),
    "PSLL": ("0", "10",  "barg"),
    # ── Temperature ──────────────────────────────────────────────────
    "TT":   ("0", "150", "°C"),
    "TIT":  ("0", "150", "°C"),
    "TI":   ("0", "150", "°C"),
    "TG":   ("0", "150", "°C"),
    "TE":   ("0", "150", "°C"),
    "TW":   ("0", "150", "°C"),
    "TSH":  ("0", "150", "°C"),
    "TSL":  ("0", "150", "°C"),
    "TSHH": ("0", "150", "°C"),
    "TSLL": ("0", "150", "°C"),
    # ── Level ────────────────────────────────────────────────────────
    "LT":   ("0", "100", "%"),
    "LIT":  ("0", "100", "%"),
    "LI":   ("0", "100", "%"),
    "LG":   ("0", "100", "%"),
    "LSH":  ("0", "100", "%"),
    "LSL":  ("0", "100", "%"),
    "LSHH": ("0", "100", "%"),
    "LSLL": ("0", "100", "%"),
    "LALL": ("0", "100", "%"),
    "LAHH": ("0", "100", "%"),
    "LAH":  ("0", "100", "%"),
    "LAL":  ("0", "100", "%"),
    # ── Flow ─────────────────────────────────────────────────────────
    "FT":   ("0", "100", "m³/h"),
    "FIT":  ("0", "100", "m³/h"),
    "FE":   ("0", "100", "m³/h"),
    "FI":   ("0", "100", "m³/h"),
    "FQ":   ("0", "100", "m³/h"),
    "FQI":  ("0", "100", "m³/h"),
    # ── Position / valve ─────────────────────────────────────────────
    "ZT":   ("0", "100", "%"),
    "ZI":   ("0", "100", "%"),
    "FZT":  ("0", "100", "%"),
    "FZI":  ("0", "100", "%"),
    # ── Analyser ─────────────────────────────────────────────────────
    "AT":   ("0", "100", "%"),
    "AIT":  ("0", "100", "%"),
    "AI":   ("0", "100", "%"),
    "AE":   ("0", "100", "%"),
    # ── Vibration ────────────────────────────────────────────────────
    "VT":   ("0", "20",  "mm/s"),
    "VSH":  ("0", "20",  "mm/s"),
    "VSL":  ("0", "20",  "mm/s"),
    "VAH":  ("0", "20",  "mm/s"),
    "VAL":  ("0", "20",  "mm/s"),
}

# ISA categories whose datasheet "Inst Range" / "Calibration Range" cells
# are conventionally left blank — final-element devices, mechanical
# protection devices, and on/off valves don't carry a measurement span.
_ADNOC_ONSHORE_RANGE_BLANK_ISA = {
    # Final elements / control valves
    "FV", "FCV", "LV", "LCV", "PV", "PCV", "TV", "TCV",
    "FY", "PY", "TY", "LY", "AY", "HY",
    # On/off & shutdown valves & solenoids
    "SDV", "BDV", "SOV", "MOV", "ROV", "XV",
    "ESV", "SSV", "SSSV", "HV",
    # Mechanical relief / orifice
    "PSV", "PRV", "PVSV", "PSE", "RO", "RD",
    # Manual switches
    "HS",
}


def _adnoc_onshore_design_range(cal_data, isa):
    """Return ``(min, max, unit)`` from drawing-derived design data.

    Priority:
      * Pressure ISA → ``(0, design_press, design_press_unit)``
      * Temperature ISA → ``(0, design_temp, design_temp_unit)``
      * Level ISA → ``(0, 100, "%")`` (always)
    Returns ``None`` if no rule matches or design data is missing.
    """
    if not cal_data or not isa:
        return None
    rule = _ADNOC_CALIBRATION_RULES.get(isa)
    if not rule:
        return None
    kind, lo, _hi = rule
    if kind == "pressure" and cal_data.get("design_press"):
        return (str(lo), str(cal_data["design_press"]),
                cal_data.get("design_press_unit") or "")
    if kind == "temperature" and cal_data.get("design_temp"):
        return (str(lo), str(cal_data["design_temp"]),
                cal_data.get("design_temp_unit") or "")
    if kind == "level":
        return (str(lo), "100", "%")
    return None


def _adnoc_onshore_resolve_range(inst, cal_data):
    """Merge drawing-derived + typical-fallback ranges, returning
    ``(min, max, unit)`` or ``None`` when the ISA is range-blank
    (final elements, on/off valves, mechanical reliefs)."""
    isa = _adnoc_onshore_resolve_isa(inst)
    if not isa or isa in _ADNOC_ONSHORE_RANGE_BLANK_ISA:
        return None
    # 1) Drawing-derived design data wins when available.
    drv = _adnoc_onshore_design_range(cal_data, isa)
    if drv:
        return drv
    # 2) Typical-range fallback table.
    return _ADNOC_ONSHORE_TYPICAL_RANGES.get(isa)


# Default alarm pattern for ISA codes that almost always carry alarms in
# ADNOC drawings. Marker = '✓' (a tick) so the column is populated even
# when no numeric setpoint is present on the drawing — matches the manual
# convention where alarm columns are flagged, not numerically populated.
_ADNOC_ALARM_DEFAULT_MARKERS = {
    # Level instruments — typically L + H alarms
    "LT":  {"alarm_l": "L",  "alarm_h": "H"},
    "LIT": {"alarm_l": "L",  "alarm_h": "H"},
    "LIC": {"alarm_l": "L",  "alarm_h": "H"},
    "LC":  {"alarm_l": "L",  "alarm_h": "H"},
    "LI":  {"alarm_l": "L"},
    "LG":  {"alarm_l": "L"},
    "LSH": {"alarm_h": "H"},
    "LSL": {"alarm_l": "L"},
    "LSHH":{"alarm_hh":"HH"},
    "LSLL":{"alarm_ll":"LL"},
    # PSV — high-high trip
    "PSV": {"alarm_hh": "HH"},
    # Pressure switches
    "PSH": {"alarm_h":  "H"},
    "PSL": {"alarm_l":  "L"},
    # Temperature switches
    "TSH": {"alarm_h":  "H"},
    "TSL": {"alarm_l":  "L"},
}


# ── Soft-coded LOCATION inference rules ──────────────────────────────────
# Determines whether an instrument sits on the parent vessel/equipment
# shell ("Vessel") or on a piping run / in the field ("Field"). The frontend
# accepts: "Field", "Vessel", "Local Panel", "Control Room".
#
#   _ADNOC_VESSEL_MOUNTED_ISA  → ISA codes that are ALWAYS vessel-mounted
#                                 in ADNOC drawings (level + sight glasses).
#   _ADNOC_VESSEL_HINT_ISA     → ISA codes that are vessel-mounted ONLY
#                                 when no line_number is present (i.e.
#                                 the device is sitting directly on the
#                                 equipment shell, not on a tap-off pipe).
#   _ADNOC_LINE_MOUNTED_ISA    → ISA codes that ALWAYS sit on a line
#                                 (orifices, control valves, flow elements).
#   _ADNOC_LOCATION_KEYWORDS   → free-text snippets in the AI's location
#                                 string mapped to canonical labels.
_ADNOC_VESSEL_MOUNTED_ISA = {"LG", "LT", "LIT", "LI", "LSH", "LSL", "LSHH", "LSLL"}
_ADNOC_VESSEL_HINT_ISA    = {"PG", "TG", "TT", "TI", "TE", "PT", "PI", "PSV"}
_ADNOC_LINE_MOUNTED_ISA   = {"FE", "FT", "FI", "FQ", "FV", "PV", "LV", "TV", "RO", "AE", "AT"}
_ADNOC_LOCATION_KEYWORDS  = (
    ("CONTROL ROOM", "Control Room"),
    ("LOCAL PANEL",  "Local Panel"),
    ("LOCAL",        "Local Panel"),
    ("PANEL",        "Local Panel"),
    ("FIELD",        "Field"),
    ("LINE",         "Field"),
    ("PIPING",       "Field"),
    ("VESSEL",       "Vessel"),
    ("DRUM",         "Vessel"),
    ("EXCHANGER",    "Vessel"),
    ("TANK",         "Vessel"),
    ("SHELL",        "Vessel"),
    ("COLUMN",       "Vessel"),
    ("REACTOR",      "Vessel"),
    ("SEPARATOR",    "Vessel"),
)


def _adnoc_resolve_location(inst, isa):
    """Return the canonical Location label for an instrument.

    Priority:
      1. AI-supplied location text → mapped via _ADNOC_LOCATION_KEYWORDS.
      2. ISA hard rule (LG / LT / LI / LSH / LSL → Vessel).
      3. ISA soft rule (PG / TG / PT / TT / TI → Vessel when no line_no).
      4. ISA hard rule (FE / FT / FV / RO → Field).
      5. Default: Field (line-mounted is the most common case in ADNOC P&IDs).
    """
    cur = (inst.get("location") or "").strip()
    if cur:
        upper_cur = cur.upper()
        for needle, canonical in _ADNOC_LOCATION_KEYWORDS:
            if needle in upper_cur:
                return canonical

    if isa in _ADNOC_VESSEL_MOUNTED_ISA:
        return "Vessel"

    line_no = (inst.get("line_number") or "").strip()
    if isa in _ADNOC_VESSEL_HINT_ISA and not line_no:
        return "Vessel"

    if isa in _ADNOC_LINE_MOUNTED_ISA:
        return "Field"

    # Default fallback — most P&ID instruments are field-mounted on lines.
    return "Field"


# ── Soft-coded I/O TYPE inference rules ─────────────────────────────────
# DCS I/O classification per ISA function letter. Convention:
#   AI    = Analog Input  (4-20 mA / HART transmitter signal into DCS)
#   AI-R  = Analog Input, Redundant — used for transmitters whose reading
#           also drives a control loop (controllers PIC/TIC/LIC/FIC).
#   AO    = Analog Output (4-20 mA to a positioner)
#   AO-R  = Analog Output, Redundant — control valves with position feedback
#   DI    = Discrete Input  (switch contact: PSH/PSL/LSH/LSL/TSH/TSL)
#   DO    = Discrete Output (solenoid / shutdown valve command)
#   DO-R  = Discrete Output, Redundant — SDV/BDV/MOV with limit-switch feedback
#   ""    = no I/O (purely local indicator: PG/TG/LG/PI/TI/LI, or PSV)
#
#   _ADNOC_IO_TYPE_BY_ISA  → ISA-only mapping (used when no controller in loop).
#   _ADNOC_IO_REDUNDANT_ISA → ISA codes that imply -R suffix (controllers,
#                              control valves, shutdown valves with feedback).
#   _ADNOC_IO_LOCAL_ISA    → ISA codes with no DCS I/O at all.
_ADNOC_IO_TYPE_BY_ISA = {
    # Analog Inputs — transmitters
    "PT":   "AI",   "TT":   "AI",   "FT":   "AI",   "LT":   "AI",
    "AT":   "AI",   "LIT":  "AI",   "PDT":  "AI",   "FIT":  "AI",
    "TE":   "AI",
    # Analog Inputs — indicating transmitters / controllers (drive loop → -R)
    "PIT":  "AI-R", "TIT":  "AI-R",
    "PIC":  "AI-R", "TIC":  "AI-R", "FIC":  "AI-R", "LIC":  "AI-R",
    "PC":   "AI-R", "TC":   "AI-R", "FC":   "AI-R", "LC":   "AI-R",
    # Analog Outputs — control valves
    "PV":   "AO-R", "TV":   "AO-R", "FV":   "AO-R", "LV":   "AO-R",
    "PCV":  "AO-R", "TCV":  "AO-R", "FCV":  "AO-R", "LCV":  "AO-R",
    "FY":   "AO",   "PY":   "AO",   "TY":   "AO",   "LY":   "AO",
    # Discrete Inputs — switches
    "PSH":  "DI",   "PSL":  "DI",   "PSHH": "DI",   "PSLL": "DI",
    "TSH":  "DI",   "TSL":  "DI",   "TSHH": "DI",   "TSLL": "DI",
    "LSH":  "DI",   "LSL":  "DI",   "LSHH": "DI",   "LSLL": "DI",
    "FSH":  "DI",   "FSL":  "DI",   "ZSH":  "DI",   "ZSL":  "DI",
    # Discrete Outputs — shutdown / blowdown / on-off valves with feedback
    "SDV":  "DO-R", "BDV":  "DO-R", "MOV":  "DO-R", "XV":   "DO-R",
    "ESD":  "DO-R",
    # Solenoids on their own → simple DO
    "SOV":  "DO",   "SV":   "DO",
}

# Local-only devices: no DCS I/O, leave column blank.
_ADNOC_IO_LOCAL_ISA = {"PG", "TG", "LG", "PI", "TI", "LI", "FI", "FG",
                       "PSV", "PRV", "RO", "FO", "RD"}


def _adnoc_resolve_io_type(inst, isa):
    """Return canonical I/O type per ISA + signal/control hints.

    Priority:
      1. Local-only ISA (PG / TG / LG / PI / TI / LI / PSV) → "" (blank).
      2. Direct ISA → I/O lookup (covers the vast majority of cases).
      3. Signal-type hint (HART / 4-20 / discrete) — last-resort fallback.
      4. Default: "" (blank — operator fills in).

    The "-R" (redundant / with feedback) suffix is applied automatically
    for controllers (xIC), control valves (xV / xCV), and shutdown valves
    (SDV / BDV / MOV / XV / ESD) — those always need return-feedback in
    a SIL-rated DCS.
    """
    if isa in _ADNOC_IO_LOCAL_ISA:
        return ""

    direct = _ADNOC_IO_TYPE_BY_ISA.get(isa)
    if direct:
        return direct

    # Fallback: probe signal_type / tag for hints.
    sig = (inst.get("signal_type") or "").upper()
    tag = (inst.get("tag_number") or "").upper()
    if "DISCRETE" in sig or "0/1" in sig or "ON/OFF" in sig:
        if any(p in tag for p in ("SDV", "BDV", "MOV", "XV", "ESD")):
            return "DO-R"
        if any(p in tag for p in ("SOV", "SV")):
            return "DO"
        return "DI"
    if "4-20" in sig or "HART" in sig or "ANALOG" in sig:
        if any(p in tag for p in ("FV", "PV", "TV", "LV", "FCV", "PCV", "TCV", "LCV")):
            return "AO-R"
        # Controllers (xIC) imply analog input feeding a loop → AI-R
        if any(p in tag for p in ("PIC", "TIC", "FIC", "LIC", "PC-", "TC-", "FC-", "LC-")):
            return "AI-R"
        return "AI"

    return ""


# ── Soft-coded ADNOC LINE-NUMBER scrape & normaliser ────────────────────
# ADNOC Gas / Onshore line-number scheme (5 dash-separated parts):
#   SIZE - UNIT(3-4 digits) - SERVICE_CODE - SEQUENCE - INSULATION
#
# Examples (from project P&ID_5.pdf):
#   3/4"-803-CHXX1-33030X-V
#   2"-803-MBWXX2-31270X-I
#   1-1/2"-803-BAXX1-31270X-I
#   8"-803-PXX2-31051Y-I
#   6"-803-LSXX2-31270X-I
#
# The pattern is intentionally permissive on the SIZE token so it accepts:
#   - integer NPS (8")
#   - fraction NPS (3/4")
#   - compound NPS (1-1/2")
#   - any inch-symbol glyph variant (",″,'',etc.)
_ADNOC_LINE_FULL_RE = re.compile(
    r'((?:\d+-)?\d+(?:/\d+)?)\s*'      # group 1: SIZE token (8 | 3/4 | 1-1/2)
    r'["″\u2033\u02BA\'\u2019\uFF02]?\s*'   # optional inch glyph
    r'[-–—]\s*'
    r'(\d{3,4})\s*[-–—]\s*'             # group 2: UNIT (3-4 digits)
    r'([A-Z]{1,6}[A-Z0-9]{0,4})\s*[-–—]\s*'  # group 3: SERVICE_CODE
    r'(\d{4,6}[A-Z]?)\s*[-–—]\s*'       # group 4: SEQUENCE (digits + opt letter)
    r'([A-Z])\b',                       # group 5: INSULATION (single letter)
)

# Canonical inch-symbol glyph used when reformatting.
_ADNOC_INCH_GLYPH = '"'


def _adnoc_extract_drawing_line_numbers(pdf_text_blob):
    """Return a deduplicated list of canonical ADNOC line numbers found
    in ``pdf_text_blob`` (preserving first-seen order).

    Each entry is normalised to:
       <size>"-<unit>-<service>-<sequence>-<insulation>
    so partial hits later can be looked up via simple substring match.
    """
    if not pdf_text_blob:
        return []
    seen = set()
    out = []
    for m in _ADNOC_LINE_FULL_RE.finditer(pdf_text_blob.upper()):
        size, unit, svc, seq, ins = m.groups()
        canonical = f'{size}{_ADNOC_INCH_GLYPH}-{unit}-{svc}-{seq}-{ins}'
        if canonical not in seen:
            seen.add(canonical)
            out.append(canonical)
    return out


def _adnoc_normalise_line_number(raw, canonical_lines):
    """Upgrade an instrument's ``line_number`` value.

    Logic (priority order):
      1. If ``raw`` already matches the full 5-part scheme, return it
         unchanged (formatting normalised — spaces removed, inch glyph
         restored).
      2. Else, if ``raw`` is a non-empty substring of any canonical line,
         return that canonical line (so AI partials get upgraded).
      3. Else return the original ``raw`` (untouched — never invent data).

    ``canonical_lines`` is the list returned by
    :func:`_adnoc_extract_drawing_line_numbers`.
    """
    if not raw:
        return raw
    txt = str(raw).strip().upper()
    if not txt or txt in ("-", "N/A", "NA", "NONE", "NULL"):
        return raw

    # Strip extraneous whitespace around dashes for matching.
    compact = re.sub(r'\s*-\s*', '-', txt)

    # 1) Exact canonical match — already 5-part. Just normalise inch glyph.
    m = _ADNOC_LINE_FULL_RE.match(compact)
    if m:
        size, unit, svc, seq, ins = m.groups()
        return f'{size}{_ADNOC_INCH_GLYPH}-{unit}-{svc}-{seq}-{ins}'

    # 2) Partial — try to upgrade by substring match against canonicals.
    if canonical_lines:
        # Drop inch glyph from both sides for permissive substring search.
        needle = compact.rstrip('"\'\u2033\u2019\u02BA').replace('"', '')
        if len(needle) >= 5:
            for cand in canonical_lines:
                cand_bare = cand.replace('"', '')
                if needle in cand_bare:
                    return cand

    # 3) No upgrade possible — return raw (never invent).
    return raw


def _adnoc_extract_drawing_calibration_data(pdf_text_blob):
    """Soft-coded scraper for design pressure / temperature / PSV setpoint
    values from a P&ID's title-block + equipment-data text. Returns a dict
    with keys:
       design_press      → float | None
       design_press_unit → canonical unit string
       design_temp       → float | None
       design_temp_unit  → "°C" | "°F"
       psv_setpoints     → [(value:float, unit:str), …] (in document order)
    """
    out = {
        "design_press": None,        "design_press_unit": "",
        "design_temp":  None,        "design_temp_unit":  "",
        "psv_setpoints": [],
    }
    if not pdf_text_blob:
        return out

    # Pressure: keep the LOWEST design pressure found (conservative —
    # matches manual where shell ratings drive vessel-side instruments).
    p_values = []
    for m in _ADNOC_DESIGN_PRESS_RE.finditer(pdf_text_blob):
        try:
            p_values.append((float(m.group(1)), _adnoc_canonical_unit(m.group(2))))
        except ValueError:
            continue
    if p_values:
        p_values.sort(key=lambda x: x[0])
        out["design_press"], out["design_press_unit"] = p_values[0]

    t_values = []
    for m in _ADNOC_DESIGN_TEMP_RE.finditer(pdf_text_blob):
        try:
            t_values.append((float(m.group(1)), _adnoc_canonical_unit(m.group(2))))
        except ValueError:
            continue
    if t_values:
        # Use the HIGHEST design temperature (worst-case service).
        t_values.sort(key=lambda x: -x[0])
        out["design_temp"], out["design_temp_unit"] = t_values[0]

    for m in _ADNOC_PSV_SET_RE.finditer(pdf_text_blob):
        try:
            out["psv_setpoints"].append(
                (float(m.group(1)), _adnoc_canonical_unit(m.group(2)))
            )
        except ValueError:
            continue

    return out



def _adnoc_extract_service_designation(line_number):
    """Return the canonical service-designation stem from an ADNOC line
    number (e.g. '2"-803-MBWXX2-31270X-I' → 'MBW'). Returns '' on no match."""
    if not line_number:
        return ""
    m = _ADNOC_LINE_NUMBER_RE.match(line_number)
    if not m:
        return ""
    raw = m.group(1).upper()
    # Try progressively shorter prefixes against the fluid map so 'MBWXX2'
    # → 'MBW', 'CHX' → 'CH', 'PXX' → 'P'.
    for n in (len(raw), 4, 3, 2, 1):
        cand = raw[:n]
        if cand in _ADNOC_FLUID_LABEL_MAP or cand in _ADNOC_DESTINATION_PHRASES:
            return cand
    return raw


def _adnoc_build_service_description(inst, eq_tag, eq_desc):
    """
    Build a manual-style service description from the instrument's context.

    Decision tree (all soft-coded above):
      • Vessel-mounted measurement (PT/PG/LT/LG/TT/TG/…):
          → "<eq_desc> <Property>"  (or with eq_tag prefix for PT)
      • PSV with a fluid line:
          → "<Fluid> to Vent (Safe Location)"
      • Line-mounted device with a destination service code (BA/BD/VE/FL):
          → "From <eq_desc> to <Destination>"
      • Line-mounted device with a fluid service code:
          → "<Fluid> To <eq_tag> <eq_desc>"
      • Control valves: append " - FCV/PCV/LCV/TCV" suffix to the loop text.
      • Fallback: ""  (caller may keep AI value or use eq description).
    """
    tag = (inst.get("tag_number") or "").upper()
    m = re.match(r"^[A-Z0-9]+-([A-Z]{1,5})-", tag)
    isa = m.group(1) if m else ""
    if not isa:
        return ""

    # Normalise equipment description to manual case (e.g. "LP STEAM GENERATOR"
    # → "LP Steam Generator"). Idempotent.
    if eq_desc:
        eq_desc = _adnoc_titlecase_service(eq_desc)

    location = (inst.get("location") or "").strip().upper()
    on_vessel = "VESSEL" in location or "DRUM" in location or "EXCHANGER" in location
    line_no = inst.get("line_number") or ""
    svc_code = _adnoc_extract_service_designation(line_no)
    fluid = _ADNOC_FLUID_LABEL_MAP.get(svc_code, "")
    destination = _ADNOC_DESTINATION_PHRASES.get(svc_code, "")
    valve_suffix = _ADNOC_VALVE_SUFFIX_MAP.get(isa, "")

    # ── Fluid inference fallbacks (when line_number missing/unparseable) ──
    if not fluid and not destination:
        # 1b) Scan the AI's existing service_description for a fluid keyword.
        #     Skip rejected boilerplate ("System 803", "Process line", etc.).
        ai_text_raw = (inst.get("service_description") or "")
        scan_text = ai_text_raw.upper() if not _adnoc_service_is_rejected(ai_text_raw) else ""
        if scan_text.strip():
            FLUID_KEYWORDS = (
                ("MBW", "MBW"),
                ("BOILER FEED WATER", "BFW"), ("BFW", "BFW"),
                ("FEED WATER", "Feed Water"),
                ("DM WATER", "DM Water"), ("DEMIN WATER", "DM Water"),
                ("LP STEAM", "LP Steam"), ("MP STEAM", "MP Steam"),
                ("HP STEAM", "HP Steam"), ("VHP STEAM", "VHP Steam"),
                ("STEAM", "Steam"),
                ("PHOSPHATE", "Phosphate"),
                ("DIESEL PRODUCT", "Diesel Product"),
                ("DIESEL", "Diesel"),
                ("COOLING WATER", "Cooling Water"),
                ("INSTRUMENT AIR", "Instrument Air"),
                ("PLANT AIR", "Plant Air"),
                ("NITROGEN", "Nitrogen"),
                ("FUEL GAS", "Fuel Gas"),
                ("ACID GAS", "Acid Gas"),
                ("SOUR WATER", "Sour Water"),
                ("LUBE OIL", "Lube Oil"),
                ("HYDROCARBON", "Hydrocarbon"),
            )
            DEST_KEYWORDS = (
                ("CONTINUOUS BLOWDOWN", "Continuous Blowdown"),
                ("BLOWDOWN", "Blowdown"),
                ("VENT", "Vent (Safe Location)"),
                ("FLARE", "Flare"),
                ("CLOSED DRAIN", "Closed Drain"),
                ("OPEN DRAIN", "Open Drain"),
            )
            for needle, label in FLUID_KEYWORDS:
                if needle in scan_text:
                    fluid = label
                    break
            if not destination:
                for needle, label in DEST_KEYWORDS:
                    if needle in scan_text:
                        destination = label
                        break

    # ── Vessel inference: LG / LT / LI almost always sit on a vessel even
    #    when location is unset; PG / TG without a line number → vessel.
    if not on_vessel and isa in ("LG", "LT", "LIT", "LI"):
        on_vessel = True
    if not on_vessel and isa in ("PG", "TG", "TT", "TI", "PT", "PI") and not line_no:
        on_vessel = True

    eq_phrase_full = f"{eq_tag} {eq_desc}".strip() if (eq_tag and eq_desc) else (
        eq_desc or eq_tag or ""
    )

    # ── 1) Vessel-mounted measurement ──
    prop = _ADNOC_VESSEL_PROPERTY_MAP.get(isa)
    if on_vessel and prop and eq_desc:
        # Pressure shown with eq tag prefix (matches manual row 20):
        # "803-E-XX1 LP Steam Generator - Pressure"
        if prop == "Pressure" and eq_tag:
            return f"{eq_tag} {eq_desc} - {prop}"
        # Level / Temperature use plain eq description form (rows 28-31).
        return f"{eq_desc} {prop}"

    # ── 2) PSV / safety-vent on a fluid line ──
    if isa == "PSV" and fluid:
        return f"{fluid} to Vent (Safe Location)"

    # ── 3) Destination service codes (blowdown / vent / flare lines) ──
    if destination and eq_desc:
        return f"From {eq_desc} to {destination}"

    # ── 4) Fluid → equipment routing (most common pattern) ──────────────
    #     Direction-aware: detect outlet/inlet hints and the presence of a
    #     sub-equipment (Superheater, Cooler, …) or a second equipment tag
    #     in the AI's text so we route to the more specific destination.
    ai_raw = (inst.get("service_description") or "")
    ai_upper = ai_raw.upper()

    # 4a) Find a sub-equipment name in the AI text (Superheater, Cooler, …).
    sub_eq_label = ""
    for needle, label in _ADNOC_SUBEQUIPMENT_LABELS:
        if needle in ai_upper:
            sub_eq_label = label
            break

    # 4b) Find a SECOND equipment tag in the AI text (different from
    #     the parent vessel). Useful for "Diesel Product from 803-E-012 A/B
    #     to LP Steam Generator".
    second_tag = ""
    eq_tag_root = (eq_tag or "").upper().split()[0] if eq_tag else ""
    for m_tag in _ADNOC_EQUIPMENT_TAG_RE.finditer(ai_upper):
        cand = m_tag.group(1).strip()
        cand_norm = re.sub(r"\s+", " ", cand)
        cand_root = cand_norm.split()[0] if cand_norm else ""
        if cand_root and cand_root != eq_tag_root:
            second_tag = cand_norm
            break

    is_outlet = any(tok in ai_upper for tok in _ADNOC_OUTLET_TOKENS)
    is_inlet = any(tok in ai_upper for tok in _ADNOC_INLET_TOKENS)

    if fluid:
        # 4c) Outlet line going to a sub-equipment within the same vessel
        #     → "<fluid> to <Sub-equipment> <eq_tag>"
        #     (matches "LP Steam to Superheater 803-E-XX1")
        #     If a second equipment tag was found, prefer it (the
        #     sub-equipment lives on a different parent — e.g. Diesel
        #     Product Cooler 803-EA-005).
        if sub_eq_label and (is_outlet or isa in ("FT", "FE", "FI")):
            target_tag = second_tag or eq_tag
            if target_tag:
                base = f"{fluid} to {sub_eq_label} {target_tag}"
                if valve_suffix:
                    return f"{base} - {valve_suffix}"
                return base

        # 4d) Cross-equipment routing where AI text shows a source tag
        #     → "<fluid> from <src_tag> to <eq_desc>"
        #     (matches "Diesel Product from 803-E-012 A/B to LP Steam Generator")
        if second_tag and eq_desc and is_outlet:
            base = f"{fluid} from {second_tag} to {eq_desc}"
            if valve_suffix:
                return f"{base} - {valve_suffix}"
            return base

        # 4e) Default inlet form → "<fluid> To <eq_tag> <eq_desc>"
        if eq_phrase_full:
            base = f"{fluid} To {eq_phrase_full}"
            if valve_suffix:
                return f"{base} - {valve_suffix}"
            return base

    # ── 5) Bare valve suffix when no fluid found but we have eq context ──
    if valve_suffix and eq_phrase_full:
        return f"To {eq_phrase_full} - {valve_suffix}"

    return ""


def _build_adnoc_gas_style_block():
    """Compact ADNOC Gas style/vocabulary reference for the AI prompt."""
    type_lines = "\n".join(
        f"  • {code:<5} → {label}" for code, label in _ADNOC_GAS_INSTRUMENT_TYPE_MAP.items()
    )
    return (
        "ADNOC GAS STYLE GUIDE — match the manual reference sheet exactly:\n"
        "1. INSTRUMENT TYPE — return the verbose descriptive label, NOT the bare ISA code:\n"
        f"{type_lines}\n"
        "2. LOOP NUMBER (column 'Loop No.') — this is the PAIRED CONTROL-SYSTEM TAG, not just the digits:\n"
        "   • For a transmitter (FT/PT/TT/LT/AT) → return the matching DCS controller/indicator\n"
        "     tag (FT→FC or FI, PT→PI, TT→TI, LT→LI or LC, AT→AI), e.g. '803-FC-XXXX'.\n"
        "   • For a control valve (FV/PV/LV/TV) → return the same controller tag as its loop driver.\n"
        "   • For field-only devices (FE, PG, TG, LG, PSV, TE, TW, RO) → return '-'.\n"
        "   • If the loop digits are not finalised on the drawing (e.g. 'XXXX' or 'XXX'),\n"
        "     PRESERVE THE PLACEHOLDER as-is — do NOT guess numbers.\n"
        "3. TAG NUMBER — preserve placeholder digits exactly:\n"
        "   • '803-FT-XXXX' must be returned as '803-FT-XXXX' (4 X's). Do not normalise to '0000'.\n"
        "4. SERVICE — short line-context phrase using verbs From/To/Inlet/Outlet/Suction/Discharge,\n"
        "   referencing the equipment by its tag. CRITICAL — match these exact patterns:\n"
        "   • Line-mounted INLET (location=Field, fluid going INTO eq):\n"
        "       '<FLUID> To <eq_tag> <eq_description>'\n"
        "       e.g. 'MBW To 803-E-XX1 LP Steam Generator'\n"
        "       e.g. 'Phosphate to 803-E-XX1 LP Steam Generator'\n"
        "   • Line-mounted control valve (FV/PV/LV/TV) on inlet: append valve suffix\n"
        "       e.g. 'MBW To 803-E-XX1 LP Steam Generator - FCV'\n"
        "   • Line-mounted OUTLET to a sub-component (Superheater / Cooler / Reboiler /\n"
        "     Economiser / KO Drum / Heat Exchanger):\n"
        "       '<FLUID> to <Sub-equipment> <eq_tag>'\n"
        "       e.g. 'LP Steam to Superheater 803-E-XX1'\n"
        "       e.g. 'Diesel Product to Diesel Product Cooler (803-EA-005)'\n"
        "   • Cross-equipment routing with a SOURCE equipment tag:\n"
        "       '<FLUID> from <src_tag> to <dest_eq_description>'\n"
        "       e.g. 'Diesel Product from 803-E-012 A/B to LP Steam Generator'\n"
        "   • Vessel-mounted Pressure (PT/PI/PG): '<eq_tag> <eq_description> - Pressure'\n"
        "       e.g. '803-E-XX1 LP Steam Generator - Pressure'\n"
        "   • Vessel-mounted Level (LT/LG/LI):  '<eq_description> Level'\n"
        "       e.g. 'LP Steam Generator Level'\n"
        "   • Vessel-mounted Temperature (TG/TT/TE/TI):  '<eq_description> Temperature'\n"
        "       e.g. 'LP Steam Generator Temperature'\n"
        "   • PSV vent line: '<FLUID> to Vent (Safe Location)'\n"
        "       e.g. 'LP Steam to Vent (Safe Location)'\n"
        "   • Blowdown / drain line: 'From <eq_description> to Continuous Blowdown'\n"
        "       e.g. 'From LP Steam Generator to Continuous Blowdown'\n"
        "   IMPORTANT — when you describe an outlet line, ALWAYS include the word\n"
        "   'OUTLET' or the destination component name (Superheater, Cooler, …) so the\n"
        "   post-processor can route correctly. NEVER return generic phrases like\n"
        "   'Process line' / 'Instrument signal' / 'Flow measurement'. If you cannot\n"
        "   determine the routing from the drawing, return an empty string.\n"
        "   FORBIDDEN OUTPUTS for SERVICE — DO NOT EVER RETURN THESE:\n"
        "     • 'System NNN', 'System 803', 'System 31270' (these are AREA codes,\n"
        "       not service descriptions — they belong in the drawing title block).\n"
        "     • 'Sys-NNN', 'Unit NNN', 'Area NNN'.\n"
        "     • Bare drawing numbers like 'TAK300171-803-PRU-B-0104'.\n"
        "     • Bare ISA codes like 'FT-XXXX' or 'PT'.\n"
        "     • Bare line numbers like '2\"-803-MBWXX2-31270X-I' (use the FLUID name\n"
        "       extracted from the line number, not the line number itself).\n"
        "     If you can only see boilerplate text and cannot identify a real fluid\n"
        "     or routing, RETURN AN EMPTY STRING for service. Empty is correct;\n"
        "     boilerplate is wrong.\n"
        "5. LOCATION — 'Field' for line-mounted, 'Vessel' for instruments mounted directly on a\n"
        "   drum/exchanger/tower (PT/PG/TG/LT/LG/LI on equipment shell).\n"
        "6. I/O TYPE — DCS terminology: 'AI' (analog in), 'AI-R' (with redundancy), 'AO-R'\n"
        "   (analog out, redundant). Use '-' for non-DCS field devices.\n"
        "7. SYSTEM — 'DCS' if the instrument has a DCS I/O, otherwise '-'.\n"
        "8. LINE NO — for vessel-mounted instruments, put the EQUIPMENT TAG (e.g. '803-E-XX1')\n"
        "   here; for line-mounted instruments, put the full piping line number\n"
        "   (e.g. '2\"-803-MBWXX2-31270X-I').\n"
    )


# Shared template body for ADNOC family (Onshore / Gas). Both categories
# follow the same drawing convention (Habshan-5 Unit 562 etc.).
_ADNOC_PROMPT_INTRO = (
    "PROJECT TEMPLATE: ADNOC P&ID Instrument Index. "
    "Tags follow the format <UNIT>-<TYPE>-<LOOP>, where UNIT is the "
    "3-digit area / unit code (e.g. 562 for Habshan-5, 803 for RR-3 Scheme), "
    "TYPE is the ISA letter group from the legend (FT, FE, LT, PT, PSV, MOV, "
    "XV, KV, XDV, etc.), and LOOP is 3-4 digits with an optional single-letter "
    "suffix. Examples of valid tags: 562-FT-1502, 562-PSV-8501A, "
    "562-VSH-7502, 803-FT-XXXX (placeholder digits), 803-LT-XXX, "
    "562-MOV-1101, 562-XDV-3201. "
    "If you read a tag without the unit prefix (e.g. just FT-1502), "
    "still return it that way — the system will prepend the unit code. "
    "Capture calibration ranges, alarm set points, line numbers, "
    "equipment numbers and I/O types when visible. Use the legend "
    "reference below to disambiguate symbols and abbreviations."
) + "\n\n" + _build_adnoc_legend_block()


# ADNOC Gas extends the shared intro with the verbose style guide.
_ADNOC_GAS_PROMPT_INTRO = _ADNOC_PROMPT_INTRO + "\n" + _build_adnoc_gas_style_block()


_ADNOC_TAG_FORMAT = {
    "pattern":           "{unit}-{type}-{loop}",
    "unit_regex":        r"^\d{3}$",
    "type_regex":        r"^[A-Z]{1,5}$",
    "loop_regex":        r"^[A-Z0-9]{2,8}[A-Z]?$",
    "loop_strict_regex": r"^\d{3,4}[A-Z]?$",
    # Placeholder tokens used on FEED / pre-FEED P&IDs where loop digits are
    # not yet finalised. Matches XXX, XXXX, XXX1, XXXXA etc. Soft-coded so
    # other ADNOC schemes can extend the placeholder vocabulary.
    "loop_placeholder_regex": r"^X{2,5}[A-Z0-9]?$",
    # When the AI returns a tag without any loop digits (e.g. 'TT-803' on a
    # FEED P&ID), substitute this placeholder so the canonical tag becomes
    # '<unit>-<type>-XXXX'. Soft-coded — change to '' to disable.
    "loop_default_placeholder": "XXXX",
    # When a tag's loop fragment does not match `loop_strict_regex`
    # (e.g. 'FI-803-9' → loop '9' is too short, or 'PG-31270X-803' →
    # loop '31270X' is too long), replace the loop with the placeholder
    # so the display tag stays in unit-first canonical order. The
    # original fragment is preserved in `instrument_remark`. Soft-coded
    # — set False to keep the original (flagged) tag instead.
    "substitute_nonstd_loop": True,
    "validate_regex":    r"^\d{3}-[A-Z]{1,5}-(?:\d{3,4}|X{2,5})[A-Z0-9]?$",
    "drop_invalid":      False,
    "flag_invalid":      True,
    "unit_from_pid_no":  True,
    "unit_min_len":      3,
    # Soft-coded vocabularies — used by _apply_tag_format to validate the
    # TYPE token against the ADNOC legend. Unknown types are still kept
    # but flagged in the remark.
    "known_types":       set(_ADNOC_INSTRUMENT_TYPES),
}


INSTRUMENT_TEMPLATES = {
    # Default — no extra fields, generic prompt (current behaviour preserved)
    "default": {
        "label": "Standard",
        "extra_fields": [],
        "empty_dash": "—",
    },
    # ADNOC Gas — adds the 17 extra columns from the manual reference sheet
    "adnoc_gas": {
        "label": "ADNOC Gas",
        "extra_fields": _ADNOC_GAS_EXTRA_FIELDS,
        "empty_dash": "-",
        "prompt_intro": _ADNOC_GAS_PROMPT_INTRO,
        "tag_format": _ADNOC_TAG_FORMAT,
    },
    # ADNOC Onshore / Offshore — placeholders, follow default schema until
    # their own legend sheets are loaded into the registry.
    "adnoc_onshore":  {"label": "ADNOC Onshore",  "extra_fields": [], "empty_dash": "—"},
    "adnoc_offshore": {"label": "ADNOC Offshore", "extra_fields": [], "empty_dash": "—"},
}


def get_template(category):
    """Resolve a template by category id with safe fallback."""
    return INSTRUMENT_TEMPLATES.get((category or "").lower(), INSTRUMENT_TEMPLATES["default"])


EXTRACTION_CONFIG = {
    # PDF rendering
    "pdf_dpi":           150,     # DPI for PDF→image conversion (150 is sufficient for A0/A1 P&IDs)
    "max_image_size":    4096,    # Max pixel dimension per image sent to Vision API
    "jpeg_quality":       90,     # JPEG compression quality (0-100)

    # AI engine priority for Vision passes (first available + not-quota-exceeded wins)
    # Supported: 'gemini', 'openai'
    # Note: 'tesseract' is handled separately at the orchestration level (not a Vision engine)
    "ai_engines":        ["gemini", "openai"],

    # Full extraction engine order — controls priority waterfall in extract_instruments
    # 1=gemini, 2=tesseract (OCR, free, no quota), 3=openai (last resort)
    "engine_order":      ["gemini", "tesseract", "openai"],

    # Gemini model
    "gemini_model":      "gemini-2.0-flash",

    # Multi-orientation passes (handles vertical / slanted text)
    "enable_rotation":   True,    # Add 90° CW and 90° CCW passes
    "rotation_angles":   [90, 270],

    # Tiled quadrant scanning (handles dense drawings with tiny tags)
    "enable_tiling":     True,
    "tile_grid":         (2, 2),
    "tile_overlap":      0.12,

    # Tesseract OCR settings
    "enable_tesseract":  True,
    "tesseract_on_all":  True,      # Also run Tesseract on vector PDFs (catches circle text)
    "tesseract_dpi":     150,       # DPI for Tesseract rendering

    # Supplementary OCR engines — run alongside Tesseract on the SAME rendered
    # pages and feed any extra tag candidates into the de-duplicated tag pool.
    # Purely additive recall booster — never replaces Tesseract.
    # Requires `easyocr` / `paddleocr` / `paddlepaddle` (already in requirements.txt).
    "enable_easyocr":    True,
    "enable_paddleocr":  True,
    # Skip the heavy ML engines on quick ad-hoc runs by setting to False; keep
    # them on for production accuracy.
    "supplementary_ocr_min_conf": 0.30,

    # Spatial word-proximity grouping (catches tags split across spans inside circles)
    "spatial_grouping":  True,
    "spatial_radius":    80,        # px at 150 DPI — approx instrument circle diameter

    # OpenAI Vision settings
    "max_tokens_primary":  16000,
    "max_tokens_tile":      8000,
    "temperature":           0.1,
    "model":           "gpt-4o",

    # Minimum text-layer tags before skipping Vision passes
    "min_text_tags":         1,

    # Categories where the full AI Vision extraction must run on every page
    # IN ADDITION TO the free-tier extractors (text-layer + Tesseract OCR
    # + EasyOCR/PaddleOCR). The vision results are union-merged into the
    # free-tier pool so all three extractor families contribute. Add a new
    # category key here to opt it in — purely additive, never replaces the
    # cheaper engines.
    "full_vision_categories": ["adnoc_onshore"],

    # Gemini rate-limit retry: sleep this many seconds then retry once before disabling
    "gemini_retry_delay":    5,
}

# ────────────────────────────────────────────────────────────────────────────
# SOFT-CODED DRAWING/P&ID NUMBER RESOLUTION
# Extract DWG/P&ID number from title-block style labels in PDF text.
# ────────────────────────────────────────────────────────────────────────────
DRAWING_NUMBER_CONFIG = {
    # Labels commonly used in title blocks. Order MATTERS — the loop in
    # `_extract_drawing_number_from_text` returns the first valid match,
    # so place the most specific / canonical client labels first.
    "label_patterns": [
        # ── Highest-priority: CAD filename stamp ───────────────────────
        # Most CAD title blocks render `FILE NAME: <drg>.dwg` on every
        # sheet. The value sits immediately after the label (well inside
        # the 140-char window) and equals the canonical drawing number,
        # so this beats ambiguous body-text cross-references that share
        # the same segmented format. The trailing ".dwg"/".pdf" is
        # naturally clipped by the value patterns at the word boundary.
        r'\bFILE\s*NAME\b',
        # ── ADNOC-style title blocks ───────────────────────────────────
        # "GROUP COMPANY DRG. NO." (a.k.a. DRG/DWG/DOC NO under the same
        # owner column) is the canonical document identifier on ADNOC
        # P&IDs. Win over the generic "DWG NO" so we don't accidentally
        # pick a contractor or supplier number from elsewhere on the
        # sheet. Punctuation (dots) and the order of the qualifier words
        # are tolerated. Edit this list to extend to other clients.
        r'\bGROUP\s*COMPANY\s*(?:DRG|DWG|DRAWING|DOC|DOCUMENT)\.?\s*(?:NO|NUMBER|#)\b',
        r'\bCOMPANY\s*(?:DRG|DWG|DRAWING|DOC|DOCUMENT)\.?\s*(?:NO|NUMBER|#)\b',
        # ── Contractor drawing number (ADNOC Gas convention) ───────────
        # ADNOC Gas title blocks carry the canonical drawing identifier
        # against "CONTR. DWG. NO." or "CONT. DWG. NO.". The fixed-width
        # negative lookbehind blocks the unrelated "ENGG. CONT. DRG NO."
        # field that appears on ADNOC Onshore P&IDs (which is a separate
        # contractor sequence we must NOT pick).
        r'(?<!ENGG\.\s)\bCONT(?:R|RACTOR)?\.?\s*(?:DWG|DRG|DRAWING|DOC|DOCUMENT)\.?\s*(?:NO|NUMBER|#)\b',
        # ── Generic title-block labels ─────────────────────────────────
        r'\b(?:DWG|DRAWING)\s*(?:NO|NUMBER|#)\b',
        r'\bP\s*&\s*ID\s*(?:NO|NUMBER|#)\b',
        r'\bP\s*ID\s*(?:NO|NUMBER|#)\b',
        r'\bDOCUMENT\s*(?:NO|NUMBER|#)\b',
        r'\bDOC\s*(?:NO|NUMBER|#)\b',
    ],
    # Candidate number formats (kept broad but engineering-oriented).
    # The leading segment must be ≥ 2 chars (anchor), but later segments
    # may be a single char to support real drawing numbers like
    # ``TAK300171-803-PRU-B-0104`` whose 4th segment is just "B". The
    # candidate-validity filter (min_length, must-have-letter+digit, must
    # contain a separator) downstream still rejects noise tokens.
    "value_patterns": [
        r'\b([A-Z0-9]{2,}(?:-[A-Z0-9]+){2,})\b',
        r'\b([A-Z]{2,}[0-9]{1,}(?:-[A-Z0-9]{1,}){2,})\b',
        r'\b([A-Z0-9]{3,}(?:[./][A-Z0-9]{2,}){2,})\b',
    ],
    # Hard filters for candidate sanity
    "min_length": 8,
    "max_length": 64,
    "window_chars": 140,
}

# Optional S3 legend-sheet discovery. Used only to enrich interpretation,
# not to replace instrument extraction from the P&ID itself.
LEGEND_S3_CONFIG = {
    "enabled": True,
    "filename_keywords": [
        "legend", "legends", "symbol", "symbols", "abbreviation", "abbr",
    ],
    "preferred_extensions": [".pdf"],
    "max_list_keys": 400,
    "max_candidate_files": 3,
    "max_text_chars": 12000,
    "max_pages_per_file": 3,
}

# ────────────────────────────────────────────────────────────────────────────
# CONTROL SYSTEM TAG (CS TAG) DETECTION CONFIG
# Identifies whether an instrument is a DCS/CS instrument or a field device
# and, for field devices, derives the expected CS controller tag.
# Tune here — no logic changes required.
# ────────────────────────────────────────────────────────────────────────────
CS_TAG_CONFIG = {
    # Full ISA-5.1 function-code prefixes (the leading letters before the dash)
    # that identify DCS / control-system instruments.
    # For these: control_system_tag = tag_number (the instrument IS a CS tag).
    "dcs_function_codes": {
        # Flow controllers / recorders
        "FIC", "FRC", "FC", "FFC", "FFIC", "FQC",
        # Pressure controllers / recorders
        "PIC", "PRC", "PC", "PDC",
        # Temperature controllers / recorders
        "TIC", "TRC", "TC", "TDIC",
        # Level controllers / recorders
        "LIC", "LRC", "LC",
        # Analysis / composition controllers
        "AIC", "ARC", "AC",
        # Generic / multi-variable controllers
        "IC", "RC", "HC", "HIC", "HRC",
        "SC", "SIC", "SRC",
        "XC", "YC", "ZC",
        "QIC", "QC", "DC", "EC", "GC", "JC", "KC", "MC", "NC", "OC", "UC", "VC", "WC",
    },
    # Mapping: field-instrument function-code suffix → controller suffix to derive.
    # key   = suffix of field instrument (after the first measured-variable letter)
    # value = controller suffix to substitute
    # e.g.  FT-3901-01 (suffix="T") → FIC-3901-01 (suffix="IC")
    "transmitter_to_controller": {
        "T":   "IC",    # Transmitter          → Indicating Controller  (FT→FIC)
        "IT":  "IC",    # Indicating Transmitter → Indicating Controller
        "E":   "C",     # Element / Sensor     → Controller            (FE→FC)
    },
    # Regex patterns that identify explicit CS/DCS tag annotations in PDF text.
    # Group 1 must capture the tag value.
    "label_patterns": [
        r'(?:CS|DCS|PLC|F&?G|ESD|SIS|SCADA)\s*[-:–]?\s*(?:TAG|NO\.?|NUM\.?)?\s*[-:–]?\s*([A-Z]{2,6}[-]\d{3,6}(?:[-][A-Z0-9]{1,4})?)',
        r'\bCONTROL(?:LER)?\s+TAG\s*[-:–]\s*([A-Z]{2,6}[-]\d{3,6}(?:[-][A-Z0-9]{1,4})?)',
    ],
    # Characters of PDF text after the CS label to search for the tag value.
    "label_window_chars": 80,
}

# ────────────────────────────────────────────────────────────────────────────
# SOFT-CODED CONTEXTUAL ENRICHMENT CONFIGURATION
# Regex patterns for extracting service context without Vision AI.
# Add/edit patterns here — no logic changes required.
# ────────────────────────────────────────────────────────────────────────────
ENRICHMENT_CONFIG = {
    # ── Piping line number patterns ──────────────────────────────────────
    # ADNOC / oil-gas format:  NPS["-symbol] - FLUID_CODE - AREA - SPEC
    # Examples: 16"-HC-3901-A2A   10-G-3901-A2A   2''-FG-1001-B3A   3/4"-IA-3901-A2A
    #
    # WHY MULTIPLE PATTERNS:
    #   • The NPS inch symbol ("  ″  ''  ʺ) encodes differently across PDF fonts.
    #   • Pattern 1: handles inch symbol between size digits and first dash.
    #   • Pattern 2: no inch symbol (size directly followed by dash).
    #   • Pattern 3: fraction pipe sizes (3/4", 1/2").
    #   • Pattern 4: partial match — SIZE-FLUID-AREA without spec class.
    #   • Pattern 5: area-first format used by SABIC and some FEED contractors.
    "line_no_re": [
        # ADNOC Gas / Onshore 5-part: SIZE - UNIT(3-4 digits) - SERVICE_CODE
        #   - SEQUENCE - INSULATION  (single trailing letter V / I / N / U / H)
        # Service codes embed an "X" placeholder for hold/typical lines, e.g.
        #   3/4"-803-CHXX1-33030X-V
        #   2"-803-MBWXX2-31270X-I
        #   1-1/2"-803-BAXX1-31270X-I    (compound NPS like "1-1/2")
        #   8"-803-PXX2-31051Y-I
        #   6"-803-LSXX2-31270X-I
        # Size group permits: integer (8), fraction (3/4), compound (1-1/2),
        # and any inch-symbol glyph after the digits.
        r'(?<!\w)((?:\d+-)?\d+(?:/\d+)?["″\u2033\u02BA\'\u2019\uFF02]?\s*[-–—]\s*\d{3,4}\s*[-–—]\s*[A-Z]{1,6}[A-Z0-9]{0,4}\s*[-–—]\s*\d{4,6}[A-Z]?\s*[-–—]\s*[A-Z])(?!\w)',
        # Line-list style 5-part format: SIZE-FLUID-SEQUENCE-CLASS-INSULATION
        r'(?<!\w)(\d+(?:\.\d+)?["″\u2033\u02BA\'\u2019\uFF02]?\s*[-–—]\s*[A-Z]{1,4}\s*[-–—]\s*\d{4,6}\s*[-–—]\s*[A-Z]\d[A-Z]\d{1,2}\s*[-–—]\s*[A-Z]{1,2})(?!\w)',
        # Line-list style 6-part format: SIZE-AREA-FLUID-SEQUENCE-CLASS-INSULATION(optional)
        r'(?<!\w)(\d{1,2}["″\u2033\u02BA\'\u2019\uFF02]?\s*[-–—]\s*\d{1,2}\s*[-–—]\s*[A-Z]{1,2}\s*[-–—]\s*\d{4}\s*[-–—]\s*[0-9][A-Z0-9]{5}(?:\s*[-–—]\s*[A-Z]{1,2})?)(?!\w)',
        # Full 4-part: SIZE[inch]-FLUID-AREA-SPEC  (inch symbol optional, many variants)
        r'(?<!\w)(\d{1,3}(?:["″\u2033\u02BA\'\u2019\uFF02]{1,2})?[-]\s*[A-Z]{1,6}\s*[-]\s*\d{3,6}\s*[-]\s*[A-Z][0-9][A-Z0-9]{0,4})(?!\w)',
        # 3-part without SPEC: SIZE-FLUID-AREA  (the spec may be on a separate annotation)
        r'(?<!\w)(\d{1,3}["″\u2033\u02BA\'\u2019]{0,2}[-][A-Z]{1,6}[-]\d{3,6})(?![-A-Z0-9])',
        # Fraction pipe sizes: 3/4"-IA-3901-A2A
        r'(?<!\w)(\d(?:[/]\d)["″\u2033]{0,2}[-][A-Z]{1,6}[-]\d{3,6}(?:[-][A-Z][0-9][A-Z0-9]{0,4})?)(?!\w)',
        # Area-first format: AREA-FLUID-SIZE  e.g. 3901-G-16 or 3901-HC-6"-A2A
        r'(?<!\w)(\d{3,6}[-][A-Z]{1,6}[-]\d{1,3}["″\u2033]{0,2}(?:[-][A-Z][0-9][A-Z0-9]{0,4})?)(?!\w)',
        # Flexible join-product (words joined with dashes by block reconstruction)
        r'\b(\d{1,3}-[A-Z]{1,6}-\d{3,6}-[A-Z][0-9][A-Z0-9]{0,4})\b',
    ],

    # Structured line-number rules borrowed from the working line-list parser.
    # These build canonical line numbers from captured components instead of
    # relying only on raw regex matches.
    "line_no_structured_rules": [
        {
            "name": "five_part_line",
            "pattern": r'(?<!\w)(\d+(?:\.\d+)?)\s*["\']?\s*[-–—]\s*([A-Z]{1,4})\s*[-–—]\s*(\d{4,6})\s*[-–—]\s*([A-Z]\d[A-Z]\d{1,2})\s*[-–—]\s*([A-Z]{1,2})(?!\w)',
        },
        {
            "name": "six_part_line",
            "pattern": r'(?<!\w)(\d{1,2})\s*["\']?\s*[-–—]\s*(\d{1,2})\s*[-–—]\s*([A-Z]{1,2})\s*[-–—]\s*(\d{4})\s*[-–—]\s*([0-9][A-Z0-9]{5})(?:\s*[-–—]\s*([A-Z]{1,2}))?(?!\w)',
        },
    ],

    # ── Equipment tag patterns ───────────────────────────────────────────
    # Whitelist: only codes that are NOT in INSTRUMENT_CATEGORIES.
    # Single-letter equipment codes: V (vessel), E (exchanger), P (pump),
    #   C (compressor/column), K (compressor alt), T (tower/tank), R (reactor), D (drum).
    # Two-letter: LP, HP, HE, VV, KO, SD, TK, AC, WH (none are instrument prefixes).
    # Three-letter: SEP, FLR, CMP, PMP, SCR, EXC, KOD, BFW.
    "equipment_re": [
        # Single-letter ISA equipment codes (safe whitelist — none in INSTRUMENT_CATEGORIES)
        r'\b([VEPCKTR][-]\d{3,5}[A-Z]?(?:[/][A-Z])?)\b',
        # Two-letter equipment codes NOT in INSTRUMENT_CATEGORIES
        r'\b((?:LP|HP|HE|VV|KO|SD|TK|AC|WH|UD|CD|FP|MD|SG)[-]\d{3,5}[A-Z]?)\b',
        # Three-letter codes
        r'\b((?:SEP|FLR|CMP|PMP|SCR|EXC|KOD|BFW|SRT|SKD|VES)[-]\d{3,5}[A-Z]?)\b',
        # Drum variant: D-3901, D-3901A
        r'\b(D[-]\d{3,5}[A-Z]?)\b',
    ],

    # ── Fail-safe position annotations ──────────────────────────────────
    "fail_safe_re": [
        r'\b(FC)\b', r'\b(FO)\b', r'\b(FL)\b',
        r'FAIL[-\s]?(CLOSE[D]?|OPEN|LAST|LOCK(?:ED)?)',
    ],

    # ── Signal / communication type ──────────────────────────────────────
    "signal_re": [
        r'(4[-–/]20\s*m[Aa])',
        r'\b(HART)\b',
        r'\b(Fieldbus|Foundation\s*Fieldbus)\b',
        r'\b(Profibus|PROFIBUS)\b',
        r'\b(Wireless|WirelessHART|ISA100)\b',
        r'\b(Pneumatic|Pneum\.?)\b',
        r'(?<![A-Z])(DI|DO|AO)(?![A-Z])',
        r'\b(24\s*VDC?)\b',
        r'\b(On[-/]Off|Discrete|Digital)\b',
    ],

    # ── Set-point value + engineering unit ──────────────────────────────
    "set_point_re": [
        r'(\d+\.?\d*\s*(?:bara|barg|bar[ag]?|kPa[ag]?|MPa[ag]?|psia|psig|psi))\b',
        r'(\d+\.?\d*\s*(?:\xb0C|\xb0F|degC|degF))\b',
        r'(\d+\.?\d*\s*(?:kg\/h|t\/h|m3\/h|Nm3\/h|MMSCFD|SCFD))\b',
        r'(\d+\.?\d*\s*%\s*(?:FS|LRV|URV|SPAN)?)\b',
        r'(\d+\.?\d*\s*(?:rpm|Hz|m\/s))\b',
    ],

    # Spatial search radius for fail-safe/signal/setpoint (pixels at 150 DPI)
    "spatial_radius": 200,
    # Larger radius for line-number and equipment-number search
    # (pipe label can be 300-500px from the instrument circle on large A0 drawings)
    "spatial_radius_context": 350,
    # Fuzzy area-code tolerance — match line/equip tags whose area code is within
    # this many integer units of the instrument's own area code.
    "area_tolerance": 5,
    # Words to exclude from auto-generated service descriptions
    "exclude_desc_words": {
        "p&id", "drawing", "revision", "sheet", "dwg", "doc", "project", "client",
        "date", "appr", "chck", "eng", "rev", "by", "ref", "scale", "nts",
        "title", "north", "south", "tag", "no.", "number", "the", "and", "for",
        "to", "of", "in", "is", "at", "on", "an", "a",
    },
    # Min word count to use proximity text as service description
    "desc_min_words": 2,
}

# ────────────────────────────────────────────────────────────────────────────
# SOFT-CODED INSTRUMENT VALIDATION / CLEAN-UP LAYER
# ----------------------------------------------------------------------------
# Addresses real-world P&ID feedback (see feature-instrument.instructions.md)
#   1. Random P&ID no (not title-block)  → per-page title-block scan
#   2. Tags inside line-numbers picked up as instruments (e.g. PG-45011 inside
#      "12"-13-PG-45011-A0JP08-F")                            → line-context filter
#   3. Tank-level labels mistakenly flagged as instruments
#      (e.g. LSH-800-300 = tank level, not an instrument)    → level-label heuristic
#   4. ISA-5.1 universal tag format: Unit-InstrumentTag-Seq  → format validator
#   5. Accessories auto-inference: TT usually has TE + TW    → accessories map
#   6. "Inline" function codes to prioritise                 → PG, PT, TG, TT …
#
# Every behaviour here can be switched off, or its threshold tuned, without
# touching any extraction regex or engine code.  Pure post-processing.
# ────────────────────────────────────────────────────────────────────────────
INSTRUMENT_VALIDATION_CONFIG = {
    "enabled": True,

    # 1 ── Per-page P&ID number extraction (from title block) -----------------
    "per_page_pid_no": {
        "enabled": True,
        # Fallback to drawing_info["pid_no"] when title-block detect fails
        "fallback_to_input": True,
    },

    # 2 ── Line-context filter ----------------------------------------------
    # If a candidate instrument tag appears as a TOKEN INSIDE a known
    # line-number pattern on the page text, drop it — it's a line spec, not
    # an instrument.  Patterns are the same as ENRICHMENT_CONFIG['line_no_re']
    # so tuning one updates both.
    "line_context_filter": {
        "enabled": True,
        # Minimum confidence before dropping: if we can't positively confirm
        # the tag lives inside a line number, keep it.
        "require_containment": True,
    },

    # 3 ── Level-label heuristic --------------------------------------------
    # Tags matching `^L[SI]?H?L?-\d{3}-\d{3}$` where the numbers are "round"
    # (like 800-300, 1000-500) are commonly tank level / nozzle labels, not
    # instruments.  Flag them via notes and expose in `warnings` field.
    "level_label_filter": {
        "enabled": True,
        # Function-code prefixes to check
        "prefixes": ["LSH", "LSL", "LSHH", "LSLL", "LI", "LG"],
        # If the trailing segment is a round multiple of this → warn
        "round_multiple": 50,
        # Action: 'warn'  → add note + warning, keep row
        #         'drop'  → remove from results
        "action": "warn",
    },

    # 4 ── ISA-5.1 universal tag-format validator ---------------------------
    # Canonical shape: <UNIT>-<INSTRUMENT_TAG>-<SEQUENCE>
    # UNIT     = 2–4 digits (loop / area / unit number)
    # TAG      = 2–6 letters (ISA function code)
    # SEQUENCE = 1–4 chars, digits optionally followed by a single letter
    "format_validator": {
        "enabled": True,
        # Regexes ordered most-specific → most-permissive.  Any match = valid.
        "valid_patterns": [
            # Standard 3-part:  FT-3901-01, PIT-2600-12A
            r"^[A-Z]{2,6}-\d{2,4}-\d{1,4}[A-Z]?$",
            # 2-part (older Gulf / ADNOC FEED): FT-3901A
            r"^[A-Z]{2,6}-\d{2,4}[A-Z]?$",
            # Unit-prefixed 4-part: 26-FT-3901-01
            r"^\d{2}-[A-Z]{2,6}-\d{2,4}-\d{1,4}[A-Z]?$",
        ],
        # When invalid → keep row but add warning + mark `format_valid=False`.
        "action": "warn",
    },

    # 5 ── Accessory auto-inference ----------------------------------------
    # Parent function code → list of accessory codes that are implied.
    # Soft-coded per ISA-5.1 best practice.  Accessories inherit the parent's
    # unit/sequence numbers and are marked `inferred=True` so engineers can
    # confirm visually on the P&ID.
    "accessories": {
        "enabled": True,
        "map": {
            # Temperature transmitter → element + thermowell
            "TT":  ["TE", "TW"],
            "TIT": ["TE", "TW"],
            # Flow transmitter → element / orifice
            "FT":  ["FE"],
            "FIT": ["FE"],
            # Pressure transmitter (often) → isolation valve manifold (not auto-created —
            # valves are not in INSTRUMENT_CATEGORIES, so we leave it off).
            # Level transmitter (DP-type) → HP/LP taps are piping, not instruments.
        },
        # Duplicate guard — skip accessory if the same code+unit+seq already exists
        "dup_guard": True,
    },

    # 6 ── Inline-instrument priority list ---------------------------------
    # Per feedback: extractions should start from basic inline instruments.
    # We mark these with `is_inline=True` and move them to the top of the
    # result list (index_no is recomputed).
    "inline_priority": {
        "enabled": True,
        "codes": ["PG", "PI", "PT", "TG", "TI", "TT", "LG", "LI", "LT", "FG", "FI", "FE"],
    },
}

# ────────────────────────────────────────────────────────────────────────────
# SOFT-CODED SMART DEFAULTS for Fail-Safe / Signal / Set-Point columns
# ----------------------------------------------------------------------------
# Post-enrichment layer that (1) **validates** the values produced by the
# proximity regex scan against the instrument's function code/category, and
# (2) fills sensible engineering defaults where the drawing does not spell
# the value out explicitly (local gauges → "Local", transmitters → "4-20mA",
# safety valves → "FO", etc.).
#
# Everything here is tunable without touching extraction / regex logic.
# Set `enabled: False` to revert to pure drawing-only values.
# ────────────────────────────────────────────────────────────────────────────
SMART_FIELD_DEFAULTS_CONFIG = {
    "enabled": True,

    # ── FAIL-SAFE rules ─────────────────────────────────────────────────
    "fail_safe": {
        # Function-code prefix → default fail-safe position
        # (ISA-5.1 / industry convention; override per project if needed)
        "by_prefix": {
            # Shutdown / ESD valves default fail-close
            "SDV":  "FC",
            "ESDV": "FC",
            "ESV":  "FC",
            "XV":   "FC",   # on/off block valve — assume FC unless marked
            "BDV":  "FO",   # blow-down valve — fail-open to relieve
            # Pressure-relief / safety
            "PSV":  "FO",   # pressure safety valve — spring-operated, treat as FO
            "PRV":  "FO",
            "PSE":  "FO",   # rupture disc
            # Control valves — conservative industry default = FC
            # (process fluid isolation on air/power failure).  Project-specific
            # services (e.g. anti-surge PCV, cooling-water FCV) may differ and
            # should be corrected by the explicit drawing callout.
            "FCV":  "FC", "FV": "FC",
            "PCV":  "FC", "PV": "FC",
            "TCV":  "FC", "TV": "FC",
            "LCV":  "FC", "LV": "FC",
            "HCV":  "FC", "HV": "FC",
            "CV":   "FC",
            # Solenoid / motor-operated — energise to open → FC on power loss
            "SOV":  "FC", "MOV": "FC",
            "ZV":   "FC",
        },
        # Category-wide fallback (used only if prefix isn't in by_prefix above)
        "by_category": {
            "Shutdown & ESD":   "FC",
            "Pressure Relief":  "FO",
            "Motor & Solenoid": "FC",
        },
        # Clear value if it was wrongly extracted for a non-valve instrument
        "clear_for_non_valve": True,
        # Function codes that qualify as "valve" — others get fail_safe cleared
        "valve_prefixes": {
            "SDV", "ESDV", "ESV", "XV", "BDV", "PSV", "PRV", "PSE",
            "FCV", "PCV", "TCV", "LCV", "HCV", "FV", "PV", "TV", "LV",
            "CV", "PV", "ZV", "MV", "TSV",
        },
    },

    # ── SIGNAL-TYPE rules ───────────────────────────────────────────────
    "signal_type": {
        # Function-code prefix → default signal type
        "by_prefix": {
            # Transmitters — 4-20mA HART is the industry default
            "PT":  "4-20mA HART", "PIT": "4-20mA HART",
            "TT":  "4-20mA HART", "TIT": "4-20mA HART",
            "FT":  "4-20mA HART", "FIT": "4-20mA HART",
            "LT":  "4-20mA HART", "LIT": "4-20mA HART",
            "AT":  "4-20mA HART", "AIT": "4-20mA HART",
            "DPT": "4-20mA HART", "DT":  "4-20mA HART",
            "VT":  "4-20mA HART", "WT":  "4-20mA HART",
            # Switches / discrete devices
            "PSH": "Discrete (DI)", "PSL": "Discrete (DI)",
            "PSHH":"Discrete (DI)", "PSLL":"Discrete (DI)",
            "TSH": "Discrete (DI)", "TSL": "Discrete (DI)",
            "LSH": "Discrete (DI)", "LSL": "Discrete (DI)",
            "LSHH":"Discrete (DI)", "LSLL":"Discrete (DI)",
            "FSH": "Discrete (DI)", "FSL": "Discrete (DI)",
            "ZSH": "Discrete (DI)", "ZSL": "Discrete (DI)",
            # Solenoids / outputs
            "SOV": "Discrete (DO)", "XY": "Discrete (DO)",
            # Valves (on/off vs modulating)
            "SDV": "Discrete (DO)", "ESDV": "Discrete (DO)",
            "XV":  "Discrete (DO)", "BDV": "Discrete (DO)",
            "FCV": "4-20mA", "PCV": "4-20mA",
            "TCV": "4-20mA", "LCV": "4-20mA",
            # Relays / computers
            "FY": "Digital", "PY": "Digital", "TY": "Digital", "LY": "Digital",
            # Local gauges / indicators / sight glasses → NO signal
            "PG":  "Local (Mechanical)", "PI":  "Local (Mechanical)",
            "TG":  "Local (Mechanical)", "TI":  "Local (Mechanical)",
            "LG":  "Local (Mechanical)", "LI":  "Local (Mechanical)",
            "FG":  "Local (Mechanical)", "FI":  "Local (Mechanical)",
            "SG":  "Local (Mechanical)",
            # Elements — upstream of a transmitter, no signal of its own
            "TE":  "RTD / Thermocouple",
            "TW":  "—",
            "FE":  "—",
            "PE":  "—",
            "LE":  "—",
        },
        # Categories that NEVER carry an electronic signal — always force "Local"
        "local_only_categories": set(),
        # If the regex pass put "4-20mA" on a local-gauge code, override it
        "validate_against_prefix": True,
    },

    # ── SET-POINT rules ─────────────────────────────────────────────────
    "set_point": {
        # When no numeric set point found on the drawing, provide a
        # unit hint based on the instrument category so the engineer
        # sees what kind of value is expected.
        "unit_hint_by_category": {
            "Pressure":              "—— bar(g)",
            "Differential Pressure": "—— mbar",
            "Temperature":           "—— °C",
            "Level":                 "—— %",
            "Flow":                  "—— m³/h",
            "Analysis":              "—— ppm",
            "Shutdown & ESD":        "Trip on alarm",
            "Pressure Relief":       "—— bar(g) (set)",
            "Motor & Solenoid":      "Energise / De-energise",
            "Position":              "Open / Closed",
        },
        # Fallback when category is unknown — derive from the first letter
        # of the ISA function code (P = Pressure, T = Temperature, etc.).
        "unit_hint_by_first_letter": {
            "P": "—— bar(g)",
            "T": "—— °C",
            "F": "—— m³/h",
            "L": "—— %",
            "A": "—— ppm",
            "D": "—— mbar",     # D = Differential
            "S": "—— rpm",      # S = Speed
            "V": "—— m/s",      # V = Vibration/Velocity
            "W": "—— kg",       # W = Weight
            "J": "—— A",        # J = Power (current)
            "I": "—— A",        # I = Current
            "E": "—— V",        # E = Voltage
            "Z": "Open / Closed",  # Z = Position
        },
        # Switches always have a set-point; mark "Field-adjustable" when missing
        "switch_prefixes": {
            "PSH", "PSL", "PSHH", "PSLL",
            "TSH", "TSL", "TSHH", "TSLL",
            "LSH", "LSL", "LSHH", "LSLL",
            "FSH", "FSL", "ZSH", "ZSL",
        },
        "switch_default": "Field-adjustable",
        # Control valves get a default when no explicit SP found on drawing
        "control_valve_prefixes": {
            "FCV", "PCV", "TCV", "LCV", "HCV",
            "FV", "PV", "TV", "LV", "HV", "CV",
        },
        "control_valve_default": "Set by DCS loop",
        # Transmitters / indicators / elements generally have no set-point
        "no_setpoint_prefixes": {
            "PT","PIT","TT","TIT","FT","FIT","LT","LIT","AT","AIT",
            "DPT","PG","PI","TG","TI","LG","LI","FG","FI","SG",
            "TE","TW","FE","PE","LE",
        },
        "no_setpoint_marker": "—",
        # Validate current value's unit matches the category;
        # if mismatch (e.g. "100 bar" on a Temperature instrument) → clear.
        "validate_units_by_category": True,
        "category_units": {
            "Pressure":              {"bar", "barg", "bara", "kpa", "mpa", "psi", "psig", "psia"},
            "Differential Pressure": {"mbar", "kpa", "pa", "inh2o", "mmh2o", "bar"},
            "Temperature":           {"°c", "°f", "degc", "degf", "k"},
            "Flow":                  {"m3/h", "nm3/h", "kg/h", "t/h", "mmscfd", "scfd", "gpm", "bpd", "lpm"},
            "Level":                 {"%", "mm", "m", "ft", "in"},
            "Analysis":              {"ppm", "ppb", "%", "mg/l"},
        },
    },
}

# Human-readable verb per instrument category — used to build default service descriptions
_SERVICE_VERB_MAP = {
    "Flow":                  "Flow Measurement",
    "Pressure":              "Pressure Measurement",
    "Temperature":           "Temperature Measurement",
    "Level":                 "Level Measurement",
    "Differential Pressure": "Differential Pressure Measurement",
    "Analysis":              "Process Analyser",
    "Safety":                "Pressure Safety Relief",
    "Shutdown & ESD":        "Shutdown / ESD Control",
    "Control Valves":        "Flow Control Valve",
    "Motor & Solenoid":      "Actuated Valve",
    "Position":              "Valve Position Monitor",
    "Restriction":           "Flow Restriction Orifice",
    "Speed":                 "Speed Measurement",
    "Vibration":             "Vibration Monitor",
    "Weight":                "Weight / Load Measurement",
    "Hand/Manual":           "Manual Indication",
    "Special":               "Special Purpose Device",
}


class InstrumentIndexService:
    """
    Extract ALL instrument tags from a P&ID drawing.
    Engine waterfall (soft-coded via EXTRACTION_CONFIG['ai_engines']):
      1. PyMuPDF text-layer  — free, instant
      2. Gemini Vision       — primary AI (free tier, 1M context)
      3. OpenAI Vision       — fallback AI (GPT-4o)
      4. Tesseract OCR       — last resort for pure scanned PDFs
    """

    def __init__(self):
        self.extraction_config  = EXTRACTION_CONFIG.copy()
        self._quota_exceeded    = False   # set True on OpenAI 429
        self._gemini_quota_exceeded = False
        self.openai_client      = self._init_openai()
        self.gemini_client      = self._init_gemini()
        self.tesseract_available = self._check_tesseract()
        # Lazy-loaded supplementary OCR engines — built on first use to keep
        # the Django boot fast. None = not loaded yet, False = load failed.
        self._easyocr_reader   = None
        self._paddleocr_reader = None

    # ────────────────────────────────────────────────────────────────────
    # Initialisation helpers
    # ────────────────────────────────────────────────────────────────────

    def _init_openai(self):
        try:
            import openai
            api_key = os.environ.get("OPENAI_API_KEY")
            if not api_key:
                logger.warning("[InstrumentIndex] OPENAI_API_KEY not set")
                return None
            client = openai.OpenAI(api_key=api_key)
            logger.info("[InstrumentIndex] ✅ OpenAI client initialised")
            return client
        except Exception as e:
            logger.warning(f"[InstrumentIndex] OpenAI init skipped: {e}")
            return None

    def _init_gemini(self):
        try:
            from google import genai
            api_key = os.environ.get("GEMINI_API_KEY") or os.environ.get("GOOGLE_API_KEY")
            if not api_key:
                logger.warning("[InstrumentIndex] GEMINI_API_KEY not set")
                return None
            client = genai.Client(api_key=api_key)
            logger.info("[InstrumentIndex] ✅ Gemini client initialised")
            return client
        except ImportError:
            logger.warning("[InstrumentIndex] google-genai not installed — Gemini disabled")
            return None
        except Exception as e:
            logger.warning(f"[InstrumentIndex] Gemini init skipped: {e}")
            return None

    def _check_tesseract(self):
        try:
            import pytesseract
            pytesseract.get_tesseract_version()
            logger.info("[InstrumentIndex] ✅ Tesseract available")
            return True
        except Exception:
            logger.info("[InstrumentIndex] Tesseract not available (optional)")
            return False

    # ────────────────────────────────────────────────────────────────────
    # Lazy-loaded supplementary OCR engines (EasyOCR + PaddleOCR)
    # Both initialise on first call only — boot stays fast for users that
    # never trigger an OCR pass.
    # ────────────────────────────────────────────────────────────────────
    def _get_easyocr_reader(self):
        if self._easyocr_reader is False:
            return None
        if self._easyocr_reader is not None:
            return self._easyocr_reader
        try:
            import easyocr
            # English only, CPU mode (matches designiq usage)
            self._easyocr_reader = easyocr.Reader(['en'], gpu=False, verbose=False)
            logger.info("[InstrumentIndex] ✅ EasyOCR initialised")
            return self._easyocr_reader
        except Exception as e:
            logger.warning(f"[InstrumentIndex] EasyOCR unavailable: {e}")
            self._easyocr_reader = False
            return None

    def _get_paddleocr_reader(self):
        if self._paddleocr_reader is False:
            return None
        if self._paddleocr_reader is not None:
            return self._paddleocr_reader
        try:
            from paddleocr import PaddleOCR
            self._paddleocr_reader = PaddleOCR(
                use_angle_cls=True, lang='en', show_log=False
            )
            logger.info("[InstrumentIndex] ✅ PaddleOCR initialised")
            return self._paddleocr_reader
        except Exception as e:
            logger.warning(f"[InstrumentIndex] PaddleOCR unavailable: {e}")
            self._paddleocr_reader = False
            return None

    def _run_supplementary_ocr(self, gray_img, page_no, seen_tags, dn, rev, instruments):
        """
        Run EasyOCR + PaddleOCR on the same preprocessed page image used by
        Tesseract. Any text they recognise is fed through `_scan_for_tags`
        so the existing dedup + regex pipeline picks up extra tags.

        Purely additive — no engine result replaces another. Recall booster
        for circle-embedded text and stylised fonts where Tesseract is weak.
        """
        cfg = self.extraction_config
        min_conf = float(cfg.get("supplementary_ocr_min_conf", 0.30))

        # Convert PIL gray image → numpy array (both libs accept this).
        try:
            import numpy as np
            np_img = np.array(gray_img.convert("RGB"))
        except Exception as ce:
            logger.debug(f"[InstrumentIndex] supplementary OCR: img convert failed p{page_no}: {ce}")
            return

        # ── EasyOCR ─────────────────────────────────────────────────────
        if cfg.get("enable_easyocr", True):
            reader = self._get_easyocr_reader()
            if reader is not None:
                try:
                    results = reader.readtext(np_img, detail=1, paragraph=False)
                    words = [
                        str(text).strip()
                        for (_box, text, conf) in results
                        if str(text).strip() and float(conf or 0) >= min_conf
                    ]
                    if words:
                        full_text = " ".join(words)
                        self._scan_for_tags(
                            full_text, seen_tags, dn, rev, instruments,
                            "Tesseract EasyOCR"   # routed through Tesseract note-mapping → "OCR text"
                        )
                except Exception as ee:
                    logger.debug(f"[InstrumentIndex] EasyOCR p{page_no} error: {ee}")

        # ── PaddleOCR ───────────────────────────────────────────────────
        if cfg.get("enable_paddleocr", True):
            reader = self._get_paddleocr_reader()
            if reader is not None:
                try:
                    # PaddleOCR returns [[box, (text, conf)], …] for v2 API
                    # or {'rec_texts': [...], 'rec_scores': [...]} for v3.
                    raw = reader.ocr(np_img, cls=True) if hasattr(reader, 'ocr') else None
                    words: list[str] = []
                    if isinstance(raw, list) and raw and isinstance(raw[0], list):
                        # v2 layout
                        for line in raw[0] or []:
                            try:
                                _box, (text, conf) = line[0], line[1]
                                if text and float(conf or 0) >= min_conf:
                                    words.append(str(text).strip())
                            except Exception:
                                continue
                    elif isinstance(raw, list) and raw and isinstance(raw[0], dict):
                        # v3 layout
                        d = raw[0]
                        for text, conf in zip(d.get('rec_texts') or [], d.get('rec_scores') or []):
                            if text and float(conf or 0) >= min_conf:
                                words.append(str(text).strip())
                    if words:
                        full_text = " ".join(words)
                        self._scan_for_tags(
                            full_text, seen_tags, dn, rev, instruments,
                            "Tesseract PaddleOCR"
                        )
                except Exception as pe:
                    logger.debug(f"[InstrumentIndex] PaddleOCR p{page_no} error: {pe}")

    # ────────────────────────────────────────────────────────────────────
    # Public entry point
    # ────────────────────────────────────────────────────────────────────

    def extract_instruments(self, pid_bytes, drawing_info, legend_context_override=None):
        """
        Extract all instrument tags from a P&ID PDF (or image).

        Extraction strategy (free-first, AI-enrichment):
          Step 1 — PyMuPDF text layer (3-pass: plain text + block reconstruction + spatial grouping)
          Step 2 — Tesseract OCR alongside text layer on ALL PDFs (catches circle-embedded tags)
          Step 3 — AI enrichment (Gemini, then OpenAI) to fill service descriptions & add missed tags

        Args:
            pid_bytes  : raw file bytes
            drawing_info: dict with drawing_number, drawing_title, revision, project_name

        Returns:
            list[dict]: instrument records
        """
        MIN_TEXT_TAGS = self.extraction_config.get("min_text_tags", 1)
        engine_order  = self.extraction_config.get("engine_order", ["gemini", "tesseract", "openai"])
        cfg = self.extraction_config

        try:
            # Resolve drawing number from title-block text when available.
            # This keeps P&ID No aligned with DWG NO per drawing.
            drawing_info = self._resolve_drawing_info_from_pdf(pid_bytes, drawing_info)
            legend_context = self._merge_legend_contexts(
                legend_context_override,
                self._load_legend_context_from_s3(drawing_info),
            )

            # ── Step 1: PyMuPDF text layer (free, always runs) ────────────
            text_instruments = self._extract_with_text_layer(pid_bytes, drawing_info)
            logger.info(f"[InstrumentIndex] Text-layer: {len(text_instruments)} tags")

            # ── Step 2: Tesseract OCR (free, runs on ALL PDFs) ────────────
            # Catches tags in instrument circles that the text layer may split or miss.
            tess_instruments: list = []
            if self.tesseract_available and cfg.get("tesseract_on_all", True):
                logger.info("[InstrumentIndex] Tesseract OCR — running alongside text layer…")
                tess_instruments = self._extract_with_tesseract(pid_bytes, drawing_info)
                logger.info(f"[InstrumentIndex] Tesseract OCR: {len(tess_instruments)} tags")

            # Merge free-tier results
            all_free = self._merge_instruments(text_instruments + tess_instruments)
            logger.info(f"[InstrumentIndex] Free-tier combined: {len(all_free)} unique tags")

            # ── Step 2.5: Contextual enrichment — pattern-based, zero AI quota ──
            # Fills: loop_number, service_description, line_number, equipment_number,
            #        fail_safe, signal_type, set_point using PDF text layer patterns.
            if all_free:
                logger.info("[InstrumentIndex] Running contextual enrichment (pattern-based)…")
                all_free = self._enrich_from_pdf_context(all_free, pid_bytes)

                # Gemini text-only enrichment (much lighter quota than Vision)
                if not self._gemini_quota_exceeded and self.gemini_client:
                    try:
                        import fitz as _fitz
                        _doc = _fitz.open(stream=pid_bytes, filetype="pdf")
                        _pdf_text = "\n".join(
                            _doc[_p].get_text("text") or "" for _p in range(len(_doc))
                        )
                        _doc.close()
                        logger.info("[InstrumentIndex] Running Gemini text enrichment…")
                        all_free = self._enrich_with_gemini_text(
                            all_free,
                            _pdf_text,
                            legend_text=legend_context.get("text", ""),
                            legend_files=legend_context.get("files", []),
                        )
                    except Exception as _etxt:
                        logger.warning(f"[InstrumentIndex] Gemini text enrichment setup error: {_etxt}")

            # ── Step 3: AI Vision (enrich / fill gaps) ────────────────────
            # Categories listed in `full_vision_categories` always get a
            # full multi-page AI vision pass merged into the free-tier
            # pool — guarantees all three extractor families (text-layer,
            # Tesseract OCR, AI Vision) contribute to the final result.
            _category = (drawing_info or {}).get("project_category") or "default"
            _force_full_vision = _category in set(cfg.get("full_vision_categories", []))

            if _force_full_vision:
                logger.info(
                    "[InstrumentIndex] category='%s' → forcing full multi-page AI vision pass",
                    _category,
                )
                vision_full = []
                seen_v = {(i.get("tag_number") or "").strip().upper() for i in all_free}
                for engine in [e for e in engine_order if e != "tesseract"]:
                    if engine == "gemini" and (self._gemini_quota_exceeded or not self.gemini_client):
                        continue
                    if engine == "openai" and (self._quota_exceeded or not self.openai_client):
                        continue
                    try:
                        pages = self._to_jpeg_pages(pid_bytes)
                    except Exception as pe:
                        logger.error(f"[InstrumentIndex] PDF→image failed: {pe}")
                        continue
                    for page_no, jpeg_page in enumerate(pages, start=1):
                        try:
                            page_insts = self._analyse_page(
                                jpeg_page, drawing_info, page_no, only_engine=engine
                            )
                        except Exception as pae:
                            logger.warning(
                                "[InstrumentIndex] %s page %d failed: %s",
                                engine, page_no, pae,
                            )
                            continue
                        for inst in page_insts:
                            tag = (inst.get("tag_number") or "").strip().upper()
                            if tag and tag not in seen_v:
                                seen_v.add(tag)
                                vision_full.append(inst)
                            elif not tag:
                                vision_full.append(inst)
                    if vision_full:
                        break  # First successful engine wins
                if vision_full:
                    all_free = self._merge_instruments(all_free + vision_full)
                    logger.info(
                        "[InstrumentIndex] Full vision merged: +%d → total %d",
                        len(vision_full), len(all_free),
                    )

            if len(all_free) >= MIN_TEXT_TAGS:
                # Good free-tier results — use AI to enrich service descriptions and find any missed tags
                enriched = []
                for eng in [e for e in engine_order if e != "tesseract"]:
                    if eng == "gemini" and (self._gemini_quota_exceeded or not self.gemini_client):
                        continue
                    if eng == "openai" and (self._quota_exceeded or not self.openai_client):
                        continue
                    try:
                        pages = self._to_jpeg_pages(pid_bytes)
                        if pages:
                            logger.info(f"[InstrumentIndex] AI enrichment via {eng}…")
                            enriched = self._vision_pass(
                                pages[0], drawing_info, 1,
                                extra_hint=(
                                    "Enrich the already-extracted tags with service descriptions, line numbers, "
                                    "signal types, and fail-safe positions. Also add any missed tags. "
                                    "Focus on ALL instrument circles/bubbles."
                                ),
                                mode=f"enrich_{eng}",
                                max_tokens=cfg["max_tokens_primary"],
                                only_engine=eng,
                            )
                    except Exception as ve:
                        logger.warning(f"[InstrumentIndex] {eng} enrichment skipped: {ve}")
                    if enriched:
                        break

                all_instruments = self._merge_instruments(all_free + enriched)

            else:
                # Sparse/no free-tier results — try full AI Vision extraction
                logger.info("[InstrumentIndex] Sparse free-tier — trying full AI Vision extraction…")
                all_instruments = list(all_free)

                for engine in [e for e in engine_order if e != "tesseract"]:
                    if engine == "gemini" and (self._gemini_quota_exceeded or not self.gemini_client):
                        logger.info("[InstrumentIndex] Gemini unavailable — skipping")
                        continue
                    if engine == "openai" and (self._quota_exceeded or not self.openai_client):
                        logger.info("[InstrumentIndex] OpenAI unavailable — skipping")
                        continue

                    logger.info(f"[InstrumentIndex] Full AI scan via {engine}…")
                    try:
                        pages = self._to_jpeg_pages(pid_bytes)
                    except Exception as pe:
                        logger.error(f"[InstrumentIndex] PDF→image failed: {pe}")
                        continue

                    vision_all: list = []
                    seen_vision: set = set()
                    for page_no, jpeg_page in enumerate(pages, start=1):
                        page_insts = self._analyse_page(jpeg_page, drawing_info, page_no, only_engine=engine)
                        for inst in page_insts:
                            tag = (inst.get("tag_number") or "").strip().upper()
                            if tag and tag not in seen_vision:
                                seen_vision.add(tag)
                                vision_all.append(inst)
                            elif not tag:
                                vision_all.append(inst)

                    if vision_all:
                        all_instruments = self._merge_instruments(all_instruments + vision_all)
                        logger.info(f"[InstrumentIndex] {engine} full scan: +{len(vision_all)} tags → total {len(all_instruments)}")
                        break

            # Sequential index numbers
            for i, inst in enumerate(all_instruments, start=1):
                inst["index_no"] = i

            # Ensure drawing number is populated
            dn = drawing_info.get("drawing_number", "")
            pid_no = drawing_info.get("pid_no") or dn
            # ── Soft-coded propagation policy ─────────────────────────
            # When the title-block resolver (`_resolve_drawing_info_from_pdf`)
            # picked up a canonical drawing/PID number, treat it as the
            # source-of-truth and OVERRIDE per-instrument `pid_no` values
            # the Vision/AI step may have guessed from a secondary label
            # (e.g. an ADNOC Gas P&ID where the LLM grabs the supplier
            # "ADNOC GAS DWG. NO." instead of the contractor "CONT. DWG.
            # NO."). Empty/N/A per-instrument values are also filled.
            # Set `_PID_NO_TITLEBLOCK_OVERRIDE = False` to revert to the
            # legacy fill-only-if-empty policy without code changes.
            _PID_NO_TITLEBLOCK_OVERRIDE = True
            _PID_NO_EMPTY_TOKENS = ("N/A", "", None)
            if dn or pid_no:
                for inst in all_instruments:
                    cur_pid = inst.get("pid_no")
                    if cur_pid in _PID_NO_EMPTY_TOKENS or (
                        _PID_NO_TITLEBLOCK_OVERRIDE and pid_no and cur_pid != pid_no
                    ):
                        inst["pid_no"] = pid_no
                    cur_dn = inst.get("drawing_number")
                    if cur_dn in _PID_NO_EMPTY_TOKENS or (
                        _PID_NO_TITLEBLOCK_OVERRIDE and dn and cur_dn != dn
                    ):
                        inst["drawing_number"] = dn

            # ── Smart validation / accessory inference / inline priority ──
            # Soft-coded via INSTRUMENT_VALIDATION_CONFIG (see top of module).
            # Purely additive: filters out false positives (tags inside line
            # numbers, tank-level labels), flags non-ISA-5.1 formats,
            # synthesises accessories (TE+TW for TT, FE for FT), and re-orders
            # inline instruments (PG/PT/TG/TT…) first.
            try:
                all_instruments = self._apply_post_validation(
                    all_instruments, pid_bytes, drawing_info
                )
            except Exception as ve:
                logger.warning(f"[InstrumentIndex] Post-validation skipped: {ve}")

            # ── Soft-coded tag-format normalisation (category-aware) ──
            try:
                all_instruments = self._apply_tag_format(all_instruments, drawing_info)
            except Exception as fe:
                logger.warning(f"[InstrumentIndex] Tag-format normalisation skipped: {fe}")

            # ── Soft-coded template field fill (category-aware) ──
            try:
                all_instruments = self._apply_template_fields(all_instruments, drawing_info, pid_bytes=pid_bytes)
            except Exception as te:
                logger.warning(f"[InstrumentIndex] Template field fill skipped: {te}")

            logger.info(f"[InstrumentIndex] ✅ Total unique instruments: {len(all_instruments)}")
            return all_instruments

        except Exception as e:
            logger.error(f"[InstrumentIndex] extract_instruments error: {e}", exc_info=True)
            return []

    # ────────────────────────────────────────────────────────────────────
    # Contextual enrichment — fills service description, line no., etc.
    # 100 % pattern-based, zero AI quota consumed.
    # ────────────────────────────────────────────────────────────────────

    # ────────────────────────────────────────────────────────────────────
    # SMART VALIDATION / CLEAN-UP LAYER
    # All soft-coded via INSTRUMENT_VALIDATION_CONFIG.  Addresses expert
    # feedback: random P&ID no, tags inside line numbers, tank-level labels,
    # tag-format checks, accessory inference, inline-instrument priority.
    # ────────────────────────────────────────────────────────────────────

    def _extract_per_page_pid_numbers(self, pid_bytes):
        """
        Return dict {page_number (1-based) → detected P&ID No} using the same
        title-block heuristics as the global detector, but scanned page by page.
        If a page has no clear candidate, the entry is omitted.
        """
        result = {}
        try:
            import fitz
        except ImportError:
            return result
        try:
            doc = fitz.open(stream=pid_bytes, filetype="pdf")
        except Exception:
            return result
        try:
            for page_idx in range(len(doc)):
                try:
                    txt = doc[page_idx].get_text("text") or ""
                    cand = self._extract_drawing_number_from_text(txt)
                    if cand:
                        result[page_idx + 1] = cand
                except Exception:
                    continue
        finally:
            doc.close()
        return result

    def _build_page_text_map(self, pid_bytes):
        """Return {page_number → full text} for line-context filtering."""
        out = {}
        try:
            import fitz
            doc = fitz.open(stream=pid_bytes, filetype="pdf")
        except Exception:
            return out
        try:
            for page_idx in range(len(doc)):
                try:
                    out[page_idx + 1] = doc[page_idx].get_text("text") or ""
                except Exception:
                    continue
        finally:
            doc.close()
        return out

    def _tag_is_inside_line_number(self, tag, page_texts):
        """
        True if `tag` appears as a substring of any line-number token on any
        page.  Uses ENRICHMENT_CONFIG['line_no_re'] so line-list tuning
        automatically applies here.
        """
        if not tag or not page_texts:
            return False
        tag_u = tag.strip().upper()
        line_patterns = [re.compile(p, re.IGNORECASE) for p in ENRICHMENT_CONFIG.get("line_no_re", [])]
        for text in page_texts.values():
            if not text:
                continue
            for lp in line_patterns:
                for m in lp.finditer(text):
                    token = m.group(0).upper()
                    # Strip separators/spaces for robust containment check
                    token_clean = re.sub(r"\s+", "", token)
                    if tag_u in token_clean:
                        # Confirm it's INSIDE the line (not the full line itself)
                        if token_clean != tag_u and len(token_clean) > len(tag_u):
                            return True
        return False

    def _is_level_label_like(self, tag):
        """
        Heuristic: tags shaped like LSH-800-300 where the trailing numbers are
        round (multiples of e.g. 50) are usually tank level / nozzle labels
        rather than instrument tags.
        """
        cfg = INSTRUMENT_VALIDATION_CONFIG.get("level_label_filter", {})
        if not cfg.get("enabled"):
            return False
        prefixes = tuple(cfg.get("prefixes", []))
        round_mult = int(cfg.get("round_multiple", 50)) or 50
        tag_u = (tag or "").upper().strip()
        m = re.match(r"^([A-Z]{1,4})-(\d{2,4})-(\d{2,4})$", tag_u)
        if not m:
            return False
        if not m.group(1) in prefixes:
            return False
        try:
            n1 = int(m.group(2))
            n2 = int(m.group(3))
        except ValueError:
            return False
        # Both numbers round multiples → strong level-label signal
        return (n1 % round_mult == 0) and (n2 % round_mult == 0)

    def _validate_tag_format(self, tag):
        """Return True if the tag matches any of the ISA-5.1 universal shapes."""
        cfg = INSTRUMENT_VALIDATION_CONFIG.get("format_validator", {})
        if not cfg.get("enabled"):
            return True
        tag_u = (tag or "").strip().upper()
        for pat in cfg.get("valid_patterns", []):
            if re.match(pat, tag_u):
                return True
        return False

    def _infer_accessories(self, instruments, dn, rev):
        """
        Given a list of extracted instruments, create accessory stub records
        (TE, TW for TT; FE for FT, etc.) per INSTRUMENT_VALIDATION_CONFIG.
        Accessories inherit unit/sequence numbers and are marked inferred=True.
        """
        cfg = INSTRUMENT_VALIDATION_CONFIG.get("accessories", {})
        if not cfg.get("enabled"):
            return []
        amap = cfg.get("map", {})
        dup_guard = cfg.get("dup_guard", True)

        existing_tags = {
            (i.get("tag_number") or "").strip().upper() for i in instruments
        }
        inferred = []
        for inst in list(instruments):
            tag = (inst.get("tag_number") or "").strip().upper()
            m = re.match(r"^([A-Z]{2,6})(-.+)$", tag)
            if not m:
                continue
            prefix = m.group(1)
            suffix = m.group(2)  # e.g. "-3901-01"
            if prefix not in amap:
                continue
            for acc_code in amap[prefix]:
                new_tag = f"{acc_code}{suffix}"
                if dup_guard and new_tag in existing_tags:
                    continue
                if acc_code not in INSTRUMENT_CATEGORIES:
                    continue
                rec = self._make_instrument_record(
                    new_tag, acc_code, INSTRUMENT_CATEGORIES[acc_code],
                    dn, rev, f"Inferred accessory of {tag}"
                )
                rec["inferred"] = True
                rec["parent_tag"] = tag
                rec["service_description"] = (
                    inst.get("service_description") or ""
                )
                rec["line_number"] = inst.get("line_number") or "N/A"
                rec["equipment_number"] = inst.get("equipment_number") or "N/A"
                rec["pid_no"] = inst.get("pid_no") or "N/A"
                rec["drawing_number"] = inst.get("drawing_number") or dn
                rec["loop_number"] = inst.get("loop_number") or self._derive_loop_number(new_tag)
                inferred.append(rec)
                existing_tags.add(new_tag)
        return inferred

    def _apply_smart_field_defaults(self, instruments):
        """
        Validate + intelligently default the Fail-Safe / Signal / Set-Point
        columns using SMART_FIELD_DEFAULTS_CONFIG.  Runs after regex-based
        enrichment so explicit drawing values always win; we only touch a
        field if it is N/A or clearly mis-assigned (e.g. 4-20mA on a local
        gauge, or a pressure set-point on a temperature instrument).
        """
        cfg = SMART_FIELD_DEFAULTS_CONFIG
        if not cfg.get("enabled", True) or not instruments:
            return instruments

        fs_cfg  = cfg.get("fail_safe", {})
        sig_cfg = cfg.get("signal_type", {})
        sp_cfg  = cfg.get("set_point", {})

        fs_prefix_map = fs_cfg.get("by_prefix", {})
        fs_cat_map    = fs_cfg.get("by_category", {})
        valve_prefixes = set(fs_cfg.get("valve_prefixes", set()))

        sig_prefix_map = sig_cfg.get("by_prefix", {})

        sp_unit_hint   = sp_cfg.get("unit_hint_by_category", {})
        sp_hint_letter = sp_cfg.get("unit_hint_by_first_letter", {})
        sp_switches    = set(sp_cfg.get("switch_prefixes", set()))
        sp_cvalves     = set(sp_cfg.get("control_valve_prefixes", set()))
        sp_no_setpt    = set(sp_cfg.get("no_setpoint_prefixes", set()))
        sp_units_cat   = sp_cfg.get("category_units", {})

        cleared_fs = cleared_sig = cleared_sp = 0
        filled_fs = filled_sig = filled_sp = 0

        for inst in instruments:
            tag = (inst.get("tag_number") or "").strip().upper()
            m = re.match(r"^([A-Z]{2,6})", tag)
            prefix = m.group(1) if m else ""
            category = inst.get("category") or ""

            # ── 1) FAIL-SAFE ─────────────────────────────────────────────
            cur_fs = (inst.get("fail_safe") or "").strip()
            is_valve = prefix in valve_prefixes
            if fs_cfg.get("clear_for_non_valve", True) and not is_valve:
                if cur_fs and cur_fs not in ("N/A", "—", "-"):
                    inst["fail_safe"] = "N/A"
                    cleared_fs += 1
                    cur_fs = "N/A"
            if cur_fs in ("", "N/A", None) and is_valve:
                default_fs = fs_prefix_map.get(prefix) or fs_cat_map.get(category)
                if default_fs:
                    inst["fail_safe"] = default_fs
                    filled_fs += 1

            # ── 2) SIGNAL TYPE ───────────────────────────────────────────
            cur_sig = (inst.get("signal_type") or "").strip()
            prefix_default_sig = sig_prefix_map.get(prefix)

            # Validate: override a clearly-wrong value on local-only devices
            if sig_cfg.get("validate_against_prefix", True) and prefix_default_sig:
                looks_local = prefix_default_sig.startswith("Local") or prefix_default_sig == "—"
                if looks_local and cur_sig and cur_sig not in ("N/A", "", "—"):
                    # Only override if current value looks like an electronic signal
                    if re.search(r"(4-20|HART|FIELD|PROFI|DI|DO|DISCRETE|DIGITAL)", cur_sig, re.IGNORECASE):
                        inst["signal_type"] = prefix_default_sig
                        cleared_sig += 1
                        cur_sig = prefix_default_sig

            if cur_sig in ("", "N/A", None) and prefix_default_sig:
                inst["signal_type"] = prefix_default_sig
                filled_sig += 1

            # ── 3) SET POINT ─────────────────────────────────────────────
            cur_sp = (inst.get("set_point") or "").strip()

            # Validate current set-point's unit matches category
            if sp_cfg.get("validate_units_by_category", True) and cur_sp and cur_sp not in ("N/A", "—"):
                expected_units = sp_units_cat.get(category)
                if expected_units:
                    sp_low = cur_sp.lower()
                    if not any(u in sp_low for u in expected_units):
                        inst["set_point"] = "N/A"
                        cleared_sp += 1
                        cur_sp = "N/A"

            if cur_sp in ("", "N/A", None):
                if prefix in sp_switches:
                    inst["set_point"] = sp_cfg.get("switch_default", "Field-adjustable")
                    filled_sp += 1
                elif prefix in sp_no_setpt:
                    inst["set_point"] = sp_cfg.get("no_setpoint_marker", "—")
                    filled_sp += 1
                elif prefix in sp_cvalves:
                    inst["set_point"] = sp_cfg.get("control_valve_default", "Set by DCS loop")
                    filled_sp += 1
                elif category in sp_unit_hint:
                    inst["set_point"] = sp_unit_hint[category]
                    filled_sp += 1
                elif prefix and prefix[0] in sp_hint_letter:
                    inst["set_point"] = sp_hint_letter[prefix[0]]
                    filled_sp += 1

        logger.info(
            f"[SmartDefaults] fail_safe: cleared={cleared_fs} filled={filled_fs} | "
            f"signal: cleared={cleared_sig} filled={filled_sig} | "
            f"set_point: cleared={cleared_sp} filled={filled_sp}"
        )
        return instruments

    # ────────────────────────────────────────────────────────────────────
    # Category-aware tag-format normaliser (soft-coded via INSTRUMENT_TEMPLATES)
    # ────────────────────────────────────────────────────────────────────
    def _apply_tag_format(self, instruments, drawing_info):
        """
        Rewrite every instrument's tag_number into the canonical format
        declared by the active template (e.g. ADNOC Gas → '<UNIT>-<TYPE>-<LOOP>').

        The AI may return tags as 'TT-803', 'PG-31270X-803', '562 FT 1502',
        'FT-1502', etc.  We tokenise on any non-alphanumeric separator,
        classify each token (unit / type / loop) using the template's regexes,
        and reassemble per the template's `pattern`.

        Falls back gracefully:
          • If the template has no `tag_format` → leave tags untouched.
          • If we can't classify all required parts → leave that record alone.
          • If unit is missing AND `unit_from_pid_no` is True → derive from
            drawing_info['pid_no'] (first numeric run of correct length).
        """
        import re as _re

        category = (drawing_info or {}).get("project_category") or "default"
        tpl = get_template(category)
        spec = tpl.get("tag_format")
        if not spec:
            return instruments

        pattern    = spec.get("pattern") or "{unit}-{type}-{loop}"
        unit_re    = _re.compile(spec.get("unit_regex", r"^\d{2,4}$"))
        type_re    = _re.compile(spec.get("type_regex", r"^[A-Z]{1,5}$"))
        loop_re    = _re.compile(spec.get("loop_regex", r"^[A-Z0-9]{1,8}$"))
        loop_strict_re = _re.compile(spec["loop_strict_regex"]) if spec.get("loop_strict_regex") else None
        loop_placeholder_re = _re.compile(spec["loop_placeholder_regex"]) if spec.get("loop_placeholder_regex") else None
        loop_default_placeholder = (spec.get("loop_default_placeholder") or "").strip().upper()
        substitute_nonstd_loop = bool(spec.get("substitute_nonstd_loop", False))
        validate_re    = _re.compile(spec["validate_regex"])    if spec.get("validate_regex")    else None
        drop_invalid   = bool(spec.get("drop_invalid", False))
        flag_invalid   = bool(spec.get("flag_invalid", False))
        unit_from_pid = spec.get("unit_from_pid_no", False)
        unit_min   = int(spec.get("unit_min_len", 3))

        # Derive default unit — priority order:
        #   1. Explicit `project_unit` from the project record (most reliable)
        #   2. Token in `project_code` matching the unit regex (e.g. "562-PID-…")
        #   3. Token / digit run in `pid_no` matching the unit regex
        # ADNOC Gas drawings often have plant numbers like "50196-500-00-30-101"
        # where the leading digits ARE NOT the unit — the unit lives in the
        # title block as "UNIT: 562". So pid_no derivation is the last resort.
        default_unit = ""
        explicit_unit = str(drawing_info.get("project_unit") or "").strip()
        if explicit_unit and unit_re.match(explicit_unit):
            default_unit = explicit_unit
        elif unit_from_pid:
            for src_key in ("project_code", "pid_no", "drawing_number"):
                src = str(drawing_info.get(src_key) or "")
                if not src:
                    continue
                # 1) Prefer a token that fully matches the unit regex
                for tok in _re.split(r"[^A-Za-z0-9]+", src):
                    if tok.isdigit() and unit_re.match(tok):
                        default_unit = tok
                        break
                if default_unit:
                    break
            # 2) Last-ditch fallback: pull the first run of `unit_min` digits
            if not default_unit:
                src = str(drawing_info.get("pid_no") or drawing_info.get("drawing_number") or "")
                m = _re.search(rf"\d{{{unit_min}}}", src)
                if m:
                    default_unit = m.group(0)

        # Counters for diagnostics
        rewritten = 0
        skipped   = 0
        unit_filled = 0
        dropped   = 0
        original_samples = []
        dropped_samples  = []

        kept = []
        for inst in instruments:
            raw = (inst.get("tag_number") or "").strip().upper()
            if not raw or raw in ("N/A", "-", ""):
                skipped += 1
                kept.append(inst)
                continue

            # Tokenise on any separator (dash, space, slash, dot, underscore)
            tokens = [t for t in _re.split(r"[^A-Z0-9]+", raw) if t]
            if not tokens:
                skipped += 1
                kept.append(inst)
                continue

            unit_tok = ""
            type_tok = ""
            loop_toks = []

            for t in tokens:
                # Placeholder tokens (e.g. XXXX) must always be treated as
                # loop digits, never as the TYPE token — even when they
                # technically match the type_regex.
                if loop_placeholder_re and loop_placeholder_re.match(t):
                    loop_toks.append(t)
                    continue
                if not unit_tok and unit_re.match(t):
                    unit_tok = t
                elif not type_tok and type_re.match(t):
                    type_tok = t
                else:
                    loop_toks.append(t)

            # Type is mandatory — if absent, leave the tag alone
            if not type_tok:
                skipped += 1
                kept.append(inst)
                continue

            # Unit fallback from drawing pid_no
            if not unit_tok:
                if default_unit:
                    unit_tok = default_unit
                    unit_filled += 1
                else:
                    skipped += 1
                    kept.append(inst)
                    continue

            loop_part = "-".join(loop_toks) if loop_toks else ""

            # Placeholder loops (XXXX, XXX, XXXXA …) bypass strict numeric
            # validation — these are intentional pre-FEED markers that the
            # engineer fills in later.
            is_placeholder_loop = bool(
                loop_placeholder_re and loop_part and loop_placeholder_re.match(loop_part)
            )

            # Strict loop check (e.g. ADNOC Gas: only 3-4 digits + optional letter)
            if loop_strict_re and loop_part and not is_placeholder_loop and not loop_strict_re.match(loop_part):
                # Try to recover a valid loop fragment from the tokens
                recovered = ""
                for t in loop_toks:
                    if loop_strict_re.match(t):
                        recovered = t
                        break
                if recovered:
                    loop_part = recovered
                elif substitute_nonstd_loop and loop_default_placeholder:
                    # Soft-coded: replace non-standard loop with the
                    # placeholder so the display tag is unit-first canonical
                    # ('FI-803-9' → '803-FI-XXXX'). Preserve original in
                    # the remark for traceability.
                    rk = inst.get("instrument_remark") or ""
                    note = f"Original loop: {loop_part}"
                    inst["instrument_remark"] = (
                        f"{rk} | {note}" if rk and rk != "-" else note
                    ).strip(" |")
                    loop_part = loop_default_placeholder
                    is_placeholder_loop = True
                else:
                    if drop_invalid:
                        dropped += 1
                        if len(dropped_samples) < 5:
                            dropped_samples.append(f"{raw} (loop='{loop_part}')")
                        continue
                    if flag_invalid:
                        rk = inst.get("instrument_remark") or ""
                        inst["instrument_remark"] = (rk + " | NON-STD FORMAT" if rk and rk != "-" else "NON-STD FORMAT").strip(" |")
                    # Even when flagged, still rewrite to unit-first canonical
                    # order so the display always shows '<unit>-<type>-...'.
                    new_tag = pattern.format(unit=unit_tok, type=type_tok, loop=loop_part)
                    if new_tag != raw:
                        if len(original_samples) < 5:
                            original_samples.append(f"{raw} → {new_tag}")
                        cs = (inst.get("control_system_tag") or "").strip().upper()
                        inst["tag_number"] = new_tag
                        if cs == raw:
                            inst["control_system_tag"] = new_tag
                        rewritten += 1
                    skipped += 1
                    kept.append(inst)
                    continue

            if not loop_part:
                # Soft-coded fallback: use the template's placeholder (e.g.
                # 'XXXX' for ADNOC Gas FEED) instead of skipping. This lets
                # 'TT-803' normalise to '803-TT-XXXX' even when the AI did
                # not return a loop number.
                if loop_default_placeholder:
                    loop_part = loop_default_placeholder
                    is_placeholder_loop = True
                else:
                    if drop_invalid:
                        dropped += 1
                        if len(dropped_samples) < 5:
                            dropped_samples.append(f"{raw} (no loop)")
                        continue
                    if flag_invalid:
                        rk = inst.get("instrument_remark") or ""
                        inst["instrument_remark"] = (rk + " | NO LOOP" if rk and rk != "-" else "NO LOOP").strip(" |")
                    skipped += 1
                    kept.append(inst)
                    continue

            new_tag = pattern.format(unit=unit_tok, type=type_tok, loop=loop_part)

            # Final validation against the canonical pattern
            if validate_re and not validate_re.match(new_tag):
                if drop_invalid:
                    dropped += 1
                    if len(dropped_samples) < 5:
                        dropped_samples.append(f"{raw} → {new_tag} (failed validate)")
                    continue
                if flag_invalid:
                    rk = inst.get("instrument_remark") or ""
                    inst["instrument_remark"] = (rk + " | NON-STD FORMAT" if rk and rk != "-" else "NON-STD FORMAT").strip(" |")
                # Still apply the unit-first rewrite so the visual order is correct.
                if new_tag != raw:
                    if len(original_samples) < 5:
                        original_samples.append(f"{raw} → {new_tag}")
                    cs = (inst.get("control_system_tag") or "").strip().upper()
                    inst["tag_number"] = new_tag
                    if cs == raw:
                        inst["control_system_tag"] = new_tag
                    rewritten += 1
                skipped += 1
                kept.append(inst)
                continue

            if new_tag != raw:
                if len(original_samples) < 5:
                    original_samples.append(f"{raw} → {new_tag}")
                inst["tag_number"] = new_tag
                rewritten += 1

                # Mirror to control_system_tag if it pointed to the old tag
                cs = (inst.get("control_system_tag") or "").strip().upper()
                if cs == raw:
                    inst["control_system_tag"] = new_tag

            kept.append(inst)

        logger.info(
            f"[TagFormat] category='{category}' pattern='{pattern}' "
            f"rewritten={rewritten} unit_filled_from_pid={unit_filled} "
            f"skipped={skipped} dropped={dropped} default_unit='{default_unit}' "
            f"samples={original_samples} dropped_samples={dropped_samples}"
        )
        return kept

    # ────────────────────────────────────────────────────────────────────
    # Category-aware template field fill (soft-coded via INSTRUMENT_TEMPLATES)
    # ────────────────────────────────────────────────────────────────────
    def _apply_template_fields(self, instruments, drawing_info, pid_bytes=None):
        """
        Ensure every record has all keys defined by the active template.
        Calls each field's derive(inst) (if provided) to compute a value when
        the AI didn't return one; otherwise falls back to the template default.
        """
        category = (drawing_info or {}).get("project_category") or "default"
        tpl = get_template(category)
        extra = tpl.get("extra_fields") or []
        if not extra:
            logger.info(f"[Template] category='{category}' → no extra fields")
            # Even when the template adds no extra fields, the
            # category-specific style normaliser must still run so
            # ADNOC Onshore tag filtering is never bypassed.
            if category == "adnoc_gas":
                self._apply_adnoc_gas_style(
                    instruments, drawing_info=drawing_info, pid_bytes=pid_bytes
                )
            elif category == "adnoc_onshore":
                self._apply_adnoc_onshore_style(
                    instruments, drawing_info=drawing_info, pid_bytes=pid_bytes
                )
            return instruments

        derived_count = {f["key"]: 0 for f in extra}
        for inst in instruments:
            for f in extra:
                key = f["key"]
                cur = inst.get(key)
                if cur not in (None, "", "N/A", "n/a"):
                    continue
                derive_fn = f.get("derive")
                value = ""
                if callable(derive_fn):
                    try:
                        value = derive_fn(inst) or ""
                    except Exception:
                        value = ""
                if not value:
                    value = f.get("default", "-")
                else:
                    derived_count[key] += 1
                inst[key] = value

        logger.info(
            f"[Template] category='{category}' tpl='{tpl.get('label')}' "
            f"records={len(instruments)} derived={derived_count}"
        )

        # ── Category-specific style normalisation ──────────────────────
        # Each category gets its own isolated branch so edits to one
        # client's pipeline can never bleed into another's. Adding a new
        # category? Add a new `elif` + dedicated `_apply_<name>_style`
        # method below — leave the existing branches untouched.
        if category == "adnoc_gas":
            # ADNOC Gas: rewrite generic ISA codes into the descriptive
            # labels used in the manual "Manual Inst Index" sheet, and
            # align the 'loop_number' column with the paired DCS tag.
            self._apply_adnoc_gas_style(
                instruments, drawing_info=drawing_info, pid_bytes=pid_bytes
            )
        elif category == "adnoc_onshore":
            # ADNOC Onshore: dedicated isolation hook. Currently a no-op
            # so the legacy 15-column extraction is preserved verbatim.
            # Add Onshore-specific normalisation rules here when needed —
            # they cannot leak into `_apply_adnoc_gas_style` or vice
            # versa.
            self._apply_adnoc_onshore_style(
                instruments, drawing_info=drawing_info, pid_bytes=pid_bytes
            )

        return instruments

    # ────────────────────────────────────────────────────────────────────
    # ADNOC ONSHORE — dedicated normaliser hook (currently a no-op).
    # Keep this separate from `_apply_adnoc_gas_style` so the two client
    # pipelines stay independently editable. All Onshore-specific
    # constants should be declared in their own `_ADNOC_ONSHORE_*`
    # module-level block above (mirroring the `_ADNOC_GAS_*` pattern).
    # ────────────────────────────────────────────────────────────────────
    def _apply_adnoc_onshore_style(self, instruments, drawing_info=None, pid_bytes=None):
        # ── Soft-coded ADNOC Onshore tag filter ────────────────────────
        # Manual datasheet shape is `<ISA>-<LOOP>-<PAGE>` (e.g.
        # ``FE-1401-10``). Anything outside that shape — equipment
        # tags, line numbers, drawing IDs, or partial two-segment forms
        # — is dropped here. Soft-coded via `_ADNOC_ONSHORE_TAG_RE` and
        # `_ADNOC_ONSHORE_VALID_ISA`; this method stays a thin wiring
        # layer.
        kept = []
        seen = set()
        rejected_examples = []
        # Resolve unit prefix once per batch (drawing_info hint → fallback).
        default_unit = ""
        try:
            if drawing_info and isinstance(drawing_info, dict):
                default_unit = (
                    str(drawing_info.get("project_unit") or "").strip()
                    or str(drawing_info.get("unit") or "").strip()
                )
        except Exception:
            default_unit = ""
        for inst in instruments:
            if not isinstance(inst, dict):
                continue
            raw_tag = inst.get("tag_number") or inst.get("tag") or ""
            canonical, _isa = _adnoc_onshore_canonicalise_tag(
                raw_tag, default_unit=default_unit or None
            )
            if not canonical:
                if len(rejected_examples) < 6:
                    rejected_examples.append(str(raw_tag)[:40])
                continue
            if canonical in seen:
                continue
            seen.add(canonical)
            inst["tag_number"] = canonical
            kept.append(inst)

        # ── Soft-coded Location fill ──────────────────────────────────
        # Onshore template has no `extra_fields`, so the generic
        # `_derive_location` never runs. Populate `location` here using
        # `_adnoc_onshore_derive_location` (idempotent — never
        # overwrites a value the AI already produced).
        loc_filled = 0
        for inst in kept:
            new_loc = _adnoc_onshore_derive_location(inst)
            if new_loc and new_loc != (inst.get("location") or "").strip():
                inst["location"] = new_loc
                loc_filled += 1
            elif not (inst.get("location") or "").strip():
                inst["location"] = new_loc
                loc_filled += 1

        # ── Soft-coded I/O Type / IS-NIS / System fill ────────────────
        # Onshore datasheet exposes three classifier columns that the AI
        # rarely produces directly. Soft-coded ISA→category maps live in
        # `_ADNOC_ONSHORE_IO_TYPE_BY_ISA`, `_ADNOC_ONSHORE_SYSTEM_*_ISA`,
        # and `_ADNOC_ONSHORE_IS_NIS_*_ISA` near the top of this module.
        # Idempotent: only fills blank/dash values; AI-supplied content
        # is preserved verbatim.
        _BLANK_TOKENS = ("", "-", "—", "N/A", "n/a", "NA", "None", "null")
        io_filled = 0
        sys_filled = 0
        is_nis_filled = 0
        for inst in kept:
            cur_io = (inst.get("io_type") or "").strip()
            if cur_io in _BLANK_TOKENS:
                new_io = _adnoc_onshore_derive_io_type(inst)
                if new_io:
                    inst["io_type"] = new_io
                    io_filled += 1

            cur_sys = (inst.get("system") or "").strip()
            if cur_sys in _BLANK_TOKENS:
                new_sys = _adnoc_onshore_derive_system(inst)
                if new_sys:
                    inst["system"] = new_sys
                    sys_filled += 1

            cur_isn = (inst.get("is_nis") or "").strip()
            if cur_isn in _BLANK_TOKENS:
                new_isn = _adnoc_onshore_derive_is_nis(inst)
                if new_isn:
                    inst["is_nis"] = new_isn
                    is_nis_filled += 1

        # ── Soft-coded Inst Range / Calibration Range fill ────────────
        # Onshore datasheet exposes six range cells per instrument.
        # Population strategy:
        #   1. Scrape design pressure / design temperature / level cues
        #      from the drawing's text once (free, deterministic).
        #   2. For each instrument, resolve via
        #      `_adnoc_onshore_resolve_range` — drawing-derived data
        #      preferred, soft-coded `_ADNOC_ONSHORE_TYPICAL_RANGES`
        #      table is the safety net.
        # Idempotent: any AI-supplied numeric span is preserved verbatim.
        cal_data = {"design_press": None, "design_press_unit": "",
                    "design_temp":  None, "design_temp_unit":  "",
                    "psv_setpoints": []}
        if pid_bytes:
            try:
                pmap = self._build_page_text_map(pid_bytes)
                pdf_blob = "\n".join(pmap.values()) if pmap else ""
                cal_data = _adnoc_extract_drawing_calibration_data(pdf_blob)
                logger.info(
                    "[ADNOC Onshore] cal data: design_press=%s %s "
                    "design_temp=%s %s",
                    cal_data.get("design_press"), cal_data.get("design_press_unit"),
                    cal_data.get("design_temp"),  cal_data.get("design_temp_unit"),
                )
            except Exception as _ce:
                logger.debug(f"[ADNOC Onshore] cal scrape skipped: {_ce}")

        inst_range_filled = 0
        cal_range_filled = 0
        for inst in kept:
            rng = _adnoc_onshore_resolve_range(inst, cal_data)
            if not rng:
                continue
            r_min, r_max, r_unit = rng
            # Inst range
            if (inst.get("inst_range_min") or "").strip() in _BLANK_TOKENS:
                inst["inst_range_min"] = r_min
                inst_range_filled += 1
            if (inst.get("inst_range_max") or "").strip() in _BLANK_TOKENS:
                inst["inst_range_max"] = r_max
                inst_range_filled += 1
            if r_unit and (inst.get("inst_range_unit") or "").strip() in _BLANK_TOKENS:
                inst["inst_range_unit"] = r_unit
                inst_range_filled += 1
            # Calibration range — defaults to the same span; refined
            # downstream when project-specific data is supplied.
            if (inst.get("calibration_min") or "").strip() in _BLANK_TOKENS:
                inst["calibration_min"] = r_min
                cal_range_filled += 1
            if (inst.get("calibration_max") or "").strip() in _BLANK_TOKENS:
                inst["calibration_max"] = r_max
                cal_range_filled += 1
            if r_unit and (inst.get("calibration_unit") or "").strip() in _BLANK_TOKENS:
                inst["calibration_unit"] = r_unit
                cal_range_filled += 1

        # Re-number `index_no` so the surviving rows stay 1..N.
        for i, inst in enumerate(kept, start=1):
            inst["index_no"] = i

        # In-place mutation — the caller ignores the return value.
        instruments.clear()
        instruments.extend(kept)

        logger.info(
            "[ADNOC Onshore] tag-filter: kept=%d unique '<ISA>-<LOOP>-<PAGE>' "
            "instruments; location_filled=%d io_filled=%d system_filled=%d "
            "is_nis_filled=%d inst_range_filled=%d cal_range_filled=%d; "
            "rejected_examples=%s",
            len(kept), loc_filled, io_filled, sys_filled, is_nis_filled,
            inst_range_filled, cal_range_filled, rejected_examples,
        )
        return instruments

    # ────────────────────────────────────────────────────────────────────
    # ADNOC GAS — descriptive style normaliser. Soft-coded via
    # _ADNOC_GAS_INSTRUMENT_TYPE_MAP. Idempotent: only fills/upgrades fields
    # that are missing or still bare ISA codes; never overwrites a richer
    # value the AI already produced.
    # ────────────────────────────────────────────────────────────────────
    def _apply_adnoc_gas_style(self, instruments, drawing_info=None, pid_bytes=None):
        type_map = _ADNOC_GAS_INSTRUMENT_TYPE_MAP
        # Field-only ISA codes whose Loop No. is always '-' in the manual
        FIELD_ONLY = {"FE", "FG", "PG", "PSV", "PSE", "TE", "TG", "TW",
                      "LG", "AE", "RO"}
        upgraded_types = 0
        loop_rewrites = 0

        # ── Soft-coded equipment-tag derivation ────────────────────────
        # When the AI didn't return `equipment_number` for the instruments
        # mounted on the drawing's primary equipment, scan all available
        # text sources for an ADNOC equipment tag (XXX-E-XXXX, XXX-V-XXXX,
        # …) and use the drawing title for the equipment description so
        # the frontend group header renders e.g. "LP STEAM GENERATOR (803-E-XX1)".
        derived_eq_tag, derived_eq_desc = self._derive_adnoc_equipment(
            instruments, drawing_info, pid_bytes=pid_bytes
        )
        eq_filled = 0
        svc_normalised = 0

        # ── Soft-coded calibration-range / PSV-setpoint scrape ────────
        # Reads design pressure, design temperature, and any "SET @ N unit"
        # callouts directly from the drawing's text. Used below to fill
        # calibration_min/max/unit and alarm_hh for PSVs when the AI
        # didn't return those fields.
        cal_data = {"design_press": None, "design_press_unit": "",
                    "design_temp": None,  "design_temp_unit": "",
                    "psv_setpoints": []}
        if pid_bytes:
            try:
                pmap = self._build_page_text_map(pid_bytes)
                pdf_blob = "\n".join(pmap.values()) if pmap else ""
                cal_data = _adnoc_extract_drawing_calibration_data(pdf_blob)
                logger.info(
                    f"[ADNOC-Cal] design_press={cal_data['design_press']} "
                    f"{cal_data['design_press_unit']} | "
                    f"design_temp={cal_data['design_temp']} "
                    f"{cal_data['design_temp_unit']} | "
                    f"psv_setpoints={cal_data['psv_setpoints']}"
                )
            except Exception as ce:
                logger.debug(f"[ADNOC-Cal] scrape skipped: {ce}")

        # ── Soft-coded line-number scrape ─────────────────────────────
        # Mines all canonical ADNOC line numbers (5-part scheme) from the
        # raw PDF text blob. The list is used below to UPGRADE instrument
        # line_number entries that are partial / formatted incorrectly
        # (e.g. AI returned "3/4-803-CHXX1" but the drawing actually says
        # "3/4\"-803-CHXX1-33030X-V"). Pure additive — never overrides a
        # value that's already canonical.
        canonical_lines = []
        if pid_bytes:
            try:
                # Re-use page-text map already built for cal_data when available
                if 'pdf_blob' not in dir() or not pdf_blob:
                    pmap = self._build_page_text_map(pid_bytes)
                    pdf_blob = "\n".join(pmap.values()) if pmap else ""
                canonical_lines = _adnoc_extract_drawing_line_numbers(pdf_blob)
                logger.info(
                    f"[ADNOC-Line] canonical_lines={len(canonical_lines)} "
                    f"sample={canonical_lines[:3]}"
                )
            except Exception as le:
                logger.debug(f"[ADNOC-Line] scrape skipped: {le}")

        cal_filled = 0
        alarm_filled = 0
        psv_idx = 0  # round-robin index over collected PSV setpoints
        location_normalised = 0
        io_filled = 0
        line_filled = 0

        for inst in instruments:
            tag = (inst.get("tag_number") or "").upper()
            # ISA TYPE token = letters between the first two dashes
            m = re.match(r"^[A-Z0-9]+-([A-Z]{1,5})-", tag)
            isa = m.group(1) if m else ""

            # 1) Upgrade instrument_type when AI returned a bare code or generic name.
            #    Variant resolver picks a context-aware label (e.g. FI on a
            #    field rotameter → "Flow Meter (Rotameter)", FE on a
            #    small-bore line → "Flow Element (Integral Orifice)").
            cur_type = (inst.get("instrument_type") or "").strip()
            default_verbose = type_map.get(isa, "")
            verbose = _adnoc_resolve_instrument_type(isa, inst, default_verbose) if isa else ""
            if verbose and (
                not cur_type
                or cur_type in ("-", "N/A", "n/a")
                or cur_type.upper() == isa
                or len(cur_type) <= len(isa) + 2
            ):
                inst["instrument_type"] = verbose
                upgraded_types += 1
            elif verbose and verbose != default_verbose and cur_type == default_verbose:
                # AI returned the generic default but a variant rule applies.
                inst["instrument_type"] = verbose
                upgraded_types += 1

            # 2) Loop No. column = canonical {UNIT}-{CTRL_ISA}-{LOOP_SEQ}
            #    (e.g. FT row → "803-FC-XXXX", PI row → "803-PI-XXXX").
            #    Field-only devices (FE/PG/TG/LG/PSV/…) → "-" per the manual.
            #    The accessor on the frontend reads `control_system_tag` first;
            #    keep both fields in sync for downstream consumers.
            cs_tag = (inst.get("control_system_tag") or "").strip()
            if isa in FIELD_ONLY:
                # Field-only devices: loop = '-' per manual convention
                inst["control_system_tag"] = "-"
                inst["loop_number"] = "-"
                loop_rewrites += 1
            else:
                # Derive the unit prefix from the instrument's own tag
                # (or the drawing's default unit when the tag is partial).
                default_unit = ""
                mu = re.match(r"^(\d{3,4})-", tag)
                if mu:
                    default_unit = mu.group(1)[:3]
                # Best existing source of a sequence: control_system_tag,
                # then loop_number — `_adnoc_normalize_loop_no` extracts
                # the digits and falls back to the placeholder.
                src_loop = cs_tag or (inst.get("loop_number") or "").strip()
                canon = _adnoc_normalize_loop_no(tag, src_loop, default_unit)
                if canon and canon != "-":
                    if (inst.get("loop_number") or "").strip() != canon:
                        loop_rewrites += 1
                    inst["loop_number"] = canon
                    inst["control_system_tag"] = canon

            # 3) Equipment number — fill from the derived drawing equipment
            #    when the AI returned blank/N/A. Also write the description
            #    into a soft-coded helper field the frontend group header
            #    can use directly.
            cur_eq = (inst.get("equipment_number") or "").strip()
            if derived_eq_tag and (not cur_eq or cur_eq.upper() in ("N/A", "NA", "-", "—", "NONE", "NULL")):
                inst["equipment_number"] = derived_eq_tag
                eq_filled += 1
            if derived_eq_desc:
                # Always set this — purely informational, used by the
                # frontend `groupHeaderLabel` for the merged header text.
                inst.setdefault("equipment_description", derived_eq_desc)

            # 4) Service description — manual-style templating.
            #    Policy (per user requirement):
            #      a) If the soft-coded templater can build a phrase from
            #         line_number + equipment + ISA context, USE IT (always
            #         wins — overrides AI).
            #      b) Else if the AI value already looks like a manual
            #         service phrase (mentions a fluid keyword OR an eq tag
            #         OR a property word like "Level"/"Pressure"), keep it
            #         (after title-casing + valve-suffix appendage).
            #      c) Otherwise CLEAR the field to empty — never fall back
            #         to a generic "Process line" / equipment-only phrase.
            cur_svc = (inst.get("service_description") or "").strip()
            built = _adnoc_build_service_description(inst, derived_eq_tag or "", derived_eq_desc or "")
            valve_suffix = _ADNOC_VALVE_SUFFIX_MAP.get(isa, "")

            if built:
                # Templater succeeded → it wins (matches manual exactly).
                if built != cur_svc:
                    inst["service_description"] = built
                    svc_normalised += 1
            else:
                # Decide whether AI value is descriptive enough to keep.
                upper_cur = cur_svc.upper()
                # First pass: drop title-block / system-code boilerplate.
                if _adnoc_service_is_rejected(cur_svc):
                    looks_descriptive = False
                else:
                    looks_descriptive = bool(cur_svc) and cur_svc not in ("-", "N/A", "NA") and (
                        re.search(r"\d{3}-[A-Z]+-[A-Z0-9X]+", upper_cur) or
                        any(w in upper_cur for w in (
                            "PRESSURE", "TEMPERATURE", "LEVEL", "FLOW",
                            "STEAM", "WATER", "GAS", "OIL", "DIESEL",
                            "VENT", "BLOWDOWN", "DRAIN", "FLARE", "SUPERHEATER",
                            "EXCHANGER", "GENERATOR", "DRUM", "VESSEL", "TOWER",
                            "COLUMN", "REACTOR", "PUMP", "COMPRESSOR", "COOLER",
                            "HEATER", "CONDENSER", "REBOILER", "SEPARATOR",
                            "ABSORBER", "STRIPPER", "PHOSPHATE", "MBW", "BFW",
                        ))
                    ) and len(cur_svc) >= 8
                if looks_descriptive:
                    normalised = _adnoc_titlecase_service(cur_svc)
                    if valve_suffix and valve_suffix not in normalised.upper():
                        normalised = f"{normalised} - {valve_suffix}"
                    if normalised != cur_svc:
                        inst["service_description"] = normalised
                        svc_normalised += 1
                else:
                    # ── Last-resort minimal phrase from drawing-derived
                    #     equipment. Not fabricated — uses only the parent
                    #     equipment we already extracted. ISA-aware:
                    #       FT/FE/FI/FV   → "To <eq_tag> <eq_desc>"
                    #       PT/PI/PG      → "<eq_tag> <eq_desc> - Pressure"
                    #       LT/LG/LI      → "<eq_desc> Level"
                    #       TT/TG/TI      → "<eq_desc> Temperature"
                    #     If we don't even have eq context, fall back to
                    #     empty (per the user's "no defaults" rule).
                    minimal = ""
                    if derived_eq_tag and derived_eq_desc:
                        eq_full = f"{derived_eq_tag} {derived_eq_desc}".strip()
                        if isa in ("FT", "FE", "FI", "FV", "FQ"):
                            base = f"To {eq_full}"
                            minimal = f"{base} - {valve_suffix}" if valve_suffix else base
                        elif isa in ("PT", "PI", "PG"):
                            minimal = f"{eq_full} - Pressure"
                        elif isa in ("LT", "LG", "LI", "LIT"):
                            minimal = f"{derived_eq_desc} Level"
                        elif isa in ("TT", "TG", "TE", "TI"):
                            minimal = f"{derived_eq_desc} Temperature"
                        elif isa == "PSV":
                            minimal = f"{eq_full} Relief"
                    if minimal:
                        if minimal != cur_svc:
                            inst["service_description"] = minimal
                            svc_normalised += 1
                    elif cur_svc:
                        # Truly nothing useful → leave blank.
                        inst["service_description"] = ""
                        svc_normalised += 1

            # ── 5) Calibration Range — fill from drawing-derived data ──
            #     Soft-coded per ISA via _ADNOC_CALIBRATION_RULES. Only
            #     overwrites blanks/dashes (never clobbers AI values).
            def _is_blank(v):
                return v is None or str(v).strip() in ("", "-", "N/A", "NA", "—")

            rule = _ADNOC_CALIBRATION_RULES.get(isa)
            if rule:
                kind, lo, hi = rule
                cur_min = inst.get("calibration_min")
                cur_max = inst.get("calibration_max")
                cur_unit = inst.get("calibration_unit")
                target_min, target_max, target_unit = None, None, None

                if kind == "pressure" and cal_data["design_press"]:
                    target_min = lo
                    target_max = cal_data["design_press"]
                    target_unit = cal_data["design_press_unit"]
                elif kind == "temperature" and cal_data["design_temp"]:
                    target_min = lo
                    target_max = cal_data["design_temp"]
                    target_unit = cal_data["design_temp_unit"]
                elif kind == "level":
                    target_min, target_max, target_unit = lo, hi, "%"

                if target_max is not None:
                    if _is_blank(cur_min):
                        inst["calibration_min"] = str(target_min)
                        cal_filled += 1
                    if _is_blank(cur_max):
                        inst["calibration_max"] = str(target_max)
                        cal_filled += 1
                    if _is_blank(cur_unit):
                        inst["calibration_unit"] = target_unit
                        cal_filled += 1

            # PSV — calibration_max + alarm_hh from "SET @ N unit" callouts
            if isa == "PSV" and cal_data["psv_setpoints"]:
                # Distribute set-pressure values across multiple PSVs in
                # document order; if more PSVs than setpoints, last value
                # is reused (drawings typically share a setpoint between A/B).
                set_val, set_unit = cal_data["psv_setpoints"][
                    min(psv_idx, len(cal_data["psv_setpoints"]) - 1)
                ]
                psv_idx += 1
                if _is_blank(inst.get("calibration_min")):
                    inst["calibration_min"] = "0"
                    cal_filled += 1
                if _is_blank(inst.get("calibration_max")):
                    inst["calibration_max"] = str(set_val)
                    cal_filled += 1
                if _is_blank(inst.get("calibration_unit")):
                    inst["calibration_unit"] = set_unit
                    cal_filled += 1
                if _is_blank(inst.get("alarm_hh")):
                    inst["alarm_hh"] = str(set_val)
                    alarm_filled += 1

            # ── 6) Alarm flags — default markers per ISA convention ──
            alarm_defaults = _ADNOC_ALARM_DEFAULT_MARKERS.get(isa, {})
            for alarm_key, marker in alarm_defaults.items():
                if _is_blank(inst.get(alarm_key)):
                    inst[alarm_key] = marker
                    alarm_filled += 1

            # ── 7) Location — canonical Field / Vessel / Local Panel ──
            #     Soft-coded via _ADNOC_VESSEL_MOUNTED_ISA + line_number hint.
            new_loc = _adnoc_resolve_location(inst, isa)
            if (inst.get("location") or "").strip() != new_loc:
                inst["location"] = new_loc
                location_normalised += 1

            # ── 8) I/O Type — canonical AI / AI-R / AO-R / DI / DO / DO-R ──
            #     Soft-coded via _ADNOC_IO_TYPE_BY_ISA + _ADNOC_IO_LOCAL_ISA.
            cur_io = (inst.get("io_type") or "").strip()
            if not cur_io or cur_io in ("-", "N/A", "NA"):
                new_io = _adnoc_resolve_io_type(inst, isa)
                if new_io:
                    inst["io_type"] = new_io
                    io_filled += 1

            # ── 9) Line Number — upgrade partials to canonical ADNOC ──
            #     Soft-coded via _adnoc_normalise_line_number(): cleans
            #     whitespace, restores inch glyph, and upgrades partial
            #     hits (e.g. "3/4-803-CHXX1") to the full canonical
            #     "3/4\"-803-CHXX1-33030X-V" found elsewhere in the PDF.
            new_line = _adnoc_normalise_line_number(
                inst.get("line_number"), canonical_lines
            )
            if new_line and new_line != (inst.get("line_number") or "").strip():
                inst["line_number"] = new_line
                line_filled += 1

        logger.info(
            f"[ADNOC-Gas style] upgraded_types={upgraded_types} "
            f"loop_rewrites={loop_rewrites} eq_filled={eq_filled} "
            f"svc_normalised={svc_normalised} "
            f"cal_filled={cal_filled} alarm_filled={alarm_filled} "
            f"location_normalised={location_normalised} "
            f"io_filled={io_filled} "
            f"line_filled={line_filled} "
            f"derived_eq_tag='{derived_eq_tag}' derived_eq_desc='{derived_eq_desc}' "
            f"records={len(instruments)}"
        )

    def _derive_adnoc_equipment(self, instruments, drawing_info, pid_bytes=None):
        """
        Soft-coded equipment-tag + description derivation for ADNOC-style
        P&IDs. Returns ``(eq_tag, eq_description)``; either may be empty.

        Strategy (priority order):
          1. Most-frequent equipment-pattern hit across all instrument
             text fields (line_number, service_description, notes,
             equipment_number itself).
          2. Single hit anywhere in those fields.
          3. Hit inside the drawing title.

        For the description:
          1. Match against `_ADNOC_GAS_EQUIPMENT_NOUN_RE` over (a) the
             drawing title and (b) every instrument's service description.
          2. Fall back to the noun mapped to the equipment-type letter
             (e.g. 'E' → 'Exchanger / Generator').

        Works for any P&ID using the ADNOC equipment-tag convention —
        the constants drive everything.
        """
        if not instruments and not drawing_info:
            return ("", "")

        drawing_info = drawing_info or {}
        title = str(drawing_info.get("drawing_title") or "")

        # ── Collect candidate text from every instrument record ──
        text_blobs: list[str] = [title]
        for inst in instruments:
            for k in ("equipment_number", "line_number",
                      "service_description", "notes",
                      "instrument_remark"):
                v = inst.get(k)
                if v:
                    text_blobs.append(str(v))

        # ── Append PDF page text (title block, equipment list table) ──
        pdf_text_blob = ""
        if pid_bytes:
            try:
                page_map = self._build_page_text_map(pid_bytes)
                pdf_text_blob = "\n".join(page_map.values())
                if pdf_text_blob:
                    text_blobs.append(pdf_text_blob)
            except Exception as e:
                logger.debug(f"[ADNOC-Eq] PDF text scan skipped: {e}")

        haystack = "  ".join(text_blobs).upper()

        # ── Find most common equipment tag matching ADNOC pattern ──
        from collections import Counter
        counts: Counter = Counter()
        for m in _ADNOC_GAS_EQUIPMENT_TAG_RE.finditer(haystack):
            tag = m.group(0)
            counts[tag] += 1

        eq_tag = ""
        eq_letter = ""
        if counts:
            eq_tag, _freq = counts.most_common(1)[0]
            mtl = re.match(r"^\d{3}-([A-Z])-", eq_tag)
            eq_letter = mtl.group(1) if mtl else ""

        # ── Description: noun phrase from title or service text ──
        eq_desc = ""
        noun_sources = [title]
        noun_sources.extend((inst.get("service_description") or "") for inst in instruments)
        if pdf_text_blob:
            # Limit to first 4000 chars so the title-block region wins out
            # over body line callouts.
            noun_sources.append(pdf_text_blob[:4000])
        for src in noun_sources:
            if not src:
                continue
            mn = _ADNOC_GAS_EQUIPMENT_NOUN_RE.search(str(src))
            if mn:
                eq_desc = mn.group(0).strip()
                break
        if not eq_desc and eq_letter:
            eq_desc = _ADNOC_GAS_EQUIPMENT_TYPE_NOUN.get(eq_letter, "")

        # Title-case for display ("LP STEAM GENERATOR" → upper-case kept;
        # frontend already upper-cases via groupHeaderLabel anyway).
        return (eq_tag, eq_desc)

    def _apply_post_validation(self, instruments, pid_bytes, drawing_info):
        """
        Orchestrate all post-processing fixes.  Pure add-on — never modifies
        extraction/regex behaviour.  Safe to disable via
        INSTRUMENT_VALIDATION_CONFIG['enabled'] = False.
        """
        if not INSTRUMENT_VALIDATION_CONFIG.get("enabled", True):
            return instruments
        if not instruments:
            return instruments

        dn = drawing_info.get("drawing_number", "")
        rev = drawing_info.get("revision", "0")

        # 1) Per-page P&ID number stamping
        ppg_cfg = INSTRUMENT_VALIDATION_CONFIG.get("per_page_pid_no", {})
        if ppg_cfg.get("enabled", True):
            pid_per_page = self._extract_per_page_pid_numbers(pid_bytes)
            if pid_per_page:
                for inst in instruments:
                    page = inst.get("page") or inst.get("page_number")
                    resolved = pid_per_page.get(page)
                    if resolved:
                        inst["pid_no"] = resolved
                    elif ppg_cfg.get("fallback_to_input", True):
                        inst.setdefault("pid_no", drawing_info.get("pid_no") or dn)

        # Pre-compute the page→text map once for the line-context filter
        page_texts = self._build_page_text_map(pid_bytes)

        # 2) Line-context filter  +  3) Level-label filter  +  4) Format check
        lc_cfg = INSTRUMENT_VALIDATION_CONFIG.get("line_context_filter", {})
        lvl_cfg = INSTRUMENT_VALIDATION_CONFIG.get("level_label_filter", {})
        filtered = []
        dropped = 0
        for inst in instruments:
            tag = (inst.get("tag_number") or "").strip().upper()
            warnings = list(inst.get("warnings") or [])

            # 2) inside line-number?
            if lc_cfg.get("enabled", True) and self._tag_is_inside_line_number(tag, page_texts):
                dropped += 1
                logger.info(
                    f"[Validator] ⛔ drop '{tag}' — appears inside line number token"
                )
                continue

            # 3) tank-level label?
            if self._is_level_label_like(tag):
                if lvl_cfg.get("action", "warn") == "drop":
                    dropped += 1
                    logger.info(
                        f"[Validator] ⛔ drop '{tag}' — tank-level label, not instrument"
                    )
                    continue
                warnings.append("Possible tank-level / nozzle label — verify manually")

            # 4) ISA-5.1 universal format check
            if not self._validate_tag_format(tag):
                warnings.append("Tag does not match ISA-5.1 Unit-Tag-Sequence format")
                inst["format_valid"] = False
            else:
                inst["format_valid"] = True

            if warnings:
                inst["warnings"] = warnings
            filtered.append(inst)

        if dropped:
            logger.info(f"[Validator] Dropped {dropped} tag(s) via line-context / level-label filters")

        # 5) Accessory auto-inference
        inferred = self._infer_accessories(filtered, dn, rev)
        if inferred:
            logger.info(f"[Validator] ➕ {len(inferred)} accessory instruments inferred")
            filtered.extend(inferred)

        # 5b) Smart defaults for Fail-Safe / Signal / Set-Point
        filtered = self._apply_smart_field_defaults(filtered)

        # 6) Inline-instrument priority re-sort
        ip_cfg = INSTRUMENT_VALIDATION_CONFIG.get("inline_priority", {})
        if ip_cfg.get("enabled", True):
            inline_codes = set(c.upper() for c in ip_cfg.get("codes", []))
            def _inline_key(inst):
                tag = (inst.get("tag_number") or "").upper()
                m = re.match(r"^([A-Z]{2,6})", tag)
                code = m.group(1) if m else ""
                is_inline = code in inline_codes
                inst["is_inline"] = is_inline
                # sort: inline first (0), then others (1); stable within group
                return (0 if is_inline else 1,)
            filtered.sort(key=_inline_key)

        # Rebuild sequential index numbers
        for i, inst in enumerate(filtered, start=1):
            inst["index_no"] = i

        return filtered

    def _enrich_from_pdf_context(self, instruments, pid_bytes):
        """
        Post-extraction contextual enrichment using the PDF text layer.

        Fills N/A fields without consuming any AI quota:
          loop_number         — derived from the tag itself (deterministic)
          service_description — category verb + area code
          line_number         — piping line tag matched by area code + spatial proximity
          equipment_number    — equipment tag matched by area code + spatial proximity
          fail_safe           — FC / FO / FL from spatial proximity
          signal_type         — 4-20mA / HART / Discrete from proximity + global scan
          set_point           — value+unit from spatial proximity

        Four-layer strategy:
          L1: Deterministic  — loop_number + default service_description (always).
          L2: Block scan     — joins block lines to reconstruct multi-span line nos.
          L3: Spatial        — finds line nos, equip tags, fail-safe, signal, setpoint
                               near each instrument's position on the page.
          L4: Area-code      — matches globally scanned line nos / equip tags by the
                               shared numeric area code (with fuzzy ±tolerance).
        """
        ec = ENRICHMENT_CONFIG

        # Pre-compile all enrichment patterns (once per call)
        line_no_pats = [re.compile(p, re.IGNORECASE) for p in ec["line_no_re"]]
        line_no_structured_rules = [
            {
                "name": rule["name"],
                "pattern": re.compile(rule["pattern"], re.IGNORECASE),
            }
            for rule in ec.get("line_no_structured_rules", [])
        ]
        equip_pats   = [re.compile(p, re.IGNORECASE) for p in ec["equipment_re"]]
        fail_pats    = [re.compile(p, re.IGNORECASE) for p in ec["fail_safe_re"]]
        signal_pats  = [re.compile(p, re.IGNORECASE) for p in ec["signal_re"]]
        sp_pats      = [re.compile(p, re.IGNORECASE) for p in ec["set_point_re"]]

        radius         = ec["spatial_radius"]
        ctx_radius     = ec.get("spatial_radius_context", 350)
        area_tolerance = ec.get("area_tolerance", 5)
        stopwords      = ec["exclude_desc_words"]
        _area_re       = re.compile(r'(\d{3,6})')

        def _first_match(text, pats):
            text = _normalize_engineering_text(text)
            for p in pats:
                m = p.search(text)
                if m:
                    return _normalize_context_match(m.group(1) if m.lastindex else m.group(0))
            return None

        def _all_matches(text, pats):
            text = _normalize_engineering_text(text)
            out = []
            for p in pats:
                out.extend(
                    _normalize_context_match(m.group(1) if m.lastindex else m.group(0))
                    for m in p.finditer(text)
                )
            return out

        def _normalize_engineering_text(text):
            if not text:
                return ''
            cleaned = str(text)
            cleaned = (cleaned
                .replace('\u2013', '-')
                .replace('\u2014', '-')
                .replace('_', '-')
                .replace('\u201c', '"')
                .replace('\u201d', '"')
                .replace('\u2033', '"')
                .replace('\u02BA', '"')
                .replace('\uFF02', '"')
                .replace('\u2018', "'")
                .replace('\u2019', "'"))
            cleaned = re.sub(r'\s*[-]\s*', '-', cleaned)
            cleaned = re.sub(r'\s+', ' ', cleaned)
            return cleaned.strip()

        def _normalize_context_match(value):
            if not value:
                return ''
            normalized = _normalize_engineering_text(value).upper()
            normalized = re.sub(r'(?<=\d)\s*(?:"|\')', '"', normalized)
            normalized = re.sub(r'"{2,}', '"', normalized)
            return normalized.strip('- ').strip()

        def _canonicalize_line_number(size, *parts):
            size_part = _normalize_context_match(size)
            if size_part and '"' not in size_part:
                size_part = f'{size_part}"'
            ordered_parts = [size_part]
            ordered_parts.extend(_normalize_context_match(part) for part in parts if part)
            return "-".join(part for part in ordered_parts if part)

        def _extract_line_numbers(text):
            candidates = _all_matches(text, line_no_pats)
            normalized_text = _normalize_engineering_text(text)

            for rule in line_no_structured_rules:
                for match in rule["pattern"].finditer(normalized_text):
                    groups = [grp.strip() for grp in match.groups() if grp is not None]
                    if rule["name"] == "five_part_line" and len(groups) == 5:
                        line_no = _canonicalize_line_number(
                            groups[0], groups[1], groups[2], groups[3], groups[4]
                        )
                        if line_no:
                            candidates.append(line_no)
                    elif rule["name"] == "six_part_line" and len(groups) >= 5:
                        line_no = _canonicalize_line_number(*groups)
                        if line_no:
                            candidates.append(line_no)

            return list(dict.fromkeys(c for c in candidates if c))

        def _pick_best_line_number(candidates, area_hint=''):
            if not candidates:
                return None
            if area_hint:
                exact = [cand for cand in candidates if _area(cand) == area_hint]
                if exact:
                    return exact[0]
                fuzzy = _fuzzy_lookup(area_hint, { _area(c): [c] for c in candidates if _area(c) })
                if fuzzy:
                    return fuzzy[0]
            return candidates[0]

        def _area(s):
            m = _area_re.search(s or '')
            return m.group(1) if m else ''

        def _fuzzy_lookup(area_key, by_area_dict):
            """Return candidates for area_key with fuzzy numeric tolerance."""
            if not area_key:
                return []
            if area_key in by_area_dict:
                return by_area_dict[area_key]
            # Numeric proximity search within tolerance
            try:
                target = int(area_key)
                best, best_dist = None, area_tolerance + 1
                for k in by_area_dict:
                    try:
                        d = abs(int(k) - target)
                        if d <= area_tolerance and d < best_dist:
                            best_dist, best = d, k
                    except ValueError:
                        pass
                return by_area_dict[best] if best else []
            except ValueError:
                return []

        def _build_join_variants(tokens, max_span=4):
            clean_tokens = [_normalize_engineering_text(tok) for tok in tokens if _normalize_engineering_text(tok)]
            if not clean_tokens:
                return []

            variants = [
                " ".join(clean_tokens),
                "-".join(clean_tokens),
                "".join(clean_tokens),
            ]

            for span in range(2, min(max_span, len(clean_tokens)) + 1):
                dash_chunks = []
                plain_chunks = []
                spaced_chunks = []
                for idx in range(len(clean_tokens) - span + 1):
                    chunk = clean_tokens[idx:idx + span]
                    dash_chunks.append("-".join(chunk))
                    plain_chunks.append("".join(chunk))
                    spaced_chunks.append(" ".join(chunk))
                variants.append(" ".join(dash_chunks))
                variants.append(" ".join(plain_chunks))
                variants.append(" ".join(spaced_chunks))

            return list(dict.fromkeys(v for v in variants if v))

        def _reconstruct_line_no(word_tokens, area_hint=''):
            """
            Reconstruct piping line number from a sequence of word tokens.

            Piping line format: {size}["]-{fluid}-{area}-{spec}
            The PDF may split "16"" and "HC-3901-A2A" into separate word tokens.
            This function tries to stitch them back together.

            Handles tokens like:
              ["16\"", "HC-3901-A2A"]       → "16-HC-3901-A2A"
              ["16\"", "HC", "3901", "A2A"] → "16-HC-3901-A2A"
              ["16", "HC-3901-A2A"]         → "16-HC-3901-A2A"
            """
            _SIZE  = re.compile(r'^\d{1,3}(?:["″\u2033\u02BA\'\u2019]{1,2})?$')
            _FLUID = re.compile(r'^[A-Z]{1,6}$', re.IGNORECASE)
            _AREA  = re.compile(r'^\d{3,6}$')
            _SPEC  = re.compile(r'^[A-Z][0-9][A-Z0-9]{0,3}$', re.IGNORECASE)
            # A combined token like "HC-3901-A2A" (fluid + rest)
            _FLUID_REST = re.compile(
                r'^([A-Z]{1,6})[-](\d{3,6})[-]([A-Z][0-9][A-Z0-9]{0,3})$', re.IGNORECASE
            )
            _FLUID_AREA = re.compile(r'^([A-Z]{1,6})[-](\d{3,6})$', re.IGNORECASE)

            words = [w.strip() for w in word_tokens if w.strip()]

            for i, w in enumerate(words):
                clean_size = re.sub(r'["″\u2033\u02BA\'\u2019]+', '', w)
                if not _SIZE.match(w) and not re.match(r'^\d{1,3}$', clean_size):
                    continue
                rest = words[i + 1:i + 5]  # look ahead up to 4 tokens

                for j, rw in enumerate(rest):
                    # Case A: next word is "fluid-area-spec"
                    m = _FLUID_REST.match(rw)
                    if m:
                        fluid, area_code, spec = m.group(1), m.group(2), m.group(3)
                        if area_hint and not area_code.startswith(area_hint[:2]):
                            continue
                        return f"{clean_size}-{fluid.upper()}-{area_code}-{spec.upper()}"

                    # Case B: next word is "fluid-area"
                    m2 = _FLUID_AREA.match(rw)
                    if m2:
                        fluid, area_code = m2.group(1), m2.group(2)
                        if area_hint and not area_code.startswith(area_hint[:2]):
                            continue
                        # Look for spec in the word after
                        if j + 1 < len(rest) and _SPEC.match(rest[j + 1]):
                            return f"{clean_size}-{fluid.upper()}-{area_code}-{rest[j+1].upper()}"
                        return f"{clean_size}-{fluid.upper()}-{area_code}"

                    # Case C: fluid/area/spec in separate tokens
                    if not _FLUID.match(rw):
                        continue
                    fluid = rw.upper()
                    remaining = rest[j + 1:]
                    for k_idx, kw in enumerate(remaining):
                        if not _AREA.match(kw):
                            continue
                        if area_hint and not kw.startswith(area_hint[:2]):
                            continue
                        area_code = kw
                        after = remaining[k_idx + 1:]
                        if after and _SPEC.match(after[0]):
                            return f"{clean_size}-{fluid}-{area_code}-{after[0].upper()}"
                        return f"{clean_size}-{fluid}-{area_code}"
            return None

        # ── L1: Deterministic (loop_number + default service description + CS tag ISA) ─
        for inst in instruments:
            tag = inst.get("tag_number", "")
            if not inst.get("loop_number") or inst["loop_number"] == "N/A":
                inst["loop_number"] = self._derive_loop_number(tag)
            if not inst.get("service_description"):
                inst["service_description"] = self._infer_service_description(
                    tag, inst.get("instrument_type", ""), inst.get("category", "")
                )
            # CS Tag L1: if the instrument IS a DCS controller, mark it immediately
            if inst.get("control_system_tag") in ("N/A", "", None):
                _, is_dcs = self._derive_cs_tag_isa(tag)
                if is_dcs:
                    inst["control_system_tag"] = tag.upper()

        # ── Open PDF ──────────────────────────────────────────────────────
        try:
            import fitz
        except ImportError:
            logger.warning("[Enrich] PyMuPDF not installed — skipping contextual enrichment")
            return instruments

        try:
            doc = fitz.open(stream=pid_bytes, filetype="pdf")
        except Exception as ex:
            logger.warning(f"[Enrich] Cannot open PDF for enrichment: {ex}")
            return instruments

        global_line_nos: list = []
        global_equip:    list = []
        global_signals:  list = []
        global_sps:      list = []
        all_words_pages: list = []

        # ── L2: Block-level scan ─────────────────────────────────────────
        # Plain text alone misses line numbers split across spans (e.g. "16""
        # as one span, "HC-3901-A2A" as the next). Block reconstruction joins
        # them with a dash, matching: 16-HC-3901-A2A → line_no regex.
        for pno in range(len(doc)):
            pg = doc[pno]

            # 2a: Full plain text
            txt = _normalize_engineering_text(pg.get_text("text") or "")
            global_line_nos.extend(_extract_line_numbers(txt))
            global_equip.extend(_all_matches(txt, equip_pats))
            global_signals.extend(_all_matches(txt, signal_pats))
            global_sps.extend(_all_matches(txt, sp_pats))

            # 2b: Word-level pair/triple joining (catches split tokens on same line)
            try:
                words_raw = pg.get_text("words") or []
                all_words_pages.append(words_raw)
                for wi_r in range(len(words_raw)):
                    for span_len in (2, 3, 4):
                        if wi_r + span_len > len(words_raw):
                            break
                        # Only join words on roughly the same text line (y-delta < 12 pts)
                        y_start = words_raw[wi_r][1]
                        y_end   = words_raw[wi_r + span_len - 1][3]
                        if abs(y_end - y_start) > 12:
                            break
                        for joiner in ("-", ""):
                            chunk = _normalize_engineering_text(joiner.join(
                                words_raw[wi_r + s][4].strip()
                                for s in range(span_len)
                            ))
                            global_line_nos.extend(_extract_line_numbers(chunk))
                            global_equip.extend(_all_matches(chunk, equip_pats))
            except Exception:
                if len(all_words_pages) <= pno:
                    all_words_pages.append([])

            # 2c: Block-line reconstruction (joins block's internal lines)
            try:
                page_dict = pg.get_text("dict")
                for block in page_dict.get("blocks", []):
                    if block.get("type") != 0:
                        continue
                    block_lines = []
                    for ln in block.get("lines", []):
                        lt = "".join(s.get("text", "") for s in ln.get("spans", [])).strip()
                        if lt:
                            block_lines.append(lt)
                    for joiner in ("-", " ", ""):
                        combined = _normalize_engineering_text(joiner.join(block_lines))
                        global_line_nos.extend(_extract_line_numbers(combined))
                        global_equip.extend(_all_matches(combined, equip_pats))
                    # Also try _reconstruct_line_no on block tokens
                    reconstructed = _reconstruct_line_no(block_lines)
                    if reconstructed:
                        global_line_nos.append(reconstructed)
            except Exception:
                pass

        # Deduplicate while preserving order
        global_line_nos = list(dict.fromkeys(ln.upper() for ln in global_line_nos if ln))
        global_equip    = list(dict.fromkeys(eq.upper() for eq in global_equip if eq))
        global_signals  = list(dict.fromkeys(global_signals))
        global_sps      = list(dict.fromkeys(global_sps))

        # Filter equipment: remove anything that IS an instrument prefix
        global_equip = [
            eq for eq in global_equip
            if not self._match_instrument_code(eq.split('-')[0].upper())[0]
        ]

        logger.info(
            f"[Enrich] Global scan: {len(global_line_nos)} line nos, "
            f"{len(global_equip)} equip tags, "
            f"{len(global_signals)} signal types, "
            f"{len(global_sps)} set points"
        )

        # ── L2 area-code lookup tables (built from global scan) ──────────
        ln_by_area: dict = {}
        for ln in global_line_nos:
            a = _area(ln)
            if a:
                ln_by_area.setdefault(a, []).append(ln)

        eq_by_area: dict = {}
        for eq in global_equip:
            a = _area(eq)
            if a:
                eq_by_area.setdefault(a, []).append(eq)

        # ── L3: Spatial proximity enrichment ─────────────────────────────
        # Locate each instrument tag in the page word-list, then search
        # within ctx_radius pixels for line numbers, equipment tags,
        # fail-safe annotations, signal types, and set points.
        tag_to_inst: dict = {}
        for inst in instruments:
            n = self._normalize_tag(inst.get("tag_number", ""))
            if n:
                tag_to_inst[n] = inst

        _valve_categories = {
            "Shutdown & ESD", "Control Valves", "Motor & Solenoid",
            "Flow", "Pressure", "Level", "Temperature",
        }

        for pno, words in enumerate(all_words_pages):
            for wi in range(len(words)):
                for span in range(1, 4):
                    if wi + span > len(words):
                        break
                    chunk_dash  = "-".join(words[wi + s][4].strip() for s in range(span))
                    chunk_plain = "".join(words[wi + s][4].strip() for s in range(span))
                    norm = self._normalize_tag(chunk_dash)
                    if norm not in tag_to_inst:
                        norm = self._normalize_tag(chunk_plain)
                    if norm not in tag_to_inst:
                        continue

                    inst = tag_to_inst[norm]
                    x0 = words[wi][0];            y0 = words[wi][1]
                    x1 = words[wi + span - 1][2]; y1 = words[wi + span - 1][3]
                    px = (x0 + x1) / 2;           py = (y0 + y1) / 2
                    tag_area = inst.get("loop_number", "") or _area(inst.get("tag_number", ""))

                    # Collect ALL nearby words with their positions (using larger radius)
                    nearby_words_sorted: list = []  # (dist, text, x, y)
                    for nw in words:
                        nt = nw[4].strip()
                        if not nt or len(nt) < 2:
                            continue
                        nx = (nw[0] + nw[2]) / 2
                        ny = (nw[1] + nw[3]) / 2
                        d  = ((nx - px) ** 2 + (ny - py) ** 2) ** 0.5
                        if d < ctx_radius:
                            nearby_words_sorted.append((d, nt, nx, ny))
                    nearby_words_sorted.sort()
                    nearby = [w[1] for w in nearby_words_sorted]
                    nearby_layout = [
                        w[1] for w in sorted(
                            nearby_words_sorted,
                            key=lambda item: (round(item[3] / 10), item[2])
                        )
                    ]
                    anchor_start = max(0, wi - 12)
                    anchor_end = min(len(words), wi + span + 12)
                    anchor_window = [
                        words[idx][4].strip()
                        for idx in range(anchor_start, anchor_end)
                        if words[idx][4].strip()
                    ]
                    nearby_variants = _build_join_variants(nearby, max_span=4)
                    nearby_layout_variants = _build_join_variants(nearby_layout, max_span=4)
                    anchor_variants = _build_join_variants(anchor_window, max_span=5)

                    # — Line number —
                    if inst.get("line_number") in ("N/A", "", None):
                        ln_found = None
                        for candidate_text in nearby_layout_variants + anchor_variants + nearby_variants:
                            ln_found = _pick_best_line_number(
                                _extract_line_numbers(candidate_text),
                                tag_area,
                            )
                            if ln_found:
                                break
                        if not ln_found:
                            ln_found = (
                                _reconstruct_line_no(nearby_layout, tag_area)
                                or _reconstruct_line_no(anchor_window, tag_area)
                                or _reconstruct_line_no(nearby, tag_area)
                            )
                        if ln_found:
                            ln_found = _normalize_context_match(ln_found)
                            inst["line_number"] = ln_found
                            # Immediately feed back into lookup table
                            la = _area(ln_found)
                            if la:
                                ln_by_area.setdefault(la, [])
                                if ln_found not in ln_by_area[la]:
                                    ln_by_area[la].append(ln_found)

                    # — Equipment number —
                    if inst.get("equipment_number") in ("N/A", "", None):
                        eq_candidates = []
                        for candidate_text in nearby_layout_variants + anchor_variants + nearby_variants:
                            eq_candidates.extend(_all_matches(candidate_text, equip_pats))
                        for eq_c in eq_candidates:
                            code = eq_c.split('-')[0].upper()
                            if not self._match_instrument_code(code)[0]:
                                inst["equipment_number"] = _normalize_context_match(eq_c)
                                ea = _area(eq_c)
                                if ea:
                                    eq_by_area.setdefault(ea, [])
                                    normalized_eq = _normalize_context_match(eq_c)
                                    if normalized_eq not in eq_by_area[ea]:
                                        eq_by_area[ea].append(normalized_eq)
                                break

                    # — Fail-safe (control/shutdown valves only) —
                    if inst.get("fail_safe") in ("N/A", "", None):
                        if inst.get("category") in _valve_categories:
                            fs_str = " ".join(w[1] for w in nearby_words_sorted if w[0] < radius)
                            fs = _first_match(fs_str, fail_pats)
                            if fs:
                                fsu = fs.upper()
                                if "CLOSE" in fsu:
                                    fs = "FC"
                                elif "OPEN" in fsu:
                                    fs = "FO"
                                elif "LAST" in fsu or "LOCK" in fsu:
                                    fs = "FL"
                                inst["fail_safe"] = fs.upper()[:3]

                    # — Signal type —
                    if inst.get("signal_type") in ("N/A", "", None):
                        sig_str = " ".join(w[1] for w in nearby_words_sorted if w[0] < radius)
                        sig = _first_match(sig_str, signal_pats)
                        if sig:
                            sigu = sig.upper()
                            if "4" in sigu and "20" in sigu:
                                sig = "4-20mA"
                            elif "HART" in sigu:
                                sig = "HART"
                            elif "FIELD" in sigu:
                                sig = "Fieldbus"
                            elif "PROFI" in sigu:
                                sig = "Profibus"
                            elif "PNEUM" in sigu:
                                sig = "Pneumatic"
                            elif any(x in sigu for x in ("DISCRETE", "ON/OFF", "ON-OFF")):
                                sig = "Discrete (0/1)"
                            elif sigu in ("AO",):
                                sig = "4-20mA"
                            elif sigu in ("DO",):
                                sig = "Discrete (0/1)"
                            inst["signal_type"] = sig

                    # — Set point —
                    if inst.get("set_point") in ("N/A", "", None):
                        sp_str = " ".join(w[1] for w in nearby_words_sorted if w[0] < radius)
                        sp = _first_match(sp_str, sp_pats)
                        if sp:
                            inst["set_point"] = sp.strip()

                    # — Control System Tag (L3: proximity label scan) —
                    if inst.get("control_system_tag") in ("N/A", "", None):
                        _cs_context = " ".join(nearby_layout[:40])
                        for _cs_pat in [re.compile(p, re.IGNORECASE)
                                        for p in CS_TAG_CONFIG["label_patterns"]]:
                            _cs_m = _cs_pat.search(_cs_context)
                            if _cs_m:
                                inst["control_system_tag"] = _cs_m.group(1).upper()
                                break

                    # — Service description: enrich default with nearby context —
                    current_desc = inst.get("service_description", "")
                    is_default = current_desc.startswith(
                        _SERVICE_VERB_MAP.get(inst.get("category", ""), "~NONE~")
                    )
                    if is_default:
                        desc_candidates = [
                            w for w in nearby[:15]
                            if len(w) > 2
                            and w.lower() not in stopwords
                            and not re.match(r'^\d+$', w)
                            and not self._match_instrument_code(w.upper())[0]
                            and not _first_match(w, line_no_pats + equip_pats + sp_pats)
                        ]
                        if len(desc_candidates) >= ec["desc_min_words"]:
                            inst["service_description"] = " ".join(desc_candidates[:5]).title()

                    break  # tag located — stop trying longer spans

        doc.close()

        # ── L4: Area-code matching (fallback for spatially unlocated tags) ─
        for inst in instruments:
            tag  = inst.get("tag_number", "")
            loop = inst.get("loop_number", "N/A")
            area = loop if loop != "N/A" else _area(tag)

            if inst.get("line_number") in ("N/A", "", None) and area:
                candidates = _fuzzy_lookup(area, ln_by_area)
                if candidates:
                    inst["line_number"] = candidates[0]

            if inst.get("equipment_number") in ("N/A", "", None) and area:
                candidates = _fuzzy_lookup(area, eq_by_area)
                if candidates:
                    inst["equipment_number"] = candidates[0]

            # Category-aware default signal type from global document signals
            if inst.get("signal_type") in ("N/A", "", None) and global_signals:
                cat = inst.get("category", "")
                for sig_raw in global_signals:
                    sigu = sig_raw.upper()
                    if "4" in sigu and "20" in sigu:
                        if cat in ("Flow", "Pressure", "Temperature", "Level",
                                   "Differential Pressure", "Analysis"):
                            inst["signal_type"] = "4-20mA"
                            break
                    if any(x in sigu for x in ("DISCRETE", "DI", "DO")):
                        if cat in ("Shutdown & ESD", "Motor & Solenoid", "Position"):
                            inst["signal_type"] = "Discrete (0/1)"
                            break

        logger.info(f"[Enrich] Complete — {len(instruments)} instruments enriched, "
                    f"{sum(1 for i in instruments if i.get('line_number') not in ('N/A','',None))} have line nos, "
                    f"{sum(1 for i in instruments if i.get('equipment_number') not in ('N/A','',None))} have equip tags")

        # ── L5: Cross-reference CS tags (field transmitters → controllers) ─
        instruments = self._cross_ref_cs_tags(instruments)

        return instruments

    def _enrich_with_gemini_text(self, instruments, pdf_text, legend_text="", legend_files=None):
        """
        Use Gemini's text-only API (NOT Vision) to infer service descriptions
        and contextual fields from the raw PDF text content.

        Why text instead of vision:
          - Text API uses far less quota than vision inference.
          - The raw PDF text already contains line numbers, equipment tags,
            service labels — Gemini just needs to correlate them to each tag.
          - A single text call replaces many expensive multi-angle vision passes.

        Only runs when the Gemini client is available and quota not exceeded.
        Skips instruments where all fields are already populated.
        """
        if not self.gemini_client or self._gemini_quota_exceeded:
            return instruments
        if not pdf_text or len(pdf_text.strip()) < 50:
            return instruments
        legend_files = legend_files or []

        # Only enrich instruments that still have gap fields
        _gap_fields = ("service_description", "line_number", "equipment_number",
                       "fail_safe", "signal_type", "set_point", "control_system_tag")
        needs_enrich = [
            inst for inst in instruments
            if any(
                not inst.get(f) or inst.get(f) == "N/A"
                or (f == "service_description" and inst.get(f, "").startswith(
                    _SERVICE_VERB_MAP.get(inst.get("category", ""), "~NONE~")
                ))
                for f in _gap_fields
            )
        ]
        if not needs_enrich:
            logger.info("[Enrich] Gemini text: all fields already populated — skipping")
            return instruments

        tag_list = [inst["tag_number"] for inst in needs_enrich if inst.get("tag_number")]
        if not tag_list:
            return instruments

        # Limit text to first 10 000 chars to avoid large token usage
        text_chunk = pdf_text[:10000]
        legend_chunk = (legend_text or "")[:6000]
        legend_block = ""
        if legend_chunk:
            legend_block = f"""

    LEGEND / SYMBOL SHEET TEXT FROM AWS S3 (supplemental context only):
    {legend_chunk}

    Use the legend-sheet context to interpret abbreviations, DCS/CS conventions,
    signal notations, fail-safe symbols, and instrument/function-code meanings.
    If legend context helps you infer a field, prefer that interpretation over guesswork.
    """

        prompt = f"""You are a senior P&ID / FEED engineer. The text below was extracted from a P&ID drawing.
Correlate EACH instrument tag to any service context visible in the text, then return a JSON object.

PDF TEXT (raw extraction — up to 10 000 chars):
{text_chunk}
    {legend_block}

Instrument tags to enrich (provide data for as many as you can):
{json.dumps(tag_list[:60])}

Return ONLY a JSON object — no markdown fences, no explanation:
{{
  "PIT-3901-01": {{
    "service_description": "Pig Receiver Inlet Pressure",
    "line_number": "10\\"-G-3901-A2A",
    "equipment_number": "LP-3901",
    "fail_safe": "N/A",
    "signal_type": "4-20mA",
    "set_point": "75 barg (PSHH)",
    "control_system_tag": "PIC-3901-01"
  }}
}}

Rules:
- Use "N/A" for any field you cannot determine from the text.
- fail_safe values: FC (fail closed), FO (fail open), FL (fail last), N/A.
- signal_type: 4-20mA, HART, Fieldbus, Profibus, Discrete (0/1), Pneumatic, N/A.
- set_point: include the number AND unit, e.g. "75 barg", "250 °C", "12000 kg/h".
- control_system_tag: the DCS/CS tag for this instrument if identifiable from context; N/A otherwise.
- Only include tags from the list above.
"""

        try:
            from google.genai import types as _gtypes
            model = self.extraction_config.get("gemini_model", "gemini-2.0-flash")
            response = self.gemini_client.models.generate_content(
                model=model,
                contents=prompt,
                config=_gtypes.GenerateContentConfig(
                    temperature=0.05,
                    max_output_tokens=8000,
                ),
            )
            raw = response.text or ""
            s = raw.find("{");  e = raw.rfind("}") + 1
            if s >= 0 and e > s:
                enrichment_map = json.loads(raw[s:e])
                if isinstance(enrichment_map, dict):
                    updated = 0
                    legend_marked_tags = set()
                    for inst in instruments:
                        tag = inst.get("tag_number", "")
                        enrichment = enrichment_map.get(tag, {})
                        if not enrichment:
                            continue
                        for field in _gap_fields:
                            val = (enrichment.get(field) or "").strip()
                            existing = (inst.get(field) or "").strip()
                            is_gap = (
                                not existing
                                or existing == "N/A"
                                or (field == "service_description"
                                    and existing.startswith(
                                        _SERVICE_VERB_MAP.get(inst.get("category", ""), "~NONE~")
                                    ))
                            )
                            if val and val != "N/A" and is_gap:
                                inst[field] = val
                                updated += 1
                                if legend_chunk:
                                    legend_marked_tags.add(tag)
                    if legend_marked_tags:
                        for inst in instruments:
                            tag = inst.get("tag_number", "")
                            if tag in legend_marked_tags:
                                inst["notes"] = self._append_note_source(
                                    inst.get("notes", ""),
                                    "Legends sheet",
                                )
                    logger.info(f"[Enrich] Gemini text enrichment: updated {updated} fields across {len(instruments)} instruments")
        except Exception as ex:
            err = str(ex)
            if "429" in err or "quota" in err.lower() or "rate" in err.lower():
                self._gemini_quota_exceeded = True
                logger.warning("[Enrich] Gemini text quota exceeded — disabling Gemini for this request")
            else:
                logger.warning(f"[Enrich] Gemini text enrichment error: {ex}")

        return instruments

    # ────────────────────────────────────────────────────────────────────
    # Text-layer extraction (PyMuPDF — free, no API quota)
    # ────────────────────────────────────────────────────────────────────

    # Strict regex: requires dash/underscore separator (e.g. FT-3901-01)
    _TAG_RE = re.compile(
        r'\b([A-Z]{1,5})'
        r'[-_]'
        r'(\d{3,6}[A-Z]?)'
        r'(?:[-_](\d{1,3}[A-Z]?))?'
        r'\b',
        re.IGNORECASE,
    )

    # Flexible regex: allows space as separator (e.g. "FT 3901-01" in a bubble)
    _TAG_RE_FLEX = re.compile(
        r'\b([A-Z]{1,5})'
        r'[\s]'
        r'(\d{3,6}[A-Z]?)'
        r'(?:[-_](\d{1,3}[A-Z]?))?'
        r'\b',
        re.IGNORECASE,
    )

    _TAG_MIN_LEN = 5

    def _match_instrument_code(self, tag_upper):
        """Return (matched_code, cfg_entry) or (None, None) for a tag string."""
        prefix = re.match(r'^([A-Z]+)', tag_upper)
        if not prefix:
            return None, None
        code = prefix.group(1)
        for length in range(len(code), 0, -1):
            candidate = code[:length]
            if candidate in INSTRUMENT_CATEGORIES:
                return candidate, INSTRUMENT_CATEGORIES[candidate]
        return None, None

    def _make_instrument_record(self, tag, matched_code, cfg_entry, dn, rev, note):
        """Build a standardised instrument dict."""
        return {
            "tag_number":          tag.upper(),
            "control_system_tag":  "N/A",
            "instrument_type":     cfg_entry["name"],
            "category":            cfg_entry["category"],
            "pid_no":              dn,
            "service_description": "",
            "line_number":         "N/A",
            "equipment_number":    "N/A",
            "loop_number":         "N/A",
            "fail_safe":           "N/A",
            "signal_type":         "N/A",
            "set_point":           "N/A",
            "drawing_number":      dn,
            "revision":            rev,
            "notes":               note,
        }

    def _derive_loop_number(self, tag):
        """
        Derive loop number from the instrument tag number.
        The numeric area/loop code is the middle segment after the first dash.
          PIT-3901-01  → 3901
          FIT-3901-08A → 3901
          SDV-3901-01  → 3901
          FT-101A      → 101
        """
        m = re.search(r'[-_](\d{3,6})', tag or '')
        return m.group(1) if m else 'N/A'

    def _derive_cs_tag_isa(self, tag):
        """
        Determine the Control System Tag from ISA-5.1 function code analysis.

        Returns (cs_tag, is_dcs_instrument):
          is_dcs_instrument=True  → the instrument IS a DCS/CS device;
                                    cs_tag == tag (it is its own CS tag)
          is_dcs_instrument=False → field instrument; cs_tag is the *expected*
                                    controller tag derived via the function-code
                                    substitution rules (may not exist on this drawing)
          cs_tag == "N/A"         → derivation not possible

        Soft-coded via CS_TAG_CONFIG — no changes to this method needed for tuning.
        """
        cfg = CS_TAG_CONFIG
        tag_upper = (tag or "").strip().upper()
        # Match:  function-code letters  +  dash+numbers+optional-suffix
        m = re.match(r'^([A-Z]{2,6})([-–]\d.*)$', tag_upper)
        if not m:
            return "N/A", False
        func_code, numeric_suffix = m.group(1), m.group(2)

        # Step 1 — Is this already a DCS controller instrument?
        if func_code in cfg["dcs_function_codes"]:
            return tag_upper, True

        # Step 2 — Is it a field transmitter/element? Derive the controller tag.
        first_letter = func_code[0]
        rest         = func_code[1:]
        for field_suffix, ctrl_suffix in cfg["transmitter_to_controller"].items():
            if rest == field_suffix:
                derived = f"{first_letter}{ctrl_suffix}{numeric_suffix}"
                return derived, False

        return "N/A", False

    def _cross_ref_cs_tags(self, instruments):
        """
        Post-processing: cross-reference field instruments against the complete
        list of instruments extracted from this drawing.

        For each field transmitter (e.g. FT-3901-01) whose control_system_tag is
        still N/A, check whether the derived controller (FIC-3901-01 or FIC-3901)
        actually appears in the instruments list.  If found → set the CS tag.
        If not found but derivation still produced a plausible controller tag,
        store the expected tag (prefixed "Expected:") so engineers know what
        controller *should* be on the drawing.

        Soft-coded via CS_TAG_CONFIG.
        """
        # Build lookup: normalised tag → instrument record
        tag_map = {
            (i.get("tag_number") or "").strip().upper(): i
            for i in instruments
        }

        for inst in instruments:
            if inst.get("control_system_tag") not in ("N/A", "", None):
                continue  # already resolved

            tag = (inst.get("tag_number") or "").strip().upper()
            derived_cs, is_dcs = self._derive_cs_tag_isa(tag)

            if is_dcs:
                # This instrument IS a DCS/CS device
                inst["control_system_tag"] = tag
                continue

            if derived_cs == "N/A":
                continue

            # Check exact match in drawing's instrument list
            if derived_cs in tag_map:
                inst["control_system_tag"] = derived_cs
                continue

            # Try abbreviated loop match (drop the last suffix)
            # e.g. try FIC-3901 when FIC-3901-01 not found
            parts = derived_cs.split("-")
            if len(parts) >= 3:
                short_cs = f"{parts[0]}-{parts[1]}"
                if short_cs in tag_map:
                    inst["control_system_tag"] = tag_map[short_cs]["tag_number"]
                    continue

            # Controller not found on this drawing — store expected tag for reference
            # (wrapped in "Expected:" so users know it's inferred, not confirmed)
            inst["control_system_tag"] = f"Expected: {derived_cs}"

        cs_found = sum(
            1 for i in instruments
            if i.get("control_system_tag") not in ("N/A", "", None)
            and not (i.get("control_system_tag") or "").startswith("Expected:")
        )
        logger.info(f"[Enrich] CS tag cross-reference: {cs_found} confirmed CS tags resolved")
        return instruments

    def _infer_service_description(self, tag, instrument_type, category):
        """
        Generate a meaningful base service description from tag + Category.
        Used as a fallback when no contextual text is available from the drawing.
        """
        loop = self._derive_loop_number(tag)
        verb = _SERVICE_VERB_MAP.get(category, instrument_type or "Measurement")
        if loop and loop != 'N/A':
            return f"{verb} — System {loop}"
        return verb

    def _resolve_drawing_info_from_pdf(self, pid_bytes, drawing_info):
        """
        Resolve drawing_number/pid_no from PDF title-block text (DWG NO / P&ID NO).
        Soft-coded via DRAWING_NUMBER_CONFIG and only overrides when a strong
        candidate is found.
        """
        info = dict(drawing_info or {})
        current_dn = (info.get("drawing_number") or "").strip()
        current_pid = (info.get("pid_no") or "").strip()

        detected = self._extract_drawing_number_from_pdf(pid_bytes)
        if not detected:
            return info

        if detected != current_dn or detected != current_pid:
            logger.info(
                f"[InstrumentIndex] Drawing number detected from title block: {detected} "
                f"(input was drawing='{current_dn or 'N/A'}', pid='{current_pid or 'N/A'}')"
            )

        info["drawing_number"] = detected
        info["pid_no"] = detected
        return info

    def _extract_drawing_number_from_pdf(self, pid_bytes):
        """Extract best DWG/P&ID number candidate from PDF text."""
        try:
            import fitz
        except ImportError:
            return ""

        try:
            doc = fitz.open(stream=pid_bytes, filetype="pdf")
        except Exception:
            return ""

        try:
            page_count = len(doc)
            # Title blocks are usually on first/last pages; scan those first.
            page_indices = list(dict.fromkeys(
                [0, 1, page_count - 2, page_count - 1] if page_count > 2 else list(range(page_count))
            ))

            text_chunks = []
            for idx in page_indices:
                if 0 <= idx < page_count:
                    text_chunks.append(doc[idx].get_text("text") or "")

            text = "\n".join(text_chunks)
            return self._extract_drawing_number_from_text(text)
        except Exception:
            return ""
        finally:
            doc.close()

    def _extract_drawing_number_from_text(self, text):
        """Label-aware extraction of drawing number from raw OCR/PDF text."""
        if not text:
            return ""

        cfg = DRAWING_NUMBER_CONFIG
        normalized = re.sub(r'\s+', ' ', text.upper())
        label_pats = [re.compile(p, re.IGNORECASE) for p in cfg.get("label_patterns", [])]
        value_pats = [re.compile(p, re.IGNORECASE) for p in cfg.get("value_patterns", [])]
        min_len = cfg.get("min_length", 8)
        max_len = cfg.get("max_length", 64)
        window = cfg.get("window_chars", 140)

        def _is_valid_candidate(val):
            if not val:
                return False
            v = val.strip().upper().strip("-:;,. ")
            if len(v) < min_len or len(v) > max_len:
                return False
            if not re.search(r'[A-Z]', v) or not re.search(r'\d', v):
                return False
            if not any(sep in v for sep in ('-', '/', '.')):
                return False
            # Filter obvious non-document tokens
            blocked = {
                "PROCESS", "INSTRUMENT", "DRAWING", "NUMBER", "REVISION",
                "SCALE", "SHEET", "TITLE", "PROJECT", "CLIENT",
            }
            if v in blocked:
                return False
            # Filter instrument-like tags (FT-3901-01 etc.)
            if re.match(r'^[A-Z]{1,5}-\d{3,6}(?:-\d{1,3}[A-Z]?)?$', v):
                return False
            return True

        # 1) Prefer value near explicit title-block labels (DWG NO, P&ID NO, ...)
        for lp in label_pats:
            for lm in lp.finditer(normalized):
                start = max(0, lm.end())
                end = min(len(normalized), lm.end() + window)
                region = normalized[start:end]
                for vp in value_pats:
                    for vm in vp.finditer(region):
                        cand = vm.group(1)
                        if _is_valid_candidate(cand):
                            return cand.strip().upper()

        # 2) Fallback: strongest global candidate in the scanned pages
        all_candidates = []
        for vp in value_pats:
            for vm in vp.finditer(normalized):
                cand = vm.group(1)
                if _is_valid_candidate(cand):
                    all_candidates.append(cand.strip().upper())

        if not all_candidates:
            return ""

        # Prefer richer segmented identifiers (more separators/segments)
        all_candidates = list(dict.fromkeys(all_candidates))
        all_candidates.sort(key=lambda s: (s.count('-') + s.count('/') + s.count('.'), len(s)), reverse=True)
        return all_candidates[0]

    def _load_legend_context_from_s3(self, drawing_info):
        """
        Optionally load related legend/symbol sheets from S3.

        This is a soft enrichment path only: the P&ID remains the primary
        extraction source, while legend sheets help interpret conventions such as
        control-system tags, signal notation, fail-safe symbols, and abbreviations.
        """
        cfg = LEGEND_S3_CONFIG
        context = {"text": "", "files": []}

        if not cfg.get("enabled", True):
            return context

        use_s3 = os.environ.get("USE_S3", "false").lower() == "true"
        s3_ready = os.environ.get("S3_READY", "false").lower() == "true"
        if not (use_s3 and s3_ready):
            return context

        try:
            from apps.core.s3_service import get_s3_service
        except Exception as ex:
            logger.warning(f"[InstrumentIndex] Legend S3 helper unavailable: {ex}")
            return context

        drawing_number = (drawing_info.get("drawing_number") or "").strip().upper()
        project_name = (drawing_info.get("project_name") or "").strip()
        base_path = (os.environ.get("PFD_BASE_PATH") or "").strip().strip("/")

        search_prefixes = []
        if base_path and project_name:
            search_prefixes.append(f"{base_path}/{project_name}/")
        if base_path:
            search_prefixes.append(f"{base_path}/")
        if project_name:
            search_prefixes.append(f"{project_name}/")
        search_prefixes.append("")

        search_prefixes = list(dict.fromkeys(p for p in search_prefixes if p is not None))
        keywords = tuple(k.lower() for k in cfg.get("filename_keywords", []))
        exts = tuple(e.lower() for e in cfg.get("preferred_extensions", [".pdf"]))
        max_keys = cfg.get("max_list_keys", 400)
        max_files = cfg.get("max_candidate_files", 3)

        s3 = get_s3_service()
        candidates = []

        for prefix in search_prefixes:
            listing = s3.list_files(prefix=prefix, max_keys=max_keys)
            if not listing.get("success"):
                continue

            for file_info in listing.get("files", []):
                key = file_info.get("key") or ""
                filename = os.path.basename(key).lower()
                if not filename.endswith(exts):
                    continue
                if not any(word in filename for word in keywords):
                    continue

                score = 0
                if "legend" in filename:
                    score += 5
                if "symbol" in filename:
                    score += 4
                if project_name and project_name.lower() in key.lower():
                    score += 2
                if drawing_number:
                    drawing_tokens = [tok for tok in re.split(r'[-_/ .]+', drawing_number.lower()) if len(tok) >= 3]
                    score += sum(1 for tok in drawing_tokens if tok in key.lower())

                candidates.append((score, key))

            if candidates:
                break

        if not candidates:
            return context

        selected_keys = []
        seen = set()
        for _, key in sorted(candidates, key=lambda item: (item[0], len(item[1])), reverse=True):
            if key in seen:
                continue
            seen.add(key)
            selected_keys.append(key)
            if len(selected_keys) >= max_files:
                break

        if not selected_keys:
            return context

        legend_chunks = []
        for key in selected_keys:
            raw_bytes = self._download_s3_object_bytes(s3, key)
            if not raw_bytes:
                continue
            legend_text = self._extract_pdf_text_bytes(
                raw_bytes,
                max_pages=cfg.get("max_pages_per_file", 3),
            )
            if legend_text:
                legend_chunks.append(f"LEGEND FILE: {os.path.basename(key)}\n{legend_text}")

        if not legend_chunks:
            return context

        context["text"] = "\n\n".join(legend_chunks)[:cfg.get("max_text_chars", 12000)]
        context["files"] = selected_keys
        logger.info(
            f"[InstrumentIndex] Loaded legend context from S3: {len(selected_keys)} file(s)"
        )
        return context

    def build_legend_context_from_uploaded_file(self, legend_bytes, filename):
        """Build legend context from an explicitly uploaded legend/symbol PDF."""
        text = self._extract_pdf_text_bytes(
            legend_bytes,
            max_pages=LEGEND_S3_CONFIG.get("max_pages_per_file", 3),
        )
        if not text:
            return {"text": "", "files": []}
        return {
            "text": text[:LEGEND_S3_CONFIG.get("max_text_chars", 12000)],
            "files": [filename],
        }

    def _merge_legend_contexts(self, primary_context, secondary_context):
        """Merge uploaded and S3 legend contexts, preferring uploaded text first."""
        primary_context = primary_context or {}
        secondary_context = secondary_context or {}

        primary_text = (primary_context.get("text") or "").strip()
        secondary_text = (secondary_context.get("text") or "").strip()
        merged_files = []
        for item in (primary_context.get("files") or []) + (secondary_context.get("files") or []):
            if item and item not in merged_files:
                merged_files.append(item)

        merged_text = "\n\n".join(chunk for chunk in (primary_text, secondary_text) if chunk)
        if merged_text:
            merged_text = merged_text[:LEGEND_S3_CONFIG.get("max_text_chars", 12000)]

        return {
            "text": merged_text,
            "files": merged_files,
        }

    def _download_s3_object_bytes(self, s3_service, s3_key):
        try:
            result = s3_service.download_file(s3_key)
            if not result.get("success"):
                return b""
            body = result.get("body")
            if not body:
                return b""
            return body.read()
        except Exception as ex:
            logger.warning(f"[InstrumentIndex] Legend S3 download failed for {s3_key}: {ex}")
            return b""

    def _extract_pdf_text_bytes(self, pdf_bytes, max_pages=3):
        if not pdf_bytes:
            return ""
        try:
            import fitz
            doc = fitz.open(stream=pdf_bytes, filetype="pdf")
            try:
                return "\n".join(
                    doc[idx].get_text("text") or ""
                    for idx in range(min(len(doc), max_pages))
                )
            finally:
                doc.close()
        except Exception as ex:
            logger.warning(f"[InstrumentIndex] Legend text extraction failed: {ex}")
            return ""

    def _append_note_source(self, note, marker):
        base = (note or "").strip()
        if not marker:
            return base
        if not base:
            return marker
        if marker in base:
            return base
        return f"{base} + {marker}"

    def _scan_for_tags(self, text, seen_tags, dn, rev, instruments, note="PDF text layer"):
        """
        Extract instrument tags from a text string using both strict and flexible regex.
        Appends new (not-yet-seen) records to instruments in-place.
        Internal note labels are mapped to user-friendly strings for the frontend.
        """
        # Map verbose internal labels → concise user-visible source labels
        _note_map = {
            "PDF text (plain)":            "PDF text layer",
            "PDF block reconstruction":    "PDF circle/bubble",
            "PDF block pair":              "PDF circle/bubble",
            "PDF spatial grouping":        "PDF circle (spatial)",
            "PDF spatial triple":          "PDF circle (spatial)",
        }
        # Tesseract notes: simplify PSM details
        if note.startswith("Tesseract"):
            if "spatial triple" in note:
                display_note = "OCR circle (3-part)"
            elif "spatial" in note:
                display_note = "OCR circle (spatial)"
            else:
                display_note = "OCR text"
        else:
            display_note = _note_map.get(note, note)

        for pattern in (self._TAG_RE, self._TAG_RE_FLEX):
            for m in pattern.finditer(text):
                # Normalise: collapse whitespace separators to dash
                raw = re.sub(r'\s+', '-', m.group(0).strip())
                full_tag = raw.upper()
                if len(full_tag) < self._TAG_MIN_LEN:
                    continue
                matched_code, cfg_entry = self._match_instrument_code(full_tag)
                if not matched_code:
                    continue
                norm = self._normalize_tag(full_tag)
                if norm in seen_tags:
                    continue
                seen_tags.add(norm)
                instruments.append(self._make_instrument_record(full_tag, matched_code, cfg_entry, dn, rev, display_note))


    def _extract_with_text_layer(self, pid_bytes, drawing_info):
        """
        Advanced 3-pass text-layer extraction using PyMuPDF.

        Pass A  — Full-page plain text regex
                  Catches complete single-span tags: "FIT-3901-08A"

        Pass B  — Block-level text reconstruction
                  Joins lines within each text block (instrument circles are one block).
                  Catches multi-line tags: ["FIT", "3901", "08A"] → "FIT-3901-08A"

        Pass C  — Spatial word-proximity grouping
                  Finds isolated prefix words (e.g. "FT") and nearby number words
                  ("3901-01") within the instrument circle diameter.
                  Catches split-span tags that no other pass can reconstruct.

        All three passes feed the same deduplication set.
        """
        instruments: list = []
        seen_tags:   set  = set()
        dn  = drawing_info.get("drawing_number", "N/A")
        rev = drawing_info.get("revision", "0")

        try:
            import fitz  # PyMuPDF
        except ImportError:
            logger.warning("[InstrumentIndex] PyMuPDF not installed — skipping text-layer extraction")
            return []

        try:
            doc = fitz.open(stream=pid_bytes, filetype="pdf")
        except Exception as e:
            logger.warning(f"[InstrumentIndex] PyMuPDF open error: {e}")
            return []

        # Pre-build a sorted prefix list (longest-first) + compiled patterns for Pass C
        all_pfx = sorted(INSTRUMENT_CATEGORIES.keys(), key=len, reverse=True)
        _pfx_exact  = re.compile(
            r'^(' + '|'.join(re.escape(p) for p in all_pfx) + r')$', re.IGNORECASE
        )
        # Number part: "3901-01A" / "3901" / "001"
        _num_part = re.compile(r'^\d{2,6}[A-Z]?(?:[-]\d{1,3}[A-Z]?)?$', re.IGNORECASE)
        # Bare suffix: "01" / "08A"
        _suffix   = re.compile(r'^\d{1,3}[A-Z]?$', re.IGNORECASE)

        use_spatial = self.extraction_config.get("spatial_grouping", True)
        radius      = self.extraction_config.get("spatial_radius", 80)

        for page_no in range(len(doc)):
            page = doc[page_no]

            # ── Pass A: Full-page plain text ─────────────────────────────
            plain = page.get_text("text") or ""
            if len(plain.strip()) > 10:
                self._scan_for_tags(plain, seen_tags, dn, rev, instruments, "PDF text (plain)")

            # ── Pass B: Block-level reconstruction ───────────────────────
            # P&ID instrument circles are typically ONE text block whose lines contain:
            #   line 1 → instrument type prefix  "FIT"
            #   line 2 → loop / area number      "3901"
            #   line 3 → suffix                  "08A"
            # Joining with "-" reconstructs the full tag.
            try:
                page_dict = page.get_text("dict")
                for block in page_dict.get("blocks", []):
                    if block.get("type") != 0:
                        continue  # skip image blocks
                    lines_text = []
                    for ln in block.get("lines", []):
                        lt = "".join(s.get("text", "") for s in ln.get("spans", [])).strip()
                        if lt:
                            lines_text.append(lt)
                    if not lines_text:
                        continue
                    # Try joining with dash, no separator, and space
                    for joiner in ("-", "", " "):
                        combined = joiner.join(lines_text)
                        if combined:
                            self._scan_for_tags(
                                combined, seen_tags, dn, rev, instruments,
                                "PDF block reconstruction"
                            )
                    # Also try consecutive-line pairs (handles 2-part split)
                    for i in range(len(lines_text) - 1):
                        for joiner in ("-", ""):
                            pair = joiner.join([lines_text[i], lines_text[i + 1]])
                            self._scan_for_tags(
                                pair, seen_tags, dn, rev, instruments,
                                "PDF block pair"
                            )
            except Exception as be:
                logger.debug(f"[InstrumentIndex] Block pass error p{page_no+1}: {be}")

            # ── Pass C: Spatial word-proximity grouping ───────────────────
            # For each word that is a pure instrument prefix, find the closest
            # number word within `radius` pixels and combine them into a tag.
            if use_spatial:
                try:
                    words = page.get_text("words")
                    # words: (x0, y0, x1, y1, text, block_no, line_no, word_no)
                    for wi, w in enumerate(words):
                        wtext = w[4].strip()
                        if not _pfx_exact.match(wtext):
                            continue
                        # centre of this prefix word
                        px = (w[0] + w[2]) / 2
                        py = (w[1] + w[3]) / 2

                        # Collect all nearby number-like words
                        candidates: list[tuple[float, str, str]] = []  # (dist, raw, joined)
                        for nw in words:
                            ntext = nw[4].strip()
                            if not (_num_part.match(ntext)):
                                continue
                            nx = (nw[0] + nw[2]) / 2
                            ny = (nw[1] + nw[3]) / 2
                            dist = ((nx - px) ** 2 + (ny - py) ** 2) ** 0.5
                            if dist < radius:
                                candidates.append((dist, ntext, f"{wtext}-{ntext}"))

                        # Also look for a bare suffix word near the best number candidate
                        candidates.sort()
                        for _, ntext, joined_tag in candidates:
                            self._scan_for_tags(
                                joined_tag, seen_tags, dn, rev, instruments,
                                "PDF spatial grouping"
                            )
                            # Check if a suffix word is also nearby
                            # (builds "FT-3901-01" from three separate words)
                            for nw2 in words:
                                nx2 = (nw2[0] + nw2[2]) / 2
                                ny2 = (nw2[1] + nw2[3]) / 2
                                if not _suffix.match(nw2[4].strip()):
                                    continue
                                dist2 = ((nx2 - px) ** 2 + (ny2 - py) ** 2) ** 0.5
                                if dist2 < radius and nw2[4].strip() != ntext:
                                    triple = f"{joined_tag}-{nw2[4].strip()}"
                                    self._scan_for_tags(
                                        triple, seen_tags, dn, rev, instruments,
                                        "PDF spatial triple"
                                    )
                except Exception as se:
                    logger.debug(f"[InstrumentIndex] Spatial pass error p{page_no+1}: {se}")

        doc.close()
        logger.info(
            f"[InstrumentIndex] Text-layer 3-pass result: {len(instruments)} unique instrument tags"
        )
        return instruments

    # ────────────────────────────────────────────────────────────────────
    # Tesseract OCR — runs on ALL PDFs (vector + scanned)
    # ────────────────────────────────────────────────────────────────────

    def _extract_with_tesseract(self, pid_bytes, drawing_info):
        """
        Tesseract OCR extraction with spatial word-proximity grouping.

        Runs on ALL PDFs (not just scanned) because even vector P&IDs benefit
        from OCR — instrument circle text is sometimes poorly extracted by
        the PDF text layer due to how AutoCAD writes character spans.

        Uses:
          - Auto-contrast preprocessing for better circle text recognition
          - Multiple PSM modes (11=sparse best for P&IDs, 6=uniform block, 3=auto)
          - Spatial grouping: prefix word + nearby number word → full tag
          - Both strict and flexible regex on OCR output
        """
        if not self.extraction_config.get("enable_tesseract", True):
            return []
        if not self.tesseract_available:
            return []

        try:
            import pytesseract
            from pytesseract import Output as TessOutput
        except ImportError:
            return []

        cfg = self.extraction_config
        dn  = drawing_info.get("drawing_number", "N/A")
        rev = drawing_info.get("revision", "0")
        instruments: list = []
        seen_tags:   set  = set()

        tess_dpi = cfg.get("tesseract_dpi", 150)
        try:
            pil_images = convert_from_bytes(pid_bytes, dpi=tess_dpi)
        except Exception as e:
            logger.warning(f"[InstrumentIndex] Tesseract: pdf2image failed: {e}")
            return []

        # Compiled patterns for spatial grouping
        all_pfx = sorted(INSTRUMENT_CATEGORIES.keys(), key=len, reverse=True)
        _pfx_exact = re.compile(
            r'^(' + '|'.join(re.escape(p) for p in all_pfx) + r')$', re.IGNORECASE
        )
        _num_part = re.compile(r'^\d{2,6}[A-Z]?(?:[-]\d{1,3}[A-Z]?)?$', re.IGNORECASE)
        _suffix   = re.compile(r'^\d{1,3}[A-Z]?$', re.IGNORECASE)
        spatial_radius = cfg.get("spatial_radius", 80) * (tess_dpi / 150)

        # PSM modes: 11=sparse (best for P&IDs),  6=uniform block,  3=fully auto
        psm_modes = [11, 6, 3]

        for page_no, img in enumerate(pil_images, start=1):
            # Pre-process: grayscale + auto-contrast improves OCR on blueprint drawings
            try:
                from PIL import ImageOps
                gray = img.convert("L")
                gray = ImageOps.autocontrast(gray, cutoff=2)
            except Exception:
                gray = img

            for psm in psm_modes:
                try:
                    tess_config = f"--psm {psm} --oem 3"
                    data = pytesseract.image_to_data(
                        gray, config=tess_config, output_type=TessOutput.DICT
                    )
                except Exception as te:
                    logger.debug(f"[InstrumentIndex] Tesseract PSM {psm} p{page_no} error: {te}")
                    continue

                # Build word list with confidence filter (>= 30 %)
                conf_threshold = 30
                word_list = [
                    (
                        data["left"][i], data["top"][i],
                        data["width"][i], data["height"][i],
                        data["text"][i],
                    )
                    for i in range(len(data["text"]))
                    if data["text"][i].strip() and int(data["conf"][i] or 0) >= conf_threshold
                ]

                # Pass A: plain-text scan of all words joined
                full_ocr_text = " ".join(w[4] for w in word_list)
                self._scan_for_tags(
                    full_ocr_text, seen_tags, dn, rev, instruments, f"Tesseract PSM {psm}"
                )

                # Pass B: spatial grouping — find prefix + nearby number words
                for wi, w in enumerate(word_list):
                    wtext = w[4].strip()
                    if not _pfx_exact.match(wtext):
                        continue
                    px = w[0] + w[2] / 2
                    py = w[1] + w[3] / 2

                    nearby: list[tuple[float, str]] = []
                    for nw in word_list:
                        ntext = nw[4].strip()
                        if not _num_part.match(ntext):
                            continue
                        nx = nw[0] + nw[2] / 2
                        ny = nw[1] + nw[3] / 2
                        dist = ((nx - px) ** 2 + (ny - py) ** 2) ** 0.5
                        if dist < spatial_radius:
                            nearby.append((dist, ntext))

                    nearby.sort()
                    for _, ntext in nearby:
                        candidate = f"{wtext.upper()}-{ntext.upper()}"
                        self._scan_for_tags(
                            candidate, seen_tags, dn, rev, instruments,
                            f"Tesseract spatial PSM {psm}"
                        )
                        # Also look for a trailing suffix word
                        for nw2 in word_list:
                            ntext2 = nw2[4].strip()
                            if not _suffix.match(ntext2) or ntext2 == ntext:
                                continue
                            nx2 = nw2[0] + nw2[2] / 2
                            ny2 = nw2[1] + nw2[3] / 2
                            dist2 = ((nx2 - px) ** 2 + (ny2 - py) ** 2) ** 0.5
                            if dist2 < spatial_radius:
                                triple = f"{candidate}-{ntext2.upper()}"
                                self._scan_for_tags(
                                    triple, seen_tags, dn, rev, instruments,
                                    f"Tesseract spatial triple PSM {psm}"
                                )

            # ── Supplementary OCR (EasyOCR + PaddleOCR) on the same page ──
            # Purely additive recall booster — runs once per page on the
            # already-preprocessed `gray` image. Soft-coded via
            # `enable_easyocr` / `enable_paddleocr` in EXTRACTION_CONFIG.
            try:
                self._run_supplementary_ocr(
                    gray, page_no, seen_tags, dn, rev, instruments
                )
            except Exception as se:
                logger.debug(f"[InstrumentIndex] supplementary OCR p{page_no} skipped: {se}")

        logger.info(f"[InstrumentIndex] Tesseract result: {len(instruments)} unique tags")
        return instruments

    # ────────────────────────────────────────────────────────────────────
    # PDF → JPEG conversion
    # ────────────────────────────────────────────────────────────────────

    def _to_jpeg_pages(self, pid_bytes):
        """Convert PDF (or image) to list of JPEG bytes, one entry per page."""
        is_pdf = pid_bytes[:4] == b"%PDF"

        if is_pdf:
            logger.info("[InstrumentIndex] Converting PDF to images…")
            try:
                pil_images = convert_from_bytes(pid_bytes, dpi=self.extraction_config["pdf_dpi"])
            except Exception as e:
                logger.error(f"[InstrumentIndex] pdf2image failed: {e}")
                # Fallback: send raw bytes as a single "page"
                return [pid_bytes]
        else:
            # Already an image
            pil_images = [Image.open(io.BytesIO(pid_bytes))]

        jpeg_pages = []
        for img in pil_images:
            jpeg_pages.append(self._pil_to_jpeg(img))
        return jpeg_pages

    def _pil_to_jpeg(self, img, max_size=None):
        """Resize + convert PIL image to JPEG bytes."""
        if max_size is None:
            max_size = self.extraction_config["max_image_size"]
        # Resize if too large
        if img.width > max_size or img.height > max_size:
            ratio = min(max_size / img.width, max_size / img.height)
            img = img.resize((int(img.width * ratio), int(img.height * ratio)), Image.Resampling.LANCZOS)
        # Flatten transparency
        if img.mode in ("RGBA", "LA", "P"):
            bg = Image.new("RGB", img.size, (255, 255, 255))
            if img.mode == "P":
                img = img.convert("RGBA")
            bg.paste(img, mask=img.split()[-1] if img.mode in ("RGBA", "LA") else None)
            img = bg
        elif img.mode != "RGB":
            img = img.convert("RGB")

        buf = io.BytesIO()
        img.save(buf, format="JPEG", quality=self.extraction_config.get("jpeg_quality", 90), optimize=True)
        return buf.getvalue()

    # ────────────────────────────────────────────────────────────────────
    # Per-page AI analysis — multi-angle, multi-pass
    # ────────────────────────────────────────────────────────────────────

    def _analyse_page(self, jpeg_bytes, drawing_info, page_no, only_engine=None):
        """
        Multi-pass extraction for one P&ID page.
          Pass 1  — Full drawing, 0°
          Pass 2+ — Full drawing rotated (vertical text)
          Tile passes — 2×2 quadrant zoom (dense drawings)

        only_engine: when set, restricts all Vision calls to that engine only.
        """
        cfg = self.extraction_config
        img = Image.open(io.BytesIO(jpeg_bytes))

        all_pass_instruments = []

        # ── Pass 1: Full drawing, normal orientation ─────────────────────
        logger.info(f"[InstrumentIndex] Page {page_no} — Pass 1 (0° full drawing) [{only_engine or 'auto'}]")
        p1 = self._vision_pass(
            jpeg_bytes, drawing_info, page_no,
            extra_hint="Standard orientation. Extract ALL instrument tags visible.",
            mode="primary", max_tokens=cfg["max_tokens_primary"], only_engine=only_engine,
        )
        logger.info(f"[InstrumentIndex] Page {page_no} — Pass 1: {len(p1)} instruments")
        all_pass_instruments.extend(p1)

        # ── Rotation passes: catch vertical / slanted tags ───────────────
        if cfg["enable_rotation"]:
            for angle in cfg["rotation_angles"]:
                rotated_img = img.rotate(-angle, expand=True)
                rot_jpeg = self._pil_to_jpeg(rotated_img)
                label = f"rotated_{angle}cw"
                hint = (
                    f"IMAGE ROTATED {angle}° CLOCKWISE. "
                    "Tags that were printed vertically now appear horizontal. "
                    "Focus on catching instrument tags along pipe runs and diagonal areas. "
                    "Do NOT re-report tags already clearly horizontal in the standard view."
                )
                logger.info(f"[InstrumentIndex] Page {page_no} — Rotation pass {angle}°CW [{only_engine or 'auto'}]")
                pr = self._vision_pass(
                    rot_jpeg, drawing_info, page_no,
                    extra_hint=hint, mode=label, max_tokens=cfg["max_tokens_primary"], only_engine=only_engine,
                )
                logger.info(f"[InstrumentIndex] Page {page_no} — Rotation {angle}°: {len(pr)} instruments")
                all_pass_instruments.extend(pr)

        # ── Tile passes: zoomed quadrant scan ────────────────────────────
        if cfg["enable_tiling"]:
            tiles = self._generate_tiles(img, cfg["tile_grid"], cfg["tile_overlap"])
            for tile_idx, tile_jpeg in enumerate(tiles):
                row = tile_idx // cfg["tile_grid"][1]
                col = tile_idx % cfg["tile_grid"][1]
                hint = (
                    f"ZOOMED TILE — Quadrant row={row+1}, col={col+1} of a {cfg['tile_grid'][0]}×{cfg['tile_grid'][1]} grid. "
                    "This is a high-resolution crop of part of the P&ID. "
                    "Extract EVERY instrument tag visible, including small or partially visible ones."
                )
                logger.info(
                    f"[InstrumentIndex] Page {page_no} — Tile ({row+1},{col+1}) [{only_engine or 'auto'}]"
                )
                pt = self._vision_pass(
                    tile_jpeg, drawing_info, page_no,
                    extra_hint=hint, mode=f"tile_r{row+1}c{col+1}",
                    max_tokens=cfg["max_tokens_tile"], only_engine=only_engine,
                )
                logger.info(f"[InstrumentIndex] Page {page_no} — Tile ({row+1},{col+1}): {len(pt)} instruments")
                all_pass_instruments.extend(pt)

        # ── Merge & deduplicate ──────────────────────────────────────────
        merged = self._merge_instruments(all_pass_instruments)

        # Enrich category from tag prefix
        for inst in merged:
            inst = self._enrich_category(inst)

        logger.info(
            f"[InstrumentIndex] Page {page_no} — "
            f"Total across all passes: {len(all_pass_instruments)}, "
            f"unique after merge: {len(merged)}"
        )
        return merged

    # ────────────────────────────────────────────────────────────────────
    # Vision pass dispatcher
    # ────────────────────────────────────────────────────────────────────

    def _vision_pass(self, jpeg_bytes, drawing_info, page_no,
                     extra_hint="", mode="primary", max_tokens=None, only_engine=None):
        """
        Try AI engines in priority order (EXTRACTION_CONFIG['ai_engines']).
        only_engine: when set, restricts to that single engine (e.g. 'gemini' or 'openai').
        Falls back to next engine on failure / quota exceeded.
        """
        if max_tokens is None:
            max_tokens = self.extraction_config["max_tokens_primary"]

        b64 = base64.b64encode(jpeg_bytes).decode("utf-8")
        if len(b64) > 20 * 1024 * 1024:
            logger.warning(f"[InstrumentIndex] {mode} image >20 MB — skipping")
            return []

        prompt = self._build_prompt(drawing_info, page_no, extra_hint=extra_hint)
        fallback_prompt = self._build_fallback_prompt(drawing_info, page_no, extra_hint=extra_hint)

        engines = self.extraction_config.get("ai_engines", ["gemini", "openai"])
        if only_engine:
            engines = [only_engine]
        for engine in engines:
            instruments = []
            if engine == "gemini":
                if self._gemini_quota_exceeded or not self.gemini_client:
                    continue
                instruments = self._call_gemini_vision(b64, prompt, f"{mode}_gemini", max_tokens)
                if not instruments:
                    instruments = self._call_gemini_vision(b64, fallback_prompt, f"{mode}_gemini_fb", max_tokens)
            elif engine == "openai":
                if self._quota_exceeded or not self.openai_client:
                    continue
                instruments = self._call_openai_vision(b64, prompt, f"{mode}_openai", max_tokens)
                if not instruments:
                    instruments = self._call_openai_vision(b64, fallback_prompt, f"{mode}_openai_fb", max_tokens)

            if instruments:
                logger.info(f"[InstrumentIndex] {mode} — {engine} returned {len(instruments)} instruments")
                return instruments

        logger.warning(f"[InstrumentIndex] {mode} — all AI engines exhausted or quota exceeded")
        return []

    # ────────────────────────────────────────────────────────────────────
    # Tile generator
    # ────────────────────────────────────────────────────────────────────

    def _generate_tiles(self, img, grid=(2, 2), overlap=0.12):
        """
        Crop a PIL image into rows×cols tiles with fractional overlap.
        Tiles are converted to JPEG bytes at full extraction config size.
        """
        rows, cols = grid
        w, h = img.size
        stride_w = w / cols
        stride_h = h / rows
        pad_w = int(w * overlap / 2)
        pad_h = int(h * overlap / 2)

        tiles = []
        for row in range(rows):
            for col in range(cols):
                left  = max(0, int(col * stride_w) - pad_w)
                upper = max(0, int(row * stride_h) - pad_h)
                right = min(w, int((col + 1) * stride_w) + pad_w)
                lower = min(h, int((row + 1) * stride_h) + pad_h)
                tile = img.crop((left, upper, right, lower))
                tiles.append(self._pil_to_jpeg(tile))
        return tiles

    # ────────────────────────────────────────────────────────────────────
    # Instrument merger / deduplicator
    # ────────────────────────────────────────────────────────────────────

    def _normalize_tag(self, tag):
        """Normalise tag number for deduplication (uppercase, stripped of spaces/dashes)."""
        return re.sub(r"[-_\s]+", "", (tag or "").upper().strip())

    def _merge_instruments(self, instruments):
        """
        Deduplicate instrument records from multiple passes.
        Rules:
        1. Same normalised tag → keep richest record (back-fill N/A fields).
        2. Partial-tag absorption: if a new tag's normalised key is a PREFIX of
           an existing longer tag (e.g. PI3700 vs PI370012), keep the longer one.
           This prevents OCR from generating both "PI-3700" and "PI-3700-12" as
           separate entries when they are the same physical instrument read at
           different precision levels.
        3. Records with no tag_number are always included as-is.
        """
        seen: dict[str, dict] = {}   # normalised_key → record
        no_tag: list = []

        def _norm_stripped(tag: str) -> str:
            """Remove ALL non-alphanumeric chars for prefix-match comparison."""
            return re.sub(r"[^A-Z0-9]", "", (tag or "").upper())

        for inst in instruments:
            raw_tag = inst.get("tag_number") or ""
            norm     = self._normalize_tag(raw_tag)
            norm_s   = _norm_stripped(raw_tag)   # for prefix checks

            if not norm:
                no_tag.append(inst)
                continue

            if norm in seen:
                # Same tag — back-fill missing fields
                existing = seen[norm]
                for key, value in inst.items():
                    if value and value != "N/A" and (
                        not existing.get(key) or existing[key] == "N/A"
                    ):
                        existing[key] = value
                continue

            # Check if the current tag is a prefix of an already-stored longer tag
            # e.g. "PI3700" is a prefix of "PI370012"
            is_prefix_of_existing = any(
                existing_norm_s.startswith(norm_s) and existing_norm_s != norm_s
                for existing_norm_s in (_norm_stripped(r.get("tag_number", "")) for r in seen.values())
            )
            if is_prefix_of_existing:
                # This candidate is less specific — skip it
                continue

            # Check if any existing tag is a prefix of this new (longer) tag
            # → replace the shorter existing tag with this more specific one
            to_remove = [
                k for k, r in seen.items()
                if norm_s.startswith(_norm_stripped(r.get("tag_number", "")))
                and _norm_stripped(r.get("tag_number", "")) != norm_s
                and _norm_stripped(r.get("tag_number", ""))  # not empty
            ]
            for k in to_remove:
                old = seen.pop(k)
                # Back-fill fields from the shorter record into the new longer one
                for key, value in old.items():
                    if value and value != "N/A" and (
                        not inst.get(key) or inst.get(key) == "N/A"
                    ):
                        inst[key] = value

            seen[norm] = dict(inst)

        return list(seen.values()) + no_tag

    def _call_vision(self, b64_image, prompt, mode_label, max_tokens=None):
        """Call gpt-4o with vision content; return parsed list or []."""
        cfg = self.extraction_config
        if max_tokens is None:
            max_tokens = cfg["max_tokens_primary"]
    # ────────────────────────────────────────────────────────────────────
    # Gemini Vision call
    # ────────────────────────────────────────────────────────────────────

    def _call_gemini_vision(self, b64_image, prompt, mode_label, max_tokens=None):
        """Call Gemini Vision (gemini-2.0-flash); return parsed list or []."""
        try:
            from google.genai import types as _gtypes
            cfg = self.extraction_config
            model = cfg.get("gemini_model", "gemini-2.0-flash")
            logger.info(f"[InstrumentIndex] Calling Gemini Vision ({mode_label}, model={model})…")

            image_bytes = base64.b64decode(b64_image)
            image_part  = _gtypes.Part.from_bytes(data=image_bytes, mime_type="image/jpeg")

            system_text = (
                "You are an expert P&ID analyst and process instrumentation engineer "
                "with 20+ years of experience reading ADNOC / oil & gas engineering drawings. "
                "You can identify ALL types of instrument symbols regardless of orientation — "
                "horizontal, vertical, rotated at any angle. "
                "Extract EVERY instrument tag you see. Return ONLY a valid JSON array."
            )

            response = self.gemini_client.models.generate_content(
                model=model,
                contents=[
                    _gtypes.Content(
                        role="user",
                        parts=[_gtypes.Part.from_text(text=prompt), image_part],
                    )
                ],
                config=_gtypes.GenerateContentConfig(
                    system_instruction=system_text,
                    temperature=cfg.get("temperature", 0.1),
                    max_output_tokens=max_tokens or cfg["max_tokens_primary"],
                ),
            )
            raw = response.text or ""
            logger.info(f"[InstrumentIndex] Gemini response {len(raw)} chars ({mode_label})")
            return self._parse_response(raw)

        except Exception as e:
            err_str = str(e)
            if "429" in err_str or "quota" in err_str.lower() or "rate" in err_str.lower():
                retry_delay = self.extraction_config.get("gemini_retry_delay", 5)
                logger.warning(
                    f"[InstrumentIndex] Gemini rate-limit ({mode_label}) — "
                    f"waiting {retry_delay}s then retrying once…"
                )
                time.sleep(retry_delay)
                # Single retry attempt
                try:
                    from google.genai import types as _gtypes
                    cfg = self.extraction_config
                    model = cfg.get("gemini_model", "gemini-2.0-flash")
                    image_bytes = base64.b64decode(b64_image)
                    image_part  = _gtypes.Part.from_bytes(data=image_bytes, mime_type="image/jpeg")
                    system_text = (
                        "You are an expert P&ID analyst and process instrumentation engineer "
                        "with 20+ years of experience reading ADNOC / oil & gas engineering drawings. "
                        "Extract EVERY instrument tag you see. Return ONLY a valid JSON array."
                    )
                    response = self.gemini_client.models.generate_content(
                        model=model,
                        contents=[_gtypes.Content(role="user", parts=[_gtypes.Part.from_text(text=prompt), image_part])],
                        config=_gtypes.GenerateContentConfig(
                            system_instruction=system_text,
                            temperature=cfg.get("temperature", 0.1),
                            max_output_tokens=max_tokens or cfg["max_tokens_primary"],
                        ),
                    )
                    raw = response.text or ""
                    logger.info(f"[InstrumentIndex] Gemini retry succeeded {len(raw)} chars ({mode_label})")
                    return self._parse_response(raw)
                except Exception as e2:
                    err2 = str(e2)
                    if "429" in err2 or "quota" in err2.lower() or "rate" in err2.lower():
                        self._gemini_quota_exceeded = True
                        logger.error(f"[InstrumentIndex] Gemini quota exhausted ({mode_label}) — Gemini disabled for this request")
                    else:
                        logger.error(f"[InstrumentIndex] Gemini retry failed ({mode_label}): {e2}")
                    return []
            else:
                logger.error(f"[InstrumentIndex] Gemini Vision error ({mode_label}): {e}")
            return []

    # ────────────────────────────────────────────────────────────────────
    # OpenAI Vision call
    # ────────────────────────────────────────────────────────────────────

    def _call_openai_vision(self, b64_image, prompt, mode_label, max_tokens=None):
        """Call OpenAI gpt-4o Vision; return parsed list or []."""
        cfg = self.extraction_config
        if max_tokens is None:
            max_tokens = cfg["max_tokens_primary"]

        if self._quota_exceeded:
            logger.warning(f"[InstrumentIndex] Skipping OpenAI call ({mode_label}) — quota exceeded")
            return []
        try:
            logger.info(f"[InstrumentIndex] Calling OpenAI Vision ({mode_label}, max_tokens={max_tokens})…")
            resp = self.openai_client.chat.completions.create(
                model=cfg.get("model", "gpt-4o"),
                messages=[
                    {
                        "role": "system",
                        "content": (
                            "You are an expert P&ID analyst and process instrumentation engineer "
                            "with 20+ years of experience reading ADNOC / oil & gas engineering drawings. "
                            "You can identify ALL types of instrument symbols regardless of orientation. "
                            "Extract EVERY instrument tag you see. Return ONLY a valid JSON array."
                        ),
                    },
                    {
                        "role": "user",
                        "content": [
                            {"type": "text", "text": prompt},
                            {
                                "type": "image_url",
                                "image_url": {
                                    "url": f"data:image/jpeg;base64,{b64_image}",
                                    "detail": "high",
                                },
                            },
                        ],
                    },
                ],
                max_tokens=max_tokens,
                temperature=cfg.get("temperature", 0.1),
            )
            raw = resp.choices[0].message.content
            finish = getattr(resp.choices[0], "finish_reason", "?")
            raw_len = len(raw or "")
            logger.info(f"[InstrumentIndex] OpenAI response {raw_len} chars finish={finish} ({mode_label})")
            if not raw or not raw.strip():
                # Empty content — surface the reason and don't try to parse
                logger.warning(
                    f"[InstrumentIndex] OpenAI returned EMPTY content ({mode_label}) "
                    f"finish_reason={finish} — likely content filter or refusal. "
                    f"Prompt length={len(prompt) if 'prompt' in locals() else '?'} chars."
                )
                return []
            return self._parse_response(raw)
        except Exception as e:
            err_str = str(e)
            if "429" in err_str or "insufficient_quota" in err_str or "rate_limit" in err_str.lower():
                self._quota_exceeded = True
                logger.error(f"[InstrumentIndex] OpenAI quota/rate-limit hit ({mode_label}) — OpenAI disabled")
            else:
                logger.error(f"[InstrumentIndex] OpenAI Vision error ({mode_label}): {e}", exc_info=True)
            return []

    # Keep _call_vision as alias for backward compatibility
    def _call_vision(self, b64_image, prompt, mode_label, max_tokens=None):
        return self._call_openai_vision(b64_image, prompt, mode_label, max_tokens)

    # ────────────────────────────────────────────────────────────────────
    # Prompt templates
    # ────────────────────────────────────────────────────────────────────

    def _build_prompt(self, drawing_info, page_no, extra_hint=""):
        type_list = ", ".join(
            f"{k} ({v['name']})" for k, v in INSTRUMENT_CATEGORIES.items()
        )
        hint_block = f"\n⚡ SCAN CONTEXT: {extra_hint}\n" if extra_hint else ""

        # ── Soft-coded category template injection ──────────────────
        tpl = get_template(drawing_info.get("project_category"))
        extra_fields = tpl.get("extra_fields") or []
        tpl_intro = tpl.get("prompt_intro") or ""

        # If an explicit project unit is provided, surface it in the prompt so
        # the AI prepends it to every tag automatically.
        explicit_unit = str(drawing_info.get("project_unit") or "").strip()
        unit_hint = ""
        if explicit_unit:
            unit_hint = (
                f"\nUNIT CODE FOR THIS PROJECT: {explicit_unit}. "
                f"Every instrument tag you return MUST start with '{explicit_unit}-' "
                f"(e.g. '{explicit_unit}-FT-1502', '{explicit_unit}-PSV-8501A').\n"
            )
        # Build the extra-field bullet list and the example record additions
        extra_field_block = ""
        extra_example_block = ""
        if extra_fields:
            extra_field_block = "\nADDITIONAL TEMPLATE FIELDS — extract when visible:\n"
            for f in extra_fields:
                extra_field_block += f"- {f['key']:<22}: {f['description']}\n"
            extra_example_block = "".join(
                f',\n    "{f["key"]}": "{f.get("default","-")}"' for f in extra_fields
            )
        tpl_intro_block = f"\n{tpl_intro}\n" if tpl_intro else ""

        return f"""
🎯 MISSION: Extract the COMPLETE Instrument Index from this P&ID drawing.
Page {page_no} — Drawing: {drawing_info.get('drawing_number', 'N/A')} — {drawing_info.get('drawing_title', 'N/A')}
Project: {drawing_info.get('project_name', 'N/A')}   Revision: {drawing_info.get('revision', '0')}
{tpl_intro_block}{unit_hint}{hint_block}
─────────────────────────────────────────────
WHAT TO EXTRACT
─────────────────────────────────────────────
Extract EVERY instrument tag visible on this drawing.  A P&ID typically has
15–60+ instruments.  Do NOT skip any.

Target instrument tag prefixes (non-exhaustive):
{type_list}

Tag format examples from ADNOC / oil & gas:
  FIT-3901-08A   TI-3901-01   PIT-3901-03   LIT-3601-01   SDV-3901-01
  SVZY-3901-03   MOV-3901-01  PSV-3901-01   RO-3901-01    XPD-3901-01

─────────────────────────────────────────────
WHERE TO LOOK — INSTRUMENT CIRCLES / BUBBLES ARE THE PRIMARY SOURCE
─────────────────────────────────────────────
In P&ID drawings, instrument tags are shown INSIDE CIRCLES (instrument bubbles).
These circles may contain:
  • A single line of text:  "FT-3901-01"
  • Two lines:              "FT" (top) and "3901-01" (bottom)
  • Three lines:            "FIT" + "3901" + "08A"
READ every circle regardless of its size, angle, or position.

Also scan:
1. Circles / bubbles on ALL process lines — horizontal AND vertical pipe runs
2. Tags printed VERTICALLY along pipe runs (rotate reading angle)
3. Circles inside control loops (dashed boxes)
4. Tags connected to equipment nozzles (any angle)
5. Instrument tables in title block or margins
6. Any isolated letter+number combination inside or near a symbol

─────────────────────────────────────────────
FIELDS TO EXTRACT PER INSTRUMENT
─────────────────────────────────────────────
For EACH instrument found, return:

- tag_number          : Full tag e.g. "PIT-3901-01" — REQUIRED
- control_system_tag  : The DCS / Control-System tag for this instrument.
                        • If the instrument bubble is HEXAGONAL or inside a SHARED-DISPLAY / DCS box → it IS a CS tag; use the same tag number.
                        • If a second tag is shown near the instrument (in a box, or labelled "CS TAG", "DCS TAG") → use that tag.
                        • If unknown → "N/A"
- instrument_type     : Full description e.g. "Pressure Indicating Transmitter"
- category            : e.g. Flow / Pressure / Temperature / Level / Safety / Shutdown & ESD / etc.
- pid_no              : P&ID drawing number (default: "{drawing_info.get('drawing_number','N/A')}")
- service_description : What the instrument measures (e.g. "Pig Receiver Inlet Pressure")
- line_number         : Process line tag where instrument is installed
- equipment_number    : Associated equipment tag (vessel, pump, compressor, etc.)
- loop_number         : Control/safety loop number if shown
- fail_safe           : Fail-safe position — "FC" (fail closed), "FO" (fail open), "FL" (fail last), "N/A"
- signal_type         : "4-20mA", "Discrete (0/1)", "HART", "Fieldbus", "Pneumatic", "N/A"
- set_point           : Alarm / trip set point if shown on drawing or in instrument list
- drawing_number      : "{drawing_info.get('drawing_number','N/A')}"
- revision            : "{drawing_info.get('revision','0')}"
- notes               : Any relevant remark, special service (H2S, NACE, SIL), or uncertainty
{extra_field_block}
─────────────────────────────────────────────
OUTPUT
─────────────────────────────────────────────
Return ONLY a JSON array — no markdown fences, no explanation text.
Example single record:
[
  {{
    "tag_number":         "PIT-3901-01",
    "control_system_tag": "PIC-3901-01",
    "instrument_type":    "Pressure Indicating Transmitter",
    "category":           "Pressure",
    "pid_no":             "{drawing_info.get('drawing_number','N/A')}",
    "service_description":"Pig Receiver Inlet Pressure",
    "line_number":        "10\\\"-G-3901-A2A",
    "equipment_number":   "LP-3901",
    "loop_number":        "3901",
    "fail_safe":          "N/A",
    "signal_type":        "4-20mA",
    "set_point":          "75 barg (PSHH)",
    "drawing_number":     "{drawing_info.get('drawing_number','N/A')}",
    "revision":           "{drawing_info.get('revision','0')}",
    "notes":              "SIL-rated loop"{extra_example_block}
  }}
]

⚠️ CRITICAL: Extract ALL instruments — including vertical/rotated text.
A response of [] or < 5 items for a process P&ID almost certainly means you missed instruments.
Scan in ALL orientations. Start response with [ and end with ].
"""

    def _build_fallback_prompt(self, drawing_info, page_no, extra_hint=""):
        """Simpler, more aggressive fallback prompt."""
        hint_block = f"\n⚡ {extra_hint}\n" if extra_hint else ""
        return f"""
EMERGENCY FALLBACK — Extract ALL instrument tags from this P&ID.
Page {page_no}  |  Drawing: {drawing_info.get('drawing_number', 'N/A')}
{hint_block}
Instructions:
1. Find EVERY circle or bubble containing a text tag on this drawing.
2. READ THE TAG even if it is printed vertically, upside-down, or at an angle.
3. Tag examples: FIT-1234, TI-56, SDV-3901-01, MOV-3901-02, LIT-101A, PT-8001, PSHH-001.
4. For each tag extract as much data as you can see.

Return JSON array only, format:
[
  {{
    "tag_number": "TAG-NO",
    "control_system_tag": "CS-TAG or N/A",
    "instrument_type": "Description",
    "category": "Category",
    "pid_no": "{drawing_info.get('drawing_number','N/A')}",
    "service_description": "what it measures",
    "line_number": "line tag or N/A",
    "equipment_number": "equipment tag or N/A",
    "loop_number": "N/A",
    "fail_safe": "N/A",
    "signal_type": "N/A",
    "set_point": "N/A",
    "drawing_number": "{drawing_info.get('drawing_number','N/A')}",
    "revision": "{drawing_info.get('revision','0')}",
    "notes": ""
  }}
]

Return ONLY the JSON array.
"""

    # ────────────────────────────────────────────────────────────────────
    # Response parsing helpers
    # ────────────────────────────────────────────────────────────────────

    def _parse_response(self, raw):
        """Extract JSON array from raw AI response text."""
        try:
            text = raw.strip()
            # Strip markdown fences
            for fence in ("```json", "```"):
                if fence in text:
                    start = text.find(fence) + len(fence)
                    end = text.find("```", start)
                    if end > start:
                        text = text[start:end].strip()
                        break

            # Find JSON array boundaries
            s = text.find("[")
            e = text.rfind("]") + 1
            if s >= 0 and e > s:
                data = json.loads(text[s:e])
                if isinstance(data, list):
                    return data

            # Try entire payload
            data = json.loads(text)
            if isinstance(data, list):
                return data
            if isinstance(data, dict) and "instruments" in data:
                return data["instruments"]

        except json.JSONDecodeError as exc:
            logger.error(f"[InstrumentIndex] JSON decode error: {exc}")
        except Exception as exc:
            logger.error(f"[InstrumentIndex] Parse error: {exc}", exc_info=True)
        return []

    def _enrich_category(self, inst):
        """Fill in instrument_type + category from INSTRUMENT_CATEGORIES if missing."""
        tag = (inst.get("tag_number") or "").strip().upper()
        if not tag:
            return inst

        # Extract function code (letters before first digit or dash-digit)
        match = re.match(r"^([A-Z]+)", tag)
        if not match:
            return inst
        code = match.group(1)

        cfg = INSTRUMENT_CATEGORIES.get(code)
        if cfg:
            if not inst.get("instrument_type"):
                inst["instrument_type"] = cfg["name"]
            if not inst.get("category"):
                inst["category"] = cfg["category"]
        return inst

    # ────────────────────────────────────────────────────────────────────
    # Excel export
    # ────────────────────────────────────────────────────────────────────

    def generate_excel(self, instruments, drawing_info):
        """
        Build an openpyxl workbook with two sheets:
          1. Instrument Index — row per instrument, category-coloured
          2. Summary          — count per category

        Returns bytes of the .xlsx file.
        """
        # ── Soft-coded schema lookup ───────────────────────────────────
        # Pick the column schema that matches the project category.
        # Falls back to the generic 16-column layout when not registered.
        category = (drawing_info or {}).get("project_category") or "default"
        schema = EXCEL_SCHEMAS.get(category)
        if schema:
            group_header, columns = schema
            use_accessor = True
        else:
            group_header, columns = None, EXCEL_COLUMNS
            use_accessor = False

        wb = openpyxl.Workbook()

        # ── Sheet 1: Instrument Index ────────────────────────────────────
        ws = wb.active
        ws.title = "Instrument Index"
        ws.sheet_view.showGridLines = True

        # Header style
        hdr_font   = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
        hdr_fill   = PatternFill("solid", fgColor="1F4E79")
        hdr_align  = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin       = Side(style="thin", color="CCCCCC")
        std_border = Border(left=thin, right=thin, top=thin, bottom=thin)

        # Title row
        ws.row_dimensions[1].height = 22
        title_cell = ws.cell(row=1, column=1, value="INSTRUMENT INDEX")
        title_cell.font = Font(name="Calibri", bold=True, size=14, color="1F4E79")
        title_cell.alignment = Alignment(horizontal="left", vertical="center")

        # Metadata row
        ws.row_dimensions[2].height = 16
        ws.cell(row=2, column=1, value=f"Drawing: {drawing_info.get('drawing_number','N/A')}")
        ws.cell(row=2, column=5, value=f"Title: {drawing_info.get('drawing_title','N/A')}")
        ws.cell(row=2, column=9, value=f"Rev: {drawing_info.get('revision','0')}")
        ws.cell(row=2, column=11, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")

        # ── Group header strip (merged cells) — only for schemas that define one
        if group_header:
            ws.row_dimensions[3].height = 22
            col_cursor = 1
            for label, span in group_header:
                start_col = col_cursor
                end_col = col_cursor + span - 1
                cell = ws.cell(row=3, column=start_col, value=label or None)
                if label:
                    cell.font = hdr_font
                    cell.fill = hdr_fill
                    cell.alignment = hdr_align
                    cell.border = std_border
                if span > 1:
                    ws.merge_cells(start_row=3, start_column=start_col,
                                   end_row=3, end_column=end_col)
                col_cursor = end_col + 1
            header_row = 4
            data_start = 5
        else:
            header_row = 4
            data_start = 5

        # Header row
        ws.row_dimensions[header_row].height = 30
        for col_idx, col_def in enumerate(columns, start=1):
            cell = ws.cell(row=header_row, column=col_idx, value=col_def["label"])
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = hdr_align
            cell.border = std_border
            ws.column_dimensions[cell.column_letter].width = col_def["width"]

        # Data rows
        for row_offset, inst in enumerate(instruments):
            row_no = data_start + row_offset
            ws.row_dimensions[row_no].height = 15

            cat_key = inst.get("category") or "Special"
            fill_hex  = CATEGORY_COLOURS.get(cat_key, "F5F5F5")
            row_fill  = PatternFill("solid", fgColor=fill_hex)
            std_font  = Font(name="Calibri", size=9)
            std_align = Alignment(vertical="center", wrap_text=False)

            for col_idx, col_def in enumerate(columns, start=1):
                if use_accessor:
                    val = col_def["accessor"](inst)
                else:
                    raw_val = inst.get(col_def["key"], "")
                    val = "" if raw_val == "N/A" else raw_val
                cell = ws.cell(row=row_no, column=col_idx, value=val)
                cell.font = std_font
                cell.fill = row_fill
                cell.alignment = std_align
                cell.border = std_border

        # Freeze header
        ws.freeze_panes = f"A{data_start}"

        # Auto-filter on header row
        ws.auto_filter.ref = (
            f"A{header_row}:{ws.cell(row=header_row, column=len(columns)).column_letter}{header_row}"
        )

        # ── Sheet 2: Summary ─────────────────────────────────────────────
        ws2 = wb.create_sheet("Summary")
        ws2.column_dimensions["A"].width = 30
        ws2.column_dimensions["B"].width = 14

        ws2.row_dimensions[1].height = 22
        ws2.cell(row=1, column=1, value="INSTRUMENT INDEX — SUMMARY").font = Font(
            bold=True, size=13, color="1F4E79"
        )
        ws2.cell(row=2, column=1, value=f"Drawing: {drawing_info.get('drawing_number','N/A')}")
        ws2.cell(row=2, column=2, value=f"Total: {len(instruments)}")

        ws2.row_dimensions[4].height = 22
        for col, hdr in [(1, "Category"), (2, "Count")]:
            c = ws2.cell(row=4, column=col, value=hdr)
            c.font = hdr_font
            c.fill = hdr_fill
            c.alignment = hdr_align
            c.border = std_border

        # Build category counts
        counts: dict[str, int] = {}
        for inst in instruments:
            cat = inst.get("category") or "Unknown"
            counts[cat] = counts.get(cat, 0) + 1

        row_n = 5
        for cat, cnt in sorted(counts.items()):
            fill_hex = CATEGORY_COLOURS.get(cat, "F5F5F5")
            for col, val in [(1, cat), (2, cnt)]:
                c = ws2.cell(row=row_n, column=col, value=val)
                c.font = Font(name="Calibri", size=10)
                c.fill = PatternFill("solid", fgColor=fill_hex)
                c.alignment = Alignment(vertical="center")
                c.border = std_border
            row_n += 1

        # Total row
        total_cell = ws2.cell(row=row_n, column=1, value="TOTAL")
        total_cell.font = Font(bold=True, size=10, color="1F4E79")
        ws2.cell(row=row_n, column=2, value=len(instruments)).font = Font(bold=True)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.getvalue()
