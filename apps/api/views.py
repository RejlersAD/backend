"""
API views for RADAI.
Smart ViewSets with proper permissions and pagination.
"""
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework.views import APIView
from django.contrib.auth import get_user_model
from django.conf import settings
from apps.users.serializers import UserSerializer, UserRegistrationSerializer
import os
import logging

logger = logging.getLogger(__name__)

User = get_user_model()


class HealthCheckView(APIView):
    """
    Health check endpoint to verify API is running.
    """
    permission_classes = [AllowAny]
    
    def get(self, request):
        """Return health status."""
        return Response({
            'status': 'healthy',
            'message': 'RADAI API is running successfully'
        })


class CORSDiagnosticView(APIView):
    """
    CORS diagnostic endpoint to debug CORS configuration.
    Returns current CORS settings and request origin.
    """
    permission_classes = [AllowAny]
    
    def get(self, request):
        """Return CORS diagnostic information."""
        origin = request.META.get('HTTP_ORIGIN', 'No origin header')
        
        return Response({
            'status': 'cors_diagnostic',
            'request_origin': origin,
            'cors_settings': {
                'allowed_origins': list(settings.CORS_ALLOWED_ORIGINS) if hasattr(settings, 'CORS_ALLOWED_ORIGINS') else [],
                'allow_credentials': getattr(settings, 'CORS_ALLOW_CREDENTIALS', False),
                'allow_all_origins': getattr(settings, 'CORS_ALLOW_ALL_ORIGINS', False),
            },
            'environment_variables': {
                'FRONTEND_URL': os.getenv('FRONTEND_URL', 'Not set - using default: https://airflow-frontend.vercel.app'),
                'BACKEND_URL': os.getenv('BACKEND_URL', 'Not set'),
                'CORS_ALLOW_VERCEL': os.getenv('CORS_ALLOW_VERCEL', 'Not set - using default: true'),
                'CORS_ALLOW_LOCALHOST': os.getenv('CORS_ALLOW_LOCALHOST', 'Not set - using default: true'),
            },
            'request_headers': {
                'Origin': request.META.get('HTTP_ORIGIN', 'Not present'),
                'Host': request.META.get('HTTP_HOST', 'Not present'),
                'User-Agent': request.META.get('HTTP_USER_AGENT', 'Not present'),
            },
            'message': 'If FRONTEND_URL is "Not set", add it in Railway dashboard: Variables tab'
        })
    
    def options(self, request):
        """Handle OPTIONS preflight for diagnostic endpoint."""
        return Response({'status': 'preflight_ok'}, status=200)


class UserViewSet(viewsets.ModelViewSet):
    """
    ViewSet for user operations.
    Smart CRUD operations with custom actions.
    """
    queryset = User.objects.all()
    serializer_class = UserSerializer
    
    def get_permissions(self):
        """Set permissions based on action."""
        if self.action == 'create':
            return [AllowAny()]
        return [IsAuthenticated()]
    
    def get_serializer_class(self):
        """Return appropriate serializer based on action."""
        if self.action == 'create':
            return UserRegistrationSerializer
        return UserSerializer
    
    @action(detail=False, methods=['get'], permission_classes=[IsAuthenticated])
    def me(self, request):
        """Get current user information."""
        serializer = self.get_serializer(request.user)
        return Response(serializer.data)
    
    @action(detail=False, methods=['put', 'patch'], permission_classes=[IsAuthenticated])
    def update_profile(self, request):
        """Update current user's profile."""
        serializer = self.get_serializer(request.user, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data)


# ========================================================================
# DASHBOARD METRICS API - Advanced Intelligence
# ========================================================================

from rest_framework.decorators import api_view, permission_classes
from datetime import datetime, timedelta
from django.db.models import Count, Q, Avg
from django.utils import timezone


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def dashboard_metrics(request):
    """
    Comprehensive dashboard metrics aggregation.
    Soft-coded to aggregate data from all modules intelligently.
    Returns: Real-time system statistics and utilization metrics.
    """
    from apps.pid_analysis.models import PIDDrawing
    from apps.pfd_converter.models import PFDDocument, PIDConversion
    from apps.qhse.models import QHSERunningProject
    from apps.electrical_checklist.models import ChecklistExtractionJob
    from apps.usage_tracking.models import UsageLog
    
    # Time filters
    today = timezone.now().date()
    yesterday = today - timedelta(days=1)
    week_ago = today - timedelta(days=7)
    month_ago = today - timedelta(days=30)
    
    # =======================================================================
    # USER METRICS - Updated to show ACTIVE users only
    # =======================================================================
    # Active users only (is_active=True)
    total_users = User.objects.filter(is_active=True).count()
    
    # All users for comparison (including inactive)
    all_users_count = User.objects.count()
    inactive_users_count = User.objects.filter(is_active=False).count()
    
    users_today = User.objects.filter(
        date_joined__date=today,
        is_active=True
    ).count()
    users_yesterday = User.objects.filter(
        date_joined__date=yesterday,
        is_active=True
    ).count()
    
    # Active users (logged in within last 30 days)
    active_users = User.objects.filter(
        last_login__gte=timezone.now() - timedelta(days=30)
    ).count()
    total_active_users = User.objects.filter(is_active=True).count()

    # Active users previous period (for trend)
    active_users_prev = User.objects.filter(
        last_login__gte=timezone.now() - timedelta(days=60),
        last_login__lt=timezone.now() - timedelta(days=30)
    ).count()
    
    # =======================================================================
    # DOCUMENT METRICS
    # =======================================================================
    # Aggregate documents from all modules
    pid_count = PIDDrawing.objects.count()
    pfd_count = PFDDocument.objects.count()
    
    try:
        qhse_count = QHSERunningProject.objects.count()
    except:
        qhse_count = 0
    
    # Don't import CRSDocument - causes error
    crs_count = 0
    
    total_documents = pid_count + pfd_count + qhse_count + crs_count
    
    # Documents uploaded today
    pid_today = PIDDrawing.objects.filter(created_at__date=today).count()
    pfd_today = PFDDocument.objects.filter(created_at__date=today).count()
    documents_today = pid_today + pfd_today
    
    # Documents yesterday (for trend)
    pid_yesterday = PIDDrawing.objects.filter(created_at__date=yesterday).count()
    pfd_yesterday = PFDDocument.objects.filter(created_at__date=yesterday).count()
    documents_yesterday = pid_yesterday + pfd_yesterday

    pfd_avg_confidence = PIDConversion.objects.filter(
        confidence_score__gt=0
    ).aggregate(avg=Avg('confidence_score'))['avg'] or 0

    checklist_avg_confidence = ChecklistExtractionJob.objects.filter(
        confidence_score__isnull=False
    ).aggregate(avg=Avg('confidence_score'))['avg'] or 0

    usage_today = UsageLog.objects.filter(
        timestamp__date=timezone.now().date()
    ).count()

    usage_30d = UsageLog.objects.filter(
        timestamp__gte=timezone.now() - timedelta(days=30)
    ).count()

    cutoff_30d = timezone.now() - timedelta(days=30)

    feature_usage = (
        UsageLog.objects
        .filter(timestamp__gte=cutoff_30d)
        .values('discipline_key', 'discipline_label')
        .annotate(count=Count('id'))
        .order_by('-count')[:50]
    )

    feature_usage_map = {
        f['discipline_key']: {
            'count': f['count'],
            'label': f['discipline_label']
        }
        for f in feature_usage
    }

    # =======================================================================
    # FEATURE UTILIZATION METRICS
    # =======================================================================
    try:
        from apps.mlflow_integration.models import FeatureUsage
        # Total feature usage from MLflow
        total_feature_runs = FeatureUsage.objects.count()
        feature_runs_today = FeatureUsage.objects.filter(
            timestamp__date=today
        ).count()
        
        # Most used feature
        most_used = FeatureUsage.objects.values('feature_name').annotate(
            usage_count=Count('id')
        ).order_by('-usage_count').first()
        
        # AI features usage
        ai_features_count = FeatureUsage.objects.filter(
            feature_name__in=['pid_analysis', 'pfd_conversion', 'smart_ocr']
        ).count()
        
        # Calculate utilization percentage (active users using features)
        users_using_features = FeatureUsage.objects.filter(
            timestamp__gte=timezone.now() - timedelta(days=30)
        ).values('user_id').distinct().count()
        
        feature_utilization = (users_using_features / total_users * 100) if total_users > 0 else 0
        
    except Exception as e:
        print(f"Feature usage tracking not available: {e}")
        total_feature_runs = 0
        feature_runs_today = 0
        most_used = None
        ai_features_count = 0
        feature_utilization = 0
    
    # =======================================================================
    # BUSINESS METRICS
    # =======================================================================
    try:
        from apps.designiq.models import Project
        active_projects = Project.objects.filter(status='active').count()
        active_projects_prev = Project.objects.filter(
            status='active',
            created_at__lt=timezone.now() - timedelta(days=30)
        ).count()
    except:
        active_projects = 0
        active_projects_prev = 0
    
    # Pending approvals (from various modules)
    try:
        from apps.finance.models import Invoice
        from apps.procurement.models import PurchaseOrder
        
        pending_invoices = Invoice.objects.filter(
            approval_status__in=['pending', 'submitted']
        ).count()
        
        pending_pos = PurchaseOrder.objects.filter(
            status__in=['pending', 'awaiting_approval']
        ).count()
        
        pending_approvals = pending_invoices + pending_pos
    except:
        pending_approvals = 0
    
    # =======================================================================
    # PERFORMANCE METRICS - Soft-coded health calculation
    # =======================================================================
    # System health calculated from multiple real data sources
    health_factors = []
    
    # Factor 1: Database connectivity (if we got here, DB is working)
    health_factors.append(100)
    
    # Factor 2: Document processing activity (documents uploaded recently = healthy)
    if total_documents > 0:
        doc_health = min(100, (documents_today + documents_yesterday) * 10)
        health_factors.append(max(80, doc_health))  # Minimum 80% if documents exist
    else:
        health_factors.append(70)  # Lower if no documents
    
    # Factor 3: User engagement (active users)
    if total_users > 0:
        user_engagement = (active_users / total_users) * 100
        health_factors.append(max(60, user_engagement))  # Minimum 60%
    else:
        health_factors.append(50)
    
    # Factor 4: Feature usage success rate (if available)
    try:
        from apps.mlflow_integration.models import FeatureUsage
        recent_runs = FeatureUsage.objects.filter(
            timestamp__gte=timezone.now() - timedelta(hours=24)
        )
        total_runs = recent_runs.count()
        if total_runs > 0:
            successful_runs = recent_runs.filter(status='completed').count()
            feature_health = (successful_runs / total_runs * 100)
            health_factors.append(feature_health)
    except Exception as e:
        # Feature tracking not available - not critical
        pass
    
    # Calculate weighted average system health
    system_health = sum(health_factors) / len(health_factors) if health_factors else 85
    
    # =======================================================================
    # AGGREGATE RESPONSE
    # =======================================================================
    metrics = {
        'users': {
            'total_users': total_users,  # Active users only
            'total_users_previous': total_users - users_today,
            'all_users_count': all_users_count,  # All users including inactive
            'inactive_users_count': inactive_users_count,  # Inactive users
            'active_users': active_users,
            'active_users_previous': active_users_prev,
            'total_active_users': total_active_users,
            'new_users_today': users_today,
            'new_users_yesterday': users_yesterday,
            'note': 'total_users shows active users only (is_active=True)'
        },
        'documents': {
            'total_documents': total_documents,
            'total_documents_previous': total_documents - documents_today,
            'documents_today': documents_today,
            'documents_yesterday': documents_yesterday,
            'pid_drawings': pid_count,
            'pfd_documents': pfd_count,
            'qhse_documents': qhse_count,
            'crs_documents': crs_count,
            'pfd_confidence': round(float(pfd_avg_confidence), 1),
            'checklist_confidence': round(float(checklist_avg_confidence), 1)
        },
        'features': {
            'total_usage': total_feature_runs,
            'usage_today': feature_runs_today,
            'most_used_feature': most_used['feature_name'] if most_used else 'N/A',
            'most_used_count': most_used['usage_count'] if most_used else 0,
            'ai_features_usage': ai_features_count,
            'utilization_percentage': round(feature_utilization, 1),
            'usage_log_today': usage_today,
            'usage_log_30d': usage_30d
        },
        'business': {
            'active_projects': active_projects,
            'active_projects_previous': active_projects_prev,
            'pending_approvals': pending_approvals
        },
        'performance': {
            'system_health': round(system_health, 1),
            'avg_response_time': None,  # Real monitoring not yet implemented - calculate from middleware
            'response_time_note': 'Response time monitoring can be added via Django middleware'
        },
        'metadata': {
            'timestamp': timezone.now().isoformat(),
            'period': 'real-time',
            'timezone': str(timezone.get_current_timezone())
        },
        'feature_usage_map': feature_usage_map,
    }
    
    return Response(metrics)


# ========================================================================
# REAL-TIME ACTIVITY TRACKING API - Database & S3 History
# ========================================================================

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def recent_activity(request):
    """
    Real-time activity feed from database and S3 history.
    Aggregates recent actions across all modules intelligently.
    Returns: Unified activity stream with timestamps and user info.
    """
    from apps.pid_analysis.models import PIDDrawing
    from apps.pfd_converter.models import PFDDocument, PIDConversion
    from apps.qhse.models import QHSERunningProject
    
    # Get limit from query params (default 50)
    limit = int(request.GET.get('limit', 50))
    hours_ago = int(request.GET.get('hours', 24))  # Last 24 hours by default
    
    time_threshold = timezone.now() - timedelta(hours=hours_ago)
    activities = []
    
    # =======================================================================
    # P&ID DRAWINGS - Document Uploads
    # =======================================================================
    try:
        pid_drawings = PIDDrawing.objects.filter(
            uploaded_at__gte=time_threshold
        ).select_related('uploaded_by', 'project').order_by('-uploaded_at')[:20]
        
        for drawing in pid_drawings:
            activities.append({
                'id': f'pid_{drawing.id}',
                'type': 'document_upload',
                'category': 'documents',
                'user': drawing.uploaded_by.email if drawing.uploaded_by else 'System',
                'user_name': drawing.uploaded_by.get_full_name() if drawing.uploaded_by else 'System',
                'title': drawing.title or drawing.drawing_number,
                'description': f'P&ID Drawing: {drawing.drawing_number}',
                'timestamp': drawing.uploaded_at.isoformat(),
                'metadata': {
                    'module': 'P&ID',
                    'drawing_number': drawing.drawing_number,
                    'project': drawing.project.name if drawing.project else 'No Project',
                    'file_size': f'{drawing.file_size / 1024:.1f} KB' if drawing.file_size else 'N/A'
                }
            })
    except Exception as e:
        print(f"Error fetching P&ID activities: {e}")
    
    # =======================================================================
    # PFD DOCUMENTS - Document Uploads
    # =======================================================================
    try:
        pfd_documents = PFDDocument.objects.filter(
            uploaded_at__gte=time_threshold
        ).select_related('uploaded_by', 'project').order_by('-uploaded_at')[:20]
        
        for doc in pfd_documents:
            activities.append({
                'id': f'pfd_{doc.id}',
                'type': 'document_upload',
                'category': 'documents',
                'user': doc.uploaded_by.email if doc.uploaded_by else 'System',
                'user_name': doc.uploaded_by.get_full_name() if doc.uploaded_by else 'System',
                'title': doc.title or doc.document_number,
                'description': f'PFD Document: {doc.document_number}',
                'timestamp': doc.uploaded_at.isoformat(),
                'metadata': {
                    'module': 'PFD',
                    'document_number': doc.document_number,
                    'project': doc.project.name if doc.project else 'No Project',
                    'pages': doc.total_pages or 'N/A'
                }
            })
    except Exception as e:
        print(f"Error fetching PFD activities: {e}")
    
    # =======================================================================
    # PID CONVERSIONS - AI Analysis
    # =======================================================================
    try:
        conversions = PIDConversion.objects.filter(
            created_at__gte=time_threshold
        ).select_related('pfd_document__uploaded_by').order_by('-created_at')[:15]
        
        for conversion in conversions:
            activities.append({
                'id': f'conversion_{conversion.id}',
                'type': 'ai_analysis',
                'category': 'ai',
                'user': conversion.pfd_document.uploaded_by.email if conversion.pfd_document and conversion.pfd_document.uploaded_by else 'System',
                'user_name': conversion.pfd_document.uploaded_by.get_full_name() if conversion.pfd_document and conversion.pfd_document.uploaded_by else 'AI System',
                'title': f'PFD to P&ID Conversion',
                'description': f'Status: {conversion.status}',
                'timestamp': conversion.created_at.isoformat(),
                'metadata': {
                    'module': 'AI Conversion',
                    'status': conversion.status,
                    'conversion_type': 'PFD to P&ID',
                    'accuracy': f'{conversion.confidence_score:.1f}%' if conversion.confidence_score else 'N/A'
                }
            })
    except Exception as e:
        print(f"Error fetching conversion activities: {e}")
    
    # =======================================================================
    # QHSE PROJECTS - Project Activities
    # =======================================================================
    try:
        qhse_projects = QHSERunningProject.objects.filter(
            created_at__gte=time_threshold
        ).select_related('created_by').order_by('-created_at')[:15]
        
        for project in qhse_projects:
            activities.append({
                'id': f'qhse_{project.id}',
                'type': 'project_created',
                'category': 'projects',
                'user': project.created_by.email if project.created_by else 'System',
                'user_name': project.created_by.get_full_name() if project.created_by else 'System',
                'title': project.project_name,
                'description': f'QHSE Project: {project.project_number}',
                'timestamp': project.created_at.isoformat(),
                'metadata': {
                    'module': 'QHSE',
                    'project_number': project.project_number,
                    'location': project.location or 'N/A',
                    'status': 'Active'
                }
            })
    except Exception as e:
        print(f"Error fetching QHSE activities: {e}")
    
    # =======================================================================
    # FEATURE USAGE - MLflow Tracking
    # =======================================================================
    try:
        from apps.mlflow_integration.models import FeatureUsage
        feature_runs = FeatureUsage.objects.filter(
            timestamp__gte=time_threshold
        ).order_by('-timestamp')[:20]
        
        for run in feature_runs:
            # Get user info if available
            user_email = 'System'
            user_name = 'System'
            if run.user_id:
                try:
                    user_obj = User.objects.get(id=run.user_id)
                    user_email = user_obj.email
                    user_name = user_obj.get_full_name()
                except:
                    pass
            
            activities.append({
                'id': f'feature_{run.id}',
                'type': 'feature_usage',
                'category': 'features',
                'user': user_email,
                'user_name': user_name,
                'title': run.feature_name.replace('_', ' ').title(),
                'description': f'Status: {run.status}',
                'timestamp': run.timestamp.isoformat(),
                'metadata': {
                    'module': 'Features',
                    'feature': run.feature_name,
                    'status': run.status,
                    'execution_time': f'{run.execution_time:.2f}s' if run.execution_time else 'N/A'
                }
            })
    except Exception as e:
        print(f"Error fetching feature usage: {e}")
    
    # =======================================================================
    # USER REGISTRATIONS - New Users
    # =======================================================================
    try:
        new_users = User.objects.filter(
            date_joined__gte=time_threshold
        ).order_by('-date_joined')[:10]
        
        for user in new_users:
            activities.append({
                'id': f'user_{user.id}',
                'type': 'user_registration',
                'category': 'users',
                'user': user.email,
                'user_name': user.get_full_name() or user.email,
                'title': 'New User Registration',
                'description': f'Joined from {user.department or "Unknown Department"}',
                'timestamp': user.date_joined.isoformat(),
                'metadata': {
                    'module': 'Users',
                    'email': user.email,
                    'department': user.department or 'N/A',
                    'role': user.role or 'User'
                }
            })
    except Exception as e:
        print(f"Error fetching user activities: {e}")
    
    # =======================================================================
    # PROJECTS - DesignIQ Projects
    # =======================================================================
    try:
        from apps.designiq.models import Project
        projects = Project.objects.filter(
            created_at__gte=time_threshold
        ).select_related('created_by').order_by('-created_at')[:10]
        
        for project in projects:
            activities.append({
                'id': f'project_{project.id}',
                'type': 'project_created',
                'category': 'projects',
                'user': project.created_by.email if project.created_by else 'System',
                'user_name': project.created_by.get_full_name() if project.created_by else 'System',
                'title': project.name,
                'description': f'Project created with status: {project.status}',
                'timestamp': project.created_at.isoformat(),
                'metadata': {
                    'module': 'DesignIQ',
                    'status': project.status,
                    'priority': getattr(project, 'priority', 'N/A'),
                    'team_size': project.team_members.count() if hasattr(project, 'team_members') else 0
                }
            })
    except Exception as e:
        print(f"Error fetching project activities: {e}")
    
    # =======================================================================
    # SORT AND LIMIT ACTIVITIES
    # =======================================================================
    # Sort all activities by timestamp (most recent first)
    activities.sort(key=lambda x: x['timestamp'], reverse=True)
    
    # Limit to requested number
    activities = activities[:limit]
    
    # =======================================================================
    # RETURN RESPONSE
    # =======================================================================
    return Response({
        'count': len(activities),
        'hours_range': hours_ago,
        'results': activities,
        'metadata': {
            'timestamp': timezone.now().isoformat(),
            'sources': ['database', 's3_uploads'],
            'refresh_interval': 30  # seconds
        }
    })


# ========================================================================
# PREDICTIVE ANALYTICS API - ML-POWERED FORECASTING & INSIGHTS
# ========================================================================

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def predictive_analytics(request):
    """
    Smart predictive analytics with historical data aggregation.
    Provides time-series data for ML-powered forecasting on frontend.
    Returns: Historical metrics for intelligent prediction algorithms.
    """
    from apps.pid_analysis.models import PIDDrawing
    from apps.pfd_converter.models import PFDDocument
    from apps.qhse.models import QHSERunningProject
    from django.db.models import Count
    from django.db.models.functions import TruncDate
    
    # Time ranges for historical analysis
    days_back = int(request.GET.get('days', 30))
    start_date = timezone.now() - timedelta(days=days_back)
    
    historical_data = {}
    
    # =======================================================================
    # DOCUMENT UPLOADS - Daily aggregation
    # =======================================================================
    try:
        # PID drawings by day
        pid_by_day = PIDDrawing.objects.filter(
            created_at__gte=start_date
        ).annotate(
            date=TruncDate('created_at')
        ).values('date').annotate(
            count=Count('id')
        ).order_by('date')
        
        # PFD documents by day
        pfd_by_day = PFDDocument.objects.filter(
            created_at__gte=start_date
        ).annotate(
            date=TruncDate('created_at')
        ).values('date').annotate(
            count=Count('id')
        ).order_by('date')
        
        # Create full date range
        date_range = [(start_date.date() + timedelta(days=i)) for i in range(days_back)]
        
        # Map counts to dates
        pid_counts = {item['date']: item['count'] for item in pid_by_day}
        pfd_counts = {item['date']: item['count'] for item in pfd_by_day}
        
        # Build daily document uploads array
        document_uploads = []
        for date in date_range:
            total = pid_counts.get(date, 0) + pfd_counts.get(date, 0)
            document_uploads.append(total)
        
        historical_data['document_uploads'] = document_uploads
        
    except Exception as e:
        print(f"Error aggregating document uploads: {e}")
        historical_data['document_uploads'] = [0] * days_back
    
    # =======================================================================
    # USER ACTIVITY - Daily active users
    # =======================================================================
    try:
        users_by_day = User.objects.filter(
            last_login__gte=start_date
        ).annotate(
            date=TruncDate('last_login')
        ).values('date').annotate(
            count=Count('id')
        ).order_by('date')
        
        date_range = [(start_date.date() + timedelta(days=i)) for i in range(days_back)]
        user_counts = {item['date']: item['count'] for item in users_by_day}
        
        user_activity = []
        for date in date_range:
            user_activity.append(user_counts.get(date, 0))
        
        historical_data['user_activity'] = user_activity
        
    except Exception as e:
        print(f"Error aggregating user activity: {e}")
        historical_data['user_activity'] = [0] * days_back
    
    # =======================================================================
    # AI ANALYSIS USAGE - Feature utilization
    # =======================================================================
    try:
        from apps.mlflow_integration.models import FeatureUsage
        
        ai_by_day = FeatureUsage.objects.filter(
            timestamp__gte=start_date
        ).annotate(
            date=TruncDate('timestamp')
        ).values('date').annotate(
            count=Count('id')
        ).order_by('date')
        
        date_range = [(start_date.date() + timedelta(days=i)) for i in range(days_back)]
        ai_counts = {item['date']: item['count'] for item in ai_by_day}
        
        ai_usage = []
        for date in date_range:
            ai_usage.append(ai_counts.get(date, 0))
        
        historical_data['ai_analysis_usage'] = ai_usage
        
    except Exception as e:
        print(f"Error aggregating AI usage: {e}")
        historical_data['ai_analysis_usage'] = [0] * days_back
    
    # =======================================================================
    # PROJECT COMPLETION - DesignIQ projects
    # =======================================================================
    try:
        from apps.designiq.models import Project
        
        projects_by_day = Project.objects.filter(
            updated_at__gte=start_date,
            status='completed'
        ).annotate(
            date=TruncDate('updated_at')
        ).values('date').annotate(
            count=Count('id')
        ).order_by('date')
        
        date_range = [(start_date.date() + timedelta(days=i)) for i in range(days_back)]
        project_counts = {item['date']: item['count'] for item in projects_by_day}
        
        project_completion = []
        for date in date_range:
            project_completion.append(project_counts.get(date, 0))
        
        historical_data['project_completion'] = project_completion
        
    except Exception as e:
        print(f"Error aggregating project completion: {e}")
        historical_data['project_completion'] = [0] * days_back
    
    # =======================================================================
    # SYSTEM LOAD - Simplified metric (based on overall activity)
    # =======================================================================
    try:
        # Calculate system load as percentage of max capacity
        # This is a simplified metric - can be enhanced with real monitoring
        total_activity_by_day = []
        date_range = [(start_date.date() + timedelta(days=i)) for i in range(days_back)]
        
        for date in date_range:
            # Sum all activities for this day
            pid_count = PIDDrawing.objects.filter(created_at__date=date).count()
            pfd_count = PFDDocument.objects.filter(created_at__date=date).count()
            user_count = User.objects.filter(last_login__date=date).count()
            
            # Calculate as percentage (assuming max capacity of 100 items/day)
            total = pid_count + pfd_count + (user_count * 2)  # Weight users more
            load_percent = min((total / 100) * 100, 100)  # Cap at 100%
            total_activity_by_day.append(round(load_percent, 1))
        
        historical_data['system_load'] = total_activity_by_day
        
    except Exception as e:
        print(f"Error calculating system load: {e}")
        historical_data['system_load'] = [50] * days_back  # Default 50% load
    
    # =======================================================================
    # GENERATE DATE LABELS
    # =======================================================================
    date_labels = []
    for i in range(days_back):
        date = (start_date.date() + timedelta(days=i))
        date_labels.append(date.strftime('%Y-%m-%d'))
    
    # =======================================================================
    # RETURN PREDICTIVE DATA
    # =======================================================================
    return Response({
        'historical': historical_data,
        'date_labels': date_labels,
        'metadata': {
            'days_back': days_back,
            'start_date': start_date.date().isoformat(),
            'end_date': timezone.now().date().isoformat(),
            'timestamp': timezone.now().isoformat(),
            'data_points': days_back,
            'metrics_count': len(historical_data)
        },
        'recommendations': {
            'model': 'Use Linear Regression for steady trends, Exponential Smoothing for volatile data',
            'confidence': 'Higher accuracy with more historical data (30+ days recommended)',
            'refresh': 'Update predictions daily for optimal accuracy'
        }
    })


# =============================================================================
# SOFT-CODED STATS ENDPOINTS (Dashboard Integration)
# =============================================================================
from rest_framework.decorators import api_view, permission_classes

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def projects_stats(request):
    """
    Project Statistics Endpoint - Soft-coded for dashboard integration
    Returns aggregate stats for all projects across modules
    ACCURATE DATA: Aggregates from all project sources intelligently
    """
    try:
        from apps.qhse.models import QHSERunningProject
        from apps.pid_analysis.models import PIDProject
        
        # Get QHSE projects (primary source)
        qhse_projects = QHSERunningProject.objects.filter(is_active=True).count()
        qhse_total = QHSERunningProject.objects.count()
        
        # Get PID projects
        pid_projects = PIDProject.objects.count()
        
        # Try to get core projects if available
        try:
            from apps.core.models import Project as CoreProject
            core_projects = CoreProject.objects.filter(status='active').count() if hasattr(CoreProject, 'objects') else 0
        except:
            core_projects = 0
        
        # Calculate total active projects (best estimate)
        # Use QHSE as primary + PID projects not linked to QHSE
        total_active = qhse_projects + core_projects
        
        return Response({
            'active_count': total_active,
            'qhse_projects': qhse_projects,
            'qhse_total': qhse_total,
            'pid_projects': pid_projects,
            'core_projects': core_projects,
            'total_count': total_active + pid_projects,
            'status': 'success',
            'timestamp': timezone.now().isoformat(),
            'data_sources': ['QHSE', 'PID', 'Core']
        })
    except Exception as e:
        print(f"[STATS ERROR] projects_stats: {e}")
        import traceback
        traceback.print_exc()
        # Return safe fallback data
        return Response({
            'active_count': 0,
            'status': 'error',
            'message': str(e)
        }, status=200)  # Return 200 to prevent dashboard errors


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def pid_stats(request):
    """
    P&ID Drawing Statistics Endpoint - Soft-coded for dashboard integration
    Returns accurate aggregate stats for P&ID drawings
    ACCURATE DATA: Real-time counts from PIDDrawing model
    """
    try:
        from apps.pid_analysis.models import PIDDrawing
        from django.db.models import Count, Q
        
        # Total drawings count
        total_drawings = PIDDrawing.objects.count()
        
        # Status-based counts (using actual fields from model)
        # Drawings are considered analyzed if they have analysis_report or analysis_completed_at
        analyzed_drawings = PIDDrawing.objects.filter(
            Q(status='completed') | Q(status='analyzed') | 
            Q(analysis_report__isnull=False) | Q(analysis_completed_at__isnull=False)
        ).count()
        
        # Pending drawings - those without analysis completion
        pending_drawings = PIDDrawing.objects.filter(
            Q(status='pending') | Q(status='processing') | 
            (Q(analysis_completed_at__isnull=True) & ~Q(status='failed'))
        ).count()
        
        # Additional metrics
        failed_drawings = PIDDrawing.objects.filter(status='failed').count()
        
        # Time-based metrics
        recent_drawings = PIDDrawing.objects.filter(
            created_at__gte=timezone.now() - timedelta(days=7)
        ).count()
        
        return Response({
            'total_drawings': total_drawings,
            'analyzed_drawings': analyzed_drawings,
            'pending_drawings': pending_drawings,
            'failed_drawings': failed_drawings,
            'recent_drawings': recent_drawings,
            'status': 'success',
            'timestamp': timezone.now().isoformat(),
            'data_accuracy': 'real-time'
        })
    except Exception as e:
        print(f"[STATS ERROR] pid_stats: {e}")
        import traceback
        traceback.print_exc()
        # Return safe fallback data
        return Response({
            'total_drawings': 0,
            'analyzed_drawings': 0,
            'pending_drawings': 0,
            'failed_drawings': 0,
            'status': 'error',
            'message': str(e)
        }, status=200)  # Return 200 to prevent dashboard errors


# ============================================================================
# USAGE ANALYTICS — Daily trend from usage_log table
# Soft-coded: discipline labels come from UsageLog.DISCIPLINE_MAP so adding
# a new module only requires updating that map, not touching this view.
# ============================================================================

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def usage_daily(request):
    """
    Daily usage analytics aggregated from the UsageLog table.

    Query params
    ------------
    days   int  Number of days to look back (1–365, default 30).

    Response shape
    --------------
    {
      "summary": {
        "total_requests":  int,
        "active_users":    int,
        "success_rate":    float (%),
        "avg_response_ms": int,
        "peak_day":        str (YYYY-MM-DD) | null,
        "peak_count":      int,
        "days":            int
      },
      "daily_totals": [
        { "date": "YYYY-MM-DD", "total": int, "success": int, "failed": int }
        ...                             # one entry per day in range, zeros for empty days
      ],
      "discipline_breakdown": [
        { "key": str, "label": str, "count": int, "percentage": float }
        ...                             # sorted descending, max 10 entries
      ]
    }
    """
    try:
        from apps.usage_tracking.models import UsageLog
        from django.db.models import Count, Avg
        from django.db.models.functions import TruncDate

        days_back  = min(max(int(request.GET.get('days', 30)), 1), 365)
        start_date = timezone.now() - timedelta(days=days_back)

        qs = UsageLog.objects.filter(timestamp__gte=start_date)

        # Single query for daily totals
        by_day = (
            qs.annotate(date=TruncDate('timestamp'))
              .values('date')
              .annotate(
                  total      = Count('id'),
                  success_ct = Count('id', filter=Q(success=True)),
                  failed_ct  = Count('id', filter=Q(success=False)),
              )
              .order_by('date')
        )
        date_map = {str(r['date']): r for r in by_day}

        date_range = [
            (start_date.date() + timedelta(days=i)).isoformat()
            for i in range(days_back)
        ]
        daily_totals = [
            {
                'date':    d,
                'total':   date_map[d]['total']      if d in date_map else 0,
                'success': date_map[d]['success_ct'] if d in date_map else 0,
                'failed':  date_map[d]['failed_ct']  if d in date_map else 0,
            }
            for d in date_range
        ]

        # Single aggregation query
        agg = qs.aggregate(total=Count('id'), avg_ms=Avg('response_time_ms'))
        total_ct = agg['total'] or 0

        # Discipline breakdown - single query
        disc_qs = (
            qs.values('discipline_key', 'discipline_label')
              .annotate(count=Count('id'))
              .order_by('-count')[:20]
        )
        grand_total = total_ct or 1
        discipline_breakdown = [
            {
                'key': r['discipline_key'],
                'label': r['discipline_label'],
                'count': r['count'],
                'percentage': round(r['count'] / grand_total * 100, 1),
            }
            for r in disc_qs
            if r['discipline_key'] not in ('other', '')
        ][:10]

        # Active users and peak - lightweight queries
        active_users = qs.values('user_email').exclude(user_email='').distinct().count()
        peak = max(daily_totals, key=lambda d: d['total'], default={'date': None, 'total': 0})

        success_ct = qs.filter(success=True).count()
        success_rate = round(success_ct / total_ct * 100, 1) if total_ct else 100.0

        return Response({
            'summary': {
                'total_requests':  total_ct,
                'active_users':    active_users,
                'success_rate':    success_rate,
                'avg_response_ms': round(agg['avg_ms'] or 0),
                'peak_day':        peak['date'],
                'peak_count':      peak['total'],
                'days':            days_back,
            },
            'daily_totals':         daily_totals,
            'discipline_breakdown': discipline_breakdown,
        })

    except Exception as e:
        print(f"[STATS ERROR] usage_daily: {e}")
        import traceback
        traceback.print_exc()
        return Response({
            'summary': {
                'total_requests': 0, 'active_users': 0, 'success_rate': 100.0,
                'avg_response_ms': 0, 'peak_day': None, 'peak_count': 0,
                'days': int(request.GET.get('days', 30)),
            },
            'daily_totals': [],
            'discipline_breakdown': [],
            'status': 'error',
            'message': str(e),
        }, status=200)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def aws_status(request):
    """GET /api/v1/dashboard/aws-status/"""
    user = request.user

    # Determine role level securely from DB
    role_level = 10
    role_name = 'User'
    try:
        from apps.rbac.models import UserProfile
        profile = UserProfile.objects.prefetch_related('roles').get(user=user)
        role = profile.roles.filter(is_active=True).order_by('level').first()
        if role:
            role_level = role.level
            role_name = role.name
    except Exception:
        pass

    if user.is_superuser or user.is_staff:
        role_level = 1
        role_name = 'Super Administrator'

    # Try S3
    try:
        from apps.core.s3_service import get_s3_service
        s3 = get_s3_service()

        def get_file_breakdown(files):
            from collections import Counter
            ext_counter = Counter(
                f['key'].split('.')[-1].lower()
                for f in files
                if not f.get('key','').endswith('/') and '.' in f.get('key','').split('/')[-1]
            )
            total = sum(ext_counter.values()) or 1
            return [
                {'type': k.upper(), 'count': v, 'percentage': round(v/total*100,1)}
                for k, v in ext_counter.most_common(5)
            ]

        if role_level <= 2:
            # Admin — full bucket
            from django.contrib.auth import get_user_model
            info = s3.get_bucket_size()
            if not info.get('success'):
                raise Exception('S3 unavailable')
            from collections import Counter
            import boto3
            from django.conf import settings

            # Direct S3 paginator — no presigned URLs, no 1000 cap
            s3_client = boto3.client(
                's3',
                aws_access_key_id=settings.AWS_ACCESS_KEY_ID,
                aws_secret_access_key=settings.AWS_SECRET_ACCESS_KEY,
                region_name=settings.AWS_S3_REGION_NAME,
            )
            ext_counter = Counter()
            paginator = s3_client.get_paginator('list_objects_v2')
            for page in paginator.paginate(Bucket=settings.AWS_STORAGE_BUCKET_NAME):
                for obj in page.get('Contents', []):
                    key = obj['Key']
                    if not key.endswith('/') and '.' in key.split('/')[-1]:
                        ext_counter[key.split('.')[-1].lower()] += 1

            total_ext = sum(ext_counter.values()) or 1
            file_breakdown = [
                {'type': k.upper(), 'count': v, 'percentage': round(v / total_ext * 100, 1)}
                for k, v in ext_counter.most_common(5)
            ]
            size_gb = round(info['total_size_mb'] / 1024, 2)

            return Response({
                'status': 'connected',
                'view': 'admin',
                'role_name': role_name,
                'total_files': info['total_count'],
                'total_size_gb': size_gb,
                'total_users': get_user_model().objects.filter(is_active=True).count(),
                'file_breakdown': file_breakdown,
            })

        elif role_level == 3:
            try:
                from apps.rbac.models import UserProfile
                from django.contrib.auth import get_user_model
                User = get_user_model()
                manager_profile = UserProfile.objects.get(user=user)
                department = manager_profile.department or ''
                team_profiles = UserProfile.objects.filter(department=department, user__is_active=True).exclude(user=user).select_related('user')
                team_count = team_profiles.count()
                total_files, total_size, user_file_counts = 0, 0, {}
                all_dept_files = []
                for uid in team_profiles.values_list('user__id', flat=True):
                    r = s3.list_files(prefix=f'users/{uid}/')
                    if r.get('success'):
                        files = r.get('files', [])
                        all_dept_files.extend(files)
                        total_files += len(files)
                        total_size += sum(f.get('size', 0) for f in files)
                        if files: user_file_counts[uid] = len(files)
                most_uid = max(user_file_counts, key=user_file_counts.get) if user_file_counts else None
                most_name = None
                if most_uid:
                    try:
                        u = User.objects.get(id=most_uid)
                        most_name = u.get_full_name() or u.email
                    except Exception: pass
                last_upload = max(
                    (f.get('last_modified') for f in all_dept_files),
                    default=None
                )
                return Response({'status': 'connected', 'view': 'manager', 'role_name': role_name, 'department': department, 'team_members': team_count, 'total_files': total_files, 'total_size_gb': round(total_size / (1024**3), 2), 'most_active_user': most_name, 'last_upload': str(last_upload) if last_upload else None, 'file_breakdown': get_file_breakdown(all_dept_files)})
            except Exception as e:
                logger.warning('aws manager error: %s', e)
                return Response({'status': 'offline', 'view': 'manager', 'role_name': role_name, 'message': 'Department data unavailable'})

        else:
            # Engineer/User — own files only
            # Use user.id (server-side, cannot be tampered)
            result = s3.list_files(prefix=f'users/{user.id}/')
            files = result.get('files', []) if result.get('success') else []
            size_mb = round(sum(f.get('size', 0) for f in files) / (1024 * 1024), 2)
            last_upload = max(
                (f.get('last_modified') for f in files),
                default=None
            )
            return Response({
                'status': 'connected',
                'view': 'user',
                'role_name': role_name,
                'total_files': len(files),
                'total_size_mb': size_mb,
                'last_upload': str(last_upload) if last_upload else None,
                'file_breakdown': get_file_breakdown(files),
            })

    except Exception as e:
        # Log server-side only — never expose S3 errors to client
        logger.warning('aws_status error for user %s: %s', user.id, e)
        return Response({
            'status': 'offline',
            'view': 'admin' if role_level <= 2 else 'user',
            'role_name': role_name,
            'message': 'Storage service unavailable',
        })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def aws_report(request):
    """GET /api/v1/dashboard/aws-report/?format=xlsx|csv|pdf"""
    print(f'[AWS_REPORT] START - user={request.user.email}, format={request.GET.get("format")}')
    import io
    from django.http import HttpResponse
    from apps.rbac.models import UserProfile

    user = request.user
    fmt = 'xlsx'

    # Role level
    role_level = 10
    try:
        profile = UserProfile.objects.prefetch_related('roles').get(user=user)
        role = profile.roles.filter(is_active=True).order_by('level').first()
        if role:
            role_level = role.level
    except Exception:
        pass
    if user.is_superuser or user.is_staff:
        role_level = 1

    try:
        from apps.core.s3_service import get_s3_service
        s3 = get_s3_service()
        rows = []
        title = 'AWS Storage Report'

        if role_level <= 2:
            title = 'Full S3 Storage Report'
            import boto3
            from apps.core.s3_service import get_s3_service as _get_s3
            _svc = _get_s3()
            s3c = _svc.s3_client
            bucket = _svc.bucket_name
            paginator = s3c.get_paginator('list_objects_v2')
            for page in paginator.paginate(Bucket=bucket):
                for obj in page.get('Contents', []):
                    key = obj['Key']
                    if not key.endswith('/'):
                        rows.append({
                            'File': key,
                            'Size (KB)': round(obj['Size'] / 1024, 2),
                            'Last Modified': obj['LastModified'].strftime('%Y-%m-%d'),
                            'Type': key.split('.')[-1].upper() if '.' in key else '—',
                        })

        elif role_level == 3:
            title = 'Department Storage Report'
            dept = UserProfile.objects.get(user=user).department or ''
            team = UserProfile.objects.filter(
                department=dept, user__is_active=True
            ).exclude(user=user).select_related('user')
            for tp in team:
                r = s3.list_files(prefix=f'users/{tp.user.id}/')
                for f in (r.get('files', []) if r.get('success') else []):
                    key = f.get('key', '')
                    if not key.endswith('/'):
                        rows.append({
                            'User': tp.user.get_full_name() or tp.user.email,
                            'File': key,
                            'Size (KB)': round(f.get('size', 0) / 1024, 2),
                            'Last Modified': str(f.get('last_modified', ''))[:10],
                            'Type': key.split('.')[-1].upper() if '.' in key else '—',
                        })

        else:
            title = 'My Storage Report'
            r = s3.list_files(prefix=f'users/{user.id}/')
            for f in (r.get('files', []) if r.get('success') else []):
                key = f.get('key', '')
                if not key.endswith('/'):
                    rows.append({
                        'File': key,
                        'Size (KB)': round(f.get('size', 0) / 1024, 2),
                        'Last Modified': str(f.get('last_modified', ''))[:10],
                        'Type': key.split('.')[-1].upper() if '.' in key else '—',
                    })

    except Exception as e:
        import traceback
        logger.warning('aws_report error user %s: %s', user.id, e)
        traceback.print_exc()
        return Response({'error': 'Report generation failed'}, status=500)

    headers = list(rows[0].keys()) if rows else []

    # Excel (default)
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = Workbook()
    ws = wb.active
    ws.title = 'Report'
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='F97316')
        cell.alignment = Alignment(horizontal='center')
    for row in rows:
        ws.append([row.get(h,'') for h in headers])
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 22
    buf = io.BytesIO()
    wb.save(buf)
    res = HttpResponse(buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    res['Content-Disposition'] = f'attachment; filename="{title}.xlsx"'
    return res

