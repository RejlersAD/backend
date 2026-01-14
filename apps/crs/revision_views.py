"""
CRS Multiple Revision Views
AI-powered revision chain management with intelligent insights
"""

from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django.db import transaction
from django.utils import timezone
from django.db.models import Q, Count, Sum, Avg
from django.shortcuts import get_object_or_404

from .revision_models import (
    CRSRevisionChain, CRSRevision, CRSCommentLink,
    CRSAIInsight, CRSRevisionActivity
)
from .models import CRSDocument, CRSComment, CRSActivity
from .revision_serializers import (
    CRSRevisionChainListSerializer, CRSRevisionChainDetailSerializer,
    CRSRevisionChainCreateSerializer, CRSRevisionSerializer,
    CRSRevisionCreateSerializer, CRSCommentLinkSerializer,
    CRSCommentLinkCreateSerializer, CRSAIInsightSerializer,
    CRSAIInsightFeedbackSerializer, CRSRevisionActivitySerializer
)
"""
CRS Multiple Revision Views
AI-powered revision chain management with intelligent insights
"""

from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django.db import transaction
from django.utils import timezone
from django.db.models import Q, Count, Sum, Avg
from django.shortcuts import get_object_or_404
from io import BytesIO

from .revision_models import (
    CRSRevisionChain, CRSRevision, CRSCommentLink,
    CRSAIInsight, CRSRevisionActivity
)
from .models import CRSDocument, CRSComment, CRSActivity
from .revision_serializers import (
    CRSRevisionChainListSerializer, CRSRevisionChainDetailSerializer,
    CRSRevisionChainCreateSerializer, CRSRevisionSerializer,
    CRSRevisionCreateSerializer, CRSCommentLinkSerializer,
    CRSCommentLinkCreateSerializer, CRSAIInsightSerializer,
    CRSAIInsightFeedbackSerializer, CRSRevisionActivitySerializer
)
from .ai_service import CRSRevisionAIService

# Use unified extractor - single source of truth
from apps.core.helpers.unified_comment_extractor import extract_reviewer_comments, convert_comments_to_dict_list

import logging
logger = logging.getLogger(__name__)


class CRSRevisionChainViewSet(viewsets.ModelViewSet):
    """
    ViewSet for CRS Revision Chain management
    Handles multi-revision tracking and AI-powered insights
    """
    
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        """Get chains with optional filtering"""
        queryset = CRSRevisionChain.objects.all()
        
        # Filter by status
        status_filter = self.request.query_params.get('status')
        if status_filter:
            queryset = queryset.filter(status=status_filter)
        
        # Filter by risk level
        risk_filter = self.request.query_params.get('risk_level')
        if risk_filter:
            queryset = queryset.filter(risk_level=risk_filter)
        
        # Filter by project
        project = self.request.query_params.get('project')
        if project:
            queryset = queryset.filter(project_name__icontains=project)
        
        # Search
        search = self.request.query_params.get('search')
        if search:
            queryset = queryset.filter(
                Q(chain_id__icontains=search) |
                Q(document_title__icontains=search) |
                Q(document_number__icontains=search) |
                Q(project_name__icontains=search)
            )
        
        return queryset.prefetch_related('revisions', 'revisions__document')
    
    def get_serializer_class(self):
        """Use detailed serializer for retrieve, list serializer for list"""
        if self.action == 'retrieve':
            return CRSRevisionChainDetailSerializer
        elif self.action in ['create', 'update', 'partial_update']:
            return CRSRevisionChainCreateSerializer
        return CRSRevisionChainListSerializer
    
    def perform_create(self, serializer):
        """Create chain and log activity"""
        chain = serializer.save(created_by=self.request.user)
        
        # Log activity
        CRSRevisionActivity.objects.create(
            chain=chain,
            action='chain_created',
            description=f'Revision chain created: {chain.chain_id}',
            performed_by=self.request.user
        )
    
    @action(detail=True, methods=['post'])
    def add_revision(self, request, pk=None):
        """
        Add a new revision to the chain
        POST /api/v1/crs/revision-chains/{id}/add_revision/
        Body: {
            "document_id": 123,
            "revision_label": "Rev 2",
            "parent_revision_id": 456,  // optional
            "submitted_date": "2025-01-15T10:00:00Z",
            "notes": "Additional comments from client"
        }
        """
        chain = self.get_object()
        
        document_id = request.data.get('document_id')
        revision_label = request.data.get('revision_label')
        parent_revision_id = request.data.get('parent_revision_id')
        submitted_date = request.data.get('submitted_date')
        notes = request.data.get('notes', '')
        
        if not document_id or not revision_label:
            return Response(
                {"error": "document_id and revision_label are required"},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            document = CRSDocument.objects.get(id=document_id)
        except CRSDocument.DoesNotExist:
            return Response(
                {"error": "Document not found"},
                status=status.HTTP_404_NOT_FOUND
            )
        
        # Check if document already has a revision
        if hasattr(document, 'revision_info'):
            return Response(
                {"error": "This document is already linked to a revision"},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Get parent revision if specified
        parent_revision = None
        if parent_revision_id:
            try:
                parent_revision = CRSRevision.objects.get(id=parent_revision_id, chain=chain)
            except CRSRevision.DoesNotExist:
                return Response(
                    {"error": "Parent revision not found in this chain"},
                    status=status.HTTP_404_NOT_FOUND
                )
        
        # Calculate revision number
        revision_number = chain.total_revisions + 1
        
        with transaction.atomic():
            # Create revision
            revision = CRSRevision.objects.create(
                chain=chain,
                document=document,
                revision_number=revision_number,
                revision_label=revision_label,
                parent_revision=parent_revision,
                submitted_date=submitted_date or timezone.now(),
                status='submitted',
                notes=notes
            )
            
            # Update chain
            chain.total_revisions += 1
            chain.current_revision_number = revision_number
            chain.save()
            
            # If there's a parent revision, detect comment links
            if parent_revision:
                self._auto_link_comments(parent_revision, revision, request.user)
            
            # Calculate AI metrics
            self._calculate_revision_ai_metrics(revision)
            self._update_chain_ai_metrics(chain)
            
            # Log activity
            CRSRevisionActivity.objects.create(
                chain=chain,
                revision=revision,
                action='revision_added',
                description=f'Added {revision_label} to chain',
                performed_by=request.user,
                new_value={'revision_number': revision_number, 'revision_label': revision_label}
            )
        
        return Response({
            "success": True,
            "message": f"Revision {revision_label} added successfully",
            "data": CRSRevisionSerializer(revision).data
        }, status=status.HTTP_201_CREATED)
    
    @action(detail=True, methods=['post'])
    def upload_and_add_revision(self, request, pk=None):
        """
        Upload PDF and automatically process + add as revision to chain
        POST /api/v1/crs/revision-chains/{id}/upload_and_add_revision/
        
        Body (multipart/form-data):
            - file: PDF file (required)
            - revision_label: string (required, e.g., "Rev 2")
            - parent_revision_id: int (optional)
            - submitted_date: datetime (optional)
            - notes: string (optional)
            - project_name: string (optional)
            - document_number: string (optional)
            - contractor: string (optional)
            - department: string (optional)
        
        This endpoint:
        1. Uploads the PDF
        2. Extracts comments using CRS extraction logic
        3. Creates a CRSDocument
        4. Adds it as a revision to the chain
        5. Auto-links comments if parent exists
        6. Runs AI analysis
        """
        from django.db import connection
        import os
        
        # Ensure database connection is healthy
        try:
            if connection.connection and connection.connection.closed:
                logger.warning("Database connection closed, reopening...")
                connection.close()
                connection.ensure_connection()
        except Exception as conn_error:
            logger.warning(f"Connection check error: {conn_error}")
        
        chain = self.get_object()
        
        # Get uploaded file
        uploaded_file = request.FILES.get('file')
        if not uploaded_file:
            return Response({
                "error": "PDF file is required",
                "success": False
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Validate file type
        if not uploaded_file.name.lower().endswith('.pdf'):
            return Response({
                "error": "Only PDF files are supported",
                "success": False
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Get parameters
        revision_label = request.data.get('revision_label')
        if not revision_label:
            return Response({
                "error": "revision_label is required",
                "success": False
            }, status=status.HTTP_400_BAD_REQUEST)
        
        parent_revision_id = request.data.get('parent_revision_id')
        submitted_date = request.data.get('submitted_date')
        notes = request.data.get('notes', '')
        
        # Metadata
        metadata = {
            'project_name': request.data.get('project_name', chain.project_name),
            'document_number': request.data.get('document_number', chain.document_number),
            'contractor': request.data.get('contractor', ''),
            'department': request.data.get('department', ''),
        }
        
        try:
            # Get parent revision if specified
            parent_revision = None
            if parent_revision_id:
                try:
                    parent_revision = CRSRevision.objects.get(id=parent_revision_id, chain=chain)
                except CRSRevision.DoesNotExist:
                    return Response({
                        "error": "Parent revision not found in this chain",
                        "success": False
                    }, status=status.HTTP_404_NOT_FOUND)
            
            # Save PDF content for processing
            pdf_content = uploaded_file.read()
            pdf_buffer = BytesIO(pdf_content)
            
            # Extract comments using UNIFIED EXTRACTOR - single source of truth
            logger.info(f"Extracting comments from PDF: {uploaded_file.name}")
            
            # Reset buffer for unified extractor
            pdf_buffer.seek(0)
            
            # Extract using unified method (same as CRS document management)
            reviewer_comments = extract_reviewer_comments(
                pdf_buffer,
                apply_cleaning=True,
                filter_reviewers=True
            )
            logger.info(f"Extracted comments: {len(reviewer_comments) if reviewer_comments else 0}")
            
            if not reviewer_comments:
                return Response({
                    "error": "No comments found in the PDF file",
                    "success": False,
                    "message": "The PDF does not contain any extractable reviewer comments"
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Convert to dictionary format for processing
            processed_comments = convert_comments_to_dict_list(reviewer_comments)
            logger.info(f"Processed comments after filtering: {len(processed_comments) if processed_comments else 0}")
            
            if not processed_comments:
                return Response({
                    "error": "No valid comments after filtering",
                    "success": False,
                    "message": "All extracted text was filtered out as technical drawing elements"
                }, status=status.HTTP_400_BAD_REQUEST)
            
            logger.info(f"Extracted {len(processed_comments)} comments from PDF")
            
            with transaction.atomic():
                # Reset buffer for saving
                pdf_buffer.seek(0)
                
                # Create CRSDocument
                document = CRSDocument.objects.create(
                    document_name=metadata['document_number'] or f"{chain.document_title} - {revision_label}",
                    document_number=metadata['document_number'] or chain.document_number,
                    revision_number=revision_label,
                    project_name=metadata['project_name'],
                    contractor_name=metadata['contractor'],
                    uploaded_by=request.user,
                    status='completed'
                )
                
                # Save PDF file to document
                from django.core.files.base import ContentFile
                document.pdf_file.save(uploaded_file.name, ContentFile(pdf_content), save=True)
                
                # Create comments
                created_comments = []
                for idx, comment_data in enumerate(processed_comments, start=1):
                    comment = CRSComment.objects.create(
                        document=document,
                        serial_number=idx,
                        comment_text=comment_data['text'],
                        page_number=comment_data['page'],
                        clause_number=comment_data.get('clause', ''),
                        comment_type=comment_data['type'],
                        status='open'
                    )
                    created_comments.append(comment)
                
                # Calculate revision number
                revision_number = chain.total_revisions + 1
                
                # Create revision
                revision = CRSRevision.objects.create(
                    chain=chain,
                    document=document,
                    revision_number=revision_number,
                    revision_label=revision_label,
                    parent_revision=parent_revision,
                    submitted_date=submitted_date or timezone.now(),
                    status='submitted',
                    notes=notes,
                    total_new_comments=len(created_comments)
                )
                
                # Update chain
                chain.total_revisions += 1
                chain.current_revision_number = revision_number
                chain.save()
                
                # If there's a parent revision, detect comment links (with error handling)
                if parent_revision:
                    try:
                        self._auto_link_comments(parent_revision, revision, request.user)
                    except Exception as link_error:
                        logger.error(f"Error auto-linking comments (non-critical): {link_error}")
                        # Continue execution - linking is not critical for upload success
                
                # Calculate AI metrics (with error handling)
                try:
                    self._calculate_revision_ai_metrics(revision)
                    self._update_chain_ai_metrics(chain)
                except Exception as ai_error:
                    logger.warning(f"Error calculating AI metrics (non-critical): {ai_error}")
                
                # Log activity (with error handling)
                try:
                    self._safe_create_activity(
                        chain=chain,
                        revision=revision,
                        action='revision_uploaded',
                        description=f'Uploaded and processed {revision_label} ({len(created_comments)} comments)',
                        performed_by=request.user,
                        new_value={
                            'revision_number': revision_number,
                            'revision_label': revision_label,
                            'total_comments': len(created_comments)
                        }
                    )
                except Exception as activity_error:
                    logger.warning(f"Could not log activity (non-critical): {activity_error}")
            
            return Response({
                "success": True,
                "message": f"Revision {revision_label} uploaded and processed successfully",
                "data": {
                    "revision": CRSRevisionSerializer(revision).data,
                    "document": {
                        "id": document.id,
                        "title": document.document_name,
                        "document_number": document.document_number
                    },
                    "comments": [  # Full ReviewerComment data structure
                        {
                            "page_number": c.get('page'),
                            "reviewer_name": c.get('reviewer', 'Not Provided'),
                            "comment_text": c.get('text', ''),
                            "comment_type": c.get('type', 'GENERAL'),
                            "discipline": c.get('discipline', 'Not Provided'),
                            "drawing_ref": c.get('section_reference', 'N/A'),
                            "status": "Open"
                        } for c in processed_comments
                    ],
                    "extraction_summary": {
                        "total_comments": len(created_comments),
                        "red_comments": sum(1 for c in processed_comments if c['type'] == 'red_comment'),
                        "yellow_boxes": sum(1 for c in processed_comments if c['type'] == 'yellow_box'),
                        "pages_with_comments": len(set(c['page'] for c in processed_comments))
                    }
                }
            }, status=status.HTTP_201_CREATED)
                
        except Exception as e:
            error_msg = str(e)
            logger.error(f"Error processing PDF upload: {error_msg}", exc_info=True)
            
            # Provide more specific error messages
            if 'ssl' in error_msg.lower() or 'eof' in error_msg.lower():
                error_detail = "Database connection issue. Please try uploading again."
            elif 'connection' in error_msg.lower():
                error_detail = "Database connection interrupted. Please retry the upload."
            else:
                error_detail = str(e)
            
            return Response({
                "error": "Failed to process PDF",
                "success": False,
                "details": error_detail,
                "retry_recommended": True
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @action(detail=True, methods=['post'])
    def analyze_chain(self, request, pk=None):
        """
        Run comprehensive AI analysis on the chain
        POST /api/v1/crs/revision-chains/{id}/analyze_chain/
        """
        chain = self.get_object()
        
        with transaction.atomic():
            # Update AI metrics
            self._update_chain_ai_metrics(chain)
            
            # Generate AI insights
            insights = self._generate_chain_insights(chain, request.user)
            
            # Log activity
            CRSRevisionActivity.objects.create(
                chain=chain,
                action='ai_insight_generated',
                description=f'AI analysis completed, generated {len(insights)} insights',
                performed_by=request.user
            )
        
        return Response({
            "success": True,
            "message": "Chain analysis completed",
            "data": {
                "chain": CRSRevisionChainDetailSerializer(chain).data,
                "insights": CRSAIInsightSerializer(insights, many=True).data
            }
        })
    
    @action(detail=False, methods=['get'])
    def dashboard_summary(self, request):
        """
        Get dashboard summary statistics
        GET /api/v1/crs/revision-chains/dashboard_summary/
        """
        total_chains = CRSRevisionChain.objects.count()
        active_chains = CRSRevisionChain.objects.filter(status='active').count()
        total_revisions = CRSRevision.objects.count()
        total_comments = CRSComment.objects.count()
        
        # Get chains with most revisions
        top_chains = CRSRevisionChain.objects.annotate(
            revision_count=Count('revisions')
        ).order_by('-revision_count')[:5]
        
        # Recent activity
        recent_activities = CRSRevisionActivity.objects.select_related(
            'chain', 'revision', 'performed_by'
        ).order_by('-performed_at')[:10]
        
        return Response({
            "summary": {
                "total_chains": total_chains,
                "active_chains": active_chains,
                "total_revisions": total_revisions,
                "total_comments": total_comments
            },
            "top_chains": CRSRevisionChainListSerializer(top_chains, many=True).data,
            "recent_activities": CRSRevisionActivitySerializer(recent_activities, many=True).data
        })
    
    @action(detail=True, methods=['get'])
    def export_excel(self, request, pk=None):
        """
        Export all revisions and comments to Excel
        GET /api/v1/crs/revision-chains/{id}/export_excel/
        
        Returns Excel file with multiple sheets:
        - Chain Summary
        - All Revisions (with metadata)
        - All Comments (from all revisions)
        - Comment Links (between revisions)
        """
        from django.http import HttpResponse
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        import io
        
        chain = self.get_object()
        revisions = chain.revisions.all().order_by('revision_number')
        
        # Create workbook
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # ===== SHEET 1: Chain Summary =====
        ws_summary = wb.create_sheet("Chain Summary")
        
        # Header style
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=12)
        
        # Chain info
        ws_summary.append(["Revision Chain Summary"])
        ws_summary['A1'].font = Font(bold=True, size=14)
        ws_summary.append([])
        
        summary_data = [
            ["Document Title", chain.document_title],
            ["Document Number", chain.document_number],
            ["Project Name", chain.project_name],
            ["Status", chain.get_status_display()],
            ["Total Revisions", chain.total_revisions],
            ["Current Revision", chain.current_revision_number],
            ["Created Date", chain.created_at.strftime("%Y-%m-%d %H:%M")],
            ["Last Updated", chain.updated_at.strftime("%Y-%m-%d %H:%M")],
            ["", ""],
            ["AI Metrics", ""],
            ["Total Comments (All Revisions)", chain.total_comments_across_revisions],
            ["Resolved Comments", chain.resolved_comments_count],
            ["Pending Comments", chain.pending_comments_count],
            ["Resolution Rate", f"{chain.resolution_rate:.1f}%"],
            ["Comment Reduction Rate", f"{chain.comment_reduction_rate:.1f}%"],
        ]
        
        for row in summary_data:
            ws_summary.append(row)
        
        # Format summary sheet
        for row in ws_summary.iter_rows(min_row=3, max_row=len(summary_data)+2, min_col=1, max_col=1):
            for cell in row:
                cell.font = Font(bold=True)
        
        ws_summary.column_dimensions['A'].width = 30
        ws_summary.column_dimensions['B'].width = 40
        
        # ===== SHEET 2: All Revisions =====
        ws_revisions = wb.create_sheet("All Revisions")
        
        # Headers
        revision_headers = [
            "Rev #", "Rev Label", "Status", "Submitted Date", 
            "Total Comments", "Red Comments", "Yellow Comments",
            "Resolved", "Pending", "Resolution %", 
            "Document Title", "Document Number", "Notes"
        ]
        ws_revisions.append(revision_headers)
        
        # Style headers
        for col_num, header in enumerate(revision_headers, 1):
            cell = ws_revisions.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Add revision data
        for rev in revisions:
            # Get comment counts by type
            comments = CRSComment.objects.filter(document=rev.document)
            red_count = comments.filter(comment_type='red_comment').count()
            yellow_count = comments.filter(comment_type='yellow_box').count()
            resolved_count = comments.filter(status__in=['resolved', 'accepted']).count()
            pending_count = comments.filter(status__in=['open', 'in_progress']).count()
            resolution_pct = (resolved_count / rev.total_comments * 100) if rev.total_comments > 0 else 0
            
            ws_revisions.append([
                rev.revision_number,
                rev.revision_label,
                rev.get_status_display(),
                rev.submitted_date.strftime("%Y-%m-%d") if rev.submitted_date else "",
                rev.total_comments,
                red_count,
                yellow_count,
                resolved_count,
                pending_count,
                f"{resolution_pct:.1f}%",
                rev.document.document_name,
                rev.document.document_number,
                rev.notes or ""
            ])
        
        # Auto-size columns
        for col_num in range(1, len(revision_headers) + 1):
            ws_revisions.column_dimensions[get_column_letter(col_num)].width = 15
        
        # ===== SHEET 3: All Comments =====
        ws_comments = wb.create_sheet("All Comments")
        
        # Headers
        comment_headers = [
            "Revision", "Comment #", "Page", "Clause", "Comment Type",
            "Comment Text", "Reviewer", "Discipline", "Status", 
            "Response", "Action Taken"
        ]
        ws_comments.append(comment_headers)
        
        # Style headers
        for col_num, header in enumerate(comment_headers, 1):
            cell = ws_comments.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Add all comments from all revisions
        row_num = 2
        for rev in revisions:
            comments = CRSComment.objects.filter(document=rev.document).order_by('comment_number')
            
            for comment in comments:
                ws_comments.append([
                    rev.revision_label,
                    comment.comment_number,
                    comment.page_number,
                    comment.clause or "",
                    comment.get_comment_type_display(),
                    comment.comment_text,
                    comment.reviewer_name or "",
                    comment.discipline or "",
                    comment.get_status_display(),
                    comment.contractor_response or "",
                    comment.action_taken or ""
                ])
                
                # Color code by comment type
                type_cell = ws_comments.cell(row=row_num, column=5)
                if comment.comment_type == 'red_comment':
                    type_cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
                elif comment.comment_type == 'yellow_box':
                    type_cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
                
                # Color code by status
                status_cell = ws_comments.cell(row=row_num, column=9)
                if comment.status in ['resolved', 'accepted']:
                    status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                elif comment.status == 'open':
                    status_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                
                row_num += 1
        
        # Auto-size columns
        for col_num in range(1, len(comment_headers) + 1):
            if col_num == 6:  # Comment Text
                ws_comments.column_dimensions[get_column_letter(col_num)].width = 50
            elif col_num in [10, 11]:  # Response, Action
                ws_comments.column_dimensions[get_column_letter(col_num)].width = 40
            else:
                ws_comments.column_dimensions[get_column_letter(col_num)].width = 15
        
        # ===== SHEET 4: Comment Links =====
        ws_links = wb.create_sheet("Comment Links")
        
        # Headers
        link_headers = [
            "Parent Rev", "Parent Comment #", "Parent Comment",
            "Child Rev", "Child Comment #", "Child Comment",
            "Link Type", "Similarity", "Status Change"
        ]
        ws_links.append(link_headers)
        
        # Style headers
        for col_num, header in enumerate(link_headers, 1):
            cell = ws_links.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Add comment links
        for rev in revisions:
            if rev.parent_revision:
                links = CRSCommentLink.objects.filter(
                    child_revision=rev
                ).select_related('parent_comment', 'child_comment')
                
                for link in links:
                    parent_status = link.parent_comment.get_status_display()
                    child_status = link.child_comment.get_status_display()
                    status_change = f"{parent_status} → {child_status}"
                    
                    ws_links.append([
                        link.parent_revision.revision_label,
                        link.parent_comment.comment_number,
                        link.parent_comment.comment_text[:100] + "..." if len(link.parent_comment.comment_text) > 100 else link.parent_comment.comment_text,
                        link.child_revision.revision_label,
                        link.child_comment.comment_number,
                        link.child_comment.comment_text[:100] + "..." if len(link.child_comment.comment_text) > 100 else link.child_comment.comment_text,
                        link.get_link_type_display(),
                        f"{link.similarity_score:.2f}",
                        status_change
                    ])
        
        # Auto-size columns
        for col_num in range(1, len(link_headers) + 1):
            if col_num in [3, 6]:  # Comment text columns
                ws_links.column_dimensions[get_column_letter(col_num)].width = 40
            else:
                ws_links.column_dimensions[get_column_letter(col_num)].width = 15
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Create response
        filename = f"CRS_Chain_{chain.document_number}_{chain.id}_Export.xlsx"
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Log activity
        CRSRevisionActivity.objects.create(
            chain=chain,
            action='data_exported',
            description=f'Exported chain data to Excel ({revisions.count()} revisions)',
            performed_by=request.user
        )
        
        return response
    
    @action(detail=True, methods=['get'])
    def export_simple(self, request, pk=None):
        """
        Smart HTML-to-Excel export with proper error handling
        GET /api/v1/crs/revision-chains/{id}/export_simple/
        
        Inspired by CRS Document Management's proven export patterns
        """
        try:
            from django.http import HttpResponse
            from django.utils.html import escape
            import logging
            
            logger = logging.getLogger(__name__)
            logger.info(f"[EXPORT] Starting export_simple for chain {pk}")
            
            # Get chain with error handling
            try:
                chain = self.get_object()
                logger.info(f"[EXPORT] Found chain: {chain.chain_id}")
            except Exception as e:
                logger.error(f"[EXPORT] Chain not found: {e}")
                return HttpResponse(
                    f"<html><body><h1>Error: Chain not found</h1><p>{escape(str(e))}</p></body></html>",
                    content_type='text/html',
                    status=404
                )
            
            # Get revisions
            revisions = chain.revisions.all().order_by('revision_number')
            revision_count = revisions.count()
            logger.info(f"[EXPORT] Found {revision_count} revisions")
            
            # Build HTML with table structure (Excel-compatible)
            html_parts = []
            
            # Header with encoding
            html_parts.append('<?xml version="1.0"?>')
            html_parts.append('<?mso-application progid="Excel.Sheet"?>')
            html_parts.append('<html xmlns:x="urn:schemas-microsoft-com:office:excel">')
            html_parts.append('<head>')
            html_parts.append('<meta http-equiv="Content-Type" content="text/html; charset=utf-8" />')
            html_parts.append('<!--[if gte mso 9]><xml>')
            html_parts.append('<x:ExcelWorkbook>')
            html_parts.append('<x:ExcelWorksheets>')
            html_parts.append('<x:ExcelWorksheet>')
            html_parts.append('<x:Name>CRS Export</x:Name>')
            html_parts.append('<x:WorksheetOptions><x:Print><x:ValidPrinterInfo/></x:Print></x:WorksheetOptions>')
            html_parts.append('</x:ExcelWorksheet>')
            html_parts.append('</x:ExcelWorksheets>')
            html_parts.append('</x:ExcelWorkbook>')
            html_parts.append('</xml><![endif]-->')
            html_parts.append('<style>')
            html_parts.append('table { border-collapse: collapse; width: 100%; margin-bottom: 20px; }')
            html_parts.append('th, td { border: 1px solid #000; padding: 8px; text-align: left; }')
            html_parts.append('th { background-color: #4472C4; color: white; font-weight: bold; }')
            html_parts.append('.summary { background-color: #E7E6E6; font-weight: bold; }')
            html_parts.append('h1 { color: #2563EB; }')
            html_parts.append('h2 { color: #4472C4; margin-top: 20px; }')
            html_parts.append('</style>')
            html_parts.append('</head>')
            html_parts.append('<body>')
            
            # Title
            html_parts.append(f'<h1>CRS Revision Chain Export</h1>')
            html_parts.append(f'<p><strong>Document:</strong> {escape(chain.document_title or "N/A")}</p>')
            html_parts.append(f'<p><strong>Export Date:</strong> {timezone.now().strftime("%Y-%m-%d %H:%M")}</p>')
            
            # Chain Information Table
            html_parts.append('<h2>Chain Information</h2>')
            html_parts.append('<table>')
            html_parts.append('<tr><th>Field</th><th>Value</th></tr>')
            
            chain_info = [
                ('Chain ID', escape(chain.chain_id or "")),
                ('Project Name', escape(chain.project_name or "")),
                ('Document Number', escape(chain.document_number or "")),
                ('Document Title', escape(chain.document_title or "")),
                ('Contractor', escape(chain.contractor_name or "N/A")),
                ('Department', escape(chain.department or "N/A")),
                ('Status', escape(chain.get_status_display())),
                ('Total Revisions', str(chain.total_revisions)),
                ('Current Revision', str(chain.current_revision_number)),
                ('Max Allowed', str(chain.max_allowed_revisions)),
            ]
            
            for field, value in chain_info:
                html_parts.append(f'<tr><td class="summary">{field}</td><td>{value}</td></tr>')
            
            html_parts.append('</table>')
            
            # Revisions Table
            html_parts.append('<h2>All Revisions</h2>')
            html_parts.append('<table>')
            html_parts.append('<tr>')
            html_parts.append('<th>Rev #</th>')
            html_parts.append('<th>Label</th>')
            html_parts.append('<th>Status</th>')
            html_parts.append('<th>Submitted Date</th>')
            html_parts.append('<th>New Comments</th>')
            html_parts.append('<th>Carryover</th>')
            html_parts.append('<th>Total</th>')
            html_parts.append('<th>Resolved</th>')
            html_parts.append('<th>Resolution %</th>')
            html_parts.append('</tr>')
            
            for rev in revisions:
                resolution_pct = 0
                if hasattr(rev, 'resolution_percentage'):
                    try:
                        resolution_pct = rev.resolution_percentage
                    except:
                        resolution_pct = 0
                
                submitted = rev.submitted_date.strftime('%Y-%m-%d') if rev.submitted_date else 'N/A'
                
                html_parts.append('<tr>')
                html_parts.append(f'<td>{rev.revision_number}</td>')
                html_parts.append(f'<td>{escape(rev.revision_label or "")}</td>')
                html_parts.append(f'<td>{escape(rev.get_status_display())}</td>')
                html_parts.append(f'<td>{submitted}</td>')
                html_parts.append(f'<td>{rev.total_new_comments}</td>')
                html_parts.append(f'<td>{rev.total_carryover_comments}</td>')
                html_parts.append(f'<td>{rev.total_comments}</td>')
                html_parts.append(f'<td>{rev.total_resolved_comments}</td>')
                html_parts.append(f'<td>{resolution_pct:.1f}%</td>')
                html_parts.append('</tr>')
            
            html_parts.append('</table>')
            
            # Comments Table
            html_parts.append('<h2>All Comments</h2>')
            html_parts.append('<table>')
            html_parts.append('<tr>')
            html_parts.append('<th>Revision</th>')
            html_parts.append('<th>Serial #</th>')
            html_parts.append('<th>Page</th>')
            html_parts.append('<th>Comment Text</th>')
            html_parts.append('<th>Type</th>')
            html_parts.append('<th>Status</th>')
            html_parts.append('<th>Clause</th>')
            html_parts.append('</tr>')
            
            total_comments = 0
            for rev in revisions:
                comments = CRSComment.objects.filter(document=rev.document).order_by('serial_number')
                comment_count = comments.count()
                total_comments += comment_count
                logger.info(f"[EXPORT] Rev {rev.revision_number}: {comment_count} comments")
                
                for comment in comments:
                    comment_text = escape(comment.comment_text or "")
                    # Truncate long comments for readability
                    if len(comment_text) > 300:
                        comment_text = comment_text[:300] + "..."
                    
                    html_parts.append('<tr>')
                    html_parts.append(f'<td>{escape(rev.revision_label or "")}</td>')
                    html_parts.append(f'<td>{comment.serial_number}</td>')
                    html_parts.append(f'<td>{comment.page_number}</td>')
                    html_parts.append(f'<td>{comment_text}</td>')
                    html_parts.append(f'<td>{escape(comment.get_comment_type_display())}</td>')
                    html_parts.append(f'<td>{escape(comment.get_status_display())}</td>')
                    html_parts.append(f'<td>{escape(comment.clause_number or "N/A")}</td>')
                    html_parts.append('</tr>')
            
            html_parts.append('</table>')
            
            # Summary footer
            html_parts.append('<hr/>')
            html_parts.append(f'<p><strong>Total Comments Across All Revisions:</strong> {total_comments}</p>')
            html_parts.append(f'<p><em>Generated by RAD AI CRS Multi-Revision System on {timezone.now().strftime("%Y-%m-%d %H:%M:%S")}</em></p>')
            
            html_parts.append('</body>')
            html_parts.append('</html>')
            
            html = '\n'.join(html_parts)
            
            logger.info(f"[EXPORT] Generated HTML export with {total_comments} comments")
            
            # Create response with proper headers for Excel
            safe_chain_id = "".join(c for c in (chain.chain_id or "export") if c.isalnum() or c in "-_")
            filename = f"CRS_{safe_chain_id}_Export.xls"
            
            response = HttpResponse(html, content_type='application/vnd.ms-excel')
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            response['X-Chain-ID'] = chain.chain_id
            response['X-Total-Revisions'] = str(revision_count)
            response['X-Total-Comments'] = str(total_comments)
            
            logger.info(f"[EXPORT] Returning {len(html)} bytes as {filename}")
            
            return response
            
        except Exception as e:
            import traceback
            logger = logging.getLogger(__name__)
            logger.error(f"[EXPORT] Unexpected error: {e}")
            logger.error(traceback.format_exc())
            
            # Return error as HTML (still downloadable for debugging)
            error_html = f"""
            <html>
            <body>
                <h1 style="color: red;">Export Error</h1>
                <p><strong>Error:</strong> {escape(str(e))}</p>
                <pre>{escape(traceback.format_exc())}</pre>
            </body>
            </html>
            """
            
            return HttpResponse(error_html, content_type='text/html', status=500)
    
    @action(detail=True, methods=['get'])
    def revision_timeline(self, request, pk=None):
        """
        Get timeline view of all revisions in the chain
        GET /api/v1/crs/revision-chains/{id}/revision_timeline/
        """
        chain = self.get_object()
        revisions = chain.revisions.all().order_by('revision_number')
        
        timeline_data = []
        for revision in revisions:
            timeline_data.append({
                'revision': CRSRevisionSerializer(revision).data,
                'activities': CRSRevisionActivitySerializer(
                    chain.activities.filter(revision=revision),
                    many=True
                ).data[:10]  # Last 10 activities per revision
            })
        
        return Response({
            "chain_id": chain.chain_id,
            "timeline": timeline_data
        })
    
    @action(detail=True, methods=['get'])
    def statistics(self, request, pk=None):
        """
        Get comprehensive statistics for the chain
        GET /api/v1/crs/revision-chains/{id}/statistics/
        """
        chain = self.get_object()
        revisions = chain.revisions.all()
        
        stats = {
            'chain_info': {
                'chain_id': chain.chain_id,
                'status': chain.status,
                'risk_level': chain.risk_level,
                'ai_risk_score': chain.ai_risk_score,
                'rejection_risk_percentage': chain.rejection_risk_percentage
            },
            'revision_stats': {
                'total_revisions': chain.total_revisions,
                'current_revision': chain.current_revision_number,
                'max_allowed': chain.max_allowed_revisions,
                'remaining_revisions': chain.max_allowed_revisions - chain.current_revision_number
            },
            'comment_stats': {
                'total_comments_all_revisions': sum(r.total_comments for r in revisions),
                'total_new_comments': sum(r.total_new_comments for r in revisions),
                'total_carryover_comments': sum(r.total_carryover_comments for r in revisions),
                'total_resolved': sum(r.total_resolved_comments for r in revisions),
                'overall_resolution_rate': 0
            },
            'revision_breakdown': []
        }
        
        # Calculate overall resolution rate
        total_comments = stats['comment_stats']['total_comments_all_revisions']
        if total_comments > 0:
            stats['comment_stats']['overall_resolution_rate'] = round(
                (stats['comment_stats']['total_resolved'] / total_comments) * 100, 2
            )
        
        # Per-revision breakdown
        for revision in revisions:
            stats['revision_breakdown'].append({
                'revision_label': revision.revision_label,
                'revision_number': revision.revision_number,
                'status': revision.status,
                'total_comments': revision.total_comments,
                'new_comments': revision.total_new_comments,
                'carryover_comments': revision.total_carryover_comments,
                'resolved_comments': revision.total_resolved_comments,
                'resolution_percentage': round((revision.total_resolved_comments / revision.total_comments * 100), 2) if revision.total_comments > 0 else 0,
                'complexity_score': revision.ai_complexity_score,
                'estimated_hours': revision.ai_estimated_resolution_time_hours
            })
        
        return Response(stats)
    
    @action(detail=False, methods=['get'])
    def dashboard_summary(self, request):
        """
        Get dashboard summary of all chains
        GET /api/v1/crs/revision-chains/dashboard_summary/
        """
        chains = self.get_queryset()
        
        summary = {
            'total_chains': chains.count(),
            'by_status': {},
            'by_risk_level': {},
            'critical_chains': [],
            'near_rejection': []
        }
        
        # Count by status
        for choice in CRSRevisionChain.STATUS_CHOICES:
            count = chains.filter(status=choice[0]).count()
            summary['by_status'][choice[1]] = count
        
        # Count by risk level
        for choice in CRSRevisionChain.RISK_LEVEL_CHOICES:
            count = chains.filter(risk_level=choice[0]).count()
            summary['by_risk_level'][choice[1]] = count
        
        # Critical chains (high/critical risk)
        critical = chains.filter(risk_level__in=['high', 'critical']).order_by('-ai_risk_score')[:10]
        summary['critical_chains'] = CRSRevisionChainListSerializer(critical, many=True).data
        
        # Near rejection (close to max revisions)
        near_rejection = [c for c in chains if c.is_near_rejection]
        summary['near_rejection'] = CRSRevisionChainListSerializer(near_rejection, many=True).data
        
        return Response(summary)
    
    # Helper methods
    
    def _auto_link_comments(self, parent_revision, child_revision, user):
        """Automatically detect and link comments between revisions using AI"""
        parent_comments = parent_revision.document.comments.all()
        child_comments = child_revision.document.comments.all()
        
        # Use AI service to detect links
        potential_links = CRSRevisionAIService.detect_comment_links(
            parent_comments, child_comments, threshold=60.0
        )
        
        # Create comment links
        for link_data in potential_links:
            CRSCommentLink.objects.create(
                source_revision=parent_revision,
                target_revision=child_revision,
                source_comment_id=link_data['source_comment_id'],
                target_comment_id=link_data['target_comment_id'],
                link_type=link_data['link_type'],
                similarity_score=link_data['similarity_score'],
                ai_detected=True,
                ai_confidence=link_data['ai_confidence'],
                created_by=user
            )
        
        # Update carryover count
        child_revision.total_carryover_comments = len(potential_links)
        child_revision.total_new_comments = child_comments.count() - len(potential_links)
        child_revision.save()
        
        # Log activity with retry logic for database connection issues
        try:
            self._safe_create_activity(
                chain=child_revision.chain,
                revision=child_revision,
                action='comment_linked',
                description=f'AI detected {len(potential_links)} comment links',
                performed_by=user,
                new_value={'links_created': len(potential_links)}
            )
        except Exception as activity_error:
            logger.warning(f"Could not create activity log (non-critical): {activity_error}")
    
    def _safe_create_activity(self, **kwargs):
        """Create activity with retry logic for database connection issues"""
        from django.db import connection
        import time
        
        max_retries = 3
        retry_delay = 1  # seconds
        
        for attempt in range(max_retries):
            try:
                # Close old connection if it's in a bad state
                if connection.connection and connection.connection.closed:
                    connection.close()
                
                CRSRevisionActivity.objects.create(**kwargs)
                return True
                
            except Exception as e:
                error_msg = str(e).lower()
                if 'ssl' in error_msg or 'eof' in error_msg or 'connection' in error_msg:
                    if attempt < max_retries - 1:
                        logger.warning(f"Database connection issue (attempt {attempt + 1}/{max_retries}): {e}")
                        time.sleep(retry_delay)
                        # Close and reopen connection
                        connection.close()
                        continue
                    else:
                        logger.error(f"Failed to create activity after {max_retries} attempts: {e}")
                        raise
                else:
                    # Non-connection error, raise immediately
                    raise
        
        return False
    
    def _calculate_revision_ai_metrics(self, revision):
        """Calculate AI metrics for a revision"""
        # Complexity score
        complexity = CRSRevisionAIService.calculate_complexity_score(revision)
        
        # Estimated resolution time
        estimated_hours = CRSRevisionAIService.estimate_resolution_time(revision)
        
        # Critical issues count
        critical_count = revision.document.comments.filter(priority__in=['high', 'critical']).count()
        
        revision.ai_complexity_score = complexity
        revision.ai_estimated_resolution_time_hours = estimated_hours
        revision.ai_critical_issues_count = critical_count
        revision.save()
    
    def _update_chain_ai_metrics(self, chain):
        """Update AI metrics for the entire chain"""
        # Calculate risk score
        risk_score = CRSRevisionAIService.calculate_risk_score(chain)
        risk_level = CRSRevisionAIService.determine_risk_level(risk_score)
        
        # Generate recommendation
        recommendation = CRSRevisionAIService.generate_risk_recommendation(chain)
        
        # Predict completion date
        predicted_date = CRSRevisionAIService.predict_completion_date(chain)
        
        chain.ai_risk_score = risk_score
        chain.risk_level = risk_level
        chain.ai_recommendation = recommendation
        chain.ai_predicted_completion_date = predicted_date
        chain.save()
    
    def _generate_chain_insights(self, chain, user):
        """Generate AI insights for the chain"""
        insights = []
        
        # Risk assessment insight
        if chain.ai_risk_score >= 50:
            insights.append(
                CRSAIInsight.objects.create(
                    chain=chain,
                    insight_type='risk_assessment',
                    severity='critical' if chain.ai_risk_score >= 75 else 'warning',
                    title=f'Risk Level: {chain.get_risk_level_display()}',
                    description=f'Current risk score is {chain.ai_risk_score}/100. {chain.ai_recommendation}',
                    confidence_score=95.0,
                    recommended_action=chain.ai_recommendation
                )
            )
        
        # Near rejection alert
        if chain.is_near_rejection:
            insights.append(
                CRSAIInsight.objects.create(
                    chain=chain,
                    insight_type='escalation_alert',
                    severity='critical',
                    title='⚠️ Approaching Maximum Revisions',
                    description=f'This chain is at revision {chain.current_revision_number} of {chain.max_allowed_revisions} allowed. Immediate action required to prevent project rejection.',
                    confidence_score=100.0,
                    recommended_action='Schedule emergency review meeting with stakeholders'
                )
            )
        
        # Pattern detection for latest revision
        latest_revision = chain.revisions.order_by('-revision_number').first()
        if latest_revision:
            comments = latest_revision.document.comments.all()
            if comments.exists():
                patterns = CRSRevisionAIService.analyze_comment_patterns(comments)
                
                if patterns.get('top_keywords'):
                    keyword_list = ', '.join([k['word'] for k in patterns['top_keywords'][:5]])
                    insights.append(
                        CRSAIInsight.objects.create(
                            chain=chain,
                            revision=latest_revision,
                            insight_type='pattern_detection',
                            severity='info',
                            title='Common Comment Themes Detected',
                            description=f'Most frequent topics in {latest_revision.revision_label}: {keyword_list}',
                            confidence_score=85.0,
                            recommended_action='Focus resolution efforts on these common themes'
                        )
                    )
        
        return insights


class CRSRevisionViewSet(viewsets.ModelViewSet):
    """
    ViewSet for individual CRS Revision management
    """
    
    queryset = CRSRevision.objects.all()
    serializer_class = CRSRevisionSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        """Filter by chain if specified"""
        queryset = super().get_queryset()
        chain_id = self.request.query_params.get('chain_id')
        if chain_id:
            queryset = queryset.filter(chain_id=chain_id)
        return queryset.select_related('document', 'chain', 'parent_revision')
    
    @action(detail=True, methods=['get'])
    def comment_links(self, request, pk=None):
        """
        Get all comment links for this revision
        GET /api/v1/crs/revisions/{id}/comment_links/
        """
        revision = self.get_object()
        
        links_from = CRSCommentLink.objects.filter(source_revision=revision)
        links_to = CRSCommentLink.objects.filter(target_revision=revision)
        
        return Response({
            'revision_label': revision.revision_label,
            'links_from_previous': CRSCommentLinkSerializer(links_from, many=True).data,
            'links_to_next': CRSCommentLinkSerializer(links_to, many=True).data
        })
    
    @action(detail=True, methods=['post'])
    def update_status(self, request, pk=None):
        """
        Update revision status
        POST /api/v1/crs/revisions/{id}/update_status/
        Body: { "status": "completed" }
        """
        revision = self.get_object()
        new_status = request.data.get('status')
        
        if not new_status:
            return Response(
                {"error": "status is required"},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        valid_statuses = [choice[0] for choice in CRSRevision.STATUS_CHOICES]
        if new_status not in valid_statuses:
            return Response(
                {"error": f"Invalid status. Must be one of: {', '.join(valid_statuses)}"},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        old_status = revision.status
        revision.status = new_status
        
        # Set completion date if status is completed
        if new_status == 'completed' and not revision.completed_date:
            revision.completed_date = timezone.now()
        
        revision.save()
        
        # Log activity
        CRSRevisionActivity.objects.create(
            chain=revision.chain,
            revision=revision,
            action='status_changed',
            description=f'Revision status changed from {old_status} to {new_status}',
            performed_by=request.user,
            old_value={'status': old_status},
            new_value={'status': new_status}
        )
        
        return Response({
            "success": True,
            "message": f"Status updated to {new_status}",
            "data": CRSRevisionSerializer(revision).data
        })


class CRSCommentLinkViewSet(viewsets.ModelViewSet):
    """
    ViewSet for managing comment links across revisions
    """
    
    queryset = CRSCommentLink.objects.all()
    serializer_class = CRSCommentLinkSerializer
    permission_classes = [IsAuthenticated]
    
    def get_serializer_class(self):
        if self.action == 'create':
            return CRSCommentLinkCreateSerializer
        return CRSCommentLinkSerializer
    
    def perform_create(self, serializer):
        """Create link and log activity"""
        link = serializer.save(created_by=self.request.user)
        
        CRSRevisionActivity.objects.create(
            chain=link.source_revision.chain,
            revision=link.target_revision,
            action='comment_linked',
            description=f'Comment #{link.source_comment.serial_number} linked to #{link.target_comment.serial_number}',
            performed_by=self.request.user
        )


class CRSAIInsightViewSet(viewsets.ReadOnlyModelViewSet):
    """
    ViewSet for viewing AI insights
    """
    
    queryset = CRSAIInsight.objects.filter(is_active=True)
    serializer_class = CRSAIInsightSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        """Filter by chain, revision, or severity"""
        queryset = super().get_queryset()
        
        chain_id = self.request.query_params.get('chain_id')
        if chain_id:
            queryset = queryset.filter(chain_id=chain_id)
        
        revision_id = self.request.query_params.get('revision_id')
        if revision_id:
            queryset = queryset.filter(revision_id=revision_id)
        
        severity = self.request.query_params.get('severity')
        if severity:
            queryset = queryset.filter(severity=severity)
        
        insight_type = self.request.query_params.get('insight_type')
        if insight_type:
            queryset = queryset.filter(insight_type=insight_type)
        
        return queryset.order_by('-created_at')
    
    @action(detail=True, methods=['post'])
    def submit_feedback(self, request, pk=None):
        """
        Submit feedback on an AI insight
        POST /api/v1/crs/ai-insights/{id}/submit_feedback/
        Body: { "was_helpful": true, "user_feedback": "Very helpful!" }
        """
        insight = self.get_object()
        serializer = CRSAIInsightFeedbackSerializer(data=request.data)
        
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        
        insight.was_helpful = serializer.validated_data['was_helpful']
        insight.user_feedback = serializer.validated_data.get('user_feedback', '')
        insight.save()
        
        return Response({
            "success": True,
            "message": "Feedback submitted successfully"
        })
    
    @action(detail=True, methods=['post'])
    def dismiss(self, request, pk=None):
        """
        Dismiss an AI insight
        POST /api/v1/crs/ai-insights/{id}/dismiss/
        """
        insight = self.get_object()
        insight.is_active = False
        insight.dismissed_by = request.user
        insight.dismissed_at = timezone.now()
        insight.save()
        
        return Response({
            "success": True,
            "message": "Insight dismissed"
        })


class CRSRevisionActivityViewSet(viewsets.ReadOnlyModelViewSet):
    """
    ViewSet for viewing revision activities
    """
    
    queryset = CRSRevisionActivity.objects.all()
    serializer_class = CRSRevisionActivitySerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        """Filter by chain or revision"""
        queryset = super().get_queryset()
        
        chain_id = self.request.query_params.get('chain_id')
        if chain_id:
            queryset = queryset.filter(chain_id=chain_id)
        
        revision_id = self.request.query_params.get('revision_id')
        if revision_id:
            queryset = queryset.filter(revision_id=revision_id)
        
        return queryset.order_by('-performed_at')
