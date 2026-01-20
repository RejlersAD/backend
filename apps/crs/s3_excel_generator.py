"""
CRS S3 Excel Generator
======================
Soft-coded module for generating and uploading CRS revision Excel files to S3
Uses intelligent fallback and error handling strategies
"""

import logging
from io import BytesIO
from datetime import datetime
from django.conf import settings
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


class CRSS3ExcelGeneratorConfig:
    """Soft-coded configuration for Excel generation and S3 upload"""
    
    # Excel styling configuration
    STYLES = {
        'header': {
            'font': Font(bold=True, size=12, color="FFFFFF"),
            'fill': PatternFill(start_color="366092", end_color="366092", fill_type="solid"),
            'alignment': Alignment(horizontal="center", vertical="center", wrap_text=True),
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        },
        'data': {
            'font': Font(size=11),
            'alignment': Alignment(horizontal="left", vertical="top", wrap_text=True),
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        }
    }
    
    # Column configuration
    COLUMNS = [
        {'name': 'Page', 'width': 8},
        {'name': 'Reviewer', 'width': 20},
        {'name': 'Comment', 'width': 50},
        {'name': 'Type', 'width': 15},
        {'name': 'Discipline', 'width': 20},
        {'name': 'Drawing Ref', 'width': 20},
        {'name': 'Status', 'width': 12}
    ]
    
    # S3 configuration
    S3_CONFIG = {
        'folder': 'crs/revisions/excel/',
        'acl': 'private',
        'content_type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        'cache_control': 'max-age=31536000',  # 1 year cache
        'enabled': getattr(settings, 'USE_S3', False) and getattr(settings, 'S3_READY', False)
    }
    
    # File naming pattern
    FILE_NAME_PATTERN = "CRS_{chain_id}_Rev{revision_number}_{timestamp}.xlsx"


class CRSS3ExcelGenerator:
    """
    Generates Excel files for CRS revisions and uploads to S3
    Uses soft-coded configuration for flexibility and maintainability
    """
    
    def __init__(self):
        self.config = CRSS3ExcelGeneratorConfig()
        self.s3_enabled = self.config.S3_CONFIG['enabled']
        
        if self.s3_enabled:
            try:
                import boto3
                self.s3_client = boto3.client('s3')
                self.bucket_name = settings.AWS_STORAGE_BUCKET_NAME
            except Exception as e:
                logger.warning(f"S3 initialization failed: {e}. Excel upload will be skipped.")
                self.s3_enabled = False
    
    def generate_and_upload_revision_excel(self, revision):
        """
        Main method: Generate Excel from revision and upload to S3
        
        Args:
            revision: CRSRevision instance
            
        Returns:
            str: S3 URL of uploaded file, or None if upload failed/disabled
        """
        try:
            # Generate Excel file in memory
            excel_buffer = self._generate_excel(revision)
            
            if not excel_buffer:
                logger.warning(f"Excel generation returned empty buffer for revision {revision.id}")
                return None
            
            # Upload to S3 if enabled
            if self.s3_enabled:
                s3_url = self._upload_to_s3(excel_buffer, revision)
                return s3_url
            else:
                logger.info(f"S3 upload disabled - Excel generated but not uploaded for revision {revision.id}")
                return None
                
        except Exception as e:
            logger.error(f"Error in generate_and_upload_revision_excel: {e}", exc_info=True)
            return None
    
    def _generate_excel(self, revision):
        """
        Generate Excel file from revision data
        
        Args:
            revision: CRSRevision instance
            
        Returns:
            BytesIO: Excel file buffer
        """
        try:
            # Create workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = f"Rev {revision.revision_number}"
            
            # Add header information
            ws['A1'] = f"CRS Revision: {revision.revision_label}"
            ws['A1'].font = Font(bold=True, size=14)
            ws.merge_cells('A1:G1')
            
            ws['A2'] = f"Document: {revision.chain.document.document_number}"
            ws['A2'].font = Font(size=11)
            ws.merge_cells('A2:G2')
            
            ws['A3'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            ws['A3'].font = Font(size=10, italic=True)
            ws.merge_cells('A3:G3')
            
            # Empty row
            current_row = 5
            
            # Add column headers
            for col_idx, column_config in enumerate(self.config.COLUMNS, start=1):
                cell = ws.cell(row=current_row, column=col_idx)
                cell.value = column_config['name']
                
                # Apply header styling
                cell.font = self.config.STYLES['header']['font']
                cell.fill = self.config.STYLES['header']['fill']
                cell.alignment = self.config.STYLES['header']['alignment']
                cell.border = self.config.STYLES['header']['border']
                
                # Set column width
                ws.column_dimensions[get_column_letter(col_idx)].width = column_config['width']
            
            current_row += 1
            
            # Add comment data
            comments = revision.comments.all().order_by('page_number', 'id')
            
            for comment in comments:
                ws.cell(row=current_row, column=1).value = comment.page_number or 'N/A'
                ws.cell(row=current_row, column=2).value = comment.reviewer_name or 'Not Provided'
                ws.cell(row=current_row, column=3).value = comment.comment_text or ''
                ws.cell(row=current_row, column=4).value = comment.comment_type or 'GENERAL'
                ws.cell(row=current_row, column=5).value = comment.discipline or 'Not Provided'
                ws.cell(row=current_row, column=6).value = comment.drawing_ref or 'N/A'
                ws.cell(row=current_row, column=7).value = comment.status or 'Open'
                
                # Apply data styling
                for col_idx in range(1, 8):
                    cell = ws.cell(row=current_row, column=col_idx)
                    cell.font = self.config.STYLES['data']['font']
                    cell.alignment = self.config.STYLES['data']['alignment']
                    cell.border = self.config.STYLES['data']['border']
                
                current_row += 1
            
            # Add summary at the bottom
            current_row += 2
            ws.cell(row=current_row, column=1).value = "Summary:"
            ws.cell(row=current_row, column=1).font = Font(bold=True)
            
            current_row += 1
            ws.cell(row=current_row, column=1).value = f"Total Comments: {comments.count()}"
            
            # Save to buffer
            excel_buffer = BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)
            
            logger.info(f"Excel generated successfully for revision {revision.id} with {comments.count()} comments")
            return excel_buffer
            
        except Exception as e:
            logger.error(f"Error generating Excel: {e}", exc_info=True)
            return None
    
    def _upload_to_s3(self, excel_buffer, revision):
        """
        Upload Excel buffer to S3
        
        Args:
            excel_buffer: BytesIO buffer containing Excel file
            revision: CRSRevision instance
            
        Returns:
            str: S3 URL or None
        """
        try:
            # Generate S3 key
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            file_name = self.config.FILE_NAME_PATTERN.format(
                chain_id=revision.chain.id,
                revision_number=revision.revision_number,
                timestamp=timestamp
            )
            s3_key = f"{self.config.S3_CONFIG['folder']}{file_name}"
            
            # Upload to S3
            self.s3_client.put_object(
                Bucket=self.bucket_name,
                Key=s3_key,
                Body=excel_buffer.getvalue(),
                ContentType=self.config.S3_CONFIG['content_type'],
                CacheControl=self.config.S3_CONFIG['cache_control'],
                ACL=self.config.S3_CONFIG['acl']
            )
            
            # Generate S3 URL
            s3_url = f"https://{self.bucket_name}.s3.{settings.AWS_S3_REGION_NAME}.amazonaws.com/{s3_key}"
            
            logger.info(f"Excel uploaded to S3: {s3_url}")
            return s3_url
            
        except Exception as e:
            logger.error(f"Error uploading to S3: {e}", exc_info=True)
            return None
    
    def generate_presigned_url(self, s3_key, expiration=3600):
        """
        Generate presigned URL for downloading Excel from S3
        
        Args:
            s3_key: S3 object key
            expiration: URL expiration in seconds (default 1 hour)
            
        Returns:
            str: Presigned URL or None
        """
        if not self.s3_enabled:
            return None
            
        try:
            url = self.s3_client.generate_presigned_url(
                'get_object',
                Params={
                    'Bucket': self.bucket_name,
                    'Key': s3_key
                },
                ExpiresIn=expiration
            )
            return url
        except Exception as e:
            logger.error(f"Error generating presigned URL: {e}", exc_info=True)
            return None
