"""
Excel exporter for an IOListDocument.

Two sheets:
  - "Comments Resolution Sheet" (5 columns + status meaning + linked tags)
  - "IO List" (40 canonical columns)

openpyxl only — no third-party templates required.
"""

from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from .services.config import (
    COMMENT_SHEET_COLUMNS, IO_LIST_CANONICAL_COLUMNS,
)


_HEADER_FILL = PatternFill(start_color='003366', end_color='003366',
                            fill_type='solid')
_HEADER_FONT = Font(color='FFFFFF', bold=True)
_CENTER      = Alignment(horizontal='center', vertical='center', wrap_text=True)


def _write_header(ws, columns):
    for idx, col in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=idx, value=col)
        cell.fill, cell.font, cell.alignment = _HEADER_FILL, _HEADER_FONT, _CENTER


def export_document_to_xlsx(document) -> bytes:
    wb = Workbook()

    # Sheet 1 — comments
    ws_c = wb.active
    ws_c.title = 'Comments Resolution Sheet'
    headers_c = COMMENT_SHEET_COLUMNS + ['status_meaning', 'page_number',
                                          'linked_tags']
    _write_header(ws_c, headers_c)
    for r_idx, c in enumerate(document.extracted_comments.all(), start=2):
        row = [
            c.s_no, c.company_comment, c.contractor_reply, c.company_decision,
            c.status_code, c.status_meaning, c.page_number,
            ', '.join(c.linked_tags or []),
        ]
        for c_idx, val in enumerate(row, start=1):
            ws_c.cell(row=r_idx, column=c_idx, value=val)

    # Sheet 2 — IO rows
    ws_r = wb.create_sheet('IO List')
    headers_r = ['tag_number', 'page_number'] + [
        c for c in IO_LIST_CANONICAL_COLUMNS if c != 'tag_number'
    ]
    _write_header(ws_r, headers_r)
    for r_idx, row in enumerate(document.extracted_rows.all(), start=2):
        d = row.data or {}
        ws_r.cell(row=r_idx, column=1, value=row.tag_number)
        ws_r.cell(row=r_idx, column=2, value=row.page_number)
        for c_idx, col in enumerate(headers_r[2:], start=3):
            ws_r.cell(row=r_idx, column=c_idx, value=d.get(col, ''))

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
