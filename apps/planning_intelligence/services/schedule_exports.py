"""Version-native schedule exports for APIs, Excel, CSV, and Primavera P6."""
from __future__ import annotations

import csv
import io
import json
import re
from types import SimpleNamespace

from django.core.serializers.json import DjangoJSONEncoder
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from ..governance_serializers import GovernanceItemSerializer, ScheduleReviewSerializer
from ..models import ActivityAssignment
from ..schedule_serializers import (
    ActivityAssignmentSerializer, ActivityRelationshipSerializer, ScheduleActivitySerializer,
    ScheduleBaselineSerializer, ScheduleResourceSerializer, ScheduleVersionSerializer,
    ScheduleWBSNodeSerializer, WorkCalendarSerializer,
)
from .export_utils import generation_to_xer_bytes
from .project_controls import build_control_dashboard


EXPORT_CONTENT_TYPES = {
    'json': 'application/json', 'csv': 'text/csv; charset=utf-8',
    'xlsx': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    'xer': 'application/octet-stream',
}


def _safe(value):
    return json.loads(json.dumps(value, cls=DjangoJSONEncoder))


def schedule_snapshot(version):
    schedule = version.schedule
    project = schedule.project
    controls = build_control_dashboard(version, schedule.data_date) if version.activities.filter(is_deleted=False).exists() else None
    return _safe({
        'schema_version': '1.0', 'exported_at': timezone.now(),
        'project': {'id': project.id, 'name': project.name, 'client': project.client, 'phase': project.phase},
        'schedule': {
            'id': schedule.id, 'name': schedule.name, 'code': schedule.code,
            'status': schedule.status, 'planned_start': schedule.planned_start, 'data_date': schedule.data_date,
        },
        'version': ScheduleVersionSerializer(version).data,
        'calendar': WorkCalendarSerializer(schedule.default_calendar).data if schedule.default_calendar else None,
        'wbs': ScheduleWBSNodeSerializer(version.wbs_nodes.filter(is_deleted=False), many=True).data,
        'activities': ScheduleActivitySerializer(version.activities.filter(is_deleted=False), many=True).data,
        'relationships': ActivityRelationshipSerializer(version.relationships.filter(is_deleted=False), many=True).data,
        'resources': ScheduleResourceSerializer(project.schedule_resources.filter(is_deleted=False), many=True).data,
        'assignments': ActivityAssignmentSerializer(ActivityAssignment.objects.filter(
            activity__version=version, activity__is_deleted=False, is_deleted=False,
        ), many=True).data,
        'baselines': ScheduleBaselineSerializer(schedule.baselines.filter(is_deleted=False), many=True).data,
        'controls': controls,
        'governance': {
            'items': GovernanceItemSerializer(version.governance_items.filter(is_deleted=False), many=True).data,
            'reviews': ScheduleReviewSerializer(version.governance_reviews.filter(is_deleted=False), many=True).data,
        },
    })


def _activities_csv(snapshot):
    stream = io.StringIO(newline='')
    fields = [
        'external_id', 'name', 'activity_type', 'duration_days', 'discipline', 'responsible_role',
        'planned_start', 'planned_finish', 'total_float_days', 'free_float_days', 'is_critical',
        'constraint_type', 'constraint_date',
    ]
    writer = csv.DictWriter(stream, fieldnames=fields)
    writer.writeheader()
    for row in snapshot['activities']:
        writer.writerow({field: row.get(field, '') for field in fields})
    return stream.getvalue().encode('utf-8-sig')


def _excel(snapshot):
    workbook = Workbook()
    workbook.remove(workbook.active)

    def add_sheet(name, rows):
        sheet = workbook.create_sheet(name[:31])
        rows = list(rows or [])
        if not rows:
            sheet.append(['No data'])
            return
        headers = list(rows[0])
        sheet.append(headers)
        for cell in sheet[1]:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill('solid', fgColor='4F46E5')
        for row in rows:
            sheet.append([
                json.dumps(row.get(header), ensure_ascii=False) if isinstance(row.get(header), (dict, list))
                else row.get(header) for header in headers
            ])
        sheet.freeze_panes = 'A2'
        sheet.auto_filter.ref = sheet.dimensions
        for column in sheet.columns:
            letter = column[0].column_letter
            sheet.column_dimensions[letter].width = min(50, max(12, max(len(str(cell.value or '')) for cell in column) + 2))

    add_sheet('Activities', snapshot['activities'])
    add_sheet('WBS', snapshot['wbs'])
    add_sheet('Relationships', snapshot['relationships'])
    add_sheet('Resources', snapshot['resources'])
    add_sheet('Assignments', snapshot['assignments'])
    add_sheet('Controls WBS', (snapshot.get('controls') or {}).get('wbs_breakdown', []))
    add_sheet('Governance', snapshot['governance']['items'])
    stream = io.BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def _xer(version, snapshot):
    parent_codes = {row['id']: row['code'] for row in snapshot['wbs']}
    updates = {row['id']: row for row in (snapshot.get('controls') or {}).get('activities', [])}
    activities = []
    for row in snapshot['activities']:
        progress = updates.get(row['id'], {})
        activities.append({
            'id': row['external_id'], 'name': row['name'],
            'wbs_code': parent_codes.get(row.get('wbs_node')),
            'original_duration_days': float(row['duration_days']),
            'start_date': row['planned_start'], 'finish_date': row['planned_finish'],
            'total_float_days': float(row['total_float_days'] or 0),
            'is_milestone': row['is_milestone'], 'predecessors': [],
            'physical_progress_pct': progress.get('physical_progress_pct', 0),
            'remaining_duration_days': progress.get('remaining_duration_days'),
            'actual_start': progress.get('actual_start'), 'actual_finish': progress.get('actual_finish'),
        })
    external_ids = {row['id']: row['external_id'] for row in snapshot['activities']}
    logic = [{
        'activity_id': external_ids.get(row['successor']),
        'predecessor_id': external_ids.get(row['predecessor']),
        'type': row['relationship_type'], 'lag_days': float(row['lag_days']),
    } for row in snapshot['relationships']]
    wbs = [{
        'code': row['code'], 'name': row['name'], 'level': row['level'],
        'parent_code': parent_codes.get(row.get('parent')),
    } for row in snapshot['wbs']]
    generation = SimpleNamespace(
        project=version.schedule.project, generated_by=version.created_by,
        wbs=wbs, activities=activities, logic_matrix=logic,
    )
    return generation_to_xer_bytes(generation)


def generate_schedule_export(version, export_format):
    if export_format not in EXPORT_CONTENT_TYPES:
        raise ValueError('Unsupported schedule export format.')
    snapshot = schedule_snapshot(version)
    if export_format == 'json':
        content = json.dumps(snapshot, ensure_ascii=False, indent=2).encode('utf-8')
    elif export_format == 'csv':
        content = _activities_csv(snapshot)
    elif export_format == 'xlsx':
        content = _excel(snapshot)
    else:
        content = _xer(version, snapshot)
    code = re.sub(r'[^A-Za-z0-9_-]+', '-', version.schedule.code).strip('-') or 'schedule'
    filename = f'{code}-v{version.version}.{export_format}'
    return content, EXPORT_CONTENT_TYPES[export_format], filename
