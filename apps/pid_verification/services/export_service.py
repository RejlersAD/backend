"""
Export Service
==============
Generates Excel and PDF quality reports from stored findings.
Uploads results to AWS S3 and returns public/presigned URLs.
"""
import io
import logging
import os
import tempfile
from typing import Optional

logger = logging.getLogger(__name__)

# Column order for the findings report
FINDINGS_COLUMNS = [
    'SL No', 'Drawing ID', 'Category', 'Rule ID',
    'Issue Observed', 'Action Required', 'Evidence',
    'Direction', 'Severity', 'Status',
]


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------

def generate_excel(document) -> Optional[bytes]:
    """
    Build an Excel workbook from all findings for `document`.
    Returns raw bytes or None on failure.
    Uses pandas + openpyxl (deterministic output: same data → same file).
    """
    try:
        import pandas as pd
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        logger.error('[Export] pandas/openpyxl not installed')
        return None

    rows = []
    for drawing in document.drawings.order_by('page_index'):
        for finding in drawing.findings.order_by('sl_no'):
            rows.append({
                'SL No':           finding.sl_no,
                'Drawing ID':      drawing.drawing_id,
                'Category':        finding.get_category_display(),
                'Rule ID':         finding.rule_id,
                'Issue Observed':  finding.issue_observed,
                'Action Required': finding.action_required,
                'Evidence':        finding.evidence,
                'Direction':       finding.direction,
                'Severity':        finding.severity.upper(),
                'Status':          finding.status.capitalize(),
            })

    df = pd.DataFrame(rows, columns=FINDINGS_COLUMNS) if rows else pd.DataFrame(columns=FINDINGS_COLUMNS)

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='P&ID QC Findings')
        wb = writer.book
        ws = writer.sheets['P&ID QC Findings']

        # Header styling
        header_fill   = PatternFill(fill_type='solid', fgColor='2F5496')
        header_font   = Font(bold=True, color='FFFFFF', size=11)
        header_align  = Alignment(horizontal='center', vertical='center', wrap_text=True)

        thin = Side(style='thin', color='000000')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for col_idx, col_name in enumerate(FINDINGS_COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill   = header_fill
            cell.font   = header_font
            cell.alignment = header_align
            cell.border = border

        # Severity colour coding
        severity_colours = {
            'CRITICAL': 'FF0000',
            'MAJOR':    'FF9900',
            'MINOR':    'FFFF00',
            'INFO':     'D9EAD3',
        }
        severity_col = FINDINGS_COLUMNS.index('Severity') + 1

        for row_idx, row_data in enumerate(rows, start=2):
            sev  = row_data['Severity'].upper()
            colour = severity_colours.get(sev, 'FFFFFF')
            fill = PatternFill(fill_type='solid', fgColor=colour)
            ws.cell(row=row_idx, column=severity_col).fill = fill
            for col_idx in range(1, len(FINDINGS_COLUMNS) + 1):
                ws.cell(row=row_idx, column=col_idx).border = border
                ws.cell(row=row_idx, column=col_idx).alignment = Alignment(wrap_text=True, vertical='top')

        # Auto-fit column widths (max 60 chars)
        for col_idx, col_name in enumerate(FINDINGS_COLUMNS, start=1):
            max_len = max(
                len(col_name),
                *[len(str(r.get(col_name, '') or '')) for r in rows] if rows else [0]
            )
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 60)

        ws.freeze_panes = 'A2'

    return buf.getvalue()


# ---------------------------------------------------------------------------
# PDF  (using reportlab)
# ---------------------------------------------------------------------------

def generate_pdf(document) -> Optional[bytes]:
    """
    Generate a PDF quality report using reportlab.
    Returns raw bytes or None on failure.
    """
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.platypus import (
            SimpleDocTemplate, Table, TableStyle,
            Paragraph, Spacer, HRFlowable,
        )
    except ImportError:
        logger.error('[Export] reportlab not installed')
        return None

    buf  = io.BytesIO()
    doc  = SimpleDocTemplate(buf, pagesize=landscape(A4),
                              leftMargin=1.5*cm, rightMargin=1.5*cm,
                              topMargin=2*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()
    story  = []

    # Title
    title_style = ParagraphStyle('Title', parent=styles['Title'],
                                 fontSize=16, spaceAfter=8, textColor=colors.HexColor('#2F5496'))
    story.append(Paragraph('P&amp;ID Quality Checker — Findings Report', title_style))
    story.append(Paragraph(f'Document: {document.file_name}', styles['Normal']))
    story.append(Spacer(1, 0.3*cm))
    story.append(HRFlowable(width='100%', thickness=1, color=colors.HexColor('#2F5496')))
    story.append(Spacer(1, 0.3*cm))

    # Table data
    header = ['SL No', 'Drawing', 'Category', 'Issue Observed', 'Action Required',
              'Evidence', 'Severity', 'Status']
    data = [header]

    for drawing in document.drawings.order_by('page_index'):
        for finding in drawing.findings.order_by('sl_no'):
            wrap = lambda t: Paragraph(str(t)[:300], styles['Normal'])
            data.append([
                str(finding.sl_no),
                drawing.drawing_id,
                finding.get_category_display(),
                wrap(finding.issue_observed),
                wrap(finding.action_required),
                wrap(finding.evidence[:120]),
                finding.severity.upper(),
                finding.status.capitalize(),
            ])

    col_widths = [1.2*cm, 3.5*cm, 2.5*cm, 7*cm, 7*cm, 4*cm, 2*cm, 2*cm]

    severity_colours = {
        'CRITICAL': colors.HexColor('#FF0000'),
        'MAJOR':    colors.HexColor('#FF9900'),
        'MINOR':    colors.HexColor('#FFFF00'),
        'INFO':     colors.HexColor('#D9EAD3'),
    }

    table = Table(data, colWidths=col_widths, repeatRows=1)
    table_style_cmds = [
        ('BACKGROUND',  (0, 0), (-1, 0), colors.HexColor('#2F5496')),
        ('TEXTCOLOR',   (0, 0), (-1, 0), colors.white),
        ('FONTNAME',    (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE',    (0, 0), (-1, 0), 8),
        ('GRID',        (0, 0), (-1, -1), 0.4, colors.grey),
        ('FONTSIZE',    (0, 1), (-1, -1), 7),
        ('VALIGN',      (0, 0), (-1, -1), 'TOP'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F2F2F2')]),
    ]

    # Apply severity colour to Severity column (index 6)
    for row_idx, row_data in enumerate(data[1:], start=1):
        sev = str(row_data[6] if len(row_data) > 6 else '').upper()
        if sev in severity_colours:
            table_style_cmds.append(
                ('BACKGROUND', (6, row_idx), (6, row_idx), severity_colours[sev])
            )

    table.setStyle(TableStyle(table_style_cmds))
    story.append(table)

    doc.build(story)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# S3 upload helper
# ---------------------------------------------------------------------------

def upload_to_s3(data: bytes, s3_key: str, content_type: str) -> str:
    """
    Upload `data` to S3 at `s3_key`.
    Returns the S3 URL string, or empty string on failure.
    """
    try:
        import boto3
        bucket = os.environ.get('AWS_STORAGE_BUCKET_NAME', '')
        region = os.environ.get('AWS_S3_REGION_NAME', 'us-east-1')
        if not bucket:
            logger.warning('[Export] AWS_STORAGE_BUCKET_NAME not set – skipping S3 upload')
            return ''
        s3 = boto3.client('s3', region_name=region)
        s3.put_object(
            Bucket=bucket,
            Key=s3_key,
            Body=data,
            ContentType=content_type,
        )
        return f'https://{bucket}.s3.{region}.amazonaws.com/{s3_key}'
    except Exception as exc:
        logger.error('[Export] S3 upload failed: %s', exc)
        return ''
