"""
Invoice Export Service - Smart Excel and PDF Generation
Generates professional exports with filtering and formatting
"""
from datetime import datetime
from decimal import Decimal
from io import BytesIO
import os

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT

from django.db.models import Q, Sum, Count
from django.utils import timezone

from ..models import Invoice, Approval, InvoiceStatus, InvoiceType
import logging

logger = logging.getLogger(__name__)


class InvoiceExportService:
    """Smart invoice export service with Excel and PDF generation"""
    
    def __init__(self):
        self.styles = getSampleStyleSheet()
        self._setup_pdf_styles()
    
    def _setup_pdf_styles(self):
        """Setup custom PDF styles"""
        self.styles.add(ParagraphStyle(
            name='CustomTitle',
            parent=self.styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#1a365d'),
            spaceAfter=30,
            alignment=TA_CENTER,
            fontName='Helvetica-Bold'
        ))
        
        self.styles.add(ParagraphStyle(
            name='SectionHeader',
            parent=self.styles['Heading2'],
            fontSize=14,
            textColor=colors.HexColor('#2c5282'),
            spaceAfter=12,
            spaceBefore=12,
            fontName='Helvetica-Bold'
        ))
    
    def export_to_excel(self, filters=None):
        """
        Generate Excel export with smart filtering and formatting
        
        Args:
            filters: dict with keys: status, invoice_type, date_from, date_to, search
        
        Returns:
            BytesIO: Excel file buffer
        """
        try:
            # Get filtered invoices
            invoices = self._get_filtered_invoices(filters)
            
            # Create workbook
            wb = Workbook()
            
            # Remove default sheet and create named sheets
            wb.remove(wb.active)
            
            # Sheet 1: Invoice List
            self._create_invoice_list_sheet(wb, invoices)
            
            # Sheet 2: Summary Statistics
            self._create_summary_sheet(wb, invoices, filters)
            
            # Sheet 3: Approval Details
            self._create_approval_sheet(wb, invoices)
            
            # Save to buffer
            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            
            return buffer
            
        except Exception as e:
            logger.error(f"Excel export failed: {e}")
            raise
    
    def export_to_pdf(self, filters=None):
        """
        Generate PDF export with professional formatting
        
        Args:
            filters: dict with keys: status, invoice_type, date_from, date_to, search
        
        Returns:
            BytesIO: PDF file buffer
        """
        try:
            # Get filtered invoices
            invoices = self._get_filtered_invoices(filters)
            
            # Create PDF buffer
            buffer = BytesIO()
            doc = SimpleDocTemplate(
                buffer,
                pagesize=landscape(A4),
                rightMargin=30,
                leftMargin=30,
                topMargin=30,
                bottomMargin=30
            )
            
            # Build content
            content = []
            
            # Title
            title = Paragraph("Invoice Export Report", self.styles['CustomTitle'])
            content.append(title)
            content.append(Spacer(1, 0.2 * inch))
            
            # Export info
            export_info = self._create_export_info(filters)
            content.append(export_info)
            content.append(Spacer(1, 0.3 * inch))
            
            # Summary section
            summary_section = self._create_pdf_summary(invoices, filters)
            content.extend(summary_section)
            content.append(PageBreak())
            
            # Invoice table
            invoice_table = self._create_pdf_invoice_table(invoices)
            content.append(invoice_table)
            
            # Build PDF
            doc.build(content)
            buffer.seek(0)
            
            return buffer
            
        except Exception as e:
            logger.error(f"PDF export failed: {e}")
            raise
    
    def _get_filtered_invoices(self, filters):
        """Apply smart filters to invoice queryset"""
        queryset = Invoice.objects.all().select_related('submitted_by').prefetch_related('approvals')
        
        if not filters:
            return queryset.order_by('-created_at')
        
        # Status filter
        if filters.get('status'):
            status_list = filters['status'] if isinstance(filters['status'], list) else [filters['status']]
            queryset = queryset.filter(status__in=status_list)
        
        # Invoice type filter
        if filters.get('invoice_type'):
            type_list = filters['invoice_type'] if isinstance(filters['invoice_type'], list) else [filters['invoice_type']]
            queryset = queryset.filter(invoice_type__in=type_list)
        
        # Date range filter
        if filters.get('date_from'):
            date_from = filters['date_from']
            if isinstance(date_from, str):
                date_from = datetime.strptime(date_from, '%Y-%m-%d').date()
            queryset = queryset.filter(created_at__date__gte=date_from)
        
        if filters.get('date_to'):
            date_to = filters['date_to']
            if isinstance(date_to, str):
                date_to = datetime.strptime(date_to, '%Y-%m-%d').date()
            queryset = queryset.filter(created_at__date__lte=date_to)
        
        # Search filter (invoice number, vendor name)
        if filters.get('search'):
            search_term = filters['search']
            queryset = queryset.filter(
                Q(invoice_number__icontains=search_term) |
                Q(vendor_name__icontains=search_term) |
                Q(email_from__icontains=search_term)
            )
        
        # Amount range filter
        if filters.get('min_amount'):
            queryset = queryset.filter(total_amount__gte=filters['min_amount'])
        
        if filters.get('max_amount'):
            queryset = queryset.filter(total_amount__lte=filters['max_amount'])
        
        return queryset.order_by('-created_at')
    
    def _create_invoice_list_sheet(self, wb, invoices):
        """Create detailed invoice list sheet"""
        ws = wb.create_sheet("Invoice List")
        
        # Define headers
        headers = [
            'Tracking ID', 'Invoice Number', 'Vendor Name', 'Amount', 'Currency',
            'Tax Amount', 'Total Amount', 'Type', 'Status', 'Confidence',
            'Submitted By', 'Created At'
        ]
        
        # Header styling
        header_fill = PatternFill(start_color='1e40af', end_color='1e40af', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Write data
        for row_num, invoice in enumerate(invoices, 2):
            ws.cell(row=row_num, column=1, value=invoice.tracking_id or 'N/A')
            ws.cell(row=row_num, column=2, value=invoice.invoice_number)
            ws.cell(row=row_num, column=3, value=invoice.vendor_name or 'N/A')
            ws.cell(row=row_num, column=4, value=float(invoice.amount) if invoice.amount else 0)
            ws.cell(row=row_num, column=5, value=invoice.currency)
            ws.cell(row=row_num, column=6, value=float(invoice.tax_amount) if invoice.tax_amount else 0)
            ws.cell(row=row_num, column=7, value=float(invoice.total_amount) if invoice.total_amount else 0)
            ws.cell(row=row_num, column=8, value=invoice.get_invoice_type_display() if invoice.invoice_type else 'N/A')
            ws.cell(row=row_num, column=9, value=invoice.get_status_display())
            ws.cell(row=row_num, column=10, value=f"{invoice.classification_confidence:.2%}" if invoice.classification_confidence else 'N/A')
            ws.cell(row=row_num, column=11, value=invoice.submitted_by.username if invoice.submitted_by else 'System')
            ws.cell(row=row_num, column=12, value=invoice.created_at.strftime('%Y-%m-%d %H:%M'))
        
        # Auto-adjust column widths
        for col_num in range(1, len(headers) + 1):
            column_letter = get_column_letter(col_num)
            max_length = len(headers[col_num - 1])
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_num, max_col=col_num):
                try:
                    cell_length = len(str(row[0].value))
                    max_length = max(max_length, cell_length)
                except:
                    pass
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
        
        # Freeze header row
        ws.freeze_panes = 'A2'
    
    def _create_summary_sheet(self, wb, invoices, filters):
        """Create summary statistics sheet"""
        ws = wb.create_sheet("Summary")
        
        # Title
        ws.cell(row=1, column=1, value="Invoice Export Summary").font = Font(bold=True, size=16)
        ws.merge_cells('A1:D1')
        
        # Export metadata
        row = 3
        ws.cell(row=row, column=1, value="Export Date:").font = Font(bold=True)
        ws.cell(row=row, column=2, value=timezone.now().strftime('%Y-%m-%d %H:%M:%S'))
        
        row += 1
        ws.cell(row=row, column=1, value="Total Records:").font = Font(bold=True)
        ws.cell(row=row, column=2, value=invoices.count())
        
        # Applied filters
        if filters:
            row += 2
            ws.cell(row=row, column=1, value="Applied Filters:").font = Font(bold=True, size=12)
            row += 1
            
            for key, value in filters.items():
                if value:
                    ws.cell(row=row, column=1, value=f"  {key.replace('_', ' ').title()}:")
                    ws.cell(row=row, column=2, value=str(value))
                    row += 1
        
        # Statistics by status
        row += 2
        ws.cell(row=row, column=1, value="Status Breakdown:").font = Font(bold=True, size=12)
        row += 1
        
        status_counts = invoices.values('status').annotate(count=Count('id')).order_by('-count')
        for item in status_counts:
            status_display = dict(InvoiceStatus.choices).get(item['status'], item['status'])
            ws.cell(row=row, column=1, value=f"  {status_display}:")
            ws.cell(row=row, column=2, value=item['count'])
            row += 1
        
        # Statistics by type
        row += 2
        ws.cell(row=row, column=1, value="Type Breakdown:").font = Font(bold=True, size=12)
        row += 1
        
        type_counts = invoices.exclude(invoice_type__isnull=True).values('invoice_type').annotate(count=Count('id')).order_by('-count')
        for item in type_counts:
            type_display = dict(InvoiceType.choices).get(item['invoice_type'], item['invoice_type'])
            ws.cell(row=row, column=1, value=f"  {type_display}:")
            ws.cell(row=row, column=2, value=item['count'])
            row += 1
        
        # Financial summary
        row += 2
        ws.cell(row=row, column=1, value="Financial Summary:").font = Font(bold=True, size=12)
        row += 1
        
        total_amount = invoices.exclude(total_amount__isnull=True).aggregate(Sum('total_amount'))['total_amount__sum'] or 0
        ws.cell(row=row, column=1, value="  Total Amount (All Currencies):")
        ws.cell(row=row, column=2, value=f"{total_amount:.2f}")
        
        # Amount by currency
        currency_totals = invoices.exclude(total_amount__isnull=True).values('currency').annotate(
            total=Sum('total_amount')
        ).order_by('-total')
        
        for item in currency_totals:
            row += 1
            ws.cell(row=row, column=1, value=f"    {item['currency']}:")
            ws.cell(row=row, column=2, value=f"{item['total']:.2f}")
        
        # Auto-adjust columns
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 30
    
    def _create_approval_sheet(self, wb, invoices):
        """Create approval details sheet"""
        ws = wb.create_sheet("Approvals")
        
        # Headers
        headers = [
            'Invoice Number', 'Approver Name', 'Approver Email', 'Level',
            'Level Name', 'Status', 'Decision', 'Comments', 'Decision Date'
        ]
        
        header_fill = PatternFill(start_color='059669', end_color='059669', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Write approval data
        row_num = 2
        for invoice in invoices:
            for approval in invoice.approvals.all():
                ws.cell(row=row_num, column=1, value=invoice.invoice_number)
                ws.cell(row=row_num, column=2, value=approval.approver_name)
                ws.cell(row=row_num, column=3, value=approval.approver_email)
                ws.cell(row=row_num, column=4, value=approval.approval_level)
                ws.cell(row=row_num, column=5, value=approval.level_name)
                ws.cell(row=row_num, column=6, value=approval.get_status_display())
                ws.cell(row=row_num, column=7, value=approval.decision or 'N/A')
                ws.cell(row=row_num, column=8, value=approval.comments or 'N/A')
                ws.cell(row=row_num, column=9, value=approval.decision_date.strftime('%Y-%m-%d %H:%M') if approval.decision_date else 'Pending')
                row_num += 1
        
        # Auto-adjust columns
        for col_num in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col_num)].width = 20
        
        ws.freeze_panes = 'A2'
    
    def _create_export_info(self, filters):
        """Create export information paragraph"""
        info_text = f"<b>Generated:</b> {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}<br/>"
        
        if filters:
            info_text += "<b>Filters Applied:</b><br/>"
            for key, value in filters.items():
                if value:
                    info_text += f"  • {key.replace('_', ' ').title()}: {value}<br/>"
        else:
            info_text += "<b>Filters:</b> None (All invoices)"
        
        return Paragraph(info_text, self.styles['Normal'])
    
    def _create_pdf_summary(self, invoices, filters):
        """Create PDF summary section"""
        content = []
        
        # Section header
        content.append(Paragraph("Summary Statistics", self.styles['SectionHeader']))
        
        # Summary data
        summary_data = [
            ['Metric', 'Value'],
            ['Total Invoices', str(invoices.count())],
        ]
        
        # Status breakdown
        status_counts = invoices.values('status').annotate(count=Count('id')).order_by('-count')
        for item in status_counts[:5]:  # Top 5 statuses
            status_display = dict(InvoiceStatus.choices).get(item['status'], item['status'])
            summary_data.append([f"  {status_display}", str(item['count'])])
        
        # Financial summary
        total_amount = invoices.exclude(total_amount__isnull=True).aggregate(Sum('total_amount'))['total_amount__sum'] or 0
        summary_data.append(['Total Amount', f"{total_amount:,.2f}"])
        
        # Create table
        summary_table = Table(summary_data, colWidths=[4*inch, 2*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e40af')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
        ]))
        
        content.append(summary_table)
        content.append(Spacer(1, 0.3 * inch))
        
        return content
    
    def _create_pdf_invoice_table(self, invoices):
        """Create PDF invoice table"""
        # Headers
        table_data = [[
            'Invoice #', 'Vendor', 'Date', 'Amount', 'Currency',
            'Type', 'Status'
        ]]
        
        # Data rows (limit to first 100 for PDF size)
        for invoice in invoices[:100]:
            table_data.append([
                invoice.invoice_number[:15],  # Truncate long numbers
                (invoice.vendor_name or 'N/A')[:20],  # Truncate long names
                invoice.invoice_date.strftime('%Y-%m-%d') if invoice.invoice_date else 'N/A',
                f"{invoice.total_amount:,.2f}" if invoice.total_amount else '0.00',
                invoice.currency,
                invoice.get_invoice_type_display()[:10] if invoice.invoice_type else 'N/A',
                invoice.get_status_display()[:15]
            ])
        
        # Create table
        invoice_table = Table(table_data, colWidths=[1*inch, 1.5*inch, 1*inch, 1*inch, 0.7*inch, 1*inch, 1.3*inch])
        invoice_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e40af')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
        ]))
        
        return invoice_table
    
    def get_export_filename(self, format='xlsx', filters=None):
        """Generate smart filename based on filters"""
        timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
        
        parts = ['invoices']
        
        if filters:
            if filters.get('status'):
                status = filters['status']
                if isinstance(status, list):
                    parts.append('multi_status')
                else:
                    parts.append(status)
            
            if filters.get('invoice_type'):
                inv_type = filters['invoice_type']
                if isinstance(inv_type, list):
                    parts.append('multi_type')
                else:
                    parts.append(inv_type)
            
            if filters.get('date_from') or filters.get('date_to'):
                parts.append('filtered')
        
        parts.append(timestamp)
        filename = '_'.join(parts)
        
        return f"{filename}.{format}"
