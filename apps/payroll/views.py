"""
Payroll Intelligence — Views
==============================
7 viewsets + 1 dashboard summary view.
All endpoints require authentication.
"""
from __future__ import annotations

import datetime
import uuid
import logging

from decimal import Decimal

from django.db.models import Sum, Count, Q, Avg
from django.utils import timezone
from rest_framework import viewsets, status
from rest_framework.decorators import action, api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.exceptions import PermissionDenied

from apps.finance.salary_models import (
    EmployeeSalaryInfo,
)
# Real payroll runs/payslips are generated through the Payroll Engine
# (apps.payroll_engine) — apps.finance.salary_models.PayrollRun/SalarySlip
# is a separate, unused-in-practice model in this deployment and is always
# empty, so the dashboard summary below reads from the engine models instead.
from apps.payroll_engine.models import (
    PayrollRun as EnginePayrollRun, Payslip as EnginePayslip,
)
from apps.payroll_engine.catalog import Status as EngineStatus

from .models import (
    PayrollValidationLog,
    PayrollAuditAlert,
    ProjectCostAllocation,
    AIInsightSnapshot,
    ChatbotMessage,
    AlertStatus,
    EmployeeLeaveRecord,
    EmployeeLeaveMonthly,
    LeaveType,
    LeaveRequest,
    LeaveRequestStatus,
    PublicHoliday,
    AttendanceOverride,
    SalaryComponent,
    EmployeeSalaryStructure,
    SalaryStructureStatus,
    SalaryHistory,
    DailyWorkLog,
    DailyWorkLogApprovalStatus,
)
from .serializers import (
    PayrollValidationLogSerializer,
    PayrollAuditAlertSerializer,
    ProjectCostAllocationSerializer,
    AIInsightSnapshotSerializer,
    ChatbotMessageSerializer,
    EmployeeLeaveRecordSerializer,
    EmployeeLeaveRecordListSerializer,
    LeaveTypeSerializer,
    LeaveRequestSerializer,
    PublicHolidaySerializer,
    AttendanceOverrideSerializer,
    SalaryComponentSerializer,
    EmployeeSalaryStructureSerializer,
    SalaryHistorySerializer,
)

logger = logging.getLogger(__name__)


# ─────────────────────────────────────────────────────────────────────────────
# 1. Dashboard Summary View
# ─────────────────────────────────────────────────────────────────────────────

class PayrollDashboardSummaryView(APIView):
    """
    Single endpoint that aggregates all KPI data for the Payroll Dashboard.
    Returns a flat dict consumed directly by the frontend KPI tiles.
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        now = timezone.now()
        current_month = now.month
        current_year  = now.year

        # ── Salary-slip KPIs (current month) ────────────────────────────────
        # PayrollRun denormalises gross/net/deductions per run, so the
        # current-month row (if generated) gives us these directly.
        current_run = EnginePayrollRun.objects.filter(
            month=current_month,
            year=current_year,
        ).first()
        slip_agg = {
            'total_gross':      current_run.total_gross if current_run else Decimal('0'),
            'total_net':        current_run.total_net if current_run else Decimal('0'),
            'total_deductions': current_run.total_deductions if current_run else Decimal('0'),
            'slip_count':       current_run.employee_count if current_run else 0,
        }

        # YTD totals
        ytd_agg = EnginePayrollRun.objects.filter(
            year=current_year,
        ).aggregate(ytd_net=Sum('total_net'))

        # "Pending approval" == payslips still in the draft stage (not yet
        # HR-approved) — the earliest actionable stage of the run workflow.
        pending_approvals = EnginePayslip.objects.filter(
            status=EngineStatus.DRAFT
        ).count()

        # ── Active employee count ────────────────────────────────────────────
        # Primary: EmployeeSalaryInfo (set when salary structures are loaded).
        # Fallback: distinct employee codes from annual leave records.
        salary_employee_count = EmployeeSalaryInfo.objects.filter(is_active=True).count()
        leave_employee_count = (
            EmployeeLeaveRecord.objects
            .filter(year=current_year)
            .exclude(employee_code__isnull=True)
            .values('employee_code')
            .distinct()
            .count()
        )
        total_employees = salary_employee_count or leave_employee_count

        avg_salary = EmployeeSalaryInfo.objects.filter(
            is_active=True
        ).aggregate(avg=Avg('basic_salary'))['avg'] or Decimal('0')

        # ── Open issues ──────────────────────────────────────────────────────
        open_validations = PayrollValidationLog.objects.filter(is_resolved=False).count()
        open_alerts      = PayrollAuditAlert.objects.filter(status=AlertStatus.OPEN).count()

        # Employees with a negative or zero leave balance are an alert
        leave_critical = (
            EmployeeLeaveRecord.objects
            .filter(year=current_year, leave_balance__lte=0)
            .count()
        )

        # ── Leave summary (annual aggregates) ───────────────────────────────
        leave_agg = EmployeeLeaveRecord.objects.filter(year=current_year).aggregate(
            total_earned=Sum('total_earned'),
            avg_balance=Avg('leave_balance'),
        )
        # "Taken" comes from actual approved LeaveRequest submissions, not the
        # EmployeeLeaveRecord.total_taken column — that field only reflects a
        # one-off HR Excel snapshot and is never updated as leave is approved
        # through the app, so it stays 0 while real leave gets taken.
        leave_taken_agg = LeaveRequest.objects.filter(
            status=LeaveRequestStatus.APPROVED,
            start_date__year=current_year,
        ).aggregate(total_taken=Sum('days_requested'))
        leave_total_taken = leave_taken_agg['total_taken'] or Decimal('0')

        leave_employees_taken = (
            LeaveRequest.objects
            .filter(status=LeaveRequestStatus.APPROVED, start_date__year=current_year)
            .exclude(employee_code__isnull=True)
            .values('employee_code')
            .distinct()
            .count()
        )

        # Current-month leave taken (from monthly breakdown table)
        current_month_leave_taken = (
            EmployeeLeaveMonthly.objects
            .filter(record__year=current_year, month=current_month)
            .aggregate(taken=Sum('taken'))['taken'] or Decimal('0')
        )

        # ── Latest payroll run ───────────────────────────────────────────────
        latest_run = EnginePayrollRun.objects.order_by('-year', '-month').first()

        return Response({
            'current_month':       current_month,
            'current_year':        current_year,
            # Employee count — accurate regardless of whether salary runs exist
            'total_employees':         total_employees,
            'salary_employees':        salary_employee_count,
            'leave_employees':         leave_employee_count,
            # Salary-slip KPIs (zero until payroll runs are processed)
            'current_month_gross': str(slip_agg['total_gross'] or 0),
            'current_month_net':   str(slip_agg['total_net'] or 0),
            'total_deductions':    str(slip_agg['total_deductions'] or 0),
            'slip_count':          slip_agg['slip_count'] or 0,
            'pending_approvals':   pending_approvals,
            'ytd_payroll':         str(ytd_agg['ytd_net'] or 0),
            'avg_basic_salary':    str(avg_salary),
            'open_validations':    open_validations,
            'open_alerts':         open_alerts + leave_critical,
            # Leave intelligence — always populated from imported leave data
            'leave_data_available':        leave_employee_count > 0,
            'leave_total_taken_ytd':       str(leave_total_taken),
            'leave_total_earned_ytd':      str(leave_agg['total_earned'] or 0),
            'leave_avg_balance':           str(round(leave_agg['avg_balance'] or 0, 2)),
            'leave_employees_taken':       leave_employees_taken,
            'leave_current_month_taken':   str(current_month_leave_taken),
            'leave_critical_alerts':       leave_critical,
            # Activity intelligence — approved daily work logs (month-to-date)
            'approved_activity_hours_mtd':  float(
                DailyWorkLog.objects
                .filter(
                    approval_status=DailyWorkLogApprovalStatus.APPROVED,
                    log_date__year=current_year,
                    log_date__month=current_month,
                )
                .aggregate(h=Sum('hours_spent'))['h'] or 0
            ),
            'approved_activity_count_mtd': DailyWorkLog.objects.filter(
                approval_status=DailyWorkLogApprovalStatus.APPROVED,
                log_date__year=current_year,
                log_date__month=current_month,
            ).count(),
            'latest_run': {
                'id':     str(latest_run.id)     if latest_run else None,
                'code':   latest_run.cycle_code  if latest_run else None,
                'month':  latest_run.month       if latest_run else None,
                'year':   latest_run.year        if latest_run else None,
                'status': latest_run.status      if latest_run else None,
            },
        })


# ─────────────────────────────────────────────────────────────────────────────
# 2. PayrollValidationLog
# ─────────────────────────────────────────────────────────────────────────────

class PayrollValidationLogViewSet(viewsets.ModelViewSet):
    queryset           = PayrollValidationLog.objects.select_related(
        'payroll_run', 'employee_salary_info__user', 'resolved_by'
    ).all()
    serializer_class   = PayrollValidationLogSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        qs = super().get_queryset()
        run_id = self.request.query_params.get('payroll_run')
        if run_id:
            qs = qs.filter(payroll_run_id=run_id)
        severity = self.request.query_params.get('severity')
        if severity:
            qs = qs.filter(severity=severity)
        is_resolved = self.request.query_params.get('is_resolved')
        if is_resolved is not None:
            qs = qs.filter(is_resolved=is_resolved.lower() == 'true')
        return qs

    @action(detail=True, methods=['post'], url_path='resolve')
    def resolve(self, request, pk=None):
        log = self.get_object()
        log.is_resolved = True
        log.resolved_by = request.user
        log.resolved_at = timezone.now()
        log.save(update_fields=['is_resolved', 'resolved_by', 'resolved_at'])
        return Response(self.get_serializer(log).data)

    @action(detail=False, methods=['post'], url_path='bulk-create')
    def bulk_create(self, request):
        """Receive a list of validation findings from the frontend rule engine."""
        items = request.data if isinstance(request.data, list) else request.data.get('items', [])
        created = []
        for item in items:
            ser = self.get_serializer(data=item)
            if ser.is_valid():
                ser.save()
                created.append(ser.data)
        return Response({'created': len(created), 'items': created}, status=status.HTTP_201_CREATED)


# ─────────────────────────────────────────────────────────────────────────────
# 3. PayrollAuditAlert
# ─────────────────────────────────────────────────────────────────────────────

class PayrollAuditAlertViewSet(viewsets.ModelViewSet):
    queryset           = PayrollAuditAlert.objects.select_related(
        'payroll_run', 'compared_to_run', 'employee_salary_info__user', 'acknowledged_by'
    ).all()
    serializer_class   = PayrollAuditAlertSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        qs = super().get_queryset()
        run_id = self.request.query_params.get('payroll_run')
        if run_id:
            qs = qs.filter(payroll_run_id=run_id)
        alert_status = self.request.query_params.get('status')
        if alert_status:
            qs = qs.filter(status=alert_status)
        severity = self.request.query_params.get('severity')
        if severity:
            qs = qs.filter(severity=severity)
        return qs

    @action(detail=True, methods=['post'], url_path='acknowledge')
    def acknowledge(self, request, pk=None):
        alert = self.get_object()
        alert.status         = AlertStatus.ACKNOWLEDGED
        alert.acknowledged_by = request.user
        alert.acknowledged_at = timezone.now()
        alert.save(update_fields=['status', 'acknowledged_by', 'acknowledged_at'])
        return Response(self.get_serializer(alert).data)

    @action(detail=True, methods=['post'], url_path='resolve')
    def resolve(self, request, pk=None):
        alert = self.get_object()
        alert.status = AlertStatus.RESOLVED
        alert.save(update_fields=['status'])
        return Response(self.get_serializer(alert).data)


# ─────────────────────────────────────────────────────────────────────────────
# 4. ProjectCostAllocation
# ─────────────────────────────────────────────────────────────────────────────

class ProjectCostAllocationViewSet(viewsets.ModelViewSet):
    queryset           = ProjectCostAllocation.objects.select_related('salary_slip').all()
    serializer_class   = ProjectCostAllocationSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        qs = super().get_queryset()
        month = self.request.query_params.get('month')
        year  = self.request.query_params.get('year')
        project = self.request.query_params.get('project_code')
        if month:
            qs = qs.filter(month=month)
        if year:
            qs = qs.filter(year=year)
        if project:
            qs = qs.filter(project_code__icontains=project)
        return qs

    @action(detail=False, methods=['get'], url_path='department-summary')
    def department_summary(self, request):
        """Aggregate cost by cost_center for charts."""
        month = request.query_params.get('month')
        year  = request.query_params.get('year')
        qs = self.get_queryset()
        if month:
            qs = qs.filter(month=month)
        if year:
            qs = qs.filter(year=year)
        summary = (
            qs.values('cost_center')
              .annotate(total_cost=Sum('allocated_cost'), total_hours=Sum('allocated_hours'))
              .order_by('-total_cost')
        )
        return Response(list(summary))


# ─────────────────────────────────────────────────────────────────────────────
# 5. AIInsightSnapshot
# ─────────────────────────────────────────────────────────────────────────────

class AIInsightSnapshotViewSet(viewsets.ModelViewSet):
    queryset           = AIInsightSnapshot.objects.select_related('employee_salary_info__user').all()
    serializer_class   = AIInsightSnapshotSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        qs = super().get_queryset()
        emp_id = self.request.query_params.get('employee_salary_info')
        month  = self.request.query_params.get('month')
        year   = self.request.query_params.get('year')
        if emp_id:
            qs = qs.filter(employee_salary_info_id=emp_id)
        if month:
            qs = qs.filter(month=month)
        if year:
            qs = qs.filter(year=year)
        return qs

    @action(detail=False, methods=['delete'], url_path='clear-expired')
    def clear_expired(self, request):
        deleted, _ = AIInsightSnapshot.objects.filter(expires_at__lt=timezone.now()).delete()
        return Response({'deleted': deleted})


# ─────────────────────────────────────────────────────────────────────────────
# 6. ChatbotMessage
# ─────────────────────────────────────────────────────────────────────────────

class ChatbotMessageViewSet(viewsets.ModelViewSet):
    queryset           = ChatbotMessage.objects.select_related('user').all()
    serializer_class   = ChatbotMessageSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        # Users can only see their own messages
        qs = super().get_queryset().filter(user=self.request.user)
        session_id = self.request.query_params.get('session_id')
        if session_id:
            qs = qs.filter(session_id=session_id)
        return qs

    def perform_create(self, serializer):
        serializer.save(user=self.request.user)


# ─────────────────────────────────────────────────────────────────────────────
# 7. Employee Leave Record ViewSet
# ─────────────────────────────────────────────────────────────────────────────

class EmployeeLeaveRecordViewSet(viewsets.ReadOnlyModelViewSet):
    """
    Read-only API for employee leave records imported from the HR Excel.
    Supports filtering by year, department, employee_code, and name search.
    Detail view includes the full monthly breakdown.
    """
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        qs = EmployeeLeaveRecord.objects.prefetch_related('monthly_breakdown').all()
        year   = self.request.query_params.get('year')
        dept   = self.request.query_params.get('department')
        code   = self.request.query_params.get('employee_code')
        search = self.request.query_params.get('search')
        if year:
            qs = qs.filter(year=year)
        if dept:
            qs = qs.filter(department__iexact=dept)
        if code:
            qs = qs.filter(employee_code=code)
        if search:
            qs = qs.filter(employee_name__icontains=search)
        branch = self.request.query_params.get('branch')
        if branch:
            qs = qs.filter(branch__iexact=branch)
        return qs.order_by('employee_name')

    def get_serializer_class(self):
        # Detail view returns monthly breakdown; list view is lightweight
        if self.action == 'retrieve':
            return EmployeeLeaveRecordSerializer
        return EmployeeLeaveRecordListSerializer


# ─────────────────────────────────────────────────────────────────────────────
# 8. LeaveType — read-only list of leave type codes
# ─────────────────────────────────────────────────────────────────────────────

class LeaveTypeViewSet(viewsets.ReadOnlyModelViewSet):
    """
    Read-only list of active leave types.
    Admins can create/edit types in the Django admin; they appear here immediately.
    """
    permission_classes = [IsAuthenticated]
    serializer_class   = LeaveTypeSerializer
    queryset           = LeaveType.objects.filter(is_active=True)


# ─────────────────────────────────────────────────────────────────────────────
# 9. LeaveRequest — full CRUD with approve / reject / cancel actions
# ─────────────────────────────────────────────────────────────────────────────

class LeaveRequestViewSet(viewsets.ModelViewSet):
    """
    Full CRUD for leave requests.
    POST to create (any auth user / HR on behalf of employee).
    Custom actions: /approve/, /reject/, /cancel/
    """
    permission_classes = [IsAuthenticated]
    serializer_class   = LeaveRequestSerializer

    # Soft-coded: RBAC role codes that always have full (all-employee) visibility
    # and final-stage approval rights, matching APPROVAL_TYPES.LEAVE.allowedRoles
    # in frontend/src/config/approvalsSystem.config.js
    HR_REVIEW_ROLE_CODES = ['hr_manager', 'hr_admin', 'super_admin', 'admin']

    @staticmethod
    def _user_employee_code(user):
        """Safely read the biometric employee_id from the user's RBAC profile."""
        try:
            return user.rbac_profile.employee_id or None
        except Exception:
            return None

    @classmethod
    def _is_hr_or_admin(cls, user):
        """True for superusers/staff or users holding an HR/Admin RBAC role."""
        if user.is_staff or user.is_superuser:
            return True
        try:
            return user.rbac_profile.roles.filter(
                code__in=cls.HR_REVIEW_ROLE_CODES, is_active=True
            ).exists()
        except Exception:
            return False

    @staticmethod
    def _is_reporting_manager(user, leave_request):
        if not leave_request.employee_id:
            return False
        try:
            from apps.hr_core.models import EmployeeMaster
            return EmployeeMaster.objects.filter(
                user_id=leave_request.employee_id, manager__user=user
            ).exists()
        except Exception:
            return False

    def get_queryset(self):
        qs = (
            LeaveRequest.objects
            .select_related('leave_type', 'employee', 'reviewed_by')
            .all()
        )
        user = self.request.user
        mine_only = self.request.query_params.get('mine', '').lower() in ('true', '1', 'yes')

        # HR / Admin roles have unrestricted visibility (they perform the
        # final Stage-2 approval regardless of reporting line); everyone else
        # sees only their own leave requests + requests submitted by their
        # direct reports (Reporting Manager Stage-1 approval).
        # `?mine=true` overrides this for HR/Admin too — used by the Employee
        # Self-Service page so an HR/Admin viewing their OWN leave data isn't
        # shown every employee's requests just because their role has broader
        # review visibility elsewhere in the app.
        if mine_only or not self._is_hr_or_admin(user):
            emp_code = self._user_employee_code(user)
            own_q = Q(employee=user)
            if emp_code:
                own_q |= Q(employee_code=emp_code)

            # Direct reports are only relevant for the Reporting-Manager queue
            # view — never included for an explicit "mine only" request.
            if not mine_only:
                try:
                    from apps.rbac.models import UserProfile
                    subordinates = UserProfile.objects.filter(manager__user=user, is_deleted=False)
                    sub_user_ids = list(
                        subordinates.exclude(user__isnull=True).values_list('user_id', flat=True)
                    )
                    sub_codes = [c for c in subordinates.values_list('employee_id', flat=True) if c]
                    if sub_user_ids:
                        own_q |= Q(employee_id__in=sub_user_ids)
                    if sub_codes:
                        own_q |= Q(employee_code__in=sub_codes)
                except Exception:
                    pass

            qs = qs.filter(own_q)

        params  = self.request.query_params
        st      = params.get('status')
        # SOFT-CODED: the Approvals dashboard (frontend/src/pages/ApprovalsPageDynamic.jsx)
        # requests the "pending" tab with a comma-separated `status__in` param
        # (e.g. `status__in=PENDING,RM_APPROVED`) so it can show both approval
        # stages at once. This param was previously ignored here, so the
        # backend returned ALL leave requests (including already-APPROVED /
        # REJECTED ones) and the frontend's own guard then rejected acting on
        # them with "Invalid status for approve: <status>". Support it.
        st_in   = params.get('status__in')
        code    = params.get('employee_code')
        year    = params.get('year')
        month   = params.get('month')
        search  = params.get('search')
        if st:
            qs = qs.filter(status__iexact=st)
        elif st_in:
            statuses = [s.strip().upper() for s in st_in.split(',') if s.strip()]
            if statuses:
                qs = qs.filter(status__in=statuses)
        if code:
            qs = qs.filter(employee_code=code)
        if year and month:
            import datetime as _dt, calendar as _cal
            y, m = int(year), int(month)
            first = _dt.date(y, m, 1)
            last  = _dt.date(y, m, _cal.monthrange(y, m)[1])
            qs = qs.filter(start_date__lte=last, end_date__gte=first)
        elif year:
            qs = qs.filter(start_date__year=int(year))
        if search:
            qs = qs.filter(employee_name__icontains=search)
        return qs.order_by('-created_at')

    def perform_create(self, serializer):
        """Auto-link the leave request to the authenticated user — unless HR/Admin
        is submitting on behalf of a different employee (a different employee_code
        was explicitly supplied), in which case link to THAT employee's own User
        account instead. Previously `employee` was force-set to the submitter
        unconditionally, so "on behalf of" submissions got mis-attributed to
        whichever HR/Admin filed them rather than the actual target employee."""
        user     = self.request.user
        emp_code = self._user_employee_code(user)
        emp_name = f'{user.first_name} {user.last_name}'.strip() or user.username
        supplied_code = serializer.validated_data.get('employee_code')

        extra = {}
        if supplied_code and supplied_code != emp_code:
            # Submitting on behalf of someone else — resolve their real User
            # account so `employee` (FK) isn't mis-attributed to the submitter.
            try:
                from apps.rbac.models import UserProfile
                target = UserProfile.objects.filter(
                    employee_id=supplied_code, is_deleted=False
                ).select_related('user').first()
                if target and target.user_id:
                    extra['employee'] = target.user
            except Exception:
                pass
        else:
            extra['employee'] = user
            if emp_code and not supplied_code:
                extra['employee_code'] = emp_code

        # Only fill denormalised fields if the caller didn’t supply them
        if not serializer.validated_data.get('employee_name'):
            extra['employee_name'] = emp_name
        leave_request = serializer.save(**extra)
        try:
            from apps.hr_core.identity import EmployeeIdentityService
            from apps.hr_core.workflows import HRWorkflowService

            canonical_employee = EmployeeIdentityService.resolve(
                leave_request.employee_id or leave_request.employee_code
            )
            if canonical_employee:
                workflow = HRWorkflowService.start(
                    'leave_request_v1',
                    'payroll.leave_request',
                    leave_request.pk,
                    employee=canonical_employee,
                    requested_by=user,
                    context={
                        'leave_type': leave_request.leave_type.code,
                        'start_date': str(leave_request.start_date),
                        'end_date': str(leave_request.end_date),
                        'days_requested': str(leave_request.days_requested),
                    },
                )
                leave_request.canonical_employee = canonical_employee
                leave_request.workflow_instance = workflow
                leave_request.save(update_fields=['canonical_employee', 'workflow_instance', 'updated_at'])
        except Exception:
            logger.exception('Unable to attach shared workflow to leave request %s', leave_request.pk)

    # Soft-coded: statuses from which the final (HR) approve/reject action is
    # allowed. A request can be finalised either directly from PENDING
    # (single-stage / HR self-service) or from RM_APPROVED (after the
    # Reporting Manager has completed Stage 1).
    FINAL_APPROVABLE_STATUSES = (LeaveRequestStatus.PENDING, LeaveRequestStatus.RM_APPROVED)

    @action(detail=True, methods=['post'], url_path='approve')
    def approve(self, request, pk=None):
        """Stage 2 (or single-stage): final approval — HR Manager / Admin."""
        req = self.get_object()
        if not self._is_hr_or_admin(request.user):
            raise PermissionDenied('Only HR can give final leave approval.')
        if req.status not in self.FINAL_APPROVABLE_STATUSES:
            return Response(
                {'error': f'Cannot approve a {req.status} request'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if req.workflow_instance_id:
            from apps.hr_core.workflows import HRWorkflowService
            if req.workflow_instance.current_stage_id and req.workflow_instance.current_stage.code == 'manager_review':
                return Response(
                    {'error': 'Reporting-manager approval is required before HR final approval.'},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            HRWorkflowService.decide(
                req.workflow_instance, request.user, 'approve', request.data.get('note', '')
            )
        req.status       = LeaveRequestStatus.APPROVED
        req.reviewed_by  = request.user
        req.reviewed_at  = timezone.now()
        req.reviewer_note = request.data.get('note', '')
        req.save()
        return Response(LeaveRequestSerializer(req).data)

    @action(detail=True, methods=['post'], url_path='reject')
    def reject(self, request, pk=None):
        """Stage 2 (or single-stage): final rejection — HR Manager / Admin."""
        req = self.get_object()
        if not self._is_hr_or_admin(request.user):
            raise PermissionDenied('Only HR can give final leave rejection.')
        if req.status not in self.FINAL_APPROVABLE_STATUSES:
            return Response(
                {'error': f'Cannot reject a {req.status} request'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if req.workflow_instance_id:
            from apps.hr_core.workflows import HRWorkflowService
            if req.workflow_instance.current_stage_id and req.workflow_instance.current_stage.code == 'manager_review':
                return Response(
                    {'error': 'Use reporting-manager rejection for the current workflow stage.'},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            HRWorkflowService.decide(
                req.workflow_instance, request.user, 'reject', request.data.get('note', '')
            )
        req.status       = LeaveRequestStatus.REJECTED
        req.reviewed_by  = request.user
        req.reviewed_at  = timezone.now()
        req.reviewer_note = request.data.get('note', '')
        req.save()
        return Response(LeaveRequestSerializer(req).data)

    @action(detail=True, methods=['post'], url_path='rm-approve')
    def rm_approve(self, request, pk=None):
        """Stage 1: Direct Reporting Manager approval (PENDING → RM_APPROVED)."""
        req = self.get_object()
        if req.status != LeaveRequestStatus.PENDING:
            return Response(
                {'error': f'Cannot approve a {req.status} request at the reporting-manager stage'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if req.workflow_instance_id:
            from apps.hr_core.workflows import HRWorkflowService
            HRWorkflowService.decide(
                req.workflow_instance, request.user, 'approve', request.data.get('note', '')
            )
        elif not self._is_reporting_manager(request.user, req):
            raise PermissionDenied('Only the employee\'s reporting manager can review this request.')
        req.status         = LeaveRequestStatus.RM_APPROVED
        req.rm_reviewed_by = request.user
        req.rm_reviewed_at = timezone.now()
        req.rm_note        = request.data.get('note', '')
        req.save()
        return Response(LeaveRequestSerializer(req).data)

    @action(detail=True, methods=['post'], url_path='rm-reject')
    def rm_reject(self, request, pk=None):
        """Stage 1: Direct Reporting Manager rejection (PENDING → RM_REJECTED)."""
        req = self.get_object()
        if req.status != LeaveRequestStatus.PENDING:
            return Response(
                {'error': f'Cannot reject a {req.status} request at the reporting-manager stage'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if req.workflow_instance_id:
            from apps.hr_core.workflows import HRWorkflowService
            HRWorkflowService.decide(
                req.workflow_instance, request.user, 'reject', request.data.get('note', '')
            )
        elif not self._is_reporting_manager(request.user, req):
            raise PermissionDenied('Only the employee\'s reporting manager can review this request.')
        req.status         = LeaveRequestStatus.RM_REJECTED
        req.rm_reviewed_by = request.user
        req.rm_reviewed_at = timezone.now()
        req.rm_note        = request.data.get('note', '')
        req.save()
        return Response(LeaveRequestSerializer(req).data)

    @action(detail=True, methods=['post'], url_path='cancel')
    def cancel(self, request, pk=None):
        req = self.get_object()
        if req.employee_id != request.user.id and not self._is_hr_or_admin(request.user):
            raise PermissionDenied('Only the employee or HR can cancel this request.')
        if req.workflow_instance_id:
            from apps.hr_core.workflows import HRWorkflowService
            HRWorkflowService.cancel(
                req.workflow_instance, request.user, request.data.get('note', '')
            )
        req.status = LeaveRequestStatus.CANCELLED
        req.save()
        return Response(LeaveRequestSerializer(req).data)


# ─────────────────────────────────────────────────────────────────────────────
# 10. Leave Calendar  — per-employee, per-day approved leave map
#     Consumed by the Summary attendance pivot table to overlay leave codes.
# ─────────────────────────────────────────────────────────────────────────────
# 10. branch_employee_codes — lightweight codes-only endpoint for branch filter
# ─────────────────────────────────────────────────────────────────────────────

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def branch_employee_codes(request):
    """
    GET /api/v1/payroll/branch-employee-codes/?branch=RIN&year=2026

    Returns a flat list of employee_code values for a given branch (and
    optionally a year).  Used by the frontend Attendance Summary tab to
    filter biometric rows by legal entity without fetching full leave records.

    Response: { "branch": "RIN", "year": 2026, "codes": ["E001", "E002", …] }
    """
    branch = request.GET.get('branch', '').upper().strip()
    year   = request.GET.get('year')

    if not branch:
        return Response({'error': 'branch parameter is required.'}, status=400)

    qs = EmployeeLeaveRecord.objects.filter(branch__iexact=branch)
    if year:
        qs = qs.filter(year=int(year))

    codes = list(qs.values_list('employee_code', flat=True).order_by('employee_code'))
    return Response({'branch': branch, 'year': int(year) if year else None, 'codes': codes})


# ─────────────────────────────────────────────────────────────────────────────

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def leave_calendar(request):
    """
    GET /api/v1/payroll/leave-calendar/?year=2026&month=6

    Returns approved leave for all employees in a given month, keyed by
    employee_code → { YYYY-MM-DD: {code, name, employee_name, color, badge_bg, badge_text, request_id} }.
    Only working days (Mon–Fri) are included.  Used by the Summary attendance
    view to overlay leave codes on the biometric attendance pivot table.
    """
    import calendar as _cal
    import datetime as _dt

    year  = int(request.GET.get('year',  timezone.now().year))
    month = int(request.GET.get('month', timezone.now().month))

    first_day = _dt.date(year, month, 1)
    last_day  = _dt.date(year, month, _cal.monthrange(year, month)[1])

    requests = (
        LeaveRequest.objects
        .select_related('leave_type')
        .filter(
            status=LeaveRequestStatus.APPROVED,
            start_date__lte=last_day,
            end_date__gte=first_day,
        )
    )

    calendar_data: dict = {}
    for req in requests:
        if not req.employee_code:
            continue
        emp = req.employee_code
        if emp not in calendar_data:
            calendar_data[emp] = {}
        # Expand date range, clamped to the requested month
        cur = max(req.start_date, first_day)
        end = min(req.end_date, last_day)
        while cur <= end:
            if cur.weekday() < 5:   # weekdays only (0=Mon…4=Fri)
                calendar_data[emp][cur.isoformat()] = {
                    'code':       req.leave_type.code,
                    'name':       req.leave_type.name,
                    'employee_name': req.employee_name or emp,
                    'color':      req.leave_type.color_hex,
                    'badge_bg':   req.leave_type.badge_bg,
                    'badge_text': req.leave_type.badge_text,
                    'request_id': str(req.id),
                }
            cur += _dt.timedelta(days=1)

    return Response({'year': year, 'month': month, 'calendar': calendar_data})


# =============================================================================
# Helper: is_hr_manager
# =============================================================================
def _is_hr_manager(user) -> bool:
    """Return True if the user holds an HR Manager (or admin) role.

    Soft-coded role codes live in apps.rbac -- we look for any role whose code
    starts with 'hr' or equals 'admin'/'superadmin'.  This keeps the check
    forward-compatible: adding a new HR sub-role in the RBAC admin will
    automatically grant access here without code changes.

    Falls back to user.is_staff / user.is_superuser so Django admin accounts
    always retain access even if RBAC is not fully configured.
    """
    if user.is_superuser or user.is_staff:
        return True
    try:
        from apps.rbac.models import UserProfile
        profile = UserProfile.objects.filter(user=user, is_deleted=False).first()
        if profile and profile.role:
            code = (profile.role.code or '').lower()
            return code.startswith('hr') or code in ('admin', 'superadmin', 'manager')
    except Exception:
        pass
    return False


# =============================================================================
# 10. PublicHoliday ViewSet
# =============================================================================

class PublicHolidayViewSet(viewsets.ModelViewSet):
    """
    CRUD for public holidays.

    List/Retrieve: any authenticated user (read-only employees can see the calendar).
    Create/Update/Deactivate: HR Manager or admin only.

    Soft-coded filter params:
      ?year=2026          filter by year (default: current year)
      ?region=AE-AZ       filter by region (default: no filter ? return all)
      ?active_only=true   return only is_active=True entries (default: true)
    """
    serializer_class   = PublicHolidaySerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        params   = self.request.query_params
        year     = int(params.get('year', datetime.date.today().year))
        region   = params.get('region', None)
        active   = params.get('active_only', 'true').lower() not in ('0', 'false', 'no')

        qs = PublicHoliday.objects.filter(date__year=year)
        if region:
            qs = qs.filter(region=region)
        if active:
            qs = qs.filter(is_active=True)
        return qs.order_by('date')

    def perform_create(self, serializer):
        if not _is_hr_manager(self.request.user):
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied('Only HR Managers can create public holidays.')
        serializer.save(
            created_by=self.request.user,
            updated_by=self.request.user,
            source='hr_added',
        )

    def perform_update(self, serializer):
        if not _is_hr_manager(self.request.user):
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied('Only HR Managers can update public holidays.')
        serializer.save(updated_by=self.request.user)

    def destroy(self, request, *args, **kwargs):
        """Override destroy to deactivate instead of hard-delete."""
        if not _is_hr_manager(request.user):
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied('Only HR Managers can deactivate public holidays.')
        obj = self.get_object()
        obj.is_active = False
        obj.updated_by = request.user
        obj.save(update_fields=['is_active', 'updated_by', 'updated_at'])
        return Response(status=status.HTTP_204_NO_CONTENT)


# =============================================================================
# 11. AttendanceOverride ViewSet
# =============================================================================

class AttendanceOverrideViewSet(viewsets.ModelViewSet):
    """
    HR Manager manual corrections for Summary pivot cells.

    List: returns all active overrides for a given month.
        ?year=2026&month=6     filter by year+month (required for Summary tab)
        ?employee_code=E001    filter by specific employee
    Retrieve/Create/Update/Deactivate: HR Manager or admin only.
    Delete is disabled -- overrides are deactivated (is_active=False) instead.

    The frontend applies overrides by building a lookup dict:
        { 'E001': { '2026-06-15': { override_hours: 8.0, reason: '...', ... } } }
    so only ONE override per (employee_code, date) is returned -- the most recent
    active one.  The endpoint returns flat rows; deduplication is in the frontend.
    """
    serializer_class   = AttendanceOverrideSerializer
    permission_classes = [IsAuthenticated]
    http_method_names  = ['get', 'post', 'patch', 'head', 'options']  # no DELETE

    def get_queryset(self):
        params        = self.request.query_params
        year          = params.get('year')
        month         = params.get('month')
        employee_code = params.get('employee_code')

        qs = AttendanceOverride.objects.filter(is_active=True)
        if year and month:
            qs = qs.filter(date__year=int(year), date__month=int(month))
        if employee_code:
            qs = qs.filter(employee_code=employee_code)
        return qs.order_by('employee_code', 'date')

    def _require_hr(self):
        if not _is_hr_manager(self.request.user):
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied('Only HR Managers can edit attendance overrides.')

    def perform_create(self, serializer):
        self._require_hr()
        # Deactivate any previous override for same (employee_code, date) to
        # ensure only one active record per cell.
        employee_code = serializer.validated_data.get('employee_code')
        date          = serializer.validated_data.get('date')
        if employee_code and date:
            AttendanceOverride.objects.filter(
                employee_code=employee_code,
                date=date,
                is_active=True,
            ).update(is_active=False)
        serializer.save(created_by=self.request.user)

    def perform_update(self, serializer):
        self._require_hr()
        serializer.save()


# =============================================================================
# Salary Management helpers
# =============================================================================

def _is_senior_hr(user) -> bool:
    """True for superuser, staff, or roles with senior-level HR access."""
    if getattr(user, 'is_superuser', False) or getattr(user, 'is_staff', False):
        return True
    profile = getattr(user, 'rbac_profile', None)
    if not profile:
        return False
    role_codes = profile.roles.filter(is_active=True).values_list('code', flat=True)
    return any(
        (code or '').lower().startswith('senior_hr')
        or (code or '').lower() in {'hr_admin', 'hr_manager', 'admin', 'super_admin', 'superadmin', 'manager'}
        for code in role_codes
    )


# =============================================================================
# 10. SalaryComponent
# =============================================================================

class SalaryComponentViewSet(viewsets.ModelViewSet):
    """
    HR Managers can create/update component types.
    Senior HR / Admin can deactivate.
    Read access for all authenticated users.
    """
    serializer_class   = SalaryComponentSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        qs = SalaryComponent.objects.all()
        active_only = self.request.query_params.get('active')
        if active_only and active_only.lower() == 'true':
            qs = qs.filter(is_active=True)
        cat = self.request.query_params.get('category')
        if cat:
            qs = qs.filter(category=cat)
        return qs

    def _require_hr(self):
        if not _is_hr_manager(self.request.user):
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied('HR Manager role required.')

    def perform_create(self, serializer):
        self._require_hr()
        serializer.save(created_by=self.request.user)

    def perform_update(self, serializer):
        self._require_hr()
        serializer.save()

    def destroy(self, request, *args, **kwargs):
        """Soft-delete (deactivate) instead of hard delete."""
        if not _is_senior_hr(request.user):
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied('Senior HR role required to deactivate components.')
        obj = self.get_object()
        obj.is_active = False
        obj.save(update_fields=['is_active'])
        return Response(status=status.HTTP_204_NO_CONTENT)


# =============================================================================
# 11. EmployeeSalaryStructure
# =============================================================================

class EmployeeSalaryStructureViewSet(viewsets.ModelViewSet):
    """
    Workflow: DRAFT -> PENDING_APPROVAL -> APPROVED | REJECTED

    Endpoints:
      POST   salary-structures/                   create (HR Manager)
      PATCH  salary-structures/{id}/              update draft
      POST   salary-structures/{id}/submit/       submit for approval
      POST   salary-structures/{id}/approve/      approve (Senior HR)
      POST   salary-structures/{id}/reject/       reject (Senior HR)
      GET    salary-structures/pending/           list pending (Senior HR)
      GET    salary-structures/summary/           one active row per employee
    """
    serializer_class   = EmployeeSalaryStructureSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        qs = EmployeeSalaryStructure.objects.all()
        emp = self.request.query_params.get('employee_code')
        if emp:
            qs = qs.filter(employee_code=emp)
        stat = self.request.query_params.get('status')
        if stat:
            qs = qs.filter(status=stat)
        active = self.request.query_params.get('active')
        if active and active.lower() == 'true':
            qs = qs.filter(is_active=True)
        return qs

    def _require_hr(self):
        if not _is_hr_manager(self.request.user):
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied('HR Manager role required.')

    def _require_senior_hr(self):
        if not _is_senior_hr(self.request.user):
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied('Senior HR role required.')

    def perform_create(self, serializer):
        self._require_hr()
        serializer.save(
            created_by=self.request.user,
            status=SalaryStructureStatus.DRAFT,
        )

    def perform_update(self, serializer):
        self._require_hr()
        obj = self.get_object()
        if obj.status not in (SalaryStructureStatus.DRAFT, SalaryStructureStatus.REJECTED):
            from rest_framework.exceptions import ValidationError
            raise ValidationError('Only DRAFT or REJECTED structures can be edited.')
        serializer.save()

    def destroy(self, request, *args, **kwargs):
        """Soft-delete (deactivate) — no hard deletes."""
        self._require_hr()
        obj = self.get_object()
        if obj.status == SalaryStructureStatus.APPROVED:
            from rest_framework.exceptions import ValidationError
            raise ValidationError('Cannot delete an APPROVED salary structure.')
        obj.is_active = False
        obj.save(update_fields=['is_active', 'updated_at'])
        return Response(status=status.HTTP_204_NO_CONTENT)

    @action(detail=True, methods=['post'], url_path='submit')
    def submit(self, request, pk=None):
        """HR Manager submits draft for approval."""
        self._require_hr()
        obj = self.get_object()
        if obj.status != SalaryStructureStatus.DRAFT:
            return Response(
                {'detail': 'Only DRAFT structures can be submitted.'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        obj.status       = SalaryStructureStatus.PENDING_APPROVAL
        obj.submitted_by = request.user
        obj.submitted_at = timezone.now()
        obj.save(update_fields=['status', 'submitted_by', 'submitted_at', 'updated_at'])
        return Response(EmployeeSalaryStructureSerializer(obj).data)

    @action(detail=True, methods=['post'], url_path='approve')
    def approve(self, request, pk=None):
        """Senior HR approves a pending structure and writes SalaryHistory."""
        self._require_senior_hr()
        obj = self.get_object()
        if obj.status != SalaryStructureStatus.PENDING_APPROVAL:
            return Response(
                {'detail': 'Only PENDING_APPROVAL structures can be approved.'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        # Deactivate prior active structure
        prev_qs = EmployeeSalaryStructure.objects.filter(
            employee_code=obj.employee_code,
            is_active=True,
            status=SalaryStructureStatus.APPROVED,
        ).exclude(pk=obj.pk)
        prev = prev_qs.first()
        previous_basic = prev.basic_salary if prev else None
        previous_net   = prev.net_salary   if prev else None
        if prev:
            prev.is_active     = False
            prev.superseded_by = obj
            prev.save(update_fields=['is_active', 'superseded_by', 'updated_at'])

        # Approve current
        obj.status      = SalaryStructureStatus.APPROVED
        obj.reviewed_by = request.user
        obj.reviewed_at = timezone.now()
        obj.reviewer_note = request.data.get('reviewer_note', '')
        obj.is_active   = True
        obj.save(update_fields=[
            'status', 'reviewed_by', 'reviewed_at', 'reviewer_note',
            'is_active', 'updated_at',
        ])

        # Compute change_percent
        change_pct = None
        if previous_net and previous_net != 0:
            change_pct = round(
                ((obj.net_salary - previous_net) / previous_net) * 100,
                2,
            )

        # Write audit history
        SalaryHistory.objects.create(
            employee_code  = obj.employee_code,
            employee_name  = obj.employee_name,
            change_date    = obj.effective_date,
            previous_basic = previous_basic,
            new_basic      = obj.basic_salary,
            previous_net   = previous_net,
            new_net        = obj.net_salary,
            change_percent = change_pct,
            change_reason  = obj.reviewer_note or 'Salary structure approved',
            structure      = obj,
            approved_by    = request.user,
        )

        return Response(EmployeeSalaryStructureSerializer(obj).data)

    @action(detail=True, methods=['post'], url_path='reject')
    def reject(self, request, pk=None):
        """Senior HR rejects a pending structure."""
        self._require_senior_hr()
        obj = self.get_object()
        if obj.status != SalaryStructureStatus.PENDING_APPROVAL:
            return Response(
                {'detail': 'Only PENDING_APPROVAL structures can be rejected.'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        obj.status        = SalaryStructureStatus.REJECTED
        obj.reviewed_by   = request.user
        obj.reviewed_at   = timezone.now()
        obj.reviewer_note = request.data.get('reviewer_note', '')
        obj.save(update_fields=[
            'status', 'reviewed_by', 'reviewed_at', 'reviewer_note', 'updated_at',
        ])
        return Response(EmployeeSalaryStructureSerializer(obj).data)

    @action(detail=False, methods=['get'], url_path='pending')
    def pending(self, request):
        """Return all structures awaiting approval (Senior HR only)."""
        self._require_senior_hr()
        qs = EmployeeSalaryStructure.objects.filter(
            status=SalaryStructureStatus.PENDING_APPROVAL,
        ).order_by('submitted_at')
        return Response(EmployeeSalaryStructureSerializer(qs, many=True).data)

    @action(detail=False, methods=['get'], url_path='summary')
    def summary(self, request):
        """
        Return one active APPROVED structure per employee (current salary).
        Optionally filter by department.
        """
        qs = EmployeeSalaryStructure.objects.filter(
            is_active=True,
            status=SalaryStructureStatus.APPROVED,
        ).order_by('employee_name')
        dept = request.query_params.get('department')
        if dept:
            qs = qs.filter(department__icontains=dept)
        return Response(EmployeeSalaryStructureSerializer(qs, many=True).data)


# =============================================================================
# 12. SalaryHistory
# =============================================================================

class SalaryHistoryViewSet(viewsets.ReadOnlyModelViewSet):
    """
    Read-only. Returns the salary change audit trail.
    Filter by employee_code or date range.
    """
    serializer_class   = SalaryHistorySerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        qs  = SalaryHistory.objects.all()
        emp = self.request.query_params.get('employee_code')
        if emp:
            qs = qs.filter(employee_code=emp)
        date_from = self.request.query_params.get('date_from')
        date_to   = self.request.query_params.get('date_to')
        if date_from:
            qs = qs.filter(change_date__gte=date_from)
        if date_to:
            qs = qs.filter(change_date__lte=date_to)
        return qs


# =============================================================================
# Annual Leave Balance Summary
# =============================================================================

def _norm_name(name: str) -> str:
    """Normalise an employee name to a stable lowercase-stripped key for fuzzy match."""
    return ' '.join((name or '').lower().split())


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def annual_leave_balance(request):
    """
    GET /api/v1/payroll/annual-leave-balance/?year=2026&month=6

    Returns the YTD leave balance for every employee as of end of *month*.
    The response carries TWO lookup indices so the frontend can match even
    when biometric employee_code format differs from HR Excel codes:

      balances          → keyed by employee_code  (primary lookup)
      balances_by_name  → keyed by normalised lowercase name (fallback)

    Each value:
      { employee_name, joining_date, carryforward,
        earned_ytd, taken_ytd, encashed_ytd, balance,
        annual_entitlement }

    Computation: 22 days/year accrual service formula (pro-rated joining month).
    taken/encashed come from the stored HR-Excel import rows; earned is always
    recomputed from the formula so it stays accurate even if monthly rows are stale.
    """
    from apps.payroll.models import EmployeeLeaveRecord
    from apps.payroll.services.leave_accrual import (
        compute_monthly_earned, compute_running_balance, _dec, ANNUAL_LEAVE_DAYS,
    )
    from datetime import date as _date

    year  = int(request.GET.get('year',  timezone.now().year))
    month = int(request.GET.get('month', timezone.now().month))

    qs = (
        EmployeeLeaveRecord.objects
        .prefetch_related('monthly_breakdown')
        .filter(year=year)
    )
    branch = request.GET.get('branch')
    if branch:
        qs = qs.filter(branch__iexact=branch)

    today  = _date.today()
    result_by_code = {}
    result_by_name = {}

    for rec in qs:
        # Build monthly taken/encashed from stored import rows (HR source of truth)
        stored = {
            r.month: r
            for r in rec.monthly_breakdown.all()
        }

        # Always recompute earned from formula; use stored taken/encashed
        monthly_rows = []
        for m in range(1, month + 1):
            earned   = compute_monthly_earned(
                rec.joining_date, year, m,
                rec.annual_entitlement or ANNUAL_LEAVE_DAYS, today,
            )
            row = stored.get(m)
            taken    = _dec(row.taken    if row else 0)
            encashed = _dec(row.encashed if row else 0)
            monthly_rows.append({'month': m, 'earned': earned, 'taken': taken, 'encashed': encashed})

        balances   = compute_running_balance(_dec(rec.carryforward), monthly_rows, up_to_month=month)
        balance_at = float(balances.get(month, 0))
        earned_ytd   = round(sum(float(r['earned'])   for r in monthly_rows), 4)
        taken_ytd    = round(sum(float(r['taken'])    for r in monthly_rows), 4)
        encashed_ytd = round(sum(float(r['encashed']) for r in monthly_rows), 4)

        payload = {
            'employee_name':      rec.employee_name,
            'joining_date':       rec.joining_date.isoformat() if rec.joining_date else None,
            'carryforward':       float(rec.carryforward),
            'earned_ytd':         earned_ytd,
            'taken_ytd':          taken_ytd,
            'encashed_ytd':       encashed_ytd,
            'balance':            round(balance_at, 4),
            'annual_entitlement': int(rec.annual_entitlement or ANNUAL_LEAVE_DAYS),
        }

        # Primary key: employee_code (string)
        if rec.employee_code:
            result_by_code[str(rec.employee_code).strip()] = payload

        # Fallback key: normalised name (handles code-format mismatches)
        norm = _norm_name(rec.employee_name)
        if norm:
            result_by_name[norm] = payload

    return Response({
        'year':             year,
        'month':            month,
        'balances':         result_by_code,    # keyed by employee_code
        'balances_by_name': result_by_name,    # keyed by normalised name
    })


# =============================================================================
# Leave Encashment
# =============================================================================

def _encashment_period(request):
    """Validate and return the requested encashment (year, month)."""
    source = request.data if request.method == 'POST' else request.query_params
    try:
        year = int(source.get('year', timezone.now().year))
        month = int(source.get('month', timezone.now().month))
    except (TypeError, ValueError):
        raise ValueError('Year and month must be numbers.')
    if not 2000 <= year <= 2100:
        raise ValueError('Year must be between 2000 and 2100.')
    if not 1 <= month <= 12:
        raise ValueError('Month must be between 1 and 12.')
    return year, month


def _require_encashment_manager(user):
    if not _is_hr_manager(user):
        from rest_framework.exceptions import PermissionDenied
        raise PermissionDenied('HR Manager role required to manage leave encashment.')


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def leave_encashment_status(request):
    """Return the immutable audit status for one encashment period."""
    _require_encashment_manager(request.user)
    try:
        year, month = _encashment_period(request)
    except ValueError as exc:
        return Response({'error': str(exc)}, status=status.HTTP_400_BAD_REQUEST)

    from apps.payroll.services.leave_encashment import get_encashment_status
    result = get_encashment_status(year, month)
    if result is None:
        return Response(
            {'detail': f'No encashment run found for {year}-{month:02d}.'},
            status=status.HTTP_404_NOT_FOUND,
        )
    return Response(result)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def leave_encashment_preview(request):
    """Preview employee days, salary rate and pay without changing data."""
    _require_encashment_manager(request.user)
    try:
        year, month = _encashment_period(request)
    except ValueError as exc:
        return Response({'error': str(exc)}, status=status.HTTP_400_BAD_REQUEST)

    from apps.payroll.services.leave_encashment import run_leave_encashment
    result = run_leave_encashment(year=year, month=month, dry_run=True)
    return Response({'year': year, 'month': month, **result})


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def leave_encashment_run(request):
    """Post a reviewed encashment once and update employee leave balances."""
    _require_encashment_manager(request.user)
    try:
        year, month = _encashment_period(request)
    except ValueError as exc:
        return Response({'error': str(exc)}, status=status.HTTP_400_BAD_REQUEST)

    from apps.payroll.services.leave_encashment import (
        EncashmentAlreadyRunError,
        run_leave_encashment,
    )
    try:
        result = run_leave_encashment(
            year=year,
            month=month,
            triggered_by_user=request.user,
        )
    except EncashmentAlreadyRunError as exc:
        return Response({'error': str(exc)}, status=status.HTTP_409_CONFLICT)
    return Response({'year': year, 'month': month, **result}, status=status.HTTP_201_CREATED)


# =============================================================================
# HR Admin: Upload leave Excel + trigger import+compute (one-shot seed)
# =============================================================================

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def sync_leave_data(request):
    """
    POST /api/v1/payroll/sync-leave-data/

    HR Manager only.  Accepts a multipart Excel upload (field: file) and:
      1. Saves it to a temp path
      2. Runs import_leave_excel  (upsert all EmployeeLeaveRecord rows)
      3. Runs compute_leave_accrual  (recompute earned + balance)

    Returns a summary dict:  { imported, updated, errors, computed, year }

    Allows seeding / refreshing production from the HR Manager's browser
    without needing Railway CLI access.
    """
    import os, tempfile, warnings
    from decimal import Decimal, InvalidOperation
    from apps.payroll.models import EmployeeLeaveRecord, EmployeeLeaveMonthly
    from apps.payroll.management.commands.import_leave_excel import (
        LEAVE_EXCEL_MAP, BRANCH_OVERRIDES, _to_dec, _to_str,
        _clean_emp_code, _clean_title,
    )
    from apps.payroll.services.leave_accrual import compute_accrual_for_record
    from django.db import transaction

    if not _is_hr_manager(request.user):
        raise PermissionDenied('HR Manager role required to sync leave data.')

    upload = request.FILES.get('file')
    if not upload:
        return Response({'error': 'No file uploaded. POST with field name "file".'}, status=400)

    year   = int(request.data.get('year',   timezone.now().year))
    branch = (request.data.get('branch') or 'RAD').upper().strip()

    # Soft-coded column map for RAD branch
    override = BRANCH_OVERRIDES.get(branch, {})
    em   = {**LEAVE_EXCEL_MAP, **{k: v for k, v in override.items() if k not in ('default_path',)}}
    src  = override.get('source_tag', LEAVE_EXCEL_MAP['source_tag'])
    mmap = em['month_map']

    try:
        import openpyxl
    except ImportError:
        return Response({'error': 'openpyxl not installed on this server.'}, status=500)

    # Save upload to a temp file
    suffix = os.path.splitext(upload.name)[1] or '.xlsx'
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=suffix)
    try:
        for chunk in upload.chunks():
            tmp.write(chunk)
        tmp.close()

        warnings.filterwarnings('ignore')
        wb = openpyxl.load_workbook(tmp.name, data_only=True)
        sheets = wb.sheetnames

        created = updated = skipped = import_errors = 0

        for sheet_name in sheets:
            ws = wb[sheet_name]
            raw_name    = ws.cell(*em['name_cell']).value
            raw_emp_no  = ws.cell(*em['emp_no_cell']).value
            raw_dept    = ws.cell(*em['dept_cell']).value
            raw_title   = ws.cell(*em['title_cell']).value
            raw_joining = ws.cell(*em['joining_cell']).value

            name = _to_str(raw_name)
            if not name or name.lower() in ('name', 'employee', ''):
                skipped += 1
                continue

            emp_code    = _clean_emp_code(raw_emp_no)
            job_title   = _clean_title(raw_emp_no, raw_title)
            department  = _to_str(raw_dept) or None
            joining_date = raw_joining.date() if hasattr(raw_joining, 'date') else None

            tr  = em['total_row']
            cfr = em['carryforward_row']
            total_earned   = _to_dec(ws.cell(tr, em['col_earned']).value)
            total_taken    = _to_dec(ws.cell(tr, em['col_taken']).value)
            total_encashed = _to_dec(ws.cell(tr, em['col_encashed']).value)
            leave_balance  = _to_dec(ws.cell(tr, em['col_balance']).value)
            carryforward   = _to_dec(ws.cell(cfr, em['col_balance']).value)

            try:
                with transaction.atomic():
                    defaults = dict(
                        employee_name=name, department=department,
                        job_title=job_title or None, joining_date=joining_date,
                        year=year, branch=branch,
                        total_earned=total_earned, total_taken=total_taken,
                        total_encashed=total_encashed, leave_balance=leave_balance,
                        carryforward=carryforward, source_file=src,
                    )
                    if emp_code:
                        rec, was_created = EmployeeLeaveRecord.objects.update_or_create(
                            employee_code=emp_code, year=year, defaults=defaults,
                        )
                    else:
                        rec, was_created = EmployeeLeaveRecord.objects.update_or_create(
                            employee_name=name, year=year,
                            defaults={**defaults, 'employee_code': None},
                        )
                    if was_created:
                        created += 1
                    else:
                        updated += 1

                    for row_idx in range(em['month_start_row'], em['month_end_row'] + 1):
                        month_label = _to_str(ws.cell(row_idx, em['col_month']).value).lower()
                        month_num   = mmap.get(month_label)
                        if month_num is None:
                            continue
                        EmployeeLeaveMonthly.objects.update_or_create(
                            record=rec, month=month_num,
                            defaults=dict(
                                earned   = _to_dec(ws.cell(row_idx, em['col_earned']).value),
                                taken    = _to_dec(ws.cell(row_idx, em['col_taken']).value),
                                encashed = _to_dec(ws.cell(row_idx, em['col_encashed']).value),
                                balance  = _to_dec(ws.cell(row_idx, em['col_balance']).value),
                            ),
                        )
            except Exception as exc:
                import_errors += 1

        # Now recompute accruals for all imported records
        recs = (
            EmployeeLeaveRecord.objects
            .prefetch_related('monthly_breakdown')
            .filter(year=year, branch__iexact=branch)
        )
        computed = computed_errors = 0
        for rec in recs:
            try:
                compute_accrual_for_record(rec, target_year=year, dry_run=False)
                computed += 1
            except Exception:
                computed_errors += 1

    finally:
        os.unlink(tmp.name)

    return Response({
        'year':            year,
        'branch':          branch,
        'sheets_found':    len(sheets),
        'created':         created,
        'updated':         updated,
        'skipped':         skipped,
        'import_errors':   import_errors,
        'computed':        computed,
        'compute_errors':  computed_errors,
    })


# =============================================================================
# DailyWorkLog ViewSet
# =============================================================================
from .models import DailyWorkLog, DailyWorkLogStatus, DailyWorkLogApprovalStatus  # noqa: E402
from .serializers import DailyWorkLogSerializer  # noqa: E402


class DailyWorkLogViewSet(viewsets.ModelViewSet):
    """
    CRUD for personal daily work logs.

    Default scope: the requesting user''s own logs.
    Staff overrides:
      * ?all=true          -> all users'' logs (staff only)
      * ?user_id=<uuid>    -> specific user''s logs (staff only)

    Date filters (compatible with above scopes):
      * ?date=YYYY-MM-DD
      * ?from=YYYY-MM-DD&to=YYYY-MM-DD
      * ?status=in_progress|done|blocked|deferred

    Custom actions:
      GET  /daily-logs/summary/        -> daily totals (hours + task count) per date
      GET  /daily-logs/export-to-s3/   -> export filtered logs as JSON to S3
      GET  /daily-logs/team/           -> latest log per user for a date (staff only)
    """
    permission_classes = [IsAuthenticated]
    serializer_class   = DailyWorkLogSerializer
    http_method_names  = ['get', 'post', 'patch', 'delete', 'head', 'options']

    def get_queryset(self):
        user   = self.request.user
        params = self.request.query_params
        qs     = DailyWorkLog.objects.select_related('user')

        # Scope
        if user.is_staff and params.get('all') == 'true':
            pass  # no user filter — all logs
        elif user.is_staff and params.get('user_id'):
            qs = qs.filter(user_id=params.get('user_id'))
        else:
            qs = qs.filter(user=user)

        # Date filters
        exact_date = params.get('date')
        from_date  = params.get('from')
        to_date    = params.get('to')
        if exact_date:
            qs = qs.filter(log_date=exact_date)
        else:
            if from_date:
                qs = qs.filter(log_date__gte=from_date)
            if to_date:
                qs = qs.filter(log_date__lte=to_date)

        # Status filter
        status_filter = params.get('status')
        if status_filter:
            qs = qs.filter(status=status_filter)

        # Approval status filter
        approval_filter = params.get('approval_status')
        if approval_filter:
            qs = qs.filter(approval_status=approval_filter)

        return qs.order_by('-log_date', '-created_at')

    def perform_create(self, serializer):
        serializer.save(user=self.request.user)

    # ── Permission helper ──────────────────────────────────────────────────
    @staticmethod
    def _can_approve(request_user, log_obj):
        """
        Returns True if request_user may approve/reject log_obj.
        Rules (OR):
          1. request_user.is_staff or is_superuser
          2. request_user is the direct manager of log_obj.user
             (UserProfile.manager FK points to a UserProfile; check if that
              UserProfile.user == request_user)
        """
        if request_user.is_staff or request_user.is_superuser:
            return True
        try:
            from apps.rbac.models import UserProfile  # noqa: F811
            return UserProfile.objects.filter(
                user=log_obj.user,
                manager__user=request_user,
                is_deleted=False,
            ).exists()
        except Exception:
            return False

    @action(detail=False, methods=['get'], url_path='summary')
    def summary(self, request):
        """Return daily task and hour totals for the current filtered scope."""
        from django.db.models import Count, Sum  # noqa: F811

        rows = (
            self.get_queryset()
            .values('log_date')
            .annotate(total_hours=Sum('hours_spent'), task_count=Count('id'))
            .order_by('log_date')
        )
        return Response([
            {
                'date': str(row['log_date']),
                'total_hours': float(row['total_hours'] or 0),
                'task_count': row['task_count'],
            }
            for row in rows
        ])

    @staticmethod
    def _notify_review_result(log, reviewer, approved, note):
        try:
            from django.conf import settings as email_settings
            from django.core.mail import send_mail

            role_label = {
                'project_manager': 'Project Manager',
                'reporting_manager': 'Reporting Manager',
            }.get(log.submitted_to_role, 'Manager')
            outcome = 'approved' if approved else 'requires revision'
            detail = (
                f'Your activity "{log.task_title}" on {log.log_date} '
                f'({log.hours_spent} hrs) {"has been approved" if approved else "was not approved"} '
                f'by your {role_label}.'
            )
            if note:
                detail += f'\n\nNote: {note}'
            send_mail(
                subject=f'[RAD AI] Your activity {outcome}',
                message=(
                    f'Hi {log.user.first_name or log.user.email},\n\n{detail}\n\n'
                    f'Reviewed by: {reviewer.get_full_name() or reviewer.email}\n'
                ),
                from_email=getattr(email_settings, 'DEFAULT_FROM_EMAIL', None),
                recipient_list=[log.user.email],
                fail_silently=True,
            )
        except Exception:
            logger.exception('Daily work log review notification failed for log %s', log.pk)

    @action(detail=True, methods=['post'], url_path='approve')
    def approve(self, request, pk=None):
        log = self.get_object()
        if not self._can_approve(request.user, log):
            return Response(
                {'error': 'You do not have permission to approve this activity.'},
                status=status.HTTP_403_FORBIDDEN,
            )
        if log.approval_status == DailyWorkLogApprovalStatus.APPROVED:
            return Response(
                {'error': 'Activity is already approved.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        note = request.data.get('note', '')
        log.approval_status = DailyWorkLogApprovalStatus.APPROVED
        log.approved_by = request.user
        log.approved_at = timezone.now()
        log.approval_note = note
        log.save(update_fields=['approval_status', 'approved_by', 'approved_at', 'approval_note'])

        try:
            from apps.finance.salary_models import SalarySlip

            slip = SalarySlip.objects.filter(
                month=log.log_date.month,
                year=log.log_date.year,
                employee_salary_info__user=log.user,
            ).first()
            if slip:
                ProjectCostAllocation.objects.create(
                    salary_slip=slip,
                    project_code=log.project_category or 'DAILY-ACTIVITY',
                    project_name=log.project_category or 'Daily Activity',
                    allocated_hours=log.hours_spent,
                    allocation_percent=Decimal('0'),
                    allocated_cost=Decimal('0'),
                    month=log.log_date.month,
                    year=log.log_date.year,
                )
        except Exception:
            logger.exception('Project cost allocation failed for approved daily log %s', log.pk)

        self._notify_review_result(log, request.user, True, note)
        return Response(DailyWorkLogSerializer(log).data)

    @action(detail=True, methods=['post'], url_path='reject')
    def reject(self, request, pk=None):
        log = self.get_object()
        if not self._can_approve(request.user, log):
            return Response(
                {'error': 'You do not have permission to reject this activity.'},
                status=status.HTTP_403_FORBIDDEN,
            )
        if log.approval_status == DailyWorkLogApprovalStatus.REJECTED:
            return Response(
                {'error': 'Activity is already rejected.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        note = request.data.get('note', '')
        log.approval_status = DailyWorkLogApprovalStatus.REJECTED
        log.approved_by = request.user
        log.approved_at = timezone.now()
        log.approval_note = note
        log.save(update_fields=['approval_status', 'approved_by', 'approved_at', 'approval_note'])
        self._notify_review_result(log, request.user, False, note)
        return Response(DailyWorkLogSerializer(log).data)

    @action(detail=False, methods=['get'], url_path='export-to-s3')
    def export_to_s3(self, request):
        from django.conf import settings as storage_settings

        bucket = getattr(storage_settings, 'AWS_STORAGE_BUCKET_NAME', '')
        if not bucket:
            return Response({'error': 'S3 storage is not configured on this environment.'}, status=503)

        import boto3
        import json

        logs = self.get_queryset()
        data = DailyWorkLogSerializer(logs, many=True).data
        now = datetime.datetime.utcnow()
        key = (
            f'daily-tracker/{request.user.id}/{now.year:04d}/{now.month:02d}/'
            f'export_{now.strftime("%Y%m%d_%H%M%S")}.json'
        )
        s3 = boto3.client(
            's3',
            region_name=getattr(storage_settings, 'AWS_S3_REGION_NAME', 'us-east-1'),
            endpoint_url=getattr(storage_settings, 'AWS_S3_ENDPOINT_URL', None),
        )
        s3.put_object(
            Bucket=bucket,
            Key=key,
            Body=json.dumps(list(data), default=str, ensure_ascii=False).encode('utf-8'),
            ContentType='application/json',
        )
        logs.update(s3_export_key=key)
        url = s3.generate_presigned_url(
            'get_object',
            Params={'Bucket': bucket, 'Key': key},
            ExpiresIn=3600,
        )
        return Response({'s3_key': key, 'url': url, 'count': len(data)})

    @action(detail=False, methods=['get'], url_path='team')
    def team(self, request):
        if not request.user.is_staff:
            return Response({'error': 'Staff access required.'}, status=403)

        from django.db.models import Max  # noqa: F811

        selected_date = request.query_params.get('date', str(datetime.date.today()))
        latest = (
            DailyWorkLog.objects.filter(log_date=selected_date)
            .values('user')
            .annotate(latest=Max('created_at'))
            .values_list('latest', flat=True)
        )
        logs = (
            DailyWorkLog.objects
            .filter(log_date=selected_date, created_at__in=latest)
            .select_related('user')
            .order_by('user__first_name', 'user__last_name')
        )
        return Response(DailyWorkLogSerializer(logs, many=True).data)


# =============================================================================
# Master Payroll Generator — Sympa + ValueFrame + RADAI attendance merge
# =============================================================================
#
# Soft-coded field alias maps: add new column synonyms here without changing
# any view logic.  All comparisons are case-insensitive and whitespace-trimmed.
#
# POST /api/v1/payroll/generate-master-payroll/
#   multipart fields:  sympa_file (opt), valueframe_file (opt), year, month
#   query param:       ?format=xlsx  →  returns binary Excel instead of JSON
#
# =============================================================================

# ── Alias tables (first match wins) ──────────────────────────────────────────
_SYMPA_ALIASES = {
    'employee_code':       ['employee no', 'employee id', 'emp no', 'emp id',
                            'personnel no', 'staff id', 'empno', 'id no', 'employee number'],
    'employee_name':       ['name', 'full name', 'employee name', 'emp name',
                            'staff name', 'employee full name'],
    'department':          ['department', 'dept', 'division', 'business unit',
                            'cost centre', 'cost center'],
    'job_title':           ['job title', 'position', 'title', 'designation', 'role'],
    'joining_date':        ['joining date', 'join date', 'hire date', 'date of joining',
                            'doj', 'start date', 'employment date', 'commencement date',
                            'employment start date'],
    'basic_salary':        ['basic salary', 'basic', 'base salary', 'monthly salary',
                            'salary', 'basic pay'],
    'housing_allowance':   ['housing allowance', 'house allowance', 'hra',
                            'housing', 'home allowance', 'accommodation allowance'],
    'transport_allowance': ['transport allowance', 'transport', 'ta',
                            'travel allowance', 'commute allowance', 'transportation'],
    'other_allowances':    ['other allowances', 'misc allowances', 'miscellaneous',
                            'additional allowances'],
    'other_pay':           ['other pay', 'other payment', 'extra pay', 'additional pay',
                            'other compensation', 'additional compensation', 'bonus pay',
                            'other emoluments'],
    'deductions':          ['deductions', 'total deductions', 'deduction', 'monthly deduction',
                            'salary deduction'],
    'deduction_details':   ['deduction details', 'deduction remarks', 'deduction notes',
                            'deduction breakdown', 'salary deduction details',
                            'deduction description'],
    'details':             ['details', 'notes', 'remarks', 'additional details',
                            'employee notes', 'comments', 'employee remarks'],
    'leave_balance':       ['annual leave balance', 'leave balance', 'remaining leave',
                            'al balance', 'leave days remaining'],
}

_VF_ALIASES = {
    'employee_code':  ['employee no', 'employee id', 'emp no', 'resource id',
                       'staff id', 'personnel no', 'resource code'],
    'employee_name':  ['name', 'full name', 'employee name', 'resource', 'resource name'],
    'project_code':   ['project code', 'project no', 'project', 'project id',
                       'proj code', 'project number'],
    'project_name':   ['project name', 'proj name', 'project description', 'project title'],
    'total_hours':    ['hours', 'total hours', 'billed hours', 'worked hours',
                       'billable hours', 'actual hours', 'logged hours'],
    'overtime_hours': ['overtime hours', 'ot hours', 'extra hours', 'overtime', 'ot'],
    'month':          ['month', 'period month', 'billing month', 'period'],
    'year':           ['year', 'period year', 'billing year'],
}


def _detect_columns(df_columns, alias_map):
    """Return {canonical_field: actual_df_column} by matching headers to aliases."""
    cols_lower = {c.strip().lower(): c for c in df_columns}
    mapping = {}
    for field, aliases in alias_map.items():
        for alias in aliases:
            if alias in cols_lower:
                mapping[field] = cols_lower[alias]
                break
    return mapping


def _parse_file(upload):
    """Parse an uploaded file (XLSX, XLS, or CSV) into a pandas DataFrame."""
    import pandas as pd
    import tempfile, os

    suffix = (os.path.splitext(upload.name)[1] or '').lower() or '.xlsx'
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=suffix)
    try:
        for chunk in upload.chunks():
            tmp.write(chunk)
        tmp.close()
        if suffix == '.csv':
            df = pd.read_csv(tmp.name, dtype=str).fillna('')
        else:
            df = pd.read_excel(tmp.name, dtype=str).fillna('')
    finally:
        try:
            os.unlink(tmp.name)
        except Exception:
            pass
    # Normalise column names: strip whitespace
    df.columns = [str(c).strip() for c in df.columns]
    return df


def _safe_dec(v):
    """Convert a string value to Decimal, returning 0 on failure."""
    try:
        return Decimal(str(v).replace(',', '').strip())
    except Exception:
        return Decimal('0')


def _safe_float(v):
    """Convert a string value to float, returning 0 on failure."""
    try:
        return float(str(v).replace(',', '').strip())
    except Exception:
        return 0.0


def _norm_code(v):
    """Normalise an employee code for matching: lowercase, stripped."""
    return str(v or '').strip().lower()


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def generate_master_payroll(request):
    """
    POST /api/v1/payroll/generate-master-payroll/

    Accepts optional Sympa and ValueFrame file uploads.
    Merges with RADAI attendance (EmployeeLeaveMonthly + DailyWorkLog) for
    the requested period and RADAI salary structures.

    Returns JSON list of master payroll rows or, when ?format=xlsx, an
    Excel binary response.

    Request multipart fields:
      sympa_file       — Sympa HR export  (XLSX / XLS / CSV)  — optional
      valueframe_file  — ValueFrame hours (XLSX / XLS / CSV)  — optional
      year             — int  (defaults to current year)
      month            — int  (defaults to current month)
    """
    import pandas as pd
    from django.http import HttpResponse

    if not _is_hr_manager(request.user):
        from rest_framework.exceptions import PermissionDenied
        raise PermissionDenied('HR Manager role required.')

    now   = timezone.now()
    year  = int(request.data.get('year',  now.year))
    month = int(request.data.get('month', now.month))
    fmt   = request.query_params.get('format', 'json').lower()

    master  = {}   # employee_code → master row dict
    stats   = {'sympa_rows': 0, 'vf_employees': 0, 'radai_rows': 0, 'matched': 0}
    warnings_list = []

    # ── 1. Parse Sympa file ───────────────────────────────────────────────────
    sympa_file = request.FILES.get('sympa_file')
    if sympa_file:
        try:
            df = _parse_file(sympa_file)
            col_map = _detect_columns(df.columns, _SYMPA_ALIASES)
            if 'employee_code' not in col_map:
                warnings_list.append('Sympa: could not detect employee code column.')
            else:
                for _, row in df.iterrows():
                    code = _norm_code(row.get(col_map['employee_code'], ''))
                    if not code:
                        continue
                    basic     = _safe_dec(row.get(col_map.get('basic_salary', ''), 0))
                    housing   = _safe_dec(row.get(col_map.get('housing_allowance', ''), 0))
                    transport = _safe_dec(row.get(col_map.get('transport_allowance', ''), 0))
                    other     = _safe_dec(row.get(col_map.get('other_allowances', ''), 0))
                    other_pay = _safe_dec(row.get(col_map.get('other_pay', ''), 0))
                    deduct    = _safe_dec(row.get(col_map.get('deductions', ''), 0))
                    total_allow = housing + transport + other
                    master[code] = {
                        'employee_code':       code,
                        'employee_name':       str(row.get(col_map.get('employee_name', ''), '')).strip() or code,
                        'department':          str(row.get(col_map.get('department', ''), '')).strip(),
                        'job_title':           str(row.get(col_map.get('job_title', ''), '')).strip(),
                        'joining_date':        str(row.get(col_map.get('joining_date', ''), '')).strip(),
                        'basic_salary':        str(basic),
                        'housing_allowance':   str(housing),
                        'transport_allowance': str(transport),
                        'other_allowances':    str(other),
                        'other_pay':           str(other_pay),
                        'total_allowances':    str(total_allow),
                        'total_deductions':    str(deduct),
                        'deduction_details':   str(row.get(col_map.get('deduction_details', ''), '')).strip(),
                        'details':             str(row.get(col_map.get('details', ''), '')).strip(),
                        'leave_balance':       str(_safe_dec(row.get(col_map.get('leave_balance', ''), 0))),
                        # ValueFrame + RADAI fields — filled in later
                        'total_hours':         '0',
                        'overtime_hours':      '0',
                        'project_breakdown':   [],
                        'days_present':        None,
                        'days_absent':         None,
                        'sources':             ['sympa'],
                        'warnings':            [],
                    }
                stats['sympa_rows'] = len(master)
        except Exception as e:
            logger.warning(f'generate_master_payroll: Sympa parse error: {e}')
            warnings_list.append(f'Sympa file error: {e}')

    # ── 2. Parse ValueFrame file ──────────────────────────────────────────────
    vf_file = request.FILES.get('valueframe_file')
    if vf_file:
        try:
            df = _parse_file(vf_file)
            col_map = _detect_columns(df.columns, _VF_ALIASES)
            if 'employee_code' not in col_map:
                warnings_list.append('ValueFrame: could not detect employee code column.')
            else:
                vf_emp_hours  = {}   # employee_code → total hours
                vf_emp_ot     = {}   # employee_code → overtime hours
                vf_emp_proj   = {}   # employee_code → [{project_code, project_name, hours}]
                vf_emp_names  = {}   # employee_code → name
                for _, row in df.iterrows():
                    code = _norm_code(row.get(col_map['employee_code'], ''))
                    if not code:
                        continue
                    # Filter by month/year if those columns are present
                    if 'month' in col_map and 'year' in col_map:
                        row_month = _safe_float(row.get(col_map['month'], 0))
                        row_year  = _safe_float(row.get(col_map['year'], 0))
                        if row_month and row_year:
                            if int(row_month) != month or int(row_year) != year:
                                continue
                    hours = _safe_float(row.get(col_map.get('total_hours', ''), 0))
                    ot    = _safe_float(row.get(col_map.get('overtime_hours', ''), 0))
                    proj_code = str(row.get(col_map.get('project_code', ''), '')).strip()
                    proj_name = str(row.get(col_map.get('project_name', ''), '')).strip()
                    vf_emp_hours[code] = vf_emp_hours.get(code, 0) + hours
                    vf_emp_ot[code]    = vf_emp_ot.get(code, 0) + ot
                    if proj_code:
                        vf_emp_proj.setdefault(code, []).append({
                            'project_code': proj_code,
                            'project_name': proj_name,
                            'hours':        round(hours, 2),
                        })
                    if 'employee_name' in col_map:
                        vf_emp_names[code] = str(row.get(col_map['employee_name'], '')).strip()
                # Merge into master
                for code in set(list(vf_emp_hours.keys())):
                    if code not in master:
                        master[code] = {
                            'employee_code':       code,
                            'employee_name':       vf_emp_names.get(code, code),
                            'department':          '',
                            'job_title':           '',
                            'joining_date':        '',
                            'basic_salary':        '0',
                            'housing_allowance':   '0',
                            'transport_allowance': '0',
                            'other_allowances':    '0',
                            'other_pay':           '0',
                            'total_allowances':    '0',
                            'total_deductions':    '0',
                            'deduction_details':   '',
                            'details':             '',
                            'leave_balance':       '0',
                            'days_present':        None,
                            'days_absent':         None,
                            'project_breakdown':   [],
                            'warnings':            ['No Sympa record found for this employee.'],
                            'sources':             [],
                        }
                    master[code]['total_hours']      = str(round(vf_emp_hours.get(code, 0), 2))
                    master[code]['overtime_hours']   = str(round(vf_emp_ot.get(code, 0), 2))
                    master[code]['project_breakdown']= vf_emp_proj.get(code, [])
                    if 'valueframe' not in master[code]['sources']:
                        master[code]['sources'].append('valueframe')
                stats['vf_employees'] = len(vf_emp_hours)
        except Exception as e:
            logger.warning(f'generate_master_payroll: ValueFrame parse error: {e}')
            warnings_list.append(f'ValueFrame file error: {e}')

    # ── 3. Merge RADAI attendance (EmployeeLeaveMonthly + DailyWorkLog) ───────
    try:
        # DailyWorkLog: aggregate approved days/hours per employee for the period
        from django.db.models import Count as DjCount
        log_agg = (
            DailyWorkLog.objects
            .filter(date__year=year, date__month=month, status='approved')
            .values('user__rbac_profile__employee_id')
            .annotate(days=DjCount('id'), hours=Sum('hours'))
        )
        for row in log_agg:
            raw_code = row.get('user__rbac_profile__employee_id') or ''
            code = _norm_code(raw_code)
            if not code:
                continue
            days  = row.get('days', 0) or 0
            if code in master:
                master[code]['days_present'] = days
                master[code]['days_absent']  = max(0, 22 - days)   # soft-coded working days default
                if 'radai' not in master[code]['sources']:
                    master[code]['sources'].append('radai')
            stats['radai_rows'] = stats.get('radai_rows', 0) + 1
    except Exception as e:
        logger.warning(f'generate_master_payroll: RADAI attendance query error: {e}')
        warnings_list.append(f'RADAI attendance partial: {e}')

    # ── 4. Overlay with existing RADAI salary structures ─────────────────────
    try:
        from .models import EmployeeSalaryStructure, SalaryStructureStatus
        active_structs = EmployeeSalaryStructure.objects.filter(
            is_active=True, status=SalaryStructureStatus.APPROVED,
        ).values('employee_code', 'employee_name', 'department', 'basic_salary',
                 'net_salary', 'total_gross', 'total_deductions')
        for s in active_structs:
            code = _norm_code(s['employee_code'])
            if code in master:
                # Prefer Sympa values if already set; use RADAI as fallback
                if master[code]['basic_salary'] == '0':
                    master[code]['basic_salary']   = str(s['basic_salary'] or 0)
                    master[code]['total_deductions']= str(s['total_deductions'] or 0)
                if not master[code]['department'] and s['department']:
                    master[code]['department'] = s['department']
                if not master[code]['employee_name'] or master[code]['employee_name'] == code:
                    master[code]['employee_name'] = s['employee_name'] or code
                if 'radai' not in master[code]['sources']:
                    master[code]['sources'].append('radai')
    except Exception as e:
        logger.warning(f'generate_master_payroll: salary structure overlay error: {e}')

    # ── 5. Build final rows ───────────────────────────────────────────────────
    rows = []
    for code, r in master.items():
        # Compute estimated gross / net including other_pay
        basic      = _safe_dec(r.get('basic_salary', 0))
        allow      = _safe_dec(r.get('total_allowances', 0))
        other_pay  = _safe_dec(r.get('other_pay', 0))
        deduct     = _safe_dec(r.get('total_deductions', 0))
        gross      = basic + allow + other_pay      # employee_salary
        final_sal  = max(Decimal('0'), gross - deduct)
        r['employee_salary'] = str(gross)
        r['final_salary']    = str(final_sal)
        # Legacy aliases kept for backward-compat JSON consumers
        r['gross_salary']    = str(gross)
        r['net_salary_est']  = str(final_sal)
        rows.append(r)
        if r['sources']:
            stats['matched'] += 1

    rows.sort(key=lambda x: (x.get('department') or '', x.get('employee_name') or ''))

    # ── 6. Persist to DB + trigger async S3 upload ────────────────────────────
    import_session = None
    try:
        from .models import MasterPayrollImport, MasterPayrollRow, MasterPayrollImportStatus
        from .tasks import upload_master_payroll_to_s3

        sympa_fn = request.FILES['sympa_file'].name      if 'sympa_file'      in request.FILES else ''
        vf_fn    = request.FILES['valueframe_file'].name if 'valueframe_file' in request.FILES else ''

        import_session = MasterPayrollImport.objects.create(
            year=year,
            month=month,
            generated_by=request.user if request.user.is_authenticated else None,
            sympa_filename=sympa_fn,
            valueframe_filename=vf_fn,
            stats=stats,
            warnings=warnings_list,
            total_rows=len(rows),
            status=MasterPayrollImportStatus.PROCESSING,
        )

        # Bulk-create rows (ignore conflicts — idempotent on re-generation)
        row_objs = [
            MasterPayrollRow(
                import_session   = import_session,
                employee_code    = r.get('employee_code', ''),
                employee_name    = r.get('employee_name', ''),
                joining_date     = r.get('joining_date', '') or '',
                total_hours      = Decimal(str(r.get('total_hours', 0) or 0)),
                employee_salary  = Decimal(str(r.get('employee_salary', 0) or 0)),
                basic_salary     = Decimal(str(r.get('basic_salary', 0) or 0)),
                total_allowances     = Decimal(str(r.get('total_allowances', 0) or 0)),
                transport_allowance  = Decimal(str(r.get('transport_allowance', 0) or 0)),
                housing_allowance    = Decimal(str(r.get('housing_allowance', 0) or 0)),
                other_allowances     = Decimal(str(r.get('other_allowances', 0) or 0)),
                other_pay            = Decimal(str(r.get('other_pay', 0) or 0)),
                details              = r.get('details', '') or '',
                total_deductions     = Decimal(str(r.get('total_deductions', 0) or 0)),
                deduction_details    = r.get('deduction_details', '') or '',
                final_salary         = Decimal(str(r.get('final_salary', 0) or 0)),
                sources              = r.get('sources', []),
                row_warnings         = r.get('warnings', []),
                raw_data             = {k: v for k, v in r.items()
                                        if k not in ('sources', 'warnings', 'project_breakdown')},
            )
            for r in rows
        ]
        MasterPayrollRow.objects.bulk_create(row_objs, ignore_conflicts=True)

        # Mark ready now; Celery will flip to 'uploaded' once S3 upload completes
        import_session.status = MasterPayrollImportStatus.READY
        import_session.save(update_fields=['status'])

        # Fire async S3 upload — non-blocking
        upload_master_payroll_to_s3.delay(str(import_session.id))

        logger.info(
            'generate_master_payroll: saved import %s (%d rows), S3 upload queued',
            import_session.id, len(row_objs),
        )
    except Exception as persist_err:
        logger.warning('generate_master_payroll: DB/S3 persist failed: %s', persist_err)
        # Non-fatal — still return the data to the user even if persistence fails

    # ── 7. Return response ────────────────────────────────────────────────────
    if fmt == 'xlsx':
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = f'Payroll Master {year}-{month:02d}'
            hdr_font = Font(bold=True, color='FFFFFF')
            hdr_fill = PatternFill('solid', fgColor='2563EB')
            thin = Side(style='thin', color='CCCCCC')
            border = Border(left=thin, right=thin, top=thin, bottom=thin)
            headers = [
                'Employee Code',        # 1
                'Employee Name',        # 2
                'Joining Date',         # 3
                'No. of Working Hours', # 4
                'Employee Salary',      # 5  (gross = basic + allow + other pay)
                'Basic',                # 6
                'Allowance',            # 7  (total allow)
                'Transportation',       # 8
                'Home Allowance',       # 9
                'Other Allowance',      # 10
                'Other Pay',            # 11
                'Details',              # 12
                'Salary Deduction',     # 13
                'Deduction Details',    # 14
                'Final Salary',         # 15  (gross – deductions)
            ]
            for col_idx, hdr in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=hdr)
                cell.font = hdr_font
                cell.fill = hdr_fill
                cell.alignment = Alignment(horizontal='center')
                cell.border = border
                ws.column_dimensions[cell.column_letter].width = max(14, len(hdr) + 4)
            for r_idx, r in enumerate(rows, 2):
                vals = [
                    r.get('employee_code', ''),                           # 1
                    r.get('employee_name', ''),                           # 2
                    r.get('joining_date', ''),                            # 3
                    float(r.get('total_hours') or 0),                    # 4
                    float(r.get('employee_salary') or 0),                # 5
                    float(r.get('basic_salary') or 0),                   # 6
                    float(r.get('total_allowances') or 0),               # 7
                    float(r.get('transport_allowance') or 0),            # 8
                    float(r.get('housing_allowance') or 0),              # 9
                    float(r.get('other_allowances') or 0),               # 10
                    float(r.get('other_pay') or 0),                      # 11
                    r.get('details') or r.get('job_title', ''),          # 12
                    float(r.get('total_deductions') or 0),               # 13
                    r.get('deduction_details', ''),                       # 14
                    float(r.get('final_salary') or 0),                   # 15
                ]
                for col_idx, val in enumerate(vals, 1):
                    cell = ws.cell(row=r_idx, column=col_idx, value=val)
                    cell.border = border
                    if isinstance(val, float):
                        cell.number_format = '#,##0.00'
            # Freeze header row
            ws.freeze_panes = 'A2'
            import io
            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            filename = f'master_payroll_{year}_{month:02d}.xlsx'
            response = HttpResponse(
                buf.read(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            )
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response
        except Exception as e:
            logger.error(f'generate_master_payroll: Excel generation error: {e}')
            return Response({'error': f'Excel generation failed: {e}'}, status=500)

    return Response({
        'year':         year,
        'month':        month,
        'generated_at': timezone.now().isoformat(),
        'import_id':    str(import_session.id) if import_session else None,
        'rows':         rows,
        'stats':        stats,
        'warnings':     warnings_list,
    })


# ─────────────────────────────────────────────────────────────────────────────
# Master Payroll History — list past import sessions
# GET /api/v1/payroll/master-payroll-history/
# ─────────────────────────────────────────────────────────────────────────────

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def master_payroll_history(request):
    """
    Returns paginated list of past MasterPayrollImport sessions.
    HR managers see all; regular users only see their own.

    Query params:
      year, month  — filter by period
      page         — 1-based page number (default 1)
      page_size    — results per page (default 20, max 100)
    """
    from .models import MasterPayrollImport

    PAGE_SIZE_DEFAULT = 20
    PAGE_SIZE_MAX     = 100

    qs = MasterPayrollImport.objects.select_related('generated_by').order_by('-year', '-month', '-generated_at')

    if not _is_hr_manager(request.user):
        qs = qs.filter(generated_by=request.user)

    year  = request.query_params.get('year')
    month = request.query_params.get('month')
    if year:
        qs = qs.filter(year=int(year))
    if month:
        qs = qs.filter(month=int(month))

    try:
        page_size = min(int(request.query_params.get('page_size', PAGE_SIZE_DEFAULT)), PAGE_SIZE_MAX)
        page      = max(int(request.query_params.get('page', 1)), 1)
    except (ValueError, TypeError):
        page_size, page = PAGE_SIZE_DEFAULT, 1

    total  = qs.count()
    offset = (page - 1) * page_size
    items  = qs[offset: offset + page_size]

    def _serialize(imp):
        return {
            'id':                   str(imp.id),
            'year':                 imp.year,
            'month':                imp.month,
            'generated_at':         imp.generated_at.isoformat(),
            'generated_by':         imp.generated_by.get_full_name() if imp.generated_by else None,
            'sympa_filename':       imp.sympa_filename,
            'valueframe_filename':  imp.valueframe_filename,
            'status':               imp.status,
            'total_rows':           imp.total_rows,
            'stats':                imp.stats,
            'has_s3':               bool(imp.s3_key),
        }

    return Response({
        'count':     total,
        'page':      page,
        'page_size': page_size,
        'results':   [_serialize(i) for i in items],
    })


# ─────────────────────────────────────────────────────────────────────────────
# Master Payroll Download — presigned S3 URL or on-the-fly Excel
# GET /api/v1/payroll/master-payroll-history/<import_id>/download/
# ─────────────────────────────────────────────────────────────────────────────

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def master_payroll_download(request, import_id):
    """
    Returns a presigned S3 URL for the stored Excel, or regenerates the
    Excel on-the-fly from DB rows if the S3 upload is still pending.
    HR managers only.
    """
    from .models import MasterPayrollImport
    from apps.payroll.storage import PayrollExportStorage, S3_AVAILABLE

    if not _is_hr_manager(request.user):
        from rest_framework.exceptions import PermissionDenied
        raise PermissionDenied('HR Manager role required.')

    try:
        session = MasterPayrollImport.objects.get(id=import_id)
    except MasterPayrollImport.DoesNotExist:
        return Response({'error': 'Import session not found.'}, status=404)

    # ── Case 1: S3 key exists → return presigned URL ──────────────────────────
    if session.s3_key and S3_AVAILABLE:
        try:
            storage = PayrollExportStorage()
            relative_key = session.s3_key.split(f'{storage.location}/', 1)[-1]
            url = storage.url(relative_key)
            return Response({
                'download_url': url,
                'source':       's3',
                'filename':     f'master_payroll_{session.year}_{session.month:02d}.xlsx',
            })
        except Exception as e:
            logger.warning('master_payroll_download: presigned URL failed: %s', e)
            # Fall through to on-the-fly generation

    # ── Case 2: Generate Excel on-the-fly from DB rows ─────────────────────────
    try:
        import io as _io
        import openpyxl
        from django.http import HttpResponse as DjangoHttpResponse
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f'Payroll Master {session.year}-{session.month:02d}'

        hdr_font = Font(bold=True, color='FFFFFF')
        hdr_fill = PatternFill('solid', fgColor='2563EB')
        thin     = Side(style='thin', color='CCCCCC')
        border   = Border(left=thin, right=thin, top=thin, bottom=thin)

        headers = [
            'Employee Code', 'Employee Name', 'Joining Date', 'No. of Working Hours',
            'Employee Salary', 'Basic', 'Allowance', 'Transportation', 'Home Allowance',
            'Other Allowance', 'Other Pay', 'Details', 'Salary Deduction',
            'Deduction Details', 'Final Salary',
        ]
        for col_idx, hdr in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=hdr)
            cell.font = hdr_font; cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal='center'); cell.border = border
            ws.column_dimensions[cell.column_letter].width = max(14, len(hdr) + 4)

        for r_idx, row in enumerate(session.rows.all().order_by('employee_name'), 2):
            vals = [
                row.employee_code,   row.employee_name,   row.joining_date or '',
                float(row.total_hours),       float(row.employee_salary),
                float(row.basic_salary),      float(row.total_allowances),
                float(row.transport_allowance), float(row.housing_allowance),
                float(row.other_allowances),  float(row.other_pay),
                row.details or '',
                float(row.total_deductions),  row.deduction_details or '',
                float(row.final_salary),
            ]
            for col_idx, val in enumerate(vals, 1):
                cell = ws.cell(row=r_idx, column=col_idx, value=val)
                cell.border = border
                if isinstance(val, float):
                    cell.number_format = '#,##0.00'

        ws.freeze_panes = 'A2'
        buf = _io.BytesIO()
        wb.save(buf); buf.seek(0)
        filename = f'master_payroll_{session.year}_{session.month:02d}.xlsx'
        response = DjangoHttpResponse(
            buf.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    except Exception as e:
        logger.error('master_payroll_download: on-the-fly generation failed: %s', e)
        return Response({'error': f'Download failed: {e}'}, status=500)


    # ── Approve ───────────────────────────────────────────────────────────
    @action(detail=True, methods=['post'], url_path='approve')
    def approve(self, request, pk=None):
        log = self.get_object()
        if not self._can_approve(request.user, log):
            return Response(
                {'error': 'You do not have permission to approve this activity.'},
                status=status.HTTP_403_FORBIDDEN,
            )
        if log.approval_status == DailyWorkLogApprovalStatus.APPROVED:
            return Response(
                {'error': 'Activity is already approved.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        note = request.data.get('note', '')
        log.approval_status = DailyWorkLogApprovalStatus.APPROVED
        log.approved_by     = request.user
        log.approved_at     = timezone.now()
        log.approval_note   = note
        log.save(update_fields=['approval_status', 'approved_by', 'approved_at', 'approval_note'])

        # ── Auto-create ProjectCostAllocation if a SalarySlip exists ───
        try:
            from apps.finance.salary_models import SalarySlip as _Slip  # noqa
            slip = _Slip.objects.filter(
                month=log.log_date.month,
                year=log.log_date.year,
            ).filter(
                # Match by user → EmployeeSalaryInfo → SalarySlip
                employee_salary_info__user=log.user,
            ).first()
            if slip:
                ProjectCostAllocation.objects.create(
                    salary_slip=slip,
                    project_code=log.project_category or 'DAILY-ACTIVITY',
                    project_name=log.project_category or 'Daily Activity',
                    allocated_hours=log.hours_spent,
                    allocation_percent=Decimal('0'),
                    allocated_cost=Decimal('0'),
                    month=log.log_date.month,
                    year=log.log_date.year,
                )
        except Exception:
            pass  # Never block the approval if cost allocation fails

        # ── Notify employee by email ───────────────────────────────────
        try:
            from django.core.mail import send_mail
            from django.conf import settings as _s
            role_label = (
                'Project Manager' if log.submitted_to_role == 'project_manager'
                else 'Reporting Manager' if log.submitted_to_role == 'reporting_manager'
                else 'Manager'
            )
            send_mail(
                subject=f'[RAD AI] Your activity log has been approved',
                message=(
                    f'Hi {log.user.first_name or log.user.email},\n\n'
                    f'Your activity "{log.task_title}" on {log.log_date} '
                    f'({log.hours_spent} hrs) has been approved by your {role_label}'
                    f'{" with note: " + note if note else "."}'
                    f'\n\nApproved by: {request.user.get_full_name() or request.user.email}\n'
                ),
                from_email=getattr(_s, 'DEFAULT_FROM_EMAIL', None),
                recipient_list=[log.user.email],
                fail_silently=True,
            )
        except Exception:
            pass

        return Response(DailyWorkLogSerializer(log).data)

    # ── Reject ────────────────────────────────────────────────────────────
    @action(detail=True, methods=['post'], url_path='reject')
    def reject(self, request, pk=None):
        log = self.get_object()
        if not self._can_approve(request.user, log):
            return Response(
                {'error': 'You do not have permission to reject this activity.'},
                status=status.HTTP_403_FORBIDDEN,
            )
        if log.approval_status == DailyWorkLogApprovalStatus.REJECTED:
            return Response(
                {'error': 'Activity is already rejected.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        note = request.data.get('note', '')
        log.approval_status = DailyWorkLogApprovalStatus.REJECTED
        log.approved_by     = request.user
        log.approved_at     = timezone.now()
        log.approval_note   = note
        log.save(update_fields=['approval_status', 'approved_by', 'approved_at', 'approval_note'])

        # ── Notify employee by email ───────────────────────────────────
        try:
            from django.core.mail import send_mail
            from django.conf import settings as _s
            role_label_r = (
                'Project Manager' if log.submitted_to_role == 'project_manager'
                else 'Reporting Manager' if log.submitted_to_role == 'reporting_manager'
                else 'Manager'
            )
            send_mail(
                subject=f'[RAD AI] Your activity log requires revision',
                message=(
                    f'Hi {log.user.first_name or log.user.email},\n\n'
                    f'Your activity "{log.task_title}" on {log.log_date} '
                    f'({log.hours_spent} hrs) was reviewed by your {role_label_r} and was not approved.\n\n'
                    f'Reason: {note or "No reason provided."}\n\n'
                    f'Please update the entry and resubmit for approval.\n\n'
                    f'Reviewed by: {request.user.get_full_name() or request.user.email}\n'
                ),
                from_email=getattr(_s, 'DEFAULT_FROM_EMAIL', None),
                recipient_list=[log.user.email],
                fail_silently=True,
            )
        except Exception:
            pass

        return Response(DailyWorkLogSerializer(log).data)

    # -- Summary: daily totals for heatmap + bar chart ---------------------
    @action(detail=False, methods=['get'], url_path='summary')
    def summary(self, request):
        qs = self.get_queryset()
        from django.db.models import Sum, Count  # noqa: F811
        rows = (
            qs
            .values('log_date')
            .annotate(total_hours=Sum('hours_spent'), task_count=Count('id'))
            .order_by('log_date')
        )
        data = [
            {
                'date':        str(r['log_date']),
                'total_hours': float(r['total_hours'] or 0),
                'task_count':  r['task_count'],
            }
            for r in rows
        ]
        return Response(data)

    # ── Export to S3 ───────────────────────────────────────────────────────
    @action(detail=False, methods=['get'], url_path='export-to-s3')
    def export_to_s3(self, request):
        from django.conf import settings as _settings  # noqa: F811
        bucket = getattr(_settings, 'AWS_STORAGE_BUCKET_NAME', '')
        if not bucket:
            return Response(
                {'error': 'S3 storage is not configured on this environment.'},
                status=503,
            )

        import boto3, json, tempfile  # noqa: E402
        from datetime import datetime as _dt  # noqa: F811

        qs   = self.get_queryset()
        data = DailyWorkLogSerializer(qs, many=True).data

        user       = request.user
        now        = _dt.utcnow()
        s3_key     = (
            f'daily-tracker/{user.id}/{now.year:04d}/{now.month:02d}/'
            f'export_{now.strftime("%Y%m%d_%H%M%S")}.json'
        )

        payload = json.dumps(list(data), default=str, ensure_ascii=False)

        region   = getattr(_settings, 'AWS_S3_REGION_NAME', 'us-east-1')
        endpoint = getattr(_settings, 'AWS_S3_ENDPOINT_URL', None)
        s3 = boto3.client(
            's3',
            region_name=region,
            endpoint_url=endpoint,
        )
        s3.put_object(
            Bucket=bucket,
            Key=s3_key,
            Body=payload.encode('utf-8'),
            ContentType='application/json',
        )

        # Mark exported logs with their S3 key
        qs.update(s3_export_key=s3_key)

        presigned_url = s3.generate_presigned_url(
            'get_object',
            Params={'Bucket': bucket, 'Key': s3_key},
            ExpiresIn=3600,
        )

        return Response({'s3_key': s3_key, 'url': presigned_url, 'count': len(data)})

    # ── Team view: latest entries per user for a given date ────────────────
    @action(detail=False, methods=['get'], url_path='team')
    def team(self, request):
        if not request.user.is_staff:
            return Response({'error': 'Staff access required.'}, status=403)

        params = request.query_params
        date   = params.get('date', str(datetime.date.today()))

        from django.db.models import Max  # noqa: F811
        # Subquery: most recent entry per user on target date
        latest = (
            DailyWorkLog.objects
            .filter(log_date=date)
            .values('user')
            .annotate(latest=Max('created_at'))
            .values_list('latest', flat=True)
        )
        logs = (
            DailyWorkLog.objects
            .filter(log_date=date, created_at__in=latest)
            .select_related('user')
            .order_by('user__first_name', 'user__last_name')
        )
        return Response(DailyWorkLogSerializer(logs, many=True).data)
