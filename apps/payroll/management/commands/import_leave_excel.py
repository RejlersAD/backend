"""
Import Leave Excel — Django Management Command
==============================================
Usage (run inside backend Docker container):

    python manage.py import_leave_excel
    python manage.py import_leave_excel --path /path/to/file.xlsx
    python manage.py import_leave_excel --year 2026 --dry-run

Reads every employee sheet from the HR leave Excel file and upserts
into payroll_employee_leave_record + payroll_employee_leave_monthly.

Soft-coded column/row positions via LEAVE_EXCEL_MAP at the top of this
file — no magic numbers in the parsing logic.
"""
from __future__ import annotations

import os
import warnings
from decimal import Decimal, InvalidOperation
from typing import Optional

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

# ─────────────────────────────────────────────────────────────────────────────
# SOFT-CODED CELL MAP  — change these if the Excel layout changes, nothing
# else in the command needs touching.
# ─────────────────────────────────────────────────────────────────────────────
LEAVE_EXCEL_MAP = {
    # Employee header cells  (row, col) — 1-indexed
    'name_cell':          (1, 2),    # B1
    'emp_no_cell':        (2, 2),    # B2
    'dept_cell':          (2, 4),    # D2
    'title_cell':         (3, 2),    # B3
    'joining_cell':       (1, 5),    # E1
    'entitlement_cell':   (6, 2),    # B6  ("22 working days")

    # Monthly summary table rows — Row 9 = header, Row 10 = carryforward
    # Rows 11-22 = Jan-Dec, Row 23 = Totals
    'carryforward_row':   10,
    'month_start_row':    11,        # Jan
    'month_end_row':      22,        # Dec
    'total_row':          23,

    # Column indices within the summary table
    'col_month':          1,         # A
    'col_earned':         2,         # B
    'col_taken':          3,         # C
    'col_encashed':       4,         # D
    'col_balance':        5,         # E

    # Month label to number mapping (handles abbreviated month names)
    'month_map': {
        'jan': 1, 'feb': 2, 'mar': 3, 'march': 3,
        'apr': 4, 'may': 5, 'jun': 6, 'june': 6,
        'jul': 7, 'aug': 8, 'sep': 9, 'oct': 10,
        'nov': 11, 'dec': 12,
    },

    # Source file tag stored in DB
    'source_tag': 'Summary Leave Calculation-RAD-Updated.xlsx',
}

# Default path used when --path is not supplied
DEFAULT_EXCEL_PATH = os.path.join(
    os.path.dirname(__file__),  # management/commands/
    '..', '..', '..', '..', '..',  # → project root
    'Documents', 'Human Resource', 'Payroll',
    'Summary Leave Calculation-RAD-Updated.xlsx',
)


def _to_dec(val, default: str = '0') -> Decimal:
    """Convert any cell value to Decimal safely."""
    if val is None:
        return Decimal(default)
    try:
        return Decimal(str(val)).quantize(Decimal('0.0001'))
    except (InvalidOperation, ValueError):
        return Decimal(default)


def _to_str(val) -> str:
    """Return string representation of a cell value, stripped."""
    if val is None:
        return ''
    return str(val).strip()


def _clean_emp_code(val) -> Optional[str]:
    """
    Normalise employee code:
    - None → None
    - Integer → string
    - String that is actually a job title (contains letters) → None
    Handles the known data-quality issue where some sheets have swapped
    emp_no and title cells.
    """
    if val is None:
        return None
    s = str(val).strip()
    if not s:
        return None
    # If it looks like a job title (has spaces or non-digit letters), skip
    digits = ''.join(c for c in s if c.isdigit())
    if len(digits) >= 4 and len(digits) / max(len(s), 1) >= 0.5:
        return digits
    return None


def _clean_title(emp_no_val, title_val) -> str:
    """
    Same swap issue: if title_val is numeric and emp_no_val is text,
    the cells are swapped — return the text value as title.
    """
    if title_val is None:
        # Maybe the title ended up in emp_no_val
        if emp_no_val and not str(emp_no_val).strip().isdigit():
            return _to_str(emp_no_val)
        return ''
    try:
        int(str(title_val).strip())
        # title_val is numeric → the real title is emp_no_val
        if emp_no_val and not str(emp_no_val).strip().isdigit():
            return _to_str(emp_no_val)
        return ''
    except (ValueError, TypeError):
        return _to_str(title_val)


# ─────────────────────────────────────────────────────────────────────────────
class Command(BaseCommand):
    help = 'Import employee leave data from HR leave Excel into PostgreSQL'

    def add_arguments(self, parser):
        parser.add_argument(
            '--path',
            default=None,
            help='Absolute path to the Excel file (default: Documents/Human Resource/Payroll/...)',
        )
        parser.add_argument(
            '--year',
            type=int,
            default=2026,
            help='Leave year to tag records with (default: 2026)',
        )
        parser.add_argument(
            '--dry-run',
            action='store_true',
            help='Parse and print without writing to the database',
        )

    def handle(self, *args, **options):
        try:
            import openpyxl
        except ImportError:
            raise CommandError('openpyxl is required: pip install openpyxl')

        # Lazy import to avoid circular imports at module load
        from apps.payroll.models import EmployeeLeaveRecord, EmployeeLeaveMonthly

        path = options['path'] or os.path.normpath(DEFAULT_EXCEL_PATH)
        year = options['year']
        dry  = options['dry_run']

        if not os.path.exists(path):
            raise CommandError(f'File not found: {path}')

        self.stdout.write(f'📂 Reading: {path}')
        warnings.filterwarnings('ignore')
        wb = openpyxl.load_workbook(path, data_only=True)
        sheets = wb.sheetnames
        self.stdout.write(f'   {len(sheets)} sheets found')

        em    = LEAVE_EXCEL_MAP
        mmap  = em['month_map']
        src   = em['source_tag']

        created = updated = skipped = errors = 0

        for sheet_name in sheets:
            ws = wb[sheet_name]

            # ── Read header ──────────────────────────────────────────────
            raw_name    = ws.cell(*em['name_cell']).value
            raw_emp_no  = ws.cell(*em['emp_no_cell']).value
            raw_dept    = ws.cell(*em['dept_cell']).value
            raw_title   = ws.cell(*em['title_cell']).value
            raw_joining = ws.cell(*em['joining_cell']).value

            name = _to_str(raw_name)
            if not name or name.lower() in ('name', 'employee', ''):
                skipped += 1
                continue  # skip header / summary sheet

            emp_code    = _clean_emp_code(raw_emp_no)
            job_title   = _clean_title(raw_emp_no, raw_title)
            department  = _to_str(raw_dept) or None
            joining_date = raw_joining.date() if hasattr(raw_joining, 'date') else None

            # ── Read totals row ──────────────────────────────────────────
            tr = em['total_row']
            total_earned   = _to_dec(ws.cell(tr, em['col_earned']).value)
            total_taken    = _to_dec(ws.cell(tr, em['col_taken']).value)
            total_encashed = _to_dec(ws.cell(tr, em['col_encashed']).value)
            leave_balance  = _to_dec(ws.cell(tr, em['col_balance']).value)

            # ── Read carryforward ────────────────────────────────────────
            cfr         = em['carryforward_row']
            carryforward = _to_dec(ws.cell(cfr, em['col_balance']).value)

            if dry:
                self.stdout.write(
                    f'  DRY  {name!r:50s}  emp={emp_code}  bal={leave_balance}'
                )
                continue

            # ── Upsert EmployeeLeaveRecord ───────────────────────────────
            try:
                with transaction.atomic():
                    defaults = dict(
                        employee_name   = name,
                        department      = department,
                        job_title       = job_title or None,
                        joining_date    = joining_date,
                        year            = year,
                        total_earned    = total_earned,
                        total_taken     = total_taken,
                        total_encashed  = total_encashed,
                        leave_balance   = leave_balance,
                        carryforward    = carryforward,
                        source_file     = src,
                    )
                    if emp_code:
                        rec, was_created = EmployeeLeaveRecord.objects.update_or_create(
                            employee_code=emp_code,
                            year=year,
                            defaults=defaults,
                        )
                    else:
                        # No code — match on name + year
                        rec, was_created = EmployeeLeaveRecord.objects.update_or_create(
                            employee_name=name,
                            year=year,
                            defaults={**defaults, 'employee_code': None},
                        )

                    if was_created:
                        created += 1
                    else:
                        updated += 1

                    # ── Upsert monthly rows ──────────────────────────────
                    for row_idx in range(em['month_start_row'], em['month_end_row'] + 1):
                        month_label = _to_str(ws.cell(row_idx, em['col_month']).value).lower()
                        month_num   = mmap.get(month_label)
                        if month_num is None:
                            continue

                        EmployeeLeaveMonthly.objects.update_or_create(
                            record=rec,
                            month=month_num,
                            defaults=dict(
                                earned   = _to_dec(ws.cell(row_idx, em['col_earned']).value),
                                taken    = _to_dec(ws.cell(row_idx, em['col_taken']).value),
                                encashed = _to_dec(ws.cell(row_idx, em['col_encashed']).value),
                                balance  = _to_dec(ws.cell(row_idx, em['col_balance']).value),
                            ),
                        )

            except Exception as exc:
                errors += 1
                self.stderr.write(f'  ERROR  {name!r}: {exc}')
                continue

        # ── Summary ──────────────────────────────────────────────────────────
        if dry:
            self.stdout.write(self.style.WARNING(
                f'\n🔍 DRY RUN — {created + updated} would be imported, {skipped} skipped'
            ))
        else:
            self.stdout.write(self.style.SUCCESS(
                f'\n✅ Done — {created} created, {updated} updated, '
                f'{skipped} skipped, {errors} errors'
            ))
