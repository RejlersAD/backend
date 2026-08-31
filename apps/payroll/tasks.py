"""
Payroll Celery Tasks
====================
Async background jobs triggered after synchronous request handling.

Current tasks:
  upload_master_payroll_to_s3 — generates the Excel from a saved MasterPayrollImport
                                  and uploads it to the payroll/exports/ S3 prefix.
                                  Called after generate_master_payroll saves to DB.
  run_monthly_leave_accrual    — credits one month's leave (annual_entitlement / 12)
                                  onto every current-year EmployeeLeaveRecord. Scheduled
                                  via apps.payroll.config.BEAT_SCHEDULE.
"""
from __future__ import annotations

import io
import logging
import uuid as uuid_lib
from datetime import timedelta
from decimal import Decimal
from zoneinfo import ZoneInfo

from celery import shared_task
from django.db import IntegrityError, transaction
from django.utils import timezone

logger = logging.getLogger(__name__)

# ── Soft-coded retry / timeout constants ─────────────────────────────────────
_MAX_RETRIES     = 3
_RETRY_BACKOFF_S = 60     # seconds between retries (exponential applied by Celery)
_TASK_TIMEOUT_S  = 300    # hard limit: 5 minutes

# Excel styling constants (keep in sync with views.py generate_master_payroll)
_HDR_COLOR   = '2563EB'   # blue-600
_HDR_FONT_CL = 'FFFFFF'


@shared_task(
    name='payroll.upload_master_payroll_to_s3',
    bind=True,
    max_retries=_MAX_RETRIES,
    default_retry_delay=_RETRY_BACKOFF_S,
    time_limit=_TASK_TIMEOUT_S,
)
def upload_master_payroll_to_s3(self, import_id: str) -> dict:
    """
    Build the 15-column master payroll Excel from the saved DB rows and
    upload it to the S3 exports bucket.

    On success: updates MasterPayrollImport.status = 'uploaded' and sets s3_key.
    On failure after retries: sets status = 'failed'.

    Args:
        import_id: UUID string of a MasterPayrollImport record.
    """
    from apps.payroll.models import MasterPayrollImport, MasterPayrollImportStatus
    from apps.payroll.storage import PayrollExportStorage, S3_AVAILABLE

    try:
        session = MasterPayrollImport.objects.get(id=import_id)
    except MasterPayrollImport.DoesNotExist:
        logger.error('upload_master_payroll_to_s3: import %s not found', import_id)
        return {'status': 'not_found', 'import_id': import_id}

    # ── 1. Build Excel in memory ──────────────────────────────────────────────
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f'Payroll Master {session.year}-{session.month:02d}'

        hdr_font  = Font(bold=True, color=_HDR_FONT_CL)
        hdr_fill  = PatternFill('solid', fgColor=_HDR_COLOR)
        thin      = Side(style='thin', color='CCCCCC')
        border    = Border(left=thin, right=thin, top=thin, bottom=thin)

        headers = [
            'Employee Code',        # 1
            'Employee Name',        # 2
            'Joining Date',         # 3
            'No. of Working Hours', # 4
            'Employee Salary',      # 5
            'Basic',                # 6
            'Allowance',            # 7
            'Transportation',       # 8
            'Home Allowance',       # 9
            'Other Allowance',      # 10
            'Other Pay',            # 11
            'Details',              # 12
            'Salary Deduction',     # 13
            'Deduction Details',    # 14
            'Final Salary',         # 15
        ]
        for col_idx, hdr in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=hdr)
            cell.font      = hdr_font
            cell.fill      = hdr_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border    = border
            ws.column_dimensions[cell.column_letter].width = max(14, len(hdr) + 4)

        rows_qs = session.rows.all().order_by('employee_name')
        for r_idx, row in enumerate(rows_qs, 2):
            vals = [
                row.employee_code,
                row.employee_name,
                row.joining_date or '',
                float(row.total_hours),
                float(row.employee_salary),
                float(row.basic_salary),
                float(row.total_allowances),
                float(row.transport_allowance),
                float(row.housing_allowance),
                float(row.other_allowances),
                float(row.other_pay),
                row.details or '',
                float(row.total_deductions),
                row.deduction_details or '',
                float(row.final_salary),
            ]
            for col_idx, val in enumerate(vals, 1):
                cell = ws.cell(row=r_idx, column=col_idx, value=val)
                cell.border = border
                if isinstance(val, float):
                    cell.number_format = '#,##0.00'

        # Freeze header row
        ws.freeze_panes = 'A2'

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        excel_bytes = buf.read()

    except Exception as exc:
        logger.exception('upload_master_payroll_to_s3: Excel build failed for %s', import_id)
        session.status = MasterPayrollImportStatus.FAILED
        session.save(update_fields=['status'])
        raise self.retry(exc=exc)

    # ── 2. Upload to S3 ───────────────────────────────────────────────────────
    if not S3_AVAILABLE:
        # S3 not configured (local dev) — mark ready, skip upload
        session.status = MasterPayrollImportStatus.READY
        session.save(update_fields=['status'])
        logger.info(
            'upload_master_payroll_to_s3: S3 not available — import %s marked ready (no upload)',
            import_id,
        )
        return {'status': 'ready_no_s3', 'import_id': import_id}

    try:
        storage    = PayrollExportStorage()
        filename   = f'master_payroll_{session.year}_{session.month:02d}_{import_id[:8]}.xlsx'
        s3_name    = storage.save(filename, io.BytesIO(excel_bytes))
        # s3_name is just the filename within the storage location prefix
        full_key   = f'{storage.location}/{s3_name}'

        session.s3_key = full_key
        session.status = MasterPayrollImportStatus.UPLOADED
        session.save(update_fields=['s3_key', 'status'])

        logger.info(
            'upload_master_payroll_to_s3: uploaded %s → s3://%s',
            filename, full_key,
        )
        return {'status': 'uploaded', 'import_id': import_id, 's3_key': full_key}

    except Exception as exc:
        logger.exception('upload_master_payroll_to_s3: S3 upload failed for %s', import_id)
        session.status = MasterPayrollImportStatus.FAILED
        session.save(update_fields=['status'])
        raise self.retry(exc=exc)


# ─────────────────────────────────────────────────────────────────────────────
# Monthly leave accrual
# ─────────────────────────────────────────────────────────────────────────────

@shared_task(
    name='payroll.run_monthly_leave_accrual',
    bind=True,
    max_retries=_MAX_RETRIES,
    default_retry_delay=_RETRY_BACKOFF_S,
    time_limit=_TASK_TIMEOUT_S,
)
def run_monthly_leave_accrual(self, triggered_by: str = 'celery_beat') -> dict:
    """
    Credit one month's leave accrual (annual_entitlement / 12) onto every
    EmployeeLeaveRecord for the current year, recompute leave_balance, and
    log the run to MonthlyLeaveAccrualLog.

    The beat schedule (apps.payroll.config.BEAT_SCHEDULE) fires this daily
    from the 28th-31st at 20:05 UTC — since crontab can't express "last day
    of the month" directly, this task re-checks the real Abu Dhabi date
    itself and only does real work when tomorrow (Asia/Dubai) is the 1st.
    That makes execution correct regardless of whatever CELERY_TIMEZONE is
    configured, and safe to invoke manually/via API on any day (it's simply
    a no-op if it isn't month-end yet, and idempotent per (year, month,
    triggered_by) if run more than once for the same period).
    """
    from apps.payroll.models import EmployeeLeaveRecord, MonthlyLeaveAccrualLog

    now_dubai = timezone.now().astimezone(ZoneInfo('Asia/Dubai'))
    tomorrow_dubai = now_dubai + timedelta(days=1)
    if tomorrow_dubai.day != 1:
        logger.debug(
            'run_monthly_leave_accrual: not month-end yet in Asia/Dubai (%s), skipping',
            now_dubai.date(),
        )
        return {'status': 'skipped', 'reason': 'not_month_end', 'checked_date': str(now_dubai.date())}

    target_year, target_month = tomorrow_dubai.year, tomorrow_dubai.month

    if MonthlyLeaveAccrualLog.objects.filter(
        year=target_year, month=target_month, triggered_by=triggered_by,
    ).exists():
        logger.info(
            'run_monthly_leave_accrual: already run for %s-%02d (triggered_by=%s), skipping',
            target_year, target_month, triggered_by,
        )
        return {'status': 'skipped', 'reason': 'already_run', 'year': target_year, 'month': target_month}

    records_processed    = 0
    monthly_accrual_used = Decimal('0')

    try:
        with transaction.atomic():
            records = EmployeeLeaveRecord.objects.filter(year=target_year).select_for_update()
            for record in records:
                entitlement = record.annual_entitlement or Decimal('22')
                accrual     = (entitlement / Decimal('12')).quantize(Decimal('0.0001'))
                monthly_accrual_used = accrual

                record.total_earned  = (record.total_earned or Decimal('0')) + accrual
                record.leave_balance = (
                    record.total_earned - record.total_taken - record.total_encashed + record.carryforward
                )
                record.save(update_fields=['total_earned', 'leave_balance'])
                records_processed += 1

            MonthlyLeaveAccrualLog.objects.create(
                year                  = target_year,
                month                 = target_month,
                triggered_by          = triggered_by,
                records_processed     = records_processed,
                records_updated       = records_processed,
                monthly_accrual_used  = monthly_accrual_used,
                status                = 'success',
            )

    except IntegrityError:
        # Another worker logged this period between our existence check and create()
        logger.warning(
            'run_monthly_leave_accrual: race on log create for %s-%02d — already recorded elsewhere',
            target_year, target_month,
        )
        return {'status': 'skipped', 'reason': 'race_already_logged', 'year': target_year, 'month': target_month}

    except Exception as exc:
        logger.exception('run_monthly_leave_accrual: failed for %s-%02d', target_year, target_month)
        try:
            MonthlyLeaveAccrualLog.objects.create(
                year          = target_year,
                month         = target_month,
                triggered_by  = triggered_by,
                records_processed = 0,
                status        = 'failed',
                error_message = str(exc)[:2000],
            )
        except IntegrityError:
            pass
        raise self.retry(exc=exc)

    logger.info(
        'run_monthly_leave_accrual: completed %s-%02d — %d record(s) updated',
        target_year, target_month, records_processed,
    )
    return {
        'status':             'success',
        'year':               target_year,
        'month':              target_month,
        'records_processed':  records_processed,
    }
