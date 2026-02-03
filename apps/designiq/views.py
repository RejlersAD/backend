"""
DesignIQ Views - AI-Powered Design Analysis API
Intelligent design verification, optimization, and recommendations
"""

from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser
from django.db.models import Q, Count
from django.utils import timezone
from django.http import HttpResponse
import logging

from .models import DesignProject, DesignAnalysis, DesignOptimization, DesignTemplate, EngineeringListItem, LIST_TYPES
from .s3_utils import s3_storage  # S3 document storage
from .serializers import (
    DesignProjectListSerializer, DesignProjectDetailSerializer,
    DesignProjectCreateSerializer, DesignAnalysisSerializer,
    DesignOptimizationSerializer, DesignTemplateSerializer,
    DesignAnalysisCreateSerializer, EngineeringListItemSerializer,
    EngineeringListItemListSerializer, ListTypeConfigSerializer
)

logger = logging.getLogger(__name__)


class DesignProjectViewSet(viewsets.ModelViewSet):
    """
    ViewSet for DesignIQ Projects
    Handles design project creation, analysis, and AI-powered insights
    """
    
    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser, JSONParser]
    
    def get_queryset(self):
        """Get projects for current user with optional filtering"""
        queryset = DesignProject.objects.all()
        
        # Filter by user unless staff
        if not self.request.user.is_staff:
            queryset = queryset.filter(created_by=self.request.user)
        
        # Filter by status
        status_filter = self.request.query_params.get('status')
        if status_filter:
            queryset = queryset.filter(status=status_filter)
        
        # Filter by design type
        design_type = self.request.query_params.get('design_type')
        if design_type:
            queryset = queryset.filter(design_type=design_type)
        
        # Search
        search = self.request.query_params.get('search')
        if search:
            queryset = queryset.filter(
                Q(project_name__icontains=search) |
                Q(description__icontains=search) |
                Q(organization__icontains=search)
            )
        
        return queryset.select_related('created_by').prefetch_related('analyses', 'optimizations')
    
    def get_serializer_class(self):
        """Use different serializers for different actions"""
        if self.action == 'retrieve':
            return DesignProjectDetailSerializer
        elif self.action in ['create', 'update', 'partial_update']:
            return DesignProjectCreateSerializer
        return DesignProjectListSerializer
    
    def perform_create(self, serializer):
        """Create project and set user"""
        project = serializer.save(created_by=self.request.user)
        logger.info(f"[DesignIQ] Project created: {project.id} by {self.request.user.email}")
    
    @action(detail=True, methods=['post'])
    def analyze(self, request, pk=None):
        """
        Trigger AI analysis for a project
        POST /api/v1/designiq/projects/{id}/analyze/
        Body: {
            "parameters": {...},  // Optional analysis parameters
            "force_reanalysis": false  // Re-analyze even if already completed
        }
        """
        project = self.get_object()
        
        if project.status == 'analyzing':
            return Response(
                {"error": "Analysis already in progress"},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            # Update status
            project.status = 'analyzing'
            project.save()
            
            # Here you would trigger your AI analysis
            # For now, we'll return a placeholder response
            # TODO: Integrate with actual AI service
            
            logger.info(f"[DesignIQ] Analysis triggered for project: {project.id}")
            
            return Response({
                "message": "Analysis started successfully",
                "project_id": str(project.id),
                "status": "analyzing"
            })
            
        except Exception as e:
            logger.error(f"[DesignIQ] Analysis error: {str(e)}")
            project.status = 'failed'
            project.error_message = str(e)
            project.save()
            
            return Response(
                {"error": str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=True, methods=['get'])
    def summary(self, request, pk=None):
        """
        Get project summary with statistics
        GET /api/v1/designiq/projects/{id}/summary/
        """
        project = self.get_object()
        
        analyses_stats = project.analyses.aggregate(
            total=Count('id'),
            critical=Count('id', filter=Q(severity='critical')),
            high=Count('id', filter=Q(severity='high')),
            resolved=Count('id', filter=Q(is_resolved=True))
        )
        
        optimizations_stats = project.optimizations.aggregate(
            total=Count('id'),
            high_impact=Count('id', filter=Q(impact='high')),
            implemented=Count('id', filter=Q(is_implemented=True))
        )
        
        return Response({
            "project": DesignProjectDetailSerializer(project).data,
            "analyses": analyses_stats,
            "optimizations": optimizations_stats,
            "summary": {
                "total_findings": analyses_stats['total'],
                "critical_issues": analyses_stats['critical'],
                "high_priority_issues": analyses_stats['high'],
                "resolution_rate": (analyses_stats['resolved'] / analyses_stats['total'] * 100) if analyses_stats['total'] > 0 else 0,
                "optimization_count": optimizations_stats['total'],
                "implemented_optimizations": optimizations_stats['implemented']
            }
        })
    
    @action(detail=False, methods=['get'])
    def dashboard(self, request):
        """
        Get dashboard statistics for all user projects
        GET /api/v1/designiq/projects/dashboard/
        """
        queryset = self.get_queryset()
        
        stats = {
            "total_projects": queryset.count(),
            "by_status": {
                "draft": queryset.filter(status='draft').count(),
                "analyzing": queryset.filter(status='analyzing').count(),
                "completed": queryset.filter(status='completed').count(),
                "failed": queryset.filter(status='failed').count(),
            },
            "by_design_type": {},
            "recent_projects": DesignProjectListSerializer(
                queryset.order_by('-created_at')[:5],
                many=True
            ).data
        }
        
        # Get counts by design type
        for choice_value, choice_label in DesignProject.DESIGN_TYPE_CHOICES:
            stats['by_design_type'][choice_value] = queryset.filter(design_type=choice_value).count()
        
        return Response(stats)


class DesignAnalysisViewSet(viewsets.ModelViewSet):
    """
    ViewSet for Design Analyses
    """
    
    permission_classes = [IsAuthenticated]
    serializer_class = DesignAnalysisSerializer
    
    def get_queryset(self):
        """Get analyses for user's projects"""
        if self.request.user.is_staff:
            return DesignAnalysis.objects.all().select_related('project', 'resolved_by')
        
        return DesignAnalysis.objects.filter(
            project__created_by=self.request.user
        ).select_related('project', 'resolved_by')
    
    def get_serializer_class(self):
        if self.action in ['create', 'update', 'partial_update']:
            return DesignAnalysisCreateSerializer
        return DesignAnalysisSerializer
    
    @action(detail=True, methods=['post'])
    def resolve(self, request, pk=None):
        """
        Mark analysis as resolved
        POST /api/v1/designiq/analyses/{id}/resolve/
        Body: {"resolution_notes": "Fixed by..."}
        """
        analysis = self.get_object()
        
        analysis.is_resolved = True
        analysis.resolved_by = request.user
        analysis.resolved_at = timezone.now()
        analysis.resolution_notes = request.data.get('resolution_notes', '')
        analysis.save()
        
        return Response(DesignAnalysisSerializer(analysis).data)


class DesignOptimizationViewSet(viewsets.ModelViewSet):
    """
    ViewSet for Design Optimizations
    """
    
    permission_classes = [IsAuthenticated]
    serializer_class = DesignOptimizationSerializer
    
    def get_queryset(self):
        """Get optimizations for user's projects"""
        if self.request.user.is_staff:
            return DesignOptimization.objects.all().select_related('project', 'implemented_by')
        
        return DesignOptimization.objects.filter(
            project__created_by=self.request.user
        ).select_related('project', 'implemented_by')
    
    @action(detail=True, methods=['post'])
    def implement(self, request, pk=None):
        """
        Mark optimization as implemented
        POST /api/v1/designiq/optimizations/{id}/implement/
        Body: {"implementation_notes": "..."}
        """
        optimization = self.get_object()
        
        optimization.is_implemented = True
        optimization.implemented_by = request.user
        optimization.implemented_at = timezone.now()
        
        if 'implementation_notes' in request.data:
            optimization.implementation_notes = request.data['implementation_notes']
        
        optimization.save()
        
        return Response(DesignOptimizationSerializer(optimization).data)


class DesignTemplateViewSet(viewsets.ModelViewSet):
    """
    ViewSet for Design Templates
    """
    
    permission_classes = [IsAuthenticated]
    serializer_class = DesignTemplateSerializer
    
    def get_queryset(self):
        """Get public templates and user's private templates"""
        queryset = DesignTemplate.objects.filter(
            Q(is_public=True) | Q(created_by=self.request.user)
        ).select_related('created_by')
        
        design_type = self.request.query_params.get('design_type')
        if design_type:
            queryset = queryset.filter(design_type=design_type)
        
        return queryset.order_by('-usage_count', 'name')
    
    @action(detail=True, methods=['post'])
    def use_template(self, request, pk=None):
        """
        Create a new project from template
        POST /api/v1/designiq/templates/{id}/use_template/
        Body: {
            "project_name": "...",
            "parameters": {...}
        }
        """
        template = self.get_object()
        
        # Increment usage count
        template.usage_count += 1
        template.save()
        
        # Create project from template
        project_data = {
            "project_name": request.data.get('project_name', f"Project from {template.name}"),
            "design_type": template.design_type,
            "description": f"Created from template: {template.name}",
            "design_parameters": request.data.get('parameters', template.template_data),
            "created_by": request.user,
        }
        
        project = DesignProject.objects.create(**project_data)
        
        return Response({
            "message": "Project created from template",
            "project": DesignProjectDetailSerializer(project).data
        }, status=status.HTTP_201_CREATED)


class EngineeringListItemViewSet(viewsets.ModelViewSet):
    """
    ViewSet for Engineering List Items
    Handles Line List, Equipment List, Tie-In List, and Alarm/Trip List
    """
    
    permission_classes = [IsAuthenticated]
    parser_classes = [JSONParser, MultiPartParser, FormParser]
    
    def get_queryset(self):
        """Get list items with filtering"""
        queryset = EngineeringListItem.objects.select_related('project', 'created_by')
        
        # Filter by list type
        list_type = self.request.query_params.get('list_type')
        if list_type:
            queryset = queryset.filter(list_type=list_type)
        
        # Filter by project
        project_id = self.request.query_params.get('project')
        if project_id:
            queryset = queryset.filter(project_id=project_id)
        
        # Filter by status
        status_filter = self.request.query_params.get('status')
        if status_filter:
            queryset = queryset.filter(status=status_filter)
        
        # Filter by validation status
        is_validated = self.request.query_params.get('is_validated')
        if is_validated is not None:
            queryset = queryset.filter(is_validated=is_validated.lower() == 'true')
        
        # Search by item tag or description
        search = self.request.query_params.get('search')
        if search:
            queryset = queryset.filter(
                Q(item_tag__icontains=search) | Q(description__icontains=search)
            )
        
        return queryset.order_by('-created_at')
    
    def get_serializer_class(self):
        """Use different serializers for list vs detail"""
        if self.action == 'list':
            return EngineeringListItemListSerializer
        return EngineeringListItemSerializer
    
    @action(detail=False, methods=['get'])
    def list_types(self, request):
        """Get available list types configuration"""
        list_types_data = [
            {
                'code': code,
                'name': config['name'],
                'icon': config['icon'],
                'description': config['description'],
                'default_fields': config['default_fields']
            }
            for code, config in LIST_TYPES.items()
        ]
        
        serializer = ListTypeConfigSerializer(list_types_data, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def stats(self, request):
        """Get statistics for list items"""
        list_type = request.query_params.get('list_type')
        project_id = request.query_params.get('project')
        
        queryset = self.get_queryset()
        if list_type:
            queryset = queryset.filter(list_type=list_type)
        if project_id:
            queryset = queryset.filter(project_id=project_id)
        
        stats = {
            'total': queryset.count(),
            'by_status': {
                'active': queryset.filter(status='active').count(),
                'pending': queryset.filter(status='pending').count(),
                'approved': queryset.filter(status='approved').count(),
                'rejected': queryset.filter(status='rejected').count(),
                'inactive': queryset.filter(status='inactive').count(),
            },
            'validated': queryset.filter(is_validated=True).count(),
            'not_validated': queryset.filter(is_validated=False).count(),
            'by_list_type': {}
        }
        
        # Count by list type
        for code, config in LIST_TYPES.items():
            count = queryset.filter(list_type=code).count()
            if count > 0:
                stats['by_list_type'][code] = {
                    'name': config['name'],
                    'count': count
                }
        
        return Response(stats)
    
    @action(detail=True, methods=['post'])
    def validate_item(self, request, pk=None):
        """Validate a list item"""
        item = self.get_object()
        
        item.is_validated = True
        item.validated_by = request.user
        item.validated_at = timezone.now()
        item.validation_notes = request.data.get('notes', '')
        item.save()
        
        return Response({
            "message": "Item validated successfully",
            "item": EngineeringListItemSerializer(item, context={'request': request}).data
        })
    
    @action(detail=True, methods=['post'])
    def bulk_import(self, request):
        """Bulk import items from CSV/Excel data"""
        items_data = request.data.get('items', [])
        list_type = request.data.get('list_type')
        project_id = request.data.get('project')
        
        if not list_type or list_type not in LIST_TYPES:
            return Response(
                {"error": "Valid list_type is required"},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        created_items = []
        errors = []
        
        for idx, item_data in enumerate(items_data):
            try:
                item = EngineeringListItem.objects.create(
                    list_type=list_type,
                    project_id=project_id if project_id else None,
                    item_tag=item_data.get('item_tag', f'ITEM-{idx+1}'),
                    description=item_data.get('description', ''),
                    data=item_data.get('data', {}),
                    status=item_data.get('status', 'active'),
                    created_by=request.user
                )
                created_items.append(item)
            except Exception as e:
                errors.append({
                    'row': idx + 1,
                    'error': str(e),
                    'data': item_data
                })
        
        return Response({
            "message": f"Imported {len(created_items)} items",
            "created": len(created_items),
            "errors": len(errors),
            "error_details": errors if errors else None
        }, status=status.HTTP_201_CREATED if created_items else status.HTTP_400_BAD_REQUEST)
    
    @action(detail=False, methods=['post'])
    def export(self, request):
        """Export list items to structured format"""
        queryset = self.get_queryset()
        list_type = request.query_params.get('list_type')
        
        if list_type:
            queryset = queryset.filter(list_type=list_type)
        
        serializer = EngineeringListItemSerializer(queryset, many=True, context={'request': request})
        
        return Response({
            "list_type": list_type,
            "count": queryset.count(),
            "items": serializer.data
        })
    
    @action(detail=False, methods=['post'], parser_classes=[MultiPartParser, FormParser])
    def upload_pid(self, request):
        """
        Upload P&ID PDF and extract line list items using OCR
        Intelligently detects line numbers in horizontal/vertical orientations
        Parses components: size, fluid, sequence, class, insulation, connections
        """
        pid_file = request.FILES.get('pid_file')
        list_type = request.data.get('list_type', 'line_list')
        
        if not pid_file:
            return Response({
                "error": "No P&ID file provided"
            }, status=status.HTTP_400_BAD_REQUEST)
        
        if not pid_file.name.endswith('.pdf'):
            return Response({
                "error": "Only PDF files are supported"
            }, status=status.HTTP_400_BAD_REQUEST)
        
        if list_type not in LIST_TYPES:
            return Response({
                "error": f"Invalid list_type. Must be one of: {', '.join(LIST_TYPES.keys())}"
            }, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            import tempfile
            import os
            from django.core.files.storage import default_storage
            from .pid_ocr_extractor_v2 import PIDLineExtractorV2
            from .models import DesignProject
            
            # Get or create project
            project, _ = DesignProject.objects.get_or_create(
                project_name="P&ID Upload Project",
                defaults={
                    'created_by': request.user,
                    'design_type': 'pid',
                    'status': 'active'
                }
            )
            
            # 🆔 GENERATE UNIQUE 4-DIGIT DOCUMENT ID
            # Find the highest existing document ID from all line items
            last_doc_id = 0
            existing_items = EngineeringListItem.objects.filter(
                data__has_key='document_id'
            ).order_by('-created_at').first()
            
            if existing_items and existing_items.data.get('document_id'):
                try:
                    # Extract the 4-digit number from format "0001-filename.pdf"
                    doc_id_str = existing_items.data['document_id'].split('-')[0]
                    last_doc_id = int(doc_id_str)
                except (ValueError, IndexError):
                    pass
            
            # Generate new 4-digit ID (increment)
            new_doc_id = last_doc_id + 1
            document_id = f"{new_doc_id:04d}-{pid_file.name}"
            
            logger.info(f"🆔 Generated Document ID: {document_id}")
            
            # Read file content once to avoid file pointer issues
            pid_file.seek(0)
            file_content = pid_file.read()
            
            # 📤 UPLOAD TO S3 (if configured)
            from io import BytesIO
            s3_file = BytesIO(file_content)
            s3_result = s3_storage.upload_document(
                file_obj=s3_file,
                document_id=document_id,
                original_filename=pid_file.name
            )
            
            if s3_result['success']:
                logger.info(f"☁️ Uploaded to S3: {s3_result['s3_key']}")
                saved_path = s3_result['s3_key']  # Use S3 key as path
                storage_type = 's3'
                s3_url = s3_result['s3_url']
            else:
                logger.warning(f"⚠️ S3 upload failed: {s3_result['error']}, falling back to local storage")
                # Fallback to local storage - recreate file object with content
                from django.core.files.uploadedfile import InMemoryUploadedFile
                local_file = BytesIO(file_content)
                local_file.seek(0)
                file_path = f"designiq/pid_uploads/{timezone.now().strftime('%Y/%m/%d')}/{document_id}"
                saved_path = default_storage.save(file_path, local_file)
                storage_type = 'local'
                s3_url = None
            
            # Save to temp file for OCR processing using the content we already read
            with tempfile.NamedTemporaryFile(delete=False, suffix='.pdf') as tmp:
                tmp.write(file_content)
                tmp_path = tmp.name
            
            try:
                # Get format preference from request
                # - include_area: for ADNOC Onshore with area format
                # - format_type: 'onshore' (default) or 'offshore' for ADNOC Offshore
                include_area = request.POST.get('include_area', 'false').lower() == 'true'
                format_type = request.POST.get('format_type', 'onshore').lower()
                
                # Extract line numbers using Multi-Engine OCR + Geometric FROM-TO Detection
                extractor = PIDLineExtractorV2()
                line_items = extractor.extract_from_pdf(tmp_path, include_area=include_area, format_type=format_type)
                table_data = extractor.format_as_table_data(line_items)
                
                # FROM-TO detection methods (in order of preference):
                # 1. Spatial Matching (research paper method)
                # 2. OpenAI Vision (AI-powered visual analysis)
                # 3. Geometric Line Detection (OpenCV + connectivity graph)
                
                logger.info(f"📊 Extracted {len(line_items)} line numbers from {pid_file.name}")
                logger.info(f"🎯 Using Multi-Engine OCR (Tesseract + EasyOCR + PaddleOCR) + Regex + Geometric Detection")
                if format_type == 'offshore':
                    logger.info(f"📍 Format: ADNOC OFFSHORE (AREA-FLUID-SIZE-PIPECLASS-SEQUENCE)")
                elif include_area:
                    logger.info(f"📍 Format: ADNOC ONSHORE WITH AREA (SIZE\"-AREA-FLUID-SEQ-PIPECLASS)")
                else:
                    logger.info(f"📍 Format: GENERAL (SIZE-FLUID-SEQ-PIPECLASS)")
                
                # Create or update EngineeringListItem for each detected line
                created_items = []
                updated_items = []
                logger.info(f"📝 Saving {len(table_data)} extracted lines to database (project={project.id if project else None}, list_type={list_type})...")
                
                for idx, line_data in enumerate(table_data):
                    try:
                        # Use update_or_create to handle duplicates gracefully
                        item_data = {
                            'description': f"{line_data['fluid_description']} Line - {line_data['size']}",
                            'status': 'pending',
                            'is_validated': False,
                            'data': {
                                'source': 'pid_ocr',
                                'filename': pid_file.name,
                                'document_id': document_id,  # 🆔 UNIQUE DOCUMENT ID
                                'document_path': saved_path,  # 📄 S3 KEY or LOCAL PATH
                                'storage_type': storage_type,  # 's3' or 'local'
                                's3_url': s3_url,  # Direct S3 URL (if applicable)
                                'upload_timestamp': timezone.now().isoformat(),
                                'format_type': format_type,
                                'include_area': include_area,
                                'page_number': line_data.get('page', 1),
                                'fluid_code': line_data['fluid_code'],
                                'fluid_description': line_data['fluid_description'],
                                'size': line_data['size'],
                                'area': line_data.get('area', ''),  # AREA field
                                'sequence_no': line_data['sequence_no'],
                                'pipr_class': line_data['pipr_class'],
                                'insulation': line_data['insulation'],
                                'from_equipment': line_data.get('from_equipment', ''),
                                'to_equipment': line_data.get('to_equipment', ''),
                                'from_line': line_data.get('from_line', ''),  # NEW: FROM line number
                                'to_line': line_data.get('to_line', ''),      # NEW: TO line number
                                'flow_detection_method': line_data.get('flow_detection_method', ''),
                                'flow_confidence': line_data.get('flow_confidence', '')
                            },
                            'attachments': [{
                                'type': 'pid_pdf',
                                'filename': pid_file.name,
                                'document_id': document_id,
                                'path': saved_path,
                                'storage_type': storage_type,
                                's3_url': s3_url,
                                'uploaded_at': timezone.now().isoformat()
                            }]
                        }
                        
                        # Only set created_by on new items
                        item, created = EngineeringListItem.objects.update_or_create(
                            list_type=list_type,
                            project=project,
                            item_tag=line_data['line_number'],
                            defaults=item_data
                        )
                        
                        # Set created_by if this is a new item
                        if created and not item.created_by:
                            item.created_by = request.user
                            item.save(update_fields=['created_by'])
                        
                        if created:
                            created_items.append(item)
                            if idx == 0:
                                logger.info(f"   ✅ First item CREATED: {item.item_tag}")
                        else:
                            updated_items.append(item)
                            if idx == 0:
                                logger.info(f"   ✅ First item UPDATED: {item.item_tag}")
                        
                    except Exception as item_err:
                        logger.error(f"❌ Failed to save item {idx+1}: {line_data.get('line_number', '?')} - Error: {str(item_err)}", exc_info=True)
                        continue
                
                total_items = len(created_items) + len(updated_items)
                logger.info(f"✅ Created {len(created_items)} new items, updated {len(updated_items)} existing items from P&ID OCR")
                
                return Response({
                    "message": "P&ID processed successfully using OCR",
                    "filename": pid_file.name,
                    "document_id": document_id,  # 🆔 RETURN DOCUMENT ID
                    "document_path": saved_path,  # 📄 RETURN FILE PATH
                    "items_created": len(created_items),
                    "items_updated": len(updated_items),
                    "total_items": total_items,
                    "extracted_lines": table_data,
                    "note": "Multi-engine OCR detection (Tesseract + EasyOCR + PaddleOCR + OpenAI GPT-4)"
                }, status=status.HTTP_201_CREATED)
                
            finally:
                # Clean up temp file
                if os.path.exists(tmp_path):
                    os.unlink(tmp_path)
            
        except Exception as e:
            logger.error(f"❌ Error uploading/processing P&ID: {str(e)}", exc_info=True)
            return Response({
                "error": f"Failed to upload P&ID: {str(e)}"
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)    
    @action(detail=False, methods=['get'], url_path='documents')
    def list_documents(self, request):
        """
        List all uploaded P&ID documents with their unique IDs
        Groups line items by document_id
        """
        try:
            list_type = request.query_params.get('list_type', 'line_list')
            
            # Get all items with document IDs
            items = EngineeringListItem.objects.filter(
                list_type=list_type,
                data__has_key='document_id'
            ).order_by('-created_at')
            
            # Group by document_id
            documents_map = {}
            for item in items:
                doc_id = item.data.get('document_id')
                if not doc_id:
                    continue
                
                if doc_id not in documents_map:
                    documents_map[doc_id] = {
                        'document_id': doc_id,
                        'filename': item.data.get('filename', 'Unknown'),
                        'original_filename': item.data.get('filename', 'Unknown'),
                        'document_path': item.data.get('document_path', ''),
                        'storage_type': item.data.get('storage_type', 'local'),
                        's3_url': item.data.get('s3_url'),
                        'upload_date': item.data.get('upload_timestamp', item.created_at.isoformat()),
                        'uploaded_by': item.created_by.get_full_name() if item.created_by else 'Unknown',
                        'format_type': item.data.get('format_type', 'general'),
                        'line_count': 0,
                        'item_ids': []
                    }
                
                documents_map[doc_id]['line_count'] += 1
                documents_map[doc_id]['item_ids'].append(item.id)
            
            documents_list = list(documents_map.values())
            
            return Response({
                'documents': documents_list,
                'total_documents': len(documents_list)
            }, status=status.HTTP_200_OK)
            
        except Exception as e:
            logger.error(f"Error listing documents: {str(e)}", exc_info=True)
            return Response({
                "error": f"Failed to list documents: {str(e)}"
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @action(detail=False, methods=['get'], url_path='documents/(?P<document_id>.+)/download')
    def download_document(self, request, document_id=None):
        """
        Download P&ID document by document_id
        Supports both S3 and local storage
        """
        try:
            logger.info(f"📥 Download request for document_id: {document_id}")
            
            # Find an item with this document_id to get the storage info
            item = EngineeringListItem.objects.filter(
                data__document_id=document_id
            ).first()
            
            if not item:
                logger.warning(f"❌ Document not found in database: {document_id}")
                return Response({
                    "error": f"Document not found: {document_id}"
                }, status=status.HTTP_404_NOT_FOUND)
            
            storage_type = item.data.get('storage_type', 'local')
            document_path = item.data.get('document_path', '')
            filename = item.data.get('filename', document_id)
            
            if storage_type == 's3':
                # Generate presigned URL for S3 (1 hour expiration)
                presigned_url = s3_storage.generate_presigned_url(
                    s3_key=document_path,
                    expiration=3600
                )
                
                if presigned_url:
                    # Return presigned URL
                    return Response({
                        'url': presigned_url,
                        'filename': filename,
                        'storage_type': 's3',
                        'expires_in': 3600
                    }, status=status.HTTP_200_OK)
                else:
                    return Response({
                        "error": "Failed to generate download URL"
                    }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
            
            else:
                # Local storage - serve file directly
                from django.core.files.storage import default_storage
                
                if not default_storage.exists(document_path):
                    return Response({
                        "error": f"Document file not found: {document_path}"
                    }, status=status.HTTP_404_NOT_FOUND)
                
                file = default_storage.open(document_path, 'rb')
                response = HttpResponse(file.read(), content_type='application/pdf')
                response['Content-Disposition'] = f'inline; filename="{filename}"'
                file.close()
                
                return response
            
        except Exception as e:
            logger.error(f"Error downloading document: {str(e)}", exc_info=True)
            return Response({
                "error": f"Failed to download document: {str(e)}"
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @action(detail=False, methods=['delete'], url_path='documents/(?P<document_id>.+)')
    def delete_document(self, request, document_id=None):
        """
        Delete all line items associated with a document ID
        """
        try:
            logger.info(f"🗑️ Delete request for document_id: {document_id}")
            
            # Find all items with this document_id
            items = EngineeringListItem.objects.filter(
                data__document_id=document_id
            )
            
            count = items.count()
            if count == 0:
                logger.warning(f"❌ No items found for document: {document_id}")
                return Response({
                    "error": f"No items found for document ID: {document_id}"
                }, status=status.HTTP_404_NOT_FOUND)
            
            # Delete all items
            items.delete()
            
            logger.info(f"🗑️ Deleted document {document_id} with {count} line items")
            
            return Response({
                "message": f"Successfully deleted document {document_id}",
                "items_deleted": count
            }, status=status.HTTP_200_OK)
            
        except Exception as e:
            logger.error(f"Error deleting document: {str(e)}", exc_info=True)
            return Response({
                "error": f"Failed to delete document: {str(e)}"
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @action(detail=False, methods=['get'], url_path='export-document-excel')
    def export_document_excel(self, request):
        """
        Export document line items to Excel (CRS multi-revision pattern)
        GET /api/v1/designiq/lists/export-document-excel/?document_id={document_id}
        
        Query Parameters:
            document_id: The document ID to export (required)
        
        Returns Excel file with line items
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
            from io import BytesIO
            
            # Get document_id from query parameters
            document_id = request.query_params.get('document_id')
            if not document_id:
                return Response({
                    "error": "document_id query parameter is required"
                }, status=status.HTTP_400_BAD_REQUEST)
            
            logger.info(f"📊 Excel export request for document: {document_id}")
            
            # Find all items with this document_id
            items = EngineeringListItem.objects.filter(
                list_type='line_list',
                data__document_id=document_id
            ).order_by('item_tag')
            
            if not items.exists():
                logger.warning(f"❌ No items found for document: {document_id}")
                return Response({
                    "error": f"No line items found for document ID: {document_id}"
                }, status=status.HTTP_404_NOT_FOUND)
            
            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Line List"
            
            # Header style
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True, size=11)
            
            # Define headers
            headers = [
                'Line Number', 'Size', 'Fluid Code', 'Fluid Description',
                'Sequence No', 'Pipe Class', 'Insulation', 'Area',
                'FROM', 'TO', 'Status', 'Validated'
            ]
            
            ws.append(headers)
            
            # Style headers
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Add data rows
            for item in items:
                ws.append([
                    item.item_tag,
                    item.data.get('size', ''),
                    item.data.get('fluid_code', ''),
                    item.data.get('fluid_description', ''),
                    item.data.get('sequence_no', ''),
                    item.data.get('pipr_class', ''),
                    item.data.get('insulation', ''),
                    item.data.get('area', ''),
                    item.data.get('from_line', ''),
                    item.data.get('to_line', ''),
                    item.status,
                    'Yes' if item.is_validated else 'No'
                ])
            
            # Auto-size columns
            for col_num in range(1, len(headers) + 1):
                col_letter = get_column_letter(col_num)
                max_length = len(headers[col_num - 1])
                for cell in ws[col_letter]:
                    try:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = min(cell_length, 50)
                    except:
                        pass
                ws.column_dimensions[col_letter].width = max_length + 2
            
            # Save to BytesIO
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            # Get filename from first item
            filename = items.first().data.get('filename', document_id)
            safe_filename = "".join(c for c in filename if c.isalnum() or c in "-_.").strip() or "line_list"
            excel_filename = f"{safe_filename}_line_list.xlsx"
            
            logger.info(f"✅ Generated Excel with {items.count()} line items")
            
            # Create response
            response = HttpResponse(
                output.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{excel_filename}"'
            response['X-Item-Count'] = str(items.count())
            
            return response
            
        except Exception as e:
            logger.error(f"Error exporting Excel: {str(e)}", exc_info=True)
            return Response({
                "error": f"Failed to export Excel: {str(e)}"
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)