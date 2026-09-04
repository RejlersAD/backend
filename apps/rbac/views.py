"""
RBAC Views - DRF ViewSets for Super Admin Dashboard
"""
from rest_framework import viewsets, status, filters
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.views import APIView
from django.utils import timezone
from django.db import transaction
from django.db.models import Q, Count, Prefetch
from django_filters.rest_framework import DjangoFilterBackend
from django.http import FileResponse

from .models import (
    Organization, Module, Permission, Role, RolePermission, RoleModule,
    UserProfile, UserRole, UserStorage, AuditLog, AccessRequest,
    Achievement, WorkExperience, SocialMediaLink,
)
from .serializers import (
    OrganizationSerializer, ModuleSerializer, PermissionSerializer,
    RoleSerializer, RoleListSerializer, RolePermissionSerializer, RoleModuleSerializer,
    UserProfileSerializer, UserProfileSelfSerializer, UserProfileListSerializer, UserRoleSerializer,
    UserStorageSerializer, AuditLogSerializer,
    UserPermissionCheckSerializer, UserModuleCheckSerializer,
    AccessRequestSerializer,
)
from .permissions import (
    IsSuperAdmin, IsAdmin, CanManageUsers, CanManageRoles, SameOrganization
)
from .utils import create_audit_log
from .s3_service import S3Service
from .pagination import FlexiblePageNumberPagination


class OrganizationViewSet(viewsets.ModelViewSet):
    """
    ViewSet for managing organizations
    Admins can view organizations, only super admin can create/edit/delete
    """
    queryset = Organization.objects.all()
    serializer_class = OrganizationSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter, DjangoFilterBackend]
    search_fields = ['name', 'code', 'primary_contact_email']
    ordering_fields = ['name', 'created_at']
    filterset_fields = ['is_active']
    
    def get_permissions(self):
        """
        Allow admins to read organizations, only super admin can modify
        """
        if self.action in ['list', 'retrieve']:
            permission_classes = [IsAuthenticated, IsAdmin]
        else:
            permission_classes = [IsAuthenticated, IsSuperAdmin]
        return [permission() for permission in permission_classes]
    
    def perform_create(self, serializer):
        org = serializer.save()
        create_audit_log(
            user=self.request.user,
            action='create',
            resource_type='Organization',
            resource_id=org.id,
            resource_repr=str(org),
            ip_address=self.request.META.get('REMOTE_ADDR'),
            user_agent=self.request.META.get('HTTP_USER_AGENT', '')
        )
    
    def perform_update(self, serializer):
        old_data = {
            'name': serializer.instance.name,
            'is_active': serializer.instance.is_active
        }
        org = serializer.save()
        create_audit_log(
            user=self.request.user,
            action='update',
            resource_type='Organization',
            resource_id=org.id,
            resource_repr=str(org),
            changes={'old': old_data, 'new': serializer.data},
            ip_address=self.request.META.get('REMOTE_ADDR'),
            user_agent=self.request.META.get('HTTP_USER_AGENT', '')
        )


class ModuleViewSet(viewsets.ModelViewSet):
    """
    ViewSet for managing application modules
    Admins can view modules, only super admin can create/edit/delete

    SOFT-CODED AUTO-SYNC: get_queryset() lazily provisions any module defined
    in rbac_config.ALL_MODULES_CATALOGUE that is missing from the DB. This
    means adding a new sub-feature's entry to ALL_MODULES_CATALOGUE makes it
    appear in Module Access (Admin ▸ Roles) immediately on next page load —
    no manual seed command or migration required. Purely additive/idempotent:
    existing Module rows, role-module assignments, and permissions are never
    modified or removed by this sync.
    """
    queryset = Module.objects.all()
    serializer_class = ModuleSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter, DjangoFilterBackend]
    search_fields = ['name', 'code']
    ordering_fields = ['order', 'name']
    filterset_fields = ['is_active']
    pagination_class = None  # Disable pagination - modules are a small dataset

    def get_queryset(self):
        self._sync_catalogue_modules()
        return super().get_queryset()

    @staticmethod
    def _sync_catalogue_modules():
        """
        Idempotently create any ALL_MODULES_CATALOGUE entry that doesn't yet
        exist in the DB (matched by code). Delegates to the shared helper in
        models.py (also used by UserProfile.get_all_modules() super_admin
        bypass) so the two call sites never drift.
        """
        from apps.rbac.models import _sync_module_catalogue
        _sync_module_catalogue()

    def get_permissions(self):
        """
        Allow admins to read modules, only super admin can modify
        """
        if self.action in ['list', 'retrieve', 'active']:
            permission_classes = [IsAuthenticated, IsAdmin]
        else:
            permission_classes = [IsAuthenticated, IsSuperAdmin]
        return [permission() for permission in permission_classes]
    
    @action(detail=False, methods=['get'], permission_classes=[IsAuthenticated, IsAdmin])
    def active(self, request):
        """Get all active modules"""
        modules = self.get_queryset().filter(is_active=True).order_by('order', 'name')
        serializer = self.get_serializer(modules, many=True)
        return Response(serializer.data)


class PermissionViewSet(viewsets.ModelViewSet):
    """
    ViewSet for managing permissions
    Only super admin can create/edit permissions
    """
    queryset = Permission.objects.select_related('module').all()
    serializer_class = PermissionSerializer
    permission_classes = [IsAuthenticated, IsSuperAdmin]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter, DjangoFilterBackend]
    search_fields = ['name', 'code']
    ordering_fields = ['module__name', 'action', 'name']
    filterset_fields = ['module', 'action', 'is_active']
    
    @action(detail=False, methods=['get'], permission_classes=[IsAuthenticated])
    def by_module(self, request):
        """Get permissions grouped by module"""
        module_id = request.query_params.get('module_id')
        if module_id:
            permissions = Permission.objects.filter(
                module_id=module_id,
                is_active=True
            )
        else:
            permissions = Permission.objects.filter(is_active=True)
        
        serializer = self.get_serializer(permissions, many=True)
        return Response(serializer.data)


class RoleViewSet(viewsets.ModelViewSet):
    """
    ViewSet for managing roles
    Only super admin can create/edit roles
    """
    # SOFT-CODED: custom_role_prefix from rbac_config — roles with this prefix are
    # per-user auto-generated roles and must not appear in the Role Management UI.
    # Change the prefix constant in rbac_config.py if the naming scheme ever changes.
    from apps.rbac.rbac_config import MODULE_ASSIGNMENT_CONFIG as _mac
    _CUSTOM_PREFIX = _mac.get('custom_role_prefix', 'custom_')

    # user_count is computed in RoleSerializer.get_user_count() from the
    # prefetched 'user_profiles' relation below, not via annotate(Count(...)) —
    # that annotation caused a PostgreSQL "column must appear in the GROUP BY
    # clause" error in production.
    queryset = Role.objects.prefetch_related('permissions', 'modules', 'modules__permissions', 'user_profiles') \
                           .filter(is_active=True) \
                           .exclude(code__startswith=_CUSTOM_PREFIX)
    permission_classes = [IsAuthenticated, CanManageRoles]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter, DjangoFilterBackend]
    search_fields = ['name', 'code']
    ordering_fields = ['level', 'name']
    filterset_fields = ['level', 'is_active']
    
    def get_serializer_class(self):
        # Use full serializer for all actions — roles are a small dataset (~10-20 rows)
        # and the list endpoint needs modules + user_count for the Role Management UI
        return RoleSerializer

    def get_permissions(self):
        # Anyone who can manage users can also view the roles list (same
        # broader check — role code OR user_mgmt module OR users.manage
        # permission — instead of the stricter role-code-only IsAdmin).
        if self.action in ['list', 'retrieve']:
            return [IsAuthenticated(), CanManageUsers()]
        if self.action in ['assign_module', 'revoke_module',
                           'assign_permission', 'revoke_permission']:
            return [IsAuthenticated(), IsAdmin()]
        return [IsAuthenticated(), CanManageRoles()]

    @staticmethod
    def _lock_from_auto_sync(role):
        """
        Once an admin manually edits a policy-governed role's modules/permissions,
        exclude it from future deploy-time ROLE_MODULE_POLICY resyncs
        (sync_role_modules) so the manual change isn't silently reverted.
        """
        from apps.rbac.rbac_config import ROLE_MODULE_POLICY
        if role.code in ROLE_MODULE_POLICY and role.auto_sync_enabled:
            role.auto_sync_enabled = False
            role.save(update_fields=['auto_sync_enabled'])

    def perform_create(self, serializer):
        role = serializer.save()
        create_audit_log(
            user=self.request.user,
            action='create',
            resource_type='Role',
            resource_id=role.id,
            resource_repr=str(role),
            ip_address=self.request.META.get('REMOTE_ADDR'),
            user_agent=self.request.META.get('HTTP_USER_AGENT', '')
        )
    
    def perform_update(self, serializer):
        role = serializer.save()
        create_audit_log(
            user=self.request.user,
            action='update',
            resource_type='Role',
            resource_id=role.id,
            resource_repr=str(role),
            ip_address=self.request.META.get('REMOTE_ADDR'),
            user_agent=self.request.META.get('HTTP_USER_AGENT', '')
        )
    
    def perform_destroy(self, instance):
        if instance.is_system_role:
            raise serializers.ValidationError("Cannot delete system roles")
        
        create_audit_log(
            user=self.request.user,
            action='delete',
            resource_type='Role',
            resource_id=instance.id,
            resource_repr=str(instance),
            ip_address=self.request.META.get('REMOTE_ADDR'),
            user_agent=self.request.META.get('HTTP_USER_AGENT', '')
        )
        instance.delete()

    @action(detail=True, methods=['post'])
    def duplicate(self, request, pk=None):
        """
        Duplicate a role — copies its module and permission grants under a
        new name/code. Super-admin only (enforced by get_permissions()).

        The duplicate is always a plain, freely-editable custom role
        (is_system_role=False), even when cloning a system role like Admin —
        this is the soft-coded way to let a super admin build a custom
        variant of any existing role without touching the original.

        POST /api/v1/rbac/roles/{id}/duplicate/
        Body: { "name": "New Role Name", "description": "optional" }
        """
        from django.utils.text import slugify

        source = self.get_object()
        new_name = (request.data.get('name') or '').strip()
        if not new_name:
            return Response(
                {'error': 'name is required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        if Role.objects.filter(name=new_name).exists():
            return Response(
                {'error': f'A role named "{new_name}" already exists.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Soft-coded unique code generation — slugify the name, disambiguate
        # with a numeric suffix on collision (mirrors custom_role_prefix style).
        base_code = slugify(new_name).replace('-', '_') or 'role'
        code = base_code
        suffix = 1
        while Role.objects.filter(code=code).exists():
            suffix += 1
            code = f'{base_code}_{suffix}'

        new_role = Role.objects.create(
            name=new_name,
            code=code,
            description=request.data.get('description', source.description),
            level=source.level,
            is_active=True,
            is_system_role=False,
            auto_sync_enabled=True,
        )

        for role_module in RoleModule.objects.filter(role=source):
            RoleModule.objects.create(
                role=new_role, module=role_module.module, granted_by=request.user
            )
        for role_permission in RolePermission.objects.filter(role=source):
            RolePermission.objects.create(
                role=new_role, permission=role_permission.permission, granted_by=request.user
            )

        create_audit_log(
            user=request.user,
            action='create',
            resource_type='Role',
            resource_id=new_role.id,
            resource_repr=str(new_role),
            metadata={'duplicated_from': source.code},
            ip_address=request.META.get('REMOTE_ADDR'),
            user_agent=request.META.get('HTTP_USER_AGENT', ''),
        )

        serializer = self.get_serializer(new_role)
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=['post'])
    def assign_permission(self, request, pk=None):
        """Assign permission to role"""
        role = self.get_object()
        permission_id = request.data.get('permission_id')
        
        if not permission_id:
            return Response(
                {'error': 'permission_id is required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            permission = Permission.objects.get(id=permission_id)
            RolePermission.objects.get_or_create(
                role=role,
                permission=permission,
                defaults={'granted_by': request.user}
            )
            self._lock_from_auto_sync(role)
            
            create_audit_log(
                user=request.user,
                action='permission_grant',
                resource_type='Role',
                resource_id=role.id,
                resource_repr=str(role),
                metadata={'permission': permission.code},
                ip_address=request.META.get('REMOTE_ADDR'),
                user_agent=request.META.get('HTTP_USER_AGENT', '')
            )
            
            return Response({'status': 'permission assigned'})
        except Permission.DoesNotExist:
            return Response(
                {'error': 'Permission not found'},
                status=status.HTTP_404_NOT_FOUND
            )
    
    @action(detail=True, methods=['post'])
    def revoke_permission(self, request, pk=None):
        """Revoke permission from role"""
        role = self.get_object()
        permission_id = request.data.get('permission_id')
        
        if not permission_id:
            return Response(
                {'error': 'permission_id is required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        deleted_count = RolePermission.objects.filter(
            role=role,
            permission_id=permission_id
        ).delete()[0]
        
        if deleted_count > 0:
            self._lock_from_auto_sync(role)
            create_audit_log(
                user=request.user,
                action='permission_revoke',
                resource_type='Role',
                resource_id=role.id,
                resource_repr=str(role),
                metadata={'permission_id': str(permission_id)},
                ip_address=request.META.get('REMOTE_ADDR'),
                user_agent=request.META.get('HTTP_USER_AGENT', '')
            )
        
        return Response({'status': 'permission revoked', 'count': deleted_count})
    
    @action(detail=True, methods=['post'])
    def assign_module(self, request, pk=None):
        """Assign module to role"""
        role = self.get_object()
        module_id = request.data.get('module_id')
        
        if not module_id:
            return Response(
                {'error': 'module_id is required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            module = Module.objects.get(id=module_id)
            RoleModule.objects.get_or_create(
                role=role,
                module=module,
                defaults={'granted_by': request.user}
            )
            self._lock_from_auto_sync(role)
            
            return Response({'status': 'module assigned'})
        except Module.DoesNotExist:
            return Response(
                {'error': 'Module not found'},
                status=status.HTTP_404_NOT_FOUND
            )
    
    @action(detail=True, methods=['post'])
    def revoke_module(self, request, pk=None):
        """Revoke module from role"""
        role = self.get_object()
        module_id = request.data.get('module_id')
        
        if not module_id:
            return Response(
                {'error': 'module_id is required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        deleted_count = RoleModule.objects.filter(
            role=role,
            module_id=module_id
        ).delete()[0]
        
        if deleted_count > 0:
            self._lock_from_auto_sync(role)
        
        return Response({'status': 'module revoked', 'count': deleted_count})

    @action(detail=False, methods=['post'], url_path='sync-default-role',
            permission_classes=[IsAuthenticated, CanManageRoles])
    def sync_default_role(self, request):
        """
        Assign the Default role to every active UserProfile that has no active
        role assignment.  Idempotent — safe to call multiple times.

        Returns:
            assigned (int)  — profiles that received the Default role now
            skipped  (int)  — profiles that already had at least one active role
            total    (int)  — total profiles inspected
        """
        from apps.rbac.rbac_config import DEFAULT_ROLE_CONFIG

        default_role_code = DEFAULT_ROLE_CONFIG['code']

        # Resolve the Default role
        try:
            default_role = Role.objects.get(code=default_role_code, is_active=True)
        except Role.DoesNotExist:
            return Response(
                {'error': f"Default role (code='{default_role_code}') not found or inactive."},
                status=status.HTTP_404_NOT_FOUND,
            )

        # Profiles that already have at least one active role
        profiles_with_role_ids = set(
            UserRole.objects.filter(role__is_active=True)
                            .values_list('user_profile_id', flat=True)
                            .distinct()
        )
        roleless_profiles = UserProfile.objects.exclude(id__in=profiles_with_role_ids)

        assigned = 0
        skipped  = len(profiles_with_role_ids)

        for profile in roleless_profiles.iterator():
            _, created = UserRole.objects.get_or_create(
                user_profile=profile,
                role=default_role,
                defaults={'is_primary': True, 'assigned_by': request.user},
            )
            if created:
                assigned += 1

        create_audit_log(
            user=request.user,
            action='bulk_assign_default_role',
            resource_type='Role',
            resource_id=default_role.id,
            resource_repr=default_role.name,
            metadata={'assigned': assigned, 'already_had_role': skipped},
            ip_address=request.META.get('REMOTE_ADDR'),
            user_agent=request.META.get('HTTP_USER_AGENT', ''),
        )

        return Response({
            'status': 'ok',
            'assigned': assigned,
            'skipped':  skipped,
            'total':    assigned + skipped,
            'message':  (
                f"Assigned Default role to {assigned} user(s). "
                f"{skipped} user(s) already had a role."
            ),
        })

    @action(detail=False, methods=['post'], url_path='flush-module-caches',
            permission_classes=[IsAuthenticated, CanManageRoles])
    def flush_module_caches(self, request):
        """
        Flush cached module/permission lists for ALL users.
        Call this after changing role membership, deactivating roles, or deploying
        RBAC fixes so that every user's next API call rebuilds from fresh DB data.
        Super-admin only. Idempotent and non-destructive.
        """
        from django.core.cache import cache

        profile_ids = UserProfile.objects.values_list('id', flat=True)
        cleared = 0
        for pid in profile_ids:
            cache.delete(f'user_modules_{pid}')
            cache.delete(f'user_permissions_{pid}')
            cleared += 1

        create_audit_log(
            user=request.user,
            action='flush_module_caches',
            resource_type='Role',
            resource_id=None,
            resource_repr='all-users',
            metadata={'profiles_cleared': cleared},
            ip_address=request.META.get('REMOTE_ADDR'),
            user_agent=request.META.get('HTTP_USER_AGENT', ''),
        )

        return Response({
            'status': 'ok',
            'profiles_cleared': cleared,
            'message': (
                f"Module & permission caches cleared for {cleared} user profile(s). "
                f"Next login / sidebar load will fetch fresh module data."
            ),
        })

    @action(detail=False, methods=['get'], url_path='access-diagnostics',
            permission_classes=[IsAuthenticated, IsSuperAdmin])
    def access_diagnostics(self, request):
        """
        Read-only diagnostic — inspect why a user may not be seeing an expected
        module, without needing direct database access. Super-admin only.

        GET /api/v1/rbac/roles/access-diagnostics/?role_code=rad403&user_email=lira.viaga@rejlers.ae
        """
        role_code = request.query_params.get('role_code', '').strip()
        user_email = request.query_params.get('user_email', '').strip()
        result = {}

        if role_code:
            role = Role.objects.filter(code=role_code).first()
            if not role:
                result['role'] = {'error': f"No role found with code '{role_code}'"}
            else:
                granted_modules = list(
                    RoleModule.objects.filter(role=role).values_list('module__code', 'module__is_active')
                )
                result['role'] = {
                    'code': role.code,
                    'name': role.name,
                    'is_active': role.is_active,
                    'is_system_role': role.is_system_role,
                    'auto_sync_enabled': role.auto_sync_enabled,
                    'granted_modules': [
                        {'code': code, 'module_is_active': active} for code, active in granted_modules
                    ],
                    'assigned_user_count': UserRole.objects.filter(role=role).count(),
                }

        if user_email:
            from django.contrib.auth import get_user_model
            User = get_user_model()
            user = User.objects.filter(email__iexact=user_email).first()
            if not user:
                result['user'] = {'error': f"No user found with email '{user_email}'"}
            else:
                profile = UserProfile.objects.filter(user=user).first()
                if not profile:
                    result['user'] = {'error': 'User has no rbac_profile'}
                else:
                    roles = UserRole.objects.filter(user_profile=profile).select_related('role')
                    from apps.rbac.rbac_config import is_module_enabled
                    accessible = profile.get_all_modules()
                    result['user'] = {
                        'email': user.email,
                        'is_superuser': user.is_superuser,
                        'is_active_profile': profile.status,
                        'assigned_roles': [
                            {
                                'code': ur.role.code, 'name': ur.role.name,
                                'is_primary': ur.is_primary, 'role_is_active': ur.role.is_active,
                            } for ur in roles
                        ],
                        'computed_accessible_modules': sorted(m.code for m in accessible),
                    }
                    if role_code:
                        result['user']['has_target_role'] = any(
                            ur.role.code == role_code for ur in roles
                        )

        if not role_code and not user_email:
            return Response(
                {'error': 'Provide role_code and/or user_email query params'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        return Response(result)


class UserProfileViewSet(viewsets.ModelViewSet):
    """
    ViewSet for managing user profiles with performance optimizations
    Admin can manage users in their organization
    Super admin can manage all users
    
    Performance Features:
    - Optimized queries (66% fewer DB queries)
    - Redis caching for large list requests (5min TTL)
    - Response time: <2s for 276 users (cached)
    
    Flexible Pagination:
    - GET /api/v1/rbac/users/ - Returns 10 users (default)
    - GET /api/v1/rbac/users/?page_size=25 - Returns 25 users
    - GET /api/v1/rbac/users/?page_size=100 - Returns 100 users
    - GET /api/v1/rbac/users/?page_size=1000 - Returns all users (cached)
    """
    permission_classes = [IsAuthenticated, CanManageUsers]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter, DjangoFilterBackend]
    search_fields = ['user__email', 'user__first_name', 'user__last_name', 'employee_id']
    ordering_fields = ['created_at', 'user__email', 'last_login_at']
    filterset_fields = ['organization', 'status', 'is_deleted']
    
    # Use custom pagination for flexible page sizes
    pagination_class = FlexiblePageNumberPagination
    
    def get_permissions(self):
        """
        Permission matrix:
        - me / change_password / engineers  → authentication only
        - create / reset_password / activate / deactivate / soft_delete
          / assign_role / revoke_role       → Super Admin only
        - Everything else (list, retrieve, partial_update) → Admin or Super Admin
        """
        SUPER_ADMIN_ONLY_ACTIONS = {
            'create', 'reset_password', 'activate', 'deactivate', 'soft_delete',
            'bulk_deactivate_by_roles', 'total_count',
        }
        # Admin (level 2+) can assign/revoke roles — but the action itself guards
        # against assigning the super_admin role without super_admin privileges
        ADMIN_ACTIONS = {'assign_role', 'revoke_role'}
        AUTH_ONLY_ACTIONS = {'me', 'profile_completeness', 'change_password', 'engineers'}

        if self.action in AUTH_ONLY_ACTIONS:
            return [IsAuthenticated()]
        if self.action in SUPER_ADMIN_ONLY_ACTIONS:
            return [IsAuthenticated(), IsSuperAdmin()]
        if self.action in ADMIN_ACTIONS:
            return [IsAuthenticated(), IsAdmin()]
        return [IsAuthenticated(), CanManageUsers()]
    
    def get_queryset(self):
        """
        Filter users based on role with optimized query performance
        
        Optimization Strategy:
        - select_related: Fetch user and organization in single query (JOIN)
        - prefetch_related: Fetch roles and userrole_set efficiently
        - Reduces N+1 query problem from 276+ queries to ~3 queries
        """
        user = self.request.user
        queryset = UserProfile.objects.select_related(
            'user', 'organization', 'manager__user',
            'user__employee_master', 'user__employee_master__manager',
            'user__employee_master__manager__user__rbac_profile',
        ).filter(is_deleted=False)

        if self.action == 'list':
            # The employee directory only needs a small subset of the very wide
            # EmployeeMaster and OffboardingRecord tables. Fetching every column
            # made the first /hr/employees request take ~50 seconds against the
            # remote database. Keep list responses lean; retrieve still returns
            # the complete profile when the user opens the detail drawer.
            from apps.onboarding.models import OffboardingRecord

            queryset = queryset.only(
                'id', 'created_at', 'updated_at', 'user_id', 'organization_id',
                'manager_id', 'canonical_employee_id', 'employee_id', 'department', 'job_title', 'phone',
                'location', 'bio', 'status', 'is_mfa_enabled', 'profile_photo',
                'last_login_at', 'is_deleted',
                'user__id', 'user__username', 'user__email', 'user__first_name',
                'user__last_name', 'user__is_active', 'user__is_staff',
                'user__is_superuser',
                'organization__id', 'organization__name',
                'manager__id', 'manager__job_title', 'manager__department',
                'manager__user__id', 'manager__user__email',
                'manager__user__first_name', 'manager__user__last_name',
                'user__employee_master__id', 'user__employee_master__user_id',
                'user__employee_master__join_date', 'user__employee_master__exit_date',
                'user__employee_master__probation_end_date',
                'user__employee_master__employment_status',
                'user__employee_master__manager_id',
                'user__employee_master__manager__id',
                'user__employee_master__manager__email',
                'user__employee_master__manager__first_name',
                'user__employee_master__manager__last_name',
                'user__employee_master__manager__designation',
                'user__employee_master__manager__job_title_uae',
                'user__employee_master__manager__department',
                'user__employee_master__manager__user__rbac_profile__id',
            ).prefetch_related(
                Prefetch(
                    'userrole_set',
                    queryset=UserRole.objects.select_related('role').only(
                        'id', 'user_profile_id', 'role_id', 'is_primary',
                        'role__id', 'role__name', 'role__code', 'role__level',
                        'role__is_active',
                    ),
                ),
                Prefetch(
                    'user__offboarding_records',
                    queryset=OffboardingRecord.objects.only(
                        'id', 'user_id', 'exit_reason', 'last_working_day',
                    ),
                ),
            )
        else:
            queryset = queryset.prefetch_related(
                'roles',
                'userrole_set__role',
                'user__offboarding_records',
            )

        # Super admin sees all
        try:
            profile = user.rbac_profile
            if not profile.roles.filter(code='super_admin', is_active=True).exists():
                # Other admins see only their organization
                queryset = queryset.filter(organization=profile.organization)
        except UserProfile.DoesNotExist:
            return UserProfile.objects.none()
        
        # Optional: filter by role code — used by Role Management UI
        role_code = self.request.query_params.get('role')
        if role_code:
            queryset = queryset.filter(roles__code=role_code, roles__is_active=True)

        # Keep pagination stable so records cannot move between API pages.
        return queryset.order_by('user__first_name', 'user__last_name', 'user__email')

    def get_serializer_class(self):
        if self.action == 'list':
            return UserProfileListSerializer
        return UserProfileSerializer

    @transaction.atomic
    def perform_create(self, serializer):
        from apps.rbac.rbac_config import DEFAULT_ROLE_CONFIG
        profile = serializer.save()
        from apps.hr_core.services import EmployeeService
        EmployeeService.sync_from_rbac_profile(profile)
        create_audit_log(
            user=self.request.user,
            action='create',
            resource_type='UserProfile',
            resource_id=profile.id,
            resource_repr=str(profile),
            ip_address=self.request.META.get('REMOTE_ADDR'),
            user_agent=self.request.META.get('HTTP_USER_AGENT', '')
        )
        # Auto-assign the Default role when no role has been set yet
        if not profile.userrole_set.filter(role__is_active=True).exists():
            try:
                default_role = Role.objects.get(
                    code=DEFAULT_ROLE_CONFIG['code'], is_active=True
                )
                UserRole.objects.get_or_create(
                    user_profile=profile,
                    role=default_role,
                    defaults={'is_primary': True, 'assigned_by': self.request.user},
                )
            except Role.DoesNotExist:
                pass
    
    @transaction.atomic
    def perform_update(self, serializer):
        profile = serializer.save()
        from apps.hr_core.services import EmployeeService
        EmployeeService.sync_from_rbac_profile(profile, self.request.data.keys())
        create_audit_log(
            user=self.request.user,
            action='update',
            resource_type='UserProfile',
            resource_id=profile.id,
            resource_repr=str(profile),
            ip_address=self.request.META.get('REMOTE_ADDR'),
            user_agent=self.request.META.get('HTTP_USER_AGENT', '')
        )
        # ── Two-way sync: when Reporting Manager changes on the Profile page,
        # propagate the name to the linked OnboardingRecord (if any) and the
        # EmployeeMaster record so all three systems stay aligned. ─────────────
        if 'manager_id' in self.request.data or 'manager' in self.request.data:
            manager_name = profile.manager.user.get_full_name() if profile.manager else ''
            try:
                from apps.onboarding.models import OnboardingRecord
                OnboardingRecord.objects.filter(user=profile.user).update(
                    reporting_manager=manager_name
                )
            except Exception:
                pass

    @action(detail=True, methods=['post'])
    def deactivate(self, request, pk=None):
        """Deactivate user"""
        profile = self.get_object()
        profile.status = 'inactive'
        profile.user.is_active = False
        profile.save()
        profile.user.save()
        from apps.hr_core.services import EmployeeService
        EmployeeService.sync_from_rbac_profile(profile, {'status', 'is_active'})
        
        create_audit_log(
            user=request.user,
            action='update',
            resource_type='UserProfile',
            resource_id=profile.id,
            resource_repr=str(profile),
            metadata={'status_change': 'active -> inactive'},
            ip_address=request.META.get('REMOTE_ADDR'),
            user_agent=request.META.get('HTTP_USER_AGENT', '')
        )
        
        return Response({'status': 'user deactivated'})
    
    @action(detail=True, methods=['post'])
    def activate(self, request, pk=None):
        """Activate user"""
        profile = self.get_object()
        profile.status = 'active'
        profile.user.is_active = True
        profile.save()
        profile.user.save()
        from apps.hr_core.services import EmployeeService
        EmployeeService.sync_from_rbac_profile(profile, {'status', 'is_active'})
        
        create_audit_log(
            user=request.user,
            action='update',
            resource_type='UserProfile',
            resource_id=profile.id,
            resource_repr=str(profile),
            metadata={'status_change': 'inactive -> active'},
            ip_address=request.META.get('REMOTE_ADDR'),
            user_agent=request.META.get('HTTP_USER_AGENT', '')
        )
        
        return Response({'status': 'user activated'})
    
    @action(detail=True, methods=['delete'])
    def soft_delete(self, request, pk=None):
        """Soft delete user"""
        profile = self.get_object()
        profile.is_deleted = True
        profile.deleted_at = timezone.now()
        profile.deleted_by = request.user
        profile.user.is_active = False
        profile.save()
        profile.user.save()
        from apps.hr_core.services import EmployeeService
        EmployeeService.sync_from_rbac_profile(profile, {'status', 'is_active'})
        
        create_audit_log(
            user=request.user,
            action='delete',
            resource_type='UserProfile',
            resource_id=profile.id,
            resource_repr=str(profile),
            ip_address=request.META.get('REMOTE_ADDR'),
            user_agent=request.META.get('HTTP_USER_AGENT', '')
        )
        
        return Response({'status': 'user soft deleted'})
    
    @action(detail=False, methods=['post'], url_path='bulk-deactivate-by-roles')
    def bulk_deactivate_by_roles(self, request):
        """
        Bulk deactivate users based on their roles
        
        SOFT-CODED: All thresholds, messages, and validation rules are configurable
        
        Request body:
        {
            "role_codes": ["engineer", "contractor"],  // OR role_ids
            "role_ids": [1, 2, 3],                     // Optional alternative
            "exclude_super_admins": true,              // Soft-coded safety flag (default: true)
            "dry_run": false                           // Preview mode (default: false)
        }
        
        Returns:
        {
            "status": "success",
            "deactivated_count": 25,
            "affected_users": [...],  // if dry_run=true
            "excluded_count": 2,       // super admins excluded
            "message": "25 users deactivated successfully"
        }
        """
        # SOFT-CODED: Configuration constants
        MAX_BULK_DEACTIVATE_LIMIT = 500  # Safety limit
        EXCLUDE_SUPER_ADMINS_BY_DEFAULT = True
        SUPER_ADMIN_ROLE_CODES = ['super_admin', 'superadmin']
        
        # Extract request data
        role_codes = request.data.get('role_codes', [])
        role_ids = request.data.get('role_ids', [])
        exclude_super_admins = request.data.get('exclude_super_admins', EXCLUDE_SUPER_ADMINS_BY_DEFAULT)
        dry_run = request.data.get('dry_run', False)
        
        # Validation
        if not role_codes and not role_ids:
            return Response(
                {'error': 'Either role_codes or role_ids must be provided'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Build queryset - find all users with any of the specified roles
        user_queryset = UserProfile.objects.filter(
            status='active',  # Only deactivate currently active users
            user__is_active=True
        ).distinct()
        
        # Filter by role codes or IDs
        if role_codes:
            user_queryset = user_queryset.filter(
                roles__code__in=role_codes,
                roles__is_active=True
            )
        elif role_ids:
            user_queryset = user_queryset.filter(
                roles__id__in=role_ids,
                roles__is_active=True
            )
        
        # SOFT-CODED: Exclude super admins if requested (safety feature)
        excluded_count = 0
        if exclude_super_admins:
            super_admin_users = user_queryset.filter(
                roles__code__in=SUPER_ADMIN_ROLE_CODES
            ).distinct()
            excluded_count = super_admin_users.count()
            user_queryset = user_queryset.exclude(
                id__in=super_admin_users.values_list('id', flat=True)
            )
        
        # Check count limit (safety)
        total_count = user_queryset.count()
        if total_count > MAX_BULK_DEACTIVATE_LIMIT:
            return Response(
                {
                    'error': f'Cannot deactivate more than {MAX_BULK_DEACTIVATE_LIMIT} users at once',
                    'requested_count': total_count,
                    'limit': MAX_BULK_DEACTIVATE_LIMIT
                },
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # DRY RUN MODE - return preview without making changes
        if dry_run:
            affected_users = []
            for profile in user_queryset.select_related('user').prefetch_related('roles'):
                affected_users.append({
                    'id': str(profile.id),
                    'email': profile.user.email,
                    'name': f"{profile.user.first_name} {profile.user.last_name}".strip() or profile.user.email,
                    'department': profile.department,
                    'roles': [r.name for r in profile.roles.filter(is_active=True)]
                })
            
            return Response({
                'status': 'dry_run',
                'would_deactivate_count': total_count,
                'excluded_super_admin_count': excluded_count,
                'affected_users': affected_users,
                'message': f'Preview: {total_count} users would be deactivated'
            })
        
        # ACTUAL DEACTIVATION
        deactivated_count = 0
        deactivated_user_ids = []
        
        for profile in user_queryset.select_related('user'):
            # Deactivate user
            profile.status = 'inactive'
            profile.user.is_active = False
            profile.save()
            profile.user.save()
            from apps.hr_core.services import EmployeeService
            EmployeeService.sync_from_rbac_profile(profile, {'status', 'is_active'})
            
            deactivated_count += 1
            deactivated_user_ids.append(str(profile.id))
            
            # Create audit log for each deactivation
            create_audit_log(
                user=request.user,
                action='bulk_deactivate',
                resource_type='UserProfile',
                resource_id=profile.id,
                resource_repr=str(profile),
                metadata={
                    'bulk_operation': True,
                    'role_codes': role_codes,
                    'role_ids': role_ids,
                    'excluded_super_admins': exclude_super_admins
                },
                ip_address=request.META.get('REMOTE_ADDR'),
                user_agent=request.META.get('HTTP_USER_AGENT', '')
            )
        
        # SOFT-CODED: Success message
        message = f'{deactivated_count} user{"s" if deactivated_count != 1 else ""} deactivated successfully'
        if excluded_count > 0:
            message += f' ({excluded_count} super admin{"s" if excluded_count != 1 else ""} excluded)'
        
        return Response({
            'status': 'success',
            'deactivated_count': deactivated_count,
            'excluded_super_admin_count': excluded_count,
            'deactivated_user_ids': deactivated_user_ids,
            'message': message
        }, status=status.HTTP_200_OK)
    
    @action(detail=True, methods=['post'], url_path='reset-password')
    def reset_password(self, request, pk=None):
        """
        Reset user password to default password
        Admin-only action for security
        """
        from django.contrib.auth.hashers import make_password
        from django.utils import timezone
        from django.conf import settings
        
        profile = self.get_object()
        user = profile.user
        
        # Default password (soft-coded in settings)
        default_password = getattr(settings, 'DEFAULT_USER_PASSWORD', 'Rejlers@123')
        
        # Set the password
        user.password = make_password(default_password)
        user.last_password_change = timezone.now()
        user.must_reset_password = True
        user.is_first_login = False
        user.save()
        
        # Set must_change_password flag
        profile.must_change_password = True
        profile.save()
        
        # Log the action
        create_audit_log(
            user=request.user,
            action='reset_password',
            resource_type='User',
            resource_id=user.id,
            resource_repr=f'{user.email}',
            changes={'reset_by': request.user.email, 'must_change_password': True},
            ip_address=request.META.get('REMOTE_ADDR'),
            user_agent=request.META.get('HTTP_USER_AGENT', '')
        )
        
        return Response({
            'status': 'password reset successfully',
            'message': f'Password has been reset to default. User must change it on next login.',
            'default_password': default_password
        })
    
    @action(detail=False, methods=['post'], url_path='change-password')
    def change_password(self, request):
        """
        Change user's own password
        Required fields: old_password, new_password
        Clears must_change_password flag on success
        """
        from django.contrib.auth.hashers import check_password, make_password
        from django.utils import timezone
        
        user = request.user
        old_password = request.data.get('old_password')
        new_password = request.data.get('new_password')
        
        # Validation
        if not old_password or not new_password:
            return Response(
                {'error': 'old_password and new_password are required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Verify old password
        if not check_password(old_password, user.password):
            return Response(
                {'error': 'Invalid old password'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Password strength validation (basic)
        if len(new_password) < 8:
            return Response(
                {'error': 'Password must be at least 8 characters long'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Update password
        user.password = make_password(new_password)
        user.last_password_change = timezone.now()
        user.must_reset_password = False
        user.is_first_login = False
        user.save()
        
        # Clear must_change_password flag
        try:
            profile = user.rbac_profile  # Fixed: use rbac_profile instead of profile
            if profile.must_change_password:
                profile.must_change_password = False
                profile.save()
        except Exception as e:
            print(f"Warning: Could not clear must_change_password flag: {e}")
        
        # Log the action
        create_audit_log(
            user=user,
            action='change_password',
            resource_type='User',
            resource_id=user.id,
            resource_repr=f'{user.email}',
            changes={'password_changed': True, 'must_change_password_cleared': True},
            ip_address=request.META.get('REMOTE_ADDR'),
            user_agent=request.META.get('HTTP_USER_AGENT', '')
        )
        
        return Response({
            'status': 'password changed successfully',
            'message': 'Your password has been updated'
        })
    
    @action(detail=True, methods=['post'])
    def assign_role(self, request, pk=None):
        """Assign role to user.

        Super Admins can assign any role including super_admin.
        Admins can assign any role except super_admin.
        
        SOFT-CODED BEHAVIOR (from rbac_config.py):
        - Automatically removes ALL custom_* roles before assignment
        - Sets new role as primary if is_primary=True (default)
        - Clears cached modules/permissions for immediate effect
        - Follows ROLE_MODULE_POLICY for module access
        """
        from apps.rbac.rbac_config import MODULE_ASSIGNMENT_CONFIG
        
        profile = self.get_object()
        role_id = request.data.get('role_id')
        is_primary = request.data.get('is_primary', True)  # Default to primary

        if not role_id:
            return Response(
                {'error': 'role_id is required'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            role = Role.objects.get(id=role_id)
        except Role.DoesNotExist:
            return Response(
                {'error': 'Role not found'},
                status=status.HTTP_404_NOT_FOUND
            )

        # Guard: only super admin may assign the super_admin role
        PROTECTED_ROLE_CODES = {'super_admin'}
        if role.code in PROTECTED_ROLE_CODES:
            requester_is_super_admin = (
                request.user.is_superuser
                or (
                    hasattr(request.user, 'rbac_profile')
                    and request.user.rbac_profile.roles.filter(
                        code='super_admin', is_active=True
                    ).exists()
                )
            )
            if not requester_is_super_admin:
                return Response(
                    {'error': 'Only Super Administrators can assign the Super Admin role.'},
                    status=status.HTTP_403_FORBIDDEN
                )

        # ── CRITICAL FIX: Remove all custom_* roles first ──────────────────────
        # Prevents stale per-user custom roles from giving unintended access
        # after admin assigns a proper system role.
        # SOFT-CODED: custom role prefix from MODULE_ASSIGNMENT_CONFIG
        custom_role_prefix = MODULE_ASSIGNMENT_CONFIG.get('custom_role_prefix', 'custom_')
        custom_roles_removed = profile.roles.filter(
            code__startswith=custom_role_prefix,
            is_active=True
        )
        removed_count = custom_roles_removed.count()
        removed_names = list(custom_roles_removed.values_list('name', flat=True))
        
        # Remove custom roles through the many-to-many relationship
        for custom_role in custom_roles_removed:
            profile.roles.remove(custom_role)
        
        if removed_count > 0:
            print(f"[RBAC] 🗑️  Removed {removed_count} custom role(s) for {profile.user.email}: {removed_names}")
        
        # ── Assign new role ─────────────────────────────────────────────────────
        user_role, created = UserRole.objects.get_or_create(
            user_profile=profile,
            role=role,
            defaults={'assigned_by': request.user, 'is_primary': is_primary}
        )
        
        # If role already existed, update is_primary if requested
        if not created and is_primary:
            # Demote all other roles to non-primary
            UserRole.objects.filter(user_profile=profile).exclude(id=user_role.id).update(is_primary=False)
            if not user_role.is_primary:
                user_role.is_primary = True
                user_role.save(update_fields=['is_primary'])

        # ── SOFT-CODED CACHE INVALIDATION ─────────────────────────────────────
        # Clear the user's cached modules and permissions so they immediately
        # see their new role's access rights without waiting for the 60s TTL.
        from django.core.cache import cache
        cache.delete(f'user_modules_{profile.id}')
        cache.delete(f'user_permissions_{profile.id}')
        print(f"[RBAC] ✅ Cleared module/permission cache for user {profile.user.email}")
        # ───────────────────────────────────────────────────────────────────────

        create_audit_log(
            user=request.user,
            action='role_assign',
            resource_type='UserProfile',
            resource_id=profile.id,
            resource_repr=str(profile),
            metadata={
                'role': role.name,
                'removed_custom_roles': removed_names if removed_count > 0 else None
            },
            ip_address=request.META.get('REMOTE_ADDR'),
            user_agent=request.META.get('HTTP_USER_AGENT', '')
        )

        return Response({
            'status': 'created' if created else 'already_exists',
            'role': role.name,
            'removed_custom_roles': removed_count
        })

    @action(detail=True, methods=['post'])
    def revoke_role(self, request, pk=None):
        """Revoke role from user.

        Super Admins can revoke any role including super_admin.
        Admins can revoke any role except super_admin.
        """
        profile = self.get_object()
        role_id = request.data.get('role_id')

        if not role_id:
            return Response(
                {'error': 'role_id is required'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Guard: only super admin may revoke the super_admin role
        PROTECTED_ROLE_CODES = {'super_admin'}
        try:
            target_role = Role.objects.get(id=role_id)
            if target_role.code in PROTECTED_ROLE_CODES:
                requester_is_super_admin = (
                    request.user.is_superuser
                    or (
                        hasattr(request.user, 'rbac_profile')
                        and request.user.rbac_profile.roles.filter(
                            code='super_admin', is_active=True
                        ).exists()
                    )
                )
                if not requester_is_super_admin:
                    return Response(
                        {'error': 'Only Super Administrators can revoke the Super Admin role.'},
                        status=status.HTTP_403_FORBIDDEN
                    )
        except Role.DoesNotExist:
            pass  # Role not found — the delete below will simply affect 0 rows

        deleted_count = UserRole.objects.filter(
            user_profile=profile,
            role_id=role_id
        ).delete()[0]
        
        if deleted_count > 0:
            # ── SOFT-CODED CACHE INVALIDATION ─────────────────────────────────────
            # Clear the user's cached modules and permissions so they immediately
            # lose access to revoked role's permissions without waiting for the 60s TTL.
            from django.core.cache import cache
            cache.delete(f'user_modules_{profile.id}')
            cache.delete(f'user_permissions_{profile.id}')
            print(f"[RBAC] ✅ Cleared module/permission cache for user {profile.user.email}")
            # ───────────────────────────────────────────────────────────────────────
            
            create_audit_log(
                user=request.user,
                action='role_revoke',
                resource_type='UserProfile',
                resource_id=profile.id,
                resource_repr=str(profile),
                metadata={'role_id': str(role_id)},
                ip_address=request.META.get('REMOTE_ADDR'),
                user_agent=request.META.get('HTTP_USER_AGENT', '')
            )
        
        return Response({'status': 'role revoked', 'count': deleted_count})

    @action(detail=True, methods=['post'])
    def set_primary_role(self, request, pk=None):
        """Mark a specific role as the primary role for this user.

        Sets the given role's UserRole.is_primary = True and all other
        UserRole entries for this user to is_primary = False.
        """
        profile = self.get_object()
        role_id = request.data.get('role_id')

        if not role_id:
            return Response(
                {'error': 'role_id is required'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            role = Role.objects.get(id=role_id, is_active=True)
        except Role.DoesNotExist:
            return Response(
                {'error': 'Role not found'},
                status=status.HTTP_404_NOT_FOUND
            )

        try:
            user_role = UserRole.objects.get(user_profile=profile, role=role)
        except UserRole.DoesNotExist:
            return Response(
                {'error': 'This role is not assigned to the user'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Demote all other roles for this user, then promote the target
        UserRole.objects.filter(user_profile=profile).update(is_primary=False)
        user_role.is_primary = True
        user_role.save(update_fields=['is_primary'])

        create_audit_log(
            user=request.user,
            action='role_set_primary',
            resource_type='UserProfile',
            resource_id=profile.id,
            resource_repr=str(profile),
            metadata={'role': role.name, 'role_id': str(role.id)},
            ip_address=request.META.get('REMOTE_ADDR'),
            user_agent=request.META.get('HTTP_USER_AGENT', '')
        )

        return Response({'status': 'primary role updated', 'role': role.name})

    @action(
        detail=False,
        methods=['post'],
        parser_classes=[MultiPartParser, FormParser],
        url_path='bulk_upload',
    )
    def bulk_upload(self, request):
        """
        Bulk upload users from CSV/Excel with Email Notifications
        Accepts CSV and Excel workbooks. Human-readable template headings are
        normalized to the API field names before validation.
        
        New Features:
        - Sends welcome email with credentials to each user
        - Aligned with registration form fields
        - Better error handling and reporting
        """
        import csv
        import io
        import json
        import re
        from openpyxl import load_workbook
        from django.contrib.auth import get_user_model
        from django.core.exceptions import ValidationError
        from django.core.validators import validate_email
        from django.db import transaction
        from django.utils.dateparse import parse_datetime
        from django.utils import timezone
        import secrets
        import string
        
        User = get_user_model()
        
        if 'file' not in request.FILES:
            return Response(
                {'error': 'No file provided'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        file = request.FILES['file']
        organization_id = request.data.get('organization_id')
        preview_only = str(request.data.get('preview_only', '')).lower() in {'1', 'true', 'yes'}
        try:
            requested_update_fields = json.loads(request.data.get('update_fields', '[]'))
        except (TypeError, json.JSONDecodeError):
            requested_update_fields = []
        allowed_update_fields = {
            'first_name', 'last_name', 'username', 'department', 'job_title',
            'phone', 'employee_id', 'location', 'status', 'role_codes',
            'organization_name', 'created_at', 'last_login_at',
        }
        update_fields = {field for field in requested_update_fields if field in allowed_update_fields}
        
        extension = file.name.lower().rsplit('.', 1)[-1] if '.' in file.name else ''
        if extension not in {'csv', 'txt', 'xlsx', 'xlsm'}:
            return Response(
                {'error': 'Invalid file format. Upload CSV or Excel (.xlsx/.xlsm).'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Get organization
        try:
            organization = Organization.objects.get(id=organization_id) if organization_id else None
        except Organization.DoesNotExist:
            return Response(
                {'error': 'Organization not found'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        def normalize_header(value):
            key = re.sub(r'[^a-z0-9]+', '_', str(value or '').strip().lower()).strip('_')
            return {
                'roles': 'role_codes',
                'organization': 'organization_name',
                'created_at': 'created_at',
                'last_login': 'last_login_at',
            }.get(key, key)

        def parse_optional_datetime(value):
            if not value:
                return None
            if hasattr(value, 'tzinfo') and hasattr(value, 'year'):
                parsed = value
            else:
                parsed = parse_datetime(str(value).strip())
            if parsed and timezone.is_naive(parsed):
                parsed = timezone.make_aware(parsed, timezone.get_current_timezone())
            return parsed

        # Parse CSV or the first worksheet in an Excel workbook.
        try:
            if extension in {'xlsx', 'xlsm'}:
                workbook = load_workbook(file, read_only=True, data_only=True)
                worksheet = workbook.active
                values = worksheet.iter_rows(values_only=True)
                raw_headers = next(values, None)
                if not raw_headers:
                    raise ValueError('The Excel worksheet is empty.')
                headers = [normalize_header(header) for header in raw_headers]
                reader = [
                    {
                        header: (value.isoformat() if hasattr(value, 'isoformat') else str(value).strip()) if value is not None else ''
                        for header, value in zip(headers, row)
                        if header
                    }
                    for row in values
                    if any(value not in (None, '') for value in row)
                ]
            else:
                decoded_file = file.read().decode('utf-8-sig')
                csv_reader = csv.DictReader(io.StringIO(decoded_file))
                reader = [
                    {normalize_header(key): value for key, value in row.items() if key is not None}
                    for row in csv_reader
                ]

            # Build lookup indexes once. This keeps a 500-row preview to a
            # handful of queries instead of several queries per spreadsheet row.
            existing_profiles = UserProfile.objects.select_related('user', 'organization').prefetch_related('userrole_set__role').filter(is_deleted=False)
            profiles_by_email = {profile.user.email.strip().casefold(): profile for profile in existing_profiles if profile.user.email}
            profiles_by_employee_id = {profile.employee_id.strip().casefold(): profile for profile in existing_profiles if profile.employee_id}
            profiles_by_username = {profile.user.username.strip().casefold(): profile for profile in existing_profiles if profile.user.username}
            username_owners = {
                username.strip().casefold(): user_id
                for user_id, username in User.objects.exclude(username='').values_list('id', 'username')
                if username and username.strip()
            }
            organizations = list(Organization.objects.all())
            organizations_by_name = {item.name.strip().casefold(): item for item in organizations}
            organizations_by_code = {item.code.strip().casefold(): item for item in organizations}
            default_organization = next((item for item in organizations if 'default' in item.name.casefold()), None) or (organizations[0] if organizations else None)
            active_roles = list(Role.objects.filter(is_active=True))
            roles_by_code = {item.code.strip().casefold(): item for item in active_roles}
            roles_by_name = {item.name.strip().casefold(): item for item in active_roles}

            def find_organization(value):
                key = str(value or '').strip().casefold()
                return organizations_by_name.get(key) or organizations_by_code.get(key)

            def find_role(value):
                key = str(value or '').strip().casefold()
                return roles_by_code.get(key) or roles_by_name.get(key)

            def find_existing_profile(row):
                email = row.get('email', '').strip()
                employee_id = row.get('employee_id', '').strip()
                username = row.get('username', '').strip()
                if email:
                    match = profiles_by_email.get(email.casefold())
                    if match:
                        return match, 'email'
                if employee_id:
                    match = profiles_by_employee_id.get(employee_id.casefold())
                    if match:
                        return match, 'employee_id'
                if username:
                    match = profiles_by_username.get(username.casefold())
                    if match:
                        return match, 'username'
                return None, None

            preview_columns = [
                {'id': 'first_name', 'label': 'First Name'},
                {'id': 'last_name', 'label': 'Last Name'},
                {'id': 'username', 'label': 'Username'},
                {'id': 'department', 'label': 'Department'},
                {'id': 'job_title', 'label': 'Job Title'},
                {'id': 'phone', 'label': 'Phone'},
                {'id': 'employee_id', 'label': 'Employee ID'},
                {'id': 'location', 'label': 'Location'},
                {'id': 'status', 'label': 'Status'},
                {'id': 'role_codes', 'label': 'Roles'},
                {'id': 'organization_name', 'label': 'Organization'},
                {'id': 'created_at', 'label': 'Created At'},
                {'id': 'last_login_at', 'label': 'Last Login'},
            ]

            if preview_only:
                preview_rows = []
                counts = {'new': 0, 'existing': 0, 'invalid': 0}
                for row_num, row in enumerate(reader, start=2):
                    email = row.get('email', '').strip()
                    if not email:
                        counts['invalid'] += 1
                        preview_rows.append({'row': row_num, 'state': 'invalid', 'email': '', 'name': '', 'match_by': None, 'changed_fields': [], 'error': 'Email is required'})
                        continue
                    validation_errors = []
                    try:
                        validate_email(email)
                    except ValidationError:
                        validation_errors.append('Invalid email address')
                    incoming_status = row.get('status', '').strip().lower().replace(' ', '_')
                    if incoming_status and incoming_status not in {choice[0] for choice in UserProfile.STATUS_CHOICES}:
                        validation_errors.append(f'Invalid status "{incoming_status}"')
                    organization_name = row.get('organization_name', '').strip()
                    if organization_name and not find_organization(organization_name):
                        validation_errors.append(f'Organization "{organization_name}" was not found')
                    role_codes = row.get('role_codes', '').strip()
                    if role_codes:
                        for token in [item.strip() for item in re.split(r'[,;]', role_codes) if item.strip()]:
                            if not find_role(token):
                                validation_errors.append(f'Role "{token}" was not found')
                    if validation_errors:
                        counts['invalid'] += 1
                        preview_rows.append({'row': row_num, 'state': 'invalid', 'email': email, 'name': '', 'match_by': None, 'changed_fields': [], 'error': '; '.join(validation_errors)})
                        continue
                    existing, match_by = find_existing_profile(row)
                    state = 'existing' if existing else 'new'
                    counts[state] += 1
                    current = {}
                    if existing:
                        current = {
                            'first_name': existing.user.first_name,
                            'last_name': existing.user.last_name,
                            'username': existing.user.username,
                            'department': existing.department,
                            'job_title': existing.job_title,
                            'phone': existing.phone,
                            'employee_id': existing.employee_id,
                            'location': existing.location,
                            'status': existing.status,
                            'role_codes': ', '.join(link.role.code for link in existing.userrole_set.all() if link.role.is_active),
                            'organization_name': existing.organization.name if existing.organization else '',
                            'created_at': existing.created_at.isoformat() if existing.created_at else '',
                            'last_login_at': existing.last_login_at.isoformat() if existing.last_login_at else '',
                        }
                    incoming = {column['id']: row.get(column['id'], '').strip() for column in preview_columns}
                    changed_fields = [field for field, value in incoming.items() if value and str(current.get(field, '')).strip().casefold() != value.casefold()] if existing else []
                    preview_rows.append({
                        'row': row_num,
                        'state': state,
                        'email': email,
                        'name': f"{row.get('first_name', '').strip()} {row.get('last_name', '').strip()}".strip(),
                        'match_by': match_by,
                        'changed_fields': changed_fields,
                        'incoming': incoming,
                        'current': current,
                        'error': None,
                    })
                return Response({
                    'preview': True,
                    'columns': preview_columns,
                    'summary': {'total': len(preview_rows), **counts},
                    'rows': preview_rows,
                })
            
            results = {
                'success': [],
                'updated': [],
                'failed': [],
                'skipped': []
            }
            welcome_email_jobs = []
            users_to_update = {}
            profiles_to_update = {}
            user_update_fields = set()
            profile_update_fields = set()
            role_replacements = {}
            
            with transaction.atomic():
                for row_num, row in enumerate(reader, start=2):  # Start at 2 (1 is header)
                    savepoint = None
                    try:
                        # Validate required fields
                        email = row.get('email', '').strip()
                        if not email:
                            results['failed'].append({
                                'row': row_num,
                                'email': email or 'N/A',
                                'error': 'Email is required'
                            })
                            continue
                        try:
                            validate_email(email)
                        except ValidationError:
                            raise ValueError('Invalid email address')
                        
                        existing_profile, matched_by = find_existing_profile(row)
                        
                        if existing_profile:
                            if update_fields:
                                user = existing_profile.user
                                user_changed = []
                                profile_changed = []

                                for field_name in ('first_name', 'last_name'):
                                    incoming_value = row.get(field_name, '').strip()
                                    if field_name in update_fields and incoming_value and getattr(user, field_name) != incoming_value:
                                        setattr(user, field_name, incoming_value)
                                        user_changed.append(field_name)

                                incoming_username = row.get('username', '').strip()
                                if 'username' in update_fields and incoming_username and incoming_username.casefold() != user.username.casefold():
                                    incoming_username_key = incoming_username.casefold()
                                    username_owner_id = username_owners.get(incoming_username_key)
                                    if username_owner_id and username_owner_id != user.id:
                                        raise ValueError(f'Username "{incoming_username}" is already in use')
                                    current_username_key = user.username.strip().casefold()
                                    if username_owners.get(current_username_key) == user.id:
                                        username_owners.pop(current_username_key)
                                    username_owners[incoming_username_key] = user.id
                                    user.username = incoming_username
                                    user_changed.append('username')

                                profile_field_map = {
                                    'department': 'department', 'job_title': 'job_title',
                                    'phone': 'phone', 'employee_id': 'employee_id',
                                    'location': 'location',
                                }
                                for upload_field, model_field in profile_field_map.items():
                                    incoming_value = row.get(upload_field, '').strip()
                                    if upload_field in update_fields and incoming_value and getattr(existing_profile, model_field) != incoming_value:
                                        setattr(existing_profile, model_field, incoming_value)
                                        profile_changed.append(model_field)

                                incoming_status = row.get('status', '').strip().lower().replace(' ', '_')
                                if 'status' in update_fields and incoming_status:
                                    allowed_statuses = {choice[0] for choice in UserProfile.STATUS_CHOICES}
                                    if incoming_status not in allowed_statuses:
                                        raise ValueError(f'Invalid status "{incoming_status}"')
                                    existing_profile.status = incoming_status
                                    profile_changed.append('status')
                                    user.is_active = incoming_status not in {'inactive', 'suspended'}
                                    if 'is_active' not in user_changed:
                                        user_changed.append('is_active')

                                organization_name = row.get('organization_name', '').strip()
                                if 'organization_name' in update_fields and organization_name:
                                    matched_organization = find_organization(organization_name)
                                    if not matched_organization:
                                        raise ValueError(f'Organization "{organization_name}" was not found')
                                    existing_profile.organization = matched_organization
                                    profile_changed.append('organization')

                                role_codes = row.get('role_codes', '').strip()
                                if 'role_codes' in update_fields and role_codes:
                                    role_tokens = [token.strip() for token in re.split(r'[,;]', role_codes) if token.strip()]
                                    resolved_roles = []
                                    unresolved_roles = []
                                    for token in role_tokens:
                                        role = find_role(token)
                                        if role and role not in resolved_roles:
                                            resolved_roles.append(role)
                                        elif not role:
                                            unresolved_roles.append(token)
                                    if unresolved_roles:
                                        raise ValueError(f'Unknown role(s): {", ".join(unresolved_roles)}')
                                    existing_role_ids = {link.role_id for link in existing_profile.userrole_set.all() if link.role.is_active}
                                    resolved_role_ids = {role.id for role in resolved_roles}
                                    if existing_role_ids != resolved_role_ids:
                                        role_replacements[existing_profile.id] = (existing_profile, resolved_roles)

                                timestamp_updates = {}
                                if 'created_at' in update_fields:
                                    created_at = parse_optional_datetime(row.get('created_at'))
                                    if created_at:
                                        timestamp_updates['created_at'] = created_at
                                if 'last_login_at' in update_fields:
                                    last_login_at = parse_optional_datetime(row.get('last_login_at'))
                                    if last_login_at:
                                        timestamp_updates['last_login_at'] = last_login_at
                                        user.last_login = last_login_at
                                        user_changed.append('last_login')
                                for field_name, field_value in timestamp_updates.items():
                                    setattr(existing_profile, field_name, field_value)
                                    profile_changed.append(field_name)

                                if user_changed:
                                    users_to_update[user.id] = user
                                    user_update_fields.update(user_changed)
                                if profile_changed:
                                    existing_profile.updated_at = timezone.now()
                                    profiles_to_update[existing_profile.id] = existing_profile
                                    profile_update_fields.update(profile_changed)
                                    profile_update_fields.add('updated_at')

                                results['updated'].append({
                                    'row': row_num,
                                    'email': email,
                                    'name': user.get_full_name(),
                                    'matched_by': matched_by,
                                    'fields': sorted(update_fields),
                                })
                                continue
                            results['skipped'].append({
                                'row': row_num,
                                'email': email,
                                'reason': 'User already exists'
                            })
                            continue
                        
                        # New records perform writes immediately, so isolate each
                        # row behind a savepoint. Existing updates are batched.
                        savepoint = transaction.savepoint()
                        # Respect an uploaded username, then make it unique when needed.
                        base_username = row.get('username', '').strip() or email.split('@')[0]
                        username = base_username
                        counter = 1
                        while username.casefold() in username_owners:
                            username = f"{base_username}{counter}"
                            counter += 1
                        
                        # Get or generate password
                        password = row.get('password', '').strip()
                        if not password:
                            # Generate secure random password
                            alphabet = string.ascii_letters + string.digits + "!@#$%^&*"
                            password = ''.join(secrets.choice(alphabet) for i in range(12))
                            # Ensure at least one of each type
                            password = secrets.choice(string.ascii_uppercase) + \
                                      secrets.choice(string.ascii_lowercase) + \
                                      secrets.choice(string.digits) + \
                                      secrets.choice("!@#$%^&*") + \
                                      password[4:]
                        
                        # Store original password for email
                        original_password = password
                        
                        # Create user
                        user = User.objects.create_user(
                            username=username,
                            email=email,
                            first_name=row.get('first_name', '').strip(),
                            last_name=row.get('last_name', '').strip(),
                            password=password
                        )
                        
                        row_organization = organization
                        organization_name = row.get('organization_name', '').strip()
                        if organization_name:
                            row_organization = find_organization(organization_name)
                            if not row_organization:
                                raise ValueError(f'Organization "{organization_name}" was not found')
                        if not row_organization:
                            row_organization = default_organization
                        if not row_organization:
                            raise ValueError('No organization is available for this employee')

                        uploaded_status = row.get('status', '').strip().lower().replace(' ', '_') or 'active'
                        allowed_statuses = {choice[0] for choice in UserProfile.STATUS_CHOICES}
                        if uploaded_status not in allowed_statuses:
                            raise ValueError(f'Invalid status "{uploaded_status}"')
                        user.is_active = uploaded_status not in {'inactive', 'suspended'}
                        user.save(update_fields=['is_active'])

                        profile = UserProfile.objects.create(
                            user=user,
                            organization=row_organization,
                            department=row.get('department', '').strip(),
                            job_title=row.get('job_title', '').strip(),
                            phone=row.get('phone', '').strip() or row.get('phone_number', '').strip(),
                            employee_id=row.get('employee_id', '').strip(),
                            location=row.get('location', '').strip(),
                            status=uploaded_status,
                        )

                        created_at = parse_optional_datetime(row.get('created_at'))
                        last_login_at = parse_optional_datetime(row.get('last_login_at'))
                        timestamp_updates = {}
                        if created_at:
                            timestamp_updates['created_at'] = created_at
                        if last_login_at:
                            timestamp_updates['last_login_at'] = last_login_at
                            user.last_login = last_login_at
                            user.save(update_fields=['last_login'])
                        if timestamp_updates:
                            UserProfile.objects.filter(pk=profile.pk).update(**timestamp_updates)
                            for field_name, field_value in timestamp_updates.items():
                                setattr(profile, field_name, field_value)
                        
                        # Assign roles
                        role_codes = row.get('role_codes', '').strip()
                        if role_codes:
                            role_tokens = [token.strip() for token in re.split(r'[,;]', role_codes) if token.strip()]
                            roles = []
                            unresolved_roles = []
                            for token in role_tokens:
                                role = find_role(token)
                                if role and role not in roles:
                                    roles.append(role)
                                elif not role:
                                    unresolved_roles.append(token)
                            if unresolved_roles:
                                raise ValueError(f'Unknown role(s): {", ".join(unresolved_roles)}')
                            for idx, role in enumerate(roles):
                                UserRole.objects.create(
                                    user_profile=profile,
                                    role=role,
                                    assigned_by=request.user,
                                    is_primary=(idx == 0)
                                )
                        else:
                            # No role specified — auto-assign the Default role
                            from apps.rbac.rbac_config import DEFAULT_ROLE_CONFIG
                            try:
                                default_role = Role.objects.get(
                                    code=DEFAULT_ROLE_CONFIG['code'], is_active=True
                                )
                                UserRole.objects.get_or_create(
                                    user_profile=profile,
                                    role=default_role,
                                    defaults={'assigned_by': request.user, 'is_primary': True},
                                )
                            except Role.DoesNotExist:
                                pass
                        
                        # Assign modules
                        module_codes = row.get('module_codes', '').strip()
                        if module_codes:
                            module_code_list = [m.strip() for m in module_codes.split(',')]
                            modules = Module.objects.filter(code__in=module_code_list, is_active=True)
                            for module in modules:
                                profile.modules.add(module)
                        
                        results['success'].append({
                            'row': row_num,
                            'email': email,
                            'name': f"{user.first_name} {user.last_name}".strip(),
                            'username': username
                        })
                        username_owners[username.casefold()] = user.id
                        
                        results['success'][-1]['email_sent'] = False
                        results['success'][-1]['email_queued'] = False
                        welcome_email_jobs.append((str(user.id), original_password, len(results['success']) - 1))
                        transaction.savepoint_commit(savepoint)
                        
                    except Exception as e:
                        if savepoint is not None:
                            if transaction.get_rollback():
                                transaction.set_rollback(False)
                            transaction.savepoint_rollback(savepoint)
                        results['failed'].append({
                            'row': row_num,
                            'email': row.get('email', 'N/A'),
                            'error': str(e)
                        })

                if users_to_update and user_update_fields:
                    User.objects.bulk_update(list(users_to_update.values()), sorted(user_update_fields), batch_size=500)
                if profiles_to_update and profile_update_fields:
                    UserProfile.objects.bulk_update(list(profiles_to_update.values()), sorted(profile_update_fields), batch_size=500)
                if role_replacements:
                    replacement_profile_ids = list(role_replacements)
                    UserRole.objects.filter(user_profile_id__in=replacement_profile_ids).delete()
                    replacement_links = []
                    for profile, replacement_roles in role_replacements.values():
                        replacement_links.extend(
                            UserRole(user_profile=profile, role=role, assigned_by=request.user, is_primary=index == 0)
                            for index, role in enumerate(replacement_roles)
                        )
                    UserRole.objects.bulk_create(replacement_links, batch_size=1000)
            
            # Queue welcome mail only after the database transaction commits.
            # SMTP latency must never hold the import HTTP request open.
            if welcome_email_jobs:
                try:
                    from apps.rbac.tasks import send_bulk_welcome_email
                    for user_id, password, result_index in welcome_email_jobs:
                        send_bulk_welcome_email.delay(user_id, password)
                        results['success'][result_index]['email_queued'] = True
                except Exception as queue_error:
                    logger.warning('Could not queue bulk welcome emails: %s', queue_error)

            # Calculate email statistics
            emails_sent = sum(1 for item in results['success'] if item.get('email_sent', False))
            emails_failed = sum(1 for item in results['success'] if not item.get('email_sent', False) and not item.get('email_queued', False))
            emails_queued = sum(1 for item in results['success'] if item.get('email_queued', False))
            
            # Create summary audit log
            create_audit_log(
                user=request.user,
                action='bulk_upload',
                resource_type='UserProfile',
                resource_id=None,
                resource_repr='Bulk User Upload',
                metadata={
                    'success_count': len(results['success']),
                    'updated_count': len(results['updated']),
                    'failed_count': len(results['failed']),
                    'skipped_count': len(results['skipped']),
                    'emails_sent': emails_sent,
                    'emails_queued': emails_queued,
                    'emails_failed': emails_failed
                },
                ip_address=request.META.get('REMOTE_ADDR'),
                user_agent=request.META.get('HTTP_USER_AGENT', '')
            )
            
            return Response({
                'message': 'Bulk upload completed successfully!',
                'summary': {
                    'total_processed': len(results['success']) + len(results['updated']) + len(results['failed']) + len(results['skipped']),
                    'successful': len(results['success']),
                    'created': len(results['success']),
                    'updated': len(results['updated']),
                    'failed': len(results['failed']),
                    'skipped': len(results['skipped']),
                    'emails_sent': emails_sent,
                    'emails_queued': emails_queued,
                    'emails_failed': emails_failed
                },
                'details': results
            }, status=status.HTTP_201_CREATED)
            
        except Exception as e:
            return Response(
                {'error': f'Failed to process file: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=False, methods=['get'])
    def download_template(self, request):
        """
        Download the Excel employee bulk-upload template.
        """
        from io import BytesIO
        from django.http import HttpResponse
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.worksheet.datavalidation import DataValidation

        headers = [
            'First Name', 'Last Name', 'Email', 'Username', 'Department',
            'Job Title', 'Phone', 'Employee ID', 'Location', 'Status', 'Roles',
            'Organization', 'Created At', 'Last Login',
        ]
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = 'Employees'
        worksheet.append(headers)
        worksheet.freeze_panes = 'A2'
        worksheet.auto_filter.ref = f'A1:N1'

        header_fill = PatternFill('solid', fgColor='1E3A8A')
        for index, cell in enumerate(worksheet[1], start=1):
            cell.fill = header_fill
            cell.font = Font(color='FFFFFF', bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            worksheet.column_dimensions[cell.column_letter].width = 18 if index not in {3, 12, 13, 14} else 25
        worksheet.row_dimensions[1].height = 24

        status_validation = DataValidation(type='list', formula1='"active,inactive,suspended,pending,on_leave"', allow_blank=True)
        status_validation.promptTitle = 'Employee status'
        status_validation.prompt = 'Choose an approved account status. Blank defaults to active.'
        worksheet.add_data_validation(status_validation)
        status_validation.add('J2:J5000')

        instructions = workbook.create_sheet('Instructions')
        instructions.append(['Column', 'Guidance'])
        guidance = {
            'First Name': 'Employee first name.',
            'Last Name': 'Employee last name.',
            'Email': 'Required and must be unique.',
            'Username': 'Optional. Blank uses the email prefix; duplicates receive a numeric suffix.',
            'Status': 'Optional: active, inactive, suspended, pending, or on_leave.',
            'Roles': 'Optional role codes or names separated by commas or semicolons.',
            'Organization': 'Optional existing organization name or code. Blank uses the default organization.',
            'Created At': 'Optional ISO date/time, for example 2026-08-19T09:00:00+04:00.',
            'Last Login': 'Optional ISO date/time.',
        }
        for header in headers:
            instructions.append([header, guidance.get(header, 'Optional employee profile value.')])
        instructions.freeze_panes = 'A2'
        instructions.column_dimensions['A'].width = 22
        instructions.column_dimensions['B'].width = 85
        for cell in instructions[1]:
            cell.fill = header_fill
            cell.font = Font(color='FFFFFF', bold=True)

        output = BytesIO()
        workbook.save(output)
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="employee_bulk_upload_template.xlsx"'
        return response

    @action(detail=False, methods=['get', 'patch'], url_path='me')
    def me(self, request):
        """Get or update current user's profile"""
        import traceback
        
        try:
            # PATCH - Update profile
            if request.method == 'PATCH':
                return self.update_my_profile(request)
            
            # GET - Retrieve profile
            # Log the request for debugging
            print(f"\n[DEBUG /rbac/users/me/] User: {request.user}")
            print(f"[DEBUG /rbac/users/me/] User authenticated: {request.user.is_authenticated}")
            print(f"[DEBUG /rbac/users/me/] User email: {getattr(request.user, 'email', 'N/A')}")
            
            profile_view = request.query_params.get('view') == 'profile'
            profile_queryset = UserProfile.objects.select_related(
                'user', 'organization', 'manager__user', 'engineer_profile'
            )
            if not profile_view:
                # Access-control consumers need the full role graph. The
                # Profile screen deliberately skips it via ?view=profile.
                profile_queryset = profile_queryset.prefetch_related(
                    'roles',
                    'roles__permissions',
                    'userrole_set__role'
                )

            profile = profile_queryset.filter(
                user=request.user,
                is_deleted=False
            ).first()
            
            print(f"[DEBUG /rbac/users/me/] Profile found: {profile is not None}")
            
            # If no profile exists, return user info without RBAC data
            if not profile:
                print(f"[DEBUG /rbac/users/me/] No RBAC profile for {request.user.email}")
                return Response({
                    'id': str(request.user.id),
                    'email': request.user.email,
                    'username': request.user.username,
                    'first_name': request.user.first_name,
                    'last_name': request.user.last_name,
                    'roles': [],
                    'organization': None,
                    'status': 'pending',
                    'message': 'RBAC profile not configured. Please contact administrator.'
                })
            
            # Try to serialize the profile
            print(f"[DEBUG /rbac/users/me/] Serializing profile...")
            serializer = (
                UserProfileSelfSerializer(profile, context={'request': request})
                if profile_view
                else self.get_serializer(profile)
            )
            data = serializer.data
            print(f"[DEBUG /rbac/users/me/] Serialization successful")
            print(f"[DEBUG /rbac/users/me/] Roles count: {len(data.get('roles', []))}")
            print(f"[DEBUG /rbac/users/me/] Phone: {data.get('phone')}, Profile Photo: {data.get('profile_photo')}")
            
            return Response(data)
            
        except Exception as e:
            # Log the full error for debugging
            print(f"\n[ERROR /rbac/users/me/] Exception occurred: {str(e)}")
            print(f"[ERROR /rbac/users/me/] Exception type: {type(e).__name__}")
            print(f"[ERROR /rbac/users/me/] Traceback:")
            traceback.print_exc()
            
            # Return safe fallback data to avoid breaking the UI
            return Response({
                'id': str(getattr(request.user, 'id', '')),
                'email': getattr(request.user, 'email', ''),
                'username': getattr(request.user, 'username', ''),
                'first_name': getattr(request.user, 'first_name', ''),
                'last_name': getattr(request.user, 'last_name', ''),
                'roles': [],
                'organization': None,
                'status': 'pending',
                'message': 'Profile temporarily unavailable'
            })

    @action(detail=False, methods=['get'], url_path='me/profile-completeness')
    def profile_completeness(self, request):
        """Return completion details for the authenticated user's own profile."""
        from apps.rbac.profile_config import get_profile_completeness

        profile = UserProfile.objects.select_related('user', 'engineer_profile', 'canonical_employee').filter(
            user=request.user,
            is_deleted=False,
        ).first()
        if not profile:
            return Response({
                'percentage': 0,
                'is_complete': False,
                'missing_fields': [{
                    'key': 'employee_profile',
                    'label': 'Employee profile record - contact HR',
                    'section': 'Profile setup',
                    'weight': 100,
                }],
                'completed_fields': 0,
                'total_fields': 1,
                'detail': 'Your employee profile has not been configured. Contact HR.',
                'profile_url': '/profile',
            })

        try:
            engineer_profile = profile.engineer_profile.to_dict()
        except Exception:
            engineer_profile = {}

        result = get_profile_completeness({
            'first_name': request.user.first_name,
            'last_name': request.user.last_name,
            'profile_photo': bool(
                profile.profile_photo
                or (profile.canonical_employee and profile.canonical_employee.photo_file_path)
            ),
            'phone': profile.phone,
            'location': profile.location,
            'bio': profile.bio,
            'department': profile.department,
            'job_title': profile.job_title,
            'engineer_profile': engineer_profile,
        })
        result['profile_url'] = '/profile'
        return Response(result)
    
    @transaction.atomic
    def update_my_profile(self, request):
        """Update current user's profile"""
        try:
            # Get user profile
            profile = UserProfile.objects.filter(
                user=request.user,
                is_deleted=False
            ).first()
            
            if not profile:
                return Response(
                    {'error': 'Profile not found'},
                    status=status.HTTP_404_NOT_FOUND
                )
            
            # Track changes for audit log
            changes = {}
            
            # Update User model fields
            user = request.user
            if 'first_name' in request.data:
                changes['first_name'] = request.data['first_name']
                user.first_name = request.data['first_name']
            if 'last_name' in request.data:
                changes['last_name'] = request.data['last_name']
                user.last_name = request.data['last_name']
            user.save()
            
            # Update UserProfile fields
            if 'phone' in request.data:
                changes['phone'] = request.data['phone']
                profile.phone = request.data['phone']
            if 'bio' in request.data:
                changes['bio'] = request.data['bio']
                profile.bio = request.data['bio']
            if 'location' in request.data:
                changes['location'] = request.data['location']
                profile.location = request.data['location']
            if 'department' in request.data:
                changes['department'] = request.data['department']
                profile.department = request.data['department']
            if 'job_title' in request.data:
                changes['job_title'] = request.data['job_title']
                profile.job_title = request.data['job_title']
            
            # Handle reporting manager assignment
            if 'manager_id' in request.data:
                manager_id = request.data['manager_id']
                if manager_id:
                    try:
                        manager_profile = UserProfile.objects.get(id=manager_id, is_deleted=False)
                        profile.manager = manager_profile
                        changes['manager'] = f"{manager_profile.user.get_full_name() or manager_profile.user.username} ({manager_id})"
                    except UserProfile.DoesNotExist:
                        pass  # Silently ignore invalid manager_id
                else:
                    # Empty string or null = clear the manager
                    profile.manager = None
                    changes['manager'] = 'cleared'
            
            # Handle profile photo upload
            if 'profile_photo' in request.FILES:
                profile.profile_photo = request.FILES['profile_photo']
                changes['profile_photo'] = 'uploaded'

            # Handle engineer_profile JSON — persisted to rbac_engineer_profiles table
            # Accepts both JSON body (dict) and FormData (JSON string)
            ep_raw = request.data.get('engineer_profile')
            if ep_raw is not None:
                import json as _json
                if isinstance(ep_raw, str):
                    try:
                        ep_raw = _json.loads(ep_raw)
                    except Exception:
                        ep_raw = None
                if isinstance(ep_raw, dict):
                    from apps.rbac.models import EngineerProfile
                    ep_obj, _ = EngineerProfile.objects.get_or_create(user_profile=profile)

                    def normalize_names(values, field_name, max_items):
                        if not isinstance(values, list):
                            raise ValueError(f'{field_name} must be a list.')
                        if len(values) > max_items:
                            raise ValueError(f'{field_name} cannot contain more than {max_items} entries.')
                        normalized = []
                        seen = set()
                        for value in values:
                            name = ' '.join(str(value or '').strip().split())
                            if not name or len(name) > 100:
                                raise ValueError(f'Each {field_name} entry must be between 1 and 100 characters.')
                            key = name.casefold()
                            if key not in seen:
                                normalized.append(name)
                                seen.add(key)
                        return normalized

                    try:
                        disciplines = normalize_names(
                            ep_raw.get('engineering_disciplines', ep_obj.engineering_disciplines),
                            'engineering discipline',
                            30,
                        )
                        raw_skills = ep_raw.get('technical_skills', ep_obj.technical_skills)
                        if not isinstance(raw_skills, list) or len(raw_skills) > 100:
                            raise ValueError('Technical skills must be a list containing no more than 100 entries.')
                        skills = []
                        seen_skills = set()
                        for skill in raw_skills:
                            if not isinstance(skill, dict):
                                raise ValueError('Each technical skill must contain a name and proficiency.')
                            name = ' '.join(str(skill.get('name', '')).strip().split())
                            if not name or len(name) > 100:
                                raise ValueError('Each technical skill name must be between 1 and 100 characters.')
                            key = name.casefold()
                            if key in seen_skills:
                                continue
                            proficiency = int(skill.get('proficiency', 3))
                            if proficiency < 1 or proficiency > 5:
                                raise ValueError('Technical skill proficiency must be between 1 and 5.')
                            skills.append({'name': name, 'proficiency': proficiency})
                            seen_skills.add(key)

                        raw_projects = ep_raw.get('current_projects', ep_obj.current_projects)
                        if not isinstance(raw_projects, list) or len(raw_projects) > 50:
                            raise ValueError('Current projects must be a list containing no more than 50 entries.')
                        from apps.hr_core.models import EmployeeMaster
                        pom_users = {
                            str(employee.user_id): employee.user
                            for employee in EmployeeMaster.objects.filter(
                                user__is_active=True,
                            ).select_related('user')
                        }
                        existing_projects = {
                            str(project.get('id')): project
                            for project in (ep_obj.current_projects or [])
                            if isinstance(project, dict) and project.get('id') not in (None, '')
                        }
                        used_project_ids = {
                            str(project.get('project_id')).strip()
                            for project in (ep_obj.current_projects or [])
                            if isinstance(project, dict) and project.get('project_id')
                        }
                        current_year = timezone.localdate().year

                        def next_project_id(project_name, project_year):
                            import re
                            suffix = re.compile(rf'-(\d{{4}})-{project_year}$')
                            sequences = []
                            for project_id in used_project_ids:
                                match = suffix.search(project_id)
                                if match:
                                    sequences.append(int(match.group(1)))
                            sequence = max(sequences, default=0) + 1
                            generated = f'{project_name}-{sequence:04d}-{project_year}'
                            while generated in used_project_ids:
                                sequence += 1
                                generated = f'{project_name}-{sequence:04d}-{project_year}'
                            used_project_ids.add(generated)
                            return generated

                        projects = []
                        for project in raw_projects:
                            if not isinstance(project, dict):
                                raise ValueError('Each current project must be an object.')
                            normalized_project = dict(project)
                            project_name = ' '.join(str(normalized_project.get('name', '')).strip().split())
                            if not project_name or len(project_name) > 150:
                                raise ValueError('Each project name must be between 1 and 150 characters.')
                            normalized_project['name'] = project_name
                            existing_project = existing_projects.get(str(normalized_project.get('id')))
                            pom_id = normalized_project.get('project_manager_id')
                            if pom_id not in (None, ''):
                                pom_user = pom_users.get(str(pom_id))
                                if not pom_user:
                                    existing_pom_id = (
                                        str(existing_project.get('project_manager_id'))
                                        if existing_project else ''
                                    )
                                    if existing_pom_id != str(pom_id):
                                        raise ValueError('The selected Project Manager is not an active employee.')
                                    normalized_project['project_manager_name'] = existing_project.get(
                                        'project_manager_name',
                                        normalized_project.get('project_manager_name', ''),
                                    )
                                    normalized_project['project_manager_email'] = existing_project.get(
                                        'project_manager_email',
                                        normalized_project.get('project_manager_email', ''),
                                    )
                                else:
                                    normalized_project['project_manager_id'] = pom_user.id
                                    normalized_project['project_manager_name'] = (
                                        pom_user.get_full_name() or pom_user.email or pom_user.username
                                    )
                                    normalized_project['project_manager_email'] = pom_user.email or ''
                            existing_project_id = (
                                str(existing_project.get('project_id')).strip()
                                if existing_project and existing_project.get('project_id') else ''
                            )
                            if existing_project_id:
                                normalized_project['project_id'] = existing_project_id
                            else:
                                start_date = str(normalized_project.get('start_date') or '')
                                project_year = (
                                    int(start_date[:4])
                                    if len(start_date) >= 4 and start_date[:4].isdigit()
                                    else current_year
                                )
                                normalized_project['project_id'] = next_project_id(
                                    project_name,
                                    project_year,
                                )
                            projects.append(normalized_project)
                    except (TypeError, ValueError) as validation_error:
                        return Response(
                            {'engineer_profile': [str(validation_error)]},
                            status=status.HTTP_400_BAD_REQUEST,
                        )

                    ep_obj.expertise_level          = ep_raw.get('expertise_level', ep_obj.expertise_level)
                    ep_obj.years_experience         = int(ep_raw.get('years_experience') or ep_obj.years_experience or 0)
                    ep_obj.engineering_disciplines  = disciplines
                    ep_obj.technical_skills         = skills
                    ep_obj.languages                = ep_raw.get('languages', ep_obj.languages)
                    ep_obj.certifications           = ep_raw.get('certifications', ep_obj.certifications)
                    ep_obj.availability_status      = ep_raw.get('availability_status', ep_obj.availability_status)
                    ep_obj.availability_percentage  = int(ep_raw.get('availability_percentage') or ep_obj.availability_percentage or 100)
                    ep_obj.next_available_date      = ep_raw.get('next_available_date') or None
                    ep_obj.max_concurrent_projects  = int(ep_raw.get('max_concurrent_projects') or ep_obj.max_concurrent_projects or 2)
                    ep_obj.preferred_project_types  = ep_raw.get('preferred_project_types', ep_obj.preferred_project_types)
                    ep_obj.current_projects         = projects
                    ep_obj.save()
                    changes['engineer_profile'] = 'updated'

            profile.save()
            
            # Create audit log (only with serializable data)
            create_audit_log(
                user=request.user,
                action='update_profile',
                resource_type='UserProfile',
                resource_id=profile.id,
                resource_repr=f'{user.email}',
                changes=changes,
                ip_address=request.META.get('REMOTE_ADDR'),
                user_agent=request.META.get('HTTP_USER_AGENT', '')
            )
            
            # Return updated profile
            serializer = (
                UserProfileSelfSerializer(profile, context={'request': request})
                if request.query_params.get('view') == 'profile'
                else self.get_serializer(profile)
            )
            response_data = serializer.data
            print(f"[DEBUG] Profile response - phone: {response_data.get('phone')}, profile_photo: {response_data.get('profile_photo')}")
            return Response(response_data)
            
        except Exception as e:
            import traceback
            print(f"[ERROR] Failed to update profile: {str(e)}")
            print(traceback.format_exc())
            return Response(
                {'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=False, methods=['get'])
    def stats(self, request):
        """Get user statistics"""
        queryset = self.get_queryset()
        
        stats = {
            'total_users': queryset.count(),
            'active_users': queryset.filter(status='active').count(),
            'inactive_users': queryset.filter(status='inactive').count(),
            'suspended_users': queryset.filter(status='suspended').count(),
            'by_organization': list(queryset.values('organization__name').annotate(count=Count('id'))),
        }
        
        return Response(stats)
    
    @action(detail=False, methods=['get'], url_path='department-choices')
    def department_choices(self, request):
        """
        Get list of department choices for dropdown
        Soft-coded constants for Oil & Gas engineering organization
        """
        from .constants import get_department_choices
        return Response({
            'departments': get_department_choices(),
            'count': len(get_department_choices())
        })
    
    @action(detail=False, methods=['get'])
    def my_features(self, request):
        """Get list of features current user has access to"""
        from .utils import get_user_accessible_features
        
        features = get_user_accessible_features(request.user)
        return Response({
            'features': list(features.values()),
            'accessible_count': sum(1 for f in features.values() if f['accessible'])
        })

    @action(detail=False, methods=['get'])
    def engineers(self, request):
        """
        List engineers in the same organisation with their competency profiles.
        Used for project team building and assignment matching.

        Query params (all optional):
          discipline     — filter by engineering discipline (case-insensitive contains)
          expertise_level — filter by level code e.g. senior, lead
          available_only  — 'true' to show only available / partial engineers
          skill           — filter by technical skill name (contains)
        """
        try:
            profile = request.user.rbac_profile
        except UserProfile.DoesNotExist:
            return Response({'engineers': [], 'count': 0})

        queryset = UserProfile.objects.select_related(
            'user', 'engineer_profile'
        ).filter(
            is_deleted=False,
            status='active',
            organization=profile.organization,
        )

        # ── Query-param filters ────────────────────────────────────────────
        discipline     = (request.query_params.get('discipline') or '').strip().lower()
        expertise_lvl  = (request.query_params.get('expertise_level') or '').strip().lower()
        available_only = request.query_params.get('available_only', '').lower() == 'true'
        skill_filter   = (request.query_params.get('skill') or '').strip().lower()

        engineers_out = []
        for up in queryset:
            try:
                ep_obj = up.engineer_profile
                ep = ep_obj.to_dict()
            except Exception:
                ep = {}

            # Availability filter
            if available_only and ep.get('availability_status', 'available') not in ('available', 'partial'):
                continue

            # Expertise level filter
            if expertise_lvl and ep.get('expertise_level', '').lower() != expertise_lvl:
                continue

            # Discipline filter (any discipline contains the search string)
            if discipline:
                disciplines_lower = [d.lower() for d in ep.get('engineering_disciplines', [])]
                if not any(discipline in d for d in disciplines_lower):
                    continue

            # Skill filter
            if skill_filter:
                skills_lower = [s.get('name', '').lower() for s in ep.get('technical_skills', [])]
                if not any(skill_filter in s for s in skills_lower):
                    continue

            # Build profile photo URL
            photo_url = None
            if up.profile_photo:
                try:
                    url = up.profile_photo.url
                    if not url.startswith('http'):
                        url = request.build_absolute_uri(url)
                    photo_url = url
                except Exception:
                    pass

            engineers_out.append({
                'id':                    str(up.id),
                'name':                  f"{up.user.first_name} {up.user.last_name}".strip() or up.user.email,
                'email':                 up.user.email,
                'job_title':             up.job_title,
                'department':            up.department,
                'location':              up.location,
                'profile_photo':         photo_url,
                'expertise_level':       ep.get('expertise_level', ''),
                'years_experience':      ep.get('years_experience', ''),
                'engineering_disciplines': ep.get('engineering_disciplines', []),
                'technical_skills':      ep.get('technical_skills', []),
                'certifications':        ep.get('certifications', []),
                'availability_status':   ep.get('availability_status', 'available'),
                'availability_percentage': ep.get('availability_percentage', 100),
                'preferred_project_types': ep.get('preferred_project_types', []),
                'languages':             ep.get('languages', []),
            })

        return Response({'engineers': engineers_out, 'count': len(engineers_out)})

    @action(detail=True, methods=['post'], url_path='assign-modules')
    def assign_modules(self, request, pk=None):
        """
        Assign modules to a user by updating their role's module access
        Body: { "module_codes": ["pid_analysis", "pfd", "qhse"] }
        """
        profile = self.get_object()
        module_codes = request.data.get('module_codes', [])

        if not module_codes:
            return Response(
                {'error': 'module_codes is required (array of module codes)'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Get modules
        modules = Module.objects.filter(code__in=module_codes, is_active=True)
        if modules.count() != len(module_codes):
            found_codes = set(modules.values_list('code', flat=True))
            missing_codes = set(module_codes) - found_codes
            return Response(
                {'error': f'Some modules not found: {list(missing_codes)}'},
                status=status.HTTP_404_NOT_FOUND
            )

        # Get user's primary role (or create custom role)
        user_roles = UserRole.objects.filter(user_profile=profile, is_primary=True)

        if not user_roles.exists():
            user_roles = UserRole.objects.filter(user_profile=profile)

        if not user_roles.exists():
            return Response(
                {'error': 'User has no roles assigned. Please assign a role first.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        assigned_count = 0
        for user_role in user_roles:
            role = user_role.role
            for module in modules:
                _, created = RoleModule.objects.get_or_create(
                    role=role,
                    module=module,
                    defaults={'granted_by': request.user}
                )
                if created:
                    assigned_count += 1

        create_audit_log(
            user=request.user,
            action='modules_assign',
            resource_type='UserProfile',
            resource_id=profile.id,
            resource_repr=str(profile),
            metadata={'module_codes': module_codes, 'assigned_count': assigned_count},
            ip_address=request.META.get('REMOTE_ADDR'),
            user_agent=request.META.get('HTTP_USER_AGENT', '')
        )

        return Response({
            'status': 'modules assigned',
            'user': profile.user.email,
            'modules': [m.name for m in modules],
            'assigned_count': assigned_count
        })

    @action(detail=False, methods=['post'], url_path='bulk-assign-modules')
    def bulk_assign_modules(self, request):
        """
        Bulk assign modules to multiple users
        Body: {
            "user_ids": ["uuid1", "uuid2"],  // or "user_emails": ["email1", "email2"]
            "module_codes": ["pid_analysis", "pfd", "qhse"],
            "all_users": true  // Optional: assign to ALL users in system
        }
        """
        user_ids = request.data.get('user_ids', [])
        user_emails = request.data.get('user_emails', [])
        module_codes = request.data.get('module_codes', [])
        all_users = request.data.get('all_users', False)
        
        if not module_codes:
            return Response(
                {'error': 'module_codes is required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # If all_users is True, ignore user_ids and user_emails
        if not all_users and not user_ids and not user_emails:
            return Response(
                {'error': 'Either user_ids, user_emails, or all_users=true is required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Get modules
        modules = Module.objects.filter(code__in=module_codes, is_active=True)
        if modules.count() != len(module_codes):
            found_codes = set(modules.values_list('code', flat=True))
            missing_codes = set(module_codes) - found_codes
            return Response(
                {'error': f'Some modules not found: {list(missing_codes)}'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        # Get user profiles
        if all_users:
            # Assign to ALL active users in the system
            profiles = UserProfile.objects.filter(is_deleted=False)
        elif user_ids:
            profiles = UserProfile.objects.filter(id__in=user_ids, is_deleted=False)
        else:
            profiles = UserProfile.objects.filter(user__email__in=user_emails, is_deleted=False)
        
        if not profiles.exists():
            return Response(
                {'error': 'No users found matching criteria'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        # Assign modules to each user's roles
        results = {
            'success': [],
            'failed': [],
            'total_assignments': 0
        }
        
        for profile in profiles:
            try:
                user_roles = UserRole.objects.filter(user_profile=profile)
                
                if not user_roles.exists():
                    results['failed'].append({
                        'user': profile.user.email,
                        'reason': 'No roles assigned'
                    })
                    continue
                
                assigned_count = 0
                for user_role in user_roles:
                    role = user_role.role
                    for module in modules:
                        role_module, created = RoleModule.objects.get_or_create(
                            role=role,
                            module=module,
                            defaults={'granted_by': request.user}
                        )
                        if created:
                            assigned_count += 1
                
                results['success'].append({
                    'user': profile.user.email,
                    'user_id': str(profile.id),
                    'modules_assigned': assigned_count
                })
                results['total_assignments'] += assigned_count
                
            except Exception as e:
                results['failed'].append({
                    'user': profile.user.email,
                    'reason': str(e)
                })
        
        # Create audit log
        create_audit_log(
            user=request.user,
            action='bulk_modules_assign',
            resource_type='UserProfile',
            resource_id=None,
            resource_repr='Bulk Module Assignment',
            metadata={
                'module_codes': module_codes,
                'success_count': len(results['success']),
                'failed_count': len(results['failed']),
                'total_assignments': results['total_assignments']
            },
            ip_address=request.META.get('REMOTE_ADDR'),
            user_agent=request.META.get('HTTP_USER_AGENT', '')
        )
        
        return Response({
            'message': 'Bulk module assignment completed',
            'summary': {
                'total_users_processed': len(results['success']) + len(results['failed']),
                'successful': len(results['success']),
                'failed': len(results['failed']),
                'total_module_assignments': results['total_assignments']
            },
            'details': results
        })
    
    @action(detail=False, methods=['get'], url_path='departments')
    def get_departments(self, request):
        """
        Get unique list of departments from UserProfile
        Returns: {
            "departments": ["Engineering", "Sales", ...]
        }
        """
        departments = UserProfile.objects.filter(
            is_deleted=False,
            department__isnull=False
        ).exclude(
            department__exact=''
        ).values_list('department', flat=True).distinct().order_by('department')
        
        return Response({
            'departments': list(departments),
            'count': len(departments)
        })
    
    @action(detail=False, methods=['get'], url_path='job-titles')
    def get_job_titles(self, request):
        """
        Get unique list of job titles from UserProfile
        Returns: {
            "job_titles": ["Engineer", "Manager", ...]
        }
        """
        job_titles = UserProfile.objects.filter(
            is_deleted=False,
            job_title__isnull=False
        ).exclude(
            job_title__exact=''
        ).values_list('job_title', flat=True).distinct().order_by('job_title')
        
        return Response({
            'job_titles': list(job_titles),
            'count': len(job_titles)
        })

    @action(detail=False, methods=['get'], url_path='total-count')
    def total_count(self, request):
        """
        Company-wide user count across ALL organizations, ignoring the
        organization-scoping normally applied in get_queryset(). Super admin
        only — this deliberately bypasses org scoping, so it must not be
        reachable by a regular org-scoped admin.
        """
        count = UserProfile.objects.filter(is_deleted=False).count()
        return Response({'count': count})


class AuditLogViewSet(viewsets.ReadOnlyModelViewSet):
    """
    ViewSet for viewing audit logs
    Read-only access for admins
    """
    queryset = AuditLog.objects.select_related('user').all()
    serializer_class = AuditLogSerializer
    permission_classes = [IsAuthenticated, IsAdmin]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter, DjangoFilterBackend]
    search_fields = ['user_email', 'action', 'resource_type']
    ordering_fields = ['-timestamp']
    filterset_fields = ['action', 'resource_type', 'success']
    
    def get_queryset(self):
        """Filter logs based on organization for non-super-admins"""
        user = self.request.user
        queryset = AuditLog.objects.all()
        
        try:
            profile = user.rbac_profile
            if not profile.roles.filter(code='super_admin', is_active=True).exists():
                # Filter to organization users
                org_user_ids = UserProfile.objects.filter(
                    organization=profile.organization
                ).values_list('user_id', flat=True)
                queryset = queryset.filter(user_id__in=org_user_ids)
        except UserProfile.DoesNotExist:
            return AuditLog.objects.none()
        
        return queryset.order_by('-timestamp')
    
    @action(detail=False, methods=['get'])
    def user_activity(self, request):
        """Get activity logs for specific user"""
        user_id = request.query_params.get('user_id')
        if not user_id:
            return Response(
                {'error': 'user_id parameter is required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        logs = self.get_queryset().filter(user_id=user_id)[:50]
        serializer = self.get_serializer(logs, many=True)
        return Response(serializer.data)


class StorageViewSet(viewsets.ModelViewSet):
    """
    ViewSet for managing file storage with S3 integration
    """
    queryset = UserStorage.objects.filter(is_deleted=False)
    serializer_class = UserStorageSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter, DjangoFilterBackend]
    search_fields = ['filename', 'file_type']
    ordering_fields = ['-created_at', 'file_size', 'download_count']
    filterset_fields = ['file_type', 'mime_type']
    
    def get_queryset(self):
        """Filter files based on user and organization"""
        user = self.request.user
        queryset = UserStorage.objects.filter(is_deleted=False)
        
        try:
            profile = user.userprofile
            # Super admins see all files
            if not profile.has_permission('file_view_all'):
                # Regular users see only their files and organization files
                queryset = queryset.filter(
                    Q(user_profile=profile) | Q(organization=profile.organization)
                )
        except UserProfile.DoesNotExist:
            return UserStorage.objects.none()
        
        return queryset.order_by('-created_at')
    
    @action(detail=False, methods=['post'])
    def generate_upload_url(self, request):
        """
        Generate pre-signed URL for uploading a file to S3
        
        POST /api/v1/rbac/storage/generate_upload_url/
        {
            "file_name": "drawing.pdf",
            "file_size": 1024000,
            "content_type": "application/pdf",
            "category": "pid_analysis",
            "tags": {"project": "ABC-123"}
        }
        """
        file_name = request.data.get('file_name')
        file_size = request.data.get('file_size')
        content_type = request.data.get('content_type')
        category = request.data.get('category', 'general')
        tags = request.data.get('tags', {})
        
        if not file_name or not file_size:
            return Response(
                {'error': 'file_name and file_size are required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            s3_service = S3Service(organization=request.user.userprofile.organization)
            result = s3_service.generate_upload_url(
                user=request.user,
                file_name=file_name,
                file_size=file_size,
                content_type=content_type,
                category=category,
                tags=tags
            )
            return Response(result, status=status.HTTP_200_OK)
        except Exception as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=True, methods=['get'])
    def download(self, request, pk=None):
        """
        Generate pre-signed URL for downloading a file from S3
        
        GET /api/v1/rbac/storage/{id}/download/
        """
        try:
            s3_service = S3Service(organization=request.user.userprofile.organization)
            result = s3_service.generate_download_url(
                storage_id=pk,
                user=request.user
            )
            return Response(result, status=status.HTTP_200_OK)
        except PermissionError as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_403_FORBIDDEN
            )
        except Exception as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=True, methods=['post'])
    def verify_upload(self, request, pk=None):
        """
        Verify that file was successfully uploaded to S3
        
        POST /api/v1/rbac/storage/{id}/verify_upload/
        {
            "checksum": "md5hash"
        }
        """
        checksum = request.data.get('checksum')
        
        try:
            s3_service = S3Service(organization=request.user.userprofile.organization)
            result = s3_service.verify_upload(
                storage_id=pk,
                user=request.user,
                checksum=checksum
            )
            return Response({'verified': result}, status=status.HTTP_200_OK)
        except Exception as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=False, methods=['get'])
    def stats(self, request):
        """
        Get storage statistics for current user
        
        GET /api/v1/rbac/storage/stats/
        """
        try:
            s3_service = S3Service(organization=request.user.userprofile.organization)
            stats = s3_service.get_storage_stats(user=request.user)
            return Response(stats, status=status.HTTP_200_OK)
        except Exception as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    def destroy(self, request, *args, **kwargs):
        """Soft delete file from storage"""
        try:
            storage = self.get_object()
            s3_service = S3Service(organization=request.user.userprofile.organization)
            s3_service.delete_file(storage_id=storage.id, user=request.user)
            return Response(status=status.HTTP_204_NO_CONTENT)
        except PermissionError as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_403_FORBIDDEN
            )
        except Exception as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


# ============================================================================
# ANALYTICS VIEWSETS - AI-Powered Admin Features
# ============================================================================

from .analytics_models import (
    SystemMetrics, UserActivityAnalytics, SecurityAlert, PredictiveInsight,
    FeatureUsageAnalytics, ErrorLogAnalytics, SystemHealthCheck
)
from .analytics_serializers import (
    SystemMetricsSerializer, UserActivityAnalyticsSerializer, SecurityAlertSerializer,
    PredictiveInsightSerializer, FeatureUsageAnalyticsSerializer, ErrorLogAnalyticsSerializer,
    SystemHealthCheckSerializer, DashboardStatsSerializer, RealTimeActivitySerializer
)
from datetime import timedelta, datetime
from django.db.models import Avg, Sum, Count, Max, Min, F


class AnalyticsDashboardViewSet(viewsets.ViewSet):
    """
    AI-Powered Analytics Dashboard
    Comprehensive admin overview with real-time insights
    """
    permission_classes = [IsAuthenticated, IsSuperAdmin]
    
    @action(detail=False, methods=['get'])
    def overview(self, request):
        """
        Get comprehensive dashboard overview
        Includes system health, user stats, security alerts, and AI insights
        """
        # Refresh analytics snapshots (TTL-gated, soft-coded intervals).
        # This only ingests live data — it does not alter the read logic below.
        try:
            from .analytics_collectors import ensure_fresh
            ensure_fresh()
        except Exception:  # never let collector failure break the dashboard
            pass

        today = timezone.now().date()
        yesterday = today - timedelta(days=1)
        
        # User Statistics
        total_users = UserProfile.objects.filter(is_deleted=False).count()
        active_today = UserActivityAnalytics.objects.filter(
            date=today, 
            login_count__gt=0
        ).count()
        
        # System Metrics (Latest)
        latest_metrics = SystemMetrics.objects.first()
        
        # Security Alerts
        active_alerts = SecurityAlert.objects.filter(status='new').count()
        critical_alerts = SecurityAlert.objects.filter(
            status='new',
            severity='critical'
        ).count()
        
        # AI Predictions
        active_predictions = PredictiveInsight.objects.filter(
            is_active=True,
            is_acknowledged=False
        ).count()
        high_impact = PredictiveInsight.objects.filter(
            is_active=True,
            is_acknowledged=False,
            impact_level='high'
        ).count()
        
        # Error Statistics
        errors_today = ErrorLogAnalytics.objects.filter(
            last_occurrence__date=today,
            status='open'
        ).count()
        critical_errors = ErrorLogAnalytics.objects.filter(
            status='open',
            severity='critical'
        ).count()
        
        # User Growth
        users_yesterday = UserProfile.objects.filter(
            created_at__date__lte=yesterday,
            is_deleted=False
        ).count()
        growth_rate = ((total_users - users_yesterday) / users_yesterday * 100) if users_yesterday > 0 else 0
        
        # System Health
        latest_health = SystemHealthCheck.objects.first()
        health_score = latest_health.health_score if latest_health else 100.0
        
        data = {
            'total_users': total_users,
            'active_users_today': active_today,
            'total_api_requests_today': latest_metrics.api_requests_count if latest_metrics else 0,
            'system_health_score': health_score,
            'avg_response_time_ms': latest_metrics.avg_response_time_ms if latest_metrics else 0,
            'success_rate_percentage': latest_metrics.success_rate_percentage if latest_metrics else 100,
            'active_connections': latest_metrics.active_connections if latest_metrics else 0,
            'active_alerts_count': active_alerts,
            'critical_alerts_count': critical_alerts,
            'active_predictions_count': active_predictions,
            'high_impact_insights_count': high_impact,
            'errors_today': errors_today,
            'critical_errors_count': critical_errors,
            'user_growth_percentage': round(growth_rate, 2),
            'engagement_trend': 'growing' if growth_rate > 0 else 'stable',
        }
        
        serializer = DashboardStatsSerializer(data)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def real_time_activity(self, request):
        """
        Get real-time activity feed
        Recent user actions, alerts, and system events
        """
        limit = int(request.query_params.get('limit', 20))
        
        # Get recent audit logs
        recent_audits = AuditLog.objects.select_related('user').order_by('-timestamp')[:limit]
        
        activities = []
        for audit in recent_audits:
            activities.append({
                'activity_type': audit.action,
                'user_email': audit.user_email,
                'description': f"{audit.action.title()} {audit.resource_type}",
                'timestamp': audit.timestamp,
                'severity': 'high' if not audit.success else 'normal',
                'metadata': {
                    'resource_id': audit.resource_id,
                    'success': audit.success,
                    'changes': audit.changes
                }
            })
        
        serializer = RealTimeActivitySerializer(activities, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def system_performance(self, request):
        """
        Get system performance metrics over time
        """
        days = int(request.query_params.get('days', 7))
        start_date = timezone.now() - timedelta(days=days)
        
        metrics = SystemMetrics.objects.filter(
            timestamp__gte=start_date
        ).order_by('timestamp').values(
            'timestamp', 'avg_response_time_ms', 'success_rate_percentage',
            'cpu_usage_percentage', 'memory_usage_mb', 'active_connections',
            'api_requests_count', 'failed_requests_count'
        )
        
        return Response(list(metrics))
    
    @action(detail=False, methods=['get'])
    def user_engagement_trends(self, request):
        """
        Get user engagement trends and patterns
        """
        days = int(request.query_params.get('days', 30))
        start_date = timezone.now().date() - timedelta(days=days)
        
        analytics = UserActivityAnalytics.objects.filter(
            date__gte=start_date
        ).values('date').annotate(
            total_logins=Sum('login_count'),
            avg_engagement=Avg('engagement_score'),
            avg_productivity=Avg('productivity_score'),
            users_with_anomalies=Count('id', filter=Q(anomaly_detected=True))
        ).order_by('date')
        
        return Response(list(analytics))


class SystemMetricsViewSet(viewsets.ReadOnlyModelViewSet):
    """
    ViewSet for system performance metrics
    Read-only for admins to monitor system health
    """
    queryset = SystemMetrics.objects.all()
    serializer_class = SystemMetricsSerializer
    permission_classes = [IsAuthenticated, IsSuperAdmin]
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ['timestamp']
    ordering = ['-timestamp']
    
    @action(detail=False, methods=['get'])
    def latest(self, request):
        """Get the most recent metrics"""
        latest = self.queryset.first()
        if latest:
            serializer = self.get_serializer(latest)
            return Response(serializer.data)
        return Response({})
    
    @action(detail=False, methods=['get'])
    def averages(self, request):
        """Get average metrics over a time period"""
        days = int(request.query_params.get('days', 7))
        start_date = timezone.now() - timedelta(days=days)
        
        averages = self.queryset.filter(timestamp__gte=start_date).aggregate(
            avg_response_time=Avg('avg_response_time_ms'),
            avg_success_rate=Avg('success_rate_percentage'),
            avg_cpu=Avg('cpu_usage_percentage'),
            avg_memory=Avg('memory_usage_mb'),
            total_requests=Sum('api_requests_count'),
            total_failed=Sum('failed_requests_count')
        )
        
        return Response(averages)


class UserActivityAnalyticsViewSet(viewsets.ReadOnlyModelViewSet):
    """
    ViewSet for user behavior analytics
    AI-powered insights into user patterns and engagement
    """
    queryset = UserActivityAnalytics.objects.select_related('user').all()
    serializer_class = UserActivityAnalyticsSerializer
    permission_classes = [IsAuthenticated, IsSuperAdmin]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['date', 'anomaly_detected', 'usage_pattern']
    search_fields = ['user__email', 'user__first_name', 'user__last_name']
    ordering = ['-date']
    
    @action(detail=False, methods=['get'])
    def top_engaged_users(self, request):
        """Get users with highest engagement scores"""
        days = int(request.query_params.get('days', 7))
        start_date = timezone.now().date() - timedelta(days=days)
        limit = int(request.query_params.get('limit', 10))
        
        top_users = self.queryset.filter(date__gte=start_date).values(
            'user__email', 'user__first_name', 'user__last_name'
        ).annotate(
            avg_engagement=Avg('engagement_score'),
            total_logins=Sum('login_count'),
            total_actions=Sum('drawings_uploaded') + Sum('analyses_completed')
        ).order_by('-avg_engagement')[:limit]
        
        return Response(list(top_users))
    
    @action(detail=False, methods=['get'])
    def anomalies(self, request):
        """Get users with detected anomalies"""
        days = int(request.query_params.get('days', 7))
        start_date = timezone.now().date() - timedelta(days=days)
        
        anomalies = self.queryset.filter(
            date__gte=start_date,
            anomaly_detected=True
        ).select_related('user').order_by('-risk_score')
        
        serializer = self.get_serializer(anomalies, many=True)
        return Response(serializer.data)
    
    @action(detail=True, methods=['get'])
    def user_timeline(self, request, pk=None):
        """Get activity timeline for a specific user"""
        user_id = pk
        days = int(request.query_params.get('days', 30))
        start_date = timezone.now().date() - timedelta(days=days)
        
        timeline = self.queryset.filter(
            user_id=user_id,
            date__gte=start_date
        ).order_by('date')
        
        serializer = self.get_serializer(timeline, many=True)
        return Response(serializer.data)


class SecurityAlertViewSet(viewsets.ModelViewSet):
    """
    ViewSet for security alerts
    AI-powered threat detection and management
    """
    queryset = SecurityAlert.objects.select_related('user', 'resolved_by').all()
    serializer_class = SecurityAlertSerializer
    permission_classes = [IsAuthenticated, IsSuperAdmin]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['severity', 'status', 'alert_type']
    search_fields = ['title', 'description', 'user__email']
    ordering = ['-detection_time']
    
    @action(detail=True, methods=['post'])
    def resolve(self, request, pk=None):
        """Mark alert as resolved"""
        alert = self.get_object()
        alert.status = 'resolved'
        alert.resolved_at = timezone.now()
        alert.resolved_by = request.user
        alert.resolution_notes = request.data.get('notes', '')
        alert.save()
        
        serializer = self.get_serializer(alert)
        return Response(serializer.data)
    
    @action(detail=True, methods=['post'])
    def investigate(self, request, pk=None):
        """Mark alert as under investigation"""
        alert = self.get_object()
        alert.status = 'investigating'
        alert.save()
        
        serializer = self.get_serializer(alert)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def critical(self, request):
        """Get all critical unresolved alerts"""
        critical_alerts = self.queryset.filter(
            severity='critical',
            status__in=['new', 'investigating']
        )
        
        serializer = self.get_serializer(critical_alerts, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def statistics(self, request):
        """Get security alert statistics"""
        days = int(request.query_params.get('days', 30))
        start_date = timezone.now() - timedelta(days=days)
        
        stats = {
            'total_alerts': self.queryset.filter(detection_time__gte=start_date).count(),
            'by_severity': dict(self.queryset.filter(
                detection_time__gte=start_date
            ).values('severity').annotate(count=Count('id')).values_list('severity', 'count')),
            'by_status': dict(self.queryset.filter(
                detection_time__gte=start_date
            ).values('status').annotate(count=Count('id')).values_list('status', 'count')),
            'resolution_time_avg_hours': 0,  # Calculate from resolved alerts
        }
        
        return Response(stats)


class PredictiveInsightViewSet(viewsets.ModelViewSet):
    """
    ViewSet for AI-generated predictions and insights
    Machine learning powered recommendations
    """
    queryset = PredictiveInsight.objects.select_related('acknowledged_by').all()
    serializer_class = PredictiveInsightSerializer
    permission_classes = [IsAuthenticated, IsSuperAdmin]
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ['insight_type', 'impact_level', 'is_active', 'is_acknowledged']
    ordering = ['-created_at']
    
    @action(detail=True, methods=['post'])
    def acknowledge(self, request, pk=None):
        """Acknowledge an insight"""
        insight = self.get_object()
        insight.is_acknowledged = True
        insight.acknowledged_by = request.user
        insight.acknowledged_at = timezone.now()
        insight.save()
        
        serializer = self.get_serializer(insight)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def pending(self, request):
        """Get unacknowledged insights"""
        pending = self.queryset.filter(
            is_active=True,
            is_acknowledged=False
        ).order_by('-confidence_score')
        
        serializer = self.get_serializer(pending, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def high_priority(self, request):
        """Get high-impact unacknowledged insights"""
        high_priority = self.queryset.filter(
            is_active=True,
            is_acknowledged=False,
            impact_level='high'
        ).order_by('-confidence_score')
        
        serializer = self.get_serializer(high_priority, many=True)
        return Response(serializer.data)


class FeatureUsageAnalyticsViewSet(viewsets.ReadOnlyModelViewSet):
    """
    ViewSet for feature usage analytics
    Track feature adoption and health
    """
    queryset = FeatureUsageAnalytics.objects.all()
    serializer_class = FeatureUsageAnalyticsSerializer
    permission_classes = [IsAuthenticated, IsSuperAdmin]
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ['feature_name', 'date', 'trend']
    ordering = ['-date']
    
    @action(detail=False, methods=['get'])
    def summary(self, request):
        """Get summary of all features"""
        days = int(request.query_params.get('days', 7))
        start_date = timezone.now().date() - timedelta(days=days)
        
        summary = self.queryset.filter(date__gte=start_date).values(
            'feature_name'
        ).annotate(
            avg_adoption_rate=Avg('adoption_rate_percentage'),
            avg_health_score=Avg('health_score'),
            total_users=Sum('active_users'),
            total_usage=Sum('total_usage_count')
        ).order_by('-total_usage')
        
        return Response(list(summary))
    
    @action(detail=False, methods=['get'])
    def trending(self, request):
        """Get trending features"""
        trending = self.queryset.filter(
            trend='growing'
        ).order_by('-growth_rate_percentage')[:10]
        
        serializer = self.get_serializer(trending, many=True)
        return Response(serializer.data)


class ErrorLogAnalyticsViewSet(viewsets.ModelViewSet):
    """
    ViewSet for error analytics
    AI-powered error tracking and root cause analysis
    """
    queryset = ErrorLogAnalytics.objects.all()
    serializer_class = ErrorLogAnalyticsSerializer
    permission_classes = [IsAuthenticated, IsSuperAdmin]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['severity', 'status', 'error_type']
    search_fields = ['error_type', 'error_message']
    ordering = ['-last_occurrence']
    
    @action(detail=True, methods=['post'])
    def mark_resolved(self, request, pk=None):
        """Mark error as resolved"""
        error = self.get_object()
        error.status = 'resolved'
        error.resolution_notes = request.data.get('notes', '')
        error.resolved_at = timezone.now()
        error.save()
        
        serializer = self.get_serializer(error)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def critical_errors(self, request):
        """Get all critical unresolved errors"""
        critical = self.queryset.filter(
            severity='critical',
            status='open'
        ).order_by('-occurrence_count')
        
        serializer = self.get_serializer(critical, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def statistics(self, request):
        """Get error statistics"""
        days = int(request.query_params.get('days', 7))
        start_date = timezone.now() - timedelta(days=days)
        
        stats = {
            'total_errors': self.queryset.filter(last_occurrence__gte=start_date).count(),
            'total_occurrences': self.queryset.filter(
                last_occurrence__gte=start_date
            ).aggregate(total=Sum('occurrence_count'))['total'] or 0,
            'by_severity': dict(self.queryset.filter(
                last_occurrence__gte=start_date
            ).values('severity').annotate(count=Count('id')).values_list('severity', 'count')),
            'affected_users': self.queryset.filter(
                last_occurrence__gte=start_date
            ).aggregate(total=Sum('affected_users_count'))['total'] or 0,
        }
        
        return Response(stats)


class SystemHealthCheckViewSet(viewsets.ReadOnlyModelViewSet):
    """
    ViewSet for system health monitoring
    Real-time system status and diagnostics
    """
    queryset = SystemHealthCheck.objects.all()
    serializer_class = SystemHealthCheckSerializer
    permission_classes = [IsAuthenticated, IsSuperAdmin]
    ordering = ['-check_time']
    
    @action(detail=False, methods=['get'])
    def latest(self, request):
        """Get latest health check"""
        latest = self.queryset.first()
        if latest:
            serializer = self.get_serializer(latest)
            return Response(serializer.data)
        return Response({})
    
    @action(detail=False, methods=['get'])
    def history(self, request):
        """Get health check history"""
        hours = int(request.query_params.get('hours', 24))
        start_time = timezone.now() - timedelta(hours=hours)
        
        history = self.queryset.filter(check_time__gte=start_time).order_by('check_time')
        serializer = self.get_serializer(history, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def component_status(self, request):
        """Get current status of all system components"""
        latest = self.queryset.first()
        if latest:
            return Response({
                'database': latest.database_status,
                'redis': latest.redis_status,
                'celery': latest.celery_status,
                'storage': latest.storage_status,
                'api': latest.api_status,
                'overall': latest.overall_status,
                'health_score': latest.health_score,
                'check_time': latest.check_time,
            })
        return Response({})


# ===========================================================================
# STANDALONE: User Export View — no ViewSet inheritance, no router dependency
# GET /api/v1/rbac/users/export/?file_format=csv   → CSV download
# GET /api/v1/rbac/users/export/?file_format=xlsx  → Excel download
# NOTE: param is 'file_format' (not 'format') to avoid DRF content-negotiation
#       interception — DRF treats ?format=xxx as a renderer override and raises
#       Http404 when no renderer matches (e.g. 'csv' is not a registered renderer).
# ===========================================================================
class UserExportView(APIView):
    permission_classes = [IsAuthenticated, CanManageUsers]

    # Soft-coded export configuration
    EXPORT_HEADERS = [
        'First Name', 'Last Name', 'Email', 'Username',
        'Department', 'Job Title', 'Phone', 'Employee ID',
        'Location', 'Status', 'Roles', 'Organization',
        'Created At', 'Last Login',
    ]
    HEADER_COLOR = '1E40AF'   # Blue header for Excel
    MAX_COL_WIDTH = 40

    def get(self, request):
        import csv
        import datetime
        from io import BytesIO
        from django.http import HttpResponse

        # 'file_format' param — intentionally NOT 'format' to avoid DRF content-negotiation
        # interception: DRF treats ?format=xxx as a renderer format override and raises
        # Http404 when no renderer with that format exists (e.g. 'csv' has no renderer).
        export_format = request.query_params.get('file_format', 'csv').lower()
        timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')

        # Build queryset scoped to the requester's organization (super_admin sees all)
        queryset = UserProfile.objects.select_related(
            'user', 'organization'
        ).prefetch_related(
            'roles'
        ).filter(is_deleted=False)

        try:
            profile = request.user.rbac_profile
            if not profile.roles.filter(code='super_admin', is_active=True).exists():
                queryset = queryset.filter(organization=profile.organization)
        except UserProfile.DoesNotExist:
            queryset = UserProfile.objects.none()

        queryset = queryset.order_by('created_at')

        def build_row(p):
            usr = p.user
            roles = ', '.join(sorted(set(r.name for r in p.roles.all())))
            created = p.created_at.strftime('%Y-%m-%d %H:%M') if p.created_at else ''
            last_login = p.last_login_at.strftime('%Y-%m-%d %H:%M') if p.last_login_at else ''
            return [
                usr.first_name or '', usr.last_name or '',
                usr.email or '', usr.username or '',
                p.department or '', p.job_title or '',
                p.phone or '', p.employee_id or '',
                p.location or '', p.status or '',
                roles,
                p.organization.name if p.organization else '',
                created, last_login,
            ]

        if export_format == 'xlsx':
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Users'

            header_fill = PatternFill(start_color=self.HEADER_COLOR, end_color=self.HEADER_COLOR, fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF')
            ws.append(self.EXPORT_HEADERS)
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')

            for p in queryset:
                ws.append(build_row(p))

            for col in ws.columns:
                max_len = max((len(str(cell.value or '')) for cell in col), default=10)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, self.MAX_COL_WIDTH)

            output = BytesIO()
            wb.save(output)
            output.seek(0)
            response = HttpResponse(
                output.read(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="users_export_{timestamp}.xlsx"'
        else:
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = f'attachment; filename="users_export_{timestamp}.csv"'
            writer = csv.writer(response)
            writer.writerow(self.EXPORT_HEADERS)
            for p in queryset:
                writer.writerow(build_row(p))

        return response


# ---------------------------------------------------------------------------
# Access Request ViewSet
# ---------------------------------------------------------------------------

class AccessRequestViewSet(viewsets.ModelViewSet):
    """
    Module access requests submitted by regular users.

    - Regular users: create and view their own requests.
    - Admins / Super Admins: view all requests; approve or deny via custom actions.
    """

    serializer_class   = AccessRequestSerializer
    permission_classes = [IsAuthenticated]
    filter_backends    = [filters.OrderingFilter, DjangoFilterBackend]
    filterset_fields   = ['status', 'module']
    ordering_fields    = ['created_at']
    ordering           = ['-created_at']

    # Soft-coded: roles that can see/manage all requests
    MANAGER_ROLE_CODES = ['super_admin', 'admin']

    def _is_manager(self, user):
        """Return True if user is superuser or holds a manager-level role."""
        if user.is_superuser:
            return True
        try:
            return user.rbac_profile.roles.filter(
                code__in=self.MANAGER_ROLE_CODES, is_active=True
            ).exists()
        except UserProfile.DoesNotExist:
            return False

    def get_queryset(self):
        user = self.request.user
        base_qs = AccessRequest.objects.select_related(
            'user_profile__user', 'module', 'reviewed_by'
        )
        if self._is_manager(user):
            return base_qs.all()
        try:
            return base_qs.filter(user_profile=user.rbac_profile)
        except UserProfile.DoesNotExist:
            return AccessRequest.objects.none()

    def perform_create(self, serializer):
        try:
            profile = self.request.user.rbac_profile
        except UserProfile.DoesNotExist:
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied('User profile not found.')

        module_id = self.request.data.get('module')
        # Guard: prevent duplicate pending requests for the same module
        if AccessRequest.objects.filter(
            user_profile=profile,
            module_id=module_id,
            status=AccessRequest.STATUS_PENDING,
        ).exists():
            from rest_framework.exceptions import ValidationError
            raise ValidationError(
                {'detail': 'A pending request for this module already exists.'}
            )
        serializer.save(user_profile=profile)

    # ------------------------------------------------------------------
    # Admin actions: approve / deny
    # ------------------------------------------------------------------

    @action(detail=True, methods=['post'])
    def approve(self, request, pk=None):
        if not self._is_manager(request.user):
            return Response(
                {'detail': 'Only admins can approve access requests.'},
                status=status.HTTP_403_FORBIDDEN,
            )
        req = self.get_object()
        if req.status != AccessRequest.STATUS_PENDING:
            return Response(
                {'detail': f'Request is already {req.status}.'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        req.status      = AccessRequest.STATUS_APPROVED
        req.reviewed_by = request.user
        req.reviewed_at = timezone.now()
        req.admin_note  = request.data.get('admin_note', '')
        req.save()

        # Grant access: ensure the user's viewer role has the module, then
        # also assign the module directly to the user's viewer UserRole.
        try:
            from django.core.cache import cache
            viewer_role = Role.objects.get(code='viewer')
            RoleModule.objects.get_or_create(role=viewer_role, module=req.module)
            # If the user already has the viewer role, the cache clear is enough
            UserRole.objects.get_or_create(
                user_profile=req.user_profile,
                role=viewer_role,
                defaults={'is_primary': False},
            )
            cache.delete(f'user_modules_{req.user_profile.id}')
        except Exception:
            pass  # Non-fatal — approval still recorded

        create_audit_log(
            user=request.user,
            action='role_assign',
            resource_type='AccessRequest',
            resource_id=req.id,
            resource_repr=str(req),
            ip_address=request.META.get('REMOTE_ADDR'),
            user_agent=request.META.get('HTTP_USER_AGENT', ''),
        )
        return Response({'status': 'approved'})

    @action(detail=True, methods=['post'])
    def deny(self, request, pk=None):
        if not self._is_manager(request.user):
            return Response(
                {'detail': 'Only admins can deny access requests.'},
                status=status.HTTP_403_FORBIDDEN,
            )
        req = self.get_object()
        if req.status != AccessRequest.STATUS_PENDING:
            return Response(
                {'detail': f'Request is already {req.status}.'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        req.status      = AccessRequest.STATUS_DENIED
        req.reviewed_by = request.user
        req.reviewed_at = timezone.now()
        req.admin_note  = request.data.get('admin_note', '')
        req.save()
        return Response({'status': 'denied'})


# ═════════════════════════════════════════════════════════════════════════════
# Enhanced Profile ViewSets — Achievements, Experience, Social Media
# ═════════════════════════════════════════════════════════════════════════════

class AchievementViewSet(viewsets.ModelViewSet):
    """
    ViewSet for user achievements and milestones.
    
    Users can manage their own achievements. Admins can view all achievements.
    Soft-coded achievement categories defined in rbac.profile_config
    """
    permission_classes = [IsAuthenticated]
    filter_backends = [filters.OrderingFilter, filters.SearchFilter, DjangoFilterBackend]
    filterset_fields = ['category', 'level', 'is_public', 'is_verified']
    search_fields = ['title', 'description', 'organization']
    ordering_fields = ['achieved_date', 'created_at', 'display_order']
    ordering = ['-achieved_date', '-created_at']
    
    def get_serializer_class(self):
        from .serializers import AchievementSerializer
        return AchievementSerializer
    
    def get_queryset(self):
        """Return user's own achievements, or all if admin."""
        from .models import Achievement
        user = self.request.user

        # Profile screens must remain scoped to the signed-in user, including
        # when that user also has staff/admin privileges.
        if self.request.query_params.get('mine', '').lower() in ('1', 'true', 'yes'):
            return Achievement.objects.filter(user_profile=user.rbac_profile)
        
        # Admins can see all achievements
        if user.is_superuser or user.is_staff:
            return Achievement.objects.select_related('user_profile__user').all()
        
        # Regular users see only their own
        try:
            return Achievement.objects.filter(user_profile=user.rbac_profile)
        except:
            return Achievement.objects.none()
    
    def perform_create(self, serializer):
        """Auto-assign current user's profile (auto-provisioned if missing)."""
        from .profile_utils import get_or_create_profile, ProfileProvisioningError
        from rest_framework.exceptions import ValidationError
        import logging
        
        logger = logging.getLogger(__name__)
        logger.info(f'[AchievementViewSet] Creating achievement for user: {self.request.user.email}')
        logger.info(f'[AchievementViewSet] Request data: {self.request.data}')

        try:
            profile = get_or_create_profile(self.request.user, source='AchievementViewSet')
            logger.info(f'[AchievementViewSet] Profile obtained: {profile.id}')
        except ProfileProvisioningError as exc:
            logger.error(f'[AchievementViewSet] Profile provisioning failed: {exc}')
            raise ValidationError({'user_profile': str(exc)})
        except Exception as exc:
            logger.exception(f'[AchievementViewSet] Unexpected error getting profile')
            # Try using request.user_profile set by middleware as fallback
            profile = getattr(self.request, 'user_profile', None)
            if profile is None:
                raise ValidationError({'detail': f'Failed to get user profile: {str(exc)}'})

        try:
            instance = serializer.save(user_profile=profile)
            logger.info(f'[AchievementViewSet] Achievement created successfully: {instance.id}')
        except Exception as exc:
            logger.exception(f'[AchievementViewSet] Failed to save achievement')
            raise
    
    @action(detail=False, methods=['get'])
    def categories(self, request):
        """Return available achievement categories."""
        from apps.rbac.profile_config import ACHIEVEMENT_CATEGORIES
        categories = [
            {
                'value': code,
                'label': config['label'],
                'icon': config['icon'],
                'color': config['color'],
                'description': config['description'],
            }
            for code, config in ACHIEVEMENT_CATEGORIES.items()
        ]
        return Response(categories)
    
    @action(detail=False, methods=['get'])
    def levels(self, request):
        """Return achievement levels for gamification."""
        from apps.rbac.profile_config import ACHIEVEMENT_LEVELS
        return Response(ACHIEVEMENT_LEVELS)


class WorkExperienceViewSet(viewsets.ModelViewSet):
    """
    ViewSet for work experience entries.
    
    Users manage their own experience timeline. Admins can view all.
    """
    permission_classes = [IsAuthenticated]
    filter_backends = [filters.OrderingFilter, filters.SearchFilter, DjangoFilterBackend]
    filterset_fields = ['employment_type', 'industry', 'is_current', 'is_public']
    search_fields = ['company_name', 'job_title', 'description', 'industry']
    ordering_fields = ['start_date', 'end_date', 'created_at', 'display_order']
    ordering = ['-is_current', '-start_date']
    
    def get_serializer_class(self):
        from .serializers import WorkExperienceSerializer
        return WorkExperienceSerializer
    
    def get_queryset(self):
        """Return user's own experience, or all if admin."""
        from .models import WorkExperience
        user = self.request.user

        if self.request.query_params.get('mine', '').lower() in ('1', 'true', 'yes'):
            return WorkExperience.objects.filter(user_profile=user.rbac_profile)
        
        # Admins can see all
        if user.is_superuser or user.is_staff:
            return WorkExperience.objects.select_related('user_profile__user').all()
        
        # Regular users see only their own
        try:
            return WorkExperience.objects.filter(user_profile=user.rbac_profile)
        except:
            return WorkExperience.objects.none()
    
    def perform_create(self, serializer):
        """Auto-assign current user's profile (auto-provisioned if missing)."""
        from .profile_utils import get_or_create_profile, ProfileProvisioningError
        from rest_framework.exceptions import ValidationError
        import logging
        
        logger = logging.getLogger(__name__)
        logger.info(f'[WorkExperienceViewSet] Creating experience for user: {self.request.user.email}')
        logger.info(f'[WorkExperienceViewSet] Request data: {self.request.data}')

        try:
            profile = get_or_create_profile(self.request.user, source='WorkExperienceViewSet')
            logger.info(f'[WorkExperienceViewSet] Profile obtained: {profile.id}')
        except ProfileProvisioningError as exc:
            logger.error(f'[WorkExperienceViewSet] Profile provisioning failed: {exc}')
            raise ValidationError({'user_profile': str(exc)})
        except Exception as exc:
            logger.exception(f'[WorkExperienceViewSet] Unexpected error getting profile')
            # Try using request.user_profile set by middleware as fallback
            profile = getattr(self.request, 'user_profile', None)
            if profile is None:
                raise ValidationError({'detail': f'Failed to get user profile: {str(exc)}'})

        try:
            instance = serializer.save(user_profile=profile)
            logger.info(f'[WorkExperienceViewSet] Experience created successfully: {instance.id}')
        except Exception as exc:
            logger.exception(f'[WorkExperienceViewSet] Failed to save experience')
            raise
    
    @action(detail=False, methods=['get'])
    def employment_types(self, request):
        """Return available employment types."""
        from apps.rbac.profile_config import EMPLOYMENT_TYPES
        return Response(EMPLOYMENT_TYPES)
    
    @action(detail=False, methods=['get'])
    def industries(self, request):
        """Return industry sectors."""
        from apps.rbac.profile_config import INDUSTRY_SECTORS
        return Response([{'value': ind, 'label': ind} for ind in INDUSTRY_SECTORS])


class SocialMediaLinkViewSet(viewsets.ModelViewSet):
    """
    ViewSet for social media and professional network links.
    
    Users manage their own social links. Admins can view all.
    Platform codes soft-coded in rbac.profile_config
    """
    permission_classes = [IsAuthenticated]
    filter_backends = [filters.OrderingFilter, DjangoFilterBackend]
    filterset_fields = ['platform', 'is_public', 'is_verified']
    ordering_fields = ['platform', 'display_order', 'created_at']
    ordering = ['display_order', 'platform']
    
    def get_serializer_class(self):
        from .serializers import SocialMediaLinkSerializer
        return SocialMediaLinkSerializer
    
    def get_queryset(self):
        """Return user's own social links, or all if admin."""
        from .models import SocialMediaLink
        user = self.request.user

        if self.request.query_params.get('mine', '').lower() in ('1', 'true', 'yes'):
            return SocialMediaLink.objects.filter(user_profile=user.rbac_profile)
        
        # Admins can see all
        if user.is_superuser or user.is_staff:
            return SocialMediaLink.objects.select_related('user_profile__user').all()
        
        # Regular users see only their own
        try:
            return SocialMediaLink.objects.filter(user_profile=user.rbac_profile)
        except:
            return SocialMediaLink.objects.none()
    
    def perform_create(self, serializer):
        """Auto-assign current user's profile (auto-provisioned if missing)."""
        from .profile_utils import get_or_create_profile, ProfileProvisioningError
        from rest_framework.exceptions import ValidationError
        import logging
        
        logger = logging.getLogger(__name__)
        logger.info(f'[SocialMediaLinkViewSet] Creating social link for user: {self.request.user.email}')
        logger.info(f'[SocialMediaLinkViewSet] Request data: {self.request.data}')

        try:
            profile = get_or_create_profile(self.request.user, source='SocialMediaLinkViewSet')
            logger.info(f'[SocialMediaLinkViewSet] Profile obtained: {profile.id}')
        except ProfileProvisioningError as exc:
            logger.error(f'[SocialMediaLinkViewSet] Profile provisioning failed: {exc}')
            raise ValidationError({'user_profile': str(exc)})
        except Exception as exc:
            logger.exception(f'[SocialMediaLinkViewSet] Unexpected error getting profile')
            # Try using request.user_profile set by middleware as fallback
            profile = getattr(self.request, 'user_profile', None)
            if profile is None:
                raise ValidationError({'detail': f'Failed to get user profile: {str(exc)}'})

        try:
            instance = serializer.save(user_profile=profile)
            logger.info(f'[SocialMediaLinkViewSet] Social link created successfully: {instance.id}')
        except Exception as exc:
            logger.exception(f'[SocialMediaLinkViewSet] Failed to save social link')
            raise
    
    @action(detail=False, methods=['get'])
    def platforms(self, request):
        """Return available social media platforms."""
        from apps.rbac.profile_config import SOCIAL_MEDIA_PLATFORMS
        platforms = [
            {
                'value': code,
                'label': config['label'],
                'icon': config['icon'],
                'color': config['color'],
                'placeholder': config['placeholder'],
            }
            for code, config in SOCIAL_MEDIA_PLATFORMS.items()
        ]
        return Response(platforms)


class ProfileDocumentViewSet(viewsets.ModelViewSet):
    """
    ViewSet for profile documents (Emirates ID, Driving License, Country ID, etc.)
    
    Users manage their own documents. Admins can view and verify all.
    Document types soft-coded in rbac.profile_config.DOCUMENT_TYPES
    Files stored in AWS S3 bucket.
    """
    permission_classes = [IsAuthenticated]
    filter_backends = [filters.OrderingFilter, DjangoFilterBackend]
    filterset_fields = ['document_type', 'verification_status', 'is_active']
    ordering_fields = ['document_type', 'created_at', 'expiry_date']
    ordering = ['-created_at']

    def _can_review_documents(self, request):
        """Use the same authorization boundary as Employee Management."""
        return bool(
            request.user.is_superuser
            or request.user.is_staff
            or CanManageUsers().has_permission(request, self)
        )

    def _require_document_reviewer(self, request):
        """Enforce RBAC ownership of organization-wide document review."""
        if not self._can_review_documents(request):
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied('Only authorized RBAC administrators can review profile documents.')

    @action(detail=False, methods=['get'], url_path='pending-verification')
    def pending_verification(self, request):
        """Return active profile documents awaiting administrator verification."""
        self._require_document_reviewer(request)
        from .models import ProfileDocument

        queryset = ProfileDocument.objects.filter(
            verification_status='pending',
            is_active=True,
        ).select_related('user_profile__user', 'verified_by').order_by('created_at')

        if request.query_params.get('count_only', '').lower() in ('1', 'true', 'yes'):
            return Response({'count': queryset.count()})

        try:
            limit = min(max(int(request.query_params.get('limit', 50)), 1), 200)
        except (TypeError, ValueError):
            limit = 50
        serializer = self.get_serializer(queryset[:limit], many=True)
        return Response({'count': queryset.count(), 'results': serializer.data})

    @action(detail=True, methods=['get'], url_path='content')
    def content(self, request, pk=None):
        """Stream a profile document inline through the authenticated API."""
        import mimetypes
        import os

        document = self.get_object()
        if not document.document_file:
            return Response(
                {'detail': 'Document file not available.'},
                status=status.HTTP_404_NOT_FOUND,
            )

        try:
            document.document_file.open('rb')
            filename = os.path.basename(document.document_file.name)
            content_type = mimetypes.guess_type(filename)[0] or 'application/octet-stream'
            response = FileResponse(
                document.document_file,
                content_type=content_type,
                as_attachment=False,
                filename=filename,
            )
            response['Cache-Control'] = 'private, no-store'
            response['X-Content-Type-Options'] = 'nosniff'
            return response
        except FileNotFoundError:
            return Response(
                {'detail': 'The stored document file could not be found.'},
                status=status.HTTP_404_NOT_FOUND,
            )

    @action(
        detail=False,
        methods=['post'],
        url_path='extract-metadata',
        parser_classes=[MultiPartParser, FormParser],
    )
    def extract_metadata(self, request):
        """Extract editable metadata from an uploaded PDF or image locally.
        
        ✅ SOFT-CODED: This endpoint is fully optional. If extraction fails
        (e.g., pytesseract not installed in production), frontend can still
        proceed with manual document upload.
        """
        uploaded_file = request.FILES.get('document_file')
        if not uploaded_file:
            return Response(
                {'document_file': ['A document file is required.']},
                status=status.HTTP_400_BAD_REQUEST,
            )

        allowed_extensions = {'.pdf', '.jpg', '.jpeg', '.png'}
        from pathlib import Path
        extension = Path(uploaded_file.name or '').suffix.lower()
        if extension not in allowed_extensions:
            return Response(
                {'document_file': ['Only PDF, JPG, JPEG, and PNG files can be analyzed.']},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if uploaded_file.size > 10 * 1024 * 1024:
            return Response(
                {'document_file': ['The document must not exceed 10 MB.']},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            # ✅ SOFT-CODED: Try importing extractor dependencies
            # If pytesseract/PIL not available in production, return graceful error
            try:
                from .profile_document_extractor import extract_profile_document_metadata
            except ImportError as import_err:
                import logging
                logging.getLogger(__name__).warning(
                    f'Document extraction dependencies not available: {import_err}'
                )
                return Response(
                    {
                        'detail': 'Automatic extraction not available on this server. Please enter document details manually.',
                        'detected_fields': [],
                        'extraction_available': False,
                    },
                    status=status.HTTP_200_OK,  # 200 so frontend continues with upload
                )
            
            result = extract_profile_document_metadata(
                uploaded_file.read(),
                uploaded_file.name,
                uploaded_file.content_type or '',
                request.data.get('document_type', ''),
            )
            result['extraction_available'] = True
            return Response(result)
        except Exception as exc:
            import logging
            logging.getLogger(__name__).exception('Profile document extraction failed')
            # ✅ SOFT-CODED: Return 200 so frontend can continue with manual upload
            return Response(
                {
                    'detail': 'Could not analyze document automatically. Please enter details manually.',
                    'extraction_error': str(exc) if request.user.is_staff else None,
                    'detected_fields': [],
                    'extraction_available': False,
                },
                status=status.HTTP_200_OK,  # 200 so frontend continues with upload
            )
    
    def get_serializer_class(self):
        from .serializers import ProfileDocumentSerializer
        return ProfileDocumentSerializer
    
    def get_queryset(self):
        """Return user's own documents, or all if admin."""
        from .models import ProfileDocument
        from rest_framework.exceptions import ValidationError

        user = self.request.user
        
        # Admins can see all documents
        if self._can_review_documents(self.request):
            queryset = ProfileDocument.objects.select_related(
                'user_profile__user', 'verified_by'
            ).all()

            # Employee detail screens must be able to scope the organization-wide
            # admin queryset to the employee being viewed. Keep accepting the old
            # parameter temporarily so existing clients do not expose every record.
            target_user_id = (
                self.request.query_params.get('user_id')
                or self.request.query_params.get('user_profile__user')
            )
            if target_user_id:
                try:
                    target_user_id = int(target_user_id)
                except (TypeError, ValueError):
                    raise ValidationError({'user_id': 'A valid employee user ID is required.'})
                queryset = queryset.filter(user_profile__user_id=target_user_id)

            return queryset
        
        # Regular users see only their own active documents
        try:
            return ProfileDocument.objects.filter(
                user_profile=user.rbac_profile,
                is_active=True
            )
        except:
            return ProfileDocument.objects.none()
    
    def perform_create(self, serializer):
        """✅ SOFT-CODED: Assign uploads to self, or to an explicitly selected employee for admins."""
        from rest_framework.exceptions import PermissionDenied, ValidationError
        from .models import ProfileDocument
        from .profile_utils import get_or_create_profile, ProfileProvisioningError
        import logging
        
        logger = logging.getLogger(__name__)

        user = self.request.user
        target_user_id = self.request.data.get('target_user_id')

        if target_user_id and str(target_user_id) != str(user.pk):
            if not (user.is_superuser or user.is_staff):
                raise PermissionDenied('Only authorized administrators can upload documents for another employee.')
            try:
                profile = UserProfile.objects.get(user_id=target_user_id)
                logger.info(f'[ProfileDocument] Admin {user.email} uploading document for user {profile.user.email}')
            except (UserProfile.DoesNotExist, ValueError, TypeError):
                raise ValidationError({'target_user_id': 'The selected employee profile does not exist.'})
        else:
            try:
                profile = get_or_create_profile(user, source='ProfileDocumentViewSet')
                logger.info(f'[ProfileDocument] User {user.email} uploading document for themselves')
            except ProfileProvisioningError as exc:
                logger.error(f'[ProfileDocument] Profile provisioning error for user {user.email}: {str(exc)}')
                profile = getattr(self.request, 'user_profile', None)
                if profile is None:
                    raise ValidationError({'detail': str(exc)})
                logger.info(f'[ProfileDocument] Using middleware profile fallback for user {user.email}')

        # Validate that document_file is provided for new uploads
        if 'document_file' not in self.request.FILES:
            logger.warning(f'[ProfileDocument] Upload attempt without file by user {user.email}')
            raise ValidationError({'document_file': 'Document file is required for new uploads'})

        # Replacing a document type must affect the target employee, not the
        # administrator performing the upload.
        document_type = serializer.validated_data.get('document_type')
        existing_docs = ProfileDocument.objects.filter(
            user_profile=profile,
            document_type=document_type,
            is_active=True
        )
        
        if existing_docs.exists():
            logger.info(f'[ProfileDocument] Replacing existing {document_type} document for user {profile.user.email}')
            existing_docs.update(is_active=False)

        try:
            serializer.save(user_profile=profile, is_active=True)
            logger.info(f'[ProfileDocument] Successfully created document {document_type} for user {profile.user.email}')
        except Exception as exc:
            logger.error(f'[ProfileDocument] Failed to save document for user {profile.user.email}: {str(exc)}')
            raise
    
    def perform_update(self, serializer):
        """✅ SOFT-CODED: Update document with enhanced validation and logging."""
        from django.utils import timezone
        import logging
        
        logger = logging.getLogger(__name__)
        
        # Log the update operation for debugging
        instance = serializer.instance
        updated_fields = list(serializer.validated_data.keys())
        logger.info(f'[ProfileDocument] Updating document {instance.id} for user {instance.user_profile.user.email}. Fields: {updated_fields}')
        
        # Check if document_file is being updated
        has_new_file = 'document_file' in serializer.validated_data and serializer.validated_data['document_file'] is not None
        
        if has_new_file:
            logger.info(f'[ProfileDocument] New file uploaded for document {instance.id}')
        
        # Auto-expire if expiry_date is in the past
        expiry_date = serializer.validated_data.get('expiry_date', instance.expiry_date)
        if expiry_date and expiry_date < timezone.now().date():
            logger.info(f'[ProfileDocument] Document {instance.id} expired on {expiry_date}, marking as expired')
            serializer.save(verification_status='expired')
        else:
            serializer.save()
        
        logger.info(f'[ProfileDocument] Successfully updated document {instance.id}')
    
    @action(detail=False, methods=['get'])
    def document_types(self, request):
        """Return available document types from soft-coded config."""
        from apps.rbac.profile_config import get_all_document_types
        return Response(get_all_document_types())
    
    @action(detail=False, methods=['get'], url_path='document-types/profile')
    def document_types_profile(self, request):
        """Return document types for profile page only (identity + health docs)."""
        from apps.rbac.profile_config import get_document_types_for_profile
        return Response(get_document_types_for_profile())
    
    @action(detail=False, methods=['get'], url_path='document-types/onboarding')
    def document_types_onboarding(self, request):
        """Return document types for onboarding page (all employment docs)."""
        from apps.rbac.profile_config import get_document_types_for_onboarding
        return Response(get_document_types_for_onboarding())
    
    @action(detail=False, methods=['get'], url_path='document-types/category/(?P<category>[^/.]+)')
    def document_types_by_category(self, request, category=None):
        """Return document types filtered by category."""
        from apps.rbac.profile_config import get_document_types_by_category
        return Response(get_document_types_by_category(category))
    
    @action(detail=True, methods=['post'], permission_classes=[IsAuthenticated])
    def verify(self, request, pk=None):
        """Admin-only: Verify a document."""
        self._require_document_reviewer(request)
        document = self.get_object()
        if document.verification_status != 'pending':
            return Response(
                {'detail': f'Document is already {document.verification_status}.'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        
        from django.utils import timezone
        
        # Update verification status
        document.verification_status = 'verified'
        document.verified_by = request.user
        document.verified_at = timezone.now()
        document.rejection_reason = ''
        document.save()

        create_audit_log(
            user=request.user,
            action='verify',
            resource_type='ProfileDocument',
            resource_id=document.id,
            resource_repr=str(document),
            changes={'verification_status': {'old': 'pending', 'new': 'verified'}},
            metadata={'document_owner': document.user_profile.user.email},
            ip_address=request.META.get('REMOTE_ADDR'),
            user_agent=request.META.get('HTTP_USER_AGENT', ''),
        )
        
        serializer = self.get_serializer(document)
        return Response(serializer.data)
    
    @action(detail=True, methods=['post'], permission_classes=[IsAuthenticated])
    def reject(self, request, pk=None):
        """Admin-only: Reject a document."""
        self._require_document_reviewer(request)
        document = self.get_object()
        if document.verification_status != 'pending':
            return Response(
                {'detail': f'Document is already {document.verification_status}.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        reason = str(request.data.get('reason', '')).strip()
        if not reason:
            return Response(
                {'reason': ['A rejection reason is required.']},
                status=status.HTTP_400_BAD_REQUEST,
            )
        
        # Update rejection
        document.verification_status = 'rejected'
        document.rejection_reason = reason
        document.save()

        create_audit_log(
            user=request.user,
            action='reject',
            resource_type='ProfileDocument',
            resource_id=document.id,
            resource_repr=str(document),
            changes={
                'verification_status': {'old': 'pending', 'new': 'rejected'},
                'rejection_reason': reason,
            },
            metadata={'document_owner': document.user_profile.user.email},
            ip_address=request.META.get('REMOTE_ADDR'),
            user_agent=request.META.get('HTTP_USER_AGENT', ''),
        )
        
        serializer = self.get_serializer(document)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def expiring_soon(self, request):
        """Get documents expiring within specified days (default 30)."""
        from django.utils import timezone
        from datetime import timedelta
        from .models import ProfileDocument
        
        days = int(request.query_params.get('days', 30))
        threshold_date = timezone.now().date() + timedelta(days=days)
        
        queryset = self.get_queryset().filter(
            expiry_date__lte=threshold_date,
            expiry_date__gte=timezone.now().date(),
            verification_status='verified',
            is_active=True
        )
        
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def my_documents(self, request):
        """Get current user's active documents grouped by type."""
        try:
            profile = request.user.rbac_profile
            from .models import ProfileDocument
            from apps.rbac.profile_config import DOCUMENT_TYPES
            
            documents_by_type = {}
            for doc_code in DOCUMENT_TYPES.keys():
                doc = ProfileDocument.objects.filter(
                    user_profile=profile,
                    document_type=doc_code,
                    is_active=True
                ).first()
                
                if doc:
                    serializer = self.get_serializer(doc)
                    documents_by_type[doc_code] = serializer.data
                else:
                    documents_by_type[doc_code] = None
            
            return Response(documents_by_type)
        except:
            return Response({})

