"""
Excel Parser Service for Electrical Technical Datasheets
Parses uploaded Excel files into structured data for quality checking
Handles multiple equipment types: UPS, VFD, Cables, NER, etc.
"""

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from typing import Dict, List, Any, Optional, Tuple
import re
from datetime import datetime
import logging

logger = logging.getLogger(__name__)


class ExcelParserService:
    """
    Service to parse electrical technical datasheet Excel files
    Extracts document control info, technical data, and metadata
    """
    
    def __init__(self, file_path: str):
        """
        Initialize parser with file path
        
        Args:
            file_path: Path to Excel file (.xlsx)
        """
        self.file_path = file_path
        self.workbook: Optional[openpyxl.Workbook] = None
        self.parsed_data: Dict[str, Any] = {}
        self.equipment_type: str = 'unknown'
        
    def parse(self) -> Dict[str, Any]:
        """
        Main parsing method - parses entire Excel file
        
        Returns:
            Dictionary containing all parsed data
        """
        try:
            self.workbook = openpyxl.load_workbook(self.file_path, data_only=True)
            
            # Parse structure
            self.parsed_data = {
                'sheet_names': self.workbook.sheetnames,
                'sheets': {},
                'document_control': {},
                'revision_history': [],
                'holds': [],
                'technical_data': {},
                'equipment_type': 'unknown',
                'parsing_metadata': {
                    'parsed_at': datetime.now().isoformat(),
                    'total_sheets': len(self.workbook.sheetnames),
                }
            }
            
            # Detect equipment type from document title
            self.equipment_type = self._detect_equipment_type()
            self.parsed_data['equipment_type'] = self.equipment_type
            
            # Parse each sheet
            for sheet_name in self.workbook.sheetnames:
                sheet = self.workbook[sheet_name]
                sheet_type = self._classify_sheet(sheet_name, sheet)
                
                if sheet_type == 'cover':
                    self.parsed_data['document_control'] = self._parse_cover_sheet(sheet)
                elif sheet_type == 'revision_history':
                    self.parsed_data['revision_history'] = self._parse_revision_history(sheet)
                elif sheet_type == 'holds':
                    self.parsed_data['holds'] = self._parse_holds_sheet(sheet)
                elif sheet_type == 'technical_data':
                    tech_data = self._parse_technical_data_sheet(sheet, sheet_name)
                    self.parsed_data['technical_data'][sheet_name] = tech_data
                    
                # Store raw sheet info
                self.parsed_data['sheets'][sheet_name] = {
                    'type': sheet_type,
                    'max_row': sheet.max_row,
                    'max_column': sheet.max_column,
                }
            
            return self.parsed_data
            
        except Exception as e:
            logger.error(f"Error parsing Excel file {self.file_path}: {str(e)}")
            raise
        finally:
            if self.workbook:
                self.workbook.close()
    
    def _detect_equipment_type(self) -> str:
        """
        Detect equipment type from document title
        
        Returns:
            Equipment type identifier
        """
        # Try to find title in first few sheets
        for sheet_name in self.workbook.sheetnames[:3]:
            sheet = self.workbook[sheet_name]
            
            for row in sheet.iter_rows(max_row=30, values_only=True):
                for cell in row:
                    if cell and isinstance(cell, str):
                        cell_upper = cell.upper()
                        
                        # Check for equipment keywords
                        if 'UPS' in cell_upper or 'UNINTERRUPTIBLE POWER' in cell_upper:
                            return 'ups'
                        elif 'VFD' in cell_upper or 'VARIABLE FREQUENCY DRIVE' in cell_upper or 'VSD' in cell_upper:
                            return 'vfd'
                        elif 'NEUTRAL EARTHING RESISTOR' in cell_upper or 'NER' in cell_upper:
                            return 'ner'
                        elif 'POWER CABLE' in cell_upper and 'LV' in cell_upper:
                            return 'power_cable'
                        elif 'CONTROL CABLE' in cell_upper:
                            return 'control_cable'
                        elif 'EARTHING CABLE' in cell_upper:
                            return 'earthing_cable'
                        elif 'TRANSFORMER' in cell_upper:
                            return 'transformer'
                        elif 'MOTOR' in cell_upper:
                            return 'motor'
                        elif 'SWITCHGEAR' in cell_upper:
                            return 'switchgear'
        
        return 'unknown'
    
    def _classify_sheet(self, sheet_name: str, sheet: Worksheet) -> str:
        """
        Classify sheet type based on name and content
        
        Args:
            sheet_name: Name of the sheet
            sheet: Worksheet object
            
        Returns:
            Sheet type classification
        """
        name_lower = sheet_name.lower()
        
        # Check by name first
        if 'cover' in name_lower or 'title' in name_lower or sheet_name == 'Sheet1':
            return 'cover'
        elif 'revision' in name_lower or 'history' in name_lower:
            return 'revision_history'
        elif 'hold' in name_lower:
            return 'holds'
        elif 'toc' in name_lower or 'table of contents' in name_lower or 'contents' in name_lower:
            return 'toc'
        elif 'note' in name_lower:
            return 'notes'
        elif 'abbreviation' in name_lower:
            return 'abbreviations'
        elif 'technical' in name_lower or 'data' in name_lower or 'specification' in name_lower:
            return 'technical_data'
        
        # Check content for technical data pattern
        if self._has_technical_data_structure(sheet):
            return 'technical_data'
        
        return 'other'
    
    def _has_technical_data_structure(self, sheet: Worksheet) -> bool:
        """
        Check if sheet has technical data table structure
        (DESCRIPTION, UNIT, SPECIFIED DESIGN DATA, VENDOR DATA columns)
        
        Args:
            sheet: Worksheet to check
            
        Returns:
            True if sheet has technical data structure
        """
        for row in sheet.iter_rows(max_row=20, values_only=True):
            row_text = ' '.join([str(cell).upper() for cell in row if cell])
            
            if 'DESCRIPTION' in row_text and ('SPECIFIED' in row_text or 'DESIGN' in row_text):
                return True
            if 'DESCRIPTION' in row_text and 'VENDOR' in row_text:
                return True
        
        return False
    
    def _parse_cover_sheet(self, sheet: Worksheet) -> Dict[str, Any]:
        """
        Parse cover sheet for document control information
        
        Args:
            sheet: Cover sheet worksheet
            
        Returns:
            Dictionary of document control fields
        """
        doc_control = {
            'company_doc_number': '',
            'contractor_doc_number': '',
            'rejlers_doc_number': '',
            'document_title': '',
            'classification_code': '',
            'revision': '',
            'doc_status': '',
            'doc_purpose': '',
            'project_name': '',
            'project_location': '',
            'agreement_number': '',
            'company': '',
        }
        
        # Scan first 50 rows for key-value pairs
        for row_idx in range(1, min(sheet.max_row + 1, 50)):
            for col_idx in range(1, min(sheet.max_column + 1, 10)):
                cell = sheet.cell(row=row_idx, column=col_idx)
                
                if not cell.value:
                    continue
                
                cell_text = str(cell.value).strip()
                cell_upper = cell_text.upper()
                
                # Try to find value in adjacent cells
                value_cell = sheet.cell(row=row_idx, column=col_idx + 1)
                value = str(value_cell.value).strip() if value_cell.value else ''
                
                # Also check below
                value_below_cell = sheet.cell(row=row_idx + 1, column=col_idx)
                value_below = str(value_below_cell.value).strip() if value_below_cell.value else ''
                
                # Extract fields
                if 'COMPANY DOCUMENT NUMBER' in cell_upper or 'COMPANY DOC' in cell_upper:
                    doc_control['company_doc_number'] = value or value_below
                elif 'CONTRACTOR DOCUMENT NUMBER' in cell_upper or 'CONTRACTOR DOC' in cell_upper:
                    doc_control['contractor_doc_number'] = value or value_below
                elif 'REJLERS' in cell_upper and 'DOCUMENT NUMBER' in cell_upper:
                    doc_control['rejlers_doc_number'] = value or value_below
                elif 'DOCUMENT TITLE' in cell_upper:
                    doc_control['document_title'] = value or value_below
                elif 'CLASSIFICATION CODE' in cell_upper:
                    doc_control['classification_code'] = value or value_below
                elif 'REVISION' in cell_upper and 'HISTORY' not in cell_upper:
                    doc_control['revision'] = value or value_below
                elif 'DOC STATUS' in cell_upper or 'DOCUMENT STATUS' in cell_upper:
                    doc_control['doc_status'] = value or value_below
                elif 'DOC PURPOSE' in cell_upper or 'DOCUMENT PURPOSE' in cell_upper:
                    doc_control['doc_purpose'] = value or value_below
                elif 'PROJECT NAME' in cell_upper or 'PROJECT' in cell_upper and 'TITLE' in cell_upper:
                    doc_control['project_name'] = value or value_below
                elif 'LOCATION' in cell_upper and len(cell_text) < 50:
                    doc_control['project_location'] = value or value_below
                elif 'AGREEMENT' in cell_upper and 'NUMBER' in cell_upper:
                    doc_control['agreement_number'] = value or value_below
                elif 'COMPANY' in cell_upper and len(cell_text) < 30:
                    if not doc_control['company']:  # Only set if not already set
                        doc_control['company'] = value or value_below
        
        return doc_control
    
    def _parse_revision_history(self, sheet: Worksheet) -> List[Dict[str, str]]:
        """
        Parse revision history sheet
        
        Args:
            sheet: Revision history worksheet
            
        Returns:
            List of revision entries
        """
        revisions = []
        
        # Find header row
        header_row_idx = None
        for row_idx in range(1, min(sheet.max_row + 1, 20)):
            row_values = [str(cell.value).upper() if cell.value else '' 
                         for cell in sheet[row_idx]]
            row_text = ' '.join(row_values)
            
            if 'REV' in row_text and 'DATE' in row_text and 'DESCRIPTION' in row_text:
                header_row_idx = row_idx
                break
        
        if not header_row_idx:
            return revisions
        
        # Find column indices
        header_row = sheet[header_row_idx]
        rev_col = date_col = desc_col = None
        
        for col_idx, cell in enumerate(header_row, start=1):
            if cell.value:
                cell_upper = str(cell.value).upper()
                if 'REV' in cell_upper and not rev_col:
                    rev_col = col_idx
                elif 'DATE' in cell_upper and not date_col:
                    date_col = col_idx
                elif 'DESCRIPTION' in cell_upper and not desc_col:
                    desc_col = col_idx
        
        # Parse data rows
        for row_idx in range(header_row_idx + 1, sheet.max_row + 1):
            rev_value = sheet.cell(row=row_idx, column=rev_col).value if rev_col else None
            date_value = sheet.cell(row=row_idx, column=date_col).value if date_col else None
            desc_value = sheet.cell(row=row_idx, column=desc_col).value if desc_col else None
            
            if rev_value or date_value or desc_value:
                revisions.append({
                    'revision': str(rev_value) if rev_value else '',
                    'date': str(date_value) if date_value else '',
                    'description': str(desc_value) if desc_value else '',
                })
        
        return revisions
    
    def _parse_holds_sheet(self, sheet: Worksheet) -> List[Dict[str, str]]:
        """
        Parse holds sheet
        
        Args:
            sheet: Holds worksheet
            
        Returns:
            List of hold entries or ["NIL"]
        """
        holds = []
        
        # Check for NIL
        for row in sheet.iter_rows(max_row=10, values_only=True):
            for cell in row:
                if cell and 'NIL' in str(cell).upper():
                    return [{'status': 'NIL'}]
        
        # Find header row and parse holds
        header_row_idx = None
        for row_idx in range(1, min(sheet.max_row + 1, 20)):
            row_values = [str(cell.value).upper() if cell.value else '' 
                         for cell in sheet[row_idx]]
            row_text = ' '.join(row_values)
            
            if 'HOLD' in row_text and ('NUMBER' in row_text or 'DESCRIPTION' in row_text):
                header_row_idx = row_idx
                break
        
        if header_row_idx:
            for row_idx in range(header_row_idx + 1, min(sheet.max_row + 1, 50)):
                row_data = [cell.value for cell in sheet[row_idx]]
                if any(row_data):
                    holds.append({
                        'data': [str(cell) if cell else '' for cell in row_data]
                    })
        
        return holds if holds else [{'status': 'NIL'}]
    
    def _parse_technical_data_sheet(self, sheet: Worksheet, sheet_name: str) -> Dict[str, Any]:
        """
        Parse technical data sheet with DESCRIPTION/UNIT/SPECIFIED/VENDOR structure
        
        Args:
            sheet: Technical data worksheet
            sheet_name: Name of the sheet
            
        Returns:
            Dictionary of parsed technical data
        """
        technical_data = {
            'sheet_name': sheet_name,
            'sections': {},
            'items': [],
            'header_row': None,
            'column_mapping': {},
        }
        
        # Find header row with DESCRIPTION, UNIT, SPECIFIED, VENDOR
        header_row_idx = None
        for row_idx in range(1, min(sheet.max_row + 1, 30)):
            row_values = [str(cell.value).upper() if cell.value else '' 
                         for cell in sheet[row_idx]]
            row_text = ' '.join(row_values)
            
            if 'DESCRIPTION' in row_text and ('SPECIFIED' in row_text or 'VENDOR' in row_text):
                header_row_idx = row_idx
                break
        
        if not header_row_idx:
            return technical_data
        
        technical_data['header_row'] = header_row_idx
        
        # Map column indices
        header_row = sheet[header_row_idx]
        for col_idx, cell in enumerate(header_row, start=1):
            if cell.value:
                cell_upper = str(cell.value).upper()
                
                if 'SL' in cell_upper and 'NO' in cell_upper:
                    technical_data['column_mapping']['sl_no'] = col_idx
                elif 'DESCRIPTION' in cell_upper:
                    technical_data['column_mapping']['description'] = col_idx
                elif 'UNIT' in cell_upper:
                    technical_data['column_mapping']['unit'] = col_idx
                elif 'SPECIFIED' in cell_upper or 'DESIGN DATA' in cell_upper:
                    technical_data['column_mapping']['specified'] = col_idx
                elif 'VENDOR' in cell_upper:
                    technical_data['column_mapping']['vendor'] = col_idx
        
        # Parse data rows
        current_section = 'GENERAL'
        
        for row_idx in range(header_row_idx + 1, sheet.max_row + 1):
            cols = technical_data['column_mapping']
            
            sl_no = sheet.cell(row=row_idx, column=cols.get('sl_no', 1)).value
            description = sheet.cell(row=row_idx, column=cols.get('description', 2)).value
            unit = sheet.cell(row=row_idx, column=cols.get('unit', 3)).value
            specified = sheet.cell(row=row_idx, column=cols.get('specified', 4)).value
            vendor = sheet.cell(row=row_idx, column=cols.get('vendor', 5)).value
            
            # Skip completely empty rows
            if not any([sl_no, description, unit, specified, vendor]):
                continue
            
            description_text = str(description).strip() if description else ''
            
            # Check if this is a section header (bold, all caps, or specific pattern)
            is_section_header = False
            if description_text and not specified and not vendor:
                if description_text.upper() == description_text and len(description_text) > 5:
                    is_section_header = True
                    current_section = description_text
            
            item = {
                'row_number': row_idx,
                'sl_no': str(sl_no) if sl_no else '',
                'description': description_text,
                'unit': str(unit).strip() if unit else '',
                'specified_design_data': str(specified).strip() if specified else '',
                'vendor_data': str(vendor).strip() if vendor else '',
                'section': current_section,
                'is_section_header': is_section_header,
            }
            
            technical_data['items'].append(item)
            
            # Group by section
            if current_section not in technical_data['sections']:
                technical_data['sections'][current_section] = []
            technical_data['sections'][current_section].append(item)
        
        return technical_data
    
    def extract_document_control_summary(self) -> Dict[str, str]:
        """
        Extract key document control fields for database storage
        
        Returns:
            Dictionary of document control fields
        """
        if not self.parsed_data:
            return {}
        
        doc_control = self.parsed_data.get('document_control', {})
        return {
            'company_doc_number': doc_control.get('company_doc_number', ''),
            'contractor_doc_number': doc_control.get('contractor_doc_number', ''),
            'rejlers_doc_number': doc_control.get('rejlers_doc_number', ''),
            'document_title': doc_control.get('document_title', ''),
            'classification_code': doc_control.get('classification_code', ''),
            'revision': doc_control.get('revision', ''),
            'doc_status': doc_control.get('doc_status', ''),
            'doc_purpose': doc_control.get('doc_purpose', ''),
            'project_name': doc_control.get('project_name', ''),
            'project_location': doc_control.get('project_location', ''),
            'agreement_number': doc_control.get('agreement_number', ''),
        }
