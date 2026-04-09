"""
11KV Switchgear Datasheet Generator from SLD Documents
Extracts equipment data and generates comprehensive datasheets
"""
import logging
import re
import json
from typing import Dict, List, Optional
from django.conf import settings
from openai import OpenAI
import PyPDF2

logger = logging.getLogger(__name__)


class SwitchgearDatasheetGenerator:
    """Generate 11KV switchgear datasheets from SLD documents"""
    
    def __init__(self):
        self.client = OpenAI(api_key=settings.OPENAI_API_KEY)
    
    def extract_text_from_pdf(self, pdf_file) -> str:
        """Extract text from uploaded PDF file"""
        try:
            # Reset file pointer to beginning
            pdf_file.seek(0)
            
            pdf_reader = PyPDF2.PdfReader(pdf_file)
            text = ""
            
            logger.info(f"[SwitchgearDatasheet] PDF has {len(pdf_reader.pages)} pages")
            
            for page_num, page in enumerate(pdf_reader.pages, 1):
                page_text = page.extract_text()
                if page_text:
                    text += page_text + "\n"
                    logger.info(f"[SwitchgearDatasheet] Page {page_num}: Extracted {len(page_text)} chars")
                else:
                    logger.warning(f"[SwitchgearDatasheet] Page {page_num}: No text extracted (might be image-based)")
            
            logger.info(f"[SwitchgearDatasheet] Total extracted text: {len(text)} characters")
            return text
        except Exception as e:
            logger.error(f"[SwitchgearDatasheet] PDF extraction error: {e}", exc_info=True)
            return ""
    
    def generate_datasheet_from_sld(self, pdf_file, project_info: Dict = None) -> Dict:
        """
        Generate comprehensive 11KV switchgear datasheet from SLD PDF
        
        Args:
            pdf_file: Uploaded PDF file
            project_info: Optional project metadata
        
        Returns:
            {
                'success': bool,
                'datasheet_rows': List[Dict],
                'summary': Dict,
                'extraction_metadata': Dict
            }
        """
        try:
            # ── Step 1: Always start with the full template ──────────────────────
            # This guarantees unit and required_data are ALWAYS populated.
            datasheet_rows = self._get_default_datasheet_template()

            # ── Step 2: Extract text from the uploaded PDF ───────────────────────
            logger.info("[SwitchgearDatasheet] Extracting text from SLD PDF...")
            doc_text = self.extract_text_from_pdf(pdf_file)

            # ── Step 3: If text available, use AI to extract vendor values ────────
            if doc_text and len(doc_text) >= 20:
                logger.info(f"[SwitchgearDatasheet] {len(doc_text)} chars extracted — running AI vendor extraction...")
                vendor_map = self._extract_vendor_data_with_ai(doc_text, project_info)
                if vendor_map:
                    merged = self._merge_vendor_data(datasheet_rows, vendor_map)
                    logger.info(f"[SwitchgearDatasheet] Merged {merged} vendor values into template")
            else:
                logger.warning("[SwitchgearDatasheet] No/insufficient text from PDF — showing template with empty vendor data")
                doc_text = ""

            summary = {
                'total_rows': len(datasheet_rows),
                'equipment_count': sum(1 for r in datasheet_rows if r.get('description', '').strip()),
                'completed_fields': sum(1 for r in datasheet_rows if r.get('vendor_data', '').strip()),
                'missing_fields': sum(1 for r in datasheet_rows if not r.get('vendor_data', '').strip()),
            }

            logger.info(f"[SwitchgearDatasheet] ✅ {summary['total_rows']} rows | {summary['completed_fields']} vendor values filled")
            return {
                'success': True,
                'datasheet_rows': datasheet_rows,
                'summary': summary,
                'extraction_metadata': {
                    'document_length': len(doc_text),
                    'project_info': project_info or {}
                }
            }

        except Exception as e:
            logger.error(f"[SwitchgearDatasheet] Error: {e}", exc_info=True)
            return {'success': False, 'error': str(e)}
    
    def _extract_datasheet_with_ai(self, sld_text: str, project_info: Dict = None) -> List[Dict]:
        """Use AI to extract structured datasheet data from SLD text"""
        
        extraction_prompt = f"""You are an expert electrical engineer specializing in 11KV switchgear systems. 
Analyze the provided Single Line Diagram (SLD) document and extract comprehensive equipment datasheet information.

PROJECT INFORMATION:
{json.dumps(project_info or {}, indent=2)}

SLD DOCUMENT CONTENT:
{sld_text[:6000]}

TASK: Extract and structure ALL equipment data into a comprehensive datasheet format with EXACTLY 6 fields per row.

The datasheet MUST follow this exact column structure (same as the physical MV switchgear datasheet form):
- SR_NO: Sequential item number (e.g. 1.0, 1.1, 2, 2.1 ...) — blank for section header rows
- DESCRIPTION: Parameter name or section header
- UNIT: Engineering unit for the parameter (e.g. kV, A, kA, Hz, ℃, mm, kg, %) — blank if not applicable
- REQUIRED_DATA: Specification/standard requirement value filled by the engineer
- VENDOR_DATA: Actual value extracted from the uploaded SLD/document — blank string if not found
- REV: Revision marker — leave as empty string unless a specific revision is noted in the document

Cover ALL the following sections:
1. GENERAL — Equipment tag, service description
2. REFERENCE — Applicable international standards (IEC 60298, IEC 60694, IEC 60255, IEC 60529), ADNOC specs
3. SITE DATA — Location, area classification, climate, altitude, min/max ambient temperature, humidity
4. GENERAL CHARACTERISTICS — Type of switchgear, circuit breaker type, standards, system voltage, frequency, phases, earthing
5. RATINGS AND SHORT CIRCUIT DATA — Rated insulation voltage, rated voltage, rated normal current, SC breaking current, peak withstand current, short time withstand current, power frequency withstand voltage, impulse withstand voltage
6. CONSTRUCTION — Type, IP rating, colour, arc fault classification
7. BUSBAR — Material, shape, busbar rating
8. CIRCUIT BREAKER — Type, operating mechanism, auxiliary supply voltage, number of operating cycles
9. CURRENT TRANSFORMER — Type, number of CT cores, CT ratio, CT class
10. VOLTAGE TRANSFORMER — Type, VT ratio, VT class
11. EARTHING — Main earthing bar, earth fault relay
12. PROTECTION & CONTROL — Protection relay type, protection functions, metering
13. AUXILIARY EQUIPMENT — Anti-condensation heater, space heater rating, lighting
14. MANUFACTURER — Name, model/type, country of origin

Return your response as a JSON array where each object has EXACTLY this structure:
{{
    "sr_no": "<sequential number or empty string for section headers>",
    "description": "<parameter name or section header>",
    "unit": "<engineering unit or empty string>",
    "required_data": "<specification requirement value>",
    "vendor_data": "<value extracted from SLD/document, or empty string>",
    "rev": ""
}}

IMPORTANT GUIDELINES:
- ALWAYS include the "unit" field — use the correct SI/engineering unit for every measured quantity
- Common units: voltage → kV, current → A, breaking current → kA, frequency → Hz, temperature → ℃, dimensions → mm, weight → kg, percentage → %
- Section header rows have blank sr_no, blank unit, blank required_data, blank vendor_data, blank rev
- Extract ACTUAL values from the SLD document for vendor_data where available
- Leave vendor_data as empty string "" when value not found in document
- Leave rev as empty string "" unless document contains a specific revision reference
- Be comprehensive — include ALL standard 11KV switchgear parameters

Return ONLY the JSON array, no additional text."""

        try:
            response = self.client.chat.completions.create(
                model="gpt-4o",
                messages=[
                    {"role": "system", "content": "You are an expert electrical engineer specializing in switchgear datasheets. Extract comprehensive equipment data and return only valid JSON."},
                    {"role": "user", "content": extraction_prompt}
                ],
                temperature=0.2,
                max_tokens=4000
            )
            
            ai_response = response.choices[0].message.content
            
            # Parse JSON response
            # Remove markdown code blocks if present
            if '```json' in ai_response:
                ai_response = ai_response.split('```json')[1].split('```')[0]
            elif '```' in ai_response:
                ai_response = ai_response.split('```')[1].split('```')[0]
            
            datasheet_rows = json.loads(ai_response.strip())
            
            # Validate structure
            if isinstance(datasheet_rows, list) and len(datasheet_rows) > 0:
                # Ensure all rows have required fields
                for i, row in enumerate(datasheet_rows):
                    if 'sr_no' not in row:
                        row['sr_no'] = str(i + 1)
                    if 'description' not in row:
                        row['description'] = ''
                    if 'unit' not in row:
                        row['unit'] = ''
                    if 'required_data' not in row:
                        row['required_data'] = ''
                    if 'vendor_data' not in row:
                        row['vendor_data'] = ''
                    if 'rev' not in row:
                        row['rev'] = ''
                    # Remove legacy 'remarks' key if present (replaced by 'rev')
                    row.pop('remarks', None)
                
                return datasheet_rows
            else:
                logger.error("[SwitchgearDatasheet] Invalid AI response structure")
                return self._get_default_datasheet_template()
                
        except json.JSONDecodeError as e:
            logger.error(f"[SwitchgearDatasheet] JSON decode error: {e}")
            logger.error(f"AI Response: {ai_response[:500]}")
            return self._get_default_datasheet_template()
        except Exception as e:
            logger.error(f"[SwitchgearDatasheet] AI extraction error: {e}")
            return self._get_default_datasheet_template()
    
    def _extract_vendor_data_with_ai(self, doc_text: str, project_info: Dict = None) -> Dict:
        """
        Use GPT-4o to extract parameter values from the uploaded document.
        Returns a dict mapping NORMALIZED_DESCRIPTION_UPPER → extracted_value.
        """
        prompt = f"""You are a senior electrical engineer specialising in MV 11kV Switchgear. Read the following technical document and extract every parameter value you can find.

DOCUMENT:
{doc_text[:8000]}

TASK:
Return a JSON object where:
- Keys   = parameter/field names in UPPERCASE (e.g. "RATED VOLTAGE", "RATED CURRENT", "MANUFACTURER")
- Values = the extracted value as a string

Focus on: equipment tags, voltages (kV), currents (A), short circuit ratings (kA), frequency (Hz), IP rating, dimensions, weights, temperatures, manufacturer names, circuit breaker type, CT/VT ratios, protection relay type.

Rules:
- Include ONLY fields actually present in the document.
- Do NOT invent or assume values.
- Return ONLY a valid JSON object — no explanation, no markdown.

Example:
{{"RATED VOLTAGE": "11", "RATED CURRENT": "1000", "MANUFACTURER": "SCHNEIDER", "FREQUENCY": "50", "SHORT CIRCUIT CURRENT": "25"}}"""

        try:
            response = self.client.chat.completions.create(
                model="gpt-4o",
                messages=[
                    {"role": "system", "content": "You are an expert electrical engineer. Extract parameter values from MV switchgear documents. Return only valid JSON."},
                    {"role": "user", "content": prompt},
                ],
                temperature=0.1,
                max_tokens=2000,
            )
            ai_response = response.choices[0].message.content.strip()
            if "```json" in ai_response:
                ai_response = ai_response.split("```json")[1].split("```")[0]
            elif "```" in ai_response:
                ai_response = ai_response.split("```")[1].split("```")[0]
            vendor_map = json.loads(ai_response.strip())
            if isinstance(vendor_map, dict):
                normalized = {k.strip().upper(): str(v).strip() for k, v in vendor_map.items() if v and str(v).strip()}
                logger.info(f"[SwitchgearDatasheet] AI extracted {len(normalized)} vendor key-value pairs")
                return normalized
            return {}
        except json.JSONDecodeError as e:
            logger.error(f"[SwitchgearDatasheet] Vendor JSON decode error: {e}")
            return {}
        except Exception as e:
            logger.error(f"[SwitchgearDatasheet] AI extraction error: {e}")
            return {}

    def _merge_vendor_data(self, template_rows: List[Dict], vendor_map: Dict) -> int:
        """
        Merge vendor_map values into template_rows by description matching.
        Returns count of rows filled.
        """
        merged = 0
        stop_words = {'AND', 'OR', 'OF', 'THE', 'WITH', 'FOR', 'AT', 'IN', 'TO', 'A', 'AN', 'BY', 'ON'}
        for row in template_rows:
            if row.get('vendor_data'):
                continue
            desc = row.get('description', '').strip().upper()
            if not desc:
                continue
            # 1. Exact match
            if desc in vendor_map:
                row['vendor_data'] = vendor_map[desc]
                merged += 1
                continue
            # 2. Containment match
            matched = False
            for ai_key, ai_val in vendor_map.items():
                if not ai_key or not ai_val:
                    continue
                if ai_key in desc or desc in ai_key:
                    row['vendor_data'] = ai_val
                    merged += 1
                    matched = True
                    break
            if matched:
                continue
            # 3. Word-overlap match (≥2 meaningful common words)
            desc_words = set(desc.split()) - stop_words
            for ai_key, ai_val in vendor_map.items():
                if not ai_key or not ai_val:
                    continue
                ai_words = set(ai_key.upper().split()) - stop_words
                if len(desc_words & ai_words) >= 2:
                    row['vendor_data'] = ai_val
                    merged += 1
                    break
        return merged

    def _get_default_datasheet_template(self) -> List[Dict]:
        """Return full ADNOC MV 11kV Switchgear datasheet template (6-column)."""
        R = lambda sr, desc, unit="", req="": {
            "sr_no": sr, "description": desc, "unit": unit,
            "required_data": req, "vendor_data": "", "rev": ""
        }
        H = lambda desc: {
            "sr_no": "", "description": desc, "unit": "",
            "required_data": "", "vendor_data": "", "rev": ""
        }
        return [
            # ── GENERAL ──────────────────────────────────────────────────────
            H("GENERAL"),
            R("1",    "EQUIPMENT TAG NO."),
            R("2",    "SERVICE",                                        "",        "CONTINUOUS"),
            R("3",    "MAKE"),

            # ── REFERENCE SPECIFICATION ──────────────────────────────────────
            H("REFERENCE SPECIFICATION"),
            R("1.1",  "DESIGN GUIDELINES",                             "",        "BGS-EU-001"),
            R("1.2",  "HIGH VOLTAGE SWITCHGEAR AND CONTROL GEAR",      "",        "BGS-EE-004"),
            R("1.3",  "INTEGRATED PROTECTION AND CONTROL SYSTEM (IPCS)","",       "BGS-EE-010"),
            R("1.3.1","GENERAL REQUIREMENT FOR IPCS",                  "",        "REFER CLAUSE NO. 14.1 OF BGS-EE-010"),
            R("1.3.2","CONFIGURATION AND COMMUNICATION OF IPCS",       "",        "AS PER CLAUSE NO. 14.2 OF BGS-EE-010"),
            R("1.3.3","IPCS NETWORK DETAILS",                          "",        "REFER CLAUSE NO. 14.3 OF BGS-EE-010"),
            R("1.4",  "POWER, CONTROL AND EARTHING CABLES",            "",        "BGS-EE-011"),
            R("1.5",  "FIELD COMMISSIONING OF ELECTRICAL INSTALLATION AND EQUIPMENT", "", "BGS-EE-020"),

            # ── REFERENCE DRAWING ────────────────────────────────────────────
            H("REFERENCE DRAWING"),
            R("2.1",  "KEY SINGLE LINE DIAGRAM – HYDROGEN EXTRACTION UNIT",      "", "15-EE-133-00001"),
            R("2.2",  "TYPICAL PROTECTION AND METERING ONE LINE DIAGRAMS: MV 11KV AND 3.3KV SWITCHGEAR", "", "RUE/EE: P3/G0409"),
            R("2.3",  "ELECTRICAL LOAD LIST – EU-H3 EXTRACTION UNIT",            "", "15-EE-359-00001"),

            # ── C. SITE DATA ─────────────────────────────────────────────────
            H("C   SITE DATA"),
            R("1",    "TYPE OF INSTALLATION",                          "",        "INDOOR IN AIR CONDITIONED BUILDING"),
            R("2",    "ATMOSPHERE",                                    "",        "INDOOR IN AIR CONDITIONED BUILDING, SAFE AREA"),
            R("3",    "DESIGN AMBIENT TEMPERATURE",                    "degC",    "60"),
            R("4",    "ALTITUDE",                                      "m",       "< 1000"),
            R("5",    "MAXIMUM RELATIVE HUMIDITY AT 43°C",             "%",       "85"),
            R("6",    "MAXIMUM RELATIVE HUMIDITY AT 54°C",             "%",       "89"),
            R("7",    "SITE CLASS DEFINITION",                         "",        "SITE CLASS C"),
            R("8",    "CRITICALITY RATING"),
            R("9",    "INSPECTION CLASS",                              "",        "1"),

            # ── D. GENERAL CHARACTERISTICS ──────────────────────────────────
            H("D   GENERAL CHARACTERISTICS"),
            R("1",    "MANUFACTURER TYPE DESIGNATION",                 "",        "B9A"),
            R("2",    "APPLICABLE INTERNATIONAL STANDARDS",            "",        "IEC 62271-200, BGS-EE-004"),
            R("3",    "FREQUENCY",                                     "Hz",      "50"),
            R("3.1",  "STEADY STATE VARIATION",                        "",        "± 1%"),
            R("3.2",  "TRANSIENT VARIATION",                           "",        "± 8%"),
            R("4.1",  "OPERATING VOLTAGE",                             "V",       "11000"),
            R("4.2",  "STEADY STATE VARIATION",                        "",        "± 16%"),
            R("4.3",  "TRANSIENT VARIATION",                           "",        "+16%, -20%"),
            R("5",    "RATED VOLTAGE",                                 "kV",      "12"),
            R("6",    "POWER FREQUENCY WITHSTAND VOLTAGE",             "kV",      "28"),
            R("7",    "LIGHTNING IMPULSE DE WITHSTAND VOLTAGE",        "kV",      "75"),
            R("8",    "MAXIMUM AMBIENT TEMPERATURE",                   "DEG C",   "40 C"),
            R("9",    "RATED CURRENT OF BUSBAR",                       "A",       "1000A (REFER NOTE 15)"),
            R("10",   "RATED CURRENT OF BRANCH BARS",                  "A",       "*** (REFER NOTE 15)"),
            R("11",   "RATED SHORT TIME CURRENT",                      "kA"),
            R("12",   "RATED COMBINATION OF SHORT CIRCUIT",            "kA"),
            R("13",   "RATED PEAK WITHSTAND CURRENT",                  "kA"),
            R("14",   "EARTH BAR CROSS SECTION",                       "SQ MM"),
            R("15",   "BUS BAR CROSS SECTION",                         "SQ MM",   "MINIMUM 240SQM"),
            R("16",   "BUS BAR FINISHING",                             "",        "FULLY INSULATED"),
            R("17",   "SYSTEM EARTHING",                               "",        "RESISTANCE EARTH"),

            # ── E. CONSTRUCTION CHARACTERISTICS ─────────────────────────────
            H("E   CONSTRUCTION CHARACTERISTICS"),
            R("1",    "TYPE OF ENCLOSURE",                             "",        "METAL CLAD"),
            R("2",    "DEGREE OF PROTECTION OF ENCLOSURE",             "",        "IP 41"),
            R("3",    "BASE FRAME",                                    "",        "BASE FRAME PROVIDED. PANEL CAN ALSO BE INSTALLED DIRECTLY TO THE CONCRETE"),
            R("4",    "CABLE ENTRY",                                   "",        "BOTTOM"),
            R("5",    "TYPE OF INCOMING",                              "",        "CABLE"),
            R("",     "INCOMING CABLE SIZE",                           "",        "4R X 1C 500 Sq.mm PER PHASE"),
            R("",     "BOTTOM CLOSING METHOD",                         "",        "AS PER BGS-EE-004 CLAUSE 14.8"),

            # ── F. DIMENSIONS OF SWITCHBOARD ────────────────────────────────
            H("F   DIMENSIONS OF SWITCHBOARD"),
            R("1",    "MAXIMUM OVERALL LENGTH",                        "mm"),
            R("2",    "MAXIMUM OVERALL HEIGHT",                        "mm"),
            R("3",    "MAXIMUM OVERALL DEPTH",                         "mm"),

            # ── G. DIMENSIONS OF SHIPPING SECTION ───────────────────────────
            H("G   DIMENSIONS OF SHIPPING SECTION"),
            R("1",    "MAXIMUM OVERALL LENGTH",                        "mm"),
            R("2",    "MAXIMUM OVERALL HEIGHT",                        "mm"),
            R("3",    "MAXIMUM OVERALL DEPTH",                         "mm"),

            # ── H. 11KV DC CONSUMPTION ───────────────────────────────────────
            H("H   11KV DC CONSUMPTION"),
            R("",     "CONTINUOUS",                                    "WATTS"),
            R("",     "PEAK OR CLOSING COIL / TRIP COIL / SPRING CHARGING", "WATTS"),

            # ── I. WEIGHTS ───────────────────────────────────────────────────
            H("I   WEIGHTS"),
            R("",     "TOTAL",                                         "kg"),
            R("",     "MAXIMUM OF REMOVABLE PARTS",                    "kg"),
            R("",     "MAXIMUM OF PANELS WITH NON-REMOVABLE PARTS",    "kg"),

            # ── J. LOSSES AND TEMPERATURE ────────────────────────────────────
            H("J   LOSSES AND TEMPERATURE"),
            R("1",    "TOTAL HEAT GENERATED BY ASSEMBLY",              "kW",      "8.8"),
            R("2",    "MAXIMUM TEMPERATURE INSIDE SECTION",            "DEG C",   "50"),

            # ── K. CIRCUIT BREAKERS ──────────────────────────────────────────
            H("K   CIRCUIT BREAKERS"),
            R("",     "MANUFACTURER",                                  "",        "SCHNEIDER ELECTRIC"),
            R("",     "TYPE (CB'S)",                                   "",        "VACUUM OR SF6 (WITH SURGE SUPPRESSOR)"),
            R("",     "OPERATING MECHANISM",                           "",        "MOTORIZED"),
            R("3",    "RATED CURRENT",                                 "A",       "*** (REFER NOTE 15)"),
            R("4",    "RATED VOLTAGE",                                 "kV",      "12"),
            R("5",    "NUMBER OF POLES"),
            R("6",    "RATED SYMMETRICAL SHORT CIRCUIT BREAK CAPACITY","kA"),
            R("",     "RATED ASYMMETRICAL SHORT CIRCUIT BREAK CAPACITY","kA",     "25"),
            R("8",    "DC COMPONENT BREAKING CAPACITY",                "kA"),
            R("9",    "RATED SHORT CIRCUIT MAKING CAPACITY",           "kA",      "63.1"),
            R("10",   "SHORT TIME RATING FOR 1 SECOND",                "kA",      "25"),
            R("11",   "POWER FREQ. WITHSTAND VOLTAGE FOR 1 MINUTE",    "kV",      "25"),
            R("12",   "1.2/50 MICRO COND IMPULSE WITHSTAND VOLTAGE",   "kV",      "75"),
            H("13   AUXILIARY CONTROL VOLTAGE"),
            R("14.1", "TRIP & CLOSING COIL",                           "V",       "110V DC"),
            R("14.2", "SPRING CHARGING MOTOR",                         "V",       "110V DC"),
            R("14.3", "MAXIMUM TRIP VOLTAGE",                          "V"),
            R("14.4", "ANTI-CONDENSATION HEATER",                      "",        "240V, 1 PHASE, 50Hz AC (REFER NOTE-7)"),
            R("16",   "CAPACITIVE SWITCH (WITH 100% MAKING CAPACITY)"),
            R("",     "NUMBER OF AUXILIARY CONTACTS (NO & NC)",        "",        "MINIMUM 2 NO & 2 NC CONTACTS REQUIRED"),

            # ── L. CURRENT TRANSFORMER ───────────────────────────────────────
            H("L   CURRENT TRANSFORMER"),
            R("1",    "MANUFACTURER",                                  "",        "TRAFINOD AND MARAVAI"),
            R("2",    "PROTECTION CT ACCURACY",                        "",        "5P20 PX"),
            R("3",    "METERING CT ACCURACY",                          "",        "0.2S"),
            R("4",    "BURDEN",                                        "Ni",      "FE 10 AND 0.5"),
            R("5",    "RATIO",                                         "",        "FE 10 AND 0.5: 1800/1"),

            # ── M. VOLTAGE TRANSFORMER ──────────────────────────────────────
            H("M   VOLTAGE TRANSFORMER"),
            R("1",    "MANUFACTURERS",                                 "",        "TRAFINOD OR EQUIVALENT"),
            R("2",    "TYPE (DRAWABLES OR FIXED)",                     "",        "WITHDRAWABLE"),
            R("3",    "BURDEN/ACCURACY",                               "Ni"),
            R("4",    "RATIO",                                         "Ni",      "1kV/100V(√3)"),
            R("5",    "RATIO"),

            # ── N. PROTECTION & METERING ─────────────────────────────────────
            H("N   PROTECTION & METERING"),
            R("1",    "MANUFACTURER",                                  "",        "SCHNEIDER ELECTRIC"),
            R("2",    "CONSTRUCTION",                                  "",        "MICROPROCESSOR BASED MULTIFUNCTION TYPE WITH INTEGRAL DISPLAY, TBE STAMPING (NOTE: ACCESS TYPE/INDICATION USING SCADA COMMUNICATION FACILITY) (REFER NOTE 3 & 8)"),
            R("3",    "TYPE",                                          "",        "EASERGY RELAY"),
            R("4",    "ANNUNCIATION",                                  "",        "NOTE: 14"),
            H("5   POWER MONITORING UNIT/PMU"),
            R("",     "PM SHALL BE WITH FOLLOWING FUNCTION"),
            R("1.2",  "MONITOR REQUIRED METERING PARAMETERS",          "",        "REQUIRED"),
            R("1.3",  "CAPTURE CURRENT AND VOLTAGE WAVEFORMS",         "",        "REQUIRED"),
            R("1.4",  "TIME-STAMPED AND RECORDING SEQUENCE OF EVENTS, BREAKER TRIPPING, LOOKUP READING WITH 1ms ACCURACY", "", "REQUIRED"),
            R("2",    "TIME STAMP CAPTURE",                            "",        "REQUIRED"),
            R("3",    "MANUFACTURER",                                  "",        "SCHNEIDER ELECTRIC"),
            H("AUTOMATIC TRANSFER (ATS) / HIGH SPEED BUS TRANSFER (HSBT)"),
            R("",     "AUTOMATIC TRANSFER (ATS)",                      "",        "ATS"),
            R("",     "PARKING CYCLE",                                 "",        "AS PER MANUFACTURER STANDARD"),

            # ── PAINTING ─────────────────────────────────────────────────────
            H("PAINTING"),
            R("1",    "COLOR OF EXTERNAL PAINTING",                    "",        "NA"),
            R("2",    "COLOR OF INTERNAL PAINTING",                    "",        "NA"),
            R("3",    "ABR/DENCEBAR INTERNAL PAINTING",                "",        "NA"),
            R("4",    "TROPICALIZATION",                               "",        "YES"),

            # ── TESTS ────────────────────────────────────────────────────────
            H("TESTS"),
            R("1",    "ROUTINE TESTS (TO BE WITNESSED AS PER CL. 71.2 OF BGS-EE-004)", "", "YES"),
            R("2",    "TYPE TESTS AS PER CL. 71.3 OF BGS-EE-004",     "",        "SEPARATE QUOTE REQUIRED"),
            R("3",    "SITE ACCEPTANCE TEST (AS PER CL. 71.3 OF BGS-EE-004)", "", "YES"),
        ]
    
    def export_to_excel(self, datasheet_rows: List[Dict], project_info: Dict = None):
        """Export datasheet to Excel with formatting"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from io import BytesIO
        
        wb = Workbook()
        ws = wb.active
        ws.title = "11KV Switchgear Datasheet"
        
        # Define styles
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        section_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
        section_font = Font(bold=True, size=10)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Add project header
        if project_info:
            ws.merge_cells('A1:E1')
            ws['A1'] = "11KV SWITCHGEAR DATASHEET"
            ws['A1'].font = Font(bold=True, size=14)
            ws['A1'].alignment = Alignment(horizontal='center')
            
            row_idx = 3
            for key, value in project_info.items():
                ws[f'A{row_idx}'] = key.replace('_', ' ').title()
                ws[f'B{row_idx}'] = value
                row_idx += 1
            
            row_idx += 1
        else:
            row_idx = 1
        
        # Add column headers — 6 columns: SR NO, DESCRIPTION, UNIT, REQUIRED DATA, VENDOR DATA, Rev
        headers = ['SR NO', 'DESCRIPTION', 'UNIT', 'REQUIRED DATA', 'VENDOR DATA', 'Rev']
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Set column widths
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 35
        ws.column_dimensions['E'].width = 30
        ws.column_dimensions['F'].width = 8
        
        row_idx += 1
        
        # Add data rows — 6 columns: sr_no, description, unit, required_data, vendor_data, rev
        for row_data in datasheet_rows:
            sr_no = row_data.get('sr_no', '')
            description = row_data.get('description', '')
            unit = row_data.get('unit', '')
            required_data = row_data.get('required_data', '')
            vendor_data = row_data.get('vendor_data', '')
            rev = row_data.get('rev', row_data.get('remarks', ''))  # fallback to remarks for legacy data
            
            # Check if this is a section header
            is_section = (sr_no == '' or sr_no is None) and description and not required_data and not vendor_data
            
            # Add cells
            ws.cell(row=row_idx, column=1, value=sr_no).border = border
            cell_desc = ws.cell(row=row_idx, column=2, value=description)
            cell_desc.border = border
            cell_desc.alignment = Alignment(wrap_text=True, vertical='top')
            
            cell_unit = ws.cell(row=row_idx, column=3, value=unit)
            cell_unit.border = border
            cell_unit.alignment = Alignment(horizontal='center', vertical='top')
            
            cell_req = ws.cell(row=row_idx, column=4, value=required_data)
            cell_req.border = border
            cell_req.alignment = Alignment(wrap_text=True, vertical='top')
            
            cell_vendor = ws.cell(row=row_idx, column=5, value=vendor_data)
            cell_vendor.border = border
            cell_vendor.alignment = Alignment(wrap_text=True, vertical='top')
            
            ws.cell(row=row_idx, column=6, value=rev).border = border
            
            # Apply section styling
            if is_section:
                for col in range(1, 7):
                    cell = ws.cell(row=row_idx, column=col)
                    cell.fill = section_fill
                    cell.font = section_font
            
            row_idx += 1
        
        # Save to BytesIO
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        return excel_buffer
