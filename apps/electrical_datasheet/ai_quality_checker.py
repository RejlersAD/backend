"""
Unified AI-Powered Quality Checker for Electrical Datasheets
=============================================================
Purpose: Single intelligent quality checker that works for ALL equipment types
Approach: Uses GPT-4 to dynamically analyze any datasheet type without hardcoded rules
Features: 
- Smart completeness analysis
- Intelligent consistency validation
- Dynamic standards compliance checking
- Context-aware recommendations
- Universal equipment type support
- Rule-based fallback when AI is unavailable
"""

import json
import re
from datetime import datetime
from typing import Dict, List, Any, Optional
from django.conf import settings
from openai import OpenAI
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework import status
from .adnoc_standards import adnoc_standards
from .equipment_types_config import (
    get_equipment_type_by_id,
    get_critical_fields,
    get_validation_rules,
    validate_field
)
from .ai_config import AI_CONFIG, should_use_fallback
from .rule_based_checker import RuleBasedQualityChecker


class UnifiedAIQualityChecker:
    """
    Unified AI-Powered Quality Checker
    Uses GPT-4 to intelligently analyze datasheets of ANY equipment type
    No hardcoded validation rules - fully dynamic and adaptable
    """
    
    def __init__(self):
        """Initialize OpenAI client with fallback support"""
        api_key = getattr(settings, 'OPENAI_API_KEY', None)
        self.ai_available = False
        self.client = None
        
        if api_key and api_key != '':
            try:
                self.client = OpenAI(api_key=api_key)
                self.model = getattr(settings, 'OPENAI_MODEL', 'gpt-4-turbo-preview')
                self.ai_available = True
            except Exception as e:
                print(f"⚠️ OpenAI initialization failed: {str(e)}")
                self.ai_available = False
        else:
            print("⚠️ OPENAI_API_KEY not configured - using rule-based fallback")
        
        # Initialize rule-based checker for fallback
        self.rule_checker = RuleBasedQualityChecker()
    
    def perform_quality_check(self, datasheet, equipment_type) -> Dict[str, Any]:
        """
        Perform comprehensive AI-powered quality check with step-by-step tracking
        
        Args:
            datasheet: ElectricalDatasheet model instance
            equipment_type: ElectricalEquipmentType model instance
            
        Returns:
            Comprehensive quality report dictionary with detailed progress
        """
        # Initialize report structure with step tracking
        report = {
            'datasheet_id': datasheet.id,
            'tag_number': datasheet.tag_number,
            'equipment_type': equipment_type.name,
            'check_timestamp': datetime.now().isoformat(),
            'overall_score': 0,
            'status': 'in_progress',
            'ai_analysis': {},
            'completeness': {},
            'consistency': {},
            'standards_compliance': {},
            'technical_validation': {},
            'issues': [],
            'warnings': [],
            'recommendations': [],
            'summary': '',
            'progress_steps': [],
            'processing_details': {
                'ai_available': self.ai_available,
                'method_used': 'pending',
                'processing_time': 0,
                'steps_completed': 0,
                'total_steps': 8
            }
        }
        
        start_time = datetime.now()
        
        try:
            # Step 1: Pre-validation checks
            self._add_step(report, "Starting quality check validation", "validating_input")
            self._validate_datasheet_input(datasheet, equipment_type, report)
            
            # Step 2: Data extraction and preparation
            self._add_step(report, "Extracting and preparing datasheet data", "preparing_data")
            form_data = self._prepare_datasheet_data(datasheet, equipment_type, report)
            
            # Step 3: AI Analysis or Fallback determination
            self._add_step(report, "Initializing analysis engine", "initializing_analysis")
            analysis_method = "AI" if self.ai_available and AI_CONFIG['enabled'] else "Rule-based"
            report['processing_details']['method_used'] = analysis_method
            
            # Step 4: Perform comprehensive analysis
            self._add_step(report, f"Running {analysis_method} quality analysis", "running_analysis")
            ai_analysis = self._analyze_with_ai(datasheet, equipment_type)
            
            # Step 5: Parse and validate analysis results
            self._add_step(report, "Processing analysis results", "processing_results")
            
            # Ensure ai_analysis is a dictionary
            if not isinstance(ai_analysis, dict):
                # If AI analysis failed and returned a string, create minimal dict structure
                ai_analysis = {
                    'error': str(ai_analysis),
                    'completeness_analysis': {},
                    'consistency_analysis': {},
                    'standards_compliance': {},
                    'technical_validation': {},
                    'quality_issues': [],
                    'quality_warnings': [],
                    'intelligent_recommendations': [],
                    'overall_assessment': {'quality_grade': 'UNKNOWN', 'readiness': 'needs_review'}
                }
            
            self._populate_report(report, ai_analysis, datasheet, equipment_type)
            
            # Step 6: Calculate comprehensive scores
            self._add_step(report, "Calculating quality scores", "calculating_scores")
            report['overall_score'] = self._calculate_overall_score(report)
            
            # Step 7: Determine final status and recommendations
            self._add_step(report, "Generating recommendations", "generating_recommendations")
            report['status'] = self._determine_status(report['overall_score'])
            
            # Step 8: Finalize report
            self._add_step(report, "Finalizing quality report", "finalizing_report")
            report['summary'] = self._generate_summary(report, ai_analysis)
            
            # Mark as completed
            report['status'] = 'completed' if report['status'] != 'error' else report['status']
            
        except Exception as e:
            self._add_step(report, f"Error encountered: {str(e)}", "error", success=False)
            report['error'] = f"Quality check failed: {str(e)}"
            report['status'] = 'error'
        
        # Calculate processing time
        end_time = datetime.now()
        report['processing_details']['processing_time'] = (end_time - start_time).total_seconds()
        
        return report
    
    def _analyze_with_ai(self, datasheet, equipment_type) -> Dict[str, Any]:
        """
        Use GPT-4 to perform intelligent analysis of the datasheet
        Falls back to rule-based checking if AI is unavailable
        """
        # Check if AI is available
        if not self.ai_available or not AI_CONFIG['enabled']:
            print("ℹ️ AI not available - using rule-based fallback")
            return self._use_fallback(datasheet, equipment_type, "AI not configured")
        
        # Prepare datasheet context
        form_data = datasheet.form_data or {}
        sections = equipment_type.sections or []
        standards = self._get_equipment_standards(equipment_type.id)
        
        # Debug logging
        print(f"DEBUG: sections type: {type(sections)}")
        print(f"DEBUG: sections content: {sections}")
        print(f"DEBUG: form_data type: {type(form_data)}")
        
        # Ensure sections is a list
        if isinstance(sections, str):
            try:
                sections = json.loads(sections)
                print(f"DEBUG: Parsed sections from JSON string")
            except json.JSONDecodeError as e:
                print(f"DEBUG: Failed to parse sections JSON: {e}")
                sections = []
        
        if not isinstance(sections, list):
            print(f"DEBUG: sections is not a list, converting to empty list")
            sections = []
        
        # Build comprehensive prompt
        analysis_prompt = self._build_analysis_prompt(
            equipment_type=equipment_type,
            form_data=form_data,
            sections=sections,
            standards=standards,
            datasheet=datasheet
        )
        
        try:
            # Call GPT-4 for analysis
            response = self.client.chat.completions.create(
                model=self.model,
                messages=[
                    {
                        "role": "system",
                        "content": self._get_system_prompt()
                    },
                    {
                        "role": "user",
                        "content": analysis_prompt
                    }
                ],
                temperature=0.2,  # Lower temperature for more consistent analysis
                max_tokens=4000,
                response_format={"type": "json_object"}
            )
            
            # Parse AI response
            ai_response = response.choices[0].message.content
            print(f"DEBUG: AI response type: {type(ai_response)}")
            print(f"DEBUG: AI response content: {ai_response[:500]}...")
            
            analysis = json.loads(ai_response)
            print(f"DEBUG: After JSON parsing, type: {type(analysis)}")
            
            # Ensure we have a dictionary
            if not isinstance(analysis, dict):
                print(f"WARNING: Expected dict but got {type(analysis)}, converting...")
                analysis = {
                    'error': f'Unexpected response type: {type(analysis)}',
                    'original_response': str(analysis)[:200],
                    'completeness_analysis': {},
                    'consistency_analysis': {},
                    'standards_compliance': {},
                    'technical_validation': {},
                    'quality_issues': [],
                    'quality_warnings': [],
                    'intelligent_recommendations': [],
                    'overall_assessment': {'quality_grade': 'UNKNOWN', 'readiness': 'needs_review'}
                }
            
            analysis['ai_used'] = True
            analysis['method'] = 'gpt-4'
            
            return analysis
            
        except Exception as e:
            error_str = str(e)
            error_code = None
            
            # Extract error code if available
            if 'insufficient_quota' in error_str.lower():
                error_code = 'insufficient_quota'
            elif 'rate_limit' in error_str.lower():
                error_code = 'rate_limit_exceeded'
            elif '429' in error_str:
                error_code = 'insufficient_quota'
            elif 'timeout' in error_str.lower():
                error_code = 'timeout'
            
            # Determine if fallback should be used
            if should_use_fallback(error_code, error_str):
                print(f"⚠️ OpenAI error ({error_code}): {error_str}")
                print("ℹ️ Using rule-based fallback...")
                return self._use_fallback(datasheet, equipment_type, error_str)
            else:
                # Re-raise if fallback not appropriate
                raise Exception(f"AI analysis failed: {error_str}")
    
    def _use_fallback(self, datasheet, equipment_type, reason) -> Dict[str, Any]:
        """
        Use rule-based fallback when AI is unavailable
        """
        try:
            # Prepare datasheet data for rule-based checking
            form_data = datasheet.form_data or {}
            
            # Add metadata
            form_data['_equipment_code'] = equipment_type.code
            form_data['_equipment_type'] = equipment_type.name
            form_data['_tag_number'] = datasheet.tag_number
            
            # Perform rule-based check
            fallback_result = self.rule_checker.check_datasheet(form_data, equipment_type)
            
            # Count actual data fields for better reporting
            non_meta_fields = [k for k in form_data.keys() if not k.startswith('_')]
            filled_fields = [k for k in non_meta_fields if form_data.get(k)]
            empty_fields = [k for k in non_meta_fields if not form_data.get(k)]
            
            # Analyze completeness in detail
            sections = equipment_type.sections or []
            expected_fields = []
            for section in sections:
                if isinstance(section, dict) and 'fields' in section:
                    expected_fields.extend(section.get('fields', []))
            
            # Create structured missing fields information
            structured_missing = []
            if expected_fields:
                for field in expected_fields:
                    if field not in form_data or not form_data.get(field):
                        structured_missing.append({
                            'field': field,
                            'label': field.replace('_', ' ').title(),
                            'section': 'Expected Fields',
                            'importance': 'HIGH',
                            'reason': 'Required for complete datasheet'
                        })
            
            # Create well documented areas list
            well_documented_areas = [f"{len(filled_fields)} data fields present"]
            if filled_fields:
                # Sample some filled fields
                sample_fields = filled_fields[:5]
                well_documented_areas.append(f"Includes: {', '.join(str(f) for f in sample_fields)}")
            
            # Create structured consistency checks from logical issues
            structured_consistency_checks = []
            for idx, warning in enumerate(fallback_result['warnings']):
                check_name = f"Consistency Check {idx + 1}"
                status = "WARNING"
                
                # Determine check type and affected fields
                affected_fields = []
                if 'standards' in warning.lower():
                    check_name = "Standards Documentation"
                    affected_fields = ['standards_compliance']
                elif 'manufacturer' in warning.lower():
                    check_name = "Manufacturer Information"
                    affected_fields = ['manufacturer']
                elif 'model' in warning.lower() or 'type' in warning.lower():
                    check_name = "Model/Type Information"
                    affected_fields = ['model_number', 'equipment_type']
                
                structured_consistency_checks.append({
                    'check_name': check_name,
                    'status': status,
                    'details': warning,
                    'affected_fields': affected_fields,
                    'severity': 'MEDIUM'
                })
            
            # Add positive checks from passed_checks
            for check in fallback_result['passed_checks'][:3]:  # Limit to top 3
                structured_consistency_checks.append({
                    'check_name': 'Data Validation',
                    'status': 'PASS',
                    'details': check,
                    'affected_fields': [],
                    'severity': 'LOW'
                })
            
            # Format result to match AI response structure that _populate_report expects
            analysis = {
                'overall_score': fallback_result['compliance_score'],
                'completeness_analysis': {
                    'overall_completion_percentage': fallback_result['compliance_score'],
                    'critical_missing_fields': structured_missing if structured_missing else fallback_result['issues'],
                    'optional_missing_fields': [],
                    'well_documented_areas': well_documented_areas,
                    'data_richness': fallback_result['compliance_score'],
                    'status': fallback_result['quality_level'],
                    'filled_count': len(filled_fields),
                    'total_count': len(non_meta_fields),
                    'expected_fields_count': len(expected_fields)
                },
                'consistency_analysis': {
                    'consistency_score': fallback_result['compliance_score'],
                    'consistency_checks': structured_consistency_checks,
                    'logical_issues': fallback_result['warnings'],
                    'cross_reference_checks': fallback_result['passed_checks'],
                    'checks_performed': len(structured_consistency_checks),
                    'checks_passed': len([c for c in structured_consistency_checks if c['status'] == 'PASS']),
                    'checks_failed': len([c for c in structured_consistency_checks if c['status'] in ['FAIL', 'WARNING']])
                },
                'standards_compliance': {
                    'compliance_score': fallback_result['compliance_score'],
                    'adnoc_compliance': [],
                    'industry_standards': fallback_result['passed_checks'],
                    'compliant_areas': fallback_result['passed_checks'],
                    'non_compliant_areas': fallback_result['issues'],
                    'recommendations': []
                },
                'technical_validation': {
                    'validation_score': fallback_result['compliance_score'],
                    'technical_checks': [],
                    'design_concerns': []
                },
                'quality_issues': [
                    {
                        'severity': 'MAJOR',
                        'category': 'DATA',
                        'issue': issue,
                        'affected_fields': [],
                        'remediation': 'Address the identified issue'
                    } for issue in fallback_result['issues']
                ],
                'quality_warnings': [
                    {
                        'type': 'INFO',
                        'message': warning,
                        'suggestion': 'Review and verify'
                    } for warning in fallback_result['warnings']
                ],
                'intelligent_recommendations': [
                    {
                        'category': 'SYSTEM',
                        'priority': 'LOW',
                        'recommendation': "✅ Quality check completed using rule-based validation",
                        'benefit': 'Basic validation performed',
                        'effort': 'N/A'
                    },
                    {
                        'category': 'SYSTEM',
                        'priority': 'MEDIUM',
                        'recommendation': f"ℹ️ AI validation unavailable: {reason[:100]}...",
                        'benefit': 'AI would provide deeper analysis',
                        'effort': 'Configure OpenAI API key'
                    },
                    {
                        'category': 'SYSTEM',
                        'priority': 'LOW',
                        'recommendation': "💡 For AI-powered analysis, check OpenAI API quota and billing",
                        'benefit': 'Enhanced intelligent recommendations',
                        'effort': 'LOW'
                    }
                ],
                'overall_assessment': {
                    'quality_grade': fallback_result['quality_level'].upper(),
                    'strengths': fallback_result['passed_checks'][:5],
                    'weaknesses': fallback_result['issues'] + fallback_result['warnings'],
                    'readiness': 'NEEDS_REVIEW' if fallback_result['issues'] else 'READY_FOR_REVIEW',
                    'summary': fallback_result['analysis']
                },
                'ai_used': False,
                'method': 'rule_based_fallback',
                'fallback_reason': reason
            }
            
            return analysis
            
        except Exception as e:
            # Last resort: return minimal analysis
            print(f"❌ Fallback also failed: {str(e)}")
            return {
                'overall_score': 50,
                'completeness': {'score': 50, 'status': 'unknown'},
                'consistency': {'score': 50},
                'standards_compliance': {'score': 50},
                'technical_validation': {'score': 50},
                'issues': [f"Quality check failed: {str(e)}"],
                'warnings': ['Unable to perform complete quality check'],
                'recommendations': ['Please review datasheet manually', 'Check system configuration'],
                'summary': f"Quality check encountered errors. Reason: {str(e)}",
                'ai_used': False,
                'method': 'error_fallback'
            }
    
    def _get_system_prompt(self) -> str:
        """Get the system prompt for AI quality checker"""
        return """You are an expert electrical engineer specializing in quality assurance for electrical equipment datasheets. You have deep knowledge of:
- ADNOC engineering standards and specifications
- IEC, IEEE, and international electrical standards
- Equipment design principles and best practices
- Technical consistency and validation requirements
- Industry safety and compliance regulations

Your task is to analyze electrical equipment datasheets and provide comprehensive quality assessments. You should:
1. Evaluate data completeness across all sections
2. Check technical consistency and logical relationships
3. Validate against ADNOC and industry standards
4. Identify potential issues, gaps, or concerns
5. Provide intelligent recommendations for improvement
6. Calculate quality scores based on multiple factors

Be thorough, precise, and provide actionable insights. Always respond in valid JSON format."""
    
    def _build_analysis_prompt(self, equipment_type, form_data, sections, standards, datasheet) -> str:
        """Build comprehensive analysis prompt for AI"""
        
        # Extract filled vs empty fields
        filled_fields = []
        empty_fields = []
        
        for section in sections:
            section_name = section.get('name', '')
            for field in section.get('fields', []):
                # Handle both string and dictionary field formats
                if isinstance(field, str):
                    # Simple string field format
                    field_id = field
                    field_label = field.replace('_', ' ').title()
                    field_type = 'text'
                    required = False
                else:
                    # Dictionary field format
                    field_id = field.get('id', '')
                    field_label = field.get('label', '')
                    field_type = field.get('type', '')
                    required = field.get('required', False)
                    
                value = form_data.get(field_id)
                
                field_info = {
                    'section': section_name,
                    'label': field_label,
                    'type': field_type,
                    'required': required,
                    'value': value
                }
                
                if value and str(value).strip():
                    filled_fields.append(field_info)
                else:
                    empty_fields.append(field_info)
        
        # Build prompt
        prompt = f"""
ELECTRICAL DATASHEET QUALITY ANALYSIS REQUEST
==============================================

EQUIPMENT INFORMATION:
- Equipment Type: {equipment_type.name}
- Equipment Code: {equipment_type.code}
- Category: {equipment_type.category}
- Tag Number: {datasheet.tag_number}
- Service Description: {datasheet.service_description}
- Location: {datasheet.location}
- Project: {datasheet.project_name} ({datasheet.project_number})

DATASHEET STRUCTURE:
Total Sections: {len(sections)}
Total Fields: {len(filled_fields) + len(empty_fields)}
Filled Fields: {len(filled_fields)}
Empty Fields: {len(empty_fields)}

FILLED DATA (Current Values):
{json.dumps(filled_fields, indent=2)}

EMPTY/MISSING FIELDS:
{json.dumps(empty_fields, indent=2)}

APPLICABLE STANDARDS:
{json.dumps(standards, indent=2)}

EQUIPMENT TYPE CONFIGURATION:
{json.dumps({
    'description': equipment_type.description,
    'standards_reference': equipment_type.standards,
    'sections': [{'name': s.get('name'), 'field_count': len(s.get('fields', []))} for s in sections]
}, indent=2)}

ANALYSIS REQUIREMENTS:
Please perform a comprehensive quality analysis and provide your response in the following JSON structure:

{{
    "completeness_analysis": {{
        "overall_completion_percentage": <0-100>,
        "critical_missing_fields": [
            {{
                "field": "<field label>",
                "section": "<section name>",
                "importance": "<HIGH/MEDIUM/LOW>",
                "reason": "<why this field is important>"
            }}
        ],
        "optional_missing_fields": ["<field labels>"],
        "well_documented_areas": ["<section names>"]
    }},
    
    "consistency_analysis": {{
        "consistency_score": <0-100>,
        "consistency_checks": [
            {{
                "check_name": "<descriptive name>",
                "status": "<PASS/FAIL/WARNING>",
                "details": "<explanation>",
                "affected_fields": ["<field names>"]
            }}
        ],
        "logical_issues": [
            {{
                "issue": "<description>",
                "severity": "<HIGH/MEDIUM/LOW>",
                "recommendation": "<how to fix>"
            }}
        ]
    }},
    
    "standards_compliance": {{
        "compliance_score": <0-100>,
        "adnoc_compliance": [
            {{
                "standard_item": "<standard requirement>",
                "status": "<COMPLIANT/NON_COMPLIANT/PARTIAL/NOT_APPLICABLE>",
                "details": "<explanation>",
                "recommendation": "<if non-compliant>"
            }}
        ],
        "industry_standards": [
            {{
                "standard": "<IEC/IEEE/etc>",
                "requirement": "<description>",
                "compliance_status": "<status>",
                "notes": "<additional info>"
            }}
        ]
    }},
    
    "technical_validation": {{
        "validation_score": <0-100>,
        "technical_checks": [
            {{
                "parameter": "<technical parameter>",
                "expected_range": "<typical range or value>",
                "actual_value": "<value from datasheet>",
                "status": "<VALID/INVALID/QUESTIONABLE>",
                "comment": "<technical assessment>"
            }}
        ],
        "design_concerns": [
            {{
                "concern": "<description>",
                "severity": "<HIGH/MEDIUM/LOW>",
                "impact": "<potential impact>",
                "recommendation": "<suggested action>"
            }}
        ]
    }},
    
    "intelligent_recommendations": [
        {{
            "category": "<COMPLETENESS/CONSISTENCY/STANDARDS/TECHNICAL/DOCUMENTATION>",
            "priority": "<HIGH/MEDIUM/LOW>",
            "recommendation": "<detailed recommendation>",
            "benefit": "<expected improvement>",
            "effort": "<LOW/MEDIUM/HIGH>"
        }}
    ],
    
    "quality_issues": [
        {{
            "severity": "<CRITICAL/MAJOR/MINOR>",
            "category": "<category>",
            "issue": "<description>",
            "affected_fields": ["<fields>"],
            "remediation": "<how to fix>"
        }}
    ],
    
    "quality_warnings": [
        {{
            "type": "<warning type>",
            "message": "<warning message>",
            "suggestion": "<what to check>"
        }}
    ],
    
    "overall_assessment": {{
        "quality_grade": "<EXCELLENT/GOOD/FAIR/POOR>",
        "strengths": ["<key strengths>"],
        "weaknesses": ["<key weaknesses>"],
        "readiness": "<READY_FOR_REVIEW/NEEDS_IMPROVEMENT/REQUIRES_MAJOR_REVISION>",
        "summary": "<2-3 sentence executive summary>"
    }}
}}

Provide thorough, actionable analysis based on engineering best practices and the specific requirements of {equipment_type.name}.
"""
        
        return prompt
    
    def _get_equipment_standards(self, equipment_type_id: str) -> Dict[str, Any]:
        """
        Get applicable standards for equipment type from soft-coded configuration
        Falls back to adnoc_standards if not found in config
        """
        try:
            # First try to get from equipment types config (soft-coded)
            equipment_config = get_equipment_type_by_id(equipment_type_id)
            if equipment_config and 'standards' in equipment_config:
                return {
                    'standards_list': equipment_config['standards'],
                    'critical_fields': equipment_config.get('critical_fields', []),
                    'validation_rules': equipment_config.get('validation_rules', {}),
                    'sections': equipment_config.get('sections', [])
                }
            
            # Fallback to legacy adnoc_standards
            return adnoc_standards.get(equipment_type_id, {})
        except Exception:
            return {}
    
    def _populate_report(self, report: Dict, ai_analysis: Dict, datasheet, equipment_type):
        """Populate report with AI analysis results"""
        
        # Store complete AI analysis
        report['ai_analysis'] = ai_analysis
        
        # Extract completeness data
        completeness = ai_analysis.get('completeness_analysis', {})
        report['completeness'] = {
            'completion_percentage': completeness.get('overall_completion_percentage', 0),
            'critical_missing': len(completeness.get('critical_missing_fields', [])),
            'critical_missing_fields': completeness.get('critical_missing_fields', []),
            'optional_missing': len(completeness.get('optional_missing_fields', [])),
            'well_documented': completeness.get('well_documented_areas', [])
        }
        
        # Extract consistency data
        consistency = ai_analysis.get('consistency_analysis', {})
        report['consistency'] = {
            'consistency_score': consistency.get('consistency_score', 0),
            'checks_performed': len(consistency.get('consistency_checks', [])),
            'checks_passed': len([c for c in consistency.get('consistency_checks', []) if c.get('status') == 'PASS']),
            'checks_failed': len([c for c in consistency.get('consistency_checks', []) if c.get('status') == 'FAIL']),
            'logical_issues': consistency.get('logical_issues', []),
            'details': consistency.get('consistency_checks', [])
        }
        
        # Extract standards compliance
        standards = ai_analysis.get('standards_compliance', {})
        report['standards_compliance'] = {
            'compliance_score': standards.get('compliance_score', 0),
            'adnoc_compliance': standards.get('adnoc_compliance', []),
            'industry_standards': standards.get('industry_standards', []),
            'non_compliant_items': len([s for s in standards.get('adnoc_compliance', []) if s.get('status') == 'NON_COMPLIANT'])
        }
        
        # Extract technical validation
        technical = ai_analysis.get('technical_validation', {})
        report['technical_validation'] = {
            'validation_score': technical.get('validation_score', 0),
            'checks': technical.get('technical_checks', []),
            'concerns': technical.get('design_concerns', []),
            'invalid_parameters': len([t for t in technical.get('technical_checks', []) if t.get('status') == 'INVALID'])
        }
        
        # Extract issues
        issues = ai_analysis.get('quality_issues', [])
        report['issues'] = [
            {
                'severity': issue.get('severity', 'MINOR'),
                'category': issue.get('category', 'GENERAL'),
                'message': issue.get('issue', ''),
                'affected_fields': issue.get('affected_fields', []),
                'remediation': issue.get('remediation', '')
            }
            for issue in issues
        ]
        
        # Extract warnings
        warnings = ai_analysis.get('quality_warnings', [])
        report['warnings'] = [
            {
                'type': warning.get('type', 'GENERAL'),
                'message': warning.get('message', ''),
                'suggestion': warning.get('suggestion', '')
            }
            for warning in warnings
        ]
        
        # Extract recommendations
        recommendations = ai_analysis.get('intelligent_recommendations', [])
        report['recommendations'] = [
            {
                'category': rec.get('category', 'GENERAL'),
                'priority': rec.get('priority', 'MEDIUM'),
                'recommendation': rec.get('recommendation', ''),
                'benefit': rec.get('benefit', ''),
                'effort': rec.get('effort', 'MEDIUM')
            }
            for rec in recommendations
        ]
    
    def _calculate_overall_score(self, report: Dict) -> float:
        """Calculate overall quality score based on all factors"""
        weights = {
            'completeness': 0.30,
            'consistency': 0.25,
            'standards_compliance': 0.25,
            'technical_validation': 0.20
        }
        
        scores = {
            'completeness': report['completeness'].get('completion_percentage', 0),
            'consistency': report['consistency'].get('consistency_score', 0),
            'standards_compliance': report['standards_compliance'].get('compliance_score', 0),
            'technical_validation': report['technical_validation'].get('validation_score', 0)
        }
        
        overall = sum(scores[key] * weights[key] for key in weights.keys())
        
        # Apply penalties for critical issues
        critical_issues = len([i for i in report['issues'] if i.get('severity') == 'CRITICAL'])
        overall = max(0, overall - (critical_issues * 5))
        
        return round(overall, 2)
    
    def _determine_status(self, overall_score: float) -> str:
        """Determine quality status based on overall score"""
        if overall_score >= 90:
            return 'excellent'
        elif overall_score >= 75:
            return 'good'
        elif overall_score >= 60:
            return 'acceptable'
        elif overall_score >= 40:
            return 'needs_improvement'
        else:
            return 'poor'
    
    def _generate_summary(self, report: Dict, ai_analysis: Dict) -> str:
        """Generate executive summary"""
        overall_assessment = ai_analysis.get('overall_assessment', {})
        
        score = report['overall_score']
        status = report['status']
        quality_grade = overall_assessment.get('quality_grade', 'N/A')
        readiness = overall_assessment.get('readiness', 'N/A')
        
        summary_parts = []
        
        # Overall status
        summary_parts.append(f"Overall Quality Score: {score}/100 ({status.upper()})")
        summary_parts.append(f"Quality Grade: {quality_grade}")
        summary_parts.append(f"Readiness: {readiness.replace('_', ' ').title()}")
        
        # Key metrics
        completion = report['completeness'].get('completion_percentage', 0)
        summary_parts.append(f"\nKey Metrics:")
        summary_parts.append(f"- Data Completeness: {completion}%")
        summary_parts.append(f"- Consistency Score: {report['consistency'].get('consistency_score', 0)}%")
        summary_parts.append(f"- Standards Compliance: {report['standards_compliance'].get('compliance_score', 0)}%")
        summary_parts.append(f"- Technical Validation: {report['technical_validation'].get('validation_score', 0)}%")
        
        # Critical items
        critical_missing = report['completeness'].get('critical_missing', 0)
        critical_issues = len([i for i in report['issues'] if i.get('severity') == 'CRITICAL'])
        
        if critical_missing > 0 or critical_issues > 0:
            summary_parts.append(f"\nAttention Required:")
            if critical_missing > 0:
                summary_parts.append(f"- {critical_missing} critical field(s) missing")
            if critical_issues > 0:
                summary_parts.append(f"- {critical_issues} critical issue(s) identified")
        
        # AI assessment summary
        ai_summary = overall_assessment.get('summary', '')
        if ai_summary:
            summary_parts.append(f"\nAI Assessment:")
            summary_parts.append(ai_summary)
        
        # Strengths
        strengths = overall_assessment.get('strengths', [])
        if strengths:
            summary_parts.append(f"\nStrengths:")
            for strength in strengths[:3]:
                summary_parts.append(f"✓ {strength}")
        
        # Weaknesses
        weaknesses = overall_assessment.get('weaknesses', [])
        if weaknesses:
            summary_parts.append(f"\nAreas for Improvement:")
            for weakness in weaknesses[:3]:
                summary_parts.append(f"⚠ {weakness}")
        
        return "\n".join(summary_parts)


    # Helper methods for enhanced step-by-step processing
    def _add_step(self, report, description, step_type, success=True):
        """Add a processing step to the report for detailed tracking"""
        step = {
            'step_number': len(report['progress_steps']) + 1,
            'description': description,
            'type': step_type,
            'timestamp': datetime.now().isoformat(),
            'success': success
        }
        report['progress_steps'].append(step)
        report['processing_details']['steps_completed'] = len(report['progress_steps'])
    
    def _validate_datasheet_input(self, datasheet, equipment_type, report):
        """Validate that the datasheet has the minimum required data for quality checking"""
        issues = []
        
        # Check if datasheet has form data
        if not datasheet.form_data:
            issues.append("No form data found in datasheet")
        
        # Check if equipment type is valid
        if not equipment_type:
            issues.append("Invalid equipment type")
        
        # Check if tag number exists
        if not datasheet.tag_number:
            issues.append("Missing tag number")
        
        # Add validation results to report
        report['validation'] = {
            'pre_check_passed': len(issues) == 0,
            'validation_issues': issues
        }
        
        if issues:
            raise Exception(f"Validation failed: {'; '.join(issues)}")
    
    def _prepare_datasheet_data(self, datasheet, equipment_type, report):
        """Extract and prepare datasheet data for analysis with enhanced error handling"""
        try:
            form_data = datasheet.form_data or {}
            
            # Add metadata for better analysis
            form_data['_equipment_code'] = equipment_type.code
            form_data['_equipment_type'] = equipment_type.name
            form_data['_tag_number'] = datasheet.tag_number
            form_data['_sections'] = equipment_type.sections or []
            
            # Count available fields
            field_count = len([k for k, v in form_data.items() if not k.startswith('_') and v])
            total_fields = len(equipment_type.sections or []) * 10  # Rough estimate
            
            report['data_preparation'] = {
                'fields_available': field_count,
                'estimated_total_fields': total_fields,
                'data_richness': min(100, (field_count / max(1, total_fields)) * 100)
            }
            
            return form_data
            
        except Exception as e:
            raise Exception(f"Data preparation failed: {str(e)}")


class AIQualityCheckerMixin:
    """
    Mixin to add unified AI-powered quality checking to ViewSets
    Replaces the old hardcoded equipment-specific quality checker
    """
    
    @action(detail=True, methods=['post'], url_path='ai-quality-check')
    def ai_quality_check(self, request, pk=None):
        """
        Perform AI-powered quality check on a datasheet with detailed progress tracking
        Works for ALL equipment types using intelligent analysis
        """
        try:
            datasheet = self.get_object()
            equipment_type = datasheet.equipment_type
            
            # Validate inputs
            if not datasheet:
                return Response(
                    {'error': 'Datasheet not found'},
                    status=status.HTTP_404_NOT_FOUND
                )
                
            if not equipment_type:
                return Response(
                    {'error': 'Equipment type not found for this datasheet'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Initialize AI quality checker
            checker = UnifiedAIQualityChecker()
            
            # Log the quality check attempt
            print(f"🔍 Starting quality check for datasheet {datasheet.id} ({datasheet.tag_number})")
            print(f"📊 Equipment type: {equipment_type.name} ({equipment_type.code})")
            print(f"🤖 AI available: {checker.ai_available}")
            
            # Perform quality check with enhanced error handling
            report = checker.perform_quality_check(datasheet, equipment_type)
            
            # Update datasheet with results
            if 'overall_score' in report and report['overall_score'] is not None:
                datasheet.compliance_score = report['overall_score']
                from django.utils import timezone
                datasheet.last_quality_check = timezone.now()
                datasheet.save(update_fields=['compliance_score', 'last_quality_check'])
                print(f"✅ Updated datasheet with score: {report['overall_score']}")
            
            # Log completion
            print(f"✨ Quality check completed with status: {report.get('status', 'unknown')}")
            
            return Response(report, status=status.HTTP_200_OK)
            
        except Exception as e:
            error_message = str(e)
            print(f"❌ Quality check failed: {error_message}")
            
            return Response(
                {
                    'error': 'Quality check failed',
                    'details': error_message,
                    'datasheet_id': pk,
                    'timestamp': datetime.now().isoformat(),
                    'troubleshooting': {
                        'common_solutions': [
                            'Ensure the datasheet has form data',
                            'Verify the equipment type is properly configured',
                            'Check if OpenAI API key is valid (if using AI)',
                            'Try again in a few moments'
                        ]
                    }
                },
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=False, methods=['post'], url_path='batch-ai-quality-check')
    def batch_ai_quality_check(self, request):
        """
        Perform AI quality check on multiple datasheets
        """
        datasheet_ids = request.data.get('datasheet_ids', [])
        
        if not datasheet_ids:
            return Response(
                {'error': 'No datasheet IDs provided'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            queryset = self.get_queryset().filter(id__in=datasheet_ids)
            checker = UnifiedAIQualityChecker()
            results = []
            
            for datasheet in queryset:
                try:
                    report = checker.perform_quality_check(datasheet, datasheet.equipment_type)
                    results.append(report)
                    
                    # Update compliance score
                    if 'overall_score' in report:
                        datasheet.compliance_score = report['overall_score']
                        datasheet.save(update_fields=['compliance_score'])
                        
                except Exception as e:
                    results.append({
                        'datasheet_id': datasheet.id,
                        'error': str(e),
                        'status': 'error'
                    })
            
            # Calculate batch summary
            batch_summary = {
                'total_checked': len(results),
                'successful': len([r for r in results if r.get('status') != 'error']),
                'failed': len([r for r in results if r.get('status') == 'error']),
                'average_score': sum(r.get('overall_score', 0) for r in results if 'overall_score' in r) / len(results) if results else 0,
                'results': results
            }
            
            return Response(batch_summary, status=status.HTTP_200_OK)
            
        except Exception as e:
            return Response(
                {
                    'error': 'Batch quality check failed',
                    'details': str(e)
                },
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=True, methods=['get'], url_path='download-quality-report-excel')
    def download_quality_report_excel(self, request, pk=None):
        """
        Download quality check report as Excel file
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            from django.http import HttpResponse
            
            datasheet = self.get_object()
            equipment_type = datasheet.equipment_type
            
            # Perform quality check to get the latest report
            checker = UnifiedAIQualityChecker()
            report = checker.perform_quality_check(datasheet, equipment_type)
            
            # Create workbook
            wb = Workbook()
            
            # Remove default sheet
            wb.remove(wb.active)
            
            # Define styles
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF', size=12)
            subheader_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
            subheader_font = Font(bold=True, size=11)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Sheet 1: Summary
            ws_summary = wb.create_sheet('Summary')
            ws_summary.column_dimensions['A'].width = 30
            ws_summary.column_dimensions['B'].width = 50
            
            row = 1
            ws_summary[f'A{row}'] = 'Quality Check Report'
            ws_summary[f'A{row}'].font = Font(bold=True, size=16)
            row += 2
            
            # Datasheet info
            ws_summary[f'A{row}'] = 'Tag Number:'
            ws_summary[f'B{row}'] = datasheet.tag_number or 'N/A'
            ws_summary[f'A{row}'].font = Font(bold=True)
            row += 1
            
            ws_summary[f'A{row}'] = 'Equipment Type:'
            ws_summary[f'B{row}'] = equipment_type.name if equipment_type else 'N/A'
            ws_summary[f'A{row}'].font = Font(bold=True)
            row += 1
            
            ws_summary[f'A{row}'] = 'Service Description:'
            ws_summary[f'B{row}'] = datasheet.service_description or 'N/A'
            ws_summary[f'A{row}'].font = Font(bold=True)
            row += 2
            
            # Overall scores
            ws_summary[f'A{row}'] = 'Overall Quality Score'
            ws_summary[f'A{row}'].font = subheader_font
            ws_summary[f'A{row}'].fill = subheader_fill
            row += 1
            
            ws_summary[f'A{row}'] = 'Score:'
            ws_summary[f'B{row}'] = f"{report.get('overall_score', 0)}/100"
            score = report.get('overall_score', 0)
            if score >= 80:
                ws_summary[f'B{row}'].fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
            elif score >= 60:
                ws_summary[f'B{row}'].fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
            else:
                ws_summary[f'B{row}'].fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
            row += 2
            
            # Component scores
            ws_summary[f'A{row}'] = 'Component Scores'
            ws_summary[f'A{row}'].font = subheader_font
            ws_summary[f'A{row}'].fill = subheader_fill
            row += 1
            
            completeness = report.get('completeness', {})
            ws_summary[f'A{row}'] = 'Completeness:'
            ws_summary[f'B{row}'] = f"{completeness.get('completion_percentage', 0)}%"
            row += 1
            
            consistency = report.get('consistency', {})
            ws_summary[f'A{row}'] = 'Consistency:'
            ws_summary[f'B{row}'] = f"{consistency.get('consistency_score', 0)}%"
            row += 1
            
            standards = report.get('standards', {})
            ws_summary[f'A{row}'] = 'Standards Compliance:'
            ws_summary[f'B{row}'] = f"{standards.get('compliance_score', 0)}%"
            row += 1
            
            technical = report.get('technical_validation', {})
            ws_summary[f'A{row}'] = 'Technical Validation:'
            ws_summary[f'B{row}'] = f"{technical.get('validation_score', 0)}%"
            row += 1
            
            # Sheet 2: Completeness
            ws_completeness = wb.create_sheet('Completeness')
            ws_completeness.column_dimensions['A'].width = 30
            ws_completeness.column_dimensions['B'].width = 30
            ws_completeness.column_dimensions['C'].width = 40
            
            row = 1
            ws_completeness[f'A{row}'] = 'Field'
            ws_completeness[f'B{row}'] = 'Section'
            ws_completeness[f'C{row}'] = 'Reason'
            for col in ['A', 'B', 'C']:
                ws_completeness[f'{col}{row}'].font = header_font
                ws_completeness[f'{col}{row}'].fill = header_fill
                ws_completeness[f'{col}{row}'].border = border
            row += 1
            
            missing_fields = completeness.get('missing_fields', [])
            for field in missing_fields:
                if isinstance(field, dict):
                    ws_completeness[f'A{row}'] = field.get('label', field.get('field', ''))
                    ws_completeness[f'B{row}'] = field.get('section', 'N/A')
                    ws_completeness[f'C{row}'] = field.get('reason', 'Missing')
                else:
                    ws_completeness[f'A{row}'] = str(field)
                    ws_completeness[f'B{row}'] = 'N/A'
                    ws_completeness[f'C{row}'] = 'Missing'
                
                for col in ['A', 'B', 'C']:
                    ws_completeness[f'{col}{row}'].border = border
                row += 1
            
            # Sheet 3: Consistency
            ws_consistency = wb.create_sheet('Consistency')
            ws_consistency.column_dimensions['A'].width = 35
            ws_consistency.column_dimensions['B'].width = 15
            ws_consistency.column_dimensions['C'].width = 50
            ws_consistency.column_dimensions['D'].width = 15
            
            row = 1
            ws_consistency[f'A{row}'] = 'Check Name'
            ws_consistency[f'B{row}'] = 'Status'
            ws_consistency[f'C{row}'] = 'Details'
            ws_consistency[f'D{row}'] = 'Severity'
            for col in ['A', 'B', 'C', 'D']:
                ws_consistency[f'{col}{row}'].font = header_font
                ws_consistency[f'{col}{row}'].fill = header_fill
                ws_consistency[f'{col}{row}'].border = border
            row += 1
            
            consistency_details = consistency.get('details', [])
            for check in consistency_details:
                ws_consistency[f'A{row}'] = check.get('check_name', 'Check')
                ws_consistency[f'B{row}'] = check.get('status', 'UNKNOWN')
                ws_consistency[f'C{row}'] = check.get('details', '')
                ws_consistency[f'D{row}'] = check.get('severity', 'MEDIUM')
                
                # Color code based on status
                status_val = check.get('status', 'UNKNOWN')
                if status_val == 'PASS':
                    ws_consistency[f'B{row}'].fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                elif status_val == 'WARNING':
                    ws_consistency[f'B{row}'].fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
                else:
                    ws_consistency[f'B{row}'].fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                
                for col in ['A', 'B', 'C', 'D']:
                    ws_consistency[f'{col}{row}'].border = border
                row += 1
            
            # Also add logical issues if present
            logical_issues = consistency.get('logical_issues', [])
            if logical_issues:
                row += 1
                ws_consistency[f'A{row}'] = 'Additional Issues'
                ws_consistency[f'A{row}'].font = subheader_font
                ws_consistency[f'A{row}'].fill = subheader_fill
                row += 1
                
                for issue in logical_issues:
                    ws_consistency[f'A{row}'] = 'Logical Issue'
                    ws_consistency[f'B{row}'] = 'WARNING'
                    ws_consistency[f'C{row}'] = issue
                    ws_consistency[f'D{row}'] = 'MEDIUM'
                    for col in ['A', 'B', 'C', 'D']:
                        ws_consistency[f'{col}{row}'].border = border
                    row += 1
            
            # Sheet 4: Standards
            ws_standards = wb.create_sheet('Standards')
            ws_standards.column_dimensions['A'].width = 35
            ws_standards.column_dimensions['B'].width = 60
            
            row = 1
            ws_standards[f'A{row}'] = 'Category'
            ws_standards[f'B{row}'] = 'Details'
            for col in ['A', 'B']:
                ws_standards[f'{col}{row}'].font = header_font
                ws_standards[f'{col}{row}'].fill = header_fill
                ws_standards[f'{col}{row}'].border = border
            row += 1
            
            # Standards compliance details
            standards_issues = standards.get('issues', [])
            for issue in standards_issues:
                if isinstance(issue, dict):
                    ws_standards[f'A{row}'] = issue.get('category', 'Issue')
                    ws_standards[f'B{row}'] = issue.get('details', '')
                else:
                    ws_standards[f'A{row}'] = 'Issue'
                    ws_standards[f'B{row}'] = str(issue)
                
                for col in ['A', 'B']:
                    ws_standards[f'{col}{row}'].border = border
                row += 1
            
            # Sheet 5: Technical Validation
            ws_technical = wb.create_sheet('Technical Validation')
            ws_technical.column_dimensions['A'].width = 35
            ws_technical.column_dimensions['B'].width = 60
            
            row = 1
            ws_technical[f'A{row}'] = 'Parameter'
            ws_technical[f'B{row}'] = 'Finding'
            for col in ['A', 'B']:
                ws_technical[f'{col}{row}'].font = header_font
                ws_technical[f'{col}{row}'].fill = header_fill
                ws_technical[f'{col}{row}'].border = border
            row += 1
            
            # Technical validation findings
            technical_findings = technical.get('findings', [])
            for finding in technical_findings:
                if isinstance(finding, dict):
                    ws_technical[f'A{row}'] = finding.get('parameter', 'Parameter')
                    ws_technical[f'B{row}'] = finding.get('finding', '')
                else:
                    ws_technical[f'A{row}'] = 'Finding'
                    ws_technical[f'B{row}'] = str(finding)
                
                for col in ['A', 'B']:
                    ws_technical[f'{col}{row}'].border = border
                row += 1
            
            # Prepare response
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            filename = f'quality_report_{datasheet.tag_number or datasheet.id}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            
            wb.save(response)
            return response
            
        except Exception as e:
            return Response(
                {'error': 'Failed to generate Excel report', 'details': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=True, methods=['get'], url_path='download-quality-report-pdf')
    def download_quality_report_pdf(self, request, pk=None):
        """
        Download quality check report as PDF file
        """
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import letter, A4
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
            from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
            from django.http import HttpResponse
            from io import BytesIO
            
            datasheet = self.get_object()
            equipment_type = datasheet.equipment_type
            
            # Perform quality check to get the latest report
            checker = UnifiedAIQualityChecker()
            report = checker.perform_quality_check(datasheet, equipment_type)
            
            # Create PDF buffer
            buffer = BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=0.5*inch, bottomMargin=0.5*inch)
            
            # Container for PDF elements
            elements = []
            
            # Styles
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=24,
                textColor=colors.HexColor('#4472C4'),
                spaceAfter=30,
                alignment=TA_CENTER
            )
            heading_style = ParagraphStyle(
                'CustomHeading',
                parent=styles['Heading2'],
                fontSize=16,
                textColor=colors.HexColor('#4472C4'),
                spaceAfter=12,
                spaceBefore=12
            )
            subheading_style = ParagraphStyle(
                'CustomSubHeading',
                parent=styles['Heading3'],
                fontSize=14,
                textColor=colors.HexColor('#5B9BD5'),
                spaceAfter=8
            )
            normal_style = styles['Normal']
            
            # Title
            elements.append(Paragraph('Quality Check Report', title_style))
            elements.append(Spacer(1, 0.2*inch))
            
            # Datasheet Information
            elements.append(Paragraph('Datasheet Information', heading_style))
            
            info_data = [
                ['Tag Number:', datasheet.tag_number or 'N/A'],
                ['Equipment Type:', equipment_type.name if equipment_type else 'N/A'],
                ['Service Description:', datasheet.service_description or 'N/A'],
                ['Report Date:', datetime.now().strftime('%Y-%m-%d %H:%M:%S')]
            ]
            info_table = Table(info_data, colWidths=[2*inch, 4*inch])
            info_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#D9E1F2')),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
                ('GRID', (0, 0), (-1, -1), 1, colors.grey)
            ]))
            elements.append(info_table)
            elements.append(Spacer(1, 0.3*inch))
            
            # Overall Score
            elements.append(Paragraph('Overall Quality Score', heading_style))
            
            overall_score = report.get('overall_score', 0)
            score_color = colors.HexColor('#C6EFCE') if overall_score >= 80 else (
                colors.HexColor('#FFEB9C') if overall_score >= 60 else colors.HexColor('#FFC7CE')
            )
            
            score_data = [
                ['Overall Score', f"{overall_score}/100"]
            ]
            score_table = Table(score_data, colWidths=[3*inch, 3*inch])
            score_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, -1), score_color),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 14),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
                ('GRID', (0, 0), (-1, -1), 2, colors.grey)
            ]))
            elements.append(score_table)
            elements.append(Spacer(1, 0.3*inch))
            
            # Component Scores
            elements.append(Paragraph('Component Scores', heading_style))
            
            completeness = report.get('completeness', {})
            consistency = report.get('consistency', {})
            standards = report.get('standards', {})
            technical = report.get('technical_validation', {})
            
            component_data = [
                ['Component', 'Score'],
                ['Completeness', f"{completeness.get('completion_percentage', 0)}%"],
                ['Consistency', f"{consistency.get('consistency_score', 0)}%"],
                ['Standards Compliance', f"{standards.get('compliance_score', 0)}%"],
                ['Technical Validation', f"{technical.get('validation_score', 0)}%"]
            ]
            component_table = Table(component_data, colWidths=[3*inch, 3*inch])
            component_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 11),
                ('FONTSIZE', (0, 1), (-1, -1), 10),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 1, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F2F2F2')])
            ]))
            elements.append(component_table)
            elements.append(PageBreak())
            
            # Completeness Details
            elements.append(Paragraph('Completeness Analysis', heading_style))
            missing_fields = completeness.get('missing_fields', [])
            
            if missing_fields:
                completeness_data = [['Field', 'Section', 'Reason']]
                for field in missing_fields[:20]:  # Limit to 20 for PDF
                    if isinstance(field, dict):
                        completeness_data.append([
                            field.get('label', field.get('field', '')),
                            field.get('section', 'N/A'),
                            field.get('reason', 'Missing')[:50]  # Truncate long reasons
                        ])
                    else:
                        completeness_data.append([str(field), 'N/A', 'Missing'])
                
                completeness_table = Table(completeness_data, colWidths=[2*inch, 2*inch, 2.5*inch])
                completeness_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 9),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                    ('GRID', (0, 0), (-1, -1), 1, colors.grey),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F2F2F2')])
                ]))
                elements.append(completeness_table)
            else:
                elements.append(Paragraph('All required fields are completed.', normal_style))
            
            elements.append(Spacer(1, 0.2*inch))
            
            # Consistency Checks
            elements.append(Paragraph('Consistency Checks', heading_style))
            consistency_details = consistency.get('details', [])
            
            if consistency_details:
                consistency_data = [['Check Name', 'Status', 'Details']]
                for check in consistency_details[:15]:  # Limit to 15 for PDF
                    status_val = check.get('status', 'UNKNOWN')
                    consistency_data.append([
                        check.get('check_name', 'Check')[:30],
                        status_val,
                        check.get('details', '')[:60]
                    ])
                
                consistency_table = Table(consistency_data, colWidths=[2*inch, 1.5*inch, 3*inch])
                consistency_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 9),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                    ('GRID', (0, 0), (-1, -1), 1, colors.grey),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F2F2F2')])
                ]))
                elements.append(consistency_table)
            
            # Add logical issues
            logical_issues = consistency.get('logical_issues', [])
            if logical_issues:
                elements.append(Spacer(1, 0.1*inch))
                elements.append(Paragraph('Additional Consistency Issues:', subheading_style))
                for issue in logical_issues[:10]:
                    elements.append(Paragraph(f'• {issue}', normal_style))
            
            if not consistency_details and not logical_issues:
                elements.append(Paragraph('No consistency issues found.', normal_style))
            
            # Build PDF
            doc.build(elements)
            
            # Prepare response
            buffer.seek(0)
            response = HttpResponse(buffer.read(), content_type='application/pdf')
            filename = f'quality_report_{datasheet.tag_number or datasheet.id}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            
            return response
            
        except Exception as e:
            return Response(
                {'error': 'Failed to generate PDF report', 'details': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
