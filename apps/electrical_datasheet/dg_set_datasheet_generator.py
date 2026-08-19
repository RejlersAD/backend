"""
Emergency Diesel Generator (EDG) Set Datasheet Generator
Extracts equipment data from EDG Sizing Calculation documents
and generates comprehensive datasheets matching the standard ADNOC form.

Columns: SI No. | DESCRIPTION | UNIT | SPECIFIED DESIGN DATA | VENDOR DATA | Rev
"""
import logging
import json
from typing import Dict, List, Optional
from django.conf import settings
from openai import OpenAI
import PyPDF2

logger = logging.getLogger(__name__)


class DGSetDatasheetGenerator:
    """Generate Emergency Diesel Generator (EDG) Set datasheets from sizing calculation documents."""

    def __init__(self):
        self.client = OpenAI(api_key=settings.OPENAI_API_KEY)

    # ──────────────────────────────────────────────────────────────────────────
    # Document Extraction (multi-format)
    # ──────────────────────────────────────────────────────────────────────────
    def extract_text_from_pdf(self, pdf_file) -> str:
        """Extract text from any supported document type."""
        from .document_extractor import extract_text
        text = extract_text(pdf_file)
        logger.info(f"[DGSetDatasheet] Extracted {len(text)} chars from {getattr(pdf_file, 'name', '?')}")
        return text

    # ──────────────────────────────────────────────────────────────────────────
    # Main entry point
    # ──────────────────────────────────────────────────────────────────────────
    def generate_datasheet_from_sizing_calc(self, pdf_file, project_info: Dict = None) -> Dict:
        """
        Generate a DG set datasheet from a sizing calculation PDF.

        Returns:
            {
                'success': bool,
                'datasheet_rows': List[Dict],   # sr_no, description, unit, required_data, vendor_data, rev
                'summary': Dict,
                'extraction_metadata': Dict
            }
        """
        try:
            logger.info("[DGSetDatasheet] Extracting text from EDG sizing calculation PDF…")
            doc_text = self.extract_text_from_pdf(pdf_file)

            if not doc_text or len(doc_text) < 20:
                logger.error(f"[DGSetDatasheet] Insufficient text: {len(doc_text) if doc_text else 0} chars")
                return {
                    'success': False,
                    'error': (
                        'Could not extract text from the PDF. '
                        'The file may be image-based or empty. '
                        'Please provide a text-based EDG sizing calculation document.'
                    )
                }

            # ── Template-first strategy ────────────────────────────────────────
            # Always start with the full template (unit + required_data always correct)
            datasheet_rows = self._get_default_datasheet_template()

            logger.info("[DGSetDatasheet] Running AI vendor data extraction…")
            vendor_map = self._extract_vendor_data_with_ai(doc_text, project_info)
            if vendor_map:
                merged = self._merge_vendor_data(datasheet_rows, vendor_map)
                logger.info(f"[DGSetDatasheet] Merged {merged} vendor values into template")

            summary = {
                'total_rows': len(datasheet_rows),
                'equipment_count': sum(1 for r in datasheet_rows if r.get('description', '').strip()),
                'completed_fields': sum(1 for r in datasheet_rows if r.get('vendor_data', '').strip()),
                'missing_fields': sum(1 for r in datasheet_rows if not r.get('vendor_data', '').strip()),
            }

            logger.info(f"[DGSetDatasheet] ✅ {summary['total_rows']} rows | {summary['completed_fields']} vendor values filled")
            return {
                'success': True,
                'datasheet_rows': datasheet_rows,
                'summary': summary,
                'extraction_metadata': {
                    'document_length': len(doc_text),
                    'project_info': project_info or {},
                }
            }

        except Exception as e:
            logger.error(f"[DGSetDatasheet] Error: {e}", exc_info=True)
            return {'success': False, 'error': str(e)}

    # ──────────────────────────────────────────────────────────────────────────
    # AI vendor data extraction + merge helpers
    # ──────────────────────────────────────────────────────────────────────────

    def _extract_vendor_data_with_ai(self, doc_text: str, project_info: Dict = None) -> Dict:
        """
        Use GPT-4o to extract parameter values from the uploaded document.
        Returns a dict mapping NORMALIZED_DESCRIPTION_UPPER → extracted_value.
        """
        prompt = f"""You are a senior electrical engineer specialising in Emergency Diesel Generator (EDG) sets. Read the following technical document and extract every parameter value you can find.

DOCUMENT:
{doc_text[:8000]}

TASK:
Return a JSON object where:
- Keys   = parameter/field names in UPPERCASE (e.g. "RATED POWER", "ENGINE MANUFACTURER", "RATED VOLTAGE")
- Values = the extracted value as a string

Focus on: equipment tags, kW/kVA ratings, voltages, currents, speeds (RPM), frequencies, fuel types, cooling types, manufacturers, engine model, alternator model, dimensions, weights, temperatures, quantities.

Rules:
- Include ONLY fields actually present in the document.
- Do NOT invent or assume values.
- Return ONLY a valid JSON object — no explanation, no markdown.

Example:
{{"RATED POWER": "500", "ENGINE MANUFACTURER": "CUMMINS", "RATED VOLTAGE": "415", "RATED FREQUENCY": "50", "POWER FACTOR": "0.8"}}"""

        try:
            response = self.client.chat.completions.create(
                model="gpt-4o",
                messages=[
                    {"role": "system", "content": "You are an expert electrical engineer. Extract parameter values from EDG/DG set documents. Return only valid JSON."},
                    {"role": "user", "content": prompt},
                ],
                temperature=0.1,
                max_tokens=2000,
            )
            ai_response = response.choices[0].message.content.strip()
            if "```json" in ai_response:
                ai_response = ai_response.split("```json")[1].split("```")[0]
            elif "```" in ai_response:
                ai_response = ai_response.split("```")[1].split("```")[0]
            vendor_map = json.loads(ai_response.strip())
            if isinstance(vendor_map, dict):
                normalized = {k.strip().upper(): str(v).strip() for k, v in vendor_map.items() if v and str(v).strip()}
                logger.info(f"[DGSetDatasheet] AI extracted {len(normalized)} vendor key-value pairs")
                return normalized
            return {}
        except json.JSONDecodeError as e:
            logger.error(f"[DGSetDatasheet] Vendor JSON decode error: {e}")
            return {}
        except Exception as e:
            logger.error(f"[DGSetDatasheet] Vendor AI extraction error: {e}")
            return {}

    def _merge_vendor_data(self, template_rows: List[Dict], vendor_map: Dict) -> int:
        """
        Merge vendor_map values into template_rows by description matching.
        Returns count of rows filled.
        """
        merged = 0
        stop_words = {'AND', 'OR', 'OF', 'THE', 'WITH', 'FOR', 'AT', 'IN', 'TO', 'A', 'AN', 'BY', 'ON'}
        for row in template_rows:
            if row.get('vendor_data'):
                continue
            desc = row.get('description', '').strip().upper()
            if not desc:
                continue
            # 1. Exact match
            if desc in vendor_map:
                row['vendor_data'] = vendor_map[desc]
                merged += 1
                continue
            # 2. Containment match
            matched = False
            for ai_key, ai_val in vendor_map.items():
                if not ai_key or not ai_val:
                    continue
                if ai_key in desc or desc in ai_key:
                    row['vendor_data'] = ai_val
                    merged += 1
                    matched = True
                    break
            if matched:
                continue
            # 3. Word-overlap match (≥2 meaningful common words)
            desc_words = set(desc.split()) - stop_words
            for ai_key, ai_val in vendor_map.items():
                if not ai_key or not ai_val:
                    continue
                ai_words = set(ai_key.split()) - stop_words
                if len(desc_words & ai_words) >= 2:
                    row['vendor_data'] = ai_val
                    merged += 1
                    break
        return merged

    def _LEGACY_extract_datasheet_with_ai(self, doc_text: str, project_info: Dict = None) -> List[Dict]:
        """LEGACY — kept for reference only. Use _extract_vendor_data_with_ai instead."""

        prompt = f"""You are a senior electrical engineer specialising in Emergency Diesel Generator (EDG) sets.
Analyse the provided EDG Sizing Calculation document and extract comprehensive datasheet information.

PROJECT INFORMATION:
{json.dumps(project_info or {}, indent=2)}

DOCUMENT CONTENT:
{doc_text[:7000]}

TASK:
Populate a standard ADNOC EDG set datasheet with EXACTLY 6 fields per row:
- SR_NO       : Sequential item number (e.g. 1, 2, A, B, C …). Blank for section-header rows.
- DESCRIPTION : Parameter name or section heading.
- UNIT        : Engineering unit (kW, kVA, V, A, Hz, RPM, ℃, %, dB, L, kg, etc.). Blank if not applicable.
- SPECIFIED_DESIGN_DATA : Required/design value (engineer-specified column).
- VENDOR_DATA  : Value extracted from the uploaded sizing-calculation document; empty string "" if not found.
- REV         : Revision marker – empty string "" unless explicitly noted.

Cover ALL sections below in order:

HEADER FIELDS (no section letter)
  Tag No., Title (EMERGENCY DIESEL GENERATOR SET), Manufacturer / Model / Country of Origin,
  Year of Manufacture, Quantity

GENERAL INFO
  Design Life, Criticality Rating, Inspection Class

REFERENCE SPECIFICATIONS
  9.1  Emergency Generator (1000kVA and Above)
  9.2  Synchronous AC Generators 1000kVA and Above
  9.3  Diesel Fuelled Compression Ignition Engines
  9.4  Fire Protection Design Philosophy
  9.5  Direct Current UPS System
  9.6  Painting
  9.7  Key Single Line Diagram
  9.8  Power, Control and Earthing Cables
  9.9  Instrument and Control Design Guideline
  9.10 Instrumentation Furnished with Package Units
  9.11 Local Control Panels

B – ENVIRONMENTAL CONDITIONS
  1. Type of Installation
  2. Atmosphere
  3. Design Ambient Temperature
  4. Altitude
  5. Minimum Ambient Temperature
  6. Maximum Relative Humidity
  7. Average Relative Humidity
  8. Degree of Protection (IP)
  9. Solar Background Radiation Heat Flux
  10. Site Class Definition

C – GENERAL TECHNICAL CHARACTERISTICS (ALTERNATOR)
  1.  Rated Voltage
  2.  Phases
  3.  Frequency
  4.  Name Plate kW / kVA
  5.  Power Factor (PF)
  6.  Speed
  7.  Rotor Construction
  8.  Armature (Stator) Insulation Class / Rise
  9.  Field (Rotor) Insulation Class / Rise
  10. Exciter Insulation Class / Rise
  11. Minimum % Overspeed
  12. Maximum Unique Equipment Vertical Thrust
  13. Bearing Type – Sleeve
  14. Bearing Type – Anti-Friction
  15. Type (Synchronous / Induction)
  16. Service
  17. Duty Type

D – AREA CLASSIFICATION
  1. Zone
  2. Group
  3. Area
  4. Temp Class
  5. Outdoor
  6. Roof Over
  7. Max Sound Pressure Level

E – UNUSUAL CONDITION
  1. Abrasive Dust
  2. External Forces & Moments
  3. Seismic Loading
  4. Corrosive Agents

F – ELECTRICAL SYSTEM CONDITION
  1. Type of System Grounding
  2. Neutral Isolation Switch
  3. 3-Phase Symmetrical Fault Current
  4. Earth Fault Ampere
  5. Electrical Phase Rotation (ABC or ACB)

G – ENGINE CHARACTERISTICS
  1. Engine Manufacturer / Model
  2. Engine Type
  3. Number of Cylinders
  4. Bore × Stroke
  5. Rated Engine Speed at Full Load
  6. Maximum Continuous Engine Power (kW)
  7. Standby Engine Power (kW)
  8. Engine Starting System
  9. Number of Starting Attempts
  10. Starting-to-Full-Load Time
  11. Fuel Consumption at Full Load
  12. Fuel Consumption at 75% Load
  13. Fuel Consumption at 50% Load
  14. Fuel Type / Grade
  15. Fuel Tank Capacity (Day Tank)
  16. Fuel Tank Autonomy
  17. Lube Oil Pressure
  18. Lube Oil Temperature
  19. Coolant Temperature (Inlet / Outlet)
  20. Radiator Cooling – Air Flow

H – GENERATOR / ALTERNATOR CHARACTERISTICS
  1. Alternator Manufacturer / Model
  2. Alternator Type (Brushless / Static Excitation)
  3. Rated Output (kVA)
  4. Rated Voltage (V)
  5. Rated Frequency (Hz)
  6. Rated Speed (RPM)
  7. Number of Poles
  8. Power Factor
  9. Efficiency at 100% Load
  10. Efficiency at 75% Load
  11. Voltage Regulation (No Load to Full Load)
  12. Short Circuit Ratio
  13. Subtransient Reactance (Xd'')
  14. Transient Reactance (Xd')
  15. Synchronous Reactance (Xd)
  16. Winding Temperature Rise (Class F Limit)
  17. Excitation System Type
  18. AVR Manufacturer / Model

I – FUEL SYSTEM
  1. Fuel System Type (Gravity / Pump)
  2. Day Tank Capacity
  3. Day Tank Material
  4. Main Fuel Tank Capacity
  5. Fuel Transfer Pump (Duty / Standby)
  6. Fuel Filter Type
  7. Fuel Level Gauge / Alarm

J – COOLING SYSTEM
  1. Cooling Type (Radiator / Remote Radiator)
  2. Cooling Fan Drive
  3. Radiator Fan Motor Rating (kW)
  4. Coolant Type
  5. Coolant Capacity
  6. Radiator Pressure Cap Setting
  7. Expansion Tank

K – LUBRICATION SYSTEM
  1. Lube Oil Pump Type
  2. Lube Oil Filter Type
  3. Lube Oil Capacity
  4. Lube Oil Grade / Specification
  5. Lube Oil Pressure (Normal Operating)
  6. Pre-Lubrication Provision

L – EXHAUST SYSTEM
  1. Exhaust Temperature at Rated Load
  2. Exhaust Back Pressure (Maximum Allowable)
  3. Exhaust Silencer Type (Industrial / Critical / Hospital)
  4. Exhaust Pipe Material
  5. Exhaust Lagging

M – CONTROL AND PROTECTION PANEL
  1. Control Panel Type (AMF / ATS / Paralleling)
  2. Enclosure Protection (IP Rating)
  3. Control Voltage
  4. Battery Charger
  5. Synchronising Facility
  6. Speed Governor Type
  7. Engine Protection: Over-Speed Trip
  8. Engine Protection: Low Oil Pressure Trip
  9. Engine Protection: High Coolant Temperature Trip
  10. Generator Protection: Over-Voltage
  11. Generator Protection: Under-Voltage
  12. Generator Protection: Over-Frequency
  13. Generator Protection: Under-Frequency
  14. Generator Protection: Reverse Power
  15. Generator Protection: Short Circuit (Overcurrent)
  16. Generator Protection: Earth Fault

N – ACOUSTIC AND VIBRATION
  1. Max Sound Pressure Level at 1 m (dB(A))
  2. Max Sound Power Level (dB(A))
  3. Vibration Isolation System
  4. Anti-Vibration Mounts

O – CIVIL / STRUCTURAL
  1. Skid / Baseframe Material
  2. Skid Dimensions (L × W × H)
  3. Total Operating Weight
  4. Anchor Bolt Size / Pattern
  5. Weatherproof Canopy / Enclosure

P – INSPECTION AND TESTING
  1. Factory Acceptance Test (FAT)
  2. Load Test Duration at Full Load
  3. Transient Response Test
  4. Noise Level Test
  5. Vibration Test
  6. Insulation Resistance Test
  7. High Voltage Test
  8. Protection Relay Testing

Return ONLY a JSON array. Each element must have exactly these keys:
  "sr_no", "description", "unit", "required_data", "vendor_data", "rev"

Rules:
- Section header rows: sr_no = "", unit = "", required_data = "", vendor_data = "", rev = ""
- Extract ACTUAL values from the document for vendor_data; use "" when not found
- required_data = standard / typical requirement value for an ADNOC EDG per BGS-MA-004 / IEC 60034
- rev = "" always unless document specifies a revision letter
- Include every parameter listed above even if vendor_data is empty
- Return ONLY the JSON array – no markdown, no explanation"""

        try:
            response = self.client.chat.completions.create(
                model="gpt-4o",
                messages=[
                    {
                        "role": "system",
                        "content": (
                            "You are an expert electrical engineer specialising in "
                            "Emergency Diesel Generator (EDG) set datasheets per ADNOC BGS-MA-004 / IEC 60034 standards. "
                            "Return only valid JSON arrays."
                        ),
                    },
                    {"role": "user", "content": prompt},
                ],
                temperature=0.1,
                max_tokens=6000,
            )

            ai_response = response.choices[0].message.content.strip()

            # Strip markdown fences if present
            if "```json" in ai_response:
                ai_response = ai_response.split("```json")[1].split("```")[0]
            elif "```" in ai_response:
                ai_response = ai_response.split("```")[1].split("```")[0]

            datasheet_rows = json.loads(ai_response.strip())

            if isinstance(datasheet_rows, list) and len(datasheet_rows) > 0:
                for row in datasheet_rows:
                    row.setdefault("sr_no", "")
                    row.setdefault("description", "")
                    row.setdefault("unit", "")
                    row.setdefault("required_data", "")
                    row.setdefault("vendor_data", "")
                    row.setdefault("rev", "")
                    row.pop("remarks", None)       # remove legacy key
                    row.pop("specified_design_data", None)  # normalise key
                logger.info(f"[DGSetDatasheet] AI returned {len(datasheet_rows)} rows")
                return datasheet_rows
            else:
                logger.error("[DGSetDatasheet] Invalid AI response structure")
                return self._get_default_datasheet_template()

        except json.JSONDecodeError as e:
            logger.error(f"[DGSetDatasheet] JSON decode error: {e}")
            return self._get_default_datasheet_template()
        except Exception as e:
            logger.error(f"[DGSetDatasheet] AI extraction error: {e}")
            return self._get_default_datasheet_template()

    # ──────────────────────────────────────────────────────────────────────────
    # Default template  (full ADNOC EDG datasheet – DS-13-EE-403-70010)
    # ──────────────────────────────────────────────────────────────────────────
    def _get_default_datasheet_template(self) -> List[Dict]:
        """Return the FULL ADNOC EDG Set datasheet template matching DS-13-EE-403-70010."""
        def R(sr, desc, unit="", req=""):
            return {"sr_no": sr, "description": desc, "unit": unit,
                    "required_data": req, "vendor_data": "", "rev": ""}
        def H(desc):
            return {"sr_no": "", "description": desc, "unit": "",
                    "required_data": "", "vendor_data": "", "rev": ""}
        return [
            # ═══════════════════════════════════════════════════════════════
            # SHEET 06/23 – GENERAL DATA
            # ═══════════════════════════════════════════════════════════════
            H("EMERGENCY DIESEL GENERATOR SET"),
            H("A   GENERAL DATA"),
            R("1",   "TAG NO.",                                       "-",        "13-EG-0101"),
            R("2",   "TITLE",                                         "-",        "EMERGENCY DIESEL GENERATOR SET"),
            R("3",   "MANUFACTURER / MODEL / COUNTRY OF ORIGIN",      "-",        "*** (AS PER COMPANY APPROVED LIST)"),
            R("4",   "YEAR OF MANUFACTURE",                           "-",        "***"),
            R("5",   "QUANTITY",                                      "No",       "1"),
            R("6",   "DESIGN LIFE",                                   "-",        "MINIMUM SERVICE LIFE 25 YEARS"),
            R("7",   "CRITICALITY RATING",                            "-",        "3"),
            R("8",   "INSPECTION CLASS",                              "-",        "1"),
            H("9    REFERENCE SPECIFICATION"),
            R("9.1", "EMERGENCY GENERATOR",                           "-",        "BGS-MA-004"),
            R("9.2", "SYNCHRONOUS AC GENERATORS 1250KVA AND ABOVE",   "-",        "BGS-MV-004"),
            R("9.3", "DIESEL FUELLED COMPRESSION IGNITION ENGINES",   "-",        "BGS-MV-003"),
            R("9.4", "FIRE PROTECTION DESIGN PHILOSOPHY",             "-",        "BGS-MU-200"),
            R("9.5", "DIRECT CURRENT UPS SYSTEM",                     "-",        "BGS-EE-009"),
            R("9.6", "PAINTING",                                      "-",        "BGS-MX-001"),
            R("9.7", "KEY SINGLE LINE DIAGRAM - HYDROGEN EXTRACTION UNIT", "-",  "13-EE-153-00001"),
            R("9.8", "POWER, CONTROL AND EARTHING CABLES",            "-",        "BGS-EE-011"),
            R("9.9", "INSTRUMENT AND CONTROL DESIGN GUIDELINE",       "-",        "BGS-IU-001"),
            R("9.10","INSTRUMENTATION FURNISHED WITH PACKAGE UNITS",  "-",        "BGS-IU-007"),
            R("9.11","LOCAL CONTROL PANELS",                          "-",        "BGS-IU-023"),

            H("B   ENVIRONMENTAL CONDITIONS"),
            R("1",  "TYPE OF INSTALLATION",                  "-",   "-OUTDOOR FOR DIESEL GENERATOR CONTAINER, -INDOOR FOR REMOTE CONTROL, PROTECTION AND MONITORING PANEL"),
            R("2",  "ATMOSPHERE",                            "-",   "SALTY, SULFUROUS AND DUSTY WITH HIGH CONCENTRATION OF WINDBORNE SAND"),
            R("3",  "DESIGN AMBIENT TEMPERATURE",            "",    "54 °C FOR DIESEL GENERATOR CONTAINER, 40°C FOR REMOTE CONTROL, PROTECTION & MONITORING PANEL"),
            R("4",  "ALTITUDE",                              "M",   ""),
            R("5",  "MINIMUM AMBIENT TEMPERATURE",           "°C",  "LESS THAN 1000m AMSL"),
            R("6",  "MAXIMUM RELATIVE HUMIDITY",             "°C",  "95% AT 43°C"),
            R("7",  "AVERAGE RELATIVE HUMIDITY",             "°C",  "60% AT 54°C"),
            R("8",  "DEGREE OF PROTECTION (IP)",             "-",   "IP55 (MINIMUM)"),
            R("9",  "SOLAR BACKGROUND RADIATION HEAT FLUX",  "-",   "95 W/M2"),
            R("10", "SITE CLASS DEFINITION",                 "-",   "SITE CLASS C"),

            H("C   GENERAL TECHNICAL CHARACTERISTICS"),
            R("1",   "RATED VOLTAGE",                                 "V",        "415 V±10 %"),
            R("2",   "PHASES",                                        "-",        "3"),
            R("3",   "FREQUENCY",                                     "Hz",       "50 Hz±2 %"),
            R("4",   "NAME PLATE kW / kVA",                           "kW / kVA", "720 kW / 900 kVA"),
            R("5",   "PF",                                            "",         "0.8"),
            R("6",   "SPEED",                                         "RPM",      "1500 rpm"),
            R("7",   "ROTOR CONSTRUCTION",                            "",         "CYLINDRICAL"),
            R("8",   "POLE",                                          "",         "4"),
            R("9",   "ARMATURE (STATOR) INSULATION CLASS / RISE",     "",         "F / B (66°C ABOVE 54°C AMBIENT)"),
            R("10",  "FIELD (ROTOR) INSULATION CLASS / RISE",         "",         "F / B (66°C ABOVE 54°C AMBIENT)"),
            R("11",  "EXCITER INSULATION CLASS / RISE",               "",         "F / B (66°C ABOVE 54°C AMBIENT)"),
            R("12",  "MINIMUM % OVERSPEED",                           "",         "120 % FOR DURATION OF 2 MIN. (REFER BGS-MA-004, CL 12.3.8)"),
            R("13",  "MAXIMUM DRIVEN EQUIPMENT VERTICAL THRUST",      "",         "***"),
            R("14",  "BEARING TYPE",                                  "",         "***"),
            R("14.1","SLEEVE",                                        "",         "***"),
            R("14.2","ANTIFRICTION",                                  "",         "***"),
            R("15",  "TYPE",                                          "",         "SYNCHRONOUS"),
            R("16",  "SERVICE",                                       "",         "CONTINUOUS FULL LOAD"),
            R("17",  "DUTY TYPE",                                     "",         "S1"),

            H("D   AREA CLASSIFICATION"),
            R("1",  "ZONE",                     "",  "NOT APPLICABLE"),
            R("2",  "GROUP",                    "",  "NOT APPLICABLE"),
            R("3",  "AREA",                     "",  "UNCLASSIFIED & REFER NOTE 7"),
            R("4",  "TEMP CLASS",               "",  "NOT APPLICABLE"),
            R("5",  "OUTDOOR",                  "",  "YES"),
            R("6",  "ROOF OVER",                "",  "NO ROOF OVER"),
            R("7",  "MAX. SOUND PRESSURE LEVEL","",  "85 dBA @ 1 m. FOR TOTAL SKID (DIESEL ENGINE & ALTERNATOR) AS PER BGS-MU-009 & BGS-MA-004"),

            H("E   UNUSUAL CONDITION"),
            R("1",  "ABRASIVE DUST",              "",  "INFORMATION WILL PROVIDE LATER"),
            R("2",  "EXTERNAL FORCES & MOMENTS",  "",  "INFORMATION WILL PROVIDE LATER"),
            R("3",  "SEISMIC LOADING",            "",  "INFORMATION WILL PROVIDE LATER"),
            R("4",  "CORROSIVE AGENTS",           "",  "INFORMATION WILL PROVIDE LATER"),

            H("F   ELECTRICAL SYSTEM CONDITION"),
            R("1",  "TYPE OF SYSTEM GROUNDING",                   "",  "REFER NOTE 8"),
            R("2",  "NEUTRAL ISOLATION SWITCH",                    "",  "YES, REQUIRED"),
            R("3",  "3 PHASE SYMMETRICAL FAULT CURRENT",           "",  "65 kA (INCLUDING DG CONTRIBUTION) FOR 3 SEC."),
            R("4",  "EARTH FAULT AMPERES",                         "",  "***"),
            R("5",  "ELECTRICAL PHASE ROTATION (ABC OR ACB)",      "",  "ABC"),

            # ═══════════════════════════════════════════════════════════════
            # SHEET 07/23 – PACKAGE COMPOSITION
            # ═══════════════════════════════════════════════════════════════
            H("G   PACKAGE COMPOSITION"),
            R("1",  "DIESEL ENGINE",                                         "",  "INSIDE CONTAINER"),
            R("2",  "AIR FILTERS",                                           "",  "INSIDE CONTAINER"),
            R("3",  "SILENCER",                                              "",  "TOP OF THE CONTAINER"),
            R("4",  "FUEL PUMPS AND FILTERATION",                            "",  "INSIDE CONTAINER"),
            R("5",  "DAY TANK WITH 24 HOURS FUEL CONSUMPTION",               "",  "DAY TANK SUPPLIED LOOSE FOR OUTDOOR INSTALLATION"),
            R("6",  "LUBRICATION SYSTEM",                                    "",  "INSIDE CONTAINER"),
            R("7",  "START-UP BATTERIES",                                    "",  "INSIDE CONTAINER"),
            R("8",  "GENERATOR",                                             "",  "INSIDE CONTAINER"),
            R("9",  "PMG AND EXCITERS",                                      "",  "INSIDE CONTAINER"),
            R("10", "AUTOMATIC VOLTAGE REGULATOR",                           "",  "REMOTE PANEL (NOTE 5)"),
            R("11", "AUTOMATIC FIRE PROTECTION SYSTEM",                      "",  "INSIDE CONTAINER & SHALL BE SUITABLE FOR 54°C TEMP."),
            R("12", "INSTRUMENT GAUGE BOARD",                                "",  "INSIDE CONTAINER & SHALL BE SUITABLE FOR 54°C TEMP."),
            R("13", "GENERATOR PROTECTION",                                  "",  "REMOTE PANEL"),
            R("14", "DG SEQUENCER AND PACKAGE CONTROL SYSTEM",               "",  "REMOTE PANEL"),
            R("15", "MIMIC PANEL AND SUPERVISORY CONTROL PANEL",             "",  "REMOTE PANEL"),
            R("16", "CONTROL BATTERIES",                                     "",  "REMOTE PANEL"),
            R("17", "BATTERY CHARGERS FOR START-UP BATTERIES",               "",  "INSIDE CONTAINER (NOTE 4)"),
            R("18", "BATTERY CHARGERS FOR CONTROL BATTERIES",                "",  "REMOTE PANEL"),
            R("19", "STARTERS FOR VARIOUS AUXILIARY MOTORS, HEATERS",        "",  "REMOTE PANEL"),
            R("20", "FIRE & GAS DETECTION SYSTEM (BGS-MU-200 & 13-PH-103-00009)","","INSIDE CONTAINER"),

            H("H   PAINTING"),
            R("1",  "MFR'S STANDARD",         "",  "AS PER BGS-MX-001"),
            R("2",  "EXTRA COAT @ SHIPMENT",   "",  "YES"),
            R("3",  "USER'S SPECIAL PAINT",   "",  "***"),

            H("I   DRIVE SYSTEM"),
            R("",   "DRIVE SYSTEM",            "",  "YES"),
            R("1",  "DIRECT CONNECTED",        "",  "CW"),
            R("2",  "ROTATION FACING DRIVE END","", "***"),

            H("J   SHIPMENT"),
            R("1",  "EXPORT",                                     "",  "YES"),
            R("2",  "EXPORT BOXING REQUIRED",                     "",  "YES"),
            R("3",  "OUTDOOR STORAGE FOR MORE THAN SIX MONTHS",   "",  "YES"),
            R("4",  "SPECIAL SHIPPING BEARING",                   "",  "***"),
            R("5",  "MOUNTED ON SKID",                            "",  "***"),

            H("K   MISCELLANEOUS"),
            R("1",  "EPOXY GROUT TYPE",                              "",  "***"),
            R("2",  "NAME PLATE MATERIAL",                           "",  "SS"),
            R("3",  "SPECIAL NAMEPLATE FOR ID",                      "",  "***"),
            R("4",  "QTY. OF SPECIAL TOOLS REQUIRED",               "",  "***"),
            R("5",  "PROOF OF NON SPARKING CORROSION RESISTANT FAN", "",  "***"),
            R("6",  "SHAFT SEALS (NON-CONDUCTIVE)",                  "",  "***"),
            R("7",  "LOCATION OF AUX. JUNCTION BOXES",              "",  "***"),
            R("8",  "SWITCHES WITH CONTACTS RATED FOR DC",           "",  "YES"),

            H("L   OTHER REQUIREMENT"),
            R("",   "VENDOR REPRESENTATIVE TO BE PRESENT DURING INITIAL EQUIPMENT ALIGNMENT CHECK", "",  "YES"),
            R("",   "GUARANTEE EFFICIENCY",                          "",  "***"),

            # ═══════════════════════════════════════════════════════════════
            # SHEETS 08-12 – DIESEL ENGINE AND AUXILIARIES
            # ═══════════════════════════════════════════════════════════════
            H("DIESEL ENGINE AND AUXILIARIES"),

            H("A   DIESEL ENGINE"),
            R("1",  "REFERENCE STANDARD FOR ENGINE",                              "",      "ISO 3046 PART 1 TO 7"),
            R("2",  "POWER OUTPUT (CDP)",                                         "kW",    "***"),
            R("3",  "ABSORBED POWER OF ENGINE DRIVEN AUXILIARIES (AS PER ISO 3046-1)", "kW","***"),
            R("4",  "FUEL CONSUMPTION (AT ISO RATING AND SERVICE POWERS)",         "gr/kWh","***"),
            R("5",  "LUBRICANT CONSUMPTION (AT ISO RATING & SERVICE POWERS)",      "gr/kWh","***"),
            R("6",  "EMISSION CONTROL (IN ACCORDANCE WITH HSE-EN-ST02)",           "",      "***"),
            R("7",  "MODEL NO. / MANUFACTURER",                                   "",      "***"),
            R("8",  "SERVICE",                                                    "",      "VITAL (ENGINE SHALL BE ABLE TO ACCEPT FULL LOAD WITHIN 10 SECONDS OF RECEIVING STARTING SIGNAL)"),
            R("9",  "PAINT SHADE",                                                "",      "AS PER BGS-MX-001"),
            R("10", "ARRANGEMENT OF CYLINDERS",                                   "",      "***"),
            R("11", "V-ANGLE",                                                    "",      "***"),
            R("12", "SPEED",                                                      "",      "***"),
            R("13", "DIRECTION OF ROTATION (VIEW AT FLYWHEEL END)",               "",      "***"),
            R("14", "BORE",                                                       "",      "***"),
            R("15", "STROKE",                                                     "",      "***"),
            R("16", "PISTON DISPLACEMENT",                                        "",      "***"),
            R("17", "PISTON SPEED",                                               "",      "***"),
            R("18", "COMPRESSION RATIO",                                          "",      "***"),
            R("19", "CYCLES",                                                     "",      "***"),
            R("20", "FIRING ORDER",                                               "",      "***"),
            R("21", "BRAKE MEAN EFFECTIVE PRESSURE",                              "",      "***"),
            R("22", "CHARGE AIR PRESSURE",                                        "",      "***"),
            R("23", "CHARGE AIR TEMPERATURE",                                     "",      "***"),
            R("24", "LUBE OIL CONSUMPTION",                                       "",      "***"),
            R("25", "CRITICAL SPEED",                                             "",      "***"),
            R("26", "FUEL FLOW TO TRANSFER PUMP",                                 "",      "***"),
            R("27", "STARTUP TIME UNTIL LOADING WITH RATED POWER",                "",      "***"),
            H("28  HEAT REJECTION TO"),
            R("28.1","EXHAUST",            "", "***"),
            R("28.2","COOLANT",            "", "***"),
            R("28.3","AFTER-COOLER COOLANT","","***"),
            R("28.4","OIL COOLER COOLANT", "", "***"),
            R("28.5","ATMOSPHERE FROM ENGINE","","***"),
            R("30", "AIR FLOW COOLING RATIO",                     "", "***"),
            R("31", "COOLING WATER FLOW RATE",                    "", "***"),
            R("32", "ENGINE OIL SUMP VOLUME (MIN/MAX)",           "", "***"),
            R("33", "MAX SURFACE TEMPERATURE OF ENTIRE ENGINE",   "", "***"),
            R("34", "DIESEL ENGINE WEIGHT",                       "", "***"),
            R("35", "WEIGHT TOTAL (INCL. FLYWHEEL)",              "", "***"),
            R("36", "ENGINE INSTRUMENT PANEL",                    "", "***"),
            R("37", "START & STOP PUSHBUTTONS",                   "", "***"),
            R("38", "ESD PUSHBUTTON",                             "", "***"),
            R("39", "PERFORMANCE CLASS OF ENGINE",                "", "***"),
            R("40", "POWER OF GENERATOR AS PER ISO 8528",         "", "***"),

            H("B   COMBUSTION AND AIR INTAKE SYSTEM"),
            H("    AIR FILTERS"),
            R("1",  "MANUFACTURER",                                           "",  "***"),
            R("2",  "SHELTER FOR RAIN INGRESS",                               "",  "***"),
            R("3",  "FILTER FOR PARTICLES GREATER THAN 10 MICROMETER (ASHRAE 52)","","YES"),
            R("4",  "MATERIAL",                                               "",  "NON CORRODING MATERIALS - ***"),
            R("4.1","STAINLESS STEEL",                                        "",  "STAINLESS STEEL - ***"),
            R("4.2","ALUMINUM ALLOY (<3% MAGNESIUM)",                         "",  "ALUMINUM ALLOY (<3% MAGNESIUM) ***"),
            R("5",  "INSTALLATION",                                           "",  "WITHIN CONTAINER BY PACKAGE VENDOR"),
            H("6   AIR FROM"),
            R("6.1","INSIDE",  "", "***"),
            R("6.2","OUTSIDE", "", "***"),
            R("7",  "MAX. AIR CLEANER RESTRICTION",    "", "***"),
            R("8",  "INTAKE MANIFOLD PRESSURE",        "", "***"),
            R("9",  "COMBUSTION AIR INLET FLOW RATE",  "", "***"),
            R("10", "TURBO-CHARGER",                   "", "***"),
            H("11  COMBUSTION AIR COOLER"),
            R("11.1","INTER-COOLER",  "", "***"),
            R("11.2","AFTER-COOLER",  "", "***"),

            H("C   EXHAUST SYSTEM"),
            H("    SILENCER"),
            R("1",  "MANUFACTURER",                                  "",  "***"),
            R("2",  "MATERIAL",                                      "",  "***"),
            R("3",  "EXHAUST DUCT (TO BE AGREED WITH CONTRACTOR)",   "",  "***"),
            R("4",  "CORROSION PROTECTION",                          "",  "***"),
            R("5",  "SURFACE TEMPERATURE",                           "",  "***"),
            R("6",  "EXHAUST PIPING TYPE",                           "",  "***"),
            H("7   EXHAUST SILENCER"),
            R("7.1","INSIDE",    "", "***"),
            R("7.2","OUTSIDE",   "", "***"),
            R("7.3","CONTAINER", "", "***"),
            R("8",  "SILENCER SPARK ARRESTER",                       "",  "***"),
            R("9",  "EXHAUST GAS TEMPERATURE",                       "",  "***"),
            R("10", "EXHAUST GAS FLOW RATE AT RATED LOAD",           "",  "***"),
            R("11", "EXHAUST GAS NOISE LEVEL",                       "",  "***"),
            R("12", "MAX BACKPRESSURE",                              "",  "***"),
            R("13", "EXHAUST MANIFOLD COOLING",                      "",  "***"),
            R("14", "TURBO-CHARGER",                                 "",  "***"),
            R("15", "TURBO-CHARGER COOLING",                         "",  "***"),
            R("16", "EXHAUST TEMPERATURE GAUGING POINTS",            "",  "***"),

            H("D   FUEL SYSTEM"),
            R("1",  "FUEL QUALITY",                           "",  "***"),
            R("2",  "FUEL SUPPLY PUMPS AND FILTERATION",      "",  "AS PER CLAUSE NO.: 13.4.2 OF BGS-MV-003"),
            R("3",  "ISOLATION VALVE",                        "",  "AS PER CLAUSE NO.: 13.4.3 OF BGS-MV-003"),
            R("4",  "FUEL SYSTEM INSTALLATION",               "",  "INSIDE CONTAINER"),
            R("5",  "DAY TANK",                               "",  "MATERIAL: STAINLESS STEEL 316L"),
            R("6",  "DAY TANK INSTALLATION",                  "",  "DAY TANK SUPPLIED LOOSE FOR OUTDOOR INSTALLATION"),
            R("7",  "DAY TANK INSTRUMENTATION",               "",  "AS PER CLAUSE NO.: 13.4.6 OF BGS-MV-003"),
            R("8",  "FULL TANK CAPACITY",                     "",  "***"),
            R("9",  "FULL-LOAD OPERATING PERIOD",             "",  "***"),
            R("10", "FUEL PIPING",                            "",  "***"),
            R("11", "TRANSFER PUMPS",                         "",  "***"),
            R("11.1","2 Nos",                                 "",  "***"),
            R("11.2","GEAR PUMP",                             "",  "***"),
            R("12", "FUEL FLOW AT RATED LOAD",                "",  "***"),
            H("13  FUEL PRESSURE"),
            R("13.1","BEFORE FILTER", "", "***"),
            R("13.2","AFTER FILTER",  "", "***"),
            R("14", "FUEL TEMPERATURE",                              "",  "***"),
            H("15  FUEL CONTROL AND MONITORING INSTRUMENTS"),
            R("15.1","LEVEL SWITCH LOW / HIGH ON FUEL",              "",  "***"),
            R("15.2","STORAGE TANK",                                 "",  "***"),
            R("15.3","LEVEL GAUGE ON FUEL STORAGE TANK",             "",  "***"),
            H("15.4 PRESSURE GAUGE"),
            R("15.4.1","BEFORE FILTER", "", "***"),
            R("15.4.2","AFTER FILTER",  "", "***"),
            R("15.5","DIFFERENTIAL-PRESSURE GAUGE",                  "",  "***"),
            R("16", "FUEL PUMP CONSUMPTION",                         "",  "***"),

            H("E   FUEL INJECTION"),
            R("",   "FUEL INJECTION",  "",  "AS PER VENDOR STANDARD"),

            H("F   LUBRICATION SYSTEM"),
            R("1",  "FILTERS",                  "",  "AS PER CLAUSE NO.: 13.6 OF BGS-MV-003"),
            R("2",  "OIL COOLING",              "",  "AS PER CLAUSE NO.: 13.6 OF BGS-MV-003"),
            R("3",  "PRE-LUBRICATION",          "",  "AS PER CLAUSE NO.: 13.6 OF BGS-MV-003"),
            R("4",  "CONDITION MONITORING",     "",  "AS PER CLAUSE NO.: 13.6 OF BGS-MV-003"),
            R("5",  "FIRE PREVENTION",          "",  "AS PER CLAUSE NO.: 13.6 OF BGS-MV-003"),
            R("6",  "OIL TYPE",                 "",  "***"),
            R("7",  "ENGINE DRIVEN GEAR PUMP",  "",  "***"),
            R("8",  "HEAT EXCHANGER",           "",  "***"),
            R("9",  "LUBE OIL FLOW TO ENGINE",  "",  "***"),
            H("10  LUBE OIL TEMPERATURE TO"),
            R("10.1","ENGINE",     "", "***"),
            R("10.2","OIL COOLER", "", "***"),
            R("11", "DIFFERENTIAL PRESSURE THROUGH FILTER",    "",  "***"),
            R("12", "ELECTRIC PRELUBRICATION PUMP",            "",  "***"),
            R("13", "LUBE OIL CONTROL AND MONITORING INSTRUMENTS","","***"),
            H("14  TEMPERATURE GAUGE OF"),
            R("14.1","ENGINE INLET",  "", "***"),
            R("14.2","ENGINE OUTLET", "", "***"),
            H("15  PRESSURE GAUGE"),
            R("15.1","BEFORE FILTER",           "", "***"),
            R("15.2","AFTER FILTER",            "", "***"),
            R("",   "DIFFERENTIAL-PRESSURE GAUGE",   "",  "***"),
            R("",   "RESERVOIR LEVEL INDICATOR",     "",  "***"),
            R("",   "REFILL VOLUME WITH FILTER CHANGE","", "***"),

            H("G   ENGINE COOLING"),
            R("1",  "COOLANT",                                   "",  "***"),
            R("2",  "METHOD OF COOLANT TEMPERATURE CONTROL",     "",  "***"),
            R("3",  "RADIATOR",                                  "",  "***"),
            R("4",  "RADIATOR FAN DRIVES",                       "",  "***"),
            R("5",  "ENGINE JACKET WATER",                       "",  "***"),
            H("6   JACKET WATER TEMPERATURE"),
            R("6.1","OUTLET", "", "***"),
            R("6.2","INLET",  "", "***"),
            R("7",  "PRESSURE",                                  "",  "***"),
            R("8",  "WATER PUMP TYPE",                           "",  "***"),
            R("9",  "RADIATOR SYSTEM CAPACITY",                  "",  "***"),
            R("10", "TOTAL AIR INLET REQUIREMENT",               "",  "***"),
            R("11", "PREHEATER",                                 "",  "***"),
            R("11.1","RATING",           "", "***"),
            R("11.2","CIRCULATING PUMP", "", "***"),
            R("12", "COOLING WATER CONTROL AND MONITORING INSTRUMENTS","","***"),
            H("13  TEMPERATURE GAUGE"),
            R("13.1","INLET",  "", "***"),
            R("13.2","OUTLET", "", "***"),
            H("14  EXPANSION TANK"),
            R("14.1","LEVEL INDICATOR",         "", "***"),
            R("14.2","LOW-END POSITION SWITCH", "", "***"),
            R("13", "CONNECTION FOR AIR FLOW TO OUTWARD BUILDING","","***"),

            H("H   STARTING SYSTEM"),
            R("1",  "ELECTRIC BATTERY STARTING SYSTEM",  "",  "YES"),
            R("2",  "BATTERY SYSTEM",                    "",  "DOUBLE SET OF BATTERY SYSTEMS WITH AUTOMATIC CHANGEOVER FACILITIES"),
            R("3.1","START ATTEMPTS WILL BE ALTERNATE BETWEEN BATTERY SETS","","START ATTEMPTS WILL BE ALTERNATE BETWEEN BATTERY SETS"),
            R("3.2","BATTERIES SHALL BE OF Ni-Cd (SEALED MAINTENANCE FREE)","","BATTERIES SHALL BE OF Ni-Cd (SEALED MAINTENANCE FREE)"),
            R("3",  "BATTERY CHARGERS",                  "",  "TO BE SUPPLIED BY VENDOR IN ACCORDANCE WITH BGS-EE-009"),
            R("4",  "BOOST VOLTAGE",                     "",  "***"),
            R("5",  "FLOAT VOLTAGE",                     "",  "***"),
            R("6",  "RECHARGING TIME OF EMPTY BATTERY",  "",  "***"),
            H("7   STARTING CONTROL AND MONITORING INSTRUMENTS"),
            R("7.1","DC VOLTMETER", "", "***"),
            R("7.2","DC AMMETER",   "", "***"),
            H("8   ALARM CONTACTS FOR"),
            R("8.1","FAILURE OF POWER SUPPLY",  "", "***"),
            R("8.2","MALFUNCTION OF CHARGER",   "", "***"),
            R("8.3","BATTERY UNDERVOLTAGE",     "", "***"),

            H("I   SPEED GOVERNING"),
            R("1",  "FREQUENCY STABILITY",  "",  "IN ACCORDANCE WITH GOVERNING CLASS A2 AS DEFINED IN ISO 3046-4"),
            R("2",  "GOVERNOR TYPE",         "",  "ELECTRONICS"),
            R("3",  "GOVERNOR CLASS",        "",  "***"),
            H("4   REMOTE CONTROL FROM"),
            R("4.1","SYNCHRONIZING UNIT", "", "***"),
            R("4.2","CONTROL CUBICLE",    "", "***"),
            R("5",  "OVERSPEED PROTECTION",  "",  "IN ACCORDANCE WITH ISO 3046-6"),
            R("5.1","SHUT OFF DEVICE WILL BE INDEPENDENT OF SPEED GOVERNOR","","SHUT OFF DEVICE WILL BE INDEPENDENT OF SPEED GOVERNOR"),

            H("J   TORSIONAL VIBRATION"),
            R("",   "TORSIONAL VIBRATION",  "",  "IN ACCORDANCE WITH ISO 3046-5 AND SPEED RANGE OF 90% TO 110%"),

            H("K   LATERAL AND VIBRATION SUPPRESSION"),
            R("",   "LATERAL AND VIBRATION SUPPRESSION",  "",  "VENDOR TO PERFORM ANALYSIS AS SPECIFIED IN 13.12 OF BGS-MV-003"),
            R("",   "ANTI VIBRATION MOUNTS SHALL BE SUPPLIED AS A PART OF DG PACKAGE","","ANTI VIBRATION MOUNTS SHALL BE SUPPLIED AS A PART OF DG PACKAGE"),

            H("L   SHAFT COUPLING AND GUARDS"),
            R("",   "SHAFT COUPLING AND GUARDS",  "",  "***"),
            R("1",  "TYPE",                        "",  "NON LUBRICATED TYPE"),

            H("M   CRANK CASE EXPLOSION RELIEF"),
            R("",   "CRANK CASE EXPLOSION RELIEF", "",  "(REQUIRED VENT / DAMPER SHALL BE PROVIDED) ***"),

            H("N   BASE PLATE"),
            R("",   "BASE PLATE",  "",  "SINGLE RIGID BASEPLATE FOR ENGINE AND DRIVEN EQUIPMENT"),

            H("O   PIPING"),
            R("",   "PIPING",  "",  "BY VENDOR IN ACCORDANCE WITH ASME B31.3"),

            H("P   INSTRUMENTATION AND CONTROL"),
            R("1",  "OPERATING MODE OF ENGINE",  "",  "MANUAL INITIATION OF AN AUTOMATIC START AND STOP SEQUENCE ***"),
            R("",   "FULLY AUTOMATIC START, LOAD ACCEPTANCE & STOP INITIATED BY A LOCAL OR REMOTE SIGNAL ***","","FULLY AUTOMATIC START, LOAD ACCEPTANCE & STOP INITIATED BY A LOCAL OR REMOTE SIGNAL ***"),
            R("2",  "REALIZED BY",               "",  "PROGRAMMABLE LOGIC CONTROLLER (DG SEQUENCER AND PACKAGE CONTROL SYSTEM) ***"),

            H("Q   MONITORING AND PROTECTION INSTRUMENTATION (VENDOR TO FURNISH THE DATA)"),
            H("    FUNCTION  |  INDICATION / ALARM / SHUTDOWN"),
            H("Q.1 GENERAL"),
            R("1",  "ENGINE SPEED",           "", "***"),
            R("2",  "HOURS RUN",              "", "***"),
            R("3",  "TURBOCHARGER SPEED",     "", "***"),
            R("4",  "ENGINE VIBRATION",       "", "***"),
            R("5",  "CYLINDER HEAD VIBRATION","", "***"),
            R("6",  "TURBOCHARGER VIBRATION", "", "***"),
            R("7",  "FAN VIBRATION",          "", "***"),
            R("8",  "START SEQUENCE",         "", "***"),
            R("9",  "START SEQUENCE FAIL",    "", "***"),
            R("10", "CONTROLS FAIL",          "", "***"),
            R("11", "MODE OF OPERATION",      "", "***"),
            R("12", "CRANKSHAFT KEYPHASOR",   "", "***"),
            H("Q.2 TEMPERATURE"),
            R("1",  "LUBE OIL TO ENGINE",              "", "***"),
            R("2",  "LUBE OIL TO COOLER",              "", "***"),
            R("3",  "LUBE OIL FROM COOLER",            "", "***"),
            R("4",  "COOLANT TO ENGINE",               "", "***"),
            R("5",  "COOLANT FROM ENGINE",             "", "***"),
            R("6",  "AIR INLET MANIFOLD",              "", "***"),
            R("7",  "EXHAUST - EACH CYLINDER",         "", "***"),
            R("8",  "EXHAUST - TURBOCHARGER INLET",    "", "***"),
            R("9",  "EXHAUST - TURBOCHARGER OUTLET",   "", "***"),
            R("10", "MAIN BEARING",                    "", "***"),
            R("11", "BIG END BEARINGS",                "", "***"),
            R("12", "TURBOCHARGER BEARINGS",           "", "***"),
            H("Q.3 PRESSURE"),
            R("1",  "LUBE OIL",            "", "***"),
            R("2",  "COOLANT",             "", "***"),
            R("3",  "AIR INLET MANIFOLD",  "", "***"),
            R("4",  "EXHAUST TO TURBOCHARGER","","***"),
            H("Q.3 DIFFERENTIAL PRESSURE"),
            R("1",  "LUBE OIL FILTER",  "", "***"),
            R("2",  "AIR FILTER",       "", "***"),
            H("Q.4 LEVEL"),
            R("1",  "LUBE OIL SUMP",       "", "***"),
            R("2",  "COOLANT",             "", "***"),
            R("3",  "AIR FILTER OIL BATH", "", "***"),
            R("4",  "FUEL DAY TANK",       "", "***"),

            # ═══════════════════════════════════════════════════════════════
            # SHEET 13/23 – ANALYSIS, SHOP INSPECTION
            # ═══════════════════════════════════════════════════════════════
            H("ANALYSIS, SHOP INSPECTION"),
            R("1",  "LATERAL CRITICAL SPEED ANALYSIS",                          "",  "***"),
            R("2",  "DESIGN REVIEW MEETING",                                    "",  "***"),
            R("3",  "TORSIONAL ANALYSIS",                                       "",  "***"),
            R("4",  "SHOP INSPECTION",                                          "",  "***"),
            R("5",  "DATA PRESERVATION",                                        "",  "***"),
            R("6",  "MATERIAL CERTIFICATION",                                   "",  "***"),
            R("7",  "ASSEMBLY MAINTENANCE AND RUNNING CLEARANCE",               "",  "***"),
            R("8",  "PAINTING DEFERRED",                                        "",  "***"),
            R("9",  "SURFACE AND SUBSURFACE INSPECTION OF PARTS",               "",  "***"),
            R("10", "RADIOGRAPHIC TEST PART",                                   "",  "***"),
            R("11", "ULTRASONIC TEST PART",                                     "",  "***"),
            R("12", "MAGNETIC PARTICLE TEST PARTS",                             "",  "***"),
            R("13", "LIQUID PENETRANT TEST PART",                               "",  "***"),
            R("14", "HARDNESS TEST PART",                                       "",  "***"),
            R("15", "OPTIONAL ASTM TEST TO BE PROPOSED BY VENDOR",              "",  "***"),
            R("16", "BALANCE [3 PLANES] [VACUUM PIT]",                          "",  "***"),
            R("17", "CHECK BALANCE WITH HALF COUPLING",                         "",  "***"),
            R("18", "INSPECTION FOR CLEANLINESS AS PER API 614",                "",  "***"),
            R("19", "REVIEW OF QUALITY CONTROL PROGRAM",                        "",  "***"),
            R("20", "STATOR INSPECTION PRIOR TO VACUUM PRESSURE IMPREGNATION",  "",  "***"),
            R("21", "RESIDUAL UNBALANCE TEST",                                  "",  "***"),
            R("22", "POLARIZATION INDEX VOLTAGE [1000V] [2500V] [5000V]",       "",  "***"),
            R("23", "VIBRATION RECORDING",                                      "",  "***"),
            R("24", "SPECIAL SURGE TEST ON SAMPLE COILS",                       "",  "***"),
            R("25", "POWER FACTOR TIP-UP TEST [WOUND STATOR] [SACRIFICIAL COILS]","","***"),
            R("26", "SEALED WINDING CONFORMANCE TEST",                          "",  "***"),
            R("27", "DC HIGH POTENTIAL TEST",                                   "",  "***"),
            R("28", "COMPLETE TEST (IEEE 115)",                                 "",  "***"),
            R("29", "EFFICIENCY",                                               "",  "***"),
            R("30", "LOCKED ROTOR",                                             "",  "***"),
            R("31", "OPEN SHORT-CIRCUIT SATURATION",                            "",  "***"),
            R("32", "HEAT RUN",                                                 "",  "***"),
            R("33", "HOT AND COLD VIBRATION",                                   "",  "***"),
            R("34", "UNBALANCED RESPONSE",                                      "",  "***"),
            R("35", "MANUFACTURER'S STANDARD SHOP AND ROUTINE TEST",           "",  "***"),
            R("36", "IMMERSION OR SPRAY TEST",                                  "",  "***"),
            R("37", "RATED ROTOR TEMP. VIBRATION TEST WHEN COMPLETE TEST NOT SPECIFIED","","***"),
            R("38", "BEARING INSPECTION AT COMPLETION OF TEST",                 "",  "***"),
            R("39", "UNBALANCE RESPONSE TEST",                                  "",  "***"),
            R("40", "BEARING HOUSING NATURAL FREQUENCY TEST",                   "",  "***"),
            R("41", "CERTIFIED DATA PRIOR TO SHIPMENT",                         "",  "***"),
            R("42", "NOISE LEVEL TEST",                                         "",  "***"),
            R("43", "OTHER",                                                    "",  "***"),

            # ═══════════════════════════════════════════════════════════════
            # SHEETS 14-16 – GENERATOR ENCLOSURE
            # ═══════════════════════════════════════════════════════════════
            H("GENERATOR ENCLOSURE"),

            H("A   GENERATOR ENCLOSURE"),
            R("1",  "DEGREE OF PROTECTION",                            "",  "IP55 (MINIMUM)"),
            R("2",  "OPEN - DRIPPROOF",                                "",  "***"),
            R("3",  "WEATHER PROTECTED (TYPE I / TYPE II)",            "",  "***"),
            R("4",  "TEFC (IP 55)",                                    "",  "CACA (IC0161)"),
            H("5   CACA TUBES"),
            R("5.1","COPPER",         "", "***"),
            R("5.2","Cu ALLOY",       "", "***"),
            R("5.3","ALUMINIUM",      "", "***"),
            R("5.4","STAINLESS STEEL","", "***"),
            R("5.5","AL ALLOY",       "", "***"),
            R("6",  "AISI 300 SERIES HARDWARE",  "",  "***"),
            R("7",  "MOUNTING",                  "",  "HORIZONTAL"),
            R("8",  "PAINT SHADE",               "",  "RAL 6011"),
            H("9   THE SYNCHRONOUS GENERATOR IS EQUIPPED WITH"),
            R("9.1","BRUSHLESS EXCITATION",             "",  "YES"),
            R("9.2","SELF EXCITATION",                  "",  "YES"),
            R("9.3","STATIC VOLTAGE REGULATOR",         "",  "YES"),
            R("9.4","ANTI-CONDENSATING HEATER",         "",  "YES"),
            R("9.5","DIODE FAILURE DETECTOR",           "",  "YES"),
            R("9.6","AUTOMATIC VOLTAGE REGULATOR",      "",  "YES"),
            R("9.7","ROTATING DIODES",                  "",  "YES"),
            R("9.8","STATIC VOLTAGE DROP COMPENSATION", "",  "YES"),
            R("9.9","2 SEPARATE EARTHING TERMINALS AT GEN. FRAME","","YES"),

            H("B   MAIN TERMINAL BOX"),
            R("1",  "BOX LOCATION",                                    "",      "AS PER BGS-MA-004 Clause No 13.1.3"),
            R("1.1","OVERSIZED BOX",                                   "",      "AS PER BGS-MA-004 Clause No 13.1.3"),
            R("2",  "CABLE SIZE",                                      "",      "TO BE PROVIDED LATER"),
            R("3",  "ENTER FROM",                                      "",      "BELOW"),
            R("3.1","ABOVE",                                           "",      "***"),
            R("3.2","BELOW",                                           "",      "YES"),
            R("3.3","SIDE (R/L)",                                      "",      "***"),
            R("4",  "MAIN TERMINAL BOX VOLTAGE / PHASE / FREQUENCY",   "V/Hz",  "415V / 3PHASE / 50Hz"),
            R("5",  "SPACE HEATERS",                                   "",      "REQUIRED"),
            R("6",  "SPACE HEATER VOLTAGE / PHASE / FREQUENCY",        "V/Hz",  "240V / 3PHASE / 50Hz"),
            R("7",  "MAX. SHEATH TEMPERATURE",                         "°C",    "5°C ABOVE AMBIENT"),
            R("8",  "FAULT WITHSTAND / SEC",                           "kA/SEC","50/3"),

            H("C   INSTRUMENT TRANSFORMERS LOCATED INSIDE DG CONTAINER (BY EPC CONTRACTOR)"),
            H("1   LINE CT's"),
            R("1.1","CT5 (87G)",   "", "***"),
            R("1.2","CORE 1",      "", "***"),
            R("1.3","RATIO Ip/Is", "", "***"),
            R("1.4","VA*",         "", "***"),
            R("1.5","CLASS",       "", "***"),
            H("2   CT4"),
            R("2.1","CORE 1",      "", "***"),
            R("2.2","RATIO Ip/Is", "", "***"),
            R("2.3","VA*",         "", "***"),
            R("2.4","CLASS",       "", "***"),
            H("3   NEUTRAL CT's"),
            R("3.1","CT2 (87G)",   "", "***"),
            R("3.2","CORE 1",      "", "***"),
            R("3.3","RATIO Ip/Is", "", "***"),
            R("3.4","VA*",         "", "***"),
            R("3.5","CLASS",       "", "***"),
            H("4   CT3"),
            R("4.1","CORE 1",      "", "***"),
            R("4.2","RATIO Ip/Is", "", "***"),
            R("4.3","VA*",         "", "***"),
            R("4.4","CLASS",       "", "***"),
            H("5   CT1"),
            R("5.1","STAR-POINT CT","","***"),
            R("5.2","RATIO Ip/Is", "", "***"),
            R("5.3","VA*",         "", "***"),
            R("5.4","CLASS",       "", "***"),
            H("6   VT's"),
            R("6.1","VT",          "", "***"),
            R("6.2","RATIO Up/Us", "", "***"),
            R("6.3","VA*",         "", "***"),
            R("6.4","CLASS",       "", "***"),

            H("D   WINDING TEMPERATURE DETECTOR"),
            R("1",  "RTD TYPE",           "",   "YES"),
            R("2",  "NO. / PHASE",        "",   "2"),
            R("3",  "RESISTANCE MATERIAL","",   "PT 100 RTD ELEMENT, 100 Ohms AT 0°C WITH OVER VOLTAGE SURGE"),
            R("4",  "WIRE",               "",   "4 WIRE"),

            H("E   BEARING TEMPERATURE DETECTOR"),
            R("1",  "PROVISION ONLY",             "", "***"),
            R("",   "PER API 670",                "", "***"),
            R("",   "MFR. STD.",                  "", "***"),
            R("2",  "RTD TYPE MATERIAL",          "", "PT 100 RTD ELEMENT, 100 Ohms AT 0°C WITH OVER VOLTAGE SURGE"),
            R("3",  "TERMINAL HEAD OR BOX",       "", "BOX"),
            R("4",  "HOUSING TEMP. INDICATOR REQUIRED","","NO"),

            # ═══════════════════════════════════════════════════════════════
            # SHEETS 16-19 – SYNCHRONOUS GENERATOR DATA (VENDOR TO FILL)
            # ═══════════════════════════════════════════════════════════════
            H("SYNCHRONOUS GENERATOR DATA TO BE FILLED BY VENDOR"),

            H("A   OPERATING CONDITIONS"),
            R("1",  "RPM",                    "", "***"),
            R("2",  "ARMATURE AMPS",          "", "***"),
            R("3",  "FIELD AMPS",             "", "***"),
            R("4",  "SPEED RPM",              "", "***"),
            R("5",  "TRIP",                   "", "***"),
            R("6",  "OVERSPEED",              "", "***"),
            R("7",  "1st CRITICAL",           "", "***"),
            R("8",  "2nd CRITICAL",           "", "***"),
            H("9   EFFICIENCY (%) AT RATED PF"),
            R("9.1","100% LOAD", "", "***"),
            R("9.2","75% LOAD",  "", "***"),
            R("9.3","50% LOAD",  "", "***"),
            R("9.4","25% LOAD",  "", "***"),
            R("10", "VOLTAGE REGULATION",                          "", "***"),
            R("11", "VOLTAGE ADJUSTING RANGE",                     "", "***"),
            R("12", "OVERLOAD CAPACITY",                           "", "***"),
            R("13", "FREQUENCY ADJUSTING RANGE",                   "", "***"),
            H("14  SPEED CONTROL"),
            R("14.1","DROOP",        "", "***"),
            R("14.2","ISOCHRONOUS",  "", "***"),
            H("15  OPERATING MODE"),
            R("15.1","GENERATOR IN ISOLATED OPERATION", "", "***"),
            R("16", "STATIC VOLTAGE DEVIATION",                                    "", "***"),
            R("17", "DYNAMIC VOLTAGE DEVIATION AT LOADING WITH NORMAL LOAD COS PHI 0.8","","***"),
            R("18", "VOLTAGE RECOVERY TIME",                                       "", "***"),
            R("19", "STATIC FREQUENCY DEVIATION",                                  "", "***"),
            R("20", "DYNAMIC FREQUENCY DEVIATION AT LOADING WITH NORMAL LOAD COS PHI 0.8","","***"),
            R("21", "FREQUENCY RECOVERY TIME",                                     "", "***"),
            R("22", "ROTOR INERTIA (J)",                                            "", "***"),
            R("23", "ROTOR WITHDRAWAL SPACE FROM GEN. HOUSING",                    "", "***"),
            R("24", "UNBALANCE LOAD",                                              "", "***"),
            R("25", "SATURATION FACTOR",                                           "", "***"),

            H("B   CHARACTERISTICS"),
            R("1",  "SHORT CIRCUIT RATIO (SCR)",                    "", "***"),
            R("2",  "X/R RATIO",                                    "", "***"),
            H("3   TELEPHONE INFLUENCE FACTOR"),
            R("3.1","BALANCED",  "", "***"),
            R("3.2","RESIDUAL",  "", "***"),
            R("4",  "ROTOR SHORT TIME THERMAL CAPACITY I2 SQT",    "", "***"),
            R("5",  "GENERATOR INERTIA CONSTANT, H (SEC)",          "", "***"),
            R("6",  "3 PHASE ARMATURE WINDING CAPACITANCE",         "mF","***"),
            R("7",  "MINIMUM 3 PHASE MOTORING POWER",               "kW","***"),

            H("C   TIME CONSTANTS (SECONDS)"),
            R("1",  "DIRECT AXIS TRANSIENT OPEN CIRCUIT (T'do)",        "", "***"),
            R("2",  "DIRECT AXIS TRANSIENT SHORT CIRCUIT (T'd)",         "", "***"),
            R("3",  "DIRECT AXIS SUB-TRANSIENT OPEN CIRCUIT (T''do)",  "", "***"),
            R("4",  "DIRECT AXIS SUB-TRANSIENT SHORT CIRCUIT (T''d)",  "", "***"),
            R("5",  "QUADRANT AXIS SUB-TRANSIENT OPEN CIRCUIT (T''qo)","","***"),
            R("6",  "QUADRANT AXIS TRANSIENT OPEN CIRCUIT (T'qo)",      "", "***"),
            R("7",  "QUADRANT AXIS TRANSIENT SHORT CIRCUIT (Tq)",        "", "***"),
            R("8",  "QUADRANT AXIS SUB-TRANSIENT SHORT CIRCUIT (T''q)","","***"),
            R("9",  "ARMATURE WINDING 3 PHASE SHORT CIRCUIT (Ta3)",      "", "***"),

            H("D   REACTANCE (%) - SATURATED / UNSATURATED"),
            R("1",  "DIRECT AXIS ARMATURE REACTANCE (Xad)",         "",  "*** / ***"),
            R("2",  "DIRECT AXIS SYNCHRONOUS (Xd)",                 "",  "*** / ***"),
            R("3",  "DIRECT AXIS TRANSIENT (X'd)",                  "",  "15% / ***"),
            R("4",  "DIRECT AXIS SUB TRANSIENT (X''d)",            "",  "7% (NO NEGATIVE TOLERANCE) / ***"),
            R("5",  "QUADRANT AXIS ARMATURE REACTANCE (Xaq)",       "",  "*** / ***"),
            R("6",  "QUADRANT AXIS SYNCHRONOUS (Xq)",               "",  "*** / ***"),
            R("7",  "QUADRANT AXIS TRANSIENT (X'q)",               "",  "*** / ***"),
            R("8",  "QUADRANT AXIS SUB TRANSIENT (X''q)",          "",  "*** / ***"),
            R("9",  "NEGATIVE SEQUENCE (X0)",                       "",  "*** / ***"),
            R("10", "ZERO SEQUENCE (X2)",                           "",  "*** / ***"),
            R("11", "ARMATURE LEAKAGE (XL)",                        "",  "*** / ***"),
            R("12", "FIELD LEAKAGE REACTANCE (Xf)",                 "",  "*** / ***"),
            R("13", "POITER REACTANCE (Xp)",                        "",  "*** / ***"),

            H("E   COLD WINDING RESISTANCES (20°C)"),
            R("1",  "ARMATURE (PER PHASE)", "Ohms", "***"),
            R("2",  "FIELD",                "Ohms", "***"),

            H("F   EXCITATION DATA"),
            R("1",  "MAKE",                                                                "", "***"),
            R("2",  "VOLTAGE REGULATOR AND EXCITER CHARACTERISTICS PER IEEE PAS-100",     "", "***"),
            R("3",  "NOTE: EXCITER AND VOLTAGE REGULATOR CONSTANTS TO BE SUPPLIED WITH PROPOSAL","",""),
            H("4   TYPE"),
            R("4.1","STATIC",         "", "***"),
            R("4.2","ROTOR BRUSHLESS","", "***"),
            H("5   POWER SOURCE"),
            R("5.1","PPT",            "", "***"),
            R("5.2","PMG",            "", "***"),
            R("5.3","RESPONSE RATIO", "", "***"),
            R("6",  "MAXIMUM CEILING EXCITATION (PU) AND TIMES",       "", "(MIN. 120% OF RATED VOLTAGE OF EXCITER)"),
            R("7",  "CURRENT / VOLTAGE AT RATED LOAD (A/V)",           "", "***"),
            R("8",  "CURRENT / VOLTAGE AT NO LOAD (A/V)",              "", "***"),
            H("9   CHARACTERISTICS (AS PER IEEE 421)"),
            R("9.1","FORWARD GAIN TIME CONSTANT",                 "", "***"),
            R("9.2","STABILIZER FEEDBACK GAIN TIME CONSTANT",     "", "***"),
            R("9.3","OPEN CIRCUIT CEILING VOLTAGE",               "", "***"),
            R("9.4","ACCURACY OF REGULATOR",                      "", "***"),

            H("G   AUTOMATIC VOLTAGE REGULATOR"),
            R("1",  "MODEL NO. / MANUFACTURER",                      "",  "***"),
            R("2",  "NO. OF CHANNELS",                               "",  "***"),
            R("3",  "NO. OF VOLTAGE INPUTS",                         "",  "***"),
            R("4",  "HARMONIC FILTER (VOLTAGE, THD HARMONIC DISTORTION & ITS LIMIT AS PER IEC 519)","","***"),

            H("H   WEIGHTS"),
            R("1",  "GENERATOR (INCLUDING ROTOR)", "kg", "***"),
            R("2",  "ROTOR",                       "kg", "***"),

            H("I   BEARINGS AND LUBRICATION"),
            H("1   BEARING TYPE"),
            R("1.1","DRIVE END",     "", "***"),
            R("1.2","NON DRIVE END", "", "***"),
            R("2",  "LUBRICANT",          "",  "***"),
            R("3",  "LUBRICANT SYSTEM",   "",  "***"),
            H("4   BEARING INSULATED"),
            R("4.1","DRIVE END",     "", "***"),
            R("4.2","NON DRIVE END", "", "***"),
            R("5",  "BEARING LIFETIME",              "",  "***"),
            R("6",  "MAX RELUBRICATION INTERVAL",    "",  "***"),

            # ═══════════════════════════════════════════════════════════════
            # SHEETS 19-20 – REMOTE CONTROL, PROTECTION AND MONITORING PANEL
            # ═══════════════════════════════════════════════════════════════
            H("REMOTE CONTROL, PROTECTION AND MONITORING PANEL"),
            H("LOCATION - SUBSTATION"),

            H("A   CONSTRUCTION FEATURES"),
            R("1",  "DEGREE OF PROTECTION",        "",  "IP41 (MINIMUM)"),
            R("2",  "INSTALLATION",                "",  "FREE STANDING"),
            R("3",  "ENCLOSURE MATERIAL",          "",  "SHEET STEEL"),
            R("4",  "CABLE ENTRY",                 "",  "FROM BOTTOM"),
            R("5",  "ENCLOSURE DIMENSION THICKNESS","", "***"),
            R("6",  "DESIGN TEMPERATURE",          "",  "40°C"),
            R("7",  "OUTER SURFACE COLOUR",        "",  "RAL 7035"),
            R("8",  "INNER SURFACE COLOUR",        "",  "WHITE"),
            R("9",  "NAME PLATE & TAGS",           "",  "WHITE WITH ENGLISH BLACK WORDING"),
            R("10", "TYPICAL SCHEMATIC DIAGRAM",   "",  "DURING DETAILED ENGINEERING"),
            R("11", "CONTROL VOLTAGE",             "",  "24V DC"),
            R("12", "CONTROL VOLTAGE DERIVED FROM","",  "24V DC CONTROL BATTERIES"),

            H("B   VOLTMETERS"),
            R("1",  "MODEL NO. / MANUFACTURER", "",  "***"),
            R("2",  "SIZE",                     "",  "***"),
            R("3",  "ACCURACY CLASS",           "",  "***"),

            H("C   VOLTAGE SELECTOR SWITCHES"),
            R("1",  "MODEL NO. / MANUFACTURER", "",  "***"),
            R("2",  "SIZE",                     "",  "***"),
            R("3",  "NO. OF POSITIONS",         "",  "***"),
            R("4",  "CONTACT RATING",           "",  "***"),

            H("D   CONTROL SELECTION SWITCHES"),
            R("1",  "MODEL NO. / MANUFACTURER", "",  "***"),
            R("2",  "SIZE",                     "",  "***"),
            R("3",  "NO. OF POSITIONS",         "",  "***"),
            R("4",  "CONTACT RATING",           "",  "***"),

            H("E   SEMAPHORE INDICATORS"),
            R("1",  "MODEL NO. / MANUFACTURER", "",  "***"),
            R("2",  "SIZE",                     "",  "***"),
            R("3",  "NO. OF POSITIONS",         "",  "3 (2 N° STATUS + 1N° FOR CONT. VOLTAGE LOSS)"),
            R("4",  "COIL RATING",              "",  "***"),

            H("F   DG SEQUENCER AND PACKAGE CONTROL SYSTEM (DG PLC)"),
            R("1",  "MODEL NO. / MANUFACTURER", "",  "***"),
            R("2",  "NO. DIGITAL INPUTS",       "",  "TO BE PROVIDED LATER"),
            R("3",  "NO. DIGITAL OUTPUTS",      "",  "TO BE PROVIDED LATER"),
            R("4",  "HARDWARE REQUIREMENT",     "",  "***"),
            R("5",  "COMMUNICATION FEATURES",   "",  "MODBUS TCP"),

            H("G   ANNUNCIATION WINDOWS"),
            R("1",  "ANNUNCIATOR FACIA WINDOW SIZE",    "",  "***"),
            R("2",  "WINDOW LEDS COLOR",                "",  "PROGRAMMABLE / SELECTABLE"),
            R("3",  "COLOR OF MIMIC / MOSAIC PANEL",    "",  "***"),
            R("4",  "MATERIAL OF MIMIC / MOSAIC PANEL", "",  "***"),
            R("5",  "SIGNAL MANAGEMENT BY",             "",  "BY DG SEQUENCER AND PACKAGE CONTROL SYSTEM"),
            R("6",  "COMPOSITION",                      "",  "TO BE PROVIDED LATER"),

            H("H   PROTECTION RELAYS"),
            R("",   "MAKE/MODEL",                       "",  "GE MULTILINE 489SR OR EQUIVALENT"),
            H("1   FUNCTION: F49/F50/F51/F46/F51N/F38-49T/F27/F51V/F86G"),
            R("2",  "REQUIRED DATA (Qty)",              "",  "TO BE PROVIDED LATER"),
            R("3",  "BY VENDOR (Qty)",                  "",  "***"),
            R("4",  "CONSTRUCTION",                     "",  "MICROPROCESSOR BASED MULTIFUNCTION TYPE WITH DIGITAL DISPLAY AND SEVERAL COMMUNICATION FACILITY"),
            R("5",  "MAKE",                             "",  "***"),
            H("6   FUNCTION: F32, F59, F81U, F51G, F87G"),
            R("8",  "REQUIRED DATA (Qty)",              "",  "TO BE PROVIDED LATER"),
            R("9",  "BY VENDOR (Qty)",                  "",  "***"),
            R("10", "CONSTRUCTION",                     "",  "MICROPROCESSOR BASED MULTIFUNCTION TYPE WITH DIGITAL DISPLAY AND SEVERAL COMMUNICATION FACILITY"),
            R("11", "MAKE",                             "",  "***"),
            R("12", "DG CONTROL & SEQUENCING",          "",  "TO BE PROVIDED LATER"),

            H("I   AUTOSYNCHRONIZER"),
            R("1",  "MODEL NO. / MANUFACTURER",  "",  "***"),
            R("2",  "AUX. CONTROL VOLTAGE",      "",  "24 VDC"),
            R("3",  "NO. OUTPUT CONTACTS",       "",  "DURING DETAIL ENGINEERING BY EPC CONTRACTOR"),
            R("4",  "TYPE OF LINE INPUT VOLTAGES","",  "TWO LINES (L1-L2)"),

            H("J   BACK-UP RELAY SYSTEM"),
            R("1",  "AS PER CL.13.4.5 OF BGS-MA-004", "",  "REQUIRED AS APPLICABLE"),
            R("2",  "AS PER CL.16.11 OF BGS-MV-004",  "",  "REQUIRED AS APPLICABLE"),

            H("K   WEIGHT & DIMENSION"),
            R("1",  "TOTAL MASS",                   "",  "***"),
            R("2",  "OVERALL DIMENSION (L × H × W)","",  "***"),
            R("3",  "OVERALL DIMENSION DRAWING",    "",  "***"),

            # ═══════════════════════════════════════════════════════════════
            # SHEET 21/23 – DIESEL GENERATOR AUXILIARY DIST. BOARD
            # ═══════════════════════════════════════════════════════════════
            H("DIESEL GENERATOR AUXILIARY DISTRIBUTION BOARD"),
            H("LOCATION - SUBSTATION"),

            H("A   CONSTRUCTION FEATURES"),
            R("1",  "DEGREE OF PROTECTION",                "",  "IP41 (MINIMUM)"),
            R("2",  "INSTALLATION",                        "",  "FREE STANDING"),
            R("3",  "ENCLOSURE MATERIAL",                  "",  "SHEET STEEL"),
            R("4",  "SWITCHGEAR MOTORS / FEEDERS FORMATION","", "NON-COMPARTMENTALIZED, FIXED UNITS"),
            R("5",  "CABLE ENTRY",                         "",  "FROM BOTTOM"),
            R("6",  "ENCLOSURE DIMENSION THICKNESS",       "",  "By Vendor"),
            R("7",  "DESIGN TEMPERATURE",                  "",  "40°C"),
            R("8",  "OUTER SURFACE COLOUR",                "",  "RAL 7035"),
            R("9",  "INNER SURFACE COLOUR",                "",  "WHITE"),
            R("10", "NAME PLATE & TAGS",                   "",  "WHITE WITH ENGLISH BLACK WORDING"),
            R("11", "TYPICAL SCHEMATIC DIAGRAM",           "",  "DURING DETAILED ENGINEERING BY EPC CONTRACTOR"),
            R("12", "SUPPLY VOLTAGE",                      "",  "415V, 3PH, 4 WIRE SYSTEM WITH SOLIDLY EARTHED"),
            R("13", "CONTROL VOLTAGE",                     "",  "240V AC, 50Hz"),
            R("14", "CONTROL VOLTAGE DERIVED FROM",        "",  "LINE AND NEUTRAL"),
            R("15", "NO. OF INCOMERS",                     "",  "ONE"),
            R("16", "NO. OF OUTGOING FEEDERS",             "",  "AS REQUIRED BY PACKAGE SYSTEM"),
            R("17", "NO. OF MOTOR FEEDERS",                "",  "AS REQUIRED BY PACKAGE SYSTEM"),

            H("B   VOLTMETERS"),
            R("1",  "MODEL NO. / MANUFACTURER", "",  "***"),
            R("2",  "SIZE",                     "",  "***"),
            R("3",  "ACCURACY CLASS",           "",  "***"),

            H("C   VOLTAGE SELECTOR SWITCHES"),
            R("1",  "MODEL NO. / MANUFACTURER", "",  "***"),
            R("2",  "SIZE",                     "",  "***"),
            R("3",  "NO. OF POSITIONS",         "",  "***"),
            R("4",  "CONTACT RATING",           "",  "***"),

            H("D   CONTROL SELECTION SWITCHES"),
            R("1",  "MODEL NO. / MANUFACTURER", "",  "***"),
            R("2",  "SIZE",                     "",  "***"),
            R("3",  "NO. OF POSITIONS",         "",  "***"),
            R("4",  "CONTACT RATING",           "",  "***"),

            # ═══════════════════════════════════════════════════════════════
            # SHEET 22/23 – INCOMING SUPPLY LINE
            # ═══════════════════════════════════════════════════════════════
            H("INCOMING SUPPLY LINE"),
            R("A",  "ITEM / TYPICAL SCHEMATIC DIAGRAM",  "",  "MCCB WITH CLOSING AND TRIPPING COIL"),
            R("B",  "RATED POWER",                       "",  "AS REQUIRED BY PACKAGE SYSTEM"),
            R("C",  "RATED CURRENT",                     "",  "***"),
            H("D   SWITCHING EQUIPMENT"),
            R("1",  "CIRCUIT BREAKER TYPE", "",   "MOULDED CASE CIRCUIT BREAKER"),
            R("2",  "RATED CURRENT",        "A",  "***"),
            H("E   INSTRUMENT TRANSFORMER"),
            R("1",  "LINE CT's",    "", "***"),
            R("2",  "CT1",           "", "***"),
            R("3",  "CORE-1",        "", "***"),
            R("4",  "RATIO Ip/Is",   "", "***"),
            R("5",  "VA*",           "", "***"),
            R("6",  "CLASS",         "", "***"),
            R("7",  "REQUIRED DATA", "", "***"),
            H("F   MEASURING TRANSFORMER"),
            R("1",  "LOCAL INSTRUMENTS", "", "***"),
            R("2",  "VOLTAGE",           "", "***"),
            R("3",  "AMPERE",            "", "***"),
            H("G   PROTECTION RELAYS"),
            H("1   FUNCTION: F50, F51, F51N, F27, F86"),
            R("1",  "CONSTRUCTION", "",  "***"),
            R("2",  "MAKE",         "",  "***"),

            # ═══════════════════════════════════════════════════════════════
            # SHEET 23/23 – MOTOR FEEDER
            # ═══════════════════════════════════════════════════════════════
            H("MOTOR FEEDER"),

            H("A   SWITCHING EQUIPMENT"),
            R("1",  "FUSE SWITCH DISCONNECTOR","",  "MANUAL OPERATING, AIR BREAK TYPE, PADLOCKABLE IN OFF POSITION"),
            R("2",  "TYPE",         "",  "***"),
            R("3",  "MAKE",         "",  "***"),
            R("4",  "MODEL RANGE",  "",  "***"),
            R("5",  "RATING",       "",  "AS PER TYPE 2 CO-ORDINATION"),
            H("6   CONTACTOR"),
            R("6.1","TYPE",         "",  "ELECTRICALLY HELD TYPE"),
            R("6.2","MAKE",         "",  "***"),
            R("6.3","MODEL RANGE",  "",  "***"),
            R("6.4","RATING",       "",  "AS PER TYPE 2 CO-ORDINATION"),

            H("B   INSTRUMENT TRANSFORMER"),
            H("1   LINE CT's"),
            R("1.1","CT1",           "", "***"),
            R("1.2","CORE-1",        "", "***"),
            R("1.3","RATIO Ip/Is",   "", "***"),
            R("1.5","VA*",           "", "***"),
            R("1.5","CLASS",         "", "***"),
            R("1.6","REQUIRED DATA", "", "***"),
            H("2   LINE CT's"),
            R("2.1","CT2",           "", "***"),
            R("2.2","CORE-1",        "", "***"),
            R("2.3","RATIO Ip/Is",   "", "***"),
            R("2.4","VA*",           "", "***"),
            R("2.5","CLASS",         "", "***"),
            R("2.6","REQUIRED DATA", "", "***"),
            H("3   LINE CT's"),
            R("3.1","CT0",           "", "***"),
            R("3.2","CORE-1",        "", "***"),
            R("3.3","RATIO Ip/Is",   "", "***"),
            R("3.4","VA*",           "", "***"),
            R("3.5","CLASS",         "", "***"),
            R("3.6","REQUIRED DATA", "", "***"),

            H("C   MEASURING TRANSFORMER"),
            H("1   LOCAL INSTRUMENTS"),
            R("1.1","VOLTAGE", "", "***"),
            R("1.2","AMPERE",  "", "***"),

            H("D   PROTECTION RELAYS"),
            H("1   FUNCTION: F51, F49, F46, F50G"),
            R("3",  "CONSTRUCTION", "", "***"),
            R("4",  "MAKE",         "", "***"),

            H("E   WEIGHT & DIMENSION"),
            R("1",  "TOTAL MASS",                   "",  "***"),
            R("2",  "OVERALL DIMENSION (L × H × W)","",  "***"),
            R("3",  "OVERALL DIMENSION DRAWING",    "",  "***"),
        ]

    # ──────────────────────────────────────────────────────────────────────────
    # Excel Export
    # ──────────────────────────────────────────────────────────────────────────
    def export_to_excel(self, datasheet_rows: List[Dict], project_info: Dict = None):
        """Export DG set datasheet to formatted Excel workbook."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from io import BytesIO

        wb = Workbook()
        ws = wb.active
        ws.title = "DG Set Datasheet"

        # Dark green header to distinguish from transformer (blue)
        header_fill  = PatternFill(start_color="1E5631", end_color="1E5631", fill_type="solid")
        header_font  = Font(color="FFFFFF", bold=True, size=10)
        section_fill = PatternFill(start_color="D5E8D4", end_color="D5E8D4", fill_type="solid")
        section_font = Font(bold=True, size=10)
        thin   = Side(style="thin")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        row_idx = 1

        # Title row
        ws.merge_cells("A1:F1")
        ws["A1"] = "EMERGENCY DIESEL GENERATOR (EDG) SET – DATASHEET"
        ws["A1"].font = Font(bold=True, size=13, color="1E5631")
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        row_idx = 3

        # Project info
        if project_info:
            for key, val in project_info.items():
                ws.cell(row=row_idx, column=1, value=key.replace("_", " ").title())
                ws.cell(row=row_idx, column=2, value=val)
                row_idx += 1
            row_idx += 1

        # Column headers – 6 columns
        col_headers = ["SI NO.", "DESCRIPTION", "UNIT", "SPECIFIED DESIGN DATA", "VENDOR DATA", "Rev"]
        col_widths  = [9, 52, 10, 35, 30, 8]
        for col_idx, (header, width) in enumerate(zip(col_headers, col_widths), 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.column_dimensions[chr(64 + col_idx)].width = width

        row_idx += 1

        # Data rows
        for row_data in datasheet_rows:
            sr_no       = row_data.get("sr_no", "")
            description = row_data.get("description", "")
            unit        = row_data.get("unit", "")
            req_data    = row_data.get("required_data", "")
            vendor_data = row_data.get("vendor_data", "")
            rev         = row_data.get("rev", row_data.get("remarks", ""))

            is_section = (not sr_no) and description and not req_data and not vendor_data

            cells  = [sr_no, description, unit, req_data, vendor_data, rev]
            aligns = ["center", "left", "center", "left", "left", "center"]

            for col_idx, (val, align) in enumerate(zip(cells, aligns), 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.border = border
                cell.alignment = Alignment(horizontal=align, vertical="top", wrap_text=True)
                if is_section:
                    cell.fill = section_fill
                    cell.font = section_font

            row_idx += 1

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf
