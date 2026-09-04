"""
Power/Distribution Transformer Datasheet Generator

Generates a transformer datasheet aligned 1:1 with the ADNOC / Borouge
template `DS-13-574-EP-00001.xlsm` (sheets `1.25MVA` and `25MVA`).

The full structural template (sections A–U, units, default specified values
and Rev codes) is **soft-coded** in
`transformer_datasheet_schema.py` — this module orchestrates extraction and
Excel rendering only.

Columns: Sl. No. | DESCRIPTION | UNIT | SPECIFIED DESIGN DATA | VENDOR DATA | Rev
"""
from __future__ import annotations

import json
import logging
from io import BytesIO
from typing import Dict, List, Optional

import PyPDF2
from django.conf import settings
from openai import OpenAI

from .transformer_datasheet_schema import (
    DOC_HEADER,
    TABLE_HEADERS,
    TABLE_COL_WIDTHS,
    VARIANT_POWER,
    VARIANT_DISTRIBUTION,
    VARIANT_DEFAULTS,
    SHEET_TITLES,
    DEFAULT_PAGINATION,
    REVISION_HISTORY,
    REVISION_FOOTER_NOTES,
    HOLD_ENTRIES,
    INDEX_ENTRIES,
    ABBREVIATIONS,
    GENERAL_NOTES,
    build_schema,
    detect_variant_from_text,
)

logger = logging.getLogger(__name__)

# ─────────────────────────────────────────────────────────────────────────────
# Soft-coded constants
# ─────────────────────────────────────────────────────────────────────────────
AI_MODEL              = "gpt-4o"
AI_TEMPERATURE        = 0.1
AI_MAX_TOKENS         = 8000
AI_DOC_TEXT_LIMIT     = 8000      # chars of document content sent to model
MIN_DOC_TEXT_LEN      = 20         # minimum extracted PDF text length

# Excel formatting constants
EXCEL_TITLE_FILL      = "1F4E79"
EXCEL_TITLE_FONT      = "FFFFFF"
EXCEL_SECTION_FILL    = "D6E4F0"
EXCEL_HEADER_FILL     = "1F4E79"


class TransformerDatasheetGenerator:
    """Generate Power/Distribution Transformer datasheets from sizing calculations."""

    def __init__(self):
        self.client = OpenAI(api_key=settings.OPENAI_API_KEY)

    # ──────────────────────────────────────────────────────────────────────
    # Document Extraction (multi-format — PDF / Excel / Word / image / …)
    # ──────────────────────────────────────────────────────────────────────
    def extract_text_from_pdf(self, pdf_file) -> str:
        """Extract text from any supported document type (kept name for back-compat)."""
        from .document_extractor import extract_text
        text = extract_text(pdf_file)
        logger.info(f"[TransformerDatasheet] Extracted {len(text)} chars from {getattr(pdf_file, 'name', '?')}")
        return text

    # ──────────────────────────────────────────────────────────────────────
    # Main entry point
    # ──────────────────────────────────────────────────────────────────────
    def generate_datasheet_from_sizing_calc(
        self, pdf_file, project_info: Optional[Dict] = None
    ) -> Dict:
        """Generate datasheet rows from a transformer sizing calculation PDF."""
        try:
            logger.info("[TransformerDatasheet] Extracting text from sizing calculation PDF…")
            doc_text = self.extract_text_from_pdf(pdf_file)

            if not doc_text or len(doc_text) < MIN_DOC_TEXT_LEN:
                logger.error(
                    f"[TransformerDatasheet] Insufficient text: "
                    f"{len(doc_text) if doc_text else 0} chars"
                )
                return {
                    "success": False,
                    "error": (
                        "Could not extract text from the PDF. The file may be image-based "
                        "or empty. Please provide a text-based transformer sizing "
                        "calculation document."
                    ),
                }

            # Detect variant (power vs distribution) from document content
            variant = detect_variant_from_text(doc_text)
            logger.info(f"[TransformerDatasheet] Detected variant: {variant}")

            # Build the structural template for the detected variant
            template_rows = build_schema(variant)

            # Use AI to fill VENDOR DATA values from the document content
            logger.info("[TransformerDatasheet] Populating vendor data via AI…")
            populated_rows = self._populate_vendor_data_with_ai(
                template_rows, doc_text, variant, project_info
            )

            summary = {
                "variant":          variant,
                "total_rows":       len(populated_rows),
                "section_rows":     sum(1 for r in populated_rows if r.get("is_section")),
                "data_rows":        sum(1 for r in populated_rows if not r.get("is_section")),
                "completed_fields": sum(
                    1 for r in populated_rows
                    if not r.get("is_section") and (r.get("vendor_data") or "").strip()
                ),
                "missing_fields":   sum(
                    1 for r in populated_rows
                    if not r.get("is_section") and not (r.get("vendor_data") or "").strip()
                ),
            }

            logger.info(
                f"[TransformerDatasheet] ✅ Generated {summary['total_rows']} rows "
                f"({summary['completed_fields']} vendor fields populated)"
            )

            return {
                "success": True,
                "datasheet_rows": populated_rows,
                "summary": summary,
                "extraction_metadata": {
                    "document_length": len(doc_text),
                    "variant":         variant,
                    "project_info":    project_info or {},
                },
            }

        except Exception as e:
            logger.error(f"[TransformerDatasheet] Error: {e}", exc_info=True)
            return {"success": False, "error": str(e)}

    # ──────────────────────────────────────────────────────────────────────
    # AI vendor-data population
    # ──────────────────────────────────────────────────────────────────────
    def _populate_vendor_data_with_ai(
        self,
        template_rows: List[Dict],
        doc_text: str,
        variant: str,
        project_info: Optional[Dict] = None,
    ) -> List[Dict]:
        """
        Send the structured template to the AI and ask it to populate ONLY the
        ``vendor_data`` field for non-section rows, by extracting matching
        values from the supplied sizing-calculation document.

        Returns the rows in the same order with vendor_data filled where found.
        """
        # Build a compact list the AI must populate (sr_no + description + unit)
        ai_targets = [
            {"sr_no": r["sr_no"], "description": r["description"], "unit": r["unit"]}
            for r in template_rows
            if not r.get("is_section")
        ]

        variant_label = (
            "POWER TRANSFORMER (e.g. 25 MVA, 33/11.5 kV)"
            if variant == VARIANT_POWER
            else "DISTRIBUTION TRANSFORMER (e.g. 1250 kVA, 11/0.433 kV)"
        )

        prompt = f"""You are a senior electrical engineer specialising in power and distribution
transformers per IEC 60076 and ADNOC / Borouge specifications.

You will receive:
  1. PROJECT INFORMATION
  2. TRANSFORMER VARIANT detected from the document
  3. The TEXT CONTENT of a Transformer Sizing Calculation document
  4. A LIST OF DATASHEET LINE-ITEMS (sr_no + description + unit)

TASK
Return a JSON ARRAY where each element corresponds — IN THE SAME ORDER — to
each item in the LIST OF DATASHEET LINE-ITEMS, with one key:
  • "vendor_data" : the value extracted FROM THE DOCUMENT for that parameter,
                    or "" (empty string) if not explicitly present.

RULES
- Do NOT invent values. If the document does not state a value, return "".
- Numeric values: include the value only (no unit, since unit is a separate column).
- For YES/NO style entries, return "YES" / "NO" / "NA" / "***" as appropriate.
- For Tag No., extract the actual transformer tag(s) found in the document.
- Preserve original casing from the document where reasonable.
- Output ONLY the JSON array — no markdown fences, no commentary.
- The array length MUST equal the number of LINE-ITEMS supplied.

PROJECT INFORMATION:
{json.dumps(project_info or {}, indent=2)}

TRANSFORMER VARIANT:
{variant_label}

DOCUMENT CONTENT (truncated):
{doc_text[:AI_DOC_TEXT_LIMIT]}

DATASHEET LINE-ITEMS (extract vendor_data for each, in order):
{json.dumps(ai_targets, ensure_ascii=False)}
"""

        try:
            response = self.client.chat.completions.create(
                model=AI_MODEL,
                messages=[
                    {
                        "role": "system",
                        "content": (
                            "You extract structured engineering data from technical "
                            "documents. Return only a valid JSON array."
                        ),
                    },
                    {"role": "user", "content": prompt},
                ],
                temperature=AI_TEMPERATURE,
                max_tokens=AI_MAX_TOKENS,
            )

            ai_response = (response.choices[0].message.content or "").strip()

            # Strip markdown fences if present
            if "```json" in ai_response:
                ai_response = ai_response.split("```json", 1)[1].split("```", 1)[0]
            elif "```" in ai_response:
                ai_response = ai_response.split("```", 1)[1].split("```", 1)[0]

            extracted = json.loads(ai_response.strip())
            if not isinstance(extracted, list):
                logger.warning("[TransformerDatasheet] AI response is not a list — keeping defaults")
                return template_rows

            # Merge vendor_data back into the template (preserve order)
            data_iter = iter(extracted)
            out: List[Dict] = []
            for row in template_rows:
                if row.get("is_section"):
                    out.append(dict(row))
                    continue
                ai_row = next(data_iter, {}) or {}
                vendor = (ai_row.get("vendor_data") or "").strip() if isinstance(ai_row, dict) else ""
                merged = dict(row)
                if vendor:
                    merged["vendor_data"] = vendor
                out.append(merged)

            return out

        except json.JSONDecodeError as e:
            logger.error(f"[TransformerDatasheet] JSON decode error: {e}")
            return template_rows
        except Exception as e:
            logger.error(f"[TransformerDatasheet] AI extraction error: {e}", exc_info=True)
            return template_rows

    # ──────────────────────────────────────────────────────────────────────
    # Fallback default template
    # ──────────────────────────────────────────────────────────────────────
    def _get_default_datasheet_template(self) -> List[Dict]:
        """Return the full structured template for the power-transformer variant."""
        return build_schema(VARIANT_POWER)

    # ──────────────────────────────────────────────────────────────────────
    # Excel Export — full multi-sheet ADNOC document (Cover / Revision /
    # Hold / Index / Datasheet / Notes) — soft-coded from the schema.
    # ──────────────────────────────────────────────────────────────────────
    def export_to_excel(
        self, datasheet_rows: List[Dict], project_info: Optional[Dict] = None
    ) -> BytesIO:
        """Render the full ADNOC datasheet workbook (6 sheets) to BytesIO."""
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

        info = project_info or {}
        pagination = {**DEFAULT_PAGINATION, **(info.get("pagination") or {})}
        rev_letter = info.get("revision", "P")

        # Shared styles ----------------------------------------------------
        thin   = Side(style="thin")
        medium = Side(style="medium")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        thick_border = Border(left=medium, right=medium, top=medium, bottom=medium)

        title_fill   = PatternFill(start_color=EXCEL_TITLE_FILL,   end_color=EXCEL_TITLE_FILL,   fill_type="solid")
        header_fill  = PatternFill(start_color=EXCEL_HEADER_FILL,  end_color=EXCEL_HEADER_FILL,  fill_type="solid")
        section_fill = PatternFill(start_color=EXCEL_SECTION_FILL, end_color=EXCEL_SECTION_FILL, fill_type="solid")
        title_font   = Font(color=EXCEL_TITLE_FONT, bold=True, size=12)
        header_font  = Font(color="FFFFFF", bold=True, size=10)
        section_font = Font(bold=True, size=10)
        bold9        = Font(bold=True, size=9)
        bold10       = Font(bold=True, size=10)
        bold11       = Font(bold=True, size=11)

        company_doc = info.get("company_doc_number")        or DOC_HEADER["company_doc_default"]
        contractor  = info.get("contractor_drawing_number") or DOC_HEADER["contractor_default"]
        rejlers     = info.get("rejlers_drawing_number")    or DOC_HEADER["rejlers_default"]

        # ── Helper: render the 7-row ADNOC document header block (A1:F7)
        def _render_header(ws, sheet_pageno: str):
            for col_idx, width in enumerate(TABLE_COL_WIDTHS, start=1):
                ws.column_dimensions[chr(64 + col_idx)].width = width

            ws.merge_cells("A1:C2")
            ws["A1"] = DOC_HEADER["company_name"]
            ws["A1"].font = bold11
            ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            ws.merge_cells("D1:E1")
            ws["D1"] = DOC_HEADER["company_doc_label"]
            ws["D1"].font = bold9
            ws["D1"].alignment = Alignment(horizontal="center", vertical="center")
            ws.merge_cells("D2:E2")
            ws["D2"] = company_doc
            ws["D2"].alignment = Alignment(horizontal="center", vertical="center")

            ws["F1"] = "Rev"
            ws["F1"].font = bold9
            ws["F1"].alignment = Alignment(horizontal="center", vertical="center")
            ws["F2"] = rev_letter
            ws["F2"].alignment = Alignment(horizontal="center", vertical="center")

            ws.merge_cells("A3:C3")
            ws["A3"] = f"LOCATION:\n{DOC_HEADER['location']}"
            ws["A3"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            ws.merge_cells("D3:E3")
            ws["D3"] = DOC_HEADER["project_title"]
            ws["D3"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws["F3"] = f"Sheet\n{sheet_pageno}"
            ws["F3"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            ws.merge_cells("A4:C4")
            ws["A4"] = DOC_HEADER["document_title"]
            ws["A4"].font = bold10
            ws["A4"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            ws.merge_cells("D4:F4")
            ws["D4"] = DOC_HEADER["contractor_label"]
            ws["D4"].font = bold9
            ws["D4"].alignment = Alignment(horizontal="center", vertical="center")
            ws.merge_cells("D5:F5")
            ws["D5"] = contractor
            ws["D5"].alignment = Alignment(horizontal="center", vertical="center")

            ws.merge_cells("D6:F6")
            ws["D6"] = DOC_HEADER["rejlers_label"]
            ws["D6"].font = bold9
            ws["D6"].alignment = Alignment(horizontal="center", vertical="center")
            ws.merge_cells("D7:F7")
            ws["D7"] = rejlers
            ws["D7"].alignment = Alignment(horizontal="center", vertical="center")

            for r in range(1, 8):
                for c in range(1, 7):
                    ws.cell(row=r, column=c).border = border

        def _banner(ws, row, text):
            """Section banner spanning A:F."""
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
            cell = ws.cell(row=row, column=1, value=text)
            cell.fill = title_fill
            cell.font = title_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            for c in range(1, 7):
                ws.cell(row=row, column=c).border = border

        # ── Workbook scaffold ─────────────────────────────────────────────
        wb = Workbook()
        wb.remove(wb.active)

        # ╔════════════════════════════════════════════════════════════════╗
        # ║ 1) COVERSHEET                                                  ║
        # ╚════════════════════════════════════════════════════════════════╝
        ws = wb.create_sheet(SHEET_TITLES["cover"])
        _render_header(ws, pagination["cover"])
        ws.merge_cells("A9:F12")
        cover_title = info.get("variant_title") or VARIANT_DEFAULTS[VARIANT_POWER]["title_line"]
        ws["A9"] = f"TECHNICAL DATASHEET FOR TRANSFORMER (POWER AND DISTRIBUTION)\n\n{cover_title}"
        ws["A9"].fill = title_fill
        ws["A9"].font = Font(color=EXCEL_TITLE_FONT, bold=True, size=14)
        ws["A9"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for r in range(9, 13):
            for c in range(1, 7):
                ws.cell(row=r, column=c).border = thick_border

        # ╔════════════════════════════════════════════════════════════════╗
        # ║ 2) REVISION                                                    ║
        # ╚════════════════════════════════════════════════════════════════╝
        ws = wb.create_sheet(SHEET_TITLES["revision"])
        _render_header(ws, pagination["revision"])
        _banner(ws, 9, "REVISION HISTORY")
        rev_headers = ["Rev. No.", "Date", "Section or Page Revised", "Revision Description"]
        ws.cell(row=10, column=1, value=rev_headers[0]).font = header_font
        ws.cell(row=10, column=1).fill = header_fill
        ws.cell(row=10, column=2, value=rev_headers[1]).font = header_font
        ws.cell(row=10, column=2).fill = header_fill
        ws.cell(row=10, column=3, value=rev_headers[2]).font = header_font
        ws.cell(row=10, column=3).fill = header_fill
        ws.merge_cells("D10:F10")
        ws.cell(row=10, column=4, value=rev_headers[3]).font = header_font
        ws.cell(row=10, column=4).fill = header_fill
        for c in range(1, 7):
            ws.cell(row=10, column=c).border = border
            ws.cell(row=10, column=c).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        r = 11
        for rev, date, section, desc in REVISION_HISTORY:
            ws.cell(row=r, column=1, value=rev).alignment = Alignment(horizontal="center")
            ws.cell(row=r, column=2, value=date).alignment = Alignment(horizontal="center")
            ws.cell(row=r, column=3, value=section).alignment = Alignment(horizontal="center")
            ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=6)
            ws.cell(row=r, column=4, value=desc).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            for c in range(1, 7):
                ws.cell(row=r, column=c).border = border
            r += 1

        r += 1
        ws.cell(row=r, column=1, value="NOTES:").font = bold10
        r += 1
        for note in REVISION_FOOTER_NOTES:
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
            cell = ws.cell(row=r, column=1, value=note)
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            r += 1

        # ╔════════════════════════════════════════════════════════════════╗
        # ║ 3) HOLD                                                        ║
        # ╚════════════════════════════════════════════════════════════════╝
        ws = wb.create_sheet(SHEET_TITLES["hold"])
        _render_header(ws, pagination["hold"])
        _banner(ws, 9, "HOLDS")
        ws.cell(row=10, column=1, value="Rev. No.").font = header_font
        ws.cell(row=10, column=1).fill = header_fill
        ws.merge_cells("B10:E10")
        ws.cell(row=10, column=2, value="Hold Description").font = header_font
        ws.cell(row=10, column=2).fill = header_fill
        ws.cell(row=10, column=6, value="Section").font = header_font
        ws.cell(row=10, column=6).fill = header_fill
        for c in range(1, 7):
            ws.cell(row=10, column=c).border = border
            ws.cell(row=10, column=c).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        r = 11
        for rev, desc, section in HOLD_ENTRIES:
            ws.cell(row=r, column=1, value=rev).alignment = Alignment(horizontal="center")
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
            ws.cell(row=r, column=2, value=desc).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            ws.cell(row=r, column=6, value=section).alignment = Alignment(horizontal="center")
            for c in range(1, 7):
                ws.cell(row=r, column=c).border = border
            r += 1

        # ╔════════════════════════════════════════════════════════════════╗
        # ║ 4) INDEX                                                       ║
        # ╚════════════════════════════════════════════════════════════════╝
        ws = wb.create_sheet(SHEET_TITLES["index"])
        _render_header(ws, pagination["index"])
        _banner(ws, 9, "TABLE OF CONTENTS")
        ws.cell(row=10, column=1, value="Sr. No.").font = header_font
        ws.cell(row=10, column=1).fill = header_fill
        ws.merge_cells("B10:E10")
        ws.cell(row=10, column=2, value="DESCRIPTION").font = header_font
        ws.cell(row=10, column=2).fill = header_fill
        ws.cell(row=10, column=6, value="SHEET").font = header_font
        ws.cell(row=10, column=6).fill = header_fill
        for c in range(1, 7):
            ws.cell(row=10, column=c).border = border
            ws.cell(row=10, column=c).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        r = 11
        for sr, desc, sheet in INDEX_ENTRIES:
            ws.cell(row=r, column=1, value=sr).alignment = Alignment(horizontal="center")
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
            ws.cell(row=r, column=2, value=desc).alignment = Alignment(horizontal="left", vertical="center")
            ws.cell(row=r, column=6, value=sheet).alignment = Alignment(horizontal="center")
            for c in range(1, 7):
                ws.cell(row=r, column=c).border = border
            r += 1

        r += 1
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        ws.cell(row=r, column=1, value="ABBREVIATIONS:").font = bold10
        r += 1
        for abbr, meaning in ABBREVIATIONS:
            ws.cell(row=r, column=1, value=abbr).font = bold9
            ws.cell(row=r, column=1).alignment = Alignment(horizontal="center")
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
            ws.cell(row=r, column=2, value=meaning).alignment = Alignment(horizontal="left", vertical="center")
            for c in range(1, 7):
                ws.cell(row=r, column=c).border = border
            r += 1

        # ╔════════════════════════════════════════════════════════════════╗
        # ║ 5) DATASHEET (variant body)                                    ║
        # ╚════════════════════════════════════════════════════════════════╝
        ws_title = info.get("data_sheet_name") or info.get("variant_title") or SHEET_TITLES["data"]
        # Excel sheet-name rules: max 31 chars, no  : \ / ? * [ ]
        import re
        ws_title = re.sub(r"[:\\/\?\*\[\]]", "-", ws_title)[:31] or SHEET_TITLES["data"]
        ws = wb.create_sheet(ws_title)
        _render_header(ws, pagination["data"])

        # Variant title (row 9)
        variant_title = info.get("variant_title", "")
        if not variant_title:
            for row in datasheet_rows:
                if (row.get("description") or "").strip().upper() == "RATING":
                    rating = (row.get("vendor_data") or row.get("required_data") or "").strip()
                    unit   = (row.get("unit") or "").strip()
                    if rating:
                        variant_title = f"{rating} {unit}".strip() + " TRANSFORMER"
                    break
            if not variant_title:
                variant_title = VARIANT_DEFAULTS[VARIANT_POWER]["title_line"]
        _banner(ws, 9, variant_title)

        # Body header (row 10)
        body_header_row = 10
        for col_idx, label in enumerate(TABLE_HEADERS, start=1):
            cell = ws.cell(row=body_header_row, column=col_idx, value=label)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Body data rows
        row_idx = body_header_row + 1
        aligns  = ["center", "left", "center", "left", "left", "center"]

        for row in datasheet_rows:
            sr_no       = row.get("sr_no", "")
            description = row.get("description", "")
            unit        = row.get("unit", "")
            req_data    = row.get("required_data", "")
            vendor_data = row.get("vendor_data", "")
            rev         = row.get("rev", "")

            if "is_section" in row:
                is_section = bool(row["is_section"])
            else:
                is_section = bool(description and not unit and not req_data and not vendor_data)

            cells = [sr_no, description, unit, req_data, vendor_data, rev]
            for col_idx, (val, align) in enumerate(zip(cells, aligns), start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.border = border
                cell.alignment = Alignment(horizontal=align, vertical="top", wrap_text=True)
                if is_section:
                    cell.fill = section_fill
                    cell.font = section_font

            row_idx += 1

        ws.freeze_panes = "A11"

        # ╔════════════════════════════════════════════════════════════════╗
        # ║ 6) NOTES                                                       ║
        # ╚════════════════════════════════════════════════════════════════╝
        ws = wb.create_sheet(SHEET_TITLES["notes"])
        _render_header(ws, pagination["notes"])
        _banner(ws, 9, "GENERAL NOTES")
        ws.cell(row=10, column=1, value="SI. NO").font = header_font
        ws.cell(row=10, column=1).fill = header_fill
        ws.merge_cells("B10:F10")
        ws.cell(row=10, column=2, value="Description").font = header_font
        ws.cell(row=10, column=2).fill = header_fill
        for c in range(1, 7):
            ws.cell(row=10, column=c).border = border
            ws.cell(row=10, column=c).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        r = 11
        for label, text in GENERAL_NOTES:
            ws.cell(row=r, column=1, value=label).alignment = Alignment(horizontal="center", vertical="top")
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
            ws.cell(row=r, column=2, value=text).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            for c in range(1, 7):
                ws.cell(row=r, column=c).border = border
            r += 1

        # ── Save ─────────────────────────────────────────────────────────
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf
