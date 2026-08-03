"""
LV Switchgear Datasheet Generator from Technical Datasheet Documents
Extracts equipment data and generates comprehensive datasheets matching
the ADNOC Technical Datasheet for LV Switchgear form.

Columns: SR NO | DESCRIPTION | UNIT | REQUIRED DATA | VENDOR DATA | Rev
"""
import logging
import json
from typing import Dict, List
from django.conf import settings
from openai import OpenAI
import PyPDF2

logger = logging.getLogger(__name__)


class LVSwitchgearDatasheetGenerator:
    """Generate LV Switchgear datasheets from Technical Datasheet documents."""

    def __init__(self):
        self.client = OpenAI(api_key=settings.OPENAI_API_KEY)

    # ──────────────────────────────────────────────────────────────────────────
    # PDF Extraction
    # ──────────────────────────────────────────────────────────────────────────
    def extract_text_from_pdf(self, pdf_file) -> str:
        """Extract all text from an uploaded PDF file."""
        try:
            pdf_file.seek(0)
            reader = PyPDF2.PdfReader(pdf_file)
            text = ""
            logger.info(f"[LVSwitchgearDatasheet] PDF has {len(reader.pages)} pages")
            for i, page in enumerate(reader.pages, 1):
                page_text = page.extract_text()
                if page_text:
                    text += page_text + "\n"
                    logger.info(f"[LVSwitchgearDatasheet] Page {i}: {len(page_text)} chars")
                else:
                    logger.warning(f"[LVSwitchgearDatasheet] Page {i}: no text (image-based?)")
            logger.info(f"[LVSwitchgearDatasheet] Total: {len(text)} chars")
            return text
        except Exception as e:
            logger.error(f"[LVSwitchgearDatasheet] PDF extraction error: {e}", exc_info=True)
            return ""

    # ──────────────────────────────────────────────────────────────────────────
    # Main entry point
    # ──────────────────────────────────────────────────────────────────────────
    def generate_datasheet_from_document(self, pdf_file, project_info: Dict = None) -> Dict:
        """
        Generate a LV Switchgear datasheet from the uploaded Technical Datasheet PDF.

        Returns:
            {
                'success': bool,
                'datasheet_rows': List[Dict],
                'summary': Dict,
                'extraction_metadata': Dict
            }
        """
        try:
            # ── Step 1: Always start with full template ───────────────────────
            datasheet_rows = self._get_default_datasheet_template()

            # ── Step 2: Extract text from PDF ─────────────────────────────────
            logger.info("[LVSwitchgearDatasheet] Extracting text from PDF...")
            doc_text = self.extract_text_from_pdf(pdf_file)

            # ── Step 3: If text available, use AI to extract vendor values ─────
            if doc_text and len(doc_text) >= 20:
                logger.info(f"[LVSwitchgearDatasheet] {len(doc_text)} chars extracted — running AI vendor extraction...")
                vendor_map = self._extract_vendor_data_with_ai(doc_text, project_info)
                if vendor_map:
                    merged = self._merge_vendor_data(datasheet_rows, vendor_map)
                    logger.info(f"[LVSwitchgearDatasheet] Merged {merged} vendor values into template")
            else:
                logger.warning("[LVSwitchgearDatasheet] No/insufficient text — showing template with empty vendor data")
                doc_text = ""

            summary = {
                'total_rows': len(datasheet_rows),
                'equipment_count': sum(1 for r in datasheet_rows if r.get('description', '').strip()),
                'completed_fields': sum(1 for r in datasheet_rows if r.get('vendor_data', '').strip()),
                'missing_fields': sum(1 for r in datasheet_rows if not r.get('vendor_data', '').strip()),
            }

            logger.info(f"[LVSwitchgearDatasheet] ✅ {summary['total_rows']} rows | {summary['completed_fields']} vendor values filled")
            return {
                'success': True,
                'datasheet_rows': datasheet_rows,
                'summary': summary,
                'extraction_metadata': {
                    'document_length': len(doc_text),
                    'project_info': project_info or {}
                }
            }

        except Exception as e:
            logger.error(f"[LVSwitchgearDatasheet] Error: {e}", exc_info=True)
            return {'success': False, 'error': str(e)}

    # ──────────────────────────────────────────────────────────────────────────
    # AI vendor data extraction + merge helpers
    # ──────────────────────────────────────────────────────────────────────────
    def _extract_vendor_data_with_ai(self, doc_text: str, project_info: Dict = None) -> Dict:
        """
        Use GPT-4o to extract parameter values from the uploaded document.
        Returns a dict mapping NORMALIZED_DESCRIPTION_UPPER → extracted_value.
        """
        prompt = f"""You are a senior electrical engineer specialising in LV (Low Voltage) Switchgear and Motor Control Centres (MCC). Read the following technical document and extract every parameter value you can find.

DOCUMENT:
{doc_text[:8000]}

TASK:
Return a JSON object where:
- Keys   = parameter/field names in UPPERCASE (e.g. "RATED VOLTAGE", "MANUFACTURER", "RATED CURRENT")
- Values = the extracted value as a string

Focus on: equipment tags, voltage ratings (V/kV), current ratings (A/kA), short circuit ratings, frequency (Hz), IP rating, dimensions, weights, temperatures, manufacturer names, circuit breaker type, CT/VT ratios, protection relay type, busbar details, contactor ratings.

Rules:
- Include ONLY fields actually present in the document.
- Do NOT invent or assume values.
- Return ONLY a valid JSON object — no explanation, no markdown.

Example:
{{"RATED VOLTAGE": "0.415", "RATED CURRENT": "1600", "MANUFACTURER": "SCHNEIDER ELECTRIC", "FREQUENCY": "50", "SHORT CIRCUIT CURRENT": "50", "IP RATING": "IP41"}}"""

        try:
            response = self.client.chat.completions.create(
                model="gpt-4o",
                messages=[
                    {"role": "system", "content": "You are an expert electrical engineer. Extract parameter values from LV switchgear/MCC technical documents. Return only valid JSON."},
                    {"role": "user", "content": prompt},
                ],
                temperature=0.1,
                max_tokens=2000,
            )
            ai_response = response.choices[0].message.content.strip()
            if "```json" in ai_response:
                ai_response = ai_response.split("```json")[1].split("```")[0]
            elif "```" in ai_response:
                ai_response = ai_response.split("```")[1].split("```")[0]
            vendor_map = json.loads(ai_response.strip())
            if isinstance(vendor_map, dict):
                normalized = {k.strip().upper(): str(v).strip() for k, v in vendor_map.items() if v and str(v).strip()}
                logger.info(f"[LVSwitchgearDatasheet] AI extracted {len(normalized)} vendor key-value pairs")
                return normalized
            return {}
        except json.JSONDecodeError as e:
            logger.error(f"[LVSwitchgearDatasheet] Vendor JSON decode error: {e}")
            return {}
        except Exception as e:
            logger.error(f"[LVSwitchgearDatasheet] AI extraction error: {e}")
            return {}

    def _merge_vendor_data(self, template_rows: List[Dict], vendor_map: Dict) -> int:
        """
        Merge vendor_map values into template_rows by description matching.
        Returns count of rows filled.
        """
        merged = 0
        stop_words = {'AND', 'OR', 'OF', 'THE', 'WITH', 'FOR', 'AT', 'IN', 'TO', 'A', 'AN', 'BY', 'ON'}
        for row in template_rows:
            if row.get('vendor_data'):
                continue
            desc = row.get('description', '').strip().upper()
            if not desc:
                continue
            # 1. Exact match
            if desc in vendor_map:
                row['vendor_data'] = vendor_map[desc]
                merged += 1
                continue
            # 2. Containment match
            matched = False
            for ai_key, ai_val in vendor_map.items():
                if not ai_key or not ai_val:
                    continue
                if ai_key in desc or desc in ai_key:
                    row['vendor_data'] = ai_val
                    merged += 1
                    matched = True
                    break
            if matched:
                continue
            # 3. Word-overlap match (≥2 meaningful common words)
            desc_words = set(desc.split()) - stop_words
            for ai_key, ai_val in vendor_map.items():
                if not ai_key or not ai_val:
                    continue
                ai_words = set(ai_key.upper().split()) - stop_words
                if len(desc_words & ai_words) >= 2:
                    row['vendor_data'] = ai_val
                    merged += 1
                    break
        return merged

    # ──────────────────────────────────────────────────────────────────────────
    # Full ADNOC LV Switchgear Template
    # ──────────────────────────────────────────────────────────────────────────
    def _get_default_datasheet_template(self) -> List[Dict]:
        """Return full ADNOC Technical Datasheet for LV Switchgear (6-column)."""
        R = lambda sr, desc, unit="", req="": {
            "sr_no": sr, "description": desc, "unit": unit,
            "required_data": req, "vendor_data": "", "rev": ""
        }
        H = lambda desc: {
            "sr_no": "", "description": desc, "unit": "",
            "required_data": "", "vendor_data": "", "rev": ""
        }
        return [
            # ── A. GENERAL ───────────────────────────────────────────────────
            H("GENERAL"),
            R("1",    "EQUIPMENT TAG NO."),
            R("2",    "TITLE",                                          "",        "AT5A, LV SWITCHBOARD"),
            R("3",    "SUBSTATION"),
            R("4",    "MAKE"),
            R("5",    "YEAR OF MANUFACTURE"),

            # ── B. REFERENCE SPECIFICATION ──────────────────────────────────
            H("REFERENCE SPECIFICATION"),
            R("1.1",  "ELECTRICAL DESIGN GUIDELINES",                  "",        "BGS-EU-001"),
            R("1.2",  "CENTRAL PADS SWITCHGEAR AND CONTROL GEAR",      "",        "BGS-EE-004"),
            R("1.3",  "APPLICABLE INTERNATIONAL STANDARDS",            "",        "IEC 61439-2"),
            R("1.4",  "INTEGRATED PROTECTION AND CONTROL SYSTEM (IPCS)","",       "BGS-EE-010"),
            R("1.5",  "POWER, CONTROL AND SWITCHING CABLES",           "",        "BGS-EE-011"),
            R("1.6",  "LV BUS TRAINERS",                               "",        "BGS-EE-012"),
            R("1.7",  "FIELD COMMISSIONING OF ELECTRICAL INSTALLATION AND EQUIPMENT", "", "BGS-EE-020"),

            # ── C. REFERENCE DRAWING ─────────────────────────────────────────
            H("REFERENCE DRAWING"),
            R("2.1",  "KEY SINGLE LINE DIAGRAM",                       "",        "15-EE-133-00001"),
            R("2.2",  "ELECTRICAL LOAD LIST",                          "",        "15-EE-359-00001"),
            R("2.3",  "ELECTRICAL DATASHEET",                          "",        "15-EE-409-00001"),

            # ── D. GENERAL CHARACTERISTICS ──────────────────────────────────
            H("D   GENERAL CHARACTERISTICS"),
            R("1",    "MANUFACTURER TYPE DESIGNATION"),
            R("2",    "APPLICABLE INTERNATIONAL STANDARDS",            "",        "IEC 61439-2, BGS-EE-004"),
            R("3",    "FREQUENCY",                                     "Hz",      "50"),
            R("3.1",  "STEADY STATE VARIATION",                        "",        "± 1%"),
            R("3.2",  "TRANSIENT VARIATION",                           "",        "± 8%"),
            R("4",    "OPERATING VOLTAGE",                             "V",       "415"),
            R("4.1",  "STEADY STATE VARIATION",                        "",        "± 16%"),
            R("4.2",  "TRANSIENT VARIATION",                           "",        "+16%, -20%"),
            R("5",    "RATED VOLTAGE",                                 "kV",      "0.433"),
            R("6",    "RATED REGULATION VOLTAGE",                      "V"),
            R("7",    "MAXIMUM AMBIENT TEMPERATURE",                   "DEG C",   "40"),
            R("8",    "NUMBER OF PHASES",                              "",        "3P + N"),
            R("9",    "MAXIMUM MOMENTARY VOLTAGE",                     "kA"),
            R("10",   "MIN MCC TRIP",                                  "V DC",    "110"),
            R("11",   "MIN / CIRCUIT",                                 "A"),
            R("12",   "RATED SHORT TIME 1 SEC",                        "kA",      "50"),
            R("13",   "NUMBER OF BRANCHES"),

            # ── E. CONSTRUCTION CHARACTERISTICS ─────────────────────────────
            H("E   CONSTRUCTION CHARACTERISTICS"),
            R("1",    "TYPE OF SWITCHGEAR",                            "",        "METAL ENCLOSED MCC"),
            R("2",    "IP RATING",                                     "",        "IP 41"),
            R("3",    "EQUIPMENT ASSEMBLY",                            "",        "FIXED"),
            R("4",    "DEGREE OF PROTECTION OF ENCLOSURE",             "",        "IP 41"),
            R("5",    "CABLE ENTRY",                                   "",        "BOTTOM"),
            R("6",    "TYPE OF AIR CANOPY"),
            R("7",    "TYPE OF BUS BARS",                              "",        "FULLY INSULATED"),

            # ── F. DIMENSIONS OF SWITCHBOARD ────────────────────────────────
            H("F   DIMENSIONS OF SWITCHBOARD"),
            R("1",    "MAXIMUM HEIGHT",                                "mm"),
            R("2",    "MAXIMUM DEPTH",                                 "mm"),
            R("3",    "WIDTH PER PANEL (STANDING SPACE REQUIRED)",     "mm"),
            R("4",    "HEIGHT OF SWITCHBOARD",                         "mm"),

            # ── G. DIMENSIONS OF SHIPPING SECTION ───────────────────────────
            H("G   DIMENSIONS OF SHIPPING SECTION"),
            R("1",    "MAXIMUM LENGTH",                                "mm"),
            R("2",    "MAXIMUM HEIGHT",                                "mm"),
            R("3",    "MAXIMUM OVERALL HEIGHT",                        "mm"),

            # ── H. DC CONSUMPTION ────────────────────────────────────────────
            H("H   DC CONSUMPTION"),
            R("",     "CONTINUOUS",                                    "WATTS"),
            R("",     "PEAK OR CLOSING COIL / TRIP COIL / SPRING CHARGING", "WATTS"),

            # ── I. LOSSES AND TEMPERATURE ────────────────────────────────────
            H("I   LOSSES AND TEMPERATURE"),
            R("1",    "TOTAL HEAT GENERATED BY ASSEMBLY",              "kW",      "8.8"),
            R("2",    "MAXIMUM TEMPERATURE INSIDE SECTION",            "DEG C",   "50"),

            # ── J. ANTI-CONDENSATION HEATER & CUBICLE LIGHTING SUPPLY ────────
            H("J   ANTI-CONDENSATION HEATER & CUBICLE LIGHTING SUPPLY"),
            R("",     "SUPPLY",                                        "",        "240V, 1 PHASE, 50Hz AC (REFER NOTE-7)"),

            # ── K. LV SWITCHGEAR CONDUCTOR ───────────────────────────────────
            H("K   LV SWITCHGEAR CONDUCTOR"),
            R("",     "MANUFACTURER"),
            R("",     "TYPE",                                          "",        "AIR INSULATED BUS BAR"),
            R("",     "RATED CURRENT",                                 "A"),
            R("",     "SHORT CIRCUIT WITHSTAND CURRENT",               "kA"),
            R("",     "BUS BAR MATERIAL",                              "",        "COPPER"),
            R("",     "BUS BAR CROSS SECTION",                         "SQ MM"),
            R("",     "BUS BAR FINISHING",                             "",        "FULLY INSULATED"),

            # ── L. CONTACTOR ─────────────────────────────────────────────────
            H("L   CONTACTOR"),
            R("",     "MANUFACTURER"),
            R("",     "TYPE"),
            R("",     "RATED CURRENT",                                 "A"),
            R("",     "MAKING CAPACITY",                               "kA"),
            R("",     "BREAKING CAPACITY",                             "kA"),
            R("",     "AUXILIARY CONTACTS (NO & NC)",                  "",        "MINIMUM 2 NO & 2 NC"),
            R("",     "COIL VOLTAGE",                                  "V",       "110V DC"),

            # ── M. MOTOR STARTERS ────────────────────────────────────────────
            H("M   MOTOR STARTERS"),
            R("",     "TYPE",                                          "",        "DOL / STAR-DELTA / VFD"),
            R("",     "RATED MOTOR CURRENT",                           "A"),
            R("",     "OVERLOAD RELAY TYPE",                           "",        "ELECTRONIC"),
            R("",     "STARTING CURRENT LIMITATION"),

            # ── N. MOTOR ─────────────────────────────────────────────────────
            H("N   MOTOR"),
            R("",     "RATED POWER",                                   "kW"),
            R("",     "RATED VOLTAGE",                                 "V",       "415"),
            R("",     "RATED CURRENT",                                 "A"),
            R("",     "POWER FACTOR",                                  ""),
            R("",     "EFFICIENCY",                                    "%"),
            R("",     "STARTING METHOD"),

            # ── O. CIRCUIT BREAKERS ──────────────────────────────────────────
            H("O   CIRCUIT BREAKERS"),
            R("",     "MANUFACTURER"),
            R("",     "TYPE (ACB / MCCB / MCB)",                       "",        "ACB / MCCB"),
            R("",     "APPLICATION"),
            R("",     "CIRCUIT APPLICATION"),
            R("",     "BREAKING CAPACITY AT MCC CUBICLE",              "kA"),
            R("",     "CIRCUIT BREAKING CAPACITY PER PHASE",           "kA"),
            R("",     "RATED BREAKING CURRENT",                        "kA/Phase"),
            R("",     "RATED CURRENT",                                 "A"),
            R("14.1", "AUXILIARY SUPPLY VOLTAGE (TRIP & CLOSING COIL)", "V",      "110V DC"),
            R("14.2", "ANTI-CONDENSATION HEATER",                      "",        "240V, 1 PHASE, 50Hz AC"),
            R("",     "NUMBER OF AUXILIARY CONTACTS (NO & NC)",        "",        "MINIMUM 2 NO & 2 NC CONTACTS REQUIRED"),

            # ── P. CURRENT TRANSFORMER ───────────────────────────────────────
            H("P   CURRENT TRANSFORMER"),
            R("1",    "MANUFACTURER"),
            R("2",    "TYPE / CORE"),
            R("3",    "PROTECTION CT ACCURACY",                        "",        "5P20"),
            R("4",    "METERING CT ACCURACY",                          "",        "0.2S"),
            R("5",    "BURDEN",                                        "VA"),
            R("6",    "RATIO",                                         "",        "1000/1"),
            R("7",    "FUNCTIONAL (MCB) LOAD AND CURRENT RATIO"),

            # ── Q. VOLTAGE TRANSFORMER ───────────────────────────────────────
            H("Q   VOLTAGE TRANSFORMER"),
            R("1",    "MANUFACTURER"),
            R("2",    "TYPE"),
            R("3",    "BURDEN / ACCURACY"),
            R("4",    "RATIO",                                         "",        "415V / 110V"),

            # ── PROTECTION RELAYS FOR INCOMER & BUS TIE ──────────────────────
            H("PROTECTION RELAYS FOR INCOMER & BUS TIE"),
            R("1",    "MANUFACTURER",                                  "",        "SCHNEIDER ELECTRIC"),
            R("2",    "CONSTRUCTION",                                  "",        "MICROPROCESSOR BASED MULTIFUNCTION TYPE WITH INTEGRAL DISPLAY"),
            R("3",    "TYPE",                                          "",        "EASERGY RELAY"),
            R("4",    "PROTECTION FUNCTIONS"),
            R("5",    "METERING FUNCTIONS"),

            # ── R. UPSTREAM TRANSFORMER / MULTIPLE CURRENT TRANSFORMER ────────
            H("R   UPSTREAM TRANSFORMER / MULTIPLE CURRENT TRANSFORMER"),
            R("1",    "LV CABLE"),
            R("2",    "BURDEN",                                        "VA"),
            R("3",    "METERING ACCURACY",                             "",        "0.2S"),
            R("4",    "LABELS"),
            R("5",    "FUNCTIONAL (MCB) LOAD AND CURRENT RATIO"),
            R("6",    "ACCESSORIES"),

            # ── SWITCHGEAR CONFIGURATION ─────────────────────────────────────
            H("SWITCHGEAR CONFIGURATION"),
            R("",     "SWITCHGEAR TYPE",                               "",        "MCC (MOTOR CONTROL CENTRE)"),
            R("",     "STARTING TREATMENT / PROCEDURE",                "",        "DOL / STAR-DELTA"),

            # ── AUTOMATIC TRANSFER (ATS) ──────────────────────────────────────
            H("AUTOMATIC TRANSFER (ATS)"),
            R("",     "AUTOMATIC TRANSFER",                            "",        "ATS"),
            R("",     "PARKING CYCLE",                                 "",        "AS PER MANUFACTURER STANDARD"),
            R("",     "COMMUNICATION DETAILS (FOR LINK TO BUS-TIE)",   "",        "AS PER BGS-EE-010"),
            R("",     "PORT / EXTENSION"),

            # ── PAINTING ─────────────────────────────────────────────────────
            H("PAINTING"),
            R("1",    "COLOR OF EXTERNAL PAINTING",                    "",        "RAL 7035 LIGHT GREY"),
            R("2",    "COLOR OF INTERNAL PAINTING",                    "",        "RAL 7035 LIGHT GREY"),
            R("3",    "TROPICALIZATION",                               "",        "YES"),

            # ── TESTS ────────────────────────────────────────────────────────
            H("TESTS"),
            R("1",    "ROUTINE TESTS (TO BE WITNESSED AS PER CL. OF BGS-EE-004)", "", "YES"),
            R("2",    "TYPE TESTS AS PER IEC 61439-2",                 "",        "SEPARATE QUOTE REQUIRED"),
            R("3",    "TEMPERATURE RISE TEST",                         "",        "YES"),
            R("4",    "DIELECTRIC TEST",                               "",        "YES"),
            R("5",    "SHORT CIRCUIT WITHSTAND TEST",                  "",        "YES"),
            R("6",    "VERIFICATION OF IP DEGREE OF PROTECTION",       "",        "YES"),
            R("7",    "SITE ACCEPTANCE TEST (AS PER BGS-EE-020)",      "",        "YES"),
            R("8",    "TYPE TEST FOR INTERNAL ARC IN ACCORDANCE WITH IEC 61439-2", "", "YES"),
        ]

    # ──────────────────────────────────────────────────────────────────────────
    # Excel Export
    # ──────────────────────────────────────────────────────────────────────────
    def export_to_excel(self, datasheet_rows: List[Dict], project_info: Dict = None):
        """Export datasheet to Excel with formatting."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from io import BytesIO

        wb = Workbook()
        ws = wb.active
        ws.title = "LV Switchgear Datasheet"

        header_fill   = PatternFill(start_color="1F5C99", end_color="1F5C99", fill_type="solid")
        header_font   = Font(color="FFFFFF", bold=True, size=11)
        section_fill  = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
        section_font  = Font(bold=True, size=10)
        thin          = Side(style='thin')
        border        = Border(left=thin, right=thin, top=thin, bottom=thin)

        if project_info:
            ws.merge_cells('A1:F1')
            ws['A1'] = "LV SWITCHGEAR TECHNICAL DATASHEET"
            ws['A1'].font = Font(bold=True, size=14)
            ws['A1'].alignment = Alignment(horizontal='center')
            row_idx = 3
            for key, value in project_info.items():
                ws[f'A{row_idx}'] = key.replace('_', ' ').title()
                ws[f'B{row_idx}'] = value
                row_idx += 1
            row_idx += 1
        else:
            row_idx = 1

        headers = ['SR NO', 'DESCRIPTION', 'UNIT', 'REQUIRED DATA', 'VENDOR DATA', 'Rev']
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=header)
            cell.fill   = header_fill
            cell.font   = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 55
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 35
        ws.column_dimensions['E'].width = 30
        ws.column_dimensions['F'].width = 8
        row_idx += 1

        for row_data in datasheet_rows:
            sr_no        = row_data.get('sr_no', '')
            description  = row_data.get('description', '')
            unit         = row_data.get('unit', '')
            required_data = row_data.get('required_data', '')
            vendor_data  = row_data.get('vendor_data', '')
            rev          = row_data.get('rev', '')

            is_section = (not sr_no) and description and not required_data and not vendor_data

            ws.cell(row=row_idx, column=1, value=sr_no).border = border
            cell_desc = ws.cell(row=row_idx, column=2, value=description)
            cell_desc.border    = border
            cell_desc.alignment = Alignment(wrap_text=True, vertical='top')

            cell_unit = ws.cell(row=row_idx, column=3, value=unit)
            cell_unit.border    = border
            cell_unit.alignment = Alignment(horizontal='center', vertical='top')

            cell_req = ws.cell(row=row_idx, column=4, value=required_data)
            cell_req.border    = border
            cell_req.alignment = Alignment(wrap_text=True, vertical='top')

            cell_vend = ws.cell(row=row_idx, column=5, value=vendor_data)
            cell_vend.border    = border
            cell_vend.alignment = Alignment(wrap_text=True, vertical='top')

            ws.cell(row=row_idx, column=6, value=rev).border = border

            if is_section:
                for col in range(1, 7):
                    cell = ws.cell(row=row_idx, column=col)
                    cell.fill = section_fill
                    cell.font = section_font

            row_idx += 1

        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        return excel_buffer
