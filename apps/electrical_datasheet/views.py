from rest_framework import viewsets, status, filters
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django_filters.rest_framework import DjangoFilterBackend
from django.utils import timezone
from django.db.models import Q, Count
import json
import os
import openai
from django.conf import settings
import logging

logger = logging.getLogger(__name__)

# ── Soft-coded configuration constants ──────────────────────────────────────
AUTO_TAG_PREFIX_LENGTH = 8   # Number of chars from equipment_type for auto-generated tag prefix
AUTO_TAG_UUID_LENGTH = 8     # Number of hex chars from UUID for auto-generated tag suffix

from .models import (
    ElectricalEquipmentType,
    ElectricalDatasheet,
    DatasheetRevisionHistory,
    DatasheetComment
)
from .serializers import (
    ElectricalEquipmentTypeSerializer,
    ElectricalDatasheetSerializer,
    ElectricalDatasheetListSerializer,
    ElectricalDatasheetCreateUpdateSerializer,
    DatasheetRevisionHistorySerializer,
    DatasheetCommentSerializer
)
from .quality_checker import QualityCheckerMixin
from .s3_service import ElectricalDatasheetS3Service
from .adnoc_standards import ADNOCStandardsManager


def load_electrical_config():
    """Load electrical datasheet configuration from JSON file"""
    try:
        config_path = os.path.join(
            os.path.dirname(__file__),
            '..',
            'process_datasheet',
            'electrical_datasheet_config.json'
        )
        with open(config_path, 'r') as f:
            return json.load(f)
    except FileNotFoundError:
        return {"equipment_types": {}}
    except json.JSONDecodeError:
        return {"equipment_types": {}}


class ElectricalEquipmentTypeViewSet(viewsets.ModelViewSet):
    """ViewSet for managing electrical equipment types"""
    queryset = ElectricalEquipmentType.objects.filter(is_active=True)
    serializer_class = ElectricalEquipmentTypeSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['name', 'code', 'category']
    ordering_fields = ['name', 'category', 'created_at']
    ordering = ['name']

    @action(detail=False, methods=['get'])
    def config(self, request):
        """Get equipment type configurations from JSON file"""
        config_data = load_electrical_config()
        return Response(config_data)

    @action(detail=True, methods=['get'])
    def configuration(self, request, pk=None):
        """Get configuration for a specific equipment type"""
        config_data = load_electrical_config()
        equipment_config = config_data.get('equipment_types', {}).get(pk)
        
        if equipment_config:
            return Response(equipment_config)
        return Response(
            {"error": "Configuration not found for this equipment type"},
            status=status.HTTP_404_NOT_FOUND
        )
    
    @action(detail=True, methods=['get'], url_path='supported-documents')
    def supported_documents(self, request, pk=None):
        """Get supported document types for a specific equipment type (e.g., transformer)"""
        config_data = load_electrical_config()
        equipment_config = config_data.get('equipment_types', {}).get(pk)
        
        if equipment_config:
            supported_docs = equipment_config.get('supported_documents', [])
            return Response({
                'equipment_type': pk,
                'equipment_name': equipment_config.get('name'),
                'supported_documents': supported_docs
            })
        return Response(
            {"error": "Configuration not found for this equipment type"},
            status=status.HTTP_404_NOT_FOUND
        )

    @action(detail=False, methods=['post'])
    def sync_from_config(self, request):
        """Sync equipment types from configuration file to database"""
        config_data = load_electrical_config()
        equipment_types = config_data.get('equipment_types', {})
        
        synced_count = 0
        created_count = 0
        updated_count = 0
        
        for eq_id, eq_data in equipment_types.items():
            equipment_type, created = ElectricalEquipmentType.objects.update_or_create(
                id=eq_id,
                defaults={
                    'name': eq_data.get('name'),
                    'code': eq_data.get('code'),
                    'description': eq_data.get('description', ''),
                    'icon': eq_data.get('icon', ''),
                    'category': eq_data.get('category', ''),
                    'standards': eq_data.get('standards', []),
                    'sections': eq_data.get('sections', []),
                    'is_active': True
                }
            )
            synced_count += 1
            if created:
                created_count += 1
            else:
                updated_count += 1
        
        return Response({
            'message': 'Equipment types synced successfully',
            'total_synced': synced_count,
            'created': created_count,
            'updated': updated_count
        })


class ElectricalDatasheetViewSet(QualityCheckerMixin, viewsets.ModelViewSet):
    """ViewSet for managing electrical datasheets with AI-powered quality checking"""
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['equipment_type', 'status', 'project_number', 'discipline']
    search_fields = ['tag_number', 'service_description', 'location', 'project_name']
    ordering_fields = ['created_at', 'updated_at', 'tag_number', 'status']
    ordering = ['-created_at']

    def get_queryset(self):
        """Get queryset with optional filtering"""
        queryset = ElectricalDatasheet.objects.filter(is_deleted=False).select_related(
            'equipment_type', 'created_by', 'updated_by', 'reviewed_by', 'approved_by'
        ).annotate(
            comments_count=Count('comments')
        )
        
        # Filter by user if requested
        user_filter = self.request.query_params.get('user_filter', None)
        if user_filter == 'my_datasheets':
            queryset = queryset.filter(created_by=self.request.user)
        
        return queryset

    def get_serializer_class(self):
        """Return appropriate serializer based on action"""
        if self.action == 'list':
            return ElectricalDatasheetListSerializer
        elif self.action in ['create', 'update', 'partial_update']:
            return ElectricalDatasheetCreateUpdateSerializer
        return ElectricalDatasheetSerializer
    
    @action(detail=False, methods=['get'])
    def statistics(self, request):
        """Get statistics for datasheets dashboard"""
        queryset = self.get_queryset()
        
        # Count by status
        status_counts = queryset.values('status').annotate(count=Count('id'))
        status_dict = {item['status']: item['count'] for item in status_counts}
        
        # Calculate average compliance score
        datasheets_with_score = queryset.exclude(compliance_score__isnull=True)
        avg_compliance = 0
        if datasheets_with_score.exists():
            total_score = sum(ds.compliance_score for ds in datasheets_with_score if ds.compliance_score)
            avg_compliance = int(total_score / datasheets_with_score.count())
        
        return Response({
            'total': queryset.count(),
            'draft': status_dict.get('draft', 0),
            'validated': status_dict.get('validated', 0),
            'in_review': status_dict.get('in_review', 0),
            'approved': status_dict.get('approved', 0),
            'rejected': status_dict.get('rejected', 0),
            'avg_compliance': avg_compliance,
        })

    def perform_create(self, serializer):
        """Create datasheet with user tracking"""
        serializer.save(
            created_by=self.request.user,
            updated_by=self.request.user
        )

    def perform_update(self, serializer):
        """Update datasheet with user tracking and revision history"""
        instance = self.get_object()
        
        # Create revision history before update
        DatasheetRevisionHistory.objects.create(
            datasheet=instance,
            revision_number=instance.revision_number,
            form_data=instance.form_data,
            status=instance.status,
            revision_notes=instance.revision_notes,
            revised_by=self.request.user
        )
        
        # Increment revision number if form_data changed
        if 'form_data' in serializer.validated_data:
            serializer.save(
                updated_by=self.request.user,
                revision_number=instance.revision_number + 1
            )
        else:
            serializer.save(updated_by=self.request.user)

    @action(detail=True, methods=['post'])
    def submit_for_review(self, request, pk=None):
        """Submit datasheet for review"""
        datasheet = self.get_object()
        
        if datasheet.status not in ['draft', 'revision_required']:
            return Response(
                {"error": "Can only submit draft or revision required datasheets"},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        datasheet.status = 'under_review'
        datasheet.updated_by = request.user
        datasheet.save()
        
        return Response({
            'message': 'Datasheet submitted for review',
            'status': datasheet.status
        })

    @action(detail=True, methods=['post'])
    def approve(self, request, pk=None):
        """Approve datasheet"""
        datasheet = self.get_object()
        
        if datasheet.status != 'under_review':
            return Response(
                {"error": "Can only approve datasheets under review"},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        datasheet.status = 'approved'
        datasheet.approved_by = request.user
        datasheet.approved_at = timezone.now()
        datasheet.updated_by = request.user
        datasheet.save()
        
        return Response({
            'message': 'Datasheet approved',
            'status': datasheet.status,
            'approved_by': request.user.get_full_name(),
            'approved_at': datasheet.approved_at
        })

    @action(detail=True, methods=['post'])
    def reject(self, request, pk=None):
        """Reject datasheet"""
        datasheet = self.get_object()
        rejection_reason = request.data.get('reason', '')
        
        if datasheet.status != 'under_review':
            return Response(
                {"error": "Can only reject datasheets under review"},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        datasheet.status = 'rejected'
        datasheet.revision_notes = rejection_reason
        datasheet.reviewed_by = request.user
        datasheet.reviewed_at = timezone.now()
        datasheet.updated_by = request.user
        datasheet.save()
        
        return Response({
            'message': 'Datasheet rejected',
            'status': datasheet.status,
            'reason': rejection_reason
        })

    @action(detail=True, methods=['post'])
    def request_revision(self, request, pk=None):
        """Request revision for datasheet"""
        datasheet = self.get_object()
        revision_notes = request.data.get('notes', '')
        
        if datasheet.status != 'under_review':
            return Response(
                {"error": "Can only request revision for datasheets under review"},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        datasheet.status = 'revision_required'
        datasheet.revision_notes = revision_notes
        datasheet.reviewed_by = request.user
        datasheet.reviewed_at = timezone.now()
        datasheet.updated_by = request.user
        datasheet.save()
        
        return Response({
            'message': 'Revision requested',
            'status': datasheet.status,
            'notes': revision_notes
        })

    @action(detail=True, methods=['get'])
    def revisions(self, request, pk=None):
        """Get revision history for datasheet"""
        datasheet = self.get_object()
        revisions = datasheet.revision_history.all()
        serializer = DatasheetRevisionHistorySerializer(revisions, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=['get', 'post'])
    def comments(self, request, pk=None):
        """Get or add comments for datasheet"""
        datasheet = self.get_object()
        
        if request.method == 'GET':
            comments = datasheet.comments.filter(parent_comment__isnull=True)
            serializer = DatasheetCommentSerializer(comments, many=True)
            return Response(serializer.data)
        
        elif request.method == 'POST':
            serializer = DatasheetCommentSerializer(data=request.data)
            if serializer.is_valid():
                serializer.save(
                    datasheet=datasheet,
                    commented_by=request.user
                )
                return Response(serializer.data, status=status.HTTP_201_CREATED)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['get'])
    def statistics(self, request):
        """Get statistics about datasheets"""
        queryset = self.get_queryset()
        
        stats = {
            'total': queryset.count(),
            'by_status': {},
            'by_equipment_type': {},
            'recent_count': queryset.filter(
                created_at__gte=timezone.now() - timezone.timedelta(days=30)
            ).count()
        }
        
        # Count by status
        for status_choice in ElectricalDatasheet.STATUS_CHOICES:
            status_code = status_choice[0]
            stats['by_status'][status_code] = queryset.filter(status=status_code).count()
        
        # Count by equipment type
        equipment_types = ElectricalEquipmentType.objects.filter(is_active=True)
        for eq_type in equipment_types:
            stats['by_equipment_type'][eq_type.name] = queryset.filter(
                equipment_type=eq_type
            ).count()
        
        return Response(stats)

    @action(detail=True, methods=['post'])
    def duplicate(self, request, pk=None):
        """Create a duplicate of the datasheet"""
        original = self.get_object()
        new_tag_number = request.data.get('tag_number')
        
        if not new_tag_number:
            return Response(
                {"error": "New tag number is required"},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Check if tag number already exists
        if ElectricalDatasheet.objects.filter(tag_number=new_tag_number).exists():
            return Response(
                {"error": "Tag number already exists"},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Create duplicate
        duplicate = ElectricalDatasheet.objects.create(
            equipment_type=original.equipment_type,
            tag_number=new_tag_number,
            service_description=original.service_description,
            location=original.location,
            form_data=original.form_data.copy(),
            status='draft',
            project_name=original.project_name,
            project_number=original.project_number,
            discipline=original.discipline,
            created_by=request.user,
            updated_by=request.user
        )
        
        serializer = self.get_serializer(duplicate)
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    def _generate_mock_validation(self, equipment_type, standards, extracted_text):
        """
        Generate intelligent mock validation when OpenAI API is not available.
        Uses rule-based validation against ADNOC standards.
        """
        # Extract key parameters from text using regex and keywords
        import re
        
        extracted_data = {
            "voltage": "NOT FOUND",
            "current": "NOT FOUND",
            "power": "NOT FOUND",
            "frequency": "NOT FOUND",
            "manufacturer": "NOT FOUND",
            "model": "NOT FOUND",
            "additional_specs": {}
        }
        
        # Try to extract voltage
        voltage_match = re.search(r'(\d+\.?\d*)\s*(kV|KV|kv)', extracted_text, re.IGNORECASE)
        if voltage_match:
            extracted_data["voltage"] = f"{voltage_match.group(1)} kV"
        
        # Try to extract current
        current_match = re.search(r'(\d+\.?\d*)\s*(A|Amp|Ampere)', extracted_text, re.IGNORECASE)
        if current_match:
            extracted_data["current"] = f"{current_match.group(1)} A"
        
        # Try to extract power
        power_match = re.search(r'(\d+\.?\d*)\s*(MVA|kVA|MW|kW)', extracted_text, re.IGNORECASE)
        if power_match:
            extracted_data["power"] = f"{power_match.group(1)} {power_match.group(2)}"
        
        # Try to extract frequency
        freq_match = re.search(r'(\d+)\s*(Hz|hz)', extracted_text, re.IGNORECASE)
        if freq_match:
            extracted_data["frequency"] = f"{freq_match.group(1)} Hz"
        
        # Generate missing data recommendations
        missing_data = []
        if extracted_data["voltage"] == "NOT FOUND":
            missing_data.append({
                "parameter": "Rated Voltage",
                "criticality": "HIGH",
                "suggested_default": "11 kV (Based on equipment type and ADNOC standards)",
                "reasoning": "11kV is the most common distribution voltage for transformers and switchgear in ADNOC specifications"
            })
        
        if extracted_data["frequency"] == "NOT FOUND":
            missing_data.append({
                "parameter": "Rated Frequency",
                "criticality": "HIGH",
                "suggested_default": "50 Hz (UAE Standard)",
                "reasoning": "50 Hz is the standard frequency in UAE and GCC region as per ADNOC requirements"
            })
        
        if extracted_data["power"] == "NOT FOUND":
            missing_data.append({
                "parameter": "Power Rating",
                "criticality": "MEDIUM",
                "suggested_default": "1000 kVA (typical for 11kV transformer)",
                "reasoning": "Based on common industrial applications and ADNOC typical specifications for medium voltage equipment"
            })
        
        # Generate validation results based on standards
        validation_results = []
        compliance_count = 0
        total_checks = 0
        
        # Check voltage against standards
        if standards and 'voltage' in standards:
            total_checks += 1
            expected_voltage = standards.get('voltage', '11kV')
            if extracted_data["voltage"] != "NOT FOUND":
                passed = expected_voltage.replace(' ', '').lower() in extracted_data["voltage"].replace(' ', '').lower()
                compliance_count += 1 if passed else 0
                validation_results.append({
                    "parameter": "Rated Voltage",
                    "passed": passed,
                    "expected": expected_voltage,
                    "found": extracted_data["voltage"],
                    "message": f"Voltage {'matches' if passed else 'does not match'} ADNOC standard",
                    "recommendation": "" if passed else f"Update voltage rating to {expected_voltage} as per ADNOC specification"
                })
        
        # Check frequency
        if standards and 'frequency' in standards:
            total_checks += 1
            expected_freq = standards.get('frequency', '50Hz')
            if extracted_data["frequency"] != "NOT FOUND":
                passed = expected_freq.replace(' ', '').lower() in extracted_data["frequency"].replace(' ', '').lower()
                compliance_count += 1 if passed else 0
                validation_results.append({
                    "parameter": "Rated Frequency",
                    "passed": passed,
                    "expected": expected_freq,
                    "found": extracted_data["frequency"],
                    "message": f"Frequency {'is compliant' if passed else 'is not compliant'} with ADNOC standard",
                    "recommendation": "" if passed else f"Frequency must be {expected_freq} for UAE region"
                })
        
        # Add general checks
        validation_results.append({
            "parameter": "IEC Standards Compliance",
            "passed": True,
            "expected": "IEC 60076 / IEC 62271",
            "found": "Assumed compliant",
            "message": "Equipment should comply with relevant IEC standards",
            "recommendation": "Verify IEC certification documentation"
        })
        compliance_count += 1
        total_checks += 1
        
        # Calculate compliance score
        compliance_score = int((compliance_count / max(total_checks, 1)) * 100) if total_checks > 0 else 75
        
        # Generate comprehensive analysis
        extracted_params = [k for k, v in extracted_data.items() if v != 'NOT FOUND' and k != 'additional_specs']
        missing_params = len(missing_data)
        
        ai_analysis = f"""
🔍 ADNOC STANDARDS VALIDATION REPORT

Equipment Type: {equipment_type.name}
Analysis Date: {__import__('datetime').datetime.now().strftime('%Y-%m-%d %H:%M:%S')}

═══════════════════════════════════════════════════════════════

📊 DOCUMENT ANALYSIS SUMMARY:
Successfully extracted {len(extracted_params)} technical parameters from the provided document.
Extracted Parameters: {', '.join(extracted_params) if extracted_params else 'None - Manual entry required'}

⚠️ MISSING DATA ASSESSMENT:
{missing_params} critical parameter(s) require attention for complete ADNOC compliance.

✅ COMPLIANCE EVALUATION:
Overall Compliance Score: {compliance_score}%
{'🎯 EXCELLENT - Document meets or exceeds ADNOC standards' if compliance_score >= 90 else 
 '✓ GOOD - Minor adjustments needed for full compliance' if compliance_score >= 70 else
 '⚠️ ATTENTION REQUIRED - Significant gaps identified in documentation'}

═══════════════════════════════════════════════════════════════

🎯 RECOMMENDED ACTIONS:
1. Review and complete all HIGH criticality missing parameters immediately
2. Verify all extracted data against manufacturer specifications
3. Ensure IEC standard certifications (IEC 60076/62271) are attached
4. Cross-reference technical specifications with ADNOC-AGES standards
5. Submit for technical review once all mandatory fields are completed

💡 QUALITY ASSURANCE NOTES:
- All suggested default values are based on ADNOC standard requirements
- Equipment ratings must match application load requirements
- Safety features and protection systems must be clearly documented
- Manufacturer test certificates and quality documentation required

📋 STANDARDS REFERENCE:
- ADNOC-AGES-SP-1030 (Transformers)
- ADNOC-AGES-SP-1031 (Switchgear)
- IEC 60076 (Power Transformers)
- IEC 62271 (High Voltage Switchgear)

═══════════════════════════════════════════════════════════════
"""
        
        return {
            "equipment_type": equipment_type.name,
            "compliance_score": compliance_score,
            "extracted_data": extracted_data,
            "missing_data": missing_data,
            "validation_results": validation_results,
            "ai_analysis": ai_analysis.strip()
        }

    @action(detail=False, methods=['post'])
    def validate_diagram(self, request):
        """
        Validate electrical diagram using AI and ADNOC standards.
        Supports PDF and image files.
        """
        if 'file' not in request.FILES:
            return Response(
                {"error": "No file provided"},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        uploaded_file = request.FILES['file']
        equipment_type_id = request.data.get('equipment_type_id')
        
        if not equipment_type_id:
            return Response(
                {"error": "equipment_type_id is required"},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            equipment_type = ElectricalEquipmentType.objects.get(id=equipment_type_id)
        except ElectricalEquipmentType.DoesNotExist:
            return Response(
                {"error": "Invalid equipment type"},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Initialize services
        s3_service = ElectricalDatasheetS3Service()
        standards_manager = ADNOCStandardsManager()
        
        # Get ADNOC standards for validation
        equipment_name = equipment_type.name.lower()
        if 'transformer' in equipment_name:
            standards = standards_manager.get_transformer_standards('11kv')
        elif 'switchgear' in equipment_name:
            standards = standards_manager.get_switchgear_standards('11kv')
        else:
            standards = {}
        
        # Extract text from file
        file_extension = uploaded_file.name.lower().split('.')[-1]
        extracted_text = ""
        
        try:
            if file_extension == 'pdf':
                extracted_text = s3_service.extract_text_from_pdf(uploaded_file)
            elif file_extension in ['png', 'jpg', 'jpeg']:
                # For images, we'll use OpenAI Vision API
                extracted_text = "Image file provided for AI vision analysis"
            else:
                return Response(
                    {"error": "Unsupported file type. Please upload PDF or image files."},
                    status=status.HTTP_400_BAD_REQUEST
                )
        except Exception as e:
            return Response(
                {"error": f"File processing failed: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
        
        # Perform AI validation and data extraction using OpenAI
        try:
            # Check if OpenAI API key is properly configured
            api_key = settings.OPENAI_API_KEY
            is_valid_key = (
                api_key and 
                len(api_key) > 50 and 
                api_key.startswith('sk-') and
                'your-' not in api_key.lower() and 
                'here' not in api_key.lower()
            )
            
            if not is_valid_key:
                # Fallback: Generate intelligent mock validation based on extracted text and standards
                validation_data = self._generate_mock_validation(
                    equipment_type=equipment_type,
                    standards=standards,
                    extracted_text=extracted_text
                )
                return Response(validation_data, status=status.HTTP_200_OK)
            
            from openai import OpenAI
            client = OpenAI(api_key=settings.OPENAI_API_KEY)
            
            validation_prompt = f"""
You are an expert electrical engineer specializing in {equipment_type.name} validation and data extraction according to ADNOC standards.

Analyze the provided diagram/datasheet and perform the following tasks:

1. EXTRACT ALL AVAILABLE DATA from the document including:
   - Equipment specifications (voltage, current, power, frequency, etc.)
   - Manufacturer details
   - Model/Serial numbers
   - Technical ratings
   - Design parameters
   - Safety features
   - Testing requirements
   
2. IDENTIFY MISSING INFORMATION that should be present for complete documentation

3. VALIDATE against ADNOC Standards:
Equipment Type: {equipment_type.name}
ADNOC Standards Reference: {json.dumps(standards, indent=2)}

Extracted Document Content: {extracted_text[:3000]}

4. For any missing data, intelligently suggest DEFAULT VALUES based on:
   - ADNOC standard requirements
   - Industry best practices
   - Equipment type typical specifications
   - Safety regulations

Please provide your response as JSON with this exact structure:
{{
    "compliance_score": <number 0-100>,
    "extracted_data": {{
        "voltage": "<value or 'NOT FOUND'>",
        "current": "<value or 'NOT FOUND'>",
        "power": "<value or 'NOT FOUND'>",
        "frequency": "<value or 'NOT FOUND'>",
        "manufacturer": "<value or 'NOT FOUND'>",
        "model": "<value or 'NOT FOUND'>",
        "additional_specs": {{}}
    }},
    "missing_data": [
        {{
            "parameter": "<parameter name>",
            "criticality": "<HIGH/MEDIUM/LOW>",
            "suggested_default": "<intelligent default value>",
            "reasoning": "<why this default is recommended>"
        }}
    ],
    "validation_results": [
        {{
            "parameter": "<parameter name>",
            "passed": <true/false>,
            "expected": "<expected value/range from ADNOC>",
            "found": "<actual value found or 'MISSING'>",
            "message": "<validation message>",
            "recommendation": "<specific recommendation>"
        }}
    ],
    "ai_analysis": "<comprehensive analysis including data extraction summary, compliance assessment, and recommendations>"
}}
"""
            
            response = client.chat.completions.create(
                model=settings.OPENAI_MODEL,
                messages=[
                    {"role": "system", "content": "You are an expert electrical engineer with deep knowledge of ADNOC standards. Extract all data, identify missing information, validate against standards, and suggest intelligent defaults for missing critical parameters."},
                    {"role": "user", "content": validation_prompt}
                ],
                temperature=0.3,
                max_tokens=3000
            )
            
            ai_response = response.choices[0].message.content
            
            # Parse AI response
            try:
                validation_data = json.loads(ai_response)
                # Add equipment type to response
                validation_data['equipment_type'] = equipment_type.name
            except json.JSONDecodeError:
                # If AI returns non-JSON, wrap it
                validation_data = {
                    "equipment_type": equipment_type.name,
                    "compliance_score": 75,
                    "validation_results": [],
                    "ai_analysis": ai_response
                }
            
            return Response(validation_data, status=status.HTTP_200_OK)
            
        except Exception as e:
            return Response(
                {"error": f"AI validation failed: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=False, methods=['post'], url_path='verify-transformer')
    def verify_transformer_datasheet(self, request):
        """
        Verify transformer datasheet against supporting engineering documents
        Reuses P&ID/PFD verification architecture
        
        POST /api/v1/electrical/datasheets/verify-transformer/
        
        FormData:
        - transformer_datasheet: Excel file
        - transformer_calculation: PDF file (Transformer Sizing Calculation - Power and Distribution)
        
        Returns:
        {
            "success": true,
            "verification_results": [
                {
                    "parameter": "Transformer Rating",
                    "datasheet_value": "2500 kVA",
                    "document_value": "2500 kVA",
                    "status": "Valid",
                    "explanation": "...",
                    "confidence": "High",
                    "source_document": "Transformer Calculation"
                },
                ...
            ],
            "summary": {
                "total_parameters": 13,
                "valid": 10,
                "mismatch": 2,
                "missing": 1
            }
        }
        """
        from .transformer_verification_service import TransformerVerificationService
        import tempfile
        import os
        
        try:
            # Validate required files
            required_files = [
                'transformer_datasheet',
                'transformer_calculation'
            ]
            
            for file_key in required_files:
                if file_key not in request.FILES:
                    return Response({
                        'success': False,
                        'error': f'Missing required file: {file_key}'
                    }, status=status.HTTP_400_BAD_REQUEST)
            
            # Save uploaded files temporarily
            temp_files = {}
            try:
                for file_key in required_files:
                    uploaded_file = request.FILES[file_key]
                    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=os.path.splitext(uploaded_file.name)[1])
                    for chunk in uploaded_file.chunks():
                        temp_file.write(chunk)
                    temp_file.close()
                    temp_files[file_key] = temp_file.name
                
                # Initialize verification service
                verification_service = TransformerVerificationService()
                
                # Step 1: Extract Excel parameters
                logger.info("[TransformerVerification] Extracting Excel parameters...")
                excel_parameters = verification_service.extract_excel_parameters(
                    temp_files['transformer_datasheet']
                )
                
                if not excel_parameters:
                    return Response({
                        'success': False,
                        'error': 'No parameters could be extracted from Excel datasheet'
                    }, status=status.HTTP_400_BAD_REQUEST)
                
                # Step 2: Extract PDF content
                logger.info("[TransformerVerification] Extracting PDF document...")
                transformer_calc_text = verification_service.extract_pdf_content(
                    temp_files['transformer_calculation'],
                    'Transformer Calculation'
                )
                
                # Step 3: Perform AI verification
                logger.info("[TransformerVerification] Running AI verification...")
                verification_results = verification_service.verify_transformer_datasheet(
                    excel_parameters,
                    transformer_calc_text
                )
                
                # Step 4: Format results
                results_dict = verification_service.format_results_as_dict(verification_results)
                
                # Calculate summary statistics
                summary = {
                    'total_parameters': len(results_dict),
                    'valid': sum(1 for r in results_dict if r['status'] == 'Valid'),
                    'mismatch': sum(1 for r in results_dict if r['status'] == 'Mismatch'),
                    'incorrect': sum(1 for r in results_dict if r['status'] == 'Incorrect'),
                    'missing': sum(1 for r in results_dict if r['status'] == 'Missing'),
                }
                
                logger.info(f"[TransformerVerification] ✅ Completed: {summary}")
                
                return Response({
                    'success': True,
                    'verification_results': results_dict,
                    'extracted_parameters': excel_parameters,
                    'summary': summary
                }, status=status.HTTP_200_OK)
                
            finally:
                # Clean up temporary files
                for temp_file_path in temp_files.values():
                    try:
                        os.unlink(temp_file_path)
                    except:
                        pass
        
        except Exception as e:
            logger.error(f"[TransformerVerification] ❌ Error: {e}", exc_info=True)
            return Response({
                'success': False,
                'error': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['post'], url_path='generate-switchgear-datasheet')
    def generate_switchgear_datasheet(self, request):
        """
        Generate 11KV Switchgear Datasheet from SLD PDF
        
        POST /api/v1/electrical-datasheet/datasheets/generate-switchgear-datasheet/
        
        FormData:
        - sld_file: PDF file (SLD for 11KV Switchgear)
        - project_name: Project name (optional)
        - drawing_number: Drawing number (optional)
        - area: Area/location (optional)
        
        Returns:
        {
            "success": true,
            "datasheet_rows": [...],
            "summary": {
                "total_rows": 75,
                "equipment_count": 65,
                "completed_fields": 40,
                "missing_fields": 25
            }
        }
        """
        from .switchgear_datasheet_generator import SwitchgearDatasheetGenerator
        from django.http import HttpResponse
        
        try:
            if 'sld_file' not in request.FILES:
                return Response({
                    'success': False,
                    'error': 'SLD file is required'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            sld_file = request.FILES['sld_file']
            
            # Validate file type (soft-coded supported formats)
            from .document_extractor import is_supported, SUPPORTED_FORMATS_LABEL
            if not is_supported(sld_file.name):
                return Response({
                    'success': False,
                    'error': f'Unsupported file type. Supported formats: {SUPPORTED_FORMATS_LABEL}'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Extract project information
            project_info = {
                'project_name': request.data.get('project_name', ''),
                'drawing_number': request.data.get('drawing_number', ''),
                'area': request.data.get('area', ''),
                'voltage_level': '11KV'
            }
            
            logger.info(f"[SwitchgearDatasheet] Generating from SLD: {sld_file.name}")
            
            # Generate datasheet
            generator = SwitchgearDatasheetGenerator()
            result = generator.generate_datasheet_from_sld(sld_file, project_info)
            
            if not result['success']:
                return Response(result, status=status.HTTP_400_BAD_REQUEST)
            
            self._persist_smart_generation(
                request=request,
                equipment_type='mv_switchgear',
                source_files=[{'role': 'sld', 'file': sld_file}],
                project_info=project_info,
                result=result,
                generator=generator,
            )
            return Response(result, status=status.HTTP_200_OK)
            
        except Exception as e:
            logger.error(f"[SwitchgearDatasheet] Error: {e}", exc_info=True)
            return Response({
                'success': False,
                'error': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @action(detail=False, methods=['post'], url_path='export-switchgear-datasheet')
    def export_switchgear_datasheet(self, request):
        """
        Export 11KV Switchgear Datasheet to Excel
        
        POST /api/v1/electrical-datasheet/datasheets/export-switchgear-datasheet/
        
        Body (JSON):
        - datasheet_rows: Array of datasheet rows
        - project_info: Project metadata
        
        Returns: Excel file download
        """
        from .switchgear_datasheet_generator import SwitchgearDatasheetGenerator
        from django.http import HttpResponse
        from datetime import datetime
        
        try:
            datasheet_rows = request.data.get('datasheet_rows', [])
            project_info = request.data.get('project_info', {})
            
            if not datasheet_rows:
                return Response({
                    'success': False,
                    'error': 'No datasheet rows provided'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            logger.info(f"[SwitchgearDatasheet] Exporting {len(datasheet_rows)} rows to Excel")
            
            # Generate Excel
            generator = SwitchgearDatasheetGenerator()
            excel_buffer = generator.export_to_excel(datasheet_rows, project_info)
            
            # Create response with Excel file
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"11KV_Switchgear_Datasheet_{timestamp}.xlsx"
            
            response = HttpResponse(
                excel_buffer.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            response['Access-Control-Expose-Headers'] = 'Content-Disposition'
            
            logger.info(f"[SwitchgearDatasheet] ✅ Excel exported: {filename}")
            return response
            
        except Exception as e:
            logger.error(f"[SwitchgearDatasheet] Export error: {e}", exc_info=True)
            return Response({
                'success': False,
                'error': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    # ─────────────────────────────────────────────────────────────────────────
    # TRANSFORMER DATASHEET ENDPOINTS
    # ─────────────────────────────────────────────────────────────────────────

    @action(detail=False, methods=['post'], url_path='generate-transformer-datasheet')
    def generate_transformer_datasheet(self, request):
        """
        Generate Power/Distribution Transformer Datasheet from a Sizing Calculation PDF.

        POST /api/v1/electrical-datasheet/datasheets/generate-transformer-datasheet/

        FormData:
        - sizing_calc_file : PDF of the Transformer Sizing Calculation document
        - project_name     : (optional)
        - drawing_number   : (optional)
        - area             : (optional)

        Returns:
        {
            "success": true,
            "datasheet_rows": [...],   # sr_no, description, unit, required_data, vendor_data, rev
            "summary": { ... }
        }
        """
        from .transformer_datasheet_generator import TransformerDatasheetGenerator

        try:
            if 'sizing_calc_file' not in request.FILES:
                return Response(
                    {'success': False, 'error': 'Transformer sizing calculation file is required'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            sizing_file = request.FILES['sizing_calc_file']

            from .document_extractor import is_supported, SUPPORTED_FORMATS_LABEL
            if not is_supported(sizing_file.name):
                return Response(
                    {'success': False, 'error': f'Unsupported file type. Supported formats: {SUPPORTED_FORMATS_LABEL}'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            project_info = {
                'project_name':   request.data.get('project_name', ''),
                'drawing_number': request.data.get('drawing_number', ''),
                'area':           request.data.get('area', ''),
                'equipment_type': 'Power / Distribution Transformer',
            }

            logger.info(f"[TransformerDatasheet] Generating from: {sizing_file.name}")

            generator = TransformerDatasheetGenerator()
            result = generator.generate_datasheet_from_sizing_calc(sizing_file, project_info)

            if not result['success']:
                return Response(result, status=status.HTTP_400_BAD_REQUEST)

            # Persist to DB + S3 (best-effort; never block success on storage failure)
            self._persist_smart_generation(
                request=request,
                equipment_type='transformer',
                source_files=[{'role': 'sizing_calc', 'file': sizing_file}],
                project_info=project_info,
                result=result,
                generator=generator,
            )
            return Response(result, status=status.HTTP_200_OK)

        except Exception as e:
            logger.error(f"[TransformerDatasheet] Error: {e}", exc_info=True)
            return Response(
                {'success': False, 'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=False, methods=['post'], url_path='export-transformer-datasheet')
    def export_transformer_datasheet(self, request):
        """
        Export Transformer Datasheet rows to a formatted Excel file.

        POST /api/v1/electrical-datasheet/datasheets/export-transformer-datasheet/

        Body (JSON):
        - datasheet_rows : array of row objects
        - project_info   : project metadata dict

        Returns: Excel file (.xlsx) download
        """
        from .transformer_datasheet_generator import TransformerDatasheetGenerator
        from django.http import HttpResponse
        from datetime import datetime

        try:
            datasheet_rows = request.data.get('datasheet_rows', [])
            project_info   = request.data.get('project_info', {})

            if not datasheet_rows:
                return Response(
                    {'success': False, 'error': 'No datasheet rows provided'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            logger.info(f"[TransformerDatasheet] Exporting {len(datasheet_rows)} rows to Excel")

            generator = TransformerDatasheetGenerator()
            excel_buffer = generator.export_to_excel(datasheet_rows, project_info)

            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename  = f"Transformer_Datasheet_{timestamp}.xlsx"

            response = HttpResponse(
                excel_buffer.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            response['Access-Control-Expose-Headers'] = 'Content-Disposition'

            logger.info(f"[TransformerDatasheet] ✅ Excel exported: {filename}")
            return response

        except Exception as e:
            logger.error(f"[TransformerDatasheet] Export error: {e}", exc_info=True)
            return Response(
                {'success': False, 'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    # ─────────────────────────────────────────────────────────────────────────
    # DG SET DATASHEET ENDPOINTS
    # ─────────────────────────────────────────────────────────────────────────

    @action(detail=False, methods=['post'], url_path='generate-dg-datasheet')
    def generate_dg_datasheet(self, request):
        """
        Generate Emergency Diesel Generator (EDG) Set Datasheet from a Sizing Calculation PDF.

        POST /api/v1/electrical-datasheet/datasheets/generate-dg-datasheet/

        FormData:
        - edg_sizing_file : PDF of the EDG Sizing Calculation document
        - project_name    : (optional)
        - drawing_number  : (optional)
        - area            : (optional)

        Returns:
        {
            "success": true,
            "datasheet_rows": [...],   # sr_no, description, unit, required_data, vendor_data, rev
            "summary": { ... }
        }
        """
        from .dg_set_datasheet_generator import DGSetDatasheetGenerator

        try:
            if 'edg_sizing_file' not in request.FILES:
                return Response(
                    {'success': False, 'error': 'EDG sizing calculation file is required'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            edg_file = request.FILES['edg_sizing_file']

            from .document_extractor import is_supported, SUPPORTED_FORMATS_LABEL
            if not is_supported(edg_file.name):
                return Response(
                    {'success': False, 'error': f'Unsupported file type. Supported formats: {SUPPORTED_FORMATS_LABEL}'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            project_info = {
                'project_name':   request.data.get('project_name', ''),
                'drawing_number': request.data.get('drawing_number', ''),
                'area':           request.data.get('area', ''),
                'equipment_type': 'Emergency Diesel Generator Set',
            }

            logger.info(f"[DGSetDatasheet] Generating from: {edg_file.name}")

            generator = DGSetDatasheetGenerator()
            result = generator.generate_datasheet_from_sizing_calc(edg_file, project_info)

            if not result['success']:
                return Response(result, status=status.HTTP_400_BAD_REQUEST)

            self._persist_smart_generation(
                request=request,
                equipment_type='dg_set',
                source_files=[{'role': 'edg_sizing', 'file': edg_file}],
                project_info=project_info,
                result=result,
                generator=generator,
            )
            return Response(result, status=status.HTTP_200_OK)

        except Exception as e:
            logger.error(f"[DGSetDatasheet] Error: {e}", exc_info=True)
            return Response(
                {'success': False, 'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=False, methods=['post'], url_path='export-dg-datasheet')
    def export_dg_datasheet(self, request):
        """
        Export DG Set Datasheet rows to a formatted Excel file.

        POST /api/v1/electrical-datasheet/datasheets/export-dg-datasheet/

        Body (JSON):
        - datasheet_rows : array of row objects
        - project_info   : project metadata dict

        Returns: Excel file (.xlsx) download
        """
        from .dg_set_datasheet_generator import DGSetDatasheetGenerator
        from django.http import HttpResponse
        from datetime import datetime

        try:
            datasheet_rows = request.data.get('datasheet_rows', [])
            project_info   = request.data.get('project_info', {})

            if not datasheet_rows:
                return Response(
                    {'success': False, 'error': 'No datasheet rows provided'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            logger.info(f"[DGSetDatasheet] Exporting {len(datasheet_rows)} rows to Excel")

            generator = DGSetDatasheetGenerator()
            excel_buffer = generator.export_to_excel(datasheet_rows, project_info)

            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename  = f"DGSet_Datasheet_{timestamp}.xlsx"

            response = HttpResponse(
                excel_buffer.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            response['Access-Control-Expose-Headers'] = 'Content-Disposition'

            logger.info(f"[DGSetDatasheet] ✅ Excel exported: {filename}")
            return response

        except Exception as e:
            logger.error(f"[DGSetDatasheet] Export error: {e}", exc_info=True)
            return Response(
                {'success': False, 'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    # LV SWITCHGEAR DATASHEET ENDPOINTS
    # ─────────────────────────────────────────────────────────────────────────

    @action(detail=False, methods=['post'], url_path='generate-lv-switchgear-datasheet')
    def generate_lv_switchgear_datasheet(self, request):
        """
        Generate LV Switchgear Datasheet from Technical Datasheet PDF.

        POST /api/v1/electrical-datasheet/datasheets/generate-lv-switchgear-datasheet/

        FormData:
        - lv_datasheet_file : PDF of the Technical Datasheet for LV Switchgear
        - project_name      : (optional)
        - drawing_number    : (optional)
        - area              : (optional)
        """
        from .lv_switchgear_datasheet_generator import LVSwitchgearDatasheetGenerator

        try:
            if 'lv_datasheet_file' not in request.FILES:
                return Response(
                    {'success': False, 'error': 'Technical Datasheet for LV Switchgear file is required'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            lv_file = request.FILES['lv_datasheet_file']

            if not lv_file.name.lower().endswith('.pdf'):
                return Response(
                    {'success': False, 'error': 'Only PDF files are supported'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            project_info = {
                'project_name':   request.data.get('project_name', ''),
                'drawing_number': request.data.get('drawing_number', ''),
                'area':           request.data.get('area', ''),
                'equipment_type': 'LV Switchgear',
            }

            logger.info(f"[LVSwitchgearDatasheet] Generating from: {lv_file.name}")

            generator = LVSwitchgearDatasheetGenerator()
            result = generator.generate_datasheet_from_document(lv_file, project_info)

            if not result['success']:
                return Response(result, status=status.HTTP_400_BAD_REQUEST)

            return Response(result, status=status.HTTP_200_OK)

        except Exception as e:
            logger.error(f"[LVSwitchgearDatasheet] Error: {e}", exc_info=True)
            return Response(
                {'success': False, 'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=False, methods=['post'], url_path='export-lv-switchgear-datasheet')
    def export_lv_switchgear_datasheet(self, request):
        """
        Export LV Switchgear Datasheet rows to a formatted Excel file.

        POST /api/v1/electrical-datasheet/datasheets/export-lv-switchgear-datasheet/

        Body (JSON):
        - datasheet_rows : array of row objects
        - project_info   : project metadata dict
        """
        from .lv_switchgear_datasheet_generator import LVSwitchgearDatasheetGenerator
        from django.http import HttpResponse
        from datetime import datetime

        try:
            datasheet_rows = request.data.get('datasheet_rows', [])
            project_info   = request.data.get('project_info', {})

            if not datasheet_rows:
                return Response(
                    {'success': False, 'error': 'No datasheet rows provided'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            logger.info(f"[LVSwitchgearDatasheet] Exporting {len(datasheet_rows)} rows to Excel")

            generator = LVSwitchgearDatasheetGenerator()
            excel_buffer = generator.export_to_excel(datasheet_rows, project_info)

            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename  = f"LV_Switchgear_Datasheet_{timestamp}.xlsx"

            response = HttpResponse(
                excel_buffer.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            response['Access-Control-Expose-Headers'] = 'Content-Disposition'

            logger.info(f"[LVSwitchgearDatasheet] ✅ Excel exported: {filename}")
            return response

        except Exception as e:
            logger.error(f"[LVSwitchgearDatasheet] Export error: {e}", exc_info=True)
            return Response(
                {'success': False, 'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=False, methods=['post'], url_path='generate-smart')
    def generate_smart_datasheet(self, request):
        """
        Smart Datasheet Generator for 6 Electrical Equipment Types
        
        POST /api/v1/electrical-datasheet/datasheets/generate-smart/
        
        FormData:
        - equipment_type: Equipment ID (transformer, dg_set, mv_switchgear, lv_switchgear, ac_ups, dc_ups)
        - files: Multiple files (PDFs, Excel, Images)
        - file_type_<filename>: File type classification for each file
        
        Returns:
        {
            "success": true,
            "datasheet_id": 123,
            "excel_url": "/electrical-datasheet/datasheets/123/download/",
            "summary": {
                "files_processed": 3,
                "fields_extracted": 45,
                "confidence": "High",
                "processing_time": "45s"
            },
            "extracted_data": {...}
        }
        """
        import tempfile
        import os
        from datetime import datetime
        
        try:
            equipment_type = request.data.get('equipment_type')
            if not equipment_type:
                return Response({
                    'success': False,
                    'error': 'Equipment type is required'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            equipment_mapping = {
                'transformer': 'transformer',
                'dg_set': 'edg',
                'mv_switchgear': 'switchgear',
                'lv_switchgear': 'lv_equipment',
                'ac_ups': 'ups',
                'dc_ups': 'ups',
            }
            
            internal_type = equipment_mapping.get(equipment_type)
            if not internal_type:
                return Response({
                    'success': False,
                    'error': f'Unsupported equipment type: {equipment_type}'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Get uploaded files
            uploaded_files = request.FILES.getlist('files')
            if not uploaded_files:
                return Response({
                    'success': False,
                    'error': 'No files uploaded'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            logger.info(f"[SmartDatasheet] Starting generation for {equipment_type}, {len(uploaded_files)} files")
            
            start_time = datetime.now()
            
            # Process files and extract data using AI
            extracted_fields = {}
            temp_files = []
            
            try:
                # Save files temporarily and process
                for uploaded_file in uploaded_files:
                    suffix = '.' + uploaded_file.name.split('.')[-1]
                    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=suffix)
                    for chunk in uploaded_file.chunks():
                        temp_file.write(chunk)
                    temp_file.close()
                    temp_files.append(temp_file.name)
                    
                    # Extract text/data from file based on type
                    if suffix.lower() == '.pdf':
                        extracted_text = self._extract_pdf_text(temp_file.name)
                        extracted_fields[uploaded_file.name] = extracted_text
                    elif suffix.lower() in ['.xlsx', '.xls']:
                        extracted_data = self._extract_excel_data(temp_file.name)
                        extracted_fields[uploaded_file.name] = extracted_data
                    elif suffix.lower() in ['.png', '.jpg', '.jpeg']:
                        extracted_text = self._extract_image_text(temp_file.name)
                        extracted_fields[uploaded_file.name] = extracted_text
                
                # Use AI to structure the extracted data into datasheet format
                structured_data = self._structure_datasheet_with_ai(
                    equipment_type=internal_type,
                    extracted_fields=extracted_fields
                )
                
                # Create datasheet record
                from .models import ElectricalDatasheet, ElectricalEquipmentType
                from .equipment_types_config import EQUIPMENT_TYPES_CONFIG

                # Auto-seed the equipment type if it doesn't exist yet
                equipment_obj = ElectricalEquipmentType.objects.filter(id=internal_type).first()
                if not equipment_obj:
                    # Find config entry and create it
                    cfg = next((c for c in EQUIPMENT_TYPES_CONFIG if c['id'] == internal_type), None)
                    if cfg:
                        equipment_obj, _ = ElectricalEquipmentType.objects.get_or_create(
                            id=internal_type,
                            defaults={
                                'name': cfg.get('name', internal_type),
                                'code': cfg.get('code', internal_type.upper()[:5]),
                                'description': cfg.get('description', ''),
                                'icon': cfg.get('icon', ''),
                                'category': cfg.get('category', 'Electrical Equipment'),
                                'standards': cfg.get('standards', []),
                                'sections': cfg.get('sections', []),
                                'is_active': True,
                            }
                        )
                        logger.info(f"[SmartDatasheet] Auto-seeded ElectricalEquipmentType: {internal_type}")
                    else:
                        # Fallback: create minimal record
                        equipment_obj, _ = ElectricalEquipmentType.objects.get_or_create(
                            id=internal_type,
                            defaults={
                                'name': internal_type.replace('_', ' ').title(),
                                'code': internal_type.upper()[:5],
                                'description': f'Auto-created for {internal_type}',
                                'category': 'Electrical Equipment',
                                'is_active': True,
                            }
                        )
                        logger.warning(f"[SmartDatasheet] No config found for {internal_type}, created minimal record")
                
                import uuid as _uuid
                auto_tag = f"{equipment_type.upper()[:AUTO_TAG_PREFIX_LENGTH]}-{_uuid.uuid4().hex[:AUTO_TAG_UUID_LENGTH].upper()}"

                datasheet = ElectricalDatasheet.objects.create(
                    equipment_type=equipment_obj,
                    tag_number=auto_tag,
                    service_description=f"Auto-generated {equipment_type.replace('_',' ').title()} datasheet",
                    location='',
                    form_data=structured_data,
                    created_by=request.user,
                    status='draft'
                )
                
                processing_time = (datetime.now() - start_time).total_seconds()
                
                logger.info(f"[SmartDatasheet] ✓ Generated datasheet ID {datasheet.id} in {processing_time}s")
                
                return Response({
                    'success': True,
                    'datasheet_id': datasheet.id,
                    'excel_url': f'/electrical-datasheet/datasheets/{datasheet.id}/download/',
                    'summary': {
                        'files_processed': len(uploaded_files),
                        'fields_extracted': len(structured_data),
                        'confidence': 'High',
                        'processing_time': f'{int(processing_time)}s'
                    },
                    'extracted_data': structured_data
                }, status=status.HTTP_200_OK)
                
            finally:
                # Clean up temporary files
                for temp_file in temp_files:
                    try:
                        os.unlink(temp_file)
                    except:
                        pass
                        
        except Exception as e:
            logger.error(f"[SmartDatasheet] ❌ Error: {e}", exc_info=True)
            return Response({
                'success': False,
                'error': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    def _extract_pdf_text(self, pdf_path):
        """Extract text from PDF file"""
        try:
            import PyPDF2
            text = ""
            with open(pdf_path, 'rb') as file:
                pdf_reader = PyPDF2.PdfReader(file)
                for page in pdf_reader.pages:
                    text += page.extract_text() + "\n"
            return text
        except Exception as e:
            logger.error(f"PDF extraction error: {e}")
            return ""
    
    def _extract_excel_data(self, excel_path):
        """Extract data from Excel file"""
        try:
            import pandas as pd
            df = pd.read_excel(excel_path, sheet_name=None)
            data = {}
            for sheet_name, sheet_df in df.items():
                data[sheet_name] = sheet_df.to_dict()
            return data
        except Exception as e:
            logger.error(f"Excel extraction error: {e}")
            return {}
    
    def _extract_image_text(self, image_path):
        """Extract text from image using OCR"""
        try:
            from PIL import Image
            import pytesseract
            image = Image.open(image_path)
            text = pytesseract.image_to_string(image)
            return text
        except Exception as e:
            logger.error(f"Image OCR error: {e}")
            return ""
    
    def _structure_datasheet_with_ai(self, equipment_type, extracted_fields):
        """Use AI to structure extracted data into proper datasheet format"""
        # This would call OpenAI/Claude to intelligently structure the data
        # For now, return extracted fields as-is
        # TODO: Implement AI structuring
        return {
            'equipment_type': equipment_type,
            'extracted_fields': extracted_fields,
            'ai_structured': False
        }

    @action(detail=False, methods=['post'], url_path='export-validation')
    def export_validation_excel(self, request):
        """
        Export validation results to Excel format.
        Advanced formatting with color-coded compliance scores.
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from django.http import HttpResponse
        from datetime import datetime
        import io
        
        validation_data = request.data
        
        if not validation_data:
            return Response(
                {"error": "No validation data provided"},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Create workbook
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create Summary Sheet
        ws_summary = wb.create_sheet("Validation Summary", 0)
        ws_summary.sheet_properties.tabColor = "1072BA"
        
        # Header styling
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Summary Sheet - Title
        ws_summary['A1'] = 'ELECTRICAL DATASHEET VALIDATION REPORT'
        ws_summary['A1'].font = Font(bold=True, size=16, color="1F4E78")
        ws_summary.merge_cells('A1:D1')
        
        # Summary Info
        row = 3
        ws_summary[f'A{row}'] = 'Generated Date:'
        ws_summary[f'B{row}'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        ws_summary[f'A{row}'].font = Font(bold=True)
        
        row += 1
        equipment_type = validation_data.get('equipment_type', 'N/A')
        ws_summary[f'A{row}'] = 'Equipment Type:'
        ws_summary[f'B{row}'] = equipment_type
        ws_summary[f'A{row}'].font = Font(bold=True)
        
        row += 1
        compliance_score = validation_data.get('compliance_score', 0)
        ws_summary[f'A{row}'] = 'Compliance Score:'
        ws_summary[f'B{row}'] = f"{compliance_score}%"
        ws_summary[f'A{row}'].font = Font(bold=True)
        
        # Color-code compliance score
        if compliance_score >= 90:
            ws_summary[f'B{row}'].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            ws_summary[f'B{row}'].font = Font(color="006100", bold=True, size=14)
        elif compliance_score >= 70:
            ws_summary[f'B{row}'].fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            ws_summary[f'B{row}'].font = Font(color="9C5700", bold=True, size=14)
        else:
            ws_summary[f'B{row}'].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            ws_summary[f'B{row}'].font = Font(color="9C0006", bold=True, size=14)
        
        # Extracted Data Sheet
        ws_extracted = wb.create_sheet("Extracted Data")
        ws_extracted.sheet_properties.tabColor = "70AD47"
        
        ws_extracted['A1'] = 'EXTRACTED EQUIPMENT DATA'
        ws_extracted['A1'].font = Font(bold=True, size=14, color="70AD47")
        ws_extracted.merge_cells('A1:C1')
        
        # Headers
        headers = ['Parameter', 'Value', 'Status']
        for col, header in enumerate(headers, 1):
            cell = ws_extracted.cell(row=3, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Extracted data
        extracted_data = validation_data.get('extracted_data', {})
        row = 4
        for param, value in extracted_data.items():
            if param != 'additional_specs':
                ws_extracted[f'A{row}'] = param.replace('_', ' ').title()
                ws_extracted[f'B{row}'] = value
                
                # Status indicator
                if value and value != 'NOT FOUND':
                    ws_extracted[f'C{row}'] = '✓ Found'
                    ws_extracted[f'C{row}'].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    ws_extracted[f'C{row}'].font = Font(color="006100", bold=True)
                else:
                    ws_extracted[f'C{row}'] = '✗ Missing'
                    ws_extracted[f'C{row}'].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    ws_extracted[f'C{row}'].font = Font(color="9C0006", bold=True)
                
                for col in range(1, 4):
                    ws_extracted.cell(row=row, column=col).border = border
                row += 1
        
        # Auto-fit columns
        for col in range(1, 4):
            ws_extracted.column_dimensions[get_column_letter(col)].width = 25
        
        # Missing Data Sheet
        ws_missing = wb.create_sheet("Missing Parameters")
        ws_missing.sheet_properties.tabColor = "FFC000"
        
        ws_missing['A1'] = 'MISSING PARAMETERS & AI RECOMMENDATIONS'
        ws_missing['A1'].font = Font(bold=True, size=14, color="FFC000")
        ws_missing.merge_cells('A1:E1')
        
        # Headers
        headers = ['Parameter', 'Criticality', 'Suggested Default', 'Reasoning', 'Action Required']
        for col, header in enumerate(headers, 1):
            cell = ws_missing.cell(row=3, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border
        
        # Missing data
        missing_data = validation_data.get('missing_data', [])
        row = 4
        for item in missing_data:
            ws_missing[f'A{row}'] = item.get('parameter', '')
            ws_missing[f'B{row}'] = item.get('criticality', '')
            ws_missing[f'C{row}'] = item.get('suggested_default', '')
            ws_missing[f'D{row}'] = item.get('reasoning', '')
            
            # Criticality color coding
            criticality = item.get('criticality', '')
            if criticality == 'HIGH':
                ws_missing[f'B{row}'].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                ws_missing[f'B{row}'].font = Font(color="9C0006", bold=True)
                ws_missing[f'E{row}'] = 'URGENT'
            elif criticality == 'MEDIUM':
                ws_missing[f'B{row}'].fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                ws_missing[f'B{row}'].font = Font(color="9C5700", bold=True)
                ws_missing[f'E{row}'] = 'Required'
            else:
                ws_missing[f'B{row}'].fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
                ws_missing[f'B{row}'].font = Font(color="1F4E78")
                ws_missing[f'E{row}'] = 'Optional'
            
            for col in range(1, 6):
                cell = ws_missing.cell(row=row, column=col)
                cell.border = border
                cell.alignment = Alignment(wrap_text=True, vertical='top')
            
            row += 1
        
        # Auto-fit columns
        ws_missing.column_dimensions['A'].width = 20
        ws_missing.column_dimensions['B'].width = 12
        ws_missing.column_dimensions['C'].width = 25
        ws_missing.column_dimensions['D'].width = 50
        ws_missing.column_dimensions['E'].width = 15
        
        # Validation Results Sheet
        ws_validation = wb.create_sheet("ADNOC Validation")
        ws_validation.sheet_properties.tabColor = "C00000"
        
        ws_validation['A1'] = 'ADNOC STANDARDS VALIDATION RESULTS'
        ws_validation['A1'].font = Font(bold=True, size=14, color="C00000")
        ws_validation.merge_cells('A1:F1')
        
        # Headers
        headers = ['Parameter', 'Status', 'Expected (ADNOC)', 'Found', 'Message', 'Recommendation']
        for col, header in enumerate(headers, 1):
            cell = ws_validation.cell(row=3, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border
        
        # Validation results
        validation_results = validation_data.get('validation_results', [])
        row = 4
        for item in validation_results:
            ws_validation[f'A{row}'] = item.get('parameter', '')
            passed = item.get('passed', False)
            ws_validation[f'B{row}'] = '✓ PASS' if passed else '✗ FAIL'
            ws_validation[f'C{row}'] = item.get('expected', '')
            ws_validation[f'D{row}'] = item.get('found', '')
            ws_validation[f'E{row}'] = item.get('message', '')
            ws_validation[f'F{row}'] = item.get('recommendation', '')
            
            # Status color coding
            if passed:
                ws_validation[f'B{row}'].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                ws_validation[f'B{row}'].font = Font(color="006100", bold=True)
            else:
                ws_validation[f'B{row}'].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                ws_validation[f'B{row}'].font = Font(color="9C0006", bold=True)
            
            for col in range(1, 7):
                cell = ws_validation.cell(row=row, column=col)
                cell.border = border
                cell.alignment = Alignment(wrap_text=True, vertical='top')
            
            row += 1
        
        # Auto-fit columns
        ws_validation.column_dimensions['A'].width = 20
        ws_validation.column_dimensions['B'].width = 12
        ws_validation.column_dimensions['C'].width = 25
        ws_validation.column_dimensions['D'].width = 25
        ws_validation.column_dimensions['E'].width = 40
        ws_validation.column_dimensions['F'].width = 40
        
        # AI Analysis Sheet
        ws_analysis = wb.create_sheet("AI Analysis")
        ws_analysis.sheet_properties.tabColor = "7030A0"
        
        ws_analysis['A1'] = 'COMPREHENSIVE AI ANALYSIS'
        ws_analysis['A1'].font = Font(bold=True, size=14, color="7030A0")
        ws_analysis.merge_cells('A1:A1')
        
        ai_analysis = validation_data.get('ai_analysis', 'No analysis available')
        ws_analysis['A3'] = ai_analysis
        ws_analysis['A3'].alignment = Alignment(wrap_text=True, vertical='top')
        ws_analysis.column_dimensions['A'].width = 120
        ws_analysis.row_dimensions[3].height = 300
        
        # Save to BytesIO
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        # Create response
        response = HttpResponse(
            excel_file.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"electrical_validation_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response
    
    @action(detail=True, methods=['post'])
    def attach_file(self, request, pk=None):
        """
        Attach additional file to existing datasheet.
        """
        datasheet = self.get_object()
        
        if 'file' not in request.FILES:
            return Response(
                {"error": "No file provided"},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        uploaded_file = request.FILES['file']
        file_type = request.data.get('file_type', 'supplementary')  # supplementary, vendor_drawing, test_report
        
        # Initialize S3 service
        s3_service = ElectricalDatasheetS3Service()
        
        # Upload file
        try:
            file_info = s3_service.upload_datasheet(
                uploaded_file,
                datasheet.equipment_type.id,
                f"{datasheet.tag_number}_{file_type}"
            )
        except Exception as e:
            return Response(
                {"error": f"File upload failed: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
        
        # Update datasheet form_data with attached file info
        form_data = datasheet.form_data or {}
        if 'attached_files' not in form_data:
            form_data['attached_files'] = []
        
        form_data['attached_files'].append({
            'file_name': file_info['file_name'],
            'file_type': file_type,
            's3_key': file_info.get('s3_key', ''),
            'local_path': file_info.get('local_path', ''),
            'uploaded_at': timezone.now().isoformat(),
            'uploaded_by': request.user.get_full_name()
        })
        
        datasheet.form_data = form_data
        datasheet.updated_by = request.user
        datasheet.save()
        
        return Response({
            'message': 'File attached successfully',
            'file_info': file_info,
            'total_attachments': len(form_data['attached_files'])
        })
    
    @action(detail=True, methods=['get'], url_path='download')
    def download_excel(self, request, pk=None):
        """
        Generate and download a generic Excel file for any smart-generated datasheet.
        Converts the form_data dict into a two-column (Field / Value) spreadsheet.
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from django.http import HttpResponse
        import io

        datasheet = self.get_object()
        form_data = datasheet.form_data or {}
        equipment_type = datasheet.equipment_type.name if datasheet.equipment_type else 'Equipment'

        wb = Workbook()
        ws = wb.active
        ws.title = "Datasheet"

        # ── Styles ────────────────────────────────────────────────
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        label_fill  = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
        label_font  = Font(name="Calibri", bold=True, size=10)
        value_font  = Font(name="Calibri", size=10)
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left_align   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
        thin = Side(style="thin", color="AAAAAA")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        # ── Title row ─────────────────────────────────────────────
        ws.merge_cells("A1:B1")
        title_cell = ws["A1"]
        title_cell.value = f"{equipment_type} — Technical Datasheet"
        title_cell.font = Font(name="Calibri", bold=True, color="FFFFFF", size=13)
        title_cell.fill = header_fill
        title_cell.alignment = center_align
        ws.row_dimensions[1].height = 28

        # ── Column headers ────────────────────────────────────────
        ws["A2"] = "Field"
        ws["B2"] = "Value"
        for cell in [ws["A2"], ws["B2"]]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = border
        ws.row_dimensions[2].height = 20

        # ── Data rows ─────────────────────────────────────────────
        row = 3
        for key, value in form_data.items():
            label = str(key).replace("_", " ").title()
            display_val = value if not isinstance(value, (dict, list)) else json.dumps(value, indent=2)

            ws.cell(row=row, column=1, value=label).font  = label_font
            ws.cell(row=row, column=1).fill      = label_fill
            ws.cell(row=row, column=1).alignment = left_align
            ws.cell(row=row, column=1).border    = border

            ws.cell(row=row, column=2, value=str(display_val) if display_val is not None else "").font = value_font
            ws.cell(row=row, column=2).alignment = left_align
            ws.cell(row=row, column=2).border    = border

            ws.row_dimensions[row].height = 18
            row += 1

        # ── Column widths ─────────────────────────────────────────
        ws.column_dimensions["A"].width = 40
        ws.column_dimensions["B"].width = 60

        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        safe_name = equipment_type.replace(" ", "_").replace("/", "-")
        response = HttpResponse(
            excel_file.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="{safe_name}_Datasheet_{datasheet.id}.xlsx"'
        return response

    @action(detail=True, methods=['get'])
    def download_file(self, request, pk=None):
        """
        Download attached file from datasheet.
        """
        datasheet = self.get_object()
        file_index = request.query_params.get('file_index', 0)
        
        try:
            file_index = int(file_index)
        except ValueError:
            return Response(
                {"error": "Invalid file_index"},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        form_data = datasheet.form_data or {}
        attached_files = form_data.get('attached_files', [])
        
        if file_index >= len(attached_files):
            return Response(
                {"error": "File not found"},
                status=status.HTTP_404_NOT_FOUND
            )
        
        file_info = attached_files[file_index]
        s3_service = ElectricalDatasheetS3Service()
        
        try:
            # Download file from S3 or local storage
            file_path = s3_service.download_datasheet(
                file_info.get('s3_key') or file_info.get('local_path')
            )
            
            return Response({
                'file_path': file_path,
                'file_name': file_info['file_name'],
                'download_url': f'/media/{file_path}' if not file_info.get('s3_key') else None
            })
        except Exception as e:
            return Response(
                {"error": f"File download failed: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    def perform_destroy(self, instance):
        """Soft delete datasheet"""
        instance.is_deleted = True
        instance.deleted_at = timezone.now()
        instance.deleted_by = self.request.user
        instance.save()

    # ════════════════════════════════════════════════════════════════════
    # SMART GENERATOR — persistence, history, edit, recheck, share, PDF
    # ════════════════════════════════════════════════════════════════════
    def _persist_smart_generation(self, *, request, equipment_type, source_files, project_info, result, generator):
        """Best-effort persist a successful smart generation (DB + S3).

        Augments `result` dict in-place with `datasheet_id` and `excel_url`.
        Never raises — storage failures are logged and ignored so the
        original API response is preserved.
        """
        try:
            from .smart_storage import persist_generation
            from .models import GeneratedDatasheet  # noqa
            rows     = result.get('datasheet_rows', []) or []
            summary  = result.get('summary', {}) or {}
            variant  = (summary.get('variant') or 'default')
            title    = project_info.get('variant_title') or project_info.get('equipment_type') or ''
            metadata = {
                'project_info':      project_info or {},
                'original_filename': source_files[0]['file'].name if source_files else '',
                'extraction_metadata': result.get('extraction_metadata', {}),
            }

            # Render Excel for permanent storage
            excel_bytes = None
            try:
                buf = generator.export_to_excel(rows, project_info)
                excel_bytes = buf.getvalue() if hasattr(buf, 'getvalue') else bytes(buf)
            except Exception as exc:
                logger.warning(f"[smart_persist] excel render failed: {exc}")

            ds, excel_url = persist_generation(
                user           = request.user,
                equipment_type = equipment_type,
                rows           = rows,
                summary        = summary,
                metadata       = metadata,
                source_files   = source_files,
                excel_bytes    = excel_bytes,
                title          = title,
                variant        = variant,
            )
            result['datasheet_id'] = str(ds.id)
            if excel_url:
                result['excel_url'] = excel_url

            # Audit log
            try:
                from apps.activity.tracker import ActivityTracker
                ActivityTracker.track(
                    activity_type = 'datasheet_generated',
                    user          = request.user,
                    description   = f"Generated {equipment_type} datasheet ({title or ds.id})",
                    category      = 'electrical_datasheet',
                    severity      = 'low',
                    success       = True,
                    details       = {'datasheet_id': str(ds.id), 'equipment_type': equipment_type},
                    request       = request,
                )
            except Exception:
                pass
        except Exception as exc:
            logger.error(f"[smart_persist] non-fatal: {exc}", exc_info=True)

    # ── helpers ────────────────────────────────────────────────────────
    def _get_owned_datasheet(self, request, datasheet_id):
        """Fetch a `GeneratedDatasheet` owned by the current user, else None."""
        from .models import GeneratedDatasheet
        try:
            return GeneratedDatasheet.objects.get(id=datasheet_id, user=request.user)
        except GeneratedDatasheet.DoesNotExist:
            return None

    # ── HISTORY ────────────────────────────────────────────────────────
    @action(detail=False, methods=['get'], url_path='generated')
    def list_generated(self, request):
        """Paginated list of the current user's generated datasheets."""
        from .models import GeneratedDatasheet
        from .smart_storage import serialize_summary

        try:
            page  = max(1, int(request.query_params.get('page', 1)))
            limit = min(100, max(1, int(request.query_params.get('limit', 20))))
        except ValueError:
            page, limit = 1, 20

        qs = GeneratedDatasheet.objects.filter(user=request.user)

        equipment = request.query_params.get('equipment_type')
        if equipment:
            qs = qs.filter(equipment_type=equipment)

        if request.query_params.get('include_archived', '').lower() not in ('1', 'true', 'yes'):
            qs = qs.filter(is_archived=False)

        search = (request.query_params.get('search') or '').strip()
        if search:
            qs = qs.filter(Q(title__icontains=search) | Q(metadata__icontains=search))

        total = qs.count()
        offset = (page - 1) * limit
        items  = qs[offset:offset + limit]

        return Response({
            'success': True,
            'data': {
                'items': [serialize_summary(d) for d in items],
                'pagination': {
                    'page':        page,
                    'limit':       limit,
                    'total_count': total,
                    'total_pages': (total + limit - 1) // limit if limit else 0,
                },
            },
        })

    @action(detail=False, methods=['get'], url_path=r'generated/(?P<datasheet_id>[0-9a-f-]+)')
    def get_generated(self, request, datasheet_id=None):
        from .smart_storage import serialize_detail
        ds = self._get_owned_datasheet(request, datasheet_id)
        if not ds:
            return Response({'success': False, 'error': 'Not found'}, status=status.HTTP_404_NOT_FOUND)
        return Response({'success': True, 'data': serialize_detail(ds)})

    @action(detail=False, methods=['delete'], url_path=r'generated/(?P<datasheet_id>[0-9a-f-]+)/archive')
    def archive_generated(self, request, datasheet_id=None):
        ds = self._get_owned_datasheet(request, datasheet_id)
        if not ds:
            return Response({'success': False, 'error': 'Not found'}, status=status.HTTP_404_NOT_FOUND)
        ds.is_archived = True
        ds.status      = 'archived'
        ds.save(update_fields=['is_archived', 'status', 'updated_at'])
        return Response({'success': True})

    # ── INLINE CELL EDIT ──────────────────────────────────────────────
    @action(detail=False, methods=['patch'], url_path=r'generated/(?P<datasheet_id>[0-9a-f-]+)/cells')
    def update_cells(self, request, datasheet_id=None):
        """Batch update editable cells (vendor_data + rev only).

        Body: {"edits": [{"row_index": int, "column_key": str, "new_value": str}, ...]}
        """
        from .models import DatasheetCellEdit, CELL_EDIT_SOURCE_CHOICES
        from .smart_storage import EDITABLE_COLUMNS, EXCEL_REGEN_DELAY_S
        from .tasks import regenerate_excel_artifact

        ds = self._get_owned_datasheet(request, datasheet_id)
        if not ds:
            return Response({'success': False, 'error': 'Not found'}, status=status.HTTP_404_NOT_FOUND)

        edits = request.data.get('edits') or []
        if not isinstance(edits, list) or not edits:
            return Response({'success': False, 'error': 'No edits provided'}, status=status.HTTP_400_BAD_REQUEST)

        rows    = list(ds.rows or [])
        applied = []
        errors  = []
        source  = (request.data.get('source') or 'manual')
        if source not in dict(CELL_EDIT_SOURCE_CHOICES):
            source = 'manual'

        for edit in edits:
            try:
                row_idx = int(edit.get('row_index'))
                col_key = str(edit.get('column_key'))
                new_val = edit.get('new_value', '')
            except (TypeError, ValueError):
                errors.append({'edit': edit, 'reason': 'invalid_payload'})
                continue
            if col_key not in EDITABLE_COLUMNS:
                errors.append({'edit': edit, 'reason': 'column_not_editable'})
                continue
            if row_idx < 0 or row_idx >= len(rows):
                errors.append({'edit': edit, 'reason': 'row_out_of_range'})
                continue
            old_val = rows[row_idx].get(col_key, '')
            if str(old_val) == str(new_val):
                continue  # no-op
            rows[row_idx][col_key] = new_val
            DatasheetCellEdit.objects.create(
                datasheet  = ds,
                user       = request.user,
                row_index  = row_idx,
                column_key = col_key,
                old_value  = '' if old_val is None else str(old_val),
                new_value  = '' if new_val is None else str(new_val),
                source     = source,
            )
            applied.append({'row_index': row_idx, 'column_key': col_key})

        if applied:
            ds.rows = rows
            # Recompute lightweight summary fields
            summary = dict(ds.summary or {})
            data_rows = [r for r in rows if not r.get('is_section')]
            summary['completed_fields'] = sum(1 for r in data_rows if (r.get('vendor_data') or '').strip())
            summary['missing_fields']   = sum(1 for r in data_rows if not (r.get('vendor_data') or '').strip())
            ds.summary = summary
            ds.save(update_fields=['rows', 'summary', 'updated_at'])
            # Coalesced regen
            try:
                regenerate_excel_artifact.apply_async(args=[str(ds.id)], countdown=EXCEL_REGEN_DELAY_S)
            except Exception as exc:
                logger.warning(f"[update_cells] could not enqueue regen: {exc}")

        return Response({'success': True, 'applied': applied, 'errors': errors, 'summary': ds.summary})

    @action(detail=False, methods=['get'], url_path=r'generated/(?P<datasheet_id>[0-9a-f-]+)/cell-edits')
    def list_cell_edits(self, request, datasheet_id=None):
        ds = self._get_owned_datasheet(request, datasheet_id)
        if not ds:
            return Response({'success': False, 'error': 'Not found'}, status=status.HTTP_404_NOT_FOUND)
        edits = ds.cell_edits.select_related('user')[:500]
        return Response({
            'success': True,
            'edits': [{
                'id':         e.id,
                'row_index':  e.row_index,
                'column_key': e.column_key,
                'old_value':  e.old_value,
                'new_value':  e.new_value,
                'source':     e.source,
                'user':       e.user.get_full_name() if e.user else 'Unknown',
                'changed_at': e.changed_at.isoformat(),
            } for e in edits],
        })

    # ── RECHECK ───────────────────────────────────────────────────────
    @action(detail=False, methods=['post'], url_path=r'generated/(?P<datasheet_id>[0-9a-f-]+)/recheck')
    def recheck(self, request, datasheet_id=None):
        """Re-run AI extraction on saved source files. Returns a diff (no auto-apply)."""
        from .smart_storage import smart_storage
        ds = self._get_owned_datasheet(request, datasheet_id)
        if not ds:
            return Response({'success': False, 'error': 'Not found'}, status=status.HTTP_404_NOT_FOUND)

        sources = ds.source_files or []
        if not sources or not sources[0].get('s3_key'):
            return Response({'success': False, 'error': 'No source files stored to recheck'}, status=status.HTTP_400_BAD_REQUEST)

        primary = sources[0]
        tmp_path = smart_storage.download_to_tempfile(primary['s3_key'], suffix=os.path.splitext(primary.get('filename', ''))[1])
        if not tmp_path:
            return Response({'success': False, 'error': 'Could not retrieve source from storage'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        try:
            # Wrap the temp file as a Django-like file object
            class _Wrap:
                def __init__(self, path, name):
                    self._fh = open(path, 'rb')
                    self.name = name
                def read(self, *a, **k):  return self._fh.read(*a, **k)
                def seek(self, *a, **k):  return self._fh.seek(*a, **k)
                def close(self):          return self._fh.close()
            wrapped = _Wrap(tmp_path, primary.get('filename', 'source'))

            project_info = (ds.metadata or {}).get('project_info', {}) or {}
            try:
                if ds.equipment_type == 'transformer':
                    from .transformer_datasheet_generator import TransformerDatasheetGenerator
                    gen = TransformerDatasheetGenerator()
                    result = gen.generate_datasheet_from_sizing_calc(wrapped, project_info)
                elif ds.equipment_type == 'dg_set':
                    from .dg_set_datasheet_generator import DGSetDatasheetGenerator
                    gen = DGSetDatasheetGenerator()
                    result = gen.generate_datasheet_from_sizing_calc(wrapped, project_info)
                elif ds.equipment_type == 'mv_switchgear':
                    from .switchgear_datasheet_generator import SwitchgearDatasheetGenerator
                    gen = SwitchgearDatasheetGenerator()
                    result = gen.generate_datasheet_from_sld(wrapped, project_info)
                else:
                    return Response({'success': False, 'error': 'Unsupported equipment type'}, status=status.HTTP_400_BAD_REQUEST)
            finally:
                wrapped.close()

            if not result.get('success'):
                return Response(result, status=status.HTTP_400_BAD_REQUEST)

            new_rows = result.get('datasheet_rows', [])
            diff = []
            for i, (old, new) in enumerate(zip(ds.rows or [], new_rows)):
                for col in ('vendor_data', 'rev'):
                    o = (old.get(col) or '').strip()
                    n = (new.get(col) or '').strip()
                    if o != n:
                        diff.append({
                            'row_index':   i,
                            'description': old.get('description', ''),
                            'column_key':  col,
                            'current':     o,
                            'extracted':   n,
                        })
            return Response({
                'success':           True,
                'diff':              diff,
                'extracted_summary': result.get('summary', {}),
                'extracted_count':   len(new_rows),
            })
        finally:
            try:
                os.unlink(tmp_path)
            except Exception:
                pass

    # ── REVISIONS ─────────────────────────────────────────────────────
    @action(detail=False, methods=['get', 'post'], url_path=r'generated/(?P<datasheet_id>[0-9a-f-]+)/snapshots')
    def snapshots(self, request, datasheet_id=None):
        from .models import GeneratedDatasheetRevision
        ds = self._get_owned_datasheet(request, datasheet_id)
        if not ds:
            return Response({'success': False, 'error': 'Not found'}, status=status.HTTP_404_NOT_FOUND)

        if request.method == 'POST':
            count = ds.snapshots.count()
            label = request.data.get('label') or f"v{count + 1}"
            note  = request.data.get('note', '')
            snap  = GeneratedDatasheetRevision.objects.create(
                datasheet      = ds,
                revision_label = label,
                rows           = ds.rows or [],
                summary        = ds.summary or {},
                note           = note,
                created_by     = request.user,
            )
            return Response({'success': True, 'snapshot': {
                'id': snap.id, 'label': snap.revision_label, 'note': snap.note,
                'created_at': snap.created_at.isoformat(),
            }})

        snaps = ds.snapshots.all()[:200]
        return Response({'success': True, 'snapshots': [{
            'id':         s.id,
            'label':      s.revision_label,
            'note':       s.note,
            'created_by': s.created_by.get_full_name() if s.created_by else 'Unknown',
            'created_at': s.created_at.isoformat(),
        } for s in snaps]})

    @action(detail=False, methods=['get'], url_path=r'generated/(?P<datasheet_id>[0-9a-f-]+)/snapshots/(?P<snap_a>\d+)/compare/(?P<snap_b>\d+)')
    def compare_snapshots(self, request, datasheet_id=None, snap_a=None, snap_b=None):
        from .models import GeneratedDatasheetRevision
        ds = self._get_owned_datasheet(request, datasheet_id)
        if not ds:
            return Response({'success': False, 'error': 'Not found'}, status=status.HTTP_404_NOT_FOUND)
        try:
            a = GeneratedDatasheetRevision.objects.get(id=snap_a, datasheet=ds)
            b = GeneratedDatasheetRevision.objects.get(id=snap_b, datasheet=ds)
        except GeneratedDatasheetRevision.DoesNotExist:
            return Response({'success': False, 'error': 'Snapshot not found'}, status=status.HTTP_404_NOT_FOUND)
        diff = []
        for i, (ra, rb) in enumerate(zip(a.rows or [], b.rows or [])):
            for col in ('vendor_data', 'rev'):
                oa, ob = (ra.get(col) or ''), (rb.get(col) or '')
                if oa != ob:
                    diff.append({'row_index': i, 'description': ra.get('description', ''), 'column_key': col, 'a': oa, 'b': ob})
        return Response({'success': True, 'a_label': a.revision_label, 'b_label': b.revision_label, 'diff': diff})

    # ── COMMENTS ──────────────────────────────────────────────────────
    @action(detail=False, methods=['get', 'post'], url_path=r'generated/(?P<datasheet_id>[0-9a-f-]+)/comments')
    def cell_comments(self, request, datasheet_id=None):
        from .models import GeneratedDatasheetCellComment
        ds = self._get_owned_datasheet(request, datasheet_id)
        if not ds:
            return Response({'success': False, 'error': 'Not found'}, status=status.HTTP_404_NOT_FOUND)

        if request.method == 'POST':
            text = (request.data.get('text') or '').strip()
            if not text:
                return Response({'success': False, 'error': 'Comment text required'}, status=status.HTTP_400_BAD_REQUEST)
            ri = request.data.get('row_index')
            ck = request.data.get('column_key', '') or ''
            c  = GeneratedDatasheetCellComment.objects.create(
                datasheet  = ds,
                user       = request.user,
                row_index  = int(ri) if ri is not None else None,
                column_key = ck,
                text       = text,
            )
            return Response({'success': True, 'comment': {
                'id': c.id, 'row_index': c.row_index, 'column_key': c.column_key,
                'text': c.text, 'user': request.user.get_full_name() or request.user.username,
                'created_at': c.created_at.isoformat(),
            }})

        comments = ds.cell_comments.select_related('user').all()
        return Response({'success': True, 'comments': [{
            'id': c.id, 'row_index': c.row_index, 'column_key': c.column_key,
            'text': c.text, 'is_resolved': c.is_resolved,
            'user': c.user.get_full_name() if c.user else 'Unknown',
            'created_at': c.created_at.isoformat(),
        } for c in comments]})

    # ── AI SUGGEST ────────────────────────────────────────────────────
    @action(detail=False, methods=['post'], url_path=r'generated/(?P<datasheet_id>[0-9a-f-]+)/suggest')
    def suggest_cell(self, request, datasheet_id=None):
        """Use AI to propose a value for a single cell, grounded in source text."""
        from .smart_storage import smart_storage
        from .document_extractor import extract_text
        from openai import OpenAI

        ds = self._get_owned_datasheet(request, datasheet_id)
        if not ds:
            return Response({'success': False, 'error': 'Not found'}, status=status.HTTP_404_NOT_FOUND)
        try:
            row_idx = int(request.data.get('row_index'))
        except (TypeError, ValueError):
            return Response({'success': False, 'error': 'row_index required'}, status=status.HTTP_400_BAD_REQUEST)
        rows = ds.rows or []
        if row_idx < 0 or row_idx >= len(rows):
            return Response({'success': False, 'error': 'row_index out of range'}, status=status.HTTP_400_BAD_REQUEST)
        row = rows[row_idx]

        sources = ds.source_files or []
        if not sources or not sources[0].get('s3_key'):
            return Response({'success': False, 'error': 'No source available'}, status=status.HTTP_400_BAD_REQUEST)
        tmp_path = smart_storage.download_to_tempfile(sources[0]['s3_key'])
        if not tmp_path:
            return Response({'success': False, 'error': 'Could not load source'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        try:
            class _Wrap:
                def __init__(self, p, n):
                    self._fh = open(p, 'rb'); self.name = n
                def read(self, *a, **k): return self._fh.read(*a, **k)
                def seek(self, *a, **k): return self._fh.seek(*a, **k)
                def close(self): self._fh.close()
            w = _Wrap(tmp_path, sources[0].get('filename', 'source'))
            try:
                doc_text = extract_text(w)
            finally:
                w.close()
        finally:
            try: os.unlink(tmp_path)
            except Exception: pass

        client = OpenAI(api_key=settings.OPENAI_API_KEY)
        prompt = (
            f"You are filling a single cell on an electrical datasheet.\n"
            f"Field: {row.get('description', '')}\n"
            f"Unit: {row.get('unit', '')}\n"
            f"Specified design data (for reference): {row.get('required_data', '')}\n\n"
            f"Source document excerpt (truncated):\n{(doc_text or '')[:6000]}\n\n"
            "Reply STRICTLY as JSON: {\"suggestion\": \"<value>\", \"confidence\": \"high|medium|low\", "
            "\"source_excerpt\": \"<short quote from source or empty>\"}. "
            "If the document does not contain enough info, return suggestion as empty string and confidence \"low\"."
        )
        try:
            resp = client.chat.completions.create(
                model='gpt-4o-mini',
                temperature=0.1,
                max_tokens=400,
                response_format={'type': 'json_object'},
                messages=[{'role': 'user', 'content': prompt}],
            )
            payload = json.loads(resp.choices[0].message.content)
        except Exception as exc:
            logger.error(f"[suggest_cell] AI failed: {exc}", exc_info=True)
            return Response({'success': False, 'error': 'AI suggestion failed'}, status=status.HTTP_502_BAD_GATEWAY)
        return Response({'success': True, **payload})

    # ── SHARE LINK ────────────────────────────────────────────────────
    @action(detail=False, methods=['post'], url_path=r'generated/(?P<datasheet_id>[0-9a-f-]+)/share')
    def create_share(self, request, datasheet_id=None):
        from .models import DatasheetShareLink
        from .smart_storage import SHARE_LINK_TTL_DAYS
        ds = self._get_owned_datasheet(request, datasheet_id)
        if not ds:
            return Response({'success': False, 'error': 'Not found'}, status=status.HTTP_404_NOT_FOUND)
        link = DatasheetShareLink.objects.create(
            datasheet  = ds,
            created_by = request.user,
            expires_at = timezone.now() + timezone.timedelta(days=SHARE_LINK_TTL_DAYS),
        )
        return Response({
            'success': True,
            'token':   link.token,
            'expires_at': link.expires_at.isoformat() if link.expires_at else None,
            'share_path': f'/share/datasheet/{link.token}',
        })

    # ── PDF EXPORT ────────────────────────────────────────────────────
    @action(detail=False, methods=['get'], url_path=r'generated/(?P<datasheet_id>[0-9a-f-]+)/excel')
    def download_generated_excel(self, request, datasheet_id=None):
        """Stream the latest Excel artifact for a datasheet."""
        from django.http import HttpResponse
        from .smart_storage import smart_storage
        ds = self._get_owned_datasheet(request, datasheet_id)
        if not ds:
            return Response({'success': False, 'error': 'Not found'}, status=status.HTTP_404_NOT_FOUND)
        # Re-render fresh from current rows if no S3
        if ds.excel_s3_key:
            data = smart_storage.download_to_bytes(ds.excel_s3_key)
        else:
            data = None
        if not data:
            try:
                from .tasks import _generator_for
                gen = _generator_for(ds.equipment_type)
                buf = gen.export_to_excel(ds.rows or [], (ds.metadata or {}).get('project_info', {}) or {})
                data = buf.getvalue() if hasattr(buf, 'getvalue') else bytes(buf)
            except Exception as exc:
                return Response({'success': False, 'error': f'Render failed: {exc}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        resp = HttpResponse(data, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        fname = (ds.title or f"datasheet_{ds.id}").replace(' ', '_') + '.xlsx'
        resp['Content-Disposition'] = f'attachment; filename="{fname}"'
        resp['Access-Control-Expose-Headers'] = 'Content-Disposition'
        return resp


class DatasheetCommentViewSet(viewsets.ModelViewSet):
    """ViewSet for managing datasheet comments"""
    queryset = DatasheetComment.objects.all()
    serializer_class = DatasheetCommentSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['datasheet', 'is_resolved']

    def perform_create(self, serializer):
        """Create comment with user tracking"""
        serializer.save(commented_by=self.request.user)

    @action(detail=True, methods=['post'])
    def resolve(self, request, pk=None):
        """Mark comment as resolved"""
        comment = self.get_object()
        comment.is_resolved = True
        comment.save()
        
        return Response({
            'message': 'Comment marked as resolved',
            'is_resolved': comment.is_resolved
        })

    @action(detail=True, methods=['post'])
    def unresolve(self, request, pk=None):
        """Mark comment as unresolved"""
        comment = self.get_object()
        comment.is_resolved = False
        comment.save()
        
        return Response({
            'message': 'Comment marked as unresolved',
            'is_resolved': comment.is_resolved
        })


class SmartSLDUploadViewSet(viewsets.ViewSet):
    """
    Smart SLD Upload and Processing Endpoint
    Uses Cost-Optimized Hybrid AI (PaddleOCR + GPT-3.5-turbo) for extraction
    
    COST OPTIMIZATION:
    - Old: GPT-4o Vision = ~$75 per 1000 pages
    - New: Hybrid approach = ~$0.50 per 1000 pages  
    - Savings: 99% cost reduction!
    
    TRANSFORMER DOCUMENT UPLOAD:
    For Transformer (Power and Distribution) equipment:
    - Instead of SLD files, upload specific calculation/criteria documents
    - Supported document types:
      1. MV Trafo Calculation (mv_trafo_calculation)
      2. Criteria (criteria)
      3. Formula (formula)
      4. LV Trafo Calculation (lv_trafo_calculation)
    - Set equipment_type='transformer' in request data
    - Each file should have doc_type_<filename>=<document_type> in request data
    
    Frontend Usage:
    ```javascript
    // For transformers, send equipment_type parameter
    const formData = new FormData();
    formData.append('equipment_type', 'transformer');
    formData.append('files', file1);
    formData.append('files', file2);
    formData.append('doc_type_' + file1.name, 'mv_trafo_calculation');
    formData.append('doc_type_' + file2.name, 'lv_trafo_calculation');
    
    // Regular SLD processing (non-transformer equipment)
    const formData = new FormData();
    formData.append('files', sldFile);
    formData.append('datasheet_transformer', 'true');
    ```
    """
    permission_classes = [IsAuthenticated]
    
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        from .sld_smart_orchestrator import SmartSLDOrchestrator
        from .cost_recommendation_system import CostRecommendationSystem
        self.orchestrator = None
        self.cost_system = CostRecommendationSystem
    
    def _process_transformer_documents(self, request, uploaded_files, project_info):
        """
        Process equipment-specific documents (Transformer, EDG, Switchgear)
        instead of SLD files for specific equipment types.
        """
        try:
            import uuid
            from .s3_service import ElectricalDatasheetS3Service
            
            equipment_type = request.data.get('equipment_type', 'transformer')
            job_id = str(uuid.uuid4())
            
            equipment_labels = {
                'edg': 'Emergency Diesel Generator',
                'switchgear_11kv': '11kV Switchgear',
                'transformer': 'Transformer'
            }
            
            logger.info(f"[EquipmentDocs] Processing {len(uploaded_files)} {equipment_labels.get(equipment_type, equipment_type)} documents, job_id: {job_id}")
            
            s3_service = ElectricalDatasheetS3Service()
            uploaded_docs = []
            
            # Process each uploaded file and categorize by document type
            for file in uploaded_files:
                doc_type = request.data.get(f'doc_type_{file.name}', 'general')
                
                # Valid document types by equipment
                valid_types = {
                    'edg': ['edg_load_list', 'dg_calculation'],
                    'switchgear_11kv': ['switchgear_sld', 'switchgear_schedule'],
                    'transformer': ['mv_trafo_calculation', 'criteria', 'formula', 'lv_trafo_calculation']
                }
                
                if equipment_type in valid_types and doc_type not in valid_types[equipment_type]:
                    doc_type = 'general'
                
                # Upload to S3/local storage
                try:
                    upload_result = s3_service.upload_datasheet(
                        file_obj=file,
                        filename=file.name,
                        equipment_type=equipment_type,
                        metadata={
                            'document_type': doc_type,
                            'project_info': project_info,
                            'job_id': job_id,
                            'uploaded_by': request.user.username
                        }
                    )
                    
                    uploaded_docs.append({
                        'filename': file.name,
                        'doc_type': doc_type,
                        's3_key': upload_result.get('s3_key') or upload_result.get('local_path'),
                        'size': file.size,
                        'content_type': file.content_type
                    })
                    
                except Exception as e:
                    logger.error(f"[EquipmentDocs] Error uploading {file.name}: {e}")
                    return Response(
                        {'error': f'Failed to upload {file.name}: {str(e)}'},
                        status=status.HTTP_500_INTERNAL_SERVER_ERROR
                    )
            
            # Return success response with uploaded document details
            response_data = {
                'success': True,
                'job_id': job_id,
                'equipment_type': equipment_type,
                'documents_uploaded': len(uploaded_docs),
                'documents': uploaded_docs,
                'project_info': project_info,
                'message': f'Successfully uploaded {len(uploaded_docs)} {equipment_labels.get(equipment_type, equipment_type)} documents'
            }
            
            logger.info(f"[EquipmentDocs] ✅ Success: {len(uploaded_docs)} documents uploaded for job {job_id}")
            return Response(response_data, status=status.HTTP_200_OK)
            
        except Exception as e:
            logger.error(f"[EquipmentDocs] ❌ Error: {e}", exc_info=True)
            return Response(
                {'error': f'Processing failed: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=False, methods=['post'], url_path='process')
    def process_sld(self, request):
        """
        Process uploaded SLD files and extract equipment using AI
        
        Request Body (multipart/form-data):
        - files: SLD files (PDF, PNG, JPG)
        - drawing_number: Drawing number
        - drawing_title: Drawing title
        - revision: Revision number
        - voltage_level: Voltage level (11KV, 33KV, etc.)
        - project_name: Project name
        - area: Area/location
        - datasheet_transformer: Boolean (include transformer datasheets)
        - datasheet_diesel_generator: Boolean (include diesel generator datasheets)
        - datasheet_switchgear_11kv: Boolean (include 11KV switchgear datasheets)
        - analysis_extract_tags: Boolean (extract equipment tags)
        - analysis_detect_types: Boolean (detect equipment types)
        - analysis_extract_specs: Boolean (extract specifications)
        - analysis_generate_datasheets: Boolean (auto-generate datasheets)
        - analysis_identify_connections: Boolean (identify electrical connections)
        
        Returns:
        - job_id: Unique job identifier
        - equipment_extracted: List of extracted equipment
        - datasheets_generated: Generated datasheet objects
        """
        try:
            from .sld_smart_orchestrator import SmartSLDOrchestrator
            
            # Get uploaded files
            uploaded_files = request.FILES.getlist('files')
            if not uploaded_files:
                return Response(
                    {'error': 'No files uploaded'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Extract project information
            project_info = {
                'drawing_number': request.data.get('drawing_number', ''),
                'drawing_title': request.data.get('drawing_title', ''),
                'revision': request.data.get('revision', 'A'),
                'voltage_level': request.data.get('voltage_level', ''),
                'project_name': request.data.get('project_name', ''),
                'area': request.data.get('area', '')
            }
            
            # Extract datasheet selection
            datasheet_selection = {
                'transformer': request.data.get('datasheet_transformer', 'false').lower() == 'true',
                'diesel_generator': request.data.get('datasheet_diesel_generator', 'false').lower() == 'true',
                'switchgear_11kv': request.data.get('datasheet_switchgear_11kv', 'false').lower() == 'true'
            }
            
            # Check if this is a equipment-specific document upload (not SLD)
            equipment_type = request.data.get('equipment_type', '')
            if equipment_type in ['edg', 'switchgear_11kv', 'transformer']:
                return self._process_transformer_documents(request, uploaded_files, project_info)

            
            # Extract analysis options
            analysis_options = {
                'extract_tags': request.data.get('analysis_extract_tags', 'true').lower() == 'true',
                'detect_equipment_types': request.data.get('analysis_detect_types', 'true').lower() == 'true',
                'extract_specifications': request.data.get('analysis_extract_specs', 'true').lower() == 'true',
                'generate_datasheets': request.data.get('analysis_generate_datasheets', 'true').lower() == 'true',
                'identify_connections': request.data.get('analysis_identify_connections', 'false').lower() == 'true'
            }
            
            logger.info(f"[SmartSLDUpload] Processing {len(uploaded_files)} files for user {request.user.username}")
            logger.info(f"[SmartSLDUpload] Datasheet selection: {datasheet_selection}")
            
            # Create orchestrator and process
            orchestrator = SmartSLDOrchestrator()
            result = orchestrator.process_smart_sld(
                uploaded_files=uploaded_files,
                project_info=project_info,
                datasheet_selection=datasheet_selection,
                analysis_options=analysis_options
            )
            
            if not result.get('success'):
                return Response(
                    {'error': result.get('error', 'Processing failed')},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Return results
            response_data = {
                'success': True,
                'job_id': result['job_id'],
                'files_processed': result['files_processed'],
                'equipment_extracted': result['equipment_extracted'],
                'equipment_by_type': result['equipment_by_type'],
                'datasheets_generated': result['datasheets_generated'],
                'extraction_method': result['extraction_method'],
                'confidence': result['confidence'],
                'cost_total': result.get('total_cost', 0.0),
                'cost_breakdown': result.get('cost_breakdown', 'N/A'),
                'message': f"Successfully extracted {result['equipment_extracted']} equipment items from {result['files_processed']} SLD files"
            }
            
            # Track usage
            if result.get('total_cost', 0) > 0:
                self.cost_system.track_usage(
                    job_id=result['job_id'],
                    method=result['extraction_method'],
                    pages_processed=sum([r.get('pages_processed', 0) for r in result.get('detailed_results', [])]),
                    cost=result.get('total_cost', 0.0)
                )
            
            logger.info(f"[SmartSLDUpload] ✅ Success: {result['equipment_extracted']} equipment, cost: ${result.get('total_cost', 0):.4f}")
            
            return Response(response_data, status=status.HTTP_200_OK)
            
        except Exception as e:
            logger.error(f"[SmartSLDUpload] ❌ Error: {e}", exc_info=True)
            return Response(
                {'error': f'Processing failed: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=False, methods=['get'], url_path='job-status/(?P<job_id>[^/.]+)')
    def job_status(self, request, job_id=None):
        """
        Get status of an SLD processing job
        """
        try:
            from .sld_smart_orchestrator import SmartSLDOrchestrator
            orchestrator = SmartSLDOrchestrator()
            job_info = orchestrator.get_job_status(job_id)
            
            if job_info is None:
                return Response(
                    {'error': 'Job not found'},
                    status=status.HTTP_404_NOT_FOUND
                )
            
            return Response(job_info, status=status.HTTP_200_OK)
            
        except Exception as e:
            logger.error(f"[SmartSLDUpload] Error getting job status: {e}")
            return Response(
                {'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=False, methods=['get'], url_path='cost-recommendations')
    def cost_recommendations(self, request):
        """
        Get cost recommendations for different extraction strategies
        
        Query params:
        - num_pages: Estimated number of pages (default: 100)
        """
        try:
            num_pages = int(request.query_params.get('num_pages', 100))
            recommendations = self.cost_system.get_recommendations(num_pages)
            
            return Response(recommendations, status=status.HTTP_200_OK)
            
        except Exception as e:
            logger.error(f"[SmartSLDUpload] Error getting recommendations: {e}")
            return Response(
                {'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=False, methods=['get'], url_path='daily-usage')
    def daily_usage(self, request):
        """
        Get today's usage statistics and costs
        """
        try:
            usage = self.cost_system.get_daily_usage()
            return Response(usage, status=status.HTTP_200_OK)
            
        except Exception as e:
            logger.error(f"[SmartSLDUpload] Error getting usage: {e}")
            return Response(
                {'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=False, methods=['post'], url_path='estimate-project-cost')
    def estimate_project_cost(self, request):
        """
        Estimate cost for processing a project
        
        Body:
        - num_drawings: Number of SLD drawings
        - pages_per_drawing: Average pages per drawing (default: 5)
        """
        try:
            num_drawings = int(request.data.get('num_drawings', 10))
            pages_per_drawing = int(request.data.get('pages_per_drawing', 5))
            
            estimate = self.cost_system.estimate_project_cost(num_drawings, pages_per_drawing)
            
            return Response(estimate, status=status.HTTP_200_OK)
            
        except Exception as e:
            logger.error(f"[SmartSLDUpload] Error estimating cost: {e}")
            return Response(
                {'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=False, methods=['get'], url_path='cost-comparison-chart')
    def cost_comparison_chart(self, request):
        """
        Get data for cost comparison visualization
        """
        try:
            chart_data = self.cost_system.get_cost_comparison_chart()
            return Response(chart_data, status=status.HTTP_200_OK)
            
        except Exception as e:
            logger.error(f"[SmartSLDUpload] Error getting chart data: {e}")
            return Response(
                {'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
            orchestrator = SmartSLDOrchestrator()
            status_data = orchestrator.get_job_status(job_id)
            
            return Response(status_data, status=status.HTTP_200_OK)
            
        except Exception as e:
            logger.error(f"[SmartSLDUpload] Error getting job status: {e}")
            return Response(
                {'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

