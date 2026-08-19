"""
Instrument Tools — Soft-coded rule engines and table generators.

This module implements three independent Generator + QC services:
  • IO List
  • Cable Block Diagram
  • Cable Schedule

Design principles
─────────────────
1. **Stateless, pure-function services.** No models, no Celery. Each public
   helper accepts already-parsed rows (list[dict]) and returns a normalised
   result dict. Persistence is deliberately out of scope so this never
   collides with the existing Instrument Index core logic.

2. **Soft-coded configuration.** Every threshold, keyword, header alias and
   rule lives in a module-level constant — never inline. Tuning the rule
   engine never requires editing function bodies.

3. **Forgiving header detection.** Customer spreadsheets rarely use canonical
   headers. The `_normalise_headers()` helper maps a wide variety of common
   aliases onto a single canonical schema per tool.
"""
from __future__ import annotations

import io
import logging
import re
from typing import Iterable, Optional

logger = logging.getLogger(__name__)

# ─── Tool identifiers (single source of truth) ───────────────────────────────
TOOL_IO_LIST       = 'io_list'
TOOL_CABLE_BLOCK   = 'cable_block_diagram'
TOOL_CABLE_SCHED   = 'cable_schedule'
SUPPORTED_TOOLS    = (TOOL_IO_LIST, TOOL_CABLE_BLOCK, TOOL_CABLE_SCHED)

# ─── Modes ───────────────────────────────────────────────────────────────────
MODE_GENERATE = 'generate'
MODE_QC       = 'qc'
SUPPORTED_MODES = (MODE_GENERATE, MODE_QC)

# ─── Soft-coded canonical schemas (one per tool) ─────────────────────────────
# Each entry: canonical_key -> list of aliases (case/space/punct insensitive).
# Adding a new alias does NOT require code changes elsewhere.
_IO_LIST_SCHEMA = {
    'tag':           ['tag', 'tag no', 'tag number', 'instrument tag', 'instrument', 'item tag'],
    'description':   ['description', 'service', 'instrument description', 'service description'],
    'signal_type':   ['signal type', 'io type', 'i/o type', 'io', 'i/o', 'signal'],
    'pid':           ['pid', 'p&id', 'pandid', 'p and id', 'pid no', 'p&id no', 'drawing'],
    'location':      ['location', 'area', 'unit', 'plant'],
    'panel':         ['panel', 'marshalling panel', 'mp', 'cabinet', 'rack'],
    'range':         ['range', 'measuring range', 'operating range'],
    'units':         ['units', 'eng units', 'engineering units', 'uom'],
    'manufacturer':  ['manufacturer', 'make', 'maker', 'vendor'],
    'model':         ['model', 'model number', 'type number'],
}

_CABLE_BLOCK_SCHEMA = {
    'system':        ['system', 'sub-system', 'subsystem'],
    'source':        ['source', 'from', 'from equipment', 'origin'],
    'destination':   ['destination', 'to', 'to equipment', 'target'],
    'cable_type':    ['cable type', 'type', 'cable category', 'category'],
    'function':      ['function', 'service', 'purpose', 'signal type'],
    'qty':           ['qty', 'quantity', 'count', 'no'],
}

_CABLE_SCHEDULE_SCHEMA = {
    'cable_tag':     ['cable tag', 'cable no', 'cable number', 'tag', 'cable id'],
    'from_tag':      ['from tag', 'from', 'origin tag', 'source tag'],
    'to_tag':        ['to tag', 'to', 'destination tag', 'target tag'],
    'cable_type':    ['cable type', 'type', 'spec'],
    'size':          ['size', 'cable size', 'csa', 'cross section'],
    'cores':         ['cores', 'no of cores', 'number of cores', 'no cores'],
    'length_m':      ['length', 'length (m)', 'length m', 'cable length', 'length_m'],
    'voltage':       ['voltage', 'rated voltage', 'voltage rating'],
    'from_panel':    ['from panel', 'origin panel', 'source panel'],
    'to_panel':      ['to panel', 'destination panel', 'target panel'],
    'gland_from':    ['gland from', 'from gland'],
    'gland_to':      ['gland to', 'to gland'],
    'tray':          ['tray', 'cable tray', 'route'],
}

_TOOL_SCHEMAS = {
    TOOL_IO_LIST:     _IO_LIST_SCHEMA,
    TOOL_CABLE_BLOCK: _CABLE_BLOCK_SCHEMA,
    TOOL_CABLE_SCHED: _CABLE_SCHEDULE_SCHEMA,
}

# ─── Soft-coded validation rules (per tool) ──────────────────────────────────
# Each rule is a dict consumed by the generic checker:
#   field    – canonical field key
#   kind     – 'required' | 'unique' | 'enum' | 'regex' | 'numeric_positive'
#   severity – 'error' | 'warning' | 'info'
#   message  – human readable
#   params   – kind-specific (enum -> values, regex -> pattern, etc.)
_IO_SIGNAL_TYPES = {
    'AI', 'AO', 'DI', 'DO', 'RTD', 'TC', 'HART', 'PULSE', 'SOE',
    'FF', 'PROFIBUS', 'MODBUS',
}

_IO_LIST_RULES = [
    {'field': 'tag',         'kind': 'required',         'severity': 'error',
     'message': 'Tag is required for every IO point.'},
    {'field': 'tag',         'kind': 'unique',           'severity': 'error',
     'message': 'Duplicate tag detected — every IO point must have a unique tag.'},
    {'field': 'description', 'kind': 'required',         'severity': 'warning',
     'message': 'Description is missing — recommended for traceability.'},
    {'field': 'signal_type', 'kind': 'required',         'severity': 'error',
     'message': 'Signal type (AI/AO/DI/DO/…) is required.'},
    {'field': 'signal_type', 'kind': 'enum',             'severity': 'warning',
     'message': 'Unrecognised signal type — confirm against the project IO convention.',
     'params': {'values': sorted(_IO_SIGNAL_TYPES)}},
    {'field': 'pid',         'kind': 'required',         'severity': 'warning',
     'message': 'P&ID reference is missing — required for design traceability.'},
]

_CABLE_BLOCK_RULES = [
    {'field': 'source',      'kind': 'required',         'severity': 'error',
     'message': 'Source (From) is required.'},
    {'field': 'destination', 'kind': 'required',         'severity': 'error',
     'message': 'Destination (To) is required.'},
    {'field': 'cable_type',  'kind': 'required',         'severity': 'warning',
     'message': 'Cable type is missing — needed for the block diagram legend.'},
    {'field': 'qty',         'kind': 'numeric_positive', 'severity': 'warning',
     'message': 'Quantity should be a positive number.'},
]

_CABLE_SCHEDULE_RULES = [
    {'field': 'cable_tag',   'kind': 'required',         'severity': 'error',
     'message': 'Cable tag is required.'},
    {'field': 'cable_tag',   'kind': 'unique',           'severity': 'error',
     'message': 'Duplicate cable tag — every cable must have a unique identifier.'},
    {'field': 'from_tag',    'kind': 'required',         'severity': 'error',
     'message': 'From tag is required.'},
    {'field': 'to_tag',      'kind': 'required',         'severity': 'error',
     'message': 'To tag is required.'},
    {'field': 'cable_type',  'kind': 'required',         'severity': 'error',
     'message': 'Cable type / spec is required.'},
    {'field': 'cores',       'kind': 'numeric_positive', 'severity': 'warning',
     'message': 'Number of cores should be a positive integer.'},
    {'field': 'length_m',    'kind': 'numeric_positive', 'severity': 'warning',
     'message': 'Cable length (m) should be a positive number.'},
]

_TOOL_RULES = {
    TOOL_IO_LIST:     _IO_LIST_RULES,
    TOOL_CABLE_BLOCK: _CABLE_BLOCK_RULES,
    TOOL_CABLE_SCHED: _CABLE_SCHEDULE_RULES,
}

# ─── Soft-coded mapping signal_type -> default cable type (for generator) ────
_SIGNAL_TO_CABLE = {
    'AI':       '1Pr x 0.5mm² ISOS',
    'AO':       '1Pr x 0.5mm² ISOS',
    'DI':       '1Pr x 0.75mm² ISOS',
    'DO':       '1Pr x 0.75mm² ISOS',
    'RTD':      '3C x 0.5mm² ISOS',
    'TC':       '1Pr x 0.5mm² Compensating',
    'HART':     '1Pr x 0.5mm² ISOS (HART)',
    'PULSE':    '1Pr x 0.75mm² ISOS',
    'FF':       'FF H1 — Type A',
    'PROFIBUS': 'Profibus PA — Type A',
    'MODBUS':   '2Pr x 0.5mm² Shielded',
}

# Soft-coded default voltage rating for instrument cables (V).
_INSTRUMENT_CABLE_VOLTAGE = '300/500V'

# Soft-coded prefix used when generating cable tags from IO tag.
_DEFAULT_CABLE_TAG_PREFIX = 'C-'

# Soft-coded fallback panel name when the IO list does not declare one.
_DEFAULT_PANEL = 'MP-01'

# Soft-coded default field cable length (m) when nothing is known.
_DEFAULT_FIELD_LENGTH_M = 50


# ─── Header normaliser ───────────────────────────────────────────────────────
def _slug(s) -> str:
    """Aggressive normaliser for header / enum comparisons."""
    return re.sub(r'[^a-z0-9]', '', str(s or '').strip().lower())


def _build_alias_index(schema: dict) -> dict:
    """Pre-compute slug → canonical mapping for fast lookups."""
    index: dict = {}
    for canonical, aliases in schema.items():
        index[_slug(canonical)] = canonical
        for a in aliases:
            index[_slug(a)] = canonical
    return index


_TOOL_ALIAS_INDEX = {tool: _build_alias_index(schema) for tool, schema in _TOOL_SCHEMAS.items()}


def _normalise_rows(tool: str, rows: list[dict]) -> list[dict]:
    """Rewrite each row to use canonical keys (per the tool's schema)."""
    if tool not in _TOOL_ALIAS_INDEX:
        raise ValueError(f'Unsupported tool: {tool!r}')
    idx = _TOOL_ALIAS_INDEX[tool]
    out: list[dict] = []
    for raw in rows or []:
        if not isinstance(raw, dict):
            continue
        norm: dict = {}
        for k, v in raw.items():
            canonical = idx.get(_slug(k))
            if canonical and (canonical not in norm or _is_empty(norm[canonical])):
                norm[canonical] = v
        out.append(norm)
    return out


def _is_empty(v) -> bool:
    if v is None:
        return True
    if isinstance(v, str):
        return v.strip() == ''
    return False


def _coerce_float(v) -> Optional[float]:
    try:
        if v is None or (isinstance(v, str) and not v.strip()):
            return None
        return float(str(v).replace(',', '').strip())
    except (TypeError, ValueError):
        return None


# ─── Generic rule runner ─────────────────────────────────────────────────────
def _run_rules(tool: str, rows: list[dict]) -> list[dict]:
    """Apply the soft-coded rule set to normalised rows.

    Returns a flat list of issue dicts:
      { row, field, kind, severity, message, value? }
    """
    rules = _TOOL_RULES.get(tool, [])
    issues: list[dict] = []
    seen_by_field: dict[str, dict[str, int]] = {}

    for idx, row in enumerate(rows, start=1):
        for rule in rules:
            field = rule['field']
            kind  = rule['kind']
            val   = row.get(field)

            if kind == 'required' and _is_empty(val):
                issues.append({
                    'row':      idx, 'field': field, 'kind': kind,
                    'severity': rule['severity'], 'message': rule['message'],
                })
                continue

            if kind == 'unique' and not _is_empty(val):
                bucket = seen_by_field.setdefault(field, {})
                key = str(val).strip().lower()
                first_seen = bucket.get(key)
                if first_seen:
                    issues.append({
                        'row':      idx, 'field': field, 'kind': kind,
                        'severity': rule['severity'],
                        'message':  rule['message'],
                        'value':    val,
                        'first_seen_row': first_seen,
                    })
                else:
                    bucket[key] = idx

            elif kind == 'enum' and not _is_empty(val):
                allowed = {_slug(x) for x in (rule.get('params') or {}).get('values', [])}
                if _slug(val) not in allowed:
                    issues.append({
                        'row':      idx, 'field': field, 'kind': kind,
                        'severity': rule['severity'],
                        'message':  rule['message'],
                        'value':    val,
                    })

            elif kind == 'regex' and not _is_empty(val):
                pattern = (rule.get('params') or {}).get('pattern')
                if pattern and not re.search(pattern, str(val)):
                    issues.append({
                        'row':      idx, 'field': field, 'kind': kind,
                        'severity': rule['severity'],
                        'message':  rule['message'],
                        'value':    val,
                    })

            elif kind == 'numeric_positive' and not _is_empty(val):
                num = _coerce_float(val)
                if num is None or num <= 0:
                    issues.append({
                        'row':      idx, 'field': field, 'kind': kind,
                        'severity': rule['severity'],
                        'message':  rule['message'],
                        'value':    val,
                    })

    return issues


def _summarise(issues: list[dict], total_rows: int) -> dict:
    by_sev = {'error': 0, 'warning': 0, 'info': 0}
    for it in issues:
        by_sev[it['severity']] = by_sev.get(it['severity'], 0) + 1
    return {
        'total_rows': total_rows,
        'errors':     by_sev['error'],
        'warnings':   by_sev['warning'],
        'info':       by_sev['info'],
        'pass':       by_sev['error'] == 0,
    }


# ─── Public API: QC ──────────────────────────────────────────────────────────
def run_qc(tool: str, rows: list[dict]) -> dict:
    """Validate the supplied rows against the tool's rule set."""
    if tool not in SUPPORTED_TOOLS:
        raise ValueError(f'Unsupported tool: {tool!r}')
    normalised = _normalise_rows(tool, rows)
    issues     = _run_rules(tool, normalised)
    return {
        'tool':       tool,
        'mode':       MODE_QC,
        'summary':    _summarise(issues, len(normalised)),
        'issues':     issues,
        'normalised': normalised,
    }


# ─── Public API: Generators ──────────────────────────────────────────────────
def generate_io_list_from_instruments(rows: list[dict]) -> list[dict]:
    """Build a canonical IO list from a free-form instrument register.

    Input rows may already be IO-shaped (tag, signal_type, …) — in that case
    we just normalise + dedupe by tag. Missing signal types are left blank
    so QC flags them rather than the generator silently inventing values.
    """
    normalised = _normalise_rows(TOOL_IO_LIST, rows)
    seen: set[str] = set()
    out: list[dict] = []
    for row in normalised:
        tag = (row.get('tag') or '').strip()
        if not tag:
            continue
        key = tag.lower()
        if key in seen:
            continue
        seen.add(key)
        out.append({
            'tag':         tag,
            'description': row.get('description') or '',
            'signal_type': (row.get('signal_type') or '').strip().upper(),
            'pid':         row.get('pid') or '',
            'location':    row.get('location') or '',
            'panel':       row.get('panel') or _DEFAULT_PANEL,
            'range':       row.get('range') or '',
            'units':       row.get('units') or '',
            'manufacturer': row.get('manufacturer') or '',
            'model':       row.get('model') or '',
        })
    return out


def generate_cable_block_from_io(io_rows: list[dict]) -> list[dict]:
    """Aggregate IO rows into cable block bundles, grouped by signal type and panel."""
    normalised = _normalise_rows(TOOL_IO_LIST, io_rows)
    buckets: dict[tuple, dict] = {}
    for row in normalised:
        sig   = (row.get('signal_type') or '').strip().upper() or 'UNCLASSIFIED'
        panel = (row.get('panel') or _DEFAULT_PANEL).strip()
        loc   = (row.get('location') or 'FIELD').strip()
        key   = (sig, panel, loc)
        if key not in buckets:
            buckets[key] = {
                'system':      f'{sig} – {loc}',
                'source':      f'FIELD / {loc}',
                'destination': panel,
                'cable_type':  _SIGNAL_TO_CABLE.get(sig, 'Multi-pair instrument cable'),
                'function':    sig,
                'qty':         0,
            }
        buckets[key]['qty'] += 1
    return list(buckets.values())


def generate_cable_schedule_from_io(io_rows: list[dict]) -> list[dict]:
    """Synthesise a per-cable schedule (one row per IO point)."""
    normalised = _normalise_rows(TOOL_IO_LIST, io_rows)
    out: list[dict] = []
    for row in normalised:
        tag = (row.get('tag') or '').strip()
        if not tag:
            continue
        sig    = (row.get('signal_type') or '').strip().upper()
        panel  = (row.get('panel') or _DEFAULT_PANEL).strip()
        out.append({
            'cable_tag':  f'{_DEFAULT_CABLE_TAG_PREFIX}{tag}',
            'from_tag':   tag,
            'to_tag':     panel,
            'cable_type': _SIGNAL_TO_CABLE.get(sig, 'Multi-pair instrument cable'),
            'size':       '0.5mm²',
            'cores':      2 if sig in ('AI', 'AO', 'HART', 'TC') else 3,
            'length_m':   _DEFAULT_FIELD_LENGTH_M,
            'voltage':    _INSTRUMENT_CABLE_VOLTAGE,
            'from_panel': '',
            'to_panel':   panel,
            'gland_from': '',
            'gland_to':   '',
            'tray':       '',
        })
    return out


_GENERATORS = {
    TOOL_IO_LIST:     generate_io_list_from_instruments,
    TOOL_CABLE_BLOCK: generate_cable_block_from_io,
    TOOL_CABLE_SCHED: generate_cable_schedule_from_io,
}


def run_generator(tool: str, rows: list[dict]) -> dict:
    """Generate a canonical table for the given tool from free-form input rows."""
    if tool not in _GENERATORS:
        raise ValueError(f'Unsupported tool: {tool!r}')
    generated = _GENERATORS[tool](rows or [])
    # Run QC over the generated rows so the UI can show a single combined view.
    qc_issues = _run_rules(tool, generated)
    return {
        'tool':       tool,
        'mode':       MODE_GENERATE,
        'columns':    list(_TOOL_SCHEMAS[tool].keys()),
        'rows':       generated,
        'summary':    _summarise(qc_issues, len(generated)),
        'issues':     qc_issues,
    }


# ─── Spreadsheet ingestion (xlsx / csv) ──────────────────────────────────────
# Soft-coded list of file extensions accepted by the parsing helpers.
_ACCEPTED_EXTS = ('.xlsx', '.xlsm', '.xls', '.csv')


def parse_uploaded_table(uploaded_file) -> list[dict]:
    """Parse an uploaded spreadsheet/CSV into a list of dicts.

    Uses openpyxl for XLSX and the stdlib `csv` module otherwise. The first
    non-empty row is treated as the header; remaining rows become dict
    values keyed by that header. Empty trailing rows are skipped.
    """
    name = (getattr(uploaded_file, 'name', '') or '').lower()
    if not name.endswith(_ACCEPTED_EXTS):
        raise ValueError(
            f'Unsupported file type: {name or "<unnamed>"}. '
            f'Accepted: {", ".join(_ACCEPTED_EXTS)}'
        )

    data = uploaded_file.read()

    if name.endswith('.csv'):
        return _parse_csv_bytes(data)
    return _parse_xlsx_bytes(data)


def _parse_csv_bytes(data: bytes) -> list[dict]:
    import csv
    text = data.decode('utf-8-sig', errors='replace')
    reader = csv.reader(io.StringIO(text))
    headers: list[str] = []
    rows: list[dict] = []
    for raw in reader:
        if not any((c or '').strip() for c in raw):
            continue
        if not headers:
            headers = [str(c).strip() for c in raw]
            continue
        rows.append({headers[i]: (raw[i] if i < len(raw) else '') for i in range(len(headers))})
    return rows


def _parse_xlsx_bytes(data: bytes) -> list[dict]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:                                       # pragma: no cover
        raise RuntimeError('openpyxl is required to parse XLSX files') from exc
    wb = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    ws = wb.active
    headers: list[str] = []
    rows: list[dict] = []
    for raw in ws.iter_rows(values_only=True):
        if not any(c not in (None, '') for c in raw):
            continue
        if not headers:
            headers = [str(c).strip() if c is not None else '' for c in raw]
            continue
        rows.append({
            headers[i]: ('' if raw[i] is None else raw[i])
            for i in range(len(headers)) if headers[i]
        })
    return rows


# ─── XLSX serialisation (for download) ───────────────────────────────────────
def rows_to_xlsx_bytes(tool: str, rows: list[dict]) -> bytes:
    """Serialise canonical rows for the given tool to an XLSX file in memory."""
    try:
        from openpyxl import Workbook
    except ImportError as exc:                                       # pragma: no cover
        raise RuntimeError('openpyxl is required to write XLSX files') from exc
    columns = list(_TOOL_SCHEMAS.get(tool, {}).keys()) or (rows[0].keys() if rows else [])
    wb = Workbook()
    ws = wb.active
    ws.title = tool[:31]
    ws.append([c.replace('_', ' ').title() for c in columns])
    for row in rows:
        ws.append([row.get(c, '') for c in columns])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
