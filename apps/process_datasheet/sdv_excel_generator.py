"""
SDV Datasheet Excel Generator
Fills the bundled template with AI-mapped data
"""
import logging
from pathlib import Path
from typing import Dict, List
import openpyxl
from io import BytesIO
from django.core.files.base import ContentFile
from apps.process_datasheet.template_manager import SDVTemplateManager

logger = logging.getLogger(__name__)


class SDVExcelGenerator:
    """
    Generate filled SDV datasheets from mapped data
    Uses the bundled Excel template and openpyxl
    """
    
    # Excel cell mappings (based on template inspection)
    CELL_MAPPINGS = {
        # Document Info
        'document_no': 'C2',
        'rev_no': 'N2',
        'date': 'N3',
        
        # General Data (Column E for values)
        'tag_no': 'E5',
        'service': 'E6',
        'pid_no': 'E7',
        'line_no': 'E8',
        'piping_class': 'I8',
        'sour_service': 'E9',
        'special_service': 'I9',
        
        # Ambient Temperature
        'ambient_temp_min': 'E10',
        'ambient_temp_max': 'G10',
        'ambient_temp_unit': 'H10',
        
        # Fluid Info
        'fluid': 'E11',
        'phase': 'G11',
        'state': 'H11',
        
        # Operating Conditions - Pressure
        'operating_pressure_normal': 'E12',
        'operating_pressure_design': 'G12',
        'pressure_unit': 'I12',
        
        # Operating Conditions - Temperature
        'operating_temp_min': 'E14',
        'operating_temp_max': 'G14',
        'operating_temp_unit': 'I14',
        
        # Design Temperature
        'design_temp_min': 'E15',
        'design_temp_max': 'G15',
        'design_temp_unit': 'I15',
        
        # Valve Details
        'shut_off_pressure': 'E16',
        'bore_detail': 'E17',
        'mech_handwheel': 'E18',
        'fail_position': 'E19',
        'valve_close_time': 'E20',
        'valve_open_time': 'H21',
        'design_pressure': 'E22',
        'seat_leakage_class': 'E23',
        'nace_requirement': 'E24',
    }
    
    def __init__(self):
        """Initialize generator"""
        logger.info("[SDVExcelGenerator] Initialized")
    
    def generate_datasheet(
        self,
        mapped_data: Dict,
        output_filename: str = None
    ) -> BytesIO:
        """
        Generate filled SDV datasheet Excel file
        
        Args:
            mapped_data: AI-mapped data from SDVDatasheetAIMapper
            output_filename: Optional filename for output
        
        Returns:
            BytesIO: Excel file in memory
        """
        logger.info("[SDVExcelGenerator] Starting datasheet generation...")
        
        try:
            # Get template
            template_path = SDVTemplateManager.get_template_path()
            logger.info(f"[SDVExcelGenerator] Loading template: {template_path.name}")
            
            # Load workbook
            wb = openpyxl.load_workbook(template_path)
            ws = wb.active
            
            # Get first valve (or iterate if multiple)
            valves = mapped_data.get('valves', [])
            
            if not valves:
                logger.warning("[SDVExcelGenerator] No valves to fill")
                raise ValueError("No valve data provided")
            
            # For now, fill first valve (can be extended for multiple sheets)
            valve_data = valves[0]
            logger.info(f"[SDVExcelGenerator] Filling data for: {valve_data.get('tag_no', 'Unknown')}")
            
            # Fill all mapped fields
            filled_count = self._fill_template(ws, valve_data)
            
            logger.info(f"[SDVExcelGenerator] ✅ Filled {filled_count} fields")
            
            # Save to memory
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            wb.close()
            
            logger.info(f"[SDVExcelGenerator] ✅ Generated Excel file ({len(output.getvalue())/1024:.2f} KB)")
            
            return output
            
        except Exception as e:
            logger.error(f"[SDVExcelGenerator] ❌ Error: {e}")
            raise
    
    def _fill_template(self, ws, valve_data: Dict) -> int:
        """
        Fill Excel template with valve data
        
        Args:
            ws: openpyxl worksheet
            valve_data: Single valve data dict
        
        Returns:
            int: Number of fields filled
        """
        filled_count = 0
        
        # Add default document info
        ws['C2'] = valve_data.get('document_no', 'RJ-AB-SDV-DS-001')
        ws['N2'] = valve_data.get('rev_no', 'A')
        ws['N3'] = valve_data.get('date', 'N/A')
        filled_count += 3
        
        # Map valve data to cells
        field_mapping = {
            'tag_no': 'E5',
            'service': 'E6',
            'pid_no': 'E7',
            'line_no': 'E8',
            'piping_class': 'I8',
            'sour_service': 'E9',
            'special_service': 'I9',
            'ambient_temp_min': 'E10',
            'ambient_temp_max': 'G10',
            'ambient_temp_unit': 'H10',
            'fluid': 'E11',
            'phase': 'G11',
            'state': 'H11',
            'operating_pressure_normal': 'E12',
            'operating_pressure_design': 'G12',
            'pressure_unit': 'I12',
            'operating_temp_min': 'E14',
            'operating_temp_max': 'G14',
            'operating_temp_unit': 'I14',
            'design_temp_min': 'E15',
            'design_temp_max': 'G15',
            'design_temp_unit': 'I15',
            'shut_off_pressure': 'E16',
            'fail_position': 'E19',
            'valve_close_time': 'E20',
            'valve_open_time': 'H21',
        }
        
        for field, cell in field_mapping.items():
            value = valve_data.get(field)
            if value is not None and value != '':
                ws[cell] = value
                filled_count += 1
                logger.debug(f"[SDVExcelGenerator] {cell} = {value}")
        
        return filled_count
    
    def generate_multiple_datasheets(
        self,
        mapped_data: Dict
    ) -> Dict[str, BytesIO]:
        """
        Generate multiple datasheets (one per valve)
        
        Args:
            mapped_data: AI-mapped data with multiple valves
        
        Returns:
            Dict mapping valve tags to Excel files
        """
        valves = mapped_data.get('valves', [])
        datasheets = {}
        
        for valve in valves:
            tag = valve.get('tag_no', 'Unknown')
            logger.info(f"[SDVExcelGenerator] Generating datasheet for {tag}")
            
            # Create single-valve data
            single_valve_data = {
                'valves': [valve],
                'confidence': valve.get('confidence', 'medium')
            }
            
            # Generate datasheet
            excel_file = self.generate_datasheet(single_valve_data)
            datasheets[tag] = excel_file
        
        logger.info(f"[SDVExcelGenerator] ✅ Generated {len(datasheets)} datasheets")
        return datasheets


# Quick test function
def test_excel_generator():
    """Test Excel generator with sample data"""
    from datetime import datetime
    
    sample_mapped_data = {
        'valves': [
            {
                'tag_no': 'SDV-100-001',
                'service': 'Main Gas Line Shutdown',
                'pid_no': 'P-100-001',
                'line_no': '6"-GA-100-1501-A2B',
                'piping_class': 'ASME B16.5 150#',
                'fluid': 'Natural Gas',
                'phase': 'Gas',
                'state': 'Supercritical',
                'operating_pressure_normal': '75',
                'operating_pressure_design': '90',
                'pressure_unit': 'barg',
                'operating_temp_min': '-10',
                'operating_temp_max': '65',
                'operating_temp_unit': '°C',
                'design_temp_min': '-20',
                'design_temp_max': '85',
                'design_temp_unit': '°C',
                'fail_position': 'FC (Fail Close)',
                'valve_close_time': '5 seconds',
                'valve_open_time': '10 seconds',
                'document_no': 'RJ-AB-001-DS-001',
                'rev_no': 'A',
                'date': datetime.now().strftime('%d-%b-%Y'),
                'confidence': 'high'
            }
        ],
        'overall_confidence': 'high'
    }
    
    generator = SDVExcelGenerator()
    excel_file = generator.generate_datasheet(sample_mapped_data)
    
    # Save to file
    output_path = Path(r'C:\Users\Abdullah.Khan\RAD_AI\SDV_AI_Generated_TEST.xlsx')
    with open(output_path, 'wb') as f:
        f.write(excel_file.getvalue())
    
    print(f"✅ Test file generated: {output_path}")
    print(f"   Size: {output_path.stat().st_size / 1024:.2f} KB")


if __name__ == "__main__":
    test_excel_generator()
