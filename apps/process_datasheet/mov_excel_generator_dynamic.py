"""
Dynamic MOV Datasheet Excel Generator

Creates Process Data Sheet for Motor Operated Valves from scratch
NO TEMPLATE REQUIRED - Builds entire structure programmatically
POPULATES ONLY SECTIONS 1 & 2 - Sections 3 & 4 left blank
"""
import logging
from typing import Dict, List
from datetime import datetime
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


class MOVExcelGeneratorDynamic:
    """
    Generate MOV datasheets from scratch
    Builds complete table structure programmatically
    Only populates Sections 1 & 2
    """
    
    def __init__(self):
        """Initialize generator"""
        logger.info("[MOVExcelGeneratorDynamic] Initialized - No template needed")
        
        # Define styles
        self.header_font = Font(bold=True, size=14, name='Arial')
        self.title_font = Font(bold=True, size=16, name='Arial')
        self.section_font = Font(bold=True, size=11, name='Arial')
        self.data_font = Font(size=10, name='Arial')
        
        self.center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        self.left_alignment = Alignment(horizontal='left', vertical='center')
        
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        self.header_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
        self.section_fill = PatternFill(start_color='E8E8E8', end_color='E8E8E8', fill_type='solid')
    
    def _apply_cell_style(self, ws, start_cell, end_cell):
        """Apply border and font to range"""
        from openpyxl.utils import range_boundaries
        min_col, min_row, max_col, max_row = range_boundaries(f"{start_cell}:{end_cell}")
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = self.thin_border
                cell.font = self.data_font
    
    def _create_header(self, ws):
        """Create document header"""
        # Row 1: COMPANY Doc No
        ws.merge_cells('A1:B1')
        ws['A1'] = 'COMPANY Doc No. :'
        ws['A1'].font = self.data_font
        ws['A1'].alignment = self.left_alignment
        ws['A1'].border = self.thin_border
        
        ws.merge_cells('C1:M1')
        ws['C1'].border = self.thin_border
        
        ws['N1'] = 'Rev. No. :'
        ws['N1'].font = self.data_font
        ws['N1'].border = self.thin_border
        
        # Row 2-3: Title
        ws.merge_cells('A2:M3')
        ws['A2'] = 'PROCESS DATA SHEET\nMOV'
        ws['A2'].font = self.title_font
        ws['A2'].alignment = self.center_alignment
        ws['A2'].border = self.thin_border
        
        ws['N2'] = 'Date :'
        ws['N2'].font = self.data_font
        ws['N2'].border = self.thin_border
        ws['N3'] = datetime.now().strftime('%d-%b-%Y')
        ws['N3'].font = self.data_font
        ws['N3'].border = self.thin_border
        
        # Row 4: Document Class
        ws.merge_cells('A4:B4')
        ws['A4'] = 'Document Class:'
        ws['A4'].font = self.data_font
        ws['A4'].alignment = self.left_alignment
        ws['A4'].border = self.thin_border
        
        ws.merge_cells('C4:N4')
        ws['C4'].border = self.thin_border
    
    def _create_section_1_general_data(self, ws, start_row, valve_data):
        """Create Section 1 - General Data (POPULATE FROM DATA)"""
        # Section header
        ws.merge_cells(f'A{start_row}:B{start_row+4}')
        ws[f'A{start_row}'] = 'SECTION 1\nGENERAL\nDATA'
        ws[f'A{start_row}'].font = self.section_font
        ws[f'A{start_row}'].alignment = self.center_alignment
        ws[f'A{start_row}'].border = self.thin_border
        ws[f'A{start_row}'].fill = self.section_fill
        
        current_row = start_row
        
        # Row 1: Tag No
        ws[f'C{current_row}'] = '1'
        ws[f'D{current_row}'] = 'Tag No'
        ws.merge_cells(f'E{current_row}:N{current_row}')
        ws[f'E{current_row}'] = valve_data.get('tag_no', '')
        self._apply_cell_style(ws, f'C{current_row}', f'N{current_row}')
        current_row += 1
        
        # Row 2: Service
        ws[f'C{current_row}'] = '2'
        ws[f'D{current_row}'] = 'Service'
        ws.merge_cells(f'E{current_row}:N{current_row}')
        ws[f'E{current_row}'] = valve_data.get('service', '')
        self._apply_cell_style(ws, f'C{current_row}', f'N{current_row}')
        current_row += 1
        
        # Row 3: P&ID No.
        ws[f'C{current_row}'] = '3'
        ws[f'D{current_row}'] = 'P&ID No.'
        ws.merge_cells(f'E{current_row}:N{current_row}')
        ws[f'E{current_row}'] = valve_data.get('pid_no', '')
        self._apply_cell_style(ws, f'C{current_row}', f'N{current_row}')
        current_row += 1
        
        # Row 4: Line No | Piping Class
        ws[f'C{current_row}'] = '4'
        ws[f'D{current_row}'] = 'Line No'
        ws.merge_cells(f'E{current_row}:H{current_row}')
        ws[f'E{current_row}'] = valve_data.get('line_no', '')
        ws[f'I{current_row}'] = 'Piping Class'
        ws.merge_cells(f'J{current_row}:N{current_row}')
        ws[f'J{current_row}'] = valve_data.get('piping_class', '')
        self._apply_cell_style(ws, f'C{current_row}', f'N{current_row}')
        current_row += 1
        
        # Row 5: Fluid | State | Phase
        ws[f'C{current_row}'] = '5'
        ws[f'D{current_row}'] = 'Fluid'
        ws.merge_cells(f'E{current_row}:G{current_row}')
        ws[f'E{current_row}'] = valve_data.get('fluid', '')
        ws[f'H{current_row}'] = 'State'
        ws.merge_cells(f'I{current_row}:J{current_row}')
        ws[f'I{current_row}'] = valve_data.get('state', '')
        ws[f'K{current_row}'] = 'Phase'
        ws.merge_cells(f'L{current_row}:N{current_row}')
        ws[f'L{current_row}'] = valve_data.get('phase', '')
        self._apply_cell_style(ws, f'C{current_row}', f'N{current_row}')
        
        return start_row + 5
    
    def _create_section_2_operating_conditions(self, ws, start_row, valve_data):
        """Create Section 2 - Operating Conditions (POPULATE FROM DATA)"""
        # Section header
        ws.merge_cells(f'A{start_row}:B{start_row+5}')
        ws[f'A{start_row}'] = 'SECTION 2\nOPERATING\nCONDITIONS'
        ws[f'A{start_row}'].font = self.section_font
        ws[f'A{start_row}'].alignment = self.center_alignment
        ws[f'A{start_row}'].border = self.thin_border
        ws[f'A{start_row}'].fill = self.section_fill
        
        current_row = start_row
        
        # Row 6: Operating Pressure | Min | Normal | Max | Unit
        ws[f'C{current_row}'] = '6'
        ws[f'D{current_row}'] = 'Operating Pressure'
        ws[f'E{current_row}'] = 'Min'
        ws[f'F{current_row}'] = valve_data.get('operating_pressure_min', '')
        ws[f'G{current_row}'] = 'Normal'
        ws[f'H{current_row}'] = valve_data.get('operating_pressure_normal', '')
        ws[f'I{current_row}'] = 'Max'
        ws[f'J{current_row}'] = valve_data.get('operating_pressure_max', '')
        ws[f'K{current_row}'] = 'Unit'
        ws.merge_cells(f'L{current_row}:N{current_row}')
        ws[f'L{current_row}'] = valve_data.get('pressure_unit', '')
        self._apply_cell_style(ws, f'C{current_row}', f'N{current_row}')
        current_row += 1
        
        # Row 7: Operating Temperature | Min | Normal | Max | Unit
        ws[f'C{current_row}'] = '7'
        ws[f'D{current_row}'] = 'Operating Temperature'
        ws[f'E{current_row}'] = 'Min'
        ws[f'F{current_row}'] = valve_data.get('operating_temp_min', '')
        ws[f'G{current_row}'] = 'Normal'
        ws[f'H{current_row}'] = valve_data.get('operating_temp_normal', '')
        ws[f'I{current_row}'] = 'Max'
        ws[f'J{current_row}'] = valve_data.get('operating_temp_max', '')
        ws[f'K{current_row}'] = 'Unit'
        ws.merge_cells(f'L{current_row}:N{current_row}')
        ws[f'L{current_row}'] = valve_data.get('operating_temp_unit', '')
        self._apply_cell_style(ws, f'C{current_row}', f'N{current_row}')
        current_row += 1
        
        # Row 8: Design Pressure | Min | Max
        ws[f'C{current_row}'] = '8'
        ws[f'D{current_row}'] = 'Design Pressure'
        ws[f'E{current_row}'] = 'Min'
        ws.merge_cells(f'F{current_row}:H{current_row}')
        ws[f'F{current_row}'] = valve_data.get('design_pressure_min', '')
        ws[f'I{current_row}'] = 'Max'
        ws.merge_cells(f'J{current_row}:N{current_row}')
        ws[f'J{current_row}'] = valve_data.get('design_pressure_max', '')
        self._apply_cell_style(ws, f'C{current_row}', f'N{current_row}')
        current_row += 1
        
        # Row 9: Design Temperature | Min | Max
        ws[f'C{current_row}'] = '9'
        ws[f'D{current_row}'] = 'Design Temperature'
        ws[f'E{current_row}'] = 'Min'
        ws.merge_cells(f'F{current_row}:H{current_row}')
        ws[f'F{current_row}'] = valve_data.get('design_temp_min', '')
        ws[f'I{current_row}'] = 'Max'
        ws.merge_cells(f'J{current_row}:N{current_row}')
        ws[f'J{current_row}'] = valve_data.get('design_temp_max', '')
        self._apply_cell_style(ws, f'C{current_row}', f'N{current_row}')
        current_row += 1
        
        # Row 10: Sour Service | Special Conditions
        ws[f'C{current_row}'] = '10'
        ws[f'D{current_row}'] = 'Sour Service'
        ws.merge_cells(f'E{current_row}:H{current_row}')
        ws[f'E{current_row}'] = valve_data.get('sour_service', '')
        ws[f'I{current_row}'] = 'Special Conditions'
        ws.merge_cells(f'J{current_row}:N{current_row}')
        ws[f'J{current_row}'] = valve_data.get('special_conditions', '')
        self._apply_cell_style(ws, f'C{current_row}', f'N{current_row}')
        current_row += 1
        
        # Row 11: Shut Off Pressure
        ws[f'C{current_row}'] = '11'
        ws[f'D{current_row}'] = 'Shut Off Pressure'
        ws.merge_cells(f'E{current_row}:N{current_row}')
        ws[f'E{current_row}'] = valve_data.get('shut_off_pressure', '')
        self._apply_cell_style(ws, f'C{current_row}', f'N{current_row}')
        
        return start_row + 6
    
    def _create_section_3_valve_details(self, ws, start_row):
        """Create Section 3 - Valve Details (LEAVE BLANK)"""
        # Section header
        ws.merge_cells(f'A{start_row}:B{start_row+1}')
        ws[f'A{start_row}'] = 'SECTION 3\nVALVE\nDETAILS'
        ws[f'A{start_row}'].font = self.section_font
        ws[f'A{start_row}'].alignment = self.center_alignment
        ws[f'A{start_row}'].border = self.thin_border
        ws[f'A{start_row}'].fill = self.section_fill
        
        current_row = start_row
        
        # Row 12: Diff. Pressure (ΔP)
        ws[f'C{current_row}'] = '12'
        ws[f'D{current_row}'] = 'Diff. Pressure (ΔP)'
        ws.merge_cells(f'E{current_row}:N{current_row}')
        ws[f'E{current_row}'] = ''  # BLANK
        self._apply_cell_style(ws, f'C{current_row}', f'N{current_row}')
        current_row += 1
        
        # Row 13: Seat Leakage Class | NACE Compliant
        ws[f'C{current_row}'] = '13'
        ws[f'D{current_row}'] = 'Seat Leakage Class'
        ws.merge_cells(f'E{current_row}:H{current_row}')
        ws[f'E{current_row}'] = ''  # BLANK
        ws[f'I{current_row}'] = 'NACE Compliant'
        ws.merge_cells(f'J{current_row}:N{current_row}')
        ws[f'J{current_row}'] = ''  # BLANK
        self._apply_cell_style(ws, f'C{current_row}', f'N{current_row}')
        
        return start_row + 2
    
    def _create_section_4_actuator_details(self, ws, start_row):
        """Create Section 4 - Actuator Details (LEAVE BLANK)"""
        # Section header
        ws.merge_cells(f'A{start_row}:B{start_row+1}')
        ws[f'A{start_row}'] = 'SECTION 4\nACTUATOR\nDETAILS'
        ws[f'A{start_row}'].font = self.section_font
        ws[f'A{start_row}'].alignment = self.center_alignment
        ws[f'A{start_row}'].border = self.thin_border
        ws[f'A{start_row}'].fill = self.section_fill
        
        current_row = start_row
        
        # Row 14: Fail Position
        ws[f'C{current_row}'] = '14'
        ws[f'D{current_row}'] = 'Fail Position'
        ws.merge_cells(f'E{current_row}:N{current_row}')
        ws[f'E{current_row}'] = ''  # BLANK
        self._apply_cell_style(ws, f'C{current_row}', f'N{current_row}')
        current_row += 1
        
        # Row 15: Valve Close Time | Valve Open Time
        ws[f'C{current_row}'] = '15'
        ws[f'D{current_row}'] = 'Valve Close Time'
        ws.merge_cells(f'E{current_row}:H{current_row}')
        ws[f'E{current_row}'] = ''  # BLANK
        ws[f'I{current_row}'] = 'Valve Open Time'
        ws.merge_cells(f'J{current_row}:N{current_row}')
        ws[f'J{current_row}'] = ''  # BLANK
        self._apply_cell_style(ws, f'C{current_row}', f'N{current_row}')
        
        return start_row + 2
    
    def generate_datasheet(self, mapped_data: Dict) -> BytesIO:
        """
        Generate MOV datasheet Excel file
        
        Args:
            mapped_data: Dictionary containing valve data from AI mapper
        
        Returns:
            BytesIO: Excel file in memory
        """
        logger.info("[MOVExcelGeneratorDynamic] 🎨 Generating datasheet...")
        
        valves = mapped_data.get('valves', [])
        
        if not valves:
            raise ValueError("No valve data to generate datasheet")
        
        # Create workbook
        wb = Workbook()
        
        # Create a sheet for each valve
        for idx, valve in enumerate(valves):
            if idx == 0:
                ws = wb.active
                ws.title = f"MOV-{idx+1}"
            else:
                ws = wb.create_sheet(title=f"MOV-{idx+1}")
            
            # Set column widths
            ws.column_dimensions['A'].width = 5
            ws.column_dimensions['B'].width = 5
            ws.column_dimensions['C'].width = 3
            ws.column_dimensions['D'].width = 18
            for col in ['E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N']:
                ws.column_dimensions[col].width = 10
            
            # Build datasheet
            self._create_header(ws)
            
            current_row = 6  # Start after header
            
            # Section 1: General Data (POPULATE)
            current_row = self._create_section_1_general_data(ws, current_row, valve)
            current_row += 1  # Spacing
            
            # Section 2: Operating Conditions (POPULATE)
            current_row = self._create_section_2_operating_conditions(ws, current_row, valve)
            current_row += 1  # Spacing
            
            # Section 3: Valve Details (BLANK)
            current_row = self._create_section_3_valve_details(ws, current_row)
            current_row += 1  # Spacing
            
            # Section 4: Actuator Details (BLANK)
            current_row = self._create_section_4_actuator_details(ws, current_row)
        
        # Save to BytesIO
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        logger.info(f"[MOVExcelGeneratorDynamic] ✅ Generated {len(valves)} datasheet(s)")
        
        return excel_buffer
