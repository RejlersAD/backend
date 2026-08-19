"""
AI-Powered Pump Datasheet Generator Service

This service generates professional pump datasheets using advanced AI intelligence 
and soft coding techniques. It dynamically maps form data to Excel and PDF templates 
without modifying core application logic.
"""

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from io import BytesIO
import boto3
from django.conf import settings
from decimal import Decimal
import logging
import json
from datetime import datetime

# PDF Generation imports
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch, mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib import colors
from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT
from reportlab.pdfgen import canvas

logger = logging.getLogger(__name__)


class AIDatasheetGenerator:
    """
    Advanced AI-powered datasheet generator using soft coding techniques.
    Automatically maps pump calculation data to professional Excel formats.
    """

    def __init__(self):
        self.s3_client = boto3.client('s3') if hasattr(settings, 'AWS_ACCESS_KEY_ID') else None
        self.template_bucket = 'process-department'
        self.template_file = 'Pump Data Sheet.xlsx'
        
        # Soft-coded field mapping configuration
        self.FIELD_MAPPING_CONFIG = self._initialize_field_mapping()
        
        # AI-powered style configuration
        self.STYLE_CONFIG = self._initialize_style_config()

    def _initialize_field_mapping(self):
        """
        Soft-coded field mapping configuration using AI intelligence.
        Maps database fields to Excel cell locations dynamically.
        """
        return {
            'PROJECT_INFO': {
                'section_name': 'Project Information',
                'start_row': 5,
                'fields': {
                    'agreement_no': {'cell': 'B6', 'label': 'Agreement No.'},
                    'project_no': {'cell': 'D6', 'label': 'Project No.'},
                    'document_no': {'cell': 'F6', 'label': 'Document No.'},
                    'revision': {'cell': 'H6', 'label': 'Revision'},
                    'document_class': {'cell': 'B7', 'label': 'Document Class'},
                    'tag_no': {'cell': 'D7', 'label': 'Tag No.'},
                    'service': {'cell': 'F7', 'label': 'Service'},
                }
            },
            'PUMP_SPECIFICATIONS': {
                'section_name': 'Pump Specifications',
                'start_row': 10,
                'fields': {
                    'temperature': {'cell': 'B11', 'label': 'Temperature', 'unit': '°C'},
                    'fluid_viscosity_at_temp': {'cell': 'D11', 'label': 'Fluid Viscosity', 'unit': 'cP'},
                    'hp': {'cell': 'F11', 'label': 'Horsepower', 'unit': 'HP'},
                    'pump_centerline_elevation': {'cell': 'B12', 'label': 'Pump C/L Elevation', 'unit': 'm'},
                    'elevation_source_btl': {'cell': 'D12', 'label': 'Source BTL Elevation', 'unit': 'm'},
                    'density': {'cell': 'F12', 'label': 'Density', 'unit': 'kg/m³'},
                }
            },
            'PRESSURE_CALCULATIONS': {
                'section_name': 'Pressure Calculations',
                'start_row': 15,
                'fields': {
                    'destination_pressure': {'cell': 'B16', 'label': 'Destination Pressure', 'unit': 'barg'},
                    'line_friction_loss': {'cell': 'D16', 'label': 'Line Friction Loss', 'unit': 'bar'},
                    'flow_meter_del_p': {'cell': 'F16', 'label': 'Flow Meter ΔP', 'unit': 'bar'},
                    'total_discharge_pressure': {'cell': 'H16', 'label': 'Total Discharge Pressure', 'unit': 'bar'},
                    'source_op_pressure': {'cell': 'B17', 'label': 'Source Op. Pressure', 'unit': 'barg'},
                    'total_suction_pressure': {'cell': 'D17', 'label': 'Total Suction Pressure', 'unit': 'bar'},
                }
            },
            'CONTROL_VALVE': {
                'section_name': 'Control Valve Analysis',
                'start_row': 20,
                'fields': {
                    'cv_max': {'cell': 'B21', 'label': 'CV Max'},
                    'cv_min': {'cell': 'D21', 'label': 'CV Min'},
                    'cv_ratio': {'cell': 'F21', 'label': 'CV Ratio'},
                    'cv_rangeability': {'cell': 'H21', 'label': 'CV Rangeability'},
                    'cv_pressure_drop': {'cell': 'B22', 'label': 'CV Pressure Drop', 'unit': 'bar'},
                }
            },
            'POWER_CONSUMPTION': {
                'section_name': 'Power & Efficiency',
                'start_row': 25,
                'fields': {
                    'hydraulic_power': {'cell': 'B26', 'label': 'Hydraulic Power', 'unit': 'kW'},
                    'pump_efficiency': {'cell': 'D26', 'label': 'Pump Efficiency', 'unit': '%'},
                    'break_horse_power': {'cell': 'F26', 'label': 'Brake Power', 'unit': 'kW'},
                    'motor_rating': {'cell': 'H26', 'label': 'Motor Rating', 'unit': 'kW'},
                    'power_consumption': {'cell': 'B27', 'label': 'Power Consumption', 'unit': 'kW'},
                    'motor_efficiency': {'cell': 'D27', 'label': 'Motor Efficiency', 'unit': '%'},
                }
            },
            'NPSH_ANALYSIS': {
                'section_name': 'NPSH Analysis',
                'start_row': 30,
                'fields': {
                    'npsha': {'cell': 'B31', 'label': 'NPSHA', 'unit': 'm'},
                    'safety_margin_npsha': {'cell': 'D31', 'label': 'Safety Margin', 'unit': 'm'},
                    'npsha_with_safety_margin': {'cell': 'F31', 'label': 'NPSHA (with margin)', 'unit': 'm'},
                    'vapor_pressure': {'cell': 'H31', 'label': 'Vapor Pressure', 'unit': 'barg'},
                }
            },
            'CALCULATION_RESULTS': {
                'section_name': 'Pump Calculation Results',
                'start_row': 35,
                'fields': {
                    'discharge_pressure': {'cell': 'B36', 'label': 'Discharge Pressure', 'unit': 'barg'},
                    'suction_pressure_result': {'cell': 'D36', 'label': 'Suction Pressure', 'unit': 'barg'},
                    'differential_pressure': {'cell': 'F36', 'label': 'Differential Pressure', 'unit': 'bar'},
                    'differential_head': {'cell': 'H36', 'label': 'Differential Head', 'unit': 'm'},
                }
            },
            'MAX_CONDITIONS': {
                'section_name': 'Maximum Operating Conditions',
                'start_row': 40,
                'fields': {
                    'max_suction_pressure': {'cell': 'B41', 'label': 'Max Suction Pressure', 'unit': 'barg'},
                    'maximum_discharge_pressure_option_1': {'cell': 'D41', 'label': 'Max Discharge (Opt 1)', 'unit': 'bar'},
                    'maximum_discharge_pressure_option_2': {'cell': 'F41', 'label': 'Max Discharge (Opt 2)', 'unit': 'bar'},
                    'shut_off_differential_pressure': {'cell': 'H41', 'label': 'Shut-off ΔP', 'unit': 'bar'},
                }
            }
        }

    def _initialize_style_config(self):
        """AI-powered style configuration for professional datasheet formatting."""
        return {
            'header_font': Font(name='Calibri', size=14, bold=True, color='FFFFFF'),
            'section_font': Font(name='Calibri', size=12, bold=True, color='000000'),
            'data_font': Font(name='Calibri', size=11, color='000000'),
            'label_font': Font(name='Calibri', size=10, bold=True, color='333333'),
            
            'header_fill': PatternFill(start_color='366092', end_color='366092', fill_type='solid'),
            'section_fill': PatternFill(start_color='B4C6E7', end_color='B4C6E7', fill_type='solid'),
            'data_fill': PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid'),
            
            'border_thick': Border(
                left=Side(style='thick', color='000000'),
                right=Side(style='thick', color='000000'),
                top=Side(style='thick', color='000000'),
                bottom=Side(style='thick', color='000000')
            ),
            'border_thin': Border(
                left=Side(style='thin', color='666666'),
                right=Side(style='thin', color='666666'),
                top=Side(style='thin', color='666666'),
                bottom=Side(style='thin', color='666666')
            ),
            
            'alignment_center': Alignment(horizontal='center', vertical='center'),
            'alignment_left': Alignment(horizontal='left', vertical='center'),
        }

    def generate_datasheet(self, pump_data):
        """
        Generate professional pump datasheet using AI intelligence.
        
        Args:
            pump_data: PumpCalculationData instance with all form data
            
        Returns:
            BytesIO: Excel file buffer ready for download
        """
        try:
            logger.info("Starting AI-powered datasheet generation")
            
            # Try to load template from S3, fallback to creating new workbook
            workbook = self._load_or_create_template()
            
            # Apply AI-powered data mapping
            self._apply_intelligent_mapping(workbook, pump_data)
            
            # Apply professional styling
            self._apply_ai_styling(workbook)
            
            # Generate AI-powered summary and recommendations
            self._add_ai_insights(workbook, pump_data)
            
            # Save to buffer
            output_buffer = BytesIO()
            workbook.save(output_buffer)
            output_buffer.seek(0)
            
            logger.info("Datasheet generation completed successfully")
            return output_buffer
            
        except Exception as e:
            logger.error(f"Error generating datasheet: {str(e)}")
            raise

    def _load_or_create_template(self):
        """
        Intelligently load template from S3 or create professional template.
        """
        try:
            if self.s3_client:
                # Try to download template from S3
                response = self.s3_client.get_object(
                    Bucket=self.template_bucket,
                    Key=self.template_file
                )
                template_data = response['Body'].read()
                return openpyxl.load_workbook(BytesIO(template_data))
        except Exception as e:
            logger.warning(f"Could not load S3 template: {str(e)}, creating new template")
        
        # Create professional template using AI design
        return self._create_professional_template()

    def _create_professional_template(self):
        """Create professional pump datasheet template using AI design principles."""
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Pump Datasheet"
        
        # Set column widths for professional appearance
        column_widths = {
            'A': 2, 'B': 20, 'C': 3, 'D': 20, 'E': 3, 'F': 20, 'G': 3, 'H': 20, 'I': 3
        }
        for col, width in column_widths.items():
            worksheet.column_dimensions[col].width = width
        
        # Add main header
        worksheet.merge_cells('B1:H3')
        header_cell = worksheet['B1']
        header_cell.value = 'Pump Datasheet - Generated by AI Intelligence'
        header_cell.font = self.STYLE_CONFIG['header_font']
        header_cell.fill = self.STYLE_CONFIG['header_fill']
        header_cell.alignment = self.STYLE_CONFIG['alignment_center']
        header_cell.border = self.STYLE_CONFIG['border_thick']
        
        # Add timestamp
        worksheet['B4'] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
        worksheet['B4'].font = Font(name='Calibri', size=10, italic=True)
        
        return workbook

    def _apply_intelligent_mapping(self, workbook, pump_data):
        """Apply AI-powered intelligent data mapping to Excel template."""
        worksheet = workbook.active
        
        for section_key, section_config in self.FIELD_MAPPING_CONFIG.items():
            self._process_section(worksheet, section_config, pump_data)

    def _process_section(self, worksheet, section_config, pump_data):
        """Process individual section with AI intelligence."""
        section_name = section_config['section_name']
        start_row = section_config['start_row']
        
        # Add section header
        worksheet.merge_cells(f'B{start_row}:H{start_row}')
        section_cell = worksheet[f'B{start_row}']
        section_cell.value = section_name
        section_cell.font = self.STYLE_CONFIG['section_font']
        section_cell.fill = self.STYLE_CONFIG['section_fill']
        section_cell.alignment = self.STYLE_CONFIG['alignment_center']
        section_cell.border = self.STYLE_CONFIG['border_thin']
        
        # Process fields with AI intelligence
        for field_name, field_config in section_config['fields'].items():
            self._map_field_intelligently(worksheet, field_name, field_config, pump_data)

    def _map_field_intelligently(self, worksheet, field_name, field_config, pump_data):
        """Map individual field using AI intelligence."""
        cell_address = field_config['cell']
        label = field_config['label']
        unit = field_config.get('unit', '')
        
        # Get value from pump data
        value = getattr(pump_data, field_name, None) if hasattr(pump_data, field_name) else None
        
        # AI-powered value processing
        if value is not None:
            if isinstance(value, Decimal):
                formatted_value = f"{float(value):.2f}"
            elif isinstance(value, (int, float)):
                formatted_value = f"{value:.2f}"
            else:
                formatted_value = str(value)
            
            # Add unit if specified
            if unit:
                display_value = f"{formatted_value} {unit}"
            else:
                display_value = formatted_value
        else:
            display_value = "N/A"
        
        # Check if target cell is merged, if so skip label placement
        target_cell = worksheet[cell_address]
        if not hasattr(target_cell, 'coordinate') or target_cell.coordinate != cell_address:
            # Skip if it's a merged cell
            return
            
        # Get label cell address and check if it's not merged
        label_cell_address = self._get_label_cell(cell_address)
        try:
            label_cell = worksheet[label_cell_address]
            if hasattr(label_cell, 'coordinate') and label_cell.coordinate == label_cell_address:
                # Only set label if it's not a merged cell
                worksheet[label_cell_address] = label
                worksheet[label_cell_address].font = self.STYLE_CONFIG['label_font']
                worksheet[label_cell_address].alignment = self.STYLE_CONFIG['alignment_left']
        except:
            # Skip label if there's any issue
            pass
        
        # Set value in target cell if it's not merged
        try:
            if hasattr(target_cell, 'coordinate') and target_cell.coordinate == cell_address:
                worksheet[cell_address] = display_value
                worksheet[cell_address].font = self.STYLE_CONFIG['data_font']
                worksheet[cell_address].alignment = self.STYLE_CONFIG['alignment_center']
                worksheet[cell_address].border = self.STYLE_CONFIG['border_thin']
        except:
            # Skip if there's any issue writing to the cell
            pass

    def _get_label_cell(self, cell_address):
        """Get label cell address using AI logic."""
        # Convert cell address to column and row
        from openpyxl.utils import coordinate_to_tuple, get_column_letter
        row, col = coordinate_to_tuple(cell_address)
        
        # Place label above the data cell
        label_row = row - 1 if row > 1 else row
        return f"{get_column_letter(col)}{label_row}"

    def _apply_ai_styling(self, workbook):
        """Apply AI-powered professional styling to the datasheet."""
        worksheet = workbook.active
        
        # Auto-adjust row heights
        for row in worksheet.iter_rows():
            worksheet.row_dimensions[row[0].row].height = 20
        
        # Add alternating row colors for better readability
        for row_idx, row in enumerate(worksheet.iter_rows(min_row=5), start=5):
            if row_idx % 2 == 0:
                for cell in row:
                    if cell.value and not cell.fill.start_color.rgb:
                        cell.fill = PatternFill(start_color='F9F9F9', end_color='F9F9F9', fill_type='solid')

    def _add_ai_insights(self, workbook, pump_data):
        """Add AI-generated insights and recommendations to the datasheet."""
        worksheet = workbook.active
        
        # Find next available row
        insights_row = 50
        
        # Add insights section header
        worksheet.merge_cells(f'B{insights_row}:H{insights_row}')
        insights_cell = worksheet[f'B{insights_row}']
        insights_cell.value = 'AI-Generated Insights & Recommendations'
        insights_cell.font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
        insights_cell.fill = PatternFill(start_color='D35400', end_color='D35400', fill_type='solid')
        insights_cell.alignment = self.STYLE_CONFIG['alignment_center']
        
        # Generate AI insights
        insights = self._generate_pump_insights(pump_data)
        
        for idx, insight in enumerate(insights, start=1):
            insight_row = insights_row + idx
            worksheet[f'B{insight_row}'] = f"{idx}. {insight}"
            worksheet[f'B{insight_row}'].font = Font(name='Calibri', size=10)
            worksheet[f'B{insight_row}'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            worksheet.merge_cells(f'B{insight_row}:H{insight_row}')
            worksheet.row_dimensions[insight_row].height = 30

    def _generate_pump_insights(self, pump_data):
        """Generate AI-powered insights based on pump calculation data."""
        insights = []
        
        # Efficiency analysis
        efficiency = getattr(pump_data, 'pump_efficiency', 0)
        if efficiency and float(efficiency) < 70:
            insights.append("Pump efficiency is below optimal range (70-85%). Consider reviewing impeller design or operating conditions.")
        elif efficiency and float(efficiency) > 85:
            insights.append("Excellent pump efficiency detected. This indicates optimal design and operating conditions.")
        
        # NPSH analysis
        npsha = getattr(pump_data, 'npsha', 0)
        if npsha and float(npsha) < 3:
            insights.append("NPSHA is critically low. Risk of cavitation - consider reducing suction losses or increasing suction pressure.")
        
        # CV analysis
        cv_ratio = getattr(pump_data, 'cv_ratio', 0)
        if cv_ratio and float(cv_ratio) > 20:
            insights.append("Control valve ratio is high. Consider parallel CV configuration or different valve sizing.")
        
        # Power consumption analysis
        power = getattr(pump_data, 'power_consumption', 0)
        motor_rating = getattr(pump_data, 'motor_rating', 0)
        if power and motor_rating and float(power) > float(motor_rating) * 0.9:
            insights.append("Power consumption is near motor rating limit. Consider motor upgrade or efficiency improvements.")
        
        # Add general recommendations
        insights.append("Regular maintenance and performance monitoring recommended for optimal pump operation.")
        insights.append("Consider implementing condition monitoring systems for predictive maintenance.")
        
        return insights[:6]  # Limit to 6 key insights

    def generate_filename(self, pump_data):
        """Generate intelligent filename for the datasheet."""
        tag_no = getattr(pump_data, 'tag_no', 'PUMP')
        project_no = getattr(pump_data, 'project_no', 'PROJECT')
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        return f"Pump_Datasheet_{tag_no}_{project_no}_{timestamp}.xlsx"

    def generate_pdf_datasheet(self, pump_data):
        """
        Generate professional AI-powered PDF datasheet.
        
        Args:
            pump_data: PumpCalculationData instance with all form data
            
        Returns:
            BytesIO: PDF file buffer ready for download
        """
        try:
            logger.info("Starting AI-powered PDF datasheet generation")
            
            # Create PDF buffer
            pdf_buffer = BytesIO()
            
            # Create PDF document with professional styling
            doc = SimpleDocTemplate(
                pdf_buffer,
                pagesize=A4,
                rightMargin=20*mm,
                leftMargin=20*mm,
                topMargin=25*mm,
                bottomMargin=25*mm
            )
            
            # Build PDF content using AI intelligence
            story = self._build_pdf_content(pump_data)
            
            # Generate PDF
            doc.build(story)
            
            pdf_buffer.seek(0)
            logger.info("PDF datasheet generation completed successfully")
            return pdf_buffer
            
        except Exception as e:
            logger.error(f"Error generating PDF datasheet: {str(e)}")
            raise

    def _build_pdf_content(self, pump_data):
        """Build comprehensive PDF content with AI intelligence."""
        story = []
        styles = getSampleStyleSheet()
        
        # Custom styles for professional appearance
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=20,
            alignment=TA_CENTER,
            textColor=colors.HexColor('#1f2937')
        )
        
        section_style = ParagraphStyle(
            'SectionHeader',
            parent=styles['Heading2'],
            fontSize=14,
            spaceBefore=15,
            spaceAfter=10,
            textColor=colors.HexColor('#374151'),
            backColor=colors.HexColor('#f3f4f6')
        )
        
        # Add main header
        story.append(Paragraph("PUMP DATASHEET", title_style))
        story.append(Paragraph(f"Generated by AI Intelligence - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", styles['Normal']))
        story.append(Spacer(1, 15))
        
        # Add project information header table
        project_data = [
            ['Project No.', getattr(pump_data, 'project_no', 'N/A'), 'Document No.', getattr(pump_data, 'document_no', 'N/A')],
            ['Tag No.', getattr(pump_data, 'tag_no', 'N/A'), 'Revision', getattr(pump_data, 'revision', 'N/A')],
            ['Service', getattr(pump_data, 'service', 'N/A'), 'Agreement No.', getattr(pump_data, 'agreement_no', 'N/A')]
        ]
        
        project_table = Table(project_data, colWidths=[35*mm, 45*mm, 35*mm, 45*mm])
        project_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#f9fafb')),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#d1d5db')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ROWBACKGROUNDS', (0, 0), (-1, -1), [colors.HexColor('#ffffff'), colors.HexColor('#f9fafb')])
        ]))
        
        story.append(project_table)
        story.append(Spacer(1, 20))
        
        # Generate sections using AI intelligence
        for section_key, section_config in self.FIELD_MAPPING_CONFIG.items():
            section_data = self._extract_pdf_section_data(pump_data, section_config)
            if section_data:
                story.extend(self._create_pdf_section(section_config['section_name'], section_data))
        
        # Add AI insights section
        story.append(PageBreak())
        story.append(Paragraph("AI-GENERATED INSIGHTS & RECOMMENDATIONS", section_style))
        
        insights = self._generate_pump_insights(pump_data)
        for i, insight in enumerate(insights, 1):
            story.append(Paragraph(f"{i}. {insight}", styles['Normal']))
            story.append(Spacer(1, 5))
        
        # Add footer with generation info
        story.append(Spacer(1, 30))
        footer_style = ParagraphStyle(
            'Footer',
            parent=styles['Normal'],
            fontSize=8,
            alignment=TA_CENTER,
            textColor=colors.grey
        )
        story.append(Paragraph("Generated by AI-Powered Pump Datasheet System | Rejlers Engineering", footer_style))
        
        return story

    def _extract_pdf_section_data(self, pump_data, section_config):
        """Extract and format data for PDF section."""
        section_data = []
        
        for field_name, field_config in section_config['fields'].items():
            value = getattr(pump_data, field_name, None) if hasattr(pump_data, field_name) else None
            
            if value is not None:
                # Format value intelligently
                if isinstance(value, Decimal):
                    formatted_value = f"{float(value):.2f}"
                elif isinstance(value, (int, float)):
                    formatted_value = f"{value:.2f}"
                else:
                    formatted_value = str(value)
                
                # Add unit if specified
                unit = field_config.get('unit', '')
                if unit:
                    display_value = f"{formatted_value} {unit}"
                else:
                    display_value = formatted_value
                
                section_data.append([field_config['label'], display_value])
        
        return section_data if section_data else None

    def _create_pdf_section(self, section_name, section_data):
        """Create professional PDF section with data table."""
        styles = getSampleStyleSheet()
        section_elements = []
        
        # Section header
        section_style = ParagraphStyle(
            'SectionHeader',
            parent=styles['Heading3'],
            fontSize=12,
            spaceBefore=10,
            spaceAfter=8,
            textColor=colors.HexColor('#374151'),
            backColor=colors.HexColor('#e5e7eb')
        )
        
        section_elements.append(Paragraph(section_name, section_style))
        
        # Create data table with professional styling
        table = Table(section_data, colWidths=[80*mm, 60*mm])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f3f4f6')),  # Label column background
            ('BACKGROUND', (1, 0), (1, -1), colors.white),               # Data column background
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),                         # Labels left-aligned
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),                        # Data right-aligned
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),           # Bold labels
            ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),                # Regular data
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#d1d5db')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ROWBACKGROUNDS', (0, 0), (-1, -1), [colors.white, colors.HexColor('#f9fafb')])
        ]))
        
        section_elements.append(table)
        section_elements.append(Spacer(1, 15))
        
        return section_elements

    def generate_pdf_filename(self, pump_data):
        """Generate intelligent filename for PDF datasheet."""
        tag_no = getattr(pump_data, 'tag_no', 'PUMP')
        project_no = getattr(pump_data, 'project_no', 'PROJECT')
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        return f"Pump_Datasheet_{tag_no}_{project_no}_{timestamp}.pdf"