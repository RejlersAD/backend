"""
Smart Pump Data Sheet Generator - Template-Based Approach
Uses actual Pump Data Sheet.xlsx template and intelligently maps form data
to preserve professional formatting and structure.
"""

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from io import BytesIO
import boto3
from django.conf import settings
from decimal import Decimal
import logging
import os
from datetime import datetime

logger = logging.getLogger(__name__)


class PumpDataSheetGenerator:
    """
    Template based pump data sheet generator
    Loads actual Pump Data Sheet.xlsx template and maps form data intelligently
    """

    def __init__(self):
        self.template_path = os.path.join(
            settings.BASE_DIR, 
            'temp_pump_analysis', 
            'pump_data_sheet_output.xlsx'
        )
        
        # Smart field mapping to Excel template cells
        # Based on actual Pump Data Sheet template structure
        self.FIELD_MAPPING = self._initialize_template_mapping()

    def _initialize_template_mapping(self):
        """
        Intelligent mapping of form fields to template cells
        Soft-coded for easy maintenance and updates
        """
        return {
            # Cover Sheet Fields
            'cover': {
                'sheet_name': 'Cover',
                'fields': {
                    'document_no': 'C7',  # Document number
                    'tag_no': 'C8',  # Tag number / Description
                    'revision': 'G7',  # Revision number
                    'date': 'G8',  # Date
                    'document_class': 'A9',  # Document class
                }
            },
            
            # Sheet 1 - Main Data Sheet
            'main_sheet': {
                'sheet_name': 'Sheet 1',
                'fields': {
                    # Project Information (rows 11-15)
                    'company_name': 'I12',
                    'site': 'I13',
                    'unit': 'AD12',
                    'service': 'AD13',
                    'no_required': 'I14',
                    'type_of_pump': 'AD14',
                    'manufacturer': 'I15',
                    'model': 'AD15',
                    
                    # Liquid Characteristics (rows 18-30)
                    'liquid_type': 'N19',  # Or service description
                    'vapor_pressure': {'max': 'N20', 'min': 'Q20'},
                    'density': {'max': 'N21', 'min': 'Q21'},
                    'viscosity': {'max': 'N22', 'min': 'Q22'},
                    'temperature': {'max': 'N23', 'min': 'Q23'},
                    
                    # Flow Rate (row 31)
                    'flow_rate': {'max': 'N31', 'min': 'Q31', 'normal': 'W31'},
                    
                    # Pressure Data (rows 32-40)
                    'suction_pressure': {'max': 'N32', 'min': 'Q32', 'normal': 'W32'},
                    'discharge_pressure': {'max': 'N33', 'min': 'Q33', 'normal': 'W33'},
                    'differential_pressure': {'max': 'N34', 'min': 'Q34', 'normal': 'W34'},
                    'differential_head': {'max': 'N35', 'min': 'Q35', 'normal': 'W35'},
                    
                    # NPSH (rows 36-37)
                    'npsh_available': {'max': 'N36', 'min': 'Q36'},
                    'npsh_required': 'N37',
                    
                    # Pump Efficiency & Power (rows 38-40)
                    'pump_efficiency': {'max': 'N38', 'min': 'Q38', 'normal': 'W38'},
                    'bhp': {'max': 'N39', 'min': 'Q39', 'normal': 'W39'},
                    'absorbed_power': {'max': 'N40', 'min': 'Q40', 'normal': 'W40'},
                    
                    # Driver/Motor Data (rows 44-51)
                    'driver_type': 'N44',
                    'motor_rating': 'N47',
                    'motor_voltage': 'N48',
                    'motor_speed': 'N49',
                    'motor_efficiency': 'N50',
                    'motor_classification': 'N51',
                    
                    # Construction Materials  (rows 56-64)
                    'casing': 'N56',
                    'impeller': 'N57',
                    'shaft': 'N58',
                    'bearings': 'N59',
                    'mechanical_seal': 'N60',
                    
                    # Notes and Remarks (rows 68+)
                    'general_notes': 'C68',
                }
            }
        }

    def generate_datasheet(self, pump_data):
        """
        Generate pump data sheet from template with intelligent field mapping
        
        Args:
            pump_data: PumpCalculationData instance
            
        Returns:
            BytesIO: Excel file buffer
        """
        try:
            logger.info(f"🔄 Generating pump datasheet for ID: {pump_data.id}")
            
            # Check if template exists, otherwise create structured datasheet
            if not os.path.exists(self.template_path):
                logger.warning("Template not found, generating structured datasheet")
                return self._generate_structured_datasheet(pump_data)
            
            # Load the template
            workbook = self._load_template()
            
            # Apply data mapping
            self._map_cover_sheet(workbook, pump_data)
            self._map_main_sheet(workbook, pump_data)
            
            # Save to buffer
            output_buffer = BytesIO()
            workbook.save(output_buffer)
            output_buffer.seek(0)
            
            logger.info("✅ Datasheet generated successfully")
            return output_buffer
            
        except Exception as e:
            logger.error(f"❌ Error generating datasheet: {str(e)}")
            raise

    def _generate_structured_datasheet(self, pump_data):
        """Generate a structured Excel datasheet when template is not available"""
        logger.info("📊 Creating structured datasheet from pump data")
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Pump Hydraulic Datasheet"
        
        # Header styling
        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        section_font = Font(bold=True, size=11)
        section_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        
        row = 1
        
        # Title
        ws.merge_cells(f'A{row}:D{row}')
        title_cell = ws[f'A{row}']
        title_cell.value = "PUMP HYDRAULIC CALCULATION DATA SHEET"
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal='center')
        row += 2
        
        # Project Information Section
        ws.merge_cells(f'A{row}:D{row}')
        section_cell = ws[f'A{row}']
        section_cell.value = "PROJECT INFORMATION"
        section_cell.font = section_font
        section_cell.fill = section_fill
        row += 1
        
        project_fields = [
            ('Agreement No', getattr(pump_data, 'agreement_no', 'N/A')),
            ('Project No', getattr(pump_data, 'project_no', 'N/A')),
            ('Document No', getattr(pump_data, 'document_no', 'N/A')),
            ('Revision', getattr(pump_data, 'revision', 'A')),
            ('Document Class', getattr(pump_data, 'document_class', 'N/A')),
            ('Tag No', getattr(pump_data, 'tag_no', 'N/A')),
            ('Service', getattr(pump_data, 'service', 'N/A')),
        ]
        
        for label, value in project_fields:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = value
            row += 1
        
        row += 1
        
        # Discharge Pressure Calculations
        ws.merge_cells(f'A{row}:D{row}')
        section_cell = ws[f'A{row}']
        section_cell.value = "DISCHARGE PRESSURE CALCULATIONS"
        section_cell.font = section_font
        section_cell.fill = section_fill
        row += 1
        
        discharge_fields = [
            ('Destination Description', getattr(pump_data, 'destination_description', 'N/A')),
            ('Flow Type', getattr(pump_data, 'flow_type', 'N/A')),
            ('Destination Pressure (bar)', self._format_number(getattr(pump_data, 'destination_pressure', None))),
            ('Destination Elevation (m)', self._format_number(getattr(pump_data, 'destination_elevation', None))),
            ('Line Friction Loss (bar)', self._format_number(getattr(pump_data, 'line_friction_loss', None))),
            ('Flow Meter Del P (bar)', self._format_number(getattr(pump_data, 'flow_meter_del_p', None))),
            ('Other Losses (bar)', self._format_number(getattr(pump_data, 'other_losses', None))),
            ('Control Valve (bar)', self._format_number(getattr(pump_data, 'control_valve', None))),
            ('Misc Item (bar)', self._format_number(getattr(pump_data, 'misc_item', None))),
            ('Contingency (bar)', self._format_number(getattr(pump_data, 'contingency', None))),
            ('TOTAL DISCHARGE PRESSURE (bar)', self._format_number(getattr(pump_data, 'total_discharge_pressure', None))),
        ]
        
        for label, value in discharge_fields:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True) if 'TOTAL' in label else Font()
            ws[f'B{row}'] = value
            if 'TOTAL' in label:
                ws[f'B{row}'].font = Font(bold=True)
                ws[f'B{row}'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            row += 1
        
        row += 1
        
        # Control Valve Delta P Check
        ws.merge_cells(f'A{row}:D{row}')
        section_cell = ws[f'A{row}']
        section_cell.value = "CONTROL VALVE DELTA P CHECK"
        section_cell.font = section_font
        section_cell.fill = section_fill
        row += 1
        
        cv_fields = [
            ('Density (kg/m³)', self._format_number(getattr(pump_data, 'density', None))),
            ('CV Max', self._format_number(getattr(pump_data, 'cv_max', None))),
            ('CV Min', self._format_number(getattr(pump_data, 'cv_min', None))),
            ('CV Ratio', self._format_number(getattr(pump_data, 'cv_ratio', None))),
            ('Total Frictional Losses (bar)', self._format_number(getattr(pump_data, 'total_frictional_losses', None))),
            ('Dynamic Losses 30%', self._format_number(getattr(pump_data, 'dynamic_losses_30_percent', None))),
            ('CV Pressure Drop (bar)', self._format_number(getattr(pump_data, 'cv_pressure_drop', None))),
            ('CV Rangeability', self._format_number(getattr(pump_data, 'cv_rangeability', None))),
            ('CV Ratio Within Range?', getattr(pump_data, 'cv_ratio_within_range', 'N/A')),
            ('CV Pr.drop@Normal Flow > 30% Fric?', getattr(pump_data, 'cv_pressure_drop_check', 'N/A')),
        ]
        
        for label, value in cv_fields:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = value
            row += 1
        
        row += 1
        
        # Suction Pressure Calculations
        ws.merge_cells(f'A{row}:D{row}')
        section_cell = ws[f'A{row}']
        section_cell.value = "SUCTION PRESSURE CALCULATIONS"
        section_cell.font = section_font
        section_cell.fill = section_fill
        row += 1
        
        suction_fields = [
            ('Source Op. Pressure (bar)', self._format_number(getattr(pump_data, 'source_op_pressure', None))),
            ('Suction EL from Pump C/L (m)', self._format_number(getattr(pump_data, 'suction_el_m', None))),
            ('Inline Inst. Losses (bar)', self._format_number(getattr(pump_data, 'inline_inst_losses', None))),
            ('Line Fric. Losses (bar)', self._format_number(getattr(pump_data, 'line_fric_losses', None))),
            ('Control Valve Suction (bar)', self._format_number(getattr(pump_data, 'control_valve_suction', None))),
            ('Misc Items Suction (bar)', self._format_number(getattr(pump_data, 'misc_items_suction', None))),
            ('Total Suction Losses (bar)', self._format_number(getattr(pump_data, 'total_suction_losses', None))),
            ('TOTAL SUCTION PRESSURE (bar)', self._format_number(getattr(pump_data, 'total_suction_pressure', None))),
        ]
        
        for label, value in suction_fields:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True) if 'TOTAL' in label else Font(bold=True)
            ws[f'B{row}'] = value
            if 'TOTAL SUCTION' in label:
                ws[f'B{row}'].font = Font(bold=True)
                ws[f'B{row}'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            row += 1
        
        row += 1
        
        # Power Consumption Per Pump
        ws.merge_cells(f'A{row}:D{row}')
        section_cell = ws[f'A{row}']
        section_cell.value = "POWER CONSUMPTION PER PUMP"
        section_cell.font = section_font
        section_cell.fill = section_fill
        row += 1
        
        power_fields = [
            ('Hydraulic Power (kW)', self._format_number(getattr(pump_data, 'hydraulic_power', None))),
            ('Pump Efficiency (%)', self._format_number(getattr(pump_data, 'pump_efficiency', None))),
            ('Break Horse Power (HP)', self._format_number(getattr(pump_data, 'break_horse_power', None))),
            ('Motor Rating (kW)', self._format_number(getattr(pump_data, 'motor_rating', None))),
            ('Motor Efficiency (%)', self._format_number(getattr(pump_data, 'motor_efficiency', None))),
            ('Power Consumption (kW)', self._format_number(getattr(pump_data, 'power_consumption', None))),
            ('Type of Motor', getattr(pump_data, 'type_of_motor', 'N/A')),
            ('Motor Classification', getattr(pump_data, 'motor_classification', 'N/A')),
        ]
        
        for label, value in power_fields:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = value
            row += 1
        
        row += 1
        
        # NPSH Availability
        ws.merge_cells(f'A{row}:D{row}')
        section_cell = ws[f'A{row}']
        section_cell.value = "NPSH AVAILABILITY"
        section_cell.font = section_font
        section_cell.fill = section_fill
        row += 1
        
        npsh_fields = [
            ('Suction Pressure (bar)', self._format_number(getattr(pump_data, 'suction_pressure_npsh', None))),
            ('Vapor Pressure (bar)', self._format_number(getattr(pump_data, 'vapor_pressure', None))),
            ('NPSHA (m)', self._format_number(getattr(pump_data, 'npsha', None))),
            ('Safety Margin NPSHA (%)', self._format_number(getattr(pump_data, 'safety_margin_npsha', None))),
            ('NPSHA with Safety Margin (m)', self._format_number(getattr(pump_data, 'npsha_with_safety_margin', None))),
        ]
        
        for label, value in npsh_fields:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = value
            row += 1
        
        row += 1
        
        # Pump Calculation Results
        ws.merge_cells(f'A{row}:D{row}')
        section_cell = ws[f'A{row}']
        section_cell.value = "PUMP CALCULATION RESULTS"
        section_cell.font = section_font
        section_cell.fill = section_fill
        row += 1
        
        result_fields = [
            ('Discharge Pressure (bar)', self._format_number(getattr(pump_data, 'discharge_pressure', None))),
            ('Suction Pressure (bar)', self._format_number(getattr(pump_data, 'suction_pressure_result', None))),
            ('Differential Pressure (bar)', self._format_number(getattr(pump_data, 'differential_pressure', None))),
            ('Differential Head (m)', self._format_number(getattr(pump_data, 'differential_head', None))),
            ('NPSHA (m)', self._format_number(getattr(pump_data, 'npsha_result', None))),
        ]
        
        for label, value in result_fields:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = value
            ws[f'B{row}'].font = Font(bold=True, color="0000FF")
            row += 1
        
        row += 1
        
        # Maximum Suction Pressure
        ws.merge_cells(f'A{row}:D{row}')
        section_cell = ws[f'A{row}']
        section_cell.value = "MAXIMUM SUCTION PRESSURE"
        section_cell.font = section_font
        section_cell.fill = section_fill
        row += 1
        
        max_suction_fields = [
            ('Suction Vessel Max Op. Pressure (bar)', self._format_number(getattr(pump_data, 'suction_vessel_max_op_pressure', None))),
            ('Suction EL from Pump C/L Max (m)', self._format_number(getattr(pump_data, 'suction_el_m', None))),
            ('TL to HHLL (m)', self._format_number(getattr(pump_data, 'tl_to_hhll_m', None))),
            ('Max Suction Pressure (bar)', self._format_number(getattr(pump_data, 'max_suction_pressure', None))),
        ]
        
        for label, value in max_suction_fields:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = value
            row += 1
        
        row += 1
        
        # Minimum Flow Conditions
        ws.merge_cells(f'A{row}:D{row}')
        section_cell = ws[f'A{row}']
        section_cell.value = "MINIMUM FLOW CONDITIONS"
        section_cell.font = section_font
        section_cell.fill = section_fill
        row += 1
        
        min_flow_fields = [
            ('Pump Minimum Flow (m³/h)', self._format_number(getattr(pump_data, 'pump_minimum_flow', None))),
            ('Fluid Density MCF (kg/m³)', self._format_number(getattr(pump_data, 'fluid_density_mcf', None))),
            ('Pump Discharge Pressure Min Flow (bar)', self._format_number(getattr(pump_data, 'pump_discharge_pressure_min_flow', None))),
            ('Destination Pressure MCF (bar)', self._format_number(getattr(pump_data, 'destination_pressure', None))),
            ('EL Destination Pump C/L (m)', self._format_number(getattr(pump_data, 'el_destination_pump_cl', None))),
            ('MCF Line Friction Losses (bar)', self._format_number(getattr(pump_data, 'mcf_line_friction_losses', None))),
            ('Flow Meter Losses (bar)', self._format_number(getattr(pump_data, 'flow_meter_losses', None))),
            ('Misc Pressure Drop MCF (bar)', self._format_number(getattr(pump_data, 'misc_pressure_drop_mcf', None))),
            ('MCF CV Pressure Drop (bar)', self._format_number(getattr(pump_data, 'mcf_cv_pressure_drop', None))),
        ]
        
        for label, value in min_flow_fields:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = value
            row += 1
        
        row += 1
        
        # Maximum Discharge Pressure
        ws.merge_cells(f'A{row}:D{row}')
        section_cell = ws[f'A{row}']
        section_cell.value = "MAXIMUM DISCHARGE PRESSURE"
        section_cell.font = section_font
        section_cell.fill = section_fill
        row += 1
        
        max_discharge_fields = [
            ('API610 Tolerance Used', getattr(pump_data, 'api_610_tolerance_used', 'N/A')),
            ('API Tolerance Factor', self._format_number(getattr(pump_data, 'api_tolerance_factor', None))),
            ('Shut Off Pressure Factor', self._format_number(getattr(pump_data, 'shut_off_pressure_factor', None))),
            ('Shut Off Differential Pressure (bar)', self._format_number(getattr(pump_data, 'shut_off_differential_pressure', None))),
            ('Maximum Discharge Pressure Option 1 (bar)', self._format_number(getattr(pump_data, 'maximum_discharge_pressure_option_1', None))),
            ('Maximum Discharge Pressure Option 2 (bar)', self._format_number(getattr(pump_data, 'maximum_discharge_pressure_option_2', None))),
        ]
        
        for label, value in max_discharge_fields:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = value
            row += 1
        
        # Auto-size columns
        ws.column_dimensions['A'].width = 50
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        
        # Save to buffer
        output_buffer = BytesIO()
        wb.save(output_buffer)
        output_buffer.seek(0)
        
        logger.info("✅ Structured datasheet generated successfully")
        return output_buffer

    def _load_template(self):
        """Load the Pump Data Sheet template"""
        if not os.path.exists(self.template_path):
            logger.warning(f"Template not found at {self.template_path}, creating basic workbook")
            return openpyxl.Workbook()
        
        logger.info(f"📂 Loading template from: {self.template_path}")
        return openpyxl.load_workbook(self.template_path)

    def _map_cover_sheet(self, workbook, pump_data):
        """Map data to Cover sheet"""
        mapping = self.FIELD_MAPPING['cover']
        
        if mapping['sheet_name'] not in workbook.sheetnames:
            logger.warning(f"Cover sheet not found in template")
            return
        
        ws = workbook[mapping['sheet_name']]
        
        # Map fields using soft-coded configuration
        field_values = {
            'document_no': getattr(pump_data, 'document_no', 'N/A'),
            'tag_no': getattr(pump_data, 'tag_no', 'N/A'),
            'revision': getattr(pump_data, 'revision', '0'),
            'date': datetime.now().strftime('%Y-%m-%d'),
            'document_class': getattr(pump_data, 'document_class', 'Confidential'),
        }
        
        for field_name, cell_address in mapping['fields'].items():
            if field_name in field_values:
                try:
                    self._set_cell_value(ws, cell_address, field_values[field_name])
                except Exception as e:
                    logger.warning(f"Could not set {field_name} in {cell_address}: {e}")

    def _map_main_sheet(self, workbook, pump_data):
        """Map data to main data sheet (Sheet 1)"""
        mapping = self.FIELD_MAPPING['main_sheet']
        
        if mapping['sheet_name'] not in workbook.sheetnames:
            logger.warning(f"Main sheet '{mapping['sheet_name']}' not found")
            return
        
        ws = workbook[mapping['sheet_name']]
        
        # Project Information
        self._set_cell_value(ws, 'I12', getattr(pump_data, 'company_name', 'Rejlers'))
        self._set_cell_value(ws, 'I13', getattr(pump_data, 'site', 'TBD'))
        self._set_cell_value(ws, 'AD12', getattr(pump_data, 'unit', 'N/A'))
        self._set_cell_value(ws, 'AD13', getattr(pump_data, 'service', 'N/A'))
        self._set_cell_value(ws, 'I14', '1')  # Number required (default 1)
        self._set_cell_value(ws, 'AD14', 'Centrifugal')  # Type
        self._set_cell_value(ws, 'I15', getattr(pump_data, 'manufacturer', 'TBD'))
        self._set_cell_value(ws, 'AD15', getattr(pump_data, 'model', 'TBD'))
        
        # Liquid Characteristics
        self._set_cell_value(ws, 'N19', getattr(pump_data, 'service', 'Process Fluid'))
        
        # Vapor Pressure
        vapor_pressure = self._format_number(getattr(pump_data, 'vapor_pressure', None))
        self._set_cell_value(ws, 'N20', vapor_pressure)
        self._set_cell_value(ws, 'Q20', vapor_pressure)
        
        # Density
        density = self._format_number(getattr(pump_data, 'density', None))
        self._set_cell_value(ws, 'N21', density)
        self._set_cell_value(ws, 'Q21', density)
        
        # Viscosity
        viscosity = self._format_number(getattr(pump_data, 'fluid_viscosity_at_temp', None))
        self._set_cell_value(ws, 'N22', viscosity)
        self._set_cell_value(ws, 'Q22', viscosity)
        
        # Temperature
        temperature = self._format_number(getattr(pump_data, 'temperature', None))
        self._set_cell_value(ws, 'N23', temperature)
        self._set_cell_value(ws, 'Q23', temperature)
        
        # Flow Rate (using flow_type to determine max/normal/min)
        flow_type = getattr(pump_data, 'flow_type', 'Normal')
        # For now, put same value in all three
        self._set_cell_value(ws, 'N31', 'Max')
        self._set_cell_value(ws, 'Q31', 'Min')
        self._set_cell_value(ws, 'W31', 'Normal')
        
        # Pressure Data
        discharge_pressure = self._format_number(getattr(pump_data, 'total_discharge_pressure', None))
        suction_pressure = self._format_number(getattr(pump_data, 'total_suction_pressure', None))
        differential_pressure = self._format_number(getattr(pump_data, 'differential_pressure', None))
        differential_head = self._format_number(getattr(pump_data, 'differential_head', None))
        
        # Suction Pressure
        self._set_cell_value(ws, 'N32', suction_pressure)
        self._set_cell_value(ws, 'Q32', suction_pressure)
        self._set_cell_value(ws, 'W32', suction_pressure)
        
        # Discharge Pressure
        self._set_cell_value(ws, 'N33', discharge_pressure)
        self._set_cell_value(ws, 'Q33', discharge_pressure)
        self._set_cell_value(ws, 'W33', discharge_pressure)
        
        # Differential Pressure
        self._set_cell_value(ws, 'N34', differential_pressure)
        self._set_cell_value(ws, 'Q34', differential_pressure)
        self._set_cell_value(ws, 'W34', differential_pressure)
        
        # Differential Head
        self._set_cell_value(ws, 'N35', differential_head)
        self._set_cell_value(ws, 'Q35', differential_head)
        self._set_cell_value(ws, 'W35', differential_head)
        
        # NPSH
        npsha = self._format_number(getattr(pump_data, 'npsha', None))
        self._set_cell_value(ws, 'N36', npsha)
        self._set_cell_value(ws, 'Q36', npsha)
        self._set_cell_value(ws, 'N37', 'TBD')  # NPSH Required (not in form)
        
        # Pump Efficiency
        pump_eff = self._format_number(getattr(pump_data, 'pump_efficiency', None))
        self._set_cell_value(ws, 'N38', pump_eff)
        self._set_cell_value(ws, 'Q38', pump_eff)
        self._set_cell_value(ws, 'W38', pump_eff)
        
        # BHP (Break Horse Power)
        bhp = self._format_number(getattr(pump_data, 'break_horse_power', None))
        self._set_cell_value(ws, 'N39', bhp)
        self._set_cell_value(ws, 'Q39', bhp)
        self._set_cell_value(ws, 'W39', bhp)
        
        # Absorbed Power
        power = self._format_number(getattr(pump_data, 'power_consumption', None))
        self._set_cell_value(ws, 'N40', power)
        self._set_cell_value(ws, 'Q40', power)
        self._set_cell_value(ws, 'W40', power)
        
        # Driver/Motor Data
        self._set_cell_value(ws, 'N44', getattr(pump_data, 'type_of_motor', 'AC Induction'))
        self._set_cell_value(ws, 'N47', self._format_number(getattr(pump_data, 'motor_rating', None)))
        self._set_cell_value(ws, 'N48', 'TBD')  # Voltage not in form
        self._set_cell_value(ws, 'N49', 'TBD')  # Speed not in form
        self._set_cell_value(ws, 'N50', self._format_number(getattr(pump_data, 'motor_efficiency', None)))
        self._set_cell_value(ws, 'N51', getattr(pump_data, 'motor_classification', 'N/A'))
        
        # General Notes
        notes_data = {
            'Agreement No': getattr(pump_data, 'agreement_no', 'N/A'),
            'Project No': getattr(pump_data, 'project_no', 'N/A'),
            'Document No': getattr(pump_data, 'document_no', 'N/A'),
        }
        notes_text = ' | '.join([f"{k}: {v}" for k, v in notes_data.items()])
        self._set_cell_value(ws, 'C68', notes_text)

    def _set_cell_value(self, worksheet, cell_address, value):
        """Safely set cell value preserving formatting"""
        try:
            cell = worksheet[cell_address]
            
            if value is not None:
                cell.value = value
            else:
                cell.value = ''
                
        except Exception as e:
            logger.warning(f"Could not set value in {cell_address}: {e}")

    def _format_number(self, value):
        """Format numeric values consistently"""
        if value is None:
            return 'N/A'
        
        if isinstance(value, (Decimal, float)):
            return f"{float(value):.2f}"
        
        if isinstance(value, int):
            return str(value)
        
        return str(value)

    def generate_filename(self, pump_data):
        """Generate intelligent filename for the datasheet"""
        doc_no = getattr(pump_data, 'document_no', 'PUMP')
        tag_no = getattr(pump_data, 'tag_no', 'DATA')
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        # Clean filename (remove special characters)
        doc_no = ''.join(c for c in doc_no if c.isalnum() or c in '-_')
        tag_no = ''.join(c for c in tag_no if c.isalnum() or c in '-_')
        
        return f"Pump_Data_Sheet_{doc_no}_{tag_no}_{timestamp}.xlsx"
