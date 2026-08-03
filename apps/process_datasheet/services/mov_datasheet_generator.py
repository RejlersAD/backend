"""
Smart MOV (Motor Operated Valve) Datasheet Generator
SOFT-CODED: Intelligent extraction and Excel generation for MOV datasheets
Uses comprehensive field configuration for P&ID data extraction
"""

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from io import BytesIO
from django.conf import settings
import logging
import os
from datetime import datetime

# Import MOV configuration
from apps.process_datasheet.config.mov_datasheet_config import MOV_DATASHEET_FIELDS, get_all_fields, get_required_fields

logger = logging.getLogger(__name__)


class MOVDatasheetGenerator:
    """
    Template-based MOV datasheet generator
    Intelligently maps detected MOV equipment data to Excel format
    """

    def __init__(self):
        self.template_path = os.path.join(
            settings.BASE_DIR,
            'MOV_Datasheet.xlsx'  # Template at root level
        )
        
        # Soft-coded field mapping for MOV datasheets
        self.FIELD_MAPPING = self._initialize_template_mapping()
        
        # Default styles for professional appearance
        self.STYLES = self._initialize_styles()

    def _initialize_template_mapping(self):
        """
        Intelligent mapping of MOV fields to Excel template
        Flexible and configurable for different template formats
        """
        return {
            # Header Information
            'header': {
                'tag_number': 'B2',
                'description': 'B3',
                'service': 'B4',
                'size': 'D2',
                'rating': 'D3',
                'date': 'F2',
            },
            
            # Valve Specifications
            'valve_specs': {
                'valve_type': 'B7',
                'body_material': 'B8',
                'trim_material': 'B9',
                'seat_material': 'B10',
                'gasket_material': 'B11',
                'bonnet_type': 'D7',
                'end_connection': 'D8',
                'face_to_face': 'D9',
            },
            
            # Operating Conditions
            'operating_conditions': {
                'fluid': 'B14',
                'temperature_min': 'B15',
                'temperature_max': 'D15',
                'pressure_min': 'B16',
                'pressure_max': 'D16',
                'flow_rate': 'B17',
                'cv': 'D17',
            },
            
            # Actuator Specifications
            'actuator': {
                'actuator_type': 'B21',
                'actuator_size': 'B22',
                'operating_voltage': 'B23',
                'operating_current': 'D23',
                'power_rating': 'B24',
                'operating_torque': 'D24',
                'fail_position': 'B25',
                'operation_time': 'D25',
            },
            
            # Accessories
            'accessories': {
                'position_indicator': 'B29',
                'limit_switches': 'B30',
                'solenoid_valve': 'B31',
                'manual_override': 'B32',
            },
            
            # Standards and Certifications
            'standards': {
                'design_standard': 'B36',
                'testing_standard': 'B37',
                'material_standard': 'B38',
            }
        }

    def _initialize_styles(self):
        """Define professional Excel styles"""
        return {
            'header': Font(name='Arial', size=12, bold=True),
            'subheader': Font(name='Arial', size=11, bold=True),
            'normal': Font(name='Arial', size=10),
            'title': Font(name='Arial', size=14, bold=True, color='FFFFFF'),
            'title_fill': PatternFill(start_color='366092', end_color='366092', fill_type='solid'),
            'section_header': Font(name='Arial', size=11, bold=True, color='FFFFFF'),
            'section_fill': PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid'),
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            ),
            'alignment_center': Alignment(horizontal='center', vertical='center'),
            'alignment_left': Alignment(horizontal='left', vertical='center', wrap_text=True),
        }

    def generate_datasheet(self, equipment_data):
        """
        Generate MOV datasheet from detected equipment
        
        Args:
            equipment_data: List of detected MOV equipment from P&ID analysis
            
        Returns:
            BytesIO: Excel file buffer
        """
        try:
            logger.info(f"🔄 Generating MOV datasheet for {len(equipment_data)} valves")
            
            # Load or create workbook
            workbook = self._load_or_create_template()
            
            # Generate sheet for each MOV or summary sheet
            if len(equipment_data) == 1:
                # Single MOV - detailed sheet
                self._create_detailed_sheet(workbook, equipment_data[0])
            else:
                # Multiple MOVs - summary + individual sheets
                self._create_summary_sheet(workbook, equipment_data)
                for mov in equipment_data:
                    self._create_detailed_sheet(workbook, mov)
            
            # Save to buffer
            output_buffer = BytesIO()
            workbook.save(output_buffer)
            output_buffer.seek(0)
            
            logger.info("✅ MOV datasheet generated successfully")
            return output_buffer
            
        except Exception as e:
            logger.error(f"❌ Error generating MOV datasheet: {str(e)}")
            raise

    def _load_or_create_template(self):
        """
        Create new workbook for MOV datasheet
        Note: Template loading disabled to avoid merged cell conflicts
        """
        logger.info("Creating new MOV datasheet workbook")
        return openpyxl.Workbook()

    def _create_summary_sheet(self, workbook, equipment_list):
        """Create summary sheet for multiple MOVs"""
        # Remove default sheet if exists
        if 'Sheet' in workbook.sheetnames:
            del workbook['Sheet']
        
        ws = workbook.create_sheet('MOV Summary', 0)
        
        # Title
        ws['A1'] = 'MOTOR OPERATED VALVE (MOV) SUMMARY'
        ws['A1'].font = self.STYLES['title']
        ws['A1'].fill = self.STYLES['title_fill']
        ws['A1'].alignment = self.STYLES['alignment_center']
        ws.merge_cells('A1:H1')
        ws.row_dimensions[1].height = 25
        
        # Date
        ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        ws['A2'].font = self.STYLES['normal']
        ws.merge_cells('A2:H2')
        
        # Headers
        headers = ['No.', 'Tag Number', 'Service', 'Size', 'Type', 'Actuator', 'Voltage', 'Status']
        row = 4
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.STYLES['section_header']
            cell.fill = self.STYLES['section_fill']
            cell.alignment = self.STYLES['alignment_center']
            cell.border = self.STYLES['border']
        
        # Data rows
        for idx, mov in enumerate(equipment_list, start=1):
            row += 1
            data = [
                idx,
                mov.get('tag_number', f'MOV-{idx:03d}'),
                mov.get('service', 'N/A'),
                mov.get('size', ''),
                mov.get('valve_type', 'Ball Valve'),
                mov.get('actuator_type', 'Electric'),
                mov.get('voltage', ''),
                'Detected'
            ]
            
            for col, value in enumerate(data, start=1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.font = self.STYLES['normal']
                cell.alignment = self.STYLES['alignment_left']
                cell.border = self.STYLES['border']
        
        # Auto-adjust column widths
        for col in range(1, 9):
            ws.column_dimensions[get_column_letter(col)].width = 15

    def _create_detailed_sheet(self, workbook, mov_data):
        """
        Create detailed sheet for individual MOV using soft-coded configuration
        Dynamically generates sections based on MOV_DATASHEET_FIELDS
        """
        tag = mov_data.get('tag_number', f'MOV-{datetime.now().strftime("%H%M%S")}')
        sheet_name = tag[:31]  # Excel sheet name limit
        
        # Create new sheet
        if sheet_name in workbook.sheetnames:
            sheet_name = f"{sheet_name[:28]}_{len(workbook.sheetnames)}"
        
        ws = workbook.create_sheet(sheet_name)
        
        # Title Section
        ws['A1'] = 'MOTOR OPERATED VALVE DATASHEET'
        ws['A1'].font = self.STYLES['title']
        ws['A1'].fill = self.STYLES['title_fill']
        ws['A1'].alignment = self.STYLES['alignment_center']
        ws.merge_cells('A1:G1')
        ws.row_dimensions[1].height = 25
        
        # Date stamp
        ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        ws['A2'].font = self.STYLES['normal']
        ws.merge_cells('A2:G2')
        
        current_row = 4
        
        # ===================================================================
        # SECTION 1: GENERAL DATA
        # ===================================================================
        section_config = MOV_DATASHEET_FIELDS['general_data']
        self._add_section_header(ws, current_row, section_config['section_name'].upper())
        current_row += 1
        
        # Tag Number, Service, P&ID No. (full width)
        for field_key in ['tag_number', 'service', 'pid_no']:
            field_config = section_config['fields'][field_key]
            label = field_config['label']
            value = mov_data.get(field_key, '')
            self._add_detail_row(ws, current_row, label, value, merge_to='G')
            current_row += 1
        
        # Line Number & Piping Class (side by side)
        self._add_split_row(ws, current_row, 
                           section_config['fields']['line_number']['label'],
                           mov_data.get('line_number', ''),
                           section_config['fields']['piping_class']['label'],
                           mov_data.get('piping_class', ''))
        current_row += 1
        
        # Fluid, State, Phase (three columns)
        self._add_triple_row(ws, current_row,
                            section_config['fields']['fluid']['label'],
                            mov_data.get('fluid', ''),
                            section_config['fields']['state']['label'],
                            mov_data.get('state', ''),
                            section_config['fields']['phase']['label'],
                            mov_data.get('phase', ''))
        current_row += 2
        
        # ===================================================================
        # SECTION 2: OPERATING CONDITIONS
        # ===================================================================
        section_config = MOV_DATASHEET_FIELDS['operating_conditions']
        self._add_section_header(ws, current_row, section_config['section_name'].upper())
        current_row += 1
        
        # Operating Pressure header row
        self._add_triple_header(ws, current_row, '6. Operating Pressure', 'Min', 'Normal', 'Maximum')
        current_row += 1
        self._add_triple_value_row(ws, current_row,
                                   mov_data.get('operating_pressure_min', '0'),
                                   mov_data.get('operating_pressure_normal', ''),
                                   mov_data.get('operating_pressure_max', ''),
                                   'bar(g)')
        current_row += 1
        
        # Operating Temperature
        self._add_triple_header(ws, current_row, '7. Operating Temperature', 'Min', 'Normal', 'Maximum')
        current_row += 1
        self._add_triple_value_row(ws, current_row,
                                   mov_data.get('operating_temperature_min', ''),
                                   mov_data.get('operating_temperature_normal', ''),
                                   mov_data.get('operating_temperature_max', ''),
                                   '°C')
        current_row += 1
        
        # Design Pressure
        self._add_triple_header(ws, current_row, '8. Design Pressure', 'Min', 'Normal', 'Maximum')
        current_row += 1
        self._add_triple_value_row(ws, current_row,
                                   mov_data.get('design_pressure_min', ''),
                                   mov_data.get('design_pressure_normal', ''),
                                   mov_data.get('design_pressure_max', ''),
                                   'bar(g)')
        current_row += 1
        
        # Design Temperature
        self._add_triple_header(ws, current_row, '9. Design Temperature', 'Min', 'Normal', 'Maximum')
        current_row += 1
        self._add_triple_value_row(ws, current_row,
                                   mov_data.get('design_temperature_min', ''),
                                   mov_data.get('design_temperature_normal', ''),
                                   mov_data.get('design_temperature_max', ''),
                                   '°C')
        current_row += 1
        
        # Source Service and Shutoff Pressure
        self._add_detail_row(ws, current_row, '10. Source Service and Special Condition',
                            mov_data.get('source_service', ''), merge_to='G')
        current_row += 1
        self._add_detail_row(ws, current_row, '11. Shut Off Pressure',
                            f"{mov_data.get('shutoff_pressure', '')} bar(g)", merge_to='G')
        current_row += 2
        
        # ===================================================================
        # SECTION 3: VALVE DETAILS
        # ===================================================================
        section_config = MOV_DATASHEET_FIELDS['valve_details']
        self._add_section_header(ws, current_row, section_config['section_name'].upper())
        current_row += 1
        
        valve_details = [
            ('12. Differential Pressure', mov_data.get('differential_pressure', ''), 'bar'),
            (section_config['fields']['valve_type']['label'], mov_data.get('valve_type', 'Ball Valve'), ''),
            (section_config['fields']['valve_size']['label'], mov_data.get('valve_size', ''), ''),
            (section_config['fields']['body_material']['label'], mov_data.get('body_material', 'Carbon Steel'), ''),
            (section_config['fields']['trim_material']['label'], mov_data.get('trim_material', 'Stainless Steel 316'), ''),
            (section_config['fields']['seat_material']['label'], mov_data.get('seat_material', 'PTFE'), ''),
            (section_config['fields']['end_connection']['label'], mov_data.get('end_connection', 'Flanged RF'), ''),
        ]
        
        for label, value, unit in valve_details:
            display_value = f"{value} {unit}" if unit and value else value
            self._add_detail_row(ws, current_row, label, display_value, merge_to='G')
            current_row += 1
        
        # Seat Leakage & NACE (side by side)
        self._add_split_row(ws, current_row,
                           '13.1 Seat Leakage Class',
                           mov_data.get('seat_leakage_class', 'Class VI'),
                           '13.2 NACE Compliant',
                           mov_data.get('nace_compliant', 'N/A'))
        current_row += 2
        
        # ===================================================================
        # SECTION 4: ACTUATOR DETAILS
        # ===================================================================
        section_config = MOV_DATASHEET_FIELDS['actuator_details']
        self._add_section_header(ws, current_row, section_config['section_name'].upper())
        current_row += 1
        
        actuator_details = [
            ('14. Fail Position', mov_data.get('fail_position', 'As-Is'), ''),
            (section_config['fields']['actuator_type']['label'], mov_data.get('actuator_type', 'Electric Motor'), ''),
            (section_config['fields']['actuator_make']['label'], mov_data.get('actuator_make', 'TBD'), ''),
            (section_config['fields']['operating_voltage']['label'], mov_data.get('operating_voltage', '415V AC, 3-Phase, 50Hz'), ''),
            (section_config['fields']['operating_current']['label'], mov_data.get('operating_current', ''), 'A'),
            (section_config['fields']['power_rating']['label'], mov_data.get('power_rating', ''), 'kW'),
            (section_config['fields']['operating_torque']['label'], mov_data.get('operating_torque', ''), 'Nm'),
        ]
        
        for label, value, unit in actuator_details:
            display_value = f"{value} {unit}" if unit and value else value
            self._add_detail_row(ws, current_row, label, display_value, merge_to='G')
            current_row += 1
        
        # Valve Close/Open Time (side by side)
        self._add_split_row(ws, current_row,
                           '15.1 Valve Close Time',
                           f"{mov_data.get('valve_close_time', '')} seconds",
                           '15.2 Valve Open Time',
                           f"{mov_data.get('valve_open_time', '')} seconds")
        current_row += 1
        
        # Accessories
        accessories_data = [
            (section_config['fields']['position_indicator']['label'], mov_data.get('position_indicator', 'Visual + Electrical')),
            (section_config['fields']['limit_switches']['label'], mov_data.get('limit_switches', 'Open/Close')),
            (section_config['fields']['manual_override']['label'], mov_data.get('manual_override', 'Handwheel')),
        ]
        
        for label, value in accessories_data:
            self._add_detail_row(ws, current_row, label, value, merge_to='G')
            current_row += 1
        
        # Auto-adjust column widths
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 18
        ws.column_dimensions['F'].width = 18
        ws.column_dimensions['G'].width = 18

    def _add_section_header(self, ws, row, text):
        """Add styled section header"""
        cell = ws.cell(row=row, column=1, value=text)
        cell.font = self.STYLES['section_header']
        cell.fill = self.STYLES['section_fill']
        cell.alignment = self.STYLES['alignment_center']
        ws.merge_cells(f'A{row}:G{row}')
        ws.row_dimensions[row].height = 20

    def _add_detail_row(self, ws, row, label, value, merge_to='B'):
        """Add a detail row with label and value"""
        # Label
        label_cell = ws.cell(row=row, column=1, value=label)
        label_cell.font = Font(name='Arial', size=10, bold=True)
        label_cell.alignment = self.STYLES['alignment_left']
        label_cell.border = self.STYLES['border']
        
        # Value
        value_cell = ws.cell(row=row, column=2, value=value)
        value_cell.font = self.STYLES['normal']
        value_cell.alignment = self.STYLES['alignment_left']
        value_cell.border = self.STYLES['border']
        
        # Merge value cells if specified
        if merge_to != 'B':
            ws.merge_cells(f'B{row}:{merge_to}{row}')
    
    def _add_split_row(self, ws, row, label1, value1, label2, value2):
        """Add a row with two label-value pairs side by side"""
        # First pair
        label_cell1 = ws.cell(row=row, column=1, value=label1)
        label_cell1.font = Font(name='Arial', size=10, bold=True)
        label_cell1.alignment = self.STYLES['alignment_left']
        label_cell1.border = self.STYLES['border']
        
        value_cell1 = ws.cell(row=row, column=2, value=value1)
        value_cell1.font = self.STYLES['normal']
        value_cell1.alignment = self.STYLES['alignment_left']
        value_cell1.border = self.STYLES['border']
        ws.merge_cells(f'B{row}:C{row}')
        
        # Second pair
        label_cell2 = ws.cell(row=row, column=4, value=label2)
        label_cell2.font = Font(name='Arial', size=10, bold=True)
        label_cell2.alignment = self.STYLES['alignment_left']
        label_cell2.border = self.STYLES['border']
        
        value_cell2 = ws.cell(row=row, column=5, value=value2)
        value_cell2.font = self.STYLES['normal']
        value_cell2.alignment = self.STYLES['alignment_left']
        value_cell2.border = self.STYLES['border']
        ws.merge_cells(f'E{row}:G{row}')
    
    def _add_triple_row(self, ws, row, label1, value1, label2, value2, label3, value3):
        """Add a row with three label-value pairs"""
        # First pair
        label_cell1 = ws.cell(row=row, column=1, value=label1)
        label_cell1.font = Font(name='Arial', size=10, bold=True)
        label_cell1.alignment = self.STYLES['alignment_left']
        label_cell1.border = self.STYLES['border']
        
        value_cell1 = ws.cell(row=row, column=2, value=value1)
        value_cell1.font = self.STYLES['normal']
        value_cell1.alignment = self.STYLES['alignment_left']
        value_cell1.border = self.STYLES['border']
        
        # Second pair
        label_cell2 = ws.cell(row=row, column=3, value=label2)
        label_cell2.font = Font(name='Arial', size=10, bold=True)
        label_cell2.alignment = self.STYLES['alignment_left']
        label_cell2.border = self.STYLES['border']
        
        value_cell2 = ws.cell(row=row, column=4, value=value2)
        value_cell2.font = self.STYLES['normal']
        value_cell2.alignment = self.STYLES['alignment_left']
        value_cell2.border = self.STYLES['border']
        
        # Third pair
        label_cell3 = ws.cell(row=row, column=5, value=label3)
        label_cell3.font = Font(name='Arial', size=10, bold=True)
        label_cell3.alignment = self.STYLES['alignment_left']
        label_cell3.border = self.STYLES['border']
        
        value_cell3 = ws.cell(row=row, column=6, value=value3)
        value_cell3.font = self.STYLES['normal']
        value_cell3.alignment = self.STYLES['alignment_left']
        value_cell3.border = self.STYLES['border']
        ws.merge_cells(f'F{row}:G{row}')
    
    def _add_triple_header(self, ws, row, main_label, sub1, sub2, sub3):
        """Add a header row for triple values (Min, Normal, Maximum)"""
        # Main label
        label_cell = ws.cell(row=row, column=1, value=main_label)
        label_cell.font = Font(name='Arial', size=10, bold=True)
        label_cell.alignment = self.STYLES['alignment_left']
        label_cell.border = self.STYLES['border']
        ws.merge_cells(f'A{row}:B{row}')
        
        # Sub headers - write to columns C, E, G (before merging)
        sub_labels = [sub1, sub2, sub3]
        columns = [3, 5, 7]  # C, E, G
        
        for sub_label, col in zip(sub_labels, columns):
            cell = ws.cell(row=row, column=col, value=sub_label)
            cell.font = Font(name='Arial', size=9, bold=True, italic=True)
            cell.alignment = self.STYLES['alignment_center']
            cell.border = self.STYLES['border']
            cell.fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
        
        # Now merge cells (C:D, E:F, G stays single)
        ws.merge_cells(f'C{row}:D{row}')
        ws.merge_cells(f'E{row}:F{row}')
        # Add border to column D and F since they're merged
        ws.cell(row=row, column=4).border = self.STYLES['border']
        ws.cell(row=row, column=6).border = self.STYLES['border']
    
    def _add_triple_value_row(self, ws, row, value1, value2, value3, unit=''):
        """Add a row with three values (Min, Normal, Maximum) with unit"""
        # Empty label cells
        ws.cell(row=row, column=1).border = self.STYLES['border']
        ws.cell(row=row, column=2).border = self.STYLES['border']
        ws.merge_cells(f'A{row}:B{row}')
        
        # Values with unit - write to columns C, E, G (before merging)
        values = [value1, value2, value3]
        columns = [3, 5, 7]  # C, E, G
        
        for value, col in zip(values, columns):
            display_value = f"{value} {unit}" if value and unit else value
            cell = ws.cell(row=row, column=col, value=display_value)
            cell.font = self.STYLES['normal']
            cell.alignment = self.STYLES['alignment_center']
            cell.border = self.STYLES['border']
        
        # Now merge cells (C:D, E:F, G stays single)
        ws.merge_cells(f'C{row}:D{row}')
        ws.merge_cells(f'E{row}:F{row}')
        # Add borders to merged cells
        ws.cell(row=row, column=4).border = self.STYLES['border']
        ws.cell(row=row, column=6).border = self.STYLES['border']


def generate_mov_datasheet_excel(equipment_list):
    """
    Convenience function to generate MOV datasheet
    
    Args:
        equipment_list: List of detected MOV equipment dictionaries
        
    Returns:
        BytesIO: Excel file buffer
    """
    generator = MOVDatasheetGenerator()
    return generator.generate_datasheet(equipment_list)
