"""
Dynamic SDV Datasheet Excel Generator

Creates Process Data Sheet for Shutdown Valves from scratch
NO TEMPLATE REQUIRED - Builds entire structure programmatically
"""
import logging
from typing import Dict, List
from datetime import datetime
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


class SDVExcelGeneratorDynamic:
    """
    Generate SDV datasheets from scratch
    Builds complete table structure programmatically
    """
    
    def __init__(self):
        """Initialize generator"""
        logger.info("[SDVExcelGeneratorDynamic] Initialized - No template needed")
        
        # Define styles
        self.header_font = Font(bold=True, size=14, name='Arial')
        self.title_font = Font(bold=True, size=16, name='Arial')
        self.section_font = Font(bold=True, size=11, name='Arial')
        self.data_font = Font(size=10, name='Arial')
        
        self.center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        self.left_alignment = Alignment(horizontal='left', vertical='center')
        
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        self.header_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
        self.section_fill = PatternFill(start_color='E8E8E8', end_color='E8E8E8', fill_type='solid')
    
    def _create_header(self, ws):
        """Create document header"""
        # Row 1: COMPANY Doc No
        ws.merge_cells('A1:B1')
        ws['A1'] = 'COMPANY Doc No. :'
        ws['A1'].font = self.data_font
        ws['A1'].alignment = self.left_alignment
        ws['A1'].border = self.thin_border
        
        ws.merge_cells('C1:M1')
        ws['C1'].border = self.thin_border
        
        ws['N1'] = 'Rev. No. :'
        ws['N1'].font = self.data_font
        ws['N1'].border = self.thin_border
        
        # Row 2-3: Title
        ws.merge_cells('A2:M3')
        ws['A2'] = 'PROCESS DATA SHEET\nSHUTDOWN VALVE'
        ws['A2'].font = self.title_font
        ws['A2'].alignment = self.center_alignment
        ws['A2'].border = self.thin_border
        
        ws['N2'] = 'Date :'
        ws['N2'].font = self.data_font
        ws['N2'].border = self.thin_border
        
        ws.merge_cells('N3:N3')
        ws['N3'] = f'Page No: 2 Of 2'
        ws['N3'].font = self.data_font
        ws['N3'].border = self.thin_border
        
        # Row 4: Document Class
        ws.merge_cells('A4:B4')
        ws['A4'] = 'Document Class:'
        ws['A4'].font = self.data_font
        ws['A4'].alignment = self.left_alignment
        ws['A4'].border = self.thin_border
        
        ws.merge_cells('C4:N4')
        ws['C4'].border = self.thin_border
    
    def _create_general_data_section(self, ws, start_row):
        """Create General Data section"""
        # Section header
        ws.merge_cells(f'A{start_row}:B{start_row+5}')
        ws[f'A{start_row}'] = 'General\nData'
        ws[f'A{start_row}'].font = self.section_font
        ws[f'A{start_row}'].alignment = self.center_alignment
        ws[f'A{start_row}'].border = self.thin_border
        ws[f'A{start_row}'].fill = self.section_fill
        
        current_row = start_row
        
        # Row 1: Tag No.
        ws[f'C{current_row}'] = '1'
        ws[f'D{current_row}'] = 'Tag No.'
        ws.merge_cells(f'E{current_row}:N{current_row}')
        self._apply_cell_style(ws, f'C{current_row}', f'N{current_row}')
        current_row += 1
        
        # Row 2: Service
        ws[f'C{current_row}'] = '2'
        ws[f'D{current_row}'] = 'Service'
        ws.merge_cells(f'E{current_row}:N{current_row}')
        self._apply_cell_style(ws, f'C{current_row}', f'N{current_row}')
        current_row += 1
        
        # Row 3: P&ID No.
        ws[f'C{current_row}'] = '3'
        ws[f'D{current_row}'] = 'P&ID No.'
        ws.merge_cells(f'E{current_row}:N{current_row}')
        self._apply_cell_style(ws, f'C{current_row}', f'N{current_row}')
        current_row += 1
        
        # Row 4: Line No. | Piping class
        ws[f'C{current_row}'] = '4'
        ws[f'D{current_row}'] = 'Line No.'
        ws.merge_cells(f'E{current_row}:H{current_row}')
        ws[f'I{current_row}'] = 'Piping class'
        ws.merge_cells(f'J{current_row}:N{current_row}')
        self._apply_cell_style(ws, f'C{current_row}', f'N{current_row}')
        current_row += 1
        
        # Row 5: Sour Service | Special Service
        ws[f'C{current_row}'] = '5'
        ws[f'D{current_row}'] = 'Sour Service'
        ws.merge_cells(f'E{current_row}:H{current_row}')
        ws[f'I{current_row}'] = 'Special Service'
        ws.merge_cells(f'J{current_row}:N{current_row}')
        self._apply_cell_style(ws, f'C{current_row}', f'N{current_row}')
        current_row += 1
        
        # Row 6: Ambient Temp | Min | Max. | Unit
        ws[f'C{current_row}'] = '6'
        ws[f'D{current_row}'] = 'Ambient Temp'
        ws[f'E{current_row}'] = 'Min'
        ws.merge_cells(f'F{current_row}:F{current_row}')
        ws[f'G{current_row}'] = 'Max.'
        ws.merge_cells(f'H{current_row}:H{current_row}')
        ws[f'I{current_row}'] = 'Unit'
        ws.merge_cells(f'J{current_row}:N{current_row}')
        self._apply_cell_style(ws, f'C{current_row}', f'N{current_row}')
        
        return current_row + 1
    
    def _create_operating_conditions_section(self, ws, start_row):
        """Create Operating Conditions section"""
        # Section header
        ws.merge_cells(f'A{start_row}:B{start_row+4}')
        ws[f'A{start_row}'] = 'Operating\nConditions'
        ws[f'A{start_row}'].font = self.section_font
        ws[f'A{start_row}'].alignment = self.center_alignment
        ws[f'A{start_row}'].border = self.thin_border
        ws[f'A{start_row}'].fill = self.section_fill
        
        current_row = start_row
        
        # Row 7: Fluid | Phase | State
        ws[f'C{current_row}'] = '7'
        ws[f'D{current_row}'] = 'Fluid'
        ws.merge_cells(f'E{current_row}:E{current_row}')
        ws[f'F{current_row}'] = 'Phase'
        ws.merge_cells(f'G{current_row}:G{current_row}')
        ws[f'H{current_row}'] = 'State'
        ws.merge_cells(f'I{current_row}:N{current_row}')
        self._apply_cell_style(ws, f'C{current_row}', f'N{current_row}')
        current_row += 1
        
        # Row 8: Press. | Normal | Design | Unit
        ws[f'C{current_row}'] = '8'
        ws[f'D{current_row}'] = 'Press.'
        ws[f'E{current_row}'] = 'Normal'
        ws.merge_cells(f'F{current_row}:F{current_row}')
        ws[f'G{current_row}'] = 'Design'
        ws.merge_cells(f'H{current_row}:H{current_row}')
        ws[f'I{current_row}'] = 'Unit'
        ws.merge_cells(f'J{current_row}:N{current_row}')
        self._apply_cell_style(ws, f'C{current_row}', f'N{current_row}')
        current_row += 1
        
        # Row 9: Temperature | Min | Max. | Unit
        ws[f'C{current_row}'] = '9'
        ws[f'D{current_row}'] = 'Temperature'
        ws[f'E{current_row}'] = 'Min'
        ws.merge_cells(f'F{current_row}:F{current_row}')
        ws[f'G{current_row}'] = 'Max.'
        ws.merge_cells(f'H{current_row}:H{current_row}')
        ws[f'I{current_row}'] = 'Unit'
        ws.merge_cells(f'J{current_row}:N{current_row}')
        self._apply_cell_style(ws, f'C{current_row}', f'N{current_row}')
        current_row += 1
        
        # Row 10: Design Temp. | Min | Max. | Unit
        ws[f'C{current_row}'] = '10'
        ws[f'D{current_row}'] = 'Design Temp.'
        ws[f'E{current_row}'] = 'Min'
        ws.merge_cells(f'F{current_row}:F{current_row}')
        ws[f'G{current_row}'] = 'Max.'
        ws.merge_cells(f'H{current_row}:H{current_row}')
        ws[f'I{current_row}'] = 'Unit'
        ws.merge_cells(f'J{current_row}:N{current_row}')
        self._apply_cell_style(ws, f'C{current_row}', f'N{current_row}')
        current_row += 1
        
        # Row 11: Shut Off Pressure
        ws[f'C{current_row}'] = '11'
        ws[f'D{current_row}'] = 'Shut Off Pressure'
        ws.merge_cells(f'E{current_row}:N{current_row}')
        self._apply_cell_style(ws, f'C{current_row}', f'N{current_row}')
        
        return current_row + 1
    
    def _create_valve_details_section(self, ws, start_row):
        """Create Valve Details section"""
        # Section header
        ws.merge_cells(f'A{start_row}:B{start_row+1}')
        ws[f'A{start_row}'] = 'Valve\nDetails'
        ws[f'A{start_row}'].font = self.section_font
        ws[f'A{start_row}'].alignment = self.center_alignment
        ws[f'A{start_row}'].border = self.thin_border
        ws[f'A{start_row}'].fill = self.section_fill
        
        current_row = start_row
        
        # Row 12: Bore Detail
        ws[f'C{current_row}'] = '12'
        ws[f'D{current_row}'] = 'Bore Detail'
        ws.merge_cells(f'E{current_row}:N{current_row}')
        self._apply_cell_style(ws, f'C{current_row}', f'N{current_row}')
        current_row += 1
        
        # Row 13: Mech. Handwheel
        ws[f'C{current_row}'] = '13'
        ws[f'D{current_row}'] = 'Mech. Handwheel'
        ws.merge_cells(f'E{current_row}:N{current_row}')
        self._apply_cell_style(ws, f'C{current_row}', f'N{current_row}')
        
        return current_row + 1
    
    def _create_actuator_details_section(self, ws, start_row):
        """Create Actuator Details section"""
        # Section header
        ws.merge_cells(f'A{start_row}:B{start_row+3}')
        ws[f'A{start_row}'] = 'Actuator\nDetails'
        ws[f'A{start_row}'].font = self.section_font
        ws[f'A{start_row}'].alignment = self.center_alignment
        ws[f'A{start_row}'].border = self.thin_border
        ws[f'A{start_row}'].fill = self.section_fill
        
        current_row = start_row
        
        # Row 14: Air Fail position
        ws[f'C{current_row}'] = '14'
        ws[f'D{current_row}'] = 'Air Fail position'
        ws.merge_cells(f'E{current_row}:N{current_row}')
        self._apply_cell_style(ws, f'C{current_row}', f'N{current_row}')
        current_row += 1
        
        # Row 15: Valve Close Time | Valve Open Time
        ws[f'C{current_row}'] = '15'
        ws[f'D{current_row}'] = 'Valve Close Time'
        ws.merge_cells(f'E{current_row}:G{current_row}')
        ws[f'H{current_row}'] = 'Valve Open Time'
        ws.merge_cells(f'I{current_row}:N{current_row}')
        self._apply_cell_style(ws, f'C{current_row}', f'N{current_row}')
        current_row += 1
        
        # Row 16: Design Pressure
        ws[f'C{current_row}'] = '16'
        ws[f'D{current_row}'] = 'Design Pressure'
        ws.merge_cells(f'E{current_row}:N{current_row}')
        self._apply_cell_style(ws, f'C{current_row}', f'N{current_row}')
        current_row += 1
        
        # Row 17: Seat Leakage Class
        ws[f'C{current_row}'] = '17'
        ws[f'D{current_row}'] = 'Seat Leakage Class'
        ws.merge_cells(f'E{current_row}:N{current_row}')
        self._apply_cell_style(ws, f'C{current_row}', f'N{current_row}')
        
        return current_row + 1
    
    def _create_accessories_section(self, ws, start_row):
        """Create Accessories section"""
        # Section header
        ws[f'A{start_row}'] = 'Accessories'
        ws[f'A{start_row}'].font = self.section_font
        ws[f'A{start_row}'].alignment = self.center_alignment
        ws[f'A{start_row}'].border = self.thin_border
        ws[f'A{start_row}'].fill = self.section_fill
        ws.merge_cells(f'A{start_row}:B{start_row}')
        
        # Row 18: NACE Requirement
        ws[f'C{start_row}'] = '18'
        ws[f'D{start_row}'] = 'NACE Requirement'
        ws.merge_cells(f'E{start_row}:N{start_row}')
        self._apply_cell_style(ws, f'C{start_row}', f'N{start_row}')
        
        return start_row + 1
    
    def _apply_cell_style(self, ws, start_cell, end_cell):
        """Apply borders and font to a range of cells"""
        for row in ws[f'{start_cell}:{end_cell}']:
            for cell in row:
                cell.border = self.thin_border
                cell.font = self.data_font
                cell.alignment = self.left_alignment
    
    def _fill_data(self, ws, valve_data: Dict):
        """Fill data into the created structure"""
        # Header data
        ws['C1'] = valve_data.get('document_no', '')
        ws['N1'] = valve_data.get('rev_no', 'A')
        ws['N2'] = valve_data.get('date', datetime.now().strftime('%d-%b-%Y'))
        
        # General Data (starting at row 5)
        ws['E5'] = valve_data.get('tag_no', '')
        ws['E6'] = valve_data.get('service', '')
        ws['E7'] = valve_data.get('pid_no', '')
        ws['E8'] = valve_data.get('line_no', '')
        ws['J8'] = valve_data.get('piping_class', '')
        ws['E9'] = valve_data.get('sour_service', '')
        ws['J9'] = valve_data.get('special_service', '')
        ws['F10'] = valve_data.get('ambient_temp_min', '')
        ws['H10'] = valve_data.get('ambient_temp_max', '')
        ws['J10'] = valve_data.get('ambient_temp_unit', '°C')
        
        # Operating Conditions (starting at row 11)
        ws['E11'] = valve_data.get('fluid', '')
        ws['G11'] = valve_data.get('phase', '')
        ws['I11'] = valve_data.get('state', '')
        ws['F12'] = valve_data.get('operating_pressure_normal', '')
        ws['H12'] = valve_data.get('operating_pressure_design', '')
        ws['J12'] = valve_data.get('pressure_unit', 'barg')
        ws['F13'] = valve_data.get('operating_temp_min', '')
        ws['H13'] = valve_data.get('operating_temp_max', '')
        ws['J13'] = valve_data.get('operating_temp_unit', '°C')
        ws['F14'] = valve_data.get('design_temp_min', '')
        ws['H14'] = valve_data.get('design_temp_max', '')
        ws['J14'] = valve_data.get('design_temp_unit', '°C')
        ws['E15'] = valve_data.get('shut_off_pressure', '')
        
        # Valve Details (starting at row 16)
        ws['E16'] = valve_data.get('bore_detail', '')
        ws['E17'] = valve_data.get('mech_handwheel', '')
        
        # Actuator Details (starting at row 18)
        ws['E18'] = valve_data.get('fail_position', '')
        ws['E19'] = valve_data.get('valve_close_time', '')
        ws['I19'] = valve_data.get('valve_open_time', '')
        ws['E20'] = valve_data.get('design_pressure', '')
        ws['E21'] = valve_data.get('seat_leakage_class', '')
        
        # Accessories (row 22)
        ws['E22'] = valve_data.get('nace_requirement', '')
    
    def generate_datasheet(self, mapped_data: Dict, output_filename: str = None) -> BytesIO:
        """
        Generate SDV datasheet Excel from scratch
        
        Args:
            mapped_data: AI-mapped data from SDVDatasheetAIMapper
            output_filename: Optional filename for output
        
        Returns:
            BytesIO: Excel file in memory
        """
        logger.info("[SDVExcelGeneratorDynamic] Creating Excel from scratch...")
        
        try:
            # Get first valve data
            valves = mapped_data.get('valves', [])
            if not valves:
                raise ValueError("No valve data to generate datasheet")
            
            valve_data = valves[0]
            logger.info(f"[SDVExcelGeneratorDynamic] Generating datasheet for valve: {valve_data.get('tag_no', 'Unknown')}")
            
            # Create new workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "SDV Datasheet"
            
            # Set column widths
            ws.column_dimensions['A'].width = 8
            ws.column_dimensions['B'].width = 8
            ws.column_dimensions['C'].width = 3
            ws.column_dimensions['D'].width = 18
            for col in ['E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N']:
                ws.column_dimensions[col].width = 10
            
            # Build structure
            self._create_header(ws)
            
            current_row = 5
            current_row = self._create_general_data_section(ws, current_row)
            current_row = self._create_operating_conditions_section(ws, current_row)
            current_row = self._create_valve_details_section(ws, current_row)
            current_row = self._create_actuator_details_section(ws, current_row)
            self._create_accessories_section(ws, current_row)
            
            # Fill with data
            self._fill_data(ws, valve_data)
            
            # Add notes row
            ws.merge_cells(f'A{current_row+1}:N{current_row+1}')
            ws[f'A{current_row+1}'] = 'Notes:'
            ws[f'A{current_row+1}'].border = self.thin_border
            
            # Save to BytesIO
            excel_buffer = BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)
            
            logger.info("[SDVExcelGeneratorDynamic] ✅ Excel generated successfully")
            return excel_buffer
            
        except Exception as e:
            logger.error(f"[SDVExcelGeneratorDynamic] ❌ Error: {str(e)}")
            raise
