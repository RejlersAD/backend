"""
Excel Generator for Pressure Instrument Datasheets
"""
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

class PressureExcelGeneratorDynamic:
    def generate(self, data):
        """Generate Excel file for pressure instruments"""
        instruments = data.get('instruments', [])
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Pressure Instruments"
        
        # Header styling
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Define headers matching the PressureInstrumentAnalyzer output
        headers = [
            "Tag Number", "P&ID No", "Line No.", "Piping Class", "Equipment No.",
            "Service", "Fluid State", "Fluid Phase", "Operating Pressure (Min)",
            "Operating Pressure (Norm)", "Operating Pressure (Max)", "Operating Temp (Min)",
            "Operating Temp (Norm)", "Operating Temp (Max)", "Operating Diff. Pressure",
            "Design Pressure (Min)", "Design Pressure (Norm)", "Design Pressure (Max)",
            "Source Service", "Special Conditions", "Density (Min)", "Density (Norm)",
            "Density (Max)", "Viscosity (Min)", "Viscosity (Norm)", "Viscosity (Max)",
            "Gauge Adaptor", "NACE Requirement", "Notes"
        ]
        
        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border
        
        # Write data
        for row_num, inst in enumerate(instruments, 2):
            ws.cell(row=row_num, column=1, value=inst.get('tag_number', '')).border = border
            ws.cell(row=row_num, column=2, value=inst.get('pid_no', '')).border = border
            ws.cell(row=row_num, column=3, value=inst.get('line_no', '')).border = border
            ws.cell(row=row_num, column=4, value=inst.get('piping_class', '')).border = border
            ws.cell(row=row_num, column=5, value=inst.get('equipment_no', '')).border = border
            ws.cell(row=row_num, column=6, value=inst.get('service', '')).border = border
            ws.cell(row=row_num, column=7, value=inst.get('fluid_state', '')).border = border
            ws.cell(row=row_num, column=8, value=inst.get('fluid_phase', '')).border = border
            ws.cell(row=row_num, column=9, value=inst.get('operating_pressure_min', '')).border = border
            ws.cell(row=row_num, column=10, value=inst.get('operating_pressure_norm', '')).border = border
            ws.cell(row=row_num, column=11, value=inst.get('operating_pressure_max', '')).border = border
            ws.cell(row=row_num, column=12, value=inst.get('operating_temp_min', '')).border = border
            ws.cell(row=row_num, column=13, value=inst.get('operating_temp_norm', '')).border = border
            ws.cell(row=row_num, column=14, value=inst.get('operating_temp_max', '')).border = border
            ws.cell(row=row_num, column=15, value=inst.get('operating_differential_pressure', '')).border = border
            ws.cell(row=row_num, column=16, value=inst.get('design_pressure_min', '')).border = border
            ws.cell(row=row_num, column=17, value=inst.get('design_pressure_norm', '')).border = border
            ws.cell(row=row_num, column=18, value=inst.get('design_pressure_max', '')).border = border
            ws.cell(row=row_num, column=19, value=inst.get('source_service', '')).border = border
            ws.cell(row=row_num, column=20, value=inst.get('special_conditions', '')).border = border
            ws.cell(row=row_num, column=21, value=inst.get('density_min', '')).border = border
            ws.cell(row=row_num, column=22, value=inst.get('density_norm', '')).border = border
            ws.cell(row=row_num, column=23, value=inst.get('density_max', '')).border = border
            ws.cell(row=row_num, column=24, value=inst.get('viscosity_min', '')).border = border
            ws.cell(row=row_num, column=25, value=inst.get('viscosity_norm', '')).border = border
            ws.cell(row=row_num, column=26, value=inst.get('viscosity_max', '')).border = border
            ws.cell(row=row_num, column=27, value=inst.get('gauge_adaptor', '')).border = border
            ws.cell(row=row_num, column=28, value=inst.get('nace_requirement', '')).border = border
            ws.cell(row=row_num, column=29, value=inst.get('notes', '')).border = border
            ws.cell(row=row_num, column=6, value=inst.get('operating_pressure', '')).border = border
            ws.cell(row=row_num, column=7, value=inst.get('design_pressure', '')).border = border
            ws.cell(row=row_num, column=8, value=inst.get('temperature', '')).border = border
            ws.cell(row=row_num, column=9, value=inst.get('piping_class', '')).border = border
            ws.cell(row=row_num, column=29, value=inst.get('notes', '')).border = border
        
        # Set column widths for better visibility
        column_widths = [15, 15, 15, 12, 15, 30, 12, 12, 10, 10, 10, 10, 10, 10, 15,
                        10, 10, 10, 20, 20, 10, 10, 10, 10, 10, 10, 15, 15, 30]
        for idx, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + idx) if idx < 27 else f'A{chr(64 + idx - 26)}'].width = width
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()
