"""Phase 1 — Excel BOQ import.

Detects the header row, maps columns via the soft-coded BOQ_HEADER_SYNONYMS,
streams each data row into EstimateLineItem and returns a summary dict.

Designed to be Celery-safe (no I/O outside the call) and idempotent: every
import creates a new Estimate row with version = max(version)+1 so previous
imports remain auditable.
"""
from __future__ import annotations

import logging
from decimal import Decimal, InvalidOperation
from typing import Dict, List, Optional, Tuple

from django.db import transaction
from django.utils import timezone

from ..config import BOQ_HEADER_SCAN_ROWS, BOQ_HEADER_SYNONYMS
from ..models import Estimate, EstimateLineItem, ProjectDocument

logger = logging.getLogger(__name__)


def _to_decimal(value, default=Decimal('0')) -> Decimal:
    if value is None or value == '':
        return default
    if isinstance(value, Decimal):
        return value
    try:
        return Decimal(str(value).strip().replace(',', ''))
    except (InvalidOperation, ValueError, AttributeError):
        return default


def _normalise(s) -> str:
    return ' '.join(str(s or '').strip().lower().split())


def _detect_header_row(rows: List[List]) -> Tuple[int, Dict[str, int]]:
    """Scan the first BOQ_HEADER_SCAN_ROWS rows and pick the row with the
    most synonym hits. Returns (header_row_index, {canonical_name: col_idx}).
    """
    best_row = -1
    best_map: Dict[str, int] = {}
    best_score = 0

    scan_limit = min(BOQ_HEADER_SCAN_ROWS, len(rows))
    for r in range(scan_limit):
        cells = [_normalise(c) for c in rows[r]]
        mapping: Dict[str, int] = {}
        for canonical, synonyms in BOQ_HEADER_SYNONYMS.items():
            for col_idx, cell in enumerate(cells):
                if not cell:
                    continue
                if any(syn in cell for syn in synonyms):
                    mapping.setdefault(canonical, col_idx)
                    break
        if len(mapping) > best_score:
            best_score = len(mapping)
            best_map = mapping
            best_row = r
        if best_score == len(BOQ_HEADER_SYNONYMS):
            break

    return best_row, best_map


def _load_workbook_rows(file_obj) -> List[List]:
    """Read an .xlsx/.xls file into a plain list-of-lists for the first sheet
    with data. Lazy-imports openpyxl so the dependency stays optional at boot.
    """
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError(
            'openpyxl is required for Excel BOQ import. '
            'Add `openpyxl` to backend/requirements.txt.'
        ) from exc

    file_obj.seek(0)
    wb = load_workbook(file_obj, data_only=True, read_only=True)
    rows: List[List] = []
    for sheet in wb.worksheets:
        for row in sheet.iter_rows(values_only=True):
            rows.append(list(row))
        if rows:
            break
    wb.close()
    return rows


@transaction.atomic
def import_boq_excel(
    *,
    project,
    file_obj,
    kind: str = 'estimate',
    title: str = '',
    notes: str = '',
    user=None,
    source_document: Optional[ProjectDocument] = None,
) -> Dict:
    """Import a BOQ Excel into a new Estimate. Returns a summary dict."""
    rows = _load_workbook_rows(file_obj)
    if not rows:
        raise ValueError('Excel file appears to be empty.')

    header_row, header_map = _detect_header_row(rows)
    if header_row < 0 or 'description' not in header_map:
        raise ValueError(
            'Could not detect BOQ header row. '
            f'Recognised columns: {list(BOQ_HEADER_SYNONYMS.keys())}'
        )

    data_rows = rows[header_row + 1:]

    next_version = (
        Estimate.objects
        .filter(project=project, kind=kind, is_deleted=False)
        .order_by('-version').values_list('version', flat=True).first()
    ) or 0
    next_version += 1

    estimate = Estimate.objects.create(
        project=project,
        version=next_version,
        kind=kind,
        source='excel',
        status='draft',
        title=title or f'BOQ v{next_version} ({timezone.now():%Y-%m-%d})',
        notes=notes,
        source_document=source_document,
        created_by=user if (user and getattr(user, 'is_authenticated', False)) else None,
    )

    items: List[EstimateLineItem] = []
    running_total = Decimal('0')
    sort_idx = 0
    skipped = 0

    for raw in data_rows:
        if not raw or all((c is None or str(c).strip() == '') for c in raw):
            continue

        def cell(canonical):
            idx = header_map.get(canonical)
            if idx is None or idx >= len(raw):
                return None
            return raw[idx]

        description = str(cell('description') or '').strip()
        wbs = str(cell('wbs') or '').strip()
        if not description and not wbs:
            skipped += 1
            continue

        qty = _to_decimal(cell('quantity'))
        rate = _to_decimal(cell('unit_rate'))
        line_total = _to_decimal(cell('line_total')) or (qty * rate).quantize(Decimal('0.01'))

        items.append(EstimateLineItem(
            estimate=estimate,
            wbs_code=wbs[:64],
            description=description,
            discipline=(str(cell('discipline') or '').strip())[:64],
            unit=(str(cell('unit') or '').strip())[:32],
            quantity=qty,
            unit_rate=rate,
            line_total=line_total,
            sort_order=sort_idx,
            source_row={str(i): (None if v is None else str(v)) for i, v in enumerate(raw)},
        ))
        running_total += line_total
        sort_idx += 1

    if items:
        EstimateLineItem.objects.bulk_create(items, batch_size=500)

    estimate.total_amount = running_total
    estimate.save(update_fields=['total_amount', 'updated_at'])

    summary = {
        'estimate_id': estimate.id,
        'version': estimate.version,
        'kind': estimate.kind,
        'header_row_index': header_row,
        'detected_columns': header_map,
        'imported_rows': len(items),
        'skipped_rows': skipped,
        'total_amount': str(running_total),
        'currency': estimate.currency,
    }
    logger.info('BOQ import complete: %s', summary)
    return summary
