"""Consolidated token-usage report generator.

Produces either an Excel workbook (3 sheets) or a PDF (summary + table + chart)
from a queryset of `PidCheckerV2UsageLog` rows.  All layout constants are
module-level to keep tuning centralised.
"""
from __future__ import annotations

from collections import defaultdict
from datetime import datetime, date
from decimal import Decimal
from io import BytesIO
from typing import Iterable

# ── Soft-coded layout ────────────────────────────────────────────────
XLSX_OVERVIEW_SHEET = 'Overview'
XLSX_RUNS_SHEET = 'Runs'
XLSX_DAILY_SHEET = 'Daily'
COST_DECIMALS = 6
TOKEN_HEADER_FILL = 'FF1F4E78'
TOKEN_HEADER_FONT_COLOR = 'FFFFFFFF'

PDF_TITLE = 'P&ID Checker V2 — Token Usage Report'
PDF_PAGE_SIZE_LANDSCAPE = True
PDF_TABLE_MAX_ROWS = 25    # top-N runs shown in PDF; Excel has full list

FEATURE_DISPLAY = {
    'line_extraction': 'Line Extraction (Vision)',
    'equipment_extraction': 'Equipment Extraction (Vision)',
    'instrument_extraction': 'Instrument Extraction (Vision)',
    'line_list_cross_check': 'Line-List Cross-Check (AI)',
    'equipment_cross_check': 'Equipment Cross-Check (AI)',
    'instrument_cross_check': 'Instrument Cross-Check (AI)',
}


def _pretty_feature(f: str) -> str:
    return FEATURE_DISPLAY.get(f, f or '—')


def _aggregate(logs) -> dict:
    total_in = total_out = total_calls = 0
    total_cost = Decimal('0')
    by_feature: dict[str, dict] = defaultdict(lambda: {'calls': 0, 'input': 0, 'output': 0, 'cost': Decimal('0')})
    by_model: dict[tuple, dict] = defaultdict(lambda: {'calls': 0, 'input': 0, 'output': 0, 'cost': Decimal('0')})
    daily: dict[date, dict] = defaultdict(lambda: {'calls': 0, 'input': 0, 'output': 0, 'cost': Decimal('0')})

    for r in logs:
        total_in += r.input_tokens
        total_out += r.output_tokens
        total_calls += r.call_count
        total_cost += r.cost_usd or Decimal('0')

        f = by_feature[r.feature or '']
        f['calls'] += r.call_count
        f['input'] += r.input_tokens
        f['output'] += r.output_tokens
        f['cost'] += r.cost_usd or Decimal('0')

        key = (r.provider or '', r.model_name or '')
        m = by_model[key]
        m['calls'] += r.call_count
        m['input'] += r.input_tokens
        m['output'] += r.output_tokens
        m['cost'] += r.cost_usd or Decimal('0')

        d = r.created_at.date() if r.created_at else date.today()
        row = daily[d]
        row['calls'] += r.call_count
        row['input'] += r.input_tokens
        row['output'] += r.output_tokens
        row['cost'] += r.cost_usd or Decimal('0')

    return {
        'total_calls': total_calls,
        'total_input': total_in,
        'total_output': total_out,
        'total_tokens': total_in + total_out,
        'total_cost': total_cost,
        'by_feature': dict(by_feature),
        'by_model': dict(by_model),
        'daily': dict(daily),
    }


# ─────────────────────────────────────────────────────────────────────
# Excel
# ─────────────────────────────────────────────────────────────────────
def build_xlsx(logs) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    header_font = Font(bold=True, color=TOKEN_HEADER_FONT_COLOR)
    header_fill = PatternFill('solid', fgColor=TOKEN_HEADER_FILL)
    center = Alignment(horizontal='center', vertical='center')

    def _style_header(ws, ncols):
        for col in range(1, ncols + 1):
            c = ws.cell(row=1, column=col)
            c.font = header_font
            c.fill = header_fill
            c.alignment = center

    agg = _aggregate(logs)

    # Overview
    ws1 = wb.active
    ws1.title = XLSX_OVERVIEW_SHEET
    ws1.append(['Metric', 'Value'])
    _style_header(ws1, 2)
    ws1.append(['Total AI calls', agg['total_calls']])
    ws1.append(['Total input tokens', agg['total_input']])
    ws1.append(['Total output tokens', agg['total_output']])
    ws1.append(['Total tokens', agg['total_tokens']])
    ws1.append(['Total cost (USD)', float(agg['total_cost'])])
    ws1.append([])
    ws1.append(['Feature', 'Calls', 'Input tokens', 'Output tokens', 'Cost (USD)'])
    hdr_row = ws1.max_row
    for col in range(1, 6):
        c = ws1.cell(row=hdr_row, column=col)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center
    for feat, v in sorted(agg['by_feature'].items()):
        ws1.append([_pretty_feature(feat), v['calls'], v['input'], v['output'], float(v['cost'])])
    ws1.append([])
    ws1.append(['Provider', 'Model', 'Calls', 'Input tokens', 'Output tokens', 'Cost (USD)'])
    hdr_row = ws1.max_row
    for col in range(1, 7):
        c = ws1.cell(row=hdr_row, column=col)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center
    for (prov, model), v in sorted(agg['by_model'].items()):
        ws1.append([prov, model, v['calls'], v['input'], v['output'], float(v['cost'])])
    for col_letter, width in (('A', 32), ('B', 24), ('C', 12), ('D', 16), ('E', 16), ('F', 14)):
        ws1.column_dimensions[col_letter].width = width

    # Runs
    ws2 = wb.create_sheet(XLSX_RUNS_SHEET)
    ws2.append(['Timestamp', 'Feature', 'Provider', 'Model', 'Calls',
                'Input tokens', 'Output tokens', 'Total tokens', 'Cost (USD)',
                'Extraction ID', 'Upload ID'])
    _style_header(ws2, 11)
    for r in logs:
        ws2.append([
            r.created_at.strftime('%Y-%m-%d %H:%M:%S') if r.created_at else '',
            _pretty_feature(r.feature),
            r.provider,
            r.model_name,
            r.call_count,
            r.input_tokens,
            r.output_tokens,
            r.total_tokens,
            float(r.cost_usd or Decimal('0')),
            str(r.related_extraction_id) if r.related_extraction_id else '',
            str(r.related_upload_id) if r.related_upload_id else '',
        ])
    for i, width in enumerate([20, 28, 12, 28, 8, 14, 14, 14, 14, 38, 38], start=1):
        ws2.column_dimensions[chr(64 + i)].width = width

    # Daily
    ws3 = wb.create_sheet(XLSX_DAILY_SHEET)
    ws3.append(['Date', 'Calls', 'Input tokens', 'Output tokens', 'Total tokens', 'Cost (USD)'])
    _style_header(ws3, 6)
    for d in sorted(agg['daily'].keys()):
        v = agg['daily'][d]
        ws3.append([d.isoformat(), v['calls'], v['input'], v['output'],
                    v['input'] + v['output'], float(v['cost'])])
    for col_letter, width in (('A', 14), ('B', 10), ('C', 16), ('D', 16), ('E', 14), ('F', 14)):
        ws3.column_dimensions[col_letter].width = width

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─────────────────────────────────────────────────────────────────────
# PDF
# ─────────────────────────────────────────────────────────────────────
def build_pdf(logs) -> bytes:
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak,
    )

    logs = list(logs)
    agg = _aggregate(logs)

    buf = BytesIO()
    page = landscape(A4) if PDF_PAGE_SIZE_LANDSCAPE else A4
    doc = SimpleDocTemplate(buf, pagesize=page,
                            leftMargin=1.2 * cm, rightMargin=1.2 * cm,
                            topMargin=1.2 * cm, bottomMargin=1.2 * cm)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=18, spaceAfter=8)
    small = ParagraphStyle('Small', parent=styles['Normal'], fontSize=9)

    story = [
        Paragraph(PDF_TITLE, title_style),
        Paragraph(f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')} · "
                  f"Runs: {len(logs)}", small),
        Spacer(1, 0.4 * cm),
    ]

    # Totals card
    tot = [
        ['Total AI calls', str(agg['total_calls'])],
        ['Total input tokens', f"{agg['total_input']:,}"],
        ['Total output tokens', f"{agg['total_output']:,}"],
        ['Total tokens', f"{agg['total_tokens']:,}"],
        ['Total cost (USD)', f"${agg['total_cost']:.{COST_DECIMALS}f}"],
    ]
    t = Table(tot, colWidths=[5 * cm, 5 * cm])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#1F4E78')),
        ('TEXTCOLOR', (0, 0), (0, -1), colors.white),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('GRID', (0, 0), (-1, -1), 0.4, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 8),
        ('RIGHTPADDING', (0, 0), (-1, -1), 8),
    ]))
    story.append(t)
    story.append(Spacer(1, 0.6 * cm))

    # Breakdown by feature
    story.append(Paragraph('<b>By feature</b>', styles['Heading3']))
    rows = [['Feature', 'Calls', 'Input', 'Output', 'Cost (USD)']]
    for feat, v in sorted(agg['by_feature'].items()):
        rows.append([_pretty_feature(feat), v['calls'], v['input'],
                     v['output'], f"${v['cost']:.{COST_DECIMALS}f}"])
    t = Table(rows, colWidths=[7 * cm, 2.2 * cm, 2.6 * cm, 2.6 * cm, 3 * cm])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E78')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('GRID', (0, 0), (-1, -1), 0.3, colors.grey),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
    ]))
    story.append(t)
    story.append(Spacer(1, 0.5 * cm))

    # Top N runs
    story.append(Paragraph(f'<b>Recent runs (top {PDF_TABLE_MAX_ROWS})</b>', styles['Heading3']))
    rows = [['Timestamp', 'Feature', 'Provider', 'Model', 'Tokens', 'Cost (USD)']]
    for r in logs[:PDF_TABLE_MAX_ROWS]:
        rows.append([
            r.created_at.strftime('%Y-%m-%d %H:%M') if r.created_at else '',
            _pretty_feature(r.feature),
            r.provider,
            (r.model_name or '')[:32],
            f"{r.total_tokens:,}",
            f"${(r.cost_usd or Decimal('0')):.{COST_DECIMALS}f}",
        ])
    t = Table(rows, colWidths=[3.5 * cm, 5.4 * cm, 2 * cm, 6 * cm, 2.4 * cm, 2.8 * cm])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E78')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('GRID', (0, 0), (-1, -1), 0.3, colors.grey),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('ALIGN', (4, 0), (-1, -1), 'RIGHT'),
    ]))
    story.append(t)

    doc.build(story)
    return buf.getvalue()
