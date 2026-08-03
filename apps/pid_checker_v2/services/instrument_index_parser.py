"""Parse an Excel Instrument Index into canonical rows.

The Instrument Index xlsx is more complex than the Line/Equipment List:
  * a few free-text title rows at the top (company / project / doc no / P&ID ref)
  * a REFERENCE DOCUMENTS block (rows 5..~7) — ignored for data extraction
  * a TWO-ROW composite header (primary + secondary label row)
  * ONE INSTRUMENT = TWO ROWS
        primary row → SL.No, Tag Number, Instrument Type, PID Number,
                      EQPT Number, Ex Class, Instrument Range (min/max/unit),
                      Calibration Range (min/max/unit), Datasheet No.,
                      Hook up Dwg No, Manufacturer, Remarks, Rev.
        secondary row → Service Description, Line No., Location, Power Supply,
                        Loop Dwg No., Location Layout No, Model
  * blank / notes rows at the bottom

We auto-detect the header by scanning for the row that contains at least
`MIN_HEADER_MATCHES` distinctive tokens (Tag Number / Instrument Type /
Datasheet / Ex Class …). We then merge each primary + secondary row into
one canonical dict.

Every threshold / regex / label is soft-coded at module level.
"""
from __future__ import annotations

import io
import logging
import re
from typing import Any, Optional

logger = logging.getLogger(__name__)


# ─── Soft-coded config ────────────────────────────────────────────────
# Column aliases — case-insensitive substring match against the header text.
# Some columns live on the PRIMARY header row, some on the SECONDARY row.
PRIMARY_COL_ALIASES: dict[str, tuple[str, ...]] = {
    'sno':             ('sl.no', 'sl. no', 's.no', 's. no', 'sno', 'sr.no'),
    'tag':             ('tag number', 'tag no', 'instrument tag'),
    'instrument_type': ('instrument type', 'type of instrument'),
    'pid_no':          ('pid number', 'p&id number', 'p&id no', 'pid no'),
    'eqpt_no':         ('eqpt number', 'equipment number', 'eqpt no'),
    'ex_class':        ('ex class', 'ex. class', 'ex-class', 'ex classification'),
    'datasheet_no':    ('datasheet no', 'data sheet no', 'ds no'),
    'hookup_dwg_no':   ('hook up dwg', 'hookup dwg', 'hook-up dwg'),
    'manufacturer':    ('manufacturer', 'make'),
    'remarks':         ('remarks', 'notes'),
    'rev':             ('rev.', 'rev', 'revision'),
}

# These columns only appear on the SECONDARY (sub) header row
SECONDARY_COL_ALIASES: dict[str, tuple[str, ...]] = {
    'service_description': ('service description', 'service'),
    'line_no':             ('line no', 'line number'),
    'location':            ('location',),
    'power_supply':        ('power supply', 'supply'),
    'loop_dwg_no':         ('loop dwg', 'loop drawing'),
    'location_layout_no':  ('location layout', 'layout no', 'layout drawing'),
    'model':               ('model',),
}

# Range columns live under a group header ("Instrument Range" / "Calibration Range")
# spanning three sub-header cells: Min | Max | Unit.
RANGE_GROUP_INSTRUMENT = ('instrument range', 'process range')
RANGE_GROUP_CALIBRATION = ('calibration range', 'cal range', 'calibrated range')

# Which source row supplies each canonical field.
PRIMARY_KEYS = frozenset({
    'sno', 'tag', 'instrument_type', 'pid_no', 'eqpt_no', 'ex_class',
    'datasheet_no', 'hookup_dwg_no', 'manufacturer', 'remarks', 'rev',
    'range_min', 'range_max', 'range_unit',
    'cal_min', 'cal_max', 'cal_unit',
})
SECONDARY_KEYS = frozenset({
    'service_description', 'line_no', 'location', 'power_supply',
    'loop_dwg_no', 'location_layout_no', 'model',
})

# Tokens used to detect the primary header row.
REQUIRED_HEADER_TOKENS = (
    'tag number', 'instrument type', 'datasheet', 'ex class', 'eqpt number',
)
MIN_HEADER_MATCHES = 2

# How far down to scan for the header before giving up
MAX_HEADER_SCAN_ROWS = 30

# When encountering these labels in column A/B, stop reading data rows
STOP_MARKERS = ('notes:', 'hold list:', 'abbreviation:', 'abbreviations:',
                'legend:', 'reference document', 'reference drawing')

# Instrument tag normalisation — uppercase + collapse whitespace.
INSTRUMENT_TAG_REGEX = re.compile(r'^[A-Z]{1,4}-\d{2,4}[A-Z]?(?:\s?[A-Z]{2})?$')

# Fields that carry useful sniff info about the source document (top matter)
SOURCE_META_KEYS = ('title', 'doc_no', 'date', 'pid_extract_ref',
                    'sheet_name', 'company', 'project')

# Words used to hint that a top-of-file row is metadata
META_HINTS_DOC_NO = ('document no', 'doc no', 'doc. no', 'document number')
META_HINTS_PROJECT = ('project:', 'project ', 'engineering for')
META_HINTS_TITLE = ('instrument', 'index')
META_HINTS_PID = ('p&id:', 'p&id no', 'p&id dwg', 'p&id drawing')


class ParseError(Exception):
    pass


def parse_instrument_index(file_bytes: bytes, filename: str = '') -> dict:
    """Parse an xlsx file into a dict with:
        {
          rows: [ {
              tag, instrument_type, service_description, pid_no, line_no,
              eqpt_no, location, ex_class, power_supply,
              range_min, range_max, range_unit,
              cal_min, cal_max, cal_unit,
              datasheet_no, loop_dwg_no, hookup_dwg_no, location_layout_no,
              manufacturer, model, remarks, rev, _excel_row,
          } ],
          columns: { canonical_key: excel_column_index },
          header_row_index: int,
          meta: { title, doc_no, date, pid_extract_ref, sheet_name, company, project },
          summary: { total, by_pid, by_type, by_eqpt },
        }
    """
    try:
        import openpyxl  # local import — avoid hard dep at module import time
    except ImportError as exc:
        raise ParseError('openpyxl is not installed on the server') from exc

    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=False)
    except Exception as exc:
        raise ParseError(f'Cannot open workbook: {exc}') from exc

    # Prefer a sheet whose name mentions "instrument", else the active sheet.
    ws = None
    for name in wb.sheetnames:
        if 'instrument' in name.lower() or 'index' in name.lower():
            ws = wb[name]
            break
    if ws is None:
        ws = wb.active

    all_rows = list(ws.iter_rows(values_only=True))
    if not all_rows:
        raise ParseError('Workbook has no rows')

    # 1) Detect primary header row
    header_row_idx = _detect_header_row(all_rows)
    if header_row_idx is None:
        raise ParseError(
            'Could not find a header row containing Tag Number / Instrument Type / Datasheet'
        )

    # 2) Map columns using the composite (primary + secondary) header pair
    primary = all_rows[header_row_idx]
    secondary = all_rows[header_row_idx + 1] if header_row_idx + 1 < len(all_rows) else ()
    columns = _map_columns(primary, secondary)

    # 3) Grab free-text meta from rows above the header
    meta = _extract_meta(all_rows[:header_row_idx])
    meta['sheet_name'] = ws.title

    # 4) Iterate data rows — two rows per instrument
    data_start = header_row_idx + 2  # skip primary + secondary header
    tag_col = columns.get('tag')
    if tag_col is None:
        raise ParseError('Tag Number column could not be located')

    rows: list[dict] = []
    i = data_start
    while i < len(all_rows):
        raw = all_rows[i]
        if _is_blank_row(raw):
            i += 1
            continue
        if _is_stop_marker(raw):
            break

        tag_val = _cell(raw, tag_col)
        if tag_val is None or str(tag_val).strip() == '':
            i += 1
            continue

        # Primary row for this instrument — only fill primary-source keys
        parsed = _parse_row(raw, columns, allowed=PRIMARY_KEYS)
        parsed['_excel_row'] = i + 1  # 1-indexed for the user

        # Secondary row (i+1) — merge secondary-source fields
        if i + 1 < len(all_rows):
            nxt = all_rows[i + 1]
            if not _is_blank_row(nxt) and not _is_stop_marker(nxt):
                nxt_tag = _cell(nxt, tag_col)
                if nxt_tag is None or str(nxt_tag).strip() == '':
                    sub = _parse_row(nxt, columns, allowed=SECONDARY_KEYS)
                    for k, v in sub.items():
                        if v not in (None, ''):
                            parsed[k] = v
                    i += 1  # consume the secondary row

        # Ensure all canonical keys are present so downstream code has stable shape
        for k in list(PRIMARY_KEYS) + list(SECONDARY_KEYS):
            parsed.setdefault(k, '')

        parsed['tag'] = _clean_tag(parsed.get('tag'))
        rows.append(parsed)
        i += 1

    summary = _summarise(rows)
    return {
        'rows': rows,
        'columns': columns,
        'header_row_index': header_row_idx + 1,  # 1-indexed
        'meta': meta,
        'summary': summary,
    }


# ═════════════════════════════════════════════════════════════════════
# Internals
# ═════════════════════════════════════════════════════════════════════

def _detect_header_row(all_rows: list[tuple]) -> Optional[int]:
    scan_upto = min(len(all_rows), MAX_HEADER_SCAN_ROWS)
    for i in range(scan_upto):
        matches = 0
        for cell in all_rows[i]:
            if cell is None:
                continue
            txt = str(cell).strip().lower()
            for token in REQUIRED_HEADER_TOKENS:
                if token in txt:
                    matches += 1
                    break
        if matches >= MIN_HEADER_MATCHES:
            return i
    return None


def _map_columns(primary: tuple, secondary: tuple) -> dict[str, int]:
    """Map canonical keys to column indices, walking both header rows."""
    columns: dict[str, int] = {}
    n = max(len(primary), len(secondary))

    # First pass — direct label matches (primary or secondary alone)
    for idx in range(n):
        p_txt = _lower(primary[idx] if idx < len(primary) else None)
        s_txt = _lower(secondary[idx] if idx < len(secondary) else None)
        combined = f'{p_txt} {s_txt}'.strip()

        for key, aliases in PRIMARY_COL_ALIASES.items():
            if key in columns:
                continue
            if any(alias in p_txt for alias in aliases):
                columns[key] = idx
                break

        for key, aliases in SECONDARY_COL_ALIASES.items():
            if key in columns:
                continue
            if any(alias in s_txt for alias in aliases):
                columns[key] = idx
                break

    # Second pass — range group triples (primary header spans 3 sub-header cells: Min | Max | Unit)
    _map_range_group(primary, secondary, RANGE_GROUP_INSTRUMENT, columns,
                     ('range_min', 'range_max', 'range_unit'))
    _map_range_group(primary, secondary, RANGE_GROUP_CALIBRATION, columns,
                     ('cal_min', 'cal_max', 'cal_unit'))
    return columns


def _map_range_group(primary: tuple, secondary: tuple, group_tokens: tuple[str, ...],
                     columns: dict[str, int], out_keys: tuple[str, str, str]) -> None:
    """Locate a range-group header (e.g. "Instrument Range") and assign the
    following Min / Max / Unit sub-header columns to the given output keys.
    """
    # Find any column whose primary or secondary text matches the group tokens
    start_idx = None
    for idx in range(len(primary)):
        p_txt = _lower(primary[idx])
        s_txt = _lower(secondary[idx] if idx < len(secondary) else None)
        if any(tok in p_txt or tok in s_txt for tok in group_tokens):
            start_idx = idx
            break
    if start_idx is None:
        return

    # Walk from start_idx forward; the very next 3 sub-header cells should be Min / Max / Unit
    found = {}
    for idx in range(start_idx, min(start_idx + 6, max(len(primary), len(secondary)))):
        s_txt = _lower(secondary[idx] if idx < len(secondary) else None)
        p_txt = _lower(primary[idx])
        label = s_txt or p_txt
        if label == 'min' and 'min' not in found:
            found['min'] = idx
        elif label == 'max' and 'max' not in found:
            found['max'] = idx
        elif label in ('unit', 'units') and 'unit' not in found:
            found['unit'] = idx

    key_min, key_max, key_unit = out_keys
    if 'min' in found:
        columns.setdefault(key_min, found['min'])
    if 'max' in found:
        columns.setdefault(key_max, found['max'])
    if 'unit' in found:
        columns.setdefault(key_unit, found['unit'])


def _lower(v: Any) -> str:
    if v is None:
        return ''
    return str(v).strip().lower()


def _is_blank_row(row: tuple) -> bool:
    return not any(c not in (None, '') for c in row)


def _is_stop_marker(row: tuple) -> bool:
    first = _cell(row, 0)
    if first is None:
        return False
    low = str(first).strip().lower()
    return any(low.startswith(m) for m in STOP_MARKERS)


def _cell(row: tuple, idx: Optional[int]) -> Any:
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def _parse_row(row: tuple, columns: dict[str, int], *, allowed: frozenset) -> dict:
    parsed: dict[str, Any] = {}
    for key, col in columns.items():
        if key not in allowed:
            continue
        v = _cell(row, col)
        parsed[key] = _clean(v)
    return parsed


def _clean(v: Any) -> Any:
    if v is None:
        return ''
    if isinstance(v, str):
        s = v.strip()
        s = s.replace('\r', ' ').replace('\n', ' / ')
        s = re.sub(r'\s+', ' ', s)
        return s
    return v


def _clean_tag(v: Any) -> str:
    if v is None:
        return ''
    s = str(v).strip().upper()
    # Collapse internal whitespace so 'LT-8019 TF' and 'PT-8003ATF' normalise consistently
    return re.sub(r'\s+', '', s)


def _extract_meta(top_rows: list[tuple]) -> dict:
    """Pick up company/project/doc-no/date/PID-extract-ref hints from the free-text header block."""
    meta = {k: '' for k in SOURCE_META_KEYS if k != 'sheet_name'}
    for row in top_rows:
        for cell in row:
            if cell is None:
                continue
            s = str(cell).strip()
            if not s:
                continue
            low = s.lower()
            if any(h in low for h in META_HINTS_DOC_NO) and not meta['doc_no']:
                meta['doc_no'] = s
            elif low.startswith('date:') and not meta['date']:
                meta['date'] = s.split(':', 1)[1].strip()
            elif any(h in low for h in META_HINTS_PID) and not meta['pid_extract_ref']:
                meta['pid_extract_ref'] = s
            elif any(h in low for h in META_HINTS_TITLE) and 'index' in low and not meta['title']:
                meta['title'] = s
            elif ('co. ltd' in low or 'company' in low.split()[:2]) and not meta['company']:
                meta['company'] = s
            elif any(h in low for h in META_HINTS_PROJECT) and not meta['project']:
                meta['project'] = s
    return meta


def _summarise(rows: list[dict]) -> dict:
    by_pid: dict[str, int] = {}
    by_type: dict[str, int] = {}
    by_eqpt: dict[str, int] = {}
    tags_ok = 0
    for r in rows:
        if r.get('tag'):
            tags_ok += 1
        pid = str(r.get('pid_no') or '').strip()
        if pid and pid != '-':
            by_pid[pid] = by_pid.get(pid, 0) + 1
        itype = str(r.get('instrument_type') or '').strip()
        if itype:
            by_type[itype] = by_type.get(itype, 0) + 1
        eqpt = str(r.get('eqpt_no') or '').strip()
        if eqpt and eqpt != '-':
            by_eqpt[eqpt] = by_eqpt.get(eqpt, 0) + 1
    return {
        'total': len(rows),
        'with_tag': tags_ok,
        'by_pid': by_pid,
        'by_type': by_type,
        'by_eqpt': by_eqpt,
    }
