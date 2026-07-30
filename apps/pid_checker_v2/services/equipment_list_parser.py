"""Parse an Excel Equipment List into canonical rows.

The Equipment List Excel typically has:
  * a few free-text title rows at the top (company / project / doc no)
  * a MULTI-ROW header (group row for MIN/MAX + field row)
  * one row per equipment item
  * a NOTES / HOLD LIST / ABBREVIATIONS block at the bottom

This parser auto-detects the header block by scanning for the row that
contains the field names we care about (EQPT. TAG No., DESCRIPTION,
DESIGN PRESS., MOC), then treats every subsequent row that has a value
in the TAG column as a data row and stops when we hit a NOTES / HOLD /
ABBREVIATION section.
"""
from __future__ import annotations

import io
import logging
import re
from typing import Any, Optional

logger = logging.getLogger(__name__)

# ─── Soft-coded config ────────────────────────────────────────────────
# Column aliases — case-insensitive substring match against the header row.
COL_ALIASES = {
    'sno':             ['s.no', 's. no', 'sno', 'sr.no'],
    'rev':             ['rev'],
    'tag':             ['eqpt. tag', 'eqpt tag', 'equipment tag', 'tag no'],
    'description':     ['description', 'service'],
    'design_flow':     ['design flowrate', 'design duty', 'volume'],
    'op_pressure':     ['oper. press', 'operating press', 'op. press'],
    'op_temp':         ['oper. temp', 'operating temp', 'op. temp'],
    'design_p_min':    ['design / set press', 'design press', 'set press'],
    'design_p_max':    ['design / set press', 'design press', 'set press'],
    'design_t_min':    ['design temp'],
    'design_t_max':    ['design temp'],
    'moc':             ['moc', 'material of construction'],
    'insulation':      ['insulation'],
    'dim_length':      ['length', 'height', 'length tl'],
    'dim_diameter':    ['diameter', 'width'],
    'motor_rating':    ['motor rating', 'motor'],
    'pid_no':          ['p&id no', 'p&id', 'pid no'],
    'qty':             ['qty', 'quantity'],
    'phase':           ['phase'],
    'remarks':         ['remarks', 'notes'],
    # Deep attribute columns (compared against P&ID Vision extraction)
    'nominal_capacity':  ['nominal capacity', 'capacity', 'nom. capacity', 'design capacity'],
    'length_tt':         ['length t/t', 'length tt', 'length (t/t)', 't/t length', 'tan-to-tan'],
    'diameter_id':       ['diameter id', 'diameter (id)', 'id diameter', 'internal diameter', 'inside diameter'],
    'material_shell':    ['material of shell', 'shell material', 'material shell', 'shell moc'],
    'material_internal': ['material of internal', 'internal material', 'material internal',
                          'internals material', 'internals moc', 'internal moc'],
    'trim':              ['trim', 'trim material', 'valve trim'],
}

# When looking for the header row, at least this many of these
# distinctive field names must appear.
REQUIRED_HEADER_TOKENS = ('eqpt. tag', 'equipment tag', 'description', 'moc', 'p&id')
MIN_HEADER_MATCHES = 2

# How far down to scan for the header before giving up
MAX_HEADER_SCAN_ROWS = 25

# When encountering these labels in column A, stop reading data rows
STOP_MARKERS = ('notes:', 'hold list:', 'abbreviation:', 'abbreviations:', 'legend:')

# Fields that carry useful sniff info about the source document (top matter)
SOURCE_META_KEYS = ('title', 'doc_no', 'date', 'pid_extract_ref', 'sheet_name', 'company', 'project')


class ParseError(Exception):
    pass


def parse_equipment_list(file_bytes: bytes, filename: str = '') -> dict:
    """Parse an xlsx file into a dict with:
        {
          rows: [ { tag, description, moc, dim_length, dim_diameter, pid_no, ... } ],
          columns: { canonical_key: excel_column_index },
          header_row_index: int,
          meta: { title, doc_no, date, pid_extract_ref, sheet_name, company, project },
          summary: { total, by_pid, by_moc },
        }
    """
    try:
        import openpyxl  # local import — avoid hard dep at module import time
    except ImportError as exc:
        raise ParseError('openpyxl is not installed on the server') from exc

    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=False)
    except Exception as exc:
        raise ParseError(f'Cannot open workbook: {exc}') from exc

    # Prefer a sheet whose name mentions "equipment list", else the active sheet.
    ws = None
    for name in wb.sheetnames:
        if 'equipment list' in name.lower() or 'equipment' in name.lower():
            ws = wb[name]
            break
    if ws is None:
        ws = wb.active

    all_rows = list(ws.iter_rows(values_only=True))
    if not all_rows:
        raise ParseError('Workbook has no rows')

    # 1) Detect header row
    header_row_idx = _detect_header_row(all_rows)
    if header_row_idx is None:
        raise ParseError(
            'Could not find a header row containing EQPT. TAG / DESCRIPTION / MOC / P&ID'
        )

    # 2) Map columns  (walk header row + optional sub-header row for MIN/MAX splits)
    header_cells = all_rows[header_row_idx]
    sub_header_cells = all_rows[header_row_idx + 1] if header_row_idx + 1 < len(all_rows) else ()
    columns = _map_columns(header_cells, sub_header_cells)

    # 3) Grab free-text meta from rows above the header
    meta = _extract_meta(all_rows[:header_row_idx])
    meta['sheet_name'] = ws.title

    # 4) Iterate data rows
    data_start = header_row_idx + 1
    if data_start < len(all_rows) and _looks_like_sub_header_row(all_rows[data_start]):
        data_start += 1

    tag_col = columns.get('tag')
    if tag_col is None:
        raise ParseError('EQPT. TAG No. column could not be located')

    rows: list[dict] = []
    for i in range(data_start, len(all_rows)):
        raw = all_rows[i]
        if _is_blank_row(raw):
            continue
        # Stop when we hit a bottom section marker in column A
        first_cell = _cell(raw, 0)
        if first_cell is not None:
            low = str(first_cell).strip().lower()
            if any(low.startswith(m) for m in STOP_MARKERS):
                break
        tag_val = _cell(raw, tag_col)
        if tag_val is None or str(tag_val).strip() == '':
            continue
        parsed = _parse_row(raw, columns)
        parsed['_excel_row'] = i + 1  # 1-indexed for the user
        parsed['tag'] = _clean_tag(parsed.get('tag'))
        rows.append(parsed)

    summary = _summarise(rows)
    return {
        'rows': rows,
        'columns': columns,
        'header_row_index': header_row_idx + 1,   # 1-indexed
        'meta': meta,
        'summary': summary,
    }


# ═════════════════════════════════════════════════════════════════════
# Internals
# ═════════════════════════════════════════════════════════════════════

def _detect_header_row(all_rows: list[tuple]) -> Optional[int]:
    scan_upto = min(len(all_rows), MAX_HEADER_SCAN_ROWS)
    for i in range(scan_upto):
        matches = 0
        for cell in all_rows[i]:
            if cell is None:
                continue
            txt = str(cell).strip().lower()
            for token in REQUIRED_HEADER_TOKENS:
                if token in txt:
                    matches += 1
                    break
        if matches >= MIN_HEADER_MATCHES:
            return i
    return None


def _map_columns(header_row: tuple, sub_header_row: tuple) -> dict[str, int]:
    """Map canonical keys to column indices, using the composite (header + sub-header) label."""
    columns: dict[str, int] = {}
    for idx, cell in enumerate(header_row):
        if cell is None and idx < len(sub_header_row):
            # inherit sub-header only
            cell = sub_header_row[idx] if idx < len(sub_header_row) else None
        if cell is None:
            continue
        header_txt = str(cell).strip().lower()
        sub_txt = ''
        if idx < len(sub_header_row) and sub_header_row[idx] is not None:
            sub_txt = str(sub_header_row[idx]).strip().lower()
        combined = f'{header_txt} {sub_txt}'.strip()
        for key, aliases in COL_ALIASES.items():
            if key in columns:
                continue  # already mapped, keep first hit
            for alias in aliases:
                if alias in combined:
                    # For MIN/MAX splits, prefer sub-header to disambiguate
                    if key.endswith('_min') and 'min' not in sub_txt and 'min' not in header_txt:
                        continue
                    if key.endswith('_max') and 'max' not in sub_txt and 'max' not in header_txt:
                        continue
                    if key in ('dim_length', 'dim_diameter'):
                        # both share the DIMENSIONS header — use sub-header to split
                        if key == 'dim_length' and not any(t in sub_txt for t in ('length', 'height')):
                            continue
                        if key == 'dim_diameter' and not any(t in sub_txt for t in ('diameter', 'width')):
                            continue
                    columns[key] = idx
                    break
    return columns


def _looks_like_sub_header_row(row: tuple) -> bool:
    """Heuristic: sub-header rows contain MIN / MAX / (mm) tokens and few full labels."""
    if not row:
        return False
    tokens = 0
    for cell in row:
        if cell is None:
            continue
        s = str(cell).strip().lower()
        if s in ('min', 'max', '(mm)', '(inch)', '(psig)', '(°f)', '(°c)'):
            tokens += 1
        elif re.match(r'^(length|diameter|width|height)', s):
            tokens += 1
    return tokens >= 2


def _is_blank_row(row: tuple) -> bool:
    return not any(c not in (None, '') for c in row)


def _cell(row: tuple, idx: Optional[int]) -> Any:
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def _parse_row(row: tuple, columns: dict[str, int]) -> dict:
    parsed: dict[str, Any] = {}
    for key, col in columns.items():
        v = _cell(row, col)
        parsed[key] = _clean(v)
    return parsed


def _clean(v: Any) -> Any:
    if v is None:
        return ''
    if isinstance(v, str):
        s = v.strip()
        s = s.replace('\r', ' ').replace('\n', ' / ')
        s = re.sub(r'\s+', ' ', s)
        return s
    return v


def _clean_tag(v: Any) -> str:
    if v is None:
        return ''
    s = str(v).strip().upper()
    return re.sub(r'\s+', '', s)


def _extract_meta(top_rows: list[tuple]) -> dict:
    """Pick up company/project/doc-no/date/PID-extract-ref hints from the free-text header block."""
    meta = {k: '' for k in SOURCE_META_KEYS if k != 'sheet_name'}
    for row in top_rows:
        for cell in row:
            if cell is None:
                continue
            s = str(cell).strip()
            if not s:
                continue
            low = s.lower()
            if 'doc no' in low or 'company doc' in low:
                meta['doc_no'] = s
            elif low.startswith('date:'):
                meta['date'] = s.split(':', 1)[1].strip()
            elif 'equipment list' in low and not meta['title']:
                meta['title'] = s
            elif ('p&id' in low or 'pertaining' in low) and 'extract' not in low and not meta['pid_extract_ref']:
                meta['pid_extract_ref'] = s
            elif 'co. ltd' in low or 'company' in low.split()[:2]:
                if not meta['company']:
                    meta['company'] = s
            elif ('project' in low or 'engineering for' in low) and not meta['project']:
                meta['project'] = s
    return meta


def _summarise(rows: list[dict]) -> dict:
    by_pid: dict[str, int] = {}
    by_moc: dict[str, int] = {}
    tags_ok = 0
    for r in rows:
        if r.get('tag'):
            tags_ok += 1
        pid = str(r.get('pid_no') or '').strip()
        if pid:
            by_pid[pid] = by_pid.get(pid, 0) + 1
        moc = str(r.get('moc') or '').strip()
        if moc:
            by_moc[moc] = by_moc.get(moc, 0) + 1
    return {
        'total': len(rows),
        'with_tag': tags_ok,
        'by_pid': by_pid,
        'by_moc': by_moc,
    }
