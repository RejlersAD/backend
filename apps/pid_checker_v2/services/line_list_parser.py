"""Parse an Excel Line List into canonical rows.

The Line List Excel typically has:
  * a few free-text title rows at the top
  * a MULTI-ROW header (group row + field row + optional units row)
  * one row per pipe line

This parser auto-detects the header block by scanning for the row that
contains the field names we care about (SIZE, SERVICE CODE, LINE NUMBER,
SPECIFICATION), then treats every subsequent row that has a value in the
SIZE column as a data row.

The composite tag  SIZE"-SERVICECODE-SPEC-SERIAL  is assembled here so
it can be diff'd directly against tags extracted from the P&ID.
"""
from __future__ import annotations

import datetime
import io
import logging
import re
from typing import Any, Optional

logger = logging.getLogger(__name__)

# ─── Soft-coded config ────────────────────────────────────────────────
# Column aliases — case-insensitive substring match against the header row.
COL_ALIASES = {
    'size':         ['size'],
    'service_code': ['service code', 'service'],
    'serial':       ['line number', 'line no', 'serial'],
    'spec':         ['specification', 'spec'],
    'deviation':    ['deviation'],
    'insulation_type':     ['insulation type', 'insulation - type'],
    'insulation_material': ['insulation - material', 'insulation material'],
    'insulation_thickness':['insulation - thickness', 'thickness'],
    'from':         ['from'],
    'to':           ['to'],
    'pid_no':       ['p&id no', 'p&id', 'pid no', 'pid'],
    'fluid_service':['fluid service'],
    'fluid_phase':  ['fluid phase', 'phase'],
    'density_mw':   ['density', 'mw'],
    'op_pressure':  ['operating pressure'],
    'op_temp':      ['operating temperature'],
    'design_p_max': ['design pressure - max', 'design pressure max'],
    'design_p_min': ['design pressure - min', 'design pressure min'],
    'design_t_max': ['design temperature - max', 'design temperature max'],
    'design_t_min': ['design temperature - min', 'design temperature min'],
    'steam_out':    ['steam out'],
    'test_pressure':['test pressure'],
    'test_medium':  ['test medium'],
    'pwht':         ['pwht'],
    'nace':         ['nace'],
    'paint_spec':   ['paint specification', 'paint spec'],
    'nde_rt':       ['nde (rt)', 'nde rt'],
    'nde_mt_pt':    ['nde (mt', 'nde mt'],
    'rho_v2':       ['rho v2', 'rho v²'],
    'notes':        ['notes'],
    'rev':          ['rev', 'revision'],
}

# When looking for the header row, at least this many of these
# distinctive field names must appear.
REQUIRED_HEADER_TOKENS = ('size', 'service code', 'specification')
MIN_HEADER_MATCHES = 2

# How far down to scan for the header before giving up
MAX_HEADER_SCAN_ROWS = 20

# The composite tag components — all four are required to build the tag
TAG_COMPONENTS = ('size', 'service_code', 'spec', 'serial')

# Fields that carry useful sniff info about the source P&ID (top matter)
SOURCE_META_KEYS = ('title', 'doc_no', 'date', 'pid_extract_ref', 'sheet_name')


class ParseError(Exception):
    pass


def parse_line_list(file_bytes: bytes, filename: str = '') -> dict:
    """Parse an xlsx file into a dict with:
        {
          rows: [ { tag, size, service_code, spec, serial, ... } ],
          columns: { canonical_key: excel_column_index },
          header_row_index: int,
          meta: { title, doc_no, date, pid_extract_ref, sheet_name },
          summary: { total, distinct_services, ... },
        }
    """
    try:
        import openpyxl  # local import — avoid hard dep at module import time
    except ImportError as exc:
        raise ParseError('openpyxl is not installed on the server') from exc

    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=False)
    except Exception as exc:
        raise ParseError(f'Cannot open workbook: {exc}') from exc

    # We support single-sheet Line Lists. If more sheets exist, prefer the
    # one whose name mentions "line list", else the active sheet.
    ws = None
    for name in wb.sheetnames:
        if 'line list' in name.lower():
            ws = wb[name]
            break
    if ws is None:
        ws = wb.active

    all_rows = list(ws.iter_rows(values_only=True))
    if not all_rows:
        raise ParseError('Workbook has no rows')

    # 1) Detect header row
    header_row_idx = _detect_header_row(all_rows)
    if header_row_idx is None:
        raise ParseError(
            'Could not find a header row containing SIZE / SERVICE CODE / SPECIFICATION'
        )

    # 2) Map columns
    header_cells = all_rows[header_row_idx]
    columns = _map_columns(header_cells)

    # 3) Grab free-text meta from rows above the header
    meta = _extract_meta(all_rows[:header_row_idx])
    meta['sheet_name'] = ws.title

    # 4) Iterate data rows (below header + units row if any)
    #    Skip empty rows and rows that have no SIZE value (they're group/label rows).
    data_start = header_row_idx + 1
    # If the row immediately below header looks like a units row, skip it.
    if data_start < len(all_rows) and _looks_like_units_row(all_rows[data_start]):
        data_start += 1

    size_col = columns.get('size')
    if size_col is None:
        raise ParseError('SIZE column could not be located')

    rows: list[dict] = []
    for i in range(data_start, len(all_rows)):
        raw = all_rows[i]
        if _is_blank_row(raw):
            continue
        size_val = _cell(raw, size_col)
        if size_val is None or str(size_val).strip() == '':
            # allow location marker rows like "MUBARRAZ ISLAND" — skip
            continue
        parsed = _parse_row(raw, columns)
        parsed['_excel_row'] = i + 1  # 1-indexed for the user
        parsed['tag'] = _build_tag(parsed)
        rows.append(parsed)

    summary = _summarise(rows)
    return {
        'rows': rows,
        'columns': columns,
        'header_row_index': header_row_idx + 1,   # 1-indexed
        'meta': meta,
        'summary': summary,
    }


# ═════════════════════════════════════════════════════════════════════
# Internals
# ═════════════════════════════════════════════════════════════════════

def _detect_header_row(all_rows: list[tuple]) -> Optional[int]:
    scan_upto = min(len(all_rows), MAX_HEADER_SCAN_ROWS)
    for i in range(scan_upto):
        matches = 0
        for cell in all_rows[i]:
            if cell is None:
                continue
            txt = str(cell).strip().lower()
            for token in REQUIRED_HEADER_TOKENS:
                if token in txt:
                    matches += 1
                    break
        if matches >= MIN_HEADER_MATCHES:
            return i
    return None


def _map_columns(header_row: tuple) -> dict[str, int]:
    columns: dict[str, int] = {}
    for idx, cell in enumerate(header_row):
        if cell is None:
            continue
        txt = str(cell).strip().lower()
        for key, aliases in COL_ALIASES.items():
            if key in columns:
                continue  # already mapped, keep first hit
            for alias in aliases:
                if alias in txt:
                    columns[key] = idx
                    break
    return columns


def _looks_like_units_row(row: tuple) -> bool:
    """Heuristic: row containing (inch), (mm), Psig, °F etc."""
    if not row:
        return False
    unit_hits = 0
    for cell in row:
        if cell is None:
            continue
        s = str(cell).strip().lower()
        if s in ('(inch)', '(mm)', 'psig', '(°f)', '(kg/m3)', '(kg/m³)', '(°c)'):
            unit_hits += 1
        elif re.fullmatch(r'\([^)]{1,6}\)', s):
            unit_hits += 1
    return unit_hits >= 2


def _is_blank_row(row: tuple) -> bool:
    return not any(c not in (None, '') for c in row)


def _cell(row: tuple, idx: Optional[int]) -> Any:
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def _parse_row(row: tuple, columns: dict[str, int]) -> dict:
    parsed: dict[str, Any] = {}
    for key, col in columns.items():
        v = _cell(row, col)
        parsed[key] = _clean(v)
    return parsed


def _clean(v: Any) -> Any:
    if v is None:
        return ''
    if isinstance(v, str):
        s = v.strip()
        # Excel newline noise in multi-P&ID cells becomes " / "
        s = s.replace('\r', ' ').replace('\n', ' / ')
        s = re.sub(r'\s+', ' ', s)
        return s
    return v


# A fractional pipe-size cell like "3/4" or "1/2" is routinely mangled by
# Excel unless the SIZE column is explicitly formatted as text — Excel's
# autocorrect reinterprets "M/D"-shaped text as a date at entry time. What
# openpyxl then hands back for that cell depends on how/when the file was
# saved and by what tool, so this has to be handled defensively rather
# than assuming one fixed shape:
#   - a live datetime/date object (the common case — Excel itself saved it
#     as a real date value)
#   - a "M/D/YYYY" or "M/D/YY" string (some export pipelines/older Excel
#     re-serialize the date as displayed text instead of a date object)
#   - a "D-Mon" / "Mon-D" / "D-Mon-YY" string (a different Excel display
#     format for the same underlying date)
#   - unicode vulgar-fraction characters (¾ ½ ¼ ⅛ ⅜ ⅝ ⅞) typed directly
#     instead of "3/4" etc. — not a date-conversion bug, but the same
#     "doesn't match the P&ID's plain-ASCII fraction" symptom
# All of these are recovered back to the plain "M/D" form a real pipe size
# uses. Only NPS-legal fraction values ever hit this path (denominators
# 2/4/8/16, numerator < denominator), so the month/day round-trip is safe.
_SIZE_MDY_RE = re.compile(r'^(\d{1,2})/(\d{1,2})/\d{2,4}$')
_SIZE_ISO_RE = re.compile(r'^\d{4}-(\d{1,2})-(\d{1,2})(?:[ T]\d{1,2}:\d{2}(?::\d{2})?)?$')
_SIZE_D_MON_RE = re.compile(r'^(\d{1,2})[-\s]([A-Za-z]{3,9})(?:[-\s]\d{2,4})?$')
_SIZE_MON_D_RE = re.compile(r'^([A-Za-z]{3,9})[-\s](\d{1,2})(?:[-\s]\d{2,4})?$')
_MONTH_NUM = {
    'jan': 1, 'feb': 2, 'mar': 3, 'apr': 4, 'may': 5, 'jun': 6,
    'jul': 7, 'aug': 8, 'sep': 9, 'oct': 10, 'nov': 11, 'dec': 12,
}
_UNICODE_FRACTIONS = {
    '½': '1/2', '¼': '1/4', '¾': '3/4', '⅛': '1/8', '⅜': '3/8',
    '⅝': '5/8', '⅞': '7/8', '⅓': '1/3', '⅔': '2/3',
}


def _normalize_size_value(v: Any) -> str:
    """Recover a plain "M/D"-shaped pipe size from any of the Excel-corrupted
    forms described above. Returns str(v).strip() unchanged when none of
    the corruption patterns match — i.e. this is purely additive and never
    changes an already-correct size like "6", "20", or "1-1/2"."""
    if isinstance(v, (datetime.datetime, datetime.date)):
        return f'{v.month}/{v.day}'
    s = str(v).strip()
    for uni, ascii_frac in _UNICODE_FRACTIONS.items():
        if uni in s:
            return s.replace(uni, ascii_frac)
    m = _SIZE_MDY_RE.match(s)
    if m:
        return f'{int(m.group(1))}/{int(m.group(2))}'
    m = _SIZE_ISO_RE.match(s)
    if m:
        return f'{int(m.group(1))}/{int(m.group(2))}'
    m = _SIZE_D_MON_RE.match(s)
    if m and m.group(2).lower()[:3] in _MONTH_NUM:
        return f'{_MONTH_NUM[m.group(2).lower()[:3]]}/{int(m.group(1))}'
    m = _SIZE_MON_D_RE.match(s)
    if m and m.group(1).lower()[:3] in _MONTH_NUM:
        return f'{_MONTH_NUM[m.group(1).lower()[:3]]}/{int(m.group(2))}'
    return s


def _build_tag(row: dict) -> str:
    """Compose  SIZE"-SERVICECODE-SPEC-SERIAL  from parsed row cells."""
    parts = []
    for key in TAG_COMPONENTS:
        v = row.get(key)
        if v is None or str(v).strip() == '':
            return ''  # cannot compose — leave blank so cross-check flags it
        s = str(v).strip()
        if key == 'size':
            s = _normalize_size_value(v)
            # ensure trailing "  (openpyxl gives us bare number or "3/4")
            s = s.rstrip('"')
            s = f'{s}"'
        elif key == 'serial':
            # serial may come back as float like 7263.0
            if isinstance(v, float) and v.is_integer():
                s = str(int(v))
        parts.append(s)
    return '-'.join(parts)


def _extract_meta(top_rows: list[tuple]) -> dict:
    """Pick up title/doc no/date/PID-extract-ref hints from the free-text header block."""
    meta = {k: '' for k in SOURCE_META_KEYS if k != 'sheet_name'}
    for row in top_rows:
        for cell in row:
            if cell is None:
                continue
            s = str(cell).strip()
            if not s:
                continue
            low = s.lower()
            if 'doc no' in low or 'company doc' in low:
                meta['doc_no'] = s
            elif low.startswith('date:'):
                meta['date'] = s.split(':', 1)[1].strip()
            elif 'line list' in low and not meta['title']:
                meta['title'] = s
            elif 'p&id' in low and 'extract' in low:
                meta['pid_extract_ref'] = s
    return meta


def _summarise(rows: list[dict]) -> dict:
    services: dict[str, int] = {}
    specs: dict[str, int] = {}
    sizes: dict[str, int] = {}
    tags_ok = 0
    for r in rows:
        if r.get('tag'):
            tags_ok += 1
        svc = str(r.get('service_code') or '').strip().upper()
        if svc:
            services[svc] = services.get(svc, 0) + 1
        spec = str(r.get('spec') or '').strip().upper()
        if spec:
            specs[spec] = specs.get(spec, 0) + 1
        sz = str(r.get('size') or '').strip().rstrip('"')
        if sz:
            sizes[sz] = sizes.get(sz, 0) + 1
    return {
        'total': len(rows),
        'with_tag': tags_ok,
        'service_codes': services,
        'specifications': specs,
        'sizes': sizes,
    }
