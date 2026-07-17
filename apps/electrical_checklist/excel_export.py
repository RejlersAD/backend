"""
Excel Export Utility for Electrical Checklist
Generates formatted Excel with extracted data and signature images

SOFT-CODED: Detects which template shape a job's data was stored in and
renders the matching layout, instead of always assuming the old 5-column
format:
  - CURRENT pipeline (handwriting extraction + Save Changes edits) stores
    `job.extracted_data['checklist_data']` in the 6-column v2 template shape
    (TEMPLATE_V2_SECTIONS / TEMPLATE_V2_COLUMNS) — this is what the frontend
    checklist page actually displays and edits.
  - LEGACY jobs (created via the old `/extract/` stub endpoint, before the
    v2 pipeline existed) still have the old `extracted_data` + `signatures`
    shape — rendered with the original 5-column layout so old exports keep
    working without a data migration.
"""
import os
import base64
import io
import tempfile
from typing import Dict, Any
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

from .config import CHECKLIST_TEMPLATE
from .template_v2_config import TEMPLATE_V2_COLUMNS, TEMPLATE_V2_SECTIONS, TEMPLATE_V2_METADATA

# Soft-coded column widths for the v2 export (keyed by column key, not index)
V2_COLUMN_WIDTHS = {
    'field_name':    38,
    'site_value':    26,
    'remarks':       22,
    'need_list':     22,
    'query':         22,
    'company_reply': 22,
}


def generate_excel_export(job) -> str:
    """
    Generate Excel file from extraction job results.

    Soft-coded to detect which template shape the job's data is stored in
    (current 6-column v2 template vs. legacy 5-column format) and render the
    matching layout.

    Args:
        job: ChecklistExtractionJob instance

    Returns:
        Path to generated Excel file
    """
    extracted = job.extracted_data or {}

    if isinstance(extracted.get('checklist_data'), dict):
        return _generate_v2_excel(job, extracted)
    return _generate_legacy_excel(job, extracted)


# ─────────────────────────────────────────────────────────────────────────────
# CURRENT template renderer — 6-column v2 template
# (field_name | site_value | remarks | need_list | query | company_reply)
# ─────────────────────────────────────────────────────────────────────────────

def _generate_v2_excel(job, extracted: Dict[str, Any]) -> str:
    checklist_data = extracted.get('checklist_data') or {}
    summary = extracted.get('summary') or {}

    wb = Workbook()
    ws = wb.active
    ws.title = "Checklist Data"

    num_cols = len(TEMPLATE_V2_COLUMNS)
    last_col_letter = get_column_letter(num_cols)

    # Styling
    title_fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
    section_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    header_font = Font(bold=True, color="000000", size=11)
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # Title
    ws.merge_cells(f'A1:{last_col_letter}1')
    title_cell = ws['A1']
    title_cell.value = TEMPLATE_V2_METADATA.get('title') or CHECKLIST_TEMPLATE.get('name', 'Checklist')
    title_cell.font = Font(bold=True, size=16, color="FFFFFF")
    title_cell.fill = title_fill
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    # Metadata
    row = 3
    meta_rows = [
        ("Extraction Job ID:", job.id),
        ("Extraction Date:", job.created_at.strftime("%Y-%m-%d %H:%M:%S")),
        ("Engineer(s):", extracted.get('engineer_name') or '-'),
        ("Fields Extracted:", f"{summary.get('fields_extracted', job.fields_extracted)} / {summary.get('total_fields', '')}"),
        ("Sections Completed:", f"{summary.get('sections_completed', '')} / {summary.get('total_sections', len(TEMPLATE_V2_SECTIONS))}"),
        ("Confidence Score:", f"{summary.get('confidence_score', job.confidence_score)}%"),
    ]
    for label, value in meta_rows:
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = Font(bold=True)
        ws.merge_cells(f'B{row}:{last_col_letter}{row}')
        ws[f'B{row}'] = value
        row += 1

    row += 2  # space before data

    # Column headers (soft-coded from TEMPLATE_V2_COLUMNS — never hardcoded)
    header_row = row
    for col_idx, col_def in enumerate(TEMPLATE_V2_COLUMNS, 1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.value = col_def['label']
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
    row += 1

    # Data rows, grouped by section — mirrors the frontend detailed view
    for section in TEMPLATE_V2_SECTIONS:
        ws.merge_cells(f'A{row}:{last_col_letter}{row}')
        sec_cell = ws.cell(row=row, column=1)
        sec_cell.value = f"{section['number']}. {section['title']}"
        sec_cell.font = Font(bold=True, size=11)
        sec_cell.fill = section_fill
        sec_cell.alignment = Alignment(horizontal='left', vertical='center')
        row += 1

        for field in section['fields']:
            field_row = checklist_data.get(field['id']) or {}
            values = [
                field_row.get('field_name', field['name']),
                field_row.get('site_value', ''),
                field_row.get('remarks', ''),
                field_row.get('need_list', ''),
                field_row.get('query', ''),
                field_row.get('company_reply', ''),
            ]
            for col_idx, value in enumerate(values, 1):
                cell = ws.cell(row=row, column=col_idx)
                cell.value = value
                cell.border = border
                cell.alignment = Alignment(vertical='center', wrap_text=True)
            row += 1

    # Column widths (soft-coded lookup by column key)
    for col_idx, col_def in enumerate(TEMPLATE_V2_COLUMNS, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = V2_COLUMN_WIDTHS.get(col_def['key'], 20)

    ws.freeze_panes = f'A{header_row + 1}'

    temp_dir = tempfile.gettempdir()
    excel_path = os.path.join(temp_dir, f'checklist_{job.id}.xlsx')
    wb.save(excel_path)
    return excel_path


# ─────────────────────────────────────────────────────────────────────────────
# LEGACY template renderer — 5-column format
# (Section | Field | Value | Confidence | Page) — kept only so checklist jobs
# created before the v2 pipeline existed can still be exported. Do not use
# this shape for new jobs.
# ─────────────────────────────────────────────────────────────────────────────

def _generate_legacy_excel(job, extracted_root: Dict[str, Any]) -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = "Checklist Data"
    
    # Styling
    header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")  # Yellow/Orange
    highlighted_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow
    header_font = Font(bold=True, color="000000", size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title
    ws.merge_cells('A1:D1')
    title_cell = ws['A1']
    title_cell.value = CHECKLIST_TEMPLATE['name']
    title_cell.font = Font(bold=True, size=16, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30
    
    # Metadata
    row = 3
    ws[f'A{row}'] = "Extraction Job ID:"
    ws[f'B{row}'] = job.id
    ws[f'A{row}'].font = Font(bold=True)
    
    row += 1
    ws[f'A{row}'] = "Extraction Date:"
    ws[f'B{row}'] = job.created_at.strftime("%Y-%m-%d %H:%M:%S")
    ws[f'A{row}'].font = Font(bold=True)
    
    row += 1
    ws[f'A{row}'] = "Fields Extracted:"
    ws[f'B{row}'] = job.fields_extracted
    ws[f'A{row}'].font = Font(bold=True)
    
    row += 1
    ws[f'A{row}'] = "Signatures Found:"
    ws[f'B{row}'] = job.signatures_found
    ws[f'A{row}'].font = Font(bold=True)
    
    row += 1
    ws[f'A{row}'] = "Confidence Score:"
    ws[f'B{row}'] = f"{job.confidence_score}%"
    ws[f'A{row}'].font = Font(bold=True)
    
    row += 3  # Space before data
    
    # Headers
    headers = ['Section', 'Field', 'Value', 'Confidence', 'Page']
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    row += 1
    
    # Data rows
    extracted_data = extracted_root.get('extracted_data', {})
    
    for section in CHECKLIST_TEMPLATE['sections']:
        if section.get('is_signature_section'):
            continue  # Handle signatures separately
        
        section_name = section['name']
        
        for field in section['fields']:
            field_key = field['key']
            field_data = extracted_data.get(field_key, {})
            
            # Section
            cell = ws.cell(row=row, column=1)
            cell.value = section_name
            cell.border = border
            
            # Field label
            cell = ws.cell(row=row, column=2)
            cell.value = field['label']
            cell.border = border
            
            # Value
            cell = ws.cell(row=row, column=3)
            cell.value = field_data.get('value', '')
            cell.border = border
            if field.get('highlighted'):
                cell.fill = highlighted_fill
            
            # Confidence
            cell = ws.cell(row=row, column=4)
            confidence = field_data.get('confidence', 0)
            cell.value = f"{confidence}%" if confidence > 0 else ''
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
            
            # Page number
            cell = ws.cell(row=row, column=5)
            cell.value = field_data.get('page_number', '')
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
            
            row += 1
    
    # Signatures section
    if job.signatures_found > 0:
        row += 2
        ws.merge_cells(f'A{row}:E{row}')
        sig_header = ws[f'A{row}']
        sig_header.value = "Signatures"
        sig_header.font = Font(bold=True, size=14, color="FFFFFF")
        sig_header.fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
        sig_header.alignment = Alignment(horizontal='center', vertical='center')
        
        row += 1
        
        signatures = extracted_root.get('signatures', [])
        for sig_idx, sig in enumerate(signatures, 1):
            ws[f'A{row}'] = f"Signature {sig_idx}"
            ws[f'B{row}'] = sig.get('associated_label', 'Unknown')
            ws[f'C{row}'] = f"Page {sig.get('page_number', 'N/A')}"
            ws[f'D{row}'] = f"Confidence: {sig.get('confidence', 0)}%"
            
            # Embed signature image if available
            if 'image_base64' in sig and sig['image_base64']:
                try:
                    img_data = base64.b64decode(sig['image_base64'])
                    img = XLImage(io.BytesIO(img_data))
                    img.width = 200
                    img.height = 100
                    ws.add_image(img, f'E{row}')
                    ws.row_dimensions[row].height = 75
                except Exception as e:
                    ws[f'E{row}'] = f"[Image error: {e}]"
            
            row += 1
    
    # Auto-adjust column widths
    for col_idx in range(1, 6):
        column_letter = get_column_letter(col_idx)
        max_length = 0
        for cell in ws[column_letter]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to temp file
    temp_dir = tempfile.gettempdir()
    excel_path = os.path.join(temp_dir, f'checklist_{job.id}.xlsx')
    wb.save(excel_path)
    
    return excel_path
