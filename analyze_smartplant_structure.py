"""
Revised analysis - Understanding SPEC vs CAT structure
SPEC = Specification rules/filters (PipingCommodityFilter, PipingCommodityMatlControlData, etc.)
CAT = Component catalog (GateValve, BallValve, Tee, Elbow, etc. - actual parts)
"""
import openpyxl
from pathlib import Path

def analyze_spec_filter_sheet(spec_file, prefix):
    """Analyze PipingCommodityFilter in SPEC file"""
    print(f"\n{'='*80}")
    print(f"1. PIPING COMMODITY FILTER (SPEC) - {prefix}")
    print(f"{'='*80}")
    
    wb = openpyxl.load_workbook(spec_file, data_only=True)
    
    if 'PipingCommodityFilter' not in wb.sheetnames:
        print("  PipingCommodityFilter sheet NOT FOUND")
        wb.close()
        return
    
    ws = wb['PipingCommodityFilter']
    
    # Get headers
    headers = [cell.value for cell in ws[1] if cell.value]
    print(f"\n  Columns ({len(headers)}): {', '.join(headers)}")
    
    # Get data rows
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(cell for cell in row):
            rows.append(row)
    
    print(f"\n  Total rows: {len(rows)}")
    
    # Sample data
    print(f"\n  Sample data (first 10 rows):")
    commodity_idx = headers.index('CommodityCode') if 'CommodityCode' in headers else 0
    shortcode_idx = headers.index('ShortCode') if 'ShortCode' in headers else None
    fluid_idx = headers.index('FluidCode') if 'FluidCode' in headers else None
    size_from_idx = headers.index('FirstSizeFrom') if 'FirstSizeFrom' in headers else None
    size_to_idx = headers.index('FirstSizeTo') if 'FirstSizeTo' in headers else None
    
    for idx, row in enumerate(rows[:10], start=1):
        commodity = row[commodity_idx] if commodity_idx < len(row) else 'N/A'
        shortcode = row[shortcode_idx] if shortcode_idx is not None and shortcode_idx < len(row) else 'N/A'
        fluid = row[fluid_idx] if fluid_idx is not None and fluid_idx < len(row) else 'N/A'
        size_from = row[size_from_idx] if size_from_idx is not None and size_from_idx < len(row) else 'N/A'
        size_to = row[size_to_idx] if size_to_idx is not None and size_to_idx < len(row) else 'N/A'
        print(f"    {idx:2d}. {commodity:40s} | {shortcode:30s} | Fluid: {fluid:10s} | Size: {size_from}-{size_to}")
    
    wb.close()


def analyze_matl_control_sheet(spec_file, prefix):
    """Analyze PipingCommodityMatlControlData in SPEC file"""
    print(f"\n{'='*80}")
    print(f"2. PIPING COMMODITY MATL CONTROL DATA (SPEC) - {prefix}")
    print(f"{'='*80}")
    
    wb = openpyxl.load_workbook(spec_file, data_only=True)
    
    sheet_name = 'PipingCommodityMatlControlData'
    
    if sheet_name not in wb.sheetnames:
        print(f"  {sheet_name} sheet NOT FOUND")
        wb.close()
        return
    
    ws = wb[sheet_name]
    
    # Get headers
    headers = [cell.value for cell in ws[1] if cell.value]
    print(f"\n  Columns ({len(headers)}): {', '.join(headers)}")
    
    # Get data rows
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(cell for cell in row):
            rows.append(row)
    
    print(f"\n  Total rows: {len(rows)}")
    
    # Check descriptions
    desc_idx = headers.index('Description') if 'Description' in headers else None
    commodity_idx = headers.index('CommodityCode') if 'CommodityCode' in headers else 0
    shortcode_idx = headers.index('ShortCode') if 'ShortCode' in headers else None
    
    print(f"\n  Current descriptions (first 15 rows):")
    print(f"  {'Row':<5} {'CommodityCode':<40} {'ShortCode':<30} {'Description':<50}")
    print(f"  {'-'*5} {'-'*40} {'-'*30} {'-'*50}")
    
    for idx, row in enumerate(rows[:15], start=1):
        commodity = str(row[commodity_idx] if commodity_idx < len(row) else 'N/A')[:40]
        shortcode = str(row[shortcode_idx] if shortcode_idx is not None and shortcode_idx < len(row) else 'N/A')[:30]
        desc = str(row[desc_idx] if desc_idx is not None and desc_idx < len(row) else 'N/A')[:50]
        print(f"  {idx:<5} {commodity:<40} {shortcode:<30} {desc:<50}")
    
    wb.close()


def analyze_cat_component_sheets(cat_file, prefix):
    """Analyze component sheets in CAT file (GateValve, BallValve, etc.)"""
    print(f"\n{'='*80}")
    print(f"3. COMPONENT/PART SHEETS (CAT) - {prefix}")
    print(f"{'='*80}")
    
    wb = openpyxl.load_workbook(cat_file, data_only=True)
    
    # Focus on valve and fitting sheets
    component_sheets = [name for name in wb.sheetnames if any(x in name for x in 
        ['Valve', 'Elbow', 'Tee', 'Reducer', 'Flange', 'Cap', 'Nipple', 'Plug', 'Coupling', 'Sockolet', 'Weldolet'])]
    
    print(f"\n  Found {len(component_sheets)} component sheets:")
    for sheet_name in component_sheets[:10]:
        print(f"    - {sheet_name}")
    if len(component_sheets) > 10:
        print(f"    ... and {len(component_sheets) - 10} more")
    
    # Analyze first few component sheets in detail
    print(f"\n  Detailed analysis of component sheets:")
    
    for sheet_name in component_sheets[:3]:  # Analyze first 3 sheets
        ws = wb[sheet_name]
        
        # Get headers
        headers = [cell.value for cell in ws[1] if cell.value]
        
        # Count rows
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if any(cell for cell in row):
                rows.append(row)
        
        print(f"\n    Sheet: {sheet_name}")
        print(f"      Columns: {', '.join(headers[:10])}{'...' if len(headers) > 10 else ''}")
        print(f"      Data rows: {len(rows)}")
        
        # Sample data
        if rows and headers:
            print(f"      Sample row 1:")
            for col_name, cell_value in zip(headers[:8], rows[0][:8]):
                if cell_value:
                    print(f"        {col_name}: {str(cell_value)[:50]}")
    
    wb.close()


def compare_spec_vs_cat(spec_file, cat_file, prefix):
    """Compare SPEC and CAT to understand the relationship"""
    print(f"\n{'='*80}")
    print(f"4. SPEC vs CAT RELATIONSHIP - {prefix}")
    print(f"{'='*80}")
    
    spec_wb = openpyxl.load_workbook(spec_file, data_only=True)
    cat_wb = openpyxl.load_workbook(cat_file, data_only=True)
    
    # Get commodity codes from SPEC PipingCommodityFilter
    if 'PipingCommodityFilter' in spec_wb.sheetnames:
        pcf_ws = spec_wb['PipingCommodityFilter']
        headers = [cell.value for cell in pcf_ws[1]]
        commodity_idx = headers.index('CommodityCode') if 'CommodityCode' in headers else None
        
        if commodity_idx is not None:
            spec_commodities = set()
            for row in pcf_ws.iter_rows(min_row=2, values_only=True):
                if row and commodity_idx < len(row) and row[commodity_idx]:
                    spec_commodities.add(str(row[commodity_idx]))
            
            print(f"\n  SPEC has {len(spec_commodities)} unique commodity codes")
            print(f"  Sample codes: {', '.join(list(spec_commodities)[:5])}")
    
    # Check if CAT component sheets reference these codes
    print(f"\n  CAT component sheets: {', '.join([s for s in cat_wb.sheetnames if 'Valve' in s or 'Tee' in s or 'Elbow' in s][:5])}")
    
    spec_wb.close()
    cat_wb.close()


def main():
    """Main analysis"""
    files = [
        ('AA1B-A3_SPEC.xlsx', 'AA1B-A3_CAT.xlsx', 'AA1B-A3'),
        ('AA2A-A3_SPEC.xlsx', 'AA2A-A3_CAT.xlsx', 'AA2A-A3'),
        ('AC2A-A3_SPEC.xlsx', 'AC2A-A3_CAT.xlsx', 'AC2A-A3'),
    ]
    
    print("="*80)
    print("SMARTPLANT 3D SPEC & CAT STRUCTURE ANALYSIS")
    print("="*80)
    print("\nKEY UNDERSTANDING:")
    print("  - SPEC file = Specification rules (filters, material control, etc.)")
    print("  - CAT file = Component catalog (actual parts like valves, fittings, etc.)")
    print("\nISSUES TO FIX:")
    print("  1. PipingCommodityFilter sheet - needs more entries (30%)")
    print("  2. PipingCommodityMatlControlData - descriptions should match CAT components (20%)")
    print("  3. CAT component sheets - should be populated from paper spec data (50%)")
    
    for spec_file, cat_file, prefix in files:
        spec_path = Path(spec_file)
        cat_path = Path(cat_file)
        
        if not spec_path.exists():
            print(f"\n{prefix} SPEC file not found")
            continue
        if not cat_path.exists():
            print(f"\n{prefix} CAT file not found")
            continue
        
        try:
            analyze_spec_filter_sheet(spec_path, prefix)
            analyze_matl_control_sheet(spec_path, prefix)
            analyze_cat_component_sheets(cat_path, prefix)
            compare_spec_vs_cat(spec_path, cat_path, prefix)
        except Exception as e:
            print(f"\nError analyzing {prefix}: {e}")
            import traceback
            traceback.print_exc()
        
        break  # Just analyze first file set for now
    
    print(f"\n{'='*80}")
    print("ANALYSIS COMPLETE")
    print(f"{'='*80}")


if __name__ == "__main__":
    main()
