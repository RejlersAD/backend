from openpyxl import load_workbook
from apps.spec_customization.services.exporters import smartplant_config as cfg

wb = load_workbook(cfg.SPEC_TEMPLATE_PATH, data_only=True)
ws = wb['PipingCommodityFilter']
print('Dimensions:', ws.dimensions, 'max_row:', ws.max_row, 'max_col:', ws.max_column)
for row_idx in range(1, min(20, ws.max_row + 1)):
    row_vals = [ws.cell(row=row_idx, column=c).value for c in range(1, min(15, ws.max_column + 1))]
    if any(v is not None for v in row_vals):
        print(f'row {row_idx}: {row_vals}')
