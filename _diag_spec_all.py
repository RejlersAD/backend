from openpyxl import load_workbook
from apps.spec_customization.services.exporters import smartplant_config as cfg

wb = load_workbook(cfg.SPEC_TEMPLATE_PATH, data_only=True)
print(f'{"Sheet":<40} {"Head marker":<15} {"Data rows"}')
print('-' * 80)
for name in wb.sheetnames:
    ws = wb[name]
    head_row = None
    for r in range(1, min(ws.max_row, 60) + 1):
        v = ws.cell(r, 1).value
        if v is not None and str(v).strip() in cfg.HEADER_MARKERS:
            head_row = r
    # Count non-empty rows past row 2
    data_rows = 0
    for r in range(3, ws.max_row + 1):
        row = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        if any(v not in (None, '') for v in row):
            data_rows += 1
    print(f'{name:<40} {str(head_row):<15} {data_rows}')
