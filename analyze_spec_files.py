"""
Analyze SmartPlant 3D SPEC and CAT Excel files to understand structure and data
"""
import openpyxl
import pandas as pd
import sys
import io
from pathlib import Path

# Fix encoding for Windows console
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

def analyze_excel_file(filepath):
    """Analyze an Excel workbook and report its structure"""
    print(f"\n{'='*80}")
    print(f"Analyzing: {filepath.name}")
    print(f"{'='*80}")
    
    wb = openpyxl.load_workbook(filepath, data_only=True)
    print(f"\n📋 Sheets: {', '.join(wb.sheetnames)}")
    
    # Focus on key sheets
    key_sheets = ['PipingCommodityFilter', 'PipingCommodityMatlControlData', 'Part']
    
    for sheet_name in key_sheets:
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f"\n{'─'*80}")
            print(f"📊 Sheet: {sheet_name}")
            print(f"{'─'*80}")
            
            # Get headers (first row)
            headers = []
            for cell in ws[1]:
                if cell.value:
                    headers.append(str(cell.value))
            
            print(f"Columns ({len(headers)}): {', '.join(headers[:10])}{'...' if len(headers) > 10 else ''}")
            
            # Count rows with data
            row_count = 0
            sample_rows = []
            for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if any(cell for cell in row):
                    row_count += 1
                    if row_count <= 3:  # Keep first 3 rows as samples
                        sample_rows.append((idx, row))
            
            print(f"Data rows: {row_count}")
            
            if sample_rows and headers:
                print(f"\n📝 Sample Data (first 3 rows):")
                for row_num, row_data in sample_rows:
                    print(f"\n  Row {row_num}:")
                    for col_name, cell_value in zip(headers, row_data):
                        if cell_value:
                            value_str = str(cell_value)[:50]  # Limit to 50 chars
                            print(f"    {col_name}: {value_str}")
    
    wb.close()


def compare_files(spec_file, cat_file):
    """Compare SPEC and CAT files to understand differences"""
    print(f"\n\n{'='*80}")
    print(f"🔍 COMPARING: {spec_file.stem} (SPEC vs CAT)")
    print(f"{'='*80}")
    
    spec_wb = openpyxl.load_workbook(spec_file, data_only=True)
    cat_wb = openpyxl.load_workbook(cat_file, data_only=True)
    
    # Compare PipingCommodityFilter
    if 'PipingCommodityFilter' in spec_wb.sheetnames and 'PipingCommodityFilter' in cat_wb.sheetnames:
        spec_ws = spec_wb['PipingCommodityFilter']
        cat_ws = cat_wb['PipingCommodityFilter']
        
        spec_rows = sum(1 for row in spec_ws.iter_rows(min_row=2, values_only=True) if any(row))
        cat_rows = sum(1 for row in cat_ws.iter_rows(min_row=2, values_only=True) if any(row))
        
        print(f"\n📊 PipingCommodityFilter:")
        print(f"  SPEC rows: {spec_rows}")
        print(f"  CAT rows:  {cat_rows}")
        print(f"  Difference: {cat_rows - spec_rows} rows")
    
    # Compare PipingCommodityMatlControlData
    if 'PipingCommodityMatlControlData' in spec_wb.sheetnames and 'PipingCommodityMatlControlData' in cat_wb.sheetnames:
        spec_ws = spec_wb['PipingCommodityMatlControlData']
        cat_ws = cat_wb['PipingCommodityMatlControlData']
        
        spec_rows = sum(1 for row in spec_ws.iter_rows(min_row=2, values_only=True) if any(row))
        cat_rows = sum(1 for row in cat_ws.iter_rows(min_row=2, values_only=True) if any(row))
        
        print(f"\n📊 PipingCommodityMatlControlData:")
        print(f"  SPEC rows: {spec_rows}")
        print(f"  CAT rows:  {cat_rows}")
        print(f"  Difference: {cat_rows - spec_rows} rows")
        
        # Sample data from CAT to see description patterns
        print(f"\n📝 Sample CAT Data (descriptions):")
        cat_ws_obj = cat_wb['PipingCommodityMatlControlData']
        headers = [cell.value for cell in cat_ws_obj[1]]
        desc_idx = headers.index('Description') if 'Description' in headers else None
        comp_idx = headers.index('CommodityCode') if 'CommodityCode' in headers else None
        
        if desc_idx is not None:
            for idx, row in enumerate(cat_ws_obj.iter_rows(min_row=2, max_row=6, values_only=True), start=2):
                if row[desc_idx]:
                    comp_code = row[comp_idx] if comp_idx is not None else 'N/A'
                    print(f"  Row {idx}: [{comp_code}] {row[desc_idx]}")
    
    # Compare Part sheet
    if 'Part' in spec_wb.sheetnames and 'Part' in cat_wb.sheetnames:
        spec_ws = spec_wb['Part']
        cat_ws = cat_wb['Part']
        
        spec_rows = sum(1 for row in spec_ws.iter_rows(min_row=2, values_only=True) if any(row))
        cat_rows = sum(1 for row in cat_ws.iter_rows(min_row=2, values_only=True) if any(row))
        
        print(f"\n📊 Part:")
        print(f"  SPEC rows: {spec_rows}")
        print(f"  CAT rows:  {cat_rows}")
        print(f"  Difference: {cat_rows - spec_rows} rows")
    
    spec_wb.close()
    cat_wb.close()


def main():
    """Main analysis function"""
    # Look for files in current directory or specified path
    search_path = Path('.')
    if len(sys.argv) > 1:
        search_path = Path(sys.argv[1])
    
    # Find all Excel files
    excel_files = list(search_path.glob('*.xlsx'))
    
    if not excel_files:
        print("❌ No Excel files found in current directory")
        print(f"Looking in: {search_path.absolute()}")
        return
    
    print(f"Found {len(excel_files)} Excel files")
    
    # Group by prefix
    spec_files = [f for f in excel_files if '_SPEC.xlsx' in f.name]
    cat_files = [f for f in excel_files if '_CAT.xlsx' in f.name]
    
    print(f"\nSPEC files: {len(spec_files)}")
    print(f"CAT files:  {len(cat_files)}")
    
    # Analyze each file
    for file in sorted(excel_files):
        try:
            analyze_excel_file(file)
        except Exception as e:
            print(f"❌ Error analyzing {file.name}: {e}")
    
    # Compare SPEC vs CAT for each pair
    for spec_file in spec_files:
        prefix = spec_file.name.replace('_SPEC.xlsx', '')
        cat_file = search_path / f"{prefix}_CAT.xlsx"
        
        if cat_file.exists():
            try:
                compare_files(spec_file, cat_file)
            except Exception as e:
                print(f"❌ Error comparing {prefix}: {e}")
    
    print(f"\n{'='*80}")
    print("✅ Analysis complete!")
    print(f"{'='*80}")


if __name__ == "__main__":
    main()
