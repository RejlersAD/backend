"""
Detailed analysis of SmartPlant 3D SPEC vs CAT Excel files
Focus on the 3 main issues:
1. PipingCommodityFilter sheet not filling properly (30%)
2. PipingCommodityMatlControlData inappropriate descriptions (20%)
3. Part sheet issues (50%)
"""
import openpyxl
import sys
from pathlib import Path

def analyze_pcf_sheet(spec_file, cat_file, prefix):
    """Analyze PipingCommodityFilter differences"""
    print(f"\n{'='*80}")
    print(f"1. PIPING COMMODITY FILTER ANALYSIS - {prefix}")
    print(f"{'='*80}")
    
    spec_wb = openpyxl.load_workbook(spec_file, data_only=True)
    cat_wb = openpyxl.load_workbook(cat_file, data_only=True)
    
    if 'PipingCommodityFilter' not in spec_wb.sheetnames:
        print("  [SPEC] PipingCommodityFilter sheet NOT FOUND")
        spec_wb.close()
        cat_wb.close()
        return
    
    if 'PipingCommodityFilter' not in cat_wb.sheetnames:
        print("  [CAT] PipingCommodityFilter sheet NOT FOUND")
        spec_wb.close()
        cat_wb.close()
        return
    
    spec_ws = spec_wb['PipingCommodityFilter']
    cat_ws = cat_wb['PipingCommodityFilter']
    
    # Get headers
    spec_headers = [cell.value for cell in spec_ws[1] if cell.value]
    cat_headers = [cell.value for cell in cat_ws[1] if cell.value]
    
    print(f"\n  Headers comparison:")
    print(f"    SPEC columns: {len(spec_headers)}")
    print(f"    CAT columns:  {len(cat_headers)}")
    
    # Count data rows
    spec_rows = []
    for row in spec_ws.iter_rows(min_row=2, values_only=True):
        if any(cell for cell in row):
            spec_rows.append(row)
    
    cat_rows = []
    for row in cat_ws.iter_rows(min_row=2, values_only=True):
        if any(cell for cell in row):
            cat_rows.append(row)
    
    print(f"\n  Data rows comparison:")
    print(f"    SPEC rows: {len(spec_rows)}")
    print(f"    CAT rows:  {len(cat_rows)}")
    print(f"    MISSING:   {len(cat_rows) - len(spec_rows)} rows ({((len(cat_rows) - len(spec_rows)) / len(cat_rows) * 100):.1f}%)")
    
    # Sample CAT data to understand structure
    print(f"\n  CAT Sample Data (first 5 rows):")
    commodity_idx = cat_headers.index('CommodityCode') if 'CommodityCode' in cat_headers else None
    shortcode_idx = cat_headers.index('ShortCode') if 'ShortCode' in cat_headers else None
    size_from_idx = cat_headers.index('FirstSizeFrom') if 'FirstSizeFrom' in cat_headers else None
    size_to_idx = cat_headers.index('FirstSizeTo') if 'FirstSizeTo' in cat_headers else None
    
    for idx, row in enumerate(cat_rows[:5], start=1):
        commodity = row[commodity_idx] if commodity_idx is not None and commodity_idx < len(row) else 'N/A'
        shortcode = row[shortcode_idx] if shortcode_idx is not None and shortcode_idx < len(row) else 'N/A'
        size_from = row[size_from_idx] if size_from_idx is not None and size_from_idx < len(row) else 'N/A'
        size_to = row[size_to_idx] if size_to_idx is not None and size_to_idx < len(row) else 'N/A'
        print(f"    Row {idx}: {commodity} | {shortcode} | Size: {size_from}-{size_to}")
    
    spec_wb.close()
    cat_wb.close()


def analyze_matl_control_sheet(spec_file, cat_file, prefix):
    """Analyze PipingCommodityMatlControlData differences"""
    print(f"\n{'='*80}")
    print(f"2. PIPING COMMODITY MATL CONTROL DATA ANALYSIS - {prefix}")
    print(f"{'='*80}")
    
    spec_wb = openpyxl.load_workbook(spec_file, data_only=True)
    cat_wb = openpyxl.load_workbook(cat_file, data_only=True)
    
    sheet_name = 'PipingCommodityMatlControlData'
    
    if sheet_name not in spec_wb.sheetnames:
        print(f"  [SPEC] {sheet_name} sheet NOT FOUND")
        spec_wb.close()
        cat_wb.close()
        return
    
    if sheet_name not in cat_wb.sheetnames:
        print(f"  [CAT] {sheet_name} sheet NOT FOUND")
        spec_wb.close()
        cat_wb.close()
        return
    
    spec_ws = spec_wb[sheet_name]
    cat_ws = cat_wb[sheet_name]
    
    # Get headers
    spec_headers = [cell.value for cell in spec_ws[1] if cell.value]
    cat_headers = [cell.value for cell in cat_ws[1] if cell.value]
    
    # Count data rows
    spec_rows = []
    for row in spec_ws.iter_rows(min_row=2, values_only=True):
        if any(cell for cell in row):
            spec_rows.append(row)
    
    cat_rows = []
    for row in cat_ws.iter_rows(min_row=2, values_only=True):
        if any(cell for cell in row):
            cat_rows.append(row)
    
    print(f"\n  Data rows comparison:")
    print(f"    SPEC rows: {len(spec_rows)}")
    print(f"    CAT rows:  {len(cat_rows)}")
    print(f"    Difference: {len(cat_rows) - len(spec_rows)} rows")
    
    # Check descriptions
    desc_idx = cat_headers.index('Description') if 'Description' in cat_headers else None
    commodity_idx = cat_headers.index('CommodityCode') if 'CommodityCode' in cat_headers else None
    
    print(f"\n  CAT Description Samples (showing issue with inappropriate descriptions):")
    for idx, row in enumerate(cat_rows[:10], start=1):
        if desc_idx is not None and commodity_idx is not None:
            commodity = row[commodity_idx] if commodity_idx < len(row) else 'N/A'
            desc = row[desc_idx] if desc_idx < len(row) else 'N/A'
            print(f"    Row {idx}: [{commodity}] => {desc}")
    
    print(f"\n  SPEC Description Samples (what we're currently generating):")
    for idx, row in enumerate(spec_rows[:10], start=1):
        if desc_idx is not None and commodity_idx is not None:
            commodity = row[commodity_idx] if commodity_idx < len(row) else 'N/A'
            desc = row[desc_idx] if desc_idx < len(row) else 'N/A'
            print(f"    Row {idx}: [{commodity}] => {desc}")
    
    spec_wb.close()
    cat_wb.close()


def analyze_part_sheet(spec_file, cat_file, prefix):
    """Analyze Part sheet differences"""
    print(f"\n{'='*80}")
    print(f"3. PART SHEET ANALYSIS - {prefix}")
    print(f"{'='*80}")
    
    spec_wb = openpyxl.load_workbook(spec_file, data_only=True)
    cat_wb = openpyxl.load_workbook(cat_file, data_only=True)
    
    if 'Part' not in spec_wb.sheetnames:
        print("  [SPEC] Part sheet NOT FOUND")
        spec_wb.close()
        cat_wb.close()
        return
    
    if 'Part' not in cat_wb.sheetnames:
        print("  [CAT] Part sheet NOT FOUND")
        spec_wb.close()
        cat_wb.close()
        return
    
    spec_ws = spec_wb['Part']
    cat_ws = cat_wb['Part']
    
    # Get headers
    spec_headers = [cell.value for cell in spec_ws[1] if cell.value]
    cat_headers = [cell.value for cell in cat_ws[1] if cell.value]
    
    print(f"\n  Headers comparison:")
    print(f"    SPEC columns: {len(spec_headers)}")
    print(f"    CAT columns:  {len(cat_headers)}")
    print(f"\n  Column names:")
    print(f"    CAT:  {', '.join(cat_headers[:15])}...")
    
    # Count data rows
    spec_rows = []
    for row in spec_ws.iter_rows(min_row=2, values_only=True):
        if any(cell for cell in row):
            spec_rows.append(row)
    
    cat_rows = []
    for row in cat_ws.iter_rows(min_row=2, values_only=True):
        if any(cell for cell in row):
            cat_rows.append(row)
    
    print(f"\n  Data rows comparison:")
    print(f"    SPEC rows: {len(spec_rows)}")
    print(f"    CAT rows:  {len(cat_rows)}")
    print(f"    MISSING:   {len(cat_rows) - len(spec_rows)} rows ({((len(cat_rows) - len(spec_rows)) / len(cat_rows) * 100):.1f}%)")
    
    # Sample CAT data
    print(f"\n  CAT Sample Data (first 5 rows):")
    for idx, row in enumerate(cat_rows[:5], start=1):
        non_empty = [f"{cat_headers[i]}: {cell}" for i, cell in enumerate(row) if cell and i < len(cat_headers)]
        print(f"    Row {idx}: {' | '.join(non_empty[:5])}")
    
    spec_wb.close()
    cat_wb.close()


def main():
    """Main analysis"""
    # File pairs to analyze
    files = [
        ('AA1B-A3_SPEC.xlsx', 'AA1B-A3_CAT.xlsx', 'AA1B-A3'),
        ('AA2A-A3_SPEC.xlsx', 'AA2A-A3_CAT.xlsx', 'AA2A-A3'),
        ('AC2A-A3_SPEC.xlsx', 'AC2A-A3_CAT.xlsx', 'AC2A-A3'),
    ]
    
    print("="*80)
    print("SMARTPLANT 3D SPEC vs CAT GAP ANALYSIS")
    print("="*80)
    print("\nISSUES TO ADDRESS:")
    print("  1. PipingCommodityFilter sheet not filling properly (30% admin work)")
    print("  2. PipingCommodityMatlControlData inappropriate descriptions (20% admin work)")
    print("  3. Part sheet data issues (50% admin work)")
    
    for spec_file, cat_file, prefix in files:
        spec_path = Path(spec_file)
        cat_path = Path(cat_file)
        
        if not spec_path.exists() or not cat_path.exists():
            print(f"\nSkipping {prefix} - files not found")
            continue
        
        try:
            analyze_pcf_sheet(spec_path, cat_path, prefix)
            analyze_matl_control_sheet(spec_path, cat_path, prefix)
            analyze_part_sheet(spec_path, cat_path, prefix)
        except Exception as e:
            print(f"\nError analyzing {prefix}: {e}")
            import traceback
            traceback.print_exc()
    
    print(f"\n{'='*80}")
    print("ANALYSIS COMPLETE")
    print(f"{'='*80}")


if __name__ == "__main__":
    main()
