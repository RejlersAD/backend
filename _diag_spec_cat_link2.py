"""
Pass-2 audit: discover the *true* SPEC ↔ CAT join key.

SP3D convention:
  * CAT sheet name = part-class ShortCode (e.g. 'WeldNeckFlange').
  * Each CAT row keys by PartNumber, with physical attributes only.
  * SPEC sheets carry the ShortCode column that points back to the CAT
    sheet name -> this is the genuine relationship.

So this pass:
  1. Dumps the headers of every CAT sheet (so we can see what columns exist).
  2. Builds the set of CAT sheet-names = part-class ShortCodes.
  3. Counts how many rows in each SPEC sheet's ShortCode column match a CAT
     sheet-name (i.e. "this commodity is realised by part-class X in CAT").
  4. Lists SPEC ShortCodes that have NO matching CAT sheet (broken refs).
"""
from __future__ import annotations
import json
from collections import Counter, defaultdict
from openpyxl import load_workbook

SPEC = "/tmp/LS1E-A3_SPEC.xlsx"
CAT  = "/tmp/LS1E-A3_CAT.xlsx"
HEADER_SCAN = 5


def detect_header(ws):
    best_r, best_n, best = 1, 0, []
    for r in range(1, min(HEADER_SCAN, ws.max_row) + 1):
        vals = [c.value for c in ws[r]]
        n = sum(1 for v in vals if v not in (None, ""))
        if n > best_n:
            best_n, best_r, best = n, r, [str(v).strip() if v else "" for v in vals]
    return best_r, best


def collect(path):
    wb = load_workbook(path, data_only=False)
    out = {}
    for sn in wb.sheetnames:
        ws = wb[sn]
        hr, hdrs = detect_header(ws)
        rows = []
        for r in range(hr + 1, ws.max_row + 1):
            row = {}
            empty = True
            for ci, h in enumerate(hdrs, start=1):
                if not h:
                    continue
                v = ws.cell(row=r, column=ci).value
                row[h] = v
                if v not in (None, ""):
                    empty = False
            if not empty:
                rows.append(row)
        out[sn] = {"headers": [h for h in hdrs if h], "rows": rows}
    wb.close()
    return out


def main():
    spec = collect(SPEC)
    cat  = collect(CAT)

    # ---- 1. CAT headers ----
    print("=" * 78)
    print("CATALOG SHEET HEADERS  (these are SP3D part-class tables)")
    print("=" * 78)
    for sn, info in cat.items():
        print(f"\n[{sn}]  ({len(info['rows'])} rows)")
        print(f"   headers: {info['headers']}")
        if info["rows"]:
            sample = info["rows"][0]
            print(f"   sample : {dict(list(sample.items())[:6])}")

    # ---- 2. CAT part-class ShortCodes = sheet names ----
    cat_classes = set(cat.keys())
    print("\n" + "=" * 78)
    print(f"CAT part-class ShortCodes ({len(cat_classes)}): {sorted(cat_classes)}")
    print("=" * 78)

    # ---- 3. For each SPEC sheet, count ShortCode hits into CAT ----
    print("\nSPEC SHEET  --  ShortCode column matches into CAT sheet names")
    print("=" * 78)
    for sn, info in spec.items():
        if "ShortCode" not in info["headers"]:
            continue
        codes = [str(r.get("ShortCode") or "").strip() for r in info["rows"]]
        codes = [c for c in codes if c]
        c_total = len(codes)
        c_match = sum(1 for c in codes if c in cat_classes)
        unique  = Counter(codes)
        unmatched = [c for c in unique if c not in cat_classes]
        print(f"\n[{sn}]  rows={c_total}  uniqueSC={len(unique)}  "
              f"matched_to_CAT={c_match}  unmatched_unique={len(unmatched)}")
        # top matched
        matched_sorted = sorted([(c, n) for c, n in unique.items() if c in cat_classes],
                                key=lambda x: -x[1])
        for c, n in matched_sorted[:10]:
            print(f"     ✓ {c:<25} x{n}")
        if unmatched:
            print(f"     unmatched ShortCodes (first 10): {unmatched[:10]}")

    # ---- 4. CAT sheets that are NEVER referenced by SPEC ----
    print("\n" + "=" * 78)
    print("CAT sheets never referenced by any SPEC.ShortCode")
    print("=" * 78)
    referenced = set()
    for sn, info in spec.items():
        if "ShortCode" not in info["headers"]:
            continue
        for r in info["rows"]:
            sc = str(r.get("ShortCode") or "").strip()
            if sc in cat_classes:
                referenced.add(sc)
    print(f"  Referenced ({len(referenced)}): {sorted(referenced)}")
    print(f"  Orphan CAT  ({len(cat_classes - referenced)}): {sorted(cat_classes - referenced)}")

    # ---- 5. Are CAT/SPEC linked via secondary keys? -----
    # Check if SPEC.PipingCommodityFilter.CommodityCode appears anywhere in CAT
    print("\n" + "=" * 78)
    print("Cross-reference: SPEC CommodityCode values vs ALL CAT cell values")
    print("=" * 78)
    pcf = spec.get("PipingCommodityFilter", {}).get("rows", [])
    cc_values = set(str(r.get("CommodityCode") or "").strip() for r in pcf)
    cc_values.discard("")
    # collect every CAT cell as string
    cat_cells = set()
    for sn, info in cat.items():
        for row in info["rows"]:
            for v in row.values():
                if v not in (None, ""):
                    cat_cells.add(str(v).strip())
    hits = cc_values & cat_cells
    print(f"  SPEC distinct CommodityCodes : {len(cc_values)}")
    print(f"  Hits in CAT cells anywhere   : {len(hits)}")
    if hits:
        print(f"  Sample hits: {list(hits)[:10]}")
    print("\nDone.")


if __name__ == "__main__":
    main()
