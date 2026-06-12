"""
Audit relationship between LS1E-A3_SPEC.xlsx and LS1E-A3_CAT.xlsx (SP3D Smart 3D).

Soft-coded inspection:
  * Lists every worksheet in both workbooks
  * Detects macros (vbaProject.bin), defined names, formulas, data validations
  * Compares headers across both files (CommodityCode is the canonical join key)
  * Reports overlap / value-set relationships between SPEC and CAT
  * Detects per-row link cardinality (1 SPEC row -> N CAT rows)
"""
from __future__ import annotations
import os
import sys
import zipfile
from collections import Counter, defaultdict
from openpyxl import load_workbook

SPEC = "/tmp/LS1E-A3_SPEC.xlsx"
CAT  = "/tmp/LS1E-A3_CAT.xlsx"

# Soft-coded SP3D join-key candidates (priority order)
JOIN_KEYS = (
    "CommodityCode",
    "ContractorCommodityCode",
    "ReportableCommodityCode",
    "ShortCode",
    "PartNumber",
)

# Soft-coded inspection limits
HEADER_SCAN_ROWS = 5          # scan first N rows for header detection
SAMPLE_VALUES    = 5          # sample N distinct values per column for preview
TRUNC            = 60         # truncate long strings


def truncate(v, n=TRUNC):
    s = str(v) if v is not None else ""
    return s if len(s) <= n else s[: n - 1] + "…"


def inspect_macros_and_defined_names(path: str) -> dict:
    """Detect VBA macros, defined names, external links via raw zip inspection."""
    info = {"macros": False, "external_links": [], "defined_names": [], "sheet_count": 0}
    with zipfile.ZipFile(path) as zf:
        names = zf.namelist()
        info["macros"] = any(n.endswith("vbaProject.bin") for n in names)
        info["external_links"] = [n for n in names if "externalLink" in n]
        info["sheet_count"] = sum(1 for n in names if n.startswith("xl/worksheets/sheet"))
    # defined names via openpyxl
    wb = load_workbook(path, read_only=False, data_only=False)
    try:
        for dn in wb.defined_names:
            try:
                info["defined_names"].append((dn, str(wb.defined_names[dn].value)))
            except Exception:
                info["defined_names"].append((dn, "<unreadable>"))
    except Exception:
        pass
    wb.close()
    return info


def detect_header_row(ws, scan_rows=HEADER_SCAN_ROWS):
    """Return (header_row_idx, headers_list) - row with most non-empty cells in first N rows."""
    best_row, best_count, best_headers = 1, 0, []
    for r in range(1, min(scan_rows, ws.max_row) + 1):
        row_vals = [c.value for c in ws[r]]
        non_empty = [v for v in row_vals if v not in (None, "")]
        if len(non_empty) > best_count:
            best_count = len(non_empty)
            best_row = r
            best_headers = [str(v).strip() if v is not None else "" for v in row_vals]
    return best_row, best_headers


def collect_sheet_data(path: str) -> dict:
    """Return {sheet_name: {headers, data_rows, formula_count, col_values}}."""
    out = {}
    wb = load_workbook(path, read_only=False, data_only=False)
    for sn in wb.sheetnames:
        ws = wb[sn]
        if ws.max_row < 1:
            continue
        header_row, headers = detect_header_row(ws)
        # collect column values
        col_values = defaultdict(list)
        formula_count = 0
        data_rows = 0
        for r in range(header_row + 1, ws.max_row + 1):
            row_has_data = False
            for ci, h in enumerate(headers, start=1):
                if not h:
                    continue
                v = ws.cell(row=r, column=ci).value
                if isinstance(v, str) and v.startswith("="):
                    formula_count += 1
                if v not in (None, ""):
                    row_has_data = True
                    col_values[h].append(v)
            if row_has_data:
                data_rows += 1
        out[sn] = {
            "header_row": header_row,
            "headers": headers,
            "data_rows": data_rows,
            "formula_count": formula_count,
            "col_values": dict(col_values),
        }
    wb.close()
    return out


def find_join_key(headers: list) -> str | None:
    norm = {h.replace(" ", "").lower(): h for h in headers if h}
    for k in JOIN_KEYS:
        if k.replace(" ", "").lower() in norm:
            return norm[k.replace(" ", "").lower()]
    return None


def report(path: str, label: str):
    print(f"\n{'=' * 78}")
    print(f"  {label}: {path}")
    print(f"{'=' * 78}")
    meta = inspect_macros_and_defined_names(path)
    print(f"  Worksheets in zip   : {meta['sheet_count']}")
    print(f"  VBA macros          : {'YES (vbaProject.bin)' if meta['macros'] else 'no'}")
    print(f"  External links      : {len(meta['external_links'])} {meta['external_links'] or ''}")
    print(f"  Defined names       : {len(meta['defined_names'])}")
    for nm, val in meta["defined_names"][:10]:
        print(f"      - {nm}  =  {truncate(val, 80)}")
    return meta


def main():
    if not (os.path.exists(SPEC) and os.path.exists(CAT)):
        print("ERROR: input files missing", SPEC, CAT)
        sys.exit(2)

    spec_meta = report(SPEC, "SPEC")
    cat_meta  = report(CAT,  "CATALOG")

    print("\n" + "=" * 78)
    print("  PER-SHEET STRUCTURE")
    print("=" * 78)
    spec_data = collect_sheet_data(SPEC)
    cat_data  = collect_sheet_data(CAT)

    def dump(label, data):
        print(f"\n--- {label} ({len(data)} sheets) ---")
        for sn, info in data.items():
            jk = find_join_key(info["headers"])
            print(
                f"  [{sn:<40}] hdr_row={info['header_row']:<2} "
                f"cols={len([h for h in info['headers'] if h]):<3} "
                f"rows={info['data_rows']:<5} "
                f"formulas={info['formula_count']:<4} "
                f"join_key={jk or '-'}"
            )

    dump("SPEC", spec_data)
    dump("CATALOG", cat_data)

    # ---- Sheet-name overlap ----
    print("\n" + "=" * 78)
    print("  SHEET-NAME OVERLAP (SPEC ∩ CAT)")
    print("=" * 78)
    common = sorted(set(spec_data) & set(cat_data))
    only_spec = sorted(set(spec_data) - set(cat_data))
    only_cat  = sorted(set(cat_data)  - set(spec_data))
    print(f"  Common sheets ({len(common)}): {common}")
    print(f"  Only in SPEC  ({len(only_spec)}): {only_spec[:10]}")
    print(f"  Only in CAT   ({len(only_cat)}): {only_cat[:10]}")

    # ---- Header overlap on common sheets ----
    print("\n" + "=" * 78)
    print("  COLUMN OVERLAP ON COMMON SHEETS")
    print("=" * 78)
    for sn in common:
        sh = set(h for h in spec_data[sn]["headers"] if h)
        ch = set(h for h in cat_data[sn]["headers"]  if h)
        print(f"\n  [{sn}] spec_cols={len(sh)} cat_cols={len(ch)} "
              f"common={len(sh & ch)} only_spec={len(sh - ch)} only_cat={len(ch - sh)}")

    # ---- KEY RELATIONSHIP: CommodityCode SPEC -> CAT ----
    print("\n" + "=" * 78)
    print("  PRIMARY RELATIONSHIP (Spec ↔ Catalog via CommodityCode)")
    print("=" * 78)

    # Collect all CommodityCodes from each side
    def all_codes(data, key="CommodityCode"):
        codes = defaultdict(list)  # code -> [(sheet, row_idx_within_sheet)]
        for sn, info in data.items():
            if key not in info["col_values"]:
                # search by case-insensitive
                actual = find_join_key(info["headers"])
                if not actual:
                    continue
                vals = info["col_values"].get(actual, [])
            else:
                vals = info["col_values"][key]
            for i, v in enumerate(vals):
                codes[str(v).strip()].append((sn, i))
        return codes

    spec_codes = all_codes(spec_data)
    cat_codes  = all_codes(cat_data)
    spec_set = set(spec_codes) - {""}
    cat_set  = set(cat_codes) - {""}
    in_both  = spec_set & cat_set
    only_in_spec = spec_set - cat_set
    only_in_cat  = cat_set - spec_set

    print(f"  Distinct CommodityCodes in SPEC   : {len(spec_set)}")
    print(f"  Distinct CommodityCodes in CAT    : {len(cat_set)}")
    print(f"  Present in BOTH (linked)          : {len(in_both)}")
    print(f"  Only in SPEC (no catalog parts!)  : {len(only_in_spec)}")
    print(f"  Only in CAT  (unused catalog)     : {len(only_in_cat)}")

    # ---- Cardinality: 1 SPEC code -> N CAT rows ----
    cat_count = Counter()
    for code, rows in cat_codes.items():
        cat_count[code] = len(rows)

    linked = [(c, cat_count[c]) for c in in_both]
    linked.sort(key=lambda x: -x[1])

    print(f"\n  Top 10 SPEC codes by # of CAT rows (size variants):")
    for code, n in linked[:10]:
        print(f"      {code:<40} -> {n:>3} catalog rows")

    if only_in_spec:
        print(f"\n  Sample SPEC codes WITHOUT catalog parts:")
        for c in list(only_in_spec)[:10]:
            print(f"      {c}")
    if only_in_cat:
        print(f"\n  Sample CAT codes WITHOUT spec reference:")
        for c in list(only_in_cat)[:10]:
            print(f"      {c}")

    # ---- Per-sheet drill-down: which SPEC sheet feeds which CAT sheet? ----
    print("\n" + "=" * 78)
    print("  SHEET-LEVEL FEED MAP  (SPEC sheet → CAT sheets that reuse its codes)")
    print("=" * 78)
    # cat codes -> {sheet: set(codes)}
    cat_sheet_codes = defaultdict(set)
    for code, rows in cat_codes.items():
        for sn, _ in rows:
            cat_sheet_codes[sn].add(code)

    for spec_sn, info in spec_data.items():
        jk = find_join_key(info["headers"])
        if not jk:
            continue
        spec_codes_here = set(str(v).strip() for v in info["col_values"].get(jk, []) if v not in (None, ""))
        if not spec_codes_here:
            continue
        matches = []
        for cat_sn, codes_in_cat in cat_sheet_codes.items():
            ov = spec_codes_here & codes_in_cat
            if ov:
                matches.append((cat_sn, len(ov), len(spec_codes_here)))
        if matches:
            matches.sort(key=lambda x: -x[1])
            top = matches[0]
            print(f"  [{spec_sn:<40}] {len(spec_codes_here):>4} codes -> "
                  f"top cat sheet '{top[0]}' ({top[1]}/{top[2]} matched)")

    print("\nDone.")


if __name__ == "__main__":
    main()
